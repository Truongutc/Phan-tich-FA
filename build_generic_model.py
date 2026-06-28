#!/usr/bin/env python3
"""
build_generic_model.py — Generic Stock Model and PDF Report Generator
"""
import os
import sys
import json
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, black, white, grey
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, KeepTogether)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

# Adjust import path
sys.path.append(os.path.dirname(__file__))
from fetch_data import fetch_all, section_to_years, section_to_quarters, get_field_map
import google_drive_uploader

def get_sector_content(sector, ticker, company_name):
    """Generate sector-specific investment content for thesis, risks, moats, pestle, comments."""
    sector_lower = (sector or "").lower()
    
    # --- BANKING ---
    if any(k in sector_lower for k in ["bank", "financial service", "tài chính", "ngân hàng"]):
        return {
            "thesis": [
                f"{company_name} là một trong những ngân hàng tư nhân hàng đầu, với hệ sinh thái tài chính toàn diện và lợi thế về nền tảng công nghệ số.",
                "Tăng trưởng tín dụng vượt ngành kết hợp với NIM cải thiện giúp LNST duy trì đà tăng trưởng ổn định.",
                "Chất lượng tài sản được kiểm soát tốt, tỷ lệ nợ xấu (NPL) ở mức thấp so với bình quân ngành.",
            ],
            "risks": [
                "Rủi ro nợ xấu tăng nếu kinh tế vĩ mô suy yếu và khả năng trả nợ của doanh nghiệp sụt giảm.",
                "Áp lực tăng vốn (CAR) có thể pha loãng EPS nếu ngân hàng phải phát hành cổ phiếu mới.",
                "Cạnh tranh gay gắt từ các ngân hàng quốc doanh và làn sóng fintech đang thu hẹp biên lãi suất.",
            ],
            "moats": {
                "Network Effect": {"score": 4, "desc": "Mạng lưới khách hàng rộng tạo hiệu ứng cộng hưởng giữa các sản phẩm tài chính."},
                "Switching Cost": {"score": 4, "desc": "Chi phí chuyển đổi cao do tích hợp sâu dịch vụ tiền gửi, vay và thanh toán."},
                "Intangible Assets": {"score": 4, "desc": "Thương hiệu ngân hàng uy tín và giấy phép hoạt động là rào cản gia nhập lớn."},
                "Cost Advantage": {"score": 3, "desc": "CASA cao giúp giảm chi phí vốn so với đối thủ."},
                "Efficient Scale": {"score": 3, "desc": "Quy mô tài sản lớn cho phép đầu tư mạnh vào hạ tầng số."},
            },
            "pestle": {
                "Political": "Chính sách tiền tệ của NHNN và định hướng ưu tiên tín dụng xanh tác động trực tiếp đến hoạt động ngân hàng.",
                "Economic": "Tăng trưởng GDP bền vững thúc đẩy nhu cầu tín dụng; lãi suất điều hành là biến số chính.",
                "Social": "Dân số trẻ và tỷ lệ người dùng ngân hàng số tăng nhanh mở ra thị trường bán lẻ tiềm năng.",
                "Technological": "Chuyển đổi số và open banking là xu hướng bắt buộc để duy trì cạnh tranh.",
                "Legal": "Thông tư 06, 16 và các quy định về an toàn vốn Basel II/III tác động đến khẩu vị rủi ro.",
                "Environmental": "Áp lực từ ESG và tín dụng xanh ngày càng ảnh hưởng đến cơ cấu danh mục cho vay.",
            },
            "comments": {
                "overall": f"{company_name} ({ticker}) duy trì vị thế cạnh tranh mạnh trong phân khúc ngân hàng tư nhân với hệ sinh thái dịch vụ tài chính đa dạng.",
                "financial": "Biên lãi ròng (NIM) và tỷ suất sinh lợi vốn chủ (ROE) là 2 chỉ số then chốt cần theo dõi. Tăng trưởng tín dụng và CASA là động lực tăng trưởng chính.",
                "valuation": "Định giá theo P/B là phương pháp phù hợp nhất cho ngân hàng. P/B dưới 1.5x là vùng hấp dẫn tích lũy dài hạn.",
            }
        }

    # --- REAL ESTATE ---
    elif any(k in sector_lower for k in ["real estate", "bất động sản", "property"]):
        return {
            "thesis": [
                f"{company_name} nắm giữ quỹ đất chiến lược tại các vùng kinh tế trọng điểm với tiềm năng tăng giá dài hạn.",
                "Danh mục dự án đa dạng từ nhà ở đến thương mại giúp giảm thiểu rủi ro tập trung.",
                "Chu kỳ phục hồi bất động sản sau giai đoạn thanh lọc tạo cơ hội tăng trưởng doanh thu bàn giao.",
            ],
            "risks": [
                "Rủi ro pháp lý và tiến độ cấp phép dự án là yếu tố gây trì hoãn doanh thu và lợi nhuận.",
                "Lãi suất cao kéo dài làm tăng chi phí tài chính và giảm sức mua của người mua nhà.",
                "Áp lực dòng tiền từ đòn bẩy tài chính cao và trái phiếu đến hạn cần tái cơ cấu.",
            ],
            "moats": {
                "Intangible Assets": {"score": 4, "desc": "Thương hiệu chủ đầu tư uy tín và quỹ đất pháp lý sạch là lợi thế khó sao chép."},
                "Efficient Scale": {"score": 3, "desc": "Quy mô lớn giúp đàm phán tốt hơn với nhà thầu và tiếp cận vốn dễ dàng hơn."},
                "Cost Advantage": {"score": 3, "desc": "Chi phí đất thấp từ giai đoạn tích lũy sớm tạo biên lợi nhuận cao hơn thị trường."},
                "Switching Cost": {"score": 2, "desc": "Người mua nhà có xu hướng trung thành với thương hiệu đã tin dùng."},
                "Network Effect": {"score": 1, "desc": "Hiệu ứng mạng lưới hạn chế trong lĩnh vực bất động sản truyền thống."},
            },
            "pestle": {
                "Political": "Luật Đất đai sửa đổi và chính sách nhà ở xã hội ảnh hưởng trực tiếp đến hoạt động kinh doanh.",
                "Economic": "Lãi suất, tỷ giá và tăng trưởng thu nhập cá nhân là các yếu tố vĩ mô then chốt.",
                "Social": "Nhu cầu nhà ở đô thị hóa tăng cao, đặc biệt phân khúc trung cấp và bình dân.",
                "Technological": "Proptech và số hóa quy trình bán hàng đang thay đổi trải nghiệm khách hàng.",
                "Legal": "Tiến độ cấp phép và thủ tục pháp lý dự án là rào cản lớn nhất của ngành.",
                "Environmental": "Tiêu chuẩn xây dựng xanh (LEED, EDGE) và quy định môi trường ngày càng chặt chẽ.",
            },
            "comments": {
                "overall": f"{company_name} ({ticker}) là chủ đầu tư bất động sản với quỹ đất đáng kể và thương hiệu đã được khẳng định.",
                "financial": "Doanh thu ghi nhận theo tiến độ bàn giao nên cần theo dõi backlog và tỷ lệ hấp thụ thực tế.",
                "valuation": "P/B và NAV là phương pháp định giá phổ biến. Discount to NAV cho thấy mức độ an toàn biên.",
            }
        }

    # --- TECHNOLOGY ---
    elif any(k in sector_lower for k in ["technology", "software", "it ", "công nghệ", "tech"]):
        return {
            "thesis": [
                f"{company_name} hưởng lợi từ làn sóng chuyển đổi số của doanh nghiệp và cơ quan nhà nước tại Việt Nam.",
                "Mô hình doanh thu lặp lại (SaaS/subscription) tạo dòng tiền ổn định và khả năng dự báo cao.",
                "Đội ngũ kỹ sư chất lượng cao với chi phí cạnh tranh so với khu vực là lợi thế xuất khẩu phần mềm.",
            ],
            "risks": [
                "Cạnh tranh từ các nền tảng công nghệ toàn cầu (AWS, Microsoft, Google) thu hẹp không gian thị trường.",
                "Rủi ro giữ chân nhân tài trong bối cảnh cạnh tranh nhân lực công nghệ toàn cầu ngày càng gay gắt.",
                "Chu kỳ đầu tư công nghệ dài, khách hàng có thể trì hoãn quyết định mua trong giai đoạn khó khăn.",
            ],
            "moats": {
                "Switching Cost": {"score": 5, "desc": "Chi phí chuyển đổi hệ thống ERP/core banking rất cao, tạo thế gắn kết lâu dài với khách hàng."},
                "Intangible Assets": {"score": 4, "desc": "Bản quyền phần mềm, IP và chứng chỉ quốc tế là rào cản gia nhập."},
                "Network Effect": {"score": 3, "desc": "Nền tảng số càng nhiều người dùng càng có giá trị và khó thay thế."},
                "Cost Advantage": {"score": 3, "desc": "Chi phí lao động IT cạnh tranh so với khu vực tạo lợi thế xuất khẩu dịch vụ."},
                "Efficient Scale": {"score": 2, "desc": "Quy mô vừa đủ để phục vụ thị trường nội địa nhưng còn hạn chế ở thị trường quốc tế."},
            },
            "pestle": {
                "Political": "Chiến lược chuyển đổi số quốc gia và chính phủ điện tử tạo nhu cầu lớn từ khu vực công.",
                "Economic": "Tăng trưởng kinh tế số và FDI vào lĩnh vực công nghệ thúc đẩy nhu cầu dịch vụ IT.",
                "Social": "Thế hệ trẻ am hiểu công nghệ thúc đẩy nhu cầu fintech, edtech và healthtech.",
                "Technological": "AI, Cloud và Big Data đang tái định nghĩa mô hình kinh doanh và tạo cơ hội mới.",
                "Legal": "Luật An ninh mạng và bảo vệ dữ liệu cá nhân ảnh hưởng đến quy trình phát triển sản phẩm.",
                "Environmental": "Năng lượng tiêu thụ của trung tâm dữ liệu và điện toán đám mây là vấn đề ESG quan trọng.",
            },
            "comments": {
                "overall": f"{company_name} ({ticker}) định vị trong phân khúc công nghệ với tiềm năng tăng trưởng từ chuyển đổi số.",
                "financial": "Biên lợi nhuận gộp (gross margin) cao và tốc độ tăng trưởng ARR là chỉ số quan trọng nhất.",
                "valuation": "P/E forward và EV/EBITDA phù hợp với doanh nghiệp công nghệ. Mức định giá phụ thuộc vào tốc độ tăng trưởng.",
            }
        }

    # --- STEEL / MATERIALS ---
    elif any(k in sector_lower for k in ["steel", "material", "mining", "metal", "thép", "khoáng sản", "nguyên vật liệu", "basic material"]):
        return {
            "thesis": [
                f"{company_name} là doanh nghiệp thép hàng đầu với lợi thế quy mô và khả năng chi phối chuỗi giá trị.",
                "Tiềm năng mở rộng công suất vượt trội giúp đón đầu nhu cầu tiêu dùng và đầu tư hạ tầng phục hồi.",
                "Cơ cấu tài chính an toàn, tỷ lệ đòn bẩy được kiểm soát và dòng tiền hoạt động ổn định.",
            ],
            "risks": [
                "Biến động giá nguyên vật liệu đầu vào (quặng sắt, than coke) làm co hẹp biên lợi nhuận gộp.",
                "Rủi ro tỷ giá và lãi suất có xu hướng tăng trong ngắn hạn ảnh hưởng đến chi phí nhập khẩu.",
                "Cạnh tranh gay gắt từ thép nhập khẩu giá rẻ (Trung Quốc) và doanh nghiệp nội địa.",
            ],
            "moats": {
                "Cost Advantage": {"score": 5, "desc": "Quy mô sản xuất khổng lồ tạo lợi thế chi phí trên mỗi tấn sản phẩm."},
                "Efficient Scale": {"score": 4, "desc": "Hệ thống lò cao tích hợp vertical giúp kiểm soát toàn bộ chuỗi sản xuất."},
                "Intangible Assets": {"score": 3, "desc": "Thương hiệu và mạng lưới phân phối rộng khắp toàn quốc."},
                "Switching Cost": {"score": 2, "desc": "Khách hàng B2B có thể chuyển nhà cung cấp tương đối dễ dàng."},
                "Network Effect": {"score": 1, "desc": "Hiệu ứng mạng lưới hạn chế trong ngành sản xuất."},
            },
            "pestle": {
                "Political": "Chính sách thuế chống phá giá và bảo hộ thép nội địa ảnh hưởng lớn đến cạnh tranh nhập khẩu.",
                "Economic": "Chu kỳ đầu tư hạ tầng, BĐS và giá thép toàn cầu là biến số quan trọng nhất.",
                "Social": "Nhu cầu xây dựng nhà ở và cơ sở hạ tầng đô thị hóa tạo thị trường nội địa lớn.",
                "Technological": "Công nghệ EAF (lò điện) thân thiện môi trường là xu hướng chuyển dịch của ngành.",
                "Legal": "Quy định phát thải CO2 và tiêu chuẩn môi trường ngày càng chặt chẽ với ngành thép.",
                "Environmental": "Ngành thép là nguồn phát thải CO2 lớn — chuyển đổi sang thép xanh là áp lực ESG dài hạn.",
            },
            "comments": {
                "overall": f"{company_name} ({ticker}) là cổ phiếu chu kỳ ngành thép — giá cổ phiếu phụ thuộc nhiều vào giá thép thế giới và chu kỳ đầu tư.",
                "financial": "Biên gộp và EBITDA là chỉ tiêu theo dõi quan trọng. Chi phí nguyên liệu/giá thép là biến số then chốt.",
                "valuation": "EV/EBITDA và P/Book là phương pháp phổ biến cho ngành vật liệu cơ bản.",
            }
        }

    # --- CONSUMER / RETAIL ---
    elif any(k in sector_lower for k in ["consumer", "retail", "food", "beverage", "bán lẻ", "tiêu dùng", "thực phẩm"]):
        return {
            "thesis": [
                f"{company_name} hưởng lợi từ xu hướng tăng tiêu dùng nội địa và mở rộng tầng lớp trung lưu tại Việt Nam.",
                "Mạng lưới phân phối rộng và thương hiệu được nhận diện cao tạo rào cản gia nhập cho đối thủ mới.",
                "Chiến lược đa dạng hóa danh mục sản phẩm và mở rộng kênh bán online giúp tăng trưởng bền vững.",
            ],
            "risks": [
                "Áp lực lạm phát làm tăng chi phí nguyên liệu và co hẹp biên lợi nhuận nếu không thể tăng giá.",
                "Cạnh tranh từ hàng ngoại nhập và thương hiệu quốc tế đang thâm nhập thị trường Việt Nam.",
                "Thay đổi thói quen tiêu dùng và xu hướng healthy làm giảm nhu cầu một số sản phẩm truyền thống.",
            ],
            "moats": {
                "Intangible Assets": {"score": 5, "desc": "Thương hiệu mạnh, được tin dùng nhiều thế hệ là lợi thế cạnh tranh bền vững."},
                "Cost Advantage": {"score": 4, "desc": "Quy mô sản xuất lớn cho phép tối ưu chi phí so với các đối thủ nhỏ hơn."},
                "Network Effect": {"score": 3, "desc": "Mạng lưới phân phối bán lẻ trải rộng tạo hàng rào cho đối thủ mới."},
                "Switching Cost": {"score": 3, "desc": "Thói quen tiêu dùng và lòng trung thành thương hiệu tạo chi phí chuyển đổi tâm lý."},
                "Efficient Scale": {"score": 3, "desc": "Hệ thống sản xuất và logistics được tối ưu hóa ở quy mô lớn."},
            },
            "pestle": {
                "Political": "Chính sách xuất khẩu nông sản và quy định an toàn thực phẩm tác động đến hoạt động.",
                "Economic": "Thu nhập khả dụng và sức mua người tiêu dùng là yếu tố vĩ mô then chốt.",
                "Social": "Xu hướng sức khỏe, organic và bền vững định hình lại nhu cầu tiêu dùng.",
                "Technological": "Thương mại điện tử và omnichannel đang thay đổi kênh phân phối truyền thống.",
                "Legal": "Quy định nhãn mác, thành phần và quảng cáo thực phẩm ngày càng chặt chẽ.",
                "Environmental": "Áp lực giảm bao bì nhựa và hướng tới chuỗi cung ứng bền vững từ nhà bán lẻ lớn.",
            },
            "comments": {
                "overall": f"{company_name} ({ticker}) là doanh nghiệp tiêu dùng nội địa với nền tảng thương hiệu vững chắc.",
                "financial": "Biên gộp, chi phí bán hàng/doanh thu (SG&A ratio) và tăng trưởng SSSG là chỉ tiêu quan trọng.",
                "valuation": "P/E và EV/EBITDA là phương pháp phổ biến. So sánh với P/E ngành tiêu dùng khu vực.",
            }
        }

    # --- DEFAULT / OTHER ---
    else:
        return {
            "thesis": [
                f"{company_name} có vị thế cạnh tranh tốt trong ngành và tiềm năng tăng trưởng dài hạn.",
                "Dòng tiền hoạt động ổn định và cơ cấu tài chính lành mạnh hỗ trợ chiến lược mở rộng.",
                "Đội ngũ quản lý có kinh nghiệm với track record thực thi tốt kế hoạch kinh doanh.",
            ],
            "risks": [
                "Rủi ro kinh tế vĩ mô và biến động lãi suất có thể ảnh hưởng đến chi phí vốn.",
                "Cạnh tranh trong ngành ngày càng gay gắt đòi hỏi đầu tư liên tục vào năng lực cạnh tranh.",
                "Rủi ro thực thi chiến lược và phụ thuộc vào một số khách hàng/sản phẩm chủ lực.",
            ],
            "moats": {
                "Intangible Assets": {"score": 3, "desc": "Thương hiệu và uy tín trong ngành được xây dựng qua nhiều năm hoạt động."},
                "Cost Advantage": {"score": 3, "desc": "Tối ưu hóa quy trình vận hành giúp duy trì biên lợi nhuận cạnh tranh."},
                "Switching Cost": {"score": 2, "desc": "Mức độ gắn kết với khách hàng ở mức trung bình."},
                "Efficient Scale": {"score": 3, "desc": "Quy mô phù hợp để vận hành hiệu quả trong thị trường nội địa."},
                "Network Effect": {"score": 2, "desc": "Hiệu ứng mạng lưới còn hạn chế trong mô hình kinh doanh hiện tại."},
            },
            "pestle": {
                "Political": "Chính sách vĩ mô và định hướng phát triển ngành của Chính phủ tác động đến môi trường kinh doanh.",
                "Economic": "Tăng trưởng GDP, lãi suất và lạm phát ảnh hưởng đến nhu cầu và chi phí hoạt động.",
                "Social": "Thay đổi xu hướng tiêu dùng và nhân khẩu học tạo cả cơ hội và thách thức.",
                "Technological": "Chuyển đổi số và tự động hóa là xu hướng quan trọng cần đầu tư bắt kịp.",
                "Legal": "Tuân thủ các quy định pháp luật và tiêu chuẩn ngành là yêu cầu cơ bản.",
                "Environmental": "Áp lực ESG và báo cáo phát triển bền vững ngày càng trở thành yêu cầu của nhà đầu tư.",
            },
            "comments": {
                "overall": f"{company_name} ({ticker}) có nền tảng kinh doanh ổn định với tiềm năng tăng trưởng trong trung dài hạn.",
                "financial": "Tỷ suất sinh lợi (ROE, ROA) và tốc độ tăng trưởng lợi nhuận là chỉ số then chốt cần theo dõi.",
                "valuation": "P/E và P/B là phương pháp định giá phổ biến, cần so sánh với mức bình quân ngành.",
            }
        }


def build_generic(ticker):
    print(f"\n=== Running Automated Generic Analysis for {ticker} ===")
    
    # 1. Fetch Data
    try:
        data = fetch_all(ticker, use_cache=True)
    except Exception as e:
        print(f"Error fetching data for {ticker}: {e}")
        return False
        
    # Check details
    details_url = f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}"
    try:
        r = google_drive_uploader.get_drive_service() # just checking if we can initialize to verify setup, but we make API request manually
        session = fetch_all.__globals__.get('_session')()
        det_r = session.get(details_url, timeout=10)
        det_data = det_r.json().get("data", {})
    except Exception as e:
        print(f"Could not fetch company details: {e}")
        det_data = {}
        
    company_name = det_data.get("enOrganName", det_data.get("organName", f"Công ty Cổ phần {ticker}"))
    current_price = det_data.get("currentPrice", 10000)
    shares = det_data.get("numberOfSharesMktCap", 1000000)
    market_cap = det_data.get("marketCap", current_price * shares)
    sector = det_data.get("sector", "General")
    
    is_recs = section_to_years(data, "INCOME_STATEMENT")
    bs_recs = section_to_years(data, "BALANCE_SHEET")
    cf_recs = section_to_years(data, "CASH_FLOW")
    
    # Determine years available
    available_years = sorted(list(set([r.get("yearReport") for r in is_recs if r.get("yearReport")])))
    if not available_years:
        print("No historical data years found!")
        return False
        
    # We want up to 5 historical years
    hist_years = available_years[-5:]
    print(f"Historical years: {hist_years}")
    
    # Helper to retrieve data safely
    def get_val(recs, year, fields):
        if not isinstance(fields, list):
            fields = [fields]
        for r in recs:
            if r.get("yearReport") == year:
                for f in fields:
                    val = r.get(f)
                    if val is not None:
                        return val
        return 0

    # Fields Mapping
    # Rev: isa3 or isa1, COGS: isa4 or isa2, GP: isa5, NPAT: isa22 or isa20, EPS: isa23
    rev_fields = ["isa3", "isa1"]
    cogs_fields = ["isa4", "isa2"]
    gp_fields = ["isa5"]
    npat_fields = ["isa22", "isa20"]
    eps_fields = ["isa23"]
    
    # BS Fields: Assets: bsa53 or bsa50, Equity: bsa78 or bsa75, Cash: bsa2 or bsa1
    assets_fields = ["bsa53", "bsa50"]
    equity_fields = ["bsa78", "bsa75"]
    cash_fields = ["bsa2", "bsa1"]
    debt_fields = ["bsa56", "bsa71"] # short + long debt
    
    # Extracted data arrays
    rev_hist = [get_val(is_recs, y, rev_fields) / 1e9 for y in hist_years]
    gp_hist = [get_val(is_recs, y, gp_fields) / 1e9 for y in hist_years]
    npat_hist = [get_val(is_recs, y, npat_fields) / 1e9 for y in hist_years]
    eps_hist = [get_val(is_recs, y, eps_fields) for y in hist_years]
    
    assets_hist = [get_val(bs_recs, y, assets_fields) / 1e9 for y in hist_years]
    equity_hist = [get_val(bs_recs, y, equity_fields) / 1e9 for y in hist_years]
    cash_hist = [get_val(bs_recs, y, cash_fields) / 1e9 for y in hist_years]
    debt_hist = [(get_val(bs_recs, y, "bsa56") + get_val(bs_recs, y, "bsa71")) / 1e9 for y in hist_years]
    
    # 2. Simple Forecast Model
    # Forecast next 3 years
    fc_years = [hist_years[-1] + 1, hist_years[-1] + 2, hist_years[-1] + 3]
    
    # Calculate simple growth averages
    rev_growths = []
    for i in range(1, len(rev_hist)):
        if rev_hist[i-1] > 0:
            rev_growths.append((rev_hist[i] / rev_hist[i-1]) - 1)
    avg_rev_growth = sum(rev_growths) / len(rev_growths) if rev_growths else 0.08
    # Clamp growth to realistic [5%, 15%]
    avg_rev_growth = max(0.05, min(0.15, avg_rev_growth))
    
    # Average GP Margin
    gp_margins = [gp_hist[i] / rev_hist[i] if rev_hist[i] > 0 else 0 for i in range(len(rev_hist))]
    avg_gp_margin = sum(gp_margins) / len(gp_margins) if gp_margins else 0.15
    
    # Average NP Margin
    np_margins = [npat_hist[i] / rev_hist[i] if rev_hist[i] > 0 else 0 for i in range(len(rev_hist))]
    avg_np_margin = sum(np_margins) / len(np_margins) if np_margins else 0.05
    
    # Forecast projections
    rev_fc = []
    gp_fc = []
    npat_fc = []
    eps_fc = []
    equity_fc = []
    
    last_rev = rev_hist[-1]
    last_equity = equity_hist[-1]
    
    for y in fc_years:
        next_rev = last_rev * (1 + avg_rev_growth)
        next_gp = next_rev * avg_gp_margin
        next_npat = next_rev * avg_np_margin
        # Basic equity expansion: Retain 70% of earnings
        next_equity = last_equity + (next_npat * 0.7)
        
        rev_fc.append(next_rev)
        gp_fc.append(next_gp)
        npat_fc.append(next_npat)
        equity_fc.append(next_equity)
        
        # Estimate EPS assuming constant share count
        est_eps = (next_npat * 1e9) / (shares if shares else 1)
        eps_fc.append(est_eps)
        
        last_rev = next_rev
        last_equity = next_equity
        
    all_years = hist_years + fc_years
    all_rev = rev_hist + rev_fc
    all_npat = npat_hist + npat_fc
    all_eps = eps_hist + eps_fc
    
    # Simple Valuation (Target P/E and P/B)
    latest_eps = eps_hist[-1] if eps_hist[-1] else 1000
    current_pe = current_price / latest_eps if latest_eps > 0 else 10
    target_pe = 10.0 # Standard defensive P/E
    target_pb = 1.2  # Standard P/B
    
    est_fair_pe = eps_fc[0] * target_pe
    
    # 3. Create Excel Model
    wb = openpyxl.Workbook()
    
    # Cover Sheet
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover["A2"] = f"BÁO CÁO PHÂN TÍCH CỔ PHIẾU {ticker}"
    ws_cover["A2"].font = Font(name="Calibri", size=18, bold=True)
    ws_cover["A3"] = company_name
    ws_cover["A3"].font = Font(name="Calibri", size=12, italic=True)
    
    ws_cover["A5"] = "Giá hiện tại:"
    ws_cover["B5"] = current_price
    ws_cover["B5"].number_format = "#,##0"
    
    ws_cover["A6"] = "Vốn hóa (tỷ):"
    ws_cover["B6"] = market_cap / 1e9
    ws_cover["B6"].number_format = "#,##0"
    
    # P&L Sheet
    ws_pl = wb.create_sheet(title="P&L")
    ws_pl.views.sheetView[0].showGridLines = True
    ws_pl.append(["Chỉ tiêu"] + [f"{y}A" for y in hist_years] + [f"{y}E" for y in fc_years])
    ws_pl.append(["Doanh thu thuần (tỷ)"] + rev_hist + rev_fc)
    ws_pl.append(["Lợi nhuận gộp (tỷ)"] + gp_hist + gp_fc)
    ws_pl.append(["Lợi nhuận sau thuế (tỷ)"] + npat_hist + npat_fc)
    ws_pl.append(["EPS (VND)"] + eps_hist + eps_fc)
    
    for row in ws_pl.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                
    # Style and save
    out_dir = os.path.join(os.path.dirname(__file__), "Bao cao", ticker)
    os.makedirs(out_dir, exist_ok=True)
    
    month_str = datetime.now().strftime("%Y-%m")
    excel_path = os.path.join(out_dir, f"{ticker}_Model_{month_str}.xlsx")
    pdf_path = os.path.join(out_dir, f"{ticker}_Phan_Tich_{month_str}.pdf")
    
    wb.save(excel_path)
    print(f"[OK] Excel saved locally at: {excel_path}")
    
    # 4. Create PDF Report using ReportLab
    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
                            
    styles = getSampleStyleSheet()
    
    # Modify default styles for safer rendering
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=HexColor("#1A365D"),
        spaceAfter=15
    )
    
    h1_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=HexColor("#2B6CB0"),
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10.5,
        leading=14,
        textColor=HexColor("#2D3748"),
        spaceAfter=10
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=white,
        alignment=TA_CENTER
    )
    
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=HexColor("#2D3748"),
        alignment=TA_CENTER
    )

    story = []
    
    # Title Page / Header
    story.append(Paragraph(f"BÁO CÁO PHÂN TÍCH DOANH NGHIỆP: {ticker}", title_style))
    story.append(Paragraph(company_name, ParagraphStyle('Subtitle', parent=body_style, fontName='Helvetica-Oblique', fontSize=12, textColor=HexColor("#4A5568"))))
    story.append(Spacer(1, 15))
    
    # Summary Table
    summary_data = [
        [Paragraph("Mã cổ phiếu", table_header_style), Paragraph("Giá hiện tại (VND)", table_header_style), Paragraph("Vốn hóa thị trường", table_header_style), Paragraph("Ngành", table_header_style)],
        [Paragraph(ticker, table_cell_style), Paragraph(f"{current_price:,.0f}", table_cell_style), Paragraph(f"{market_cap/1e9:,.1f} tỷ VND", table_cell_style), Paragraph(sector, table_cell_style)]
    ]
    t = Table(summary_data, colWidths=[35*mm, 45*mm, 45*mm, 45*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1A365D")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#CBD5E0")),
    ]))
    story.append(t)
    story.append(Spacer(1, 20))
    
    # 1. Luận điểm đầu tư
    story.append(Paragraph("1. Luận điểm đầu tư chính", h1_style))
    story.append(Paragraph(f"Cổ phiếu {ticker} ({company_name}) đang giao dịch tại mức giá {current_price:,.0f} VND. Dựa trên phân tích kết quả kinh doanh lịch sử và các giả định tăng trưởng thận trọng, chúng tôi đánh giá doanh nghiệp có tiềm năng duy trì đà tăng trưởng doanh thu ở mức {avg_rev_growth*100:.1f}% mỗi năm trong giai đoạn tới nhờ vào sự phục hồi chung của ngành {sector}.", body_style))
    
    # 2. Tóm tắt tài chính
    story.append(Paragraph("2. Tóm tắt số liệu tài chính", h1_style))
    
    # Financial Table Headers
    fin_headers = [Paragraph("Chỉ tiêu", table_header_style)] + [Paragraph(f"{y}A", table_header_style) for y in hist_years] + [Paragraph(f"{y}E", table_header_style) for y in fc_years]
    
    row_rev = [Paragraph("Doanh thu (tỷ)", table_cell_style)] + [Paragraph(f"{v:,.1f}", table_cell_style) for v in all_rev]
    row_npat = [Paragraph("LNST (tỷ)", table_cell_style)] + [Paragraph(f"{v:,.1f}", table_cell_style) for v in all_npat]
    row_eps = [Paragraph("EPS (VND)", table_cell_style)] + [Paragraph(f"{v:,.0f}", table_cell_style) for v in all_eps]
    
    fin_table_data = [fin_headers, row_rev, row_npat, row_eps]
    
    # Calculate column widths: first col is wider, rest are equal
    num_cols = len(all_years) + 1
    col_widths = [45*mm] + [125*mm / len(all_years)] * len(all_years)
    
    t_fin = Table(fin_table_data, colWidths=col_widths)
    t_fin.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor("#2B6CB0")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#E2E8F0")),
    ]))
    story.append(t_fin)
    story.append(Spacer(1, 20))
    
    # 3. Định giá & Rủi ro
    story.append(Paragraph("3. Định giá & Khuyến nghị", h1_style))
    story.append(Paragraph(f"Áp dụng phương pháp định giá P/E mục tiêu ở mức {target_pe}x đối với EPS dự phóng năm {fc_years[0]}, giá trị hợp lý của cổ phiếu {ticker} ước tính đạt khoảng {est_fair_pe:,.0f} VND/CP.", body_style))
    
    story.append(Paragraph("4. Rủi ro đầu tư", h1_style))
    story.append(Paragraph("- Rủi ro biến động tỷ giá và lãi suất tăng ảnh hưởng chi phí vốn.", body_style))
    story.append(Paragraph("- Rủi ro cạnh tranh gia tăng trong phân khúc thị trường nội địa.", body_style))
    
    doc.build(story)
    print(f"[OK] PDF saved locally at: {pdf_path}")
    
    # 5. Upload to Google Drive
    gdrive_excel_url = None
    gdrive_pdf_url = None
    
    try:
        print("[GDrive] Starting upload to Google Drive...")
        _, gdrive_excel_url = google_drive_uploader.upload_file(excel_path)
        _, gdrive_pdf_url = google_drive_uploader.upload_file(pdf_path)
    except Exception as e:
        print(f"[GDrive] Upload failed: {e}")
        
    # Fetch historical ratios for the years
    pe_hist_vals = []
    pb_hist_vals = []
    pe_quarters = []
    pb_quarters = []
    quarter_labels = []
    try:
        r_ratios = session.get(f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial", timeout=10)
        ratios_data = r_ratios.json().get("data", [])
        
        # All quarters sorted
        all_quarters = sorted([x for x in ratios_data if x.get("quarter") in (1,2,3,4)], key=lambda x: (x.get("year", 0), x.get("quarter", 0)))
        pe_quarters = [round(x.get("pe", 0) or 0, 2) for x in all_quarters]
        pb_quarters = [round(x.get("pb", 0) or 0, 2) for x in all_quarters]
        quarter_labels = [f"{x.get('year')}-Q{x.get('quarter')}" for x in all_quarters]

        annual_ratios = sorted([x for x in ratios_data if x.get("quarter") == 4], key=lambda x: x.get("year", 0))[-5:]
        pe_hist_vals = [round(x.get("pe", 0) or 0, 1) for x in annual_ratios]
        pb_hist_vals = [round(x.get("pb", 0) or 0, 2) for x in annual_ratios]
    except Exception as e:
        print(f"[WARN] Could not fetch ratio history: {e}")

    # 6. Save JSON Summary data
    sector_content = get_sector_content(sector, ticker, company_name)
    
    # ── Quarterly Data Processing ──
    is_q_recs = section_to_quarters(data, "INCOME_STATEMENT")
    bs_q_recs = section_to_quarters(data, "BALANCE_SHEET")
    
    q_keys = []
    for r in is_q_recs:
        yr = r.get("yearReport")
        qt = r.get("quarter") or r.get("lengthReport")
        if yr and qt in (1, 2, 3, 4):
            q_keys.append(f"{yr}-Q{qt}")
    q_keys = sorted(list(set(q_keys)))
    
    income_quarterly_records = []
    q_labels = []
    q_nims, q_ldrs, q_casas, q_npls = [], [], [], [] # nim, ldr, casa, npl maps to GP margin, Debt/Equity, ROA, Asset Turnover for non-banks
    
    is_bank = ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in (sector or "").lower()
    
    for q_key in q_keys[-12:]:
        yr, qt = map(int, q_key.split("-Q"))
        
        q_is = {}
        for r in is_q_recs:
            r_qt = r.get("quarter") or r.get("lengthReport")
            if r.get("yearReport") == yr and r_qt == qt:
                q_is = r
                break
                
        q_bs = {}
        for r in bs_q_recs:
            r_qt = r.get("quarter") or r.get("lengthReport")
            if r.get("yearReport") == yr and r_qt == qt:
                q_bs = r
                break
        
        if is_bank:
            nii_q = q_is.get("isb27", q_is.get("isb22", 0)) or 0
            npat_q = q_is.get("isa22", q_is.get("isa20", 0)) or 0
            income_quarterly_records.append({
                "yearReport": yr,
                "quarter": qt,
                "nii": round(nii_q / 1e9, 2),
                "npat": round(npat_q / 1e9, 2),
            })
            
            q_loans = q_bs.get("bsb103", q_bs.get("bsb24", 0)) or 0
            q_deps = q_bs.get("bsb113", q_bs.get("bsb33", 0)) or 0
            q_papers = q_bs.get("bsb116", q_bs.get("bsb34", 0)) or 0
            q_casa = q_bs.get("bsb114", q_bs.get("bsb31", 0)) or 0
            q_npl = q_bs.get("bsb105", q_bs.get("bsb27", 0)) or 0
            
            q_nim = (nii_q * 4) / ((q_loans + q_papers) or 1)
            q_ldr = q_loans / ((q_deps + q_papers) or 1)
            q_casa = q_casa / (q_deps or 1)
            q_npl = q_npl / (q_loans or 1)
            
            q_nims.append(max(0.01, min(0.10, q_nim)))
            q_ldrs.append(max(0.40, min(1.30, q_ldr)))
            q_casas.append(max(0.05, min(0.60, q_casa)))
            q_npls.append(max(0.001, min(0.10, q_npl)))
        else:
            rev_q = q_is.get("isa3", q_is.get("isa1", 0)) or 0
            npat_q = q_is.get("isa22", q_is.get("isa20", 0)) or 0
            gp_q = q_is.get("isa5", 0) or 0
            if not gp_q:
                cogs_q = abs(q_is.get("isa4", 0) or 0)
                gp_q = rev_q - cogs_q if rev_q > cogs_q else rev_q * 0.25
                
            income_quarterly_records.append({
                "yearReport": yr,
                "quarter": qt,
                "nii": round(rev_q / 1e9, 2),
                "npat": round(npat_q / 1e9, 2),
            })
            
            q_debt = (q_bs.get("bsa56", 0) or 0) + (q_bs.get("bsa71", 0) or 0)
            q_equity = q_bs.get("bsa78", q_bs.get("bsa75", 1e9)) or 1e9
            q_assets = q_bs.get("bsa53", q_bs.get("bsa50", 1e9)) or 1e9
            
            q_gp_margin = gp_q / (rev_q or 1)
            q_debt_equity = q_debt / (q_equity or 1)
            q_roa = (npat_q * 4) / q_assets
            q_asset_turnover = (rev_q * 4) / q_assets
            
            q_nims.append(max(0.0, min(1.0, q_gp_margin)))
            q_ldrs.append(max(0.0, min(5.0, q_debt_equity)))
            q_casas.append(max(-0.5, min(1.0, q_roa)))
            q_npls.append(max(0.0, min(10.0, q_asset_turnover)))
            
        q_labels.append(q_key)
        
    summary_json = {
        "ticker": ticker,
        "companyName": company_name,
        "sector": sector,
        "currentPrice": current_price,
        "marketCap": market_cap,
        "shares": shares,
        "gdriveExcelUrl": gdrive_excel_url,
        "gdrivePdfUrl": gdrive_pdf_url,
        "pe_hist": pe_hist_vals,
        "pb_hist": pb_hist_vals,
        "pe_quarters": pe_quarters,
        "pb_quarters": pb_quarters,
        "quarter_labels": quarter_labels,
        "income_quarterly": income_quarterly_records,
        "ratios_quarterly": {
            "quarters": q_labels,
            "nim": [round(x, 4) for x in q_nims],
            "ldr": [round(x, 4) for x in q_ldrs],
            "casa": [round(x, 4) for x in q_casas],
            "npl": [round(x, 4) for x in q_npls]
        },

        # Sector-aware qualitative content
        "thesis": sector_content.get("thesis", []),
        "risks": sector_content.get("risks", []),
        "moats": sector_content.get("moats", {}),
        "pestle": sector_content.get("pestle", {}),
        "comments": sector_content.get("comments", {}),

        # Self-calculated ratios (avoiding Vietcap ratio errors where possible)
        "ratios": {
            "nim": [round(((get_val(is_recs, y, ["isb27", "isb22"]) / 1e9) / (get_val(bs_recs, y, assets_fields) / 1e9 or 1)), 4) if y in hist_years else 0.035 for y in all_years] if (ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in sector.lower()) else [],
            "roe": [round(npat_hist[i]/equity_hist[i], 4) if i < len(npat_hist) else 0.15 for i in range(len(all_years))],
            "roa": [round(npat_hist[i]/assets_hist[i], 4) if i < len(npat_hist) else 0.02 for i in range(len(all_years))],
            "npl": [round((get_val(bs_recs, y, ["bsb105", "bsb27"]) or 0) / (get_val(bs_recs, y, ["bsb103", "bsb24"]) or 1), 4) if y in hist_years else 0.012 for y in all_years] if (ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in sector.lower()) else [],
            "ldr": [round((get_val(bs_recs, y, ["bsb103", "bsb24"]) / 1e9) / (((get_val(bs_recs, y, ["bsb113", "bsb33"]) or 0) + (get_val(bs_recs, y, ["bsb116", "bsb34", "bsb36"]) or 0)) / 1e9 or 1), 3) if y in hist_years else 0.83 for y in all_years] if (ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in sector.lower()) else [],
            "casa": [round((get_val(bs_recs, y, ["bsb114", "bsb31"]) or 0) / (get_val(bs_recs, y, ["bsb113", "bsb33"]) or 1), 3) if y in hist_years else 0.325 for y in all_years] if (ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in sector.lower()) else [],
            "gross_margin": [round(gp_hist[i]/rev_hist[i], 4) if i < len(gp_hist) else avg_gp_margin for i in range(len(all_years))] if not (ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in sector.lower()) else [],
            "debt_to_equity": [round(debt_hist[i]/equity_hist[i], 2) if i < len(debt_hist) else 0.5 for i in range(len(all_years))] if not (ticker in ["TCB", "VCB", "MBB", "BID", "CTG", "ACB", "VPB"] or "bank" in sector.lower()) else [],
        },
        "data": {
            "years": all_years,
            "revenue": all_rev,
            "npat": all_npat,
            "eps": all_eps,
            "equity": equity_hist + equity_fc
        }
    }

    data_dir = os.path.join(os.path.dirname(__file__), "data")
    os.makedirs(data_dir, exist_ok=True)
    json_path = os.path.join(data_dir, f"{ticker}.json")
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary_json, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON Summary saved at: {json_path}")
    
    return True


if __name__ == "__main__":
    if len(sys.argv) > 1:
        ticker = sys.argv[1].upper()
        build_generic(ticker)
    else:
        print("Please provide a stock ticker. Example: python build_generic_model.py TCB")
