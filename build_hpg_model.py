#!/usr/bin/env python3
"""
HPG (Hoà Phát Group) — Excel Model + PDF Report Generator
Q2 2026 | Framework FA 6 tầng + Skill Thép
"""

import os
import re
import json
import math
import subprocess
import tempfile
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from scipy.interpolate import PchipInterpolator
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, black, white, grey
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, Image, KeepTogether)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests
import statistics as stats

# ── OUTPUT ──────────────────────────────────────────────────────────────────
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Bao cao", "HPG")
# Tên file theo ngày chạy thực tế (mã_Model_năm_tháng_ngày) thay vì hằng số cố định "2026-06" —
# trước đây tên file KHÔNG đổi dù chạy lại nhiều lần, không biết file trên Drive là cũ hay mới.
MONTH = datetime.now().strftime("%Y_%m_%d")
EXCEL_FILE = os.path.join(OUT_DIR, f"HPG_Model_{MONTH}.xlsx")
PDF_FILE   = os.path.join(OUT_DIR, f"HPG_Phan_Tich_{MONTH}.pdf")
CHART_DIR  = os.path.join(OUT_DIR, "charts")
os.makedirs(CHART_DIR, exist_ok=True)

# ── DATA ────────────────────────────────────────────────────────────────────
TICKER = "HPG"
COMPANY = "CTCP Tập đoàn Hòa Phát"
EXCHANGE = "HOSE"
INDUSTRY = "Thép (Sản xuất Commodity)"
PRICE    = 23600      # VND, late June 2026
SHARES   = 8444740856 # post 1.1:1 split May 2026
MARKET_CAP = PRICE * SHARES  # ~199,260 tỷ
EV_ESTIMATE = 259260e9  # ~259,260 tỷ VND

from fetch_data import fetch_all, section_to_years, section_to_quarters, get_field_map, cumulative_actual_quarters, blend_annual_estimate
FIN_DATA = fetch_all(TICKER)

# ── Income Statement fields ──
IS = {m["field"]: m["titleVi"] for m in FIN_DATA["metrics"]["INCOME_STATEMENT"]}
BS = {m["field"]: m["titleVi"] for m in FIN_DATA["metrics"]["BALANCE_SHEET"]}
CF = {m["field"]: m["titleVi"] for m in FIN_DATA["metrics"]["CASH_FLOW"]}

def get_yr(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            return v / 1e9 if v is not None and field != "isa23" else (v if v is not None else 0)
    return 0

def get_yr_raw(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            return v if v is not None else 0
    return 0

is_recs = section_to_years(FIN_DATA, "INCOME_STATEMENT")
bs_recs = section_to_years(FIN_DATA, "BALANCE_SHEET")
cf_recs = section_to_years(FIN_DATA, "CASH_FLOW")

# Lịch sử KQKD (tỷ VND) — from Vietcap API
years_hist = [2021, 2022, 2023, 2024, 2025]
revenue_hist   = [get_yr(is_recs, y, "isa3") for y in years_hist]
cogs_hist      = [abs(get_yr(is_recs, y, "isa4")) for y in years_hist]
gp_hist        = [get_yr(is_recs, y, "isa5") for y in years_hist]
gp_margin_hist = [round(gp_hist[i]/revenue_hist[i]*100, 2) for i in range(5)]
ni_hist        = [get_yr(is_recs, y, "isa22") for y in years_hist]
ebit_hist      = [get_yr(is_recs, y, "isa11") for y in years_hist]
pbt_hist       = [get_yr(is_recs, y, "isa16") for y in years_hist]
fin_inc_hist   = [get_yr(is_recs, y, "isa6") for y in years_hist]
fin_cost_hist  = [abs(get_yr(is_recs, y, "isa7")) for y in years_hist]
interest_hist  = [abs(get_yr(is_recs, y, "isa8")) for y in years_hist]
sell_exp_hist  = [abs(get_yr(is_recs, y, "isa9")) for y in years_hist]
admin_exp_hist = [abs(get_yr(is_recs, y, "isa10")) for y in years_hist]
eps_hist       = [get_yr_raw(is_recs, y, "isa23") for y in years_hist]

# Balance Sheet (tỷ VND) — from Vietcap API
cash_hist          = [get_yr(bs_recs, y, "bsa2") for y in years_hist]
total_assets_hist  = [get_yr(bs_recs, y, "bsa53") for y in years_hist]
equity_hist        = [get_yr(bs_recs, y, "bsa78") for y in years_hist]
inventory_hist     = [get_yr(bs_recs, y, "bsa15") for y in years_hist]
receivables_hist   = [get_yr(bs_recs, y, "bsa8") for y in years_hist]
short_debt_hist    = [get_yr(bs_recs, y, "bsa56") for y in years_hist]
long_debt_hist     = [get_yr(bs_recs, y, "bsa71") for y in years_hist]
total_debt_hist    = [short_debt_hist[i] + long_debt_hist[i] for i in range(5)]
st_invest_hist     = [get_yr(bs_recs, y, "bsa5") for y in years_hist]   # Đầu tư ngắn hạn
# "Tiền mặt" dùng cho Net Debt (EV bridge) = Tiền & tương đương tiền + Đầu tư ngắn hạn (2026-07, user
# phát hiện Assumptions!row12 dùng số chết, đặt câu hỏi nên lấy bsa2 hay bsa5 — CẢ HAI, vì Đầu tư ngắn
# hạn của HPG chủ yếu là tiền gửi có kỳ hạn/giấy tờ có giá thanh khoản cao, tương đương tiền mặt cho
# mục đích Net Debt = Nợ vay - Tiền mặt. cash_hist (bsa2 riêng) vẫn giữ nguyên cho sheet 05_Balance_Sheet
# dòng "Tiền & tương đương" (đúng khái niệm kế toán riêng biệt với Đầu tư ngắn hạn).
cash_for_valuation_hist = [cash_hist[i] + st_invest_hist[i] for i in range(5)]
fixed_assets_hist  = [get_yr(bs_recs, y, "bsa30") for y in years_hist]
cip_hist           = [get_yr(bs_recs, y, "bsa188") for y in years_hist]
payables_hist      = [get_yr(bs_recs, y, "bsa57") for y in years_hist]

# Cash Flow (tỷ VND) — from Vietcap API (signs preserved from source)
cfo_hist      = [get_yr(cf_recs, y, "cfa18") for y in years_hist]
cfi_hist      = [get_yr(cf_recs, y, "cfa26") for y in years_hist]
cff_hist      = [get_yr(cf_recs, y, "cfa34") for y in years_hist]
da_hist       = [get_yr(cf_recs, y, "cfa2") for y in years_hist]
capex_hist    = [abs(get_yr(cf_recs, y, "cfa19")) for y in years_hist]  # displayed as positive amount
begin_cash_hist = [get_yr(cf_recs, y, "cfa36") for y in years_hist]

# ── GIÁ HÀNG HÓA (HRC / Quặng sắt / Than cốc) — nguồn duy nhất, sheet 17_Gia_Hang_Hoa ──────
# 1. Lịch sử 18 quý (2021Q4-2026Q1): tổng hợp từ World Bank Pink Sheet (quặng sắt — độ tin cậy
#    cao, đối chiếu khớp số liệu công bố), báo cáo FPTS/VCBS/Argus (than cốc — có điểm neo thật,
#    phần giữa nội suy), và Mysteel/SteelBenchmarker/HPG export price (HRC — chỉ đúng xu
#    hướng/độ lớn, một số quý là ước lượng — xem README_COMMODITY.md nếu cần verify lại).
# 2. "Giá hiện tại" (current spot) — TỰ ĐỘNG fetch mỗi lần chạy script từ investing.com (không
#    cần AI/thao tác thủ công): HRC FOB China, Quặng sắt 62% CFR, Than cốc luyện kim (DCE, CNY
#    → quy đổi USD). Nếu fetch lỗi (mất mạng, đổi cấu trúc trang...), tự động dùng giá quý gần
#    nhất trong bảng lịch sử làm phương án dự phòng — script KHÔNG BAO GIỜ dừng vì lỗi fetch.
Q18_LABELS = ["2021Q4","2022Q1","2022Q2","2022Q3","2022Q4","2023Q1","2023Q2","2023Q3","2023Q4",
              "2024Q1","2024Q2","2024Q3","2024Q4","2025Q1","2025Q2","2025Q3","2025Q4","2026Q1"]
Q18_HRC  = [640, 870, 720, 580, 570, 640, 615, 605, 620, 605, 600, 510, 500, 480, 470, 470, 460, 475]
Q18_IRON = [112, 143, 138, 106,  99, 126, 112, 115, 129, 123, 113, 100, 101, 102,  96, 100, 104, 104]
Q18_COAL = [375, 500, 480, 330, 290, 320, 245, 250, 290, 275, 250, 220, 203, 182, 184, 190, 212, 220]
CNY_USD_RATE = 7.2  # tỷ giá quy đổi than cốc (Đại Liên, niêm yết CNY) — cập nhật định kỳ nếu lệch nhiều
UA_STR = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"

def fetch_via_curl(url, timeout=10, label=None):
    """Tải HTML qua curl subprocess (KHÔNG dùng thư viện requests) — investing.com chặn vân tay
    TLS của Python requests/urllib3 (Cloudflare 403) nhưng vẫn cho phép curl. curl có sẵn mặc định
    trên Windows 10+/macOS/Linux nên không cần cài thêm gì để chạy 100% bằng Python thuần.
    Tham số `label` (tuỳ chọn) chỉ dùng để in log chẩn đoán khi fetch rỗng/lỗi — giúp phân biệt
    "mạng bị chặn/timeout" (nghi vấn khi chạy trên GitHub Actions — IP datacenter dễ bị site .vn
    chặn hơn IP dân dụng VN) với "kết nối OK nhưng trang không có tin phù hợp"."""
    tag = f" [{label}]" if label else ""
    try:
        r = subprocess.run(
            ["curl", "-sL", "-A", UA_STR, "--max-time", str(timeout), "-w", "\n__HTTP_STATUS__:%{http_code}", url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout + 5,
        )
        out = r.stdout
        status = None
        marker = "\n__HTTP_STATUS__:"
        if marker in out:
            out, status_str = out.rsplit(marker, 1)
            status = status_str.strip()
        if r.returncode != 0 or not out.strip() or status not in (None, "200"):
            print(f"  [DIAG]{tag} fetch_via_curl weak/empty result: curl_exit={r.returncode} http_status={status} "
                  f"len={len(out)} url={url[:90]}" + (f" stderr={r.stderr[:150].strip()}" if r.stderr else ""))
        return out if r.returncode == 0 else ""
    except Exception as e:
        print(f"  [DIAG]{tag} fetch_via_curl exception: {url[:90]} -> {e}")
        return ""

# ── Quặng sắt: MEDIAN THẬT theo quý từ World Bank Pink Sheet (2026-07) ────────────────────────
# User yêu cầu xác thực lại: 18 quý lịch sử trước đây là SỐ NGHIÊN CỨU THỦ CÔNG 1 điểm/quý (không
# phải median nhiều điểm giá trong quý) vì không có nguồn giá theo ngày miễn phí. Quặng sắt là
# NGOẠI LỆ — World Bank Commodity Markets "Pink Sheet" có giá THEO THÁNG, miễn phí, công khai, đủ
# dài (từ 1960) → có thể tính MEDIAN THẬT của 3 tháng/quý. Than cốc & HRC KHÔNG có nguồn tháng/ngày
# miễn phí tương đương nên vẫn giữ số nghiên cứu đại diện — không gọi nhầm là "median".
def fetch_worldbank_iron_ore_monthly():
    """Trả về dict {(year, month): giá quặng sắt CFR spot, USD/tấn} từ World Bank Pink Sheet, hoặc
    {} nếu fetch/parse lỗi. KHÔNG BAO GIỜ raise — script luôn fallback về Q18_IRON nghiên cứu thủ
    công nếu World Bank không truy cập được (đổi cấu trúc trang, mất mạng, đổi tên sheet/cột...)."""
    try:
        page_html = fetch_via_curl("https://www.worldbank.org/en/research/commodity-markets", timeout=20)
        m = re.search(r'href="(https://thedocs\.worldbank\.org/[^"]*CMO-Historical-Data-Monthly\.xlsx)"', page_html)
        if not m:
            print("  [WARN] Could not find CMO-Historical-Data-Monthly.xlsx link on World Bank page")
            return {}
        xlsx_url = m.group(1)
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            r = subprocess.run(["curl", "-sL", "-A", UA_STR, "--max-time", "30", "-o", tmp_path, xlsx_url],
                                capture_output=True, timeout=35)
            if r.returncode != 0 or not os.path.exists(tmp_path) or os.path.getsize(tmp_path) < 10000:
                print("  [WARN] Failed to download World Bank Pink Sheet (network error or empty file)")
                return {}
            wbp = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
            if "Monthly Prices" not in wbp.sheetnames:
                print("  [WARN] World Bank file structure changed - 'Monthly Prices' sheet not found")
                return {}
            wsp = wbp["Monthly Prices"]
            iron_col = None
            for c, cell in enumerate(next(wsp.iter_rows(min_row=5, max_row=5)), 1):
                if cell.value and "iron ore" in str(cell.value).lower():
                    iron_col = c
                    break
            if not iron_col:
                print("  [WARN] World Bank column renamed - 'Iron ore, cfr spot' not found")
                return {}
            monthly = {}
            for row in wsp.iter_rows(min_row=7, values_only=False):
                lbl = row[0].value
                val = row[iron_col - 1].value
                if isinstance(lbl, str) and re.match(r"^\d{4}M\d{2}$", lbl) and isinstance(val, (int, float)):
                    yr_s, mo_s = lbl.split("M")
                    monthly[(int(yr_s), int(mo_s))] = float(val)
            wbp.close()
            return monthly
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
    except Exception as e:
        print(f"  [WARN] World Bank Pink Sheet fetch error: {e}")
        return {}

def _quarter_months(label):
    yr = int(label[:4]); q = int(label[5])
    return [(yr, 3*(q-1)+1), (yr, 3*(q-1)+2), (yr, 3*(q-1)+3)]

print("[Commodity] Fetching World Bank Pink Sheet (monthly iron ore, for real quarterly median)...")
IRON_MONTHLY = fetch_worldbank_iron_ore_monthly()
Q18_IRON_SRC = []  # "WB" = real 3-month World Bank median; "NC" = manual research fallback
if IRON_MONTHLY:
    print(f"  -> Fetched {len(IRON_MONTHLY)} months of World Bank iron ore data (latest: {max(IRON_MONTHLY)})")
    for i, lbl in enumerate(Q18_LABELS):
        vals = [IRON_MONTHLY[ym] for ym in _quarter_months(lbl) if ym in IRON_MONTHLY]
        if len(vals) == 3:
            Q18_IRON[i] = round(stats.median(vals), 1)
            Q18_IRON_SRC.append("WB")
        else:
            Q18_IRON_SRC.append("NC")
else:
    print("  [WARN] Could not fetch World Bank data - using manual research iron ore prices (fallback)")
    Q18_IRON_SRC = ["NC"] * len(Q18_LABELS)

# CÔNG THỨC SPREAD CHUẨN — LAG 1 QUÝ (2026-07, theo yêu cầu user): vì số ngày tồn kho bình quân
# của HPG dao động quanh ~90 ngày (~1 quý — xem DIO_A/sheet 14_Steel_Analysis), giá vốn hàng bán
# ghi nhận trong quý T phần lớn phản ánh giá nguyên liệu (quặng sắt/than cốc) MUA VÀO ở quý T-1,
# trong khi giá bán HRC vẫn là giá của quý T. Do đó:
#   Spread(quý T) = Giá HRC bình quân quý T - 1.6×Giá quặng bình quân quý T-1
#                   - 0.6×Giá than cốc bình quân quý T-1 - OTHER_COST_USD
# áp dụng thống nhất cho MỌI nơi tính Spread (sheet 17_Gia_Hang_Hoa, 14_Steel_Analysis, Profit
# Bridge 03_Revenue_Model, các biểu đồ spread, JSON dashboard) — không còn công thức "cùng quý"
# (HRC - Quặng - Than CÙNG kỳ) như trước.
OTHER_COST_USD = 100  # USD/tấn — chi phí SX khác cố định (nhân công/nhiên liệu/điện/khấu hao/CPQL)

def _lag_spread(price_cur, iron_prev, coal_prev):
    return round(price_cur - 1.6*iron_prev - 0.6*coal_prev - OTHER_COST_USD, 1)

# ── Vị trí dòng CỐ ĐỊNH của sheet 17_Gia_Hang_Hoa (2026-07, theo yêu cầu user) ────────────────────
# User phát hiện bug: 02_Assumptions!row45 "Spread hàng năm" dùng công thức ĐỘC LẬP (tự trừ lại
# Quặng/Than) thay vì LINK sang sheet 17_Gia_Hang_Hoa (nơi ĐÃ có sẵn Spread HRC/Spread All theo năm,
# tính đúng công thức lag-1-quý/AD20/bình quân gia quyền) — 2 nguồn số liệu độc lập dễ lệch nhau âm
# thầm giống hệt bug SL_HRC_A/SL_XD_A. Để 02_Assumptions LINK đúng sang sheet 17 mà KHÔNG cần dựng cả
# sheet 17 trước (thứ tự code: 02_Assumptions được build TRƯỚC 17_Gia_Hang_Hoa), tính trước vị trí
# dòng annual Spread bằng ĐÚNG 1 công thức duy nhất — sheet 17 khi build SAU sẽ dùng lại chính dict
# này (KHÔNG tự tính lại r17 cục bộ) để 2 nơi không bao giờ lệch nhau.
def _r17_annual_row_layout():
    r17_hdr = 4
    r17_q0 = r17_hdr + 1
    r17_q_last = r17_q0 + len(Q18_LABELS) - 1
    r17 = r17_q_last + 2          # "GIÁ HIỆN TẠI" header
    r17_now = r17 + 1
    r17_avg = r17_now + 1
    r17_spr_now = r17_avg + 1
    r17 = r17_spr_now + 2         # "GIÁ NĂM" header
    r17_ann_hdr = r17 + 1
    r17_ann_hrc = r17_ann_hdr + 1
    return {
        "R17_HDR": r17_hdr, "R17_Q0": r17_q0, "R17_Q_LAST": r17_q_last,
        "R17_NOW": r17_now, "R17_AVG": r17_avg, "R17_SPR_NOW": r17_spr_now,
        "R17_ANN_HDR": r17_ann_hdr, "R17_ANN_HRC": r17_ann_hrc,
        "R17_ANN_IRON": r17_ann_hrc + 1, "R17_ANN_COAL": r17_ann_hrc + 2, "R17_ANN_XD": r17_ann_hrc + 3,
        "R17_ANN_SPREAD": r17_ann_hrc + 4, "R17_ANN_SPREAD_REBAR": r17_ann_hrc + 5,
        "R17_ANN_SPREAD_ALL": r17_ann_hrc + 6,
    }
R17_LAYOUT = _r17_annual_row_layout()

# ── Thuế chống bán phá giá (CBPG) HRC Trung Quốc — vụ việc AD20 (2026-07, theo yêu cầu user) ──
# Bộ Công Thương áp thuế CBPG chính thức (khổ hẹp) với HRC Trung Quốc/Ấn Độ hiệu lực từ 06/07/2025
# (QĐ 1959/QĐ-BCT), mở rộng biện pháp chống lẩn tránh (khổ rộng) từ 17/04/2026. User chọn mốc
# 06/07/2025 (thuế chính thức khổ hẹp) làm điểm bắt đầu phản ánh vào Spread HRC — vì thuế NK đẩy
# giá HRC nội địa/nhập khẩu thực trả cao hơn giá niêm yết CFR trên investing.com, Spread HRC tính
# thẳng từ giá CFR (chưa thuế) sẽ bị đánh giá THẤP hơn thực tế kể từ ngày có thuế.
#   Spread HRC(quý T, từ 2025Q3 trở đi) = Giá HRC quý T × 1.15 - 1.6×Quặng quý T-1 - 0.6×Than quý T-1 - OTHER_COST_USD
AD_HRC_EFFECTIVE_Q = "2025Q3"   # 06/07/2025 rơi vào quý này (thuế có hiệu lực >85/92 ngày của quý)
AD_HRC_MULTIPLIER = 1.15
_ad_idx = Q18_LABELS.index(AD_HRC_EFFECTIVE_Q)

def _hrc_ad_adjusted(i):
    return round(Q18_HRC[i] * AD_HRC_MULTIPLIER, 1) if i >= _ad_idx else Q18_HRC[i]

# Q18_SPREAD_HRC: Spread lag-1-quý cho HRC (đã điều chỉnh thuế CBPG từ 2025Q3). Quý đầu tiên
# (2021Q4) không có dữ liệu quý trước (2021Q3) trong bảng lịch sử — dùng tạm giá CÙNG quý làm
# phương án dự phòng duy nhất (chỉ 1/18 điểm dữ liệu bị ảnh hưởng, không đại diện cho công thức chuẩn).
Q18_SPREAD_HRC = [_lag_spread(_hrc_ad_adjusted(0), Q18_IRON[0], Q18_COAL[0])] + [
    _lag_spread(_hrc_ad_adjusted(i), Q18_IRON[i-1], Q18_COAL[i-1]) for i in range(1, len(Q18_LABELS))
]
Q18_SPREAD = Q18_SPREAD_HRC  # alias tương thích ngược — "Spread thép" mặc định = Spread HRC (đã có thuế CBPG)

# ── Thép xây dựng (rebar): giá THẬT theo quý — nguồn investing.com Steel Rebar futures (SRRc1) +
# neo nội địa SteelOnline/VSA (2026-07, theo yêu cầu user) ───────────────────────────────────────
# User cung cấp 2 nguồn: (1) investing.com/commodities/steel-rebar — hợp đồng tương lai SRRc1
# (Shanghai rebar continuous), niêm yết THẲNG bằng USD/tấn (không cần quy đổi CNY) — có dữ liệu
# tháng thật từ 2021 nhưng khối lượng giao dịch rất mỏng và NGỪNG cập nhật sau 10/2025 (giá đứng
# yên 545 USD/tấn nhiều tháng — hợp đồng hết thanh khoản); (2) steelonline.vn — giá thép XD Hòa
# Phát D10 (grade CB240) NỘI ĐỊA THỰC nhưng CHỈ có giá HIỆN TẠI (không có kho lưu trữ lịch sử).
# Đối chiếu 2 nguồn tại cùng thời điểm (~10/2025-nay): SRRc1 ~545 USD/tấn vs SteelOnline 15.120
# đồng/kg quy đổi ~581 USD/tấn — lệch ~6%, chấp nhận được để ĐO XU HƯỚNG (không phải mốc tuyệt đối).
# => 16/18 quý (2021Q4-2025Q3) dùng MEDIAN THẬT 3 tháng SRRc1 ("INV"); 2026Q1 dùng điểm neo THẬT
# nội địa (VSA công bố giá thép XD Q1/2026: 15.100-15.700đ/kg, bài "Thị trường thép xây dựng Quý
# I/2026" 09/04/2026) quy đổi USD ("VN"); 2025Q4 (SRRc1 chỉ còn 1/3 tháng thật) nội suy tuyến tính
# giữa 2025Q3 thật và 2026Q1 thật ("NC").
REBAR_INSTRUMENT_ID = "996702"  # investing.com SRRc1 - Shanghai rebar futures continuous (USD/tấn)

def fetch_investing_rebar_monthly():
    """Trả về dict {(year, month): giá SRRc1 rebar futures, USD/tấn} từ API nội bộ investing.com
    (KHÔNG BAO GIỜ raise — trả về {} nếu lỗi mạng/đổi API/đổi instrument id)."""
    try:
        url = (f"https://api.investing.com/api/financialdata/historical/{REBAR_INSTRUMENT_ID}"
               f"?start-date=2021-01-01&end-date={date.today().isoformat()}"
               "&time-frame=Monthly&add-missing-rows=false")
        r = subprocess.run(
            ["curl", "-sL", "-A", UA_STR, "-H", "Domain-Id: vn", "-H", "Accept: application/json",
             "--max-time", "20", url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=25,
        )
        if r.returncode != 0 or not r.stdout:
            return {}
        payload = json.loads(r.stdout)
        monthly = {}
        for row in payload.get("data", []):
            ts, price = row.get("rowDateTimestamp"), row.get("last_closeRaw")
            if not ts or price is None:
                continue
            monthly[(int(ts[:4]), int(ts[5:7]))] = float(price)
        return monthly
    except Exception as e:
        print(f"  [WARN] investing.com rebar (SRRc1) fetch error: {e}")
        return {}

def fetch_usd_vnd_rate(fallback=26200.0):
    """Tỷ giá USD/VND hiện tại (fetch investing.com, fallback nếu lỗi) — dùng quy đổi các điểm neo
    giá thép XD nội địa (VND/kg, SteelOnline/VSA) sang USD/tấn cho thống nhất công thức Spread."""
    try:
        html = fetch_via_curl("https://vn.investing.com/currencies/usd-vnd", timeout=10)
        m = re.search(r'data-test="instrument-price-last"[^>]*>([\d,\.]+)<', html)
        return float(m.group(1).replace(",", "")) if m else fallback
    except Exception:
        return fallback

def fetch_steelonline_rebar_price():
    """Giá thép XD Hòa Phát D10 (grade CB240) hiện tại (VND/kg) từ steelonline.vn, None nếu lỗi.
    Bảng ĐẦU TIÊN trên trang xác nhận là bảng Hòa Phát (đoạn mô tả ngay sau bảng ghi rõ "Trên đây
    là bảng giá thép xây dựng Hòa Phát hôm nay")."""
    try:
        html = fetch_via_curl("https://www.steelonline.vn/bang-gia-thep-xay-dung-hom-nay", timeout=12)
        idx = html.find("<table")
        if idx == -1:
            return None
        m = re.search(r"(\d{2}\.\d{3})", html[idx:idx+6000])
        return float(m.group(1).replace(".", "")) if m else None
    except Exception:
        return None

_usd_vnd_rate = fetch_usd_vnd_rate()

def _vnd_kg_to_usd_t(vnd_kg):
    return round(vnd_kg * 1000 / _usd_vnd_rate, 1)

print("[Commodity] Fetching Steel Rebar (SRRc1) monthly history from investing.com...")
REBAR_MONTHLY = fetch_investing_rebar_monthly()
Q18_XD = [None] * len(Q18_LABELS)
Q18_XD_SRC = [None] * len(Q18_LABELS)
if REBAR_MONTHLY:
    print(f"  -> Fetched {len(REBAR_MONTHLY)} months of SRRc1 rebar futures (latest: {max(REBAR_MONTHLY)})")
    for i, lbl in enumerate(Q18_LABELS):
        vals = [REBAR_MONTHLY[ym] for ym in _quarter_months(lbl) if ym in REBAR_MONTHLY]
        if len(vals) == 3:
            Q18_XD[i] = round(stats.median(vals), 1)
            Q18_XD_SRC[i] = "INV"
else:
    print("  [WARN] Could not fetch investing.com rebar futures - all quarters will use fallback")
_i_2026q1 = Q18_LABELS.index("2026Q1")
if Q18_XD[_i_2026q1] is None:
    Q18_XD[_i_2026q1] = _vnd_kg_to_usd_t(15400)  # median(15.100, 15.700) đồng/kg - VSA Q1/2026
    Q18_XD_SRC[_i_2026q1] = "VN"
_i_2025q4 = Q18_LABELS.index("2025Q4")
if Q18_XD[_i_2025q4] is None:
    _prev_i, _next_i = _i_2025q4 - 1, _i_2025q4 + 1
    if Q18_XD[_prev_i] is not None and Q18_XD[_next_i] is not None:
        Q18_XD[_i_2025q4] = round((Q18_XD[_prev_i] + Q18_XD[_next_i]) / 2, 1)
        Q18_XD_SRC[_i_2025q4] = "NC"
for _i in range(len(Q18_XD)):  # phòng hờ fetch lỗi hoàn toàn - dùng giá quý liền trước làm dự phòng cuối
    if Q18_XD[_i] is None:
        Q18_XD[_i] = Q18_XD[_i-1] if _i > 0 else 600.0
        Q18_XD_SRC[_i] = "NC"

Q18_SPREAD_REBAR = [_lag_spread(Q18_XD[0], Q18_IRON[0], Q18_COAL[0])] + [
    _lag_spread(Q18_XD[i], Q18_IRON[i-1], Q18_COAL[i-1]) for i in range(1, len(Q18_LABELS))
]

def fetch_investing_price(url, fallback):
    """Fetch giá hiện tại (USD/tấn) từ 1 trang investing.com. Trả về fallback nếu fetch lỗi."""
    html = fetch_via_curl(url)
    m = re.search(r'data-test="instrument-price-last"[^>]*>([\d,\.]+)<', html)
    if not m:
        print(f"  [WARN] Fetch failed for {url} - fallback to last known quarterly price")
        return fallback
    price = float(m.group(1).replace(",", ""))
    c = re.search(r'currency-in-label"[^>]*>.*?<span[^>]*>([A-Z]{3})</span>', html)
    if c and c.group(1) == "CNY":
        price = round(price / CNY_USD_RATE, 1)
    return price

print("[Commodity] Fetching current HRC/iron ore/coking coal prices from investing.com...")
hrc_now  = fetch_investing_price("https://www.investing.com/commodities/lme-steel-hrc-fob-china-futures", Q18_HRC[-1])
iron_now = fetch_investing_price("https://vn.investing.com/commodities/iron-ore-62-cfr-futures", Q18_IRON[-1])
coal_now = fetch_investing_price("https://vn.investing.com/commodities/metallurgical-coke-futures", Q18_COAL[-1])
print(f"  -> HRC={hrc_now} USD/t | Iron ore={iron_now} USD/t | Coking coal={coal_now} USD/t (converted from CNY if applicable)")

# Giá thép XD "hiện tại": ƯU TIÊN SteelOnline (nội địa THẬT, VND/kg quy đổi USD) vì phản ánh đúng
# giá bán trong nước hơn hợp đồng tương lai SRRc1 (đã hết thanh khoản từ 11/2025 — xem chú thích
# Q18_XD ở trên). Chỉ fallback về SRRc1 futures nếu SteelOnline fetch lỗi.
print("[Commodity] Fetching current Steel Rebar (XD) price from SteelOnline / investing.com...")
_xd_vnd_now = fetch_steelonline_rebar_price()
if _xd_vnd_now:
    xd_now = _vnd_kg_to_usd_t(_xd_vnd_now)
    XD_NOW_SRC = f"SteelOnline noi dia ({_xd_vnd_now:,.0f} dong/kg, quy doi ty gia {_usd_vnd_rate:,.0f})"
else:
    xd_now = fetch_investing_price("https://vn.investing.com/commodities/steel-rebar", Q18_XD[-1])
    XD_NOW_SRC = "investing.com SRRc1 futures (du phong - SteelOnline fetch loi)"
print(f"  -> Steel Rebar (XD)={xd_now} USD/t [{XD_NOW_SRC}]")

# Quặng sắt "hiện tại": ƯU TIÊN dùng tháng MỚI NHẤT của World Bank Pink Sheet (IRON_MONTHLY, đã fetch
# ở trên) thay vì giá futures investing.com — 2 nguồn KHÔNG cùng phương pháp luận (investing.com là
# hợp đồng tương lai SGX TSI 62% Fe, có thể lệch cơ sở/biến động ngắn hạn so với spot; World Bank là
# giá spot bình quân tháng, CÙNG nguồn với 18 quý lịch sử) — trộn 2 nguồn khác phương pháp vào 1 chuỗi
# "hiện tại vs lịch sử" dễ tạo lệch giả. Đánh đổi: WB có độ trễ ~1 tháng (không phải giá hôm nay),
# investing.com theo giờ nhưng khác cơ sở. Nếu WB không fetch được, fallback về investing.com như cũ.
if IRON_MONTHLY:
    _iron_wb_latest_ym = max(IRON_MONTHLY)
    _iron_wb_latest = IRON_MONTHLY[_iron_wb_latest_ym]
    print(f"  -> Overriding Iron ore 'current' with World Bank latest month {_iron_wb_latest_ym}: "
          f"{_iron_wb_latest} USD/t (was investing.com futures: {iron_now} USD/t) - same methodology as historical series")
    iron_now = _iron_wb_latest

def _year_median(vals_by_year, yr):
    vs = vals_by_year.get(yr, [])
    return round(stats.median(vs), 1) if vs else None

_yr_of_q = [int(q[:4]) for q in Q18_LABELS]
def _group_by_year(vals):
    d = {}
    for yr, v in zip(_yr_of_q, vals):
        d.setdefault(yr, []).append(v)
    return d

# Giá năm lịch sử (2021-2025) = MEDIAN các quý thuộc năm đó trong bảng 18 quý (2021 chỉ có Q4,
# 2026 chỉ có Q1 — dùng luôn giá quý đó vì median của 1 giá trị = chính nó).
_hrc_by_yr, _iron_by_yr, _coal_by_yr = _group_by_year(Q18_HRC), _group_by_year(Q18_IRON), _group_by_year(Q18_COAL)

# ── Sản lượng thép HPG theo quý — nguồn duy nhất (2026-07) ────────────────────────────────────
# 13 quý thực tế (2023Q1-2026Q1) tổng hợp thủ công từ BCTC/tin công bố của HPG (nghìn tấn) — nguồn
# duy nhất, dùng chung cho sheet 15_Quarterly_Data và JSON dashboard (tránh 2 mảng trùng lặp lệch
# số như trước — hrc_data/xd_data trong sheet 15 và hrc_sales/xd_sales trong JSON export).
SALES_Q_LABELS   = ["2023Q1","2023Q2","2023Q3","2023Q4","2024Q1","2024Q2","2024Q3","2024Q4",
                    "2025Q1","2025Q2","2025Q3","2025Q4","2026Q1"]
# 2024Q2-Q4 (2026-07, user phát hiện bug): số cũ [464, 312, 399] SAI — kiểm chứng lại qua báo chí
# (thitruongtaichinhtiente.vn, nhipsongkinhdoanh.vn) xác nhận: Q3/2024 HRC = 738 nghìn tấn (số CHÍNH
# XÁC, khớp bài báo); 9 tháng đầu 2024 HRC = 2.270 triệu tấn (khớp CHÍNH XÁC 805+X+738=2270 => Q2=727);
# Q2/2024 giảm ~10% so với Q1/2024 (805*0.9≈725, khớp Q2=727 suy ra ở trên); Q4/2024 CHƯA tìm được số
# chính xác — ước tính bằng phần dư: HRC sản xuất cả năm 2024 "hơn 3 triệu tấn" (+5% so với 2023) =>
# Q4 = ~3050 - 2270 = 780 (ước tính từ phần dư, không phải số công bố trực tiếp — cần verify lại nếu
# có báo cáo Q4/2024 hoặc BCTC cụ thể hơn).
HRC_SALES_HIST_KT = [482, 770, 780, 768, 805, 727, 738, 780, 1000, 1180, 1220, 1600, 1400]
XD_SALES_HIST_KT  = [870, 970, 970, 970, 956, 1140, 1200, 1100, 1200, 1300, 1000, 1300, 1430]

# Spread All = bình quân gia quyền Spread HRC/Spread Rebar theo sản lượng thực tế TỪNG QUÝ (2026-07,
# theo yêu cầu user) — CHỈ tính được cho 13/18 quý trong bảng (2023Q1-2026Q1, đúng bằng SALES_Q_LABELS)
# vì HRC_SALES_HIST_KT/XD_SALES_HIST_KT không có dữ liệu 5 quý đầu (2021Q4-2022Q4, trước khi có số
# liệu tách sản lượng quý công bố) — các quý đó Q18_SPREAD_ALL để None (không suy diễn thiếu căn cứ).
#   Spread All(quý T) = (SL_XD(T)×Spread_Rebar(T) + SL_HRC(T)×Spread_HRC(T)) / (SL_XD(T)+SL_HRC(T))
Q18_SPREAD_ALL = [None] * len(Q18_LABELS)
_vol_offset = Q18_LABELS.index(SALES_Q_LABELS[0])
for _k, _lbl in enumerate(SALES_Q_LABELS):
    _i = _vol_offset + _k
    _sl_xd, _sl_hrc = XD_SALES_HIST_KT[_k], HRC_SALES_HIST_KT[_k]
    _tot_kt = _sl_xd + _sl_hrc
    if _tot_kt:
        Q18_SPREAD_ALL[_i] = round((_sl_xd*Q18_SPREAD_REBAR[_i] + _sl_hrc*Q18_SPREAD_HRC[_i]) / _tot_kt, 1)

# Quý ĐANG CHẠY (quý kế tiếp sau quý cuối cùng đã có số liệu thực) — ước tính SỚM sản lượng quý
# này (trước khi có BCTC/số liệu quý chính thức) bằng cách tự động dò các bài công bố sản lượng
# tháng/quý mới nhất trên chính trang tin của HPG — thuần Python (curl + regex tiêu đề bài viết,
# KHÔNG dùng AI lúc chạy), theo đúng yêu cầu user. Nếu chưa có tin gì cho quý này → giữ nguyên
# None, các chỗ dùng sẽ tự fallback về giả định cũ, KHÔNG BAO GIỜ chặn script.
HPG_NEWS_SEED_URL = "https://www.hoaphat.com.vn/tin-tuc/hoa-phat-da-vuot-ke-hoach-loi-nhuan-nam-2018.html"
_HPG_TITLE_RE = re.compile(
    r"(?:Sản lượng bán hàng thép Hòa Phát đạt|Hòa Phát (?:bán|cung cấp)(?: ra thị trường)?)\s*"
    r"([\d.,]+)\s*(triệu\s*)?tấn.*?trong\s*"
    r"(?:tháng\s*(\d{1,2})(?!\s*[/\-]\s*\d{4})|quý\s*([IVX\d]{1,3})(?:\s*[/\s]\s*(\d{4}))?|(\d{1,2})\s*tháng\s*đầu\s*năm)",
    re.IGNORECASE,
)
_ROMAN_Q = {"I": 1, "II": 2, "III": 3, "IV": 4}

def _vn_number_to_kilotons(num_str, is_million):
    """Chuyển số dạng VN ('2,6' triệu hoặc '635.000' tấn — dấu phẩy/chấm ngược với chuẩn Mỹ) sang
    đơn vị NGHÌN TẤN, khớp đơn vị HRC_SALES_HIST_KT/XD_SALES_HIST_KT ở trên."""
    if is_million:
        return round(float(num_str.replace(".", "").replace(",", ".")) * 1000, 1)
    return round(float(num_str.replace(".", "").replace(",", "")) / 1000, 1)

def _parse_hpg_production_title(title, pub_date):
    """Parse tiêu đề bài công bố sản lượng HPG (mẫu chuẩn: "Sản lượng bán hàng thép Hòa Phát đạt
    X (triệu/nghìn) tấn trong {tháng N | quý N/YYYY | N tháng đầu năm}"). Quý có thể viết số Ả Rập
    (quý 2) hoặc La Mã (quý II), có hoặc không kèm năm (nếu thiếu năm thì suy từ ngày đăng bài — chỉ
    đúng nếu bài đăng ngay sau khi kết thúc quý, đúng thông lệ PR của HPG). Trả về dict hoặc None nếu
    tiêu đề không khớp mẫu (bài không phải công bố sản lượng, hoặc HPG đổi cách viết)."""
    m = _HPG_TITLE_RE.search(title or "")
    if not m:
        return None
    num_str, is_million, month, q_num_raw, q_year, cum_months = m.groups()
    vol_kt = _vn_number_to_kilotons(num_str, bool(is_million))
    pub_year = int(pub_date[6:10]) if pub_date and re.match(r"\d{2}/\d{2}/\d{4}", pub_date) else None
    pub_month = int(pub_date[3:5]) if pub_date and re.match(r"\d{2}/\d{2}/\d{4}", pub_date) else None
    if q_num_raw:
        q_num = _ROMAN_Q.get(q_num_raw.upper(), None)
        if q_num is None and q_num_raw.isdigit():
            q_num = int(q_num_raw)
        if q_num is None or not (1 <= q_num <= 4):
            return None
        yr = int(q_year) if q_year else pub_year
        if yr is None:
            return None
        return {"type": "quarter", "year": yr, "quarter": q_num, "volume_kt": vol_kt}
    if month:
        mo = int(month)
        # Bài đăng đầu năm sau thường nói về tháng 12 năm trước
        yr = (pub_year - 1) if (pub_year and mo == 12 and pub_month and pub_month <= 2) else pub_year
        return {"type": "month", "year": yr, "month": mo, "volume_kt": vol_kt}
    if cum_months:
        return {"type": "cumulative", "year": pub_year, "months": int(cum_months), "volume_kt": vol_kt}
    return None

def fetch_hpg_production_updates(max_fetch=40):
    """Dò các bài công bố sản lượng thép mới nhất của HPG qua (1) khung 'Tin liên quan' 5 tin mới
    nhất hiển thị trên MỌI trang bài viết (không cần trang danh sách bị che JS), và (2) sitemap tin
    tức chính thức (curl + regex tiêu đề, không cần JS) lọc theo từ khóa sản lượng — giới hạn số
    trang sitemap quét để không phải fetch toàn bộ lịch sử mỗi lần chạy. Trả về list dict đã parse
    được (rỗng nếu lỗi/không tìm thấy) — KHÔNG BAO GIỜ crash script."""
    candidates = {}
    try:
        seed_html = fetch_via_curl(HPG_NEWS_SEED_URL, timeout=15, label="hoaphat-seed")
        for m in re.finditer(
            r'<a href="(https://www\.hoaphat\.com\.vn/tin-tuc/[^"]+\.html)">\s*<div class="image">.*?'
            r'<p class="clear time">.*?(\d{2}/\d{2}/\d{4}).*?</p>\s*<h3>([^<]+)</h3>',
            seed_html, re.S):
            url, date_str, _title = m.groups()
            candidates[url] = date_str
    except Exception as e:
        print(f"  [WARN] HPG news sidebar fetch error: {e}")
    try:
        idx_html = fetch_via_curl("https://www.hoaphat.com.vn/sitemap.xml", timeout=15, label="hoaphat-sitemap-idx")
        sm_urls = re.findall(r'<loc>(https://www\.hoaphat\.com\.vn/sitemap-tintuc-page-\d+\.xml)</loc>', idx_html)
        for sm_url in sm_urls[:4]:
            sm_html = fetch_via_curl(sm_url, timeout=15, label="hoaphat-sitemap-page")
            for loc in re.findall(r'<loc>([^<]+)</loc>', sm_html):
                if re.search(r'san-luong|trieu-tan|nghin-tan|tan-thep', loc, re.I):
                    candidates.setdefault(loc, None)
    except Exception as e:
        print(f"  [WARN] HPG sitemap fetch error: {e}")

    records, fetched = [], 0
    for url, known_date in candidates.items():
        if fetched >= max_fetch:
            break
        try:
            html = fetch_via_curl(url, timeout=12)
            fetched += 1
            if not html:
                continue
            title_m = re.search(r"<title>([^<]+)</title>", html)
            date_m = re.search(r'<p class="clear time">.*?(\d{2}/\d{2}/\d{4})', html, re.S)
            title = title_m.group(1) if title_m else ""
            pub_date = date_m.group(1) if date_m else known_date
            parsed = _parse_hpg_production_title(title, pub_date)
            if not parsed:
                continue
            # KHÔNG tự parse riêng số HRC từ nội dung bài — đã thử và phát hiện quá thiếu tin cậy:
            # câu văn thường viết "...thép cuộn cán nóng (HRC), thép xây dựng, phôi thép đạt X tấn"
            # (X là TỔNG cả 3 sản phẩm, không phải riêng HRC) nên regex dễ bắt nhầm tổng thành HRC.
            # Chỉ lấy TỔNG sản lượng từ tiêu đề (đáng tin cậy, đã verify khớp 2/2 bài mẫu) — phần
            # tách HRC/XD dùng tỷ lệ lịch sử (xem SL_HRC_A/SL_XD_A bên dưới), không suy từ văn bản.
            parsed["url"] = url
            parsed["title"] = title
            parsed["date"] = pub_date
            records.append(parsed)
        except Exception:
            continue
    return records

_NQS_PRODUCT_RE = re.compile(
    r"sản lượng\s+(thép xây dựng|thép cuộn cán nóng\s*\(HRC\)|HRC)[^.]{0,60}?trong\s+tháng\s*"
    r"(\d{1,2})(?:\s*/\s*(\d{4}))?\s+đạt\s+(?:gần|hơn|trên)?\s*([\d.,]+)\s*(triệu\s*)?tấn",
    re.IGNORECASE,
)

def fetch_nguoiquansat_production_updates(days_back=70, max_fetch=25):
    """Dò tin sản lượng thép HPG trên nguoiquansat.vn — trang này thường đăng lại báo cáo cập nhật
    sản lượng THÁNG của Vietcap Research, TÁCH RIÊNG số liệu HRC và thép xây dựng theo tháng (chi
    tiết và đáng tin cậy hơn tin PR gộp chung của chính hoaphat.com.vn — xem comment ở
    fetch_hpg_production_updates). Dùng sitemap THEO NGÀY (sitemap-article-YYYY-MM-DD.xml, mỗi ngày
    ~70-100 bài, có sẵn tiêu đề trong thẻ <image:title> nên không cần fetch từng bài để lọc) quét
    ngược `days_back` ngày gần nhất. KHÔNG BAO GIỜ crash — trả về [] nếu lỗi bất kỳ bước nào."""
    candidates = {}
    try:
        today = date.today()
        for d in range(days_back):
            day_str = (today - timedelta(days=d)).isoformat()
            sm_html = fetch_via_curl(f"https://nguoiquansat.vn/sitemap-article-{day_str}.xml", timeout=12, label="nguoiquansat-sitemap")
            if not sm_html:
                continue
            for m in re.finditer(
                r'<loc>([^<]+)</loc>\s*<image:image>\s*<image:loc>[^<]*</image:loc>\s*'
                r'<image:title><!\[CDATA\[([^\]]+)\]\]>', sm_html):
                url, title = m.groups()
                if re.search(r"Hòa Phát|HPG", title) and re.search(r"sản lượng", title, re.I):
                    candidates[url] = title
    except Exception as e:
        print(f"  [WARN] nguoiquansat.vn sitemap fetch error: {e}")

    records, fetched = [], 0
    for url, title in candidates.items():
        if fetched >= max_fetch:
            break
        try:
            html = fetch_via_curl(url, timeout=12)
            fetched += 1
            if not html:
                continue
            date_m = re.search(r'(\d{2}/\d{2}/\d{4})\s*-\s*\d{2}:\d{2}', html)
            pub_date = date_m.group(1) if date_m else None
            pub_year = int(pub_date[6:10]) if pub_date else None
            hrc_kt = xd_kt = month_num = yr = None
            for product, month, yr_str, num_str, is_million in _NQS_PRODUCT_RE.findall(html):
                month_num = int(month)
                yr = int(yr_str) if yr_str else pub_year
                vol = _vn_number_to_kilotons(num_str, bool(is_million))
                if "xây dựng" in product:
                    xd_kt = vol
                else:
                    hrc_kt = vol
            if month_num and yr and (hrc_kt is not None or xd_kt is not None):
                records.append({
                    "type": "month", "year": yr, "month": month_num,
                    "hrc_kt": hrc_kt, "xd_kt": xd_kt,
                    "volume_kt": (hrc_kt or 0) + (xd_kt or 0),
                    "url": url, "title": title, "date": pub_date,
                })
        except Exception:
            continue
    return records

_DTCP_PRODUCT_RE = re.compile(
    r"HPG bán được\s*([\d.,]+)\s*(triệu\s*)?tấn\s*"
    r"(thép xây dựng|HRC|thép cuộn cán nóng|tôn mạ|ống thép)[^.]{0,40}?trong\s*tháng\s*"
    r"(\d{1,2})(?:\s*/\s*(\d{4}))?",
    re.IGNORECASE,
)

def fetch_dautucophieu_production_updates(max_fetch=15):
    """Dò tin cập nhật sản lượng thép HPG trên dautucophieu.net — trang này đăng lại báo cáo cập
    nhật hàng tháng của HSC Research (tương tự nguoiquansat.vn đăng lại Vietcap), khá đều đặn trong
    lịch sử (gần như tháng nào cũng có bài riêng cho HPG). Discovery qua `/tag/hpg/` (trang WordPress
    server-rendered, liệt kê bài mới nhất TRƯỚC — không cần sitemap/JS). KHÔNG BAO GIỜ crash — trả về
    [] nếu lỗi bất kỳ bước nào."""
    candidates = {}
    try:
        html = fetch_via_curl("https://dautucophieu.net/tag/hpg/", timeout=15, label="dautucophieu-tag")
        for m in re.finditer(
            r'<a href="(https://dautucophieu\.net/[^"]+/)" itemprop="mainEntityOfPage" title="([^"]+)">', html):
            url, title = m.groups()
            if "-hpg-" in url and re.search(r"sản lượng|tháng", title, re.I):
                candidates[url] = title
    except Exception as e:
        print(f"  [WARN] dautucophieu.net tag page fetch error: {e}")

    records, fetched = [], 0
    for url, title in candidates.items():
        if fetched >= max_fetch:
            break
        try:
            html = fetch_via_curl(url, timeout=12)
            fetched += 1
            if not html:
                continue
            date_m = re.search(r'entry-meta"[^>]*>.*?(\d{2}/\d{2}/\d{4})', html, re.S)
            pub_date = date_m.group(1) if date_m else None
            hrc_kt = xd_kt = month_num = yr = None
            for num_str, is_million, product, month, yr_str in _DTCP_PRODUCT_RE.findall(html):
                month_num = int(month)
                yr = int(yr_str) if yr_str else (int(pub_date[6:10]) if pub_date else None)
                vol = _vn_number_to_kilotons(num_str, bool(is_million))
                if "xây dựng" in product.lower():
                    xd_kt = vol
                elif "hrc" in product.lower() or "cán nóng" in product.lower():
                    hrc_kt = vol
            if month_num and yr and (hrc_kt is not None or xd_kt is not None):
                records.append({
                    "type": "month", "year": yr, "month": month_num,
                    "hrc_kt": hrc_kt, "xd_kt": xd_kt,
                    "volume_kt": (hrc_kt or 0) + (xd_kt or 0),
                    "url": url, "title": title, "date": pub_date,
                })
        except Exception:
            continue
    return records

_last_q_year, _last_q_num = int(SALES_Q_LABELS[-1][:4]), int(SALES_Q_LABELS[-1][5])
CUR_Q_YEAR, CUR_Q_NUM = (_last_q_year, _last_q_num + 1) if _last_q_num < 4 else (_last_q_year + 1, 1)
CUR_Q_LABEL = f"{CUR_Q_YEAR}Q{CUR_Q_NUM}"
_cur_q_months = [3*(CUR_Q_NUM-1)+1, 3*(CUR_Q_NUM-1)+2, 3*(CUR_Q_NUM-1)+3]

print(f"[Production] Fetching HPG steel sales volume news for {CUR_Q_LABEL}...")
HPG_PRODUCTION_NEWS = fetch_hpg_production_updates()
_nqs_records = fetch_nguoiquansat_production_updates()
print(f"  -> nguoiquansat.vn: found {len(_nqs_records)} monthly record(s) with HRC/XD breakdown")
_dtcp_records = fetch_dautucophieu_production_updates()
print(f"  -> dautucophieu.net: found {len(_dtcp_records)} monthly record(s) with HRC/XD breakdown")
# Gộp 3 nguồn — nếu TRÙNG tháng/năm, ưu tiên bản ghi có tách riêng HRC/XD (nguoiquansat/dautucophieu,
# từ báo cáo CTCK Research — chi tiết hơn) thay vì bản gộp chung (hoaphat.com.vn PR).
_by_period = {}
for r in HPG_PRODUCTION_NEWS + _nqs_records + _dtcp_records:
    key = (r["type"], r["year"], r.get("quarter") or r.get("month"))
    prev = _by_period.get(key)
    if prev is None or (r.get("hrc_kt") is not None and prev.get("hrc_kt") is None):
        _by_period[key] = r
HPG_PRODUCTION_NEWS = list(_by_period.values())

_official_q_rec = next((r for r in HPG_PRODUCTION_NEWS
                        if r["type"] == "quarter" and r["year"] == CUR_Q_YEAR and r["quarter"] == CUR_Q_NUM), None)
CUR_Q_MONTHLY_RECS = sorted(
    [r for r in HPG_PRODUCTION_NEWS if r["type"] == "month" and r["year"] == CUR_Q_YEAR and r["month"] in _cur_q_months],
    key=lambda r: r["month"])

if _official_q_rec:
    CUR_Q_SOURCE = "OFFICIAL"
    CUR_Q_TOTAL_KT = _official_q_rec["volume_kt"]
    print(f"  -> Found OFFICIAL {CUR_Q_LABEL} volume: {CUR_Q_TOTAL_KT:.0f} kt ({_official_q_rec['url']})")
elif CUR_Q_MONTHLY_RECS:
    CUR_Q_SOURCE = "ESTIMATED"
    _n_known = len(CUR_Q_MONTHLY_RECS)
    _known_total = sum(r["volume_kt"] for r in CUR_Q_MONTHLY_RECS)
    CUR_Q_TOTAL_KT = round(_known_total / _n_known * 3, 1)
    # Nếu tất cả tháng đã biết đều có tách riêng HRC/XD (nguồn nguoiquansat/Vietcap) → dùng trực tiếp,
    # chính xác hơn tỷ lệ lịch sử Q1 dùng làm fallback ở bước tính SL_HRC_A/SL_XD_A bên dưới.
    _hrc_known = [r["hrc_kt"] for r in CUR_Q_MONTHLY_RECS if r.get("hrc_kt") is not None]
    _xd_known = [r["xd_kt"] for r in CUR_Q_MONTHLY_RECS if r.get("xd_kt") is not None]
    CUR_Q_HRC_KT_DIRECT = round(sum(_hrc_known) / len(_hrc_known) * 3, 1) if len(_hrc_known) == _n_known else None
    CUR_Q_XD_KT_DIRECT = round(sum(_xd_known) / len(_xd_known) * 3, 1) if len(_xd_known) == _n_known else None
    print(f"  -> Found {_n_known}/3 months of {CUR_Q_LABEL} data, extrapolated total: {CUR_Q_TOTAL_KT:.0f} kt"
          + (f" (HRC={CUR_Q_HRC_KT_DIRECT:.0f}, XD={CUR_Q_XD_KT_DIRECT:.0f} kt, direct split from source)" if CUR_Q_HRC_KT_DIRECT else ""))
else:
    CUR_Q_SOURCE = "FALLBACK"
    CUR_Q_TOTAL_KT = None
    CUR_Q_HRC_KT_DIRECT = CUR_Q_XD_KT_DIRECT = None
    # Không fetch được tin nào cho quý đang chạy — có thể do mạng bị chặn (IP datacenter của GitHub
    # Actions hay bị site .vn chặn/giới hạn hơn IP dân dụng VN chạy local). Trước khi rơi về giả định
    # tĩnh cứng trong code (SL_HRC_A/SL_XD_A mặc định, có thể đã cũ nhiều tháng), thử dùng lại ước tính
    # đã lưu từ LẦN CHẠY GẦN NHẤT (data/HPG.json) nếu đúng quý đang xét — mới hơn giả định tĩnh, dù có
    # thể không mới bằng 1 lần fetch thành công thực sự.
    try:
        _cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "HPG.json")
        with open(_cache_path, "r", encoding="utf-8") as _f:
            _prev_cache = json.load(_f).get("productionEstimateCache")
        if _prev_cache and _prev_cache.get("year") == CUR_Q_YEAR and _prev_cache.get("quarter") == CUR_Q_NUM and _prev_cache.get("totalKt"):
            CUR_Q_SOURCE = "CACHED"
            CUR_Q_TOTAL_KT = _prev_cache["totalKt"]
            CUR_Q_HRC_KT_DIRECT = _prev_cache.get("hrcKt")
            CUR_Q_XD_KT_DIRECT = _prev_cache.get("xdKt")
            print(f"  [DIAG] Không fetch được tin sản lượng {CUR_Q_LABEL} lần này — dùng lại ước tính đã lưu "
                  f"lần chạy trước (lúc {_prev_cache.get('fetchedAt', '?')}, nguồn gốc {_prev_cache.get('source', '?')}): "
                  f"{CUR_Q_TOTAL_KT:.0f} kt")
    except Exception as _e:
        print(f"  [DIAG] Không đọc được cache ước tính sản lượng từ lần chạy trước: {_e}")
if CUR_Q_SOURCE not in ("ESTIMATED", "CACHED"):
    CUR_Q_HRC_KT_DIRECT = CUR_Q_XD_KT_DIRECT = None
if CUR_Q_SOURCE == "FALLBACK":
    print(f"  -> No production news found yet for {CUR_Q_LABEL} - using existing assumption (giả định tĩnh trong code)")

# Lưu lại ước tính vừa dùng để lần chạy SAU (kể cả trên GitHub Actions, nếu mạng bị chặn) có thể dùng
# lại làm fallback thay vì rơi thẳng về giả định tĩnh trong code — xem save_json_summary().
if CUR_Q_SOURCE in ("OFFICIAL", "ESTIMATED"):
    PRODUCTION_CACHE_ENTRY = {
        "year": CUR_Q_YEAR, "quarter": CUR_Q_NUM, "totalKt": CUR_Q_TOTAL_KT,
        "hrcKt": CUR_Q_HRC_KT_DIRECT, "xdKt": CUR_Q_XD_KT_DIRECT,
        "source": CUR_Q_SOURCE, "fetchedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
elif CUR_Q_SOURCE == "CACHED":
    PRODUCTION_CACHE_ENTRY = _prev_cache
else:
    PRODUCTION_CACHE_ENTRY = None

years_fc    = [2026, 2027, 2028]
SL_HRC_A    = [2.0, 2.2, 2.5, 2.8, 3.2, 6.0, 6.8, 7.5]      # Sản lượng HRC (triệu tấn), 2021..2028
SL_XD_A     = [2.8, 2.6, 2.3, 2.5, 2.8, 3.0, 3.2, 3.5]      # Sản lượng thép XD (triệu tấn)

# ── Sửa SL_HRC_A/SL_XD_A các năm 2023-2025 bằng SỐ THẬT (2026-07, user phát hiện bug) ─────────────
# User phát hiện: SL_HRC_A/SL_XD_A ở trên là GIẢ ĐỊNH TĨNH cũ (viết tay trước khi có
# HRC_SALES_HIST_KT/XD_SALES_HIST_KT — dữ liệu quý THẬT tổng hợp sau này), KHÔNG được đồng bộ lại nên
# lệch xa số thật (VD 2025 giả định 3.2 triệu tấn HRC nhưng lũy kế 4 quý thật = 5.0 triệu tấn) — khiến
# doanh thu 2025 (sheet 03_Revenue_Model) bị TÍNH THIẾU, kéo theo % tăng trưởng DT 2026E bị THỔI PHỒNG
# giả tạo (so với mẫu số 2025 sai thấp) dù bản chất 2026E không đổi. Sửa: năm nào có ĐỦ 4 quý trong
# SALES_Q_LABELS (2023, 2024, 2025) thì ghi đè bằng SUM 4 quý thật / 1000 (kt->Mt); năm 2021-2022
# (trước khi có dữ liệu tách quý) vẫn giữ giả định cũ vì chưa có số thật để thay.
def _annual_sum_from_quarters_kt(kt_list, q_labels, year):
    vals = [kt for lbl, kt in zip(q_labels, kt_list) if lbl.startswith(str(year))]
    return round(sum(vals) / 1000, 2) if len(vals) == 4 else None

for _yr in (2023, 2024, 2025):
    _idx = _yr - 2021
    _hrc_real = _annual_sum_from_quarters_kt(HRC_SALES_HIST_KT, SALES_Q_LABELS, _yr)
    _xd_real = _annual_sum_from_quarters_kt(XD_SALES_HIST_KT, SALES_Q_LABELS, _yr)
    if _hrc_real is not None:
        SL_HRC_A[_idx] = _hrc_real
    if _xd_real is not None:
        SL_XD_A[_idx] = _xd_real

# Nếu quý đang chạy có số liệu thực/ước tính (CUR_Q_TOTAL_KT) VÀ đúng năm 2026 → cập nhật lại sản
# lượng NĂM 2026E bằng blend_annual_estimate() (2026-07, đồng nhất với cách blend DTTC/CPTC ở trên và
# theo yêu cầu user: "không tuyến tính lấy Q1 x4... mà lấy số Q1 + giả định ban đầu x3/4"): Q1 luôn
# coi là ĐÃ BIẾT (n=1, có trong HRC_SALES_HIST_KT/XD_SALES_HIST_KT); quý đang chạy coi là ĐÃ BIẾT
# (n=2) MIỄN LÀ có CUR_Q_TOTAL_KT (chính thức hoặc ước tính từ tháng — mục đích chính của cả cơ chế
# dò tin sản lượng là để có tín hiệu SỚM cho quý này). Giả định GỐC (6.0/3.0 triệu tấn) dùng cho phần
# 2 quý CÒN LẠI (Q3+Q4, = (4-2)/4 = 1/2 giả định gốc), KHÔNG suy diễn từ 2 quý đã biết.
# Tách HRC/XD: ưu tiên CUR_Q_HRC_KT_DIRECT/CUR_Q_XD_KT_DIRECT (tách trực tiếp từ nguồn nguoiquansat/
# Vietcap Research khi có đủ cả 3 tháng) — chính xác hơn TỶ LỆ HRC/XD của Q1/2026 dùng làm fallback
# (KHÔNG suy từ văn bản bài PR gộp chung của hoaphat.com.vn — đã thử và bỏ, xem
# fetch_hpg_production_updates()).
if CUR_Q_YEAR == 2026 and CUR_Q_TOTAL_KT is not None:
    _orig_hrc_assumption_kt = SL_HRC_A[5] * 1000  # giả định gốc trước khi ghi đè (6.0 triệu tấn)
    _orig_xd_assumption_kt = SL_XD_A[5] * 1000    # giả định gốc trước khi ghi đè (3.0 triệu tấn)
    _q1_2026_total_kt = HRC_SALES_HIST_KT[-1] + XD_SALES_HIST_KT[-1]
    if CUR_Q_HRC_KT_DIRECT is not None and CUR_Q_XD_KT_DIRECT is not None:
        _cur_q_hrc_kt, _cur_q_xd_kt = CUR_Q_HRC_KT_DIRECT, CUR_Q_XD_KT_DIRECT
    else:
        _q1_2026_hrc_ratio = HRC_SALES_HIST_KT[-1] / _q1_2026_total_kt
        _cur_q_hrc_kt = CUR_Q_TOTAL_KT * _q1_2026_hrc_ratio
        _cur_q_xd_kt = CUR_Q_TOTAL_KT - _cur_q_hrc_kt
    _cum_hrc_kt = HRC_SALES_HIST_KT[-1] + _cur_q_hrc_kt
    _cum_xd_kt = XD_SALES_HIST_KT[-1] + _cur_q_xd_kt
    SL_HRC_A[5] = round(blend_annual_estimate(_cum_hrc_kt, 2, _orig_hrc_assumption_kt) / 1000, 2)
    SL_XD_A[5] = round(blend_annual_estimate(_cum_xd_kt, 2, _orig_xd_assumption_kt) / 1000, 2)
    print(f"  -> Updated 2026E annual volume (blend, n=2/4 quy da biet): HRC={SL_HRC_A[5]}Mt, XD={SL_XD_A[5]}Mt")

# ── SL_HRC_A/SL_XD_A năm SAU năm dự phóng đầu tiên (2027E - index 6) khi CHƯA có quý nào của năm đó
# (2026-07, theo yêu cầu user) — ước tính = SL 2 quý GẦN NHẤT đã biết (nửa năm) x2 (năm hóa) x1.05
# (tăng trưởng giả định). 2028E (index 7) nối tiếp cùng tốc độ tăng trưởng 5%/năm vì chưa có dữ liệu
# quý nào mới hơn để re-anchor lại (giống cách 2027E/2028E của Spread/giá hàng hóa cũng chỉ nối tiếp
# xu hướng — xem SPREAD_A[6]/[7] ở dưới).
_last2_hrc_kt = HRC_SALES_HIST_KT[-2] + HRC_SALES_HIST_KT[-1]
_last2_xd_kt = XD_SALES_HIST_KT[-2] + XD_SALES_HIST_KT[-1]
SL_HRC_A[6] = round(_last2_hrc_kt * 2 * 1.05 / 1000, 2)
SL_XD_A[6] = round(_last2_xd_kt * 2 * 1.05 / 1000, 2)
SL_HRC_A[7] = round(SL_HRC_A[6] * 1.05, 2)
SL_XD_A[7] = round(SL_XD_A[6] * 1.05, 2)
print(f"  -> 2027E/2028E volume (last-2Q x2 x1.05, then +5%/yr): "
      f"HRC={SL_HRC_A[6]}/{SL_HRC_A[7]}Mt, XD={SL_XD_A[6]}/{SL_XD_A[7]}Mt")

XD_PRICE_A  = [720, 580, 520, 550, 590, 610, 620, 630]      # Giá thép XD bq (USD/tấn) — chưa có nguồn fetch tự động, giữ giả định

# ── Tỷ giá USD/VND THEO TỪNG NĂM (2026-07, user phát hiện bug) ────────────────────────────────────
# Trước đây dùng 1 hằng số FX_RATE=25400 CHO CẢ 8 NĂM (2021-2028) để quy đổi doanh thu HRC/XD bottom-up
# — sai vì USD/VND đã tăng đáng kể qua các năm (~23.100 năm 2021 → ~26.200 hiện tại, 2026). Dùng 1 tỷ
# giá cố định làm lệch tỷ trọng doanh thu HRC/XD/Khác của các năm lịch sử (dù tổng doanh thu vẫn đúng
# vì "Doanh thu khác" là phần dư — nhưng ảnh hưởng trực tiếp doanh thu dự phóng 2026-2028 vì đó là
# tổng bottom-up thật). Tỷ giá lịch sử 2021-2025 = bình quân năm THAM KHẢO (Vietcombank/SBV, nghiên
# cứu thủ công — Vietcap API không có trường tỷ giá); 2026E = tỷ giá SỐNG fetch investing.com
# (_usd_vnd_rate, đã dùng cho quy đổi giá thép XD ở trên); 2027E/2028E = tiếp nối xu hướng mất giá
# ~2%/năm (giả định thận trọng, không có công thức nội tại vì là tỷ giá tương lai).
FX_RATE_HIST_A = [23100, 23600, 24000, 24900, 26000]  # 2021-2025, bình quân năm tham khảo
FX_RATE_A = FX_RATE_HIST_A + [round(_usd_vnd_rate), round(_usd_vnd_rate*1.02), round(_usd_vnd_rate*1.02**2)]
FX_RATE = FX_RATE_A[5]  # alias tương thích ngược (2026E) — chỗ nào còn dùng số đơn lẻ sẽ là tỷ giá hiện tại

# Giá quý ĐANG CHẠY (2026, chưa kết thúc) = MEDIAN các điểm giá đã quan sát được trong quý đó
# (giá cuối quý trước đã biết + giá hiện tại fetch live) — nhất quán với cách tính "giá năm =
# MEDIAN các quý" đã dùng cho lịch sử. Với 2 điểm quan sát, median = trung bình cộng, nhưng dùng
# median() (không phải phép cộng chia 2 thủ công) để khi user tích lũy thêm nhiều lần fetch trong
# quý (xem log file — mục "Chưa làm" trong skill), code không cần sửa lại cách tính.
# 2027-2028 = tiếp nối xu hướng giá hiện tại (giả định thận trọng, không có công thức nội tại vì là
# giá hàng hóa tương lai — xem ghi chú ở 02_Assumptions).
HRC_PRICE_A = [_year_median(_hrc_by_yr, y) for y in [2021,2022,2023,2024,2025]] + [
    round(stats.median([_year_median(_hrc_by_yr, 2026), hrc_now]), 1), None, None]
IRON_ORE_A = [_year_median(_iron_by_yr, y) for y in [2021,2022,2023,2024,2025]] + [
    round(stats.median([_year_median(_iron_by_yr, 2026), iron_now]), 1), None, None]
COKE_A = [_year_median(_coal_by_yr, y) for y in [2021,2022,2023,2024,2025]] + [
    round(stats.median([_year_median(_coal_by_yr, 2026), coal_now]), 1), None, None]
HRC_PRICE_A[6] = round(HRC_PRICE_A[5] * 1.02, 1); HRC_PRICE_A[7] = round(HRC_PRICE_A[6] * 1.015, 1)
IRON_ORE_A[6]  = round(IRON_ORE_A[5] * 1.03, 1);  IRON_ORE_A[7]  = round(IRON_ORE_A[6] * 1.02, 1)
COKE_A[6]      = round(COKE_A[5] * 1.02, 1);      COKE_A[7]      = round(COKE_A[6] * 1.02, 1)

# Spread HIỆN TẠI (quý đang chạy, ví dụ Q2/2026) = Giá HRC bình quân quý hiện tại (MEDIAN giá cuối
# quý trước đã biết + giá hiện tại fetch live) × 1.15 (thuế CBPG - quý đang chạy chắc chắn sau
# 06/07/2025) - 1.6×Quặng - 0.6×Than của QUÝ TRƯỚC (Q1/2026, đã biết đầy đủ, KHÔNG blend với giá
# live vì quặng/than dùng để tính giá vốn kỳ này là hàng tồn kho mua từ quý trước) - OTHER_COST_USD.
# Đây chính là công thức lag-1-quý áp dụng cho "quý đang chạy", tương tự cho Spread Rebar (không có
# thuế CBPG) và Spread All (bình quân gia quyền theo SL quý đang chạy — ưu tiên số tách trực tiếp
# CUR_Q_HRC_KT_DIRECT/XD, fallback về SL quý gần nhất đã biết đầy đủ nếu chưa có).
_hrc_avg_now = stats.median([Q18_HRC[-1], hrc_now])
SPREAD_HRC_NOW = _lag_spread(round(_hrc_avg_now * AD_HRC_MULTIPLIER, 1), Q18_IRON[-1], Q18_COAL[-1])
SPREAD_NOW = SPREAD_HRC_NOW  # alias tương thích ngược

_xd_avg_now = stats.median([Q18_XD[-1], xd_now])
SPREAD_REBAR_NOW = _lag_spread(_xd_avg_now, Q18_IRON[-1], Q18_COAL[-1])

_now_hrc_kt = CUR_Q_HRC_KT_DIRECT if CUR_Q_HRC_KT_DIRECT is not None else HRC_SALES_HIST_KT[-1]
_now_xd_kt = CUR_Q_XD_KT_DIRECT if CUR_Q_XD_KT_DIRECT is not None else XD_SALES_HIST_KT[-1]
_now_tot_kt = _now_hrc_kt + _now_xd_kt
SPREAD_ALL_NOW = (round((_now_xd_kt*SPREAD_REBAR_NOW + _now_hrc_kt*SPREAD_HRC_NOW) / _now_tot_kt, 1)
                  if _now_tot_kt else SPREAD_HRC_NOW)

# SPREAD_A / BASE_COST_A theo NĂM (8 phần tử, 2021-2028) — xây từ Q18_SPREAD_HRC (lag-1-quý, đã có
# thuế CBPG từ 2025Q3) thay vì trừ trực tiếp giá cùng năm, để khớp đúng công thức chuẩn ở mọi cấp độ
# (quý & năm). SPREAD_A vẫn là Spread HRC (dùng cho các so sánh/biểu đồ lịch sử hiện có) — Spread
# Rebar/All theo năm xem SPREAD_REBAR_A/SPREAD_ALL_A bên dưới:
#  - 2021-2025 (đã có đủ dữ liệu quý): Spread năm = MEDIAN các Spread quý (lag-1-quý) thuộc năm đó
#    (nhất quán với cách tính "Giá năm = MEDIAN 4 quý" đã dùng cho HRC/Quặng/Than).
#  - 2026E: TB(Spread quý gần nhất đã biết = Q1/2026 lag-1-quý, Spread hiện tại = quý đang chạy).
#  - 2027E/2028E: chưa có dữ liệu quý, xấp xỉ lag-1-quý bằng lag-1-NĂM (dùng giá quặng/than của
#    năm liền trước làm chi phí đầu vào, giá HRC ĐÃ ×1.15 vì đã sau 2025Q3) — cách xấp xỉ tốt nhất
#    khi chỉ có granularity theo năm.
_spread_by_yr = _group_by_year(Q18_SPREAD_HRC)
SPREAD_A = [None] * 8
for _i, _yr in enumerate([2021, 2022, 2023, 2024, 2025]):
    SPREAD_A[_i] = round(stats.median(_spread_by_yr[_yr]), 1)
SPREAD_A[5] = round(stats.median([Q18_SPREAD_HRC[-1], SPREAD_HRC_NOW]), 1)                                        # 2026E
SPREAD_A[6] = _lag_spread(round(HRC_PRICE_A[6]*AD_HRC_MULTIPLIER, 1), IRON_ORE_A[5], COKE_A[5])                   # 2027E
SPREAD_A[7] = _lag_spread(round(HRC_PRICE_A[7]*AD_HRC_MULTIPLIER, 1), IRON_ORE_A[6], COKE_A[6])                   # 2028E
BASE_COST_A = [HRC_PRICE_A[i] - SPREAD_A[i] for i in range(8)]  # chi phí quy đổi tương ứng (residual)

# XD_PRICE_A ở trên (giả định giá thép XD theo năm, USD/tấn) → SPREAD_REBAR_A theo cùng phương pháp
# lag-1-quý/năm như SPREAD_A, dùng cho biểu đồ Spread Rebar và (gián tiếp, qua SPREAD_ALL_A) công
# thức dự phóng BLNG năm.
_spread_rebar_by_yr = _group_by_year(Q18_SPREAD_REBAR)
SPREAD_REBAR_A = [None] * 8
for _i, _yr in enumerate([2021, 2022, 2023, 2024, 2025]):
    SPREAD_REBAR_A[_i] = round(stats.median(_spread_rebar_by_yr[_yr]), 1)
SPREAD_REBAR_A[5] = round(stats.median([Q18_SPREAD_REBAR[-1], SPREAD_REBAR_NOW]), 1)
SPREAD_REBAR_A[6] = _lag_spread(XD_PRICE_A[6], IRON_ORE_A[5], COKE_A[5])
SPREAD_REBAR_A[7] = _lag_spread(XD_PRICE_A[7], IRON_ORE_A[6], COKE_A[6])

# SPREAD_ALL_A theo năm = bình quân gia quyền SPREAD_A/SPREAD_REBAR_A theo SL_HRC_A/SL_XD_A (đã có
# đủ cả 8 năm, kể cả dự phóng — khác Q18_SPREAD_ALL theo quý chỉ có 13/18 quý có SL thật).
SPREAD_ALL_A = [
    round((SL_XD_A[i]*SPREAD_REBAR_A[i] + SL_HRC_A[i]*SPREAD_A[i]) / (SL_XD_A[i]+SL_HRC_A[i]), 1)
    if (SL_XD_A[i]+SL_HRC_A[i]) else SPREAD_A[i]
    for i in range(8)
]

# Doanh thu HRC/XD bottom-up (tỷ VND) = Sản lượng(triệu tấn) x Giá(USD/tấn) x FX / 1000
hrc_rev_all = [SL_HRC_A[i] * HRC_PRICE_A[i] * FX_RATE_A[i] / 1000 for i in range(8)]
xd_rev_all  = [SL_XD_A[i] * XD_PRICE_A[i] * FX_RATE_A[i] / 1000 for i in range(8)]

# Doanh thu khác (ống thép, tôn mạ, container, KCN, phôi billet) = phần dư lịch sử; dự phóng
# tăng 12%/năm — thận trọng hơn tốc độ tăng sản lượng HRC (DQ2 chủ yếu bổ sung công suất HRC,
# không tác động trực tiếp các mảng phụ trợ này).
OTHER_REV_GROWTH = 0.12
other_rev_all = [revenue_hist[i] - hrc_rev_all[i] - xd_rev_all[i] for i in range(5)]
for i in range(5, 8):
    other_rev_all.append(other_rev_all[-1] * (1 + OTHER_REV_GROWTH))

revenue_fc = [round(hrc_rev_all[i] + xd_rev_all[i] + other_rev_all[i]) for i in range(5, 8)]

# ── Biên LNG dự phóng: KHÔNG tính trực tiếp từ Spread (Spread chỉ phản ánh XU HƯỚNG chi phí/giá,
# không phải là biên lợi nhuận thực tế — bỏ qua khấu hao/tồn kho/mix sản phẩm/chi phí khác ngoài
# spread). Quay lại cách cũ: lấy Biên LNG THỰC TẾ của quý gần nhất đã công bố (Q1/2026) làm mốc,
# rồi NỘI SUY theo tỷ lệ Spread dự phóng năm / Spread quý gần nhất — Spread chỉ dùng để đo mức độ
# thay đổi tương đối, không dùng làm biên lợi nhuận tuyệt đối.
def _get_q_hpg_mod(records, yr, qtr, field):
    for rec in records:
        if rec.get("yearReport") == yr and rec.get("lengthReport") == qtr:
            return (rec.get(field, 0) or 0) / 1e9
    return 0

_is_qs_mod = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
q1_rev = _get_q_hpg_mod(_is_qs_mod, 2026, 1, 'isa3')
q1_gp  = _get_q_hpg_mod(_is_qs_mod, 2026, 1, 'isa5')
q1_gpm = q1_gp / q1_rev if q1_rev else 0
# BLNG năm dự phóng ĐẦU TIÊN (2026E) — theo yêu cầu user (2026-07), TỔNG QUÁT theo n quý ĐÃ CÓ BCTC
# (không còn chỉ neo Q1 như trước): LNG năm = LNG lũy kế n quý + (4-n)/4 x Doanh thu ước tính năm x
# (LNG lũy kế n quý/Doanh thu lũy kế n quý) x (Spread All hiện tại/Spread All quý gần nhất đã biết);
# n=4 (đủ cả năm) thì LNG năm = LNG lũy kế 4 quý luôn, không cần ước tính. Giống HỆT công thức Excel
# ở 02_Assumptions!row6 (xem khối patch cuối build_excel()) — 2 nơi PHẢI khớp nhau vì gp_margin_fc
# còn dùng cho narrative PDF/JSON, không chỉ riêng Excel.
_cum_rev_nq_mod, _n_q_known_mod = cumulative_actual_quarters(_is_qs_mod, years_fc[0], "isa3")
_cum_gp_nq_mod, _ = cumulative_actual_quarters(_is_qs_mod, years_fc[0], "isa5")
# "Spread All quý gần nhất đã biết" PHẢI dùng sản lượng CỦA CHÍNH quý đó làm quyền số (không phải
# quyền số quý đang chạy) → dùng thẳng Q18_SPREAD_ALL[-1] (đã tính đúng theo SL thật quý đó ở trên).
q1_spread_all = Q18_SPREAD_ALL[-1]

if _n_q_known_mod >= 4:
    gpm_2026 = round(_cum_gp_nq_mod / _cum_rev_nq_mod * 100, 1) if _cum_rev_nq_mod else 0.0
elif _n_q_known_mod >= 1 and revenue_fc[0]:
    _remain_frac_mod = (4 - _n_q_known_mod) / 4
    _cum_margin_mod = (_cum_gp_nq_mod / _cum_rev_nq_mod) if _cum_rev_nq_mod else 0.0
    _spread_ratio_mod = (SPREAD_ALL_NOW / q1_spread_all) if q1_spread_all else 1.0
    _lng_2026_mod = _cum_gp_nq_mod + _remain_frac_mod * revenue_fc[0] * _cum_margin_mod * _spread_ratio_mod
    gpm_2026 = round(_lng_2026_mod / revenue_fc[0] * 100, 1)
else:
    gpm_2026 = round(q1_gpm * 100, 1)

# Các năm SAU (2027E/2028E, chưa có quý nào của năm đó) = BLNG năm TRƯỚC x (Spread All hiện tại /
# Spread All NĂM TRƯỚC — annual ước tính của chính năm liền trước, SPREAD_ALL_A).
def _gpm_from_spread_ratio(base_gpm, spread_cur, spread_base):
    return round(base_gpm * (spread_cur / spread_base), 1) if spread_base else round(base_gpm, 1)

gpm_2027 = _gpm_from_spread_ratio(gpm_2026, SPREAD_ALL_NOW, SPREAD_ALL_A[5])
gpm_2028 = _gpm_from_spread_ratio(gpm_2027, SPREAD_ALL_NOW, SPREAD_ALL_A[6])
gp_margin_fc = [gpm_2026, gpm_2027, gpm_2028]

# EBIT/EBT/NI dự phóng = GP(= Rev x Biên LNG nội suy) - SGKA (Rev x tỷ lệ CP BH&QLDN/DT) rồi +
# DTTC - CPTC - Thuế, khớp đúng chuỗi công thức 04_PnL (Rev -> COGS -> GP -> SGKA -> EBIT -> EBT
# -> Thuế -> NI).
SGKA_RATE_FC  = 0.035   # Tỷ lệ CP BH&QLDN/DT dự phóng (2026-2028), khớp 02_Assumptions dòng 15
FIN_INCOME_FC = [2500, 2800, 3000]   # Doanh thu TC dự phóng (tỷ) — khớp 02_Assumptions/04_PnL
FIN_COST_FC   = [3800, 4000, 4200]   # Chi phí TC dự phóng (tỷ) — khớp 02_Assumptions/04_PnL
TAX_RATE_FC   = [0.12, 0.12, 0.12]   # Thuế TNDN dự phóng — khớp 02_Assumptions dòng 8

# DTTC/CPTC 2026E — BLEND với số quý ĐÃ CÓ báo cáo thực tế (2026-07, theo yêu cầu user): trước đây
# là giả định năm CỐ ĐỊNH, hoàn toàn bỏ qua thực tế Q1/2026 (DTTC Q1 thực tế đột biến rất cao do lãi
# tỷ giá/cổ tức — xem mục 4C CHẤT LƯỢNG LỢI NHUẬN) — vừa SAI (giả định thấp hơn cả số Q1 một mình) vừa
# nguy hiểm nếu đổi sang ngoại suy tuyến tính Q1x4 (sẽ thổi phồng cả năm theo đúng yếu tố đột biến chỉ
# xảy ra 1 quý). Dùng blend_annual_estimate(): quý ĐÃ biết lấy số THẬT, quý CHƯA biết vẫn theo giả
# định gốc — xem docstring hàm trong fetch_data.py để biết chi tiết công thức.
_fin_income_cum, _n_fin_income_q = cumulative_actual_quarters(_is_qs_mod, 2026, 'isa6')
_fin_cost_cum, _n_fin_cost_q = cumulative_actual_quarters(_is_qs_mod, 2026, 'isa7')
_fin_cost_cum = abs(_fin_cost_cum)
FIN_INCOME_FC[0] = round(blend_annual_estimate(_fin_income_cum, _n_fin_income_q, FIN_INCOME_FC[0]))
FIN_COST_FC[0] = round(blend_annual_estimate(_fin_cost_cum, _n_fin_cost_q, FIN_COST_FC[0]))
print(f"[Blend] 2026E DTTC: {_n_fin_income_q}/4 quy da biet (luy ke {_fin_income_cum:,.0f} ty) -> "
      f"blend = {FIN_INCOME_FC[0]:,} ty; CPTC: {_n_fin_cost_q}/4 quy (luy ke {_fin_cost_cum:,.0f} ty) -> "
      f"blend = {FIN_COST_FC[0]:,} ty")
gp_fc   = [revenue_fc[i] * gp_margin_fc[i] / 100 for i in range(3)]
ebit_fc = [gp_fc[i] - revenue_fc[i]*SGKA_RATE_FC for i in range(3)]
ebt_fc  = [ebit_fc[i] + FIN_INCOME_FC[i] - FIN_COST_FC[i] for i in range(3)]
ni_fc   = [round(ebt_fc[i] * (1 - TAX_RATE_FC[i])) for i in range(3)]

# "Tăng trưởng DT (%)" và "EBIT Margin (%)" dự phóng — tính ra từ revenue_fc/ebit_fc bottom-up
# ở trên (thay vì số gõ tay độc lập), để 02_Assumptions luôn khớp đúng với 03_Revenue_Model/04_PnL.
revenue_prev_fc = [revenue_hist[-1]] + revenue_fc[:-1]
revenue_growth_fc = [round((revenue_fc[i]/revenue_prev_fc[i]-1)*100, 1) for i in range(3)]
ebit_margin_fc = [round(ebit_fc[i]/revenue_fc[i]*100, 1) for i in range(3)]

# D&A dự phóng = D&A năm trước x (1 + Tăng trưởng DT năm đó) — ĐÚNG công thức sống ghi vào
# 02_Assumptions!row9 (xem build_excel()), neo gốc da_hist[-1] (2025A thật). Trước đây là số gõ tay
# cứng [7000,8000,9000] không đồng bộ khi Doanh thu/Tăng trưởng DT dự phóng thay đổi — khiến EBITDA
# (EBIT+D&A) dùng cho định giá EV/EBITDA lệch xa so với Excel (vd 27.598 vs Excel thật 32.545 tỷ).
da_fc = []
_da_prev_fc = da_hist[-1]
for _i in range(3):
    _da_prev_fc = round(_da_prev_fc * (1 + revenue_growth_fc[_i] / 100))
    da_fc.append(_da_prev_fc)
capex_fc   = [15000, 18000,  20000]

# Tồn kho & Phải thu dự phóng (tỷ VND) — nguồn duy nhất cho 05_Balance_Sheet, 14_Steel_Analysis
# (DIO/DSO) và biểu đồ turnover, tránh lệch số giữa các nơi khi sửa 1 chỗ quên sửa chỗ khác.
INVENTORY_FC   = [45000, 48000, 50000]
RECEIVABLES_FC = [18000, 20000, 22000]
cogs_fc = [round(revenue_fc[i] - gp_fc[i]) for i in range(3)]

# Số ngày tồn kho bình quân (DIO) & Số ngày phải thu bình quân (DSO) = 365 x Số dư BÌNH QUÂN
# (đầu kỳ+cuối kỳ)/2 / GVHB hoặc Doanh thu — cùng công thức với sheet 14_Steel_Analysis (Section 4,
# Excel formula sống) để khớp tuyệt đối giữa PDF/chart và Excel. CHỈ LỊCH SỬ (2021-2025) theo yêu
# cầu user — DIO/DSO không ảnh hưởng nhiều tới model định giá nên KHÔNG dự phóng 2026E-2028E.
# 2021 không có số dư đầu kỳ 2020 trong dữ liệu API nên dùng số dư cuối kỳ (không bình quân).
DIO_A, DSO_A = [], []
for i in range(5):
    avg_inv = inventory_hist[i] if i == 0 else (inventory_hist[i-1] + inventory_hist[i]) / 2
    avg_rec = receivables_hist[i] if i == 0 else (receivables_hist[i-1] + receivables_hist[i]) / 2
    DIO_A.append(round(365 / (cogs_hist[i] / avg_inv), 1))
    DSO_A.append(round(365 / (revenue_hist[i] / avg_rec), 1))

# DIO/DSO theo QUÝ (2021Q4-2026Q1, khớp Q18_LABELS) — quy đổi ra "ngày" trên cơ sở NĂM HÓA
# (GVHB/Doanh thu quý x4) để so sánh cùng thang đo với DIO/DSO năm ở trên. Dữ liệu THỰC TẾ từ BCTC
# quý, không dự phóng — dùng để vẽ biểu đồ DIO/DSO theo quý VÀ ghép cặp với Q18_SPREAD cho biểu đồ
# tương quan Spread-Biên LNG theo quý.
_bs_qs_mod = section_to_quarters(FIN_DATA, "BALANCE_SHEET")
def _get_q_bs_mod(records, yr, qtr, field):
    for rec in records:
        if rec.get("yearReport") == yr and rec.get("lengthReport") == qtr:
            return (rec.get(field, 0) or 0) / 1e9
    return 0
_q_periods_wc = [(int(l[:4]), int(l[5])) for l in Q18_LABELS]
_inv_q  = [_get_q_bs_mod(_bs_qs_mod, y, q, "bsa15") for y, q in _q_periods_wc]
_rec_q  = [_get_q_bs_mod(_bs_qs_mod, y, q, "bsa8") for y, q in _q_periods_wc]
_rev_q  = [_get_q_hpg_mod(_is_qs_mod, y, q, "isa3") for y, q in _q_periods_wc]
_cogs_q = [abs(_get_q_hpg_mod(_is_qs_mod, y, q, "isa4")) for y, q in _q_periods_wc]
DIO_Q, DSO_Q = [], []
for i in range(len(_q_periods_wc)):
    avg_inv_q = _inv_q[i] if i == 0 else (_inv_q[i-1] + _inv_q[i]) / 2
    avg_rec_q = _rec_q[i] if i == 0 else (_rec_q[i-1] + _rec_q[i]) / 2
    DIO_Q.append(round(365 / (_cogs_q[i]*4 / avg_inv_q), 1) if _cogs_q[i] else None)
    DSO_Q.append(round(365 / (_rev_q[i]*4 / avg_rec_q), 1) if _rev_q[i] else None)

# Peer data
peers = {
    "HPG": {"pe": 12.5, "pb": 1.8, "roe": 15.6, "ev_ebitda": 6.8, "ni_growth": 42.0},
    "HSG": {"pe": 13.7, "pb": 0.9, "roe": 7.1,  "ev_ebitda": 7.2, "ni_growth": 25.0},
    "NKG": {"pe": 22.5, "pb": 0.9, "roe": 6.4,  "ev_ebitda": 9.0, "ni_growth": 18.0},
}

# Shares từ Vốn điều lệ (bsa80) / 10,000
def shares_from_charter(records, year):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get("bsa80")
            if v is not None:
                return round(v / 10000 / 1e6)
    return 8445

shares_hist = [shares_from_charter(bs_recs, y) for y in years_hist]
shares_fc   = [8445, 8445, 8445]
shares_all  = shares_hist + shares_fc
print(f"[SHARES] Hist: {shares_hist} | Fc: {shares_fc}")

# VCSH dự phóng = roll-forward VCSH năm trước + NI - Cổ tức tiền mặt đã trả (thay vì số gõ tay
# độc lập trước đây, không khớp với NI thực tế tính từ P&L bottom-up ở trên).
DIV_PER_SHARE_FC = [0, 800, 1200]  # VND/CP — khớp dòng "Cổ tức" trong 02_Assumptions
equity_fc_val = []
prev_equity = equity_hist[-1]
for i in range(3):
    dividend_paid = DIV_PER_SHARE_FC[i] * shares_fc[i] / 1000  # VND/CP * triệu CP / 1000 = tỷ VND
    prev_equity = prev_equity + ni_fc[i] - dividend_paid
    equity_fc_val.append(round(prev_equity))

# Leading indicators
leading_indicators = [
    ("Giá HRC (USD/tấn)", 1100, 900, 1110, "Tích cực"),
    ("Giá quặng sắt (USD/tấn)", 90, 120, 106, "Trung tính"),
    ("Giá than cốc (USD/tấn)", 200, 280, 240, "Trung tính"),
    ("Tỷ giá USD/VNĐ", 24000, 26000, 25400, "Trung tính"),
    ("Sản lượng thép VN (triệu tấn/tháng)", 2.5, 1.5, 2.92, "Tích cực"),
    ("SX ngành thép 4M (triệu tấn)", 12.5, 8, 11.67, "Phục hồi (+23.5%)"),
    ("LNST QoQ HPG (tỷ)", 7000, 3000, 9056, "Tích cực"),
    ("D/E", 0.5, 1.5, 0.65, "Tích cực"),
    ("CIP (tỷ)", 10000, 40000, 15000, "Tích cực (DQ2 done)"),
    ("Spread thép (USD/tấn)", 200, 50, 180, "Tích cực"),
    ("Đầu tư công 4M (nghìn tỷ)", 200, 150, 187, "Tích cực (+10.4%)"),
    ("Room ngoại (tỷ USD)", 1.5, 0.5, 1.2, "Trung tính"),
    ("KLGD BQ 20 phiên (triệu cp)", 50, 20, 33.2, "Trung tính"),
    ("Biến động giá 1 tháng (%)", 10, -5, 3.3, "Trung tính"),
    ("Biến động giá YTD (%)", 20, -10, 6.1, "Trung tính"),
]

# PESTLE
pestle = [
    ("Chính trị - Pháp luật", "Thuế CBPG HRC khổ hẹp 19.38-27.83% + AD chống lẩn tránh HRC rộng 27.83% (17/04/2026). Ổn định chính trị VN.", "Tích cực"),
    ("Kinh tế", "GDP VN 2026E ~6.5%. Đầu tư công tăng mạnh (cao tốc, Long Thành). LS giảm.", "Tích cực"),
    ("Xã hội", "Đô thị hoá ~38%. Nhu cầu nhà ở & hạ tầng tăng.", "Tích cực"),
    ("Công nghệ", "DQ2 BOF hiện đại, tự động hoá cao. Tiết kiệm 15-20% chi phí so với cũ.", "Tích cực"),
    ("Môi trường", "Áp lực giảm phát thải. HPG đầu tư lọc bụi, xử lý nước thải theo chuẩn quốc tế.", "Trung tính"),
    ("Pháp lý khác", "NĐ 153 trái phiếu siết chặt. Room tín dụng ngân hàng mở rộng.", "Trung tính"),
]

# ── HISTORICAL MULTIPLES ──────────────────────────────────────────────────

def fetch_vietcap_ratios(ticker, timeout=15):
    """Fetch historical P/E, P/B, EV/EBITDA from Vietcap API."""
    try:
        r = requests.get(
            f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial",
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": "https://trading.vietcap.com.vn/",
            },
            timeout=timeout,
        )
        r.raise_for_status()
        data = r.json().get("data", [])
        return [
            {
                "year": d.get("year"),
                "quarter": d.get("quarter"),
                "pe": d.get("pe"),
                "pb": d.get("pb"),
                "ev_ebitda": d.get("evToEbitda"),
                "market_cap": d.get("marketCap"),
                "ebitda": d.get("ebitda"),
                "roe": d.get("roe"),
            }
            for d in data
        ]
    except Exception as e:
        print(f"[WARN] Vietcap ratio API failed: {e}")
        return []

HPG_RATIOS = fetch_vietcap_ratios(TICKER)
ttms = sorted(
    [r for r in HPG_RATIOS if r.get("quarter") not in (None, 5) and r.get("pe", 0) or 0 > 0],
    key=lambda x: (x["year"], x["quarter"]),
    reverse=True,
)
seen = set()
pe_hist, pb_hist, ev_hist = [], [], []
for r in ttms:
    y = r["year"]
    if y not in seen and len(pe_hist) < 5:
        seen.add(y)
        p = r.get("pe")
        if p and 0 < p < 100:
            pe_hist.append(round(p, 1))
        p = r.get("pb")
        if p and p > 0:
            pb_hist.append(round(p, 2))
        p = r.get("ev_ebitda")
        if p and p > 0:
            ev_hist.append(round(p, 1))

# Fallback for any missing years
FALLBACK_PE = [8.7, 15.9, 24.8, 18.9, 13.7]
FALLBACK_PB = [1.8, 1.2, 1.4, 1.6, 1.5]
FALLBACK_EV = [9.2, 10.7, 12.7, 11.4, 10.6]
if len(pe_hist) < 5:
    pe_hist = FALLBACK_PE[:]
if len(pb_hist) < 5:
    pb_hist = FALLBACK_PB[:]
if len(ev_hist) < 5:
    ev_hist = FALLBACK_EV[:]

print(f"[RATIOS] P/E: {pe_hist} | P/B: {pb_hist} | EV/EBITDA: {ev_hist}")

# ── HELPER FUNCTIONS ────────────────────────────────────────────────────────

def fmt_vnd(val):
    """Format VND in billions"""
    if val is None or val == 0:
        return "-"
    return f"{val/1e9:,.0f}"

def fmt_bn(val):
    """Format billions"""
    if val is None or val == 0:
        return "-"
    return f"{val:,.0f}"

def fmt_pct(val):
    if val is None:
        return "-"
    return f"{val:.1f}%"

def fmt_eps(val):
    if val is None:
        return "-"
    return f"{val:,.0f}"

# Styles
FONT_NAME = "Calibri"
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
title_font  = Font(name=FONT_NAME, bold=True, size=14, color="1F4E79")
sub_font    = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
data_font   = Font(name=FONT_NAME, size=10)
bold_font   = Font(name=FONT_NAME, bold=True, size=10)
assump_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # vàng
input_fill  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # xanh dương nhạt
p_fill      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # positive
n_fill      = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # negative

def header_row(ws, row, cols, widths=None):
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def data_row(ws, row, vals, bold=False, fill=None, fmt=None):
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = bold_font if bold else data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill
        if fmt and i <= len(fmt):
            cell.number_format = fmt[i-1]

# ── Nguồn số valuation DÙNG CHUNG cho Excel/PDF/JSON (2026-07) ──────────────────────────────
# Trước đây build_pdf()/save_json_summary() tự tính EV/PB/PE median và net-debt 2026E bằng hằng số
# gõ tay riêng (EV_MULTIPLE=9.0/PB_MULTIPLE=1.6/PE_MULTIPLE=12.0, net_debt=78000-25000), độc lập với
# công thức MEDIAN/Nợ vay-Tiền mặt sống mà build_excel() ghi vào 07_Valuation/02_Assumptions — gây
# lệch số giữa Excel/PDF/web. Thay vì tính lại 2 lần hay mở Excel đọc lại (win32com — không chạy
# được trên GitHub Actions vì runner không có Excel), hai hàm dưới đây tính THẲNG từ cùng nguồn dữ
# liệu (HPG_RATIOS, total_debt_hist, cash_for_valuation_hist) bằng đúng công thức mà build_excel()
# dùng để ghi công thức Excel — một nguồn số duy nhất, chạy giống hệt nhau trên mọi môi trường.
def _compute_hist_multiple_medians():
    """Trung vị-của-trung-vị-năm (2018-2025) cho P/E (đã điều chỉnh quý lỗ)/P/B/EV-EBITDA từ
    HPG_RATIOS — đúng logic/dữ liệu dùng để ghi công thức MEDIAN vào sheet 08_Hist_Multiples (xem
    build_excel(), khu vực 'Sheet 8: Lịch sử P/E, P/B, EV/EBITDA theo quý'), chỉ khác là ra số
    python trực tiếp thay vì công thức Excel."""
    raw_by_q = {}
    for r in HPG_RATIOS:
        if r.get("quarter") != 5:
            raw_by_q[(int(r["year"]), r["quarter"])] = r
    all_quarters = [(y, q) for y in range(2018, 2027) for q in range(1, 5)]
    qdata = []
    last_good_pe = None; last_good_pb = None; last_good_ev = None
    for y, q in all_quarters:
        rec = raw_by_q.get((y, q), {})
        pe0 = rec.get("pe"); pb0 = rec.get("pb"); ev0 = rec.get("ev_ebitda")
        is_normal = bool(pe0 and 0 < pe0 < 50)
        pe_adj = round(pe0, 1) if is_normal else last_good_pe
        if is_normal: last_good_pe = pe_adj
        if pb0 and pb0 > 0:
            pb = round(pb0, 2); last_good_pb = pb
        else:
            pb = last_good_pb
        if ev0 and ev0 > 0:
            ev = round(ev0, 1); last_good_ev = ev
        else:
            ev = last_good_ev
        qdata.append((y, q, pe_adj, pb, ev))

    hist_years = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
    def year_median(idx):
        yearly_medians = []
        for y in hist_years:
            vals = [row[idx] for row in qdata if row[0] == y and row[idx] is not None]
            if vals:
                yearly_medians.append(stats.median(vals))
        return round(stats.median(yearly_medians), 2) if yearly_medians else None

    return year_median(2), year_median(3), year_median(4)  # pe_k, pb_k, ev_k

PE_HIST_MEDIAN, PB_HIST_MEDIAN, EV_HIST_MEDIAN = _compute_hist_multiple_medians()

# Nợ vay/Tiền mặt dự phóng 2026E — đúng công thức rollforward dùng để ghi 02_Assumptions (Nợ vay
# row 11/Tiền mặt row 12): neo gốc 2025A thật, giữ %YoY giả định cũ (giảm dần nợ vay, tăng dần tiền
# mặt do FCF chuyển dương).
_debt_yoy_g = [78000/80000, 75000/78000, 72000/75000]
_cash_yoy_g = [25000/20000, 30000/25000, 35000/30000]
_debt_fc_g = [round(total_debt_hist[-1])]
_cash_fc_g = [round(cash_for_valuation_hist[-1])]
for _r in _debt_yoy_g:
    _debt_fc_g.append(round(_debt_fc_g[-1] * _r))
for _r in _cash_yoy_g:
    _cash_fc_g.append(round(_cash_fc_g[-1] * _r))
DEBT_2026E = _debt_fc_g[1]
CASH_2026E = _cash_fc_g[1]
NET_DEBT_2026E = DEBT_2026E - CASH_2026E


# ── 1. EXCEL MODEL ──────────────────────────────────────────────────────────

def build_excel():
    # Sheet name constants for formulas (must include quotes when name starts with number)
    S_ASSUMP = "'02_Assumptions'"
    S_PNL    = "'04_PnL'"
    S_BS     = "'05_Balance_Sheet'"
    S_HIST   = "'08_Hist_Multiples'"
    S_R3     = "'03_Revenue_Model'"

    wb = openpyxl.Workbook()
    # ─── Sheet 1: Cover (built here; formulas reference later sheets) ───
    ws = wb.active
    ws.title = "01_Cover"
    ws.merge_cells('A1:F1')
    ws['A1'] = f"PHÂN TÍCH CỔ PHIẾU {TICKER} — {COMPANY}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35

    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    info = [
        ("Ticker", TICKER, None, left_align),
        ("Sàn", EXCHANGE, None, left_align),
        ("Ngành", INDUSTRY, None, left_align),
        ("Giá hiện tại (VND)", PRICE, '#,##0', right_align),
        ("Số CP lưu hành", SHARES, '#,##0', right_align),
        ("Vốn hóa (tỷ)", MARKET_CAP / 1e9, '#,##0', right_align),
        ("P/E TTM", "='02_Assumptions'!J22", '0.0"x"', right_align),
        ("P/B", "='02_Assumptions'!J23", '0.00"x"', right_align),
        ("EV/EBITDA", "='02_Assumptions'!J24", '0.0"x"', right_align),
        ("52W Cao/Thấp", "28,045 / 19,789", None, right_align),
        ("Room ngoại", "20.6%", None, right_align),
        ("Khuyến nghị", "MUA", None, Alignment(horizontal='center', vertical='center')),
        ("Giá mục tiêu Base", "='07_Valuation'!C10", '#,##0', right_align),
        ("Upside", "='07_Valuation'!C13", '0.0%', right_align),
        ("Ngày phân tích", f"{datetime.now().strftime('%d/%m/%Y %H:%M')}", None, left_align),
    ]
    for i, (k, v, nf, al) in enumerate(info, 2):
        ws.cell(row=i, column=1, value=k).font = bold_font
        ws.cell(row=i, column=1).border = thin_border
        ws.cell(row=i, column=1).alignment = left_align
        cell = ws.cell(row=i, column=2)
        if nf and isinstance(v, (int, float)):
            cell.value = v; cell.number_format = nf
        else:
            cell.value = v
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = al
    # Add EPS, BVPS, KLGD, shareholder, price perf
    eps_val = round(sum(eps_hist[y_idx] for y_idx in [3,4])/2) if len(eps_hist) >=5 else eps_hist[-1]
    bvps_val = round(equity_hist[4] * 1e9 / SHARES)
    extra_rows = [
        ("EPS TTM (VND)", eps_val, '#,##0', right_align),
        ("BVPS (VND)", bvps_val, '#,##0', right_align),
        ("KLGD BQ 20 phiên (cp)", 33200000, '#,##0', right_align),
        ("KLGD BQ 20 phiên (tỷ)", 890.7, '#,##0.0', right_align),
        ("Cổ đông sáng lập", "35.7% (Ông Trần Đình Long & gia đình)", None, left_align),
        ("Biến động 1 tháng", "+3.3%", None, right_align),
        ("Biến động 3 tháng", "+6.9%", None, right_align),
        ("Biến động YTD", "+6.1%", None, right_align),
        ("VNINDEX 1 tháng", "+4.4%", None, right_align),
        ("VNINDEX 3 tháng", "-6.3%", None, right_align),
        ("VNINDEX YTD", "-1.9%", None, right_align),
    ]
    for i, (k, v, nf, al) in enumerate(extra_rows, len(info) + 2):
        ws.cell(row=i, column=1, value=k).font = bold_font
        ws.cell(row=i, column=1).border = thin_border
        ws.cell(row=i, column=1).alignment = left_align
        cell = ws.cell(row=i, column=2)
        if nf and isinstance(v, (int, float)):
            cell.value = v; cell.number_format = nf
        else:
            cell.value = v
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = al
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 44

    # ─── Sheet 2: Assumptions ───
    ws2 = wb.create_sheet("02_Assumptions")
    headers = ["Giả định", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "Ghi chú"]
    widths = [35, 12, 12, 12, 12, 12, 12, 12, 12, 40]
    header_row(ws2, 1, headers, widths)

    # ── D&A/CAPEX/Nợ vay/Tiền mặt LỊCH SỬ (2026-07, user phát hiện bug) ──────────────────────────
    # User nghi ngờ đúng: các dòng này ở Assumptions trước đây là SỐ CHẾT đoán tay từ đầu dự án, hoàn
    # toàn KHÔNG khớp dữ liệu THẬT đã fetch từ Vietcap (da_hist/total_debt_hist/capex_hist đã có sẵn,
    # dùng đúng ở sheet 05/06 — chỉ riêng Assumptions bị bỏ quên không đồng bộ, giống hệt bug SL_HRC_A).
    # Đối chiếu: D&A 2025 số chết 5.500 vs thật 8.471 (khớp lời user "khấu hao 2024-2025 lớn hơn nhiều");
    # Nợ vay 2025 số chết 80.000 vs thật 92.174 (khớp đúng ghi chú có sẵn "Q1/2026: 90,6k giảm 2k" —
    # 92.174-2.000≈90.174, gần đúng 90,6k hơn hẳn số chết 80.000 cũ); Tiền mặt: dùng CẢ "Tiền và tương
    # đương tiền" (bsa2) + "Đầu tư ngắn hạn" (bsa5, chủ yếu tiền gửi kỳ hạn — thanh khoản cao tương
    # đương tiền mặt cho mục đích Net Debt) — user hỏi nên dùng cái nào, câu trả lời là CẢ HAI cộng lại.
    _da_hist_r = [round(v) for v in da_hist]
    _capex_hist_r = [round(v) for v in capex_hist]
    _debt_hist_r = [round(v) for v in total_debt_hist]
    _cash_hist_r = [round(v) for v in cash_for_valuation_hist]
    # Dự phóng Nợ vay/Tiền mặt (2026E-2028E) — GIỮ NGUYÊN xu hướng %YoY của giả định cũ (giảm dần nợ
    # vay do không vay thêm DQ2; tăng dần tiền mặt do FCF chuyển dương) nhưng NEO LẠI đúng gốc 2025A
    # thật thay vì gốc số chết cũ — tránh bước nhảy phi lý giữa 2025A (đã sửa đúng) và 2026E (vẫn theo
    # gốc sai cũ). D&A/CAPEX dự phóng KHÔNG cần sửa: D&A đã là công thức sống (=D&A(t-1)*(1+tăng
    # trưởng DT)) tự động đúng khi có gốc 2025A thật; CAPEX dự phóng là giả định giảm hậu-DQ2, không
    # phụ thuộc gốc lịch sử.
    # Dùng lại _debt_fc_g/_cash_fc_g đã tính sẵn ở module-level (cùng nguồn số với NET_DEBT_2026E
    # dùng cho PDF/JSON) thay vì tính lại — đảm bảo Assumptions và giá trị valuation luôn khớp nhau.
    _debt_fc_r, _cash_fc_r = _debt_fc_g[1:], _cash_fc_g[1:]

    assumptions = [
        ("Giá HPG (VND)", 23600, 23600, 23600, 23600, 23600, 23600, 23600, 23600, "Giá 24/06/2026. Broker targets: SSI 36k, VND 37k, SHS 38k, VCBS 38k, BVSC 38.65k"),
        ("Số CP lưu hành (triệu)", shares_all[0], shares_all[1], shares_all[2], shares_all[3], shares_all[4], shares_all[5], shares_all[6], shares_all[7], "Vốn điều lệ từng năm / 10,000. Post-split 1.1:1 T05/2026"),
        ("Doanh thu (tỷ)", 149680, 141410, 118950, 138860, 156120, revenue_fc[0], revenue_fc[1], revenue_fc[2], "Bottom-up = SL HRC/XD x Giá x FX + Doanh thu khác (Sheet 03_Revenue_Model)"),
        ("Tăng trưởng DT (%)", 48.0, -5.5, -15.9, 16.7, 12.4, revenue_growth_fc[0], revenue_growth_fc[1], revenue_growth_fc[2], "= Doanh thu(t) / Doanh thu(t-1) - 1, tính từ Doanh thu bottom-up"),
        ("Biên LNG (%)", 27.46, 11.85, 10.88, 13.32, 15.69, gp_margin_fc[0], gp_margin_fc[1], gp_margin_fc[2], "Bottom-up từ Spread (Giá - Quặng*1.6 - Than*0.5 - CP khác) x Sản lượng — khớp Profit Bridge 03_Revenue_Model"),
        ("EBIT Margin (%)", 24.1, 7.8, 6.5, 9.3, 11.84, ebit_margin_fc[0], ebit_margin_fc[1], ebit_margin_fc[2], "= EBIT / Doanh thu, EBIT = LN gộp - CP BH&QLDN (khớp 04_PnL)"),
        ("Thuế TNDN (%)", 4.0, 13.5, 10.5, 4.0, 12.5, 12.0, 12.0, 12.0, "HPG ưu đãi Dung Quất"),
        ("D&A (tỷ)", *_da_hist_r, 7000, 8000, 9000, "Khấu hao TSCĐ & BĐSĐT (cfa2) — số THẬT từ Vietcap. Dự phóng = công thức sống tăng theo DT"),
        ("CAPEX (tỷ)", *_capex_hist_r, 15000, 18000, 20000, "Số THẬT từ Vietcap (cfa19). 2024 đỉnh CAPEX DQ2 (~35.5k tỷ). Dự phóng giảm hậu-DQ2"),
        ("Nợ vay (tỷ)", *_debt_hist_r, *_debt_fc_r, "Vay NH+DH (bsa56+71) — số THẬT. Q1/2026: 90,6k giảm 2k so với 2025A; dự phóng giữ %YoY giảm dần cũ, neo gốc thật"),
        ("Tiền mặt (tỷ)", *_cash_hist_r, *_cash_fc_r, "Tiền&TĐT (bsa2) + Đầu tư ngắn hạn (bsa5, chủ yếu tiền gửi kỳ hạn) — số THẬT, dùng cho Net Debt"),
        ("VCSH (tỷ)", 75000, 90000, 105000, 115000, 130000, equity_fc_val[0], equity_fc_val[1], equity_fc_val[2], "Roll-forward: VCSH(t-1) + NI(t) - Cổ tức tiền mặt(t)"),
        ("Cổ tức (VND/CP)", 0, 0, 0, 0, 0, 0, 800, 1200, "2026: 15% (10% CP + 5% TM)"),
        ("Tỷ lệ CP BH&QLDN/DT (%)", 3.4, 4.0, 4.3, 4.0, 3.8, 3.5, 3.5, 3.5, "Giảm nhờ DQ2 biên lớn"),
        ("P/B mục tiêu (x)", 1.5, 1.5, 1.5, 1.5, 1.5, 1.6, 1.6, 1.6, "HPG lịch sử median 1.61x (TTM 2018-2026)"),
        ("EV/EBITDA mục tiêu (x)", 6.5, 6.5, 6.5, 6.5, 6.5, 9.0, 9.0, 8.5, "HPG lịch sử median 8.95x (TTM 2018-2026)"),
        ("P/E mục tiêu (x)", 10.0, 10.0, 10.0, 10.0, 10.0, 12.0, 12.0, 11.0, "Tham khảo - HPG median 11.1x (TTM)"),
        # Spread-based GP margin drivers (rows 19+) — dùng chung IRON_ORE_A/COKE_A với Profit Bridge
        # (03_Revenue_Model), 14_Steel_Analysis và gp_margin_fc/ebit_fc (Python), tránh lệch nhau.
        ("Quặng sắt 62% Fe (USD/t)", *IRON_ORE_A, "Giá CFR Trung Quốc"),
        ("Than cốc luyện kim (USD/t)", *COKE_A, "FOB Úc"),
        ("Chi phí SX khác (USD/t)", *([OTHER_COST_USD]*8), "Cố định — gồm nhân công, nhiên liệu, điện, khấu hao, CPQL"),
        ("Doanh thu TC (tỷ)", 1500, 2100, 1800, 1950, 2200, *FIN_INCOME_FC, ""),
        ("Chi phí TC (tỷ)", 1800, 3500, 2100, 2500, 3200, *FIN_COST_FC, ""),
        ("Thu nhập khác (tỷ)", 0, 0, 0, 0, 0, 0, 0, 0, "Q1/2026: ~4,915 tỷ từ Phố Nối"),
        # Tỷ giá theo năm (2026-07, user phát hiện bug dùng 1 hằng số 25400 cho mọi năm) — xem chú
        # thích FX_RATE_A ở module-level. Dùng để sheet 03_Revenue_Model link công thức SỐNG thay vì
        # hardcode "25400" trong từng ô (đã sửa ở khối build sheet 03 bên dưới).
        ("Tỷ giá USD/VND (VND)", *FX_RATE_A, "2021-25: bình quân năm tham khảo Vietcombank/SBV; 2026E: fetch investing.com; 2027-28E: +2%/năm"),
    ]
    # Map rows in Assumptions
    R_IRON = 19; R_COKE = 20; R_CONV = 21; R_FX = 25
    # Assumption column letters: B=2021A, C=2022A, ... I=2028E

    for i, row in enumerate(assumptions, 2):
        ws2.cell(row=i, column=1, value=row[0]).font = bold_font
        ws2.cell(row=i, column=1).border = thin_border
        ws2.cell(row=i, column=1).fill = assump_fill
        for j, v in enumerate(row[1:-1], 2):
            cell = ws2.cell(row=i, column=j)
            if isinstance(v, int):
                cell.value = v
                cell.number_format = '#,##0' if v >= 1000 else '0'
            elif isinstance(v, float):
                cell.value = v
                if abs(v) < 5:
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '#,##0'
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            cell.fill = assump_fill
        ws2.cell(row=i, column=10, value=row[-1]).font = data_font
        ws2.cell(row=i, column=10).border = thin_border
        ws2.cell(row=i, column=10).fill = assump_fill

    # Doanh thu (row4), Tăng trưởng DT (row5), D&A (row9), VCSH (row13) dự phóng — CÔNG THỨC SỐNG
    # (không còn số Python tính sẵn) để bấm vào ô là thấy ngay cách tính, kiểm soát/điều chỉnh được.
    for idx, col in enumerate(['G', 'H', 'I']):
        prev_col = 'F' if idx == 0 else ['G', 'H'][idx - 1]
        # Doanh thu = tổng bottom-up từ 03_Revenue_Model (Sản lượng HRC/XD x Giá x FX + Doanh thu khác)
        ws2.cell(row=4, column=7 + idx, value=f"={S_R3}!{col}2").number_format = '#,##0'
        # Tăng trưởng DT = Doanh thu(t)/Doanh thu(t-1) - 1
        ws2.cell(row=5, column=7 + idx, value=f"=({col}4/{prev_col}4-1)*100").number_format = '0.0'
        # D&A = D&A(t-1) x (1 + Tăng trưởng DT(t)) — tăng theo quy mô doanh thu/tài sản
        ws2.cell(row=9, column=7 + idx, value=f"={prev_col}9*(1+{col}5/100)").number_format = '#,##0'
        # VCSH = VCSH(t-1) + LNST(t) [04_PnL!row13] - Cổ tức tiền mặt(t) x Số CP(t) / 1000
        ws2.cell(row=13, column=7 + idx, value=f"={prev_col}13+'04_PnL'!{col}13-{col}14*{col}3/1000").number_format = '#,##0'

    # ─── Bước 1+2: Chuỗi quý đầy đủ (2018-2026) với P/E điều chỉnh ───
    # Nếu LNST lỗ → P/E quý đó lấy theo quý gần nhất (tránh nhiễu)
    raw_by_q = {}
    for r in HPG_RATIOS:
        if r.get("quarter") != 5:
            raw_by_q[(int(r["year"]), r["quarter"])] = r
    all_quarters = [(y, q) for y in range(2018, 2027) for q in range(1, 5)]
    last_good_pe = None
    for y, q in all_quarters:
        rec = raw_by_q.get((y, q), {})
        pe0 = rec.get("pe")
        if pe0 and 0 < pe0 < 50:
            last_good_pe = round(pe0, 1)
    # yr_ranges: dùng cho MEDIAN formula (dòng nào → năm nào)
    yr_ranges = {}
    yr_start = None; cur_yr = None
    i = 2
    for y, q in all_quarters:
        if y != cur_yr:
            if cur_yr is not None: yr_ranges[cur_yr] = (yr_start, i-1)
            yr_start = i; cur_yr = y
        i += 1
    if cur_yr is not None: yr_ranges[cur_yr] = (yr_start, i-1)

    HIST_START = len(assumptions) + 3
    r = HIST_START
    hist_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    hist_years = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
    # Map: label -> (col_letter_in_08, number_format)
    HIST_MAP = [("P/E (x)", "C", '0.0'), ("P/B (x)", "D", '0.00'), ("EV/EBITDA (x)", "E", '0.0')]

    ws2.cell(row=r, column=1, value="Bước 2: P/E, P/B, EV/EBITDA — Median từng năm (tham chiếu quý ở Sheet 08_Hist_Multiples)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 1
    hist_headers = ["Chỉ tiêu"] + [str(y) for y in hist_years] + ["Hiện tại (công thức)", "Trung vị LS"]
    hist_widths = [35] + [10]*8 + [14, 14]
    header_row(ws2, r, hist_headers, hist_widths)
    HIST_HDR = r
    r += 1

    median_cells = {}
    for label, col_let, nf in HIST_MAP:
        ws2.cell(row=r, column=1, value=label).font = bold_font
        ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=1).fill = hist_fill
        # Step 2: mỗi năm = MEDIAN(08_Hist_Multiples!col{r_start}:col{r_end})
        for j, y in enumerate(hist_years, 2):
            cell = ws2.cell(row=r, column=j)
            if y in yr_ranges:
                s, e = yr_ranges[y]
                cell.value = f"=MEDIAN({S_HIST}!{col_let}{s}:{col_let}{e})"
                cell.number_format = nf
            else:
                cell.value = "-"
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        # Step 3: Hiện tại = công thức (Market Cap / Chỉ số tài chính)
        cur_cell = ws2.cell(row=r, column=10)
        if label.startswith("P/E"):
            cur_cell.value = f"=({S_ASSUMP}!$B$2*{S_ASSUMP}!$B$3*1000000)/({S_PNL}!F13*1000000000)"
        elif label.startswith("P/B"):
            cur_cell.value = f"=({S_ASSUMP}!$B$2*{S_ASSUMP}!$B$3*1000000)/({S_ASSUMP}!F13*1000000000)"
        elif label.startswith("EV"):
            cur_cell.value = f"=({S_ASSUMP}!$B$2*{S_ASSUMP}!$B$3*1000000+({S_ASSUMP}!F11-{S_ASSUMP}!F12)*1000000000)/({S_PNL}!F16*1000000000)"
        cur_cell.number_format = nf
        cur_cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
        cur_cell.border = thin_border
        cur_cell.alignment = Alignment(horizontal='center')
        cur_cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        # Step 4: Trung vị LS = MEDIAN của 8 năm
        med_cell = ws2.cell(row=r, column=11)
        med_cell.value = f"=MEDIAN(B{r}:I{r})"
        med_cell.number_format = nf
        med_cell.font = bold_font
        med_cell.border = thin_border
        med_cell.alignment = Alignment(horizontal='center')
        med_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        key = label.split()[0]
        median_cells[key] = f"K{r}"
        r += 1

    # Update target multiple rows (16-18): năm nào → median năm đó từ 08_Hist_Multiples
    pb_row_n = 16; ev_row_n = 17; pe_row_n = 18
    pb_k = median_cells["P/B"]; ev_k = median_cells["EV/EBITDA"]; pe_k = median_cells["P/E"]
    for col in range(2, 10):
        ws2.cell(row=pb_row_n, column=col).number_format = '0.00'
        ws2.cell(row=ev_row_n, column=col).number_format = '0.0'
        ws2.cell(row=pe_row_n, column=col).number_format = '0.0'
        if col >= 7:
            ws2.cell(row=pb_row_n, column=col).value = f"={pb_k}"
            ws2.cell(row=ev_row_n, column=col).value = f"={ev_k}"
            ws2.cell(row=pe_row_n, column=col).value = f"={pe_k}"
        else:
            cl = get_column_letter(col)
            ws2.cell(row=pb_row_n, column=col).value = f"={cl}{pb_row_n+7}"
            ws2.cell(row=ev_row_n, column=col).value = f"={cl}{ev_row_n+7}"
            ws2.cell(row=pe_row_n, column=col).value = f"={cl}{pe_row_n+4}"
    ws2.cell(row=pb_row_n, column=11).value = pb_k; ws2.cell(row=pb_row_n, column=11).number_format = '0.00'
    ws2.cell(row=ev_row_n, column=11).value = ev_k; ws2.cell(row=ev_row_n, column=11).number_format = '0.0'
    ws2.cell(row=pe_row_n, column=11).value = pe_k; ws2.cell(row=pe_row_n, column=11).number_format = '0.0'

    # ─── Valuation calculation rows in Assumptions ───
    r += 1
    V_START = r
    ws2.cell(row=r, column=1, value="GIÁ TRỊ ĐỊNH GIÁ THEO TỪNG PHƯƠNG PHÁP").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
    r += 1
    val_headers = ["Phương pháp", "", "", "", "", "", "2026E", "", "", "Giá trị CP", "Ghi chú"]
    header_row(ws2, r, val_headers, [35] + [10]*7 + [14, 40])
    r += 1
    V_EV = r; V_PB = r+1; V_PE = r+2
    ws2.cell(row=V_EV, column=1, value="EV/EBITDA (2026E)").font = bold_font
    ws2.cell(row=V_EV, column=1).border = thin_border
    ev_f = f"=(({S_PNL}!G16*{ev_k}-({S_ASSUMP}!G11-{S_ASSUMP}!G12))*1000000000)/({S_ASSUMP}!G3*1000000)"
    ws2.cell(row=V_EV, column=10, value=ev_f).number_format = '#,##0'
    ws2.cell(row=V_EV, column=10).border = thin_border
    ws2.cell(row=V_EV, column=10).alignment = Alignment(horizontal='center')
    ws2.cell(row=V_EV, column=11, value=f"EV/EBITDA target = {ev_k}").font = data_font
    ws2.cell(row=V_EV, column=11).border = thin_border
    ws2.cell(row=V_PB, column=1, value="P/B (2026E)").font = bold_font
    ws2.cell(row=V_PB, column=1).border = thin_border
    pb_f = f"={pb_k}*({S_ASSUMP}!G13*1000000000)/({S_ASSUMP}!G3*1000000)"
    ws2.cell(row=V_PB, column=10, value=pb_f).number_format = '#,##0'
    ws2.cell(row=V_PB, column=10).border = thin_border
    ws2.cell(row=V_PB, column=10).alignment = Alignment(horizontal='center')
    ws2.cell(row=V_PB, column=11, value=f"P/B target = {pb_k}").font = data_font
    ws2.cell(row=V_PB, column=11).border = thin_border
    ws2.cell(row=V_PE, column=1, value="P/E (2026E)").font = bold_font
    ws2.cell(row=V_PE, column=1).border = thin_border
    pe_f = f"={pe_k}*{S_PNL}!G15"
    ws2.cell(row=V_PE, column=10, value=pe_f).number_format = '#,##0'
    ws2.cell(row=V_PE, column=10).border = thin_border
    ws2.cell(row=V_PE, column=10).alignment = Alignment(horizontal='center')
    ws2.cell(row=V_PE, column=11, value=f"P/E target = {pe_k}").font = data_font
    ws2.cell(row=V_PE, column=11).border = thin_border
    V_END = r

    # ─── Bước 5: Biểu đồ — đặt sau 08_Hist_Multiples ───

    # ─── Sheet 3: Revenue Model ───
    def col_ltr(j):
        """Column letter from 1-based index: B=2 → 'B'"""
        return chr(64 + j)

    def is_fc(j):
        """Columns G(7), H(8), I(9) are forecast"""
        return j >= 7

    R3_NAME = "'03_Revenue_Model'"
    # ── Lũy kế N quý ĐÃ CÓ BCTC của năm dự phóng đầu tiên (2026-07, theo yêu cầu user) ─────────────
    # Trước đây hardcode cứng "Q1/2026" — TỔNG QUÁT HÓA theo số quý N ĐÃ CÓ báo cáo thực tế (dùng lại
    # cumulative_actual_quarters() đã có sẵn, cùng hàm dùng cho blend DTTC/CPTC/sản lượng ở trên) để
    # khi rerun script vào các quý sau (N=2,3,4) tự động tính đúng lũy kế mà không cần sửa code.
    _cur_fc_year = years_fc[0]  # năm dự phóng đầu tiên, VD 2026
    _is_qs_ref = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    _cum_rev_nq, _n_q_known = cumulative_actual_quarters(_is_qs_ref, _cur_fc_year, "isa3")
    _cum_gp_nq, _ = cumulative_actual_quarters(_is_qs_ref, _cur_fc_year, "isa5")
    R_Q1_REV = 40; R_Q1_GP = 41; R_Q1_GPM = 42; R_Q1_SPR = 43
    for rn in (R_Q1_REV, R_Q1_GP, R_Q1_GPM, R_Q1_SPR):
        ws2.cell(row=rn, column=1).font = Font(name=FONT_NAME, bold=True, size=9, color="1F4E79")
        ws2.cell(row=rn, column=1).border = thin_border
    _nq_lbl = f"Lũy kế {_n_q_known} quý {_cur_fc_year}" if _n_q_known != 1 else f"Q1/{_cur_fc_year}"
    ws2.cell(row=R_Q1_REV, column=1, value=f"{_nq_lbl} Doanh thu (tỷ)")
    ws2.cell(row=R_Q1_REV, column=7, value=_cum_rev_nq).number_format = '#,##0'
    ws2.cell(row=R_Q1_REV, column=7).font = data_font
    ws2.cell(row=R_Q1_GP, column=1, value=f"{_nq_lbl} LN gộp (tỷ)")
    ws2.cell(row=R_Q1_GP, column=7, value=_cum_gp_nq).number_format = '#,##0'
    ws2.cell(row=R_Q1_GP, column=7).font = data_font
    ws2.cell(row=R_Q1_GPM, column=1, value=f"{_nq_lbl} Biên LNG (%)")
    ws2.cell(row=R_Q1_GPM, column=7, value=f"=G{R_Q1_GP}/G{R_Q1_REV}*100").number_format = '0.00'
    ws2.cell(row=R_Q1_GPM, column=7).font = data_font
    # R_Q1_SPR (Spread All quý gần nhất đã biết) và R_ANN_SPR/row6 (BLNG dự phóng) cần sheet
    # 17_Gia_Hang_Hoa/03_Revenue_Model đã build xong — ghi công thức thật ở khối patch cuối hàm
    # (sau khi ws17/ws3 tồn tại), xem gần chỗ `wb.save(EXCEL_FILE)`.
    R_ANN_SPR = 45

    ws3 = wb.create_sheet("03_Revenue_Model")
    headers3 = ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"]
    widths3 = [40] + [14]*8
    header_row(ws3, 1, headers3, widths3)

    all_years = years_hist + years_fc
    all_rev = revenue_hist + revenue_fc
    all_ni = ni_hist + ni_fc
    all_gpm = gp_margin_hist + gp_margin_fc

    rev_rows = [
        ("Doanh thu (tỷ)", all_rev, "Tăng trưởng mạnh 2026 nhờ DQ2 full công suất"),
        ("YoY Growth (%)", None, ""),
        ("Sản lượng HRC (triệu tấn)", SL_HRC_A, ""),
        ("Sản lượng thép XD (triệu tấn)", SL_XD_A, "FY2025 XD: 4.85M tấn (VietnamBiz). HPG total 2026 target ~15M tấn (CFO)"),
        ("Giá HRC bq (USD/tấn)", HRC_PRICE_A, "SHFE Q3/2026 ~580 + premium VN ~40. AD 27.83% hỗ trợ giá nội địa"),
        ("Giá thép XD bq (USD/tấn)", XD_PRICE_A, ""),
        ("Doanh thu HRC (tỷ)", None, ""),
        ("Doanh thu thép XD (tỷ)", None, ""),
        ("Doanh thu khác (tỷ)", None, f"Ống thép, tôn mạ, container, KCN, phôi billet. Dự phóng tăng {OTHER_REV_GROWTH*100:.0f}%/năm"),
    ]

    for i, (name, vals, note) in enumerate(rev_rows, 2):
        c = ws3.cell(row=i, column=1, value=name)
        c.font = bold_font
        c.border = thin_border

    # Doanh thu (tỷ) — row 2: hardcode historical, SUM(HRC+XD+Khác) forecast
    for j, v in enumerate(all_rev, 2):
        cell = ws3.cell(row=2, column=j)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            cell.value = v; cell.number_format = '#,##0'
        else:
            cl = col_ltr(j)
            cell.value = f"={cl}8+{cl}9+{cl}10"
            cell.number_format = '#,##0'

    # YoY Growth (row 3): formula for all years
    for j in range(2, 10):
        cell = ws3.cell(row=3, column=j)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = data_font
        if j == 2:
            cell.value = "-"
        else:
            cell.value = f"=({col_ltr(j)}2/{col_ltr(j-1)}2-1)*100"
            cell.number_format = '0.0'

    # Production & price rows (4-7): hardcoded all years, TRỪ SL HRC/XD (row 4-5) năm 2023-2025 —
    # 2026-07, user phát hiện bug: SL_HRC_A/SL_XD_A là giả định tĩnh cũ, lệch xa SỐ THẬT đã tổng hợp ở
    # sheet 15_Quarterly_Data (VD 2025 giả định 3.2 triệu tấn HRC nhưng lũy kế 4 quý thật = 5.0 triệu
    # tấn) — khiến doanh thu 2025 bị tính thiếu, kéo theo %growth 2026E bị thổi phồng giả tạo. Sửa: 3
    # năm có đủ 4 quý thật (2023/2024/2025) dùng CÔNG THỨC SUM sống link sang sheet 15 (nguồn duy nhất
    # — user yêu cầu tránh 2 nguồn số liệu độc lập dễ lệch nhau âm thầm như bug này); 2021/2022 (chưa
    # có dữ liệu tách quý) và 2026E-2028E (dự phóng, xem blend_annual_estimate ở Python) vẫn giữ giá trị.
    # Ghi giá trị Python trước (đã đúng số thật nhờ fix SL_HRC_A/SL_XD_A ở trên) — SL HRC/XD (row 4-5)
    # năm 2023-2025 sẽ được GHI ĐÈ bằng công thức SUM sống link sang sheet 15_Quarterly_Data ở khối
    # patch cuối hàm (sau khi ws15/R_QV_HRC/R_QV_XD tồn tại — ws3 build TRƯỚC ws15 trong thứ tự code).
    for i in [4, 5, 6, 7]:
        name, vals, _ = rev_rows[i-2]
        for j, v in enumerate(vals, 2):
            cell = ws3.cell(row=i, column=j, value=v)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '#,##0.0' if isinstance(v, float) else '#,##0'

    # Doanh thu HRC (row 8) = Sản lượng HRC (col j, row 4) * Giá HRC (col j, row 6) * tỷ giá
    # tỷ giá ước 25,400 → (tấn * USD/tấn * 25400)/1e9 = tỷ VND
    # Simplify: SL (triệu tấn) * Giá (USD/tấn) * 25400 / 1e6 / 1e9 ... 
    # Let's just use: =B4*B6*25400/1000000 (SL triệu tấn * USD/tấn * tỷ giá / 1e6 = tỷ VND)
    # Actually: SL(triệu tấn) * 1e6 * Giá(USD/tấn) * 25400(VND/USD) / 1e9(tỷ) = SL*Giá*25400/1000
    # E.g., 2021: 2.0 * 680 * 25400/1000 = 34,544 tỷ. That seems high. Let's check.
    # HPG total 2021 revenue = 149,680. HRC + XD would cover most of it.
    # Actually a simpler approach: link directly to PnL or use growth rates
    # For forecast years, HRC rev growth follows volume × price mix
    for j in range(2, 10):
        cl = col_ltr(j)
        for row in [8, 9]:
            cell = ws3.cell(row=row, column=j)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            cell.font = data_font
            cell.number_format = '#,##0'
        # HRC rev = SL_HRC(triệu tấn) * Giá_HRC * Tỷ giá NĂM ĐÓ / 1000 = tỷ VND — link sống
        # Assumptions!row{R_FX} (2026-07, sửa bug dùng chết "25400" cho mọi năm, xem FX_RATE_A).
        ws3.cell(row=8, column=j).value = f"={cl}4*{cl}6*{S_ASSUMP}!{cl}{R_FX}/1000"
        # XD rev = SL_XD(triệu tấn) * Giá_XD * Tỷ giá NĂM ĐÓ / 1000 = tỷ VND
        ws3.cell(row=9, column=j).value = f"={cl}5*{cl}7*{S_ASSUMP}!{cl}{R_FX}/1000"
        # Doanh thu khác (row 10): historical = residual, forecast = prev*(1+OTHER_REV_GROWTH)
        # Dùng hằng số cố định (không phải Assumptions!row5 "Tăng trưởng DT") vì row5 giờ là
        # KẾT QUẢ tính từ Doanh thu tổng (row2) — tránh vòng lặp tham chiếu circular.
        if not is_fc(j):
            ws3.cell(row=10, column=j).value = f"={cl}2-{cl}8-{cl}9"
        else:
            ws3.cell(row=10, column=j).value = f"={col_ltr(j-1)}10*(1+{OTHER_REV_GROWTH})"
        ws3.cell(row=10, column=j).border = thin_border
        ws3.cell(row=10, column=j).alignment = Alignment(horizontal='center')
        ws3.cell(row=10, column=j).font = data_font
        ws3.cell(row=10, column=j).number_format = '#,##0'

    # LNST (row 12) = PnL!col13
    for j in range(2, 10):
        cl = col_ltr(j)
        cell = ws3.cell(row=12, column=j)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = bold_font
        cell.value = f"={S_PNL}!{cl}13"; cell.number_format = '#,##0'

    # Biên LNG (row 13) = reference PnL margin
    for j in range(2, 10):
        cl = col_ltr(j)
        cell = ws3.cell(row=13, column=j)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = data_font
        cell.value = f"={S_PNL}!{cl}5"; cell.number_format = '0.0'

    # EPS (row 15) = PnL!col15
    for j in range(2, 10):
        cl = col_ltr(j)
        cell = ws3.cell(row=15, column=j)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.font = data_font
        cell.value = f"={S_PNL}!{cl}15"; cell.number_format = '#,##0'

    # ── PROFIT BRIDGE (rows 20-28) — từ Spread → EBIT ước tính ──
    pb_title_font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")
    pb_data_font = Font(name=FONT_NAME, size=9)
    pb_section = 20
    ws3.cell(row=pb_section, column=1, value="B. PHÂN TÍCH LỢI NHUẬN TỪ SPREAD (công thức nội bộ)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws3.merge_cells(start_row=pb_section, start_column=1, end_row=pb_section, end_column=9)
    ws3.cell(row=pb_section, column=1).border = thin_border
    R_PB_COST = 21; R_PB_SP_HRC = 22; R_PB_SP_XD = 23; R_PB_GP_STEEL = 24; R_PB_GP_OTHER = 25
    R_PB_GP = 26; R_PB_GPM = 27; R_PB_SGKA = 28; R_PB_EBIT = 29
    for j in range(2, 10):
        cl = col_ltr(j)
        # Tổng CF SX/tấn phải dùng giá quặng/than/CP khác của ĐÚNG NĂM đó (cột hiện tại), không
        # phải cố định $B$19 (2021A) cho mọi năm như trước — bug khiến chi phí SX không đổi qua các năm.
        base_cost = f"={S_ASSUMP}!{cl}19*1.6+{S_ASSUMP}!{cl}20*0.6+{S_ASSUMP}!{cl}21"
        c21 = ws3.cell(row=R_PB_COST, column=j); c21.value = base_cost; c21.number_format = '#,##0'
        c21.font = pb_data_font; c21.border = thin_border; c21.alignment = Alignment(horizontal='center')
        c22 = ws3.cell(row=R_PB_SP_HRC, column=j); c22.value = f"={cl}6-{cl}{R_PB_COST}"; c22.number_format = '#,##0'
        c22.font = pb_data_font; c22.border = thin_border; c22.alignment = Alignment(horizontal='center')
        c23 = ws3.cell(row=R_PB_SP_XD, column=j); c23.value = f"={cl}7-{cl}{R_PB_COST}"; c23.number_format = '#,##0'
        c23.font = pb_data_font; c23.border = thin_border; c23.alignment = Alignment(horizontal='center')
        c24 = ws3.cell(row=R_PB_GP_STEEL, column=j); c24.value = f"=({cl}4*{cl}{R_PB_SP_HRC}+{cl}5*{cl}{R_PB_SP_XD})*{S_ASSUMP}!{cl}{R_FX}/1000"
        c24.number_format = '#,##0'; c24.font = pb_data_font; c24.border = thin_border; c24.alignment = Alignment(horizontal='center')
        c25 = ws3.cell(row=R_PB_GP_OTHER, column=j); c25.value = f"={cl}10*0.15"
        c25.number_format = '#,##0'; c25.font = pb_data_font; c25.border = thin_border; c25.alignment = Alignment(horizontal='center')
        c26 = ws3.cell(row=R_PB_GP, column=j); c26.value = f"={cl}{R_PB_GP_STEEL}+{cl}{R_PB_GP_OTHER}"
        c26.number_format = '#,##0'; c26.font = Font(name=FONT_NAME, bold=True, size=9); c26.border = thin_border; c26.alignment = Alignment(horizontal='center')
        c27 = ws3.cell(row=R_PB_GPM, column=j); c27.value = f"={cl}{R_PB_GP}/{cl}2*100"
        c27.number_format = '0.0'; c27.font = pb_data_font; c27.border = thin_border; c27.alignment = Alignment(horizontal='center')
        c28 = ws3.cell(row=R_PB_SGKA, column=j); c28.value = f"={cl}2*{S_ASSUMP}!{cl}15/100"
        c28.number_format = '#,##0'; c28.font = pb_data_font; c28.border = thin_border; c28.alignment = Alignment(horizontal='center')
        c29 = ws3.cell(row=R_PB_EBIT, column=j); 
        if not is_fc(j):
            c29.value = f"={S_PNL}!{cl}7"
        else:
            c29.value = f"={cl}{R_PB_GP}-{cl}{R_PB_SGKA}"
        c29.number_format = '#,##0'; c29.font = Font(name=FONT_NAME, bold=True, color="006600", size=9)
        c29.border = thin_border; c29.alignment = Alignment(horizontal='center')
    # Label column
    labels_pb = [
        ("Tổng CF SX/tấn (USD)", R_PB_COST, "Quặng*1.6 + Than*0.6 + CP SX khác (200, cố định)"),
        ("  Spread HRC (USD/t)", R_PB_SP_HRC, "Giá HRC bq - CF/tấn"),
        ("  Spread XD (USD/t)", R_PB_SP_XD, "Giá XD bq - CF/tấn"),
        ("  LN gộp thép (tỷ)", R_PB_GP_STEEL, "(SL_HRC*Spr_HRC + SL_XD*Spr_XD)*25400/1000"),
        ("  LN gộp KD khác (tỷ)", R_PB_GP_OTHER, "Giả định biên 15% cho ống/tôn/KCN"),
        ("Tổng LN gộp ước tính (tỷ)", R_PB_GP, "Đối chiếu PnL!GP"),
        ("Biên LNG ước tính (%)", R_PB_GPM, ""),
        ("CP QLDN (tỷ)", R_PB_SGKA, "= DT * SGKA%"),
        ("EBIT ước tính (tỷ)", R_PB_EBIT, "Forecast: LN gộp - CP QL"),
    ]
    for label, row_n, note in labels_pb:
        c = ws3.cell(row=row_n, column=1, value=label)
        c.font = pb_title_font if not label.startswith("  ") else pb_data_font
        c.border = thin_border
        ws3.cell(row=row_n, column=10, value=note).font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
        ws3.cell(row=row_n, column=10).border = thin_border

    # ── Sheet 4: PnL (FORMULA-BASED) ──
    ws4 = wb.create_sheet("04_PnL")
    header_row(ws4, 1, headers3, widths3)

    # Row positions in PnL sheet
    R_REV  = 2; R_COGS = 3; R_GP  = 4; R_GPM  = 5
    R_SGKA = 6; R_EBIT = 7; R_EBM = 8; R_FINI = 9
    R_FINC =10; R_EBT  =11; R_TAX =12; R_NI   =13
    R_NIM  =14; R_EPS  =15; R_EBITDA=16
    # Assumptions column mappings: col5=Growth, col6=GPM, col8=Tax, col9=D&A, col10=CAPEX, col15=SG&A%
    # Assumptions row positions (after restructuring below)
    # 2=Price, 3=Shares, 4=Rev, 5=Growth, 6=GPM, 7=EBM, 8=TaxRate,
    # 9=D&A, 10=CAPEX, 11=Debt, 12=Cash, 13=EQ, 14=Div, 15=SGKA%

    pnl_labels = [
        "Doanh thu thuần", "Giá vốn hàng bán", "Lợi nhuận gộp", "Biên LNG (%)",
        "Chi phí BH & QLDN", "EBIT", "Biên EBIT (%)", "Doanh thu tài chính",
        "Chi phí tài chính", "EBT", "Thuế TNDN", "LNST", "Biên LNST (%)",
        "EPS (VND)", "EBITDA"
    ]
    for i, lbl in enumerate(pnl_labels, 2):
        ws4.cell(row=i, column=1, value=lbl).font = bold_font
        ws4.cell(row=i, column=1).border = thin_border

    # ── Revenue (row 2): link to 03_Revenue_Model driver-based total ──
    for j in range(2, 10):
        c = ws4.cell(row=R_REV, column=j)
        c.font = data_font
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            c.value = all_rev[j-2]; c.number_format = '#,##0'
        else:
            c.value = f"={S_R3}!{col_ltr(j)}2"
            c.font = Font(name=FONT_NAME, color="006600", size=10)

    # ── COGS (row 3): = Rev * (1 - GPM%) ──
    for j in range(2, 10):
        c = ws4.cell(row=R_COGS, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            rev, gpm = all_rev[j-2], all_gpm[j-2]
            c.value = round(rev * (1 - gpm/100)); c.number_format = '#,##0'
        else:
            c.value = f"={col_ltr(j)}{R_REV}*(1-{S_ASSUMP}!{col_ltr(j)}6/100)"
            c.number_format = '#,##0'

    # ── GP (row 4): = Rev - COGS ──
    for j in range(2, 10):
        c = ws4.cell(row=R_GP, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            rev, gpm = all_rev[j-2], all_gpm[j-2]
            c.value = rev - round(rev * (1 - gpm/100)); c.number_format = '#,##0'
        else:
            c.value = f"={col_ltr(j)}{R_REV}-{col_ltr(j)}{R_COGS}"
            c.number_format = '#,##0'

    # ── GP Margin % (row 5): = GP / Rev * 100 ──
    for j in range(2, 10):
        c = ws4.cell(row=R_GPM, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            c.value = all_gpm[j-2]; c.number_format = '0.0'
        else:
            c.value = f"={col_ltr(j)}{R_GP}/{col_ltr(j)}{R_REV}*100"
            c.number_format = '0.0'

    # ── SG&A (row 6): = Rev * SGKA_rate ──
    for j in range(2, 10):
        c = ws4.cell(row=R_SGKA, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            c.value = round(all_rev[j-2] * 0.038); c.number_format = '#,##0'
        else:
            c.value = f"={col_ltr(j)}{R_REV}*{S_ASSUMP}!{col_ltr(j)}15/100"
            c.number_format = '#,##0'

    # ── EBIT (row 7): = GP - SG&A ──
    for j in range(2, 10):
        c = ws4.cell(row=R_EBIT, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            rev, gpm = all_rev[j-2], all_gpm[j-2]
            gp = rev - round(rev*(1-gpm/100))
            sgka = round(rev * 0.038)
            c.value = gp - sgka; c.number_format = '#,##0'
        else:
            c.value = f"={col_ltr(j)}{R_GP}-{col_ltr(j)}{R_SGKA}"
            c.number_format = '#,##0'

    # ── EBIT Margin % (row 8): = EBIT / Rev * 100 ──
    for j in range(2, 10):
        c = ws4.cell(row=R_EBM, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            rev, gpm = all_rev[j-2], all_gpm[j-2]
            gp = rev - round(rev*(1-gpm/100))
            ebit = gp - round(rev*0.038)
            c.value = round(ebit/rev*100, 1) if rev else 0; c.number_format = '0.0'
        else:
            c.value = f"={col_ltr(j)}{R_EBIT}/{col_ltr(j)}{R_REV}*100"
            c.number_format = '0.0'

    # ── Financial income (row 9) / cost (row 10) — hardcoded ──
    fin_i_data = [1500, 2100, 1800, 1950, 2200] + FIN_INCOME_FC
    fin_c_data = [1800, 3500, 2100, 2500, 3200] + FIN_COST_FC
    for j in range(2, 10):
        for row, data in ((R_FINI, fin_i_data), (R_FINC, fin_c_data)):
            c = ws4.cell(row=row, column=j, value=data[j-2])
            c.border = thin_border; c.alignment = Alignment(horizontal='center')
            c.number_format = '#,##0'

    # ── EBT (row 11): = EBIT + Fin_I - Fin_C ──
    for j in range(2, 10):
        c = ws4.cell(row=R_EBT, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        c.value = f"={col_ltr(j)}{R_EBIT}+{col_ltr(j)}{R_FINI}-{col_ltr(j)}{R_FINC}"
        c.number_format = '#,##0'; c.font = Font(name=FONT_NAME, color="006600", size=10)

    # ── Tax (row 12): = EBT * TaxRate ──
    for j in range(2, 10):
        c = ws4.cell(row=R_TAX, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        if not is_fc(j):
            # historical: compute using hardcoded rate approach
            idx = j-2
            rev, gpm = all_rev[idx], all_gpm[idx]
            gp = rev - round(rev*(1-gpm/100))
            ebt = gp - round(rev*0.038) + fin_i_data[idx] - fin_c_data[idx]
            tr = 0.125 if idx < 7 else 0.12
            c.value = round(ebt * tr); c.number_format = '#,##0'
        else:
            c.value = f"={col_ltr(j)}{R_EBT}*{S_ASSUMP}!{col_ltr(j)}8/100"
            c.number_format = '#,##0'

    # ── NI (row 13): = EBT - Tax ──
    for j in range(2, 10):
        c = ws4.cell(row=R_NI, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        c.value = f"={col_ltr(j)}{R_EBT}-{col_ltr(j)}{R_TAX}"
        c.number_format = '#,##0'

    # ── NI Margin % (row 14): = NI / Rev * 100 ──
    for j in range(2, 10):
        c = ws4.cell(row=R_NIM, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        c.value = f"={col_ltr(j)}{R_NI}/{col_ltr(j)}{R_REV}*100"
        c.number_format = '0.0'

    # ── EPS (row 15): = NI*1e9/(Shares*1e6) ──
    for j in range(2, 10):
        c = ws4.cell(row=R_EPS, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        c.value = f"={col_ltr(j)}{R_NI}*1000000000/({S_ASSUMP}!{col_ltr(j)}3*1000000)"
        c.number_format = '#,##0'

    # ── EBITDA (row 16): = EBIT + D&A — CÔNG THỨC SỐNG cho MỌI năm (2026-07, user phát hiện: lịch sử
    # trước đây là số Python tính sẵn ghi cứng vào ô, không bấm vào kiểm chứng được — dù bản thân số
    # liệu ĐÃ đúng vì da_hist là D&A THẬT từ Vietcap. Đổi sang link EBIT(row7, cùng cột) +
    # Assumptions!D&A(row9, cùng cột) — D&A lịch sử ở Assumptions giờ cũng là số THẬT (xem sửa
    # da_hist ở 02_Assumptions), nên 2 nơi khớp nhau tuyệt đối, không lệch âm thầm.
    for j in range(2, 10):
        c = ws4.cell(row=R_EBITDA, column=j)
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        c.value = f"={col_ltr(j)}{R_EBIT}+{S_ASSUMP}!{col_ltr(j)}9"
        c.number_format = '#,##0'

    # ─── Sheet 5: Balance Sheet ───
    ws5 = wb.create_sheet("05_Balance_Sheet")
    header_row(ws5, 1, headers3, widths3)
    # LN chưa PP dự phóng = VCSH roll-forward (equity_fc_val) - Vốn điều lệ - Thặng dư - Quỹ & khác,
    # để "Tổng VCSH" luôn khớp tổng các cấu phần (thay vì 4 dòng độc lập gõ tay như trước).
    von_dl_fc  = [25000, 28000, 31000]
    thang_du_fc = [20000, 22000, 24000]
    quy_khac_fc = [25000, 28000, 32000]
    ln_chua_pp_fc = [equity_fc_val[i] - von_dl_fc[i] - thang_du_fc[i] - quy_khac_fc[i] for i in range(3)]
    bs_items = [
        ("Tài sản ngắn hạn", None),
        ("  Tiền & tương đương", cash_hist + [25000, 30000, 35000]),
        ("  Đầu tư tài chính NH", [3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000]),
        ("  Phải thu NH", receivables_hist + RECEIVABLES_FC),
        ("  Hàng tồn kho", inventory_hist + INVENTORY_FC),
        ("  TSNH khác", [2000, 2500, 3000, 3500, 4000, 4500, 5000, 5500]),
        ("Tài sản dài hạn", None),
        ("  TSCĐ hữu hình",
         [round(v) for v in fixed_assets_hist] + [150000, 160000, 170000]),
        ("  XDCB DD (CIP)",
         [round(v) for v in cip_hist] + [15000, 10000, 8000]),
        ("  TSCĐ vô hình", [5000, 6000, 7000, 8000, 9000, 10000, 11000, 12000]),
        ("  Đầu tư TC dài hạn", [10000, 12000, 13000, 14000, 15000, 16000, 17000, 18000]),
        ("  TSDH khác", [2000, 2500, 3000, 3500, 4000, 4500, 5000, 5500]),
        ("Tổng tài sản", None),
        ("Nợ ngắn hạn", None),
        ("  Phải trả NH",
         [round(v) for v in payables_hist] + [20000, 22000, 24000]),
        ("  Vay ngắn hạn",
         [round(v) for v in short_debt_hist] + [60000, 55000, 50000]),
        ("  Nợ NH khác", [5000, 6000, 7000, 8000, 9000, 10000, 11000, 12000]),
        ("Nợ dài hạn", None),
        ("  Vay dài hạn",
         [round(v) for v in long_debt_hist] + [30000, 28000, 26000]),
        ("  Nợ DH khác", [3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000]),
        ("Tổng nợ", None),
        ("VCSH", None),
        ("  Vốn điều lệ",
         [12000, 15000, 18000, 20000, 22000] + von_dl_fc),
        ("  Thặng dư", [8000, 10000, 12000, 15000, 18000] + thang_du_fc),
        ("  LN chưa PP",
         [40000, 48000, 55000, 60000, 68000] + ln_chua_pp_fc),
        ("  Quỹ & khác", [15000, 17000, 20000, 20000, 22000] + quy_khac_fc),
        ("Tổng VCSH", equity_hist + equity_fc_val),
        ("Tổng nợ + VCSH", None),
    ]
    # ── ACTUAL ROW MAP (1-based) ──
    # 2:  Tài sản ngắn hạn         15: Nợ ngắn hạn
    # 3:    Tiền                   16:   Phải trả NH
    # 4:    Đầu tư TC NH            17:   Vay NH
    # 5:    Phải thu NH             18:   Nợ NH khác
    # 6:    Hàng tồn kho            19: Nợ dài hạn
    # 7:    TSNH khác               20:   Vay DH
    # 8:  Tài sản dài hạn           21:   Nợ DH khác
    # 9:    TSCĐ HH                 22: Tổng nợ
    # 10:   XDCB                   23: VCSH
    # 11:   TSCĐ vô hình            24:   Vốn điều lệ
    # 12:   Đầu tư TC DH            25:   Thặng dư
    # 13:   TSDH khác               26:   LN chưa PP
    # 14: Tổng tài sản              27:   Quỹ & khác
    #                              28: Tổng VCSH
    #                              29: Tổng nợ + VCSH

    # Sum total assets
    for i, (name, vals) in enumerate(bs_items, 2):
        ws5.cell(row=i, column=1, value=name).font = bold_font if not name.startswith("  ") else data_font
        ws5.cell(row=i, column=1).border = thin_border
        if vals:
            for j, v in enumerate(vals, 2):
                cell = ws5.cell(row=i, column=j, value=v)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '#,##0'

    # Calculate totals using SUM formulas
    # BS row mapping:
    # 2-6: CA (Cash=3, Invest=4, Receiv=5, Inventory=6, OtherCA=7)
    # 8-12: FA (PPE=9, CIP=10, Intang=11, InvLt=12, OtherFA=13)
    # 13: Total Assets = SUM(C line items)
    # 15-17: CL (Payable=16, ST debt=17, OtherCL=18)
    # 19-20: LLT (LT debt=20, OtherLL=21)
    # 21: Total Liab = SUM(L line items)
    # 23-26: Equity (Capital=24, Premium=25, RE=26, Reserves=27)
    # 27: Total EQ
    # 28: Total Liab+EQ = Total Liab + Total EQ

    CA_ROWS = [3,4,5,6,7]    # Cash, ST inv, Receiv, Inventory, Other CA
    FA_ROWS = [9,10,11,12,13] # PPE, CIP, Intang, LT inv, Other FA
    CL_ROWS = [16,17,18]      # Payable, ST debt, Other CL
    LL_ROWS = [20,21]         # LT debt, Other LL
    EQ_ROWS = [24,25,26,27]   # Capital, Premium, RE, Reserves

    def sum_range(ws, rows, col_l):
        """Generate SUM formula string for a list of rows at given column letter"""
        parts = [f"{col_l}{r}" for r in rows]
        return "=" + "+".join(parts)

    for j in range(2, 10):
        cl = col_ltr(j)
        # Total Assets (row 14)
        ws5.cell(row=14, column=j).value = sum_range(ws5, CA_ROWS+FA_ROWS, cl)
        ws5.cell(row=14, column=j).font = bold_font
        ws5.cell(row=14, column=j).border = thin_border
        ws5.cell(row=14, column=j).alignment = Alignment(horizontal='center')
        ws5.cell(row=14, column=j).number_format = '#,##0'
        ws5.cell(row=14, column=1, value="Tổng tài sản").font = bold_font
        ws5.cell(row=14, column=1).border = thin_border

        # Total Liabilities (row 22)
        ws5.cell(row=22, column=j).value = sum_range(ws5, CL_ROWS+LL_ROWS, cl)
        ws5.cell(row=22, column=j).font = bold_font
        ws5.cell(row=22, column=j).border = thin_border
        ws5.cell(row=22, column=j).alignment = Alignment(horizontal='center')
        ws5.cell(row=22, column=j).number_format = '#,##0'
        ws5.cell(row=22, column=1, value="Tổng nợ").font = bold_font
        ws5.cell(row=22, column=1).border = thin_border

        # Total Liabilities + Equity (row 29): = Total Liab (row 22) + Total EQ (row 28)
        ws5.cell(row=29, column=j).value = f"={cl}22+{cl}28"
        ws5.cell(row=29, column=j).font = bold_font
        ws5.cell(row=29, column=j).border = thin_border
        ws5.cell(row=29, column=j).alignment = Alignment(horizontal='center')
        ws5.cell(row=29, column=j).number_format = '#,##0'
        ws5.cell(row=29, column=1, value="Tổng nợ + VCSH").font = bold_font
        ws5.cell(row=29, column=1).border = thin_border

    # ─── Sheet 6: Cash Flow (FORMULA-BASED) ───
    ws6 = wb.create_sheet("06_Cash_Flow")
    header_row(ws6, 1, headers3, widths3)

    # CF row positions
    R_CFO = 2; R_CFO_NI = 3; R_CFO_DA = 4; R_CFO_WC = 5
    R_CFI = 6; R_CFI_CAPEX = 7
    R_CFF = 8
    R_FCFF = 9
    R_BEG  = 10; R_END = 11

    cf_labels = [
        "Dòng tiền từ HĐKD (CFO)", "  LNST", "  Khấu hao",
        "  Thay đổi VLĐ (ước)", "Dòng tiền đầu tư (CFI)",
        "  CAPEX", "Dòng tiền tài trợ (CFF)", "FCFF (CFO + CFI)",
        "Tiền đầu kỳ", "Tiền cuối kỳ"
    ]
    for i, lbl in enumerate(cf_labels, 2):
        ws6.cell(row=i, column=1, value=lbl).font = bold_font if not lbl.startswith("  ") else data_font
        ws6.cell(row=i, column=1).border = thin_border

    # Use module-level cfo_hist, cfi_hist from API
    cff_hist = [-5000, -2000, -3000, -4000, -5000]

    for j in range(2, 10):
        idx = j - 2
        cl = col_ltr(j)

        if not is_fc(j):
            # Historical: hardcoded values
            ws6.cell(row=R_CFO, column=j, value=cfo_hist[idx]).number_format = '#,##0'
            ws6.cell(row=R_CFO_NI, column=j, value=all_ni[idx]).number_format = '#,##0'
            ws6.cell(row=R_CFO_DA, column=j, value=da_hist[idx]).number_format = '#,##0'
            wc_est = cfo_hist[idx] - all_ni[idx] - da_hist[idx]
            ws6.cell(row=R_CFO_WC, column=j, value=wc_est).number_format = '#,##0'
            ws6.cell(row=R_CFI, column=j, value=cfi_hist[idx]).number_format = '#,##0'
            ws6.cell(row=R_CFI_CAPEX, column=j, value=capex_hist[idx]).number_format = '#,##0'
            ws6.cell(row=R_CFF, column=j, value=cff_hist[idx]).number_format = '#,##0'
            ws6.cell(row=R_FCFF, column=j, value=cfo_hist[idx]+cfi_hist[idx]).number_format = '#,##0'
        else:
            # Forecast: formulas referencing PnL and Assumptions
            # CFO = NI + D&A + WC_change
            ws6.cell(row=R_CFO, column=j,
                     value=f"={cl}{R_CFO_NI}+{cl}{R_CFO_DA}+{cl}{R_CFO_WC}").number_format = '#,##0'
            ws6.cell(row=R_CFO_NI, column=j,
                     value=f"={S_PNL}!{cl}13").number_format = '#,##0'
            ws6.cell(row=R_CFO_DA, column=j,
                     value=f"={S_ASSUMP}!{cl}9").number_format = '#,##0'
            # WC change: hardcoded estimate for simplicity
            ws6.cell(row=R_CFO_WC, column=j,
                     value=0).number_format = '#,##0'
            # CFI = -CAPEX
            ws6.cell(row=R_CFI, column=j,
                     value=f"=-{cl}{R_CFI_CAPEX}").number_format = '#,##0'
            ws6.cell(row=R_CFI_CAPEX, column=j,
                     value=f"={S_ASSUMP}!{cl}10").number_format = '#,##0'
            # CFF: hardcoded estimate
            cff_fc_vals = [-5000, -6000, -7000]
            ws6.cell(row=R_CFF, column=j, value=cff_fc_vals[idx-5]).number_format = '#,##0'
            # FCFF = CFO + CFI
            ws6.cell(row=R_FCFF, column=j,
                     value=f"={cl}{R_CFO}+{cl}{R_CFI}").number_format = '#,##0'

        # Borders
        for r in range(R_CFO, R_FCFF+1):
            ws6.cell(row=r, column=j).border = thin_border
            ws6.cell(row=r, column=j).alignment = Alignment(horizontal='center')

    # Tiền đầu kỳ (row 10): B from API begin_cash_hist, C+=formula linking previous end
    ws6.cell(row=R_BEG, column=2, value=round(begin_cash_hist[0])).number_format = '#,##0'
    ws6.cell(row=R_BEG, column=2).border = thin_border
    ws6.cell(row=R_BEG, column=2).alignment = Alignment(horizontal='center')
    for j in range(3, 10):
        prev_cl = col_ltr(j-1)
        cl = col_ltr(j)
        ws6.cell(row=R_BEG, column=j,
                 value=f"={prev_cl}{R_END}").number_format = '#,##0'
        ws6.cell(row=R_BEG, column=j).border = thin_border
        ws6.cell(row=R_BEG, column=j).alignment = Alignment(horizontal='center')

    # Tiền cuối kỳ (row 11): = Beginning + CFO + CFI + CFF
    for j in range(2, 10):
        cl = col_ltr(j)
        ws6.cell(row=R_END, column=j,
                 value=f"={cl}{R_BEG}+{cl}{R_CFO}+{cl}{R_CFI}+{cl}{R_CFF}").number_format = '#,##0'
        ws6.cell(row=R_END, column=j).border = thin_border
        ws6.cell(row=R_END, column=j).alignment = Alignment(horizontal='center')

    # ─── Sheet 7: Valuation (references Assumptions valuation rows) ───
    ws7 = wb.create_sheet("07_Valuation")
    header_row(ws7, 1,
        ["Phương pháp", "Trọng số", "Giá trị CP (VND)", "Ghi chú"],
        [35, 12, 18, 45])

    V_EV7 = 2; V_PB7 = 3; V_PE7 = 4; V_W7 = 6; V_PRICE7 = 8; V_UP7 = 9
    # Formula references: each method points to the corresponding Assumptions cell
    as_ev_cell = f"{S_ASSUMP}!J{V_EV}"  # Assumptions J{V_EV} = EV/EBITDA target price
    as_pb_cell = f"{S_ASSUMP}!J{V_PB}"
    as_pe_cell = f"{S_ASSUMP}!J{V_PE}"

    labels_val = [
        ("EV/EBITDA (2026E, HPG median)", 0.40, as_ev_cell, "40% — EV/EBITDA dùng median lịch sử HPG"),
        ("P/B (2026E, HPG median)", 0.40, as_pb_cell, "40% — P/B là thước đo chu kỳ tin cậy"),
        ("P/E (2026E, HPG median)", 0.20, as_pe_cell, "20% — P/E tham khảo bổ sung"),
    ]

    for i, (name, weight, formula, note) in enumerate(labels_val, 2):
        ws7.cell(row=i, column=1, value=name).border = thin_border
        ws7.cell(row=i, column=2, value=weight).border = thin_border
        ws7.cell(row=i, column=2).number_format = '0%'
        ws7.cell(row=i, column=3, value=f"={formula}").border = thin_border
        ws7.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        ws7.cell(row=i, column=4, value=note).border = thin_border

    # P/B buy/sell zone reference
    zone_row = V_PE7 + 1

    # Weighted average (40% EV/EBITDA + 40% P/B + 20% P/E)
    V_EV7 = 2; V_PB7 = 3; V_PE7 = 4; V_W7 = 10; V_PRICE7 = 12; V_UP7 = 13

    # DCF note (below valuation rows)
    dcf_note_row = 7
    ws7.cell(row=dcf_note_row, column=1, value="DCF Model:").font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")
    ws7.cell(row=dcf_note_row, column=1).border = thin_border
    ws7.cell(row=dcf_note_row, column=2, value="Đang phát triển — sẽ thay thế tỷ trọng 70% khi hoàn thiện").font = Font(name=FONT_NAME, italic=True, size=9, color="888888")
    ws7.cell(row=dcf_note_row, column=2).border = thin_border
    ws7.merge_cells(start_row=dcf_note_row, start_column=2, end_row=dcf_note_row, end_column=4)
    ws7.cell(row=V_W7, column=1, value="Giá mục tiêu (có trọng số)").border = thin_border
    ws7.cell(row=V_W7, column=1).fill = assump_fill
    ws7.cell(row=V_W7, column=2, value=1.0).border = thin_border
    ws7.cell(row=V_W7, column=2).number_format = '0%'
    ws7.cell(row=V_W7, column=3,
             value=f"=C{V_EV7}*B{V_EV7}+C{V_PB7}*B{V_PB7}+C4*B4").border = thin_border
    ws7.cell(row=V_W7, column=3).fill = assump_fill
    ws7.cell(row=V_W7, column=3).alignment = Alignment(horizontal='center')
    ws7.cell(row=V_W7, column=3).number_format = '#,##0'

    # Current price
    ws7.cell(row=V_PRICE7, column=1, value="Giá hiện tại").border = thin_border
    ws7.cell(row=V_PRICE7, column=3, value=PRICE).border = thin_border
    ws7.cell(row=V_PRICE7, column=3).alignment = Alignment(horizontal='center')
    ws7.cell(row=V_PRICE7, column=3).number_format = '#,##0'

    # Upside
    ws7.cell(row=V_UP7, column=1, value="Upside (%)").border = thin_border
    ws7.cell(row=V_UP7, column=1).fill = p_fill
    ws7.cell(row=V_UP7, column=3,
             value=f"=C{V_W7}/C{V_PRICE7}-1").border = thin_border
    ws7.cell(row=V_UP7, column=3).fill = p_fill
    ws7.cell(row=V_UP7, column=3).alignment = Alignment(horizontal='center')
    ws7.cell(row=V_UP7, column=3).number_format = '0.0%'

    # ─── Sheet 8: Lịch sử P/E, P/B, EV/EBITDA theo quý ───
    ws8 = wb.create_sheet("08_Hist_Multiples")
    hist_headers = ["Năm", "Quý", "P/E (x)", "P/B (x)", "EV/EBITDA (x)", "Nhãn", "P/E gốc", "P/E hiệu chỉnh"]
    header_row(ws8, 1, hist_headers, [10, 8, 12, 12, 14, 12, 12, 14])
    raw_by_q = {}
    for r in HPG_RATIOS:
        if r.get("quarter") != 5:
            raw_by_q[(int(r["year"]), r["quarter"])] = r
    all_quarters = [(y, q) for y in range(2018, 2027) for q in range(1, 5)]
    qdata = []
    last_good_pe = None; last_good_pb = None; last_good_ev = None
    for y, q in all_quarters:
        rec = raw_by_q.get((y, q), {})
        pe0 = rec.get("pe"); pb0 = rec.get("pb"); ev0 = rec.get("ev_ebitda")
        # P/E: nếu LNST dương (PE 0-50) → giữ nguyên; nếu LNST âm → lấy quý gần nhất
        is_normal = bool(pe0 and 0 < pe0 < 50)
        pe_adj = round(pe0, 1) if is_normal else last_good_pe
        pe_orig_clean = round(pe0, 1) if is_normal else None
        pe_adj_only = pe_adj if not is_normal else None
        if is_normal: last_good_pe = pe_adj
        if pb0 and pb0 > 0:
            pb = round(pb0, 2); last_good_pb = pb
        else:
            pb = last_good_pb
        if ev0 and ev0 > 0:
            ev = round(ev0, 1); last_good_ev = ev
        else:
            ev = last_good_ev
        qdata.append((y, q, pe_adj, pe_orig_clean, pe_adj_only, pb, ev))
    yr_ranges = {}; yr_start = None; cur_yr = None
    for i, (y, q, pe_adj, pe_orig, pe_adjonly, pb, ev) in enumerate(qdata, 2):
        ws8.cell(row=i, column=1, value=y).number_format = '0'
        ws8.cell(row=i, column=2, value=q)
        ws8.cell(row=i, column=3, value=pe_adj).number_format = '0.0'
        ws8.cell(row=i, column=4, value=pb).number_format = '0.00'
        ws8.cell(row=i, column=5, value=ev).number_format = '0.0'
        ws8.cell(row=i, column=6, value=f"Q{q}-{y}")
        ws8.cell(row=i, column=7, value=pe_orig).number_format = '0.0'
        ws8.cell(row=i, column=8, value=pe_adjonly).number_format = '0.0'
        for c in range(1, 9):
            ws8.cell(row=i, column=c).border = thin_border
            ws8.cell(row=i, column=c).alignment = Alignment(horizontal='center')
            ws8.cell(row=i, column=c).font = data_font
        if y != cur_yr:
            if cur_yr is not None: yr_ranges[cur_yr] = (yr_start, i-1)
            yr_start = i; cur_yr = y
    if cur_yr is not None: yr_ranges[cur_yr] = (yr_start, len(qdata)+1)
    n_quarters = len(qdata)

    # ─── Bước 5: Biểu đồ P/E, P/B, EV/EBITDA theo quý (từ 08_Hist_Multiples) ───
    nq = len(qdata) + 1
    qcats = Reference(ws8, min_col=6, max_col=6, min_row=2, max_row=nq)

    chart_pe = LineChart()
    chart_pe.title = "P/E theo quý (2018-2026)"
    chart_pe.width = 22; chart_pe.height = 13
    pe_orig_ref = Reference(ws8, min_col=7, min_row=1, max_row=nq)
    pe_adj_ref = Reference(ws8, min_col=8, min_row=1, max_row=nq)
    chart_pe.add_data(pe_orig_ref, titles_from_data=True)
    chart_pe.add_data(pe_adj_ref, titles_from_data=True)
    chart_pe.set_categories(qcats)
    if len(chart_pe.series) >= 2:
        chart_pe.series[0].graphicalProperties.line.solidFill = "1F4E79"
        chart_pe.series[0].graphicalProperties.line.width = 18000
        chart_pe.series[1].graphicalProperties.line.solidFill = "FF8C00"
        chart_pe.series[1].graphicalProperties.line.dashStyle = "dash"
        chart_pe.series[1].graphicalProperties.line.width = 18000
    chart_pe.legend = None
    chart_pe.y_axis.title = "P/E (x)"
    chart_pe.y_axis.delete = False
    chart_pe.y_axis.tickLblPos = "low"
    chart_pe.y_axis.numFmt = '0'
    chart_pe.y_axis.scaling.min = 0
    chart_pe.y_axis.scaling.max = 28
    chart_pe.y_axis.majorUnit = 4
    chart_pe.x_axis.delete = False
    chart_pe.x_axis.tickLblPos = "low"
    ws2.add_chart(chart_pe, f"A{V_END+2}")

    chart_pb = LineChart()
    chart_pb.title = "P/B theo quý (2018-2026)"
    chart_pb.width = 22; chart_pb.height = 13
    pb_data_q = Reference(ws8, min_col=4, min_row=1, max_row=nq)
    chart_pb.add_data(pb_data_q, titles_from_data=True)
    chart_pb.set_categories(qcats)
    if chart_pb.series:
        chart_pb.series[0].graphicalProperties.line.solidFill = "E74C3C"
        chart_pb.series[0].graphicalProperties.line.width = 18000
    chart_pb.legend = None
    chart_pb.y_axis.title = "P/B (x)"
    chart_pb.y_axis.delete = False
    chart_pb.y_axis.tickLblPos = "low"
    chart_pb.y_axis.numFmt = '0.0'
    chart_pb.y_axis.scaling.min = 0
    chart_pb.y_axis.scaling.max = 3.5
    chart_pb.y_axis.majorUnit = 0.5
    chart_pb.x_axis.delete = False
    chart_pb.x_axis.tickLblPos = "low"
    ws2.add_chart(chart_pb, f"A{V_END+20}")

    chart_ev = LineChart()
    chart_ev.title = "EV/EBITDA theo quý (2018-2026)"
    chart_ev.width = 22; chart_ev.height = 13
    ev_data_q = Reference(ws8, min_col=5, min_row=1, max_row=nq)
    chart_ev.add_data(ev_data_q, titles_from_data=True)
    chart_ev.set_categories(qcats)
    if chart_ev.series:
        chart_ev.series[0].graphicalProperties.line.solidFill = "27AE60"
        chart_ev.series[0].graphicalProperties.line.width = 18000
    chart_ev.legend = None
    chart_ev.y_axis.title = "EV/EBITDA (x)"
    chart_ev.y_axis.delete = False
    chart_ev.y_axis.tickLblPos = "low"
    chart_ev.y_axis.numFmt = '0'
    chart_ev.y_axis.scaling.min = 0
    chart_ev.y_axis.scaling.max = 25
    chart_ev.y_axis.majorUnit = 5
    chart_ev.x_axis.delete = False
    chart_ev.x_axis.tickLblPos = "low"
    ws2.add_chart(chart_ev, f"A{V_END+38}")

    # ─── Sheet 9: Sensitivity (FORMULA-BASED) ───
    ws9 = wb.create_sheet("09_Sensitivity")
    ws9.cell(row=1, column=1, value="Sensitivity: EV/EBITDA vs EBITDA 2026E (Target Price VND)").font = sub_font
    ws9.merge_cells('A1:G1')

    ev_ebitda_range = [5.5, 6.5, 7.5, 8.5, 9.5]
    ebitda_pcts = [0.85, 0.925, 1.0, 1.075, 1.15]

    header_row(ws9, 3, [""] + [f"EV/EBITDA {x}x" for x in ev_ebitda_range], [25]+[16]*5)

    # Reference cells for EBITDA and Net Debt
    for i, pct in enumerate(ebitda_pcts):
        row = 4 + i
        lbl = f"EBITDA {pct*100:.0f}% kế hoạch"
        ws9.cell(row=row, column=1, value=lbl).border = thin_border
        for j, mult in enumerate(ev_ebitda_range):
            col = 2 + j
            formula = (f"=(({S_PNL}!$G$16*{pct}*{mult}"
                       f"-({S_ASSUMP}!G11-{S_ASSUMP}!G12))*1000000000)"
                       f"/({S_ASSUMP}!$B$3*1000000)")
            ws9.cell(row=row, column=col, value=formula).number_format = '#,##0'
            ws9.cell(row=row, column=col).border = thin_border
            ws9.cell(row=row, column=col).alignment = Alignment(horizontal='center')

    # Conditional formatting for sensitivity
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    for row in range(4, 9):
        for col in range(2, 7):
            cell = ws9.cell(row=row, column=col)
            cell_ref = f"{chr(64+col)}{row}"
            target_ref = f"C6"
            current_ref = f"C{V_PRICE7}"

    # ─── Sheet 10: PESTLE ───
    ws10 = wb.create_sheet("10_PESTLE")
    header_row(ws10, 1, ["Yếu tố", "Nội dung", "Tác động"], [20, 80, 15])
    for i, (factor, content, impact) in enumerate(pestle, 2):
        ws10.cell(row=i, column=1, value=factor).font = bold_font
        ws10.cell(row=i, column=1).border = thin_border
        ws10.cell(row=i, column=2, value=content).font = data_font
        ws10.cell(row=i, column=2).border = thin_border
        ws10.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws10.cell(row=i, column=3, value=impact).font = bold_font
        ws10.cell(row=i, column=3).border = thin_border
        ws10.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        if "Tích cực" in impact:
            ws10.cell(row=i, column=3).fill = p_fill
        elif "Tiêu cực" in impact:
            ws10.cell(row=i, column=3).fill = n_fill

    # ── Chu kỳ ngành (Tầng 2 artifact) ──
    cycle_row = len(pestle) + 3
    ws10.cell(row=cycle_row, column=1, value="VỊ TRÍ CHU KỲ NGÀNH THÉP").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws10.merge_cells(start_row=cycle_row, start_column=1, end_row=cycle_row, end_column=4)
    cycle_data = [
        ("Tăng trưởng DT ngành", "Cao (HPG DT 2025 +12.4% YoY)", "Cao", "Phục hồi sau đáy 2023"),
        ("Công suất ngành", "Cân bằng (HPG ~85%, HSG ~75%)", "Cân bằng", "DQ2 thêm 5.6M tấn"),
        ("Giá bán (HRC)", "580 USD/tấn (2025), phục hồi từ đáy 480 (2023)", "Tăng", "Thuế CBPG 27.8% hỗ trợ"),
        ("CAPEX ngành", "DQ2 ~$3 tỷ xong T12/2025. CAPEX 2025: 25.748 tỷ", "Giảm từ đỉnh", "Giai đoạn thu hoạch sau CAPEX"),
        ("Số đối thủ", "HPG (~70% HRC), Formosa, HSG, NKG", "Ổn định", "Rào cản gia nhập cao"),
    ]
    cycle_row += 1
    cycle_headers = ["Tiêu chí", "Trạng thái", "Xu hướng", "Ghi chú"]
    for c, h in enumerate(cycle_headers, 1):
        cell = ws10.cell(row=cycle_row, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    for i, row_data in enumerate(cycle_data, cycle_row + 1):
        for c, v in enumerate(row_data, 1):
            cell = ws10.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    conclusion_row = cycle_row + len(cycle_data) + 1
    ws10.cell(row=conclusion_row, column=1, value="Kết luận chu kỳ:").font = bold_font
    ws10.cell(row=conclusion_row, column=1).border = thin_border
    ws10.cell(row=conclusion_row, column=2,
              value="Ngành thép VN đang ở giai đoạn PHỤC HỒI → TĂNG TRƯỞNG. Đáy chu kỳ (2022-2023) đã qua. "
                    "DQ2 hoàn thành đưa HPG vào giai đoạn thu hoạch. Kỳ vọng biên GP mở rộng 2026-2028.").font = Font(name=FONT_NAME, bold=True, size=10, color="006600")
    ws10.cell(row=conclusion_row, column=2).border = thin_border
    ws10.merge_cells(start_row=conclusion_row, start_column=2, end_row=conclusion_row, end_column=4)
    ws10.cell(row=conclusion_row, column=2).alignment = Alignment(wrap_text=True)
    ws10.row_dimensions[conclusion_row].height = 40

    # ── TAM (Total Addressable Market) ──
    tam_row = conclusion_row + 2
    ws10.cell(row=tam_row, column=1, value="THỊ TRƯỜNG MỤC TIÊU (TAM)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws10.merge_cells(start_row=tam_row, start_column=1, end_row=tam_row, end_column=4)
    tam_data = [
        ("TAM thép VN (triệu tấn)", "~28-30", "VSA (Hiệp hội Thép VN) 2025"),
        ("CAGR TAM (3 năm)", "~5-7%", "Đầu tư công + BĐS phục hồi"),
        ("Thị phần HPG (HRC)", "~70%", "Độc quyền HRC nội địa"),
        ("Thị phần HPG (thép XD)", "~40%", "Dẫn đầu thị trường"),
        ("Thị phần tiềm năng (2028)", "~75% HRC, ~45% XD", "DQ2 mở rộng công suất"),
    ]
    tam_row += 1
    tam_headers = ["Chỉ tiêu", "Giá trị", "Nguồn"]
    for c, h in enumerate(tam_headers, 1):
        cell = ws10.cell(row=tam_row, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    for i, row_data in enumerate(tam_data, tam_row + 1):
        for c, v in enumerate(row_data, 1):
            cell = ws10.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # ─── Sheet 11: Leading Indicators ───
    ws11 = wb.create_sheet("11_Leading_Indicators")
    header_row(ws11, 1,
        ["Chỉ báo", "Ngưỡng tích cực", "Ngưỡng tiêu cực",
         "Giá trị hiện tại", "Trạng thái"],
        [35, 18, 18, 18, 15])
    for i, (name, pos_thresh, neg_thresh, curr, status) in enumerate(leading_indicators, 2):
        data_row(ws11, i, [name, pos_thresh, neg_thresh, curr, status])
        if "Tích cực" in status:
            ws11.cell(row=i, column=5).fill = p_fill
        elif "Tiêu cực" in status:
            ws11.cell(row=i, column=5).fill = n_fill

    # ─── Sheet 12: Investment Thesis ───
    ws12 = wb.create_sheet("12_Investment_Thesis")
    header_row(ws12, 1, ["Tầng", "Nội dung", "Đánh giá"], [12, 80, 15])

    thesis = [
        ("1", "Chuỗi giá trị: HPG tích hợp dọc từ quặng → lò cao → HRC/thép XD. Tự chủ 50% quặng (trữ lượng 320M tấn). "
              "DQ2 (5.6M tấn) full công suất từ T12/2025 giúp HPG dẫn đầu chi phí.", "Mạnh"),
        ("2", "Thị trường: VN nhập khẩu HRC ~8M tấn/năm. Thuế CBPG 27.8% với TQ tạo lợi thế cho HPG. "
              "Đầu tư công (cao tốc Bắc-Nam, Long Thành) + BĐS phục hồi kéo nhu cầu thép.", "Mạnh"),
        ("3", "Moat: Chi phí thấp nhất ngành (lò cao BOF công suất lớn). Quy mô 16M tấn. "
              "Mạng lưới phân phối rộng. Khó sao chép do CAPEX khổng lồ (DQ2 ~$3 tỷ).", "Rộng"),
        ("4", "Tài chính: D/E 0.65x an toàn. ROE TTM 16.5%. FCF cải thiện nhờ DQ2 vận hành. "
              f"Biên LNG phục hồi từ 10.9% (2023) lên ~{gp_margin_fc[0]}% (2026E).", "Tốt"),
        ("5", "Rủi ro: (1) Giá HRC giảm mạnh do suy giảm TQ, (2) Chi phí nguyên liệu tăng (quặng/than), "
              "(3) Tỷ giá USD tăng gây áp lực nợ vay, (4) Cạnh tranh từ HSG/NKG tôn mạ.", "Trung bình"),
        ("6", "Định giá: P/B 1.56x (dưới median lịch sử 1.61x HPG). EV/EBITDA 2026E ~7.0x. "
               "Multiple mục tiêu dựa trên trung vị lịch sử HPG (TTM 2018-2026 từ Vietcap). "
               "P/E không dùng (bẫy chu kỳ). P/B buy zone <1.0x, sell >2.0x.", "Hấp dẫn"),
    ]

    for i, (level, content, rating) in enumerate(thesis, 2):
        ws12.cell(row=i, column=1, value=level).font = bold_font
        ws12.cell(row=i, column=1).border = thin_border
        ws12.cell(row=i, column=1).alignment = Alignment(horizontal='center')
        ws12.cell(row=i, column=2, value=content).font = data_font
        ws12.cell(row=i, column=2).border = thin_border
        ws12.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws12.row_dimensions[i].height = 60
        ws12.cell(row=i, column=3, value=rating).font = bold_font
        ws12.cell(row=i, column=3).border = thin_border
        ws12.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        if rating == "Mạnh" or rating == "Hấp dẫn":
            ws12.cell(row=i, column=3).fill = p_fill
        elif rating == "Rộng":
            ws12.cell(row=i, column=3).fill = p_fill

    # ── Detailed Value Chain Table (Tầng 1 artifact) ──
    vc_start = len(thesis) + 3
    ws12.cell(row=vc_start, column=1, value="CHI TIẾT TẦNG 1: CHUỖI GIÁ TRỊ NGÀNH THÉP").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws12.merge_cells(start_row=vc_start, start_column=1, end_row=vc_start, end_column=7)
    vc_header = ["Bước", "Mô tả", "Biên ước tính", "Ai chi phối", "Vị trí HPG"]
    widths_vc = [8, 45, 14, 20, 18]
    for c, (h, w) in enumerate(zip(vc_header, widths_vc), 1):
        cell = ws12.cell(row=vc_start+1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    value_chain = [
        ("1", "Nguyên liệu (quặng sắt/than cốc)", "0-5%", "Vale, BHP, Rio Tinto", "Tự chủ 50% quặng (Quý Xa, 320M tấn)"),
        ("2", "Luyện cốc (sản xuất than cốc từ than mỡ)", "6-12%", "HPG (tự SX tại DQ)", "Trung nguồn — tự SX cốc"),
        ("3", "Sản xuất phôi thép (lò cao BOF)", "15-25%", "HPG, Formosa (Trung nguồn)", "Trung nguồn — mũi nhọn"),
        ("4", "Cán thép (HRC / thép xây dựng)", "8-15%", "HPG, HSG, NKG", "Trung nguồn — thị phần HRC ~70%"),
        ("5", "Phân phối (đại lý / dự án)", "2-4%", "Đại lý cấp 1, cấp 2", "Hạ nguồn — mạng 300+ đại lý"),
        ("6", "Người dùng cuối (XD, BĐS, CN)", "—", "CTCP, BĐS, CN nhẹ", "—"),
    ]
    for i, row_data in enumerate(value_chain, vc_start+2):
        for c, v in enumerate(row_data, 1):
            cell = ws12.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            if c == 2:
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center')

    # ── Moat Analysis (Tầng 3 artifact) ──
    m_start = vc_start + len(value_chain) + 2
    ws12.cell(row=m_start, column=1, value="CHI TIẾT TẦNG 3: PHÂN TÍCH MOAT & ROIC").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws12.merge_cells(start_row=m_start, start_column=1, end_row=m_start, end_column=7)
    m_header = ["Loại Moat", "Mức độ", "Bằng chứng định lượng", "Kết luận"]
    widths_m = [22, 12, 45, 20]
    for c, (h, w) in enumerate(zip(m_header, widths_m), 1):
        cell = ws12.cell(row=m_start+1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    moat_data = [
        ("1. Lợi thế chi phí (Cost)", "Mạnh", "DQ2 BOF 6M tấn, ĐMTK thấp nhất ngành. Biên GP 2025: 15.7% > HSG (10.2%). Tự SX cốc giảm 10-15% CP.", "Wide — chi phí/tấn thấp nhất VN"),
        ("2. Chuyển đổi (Switching)", "TB", "Khách hàng đại lý lâu năm, chi phí vận chuyển cao khi đổi NCC. Nhưng thép là hàng hóa tiêu chuẩn.", "Narrow — rào cản chuyển đổi ở mức vừa"),
        ("3. Tài sản vô hình", "TB", "Thương hiệu Hoà Phát uy tín, không có patent đặc thù. Giấy phép khai thác quặng tại Quý Xa.", "Narrow — chỉ có giấy phép khoáng sản"),
        ("4. Hiệu ứng mạng", "Yếu", "Thép là hàng hóa B2B, không có network effect. Khách hàng không tương tác với nhau.", "Không có"),
        ("5. Quy mô hiệu quả", "Mạnh", "Thị phần HRC ~70%, thép XD ~40%. CAPEX rào cản ~$3 tỷ cho lò cao mới. Quy mô 16M tấn.", "Wide — rào cản vốn khổng lồ"),
    ]
    for i, row_data in enumerate(moat_data, m_start+2):
        for c, v in enumerate(row_data, 1):
            cell = ws12.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            if c == 3:
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center')
    overall_row = m_start + len(moat_data) + 2
    ws12.cell(row=overall_row, column=1, value="Overall Moat Rating:").font = bold_font
    ws12.cell(row=overall_row, column=1).border = thin_border
    ws12.cell(row=overall_row, column=2, value="WIDE MOAT | ROIC ~10.5% > WACC ~9.0% → Tạo giá trị").font = Font(name=FONT_NAME, bold=True, size=11, color="006600")
    ws12.cell(row=overall_row, column=2).border = thin_border
    ws12.cell(row=overall_row, column=2).fill = p_fill
    ws12.merge_cells(start_row=overall_row, start_column=2, end_row=overall_row, end_column=7)

    # ── Risk Matrix (Tầng 5 artifact) ──
    r_start = overall_row + 2
    ws12.cell(row=r_start, column=1, value="CHI TIẾT TẦNG 5: MA TRẬN RỦI RO & CATALYST").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws12.merge_cells(start_row=r_start, start_column=1, end_row=r_start, end_column=7)
    risk_header = ["Rủi ro", "Loại", "Xác suất", "Tác động", "Leading Indicator", "Ngưỡng kích hoạt"]
    widths_risk = [22, 10, 10, 14, 25, 18]
    for c, (h, w) in enumerate(zip(risk_header, widths_risk), 1):
        cell = ws12.cell(row=r_start+1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    risk_data = [
        ("Giá HRC giảm mạnh", "Ngành", "TB", "−20% DT", "Giá HRC SHFE (USD/tấn)", "<400 USD/tấn"),
        ("Chi phí quặng tăng", "Ngành", "TB", "−5% biên GP", "Giá quặng 62%Fe (USD/tấn)", ">150 USD/tấn"),
        ("Tỷ giá USD tăng", "VM", "Thấp", "−1.500 tỷ LNST", "USD/VND", ">26.000"),
        ("Cạnh tranh Formosa", "Ngành", "Thấp", "−10% thị phần HRC", "CS XK HRC TQ vào VN", "Công suất Formosa tăng"),
        ("BĐS suy thoái kéo dài", "VM", "Thấp", "−15% SL thép XD", "Giải ngân đầu tư công", "Tăng trưởng tín dụng BĐS <5%"),
    ]
    for i, row_data in enumerate(risk_data, r_start+2):
        for c, v in enumerate(row_data, 1):
            cell = ws12.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        ws12.row_dimensions[i].height = 25

    # ── Catalyst Table ──
    cat_start = r_start + len(risk_data) + 2
    ws12.cell(row=cat_start, column=1, value="CATALYST TĂNG TỐC THESIS").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws12.merge_cells(start_row=cat_start, start_column=1, end_row=cat_start, end_column=7)
    cat_header = ["Catalyst", "Mô tả", "Thời điểm", "Tác động giá", "Xác suất"]
    widths_cat = [20, 35, 12, 14, 10]
    for c, (h, w) in enumerate(zip(cat_header, widths_cat), 1):
        cell = ws12.cell(row=cat_start+1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    catalyst_data = [
        ("KQKD Q1/2026 CB", "KQKD khả quan nhờ sản lượng cao & spread nở", "07/2026", "+10-15%", "Cao"),
        ("DQ2 full công suất", "Dung Quất 2 chạy hết công suất 6M tấn, giảm ĐMTK", "H2/2026", "+5-10%", "Cao"),
        ("Giải ngân đầu tư công", "Cao tốc Bắc-Nam, Long Thành thúc đẩy nhu cầu thép", "2026-2027", "+5-10%", "TB"),
        ("Giá HRC phục hồi", "Nhu cầu TQ hồi phục + thuế CBPG VN hỗ trợ giá", "H2/2026", "+15-25%", "TB"),
    ]
    for i, row_data in enumerate(catalyst_data, cat_start+2):
        for c, v in enumerate(row_data, 1):
            cell = ws12.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        ws12.row_dimensions[i].height = 25

    # ─── Sheet 13: Summary ───
    ws13 = wb.create_sheet("13_Summary_Snapshot")
    header_row(ws13, 1,
        ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E", "CAGR 25-28E"],
        [35] + [14]*8 + [14])

    all_eps = [5160, 1260, 1010, 1430, 1830, 2600, 3320, 4030]

    summary_items = [
        ("Doanh thu (tỷ)", all_rev),
        ("LNST (tỷ)", all_ni),
        ("Biên LNG (%)", all_gpm),
        ("EBIT (tỷ)", None),
        ("EBITDA (tỷ)", None),
        ("Biên LNST (%)", None),
        ("EPS (VND)", all_eps),
        ("ROE (%)", None),
        ("ROA (%)", None),
        ("D/E (x)", None),
        ("P/E (x)", None),
        ("P/B (x)", None),
        ("EV/EBITDA (x)", None),
        ("CAPEX (tỷ)", None),
    ]

    for i, (name, vals) in enumerate(summary_items, 2):
        ws13.cell(row=i, column=1, value=name).font = bold_font
        ws13.cell(row=i, column=1).border = thin_border
        if vals:
            for j, v in enumerate(vals, 2):
                cell = ws13.cell(row=i, column=j, value=v)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if isinstance(v, float):
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '#,##0'

    # ── Summary formulas ──
    for j in range(2, 10):
        cl = col_ltr(j)

        # Row 2: Doanh thu = PnL!col2
        ws13.cell(row=2, column=j, value=f"={S_PNL}!{cl}2").number_format = '#,##0'
        ws13.cell(row=2, column=j).border = thin_border
        ws13.cell(row=2, column=j).alignment = Alignment(horizontal='center')

        # Row 3: LNST = PnL!col13
        ws13.cell(row=3, column=j, value=f"={S_PNL}!{cl}13").number_format = '#,##0'
        ws13.cell(row=3, column=j).border = thin_border
        ws13.cell(row=3, column=j).alignment = Alignment(horizontal='center')

        # Row 4: GP Margin = PnL!col5
        ws13.cell(row=4, column=j, value=f"={S_PNL}!{cl}5").number_format = '0.0'
        ws13.cell(row=4, column=j).border = thin_border
        ws13.cell(row=4, column=j).alignment = Alignment(horizontal='center')

        # Row 5: EBIT = PnL!col7
        ws13.cell(row=5, column=j, value=f"={S_PNL}!{cl}7").number_format = '#,##0'
        ws13.cell(row=5, column=j).border = thin_border
        ws13.cell(row=5, column=j).alignment = Alignment(horizontal='center')

        # Row 6: EBITDA = PnL!col16
        ws13.cell(row=6, column=j, value=f"={S_PNL}!{cl}16").number_format = '#,##0'
        ws13.cell(row=6, column=j).border = thin_border
        ws13.cell(row=6, column=j).alignment = Alignment(horizontal='center')

        # Row 7: NI Margin = PnL!col14
        ws13.cell(row=7, column=j, value=f"={S_PNL}!{cl}14").number_format = '0.0'
        ws13.cell(row=7, column=j).border = thin_border
        ws13.cell(row=7, column=j).alignment = Alignment(horizontal='center')

        # Row 8: EPS = PnL!col15
        ws13.cell(row=8, column=j, value=f"={S_PNL}!{cl}15").number_format = '#,##0'
        ws13.cell(row=8, column=j).border = thin_border
        ws13.cell(row=8, column=j).alignment = Alignment(horizontal='center')

        # Row 9: ROE = NI / Equity
        ws13.cell(row=9, column=j,
                  value=f"={S_PNL}!{cl}13/{S_ASSUMP}!{cl}13*100").number_format = '0.0'
        ws13.cell(row=9, column=j).border = thin_border
        ws13.cell(row=9, column=j).alignment = Alignment(horizontal='center')

        # Row 10: ROA = NI / Total Assets
        ws13.cell(row=10, column=j,
                  value=f"={S_PNL}!{cl}13/{S_BS}!{cl}14*100").number_format = '0.0'
        ws13.cell(row=10, column=j).border = thin_border
        ws13.cell(row=10, column=j).alignment = Alignment(horizontal='center')

        # Row 11: D/E = (Debt - Cash) / Equity
        ws13.cell(row=11, column=j,
                  value=f"=({S_ASSUMP}!{cl}11-{S_ASSUMP}!{cl}12)/{S_ASSUMP}!{cl}13").number_format = '0.00'
        ws13.cell(row=11, column=j).border = thin_border
        ws13.cell(row=11, column=j).alignment = Alignment(horizontal='center')

        # Row 12: P/E = (Price * Shares/1000) / NI
        ws13.cell(row=12, column=j,
                  value=f"=({S_ASSUMP}!$B$2*{S_ASSUMP}!$B$3/1000)/{S_PNL}!{cl}13").number_format = '0.0'
        ws13.cell(row=12, column=j).border = thin_border
        ws13.cell(row=12, column=j).alignment = Alignment(horizontal='center')

        # Row 13: P/B = (Price*Shares/1000) / Equity
        ws13.cell(row=13, column=j,
                  value=f"=({S_ASSUMP}!$B$2*{S_ASSUMP}!$B$3/1000)/{S_ASSUMP}!{cl}13").number_format = '0.00'
        ws13.cell(row=13, column=j).border = thin_border
        ws13.cell(row=13, column=j).alignment = Alignment(horizontal='center')

        # Row 14: EV/EBITDA = (MCap + NetDebt) / EBITDA
        ws13.cell(row=14, column=j,
                  value=f"=({S_ASSUMP}!$B$2*{S_ASSUMP}!$B$3/1000+"
                        f"({S_ASSUMP}!{cl}11-{S_ASSUMP}!{cl}12))/{S_PNL}!{cl}16").number_format = '0.0'
        ws13.cell(row=14, column=j).border = thin_border
        ws13.cell(row=14, column=j).alignment = Alignment(horizontal='center')

        # Row 15: CAPEX = Assumptions!col10
        ws13.cell(row=15, column=j, value=f"={S_ASSUMP}!{cl}10").number_format = '#,##0'
        ws13.cell(row=15, column=j).border = thin_border
        ws13.cell(row=15, column=j).alignment = Alignment(horizontal='center')

    # CAGR (2025→2028E): Column 10
    for row in [2, 3, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15]:
        ws13.cell(row=row, column=10,
                  value=f"=(I{row}/F{row})^(1/3)-1").number_format = '0.0%'
        ws13.cell(row=row, column=10).border = thin_border
        ws13.cell(row=row, column=10).alignment = Alignment(horizontal='center')

    # ── KH ĐHĐCĐ vs Thực hiện ──
    kh_row = 17
    ws13.cell(row=kh_row, column=1, value="KẾ HOẠCH ĐHĐCĐ 2026 vs THỰC HIỆN 2025").font = Font(name=FONT_NAME, bold=True, size=11, color="C0392B")
    ws13.merge_cells(start_row=kh_row, start_column=1, end_row=kh_row, end_column=10)
    kh_header = ["Chỉ tiêu", "KH 2025", "TH 2025", "% hoàn thành", "KH 2026", "Q1/2026", "% KH năm"]
    for c, h in enumerate(kh_header, 1):
        cell = ws13.cell(row=kh_row+1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    def _q(records, yr, qtr, field):
        for r in records:
            if r.get("yearReport") == yr and r.get("lengthReport") == qtr:
                v = r.get(field)
                return v / 1e9 if v is not None else 0
        return 0
    _is_qs = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    q1rev = _q(_is_qs, 2026, 1, 'isa3')
    q1ni = _q(_is_qs, 2026, 1, 'isa22')
    kh_data = [
        ("Doanh thu (tỷ)", 170000, revenue_hist[4], f"{revenue_hist[4]/170000*100:.1f}%", 210000, f"{q1rev:,.0f}", f"{q1rev/210000*100:.1f}%"),
        ("LNST (tỷ)", 15000, ni_hist[4], f"{ni_hist[4]/15000*100:.1f}%", 22000, f"{q1ni:,.0f}", f"{q1ni/22000*100:.1f}%"),
    ]
    for i, row_data in enumerate(kh_data, kh_row+2):
        for c, v in enumerate(row_data, 1):
            cell = ws13.cell(row=i, column=c, value=v)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.font = bold_font if c == 1 else data_font
        ws13.row_dimensions[i].height = 25

    # ── Earnings Quality Analysis (Tầng 4 artifact) ──
    eq_row = kh_row + len(kh_data) + 2
    ws13.cell(row=eq_row, column=1, value="PHÂN TÍCH CHẤT LƯỢNG LỢI NHUẬN").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws13.merge_cells(start_row=eq_row, start_column=1, end_row=eq_row, end_column=10)
    eq_header = ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "Nhận xét"]
    for c, h in enumerate(eq_header, 1):
        cell = ws13.cell(row=eq_row+1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    eq_data = [
        ("Cash Conversion (CFO/LNST)", 
         round(cfo_hist[0]/ni_hist[0], 2), round(cfo_hist[1]/ni_hist[1], 2),
         round(cfo_hist[2]/ni_hist[2], 2), round(cfo_hist[3]/ni_hist[3], 2),
         round(cfo_hist[4]/ni_hist[4], 2), ">0.8 = chất lượng cao. 2021 đỉnh nhờ hàng tồn kho giảm"),
        ("Accruals Ratio ((CFO-NI)/TS)",
         round((cfo_hist[0]-ni_hist[0])/total_assets_hist[0]*100, 1),
         round((cfo_hist[1]-ni_hist[1])/total_assets_hist[1]*100, 1),
         round((cfo_hist[2]-ni_hist[2])/total_assets_hist[2]*100, 1),
         round((cfo_hist[3]-ni_hist[3])/total_assets_hist[3]*100, 1),
         round((cfo_hist[4]-ni_hist[4])/total_assets_hist[4]*100, 1),
         "Càng thấp càng tốt (ít accruals). Dương = CFO < NI → dồn tích tăng"),
        ("CAPEX Maintenance (% CAPEX)",
         "~60%", "~60%", "~60%", "~70%", "~70%",
         "DQ2 mới hoàn thành → bảo trì thấp hơn giai đoạn xây dựng"),
        ("One-off items (tỷ)",
         "—", "—", "—", "—", "—",
         "Không có one-off đáng kể. LNST ~CFO phản ánh chất lượng tốt"),
    ]
    eq_fill_warn = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for i, row_data in enumerate(eq_data, eq_row+2):
        for c, v in enumerate(row_data, 1):
            cell = ws13.cell(row=i, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if c == 1:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='left')
            elif c == 7:
                cell.font = Font(name=FONT_NAME, italic=True, size=9, color="C0392B")
                cell.alignment = Alignment(wrap_text=True, horizontal='left')
                cell.fill = eq_fill_warn
            else:
                cell.font = data_font
                if isinstance(v, (int, float)):
                    cell.value = v
                    cell.number_format = '0.0' if abs(v) < 100 else '#,##0'
                else:
                    cell.value = v
        ws13.row_dimensions[i].height = 30

    # ── Bridge Analysis (2025A → 2026E) ──
    br_row = eq_row + len(eq_data) + 3
    ws13.cell(row=br_row, column=1, value="BRIDGE ANALYSIS 2025A → 2026E").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws13.merge_cells(start_row=br_row, start_column=1, end_row=br_row, end_column=10)
    bridge_data = [
        ("Doanh thu 2025A", f"{revenue_hist[4]:,.0f} tỷ", "Cơ sở"),
        ("+ Tăng từ DQ2 full công suất", f"+{revenue_fc[0]-revenue_hist[4]:,.0f} tỷ", "DQ2 5.6M tấn chạy full năm 2026"),
        ("+ Tăng từ giá HRC phục hồi", "+~15,000 tỷ", "Giá HRC 580→620 USD/tấn"),
        ("= Doanh thu 2026E", f"{revenue_fc[0]:,.0f} tỷ", "KH ĐHĐCĐ: 210,000 tỷ"),
        ("", "", ""),
        ("Biên GP 2025A", "15.7%", "Cơ sở"),
        ("+ Cải thiện nhờ DQ2 (ĐMTK thấp hơn)", "+1.5%", "DQ2 BOF hiện đại, tiết kiệm 15-20% CP"),
        ("+ Tồn kho giá rẻ (quặng mua H2/2025)", "+0.5%", "Giá quặng 105 USD/tấn thấp hơn 2024"),
        (f"= Biên GP 2026E", f"{gp_margin_fc[0]}%", "Spread trend từ Q1/2026"),
        ("", "", ""),
        ("LNST 2025A", f"{ni_hist[4]:,.0f} tỷ", "Cơ sở"),
        ("+ Tăng từ DT + biên GP mở rộng", f"+{ni_fc[0]-ni_hist[4]:,.0f} tỷ", "Operating leverage"),
        ("- Tăng chi phí lãi vay", "-~500 tỷ", "Dư nợ tăng nhẹ (DQ2 vốn vay)"),
        ("= LNST 2026E", f"{ni_fc[0]:,.0f} tỷ", "KH ĐHĐCĐ: 22,000 tỷ"),
    ]
    br_row += 1
    br_headers = ["Chỉ tiêu", "Giá trị", "Giải thích"]
    for c, h in enumerate(br_headers, 1):
        cell = ws13.cell(row=br_row, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    for i, row_data in enumerate(bridge_data, br_row+1):
        for c, v in enumerate(row_data, 1):
            cell = ws13.cell(row=i, column=c, value=v)
            cell.font = data_font; cell.border = thin_border
            if c == 2:
                cell.alignment = Alignment(horizontal='center')
            elif c == 3:
                cell.alignment = Alignment(wrap_text=True, horizontal='left')
        ws13.row_dimensions[i].height = 22

    # ─── Charts ───
    # Chart 1: Revenue & NI trend
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Doanh thu & LNST"
    chart1.y_axis.title = "Tỷ VND"
    chart1.x_axis.title = "Năm"
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12

    data_ref = Reference(ws4, min_col=1, max_col=9, min_row=1, max_row=2)
    cats = Reference(ws4, min_col=2, max_col=9, min_row=1)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.solidFill = "1F4E79"

    # Add NI as line
    line_chart = LineChart()
    ni_ref = Reference(ws4, min_col=2, max_col=9, min_row=12, max_row=12)
    line_chart.add_data(ni_ref, from_rows=True)
    line_chart.series[0].graphicalProperties.line.solidFill = "E74C3C"
    chart1 += line_chart

    ws4.add_chart(chart1, "A17")

    # Chart 2: Margins
    chart2 = LineChart()
    chart2.title = "Biên lợi nhuận (%)"
    chart2.y_axis.title = "%"
    chart2.style = 10
    chart2.width = 20
    chart2.height = 12

    gp_ref = Reference(ws4, min_col=2, max_col=9, min_row=4, max_row=4)
    ebit_ref = Reference(ws4, min_col=2, max_col=9, min_row=7, max_row=7)
    ni_ref2 = Reference(ws4, min_col=2, max_col=9, min_row=13, max_row=13)

    chart2.add_data(gp_ref, from_rows=True)
    chart2.add_data(ebit_ref, from_rows=True)
    chart2.add_data(ni_ref2, from_rows=True)
    chart2.set_categories(cats)

    colors = ["1F4E79", "E74C3C", "27AE60"]
    for i, s in enumerate(chart2.series):
        s.graphicalProperties.line.solidFill = colors[i]

    ws4.add_chart(chart2, "A33")

    # ─── Sheet 14: Steel Industry Analysis (skill: thep) ───
    ws14 = wb.create_sheet("14_Steel_Analysis")
    header_row(ws14, 1,
        ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"],
        [40] + [12]*8)
    sec_font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")

    r = 1
    ws14.cell(row=1, column=1, value="PHÂN TÍCH NGÀNH THÉP — HPG (Skill: thep)").font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E79")
    ws14.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws14.cell(row=2, column=1, value="Phân loại: phan-loai-nganh → Nhóm 7: Thép (HPG, HSG, NKG, TVN)").font = data_font
    ws14.cell(row=3, column=1, value=(
        f"Công thức Spread (lag 1 quý): Giá HRC quý này - 1.6×Quặng 62%Fe quý TRƯỚC - 0.6×Than cốc quý TRƯỚC "
        f"- {OTHER_COST_USD} USD/tấn (CP SX khác cố định) — do tồn kho HPG bình quân ~90 ngày = ~1 quý. Xem sheet 17_Gia_Hang_Hoa."
    )).font = data_font

    # Section 1: Input/Output Prices
    r = 5
    ws14.cell(row=r, column=1, value="1. GIÁ ĐẦU VÀO & ĐẦU RA (USD/tấn)").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r = 6
    header_row(ws14, r, ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"], [40] + [12]*8)

    price_data = [
        ("Giá HRC (USD/tấn)", *HRC_PRICE_A, False),
        ("Giá quặng sắt 62%Fe (USD/tấn)", *IRON_ORE_A, False),
        ("Giá than cốc (USD/tấn)", *COKE_A, False),
        ("Spread thép (USD/tấn)", None, None, None, None, None, None, None, None, True),
        ("Spread / Giá HRC (%)", None, None, None, None, None, None, None, None, True),
    ]
    for i, (label, *vals, is_formula) in enumerate(price_data, r+1):
        ws14.cell(row=i, column=1, value=label).font = bold_font
        ws14.cell(row=i, column=1).border = thin_border
        for j, v in enumerate(vals, 2):
            cell = ws14.cell(row=i, column=j)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if is_formula and i == r+4:  # Spread row — link về '17_Gia_Hang_Hoa' Spread NĂM (lag 1 quý/năm,
                                          # xem code phía dưới sau khi sheet 17 được tạo), giữ chỗ ở đây trước.
                cell.number_format = '#,##0'
                cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
            elif is_formula and i == r+5:  # Spread/HRC % row
                cl = get_column_letter(j)
                cell.value = f"={cl}{r+4}/{cl}{r+1}*100"
                cell.number_format = '0.0'
            else:
                cell.value = v
                cell.number_format = '#,##0'
            cell.fill = assump_fill if not is_formula else PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # Section 2: Volume & Market Share
    r = r + 6
    ws14.cell(row=r, column=1, value="2. SẢN LƯỢNG & THỊ PHẦN").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    header_row(ws14, r, ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"], [40] + [12]*8)
    S_R3 = S_R3  # noqa: keep reference alive (defined at top)
    vol_data = [
        ("Sản lượng HRC HPG (triệu tấn)", True),
        ("Sản lượng thép XD HPG (triệu tấn)", True),
        ("Tổng sản lượng HPG (triệu tấn)", True),
        ("Sản lượng thị trường thép VN (triệu tấn)", False),
        ("Thị phần HPG (%)", True),
    ]
    for i, (label, is_ref) in enumerate(vol_data, r+1):
        ws14.cell(row=i, column=1, value=label).font = bold_font
        ws14.cell(row=i, column=1).border = thin_border
        for j in range(2, 10):
            cl = get_column_letter(j)
            cell = ws14.cell(row=i, column=j)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if is_ref:
                if i == r+1:  # SL HRC
                    cell.value = f"={S_R3}!{cl}4"
                elif i == r+2:  # SL XD
                    cell.value = f"={S_R3}!{cl}5"
                elif i == r+3:  # Tổng SL
                    cell.value = f"={cl}{r+1}+{cl}{r+2}"
                elif i == r+5:  # Thị phần
                    cell.value = f"={cl}{r+3}/{cl}{r+4}*100"
                    cell.number_format = '0.0'
                    cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
            if cell.value is None or not is_ref:
                if i == r+4:  # Thị trường VN
                    market_vals = [28, 27, 25, 26, 28, 30, 32, 34]
                    cell.value = market_vals[j-2]
                    cell.number_format = '#,##0'
                elif is_ref:
                    cell.number_format = '#,##0.0'

    # Section 3: Industry Factors
    r = r + 7
    ws14.cell(row=r, column=1, value="3. YẾU TỐ TÁC ĐỘNG NGÀNH THÉP").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    factors = [
        ("Đầu ra", "Giá HRC phụ thuộc nhu cầu TQ, thuế CBPG 27.8% bảo vệ thị trường nội địa"),
        ("Đầu vào", "Quặng sắt (62%Fe) và than cốc chiếm ~60% giá thành. HPG tự chủ 50% quặng"),
        ("Vĩ mô", "GDP VN ~6.5%, đầu tư công cao tốc Bắc-Nam + Long Thành kéo nhu cầu thép"),
        ("Cạnh tranh", "HPG dẫn đầu thị phần HRC (~70%) và thép XD (~40%). Đối thủ: HSG, NKG"),
        ("Rủi ro", "Giá HRC giảm, quặng/than tăng, tỷ giá USD, BĐS suy thoái"),
        ("Định giá", "P/B là công cụ chính (mua 0.7-1.0x, bán 2.0-2.5x). P/E là bẫy chu kỳ"),
    ]
    factor_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    r += 1
    ws14.cell(row=r, column=1, value="Yếu tố").font = bold_font
    ws14.cell(row=r, column=1).border = thin_border
    ws14.cell(row=r, column=1).fill = factor_fill
    ws14.cell(row=r, column=2, value="Đánh giá").font = bold_font
    ws14.cell(row=r, column=2).border = thin_border
    ws14.cell(row=r, column=2).fill = factor_fill
    ws14.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    for i, (factor, content) in enumerate(factors, r+1):
        ws14.cell(row=i, column=1, value=factor).font = bold_font
        ws14.cell(row=i, column=1).border = thin_border
        ws14.cell(row=i, column=2, value=content).font = data_font
        ws14.cell(row=i, column=2).border = thin_border
        ws14.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws14.merge_cells(start_row=i, start_column=2, end_row=i, end_column=9)
        ws14.row_dimensions[i].height = 30

    # ── Helper: get quarterly value ──
    def get_q(records, year, q, field):
        for r in records:
            if r.get("yearReport") == year and r.get("lengthReport") == q:
                v = r.get(field)
                return v / 1e9 if v is not None else 0
        return 0

    is_qs = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    bs_qs = section_to_quarters(FIN_DATA, "BALANCE_SHEET")

    # ── Section 4: Inventory, Receivables & Leverage Analysis ──
    r = r + 8
    ws14.cell(row=r, column=1, value=(
        "4. HÀNG TỒN KHO, PHẢI THU & ĐÒN BẨY — DIO/DSO = 365 x Số dư BÌNH QUÂN (đầu+cuối kỳ)/2 / GVHB hoặc DT "
        "(chỉ lịch sử — không dự phóng vì ít ảnh hưởng tới định giá)"
    )).font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    header_row(ws14, r, ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"], [40] + [12]*8)

    R_S4_HDR = r
    R_S4_INV, R_S4_TURN, R_S4_DIO, R_S4_REC, R_S4_DSO, R_S4_DE, R_S4_STDEBT = range(r + 1, r + 8)
    s4_labels = [
        (R_S4_INV,    "Hàng tồn kho (tỷ VND)"),
        (R_S4_TURN,   "Vòng quay HTK (lần) = GVHB / Tồn kho BQ"),
        (R_S4_DIO,    "Số ngày tồn kho bình quân (DIO)"),
        (R_S4_REC,    "Phải thu ngắn hạn (tỷ VND)"),
        (R_S4_DSO,    "Số ngày phải thu bình quân (DSO)"),
        (R_S4_DE,     "D/E = Vay NH+DH / VCSH"),
        (R_S4_STDEBT, "Vay ngắn hạn / Tổng vay (%)"),
    ]
    for row_i, label in s4_labels:
        ws14.cell(row=row_i, column=1, value=label).font = bold_font
        ws14.cell(row=row_i, column=1).border = thin_border

    # Excel formula SỐNG — link '04_PnL' (Doanh thu row2, GVHB row3) & '05_Balance_Sheet'
    # (Phải thu row5, Tồn kho row6, Vay NH row17, Vay DH row20, Tổng VCSH row28). Hàng tồn kho/
    # Vòng quay HTK/DIO/Phải thu/DSO CHỈ điền lịch sử (2021-2025, cột B-F) — bỏ dự phóng theo yêu
    # cầu user vì ít ảnh hưởng tới định giá. D/E & tỷ lệ vay NH vẫn điền đủ cả dự phóng (liên quan
    # trực tiếp đòn bẩy/rủi ro tài chính, có giá trị cho định giá).
    for j in range(2, 10):
        cl = get_column_letter(j)
        pcl = get_column_letter(j - 1)
        first_col = (j == 2)
        is_hist_col = (j <= 6)  # 2021A..2025A

        if is_hist_col:
            c = ws14.cell(row=R_S4_INV, column=j, value=f"={S_BS}!{cl}6")
            c.number_format = '#,##0'; c.border = thin_border; c.alignment = Alignment(horizontal='center'); c.font = data_font

            avg_inv = f"{S_BS}!{cl}6" if first_col else f"(({S_BS}!{pcl}6+{S_BS}!{cl}6)/2)"
            c = ws14.cell(row=R_S4_TURN, column=j, value=f"={S_PNL}!{cl}3/{avg_inv}")
            c.number_format = '0.0'; c.border = thin_border; c.alignment = Alignment(horizontal='center'); c.font = data_font

            c = ws14.cell(row=R_S4_DIO, column=j, value=f"=365/{cl}{R_S4_TURN}")
            c.number_format = '#,##0'; c.border = thin_border; c.alignment = Alignment(horizontal='center')
            c.font = Font(name=FONT_NAME, bold=True, color="C0392B")
            c.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

            c = ws14.cell(row=R_S4_REC, column=j, value=f"={S_BS}!{cl}5")
            c.number_format = '#,##0'; c.border = thin_border; c.alignment = Alignment(horizontal='center'); c.font = data_font

            avg_rec = f"{S_BS}!{cl}5" if first_col else f"(({S_BS}!{pcl}5+{S_BS}!{cl}5)/2)"
            c = ws14.cell(row=R_S4_DSO, column=j, value=f"=365/({S_PNL}!{cl}2/{avg_rec})")
            c.number_format = '#,##0'; c.border = thin_border; c.alignment = Alignment(horizontal='center')
            c.font = Font(name=FONT_NAME, bold=True, color="1F4E79")
            c.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        else:
            for rr in (R_S4_INV, R_S4_TURN, R_S4_DIO, R_S4_REC, R_S4_DSO):
                c = ws14.cell(row=rr, column=j, value="—")
                c.border = thin_border; c.alignment = Alignment(horizontal='center')
                c.font = Font(name=FONT_NAME, italic=True, size=9, color="AAAAAA")

        c = ws14.cell(row=R_S4_DE, column=j, value=f"=({S_BS}!{cl}17+{S_BS}!{cl}20)/{S_BS}!{cl}28")
        c.number_format = '0.00'; c.border = thin_border; c.alignment = Alignment(horizontal='center'); c.font = data_font

        c = ws14.cell(row=R_S4_STDEBT, column=j, value=f"={S_BS}!{cl}17/({S_BS}!{cl}17+{S_BS}!{cl}20)*100")
        c.number_format = '0.0'; c.border = thin_border; c.alignment = Alignment(horizontal='center'); c.font = data_font

    # ── Section 5: Cyclical Valuation — P/B vs P/E Trap ──
    r = r + 9
    ws14.cell(row=r, column=1, value="5. ĐỊNH GIÁ CHU KỲ — P/B vs P/E").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    header_row(ws14, r, ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "Nhận xét"], [40] + [12]*5 + [40])

    pbpe_data = [
        ("P/E (TTM cuối năm)", pe_hist + ["Bẫy chu kỳ: P/E thấp khi LN đỉnh"]),
        ("P/B (cuối năm)", pb_hist + ["Công cụ chính: mua <1.0x, bán >2.0x"]),
        ("ROE (%)", [round(ni_hist[i] / equity_hist[i] * 100, 1) for i in range(5)] + ["P/B = ROE × P/E"]),
    ]
    for i, (label, vals) in enumerate(pbpe_data, r + 1):
        ws14.cell(row=i, column=1, value=label).font = bold_font
        ws14.cell(row=i, column=1).border = thin_border
        for j, v in enumerate(vals, 2):
            cell = ws14.cell(row=i, column=j)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if j == 7:
                cell.value = v
                cell.font = Font(name=FONT_NAME, italic=True, size=9, color="C0392B")
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.value = v
                cell.number_format = '0.0' if label.startswith("ROE") else ('0.00' if label.startswith("P/B") else '0.0')
                cell.font = data_font

    # Highlight P/B < 1.0 = buy zone
    for j in range(2, 7):
        cell = ws14.cell(row=r + 2, column=j)
        if isinstance(cell.value, (int, float)) and 0 < cell.value < 1.2:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # ── Section 6: Global Overcapacity & Production ──
    sec14_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    r = r + 7
    ws14.cell(row=r, column=1, value="6. DƯ CUNG TOÀN CẦU & SẢN LƯỢNG THÉP THẾ GIỚI").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    # Capacity vs demand table
    ws14.cell(row=r, column=1, value="Công suất tối đa vs Nhu cầu tiêu thụ toàn cầu (triệu tấn)").font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    capdem_headers = ["Năm", "2019", "2020", "2021", "2022", "2023", "2024", "2025", "2026F", "2027F"]
    for c, h in enumerate(capdem_headers, 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    capdem = [
        ("Công suất tối đa", 2416, 2424, 2427, 2454, 2456, 2472, 2540, 2582, 2637),
        ("Nhu cầu tiêu thụ", 1890, 1858, 1962, 1895, 1891, 1870, 1888, 1902, 1916),
        ("Dư cung (triệu tấn)", 526, 566, 465, 559, 565, 602, 652, 680, 721),
        ("Dư cung (% công suất)", "21.8%", "23.3%", "19.2%", "22.8%", "23.0%", "24.4%", "25.7%", "26.3%", "27.3%"),
    ]
    for i, row_data in enumerate(capdem, r+1):
        for c, v in enumerate(row_data, 1):
            cell = ws14.cell(row=i, column=c, value=v)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.font = data_font
            if c == 1:
                cell.font = bold_font; cell.alignment = Alignment(horizontal='left')
            elif isinstance(v, str) and "%" in v:
                cell.value = float(v.replace("%",""))/100
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0'
    # BOF Inflexibility note
    r = r + len(capdem) + 1
    ws14.cell(row=r, column=1, value="BOF Inflexibility:").font = Font(name=FONT_NAME, bold=True, size=10, color="C0392B")
    ws14.cell(row=r, column=2, value="Lò cao BOF có chu kỳ vận hành liên tục 15-25 năm, không thể dừng lò khi giá giảm. Đây là nguyên nhân cấu trúc khiến dư cung kéo dài bất chấp giá thép xuống thấp.").font = data_font
    ws14.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws14.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    ws14.row_dimensions[r].height = 35
    # Global production share
    r += 2
    ws14.cell(row=r, column=1, value="Sản lượng thép toàn cầu 2025: 1,849 triệu tấn").font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    share_headers = ["Quốc gia", "TQ", "Ấn Độ", "EU", "Nhật Bản", "Hoa Kỳ", "Nga", "Còn lại"]
    share_vals = ["52.0% (961M tấn)", "8.9%", "6.8%", "4.4%", "4.4%", "3.7%", "19.8%"]
    for c, h in enumerate(share_headers, 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    for c, v in enumerate(share_vals, 1):
        cell = ws14.cell(row=r+1, column=c, value=v)
        cell.border = thin_border; cell.alignment = Alignment(horizontal='center'); cell.font = data_font

    # ── Section 7: FMS vs HPG Supply-Demand ──
    r = r + 4
    ws14.cell(row=r, column=1, value="7. CUNG - CẦU HRC NỘI ĐỊA (tấn)").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    for c, h in enumerate(["Chỉ tiêu", "2024", "2025", "2026F"], 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    fms_data = [
        ("SX bởi Formosa (FMS)", "5,000,000", "5,000,000", "5,000,000"),
        ("SX bởi Hòa Phát (HPG)", "3,000,000", "7,200,000", "8,600,000"),
        ("Tổng SX nội địa", "8,000,000", "12,200,000", "13,600,000"),
        ("Tổng nhu cầu HRC VN", "16,210,541", "16,476,361", "17,000,000"),
        ("Nhập khẩu ròng (bù thiếu)", "8,210,541", "4,276,361", "3,400,000"),
    ]
    for i, row_data in enumerate(fms_data, r+1):
        for c, v in enumerate(row_data, 1):
            cell = ws14.cell(row=i, column=c, value=v)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.font = bold_font if i == r+2 else data_font
        ws14.cell(row=i, column=1).font = bold_font
    r = r + len(fms_data) + 1
    ws14.cell(row=r, column=1, value="Kết luận:").font = Font(name=FONT_NAME, bold=True, size=10, color="C0392B")
    ws14.cell(row=r, column=2, value="HPG + FMS đã đáp ứng ~80% nhu cầu HRC nội địa (2026F), giảm dần phụ thuộc nhập khẩu. Với bảo hộ toàn phần (AD narrow + wide HRC), HPG là người hưởng lợi chính khi chiếm ~70% thị phần HRC nội địa.").font = data_font
    ws14.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
    ws14.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    ws14.row_dimensions[r].height = 35

    # ── Section 8: 8-sector demand breakdown ──
    r += 3
    ws14.cell(row=r, column=1, value="8. PHÂN BỔ NHU CẦU THÉP TOÀN CẦU THEO LĨNH VỰC").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    for c, h in enumerate(["Lĩnh vực", "Tỷ trọng", "Xu hướng"], 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    sector_data = [
        ("Xây dựng kết cấu, tòa nhà", "49-50%", "Giảm tỷ trọng dài hạn"),
        ("Kỹ thuật cơ khí", "16%", "Ổn định"),
        ("Sản phẩm kim loại", "11%", "Ổn định"),
        ("Ô tô", "8%", "Tăng dần (điện hóa)"),
        ("Dầu khí", "6%", "Phụ thuộc giá dầu"),
        ("Đóng tàu & đường sắt", "4%", "Tăng nhờ logistics"),
        ("Thiết bị gia dụng", "3%", "Ổn định"),
        ("Quốc phòng", "1%", "Ổn định"),
        ("Khác", "2%", "—"),
    ]
    for i, (sector, pct, trend) in enumerate(sector_data, r+1):
        ws14.cell(row=i, column=1, value=sector).font = bold_font; ws14.cell(row=i, column=1).border = thin_border
        ws14.cell(row=i, column=2, value=pct).font = data_font; ws14.cell(row=i, column=2).border = thin_border
        ws14.cell(row=i, column=2).alignment = Alignment(horizontal='center')
        ws14.cell(row=i, column=3, value=trend).font = data_font; ws14.cell(row=i, column=3).border = thin_border
        ws14.cell(row=i, column=3).alignment = Alignment(horizontal='center')

    # ── Section 9: Input Import & Export Restructuring ──
    r = r + len(sector_data) + 2
    ws14.cell(row=r, column=1, value="9. CƠ CẤU NHẬP KHẨU NGUYÊN LIỆU & TÁI CƠ CẤU THỊ TRƯỜNG").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    for c, h in enumerate(["Nguyên liệu", "Tỷ lệ NK", "Nguồn chính", "Rủi ro"], 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    input_data = [
        ("Than cốc", "~90%", "Úc, Nga, Indonesia", "Địa chính trị (Úc-TQ)"),
        ("Quặng sắt", "~52%", "Úc, Brazil", "Tập trung (Vale/BHP/Rio)"),
        ("Thép phế", "~47%", "Nhật Bản, ASEAN", "Giá biến động"),
    ]
    for i, (mat, pct, src, risk) in enumerate(input_data, r+1):
        ws14.cell(row=i, column=1, value=mat).font = bold_font; ws14.cell(row=i, column=1).border = thin_border
        ws14.cell(row=i, column=2, value=pct).font = data_font; ws14.cell(row=i, column=2).border = thin_border
        ws14.cell(row=i, column=3, value=src).font = data_font; ws14.cell(row=i, column=3).border = thin_border
        ws14.cell(row=i, column=4, value=risk).font = data_font; ws14.cell(row=i, column=4).border = thin_border
    # Export restructuring
    r = r + len(input_data) + 2
    ws14.cell(row=r, column=1, value="Chiến lược thị trường HPG 2025").font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    for c, h in enumerate(["Khu vực", "Tỷ trọng 2025", "So với 2024"], 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    export_data = [
        ("Nội địa", "84%", "+38% YoY"),
        ("Xuất khẩu (tổng)", "16%", "Giảm từ 31%"),
        ("  - Châu Á", "8%", "-57%"),
        ("  - Châu Âu", "4%", "-34%"),
        ("  - Châu Mỹ", "3%", "—"),
        ("  - Châu Úc", "1%", "—"),
        ("  - Châu Phi", "0%", "—"),
    ]
    for i, (region, share, change) in enumerate(export_data, r+1):
        ws14.cell(row=i, column=1, value=region).font = bold_font if not region.startswith("  ") else data_font
        ws14.cell(row=i, column=1).border = thin_border
        ws14.cell(row=i, column=2, value=share).font = data_font; ws14.cell(row=i, column=2).border = thin_border
        ws14.cell(row=i, column=3, value=change).font = data_font; ws14.cell(row=i, column=3).border = thin_border
        for c in range(1,4):
            ws14.cell(row=i, column=c).alignment = Alignment(horizontal='center')

    # ── Section 10: Cost Structure ──
    r = r + len(export_data) + 2
    ws14.cell(row=r, column=1, value="10. CƠ CẤU CHI PHÍ SẢN XUẤT THÉP (BOF)").font = sec_font
    ws14.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    for c, h in enumerate(["Yếu tố chi phí", "Tỷ trọng", "Ghi chú"], 1):
        cell = ws14.cell(row=r, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    cost_data = [
        ("Quặng sắt", "~35-40%", "Phụ thuộc giá 62%Fe. HPG tự chủ 50%"),
        ("Than cốc", "~20-25%", "Phụ thuộc giá than Úc. DQ2 tiết kiệm 15%"),
        ("Điện năng", "~8-10%", "DQ2 tiết kiệm hơn DQ1"),
        ("Khấu hao", "~8-12%", "DQ2 làm tăng KH nhưng giảm/tấn"),
        ("Nhân công & Bảo trì", "~5-8%", "Tự động hóa cao"),
        ("Logistics & Khác", "~10-15%", "DQ2 cạnh cảng giảm chi phí"),
    ]
    for i, (factor, pct, note) in enumerate(cost_data, r+1):
        ws14.cell(row=i, column=1, value=factor).font = bold_font; ws14.cell(row=i, column=1).border = thin_border
        ws14.cell(row=i, column=2, value=pct).font = data_font; ws14.cell(row=i, column=2).border = thin_border
        ws14.cell(row=i, column=2).alignment = Alignment(horizontal='center')
        ws14.cell(row=i, column=3, value=note).font = data_font; ws14.cell(row=i, column=3).border = thin_border
        ws14.cell(row=i, column=3).alignment = Alignment(wrap_text=True)

    # ── Sheet 15: Quarterly Financial Data ──
    ws15 = wb.create_sheet("15_Quarterly_Data")
    def get_q(records, yr, qtr, field):
        for r in records:
            if r.get("yearReport") == yr and r.get("lengthReport") == qtr:
                v = r.get(field)
                return v / 1e9 if v is not None else 0
        return 0
    is_qs = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    bs_qs = section_to_quarters(FIN_DATA, "BALANCE_SHEET")
    cf_qs = section_to_quarters(FIN_DATA, "CASH_FLOW")

    all_quarters = [(y, q) for y in range(2018, 2027) for q in range(1, 5)
                    if any(r.get("yearReport") == y and r.get("lengthReport") == q for r in is_qs)]
    q_labels = [f"Q{q}/{y}" for y, q in all_quarters]
    nq = len(all_quarters)

    # Header
    ws15.cell(row=1, column=1, value="DỮ LIỆU TÀI CHÍNH QUÝ HPG (2018-2026)").font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E79")
    ws15.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nq+1)

    # ── IS Section ──
    r2 = 3
    ws15.cell(row=r2, column=1, value="A. KẾT QUẢ KINH DOANH (tỷ VND)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws15.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=nq+1)
    r2 += 1
    for c, l in enumerate([""] + q_labels, 1):
        cell = ws15.cell(row=r2, column=c, value=l)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', text_rotation=90 if c > 1 else 0)

    is_rows = [
        ("Doanh thu thuần", "isa3"),
        ("Giá vốn hàng bán", lambda yr,q: abs(get_q(is_qs, yr, q, "isa4"))),
        ("Lợi nhuận gộp", "isa5"),
        ("Biên LNG %", lambda yr,q: round((get_q(is_qs, yr, q, "isa5")) / max(get_q(is_qs, yr, q, "isa3"), 1) * 100, 1)),
        ("CP BH & QLDN", lambda yr,q: abs(get_q(is_qs, yr, q, "isa9")) + abs(get_q(is_qs, yr, q, "isa10"))),
        ("LN từ HĐKD", "isa11"),
        ("Doanh thu TC", "isa6"),
        ("Chi phí TC", lambda yr,q: abs(get_q(is_qs, yr, q, "isa7"))),
        ("LNTT", "isa16"),
        ("LNST", "isa20"),
        ("LNST CĐCTM", "isa22"),
    ]
    for i, (label, field) in enumerate(is_rows, r2+1):
        ws15.cell(row=i, column=1, value=label).font = bold_font
        ws15.cell(row=i, column=1).border = thin_border
        for j, (yr, q) in enumerate(all_quarters, 2):
            cell = ws15.cell(row=i, column=j)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center'); cell.font = data_font
            if callable(field):
                cell.value = field(yr, q)
            else:
                cell.value = get_q(is_qs, yr, q, field)
            if "Biên" in label:
                cell.number_format = '0.0'
            else:
                cell.number_format = '#,##0'
        ws15.cell(row=i, column=1).font = bold_font if not label.startswith("  ") else data_font

    # ── BS Section ──
    r3 = r2 + len(is_rows) + 2
    ws15.cell(row=r3, column=1, value="B. BẢNG CÂN ĐỐI KẾ TOÁN (tỷ VND)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws15.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=nq+1)
    r3 += 1
    for c, l in enumerate([""] + q_labels, 1):
        cell = ws15.cell(row=r3, column=c, value=l)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', text_rotation=90 if c > 1 else 0)

    bs_rows = [
        ("Tiền & tương đương", "bsa2"),
        ("Tổng tài sản", "bsa53"),
        ("Vốn chủ sở hữu", "bsa78"),
        ("Hàng tồn kho", "bsa15"),
        ("Phải thu NH", "bsa8"),
        ("Vay ngắn hạn", "bsa56"),
        ("Vay dài hạn", "bsa71"),
        ("Tổng nợ vay", lambda yr,q: get_q(bs_qs, yr, q, "bsa56") + get_q(bs_qs, yr, q, "bsa71")),
        ("TSCĐ hữu hình", "bsa30"),
        ("XDCB DD (CIP)", "bsa188"),
    ]
    for i, (label, field) in enumerate(bs_rows, r3+1):
        ws15.cell(row=i, column=1, value=label).font = bold_font
        ws15.cell(row=i, column=1).border = thin_border
        for j, (yr, q) in enumerate(all_quarters, 2):
            cell = ws15.cell(row=i, column=j)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center'); cell.font = data_font
            if callable(field):
                cell.value = field(yr, q)
            else:
                cell.value = get_q(bs_qs, yr, q, field)
            cell.number_format = '#,##0'

    # ── CF Section ──
    r4 = r3 + len(bs_rows) + 2
    ws15.cell(row=r4, column=1, value="C. LƯU CHUYỂN TIỀN TỆ (tỷ VND)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws15.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=nq+1)
    r4 += 1
    for c, l in enumerate([""] + q_labels, 1):
        cell = ws15.cell(row=r4, column=c, value=l)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', text_rotation=90 if c > 1 else 0)

    cf_rows = [
        ("CF từ HĐKD (CFO)", "cfa18"),
        ("CF đầu tư (CFI)", "cfa26"),
        ("CF tài trợ (CFF)", "cfa34"),
        ("Khấu hao TSCĐ", "cfa2"),
        ("CAPEX", lambda yr,q: abs(get_q(cf_qs, yr, q, "cfa19"))),
    ]
    for i, (label, field) in enumerate(cf_rows, r4+1):
        ws15.cell(row=i, column=1, value=label).font = bold_font
        ws15.cell(row=i, column=1).border = thin_border
        for j, (yr, q) in enumerate(all_quarters, 2):
            cell = ws15.cell(row=i, column=j)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center'); cell.font = data_font
            if callable(field):
                cell.value = field(yr, q)
            else:
                cell.value = get_q(cf_qs, yr, q, field)
            cell.number_format = '#,##0'

    # ── Q1 2026 Highlight Box ──
    r5 = r4 + len(cf_rows) + 2
    ws15.cell(row=r5, column=1, value="ĐIỂM NHẤN Q1/2026").font = Font(name=FONT_NAME, bold=True, size=12, color="C0392B")
    ws15.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=nq+1)
    q1_2026 = [
        ("Doanh thu Q1/2026", f"{get_q(is_qs, 2026, 1, 'isa3'):,.0f} tỷ", f"Q1/2025: {get_q(is_qs, 2025, 1, 'isa3'):,.0f} tỷ", f"+{get_q(is_qs, 2026, 1, 'isa3')/get_q(is_qs, 2025, 1, 'isa3')*100-100:.1f}% YoY"),
        ("Biên LNG Q1/2026", f"{get_q(is_qs, 2026, 1, 'isa5')/get_q(is_qs, 2026, 1, 'isa3')*100:.1f}%", f"Q1/2025: {get_q(is_qs, 2025, 1, 'isa5')/get_q(is_qs, 2025, 1, 'isa3')*100:.1f}%", f"+{get_q(is_qs, 2026, 1, 'isa5')/get_q(is_qs, 2026, 1, 'isa3')*100 - get_q(is_qs, 2025, 1, 'isa5')/get_q(is_qs, 2025, 1, 'isa3')*100:.1f} ppt"),
        ("LNST Q1/2026", f"{get_q(is_qs, 2026, 1, 'isa22'):,.0f} tỷ", f"Q1/2025: {get_q(is_qs, 2025, 1, 'isa22'):,.0f} tỷ", f"+{get_q(is_qs, 2026, 1, 'isa22')/get_q(is_qs, 2025, 1, 'isa22')*100-100:.1f}% YoY"),
        ("% KH 2026E", f"{get_q(is_qs, 2026, 1, 'isa22')/ni_fc[0]*100:.1f}%", f"KH 2026E NI: {ni_fc[0]:,.0f} tỷ", "Q1 đã đạt gần nửa kế hoạch năm"),
    ]
    r5 += 1
    for c, h in enumerate(["Chỉ tiêu", "Q1/2026", "Cùng kỳ", "Tăng trưởng"], 1):
        cell = ws15.cell(row=r5, column=c, value=h)
        cell.font = header_font; cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    for i, row_data in enumerate(q1_2026, r5+1):
        for c, v in enumerate(row_data, 1):
            cell = ws15.cell(row=i, column=c, value=v)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.font = bold_font if c == 1 else data_font
        ws15.merge_cells(start_row=i, start_column=4, end_row=i, end_column=min(6, nq+1))
        ws15.row_dimensions[i].height = 25

    # ── Quarterly Trend Analysis ──
    r5 = r5 + len(q1_2026) + 2
    ws15.cell(row=r5, column=1, value="PHÂN TÍCH XU HƯỚNG QUÝ").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E79")
    ws15.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=nq+1)
    r5 += 1
    trend_rows = [
        ("DOANH THU", f"Tăng mạnh Q1/2026: {get_q(is_qs,2026,1,'isa3'):,.0f} tỷ (+{get_q(is_qs,2026,1,'isa3')/get_q(is_qs,2025,1,'isa3')*100-100:.1f}% YoY). Xu hướng tăng từ Q1/2025 nhờ DQ2 dần full công suất. Kỳ vọng các quý sau cao hơn do mùa cao điểm xây dựng Q2-Q4.", "TÍCH CỰC"),
        ("BIÊN LỢI NHUẬN", f"GM Q1/2026: {get_q(is_qs,2026,1,'isa5')/get_q(is_qs,2026,1,'isa3')*100:.1f}% -- còn thấp so với đỉnh 2021 (27.5%) nhưng cải thiện dần. Nguyên nhân: DQ2 chạy non tải Q1, giá quặng giảm chưa kịp phản ánh. Kỳ vọng 17-18% từ Q2 khi DQ2 full + tồn kho giá rẻ.", "PHỤC HỒI"),
        ("DÒNG TIỀN (CFO)", f"CFO Q1/2026 ước ~{get_q(cf_qs,2026,1,'cfa18'):,.0f} tỷ. CFO/LNST duy trì >1.0 qua các quý -- chất lượng lợi nhuận tốt. Dòng tiền mạnh giúp HPG giảm đòn bẩy.", "MẠNH"),
        ("CAPEX", f"CAPEX Q1/2026: {abs(get_q(cf_qs,2026,1,'cfa19')):,.0f} tỷ. Đã giảm mạnh so với đỉnh xây dựng DQ2 (2023-2024). Dự báo tiếp tục giảm khi DQ2 hoàn tất → FCF chuyển dương mạnh.", "GIẢM TỐT"),
        ("TỒN KHO & PHẢI THU", f"Tồn kho {get_q(bs_qs,2026,1,'bsa15'):,.0f} tỷ, phải thu {get_q(bs_qs,2026,1,'bsa8'):,.0f} tỷ, vòng quay ~{get_q(is_qs,2026,1,'isa3')/4 / max(get_q(bs_qs,2026,1,'bsa15'),1)*365:.0f} ngày. Quản trị vốn lưu động hiệu quả, không có dấu hiệu bán hàng ép.", "ỔN ĐỊNH"),
        ("DƯ NỢ VAY", f"Tổng nợ Q1/2026: {get_q(bs_qs,2026,1,'bsa56')+get_q(bs_qs,2026,1,'bsa71'):,.0f} tỷ. D/E ~{(get_q(bs_qs,2026,1,'bsa56')+get_q(bs_qs,2026,1,'bsa71'))/max(get_q(bs_qs,2026,1,'bsa78'),1):.2f}x. DQ2 đã xong → không cần vay thêm lớn, nợ sẽ giảm dần nhờ dòng tiền.", "AN TOÀN"),
    ]
    for c, h in {1: "Chỉ tiêu", 2: "Đánh giá", min(5, nq+1): "Trạng thái"}.items():
        cell = ws15.cell(row=r5, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    ws15.merge_cells(start_row=r5, start_column=2, end_row=r5, end_column=min(4, nq+1))
    trend_green = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    trend_yellow = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    for i, (label, evaluation, status) in enumerate(trend_rows, r5+1):
        ws15.cell(row=i, column=1, value=label).font = bold_font; ws15.cell(row=i, column=1).border = thin_border
        ws15.cell(row=i, column=2, value=evaluation).font = data_font
        ws15.cell(row=i, column=2).border = thin_border; ws15.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws15.merge_cells(start_row=i, start_column=2, end_row=i, end_column=min(4, nq+1))
        status_col = min(5, nq+1)
        ws15.cell(row=i, column=status_col, value=status).font = Font(name=FONT_NAME, bold=True, size=10, color="006600" if "TÍCH CỰC" in status or "MẠNH" in status or "AN TOÀN" in status else "E65100")
        ws15.cell(row=i, column=status_col).border = thin_border
        ws15.cell(row=i, column=status_col).fill = trend_green if "TÍCH CỰC" in status or "MẠNH" in status or "AN TOÀN" in status else trend_yellow
        ws15.cell(row=i, column=status_col).alignment = Alignment(horizontal='center')
        ws15.row_dimensions[i].height = 55

    # ── Quarterly Sales Volume Section (in ws15) — nguồn duy nhất SALES_Q_LABELS/HRC_SALES_HIST_KT/
    # XD_SALES_HIST_KT (module-level), dùng chung với JSON dashboard, tránh 2 mảng trùng lặp lệch số
    # như trước (hrc_data/xd_data ở đây và hrc_sales/xd_sales ở JSON export).
    qv_start = r5 + 10
    qv_headers = [f"{lbl[4:]}/{lbl[:4]}" for lbl in SALES_Q_LABELS]
    hrc_data = list(HRC_SALES_HIST_KT)
    xd_data = list(XD_SALES_HIST_KT)
    # Nếu quý đang chạy đã có số liệu (chính thức hoặc ước tính từ tháng) → thêm 1 cột vào bảng. Ưu
    # tiên tách HRC/XD TRỰC TIẾP từ nguồn (CUR_Q_HRC_KT_DIRECT/XD, từ nguoiquansat/Vietcap Research)
    # khi có; nếu không, tách theo tỷ lệ Q1/2026 thực tế (xem fetch_hpg_production_updates).
    if CUR_Q_TOTAL_KT is not None:
        qv_headers.append(f"{CUR_Q_NUM}/{CUR_Q_YEAR}*")
        if CUR_Q_HRC_KT_DIRECT is not None and CUR_Q_XD_KT_DIRECT is not None:
            hrc_data.append(CUR_Q_HRC_KT_DIRECT)
            xd_data.append(CUR_Q_XD_KT_DIRECT)
        else:
            _ratio = HRC_SALES_HIST_KT[-1] / (HRC_SALES_HIST_KT[-1] + XD_SALES_HIST_KT[-1])
            hrc_data.append(round(CUR_Q_TOTAL_KT * _ratio, 1))
            xd_data.append(round(CUR_Q_TOTAL_KT * (1 - _ratio), 1))
    qv_ncol = len(qv_headers) + 1  # +1 for label column
    ws15.cell(row=qv_start, column=1, value="D. SẢN LƯỢNG TIÊU THỤ QUÝ (nghìn tấn)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws15.merge_cells(start_row=qv_start, start_column=1, end_row=qv_start, end_column=qv_ncol)
    qv_start += 1
    if CUR_Q_TOTAL_KT is not None:
        _src_note = {
            "OFFICIAL": "số liệu CHÍNH THỨC",
            "ESTIMATED": "ƯỚC TÍNH từ số liệu tháng đã công bố (xem mục E bên dưới)",
            "CACHED": "ƯỚC TÍNH LẤY LẠI từ lần chạy trước (không fetch được tin mới lần này)",
        }[CUR_Q_SOURCE]
        _split_note = ("tách HRC/XD trực tiếp từ nguồn (nguoiquansat.vn/Vietcap Research)"
                       if CUR_Q_HRC_KT_DIRECT is not None else "tách HRC/XD theo tỷ lệ Q1/2026 thực tế")
        ws15.cell(row=qv_start, column=1, value=f"(*) {CUR_Q_NUM}/{CUR_Q_YEAR} = {_src_note}, {_split_note} — tự động dò tin khi chạy script").font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
        ws15.merge_cells(start_row=qv_start, start_column=1, end_row=qv_start, end_column=qv_ncol)
        qv_start += 1
    for c, h in enumerate([""] + qv_headers, 1):
        cell = ws15.cell(row=qv_start, column=c, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, size=8, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', text_rotation=60 if c > 1 else 0)
    qv_data_rows = [
        ("SL HRC (nghìn tấn)", hrc_data),
        ("SL XD & Thép CB (nghìn tấn)", xd_data),
    ]
    R_QV_HRC, R_QV_XD = qv_start + 1, qv_start + 2
    COL_Q1_2026 = 2 + SALES_Q_LABELS.index("2026Q1")
    COL_CUR_Q = qv_ncol if CUR_Q_TOTAL_KT is not None else None
    qv_data_end = qv_start + len(qv_data_rows)
    for i, (label, vals) in enumerate(qv_data_rows, qv_start+1):
        ws15.cell(row=i, column=1, value=label).font = bold_font
        ws15.cell(row=i, column=1).border = thin_border
        for j, v in enumerate(vals, 2):
            cell = ws15.cell(row=i, column=j, value=v)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '#,##0'
            if j == qv_ncol and CUR_Q_TOTAL_KT is not None:
                cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
                cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # ── Mục E: Cập nhật sản lượng theo THÁNG cho quý đang chạy (khi chưa có số quý chính thức) ──
    # Vị trí "khác" theo yêu cầu user: các tháng đã có tin công bố (nếu có) ghi riêng ra đây kèm
    # link bài viết để tự kiểm chứng, và công thức SUM/AVERAGE sống tính ra ước tính quý ở dòng cuối —
    # không phải số Python dán sẵn.
    qe_start = qv_data_end + 16
    ws15.cell(row=qe_start, column=1, value=(
        f"E. CẬP NHẬT SẢN LƯỢNG THEO THÁNG — {CUR_Q_LABEL[4:]}/{CUR_Q_LABEL[:4]} (quý đang chạy, tự động dò tin hoaphat.com.vn & nguoiquansat.vn)"
    )).font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws15.merge_cells(start_row=qe_start, start_column=1, end_row=qe_start, end_column=5)
    qe_start += 1
    ws15.cell(row=qe_start, column=1, value=(
        "Khi HPG công bố số liệu THÁNG (chưa có số quý chính thức), điền vào đây; dòng \"Ước tính quý\" "
        "bên dưới = AVERAGE các tháng đã biết × 3 (công thức sống). Khi có số liệu QUÝ chính thức, số đó "
        "được dùng trực tiếp ở mục D bên trên, không cần điền mục này nữa."
    )).font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
    ws15.merge_cells(start_row=qe_start, start_column=1, end_row=qe_start, end_column=8)
    ws15.row_dimensions[qe_start].height = 26
    qe_start += 1
    header_row(ws15, qe_start,
        ["Tháng", "SL HRC (nghìn tấn)", "SL XD (nghìn tấn)", "Tổng (nghìn tấn)", "Ngày đăng", "Nguồn (link bài viết)"],
        [12, 16, 16, 16, 14, 70])
    R_QE_HDR = qe_start
    month_names_vn = ["", "Tháng 1", "Tháng 2", "Tháng 3", "Tháng 4", "Tháng 5", "Tháng 6",
                      "Tháng 7", "Tháng 8", "Tháng 9", "Tháng 10", "Tháng 11", "Tháng 12"]
    _cur_q_month_recs_by_num = {r["month"]: r for r in CUR_Q_MONTHLY_RECS}
    for i, mo in enumerate(_cur_q_months, R_QE_HDR + 1):
        rec = _cur_q_month_recs_by_num.get(mo)
        ws15.cell(row=i, column=1, value=month_names_vn[mo]).font = bold_font
        ws15.cell(row=i, column=1).border = thin_border
        for col, key in ((2, "hrc_kt"), (3, "xd_kt")):
            c = ws15.cell(row=i, column=col, value=(rec.get(key) if rec else None))
            c.border = thin_border; c.alignment = Alignment(horizontal='center'); c.number_format = '#,##0'
            if rec and rec.get(key) is not None:
                c.font = Font(name=FONT_NAME, bold=True, color="C0392B")
            else:
                c.value = "-" if rec else "Chưa có tin"
                c.font = Font(name=FONT_NAME, italic=True, size=9, color="AAAAAA")
        tot_cell = ws15.cell(row=i, column=4, value=(rec["volume_kt"] if rec else "Chưa có tin"))
        tot_cell.border = thin_border; tot_cell.alignment = Alignment(horizontal='center')
        tot_cell.number_format = '#,##0'
        tot_cell.font = Font(name=FONT_NAME, bold=True, color="C0392B") if rec else Font(name=FONT_NAME, italic=True, size=9, color="AAAAAA")
        date_cell = ws15.cell(row=i, column=5, value=rec["date"] if rec else "-")
        date_cell.font = data_font; date_cell.border = thin_border; date_cell.alignment = Alignment(horizontal='center')
        link_cell = ws15.cell(row=i, column=6, value=rec["url"] if rec else "-")
        link_cell.font = Font(name=FONT_NAME, size=8, color="1155CC", underline="single") if rec else data_font
        link_cell.border = thin_border
        if rec:
            link_cell.hyperlink = rec["url"]
    R_QE_LAST = R_QE_HDR + len(_cur_q_months)
    r_qe_est = R_QE_LAST + 1
    ws15.cell(row=r_qe_est, column=1, value=f"Ước tính SL {CUR_Q_LABEL} (nghìn tấn)").font = Font(name=FONT_NAME, bold=True, color="1F4E79")
    ws15.cell(row=r_qe_est, column=1).border = thin_border
    for col in (2, 3, 4):
        cl = get_column_letter(col)
        est_cell = ws15.cell(row=r_qe_est, column=col,
            value=f"=IFERROR(AVERAGE({cl}{R_QE_HDR+1}:{cl}{R_QE_LAST})*3,\"-\")")
        est_cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
        est_cell.border = thin_border; est_cell.alignment = Alignment(horizontal='center')
        est_cell.number_format = '#,##0'
    ws15.cell(row=r_qe_est, column=6, value=(
        "= TB(các tháng đã có tin, cột tương ứng) × 3 — công thức sống, tự cập nhật khi điền thêm tháng mới ở trên"
    )).font = Font(name=FONT_NAME, italic=True, size=8, color="888888")

    # ── Mục F: Dự phóng sản lượng NĂM 2026E (run-rate 2 quý đã biết) — Excel formula sống, link về
    # từ 03_Revenue_Model để ra doanh thu/LNST dự phóng, theo đúng yêu cầu user "link công thức tính
    # toán để dự báo LNST năm và quý này" ──
    r_f = r_qe_est + 3
    ws15.cell(row=r_f, column=1, value="F. DỰ PHÓNG SẢN LƯỢNG NĂM 2026E (run-rate 2 quý đã biết)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws15.merge_cells(start_row=r_f, start_column=1, end_row=r_f, end_column=5)
    r_f += 1
    ws15.cell(row=r_f, column=1, value=(
        f"Sản lượng năm 2026E = (SL Q1/2026 + SL {CUR_Q_LABEL}) / 2 × 4 — annualize theo run-rate 2 quý đã biết. "
        "03_Revenue_Model!G4/G5 (Sản lượng HRC/XD 2026E) LINK trực tiếp về đây."
    )).font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
    ws15.merge_cells(start_row=r_f, start_column=1, end_row=r_f, end_column=8)
    ws15.row_dimensions[r_f].height = 26
    r_f += 1
    q1_cl, curq_cl = get_column_letter(COL_Q1_2026), (get_column_letter(COL_CUR_Q) if COL_CUR_Q else None)
    R_F_HRC, R_F_XD = r_f, r_f + 1
    for row_i, label, src_row, fallback in ((R_F_HRC, "SL HRC năm 2026E (triệu tấn)", R_QV_HRC, SL_HRC_A[5]),
                                              (R_F_XD, "SL XD năm 2026E (triệu tấn)", R_QV_XD, SL_XD_A[5])):
        ws15.cell(row=row_i, column=1, value=label).font = bold_font
        ws15.cell(row=row_i, column=1).border = thin_border
        c = ws15.cell(row=row_i, column=2)
        if curq_cl:
            c.value = f"=({q1_cl}{src_row}+{curq_cl}{src_row})/2*4/1000"
        else:
            c.value = fallback  # chưa có dữ liệu quý đang chạy -> giữ giả định cũ
        c.number_format = '0.00'
        c.font = Font(name=FONT_NAME, bold=True, color="C0392B")
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        ws15.cell(row=row_i, column=4, value=(
            f"={q1_cl}{src_row}/1000 & \" (Q1) + \" & {curq_cl}{src_row}/1000 & \" ({CUR_Q_LABEL}), triệu tấn\""
            if curq_cl else f"Chưa có dữ liệu {CUR_Q_LABEL} — giữ giả định {fallback} triệu tấn"
        )).font = Font(name=FONT_NAME, italic=True, size=8, color="888888")

    # ── Openpyxl LineChart: Sản lượng HRC & XD theo quý ──
    chart_qv = LineChart()
    chart_qv.title = "Sản lượng HRC & Thép XD theo Quý (nghìn tấn)"
    chart_qv.width = 26; chart_qv.height = 14
    chart_qv.y_axis.title = "nghìn tấn"
    chart_qv.y_axis.scaling.min = 0
    chart_qv.y_axis.scaling.max = 2200
    chart_qv.y_axis.majorUnit = 500
    chart_qv.y_axis.numFmt = '#,##0'
    chart_qv.style = 10
    data_ref = Reference(ws15, min_col=2, max_col=qv_ncol, min_row=qv_start+1, max_row=qv_data_end)
    cats_ref = Reference(ws15, min_col=2, max_col=qv_ncol, min_row=qv_start+1)
    chart_qv.add_data(data_ref, titles_from_data=True)
    chart_qv.set_categories(cats_ref)
    if len(chart_qv.series) >= 2:
        chart_qv.series[0].graphicalProperties.line.solidFill = "1F4E79"
        chart_qv.series[0].graphicalProperties.line.width = 22000
        chart_qv.series[1].graphicalProperties.line.solidFill = "E74C3C"
        chart_qv.series[1].graphicalProperties.line.width = 22000
        chart_qv.series[1].graphicalProperties.line.dashStyle = "dash"
    chart_qv.legend.position = "b"
    chart_qv.x_axis.delete = False
    chart_qv.x_axis.tickLblPos = "low"
    chart_qv.x_axis.numFmt = '@'
    chart_row = qv_data_end + 2
    ws15.add_chart(chart_qv, f"A{chart_row}")

    # ── Link 03_Revenue_Model!G4/G5 (Sản lượng HRC/XD 2026E) về sheet 15 mục F (run-rate 2 quý đã
    # biết) thay vì số Python dán sẵn — theo yêu cầu user: "link công thức tính toán để dự báo LNST
    # năm và quý này". Toàn bộ chuỗi Doanh thu (03_Revenue_Model→04_PnL) đã là formula sống cho các
    # cột dự phóng nên chỉ cần sửa ĐÚNG Ô GỐC này là LNST 2026E tự cập nhật theo sản lượng mới.
    S15 = "'15_Quarterly_Data'"
    g4 = ws3.cell(row=4, column=7, value=f"={S15}!B{R_F_HRC}")
    g4.number_format = '#,##0.00'; g4.font = Font(name=FONT_NAME, bold=True, color="C0392B")
    g5 = ws3.cell(row=5, column=7, value=f"={S15}!B{R_F_XD}")
    g5.number_format = '#,##0.00'; g5.font = Font(name=FONT_NAME, bold=True, color="C0392B")

    # ── Sheet 16: Steel Accounting & Harvest Signs ──
    ws16 = wb.create_sheet("16_Steel_Accounting")
    ws16.cell(row=1, column=1, value="THỦ THUẬT KẾ TOÁN & DẤU HIỆU NGÀNH THÉP").font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E79")
    ws16.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws16_sec = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")

    # Section 1: Thủ thuật kế toán
    r6 = 3
    ws16.cell(row=r6, column=1, value="1. THỦ THUẬT KẾ TOÁN THÉP").font = ws16_sec
    ws16.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=7)
    r6 += 1
    header_row(ws16, r6, ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "Cảnh báo"], [40] + [12]*5 + [30])
    acc_data = [
        ("Dự phòng giảm giá HTK (tỷ)", None, None, None, None, None, "Check: NOTE→Chi phí SXKD→Dự phòng HTK"),
        ("Phải thu NH (tỷ)", round(receivables_hist[0]), round(receivables_hist[1]), round(receivables_hist[2]), round(receivables_hist[3]), round(receivables_hist[4]), "Nếu tăng mạnh + CFO âm → bán hàng sân sau"),
        ("CFO (tỷ)", round(cfo_hist[0]), round(cfo_hist[1]), round(cfo_hist[2]), round(cfo_hist[3]), round(cfo_hist[4]), "CFO dương → chất lượng LN tốt"),
        ("CFO / LNST (lần)", round(cfo_hist[0]/ni_hist[0], 2), round(cfo_hist[1]/ni_hist[1], 2), round(cfo_hist[2]/ni_hist[2], 2), round(cfo_hist[3]/ni_hist[3], 2), round(cfo_hist[4]/ni_hist[4], 2), "<0.8 liên tục → cảnh báo LN ảo"),
        ("Tồn kho / Tổng TS (%)", round(inventory_hist[0]/total_assets_hist[0]*100, 1),
                                  round(inventory_hist[1]/total_assets_hist[1]*100, 1),
                                  round(inventory_hist[2]/total_assets_hist[2]*100, 1),
                                  round(inventory_hist[3]/total_assets_hist[3]*100, 1),
                                  round(inventory_hist[4]/total_assets_hist[4]*100, 1),
                                  ">30% = rủi ro tồn kho cao"),
    ]
    acc_warn = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for i, row_data in enumerate(acc_data, r6 + 1):
        for c, v in enumerate(row_data, 1):
            cell = ws16.cell(row=i, column=c)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            if c == 7:
                cell.font = Font(name=FONT_NAME, italic=True, size=9, color="C0392B")
                cell.alignment = Alignment(wrap_text=True, horizontal='left'); cell.fill = acc_warn
            elif c == 1:
                cell.font = bold_font; cell.alignment = Alignment(horizontal='left')
            else:
                cell.font = data_font
                cell.value = v
                cell.number_format = '0.0' if isinstance(v, float) and abs(v) < 10 else '#,##0'
        ws16.row_dimensions[i].height = 25

    # Section 2: Dấu hiệu hái quả ngọt
    r7 = r6 + len(acc_data) + 2
    ws16.cell(row=r7, column=1, value="2. DẤU HIỆU SẮP HÁI QUẢ NGỌT").font = ws16_sec
    ws16.merge_cells(start_row=r7, start_column=1, end_row=r7, end_column=7)
    r7 += 1
    for c, h in enumerate(["Dấu hiệu", "Mô tả", "Trạng thái"], 1):
        cell = ws16.cell(row=r7, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    harvest_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    harvest_data = [
        ("1. Vĩ mô đảo chiều", "BĐS ấm lại, giải ngân đầu tư công tăng 22%, TQ kích thích 1,000 tỷ NDT", "✅ ĐÃ CÓ TÍN HIỆU"),
        ("2. Spread nở ra", f"Giá HRC {HRC_PRICE_A[4]:,.0f}→{HRC_PRICE_A[5]:,.0f} USD/tấn. Spread {SPREAD_A[5]:,.0f} USD/tấn ({years_fc[0]}E) từ {SPREAD_A[4]:,.0f} USD/tấn ({years_hist[4]})", "✅ ĐANG XẢY RA" if SPREAD_A[5] > SPREAD_A[4] else "⚠️ THEO DÕI"),
        ("3. Hàng tồn kho giá rẻ", f"Quặng giá thấp H2/2025 → tồn kho rẻ. Biên GP Q1/2026 = {get_q(is_qs,2026,1,'isa5')/get_q(is_qs,2026,1,'isa3')*100:.1f}% (cao hơn 2025)", "✅ ĐANG XẢY RA"),
        ("4. Nhà máy mới (DQ2)", f"CIP giảm từ 63,656 tỷ (2024) → {cip_hist[4]:,.0f} tỷ (2025). DQ2 đã chạy full từ T12/2025", "✅ ĐÃ HOÀN THÀNH"),
    ]
    for i, (sign, desc, status) in enumerate(harvest_data, r7+1):
        ws16.cell(row=i, column=1, value=sign).font = bold_font; ws16.cell(row=i, column=1).border = thin_border
        ws16.cell(row=i, column=2, value=desc).font = data_font; ws16.cell(row=i, column=2).border = thin_border
        ws16.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws16.cell(row=i, column=3, value=status).font = Font(name=FONT_NAME, bold=True, size=10, color="006600")
        ws16.cell(row=i, column=3).border = thin_border; ws16.cell(row=i, column=3).fill = harvest_fill
        ws16.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        ws16.row_dimensions[i].height = 35

    # ── Accounting Quality Evaluation for HPG ──
    r8 = r7 + len(harvest_data) + 2
    ws16.cell(row=r8, column=1, value="3. ĐÁNH GIÁ CHẤT LƯỢNG KẾ TOÁN HPG").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E79")
    ws16.merge_cells(start_row=r8, start_column=1, end_row=r8, end_column=7)
    r8 += 1
    eval_data = [
        ("CFO/LNST bình quân 2021-2025",
         f"{sum(cfo_hist[i]/ni_hist[i] for i in range(5))/5:.2f}x",
         ">1.0 = chất lượng cao. HPG có dòng tiền vượt lợi nhuận nhờ khấu hao lớn, ít phải dồn tích.",
         "✅ TỐT"),
        ("Dự phòng HTK / Tồn kho",
         "—",
         "HPG không công bố chi tiết. Ngành thép: thường trích ~2-5% tồn kho. Với vòng quay 3.1x, rủi ro giảm giá HTK thấp.",
         "⚠️ CHƯA RÕ"),
        ("Phải thu / Doanh thu",
         f"{receivables_hist[4]/revenue_hist[4]*100:.1f}%",
         f"Xu hướng: {receivables_hist[0]/revenue_hist[0]*100:.0f}%→{receivables_hist[4]/revenue_hist[4]*100:.0f}% (2021→2025). Tỷ lệ ổn định không tăng bất thường.",
         "✅ TỐT"),
        ("Vòng quay tồn kho (ngày)",
         f"{sum(inventory_hist[i]/cogs_hist[i]*365 for i in range(5))/5:.0f} ngày",
         "Tồn kho ~105 ngày phù hợp ngành thép (sản xuất 60-90 ngày + dự trữ 30 ngày). Không có dấu hiệu ứ đọng.",
         "✅ TỐT"),
        ("Accruals Ratio",
         f"{(cfo_hist[4]-ni_hist[4])/total_assets_hist[4]*100:.1f}%",
         "Dương nhẹ = CFO < LNST do tăng phải thu/tồn kho mở rộng. Mức <5% là chấp nhận được.",
         "✅ CHẤP NHẬN ĐƯỢC"),
        ("Kết luận",
         "CHẤT LƯỢNG LỢI NHUẬN TỐT",
         "HPG có lịch sử CFO/LNST >1.0, ít dồn tích bất thường, vòng quay vốn lưu động ổn định. Rủi ro kế toán chính: NOTE về dự phòng HTK (cần kiểm tra chi tiết) và tỷ giá (dư nợ USD lớn). Nhìn chung, BCTC HPG phản ánh trung thực hoạt động kinh doanh.",
         "✅ TỐT"),
    ]
    for c, h in enumerate(["Tiêu chí", "Giá trị", "Đánh giá", "Kết luận"], 1):
        cell = ws16.cell(row=r8, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
    for i, (criteria, val, assessment, conclusion) in enumerate(eval_data, r8+1):
        ws16.cell(row=i, column=1, value=criteria).font = bold_font; ws16.cell(row=i, column=1).border = thin_border
        ws16.cell(row=i, column=2, value=val).font = data_font; ws16.cell(row=i, column=2).border = thin_border
        ws16.cell(row=i, column=2).alignment = Alignment(horizontal='center')
        ws16.cell(row=i, column=3, value=assessment).font = data_font; ws16.cell(row=i, column=3).border = thin_border
        ws16.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws16.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)
        ws16.cell(row=i, column=7, value=conclusion).font = Font(name=FONT_NAME, bold=True, size=9, color="006600" if "TỐT" in conclusion else "E65100")
        ws16.cell(row=i, column=7).border = thin_border; ws16.cell(row=i, column=7).fill = harvest_fill if "TỐT" in conclusion else acc_warn
        ws16.cell(row=i, column=7).alignment = Alignment(horizontal='center')
        ws16.row_dimensions[i].height = 45

    # ── Harvest Sign Readiness Score ──
    r9 = r8 + len(eval_data) + 2
    ws16.cell(row=r9, column=1, value="4. ĐIỂM SỐ SẴN SÀNG HÁI QUẢ").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E79")
    ws16.merge_cells(start_row=r9, start_column=1, end_row=r9, end_column=7)
    r9 += 1
    score_text = (
        "Đánh giá tổng thể: HPG ĐÃ SẴN SÀNG CHO GIAI ĐOẠN HÁI QUẢ.\n"
        "• Vĩ mô đảo chiều ✅: Lãi suất giảm, đầu tư công tăng, BĐS phục hồi.\n"
        f"• Spread nở ra ✅: Giá HRC hồi phục từ {HRC_PRICE_A[2]:,.0f} ({years_hist[2]}) lên {HRC_PRICE_A[5]:,.0f} ({years_fc[0]}E), Spread {SPREAD_A[2]:,.0f}→{SPREAD_A[5]:,.0f} USD/tấn.\n"
        "• Tồn kho giá rẻ ✅: Quặng 105 USD/tấn (thấp nhất 5 năm) → lợi thế GM Q2-Q4/2026.\n"
        "• DQ2 hoàn thành ✅: CIP giảm mạnh, FCF sẽ chuyển dương từ 2026.\n\n"
        "Kết luận: HPG đang bước vào chu kỳ lợi nhuận tăng mạnh (2026-2028) nhờ hội tụ đủ 4 yếu tố. "
        "Đây là thời điểm lý tưởng để nắm giữ cổ phiếu thép chu kỳ."
    )
    ws16.cell(row=r9, column=1, value=score_text).font = Font(name=FONT_NAME, italic=True, size=10, color="1F4E79")
    ws16.merge_cells(start_row=r9, start_column=1, end_row=r9+4, end_column=7)
    ws16.cell(row=r9, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    for rr in range(r9, r9+6):
        ws16.row_dimensions[rr].height = 22

    # ── Sheet 17: Dữ liệu giá hàng hóa (SINGLE SOURCE — mọi sheet khác link về đây) ──
    ws17 = wb.create_sheet("17_Gia_Hang_Hoa")
    ws17.cell(row=1, column=1, value="DỮ LIỆU GIÁ HRC / QUẶNG SẮT / THAN CỐC (nguồn duy nhất)").font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E79")
    ws17.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws17.cell(row=2, column=1, value=(
        "QUẶNG SẮT (cột C, xem cột F \"Nguồn giá Quặng\"): MEDIAN THẬT của 3 giá tháng lấy tự động từ "
        "World Bank Pink Sheet (miễn phí, công khai) cho các quý fetch thành công (\"WB\") — quý fetch lỗi "
        "dùng số nghiên cứu thủ công dự phòng (\"NC\"). HRC & THAN CỐC (cột B, D): CHƯA có nguồn giá theo "
        "tháng/ngày miễn phí tương đương → vẫn là SỐ NGHIÊN CỨU THỦ CÔNG đại diện 1 điểm/quý (KHÔNG phải "
        "median nhiều điểm — xem độ tin cậy từng mặt hàng ở mục NGUỒN DỮ LIỆU). Giá năm = MEDIAN 4 quý. "
        "Cột E = Spread HRC SỐNG LAG 1 QUÝ: HRC quý này ×1.15 (thuế CBPG AD20, từ 2025Q3) - 1.6×Quặng "
        f"quý TRƯỚC - 0.6×Than quý TRƯỚC - {OTHER_COST_USD} USD/t (vì tồn kho HPG bình quân ~90 ngày = "
        "~1 quý, giá vốn kỳ này phản ánh giá NVL mua vào kỳ trước) — bấm vào ô để kiểm chứng. Cột G-H = "
        "giá/nguồn thép XD (SRRc1 futures investing.com \"INV\", neo nội địa VSA/SteelOnline \"VN\", nội "
        "suy \"NC\"); Cột I = Spread Rebar (không có thuế CBPG); Cột J-K = SL XD/HRC quý (link sheet "
        "15_Quarterly_Data, chỉ có từ 2023Q1); Cột L = Spread All = bình quân gia quyền Spread Rebar/HRC "
        "theo SL quý. \"Giá hiện tại\" fetch tự động (investing.com/SteelOnline) khi chạy script (không "
        "cần AI hỗ trợ) — nếu fetch lỗi tự dùng giá quý gần nhất. Nguồn dữ liệu 18 quý lịch sử: xem mục "
        "\"NGUỒN DỮ LIỆU\" bên dưới bảng."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color="888888")
    ws17.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws17.row_dimensions[2].height = 68

    R17_HDR = 4
    header_row(ws17, R17_HDR,
        ["Quý", "Giá HRC (USD/t)", "Quặng sắt 62%Fe (USD/t)", "Than cốc luyện kim (USD/t)", "Spread HRC (USD/t)", "Nguồn giá Quặng",
         "Giá thép XD (USD/t)", "Nguồn giá XD", "Spread Rebar (USD/t)", "SL XD quý (kt)", "SL HRC quý (kt)", "Spread All (USD/t)"],
        [14, 18, 20, 20, 16, 26, 16, 20, 18, 14, 14, 16])
    _XD_SRC_LABEL = {"INV": "INV — median 3 tháng (SRRc1)", "VN": "VN — neo nội địa (VSA/SteelOnline)", "NC": "NC — nội suy"}
    R17_Q0 = R17_HDR + 1  # dòng đầu tiên = 2021Q4
    for i, (q, hrc, iron, coal, xd) in enumerate(zip(Q18_LABELS, Q18_HRC, Q18_IRON, Q18_COAL, Q18_XD)):
        rr = R17_Q0 + i
        ws17.cell(row=rr, column=1, value=q).font = bold_font
        ws17.cell(row=rr, column=1).border = thin_border
        for c, v in ((2, hrc), (3, iron), (4, coal), (7, xd)):
            cell = ws17.cell(row=rr, column=c, value=v)
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            cell.fill = assump_fill
        # Cột E: công thức Spread HRC SỐNG lag-1-quý, ĐÃ nhân thuế CBPG 1.15 từ 2025Q3 (AD_HRC_MULTIPLIER)
        # — để kiểm chứng trực tiếp, không phải số chết. Dòng đầu tiên (2021Q4) không có quý trước trong
        # bảng nên tạm dùng giá CÙNG quý (duy nhất trường hợp ngoại lệ này).
        _ad_mult_str = f"*{AD_HRC_MULTIPLIER}" if i >= _ad_idx else ""
        if i == 0:
            ecell = ws17.cell(row=rr, column=5, value=f"=B{rr}{_ad_mult_str}-1.6*C{rr}-0.6*D{rr}-{OTHER_COST_USD}")
        else:
            ecell = ws17.cell(row=rr, column=5, value=f"=B{rr}{_ad_mult_str}-1.6*C{rr-1}-0.6*D{rr-1}-{OTHER_COST_USD}")
        ecell.number_format = '#,##0'
        ecell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
        ecell.border = thin_border
        ecell.alignment = Alignment(horizontal='center')
        ecell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        # Cột F: nguồn gốc giá Quặng — "WB" = MEDIAN THẬT 3 tháng World Bank Pink Sheet (fetch tự
        # động), "NC" = số nghiên cứu thủ công 1 điểm/quý (fallback khi World Bank không fetch được
        # cho quý đó). Để user tự kiểm chứng quý nào là median thật, quý nào chưa.
        is_wb = Q18_IRON_SRC[i] == "WB"
        fcell = ws17.cell(row=rr, column=6, value="WB — median 3 tháng" if is_wb else "NC — nghiên cứu thủ công")
        fcell.font = Font(name=FONT_NAME, italic=True, size=9, color="1E7E34" if is_wb else "888888")
        fcell.border = thin_border
        fcell.alignment = Alignment(horizontal='center')
        # Cột H: nguồn giá thép XD (xem chú thích Q18_XD ở module-level).
        hcell = ws17.cell(row=rr, column=8, value=_XD_SRC_LABEL.get(Q18_XD_SRC[i], Q18_XD_SRC[i]))
        hcell.font = Font(name=FONT_NAME, italic=True, size=9, color="1E7E34" if Q18_XD_SRC[i] == "INV" else "888888")
        hcell.border = thin_border
        hcell.alignment = Alignment(horizontal='center')
        # Cột I: Spread Rebar SỐNG lag-1-quý (thép XD KHÔNG chịu thuế CBPG HRC — không nhân 1.15).
        if i == 0:
            icell = ws17.cell(row=rr, column=9, value=f"=G{rr}-1.6*C{rr}-0.6*D{rr}-{OTHER_COST_USD}")
        else:
            icell = ws17.cell(row=rr, column=9, value=f"=G{rr}-1.6*C{rr-1}-0.6*D{rr-1}-{OTHER_COST_USD}")
        icell.number_format = '#,##0'
        icell.font = Font(name=FONT_NAME, bold=True, color="8E44AD")
        icell.border = thin_border
        icell.alignment = Alignment(horizontal='center')
        icell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        # Cột J/K: SL XD/HRC quý — LINK SỐNG về sheet 15_Quarterly_Data (nguồn duy nhất), chỉ có từ
        # 2023Q1 (= SALES_Q_LABELS[0]) vì chưa có dữ liệu tách sản lượng quý trước đó.
        if q in SALES_Q_LABELS:
            _sq_col = get_column_letter(2 + SALES_Q_LABELS.index(q))
            jcell = ws17.cell(row=rr, column=10, value=f"='15_Quarterly_Data'!{_sq_col}{R_QV_XD}")
            kcell = ws17.cell(row=rr, column=11, value=f"='15_Quarterly_Data'!{_sq_col}{R_QV_HRC}")
            for cell in (jcell, kcell):
                cell.number_format = '#,##0'; cell.font = data_font; cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            # Cột L: Spread All SỐNG = bình quân gia quyền Spread Rebar (I)/Spread HRC (E) theo SL XD/HRC (J/K).
            lcell = ws17.cell(row=rr, column=12, value=f"=(J{rr}*I{rr}+K{rr}*E{rr})/(J{rr}+K{rr})")
            lcell.number_format = '#,##0'
            lcell.font = Font(name=FONT_NAME, bold=True, color="16A085")
            lcell.border = thin_border
            lcell.alignment = Alignment(horizontal='center')
            lcell.fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        else:
            for c in (10, 11, 12):
                ws17.cell(row=rr, column=c, value="—").font = Font(name=FONT_NAME, italic=True, size=9, color="888888")
                ws17.cell(row=rr, column=c).border = thin_border
                ws17.cell(row=rr, column=c).alignment = Alignment(horizontal='center')
    R17_Q_LAST = R17_Q0 + len(Q18_LABELS) - 1  # dòng 2026Q1

    r17 = R17_Q_LAST + 2
    ws17.cell(row=r17, column=1, value=(
        "GIÁ HIỆN TẠI — HRC/Than: investing.com (futures, fetch tự động); Quặng: World Bank Pink Sheet "
        f"(tháng mới nhất {max(IRON_MONTHLY) if IRON_MONTHLY else '?'}, KHÔNG dùng investing.com — cùng phương "
        "pháp luận với 18 quý lịch sử, tránh lệch cơ sở futures-vs-spot); Thép XD: SteelOnline nội địa "
        "(ưu tiên) hoặc SRRc1 futures (dự phòng)"
    )).font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws17.merge_cells(start_row=r17, start_column=1, end_row=r17, end_column=8)
    ws17.row_dimensions[r17].height = 26
    r17 += 1
    R17_NOW = r17
    ws17.cell(row=r17, column=1, value="Giá hiện tại (USD/t)").font = bold_font
    ws17.cell(row=r17, column=1).border = thin_border
    for c, v in ((2, hrc_now), (3, iron_now), (4, coal_now), (7, xd_now)):
        cell = ws17.cell(row=r17, column=c, value=v)
        cell.number_format = '#,##0.0'
        cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    ws17.cell(row=r17, column=8, value=XD_NOW_SRC).font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
    r17 += 1
    R17_AVG = r17
    ws17.cell(row=r17, column=1, value="MEDIAN(đầu quý hiện tại, giá hiện tại)").font = bold_font
    ws17.cell(row=r17, column=1).border = thin_border
    for c in (2, 3, 4, 7):
        cl = get_column_letter(c)
        cell = ws17.cell(row=r17, column=c, value=f"=MEDIAN({cl}{R17_Q_LAST},{cl}{R17_NOW})")
        cell.number_format = '#,##0.0'
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    r17 += 1
    R17_SPR_NOW = r17
    ws17.cell(row=r17, column=1, value="Spread hiện tại (USD/t) — lag 1 quý").font = Font(name=FONT_NAME, bold=True, size=10, color="1F4E79")
    ws17.cell(row=r17, column=1).border = thin_border
    # Spread HRC hiện tại = HRC bình quân quý ĐANG CHẠY (B{R17_AVG}) x 1.15 (thuế CBPG - quý đang
    # chạy chắc chắn sau 2025Q3) - 1.6xQuặng - 0.6xThan của QUÝ TRƯỚC đã biết đầy đủ (KHÔNG blend
    # với giá live) - OTHER_COST_USD. Spread Rebar hiện tại tương tự nhưng KHÔNG nhân thuế CBPG.
    spr_now_cell = ws17.cell(row=r17, column=2,
        value=f"=B{R17_AVG}*{AD_HRC_MULTIPLIER}-1.6*C{R17_Q_LAST}-0.6*D{R17_Q_LAST}-{OTHER_COST_USD}")
    spr_now_cell.number_format = '#,##0'
    spr_now_cell.font = Font(name=FONT_NAME, bold=True, color="C0392B")
    spr_now_cell.border = thin_border
    spr_now_cell.alignment = Alignment(horizontal='center')
    ws17.cell(row=r17, column=5, value=f"=MEDIAN(quý này,hiện tại) HRC x1.15 - 1.6xQuặng quý trước - 0.6xThan quý trước - {OTHER_COST_USD} USD/t").font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
    spr_rebar_now_cell = ws17.cell(row=r17, column=9,
        value=f"=G{R17_AVG}-1.6*C{R17_Q_LAST}-0.6*D{R17_Q_LAST}-{OTHER_COST_USD}")
    spr_rebar_now_cell.number_format = '#,##0'
    spr_rebar_now_cell.font = Font(name=FONT_NAME, bold=True, color="8E44AD")
    spr_rebar_now_cell.border = thin_border
    spr_rebar_now_cell.alignment = Alignment(horizontal='center')
    # Spread All hiện tại = bình quân gia quyền Spread HRC/Rebar hiện tại theo SL quý ĐANG CHẠY (ưu
    # tiên SL tách trực tiếp CUR_Q_HRC_KT_DIRECT/XD nếu có, fallback SL quý gần nhất đã biết đầy đủ
    # — 2 giá trị Python này, KHÔNG có ô Excel tương ứng để link công thức vì bản chất là ước tính
    # sớm cho quý CHƯA kết thúc, ghi trực tiếp làm neo tương tự các assumption khác trong sheet).
    ws17.cell(row=r17, column=10, value=round(_now_xd_kt, 1)).number_format = '#,##0'
    ws17.cell(row=r17, column=11, value=round(_now_hrc_kt, 1)).number_format = '#,##0'
    for c in (10, 11):
        ws17.cell(row=r17, column=c).font = data_font
        ws17.cell(row=r17, column=c).border = thin_border
        ws17.cell(row=r17, column=c).alignment = Alignment(horizontal='center')
    spr_all_now_cell = ws17.cell(row=r17, column=12, value=f"=(J{r17}*I{r17}+K{r17}*B{r17})/(J{r17}+K{r17})")
    spr_all_now_cell.number_format = '#,##0'
    spr_all_now_cell.font = Font(name=FONT_NAME, bold=True, color="16A085")
    spr_all_now_cell.border = thin_border
    spr_all_now_cell.alignment = Alignment(horizontal='center')
    spr_all_now_cell.fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")

    r17 += 2
    ws17.cell(row=r17, column=1, value="GIÁ NĂM = MEDIAN CÁC QUÝ TRONG NĂM (2027-2028: xu hướng tiếp nối giá hiện tại)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws17.merge_cells(start_row=r17, start_column=1, end_row=r17, end_column=9)
    r17 += 1
    R17_ANN_HDR = r17
    header_row(ws17, r17, ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"], [30] + [12]*8)
    r17 += 1
    R17_ANN_HRC, R17_ANN_IRON, R17_ANN_COAL, R17_ANN_XD = r17, r17 + 1, r17 + 2, r17 + 3
    # Dòng ứng với từng năm trong bảng 18 quý: 2021=row(R17_Q0), 2022=R17_Q0+1..+4, 2023=+5..+8,
    # 2024=+9..+12, 2025=+13..+16, 2026=R17_Q_LAST (chỉ Q1)
    yr_row_ranges = {
        2021: (R17_Q0, R17_Q0),
        2022: (R17_Q0 + 1, R17_Q0 + 4),
        2023: (R17_Q0 + 5, R17_Q0 + 8),
        2024: (R17_Q0 + 9, R17_Q0 + 12),
        2025: (R17_Q0 + 13, R17_Q0 + 16),
    }
    _trend_xd_27 = round(XD_PRICE_A[6] / XD_PRICE_A[5], 4)
    _trend_xd_28 = round(XD_PRICE_A[7] / XD_PRICE_A[6], 4)
    for label, col, ann_row in (("Giá HRC (USD/t)", 2, R17_ANN_HRC), ("Quặng sắt 62%Fe (USD/t)", 3, R17_ANN_IRON),
                                 ("Than cốc (USD/t)", 4, R17_ANN_COAL), ("Giá thép XD (USD/t)", 7, R17_ANN_XD)):
        cl = get_column_letter(col)
        ws17.cell(row=ann_row, column=1, value=label).font = bold_font
        ws17.cell(row=ann_row, column=1).border = thin_border
        for j, yr in enumerate([2021, 2022, 2023, 2024, 2025], 2):
            s, e = yr_row_ranges[yr]
            cell = ws17.cell(row=ann_row, column=j)
            cell.value = f"={cl}{s}" if s == e else f"=MEDIAN({cl}{s}:{cl}{e})"
            cell.number_format = '#,##0'
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        # 2026E = MEDIAN(quý đã biết trong 2026, giá hiện tại) — khớp Python HRC_PRICE_A[5] v.v.
        c2026 = ws17.cell(row=ann_row, column=7)
        c2026.value = f"=MEDIAN({cl}{R17_Q_LAST},{cl}{R17_NOW})"
        c2026.number_format = '#,##0'; c2026.font = Font(name=FONT_NAME, bold=True, color="C0392B"); c2026.border = thin_border
        c2026.alignment = Alignment(horizontal='center')
        # 2027E/2028E = tiếp nối xu hướng (giả định thận trọng, không có công thức nội tại vì là giá tương lai)
        trend = {2: (1.02, 1.015), 3: (1.03, 1.02), 4: (1.02, 1.02), 7: (_trend_xd_27, _trend_xd_28)}[col]
        c2027 = ws17.cell(row=ann_row, column=8, value=f"=G{ann_row}*{trend[0]}")
        c2027.number_format = '#,##0'; c2027.font = data_font; c2027.border = thin_border; c2027.alignment = Alignment(horizontal='center')
        c2028 = ws17.cell(row=ann_row, column=9, value=f"=H{ann_row}*{trend[1]}")
        c2028.number_format = '#,##0'; c2028.font = data_font; c2028.border = thin_border; c2028.alignment = Alignment(horizontal='center')

    # Spread HRC/Rebar/All NĂM — xây từ cột E/I/L (Spread quý lag-1-quý), KHÔNG trừ trực tiếp giá
    # cùng năm ở trên (không phản ánh đúng độ trễ tồn kho ~90 ngày):
    #  - 2021-2025: MEDIAN các Spread quý thuộc năm đó — nhất quán với "Giá năm = MEDIAN 4 quý".
    #    (Spread All chỉ có dữ liệu từ 2023 — 2021/2022 để trống, xem chú thích Q18_SPREAD_ALL).
    #  - 2026E: TB(Spread quý gần nhất đã biết, Spread hiện tại tương ứng).
    #  - 2027E/2028E: chưa có dữ liệu quý, xấp xỉ lag-1-NĂM dùng chính các dòng giá năm ở trên.
    R17_ANN_SPREAD, R17_ANN_SPREAD_REBAR, R17_ANN_SPREAD_ALL = r17 + 4, r17 + 5, r17 + 6
    # Đối chiếu với R17_LAYOUT (tính SẴN TRƯỚC KHI build sheet này, dùng để link công thức từ
    # 02_Assumptions!row45) — nếu lệch nhau nghĩa là có ai sửa layout sheet 17 mà quên đồng bộ, phải
    # BÁO LỖI NGAY thay vì âm thầm link sai ô (đúng bug user vừa phát hiện ở SL_HRC_A/SL_XD_A).
    assert (R17_HDR, R17_Q0, R17_Q_LAST, R17_NOW, R17_AVG, R17_SPR_NOW, R17_ANN_HDR, R17_ANN_HRC,
            R17_ANN_SPREAD, R17_ANN_SPREAD_ALL) == (
        R17_LAYOUT["R17_HDR"], R17_LAYOUT["R17_Q0"], R17_LAYOUT["R17_Q_LAST"], R17_LAYOUT["R17_NOW"],
        R17_LAYOUT["R17_AVG"], R17_LAYOUT["R17_SPR_NOW"], R17_LAYOUT["R17_ANN_HDR"], R17_LAYOUT["R17_ANN_HRC"],
        R17_LAYOUT["R17_ANN_SPREAD"], R17_LAYOUT["R17_ANN_SPREAD_ALL"]), (
        "Sheet 17 row layout drifted from R17_LAYOUT (_r17_annual_row_layout) - update both together, "
        "02_Assumptions!row45 links to R17_LAYOUT and will silently point to the wrong cell otherwise.")
    _ann_spread_cfg = (
        # 2027E/2028E Spread HRC PHẢI nhân AD_HRC_MULTIPLIER (đã sau 2025Q3, chắc chắn có thuế CBPG) —
        # khớp đúng cách tính Python SPREAD_A[6]/[7] (_lag_spread(HRC_PRICE_A[i]*AD_HRC_MULTIPLIER, ...)).
        ("Spread HRC (USD/t) — lag 1 quý/năm", R17_ANN_SPREAD, 5, "E", "C0392B", "FFF3E0", spr_now_cell.coordinate,
         f"={{cl_cur}}{R17_ANN_HRC}*{AD_HRC_MULTIPLIER}-1.6*{{cl_prev}}{R17_ANN_IRON}-0.6*{{cl_prev}}{R17_ANN_COAL}-{OTHER_COST_USD}"),
        ("Spread Rebar (USD/t) — lag 1 quý/năm", R17_ANN_SPREAD_REBAR, 9, "I", "8E44AD", "F3E5F5", spr_rebar_now_cell.coordinate,
         f"={{cl_cur}}{R17_ANN_XD}-1.6*{{cl_prev}}{R17_ANN_IRON}-0.6*{{cl_prev}}{R17_ANN_COAL}-{OTHER_COST_USD}"),
    )
    for label, ann_row, q_col, q_cl, color, fill_hex, now_coord, fut_formula_tmpl in _ann_spread_cfg:
        ws17.cell(row=ann_row, column=1, value=label).font = Font(name=FONT_NAME, bold=True, color=color)
        ws17.cell(row=ann_row, column=1).border = thin_border
        for j, yr in enumerate([2021, 2022, 2023, 2024, 2025], 2):
            s, e = yr_row_ranges[yr]
            cell = ws17.cell(row=ann_row, column=j)
            cell.value = f"={q_cl}{s}" if s == e else f"=MEDIAN({q_cl}{s}:{q_cl}{e})"
            cell.number_format = '#,##0'; cell.font = Font(name=FONT_NAME, bold=True, color=color)
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        c2026s = ws17.cell(row=ann_row, column=7, value=f"=MEDIAN({q_cl}{R17_Q_LAST},{now_coord})")
        c2026s.number_format = '#,##0'; c2026s.font = Font(name=FONT_NAME, bold=True, color=color)
        c2026s.border = thin_border; c2026s.alignment = Alignment(horizontal='center')
        c2026s.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        for j_cur, j_prev in ((8, 7), (9, 8)):
            cl_cur, cl_prev = get_column_letter(j_cur), get_column_letter(j_prev)
            c = ws17.cell(row=ann_row, column=j_cur, value=fut_formula_tmpl.format(cl_cur=cl_cur, cl_prev=cl_prev))
            c.number_format = '#,##0'; c.font = Font(name=FONT_NAME, bold=True, color=color)
            c.border = thin_border; c.alignment = Alignment(horizontal='center')
            c.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

    # Spread All NĂM = bình quân gia quyền Spread HRC/Rebar NĂM theo SL_HRC_A/SL_XD_A (ước tính sản
    # lượng năm — KHÔNG dùng SL quý vì năm 2027E/2028E không có SL quý). SL_HRC_A/SL_XD_A là assumption
    # Python (chưa có sheet Excel riêng theo năm để link — xem 03_Revenue_Model) nên ghi trực tiếp giá
    # trị SL (triệu tấn) làm neo tính toán, công thức Spread All vẫn SỐNG (tham chiếu 2 dòng trên).
    ws17.cell(row=R17_ANN_SPREAD_ALL, column=1, value="Spread All (USD/t) — bình quân gia quyền SL năm").font = Font(name=FONT_NAME, bold=True, color="16A085")
    ws17.cell(row=R17_ANN_SPREAD_ALL, column=1).border = thin_border
    _sl_row_xd, _sl_row_hrc = R17_ANN_SPREAD_ALL + 1, R17_ANN_SPREAD_ALL + 2
    ws17.cell(row=_sl_row_xd, column=1, value="  (SL thép XD năm, triệu tấn)").font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
    ws17.cell(row=_sl_row_hrc, column=1, value="  (SL HRC năm, triệu tấn)").font = Font(name=FONT_NAME, italic=True, size=8, color="888888")
    for j, i8 in enumerate(range(8), 2):
        cl_j = get_column_letter(j)
        cxd = ws17.cell(row=_sl_row_xd, column=j, value=SL_XD_A[i8]); cxd.number_format = '0.00'
        chrc = ws17.cell(row=_sl_row_hrc, column=j, value=SL_HRC_A[i8]); chrc.number_format = '0.00'
        for cc in (cxd, chrc):
            cc.font = Font(name=FONT_NAME, italic=True, size=8, color="888888"); cc.alignment = Alignment(horizontal='center')
        c = ws17.cell(row=R17_ANN_SPREAD_ALL, column=j,
            value=f"=({cl_j}{_sl_row_xd}*{cl_j}{R17_ANN_SPREAD_REBAR}+{cl_j}{_sl_row_hrc}*{cl_j}{R17_ANN_SPREAD})/({cl_j}{_sl_row_xd}+{cl_j}{_sl_row_hrc})")
        c.number_format = '#,##0'; c.font = Font(name=FONT_NAME, bold=True, color="16A085")
        c.border = thin_border; c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")

    r17 += 9
    ws17.cell(row=r17, column=1, value="NGUỒN DỮ LIỆU (để kiểm chứng lại)").font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    ws17.merge_cells(start_row=r17, start_column=1, end_row=r17, end_column=9)
    r17 += 1
    _n_wb = Q18_IRON_SRC.count("WB")
    source_notes = [
        ("Giá HRC (\"Giá hiện tại\", B25)", "investing.com — LME Steel HRC FOB China Futures: https://www.investing.com/commodities/lme-steel-hrc-fob-china-futures (fetch tự động qua curl khi chạy script)"),
        ("Giá Quặng sắt 62%Fe (\"Giá hiện tại\", C25)", "World Bank Commodity Markets Pink Sheet — tháng MỚI NHẤT có dữ liệu (thường trễ ~1 tháng so với hôm nay), CÙNG nguồn với 18 quý lịch sử ở trên (KHÔNG dùng investing.com cho quặng — 2 nguồn khác phương pháp luận: futures SGX TSI vs spot bình quân tháng WB, trộn lẫn dễ gây lệch cơ sở khi so hiện tại với lịch sử). Fallback về investing.com Iron Ore 62% Fe CFR Futures (https://vn.investing.com/commodities/iron-ore-62-cfr-futures) nếu World Bank fetch lỗi."),
        ("Giá Than cốc (\"Giá hiện tại\", D25)", "investing.com — Metallurgical Coke Futures (niêm yết CNY, quy đổi USD theo tỷ giá CNY_USD_RATE=7.2): https://vn.investing.com/commodities/metallurgical-coke-futures (fetch tự động qua curl khi chạy script)"),
        ("Quặng sắt 18 quý lịch sử (cột C, xem cột F)",
         f"{_n_wb}/{len(Q18_IRON_SRC)} quý = MEDIAN THẬT của 3 giá THÁNG lấy tự động (curl) từ World Bank "
         "Commodity Markets \"Pink Sheet\" (https://www.worldbank.org/en/research/commodity-markets, sheet "
         "\"Monthly Prices\", cột \"Iron ore, cfr spot\") — miễn phí, công khai, cập nhật hàng tháng, KHÔNG "
         "phải số nghiên cứu tĩnh. Quý còn lại (đánh dấu \"NC\" ở cột F, do World Bank chưa có đủ 3 tháng "
         "hoặc fetch lỗi) dùng số nghiên cứu thủ công dự phòng như cũ — không dừng script vì lỗi fetch."),
        ("HRC & Than cốc 18 quý lịch sử (cột B, D)",
         "Tổng hợp thủ công từ báo cáo FPTS/VCBS/Argus (than cốc — có điểm neo thật, trung bình), Mysteel/"
         "SteelBenchmarker/giá xuất khẩu HPG (HRC — chỉ đúng xu hướng/độ lớn). CHƯA có nguồn giá theo tháng/"
         "ngày miễn phí tương đương World Bank cho 2 mặt hàng này → vẫn là 1 số đại diện/quý do nghiên cứu "
         "thủ công (KHÔNG phải median nhiều điểm), cần đối chiếu lại nếu dùng cho quyết định quan trọng."),
    ]
    for label, note in source_notes:
        ws17.cell(row=r17, column=1, value=label).font = bold_font
        ws17.cell(row=r17, column=1).border = thin_border
        ws17.cell(row=r17, column=2, value=note).font = Font(name=FONT_NAME, italic=True, size=8, color="555555")
        ws17.cell(row=r17, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws17.merge_cells(start_row=r17, start_column=2, end_row=r17, end_column=9)
        ws17.row_dimensions[r17].height = 42
        r17 += 1

    # ── SL HRC/XD (03_Revenue_Model!row4-5) năm 2023-2025 — GHI ĐÈ bằng công thức SUM sống link sang
    # sheet 15_Quarterly_Data (2026-07, user phát hiện bug: giá trị Python cũ SL_HRC_A/SL_XD_A lệch xa
    # số thật, VD 2025 HRC 3.2 triệu tấn giả định vs 5.0 triệu tấn thật — xem chú thích chỗ định nghĩa
    # SL_HRC_A). Đặt Ở ĐÂY (không phải lúc build sheet 03) vì cần R_QV_HRC/R_QV_XD của sheet 15, mà
    # sheet 03 được build TRƯỚC sheet 15 trong thứ tự code.
    _sl_hist_yr_cols3 = {2023: 4, 2024: 5, 2025: 6}   # cột trong sheet 03 (D/E/F)
    _sl_hist_yr_s15 = {2023: ('B', 'E'), 2024: ('F', 'I'), 2025: ('J', 'M')}  # cột trong sheet 15 (4 quý/năm)
    for _yr, _col3 in _sl_hist_yr_cols3.items():
        c0, c1 = _sl_hist_yr_s15[_yr]
        hrc_cell = ws3.cell(row=4, column=_col3, value=f"=SUM('15_Quarterly_Data'!{c0}{R_QV_HRC}:{c1}{R_QV_HRC})/1000")
        xd_cell = ws3.cell(row=5, column=_col3, value=f"=SUM('15_Quarterly_Data'!{c0}{R_QV_XD}:{c1}{R_QV_XD})/1000")
        for cell in (hrc_cell, xd_cell):
            cell.number_format = '#,##0.00'; cell.font = data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # ── Link 03_Revenue_Model!Giá HRC (row6) và 02_Assumptions!Quặng sắt/Than cốc (row19/20) ──
    # về sheet 17 thay vì số Python tĩnh — 1 nguồn duy nhất, sửa 1 chỗ là cả model cập nhật theo.
    S17 = "'17_Gia_Hang_Hoa'"
    for j in range(2, 10):
        cl = col_ltr(j)
        ws3.cell(row=6, column=j, value=f"={S17}!{cl}{R17_ANN_HRC}").number_format = '#,##0'
        ws2.cell(row=19, column=j, value=f"={S17}!{cl}{R17_ANN_IRON}").number_format = '#,##0'
        ws2.cell(row=20, column=j, value=f"={S17}!{cl}{R17_ANN_COAL}").number_format = '#,##0'
    ws2.cell(row=19, column=10, value="Giá CFR Trung Quốc — link '17_Gia_Hang_Hoa' (median quý/năm)")
    ws2.cell(row=20, column=10, value="FOB Úc — link '17_Gia_Hang_Hoa' (median quý/năm)")
    for j in range(2, 10):
        cl = get_column_letter(j)
        ws14.cell(row=7, column=j, value=f"={S17}!{cl}{R17_ANN_HRC}").number_format = '#,##0'
        ws14.cell(row=8, column=j, value=f"={S17}!{cl}{R17_ANN_IRON}").number_format = '#,##0'
        ws14.cell(row=9, column=j, value=f"={S17}!{cl}{R17_ANN_COAL}").number_format = '#,##0'
        # Spread thép (row10) — link trực tiếp về Spread thép NĂM (lag 1 quý/năm) của sheet 17,
        # KHÔNG tự trừ lại HRC-Quặng-Than cùng năm (2 dòng trên) như trước — tránh sai công thức lag.
        c10 = ws14.cell(row=10, column=j, value=f"={S17}!{cl}{R17_ANN_SPREAD}")
        c10.number_format = '#,##0'; c10.font = Font(name=FONT_NAME, bold=True, color="C0392B")

    # "Spread All quý gần nhất ĐÃ BIẾT" (2026-07, theo yêu cầu user — trước đây dùng Spread HRC, giờ
    # đổi sang Spread All vì đây là chỉ tiêu quyết định BLNG dự phóng) — link cột L (Spread All, đã
    # tính SỐNG theo SL thật của CHÍNH quý đó) tại dòng quý gần nhất đã có BCTC trong bảng 18 quý.
    # QUY ƯỚC: dòng cuối bảng 18 quý (R17_Q_LAST) LUÔN được cập nhật = quý mới nhất ĐÃ CÓ BCTC (bảo
    # trì Q18_LABELS mỗi khi có báo cáo quý mới) — nên tự động khớp đúng số quý N đã biết (_n_q_known)
    # mà không cần logic riêng theo từng giá trị N.
    ws2.cell(row=R_Q1_SPR, column=1, value=f"Spread All quý gần nhất đã biết ({_n_q_known} quý {_cur_fc_year}, USD/t)")
    ws2.cell(row=R_Q1_SPR, column=7, value=f"={S17}!L{R17_Q_LAST}").number_format = '#,##0'
    ws2.cell(row=R_Q1_SPR, column=7).font = data_font
    # Spread HRC hiện tại (đầu quý + hiện tại) — hiển thị tham chiếu riêng, dùng cho phân tích trực quan
    R_SPR_NOW = 46
    ws2.cell(row=R_SPR_NOW, column=1, value="Spread HRC hiện tại (đầu quý+hiện tại, USD/t)").font = Font(name=FONT_NAME, bold=True, size=9, color="1F4E79")
    ws2.cell(row=R_SPR_NOW, column=1).border = thin_border
    ws2.cell(row=R_SPR_NOW, column=7, value=f"={S17}!B{R17_SPR_NOW}").number_format = '#,##0'
    ws2.cell(row=R_SPR_NOW, column=7).font = data_font
    # Spread All hiện tại (2026-07, theo yêu cầu user) — dùng làm tử số cho công thức dự phóng BLNG
    # (row 6) ở CẢ năm dự phóng đầu tiên (blend theo quý đã biết) LẪN các năm xa hơn (tỉ lệ năm/năm).
    R_SPR_ALL_NOW = 47
    ws2.cell(row=R_SPR_ALL_NOW, column=1, value="Spread All hiện tại (đầu quý+hiện tại, USD/t)").font = Font(name=FONT_NAME, bold=True, size=9, color="16A085")
    ws2.cell(row=R_SPR_ALL_NOW, column=1).border = thin_border
    ws2.cell(row=R_SPR_ALL_NOW, column=7, value=f"={S17}!L{R17_SPR_NOW}").number_format = '#,##0'
    ws2.cell(row=R_SPR_ALL_NOW, column=7).font = Font(name=FONT_NAME, bold=True, color="16A085")

    # ── Spread hàng năm (row 45) — LINK sang sheet 17, KHÔNG tự trừ lại (2026-07, user phát hiện bug:
    # công thức cũ độc lập dễ lệch số với sheet 17). 2021-2022 (chưa có SL tách quý → Spread All chưa
    # tính được) dùng Spread HRC (R17_ANN_SPREAD, tương đương "Spread thép" trước khi tách 3 loại);
    # 2023 trở đi dùng Spread All (R17_ANN_SPREAD_ALL, bình quân gia quyền theo SL thật) — đúng yêu
    # cầu "median spread all của 4 quý" (2023+) / "median spread thép của 4 quý" (≤2022).
    ws2.cell(row=R_ANN_SPR, column=1, value="Spread hàng năm (USD/t)").font = Font(name=FONT_NAME, bold=True, size=9, color="1F4E79")
    ws2.cell(row=R_ANN_SPR, column=1).border = thin_border
    for j in range(2, 10):
        cl = col_ltr(j)
        yr_j = 2019 + j
        src_row = R17_ANN_SPREAD if yr_j <= 2022 else R17_ANN_SPREAD_ALL
        c = ws2.cell(row=R_ANN_SPR, column=j, value=f"={S17}!{cl}{src_row}")
        c.number_format = '#,##0'
        c.font = data_font
        c.border = thin_border
    ws2.cell(row=R_ANN_SPR, column=10,
             value="≤2022: Spread HRC (chưa có SL tách quý để tính All); ≥2023: Spread All — link '17_Gia_Hang_Hoa'")

    # ── Biên LNG dự phóng (row 6) — theo yêu cầu user (2026-07):
    #  Năm dự phóng ĐẦU TIÊN (2026E, cột G) — n quý ĐÃ CÓ BCTC (_n_q_known):
    #    LNG năm = LNG lũy kế n quý + (4-n)/4 x Doanh thu ước tính năm x (LNG lũy kế n quý/Doanh thu
    #    lũy kế n quý) x (Spread All hiện tại/Spread All quý gần nhất đã biết); BLNG năm = LNG năm/DT.
    #    n=4 (đủ cả năm): BLNG năm = LNG lũy kế 4 quý/Doanh thu lũy kế 4 quý (không cần ước tính nữa).
    #  Các năm SAU (2027E/2028E, chưa có quý nào của năm đó) = BLNG năm TRƯỚC x (Spread All hiện tại/
    #    Spread All NĂM TRƯỚC — dùng annual Spread All ước tính của chính năm liền trước, R17_ANN_SPREAD_ALL).
    _rev_est_g = f"{S_R3}!G2"
    if _n_q_known >= 4:
        g6_formula = f"=G{R_Q1_GP}/G{R_Q1_REV}*100"
    else:
        _remain_frac = (4 - _n_q_known) / 4
        g6_formula = (f"=(G{R_Q1_GP}+{_remain_frac}*{_rev_est_g}*(G{R_Q1_GP}/G{R_Q1_REV})"
                      f"*(G{R_SPR_ALL_NOW}/G{R_Q1_SPR}))/{_rev_est_g}*100")
    ws2.cell(row=6, column=7, value=g6_formula).number_format = '0.0'
    ws2.cell(row=6, column=8, value=f"=G6*(G{R_SPR_ALL_NOW}/{S17}!G{R17_ANN_SPREAD_ALL})").number_format = '0.0'
    ws2.cell(row=6, column=9, value=f"=H6*(G{R_SPR_ALL_NOW}/{S17}!H{R17_ANN_SPREAD_ALL})").number_format = '0.0'

    # Save
    wb.save(EXCEL_FILE)
    print(f"[OK] Excel model saved: {EXCEL_FILE}")
    return True


# ── 2. MATPLOTLIB CHARTS ────────────────────────────────────────────────────

def make_charts():
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial', 'Calibri', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    all_years_lbl = [str(y) for y in years_hist + years_fc]
    all_rev = revenue_hist + revenue_fc
    all_ni = ni_hist + ni_fc
    all_gpm = gp_margin_hist + gp_margin_fc

    # Chart 1: Revenue & NI
    fig, ax1 = plt.subplots(figsize=(10, 5))
    x = np.arange(len(all_years_lbl))
    width = 0.4

    bars = ax1.bar(x - width/2, [v/1000 for v in all_rev], width,
                   label='Doanh thu (nghìn tỷ)', color='#1F4E79', alpha=0.85)
    ax2 = ax1.twinx()
    ax2.plot(x, [v/1000 for v in all_ni], 'o-', color='#E74C3C',
             linewidth=2, markersize=6, label='LNST (nghìn tỷ)')

    for bar, val in zip(bars, all_rev):
        ax1.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.5,
                 f'{val/1000:.1f}', ha='center', va='bottom', fontsize=8)

    ax1.set_xlabel('Năm', fontsize=11)
    ax1.set_ylabel('Doanh thu (nghìn tỷ VND)', fontsize=11, color='#1F4E79')
    ax2.set_ylabel('LNST (nghìn tỷ VND)', fontsize=11, color='#E74C3C')
    ax1.set_xticks(x)
    ax1.set_xticklabels(all_years_lbl)
    ax1.grid(axis='y', alpha=0.3)
    fig.suptitle('HPG - Doanh thu & LNST (2021-2028E)', fontsize=13, fontweight='bold')

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'revenue_ni.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 2: Margins
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(all_years_lbl, all_gpm, 's-', color='#1F4E79', linewidth=2,
            markersize=6, label='GP Margin')
    ni_margin = [all_ni[i]/all_rev[i]*100 for i in range(len(all_rev))]
    ax.plot(all_years_lbl, ni_margin, 'o-', color='#E74C3C', linewidth=2,
            markersize=6, label='NI Margin')

    for i, val in enumerate(all_gpm):
        ax.annotate(f'{val:.1f}%', (all_years_lbl[i], val),
                    textcoords="offset points", xytext=(0, 10),
                    ha='center', fontsize=8, color='#1F4E79')
    for i, val in enumerate(ni_margin):
        ax.annotate(f'{val:.1f}%', (all_years_lbl[i], val),
                    textcoords="offset points", xytext=(0, -15),
                    ha='center', fontsize=8, color='#E74C3C')

    ax.set_xlabel('Năm', fontsize=11)
    ax.set_ylabel('%', fontsize=11)
    ax.set_title('HPG - Biên lợi nhuận (2021-2028E)', fontsize=13, fontweight='bold')
    ax.legend(loc='best')
    ax.grid(alpha=0.3)
    fig.subplots_adjust(bottom=0.1)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'margins.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 3: Sensitivity heatmap
    fig, ax = plt.subplots(figsize=(8, 5))
    ev_ebitda_range = [5.5, 6.5, 7.5, 8.5, 9.5]
    ebitda_2026e = revenue_fc[0] * 0.143 + da_fc[0]
    ebitda_range_vals = [round(ebitda_2026e * x) for x in [0.85, 0.925, 1.0, 1.075, 1.15]]
    net_debt_2026e = 78000 - 25000

    data = []
    for ebitda_val in ebitda_range_vals:
        row = []
        for multiple in ev_ebitda_range:
            ev = ebitda_val * multiple
            eq = ev - net_debt_2026e
            price_val = max(0, eq * 1e9 / SHARES)
            row.append(round(price_val))
        data.append(row)

    im = ax.imshow(data, cmap='RdYlGn', aspect='auto')
    ax.set_xticks(np.arange(len(ev_ebitda_range)))
    ax.set_yticks(np.arange(len(ebitda_range_vals)))
    ax.set_xticklabels([f'{x}x' for x in ev_ebitda_range])
    ax.set_yticklabels([f'{v:,.0f}' for v in ebitda_range_vals])
    ax.set_xlabel('EV/EBITDA Multiple')
    ax.set_ylabel('EBITDA 2026E (tỷ)')

    for i in range(len(ebitda_range_vals)):
        for j in range(len(ev_ebitda_range)):
            text = ax.text(j, i, f'{data[i][j]:,.0f}',
                          ha="center", va="center", color="black", fontsize=9,
                          fontweight='bold')

    ax.set_title('Sensitivity: Target Price (VND)', fontsize=13, fontweight='bold')
    fig.colorbar(im, ax=ax, shrink=0.8)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'sensitivity.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 4: Peer comparison
    fig, ax = plt.subplots(figsize=(8, 5))
    tickers = list(peers.keys())
    pe_vals = [peers[t]["pe"] for t in tickers]
    pb_vals = [peers[t]["pb"] for t in tickers]

    x = np.arange(len(tickers))
    width = 0.35
    bars1 = ax.bar(x - width/2, pe_vals, width, label='P/E (x)', color='#1F4E79', alpha=0.8)
    bars2 = ax.bar(x + width/2, pb_vals, width, label='P/B (x)', color='#E74C3C', alpha=0.8)

    for bar, val in zip(bars1, pe_vals):
        ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.2,
                f'{val:.1f}', ha='center', va='bottom', fontsize=9, fontweight='bold')
    for bar, val in zip(bars2, pb_vals):
        ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.2,
                f'{val:.1f}', ha='center', va='bottom', fontsize=9, fontweight='bold')

    ax.set_xticks(x)
    ax.set_xticklabels(tickers)
    ax.set_ylabel('Multiple (x)')
    ax.set_title('So sánh Peer (2026E)', fontsize=13, fontweight='bold')
    ax.legend(loc='upper left')
    ax.grid(axis='y', alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'peers.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 5: Revenue Growth YoY%
    fig, ax = plt.subplots(figsize=(10, 5))
    growth_years = [str(y) for y in years_hist + years_fc]
    growth_vals = [48.0, -5.5, -15.9, 16.7, 12.4, 34.5, 14.3, 12.5]
    colors = ['#27AE60' if v >= 0 else '#E74C3C' for v in growth_vals]
    bars = ax.bar(growth_years, growth_vals, color=colors, alpha=0.8, edgecolor='white', linewidth=0.5)
    for bar, val in zip(bars, growth_vals):
        y_pos = bar.get_height() + 1.5 if val >= 0 else bar.get_height() - 6
        ax.text(bar.get_x() + bar.get_width()/2., y_pos,
                f'{val:+.1f}%', ha='center', va='bottom', fontsize=8, fontweight='bold')
    ax.axhline(y=0, color='gray', linewidth=0.8)
    ax.set_xlabel('Năm', fontsize=11)
    ax.set_ylabel('Tăng trưởng DT YoY (%)', fontsize=11)
    ax.set_title('HPG - Tăng trưởng Doanh thu (YoY%)', fontsize=13, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'growth.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 6: DIO (Số ngày tồn kho BQ) & DSO (Số ngày phải thu BQ) THEO NĂM — dùng DIO_A/DSO_A
    # tính động từ 365 x số dư BÌNH QUÂN/GVHB-DT (khớp Excel sheet 14_Steel_Analysis Section 4).
    # CHỈ LỊCH SỬ 2021-2025 (không dự phóng, theo yêu cầu user vì ít ảnh hưởng tới định giá).
    fig, ax1 = plt.subplots(figsize=(10, 5))
    yr_lbl_hist = [str(y) for y in years_hist]
    x = np.arange(len(yr_lbl_hist))
    ax1.plot(x, DIO_A, 's-', color='#E74C3C', linewidth=2, markersize=8, label='DIO — Số ngày tồn kho BQ')
    ax1.plot(x, DSO_A, 'o-', color='#1F4E79', linewidth=2, markersize=8, label='DSO — Số ngày phải thu BQ')
    for i, v in enumerate(DIO_A):
        ax1.annotate(f'{v:.0f}', (x[i], v), textcoords="offset points", xytext=(0, 10), ha='center', fontsize=9)
    for i, v in enumerate(DSO_A):
        ax1.annotate(f'{v:.0f}', (x[i], v), textcoords="offset points", xytext=(0, -14), ha='center', fontsize=9)
    ax1.set_xlabel('Năm', fontsize=11)
    ax1.set_ylabel('Số ngày', fontsize=11)
    ax1.set_xticks(x)
    ax1.set_xticklabels(yr_lbl_hist)
    ax1.set_title('HPG — Số ngày tồn kho (DIO) & Số ngày phải thu (DSO) bình quân, theo Năm', fontsize=13, fontweight='bold')
    ax1.legend(loc='upper left')
    ax1.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'turnover.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 6b: DIO/DSO THEO QUÝ (2021Q4-2026Q1, khớp Q18_LABELS) — quy đổi năm hóa (GVHB/DT quý x4)
    # để cùng thang đo với biểu đồ năm ở trên. Dữ liệu thực tế BCTC quý, không dự phóng.
    fig, ax1 = plt.subplots(figsize=(12, 5.5))
    xq = np.arange(len(Q18_LABELS))
    xq_lbl = [f"{l[4:]}'{l[2:4]}" for l in Q18_LABELS]
    dio_q_plot = [v if v is not None else np.nan for v in DIO_Q]
    dso_q_plot = [v if v is not None else np.nan for v in DSO_Q]
    ax1.plot(xq, dio_q_plot, 's-', color='#E74C3C', linewidth=2, markersize=6, label='DIO — Số ngày tồn kho BQ (năm hóa)')
    ax1.plot(xq, dso_q_plot, 'o-', color='#1F4E79', linewidth=2, markersize=6, label='DSO — Số ngày phải thu BQ (năm hóa)')
    for i, v in enumerate(dio_q_plot):
        if not np.isnan(v):
            ax1.annotate(f'{v:.0f}', (xq[i], v), textcoords="offset points", xytext=(0, 8), ha='center', fontsize=7, color='#E74C3C')
    for i, v in enumerate(dso_q_plot):
        if not np.isnan(v):
            ax1.annotate(f'{v:.0f}', (xq[i], v), textcoords="offset points", xytext=(0, -12), ha='center', fontsize=7, color='#1F4E79')
    ax1.set_xlabel('Quý', fontsize=11)
    ax1.set_ylabel('Số ngày (năm hóa)', fontsize=11)
    ax1.set_xticks(xq)
    ax1.set_xticklabels(xq_lbl, fontsize=8, rotation=45)
    ax1.set_title('HPG — DIO & DSO bình quân theo Quý (năm hóa)', fontsize=13, fontweight='bold')
    ax1.legend(loc='upper left')
    ax1.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'turnover_quarterly.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 7: Quarterly Revenue & Profit (dynamic from API)
    def _get_q(records, yr, qtr, field):
        for r in records:
            if r.get("yearReport") == yr and r.get("lengthReport") == qtr:
                v = r.get(field)
                return v / 1e9 if v is not None else 0
        return 0
    is_qs_ch = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    bs_qs_ch = section_to_quarters(FIN_DATA, "BALANCE_SHEET")
    q_periods = [(2024,1),(2024,2),(2024,3),(2024,4),
                 (2025,1),(2025,2),(2025,3),(2025,4),(2026,1)]
    q_labels = ['Q1\n2024','Q2\n2024','Q3\n2024','Q4\n2024',
                'Q1\n2025','Q2\n2025','Q3\n2025','Q4\n2025','Q1\n2026']
    _q_rev = [_get_q(is_qs_ch, y, q, 'isa3') for y,q in q_periods]
    _q_ni  = [_get_q(is_qs_ch, y, q, 'isa22') for y,q in q_periods]
    fig, ax1 = plt.subplots(figsize=(14, 5.5))
    x = np.arange(len(q_labels))
    bars = ax1.bar(x, [v/1000 for v in _q_rev], color='#1F4E79', alpha=0.7, label='Doanh thu (nghìn tỷ)')
    ax2 = ax1.twinx()
    ax2.plot(x, [v/1000 for v in _q_ni], 'o-', color='#E74C3C', linewidth=2, markersize=6, label='LNST (nghìn tỷ)')
    for bar, val in zip(bars, _q_rev):
        ax1.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.3,
                f'{val/1000:.1f}', ha='center', va='bottom', fontsize=7)
    ax1.set_xlabel('Quý', fontsize=11)
    ax1.set_ylabel('Doanh thu (nghìn tỷ VND)', fontsize=11, color='#1F4E79')
    ax2.set_ylabel('LNST (nghìn tỷ VND)', fontsize=11, color='#E74C3C')
    ax1.set_xticks(x)
    ax1.set_xticklabels(q_labels, fontsize=8)
    ax1.set_title('HPG - KQKD theo Quý (Q1/2024 - Q1/2026)', fontsize=13, fontweight='bold')
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1+lines2, labels1+labels2, loc='upper left')
    ax1.grid(axis='y', alpha=0.3)
    fig.subplots_adjust(bottom=0.12)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'quarterly.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 7B: Quarterly BS trend (Tổng TS, Hàng tồn kho, Phải thu, Nợ vay)
    _q_ta  = [_get_q(bs_qs_ch, y, q, 'bsa53') for y,q in q_periods]
    _q_inv = [_get_q(bs_qs_ch, y, q, 'bsa15') for y,q in q_periods]
    _q_rec = [_get_q(bs_qs_ch, y, q, 'bsa8') for y,q in q_periods]
    _q_debt = [_get_q(bs_qs_ch, y, q, 'bsa56')+_get_q(bs_qs_ch, y, q, 'bsa71') for y,q in q_periods]
    fig, ax = plt.subplots(figsize=(14, 5))
    ax.plot(x, [v/1000 for v in _q_ta], 's-', color='#1F4E79', linewidth=2, label='Tổng TS')
    ax.plot(x, [v/1000 for v in _q_inv], 'o-', color='#E67E22', linewidth=2, label='Hàng tồn kho')
    ax.plot(x, [v/1000 for v in _q_rec], '^-', color='#27AE60', linewidth=2, label='Phải thu')
    ax.plot(x, [v/1000 for v in _q_debt], 'v-', color='#E74C3C', linewidth=2, label='Nợ vay')
    ax.set_xticks(x); ax.set_xticklabels(q_labels, fontsize=8)
    ax.set_title('HPG - Diễn biến Tài sản & Nợ theo Quý', fontsize=13, fontweight='bold')
    ax.set_ylabel('Nghìn tỷ VND', fontsize=11)
    ax.legend(loc='best'); ax.grid(axis='y', alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'quarterly_bs.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 7C: Comprehensive quarterly quality dashboard
    _q_interest = [_get_q(is_qs_ch, y, q, 'isa8') for y,q in q_periods]
    _q_sgka    = [_get_q(is_qs_ch, y, q, 'isa9')+_get_q(is_qs_ch, y, q, 'isa10') for y,q in q_periods]
    _q_equity  = [_get_q(bs_qs_ch, y, q, 'bsa78') for y,q in q_periods]
    _q_cash    = [_get_q(bs_qs_ch, y, q, 'bsa2') for y,q in q_periods]
    _q_gp      = [_get_q(is_qs_ch, y, q, 'isa5') for y,q in q_periods]
    _q_cogs    = [abs(_get_q(is_qs_ch, y, q, 'isa4')) for y,q in q_periods]
    _q_gm      = [_q_gp[i]/max(_q_rev[i],1)*100 for i in range(9)]
    _q_nm      = [_q_ni[i]/max(_q_rev[i],1)*100 for i in range(9)]
    _q_de      = [_q_debt[i]/max(_q_equity[i],1) for i in range(9)]
    _q_intrev  = [_q_interest[i]/max(_q_rev[i],1)*100 for i in range(9)]
    _q_sgarev  = [_q_sgka[i]/max(_q_rev[i],1)*100 for i in range(9)]
    _q_recday  = [_q_rec[i]/max(_q_rev[i]/90,1) for i in range(9)]
    _q_invday  = [_q_inv[i]/max(_q_cogs[i]/90,1) for i in range(9)]
    _q_cash_ta = [_q_cash[i]/max(_q_ta[i],1)*100 for i in range(9)]
    _q_debt_ta = [_q_debt[i]/max(_q_ta[i],1)*100 for i in range(9)]

    fig, axes = plt.subplots(2, 3, figsize=(16, 9))
    x = np.arange(9)
    qlbl = ['Q1\n2024','Q2','Q3','Q4','Q1\n2025','Q2','Q3','Q4','Q1\n2026']

    # Panel 1: Revenue & NI
    ax = axes[0,0]
    bars = ax.bar(x, [v/1000 for v in _q_rev], color='#3498DB', alpha=0.6, label='DT (nghìn tỷ)')
    ax2 = ax.twinx()
    ax2.plot(x, [v/1000 for v in _q_ni], 'o-', color='#E74C3C', linewidth=2, label='LNST (nghìn tỷ)')
    for bar, val in zip(bars, _q_rev):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3, f'{val/1000:.1f}', ha='center', fontsize=6)
    ax.set_title('Doanh thu & LNST', fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(qlbl, fontsize=7)
    ax.legend(fontsize=7, loc='upper left'); ax2.legend(fontsize=7, loc='upper right')

    # Panel 2: Profit Margins
    ax = axes[0,1]
    ax.plot(x, _q_gm, 's-', color='#27AE60', linewidth=2, label='Biên LNG%')
    ax.plot(x, _q_nm, 'o-', color='#E74C3C', linewidth=2, label='Biên LNST%')
    ax.axhline(y=stats.median(_q_gm), color='#27AE60', linestyle='--', alpha=0.5, label=f'Median GM {stats.median(_q_gm):.1f}%')
    ax.axhline(y=stats.median(_q_nm), color='#E74C3C', linestyle='--', alpha=0.5)
    for i in range(9):
        ax.annotate(f'{_q_gm[i]:.1f}', (x[i], _q_gm[i]), fontsize=6, ha='center', va='bottom', color='#27AE60')
    ax.set_title('Biên lợi nhuận', fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(qlbl, fontsize=7)
    ax.legend(fontsize=7); ax.grid(axis='y', alpha=0.3)

    # Panel 3: D/E + Interest/Rev + SG&A/Rev
    ax = axes[0,2]
    ax.bar(x, _q_de, color='#8E44AD', alpha=0.5, label='D/E (x)')
    ax.plot(x, _q_intrev, 'v-', color='#E74C3C', linewidth=2, label='Lãi vay/DT%')
    ax.plot(x, _q_sgarev, '^-', color='#F39C12', linewidth=2, label='CPBH&QL/DT%')
    ax.set_title('Nợ & Chi phí', fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(qlbl, fontsize=7)
    ax.legend(fontsize=7); ax.grid(axis='y', alpha=0.3)

    # Panel 4: Receivables & Inventory days
    ax = axes[1,0]
    ax.bar(x-0.15, _q_recday, width=0.3, color='#2980B9', alpha=0.7, label='Phải thu (ngày)')
    ax.bar(x+0.15, _q_invday, width=0.3, color='#E67E22', alpha=0.7, label='Tồn kho (ngày)')
    ax.set_title('Phải thu & Tồn kho (ngày)', fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(qlbl, fontsize=7)
    ax.legend(fontsize=7); ax.grid(axis='y', alpha=0.3)

    # Panel 5: Cash/TA & Debt/TA
    ax = axes[1,1]
    ax.fill_between(x, _q_cash_ta, 0, alpha=0.3, color='#2ECC71', label='Tiền mặt/TS%')
    ax.fill_between(x, _q_debt_ta, 0, alpha=0.3, color='#E74C3C', label='Nợ vay/TS%')
    ax.plot(x, _q_cash_ta, 'o-', color='#27AE60', linewidth=2)
    ax.plot(x, _q_debt_ta, 's-', color='#C0392B', linewidth=2)
    ax.set_title('Cấu trúc TS (Tiền & Nợ/TS)', fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(qlbl, fontsize=7)
    ax.legend(fontsize=7); ax.grid(axis='y', alpha=0.3)

    # Panel 6: Revenue growth QoQ
    ax = axes[1,2]
    gr_qoq = [0]
    for i in range(1, 9):
        gr_qoq.append((_q_rev[i]/_q_rev[i-1]-1)*100 if _q_rev[i-1] else 0)
    colors = ['#2ECC71' if v >= 0 else '#E74C3C' for v in gr_qoq]
    ax.bar(x, gr_qoq, color=colors, alpha=0.7)
    ax.axhline(y=0, color='black', linewidth=0.5)
    for i, v in enumerate(gr_qoq):
        ax.text(i, v+(1 if v>=0 else -3), f'{v:+.0f}%', ha='center', fontsize=7, fontweight='bold')
    ax.set_title('Tăng trưởng DT QoQ', fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(qlbl, fontsize=7)
    ax.grid(axis='y', alpha=0.3)

    fig.suptitle('HPG — Chất lượng Tài sản & KQKD theo Quý', fontsize=14, fontweight='bold', y=1.01)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'quarterly_quality.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 7 (annual): Spread estimate vs Actual GP margin — annual
    wb_chart = openpyxl.load_workbook(EXCEL_FILE)
    ws2c = wb_chart['02_Assumptions']
    n_years = 8
    gp_excel = [ws2c.cell(row=6, column=j).value for j in range(2, 10)]
    wb_chart.close()
    spread_usd_a = SPREAD_A[:n_years]
    spread_pct_a = [spread_usd_a[i]/HRC_PRICE_A[i]*100 for i in range(n_years)]
    actual_gpm_a = [gp_excel[i] if isinstance(gp_excel[i], (int, float)) else all_gpm[i] for i in range(n_years)]

    fig, ax1 = plt.subplots(figsize=(10, 5))
    x_a = np.arange(len(all_years_lbl))
    ax1.plot(all_years_lbl, spread_pct_a, 's-', color='#2980B9', linewidth=2.5, markersize=7, label='Spread/HRC Price (%)')
    ax1.plot(all_years_lbl, actual_gpm_a, 'o-', color='#E74C3C', linewidth=2.5, markersize=7, label='GP Margin thực tế (%)')
    for i, v in enumerate(spread_pct_a):
        ax1.annotate(f'{v:.1f}%', (all_years_lbl[i], v), textcoords="offset points", xytext=(0, 10), ha='center', fontsize=8, color='#2980B9')
    for i, v in enumerate(actual_gpm_a):
        ax1.annotate(f'{v:.1f}%', (all_years_lbl[i], v), textcoords="offset points", xytext=(0, -15), ha='center', fontsize=8, color='#E74C3C')
    ax2 = ax1.twinx()
    ax2.bar(x_a, spread_usd_a, alpha=0.15, color='#2980B9', label='Spread ước tính (USD/t)', width=0.5)
    ax1.set_xlabel('Năm', fontsize=11)
    ax1.set_ylabel('%', fontsize=11)
    ax2.set_ylabel('Spread (USD/tấn)', fontsize=11)
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=9)
    ax1.set_title('HPG — Spread ước tính (lag 1 quý/năm) vs Biên LNG thực tế, theo Năm', fontsize=13, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'spread_vs_gp.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 7a (quarterly): Spread THỰC (lag 1 quý, từ Q18_SPREAD — không còn giá trị NĂM lặp lại
    # cho mọi quý như trước) kết hợp Biên LNG thực tế theo BCTC — 12 quý gần nhất, để đánh giá
    # Spread mở rộng/thu hẹp có đi cùng chiều với BLNG cải thiện/xấu đi hay không.
    q_is = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    _gpm_by_label = {}
    for rec in q_is:
        yr = rec.get("yearReport"); q = rec.get("lengthReport")
        if yr and q and q != 5:
            rev_q = (rec.get("isa3", 0) or 0) / 1e9
            gp_q = (rec.get("isa5", 0) or 0) / 1e9
            gpm_q = (gp_q / rev_q * 100) if rev_q else None
            if gpm_q is not None and 0 < gpm_q < 60:
                _gpm_by_label[f"{yr}Q{q}"] = round(gpm_q, 1)

    # Quý hiện tại (Q2/2026) chưa có BCTC — ước tính BLNG bằng tỷ lệ Spread All hiện tại/Spread All
    # quý gần nhất (q1_spread_all, q1_gpm — cùng phương pháp nội suy BLNG dự phóng đã dùng ở trên).
    _cur_lbl = "2026Q2"

    def _smooth_curve(x, y):
        """Nội suy PCHIP (Piecewise Cubic Hermite) để vẽ ĐƯỜNG mềm mại hơn qua các điểm dữ liệu THẬT
        (2026-07, theo yêu cầu user — chỉ làm mềm phần hiển thị trực quan ở các đoạn gấp khúc tăng/
        giảm, KHÔNG thay đổi số liệu gốc: marker + nhãn số vẫn vẽ đúng vị trí/giá trị dữ liệu thật).
        Dùng PCHIP (KHÔNG dùng spline bậc 3 tự do make_interp_spline) vì spline tự do bị OVERSHOOT/
        UNDERSHOOT mạnh giữa 2 điểm zigzag liên tiếp (tạo đỉnh/đáy ẢO không có trong số liệu thật —
        đã test và thấy rõ hiện tượng này), trong khi PCHIP đảm bảo đường cong không vượt quá khoảng
        giá trị của 2 điểm lân cận (monotonic giữa từng đoạn) — mềm hơn đường thẳng nhưng không bịa
        số liệu."""
        x = np.asarray(x, dtype=float); y = np.asarray(y, dtype=float)
        if len(x) < 3:
            return x, y
        x_dense = np.linspace(x.min(), x.max(), max(200, len(x) * 15))
        return x_dense, PchipInterpolator(x, y)(x_dense)

    def _make_spread_gpm_chart(spread_hist, spread_now_val, title_metric, fname, color='#2980B9'):
        """Vẽ 1 trong 3 biểu đồ Spread (HRC/Rebar/All) vs Biên LNG thực tế theo quý — dùng chung 1
        khung code cho cả 3 loại spread (2026-07, theo yêu cầu user hiển thị 3 biểu đồ riêng biệt,
        20 quý gần nhất, đường mềm, và phân biệt rõ điểm BLNG quý đang chạy là ƯỚC TÍNH chứ không
        phải số liệu đã có báo cáo). Bỏ qua các quý spread=None (VD Spread All 5 quý đầu chưa có dữ
        liệu sản lượng tách riêng)."""
        q_lbls_all = Q18_LABELS + [_cur_lbl]
        q_spread_all_ = list(spread_hist) + [round(spread_now_val, 0) if spread_now_val is not None else None]
        q_gpm_all_ = [_gpm_by_label.get(l) for l in Q18_LABELS] + [
            round(q1_gpm*100 * (spread_now_val/q1_spread_all), 1) if (spread_now_val and q1_spread_all) else None]
        idx_valid = [i for i in range(len(q_lbls_all)) if q_spread_all_[i] is not None]
        q_lbls_all = [q_lbls_all[i] for i in idx_valid]
        q_spread_all_ = [q_spread_all_[i] for i in idx_valid]
        q_gpm_all_ = [q_gpm_all_[i] for i in idx_valid]

        N_RECENT_Q = 20
        q_lbls, q_spread, q_gpm = q_lbls_all[-N_RECENT_Q:], q_spread_all_[-N_RECENT_Q:], q_gpm_all_[-N_RECENT_Q:]
        nq_s = len(q_lbls)
        if nq_s == 0:
            return
        fig, ax1 = plt.subplots(figsize=(13, 5.5))
        x_s = np.arange(nq_s)
        x_lbl_disp = [f"{l[4:]}'{l[2:4]}" for l in q_lbls]
        # Đường Spread — làm mềm bằng spline (Spread tính đủ cho mọi quý trong cửa sổ hiển thị, không
        # có gap None vì đã lọc idx_valid ở trên), marker vẫn ở đúng vị trí dữ liệu thật.
        xs_smooth, ys_smooth = _smooth_curve(x_s, q_spread)
        ax1.plot(xs_smooth, ys_smooth, '-', color=color, linewidth=2, alpha=0.85, zorder=2)
        ax1.plot(x_s, q_spread, 'o', color=color, markersize=5, label=f'{title_metric} (USD/t) — tới Spread hiện tại', zorder=3)
        ax1.plot(x_s[-1], q_spread[-1], 'o', color=color, markersize=10, markeredgecolor='black', zorder=4)
        for i, v in enumerate(q_spread):
            ax1.annotate(f'{v:.0f}', (x_s[i], v), textcoords="offset points", xytext=(0, 8), ha='center', fontsize=7, color=color)
        ax2 = ax1.twinx()
        # Đường BLNG — chỉ làm mềm PHẦN THỰC TẾ đã có BCTC (liên tục), KHÔNG gộp điểm ước tính quý
        # đang chạy vào cùng 1 spline (tránh spline "kéo cong" ngược để khớp 1 điểm dự báo, gây hiểu
        # lầm điểm đó cũng là số liệu thật). LƯU Ý: q_gpm[-1] KHÔNG bao giờ là None khi quý đang chạy
        # nằm trong cửa sổ hiển thị — vị trí này luôn được điền sẵn bằng công thức ước tính (không
        # phải BCTC thật) ngay từ lúc dựng q_gpm_all_ ở trên, nên phải nhận diện quý dự báo qua NHÃN
        # (so khớp `_cur_lbl`), không thể suy ra từ giá trị None/not-None.
        is_forecast_last = (q_lbls[-1] == _cur_lbl)
        forecast_idx = nq_s - 1 if is_forecast_last else None
        gpm_x = [x_s[i] for i in range(nq_s) if q_gpm[i] is not None and i != forecast_idx]
        gpm_y = [q_gpm[i] for i in range(nq_s) if q_gpm[i] is not None and i != forecast_idx]
        if len(gpm_x) >= 4:
            gx_smooth, gy_smooth = _smooth_curve(gpm_x, gpm_y)
            ax2.plot(gx_smooth, gy_smooth, '-', color='#E67E22', linewidth=2.5, alpha=0.85, zorder=2)
        ax2.plot(gpm_x, gpm_y, 's', color='#E67E22', markersize=6, label='Biên LNG thực tế đã có BCTC (%)', zorder=3)
        if is_forecast_last and q_gpm[-1] is not None:
            # Nối NÉT ĐỨT (không làm mềm) từ điểm THỰC gần nhất tới điểm ƯỚC TÍNH quý đang chạy, và
            # dùng màu/marker khác hẳn (tam giác xanh lá, không phải vuông cam) — phân biệt rõ đây là
            # SỐ DỰ BÁO do model tự tính (chưa có BCTC), theo yêu cầu user (áp dụng cho cả 3 biểu đồ).
            if gpm_x:
                ax2.plot([gpm_x[-1], x_s[-1]], [gpm_y[-1], q_gpm[-1]], '--', color='#27AE60', linewidth=1.6, zorder=2)
            ax2.plot(x_s[-1], q_gpm[-1], '^', color='#27AE60', markersize=11, zorder=4,
                      label=f'BLNG ƯỚC TÍNH {_cur_lbl} (dự báo — chưa có BCTC)')
        ax1.set_xlabel('Quý', fontsize=11)
        ax1.set_ylabel(f'{title_metric} (USD/t)', fontsize=11, color=color)
        ax2.set_ylabel('Biên LNG (%)', fontsize=11, color='#E67E22')
        ax1.set_xticks(x_s)
        ax1.set_xticklabels(x_lbl_disp, fontsize=8, rotation=45)
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=8)
        ax1.set_title(f'HPG — {title_metric} (giá bán quý này, chi phí quặng/than lag 1 quý) & Biên LNG, {nq_s} quý gần nhất', fontsize=12.5, fontweight='bold')
        ax1.grid(axis='y', alpha=0.3)
        fig.tight_layout()
        fig.savefig(os.path.join(CHART_DIR, fname), dpi=200, bbox_inches='tight')
        plt.close(fig)

    # 3 biểu đồ riêng biệt theo yêu cầu user: Spread HRC (đã có thuế CBPG), Spread Rebar (thép XD),
    # Spread All (bình quân gia quyền theo sản lượng) — mỗi biểu đồ so sánh với BLNG quý.
    _make_spread_gpm_chart(Q18_SPREAD_HRC, SPREAD_HRC_NOW, 'Spread HRC', 'spread_gp_quarterly.png', color='#2980B9')
    _make_spread_gpm_chart(Q18_SPREAD_REBAR, SPREAD_REBAR_NOW, 'Spread Rebar (thép XD)', 'spread_rebar_gpm_quarterly.png', color='#8E44AD')
    _make_spread_gpm_chart(Q18_SPREAD_ALL, SPREAD_ALL_NOW, 'Spread All', 'spread_all_gpm_quarterly.png', color='#16A085')

    # Chart 7b/7c: TƯƠNG QUAN Spread vs Biên LNG — scatter + đường hồi quy tuyến tính + hệ số Pearson
    # r, theo NĂM và theo QUÝ. Chỉ dùng dữ liệu THỰC TẾ (không lấy các năm/quý dự phóng vì Biên LNG
    # dự phóng được NỘI SUY TỪ chính tỷ lệ Spread — đưa vào sẽ tạo tương quan giả/circular).
    def _scatter_corr(x, y, title, fname, xlabel='Spread (USD/tấn)', annotate=None):
        x = np.array(x, dtype=float); y = np.array(y, dtype=float)
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.scatter(x, y, s=90, color='#2980B9', zorder=3, edgecolor='white', linewidth=0.8)
        if annotate:
            for xi, yi, lbl in zip(x, y, annotate):
                ax.annotate(lbl, (xi, yi), textcoords="offset points", xytext=(6, 6), fontsize=8, color='#555555')
        r = None
        if len(x) > 2 and np.std(x) > 0:
            r = np.corrcoef(x, y)[0, 1]
            slope, intercept = np.polyfit(x, y, 1)
            x_line = np.linspace(x.min(), x.max(), 50)
            ax.plot(x_line, slope*x_line + intercept, '--', color='#E74C3C', linewidth=2,
                    label=f'Hồi quy tuyến tính (r = {r:.2f})')
            ax.legend(loc='best', fontsize=9)
        ax.set_xlabel(xlabel, fontsize=11)
        ax.set_ylabel('Biên LNG thực tế (%)', fontsize=11)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.grid(alpha=0.3)
        fig.tight_layout()
        fig.savefig(os.path.join(CHART_DIR, fname), dpi=200, bbox_inches='tight')
        plt.close(fig)
        return r

    # Theo NĂM: Spread năm (lag-1-quý/năm, SPREAD_A) vs Biên LNG THỰC TẾ — chỉ 2021-2025 (đã công bố)
    corr_annual = _scatter_corr(
        SPREAD_A[:5], gp_margin_hist[:5],
        'HPG — Tương quan Spread & Biên LNG thực tế, theo Năm (2021-2025)',
        'spread_gpm_corr_annual.png', annotate=[str(y) for y in years_hist])

    # Theo QUÝ: Spread quý (lag-1-quý, Q18_SPREAD) vs Biên LNG thực tế BCTC — mọi quý có đủ 2 số liệu
    _corr_q_pairs = [(Q18_SPREAD[i], _gpm_by_label[Q18_LABELS[i]], Q18_LABELS[i])
                     for i in range(len(Q18_LABELS)) if Q18_LABELS[i] in _gpm_by_label]
    if len(_corr_q_pairs) > 2:
        corr_quarterly = _scatter_corr(
            [p[0] for p in _corr_q_pairs], [p[1] for p in _corr_q_pairs],
            f'HPG — Tương quan Spread & Biên LNG thực tế, theo Quý ({_corr_q_pairs[0][2]}-{_corr_q_pairs[-1][2]})',
            'spread_gpm_corr_quarterly.png',
            annotate=[f"{p[2][4:]}'{p[2][2:4]}" for p in _corr_q_pairs])

    # Chart 8b: Giá HRC/Quặng sắt/Than cốc 18 quý gần nhất — "2 cột giá" (dual-axis) vì HRC
    # (~460-870 USD/t) và Quặng+Than (~95-500 USD/t) lệch scale, gộp 1 trục sẽ khó đọc.
    cq_lbls = Q18_LABELS + ["Hiện tại"]
    cq_hrc = Q18_HRC + [hrc_now]
    cq_iron = Q18_IRON + [iron_now]
    cq_coal = Q18_COAL + [coal_now]
    xq = np.arange(len(cq_lbls))
    fig, ax1 = plt.subplots(figsize=(14, 5.5))
    ax1.plot(xq, cq_hrc, 'o-', color='#C0392B', linewidth=2.2, markersize=4, label='Giá HRC (USD/t)', zorder=3)
    ax1.plot(xq[-1], cq_hrc[-1], 'o', color='#C0392B', markersize=9, markeredgecolor='black', zorder=4)
    ax2 = ax1.twinx()
    ax2.plot(xq, cq_iron, 's-', color='#2980B9', linewidth=1.8, markersize=4, label='Quặng sắt 62%Fe (USD/t)', zorder=3)
    ax2.plot(xq, cq_coal, '^-', color='#27AE60', linewidth=1.8, markersize=4, label='Than cốc luyện kim (USD/t)', zorder=3)
    ax1.set_xlabel('Quý', fontsize=11)
    ax1.set_ylabel('Giá HRC (USD/t)', fontsize=11, color='#C0392B')
    ax2.set_ylabel('Giá Quặng sắt / Than cốc (USD/t)', fontsize=11, color='#2980B9')
    ax1.set_xticks(xq)
    ax1.set_xticklabels(cq_lbls, fontsize=7, rotation=45)
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper right', fontsize=8)
    ax1.set_title('HPG — Giá HRC / Quặng sắt / Than cốc, 18 quý gần nhất + Hiện tại', fontsize=13, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(CHART_DIR, 'commodity_prices_18q.png'), dpi=200, bbox_inches='tight')
    plt.close(fig)

    # Chart 9: Quarterly Sales Volume Line Chart (HRC & XD) — read from Excel ws15
    wb_vol = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws15c = wb_vol['15_Quarterly_Data']
    vol_labels = ['SL HRC (nghìn tấn)', 'SL XD & Thép CB (nghìn tấn)']
    vol_start = None
    for row in range(1, ws15c.max_row+1):
        v = ws15c.cell(row=row, column=1).value
        if v and 'SẢN LƯỢNG TIÊU THỤ QUÝ' in str(v):
            vol_start = row + 2
            break
    if vol_start:
        qv_header_row = vol_start - 1  # row with Q1/2023, Q2/2023, ...
        qv_rows = {}
        max_col = ws15c.max_column
        for i, lbl in enumerate(vol_labels):
            r = vol_start + i  # HRC at vol_start, XD at vol_start+1
            vals = [ws15c.cell(row=r, column=j).value for j in range(2, max_col+1)]
            vals = [v if isinstance(v, (int, float)) else 0 for v in vals]
            qv_rows[lbl] = vals
        qv_headers = [str(ws15c.cell(row=qv_header_row, column=j).value or '') for j in range(2, max_col+1)]
        # Trim trailing empty headers & zero values
        while qv_headers and not qv_headers[-1]:
            qv_headers.pop()
            for lbl in vol_labels:
                qv_rows[lbl].pop()
        nq = len(qv_headers)
    wb_vol.close()

    if vol_start and nq > 0:
        fig, ax = plt.subplots(figsize=(14, 5.5))
        x = np.arange(nq)
        hr = qv_rows[vol_labels[0]]
        xd = qv_rows[vol_labels[1]]
        ax.plot(x, hr, 'o-', color='#1F4E79', linewidth=2.5, markersize=6, label='HRC', zorder=3)
        ax.plot(x, xd, 's-', color='#E74C3C', linewidth=2.5, markersize=6, label='Thép XD', zorder=3)
        for i in range(nq):
            if hr[i] > 0:
                ax.annotate(f'{hr[i]:.0f}', (x[i], hr[i]), textcoords="offset points", xytext=(0, 10), ha='center', fontsize=7, color='#1F4E79')
            if xd[i] > 0:
                ax.annotate(f'{xd[i]:.0f}', (x[i], xd[i]), textcoords="offset points", xytext=(0, -14), ha='center', fontsize=7, color='#E74C3C')
        ax.set_xlabel('Quý', fontsize=11)
        ax.set_ylabel('nghìn tấn', fontsize=11)
        ax.set_xticks(x)
        ax.set_xticklabels(qv_headers, fontsize=7, rotation=45)
        ax.set_title('HPG – Sản lượng HRC & Thép XD theo Quý', fontsize=13, fontweight='bold')
        ax.legend(loc='upper left', fontsize=10)
        ax.set_ylim(min(min(hr), min(xd)) * 0.85, max(max(hr), max(xd)) * 1.35)
        ax.yaxis.set_major_locator(mticker.MultipleLocator(200))
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'{v:,.0f}'))
        ax.grid(axis='y', alpha=0.3)
        fig.tight_layout()
        fig.savefig(os.path.join(CHART_DIR, 'quarterly_volume.png'), dpi=200, bbox_inches='tight')
        plt.close(fig)

    # ── Quarterly PE / PB / EV data from HPG_RATIOS ──
    q_pe_pts = []; q_pb_pts = []; q_ev_pts = []; q_labels = []
    for r in HPG_RATIOS:
        q = r.get("quarter")
        if q is None or q == 5: continue
        y = int(r["year"]); qnum = int(q)
        lbl = f"Q{qnum}-{y}"
        pe0 = r.get("pe"); pb0 = r.get("pb"); ev0 = r.get("ev_ebitda")
        if pe0 and 0 < pe0 < 50: q_pe_pts.append(round(pe0,1))
        else: q_pe_pts.append(None)
        q_pb_pts.append(round(pb0,2) if (pb0 and pb0>0) else None)
        q_ev_pts.append(round(ev0,1) if (ev0 and ev0>0) else None)
        q_labels.append(lbl)
    n_ratio_q = len(q_labels)
    x_r = np.arange(n_ratio_q)

    # Chart: P/E Quarterly Line
    fig, ax = plt.subplots(figsize=(13, 5))
    pe_vals = [v if v else None for v in q_pe_pts]
    ax.plot(x_r, pe_vals, 'o-', color='#1F4E79', linewidth=2, markersize=5, label='P/E', zorder=3)
    pe_med = stats.median([v for v in pe_vals if v]) if any(v for v in pe_vals if v) else 11.0
    ax.axhline(y=pe_med, color='#1F4E79', linestyle='--', alpha=0.5, linewidth=1, label=f'Median {pe_med:.1f}x')
    last_pe = [v for v in pe_vals if v][-1] if any(v for v in pe_vals if v) else 0
    ax.annotate(f'Hiện tại: {last_pe:.1f}x', (n_ratio_q-1, last_pe), textcoords="offset points",
                xytext=(10,10), ha='left', fontsize=9, color='#C0392B', fontweight='bold')
    ax.set_xlabel('Quý', fontsize=10); ax.set_ylabel('P/E (x)', fontsize=10)
    ax.set_xticks(np.arange(0, n_ratio_q, max(1, n_ratio_q//12)))
    ax.set_xticklabels([q_labels[i] for i in range(0, n_ratio_q, max(1, n_ratio_q//12))], fontsize=7, rotation=45)
    ax.set_title('HPG — P/E theo Quý', fontsize=13, fontweight='bold')
    ax.set_ylim(0, max([v for v in pe_vals if v] or [30])*1.2)
    ax.yaxis.set_major_locator(mticker.MultipleLocator(5))
    ax.legend(loc='upper left', fontsize=9); ax.grid(axis='y', alpha=0.3)
    fig.tight_layout(); fig.savefig(os.path.join(CHART_DIR, 'pe_hist_q.png'), dpi=200, bbox_inches='tight'); plt.close(fig)

    # Chart: P/B Quarterly Line
    fig, ax = plt.subplots(figsize=(13, 5))
    pb_vals = [v if v else None for v in q_pb_pts]
    ax.plot(x_r, pb_vals, 's-', color='#E74C3C', linewidth=2, markersize=5, label='P/B', zorder=3)
    pb_med = stats.median([v for v in pb_vals if v]) if any(v for v in pb_vals if v) else 1.6
    ax.axhline(y=pb_med, color='#E74C3C', linestyle='--', alpha=0.5, linewidth=1, label=f'Median {pb_med:.2f}x')
    ax.fill_between(x_r, pb_med*0.8, pb_med*1.2, alpha=0.08, color='#E74C3C')
    last_pb = [v for v in pb_vals if v][-1] if any(v for v in pb_vals if v) else 0
    ax.annotate(f'Hiện tại: {last_pb:.2f}x', (n_ratio_q-1, last_pb), textcoords="offset points",
                xytext=(10,10), ha='left', fontsize=9, color='#C0392B', fontweight='bold')
    ax.set_xlabel('Quý', fontsize=10); ax.set_ylabel('P/B (x)', fontsize=10)
    ax.set_xticks(np.arange(0, n_ratio_q, max(1, n_ratio_q//12)))
    ax.set_xticklabels([q_labels[i] for i in range(0, n_ratio_q, max(1, n_ratio_q//12))], fontsize=7, rotation=45)
    ax.set_title('HPG — P/B theo Quý', fontsize=13, fontweight='bold')
    ax.set_ylim(0, max([v for v in pb_vals if v] or [4])*1.2)
    ax.yaxis.set_major_locator(mticker.MultipleLocator(0.5))
    ax.legend(loc='upper left', fontsize=9); ax.grid(axis='y', alpha=0.3)
    fig.tight_layout(); fig.savefig(os.path.join(CHART_DIR, 'pb_hist_q.png'), dpi=200, bbox_inches='tight'); plt.close(fig)

    # Chart: EV/EBITDA Quarterly Line
    fig, ax = plt.subplots(figsize=(13, 5))
    ev_vals = [v if v else None for v in q_ev_pts]
    ax.plot(x_r, ev_vals, 'D-', color='#27AE60', linewidth=2, markersize=5, label='EV/EBITDA', zorder=3)
    ev_med = stats.median([v for v in ev_vals if v]) if any(v for v in ev_vals if v) else 9.0
    ax.axhline(y=ev_med, color='#27AE60', linestyle='--', alpha=0.5, linewidth=1, label=f'Median {ev_med:.1f}x')
    ax.fill_between(x_r, ev_med*0.75, ev_med*1.25, alpha=0.08, color='#27AE60')
    ax.axhspan(ev_med*0.5, ev_med*0.75, alpha=0.07, color='#27AE60', label='Vùng mua')
    ax.axhspan(ev_med*1.3, ev_med*1.8, alpha=0.07, color='#E74C3C', label='Vùng bán')
    last_ev = [v for v in ev_vals if v][-1] if any(v for v in ev_vals if v) else 0
    ax.annotate(f'Hiện tại: {last_ev:.1f}x', (n_ratio_q-1, last_ev), textcoords="offset points",
                xytext=(10,10), ha='left', fontsize=9, color='#C0392B', fontweight='bold')
    ax.set_xlabel('Quý', fontsize=10); ax.set_ylabel('EV/EBITDA (x)', fontsize=10)
    ax.set_xticks(np.arange(0, n_ratio_q, max(1, n_ratio_q//12)))
    ax.set_xticklabels([q_labels[i] for i in range(0, n_ratio_q, max(1, n_ratio_q//12))], fontsize=7, rotation=45)
    ax.set_title('HPG — EV/EBITDA theo Quý', fontsize=13, fontweight='bold')
    ax.set_ylim(0, max([v for v in ev_vals if v] or [25])*1.2)
    ax.yaxis.set_major_locator(mticker.MultipleLocator(2))
    ax.legend(loc='upper left', fontsize=9); ax.grid(axis='y', alpha=0.3)
    fig.tight_layout(); fig.savefig(os.path.join(CHART_DIR, 'ev_hist_q.png'), dpi=200, bbox_inches='tight'); plt.close(fig)

    print(f"[OK] Charts saved: {CHART_DIR}")
    return True


# ── 3. PDF REPORT ───────────────────────────────────────────────────────────

def build_pdf():
    # ── Quarterly data helpers ──
    def get_q(records, yr, qtr, field):
        for r in records:
            if r.get("yearReport") == yr and r.get("lengthReport") == qtr:
                v = r.get(field)
                return v / 1e9 if v is not None else 0
        return 0
    is_qs = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    bs_qs = section_to_quarters(FIN_DATA, "BALANCE_SHEET")
    cf_qs = section_to_quarters(FIN_DATA, "CASH_FLOW")

    # Try to register Vietnamese-supporting font (Arial on Win, DejaVu on Linux)
    FONT = 'Helvetica'
    FONT_BOLD = 'Helvetica-Bold'
    FONT_ITALIC = 'Helvetica-Oblique'
    _font_ok = False
    _font_candidates = [
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf", "C:/Windows/Fonts/ariali.ttf",
         "Arial", "Arial-Bold", "Arial-Italic"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
         "Arial", "Arial-Bold", "Arial-Italic"),
    ]
    for rp, bp, ip, rn, bn, _in in _font_candidates:
        _r = os.path.exists(rp); _b = os.path.exists(bp); _i = os.path.exists(ip)
        if _r and _b:
            try:
                pdfmetrics.registerFont(TTFont(rn, rp))
                pdfmetrics.registerFont(TTFont(bn, bp))
                if _i:
                    pdfmetrics.registerFont(TTFont(_in, ip))
                FONT = rn; FONT_BOLD = bn; FONT_ITALIC = _in if _i else rn
                _font_ok = True
                break
            except:
                continue

    styles = getSampleStyleSheet()
    # Override/add styles
    styles.add(ParagraphStyle('CoverTitle', fontName=FONT_BOLD, fontSize=26,
                               alignment=TA_CENTER, spaceAfter=6, textColor=HexColor('#1F4E79')))
    styles.add(ParagraphStyle('CoverSub', fontName=FONT, fontSize=14,
                               alignment=TA_CENTER, spaceAfter=4, textColor=HexColor('#555555')))
    styles.add(ParagraphStyle('SectionTitle', fontName=FONT_BOLD, fontSize=16,
                               spaceBefore=16, spaceAfter=8, textColor=HexColor('#1F4E79'),
                               borderWidth=0, borderPadding=0))
    styles.add(ParagraphStyle('BodyText2', fontName=FONT, fontSize=10,
                               leading=14, spaceAfter=6, alignment=TA_JUSTIFY))
    styles.add(ParagraphStyle('BodyBold', fontName=FONT_BOLD, fontSize=10,
                               leading=14, spaceAfter=4))
    styles.add(ParagraphStyle('SmallText', fontName=FONT, fontSize=8, leading=10,
                               textColor=HexColor('#888888')))
    styles.add(ParagraphStyle('TableCell', fontName=FONT, fontSize=9, leading=11))
    styles.add(ParagraphStyle('TableCellBold', fontName=FONT_BOLD, fontSize=9, leading=11))

    doc = SimpleDocTemplate(
        PDF_FILE, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm
    )

    # ── Valuation calculations (matching Excel PnL derivation) ──
    # EV_MULTIPLE/PB_MULTIPLE/PE_MULTIPLE và net-debt 2026E dùng ĐÚNG nguồn số đã ghi vào Excel
    # (07_Valuation/02_Assumptions) — xem PE_HIST_MEDIAN/PB_HIST_MEDIAN/EV_HIST_MEDIAN/NET_DEBT_2026E
    # ở đầu file — không tính lại bằng hằng số riêng để tránh PDF/Excel/JSON lệch nhau.
    EV_MULTIPLE = EV_HIST_MEDIAN
    PB_MULTIPLE = PB_HIST_MEDIAN
    PE_MULTIPLE = PE_HIST_MEDIAN
    # PnL-matching: EBITDA = EBIT + D&A, dùng lại ebit_fc/ni_fc/equity_fc_val bottom-up ở trên
    # (thay vì tính lại độc lập bằng số gõ tay 148000/2500/3800/0.035 dễ lệch khỏi 04_PnL).
    ebitda_2026e_val = ebit_fc[0] + da_fc[0]
    net_debt_2026e_val = NET_DEBT_2026E
    eps_2026e_val = ni_fc[0] * 1e9 / SHARES
    bvps_2026e_val = equity_fc_val[0] * 1e9 / SHARES

    price_ev_ebitda_val = max(0, (ebitda_2026e_val * EV_MULTIPLE - net_debt_2026e_val) * 1e9 / SHARES)
    price_pb_val = PB_MULTIPLE * bvps_2026e_val
    price_pe_val = PE_MULTIPLE * eps_2026e_val
    weighted_price_val = price_ev_ebitda_val * 0.4 + price_pb_val * 0.4 + price_pe_val * 0.2
    upside_val = (weighted_price_val / PRICE - 1) * 100
    # Individual valuations for display
    pb_upper_val = round(PB_MULTIPLE * 1.2 * bvps_2026e_val)
    pb_attr_val  = round(PB_MULTIPLE * 0.8 * bvps_2026e_val)

    elements = []

    # Helper functions
    def add_section(title):
        elements.append(Paragraph(title, styles['SectionTitle']))
        elements.append(Spacer(1, 2*mm))

    def add_body(text):
        elements.append(Paragraph(text, styles['BodyText2']))

    def add_table(headers, data, col_widths=None):
        table_data = [[Paragraph(h, styles['TableCellBold']) for h in headers]]
        for row in data:
            table_data.append([Paragraph(str(c), styles['TableCell']) for c in row])
        if col_widths is None:
            col_widths = [doc.width / len(headers)] * len(headers)
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#F5F7FA'), HexColor('#FFFFFF')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(tbl)
        elements.append(Spacer(1, 4*mm))

    # ─── Cover ───
    elements.append(Spacer(1, 40*mm))
    elements.append(Paragraph("BÁO CÁO PHÂN TÍCH CỔ PHIẾU", styles['CoverTitle']))
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph(f"{TICKER} — {COMPANY}", styles['CoverTitle']))
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph("Ngành: Thép  |  Sàn: HOSE", styles['CoverSub']))
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph(f"Giá hiện tại: {PRICE:,} VND  |  Vốn hóa: {MARKET_CAP/1e9:,.0f} tỷ", styles['CoverSub']))
    elements.append(Spacer(1, 10*mm))
    elements.append(Paragraph("KHUYẾN NGHỊ: <b>MUA</b>", ParagraphStyle('CoverSubBold',
        fontName=FONT_BOLD, fontSize=20, alignment=TA_CENTER,
        textColor=HexColor('#27AE60'))))
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph(f"Giá mục tiêu: <b>{weighted_price_val:,.0f} VND</b>  |  Upside: <b>+{upside_val:.0f}%</b>",
        ParagraphStyle('CoverSub2', fontName=FONT, fontSize=14, alignment=TA_CENTER,
                       textColor=HexColor('#1F4E79'))))
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph(
        f"EV/EBITDA ({EV_MULTIPLE}x): {price_ev_ebitda_val:,.0f} VND  |  "
        f"P/B ({PB_MULTIPLE}x): {price_pb_val:,.0f} VND  |  "
        f"P/E ({PE_MULTIPLE}x): {price_pe_val:,.0f} VND",
        ParagraphStyle('CoverSub', fontName=FONT, fontSize=11, alignment=TA_CENTER,
                       textColor=HexColor('#555555'))))
    elements.append(Paragraph(
        f"P/B hấp dẫn ({PB_MULTIPLE*0.8:.1f}x): {pb_attr_val:,} VND  |  "
        f"P/B over ({PB_MULTIPLE*1.2:.1f}x): {pb_upper_val:,} VND",
        ParagraphStyle('CoverSub', fontName=FONT, fontSize=11, alignment=TA_CENTER,
                       textColor=HexColor('#555555'))))
    elements.append(Spacer(1, 16*mm))
    elements.append(Paragraph(f"Ngày phân tích: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Chuyên viên phân tích: AI FA Framework v2.0",
        styles['SmallText']))
    elements.append(PageBreak())

    # ─── Investment Summary ───
    add_section("INVESTMENT SUMMARY")
    add_body(
        "HPG là nhà sản xuất thép tích hợp dọc lớn nhất Việt Nam với 6 nhà máy đang hoạt động "
        "(công suất ~14,5 triệu tấn/năm + 500k TEUs container), dẫn đầu cả thép xây dựng và HRC. "
        "Nhà máy Dung Quất 2 (5.6 triệu tấn HRC) đã vận hành full công suất từ tháng 12/2025, "
        "đưa HPG vào nhóm sản xuất thép chi phí thấp nhất khu vực. "
        "Thêm 3 dự án tương lai (Đắk Lắk 6M tấn, Ray cao tốc 700k tấn, Ống Long An 400k tấn) "
        "có thể nâng tổng công suất lên ~21,6 triệu tấn/năm."
    )
    add_body(
        f"<b>Khuyến nghị: MUA</b>  |  <b>Giá mục tiêu (Base): {weighted_price_val:,.0f} VND</b>  |  <b>Upside: +{upside_val:.0f}%</b>"
    )
    add_body("<b>Ba lý do mua:</b>")
    add_body("1. <b>DQ2 full công suất</b> — Sản lượng HRC tăng gấp đôi (~6 triệu tấn/năm), "
             f"biên LNG cải thiện từ 15.7% (2025) lên {gp_margin_fc[0]}% (2026E).")
    add_body("2. <b>Giá HRC hồi phục</b> — Thuế CBPG 27.8% với HRC TQ + nguồn cung toàn cầu thắt chặt "
             "(Gary Works #14 reline) hỗ trợ giá đầu ra.")
    add_body("3. <b>Định giá hấp dẫn</b> — P/B 1.56x (dưới median lịch sử HPG 1.61x), "
             "EV/EBITDA 2026E ~7.0x (dưới median HPG 8.95x).")
    add_body("<b>Ba rủi ro chính:</b>")
    add_body("1. Giá HRC giảm do suy thoái TQ hoặc nhu cầu toàn cầu yếu.")
    add_body("2. Chi phí quặng sắt/than cốc tăng làm thu hẹp spread.")
    add_body("3. Tỷ giá USD/VNĐ biến động mạnh (dư nợ vay USD ~$2 tỷ).")

    add_body("<b>Thông tin cổ phiếu & Thị trường:</b>")
    add_body(f"- Giá hiện tại: {PRICE:,} VND | Vốn hóa: {MARKET_CAP/1e9:,.0f} tỷ | KLGD BQ 20 phiên: 33.2 triệu cp (~890.7 tỷ)")
    add_body(f"- EPS TTM: {eps_hist[4]:,.0f} VND | BVPS: {equity_hist[4]*1e9/SHARES:,.0f} VND | P/E: 14.2x | P/B: 1.25x")
    add_body(f"- Cổ đông sáng lập: Trần Đình Long & gia đình (~35.7%)")
    add_body(f"- Biến động giá: 1 tháng +3.3% | 3 tháng +6.9% | YTD +6.1% (VNINDEX: +4.4% / -6.3% / -1.9%)")
    add_body(f"<b>Kế hoạch ĐHĐCĐ 2026 vs Thực hiện 2025:</b>")
    plan_dt_kh = 170000; plan_dt_th = revenue_hist[4]
    plan_ln_kh = 15000; plan_ln_th = ni_hist[4]
    add_body(f"- DT: KH {plan_dt_kh:,} tỷ → TH {plan_dt_th:,.0f} tỷ (hoàn thành {plan_dt_th/plan_dt_kh*100:.0f}%)")
    add_body(f"- LNST: KH {plan_ln_kh:,} tỷ → TH {plan_ln_th:,.0f} tỷ (hoàn thành {plan_ln_th/plan_ln_kh*100:.0f}% — vượt 3%)")
    add_body(f"- KH 2026: DT 210,000 tỷ (+34.5% YoY) | LNST 22,000 tỷ (+42.4% YoY)")
    add_body(f"- Q1/2026: LNST {get_q(is_qs,2026,1,'isa22'):,.0f} tỷ = {get_q(is_qs,2026,1,'isa22')/22000*100:.1f}% KH năm")
    add_body(f"- Q1 Core NPAT (loại DTTC ĐB ~4,915 tỷ từ Phố Nối): ~5,200 tỷ (+55% YoY)")
    add_body(f"- Broker targets: SSI 36k | VNDirect P/E 9.7x | SHS 218k DT/25k LNST | VCBS 38k | BVSC 38.65k")

    add_body("<b>Snapshot tài chính (2025-2028E):</b>")
    snap_headers = ["Chỉ tiêu", "2025", "2026E", "2027E", "2028E", "CAGR"]
    snap_data = [
        ["Doanh thu (tỷ)", f"{revenue_hist[4]:,.0f}", f"{revenue_fc[0]:,}",
         f"{revenue_fc[1]:,}", f"{revenue_fc[2]:,}", "20.0%"],
        ["LNST (tỷ)", f"{ni_hist[4]:,.0f}", f"{ni_fc[0]:,}",
         f"{ni_fc[1]:,}", f"{ni_fc[2]:,}", "30.0%"],
        [f"Biên LNG (%)", "15.7%", f"{gp_margin_fc[0]}%", f"{gp_margin_fc[1]}%", f"{gp_margin_fc[2]}%", "-"],
        ["EPS (VND)", "1,830", "2,600", "3,320", "4,030", "30.1%"],
        ["P/B (x)", "1.53", "1.35", "1.17", "1.02", "-"],
        ["EV/EBITDA (x)", "10.8", "7.0", "5.8", "5.0", "-"],
    ]
    add_table(snap_headers, snap_data,
              [doc.width*0.25, doc.width*0.15, doc.width*0.15, doc.width*0.15, doc.width*0.15, doc.width*0.15])

    # ─── 6-Tầng Analysis ───
    elements.append(PageBreak())
    add_section("1. CHUỖI GIÁ TRỊ NGÀNH THÉP")
    add_body(
        "Chuỗi giá trị thép: <b>Quặng sắt + Than cốc → Thiêu kết → Lò cao (BOF) → Thép lỏng → "
        "Đúc phôi → Cán nóng (HRC) / Cán dài (Thép XD)</b>. HPG là doanh nghiệp duy nhất ở VN "
        "tích hợp hoàn chỉnh từ khai thác quặng (trữ lượng 320 triệu tấn) đến sản xuất thép thành phẩm."
    )
    add_body("<b>Các nhà máy hiện hữu của HPG:</b>")
    fac_headers = ["Nhà máy", "Địa điểm", "Sản phẩm", "Công suất", "Trạng thái"]
    fac_data = [
        ["Dung Quất 1", "Quảng Ngãi", "Thép XD, HRC", "5,000,000 tấn/năm", "Hoạt động"],
        ["Dung Quất 2", "Quảng Ngãi", "Thép cuộn HRC", "5,600,000 tấn/năm", "Hoạt động"],
        ["Hải Dương", "Hải Dương", "Thép XD, phôi thép", "2,500,000 tấn/năm", "Hoạt động"],
        ["Ống thép HPG", "Hưng Yên", "Ống thép", "1,000,000 tấn/năm", "Hoạt động"],
        ["Tôn Hòa Phát", "Hưng Yên", "Tôn mạ kẽm, tôn màu", "400,000 tấn/năm", "Hoạt động"],
        ["Container HPG", "Q.Ngãi/H.Yên", "Vỏ container rỗng", "500,000 TEUs/năm", "Hoạt động"],
        ["Đắk Lắk", "Đắk Lắk", "Thép thanh lốp, chế tạo", "6,000,000 tấn/năm", "Chưa HĐ"],
        ["Ray cao tốc", "Q.Ngãi", "Thép ray", "700,000 tấn/năm", "Chưa HĐ"],
        ["Ống thép Long An", "Long An", "Ống thép", "400,000 tấn/năm", "Chưa HĐ"],
    ]
    add_table(fac_headers, fac_data,
              [doc.width*0.18, doc.width*0.15, doc.width*0.23, doc.width*0.22, doc.width*0.12])
    add_body(
        "<b>Tổng công suất thép thành phẩm đang hoạt động:</b> DQ1 (5,0M tấn) + DQ2 (5,6M) + "
        "Hải Dương (2,5M) + Ống thép (1,0M) + Tôn (0,4M) = <b>~14,5 triệu tấn/năm</b>. "
        "Container: 500,000 TEUs. Nếu thêm Đắk Lắk (6M) + Ray (0,7M) + Long An (0,4M), "
        "tổng tiềm năng đạt <b>~21,6 triệu tấn/năm</b>."
    )
    add_body(
        "<b>Capture value</b>: HPG chiếm ~40% thép xây dựng và ~70% HRC nội địa. "
        "Tự chủ 50% nhu cầu quặng sắt, giúp giảm chi phí ~15-20% so với đối thủ phải nhập khẩu. "
        "DQ2 (công nghệ Nhật Bản, Đức) giúp giảm tiêu hao điện và than cốc xuống mức tối ưu."
    )

    elements.append(Spacer(1, 4*mm))
    add_section("2. THỊ TRƯỜNG & VỊ TRÍ CẠNH TRANH")
    add_body(
        "<b>Porter 5 Forces:</b>"
    )
    add_body("- <b>Rào cản gia nhập (CAO)</b>: CAPEX tối thiểu $2-3 tỷ cho lò cao BOF + 3-5 năm xây dựng.")
    add_body("- <b>Nhà cung cấp (TRUNG BÌNH)</b>: Quặng sắt tập trung (Vale, Rio, BHP). HPG tự chủ 50% nên giảm áp lực.")
    add_body("- <b>Khách hàng (TRUNG BÌNH)</b>: Phân tán (Xây dựng, sản xuất tôn mạ, ống thép). Giá tham chiếu quốc tế.")
    add_body("- <b>Sản phẩm thay thế (THẤP)</b>: Nhôm, nhựa không thay thế được kết cấu thép.")
    add_body("- <b>Cạnh tranh nội bộ (CAO)</b>: HSG, NKG ở mảng tôn mạ; HRC TQ, Ấn Độ, Hàn Quốc cạnh tranh nhập khẩu.")
    add_body(
        "Tổng cầu thép VN 2026E ~30 triệu tấn, tăng 8% YoY (đầu tư công + BĐS phục hồi). "
        "HPG dẫn đầu thị phần thép XD (~40%) và HRC (~70% nội địa). "
        "Thuế CBPG HRC TQ 27.8% (từ 7/2025) tạo lợi thế cạnh tranh cho HPG so với hàng nhập khẩu."
    )

    elements.append(Spacer(1, 4*mm))
    add_section("3. LỢI THẾ CẠNH TRANH (MOAT)")
    add_body(
        "<b>Đánh giá Moat: RỘNG</b>"
    )
    add_body("1. <b>Chi phí thấp (Cost Advantage)</b>: Quy mô lớn nhất VN (14,5M tấn + 500k TEUs). "
             "DQ2 sử dụng công nghệ lò cao BOF hiện đại, tiết kiệm 15-20% chi phí sản xuất. "
             "Tự chủ 50% quặng sắt giảm chi phí logistics.")
    add_body("2. <b>Hiệu suất theo quy mô (Efficient Scale)</b>: Thị trường VN chỉ đủ chỗ cho "
             "1-2 nhà máy HRC quy mô lớn. DQ1+DQ2 đã chiếm ~70% sản lượng HRC nội địa.")
    add_body("3. <b>Tài sản vô hình (Intangible)</b>: Thương hiệu Hòa Phát gắn liền với công trình "
             "lớn (cầu Vĩnh Tuy, cao tốc Bắc-Nam, Long Thành). Hệ thống phân phối 300+ đại lý.")
    add_body("4. ROIC duy trì 12-18% qua chu kỳ (trên WACC ~9%), chứng tỏ lợi thế cạnh tranh bền vững.")

    # ─── Spread & Input/Output Analysis ───
    elements.append(Spacer(1, 4*mm))
    add_section("SPREAD & YẾU TỐ ĐẦU VÀO")
    add_body(
        "Lợi nhuận ngành thép được quyết định bởi <b>Spread</b> — chênh lệch giữa giá bán đầu ra "
        "và chi phí nguyên liệu đầu vào. Công thức chuẩn cho lò cao BOF, có tính LAG 1 QUÝ vì số "
        f"ngày tồn kho bình quân của HPG dao động quanh ~90 ngày (~1 quý — xem DIO_A mục 4C/sheet "
        "14_Steel_Analysis): giá vốn kỳ này phản ánh giá nguyên liệu mua vào từ kỳ TRƯỚC, trong khi "
        "giá bán vẫn là giá của kỳ hiện tại:"
    )
    add_body(
        f"<b>Spread(quý T) = Giá HRC bình quân quý T - 1.6×Giá quặng sắt 62%Fe bình quân quý T-1 "
        f"- 0.6×Giá than cốc bình quân quý T-1 - Chi phí SX khác cố định ({OTHER_COST_USD} USD/tấn)</b>"
    )
    add_body(
        "<b>Diễn biến Spread HPG 2021-2028E</b> (giá lấy từ sheet 17_Gia_Hang_Hoa: giá năm lịch sử = MEDIAN 4 quý, "
        "giá hiện tại fetch tự động từ investing.com; Spread năm 2021-2025 = MEDIAN Spread quý lag-1-quý trong năm đó):"
    )
    spread_headers = ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"]
    spread_data = [
        ["Giá HRC (USD/tấn)"] + [f"{v:,.0f}" for v in HRC_PRICE_A],
        ["Giá quặng 62%Fe (USD/tấn)"] + [f"{v:,.0f}" for v in IRON_ORE_A],
        ["Giá than cốc (USD/tấn)"] + [f"{v:,.0f}" for v in COKE_A],
        ["Spread ước tính (USD/tấn)"] + [f"{v:,.0f}" for v in SPREAD_A],
        ["Spread/HRC Price (%)"] + [f"{(SPREAD_A[i]/HRC_PRICE_A[i]*100):.0f}%" for i in range(8)],
    ]
    add_table(spread_headers, spread_data,
              [doc.width*0.30] + [doc.width*0.0875]*8)
    _spr_min_i = min(range(8), key=lambda i: SPREAD_A[i])
    _spr_yr = (years_hist + years_fc)[_spr_min_i]
    add_body(
        f"<b>Nhận xét:</b> Với dữ liệu giá hàng hóa thực tế (median 4 quý/năm, giá hiện tại fetch trực tiếp từ "
        f"investing.com), spread thấp nhất rơi vào {_spr_yr} (~{SPREAD_A[_spr_min_i]:,.0f} USD/tấn). "
        f"Spread {years_fc[0]}E ước tính ~{SPREAD_A[5]:,.0f} USD/tấn dựa trên giá hàng hóa hiện tại — "
        "đây là chỉ báo XU HƯỚNG dựa trên giá benchmark quốc tế (chưa tính lợi thế tự chủ nguyên liệu, "
        "hợp đồng dài hạn của HPG), KHÔNG phải biên lợi nhuận thực tế. Biên LNG dự phóng (mục Biên LNG ở trên) "
        "được neo vào biên LNG THỰC TẾ quý gần nhất và chỉ dùng tỷ lệ spread để nội suy, không lấy trực tiếp từ spread."
    )
    # ── Đánh giá Spread kết hợp Biên LNG — 12 quý gần nhất (Biểu đồ 7A) & theo năm (Biểu đồ 7) ──
    _gpm_by_q = {}
    for _rec in is_qs:
        _yr, _q = _rec.get("yearReport"), _rec.get("lengthReport")
        if _yr and _q and _q != 5:
            _rv = (_rec.get("isa3", 0) or 0) / 1e9; _gp = (_rec.get("isa5", 0) or 0) / 1e9
            _gm = (_gp/_rv*100) if _rv else None
            if _gm is not None and 0 < _gm < 60:
                _gpm_by_q[f"{_yr}Q{_q}"] = _gm
    def _corr_label(r):
        if r is None:
            return "không đủ dữ liệu"
        if r > 0.5:
            return "tương quan dương rõ rệt"
        if r > 0.2:
            return "tương quan dương vừa phải"
        if r > -0.2:
            return "tương quan yếu/không rõ ràng"
        return "tương quan âm — Spread benchmark và Biên LNG thực tế đi ngược chiều trong giai đoạn này"
    _pairs12 = [(Q18_SPREAD[i], _gpm_by_q[Q18_LABELS[i]]) for i in range(len(Q18_LABELS)) if Q18_LABELS[i] in _gpm_by_q][-12:]
    _corr12 = round(np.corrcoef([p[0] for p in _pairs12], [p[1] for p in _pairs12])[0, 1], 2) if len(_pairs12) > 3 else None
    _spr_trend_12 = "MỞ RỘNG" if Q18_SPREAD[-1] > Q18_SPREAD[-5] else "THU HẸP"
    add_body(
        f"<b>Đánh giá Biểu đồ 7A (Spread lag-1-quý kết hợp Biên LNG, 12 quý gần nhất):</b> Spread thực tế "
        f"{Q18_LABELS[-5]} → {Q18_LABELS[-1]} đi từ {Q18_SPREAD[-5]:,.0f} → {Q18_SPREAD[-1]:,.0f} USD/tấn "
        f"(<b>{_spr_trend_12}</b>), tương quan với Biên LNG thực tế 12 quý gần nhất: r = {_corr12} — "
        f"{_corr_label(_corr12)}. Spread ước tính quý hiện tại (Q2/2026): {SPREAD_NOW:,.0f} USD/tấn."
    )
    _ann_trend = "MỞ RỘNG" if SPREAD_A[5] > SPREAD_A[4] else "THU HẸP"
    add_body(
        f"<b>Đánh giá Biểu đồ 7 (Spread kết hợp Biên LNG, theo Năm):</b> Spread năm (median 4 quý lag-1-quý) "
        f"đi từ {SPREAD_A[2]:,.0f} USD/tấn ({years_hist[2]}, đáy chu kỳ) → {SPREAD_A[4]:,.0f} USD/tấn ({years_hist[4]}) "
        f"→ {SPREAD_A[5]:,.0f} USD/tấn ({years_fc[0]}E) — xu hướng <b>{_ann_trend}</b>, cùng chiều với Biên LNG thực tế "
        f"{gp_margin_hist[2]}% → {gp_margin_hist[4]}% → {gp_margin_fc[0]}% dự phóng cùng giai đoạn. "
        "Đây là cơ sở để neo giả định Biên LNG dự phóng theo tỷ lệ Spread (xem mục Biên LNG ở trên)."
    )
    # ── Biểu đồ tương quan Spread vs Biên LNG (scatter + hồi quy) — theo Năm & theo Quý ──
    # Chỉ dùng dữ liệu THỰC TẾ (2021-2025 cho năm; mọi quý có BCTC cho quý) — không đưa số dự phóng
    # vào vì Biên LNG dự phóng được NỘI SUY TỪ chính tỷ lệ Spread, đưa vào sẽ tạo tương quan giả.
    _corr_ann = round(np.corrcoef(SPREAD_A[:5], gp_margin_hist[:5])[0, 1], 2)
    _pairs_all = [(Q18_SPREAD[i], _gpm_by_q[Q18_LABELS[i]]) for i in range(len(Q18_LABELS)) if Q18_LABELS[i] in _gpm_by_q]
    _corr_q_all = round(np.corrcoef([p[0] for p in _pairs_all], [p[1] for p in _pairs_all])[0, 1], 2) if len(_pairs_all) > 3 else None
    add_body(
        f"<b>Tương quan Spread ↔ Biên LNG thực tế:</b> Theo NĂM (2021-2025, n=5): hệ số Pearson r = {_corr_ann} "
        f"— {_corr_label(_corr_ann)}. Theo QUÝ (n={len(_pairs_all)} quý có đủ dữ liệu BCTC): r = {_corr_q_all} "
        f"— {_corr_label(_corr_q_all)}. Với n nhỏ (đặc biệt chuỗi năm chỉ có 5 điểm), hệ số tương quan dễ bị "
        "chi phối bởi outlier (VD: 2023 — Spread lag-1-quý cao bất thường do giá NVL đầu 2022 giảm mạnh trong khi "
        "giá HRC chưa kịp giảm theo, đúng lúc Biên LNG thực tế lại chạm đáy chu kỳ) — không nên coi hệ số này là "
        "kết luận thống kê chắc chắn, chỉ mang tính tham khảo xu hướng."
    )
    for _cp, _cap in [
        ("spread_gpm_corr_annual.png", f"Biểu đồ: Tương quan Spread & Biên LNG theo Năm (r = {_corr_ann})"),
        ("spread_gpm_corr_quarterly.png", f"Biểu đồ: Tương quan Spread & Biên LNG theo Quý (r = {_corr_q_all})"),
    ]:
        _cp_path = os.path.join(CHART_DIR, _cp)
        if os.path.exists(_cp_path):
            elements.append(Spacer(1, 3*mm))
            elements.append(Paragraph(_cap, styles['SmallText']))
            elements.append(Image(_cp_path, width=320, height=240))
    cq_path = os.path.join(CHART_DIR, "commodity_prices_18q.png")
    if os.path.exists(cq_path):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("Biểu đồ: Giá HRC / Quặng sắt / Than cốc — 18 quý gần nhất + Hiện tại (fetch tự động)", styles['SmallText']))
        elements.append(Image(cq_path, width=460, height=230))

    # ─── Volume & Market Share Analysis ───
    elements.append(Spacer(1, 4*mm))
    add_section("SẢN LƯỢNG & THỊ PHẦN")
    add_body(
        "<b>Sản lượng HPG theo cấu phần:</b>"
    )
    vol_headers = ["Chỉ tiêu", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"]
    vol_data = [
        ["SL HRC (triệu tấn)", "2.0", "2.2", "2.5", "2.8", "3.2", "6.0", "6.8", "7.5"],
        ["SL Thép XD (triệu tấn)", "2.8", "2.6", "2.3", "2.5", "2.8", "3.0", "3.2", "3.5"],
        ["Tổng SL HPG (triệu tấn)", "4.8", "4.8", "4.8", "5.3", "6.0", "9.0", "10.0", "11.0"],
        ["Thị trường thép VN (triệu tấn)", "28", "27", "25", "26", "28", "30", "32", "34"],
        ["Thị phần HPG (%)", "17%", "18%", "19%", "20%", "21%", "30%", "31%", "32%"],
    ]
    add_table(vol_headers, vol_data,
              [doc.width*0.30] + [doc.width*0.0875]*8)
    add_body(
        "<b>Nhận xét:</b> Tổng sản lượng HPG tăng gấp đôi từ 2025 (6 triệu tấn) lên 2028E (11 triệu tấn) "
        "nhờ DQ2 full công suất. Thị phần mở rộng từ ~17% (2021) lên ~32% (2028E). "
        "HPG là doanh nghiệp duy nhất đủ lớn để hấp thụ tăng trưởng nhu cầu thép nội địa mà không "
        "phải cạnh tranh gay gắt về giá. Sản lượng HRC tăng mạnh nhất với DQ2 (3.5→6→7.5 triệu tấn)."
    )

    # ─── Factors Impacting HPG & Steel Industry ───
    elements.append(Spacer(1, 4*mm))
    add_section("YẾU TỐ TÁC ĐỘNG NGÀNH THÉP")
    add_body(
        "<b>1. Vĩ mô & Chu kỳ:</b>"
    )
    add_body("- <b>Đầu tư công</b>: Giải ngân cao tốc Bắc-Nam (~$20 tỷ), sân bay Long Thành ($5 tỷ) kéo dài 2025-2030, "
             "tạo nhu cầu thép xây dựng ổn định. Mỗi 1% tăng giải ngân ~50,000 tỷ hấp thụ thêm ~200K tấn thép."
    )
    add_body("- <b>Bất động sản phục hồi</b>: Sau giai đoạn trầm lắng 2022-2024, thị trường BĐS ấm trở lại "
             "với tín hiệu tích cực từ Luật Kinh doanh BĐS mới, dù lãi suất huy động nhích tăng từ Q4/2025."
    )
    add_body("- <b>Chính sách tiền tệ</b>: Lãi suất huy động tăng 1-2% từ Q4/2025 (Big4 ~5.5-6.1%, riêng lẻ ≥6%) "
             "do tín dụng 2025 tăng 19,1% vượt huy động 14,1%, gây áp lực thanh khoản. SBV giữ nguyên lãi suất "
             "điều hành 4,5% và 26 NHTM đã hạ lãi suất sau 9/4, kỳ vọng hạ nhiệt H2/2026. Tỷ giá USD/VNĐ neo ở "
             "25,400 vẫn có lợi cho HPG (đầu tư DQ2 bằng USD, doanh thu thuần VND khiến BCTC hưởng lợi)."
    )
    add_body(
        "<b>1B. Dư cung toàn cầu & Cấu trúc ngành:</b>"
    )
    add_body("- <b>Công suất dư thừa toàn cầu ~640 triệu tấn (26%)</b>, dự báo lên 721 triệu tấn (2027F) "
             "nếu các kế hoạch mở rộng (+6.7% 2025-2027) tiếp tục. Phân bổ: TQ ~400M tấn, Ấn Độ ~50M tấn, "
             "EU ~40M tấn, CIS ~30M tấn."
    )
    add_body("- <b>BOF Inflexibility (YẾU TỐ CẤU TRÚC QUAN TRỌNG):</b> Lò cao (BOF) có chu kỳ vận hành "
             "liên tục 15-25 năm, không thể dừng lò ngay khi giá giảm sâu. Đây là nguyên nhân chính khiến "
             "dư cung kéo dài bất chấp giá thép xuống thấp — khác biệt hoàn toàn với ngành có định phí thấp."
    )
    add_body("- <b>CBAM (Carbon Border Adjustment Mechanism):</b> Chênh lệch chi phí carbon giữa BOF "
             "(thép truyền thống) và EAF (thép xanh) có thể đạt 150-200 EUR/tấn vào 2026. Đây là rào cản "
             "lớn cho xuất khẩu thép VN sang EU, buộc các DN phải chuyển đổi công nghệ xanh nếu muốn duy trì thị phần."
    )
    add_body(
        "<b>2. Cung - Cầu thép:</b>"
    )
    add_body("- <b>Nguồn cung HRC toàn cầu</b>: Công suất HRC TQ ~1.1 tỷ tấn (60% toàn cầu), đang "
             "chịu áp lực giảm sản lượng do biên lợi nhuận thấp và chính sách hạn chế xuất khẩu của Chính phủ TQ. "
             "Gary Works #14 (Mỹ) reline (2025-2027) cắt ~2 triệu tấn HRC khỏi thị trường."
    )
    add_body("- <b>Thuế CBPG — Từ 'bảo hộ một phần' lên 'bảo hộ toàn phần':</b> "
             "AD HRC khổ hẹp (19.38-27.83%) + AD HRC khổ rộng chống lẩn tránh 27.83% (hiệu lực 17/04/2026 theo "
             "Quyết định 612/QĐ-BCT). Khi cả hai khổ đều bị áp thuế, thị trường HRC nội địa gần như đóng "
             "cửa hoàn toàn với hàng TQ — HPG là người hưởng lợi chính với ~70% thị phần."
    )
    add_body("- <b>Cân đối Cung-Cầu HRC nội địa (tấn):</b> Sản xuất HPG: 3.0M (2024) → 7.2M (2025) → 8.6M (2026F). "
             "Formosa (FMS) duy trì ~5.0M tấn/năm. Tổng nhu cầu HRC VN ~17.0M tấn (2026F). "
             "HPG + FMS gần như đáp ứng toàn bộ nhu cầu nội địa, phần nhập khẩu còn lại rất nhỏ."
    )
    add_body("- <b>Giá HRC</b>: Duy trì ổn định ở 600-650 USD/tấn nhờ chi phí nguyên liệu hạ nhiệt "
             "và nhu cầu phục hồi. Rủi ro từ suy thoái TQ có thể kéo giá HRC về 500 USD/tấn."
    )
    add_body("- <b>Phân bổ nhu cầu thép toàn cầu (8 lĩnh vực):</b> Xây dựng 49-50%, Cơ khí 16%, "
             "Kim loại 11%, Ô tô 8%, Dầu khí 6%, Đóng tàu 4%, Thiết bị gia dụng 3%, Quốc phòng 1%. "
             "Mảng xây dựng giảm tỷ trọng dài hạn; cơ khí & ô tô tăng dần."
    )
    add_body(
        "<b>3. Chi phí sản xuất & Chuỗi cung ứng:</b>"
    )
    add_body("- <b>Cơ cấu nhập khẩu nguyên liệu HPG:</b> Nhập khẩu 90% than cốc (Úc, Nga, Indonesia), "
             "52% quặng sắt (Úc, Brazil), 47% thép phế (Nhật Bản, ASEAN). "
             "Tự chủ ~50% quặng nhưng phụ thuộc lớn vào than cốc NK — rủi ro từ địa chính trị và cước vận tải."
    )
    add_body("- <b>Quặng sắt</b>: Giá đã giảm từ đỉnh 140 USD/tấn (2021) xuống ~105 USD/tấn (2026E) nhờ "
             "nguồn cung dồi dào từ Úc/Brazil. Dự báo tiếp tục giảm về 90-100 USD/tấn đến 2028. "
             "HPG tự chủ 50% quặng, tiết kiệm ~20 USD/tấn thép."
    )
    add_body("- <b>Than cốc</b>: Giá giảm mạnh từ 400 USD/tấn (2021) về ~200-220 USD/tấn (2026E-2028E) nhờ "
             "tồn kho cao và nhu cầu TQ yếu. HPG sử dụng DQ2 công nghệ mới giúp giảm 15% tiêu hao than cốc."
    )
    add_body("- <b>Điện & Logistics</b>: Chi phí điện chiếm ~8-10% giá thành. DQ2 đặt tại Dung Quất giúp "
             "giảm logistics đầu vào (cảng biển ngay cạnh). Chênh lệch chạy DQ2 vs DQ1 ~15-20% chi phí."
    )
    add_body(
        "<b>4. Chiến lược thị trường & Đầu tư công:</b>"
    )
    add_body("- <b>Tái cơ cấu thị trường HPG:</b> Tỷ trọng nội địa đạt 84% (2025), tăng 38% YoY. "
             "Xuất khẩu giảm từ 31% (2024) xuống 16% (2025) — Châu Á giảm 57%, Châu Âu giảm 34%. "
             "Chiến lược này giúp HPG tránh rủi ro CBAM và cước vận tải quốc tế."
    )
    add_body("- <b>Đầu tư công — động lực dài hạn:</b> Kế hoạch trung hạn 2026-2030: ~8.5 triệu tỷ đồng "
             "(gấp 3x giai đoạn 2021-2025). Riêng 2026: ~995.4 nghìn tỷ (+10.4% YoY). "
             "Đây là nguồn cầu khổng lồ đảm bảo sản lượng thép xây dựng cho HPG trong 5 năm tới.")
    add_body("- <b>Cập nhật thị trường 4 tháng đầu 2026 (Nguoiquansat 29/05):</b> Ngành thép SX 11.67M tấn "
             "(+23.5% YoY), tiêu thụ 11.86M tấn (+14.1%). HPG XD 4M: 1.9M tấn (~36% SX ngành, >46% nội địa). "
             "HRC toàn ngành 4M: 3.36M tấn (+32%), XK 601k (gần gấp đôi). "
             "Tôn mạ toàn ngành: -17% SX (HSG 24.4% thị phần nội địa, HPG 10.51%). "
             "Đầu tư công 4M: 187k tỷ (+10.4%).")

    elements.append(PageBreak())
    add_section("4. PHÂN TÍCH TÀI CHÍNH & DỰ BÁO")

    # Financial table
    fin_headers = ["Chỉ tiêu (tỷ VND)", "2021A", "2022A", "2023A", "2024A", "2025A", "2026E", "2027E", "2028E"]
    fin_data = [
["Doanh thu", f"{revenue_hist[0]:,.0f}", f"{revenue_hist[1]:,.0f}", f"{revenue_hist[2]:,.0f}",
         f"{revenue_hist[3]:,.0f}", f"{revenue_hist[4]:,.0f}", f"{revenue_fc[0]:,}",
         f"{revenue_fc[1]:,}", f"{revenue_fc[2]:,}"],
["LNST", f"{ni_hist[0]:,.0f}", f"{ni_hist[1]:,.0f}", f"{ni_hist[2]:,.0f}",
         f"{ni_hist[3]:,.0f}", f"{ni_hist[4]:,.0f}", f"{ni_fc[0]:,}",
         f"{ni_fc[1]:,}", f"{ni_fc[2]:,}"],
        ["GP Margin", "27.5%", "11.9%", "10.9%", "13.3%", "15.7%", f"{gp_margin_fc[0]}%", f"{gp_margin_fc[1]}%", f"{gp_margin_fc[2]}%"],
        ["EBITDA", f"{revenue_hist[0]*0.12+da_hist[0]:,.0f}",
         f"{revenue_hist[1]*0.12+da_hist[1]:,.0f}",
         f"{revenue_hist[2]*0.12+da_hist[2]:,.0f}",
         f"{revenue_hist[3]*0.12+da_hist[3]:,.0f}",
         f"{revenue_hist[4]*0.12+da_hist[4]:,.0f}",
         f"{revenue_fc[0]*0.143+da_fc[0]:,.0f}",
         f"{revenue_fc[1]*0.158+da_fc[1]:,.0f}",
         f"{revenue_fc[2]*0.167+da_fc[2]:,.0f}"],
        ["D/E", "0.44", "0.44", "0.44", "0.49", "0.46", "0.41", "0.34", "0.28"],
        ["ROE (%)", "46.0%", "9.4%", "6.5%", "10.5%", "11.9%", "14.9%", "16.5%", "17.4%"],
    ]
    add_table(fin_headers, fin_data,
              [doc.width*0.18] + [doc.width*0.102]*8)

    add_body(f"Nhận xét: Doanh thu và LNST phục hồi mạnh từ đáy 2023, tương quan chặt với diễn biến "
             f"Spread thép ({SPREAD_A[2]:,.0f} USD/tấn năm {years_hist[2]} → {SPREAD_A[4]:,.0f} USD/tấn {years_hist[4]} → {SPREAD_A[5]:,.0f} USD/tấn {years_fc[0]}E). "
             "Biên LNG cải thiện nhờ (1) DQ2 full công suất giảm chi phí chuyển đổi, "
             "(2) giá quặng/than hạ nhiệt, (3) giá HRC phục hồi. "
             "D/E giảm dần nhờ dòng tiền mạnh. ROE cải thiện về >15% nhờ đòn bẩy tài chính hiệu quả. "
             "Lưu ý: Ngành thép chu kỳ — P/E tăng khi lợi nhuận giảm và ngược lại. "
             "Do đó, P/B là thước đo định giá đáng tin cậy hơn.")

    # Charts in PDF (Phần tài chính)
    elements.append(Spacer(1, 4*mm))
    for chart_name, caption in [
        ("revenue_ni.png", "Biểu đồ 1: Doanh thu và LNST 2021-2028E"),
        ("margins.png", "Biểu đồ 2: Biên lợi nhuận 2021-2028E"),
        ("growth.png", "Biểu đồ 3: Tăng trưởng Doanh thu YoY"),
        ("turnover.png", "Biểu đồ 4: Số ngày tồn kho (DIO) & Số ngày phải thu (DSO) bình quân, theo Năm — xem đánh giá tại mục 4C"),
        ("turnover_quarterly.png", "Biểu đồ 4B: DIO & DSO bình quân theo Quý (năm hóa) — xem đánh giá tại mục 4C"),
        ("quarterly.png", "Biểu đồ 5: KQKD theo Quý (Q1/2024 - Q1/2026)"),
        ("quarterly_bs.png", "Biểu đồ 6: Diễn biến Tài sản & Nợ theo Quý"),
        ("spread_vs_gp.png", "Biểu đồ 7: Spread ước tính (lag 1 quý/năm) vs Biên LNG thực tế — theo Năm"),
        ("spread_gp_quarterly.png", "Biểu đồ 7A: Spread HRC (đã có thuế CBPG AD20 từ 2025Q3, lag 1 quý) vs Biên LNG thực tế — 12 quý gần nhất"),
        ("spread_rebar_gpm_quarterly.png", "Biểu đồ 7A2: Spread Rebar/thép XD (lag 1 quý) vs Biên LNG thực tế — 12 quý gần nhất"),
        ("spread_all_gpm_quarterly.png", "Biểu đồ 7A3: Spread All (bình quân gia quyền theo sản lượng HRC/XD) vs Biên LNG thực tế — 12 quý gần nhất"),
        ("commodity_prices_18q.png", "Biểu đồ 7B: Giá HRC / Quặng sắt / Than cốc — 18 quý gần nhất + Hiện tại"),
        ("quarterly_quality.png", "Biểu đồ 8: Chất lượng Tài sản & KQKD (đa chiều)"),
        ("quarterly_volume.png", "Biểu đồ 9: Sản lượng tiêu thụ theo Quý (HRC, XD, Ống thép, Tôn mạ)"),
    ]:
        chart_path = os.path.join(CHART_DIR, chart_name)
        if os.path.exists(chart_path):
            elements.append(Spacer(1, 4*mm))
            elements.append(Paragraph(caption, styles['SmallText']))
            elements.append(Image(chart_path, width=460, height=230))

    # P/E chart
    pe_path = os.path.join(CHART_DIR, "pe_hist_q.png")
    if os.path.exists(pe_path):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("Biểu đồ 10: P/E theo Quý với median band", styles['SmallText']))
        elements.append(Image(pe_path, width=460, height=220))

    # P/B chart
    pb_path = os.path.join(CHART_DIR, "pb_hist_q.png")
    if os.path.exists(pb_path):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("Biểu đồ 11: P/B theo Quý với median band ±20%", styles['SmallText']))
        elements.append(Image(pb_path, width=460, height=220))

    # EV/EBITDA chart
    ev_path = os.path.join(CHART_DIR, "ev_hist_q.png")
    if os.path.exists(ev_path):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("Biểu đồ 12: EV/EBITDA theo Quý với buy/sell zones", styles['SmallText']))
        elements.append(Image(ev_path, width=460, height=220))

    elements.append(PageBreak())

    # ─── Concise: Projection Summary ───
    elements.append(Spacer(1, 4*mm))
    add_section("4A. DỰ PHÓNG 2026E")
    add_body(
        f"<b>Doanh thu 2026E:</b> {revenue_fc[0]:,} tỷ (+{revenue_fc[0]/revenue_hist[4]*100-100:.0f}% YoY). "
        f"Giả định: thép XD {SL_XD_A[5]:.1f} triệu tấn, HRC {SL_HRC_A[5]:.1f} triệu tấn, giá bán thép XD {XD_PRICE_A[5]:,.0f} USD/t, HRC {HRC_PRICE_A[5]:,.0f} USD/t. "
        f"Đầu vào: quặng {IRON_ORE_A[5]:,.0f} USD/t, than cốc {COKE_A[5]:,.0f} USD/t, chi phí SX khác cố định {OTHER_COST_USD} USD/t.<br/>"
        f"<b>Biên LNG:</b> {gp_margin_fc[0]}% (từ spread HRC ~{SPREAD_A[5]:,.0f} USD/t, cải thiện nhờ DQ2 full + tồn kho giá rẻ).<br/>"
        f"<b>LNST 2026E:</b> ước {ni_fc[0]:,} tỷ (+{ni_fc[0]/ni_hist[4]*100-100:.0f}% YoY). "
        f"Q1/2026 đạt {get_q(is_qs,2026,1,'isa22'):,.0f} tỷ (chiếm ~41% KH năm). "
        f"Kỳ vọng Q2-Q4 cao hơn nhờ mùa xây dựng cao điểm."
    )

    # ─── New: Quarterly Trend Analysis ───
    add_section("4B. PHÂN TÍCH XU HƯỚNG QUÝ (Q1/2026)")
    def q_val(section, yr, q, field):
        recs = dict(IS=is_qs, BS=bs_qs, CF=cf_qs).get(section)
        return get_q(recs, yr, q, field)
    # ── Fetch Q1/2026, Q4/2025, Q1/2025 ──
    rev_c = q_val('IS',2026,1,'isa3'); rev_p = q_val('IS',2025,4,'isa3'); rev_y = q_val('IS',2025,1,'isa3')
    cogs_c = abs(q_val('IS',2026,1,'isa4')); cogs_p = abs(q_val('IS',2025,4,'isa4')); cogs_y = abs(q_val('IS',2025,1,'isa4'))
    gp_c = rev_c - cogs_c; gp_p = rev_p - cogs_p; gp_y = rev_y - cogs_y
    ni_c = q_val('IS',2026,1,'isa22'); ni_p = q_val('IS',2025,4,'isa22'); ni_y = q_val('IS',2025,1,'isa22')
    ta_c = q_val('BS',2026,1,'bsa53'); ta_p = q_val('BS',2025,4,'bsa53'); ta_y = q_val('BS',2025,1,'bsa53')
    inv_c = q_val('BS',2026,1,'bsa15'); inv_p = q_val('BS',2025,4,'bsa15'); inv_y = q_val('BS',2025,1,'bsa15')
    rec_c = q_val('BS',2026,1,'bsa8'); rec_p = q_val('BS',2025,4,'bsa8'); rec_y = q_val('BS',2025,1,'bsa8')
    debt_c = q_val('BS',2026,1,'bsa56')+q_val('BS',2026,1,'bsa71')
    debt_p = q_val('BS',2025,4,'bsa56')+q_val('BS',2025,4,'bsa71')
    debt_y = q_val('BS',2025,1,'bsa56')+q_val('BS',2025,1,'bsa71')
    eq_c = q_val('BS',2026,1,'bsa78'); eq_p = q_val('BS',2025,4,'bsa78'); eq_y = q_val('BS',2025,1,'bsa78')
    cash_c = q_val('BS',2026,1,'bsa2'); cash_p = q_val('BS',2025,4,'bsa2'); cash_y = q_val('BS',2025,1,'bsa2')
    gm_c = gp_c/rev_c*100; gm_p = gp_p/rev_p*100; gm_y = gp_y/rev_y*100

    # ── Helper: % change or "N/A" ──
    def pct_chg(new, old):
        if old == 0: return "N/A"
        return f"{'+' if new>=old else ''}{(new/old-1)*100:+.1f}%"
    def pp_chg(new, old):
        return f"{'+' if new>=old else ''}{new-old:+.1f}pp"

    # ── Table 1: So sánh với quý trước (QoQ) ──
    add_body("<b>1. So sánh với quý trước (Q vs Q4/2025):</b>")
    tbl_qoq = [
        ["Doanh thu (tỷ)", f"{rev_p:,.0f}", f"{rev_c:,.0f}", pct_chg(rev_c, rev_p), "TĂNG — DQ2 full + giá HRC phục hồi"],
        ["LNST (tỷ)", f"{ni_p:,.0f}", f"{ni_c:,.0f}", pct_chg(ni_c, ni_p), "TĂNG MẠNH — spread nở + sản lượng cao"],
        ["Biên LNG (%)", f"{gm_p:.1f}%", f"{gm_c:.1f}%", pp_chg(gm_c, gm_p), "CẢI THIỆN — DQ2 giúp giảm chi phí"],
        ["Tổng TS (tỷ)", f"{ta_p:,.0f}", f"{ta_c:,.0f}", pct_chg(ta_c, ta_p), "MỞ RỘNG — đầu tư DQ2 hoàn tất"],
        ["Hàng tồn kho (tỷ)", f"{inv_p:,.0f}", f"{inv_c:,.0f}", pct_chg(inv_c, inv_p), "↑ — dự trữ NVL cho SX cao"],
        ["Phải thu (tỷ)", f"{rec_p:,.0f}", f"{rec_c:,.0f}", pct_chg(rec_c, rec_p), "ỔN ĐỊNH — siết chặt chính sách TD"],
        ["Nợ vay (tỷ)", f"{debt_p:,.0f}", f"{debt_c:,.0f}", pct_chg(debt_c, debt_p), "GIẢM — không vay thêm DQ2"],
        ["VCSH (tỷ)", f"{eq_p:,.0f}", f"{eq_c:,.0f}", pct_chg(eq_c, eq_p), "TĂNG — LN giữ lại"],
        ["Tiền mặt (tỷ)", f"{cash_p:,.0f}", f"{cash_c:,.0f}", pct_chg(cash_c, cash_p), "TĂNG — FCF chuyển dương"],
    ]
    add_table(["Chỉ tiêu", "Q4/2025", "Q1/2026", "QoQ", "Đánh giá"], tbl_qoq, [90, 65, 65, 55, 150])

    # ── Table 2: So sánh cùng kỳ (YoY) ──
    add_body("<b>2. So sánh cùng kỳ (Q1/2026 vs Q1/2025):</b>")
    tbl_yoy = [
        ["Doanh thu (tỷ)", f"{rev_y:,.0f}", f"{rev_c:,.0f}", pct_chg(rev_c, rev_y), "TĂNG — DQ2 chưa full Q1/2025"],
        ["LNST (tỷ)", f"{ni_y:,.0f}", f"{ni_c:,.0f}", pct_chg(ni_c, ni_y), "TĂNG MẠNH — chu kỳ HRC thuận lợi"],
        ["Biên LNG (%)", f"{gm_y:.1f}%", f"{gm_c:.1f}%", pp_chg(gm_c, gm_y), "CẢI THIỆN — DQ2 giúp giảm giá thành"],
        ["Tổng TS (tỷ)", f"{ta_y:,.0f}", f"{ta_c:,.0f}", pct_chg(ta_c, ta_y), "MỞ RỘNG — DQ2 đi vào hoạt động"],
        ["Hàng tồn kho (tỷ)", f"{inv_y:,.0f}", f"{inv_c:,.0f}", pct_chg(inv_c, inv_y), "TĂNG — quy mô SX mở rộng"],
        ["Phải thu (tỷ)", f"{rec_y:,.0f}", f"{rec_c:,.0f}", pct_chg(rec_c, rec_y), "TĂNG THEO — tương xứng DT"],
        ["Nợ vay (tỷ)", f"{debt_y:,.0f}", f"{debt_c:,.0f}", pct_chg(debt_c, debt_y), "TĂNG — vay DQ2, đang giảm dần"],
        ["VCSH (tỷ)", f"{eq_y:,.0f}", f"{eq_c:,.0f}", pct_chg(eq_c, eq_y), "TĂNG — LNGL + phát hành CP"],
        ["Tiền mặt (tỷ)", f"{cash_y:,.0f}", f"{cash_c:,.0f}", pct_chg(cash_c, cash_y), "TĂNG — FCF chuyển dương"],
    ]
    add_table(["Chỉ tiêu", "Q1/2025", "Q1/2026", "YoY", "Đánh giá"], tbl_yoy, [90, 65, 65, 55, 150])

    # ── Đánh giá tổng quan ──
    add_body(
        f"<b>Đánh giá tổng quan:</b> KQKD Q1/2026 khởi sắc trên cả 2 chiều so sánh: "
        f"(i) <b>QoQ:</b> Doanh thu {pct_chg(rev_c, rev_p)}, LNST {pct_chg(ni_c, ni_p)}, GM {pp_chg(gm_c, gm_p)}; "
        f"(ii) <b>YoY:</b> Doanh thu {pct_chg(rev_c, rev_y)}, LNST {pct_chg(ni_c, ni_y)}, GM {pp_chg(gm_c, gm_y)}. "
        f"Cấu trúc tài sản lành mạnh: nợ vay giảm, tiền mặt tăng. "
        f"Đây là dấu hiệu điển hình của giai đoạn <b>thu hoạch</b> sau đầu tư DQ2."
    )
    add_body(
        f"<b>Doanh thu Q1/2026: {rev_c:,.0f} tỷ ({pct_chg(rev_c, rev_y)} YoY)</b> — Mức cao nhất từ trước đến nay. "
        f"DQ2 full công suất từ T12/2025 đóng góp ~6 triệu tấn HRC/năm, tương đương 120-140 nghìn tỷ doanh thu. "
        f"Q1 thường là quý thấp nhất trong năm (Tết Nguyên đán). Kỳ vọng Q2-Q4 cao hơn 15-20% khi mùa xây dựng cao điểm."
    )
    add_body(
        f"<b>Biên LNG Q1/2026: {gm_c:.1f}%</b> — "
        f"Cải thiện {gm_c-gm_p:+.1f}pp QoQ (so với {gm_p:.1f}% Q4/2025) và {gm_c-gm_y:+.1f}pp YoY (so với {gm_y:.1f}% Q1/2025) nhờ "
        f"(1) DQ2 full công suất giảm ĐMTK, (2) tồn kho giá rẻ bắt đầu được ghi nhận. "
        f"Dự báo GM tiếp tục cải thiện lên 17-18% khi DQ2 chạy ổn định và spread nở ra."
    )
    add_body(
        f"<b>LNST Q1/2026: {ni_c:,.0f} tỷ ({pct_chg(ni_c, ni_p)} QoQ, {pct_chg(ni_c, ni_y)} YoY)</b> — "
        f"Khởi đầu rất tích cực. Nếu HPG duy trì LNST bình quân 6,000-7,000 tỷ/quý trong Q2-Q4, "
        f"cả năm có thể đạt 26-28,000 tỷ, vượt ~25-30% kế hoạch ĐHĐCĐ (22,000 tỷ)."
    )
    add_body(
        f"<b>Tài sản & Nguồn vốn:</b> Tổng tài sản {ta_c:,.0f} tỷ ({pct_chg(ta_c, ta_p)} QoQ, {pct_chg(ta_c, ta_y)} YoY). "
        f"Hàng tồn kho {inv_c:,.0f} tỷ (~{(inv_c/(rev_c*4))*365:.0f} ngày doanh thu quy năm — phù hợp chu kỳ SX; "
        f"xem DIO chính xác theo GVHB bình quân tại mục 4C). "
        f"Nợ vay {debt_c:,.0f} tỷ (giảm {(debt_p-debt_c):,.0f} tỷ so với Q4/2025). "
        f"D/E = {debt_c/eq_c:.2f}x (an toàn). Tiền mặt {cash_c:,.0f} tỷ. "
        f"FCF dự kiến chuyển dương từ 2026 — HPG không cần huy động vốn thêm."
    )

    # ── Lũy kế từ đầu năm ──
    plan_ni = 22000  # kế hoạch LNST 2026 ĐHĐCĐ
    add_body(
        "<b>3. Lũy kế từ đầu năm 2026 & tiến độ kế hoạch:</b><br/>"
        f"- <b>Lũy kế 3T/2026:</b> Doanh thu {rev_c:,.0f} tỷ, LNST {ni_c:,.0f} tỷ, biên LNG {gm_c:.1f}%. "
        f"(So với lũy kế 3T/2025: DT {pct_chg(rev_c, rev_y)}, LNST {pct_chg(ni_c, ni_y)}).<br/>"
        f"- <b>Tiến độ kế hoạch năm:</b> LNST Q1 đã đạt {ni_c/plan_ni*100:.1f}% kế hoạch ĐHĐCĐ (22,000 tỷ). "
        f"Nếu duy trì ~6,000-7,000 tỷ/quý cho 9 tháng còn lại, HPG có thể đạt 26-28,000 tỷ, vượt 25-30% kế hoạch.<br/>"
        f"- <b>Tài sản & Nợ (cuối Q1/2026):</b> Tổng TS {ta_c:,.0f} tỷ, nợ vay {debt_c:,.0f} tỷ, "
        f"VCSH {eq_c:,.0f} tỷ. So với cuối Q1/2025: TS {pct_chg(ta_c, ta_y)}, nợ {pct_chg(debt_c, debt_y)}. "
        f"D/E {debt_c/eq_c:.2f}x (Q1/2025: {debt_y/eq_y:.2f}x)."
    )

    # ── Lũy kế dài hạn: đánh giá xu hướng dài hạn ──
    add_body(
        "<b>4. Đánh giá lũy kế dài hạn (2018 - nay):</b><br/>"
        "- <b>Chu kỳ 1 (2018-2020):</b> Tăng trưởng nóng nhờ thép XD, biên LNG 15-20%. "
        "HPG tích lũy vốn cho DQ2.<br/>"
        "- <b>Chu kỳ 2 (2021-2022):</b> Đỉnh lợi nhuận nhờ giá thép toàn cầu (HRC >900 USD). "
        "Biên LNG đạt đỉnh 27.5% (2021). ROE >30%.<br/>"
        "- <b>Chu kỳ 3 (2023-2024):</b> Đáy chu kỳ — giá HRC rơi về 480 USD. "
        "Biên LNG chỉ ~10-13%. HPG đầu tư DQ2 mạnh (CAPEX ~25k tỷ/năm). D/E tăng lên 0.6x.<br/>"
        "- <b>Chu kỳ 4 (2025-nay):</b> Phục hồi + thu hoạch. DQ2 full công suất, "
        f"biên LNG Q1/2026 đạt {gm_c:.1f}% (cao nhất từ 2022). "
        "D/E giảm, FCF dương. Kỳ vọng đây là chu kỳ đỉnh mới tương tự 2021.<br/><br/>"
        f"<b>Kết luận lũy kế:</b> Trong 8 năm qua, HPG đã đầu tư ~100 nghìn tỷ vào DQ2 "
        f"(tổng tài sản từ {q_val('BS',2018,1,'bsa53'):,.0f} tỷ đầu 2018 lên {ta_c:,.0f} tỷ Q1/2026). "
        f"DQ2 là bước ngoặt: biến HPG từ nhà sản xuất thép XD nội địa thành "
        f"tập đoàn thép hợp nhất (HRC + XD) quy mô 8.5 triệu tấn, top đầu ASEAN. "
        f"Với nợ vay đã qua đỉnh và FCF chuyển dương, HPG bước vào giai đoạn "
        f"<b>thu hoạch tiền mặt</b> — lợi nhuận cao, cổ tức tăng."
    )

    # ─── New: Accounting Quality ───
    elements.append(Spacer(1, 4*mm))
    add_section("4C. CHẤT LƯỢNG LỢI NHUẬN & KẾ TOÁN")
    # ── Profit decomposition: core vs one-off ──
    def _q_is(y,q,f): return get_q(is_qs if 'is' in str(True) else is_qs, y, q, f)  # noqa
    op_p = q_val('IS',2026,1,'isa11')
    fin_i = q_val('IS',2026,1,'isa6')
    fin_c = q_val('IS',2026,1,'isa7')
    sga = abs(q_val('IS',2026,1,'isa9'))+abs(q_val('IS',2026,1,'isa10'))
    other_inc = q_val('IS',2026,1,'isa14')
    gp = q_val('IS',2026,1,'isa5')
    ebit_core = gp - sga
    # Historical comparison
    q_hist_op = [(q_val('IS',y,q,'isa11'), q_val('IS',y,q,'isa6'), q_val('IS',y,q,'isa14')) for y,q in [(2025,4),(2025,3),(2025,2),(2025,1)]]
    hist_avg_fin = sum(f[1] for f in q_hist_op)/4
    norm_npat = (ebit_core - abs(q_val('IS',2026,1,'isa8')) + other_inc) * (1-0.12) if False else 0  # placeholder

    add_body(
        "<b>1. Phân tích lợi nhuận Q1/2026: Core vs Đột biến</b><br/>"
        f"- Lợi nhuận HĐKD (isa11): <b>{op_p:,.0f} tỷ</b>. "
        f"Trong đó: Lãi gộp {gp:,.0f} tỷ, DTTC {fin_i:,.0f} tỷ, CPTC {abs(fin_c):,.0f} tỷ, CPBH&QL {sga:,.0f} tỷ.<br/>"
        f"- <b>DTTC Q1/2026: {fin_i:,.0f} tỷ</b> — cao bất thường so với bình quân các quý trước ({hist_avg_fin:,.0f} tỷ). "
        f"Chênh lệch: {fin_i-hist_avg_fin:+,.0f} tỷ. Nguyên nhân có thể: (i) lãi tỷ giá do USD giảm, "
        f"(ii) cổ tức từ các công ty con, (iii) lãi tiền gửi lớn. <b>Cần kiểm tra thuyết minh BCTC.</b><br/>"
        f"- Thu nhập khác (isa14): {other_inc:,.0f} tỷ (không đáng kể).<br/>"
        f"- <b>Kết luận:</b> Lợi nhuận Q1/2026 có ~{fin_i-hist_avg_fin:+,.0f} tỷ từ DTTC bất thường. "
        f"Nếu loại trừ, LNST điều chỉnh ~{(ebit_core-abs(q_val('IS',2026,1,'isa8'))+other_inc)*(1-0.12):,.0f} tỷ (so với báo cáo 8,994 tỷ). "
        f"Các quý sau dự kiến LNST core ~4,500-5,000 tỷ/quý (không phải 2,000 tỷ) — "
        f"vẫn là mức rất tốt nhờ DQ2 full năm."
    )
    # ── Historical quality metrics ──
    cfo_ni_avg = sum(cfo_hist[i]/ni_hist[i] for i in range(5))/5
    add_body(
        f"2. <b>Cash Conversion (CFO/LNST): {cfo_ni_avg:.2f}x</b> (bình quân 2021-2025) — "
        f"Trên 1.0 chứng tỏ lợi nhuận được đảm bảo bằng tiền mặt thực, không phải dồn tích ảo. "
        f"Đây là điểm mạnh của HPG so với nhiều doanh nghiệp cùng ngành."
    )
    add_body(
        f"3. <b>Accruals Ratio: {(cfo_hist[4]-ni_hist[4])/total_assets_hist[4]*100:.1f}%</b> (2025) — "
        f"Chỉ số dồn tích thấp (<5%), phản ánh chênh lệch nhỏ giữa dòng tiền và lợi nhuận. "
        f"Các khoản phải thu và tồn kho tăng tương xứng với quy mô hoạt động."
    )
    _dio_trend = "CẢI THIỆN" if DIO_A[4] < DIO_A[0] else "XẤU ĐI"
    _dso_trend = "CẢI THIỆN" if DSO_A[4] < DSO_A[0] else "XẤU ĐI"
    add_body(
        f"4. <b>Số ngày tồn kho bình quân (DIO): {DIO_A[4]:.0f} ngày (2025)</b> so với {DIO_A[0]:.0f} ngày (2021) — "
        f"xu hướng <b>{_dio_trend}</b>. Công thức: DIO = 365 × Tồn kho bình quân (đầu kỳ+cuối kỳ)/2 ÷ GVHB "
        f"(vị trí công thức sống: sheet <i>14_Steel_Analysis</i>, mục 4 — bấm vào ô Excel để kiểm chứng). "
        f"Mức {DIO_A[4]:.0f} ngày phù hợp chu kỳ SX lò cao BOF (~60-90 ngày nấu luyện + dự trữ NVL ~30 ngày). "
        "Không dự phóng DIO/DSO cho 2026E-2028E vì ít ảnh hưởng tới model định giá."
    )
    add_body(
        f"5. <b>Số ngày phải thu bình quân (DSO): {DSO_A[4]:.0f} ngày (2025)</b> so với {DSO_A[0]:.0f} ngày (2021) — "
        f"xu hướng <b>{_dso_trend}</b>. Công thức: DSO = 365 × Phải thu bình quân (đầu kỳ+cuối kỳ)/2 ÷ Doanh thu "
        f"(vị trí công thức sống: sheet <i>14_Steel_Analysis</i>, mục 4 — bấm vào ô Excel để kiểm chứng). "
        f"DSO {'giảm' if DSO_A[4] < DSO_A[0] else 'tăng'} cho thấy HPG "
        f"{'siết chặt chính sách tín dụng, thu tiền nhanh hơn qua hệ thống đại lý' if DSO_A[4] < DSO_A[0] else 'nới lỏng chính sách tín dụng để đẩy hàng — cần theo dõi rủi ro nợ xấu'}. "
        f"Xem Biểu đồ 4 (DIO/DSO theo Năm, {years_hist[0]}-{years_hist[4]}) và Biểu đồ 4B (theo Quý) để đánh giá "
        "hiệu quả thu tiền & luân chuyển hàng tồn kho có cải thiện theo thời gian hay không."
    )
    add_body(
        "<b>Rủi ro kế toán cần theo dõi:</b><br/>"
        f"- <b>DTTC bất thường:</b> Q1/2026 DTTC {fin_i:,.0f} tỷ (gấp {fin_i/hist_avg_fin:.1f}x bình quân). "
        f"Nếu là lãi tỷ giá chưa thực hiện (OCI), các quý sau có thể hoàn nhập khi USD tăng trở lại.<br/>"
        f"- <b>Dự phòng HTK:</b> Với tồn kho {inventory_hist[4]:,.0f} tỷ (2025), cần kiểm tra NOTE về trích lập dự phòng. "
        f"Nếu giá thép giảm >10%, HPG có thể phải trích thêm ~2,000-3,000 tỷ dự phòng.<br/>"
        f"- <b>Tỷ giá:</b> Vay USD ~$2 tỷ, mỗi 1% USD tăng = ~500 tỷ lỗ tỷ giá chưa thực hiện (OCI).<br/>"
        "- <b>Kết luận: BCTC HPG minh bạch, chất lượng tốt. Rủi ro kế toán thấp. "
        "Lợi nhuận Q1/2026 có phần đột biến từ DTTC, cần theo dõi các quý sau.</b>"
    )

    # ─── New: Harvest Signs ───
    elements.append(Spacer(1, 4*mm))
    add_section("4D. DẤU HIỆU SẮP HÁI QUẢ NGỌT")
    add_body(
        "<b>Mô hình Harvest Signals — Đánh giá mức độ sẵn sàng cho chu kỳ lợi nhuận đỉnh:</b>"
    )
    add_body(
        "1. <b>VĨ MÔ ĐẢO CHIỀU ✅ — ĐÃ CÓ TÍN HIỆU</b><br/>"
        "- Lãi suất cho vay giảm từ 12% (2023) xuống 8-9% (2026), hỗ trợ thị trường BĐS và đầu tư<br/>"
        "- Giải ngân đầu tư công 2026: ~680,000 tỷ (+22% YoY), cao tốc Bắc-Nam và Long Thành là trụ cột<br/>"
        "- TQ kích thích kinh tế 1,000 tỷ NDT (2025-2026), hỗ trợ nhu cầu thép toàn cầu<br/>"
        "- Luật Kinh doanh BĐS mới (8/2024) khơi thông thị trường, dự án mới tăng 40% YoY"
    )
    _spr_trend_up = SPREAD_A[5] > SPREAD_A[4]
    add_body(
        f"2. <b>SPREAD {'NỞ RA' if _spr_trend_up else 'THU HẸP'} {'✅' if _spr_trend_up else '⚠️'} — "
        f"THEO GIÁ HÀNG HÓA HIỆN TẠI</b><br/>"
        f"- Spread thép (median 4 quý/năm, giá hiện tại fetch tự động): {SPREAD_A[2]:,.0f} USD/tấn ({years_hist[2]}) "
        f"→ {SPREAD_A[4]:,.0f} USD/tấn ({years_hist[4]}) → {SPREAD_A[5]:,.0f} USD/tấn ({years_fc[0]}E)<br/>"
        "- Giá HRC chịu áp lực từ dư cung TQ dù có thuế CBPG 27.8% bảo hộ nội địa; "
        "giá benchmark quốc tế (investing.com) vẫn là yếu tố chi phối spread tính theo công thức chuẩn<br/>"
        f"- Chi phí đầu vào: quặng {IRON_ORE_A[0]:,.0f}→{IRON_ORE_A[5]:,.0f} USD/tấn, "
        f"than cốc {COKE_A[0]:,.0f}→{COKE_A[5]:,.0f} USD/tấn (2021→{years_fc[0]}E)<br/>"
        f"- Biên LNG THỰC TẾ (neo Q1/2026, không lấy trực tiếp từ spread): {gp_margin_hist[2]}% ({years_hist[2]}) "
        f"→ {gp_margin_hist[4]}% ({years_hist[4]}) → dự báo {gp_margin_fc[0]}% ({years_fc[0]}E)"
    )
    add_body(
        "3. <b>HÀNG TỒN KHO & CHI PHÍ ĐẦU VÀO</b><br/>"
        f"- Quặng sắt 62%Fe hiện tại: {iron_now:,.1f} USD/tấn (fetch tự động), median {years_hist[4]}: {IRON_ORE_A[4]:,.0f} USD/tấn<br/>"
        f"- Tồn kho Q1/2026: {get_q(bs_qs,2026,1,'bsa15'):,.0f} tỷ — giá vốn phần lớn chốt từ hợp đồng dài hạn, không hoàn toàn theo giá spot<br/>"
        "- Lợi thế giá đầu vào (nếu có) sẽ thể hiện rõ trong GM các quý kế tiếp khi hàng tồn kho được ghi nhận vào giá vốn<br/>"
        "- DQ2 tiêu hao 1.6 tấn quặng + 0.6 tấn than cốc cho 1 tấn HRC → nếu cả quặng và than cùng giảm 10 USD/tấn "
        "= spread tăng ~22 USD/tấn (1.6×10 + 0.6×10)"
    )
    add_body(
        "4. <b>NHÀ MÁY MỚI (DQ2) ✅ — ĐÃ HOÀN THÀNH</b><br/>"
        f"- DQ2 (5.6 triệu tấn HRC) chạy full công suất từ T12/2025<br/>"
        f"- CIP giảm từ 63,656 tỷ (2024) → {cip_hist[4]:,.0f} tỷ (2025) → tiếp tục giảm 2026<br/>"
        "- DQ2 sử dụng công nghệ BOF Nhật Bản, tiêu hao năng lượng thấp hơn DQ1 15%<br/>"
        "- Tổng công suất HPG đang hoạt động đạt ~14,5 triệu tấn/năm (6 nhà máy) + 500k TEUs container. "
        "Thêm 3 dự án tương lai (Đắk Lắk 6M, Ray cao tốc 700k, Ống Long An 400k) sẽ nâng lên ~21,6 triệu tấn. "
        "Lớn nhất Đông Nam Á."
    )
    add_body(
        "<b>KẾT LUẬN HARVEST SIGNALS: HPG ĐÃ HỘI TỤ ĐỦ 4 YẾU TỐ</b><br/>"
        "Đây là thời điểm lý tưởng để nắm giữ HPG trong 12-24 tháng tới khi (i) sản lượng tăng gấp đôi nhờ DQ2, "
        "(ii) biên lợi nhuận mở rộng nhờ chi phí thấp, (iii) định giá P/B 1.56x còn hấp dẫn so với median lịch sử. "
        "Các quý Q2-Q4/2026 sẽ là thời điểm HPG ghi nhận lợi nhuận kỷ lục, tạo động lực cho cổ phiếu."
    )
    add_body(
        "<b>5. Diễn biến tiêu thụ sản phẩm & Vị trí chu kỳ:</b><br/>"
        f"- <b>HRC:</b> Sản lượng tăng từ 2.0 triệu tấn (2021) → 3.2 triệu tấn (2025) → 6.0 triệu tấn (2026E). "
        f"Nhu cầu HRC nội địa ~8 triệu tấn (2026E). HPG + FMS cung cấp ~80%. Với thuế CBPG 27.8% HRC TQ, "
        f"HPG gần như độc quyền thị trường HRC nội địa. <b>Chưa qua đỉnh — sản lượng mới bắt đầu tăng tốc.</b><br/>"
        f"- <b>Thép XD:</b> Sản lượng ổn định ~2.5-3.0 triệu tấn/năm. Động lực: đầu tư công cao tốc, "
        f"BĐS phục hồi. Mùa cao điểm xây dựng Q2-Q4 giúp sản lượng tăng 15-20% so với Q1.<br/>"
        f"- <b>Giá HRC:</b> {hrc_now:,.0f} USD/tấn hiện tại (fetch tự động), {Q18_HRC[-1]:,.0f} USD/tấn Q1/2026 — "
        f"còn cách xa đỉnh {max(Q18_HRC):,.0f} USD ({Q18_LABELS[Q18_HRC.index(max(Q18_HRC))]}). "
        f"<b>Giá đang ở vùng thấp của biên độ 18 quý gần nhất, dư cung TQ vẫn là lực cản chính.</b><br/>"
        f"- <b>Chu kỳ thép toàn cầu:</b> Đáy giá 2025-2026 do dư cung TQ ~640 triệu tấn kéo dài. "
        f"Kịch bản cơ sở dùng giá hàng hóa hiện tại (xem sheet 17_Gia_Hang_Hoa), không giả định giá phục hồi mạnh.<br/>"
        f"- <b>Kết luận chu kỳ: HPG đang ở <u>đầu pha tăng tốc</u> của chu kỳ lợi nhuận.</b> "
        f"Sản lượng chưa qua đỉnh (DQ2 chỉ mới full 6 tháng), giá HRC chưa qua đỉnh (còn cách 2021 ~45%), "
        f"định giá chưa qua đỉnh (P/B 1.56x vs đỉnh 2021 là 2.8x). "
        f"Rủi ro chính: dư cung TQ kéo giá HRC xuống dưới 550 USD, làm giảm spread và GM về 12-14%."
    )

    elements.append(Spacer(1, 4*mm))
    add_section("5. ĐỊNH GIÁ")
    add_body(
        "Multiple dựa trên <b>trung vị lịch sử HPG</b> (TTM 2018-2026 từ Vietcap). "
        f"EPS 2026E: {eps_2026e_val:,} VND. BVPS 2026E: {bvps_2026e_val:,.0f} VND."
    )

    pe_price = PE_MULTIPLE * eps_2026e_val
    pb_price = round(PB_MULTIPLE * bvps_2026e_val)
    pb_upper = round(PB_MULTIPLE * 1.2 * bvps_2026e_val)
    pb_attr  = round(PB_MULTIPLE * 0.8 * bvps_2026e_val)
    ev_price = round((ebitda_2026e_val * EV_MULTIPLE - net_debt_2026e_val) * 1e9 / SHARES)
    target_price = round(ev_price * 0.4 + pb_price * 0.4 + pe_price * 0.2)
    add_body(
        f"- <b>EV/EBITDA ({EV_MULTIPLE}x, HPG median)</b> → <b>{ev_price:,} VND</b><br/>"
        f"- <b>P/B ({PB_MULTIPLE}x, HPG median)</b> → <b>{pb_price:,} VND</b>  |  "
        f"P/B cao ({PB_MULTIPLE*1.2:.1f}x) → <b>{pb_upper:,} VND</b>  |  "
        f"P/B hấp dẫn ({PB_MULTIPLE*0.8:.1f}x) → <b>{pb_attr:,} VND</b><br/>"
        f"- <b>P/E ({PE_MULTIPLE}x, HPG median)</b> → <b>{pe_price:,} VND</b>"
    )
    add_body(
        f"<b>Giá mục tiêu (40% EV/EBITDA + 40% P/B + 20% P/E): {target_price:,} VND</b>  |  "
        f"Giá hiện tại: {PRICE:,} VND  |  Upside: <b>+{round((target_price/PRICE-1)*100):.0f}%</b>"
    )
    add_body(f"<b>Tham chiếu broker:</b> SSI 36k | VNDirect ~36k | SHS 38k | VCBS 38k | BVSC 38.65k.")

    # Sensitivity chart
    sens_path = os.path.join(CHART_DIR, "sensitivity.png")
    if os.path.exists(sens_path):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("Biểu đồ 13: Sensitivity EV/EBITDA × EBITDA", styles['SmallText']))
        elements.append(Image(sens_path, width=460, height=280))

    elements.append(PageBreak())
    add_section("6. RỦI RO & CATALYSTS")
    add_body("<b>Risk Matrix:</b>")

    risk_headers = ["Rủi ro", "Xác suất", "Tác động", "Chỉ báo dẫn dắt"]
    risk_data = [
        ["Giá HRC giảm (suy thoái TQ)", "Trung bình", "Cao",
         "Giá HRC SHFE, PMI TQ, sản lượng thép TQ"],
        ["Chi phí quặng/than tăng", "Cao", "Trung bình",
         "Giá quặng 62% Fe, than cốc FOB Úc"],
        ["Tỷ giá USD tăng mạnh", "Trung bình", "Cao",
         "USD/VNĐ, D/E, lãi suất Fed"],
        ["Cạnh tranh HRC giá rẻ", "Thấp", "Trung bình",
         "Thuế CBPG, giá HRC Ấn Độ/Hàn"],
        ["BĐS VN suy thoái kéo dài", "Thấp", "Cao",
         "Tồn kho BĐS, giải ngân đầu tư công"],
    ]
    add_table(risk_headers, risk_data,
              [doc.width*0.30, doc.width*0.18, doc.width*0.18, doc.width*0.34])

    add_body("<b>Catalysts (6-12 tháng tới):</b>")
    add_body("1. <b>KQKD Q2/Q3 2026</b> — DQ2 full năm sẽ thể hiện rõ trong các quý tới. "
             "LNST mỗi quý có thể đạt 6,000-7,000 tỷ.")
    add_body("2. <b>Công bố DQ3</b> — HPG đang nghiên cứu DQ3 (5.6M tấn, vốn ~$3 tỷ). "
             "Phê duyệt đầu tư sẽ là catalyst lớn.")
    add_body("3. <b>Chính sách đầu tư công</b> — Giải ngân cao tốc Bắc-Nam, Long Thành "
             "sẽ đẩy nhu cầu thép xây dựng 6-12 tháng cuối 2026.")
    add_body("4. <b>Cổ tức 2026</b> — 15% (10% CP + 5% TM) là điểm hấp dẫn cho dòng tiền.")

    # Peer comparison
    elements.append(Spacer(1, 4*mm))
    add_section("SO SÁNH PEER")
    peer_headers = ["Ticker", "P/E (x)", "P/B (x)", "ROE (%)", "EV/EBITDA (x)", "Tăng trưởng LN 26E (%)"]
    peer_data = [
        ["HPG", "12.5", "1.80", "15.6", "6.8", "42.0"],
        ["HSG", "13.7", "0.90", "7.1", "7.2", "25.0"],
        ["NKG", "22.5", "0.90", "6.4", "9.0", "18.0"],
    ]
    add_table(peer_headers, peer_data,
              [doc.width*0.10, doc.width*0.15, doc.width*0.15, doc.width*0.15,
               doc.width*0.15, doc.width*0.20])

    peers_path = os.path.join(CHART_DIR, "peers.png")
    if os.path.exists(peers_path):
        elements.append(Spacer(1, 4*mm))
        elements.append(Paragraph("Biểu đồ 14: So sánh định giá Peer 2026E", styles['SmallText']))
        elements.append(Image(peers_path, width=400, height=250))

    elements.append(PageBreak())
    add_section("7. KẾT LUẬN")
    add_body(
        f"<b>KHUYẾN NGHỊ: MUA — Giá mục tiêu {weighted_price_val:,.0f} VND (+{upside_val:.0f}%)</b>"
    )
    add_body(
        "HPG đang ở điểm uốn chu kỳ lợi nhuận: Dung Quất 2 full công suất, giá HRC hồi phục, "
        "thuế CBPG bảo vệ thị trường nội địa. Với P/B 1.56x (dưới median lịch sử HPG 1.61x) và "
        "EV/EBITDA forward ~7.0x (dưới median HPG 8.95x), định giá còn hấp dẫn so với chính lịch sử giao dịch."
    )
    add_body(
        "<b>Leading indicators cần theo dõi:</b>"
    )
    li_headers = ["Chỉ báo", "Ngưỡng tích cực", "Hiện tại", "Trạng thái"]
    li_data = [
        ["Giá HRC (USD/tấn)", ">600", f"{hrc_now:,.0f} (fetch tự động)", "TÍCH CỰC" if hrc_now > 600 else "THEO DÕI"],
        ["Spread thép (USD/tấn)", ">50", f"{SPREAD_A[5]:,.0f} ({years_fc[0]}E)", "TÍCH CỰC" if SPREAD_A[5] > 50 else "THEO DÕI"],
        ["LNST QoQ (tỷ)", ">7,000", "9,056 (Q1)", "TÍCH CỰC"],
        ["D/E", "<0.7", "0.65", "AN TOÀN"],
        ["CIP (tỷ)", "<10,000", "~15,000", "GIẢM TỐT"],
    ]
    add_table(li_headers, li_data,
              [doc.width*0.25, doc.width*0.22, doc.width*0.22, doc.width*0.31])

    add_body("<b>Tổng quan 6 tầng:</b><br/>")
    add_body("✅ Tầng 1 (Chuỗi GT): Mạnh — Tích hợp dọc, tự chủ 50% quặng<br/>"
             "✅ Tầng 2 (Thị trường): Mạnh — Dẫn đầu thị phần, CBPG bảo vệ<br/>"
             "✅ Tầng 3 (Moat): Rộng — Chi phí thấp, quy mô, thương hiệu<br/>"
             "✅ Tầng 4 (Tài chính): Tốt — D/E 0.65x, ROE 14.9% 2026E<br/>"
             "⚠️ Tầng 5 (Rủi ro): TB — Giá HRC, tỷ giá, chi phí đầu vào<br/>"
             f"✅ Tầng 6 (Định giá): Hấp dẫn — P/B 1.56x (dưới median {PB_MULTIPLE}x HPG), EV/EBITDA ~7.0x (dưới median {EV_MULTIPLE}x HPG)")

    add_body(f"<br/><b>Quan điểm đầu tư:</b> Mua tích lũy ở vùng giá {int(weighted_price_val*0.8):,}-{int(weighted_price_val*0.87):,} VND cho mục tiêu "
             f"{weighted_price_val:,.0f} VND (+{upside_val:.0f}%) trong 12 tháng. Nếu giá HRC duy trì trên 600 USD/tấn, kết quả kinh doanh "
             "có thể vượt kế hoạch ĐHĐCĐ (22,000 tỷ LNST).")

    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph(
        "Báo cáo này được tạo bởi AI FA Framework v2.0. Dữ liệu lấy từ Cafef, Vietstock, "
        "Trading Economics, HOSE, VSA. Ngày 24/06/2026. "
        "Đây không phải tư vấn đầu tư tài chính.",
        styles['SmallText']
    ))

    # Build PDF
    doc.build(elements)
    print(f"[OK] PDF report saved: {PDF_FILE}")
    return True


# ── JSON EXPORT (for web dashboard) ───────────────────────────────────────

def save_json_summary():
    import json
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(data_dir, exist_ok=True)

    eps_fc = [round(ni_fc[i] * 1e9 / (shares_fc[i] * 1e6)) for i in range(3)]
    all_years = years_hist + years_fc
    all_rev = revenue_hist + revenue_fc
    all_npat = ni_hist + ni_fc
    all_eps = eps_hist + eps_fc
    all_equity = equity_hist + equity_fc_val

    # Ratio calculations
    all_gpm = [round(g/100, 4) for g in gp_margin_hist] + [round(g/100, 4) for g in gp_margin_fc]
    all_roe = [round(ni_hist[i]/equity_hist[i], 4) for i in range(5)] + [round(ni_fc[i]/equity_fc_val[i], 4) for i in range(3)]
    debt_fc = [78000, 75000, 72000]
    all_de = [round(total_debt_hist[i]/equity_hist[i], 4) for i in range(5)] + [round(debt_fc[i]/equity_fc_val[i], 4) for i in range(3)]
    all_ebit = ebit_hist + [round(ebit_fc[i]) for i in range(3)]
    all_ebit_margin = [round(all_ebit[i]/all_rev[i], 4) for i in range(8)]

    # ── Quarterly data for steel charts ──
    is_qs = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
    bs_qs = section_to_quarters(FIN_DATA, "BALANCE_SHEET")
    def _get_q(records, yr, qtr, field):
        for r in records:
            if r.get("yearReport") == yr and r.get("lengthReport") == qtr:
                v = r.get(field)
                return v / 1e9 if v is not None else 0
        return 0
    sl_pairs = [(2023,1),(2023,2),(2023,3),(2023,4),(2024,1),(2024,2),(2024,3),(2024,4),
                (2025,1),(2025,2),(2025,3),(2025,4),(2026,1)]
    sl_labels = []; sl_rev = []; sl_gp = []; sl_gpm = []; sl_sgka = []; sl_ni = []; sl_ebit = []
    sl_sell = []; sl_admin = []; sl_sgka_ratio = []
    for yr, qnum in sl_pairs:
        rev = _get_q(is_qs, yr, qnum, 'isa3'); gp = _get_q(is_qs, yr, qnum, 'isa5')
        sell = abs(_get_q(is_qs, yr, qnum, 'isa9')); admin = abs(_get_q(is_qs, yr, qnum, 'isa10'))
        ni = _get_q(is_qs, yr, qnum, 'isa22'); ebit = _get_q(is_qs, yr, qnum, 'isa11')
        sl_labels.append(f"Q{qnum}/{yr}"); sl_rev.append(round(rev)); sl_gp.append(round(gp))
        sl_gpm.append(round(gp/rev*100, 1) if rev else 0)
        sl_sell.append(round(sell)); sl_admin.append(round(admin))
        sl_sgka.append(round(sell+admin)); sl_sgka_ratio.append(round((sell+admin)/rev*100, 1) if rev else 0)
        sl_ni.append(round(ni)); sl_ebit.append(round(ebit))
    # Balance sheet by quarter
    sl_assets = []; sl_inv = []; sl_rec = []; sl_debt = []
    for yr, qnum in sl_pairs:
        a = _get_q(bs_qs, yr, qnum, 'bsa53'); inv = _get_q(bs_qs, yr, qnum, 'bsa15')
        rec = _get_q(bs_qs, yr, qnum, 'bsa8'); sd = _get_q(bs_qs, yr, qnum, 'bsa56')
        ld = _get_q(bs_qs, yr, qnum, 'bsa71')
        sl_assets.append(round(a)); sl_inv.append(round(inv)); sl_rec.append(round(rec))
        sl_debt.append(round(sd+ld))
    # Spread & HRC price by quarter — dùng đúng Spread lag-1-quý THỰC TẾ của quý đó (Q18_SPREAD),
    # KHÔNG lấy giá trị Spread NĂM lặp lại cho mọi quý trong năm như trước (sai lệch cấu trúc dữ
    # liệu chart, khiến biểu đồ dashboard web trông giống bậc thang thay vì diễn biến quý thực).
    _q18_spread_by_lbl = dict(zip(Q18_LABELS, Q18_SPREAD))
    _q18_hrc_by_lbl = dict(zip(Q18_LABELS, Q18_HRC))
    # Spread Rebar/All theo quý (2026-07, theo yêu cầu user hiển thị 3 loại spread trên web) — Spread
    # All = None cho 5 quý đầu (2021Q4-2022Q4, chưa có SL tách riêng — xem chú thích Q18_SPREAD_ALL).
    _q18_spread_rebar_by_lbl = dict(zip(Q18_LABELS, Q18_SPREAD_REBAR))
    _q18_spread_all_by_lbl = dict(zip(Q18_LABELS, Q18_SPREAD_ALL))
    _q18_xd_by_lbl = dict(zip(Q18_LABELS, Q18_XD))
    sl_spread = []; sl_hrc = []; sl_spread_rebar = []; sl_spread_all = []; sl_xd_price = []
    for yr, qnum in sl_pairs:
        lbl = f"{yr}Q{qnum}"
        sl_spread.append(round(_q18_spread_by_lbl.get(lbl, SPREAD_NOW)))
        sl_hrc.append(_q18_hrc_by_lbl.get(lbl, hrc_now))
        sl_spread_rebar.append(round(_q18_spread_rebar_by_lbl.get(lbl, SPREAD_REBAR_NOW)))
        _sa = _q18_spread_all_by_lbl.get(lbl)
        sl_spread_all.append(round(_sa) if _sa is not None else None)
        sl_xd_price.append(_q18_xd_by_lbl.get(lbl, xd_now))
    # Nguồn duy nhất HRC_SALES_HIST_KT/XD_SALES_HIST_KT (module-level, dùng chung với sheet 15
    # Excel) — tránh 2 mảng trùng lặp lệch số như trước. Giữ đúng độ dài khớp sl_pairs/sl_labels
    # (không thêm quý đang chạy vào đây — xem sheet 15 Excel mục D/E/F cho số liệu quý đang chạy).
    hrc_sales = list(HRC_SALES_HIST_KT)
    xd_sales = list(XD_SALES_HIST_KT)
    # Annual tables data (matching PDF tables)
    ann_labels = ["2021","2022","2023","2024","2025","2026E","2027E","2028E"]
    ann_hrc_price = HRC_PRICE_A[:]
    ann_iron_ore = IRON_ORE_A[:]
    ann_coke = COKE_A[:]
    ann_spread = SPREAD_A[:]
    ann_spread_pct = [round(SPREAD_A[i]/HRC_PRICE_A[i]*100, 0) if HRC_PRICE_A[i] else 0 for i in range(8)]
    # Spread Rebar/All + giá thép XD theo NĂM (2026-07, theo yêu cầu user hiển thị 3 loại spread) —
    # cùng phương pháp median-quý/blend-hiện tại đã dùng cho HRC_PRICE_A/SPREAD_A ở trên.
    _xd_by_yr = _group_by_year(Q18_XD)
    ann_xd_price = [_year_median(_xd_by_yr, y) for y in [2021,2022,2023,2024,2025]] + [
        round(stats.median([_year_median(_xd_by_yr, 2026), xd_now]), 1), XD_PRICE_A[6], XD_PRICE_A[7]]
    ann_spread_rebar = SPREAD_REBAR_A[:]
    ann_spread_all = SPREAD_ALL_A[:]
    ann_hrc_vol = [2.0,2.2,2.5,2.8,3.2,6.0,6.8,7.5]
    ann_xd_vol = [2.8,2.6,2.3,2.5,2.8,3.0,3.2,3.5]
    ann_total_vol = [ann_hrc_vol[i]+ann_xd_vol[i] for i in range(8)]
    ann_market_size = [28,27,25,26,28,30,32,34]
    ann_market_share = [round(ann_total_vol[i]/ann_market_size[i]*100, 0) for i in range(8)]
    # Peer comparison
    peer_data = [
        {"ticker":"HPG","pe":12.5,"pb":1.80,"roe":15.6,"ev_ebitda":6.8,"ni_growth":42.0},
        {"ticker":"HSG","pe":13.7,"pb":0.90,"roe":7.1,"ev_ebitda":7.2,"ni_growth":25.0},
        {"ticker":"NKG","pe":22.5,"pb":0.90,"roe":6.4,"ev_ebitda":9.0,"ni_growth":18.0},
    ]
    # Factory list
    factories = [
        {"name":"Dung Quất 1","location":"Quảng Ngãi","product":"HRC, Thép XD, Ống thép","capacity":"5.0","status":"Hoạt động"},
        {"name":"Dung Quất 2","location":"Quảng Ngãi","product":"HRC","capacity":"5.6","status":"Hoạt động (T12/2025)"},
        {"name":"Hải Dương","location":"Hải Dương","product":"Thép XD, Phôi thép","capacity":"2.5","status":"Hoạt động"},
        {"name":"Ống thép HPG","location":"Hải Dương","product":"Ống thép đen, ống thép mạ kẽm","capacity":"1.0","status":"Hoạt động"},
        {"name":"Tôn Hòa Phát","location":"Hải Dương","product":"Tôn mạ màu, tôn mạ kẽm","capacity":"0.4","status":"Hoạt động"},
        {"name":"Container HPG","location":"Bà Rịa-Vũng Tàu","product":"Container 20', 40'","capacity":"0.5","status":"Hoạt động"},
        {"name":"Đắk Lắk","location":"Đắk Lắk","product":"Thép XD","capacity":"6.0","status":"Chưa HĐ (KH 2028+)"},
        {"name":"Ray cao tốc","location":"Hưng Yên","product":"Ray đường sắt cao tốc","capacity":"0.7","status":"Chưa HĐ"},
        {"name":"Ống thép Long An","location":"Long An","product":"Ống thép","capacity":"0.4","status":"Chưa HĐ"},
    ]

    # PnL-matching derivation (same as Excel 04_PnL) — dùng lại ebit_fc/ni_fc bottom-up ở trên.
    # ev_mul/pb_mul/pe_mul và net_debt_26 dùng ĐÚNG nguồn số đã ghi vào Excel (xem PE_HIST_MEDIAN/
    # PB_HIST_MEDIAN/EV_HIST_MEDIAN/NET_DEBT_2026E ở đầu file) — không tính lại riêng cho JSON/web.
    ev_mul = EV_HIST_MEDIAN; pb_mul = PB_HIST_MEDIAN; pe_mul = PE_HIST_MEDIAN
    ebitda_26 = ebit_fc[0] + da_fc[0]
    net_debt_26 = NET_DEBT_2026E
    eps_26 = round(ni_fc[0] * 1e9 / SHARES)
    bvps_26 = round(equity_fc_val[0] * 1e9 / SHARES)
    ev_price = max(0, (ebitda_26 * ev_mul - net_debt_26) * 1e9 / SHARES)
    pb_price = round(pb_mul * bvps_26)
    pe_price = round(pe_mul * eps_26)
    target_price = round(ev_price * 0.4 + pb_price * 0.4 + pe_price * 0.2)
    upside = round((target_price / PRICE - 1) * 100, 1)
    pb_upper = round(pb_mul * 1.2 * bvps_26)
    pb_attr  = round(pb_mul * 0.8 * bvps_26)

    # PE/PB history arrays
    pe_arr = []
    pb_arr = []
    q_labels = []
    for r in HPG_RATIOS:
        q = r.get("quarter")
        if q is None or q == 5: continue
        pe0 = r.get("pe"); pb0 = r.get("pb")
        pe_arr.append(round(pe0, 1) if pe0 and 0 < pe0 < 50 else None)
        pb_arr.append(round(pb0, 2) if pb0 and pb0 > 0 else None)
        q_labels.append(f"{int(r['year'])}-Q{int(q)}")

    # Dùng ĐÚNG PE_HIST_MEDIAN/PB_HIST_MEDIAN (nguồn số dùng chung cho Excel/PDF ở đầu file) thay vì
    # tính median phẳng riêng ở đây — trước đây 2 cách tính khác nhau khiến "P/E Mục tiêu" trên web
    # (9.7x) lệch với "P/E" hiển thị trong PDF/Excel (10.3x) dù cùng ý nghĩa "trung vị lịch sử".
    pb_median = PB_HIST_MEDIAN
    pe_median = PE_HIST_MEDIAN

    # Valuation scenarios
    bear_price = round(pb_mul * 0.8 * bvps_26)
    base_price = target_price
    bull_price = round(pb_mul * 1.2 * bvps_26)

    summary = {
        "ticker": TICKER,
        "companyName": COMPANY,
        "sector": "Basic Resources",
        "currentPrice": PRICE,
        "marketCap": MARKET_CAP,
        "shares": SHARES,
        "gdrivePdfUrl": None,
        "gdriveExcelUrl": None,
        "productionEstimateCache": PRODUCTION_CACHE_ENTRY,
        "lastUpdated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "data": {
            "years": all_years,
            "revenue": [round(v, 1) for v in all_rev],
            "npat": [round(v, 1) for v in all_npat],
            "eps": [round(v) for v in all_eps],
            "equity": [round(v, 1) for v in all_equity],
        },
        "ratios": {
            "gross_margin": all_gpm,
            "roe": all_roe,
            "debt_to_equity": all_de,
            "ebitMargin": all_ebit_margin,
            "nim": [],
            "npl": [],
            "ldr": [],
            "casa": [],
        },
        "valuation": {
            "weightedTarget": target_price,
            "upside": upside,
            "recommend": "MUA",
            "bear": bear_price,
            "base": base_price,
            "bull": bull_price,
            "COE": 12.0,
            "evEbitdaTarget": round(ev_price),
            "pbAttractive": round(pb_mul * 0.8, 2),
            "pbAttractivePrice": bear_price,
            "pbMedian": pb_median,
            "pbOver": round(pb_mul * 1.2, 2),
            "peMedian": pe_median,
            "peTarget": round(pe_price),
            "evWeight": 40,
            "pbWeight": 40,
            "peWeight": 20,
            "bvpsBase": round(bvps_26),
        },
        "thesis": [
            "HPG là nhà sản xuất thép tích hợp dọc lớn nhất VN với 6 nhà máy (~14,5 triệu tấn/năm). Dung Quất 2 (5.6 triệu tấn HRC) full công suất từ T12/2025, đưa HPG vào nhóm chi phí thấp nhất khu vực. Thuế CBPG 27.8% với HRC Trung Quốc bảo vệ thị trường nội địa.",
            f"Biên LNG dự phóng neo vào biên LNG THỰC TẾ quý gần nhất, nội suy theo tỷ lệ spread (không lấy trực tiếp spread làm biên). GP margin dự phóng từ {gp_margin_hist[4]}% ({years_hist[4]}) lên {gp_margin_fc[0]}% ({years_fc[0]}E). Sản lượng HRC tăng gấp đôi nhờ DQ2 full năm.",
            "Định giá hấp dẫn: P/B 1.56x (dưới median 1.61x), EV/EBITDA 2026E ~7.0x (dưới median 8.95x). Giá mục tiêu 40% EV/EBITDA + 40% P/B + 20% P/E cho upside +42%."
        ],
        "risks": [
            f"Giá HRC ({hrc_now:,.0f} USD/tấn hiện tại) chịu áp lực dư cung toàn cầu ~640 triệu tấn, kéo spread thu hẹp (spread {years_fc[0]}E ước ~{SPREAD_A[5]:,.0f} USD/tấn theo giá hàng hóa hiện tại).",
            "Chi phí quặng sắt/than cốc tăng làm giảm spread. Mỗi 10 USD tăng quặng = ~16 USD/tấn giảm spread.",
            "Tỷ giá USD/VNĐ biến động mạnh — dư nợ vay USD ~$3 tỷ, ảnh hưởng đến chi phí tài chính."
        ],
        "moats": {
            "Cost Advantage": {"score": 4, "desc": "Quy mô 14,5 triệu tấn/năm, DQ2 công nghệ BOF Nhật Bản, tiết kiệm 15-20% chi phí so với đối thủ."},
            "Efficient Scale": {"score": 4, "desc": "Dẫn đầu thị phần HRC (~70%) và thép XD (~40%). DQ2 tăng tổng công suất lên ~14,5 triệu tấn."},
            "Intangible Assets": {"score": 3, "desc": "Thương hiệu Hòa Phát uy tín trong ngành thép VN, quan hệ đại lý lâu năm."},
            "Switching Cost": {"score": 2, "desc": "Thép là hàng hóa tiêu chuẩn, khách hàng có thể chuyển đổi nhà cung cấp nhưng chi phí logistics cao."},
            "Network Effect": {"score": 1, "desc": "Hiệu ứng mạng lưới hạn chế trong ngành thép."},
        },
        "pestle": {
            "Political": "Thuế CBPG HRC 27.83% bảo vệ thị trường nội địa. Ổn định chính trị VN tạo môi trường đầu tư thuận lợi.",
            "Economic": "GDP VN 2026E ~6.5%. Đầu tư công kế hoạch trung hạn 8.5 triệu tỷ (2026-2030). Lãi suất giảm hỗ trợ BĐS.",
            "Social": "Đô thị hoá ~38%, nhu cầu nhà ở và hạ tầng tăng mạnh, thúc đẩy tiêu thụ thép.",
            "Technological": "DQ2 công nghệ BOF Nhật Bản, tiêu hao năng lượng thấp hơn DQ1 15%. HPG đầu tư tự động hóa.",
            "Legal": "CBAM EU buộc DN thép VN chuyển đổi xanh. NĐ 153 siết chặt trái phiếu doanh nghiệp.",
            "Environmental": "Áp lực giảm phát thải carbon. HPG đầu tư xử lý nước thải, lọc bụi đạt chuẩn quốc tế.",
        },
        "comments": {
            "overall": "HPG là doanh nghiệp thép tích hợp dọc hàng đầu VN với lợi thế quy mô và công nghệ vượt trội. Chu kỳ tăng trưởng mới đang bắt đầu nhờ DQ2 full công suất và bảo hộ thương mại HRC.",
            "financial": "Doanh thu và LNST phục hồi mạnh từ đáy 2023. Biên LNG cải thiện nhờ spread mở rộng. CFO/LNST >1.0 phản ánh chất lượng lợi nhuận tốt. D/E ~0.65 là an toàn.",
            "valuation": "P/B 1.56x dưới median lịch sử 1.61x. EV/EBITDA 2026E ~7.0x thấp hơn median 8.95x. Kết hợp 40% EV/EBITDA + 40% P/B + 20% P/E cho upside +42%.",
        },
        "quarterly": {
            "labels": sl_labels,
            "revenue": sl_rev, "grossProfit": sl_gp, "gpMargin": sl_gpm,
            "sellExpense": sl_sell, "adminExpense": sl_admin,
            "sgka": sl_sgka, "sgkaRatio": sl_sgka_ratio,
            "netIncome": sl_ni, "ebit": sl_ebit,
            "totalAssets": sl_assets, "inventory": sl_inv,
            "receivables": sl_rec, "totalDebt": sl_debt,
            "spreadUsd": sl_spread, "hrcPrice": sl_hrc,
            "spreadRebarUsd": sl_spread_rebar, "spreadAllUsd": sl_spread_all, "xdPrice": sl_xd_price,
            "hrcSales": hrc_sales, "xdSales": xd_sales,
        },
        "annualTables": {
            "labels": ann_labels,
            "hrcPrice": ann_hrc_price, "ironOre": ann_iron_ore, "coke": ann_coke,
            "spreadUsd": ann_spread, "spreadPct": ann_spread_pct,
            "xdPrice": ann_xd_price, "spreadRebarUsd": ann_spread_rebar, "spreadAllUsd": ann_spread_all,
            "hrcVol": ann_hrc_vol, "xdVol": ann_xd_vol,
            "totalVol": ann_total_vol, "marketSize": ann_market_size, "marketShare": ann_market_share,
        },
        "commodityQuarterly": {
            "labels": Q18_LABELS,
            "hrcPrice": Q18_HRC, "ironOre": Q18_IRON, "cokingCoal": Q18_COAL, "xdPrice": Q18_XD,
            "current": {"hrcPrice": hrc_now, "ironOre": iron_now, "cokingCoal": coal_now, "xdPrice": xd_now},
            "spreadHrcUsd": Q18_SPREAD_HRC, "spreadRebarUsd": Q18_SPREAD_REBAR, "spreadAllUsd": Q18_SPREAD_ALL,
            "spreadNow": round(SPREAD_NOW, 1),
            "spreadHrcNow": round(SPREAD_HRC_NOW, 1), "spreadRebarNow": round(SPREAD_REBAR_NOW, 1),
            "spreadAllNow": round(SPREAD_ALL_NOW, 1),
            "adDutyEffectiveQuarter": AD_HRC_EFFECTIVE_Q, "adDutyMultiplier": AD_HRC_MULTIPLIER,
            "note": "Gia quy = TB dau/giua/cuoi quy. Gia hien tai fetch tu dong tu investing.com/SteelOnline khi chay script. Spread HRC da nhan thue CBPG AD20 tu quy hieu luc.",
        },
        "factories": factories,
        "peers": peer_data,
        "pe_hist": [round(v, 1) for v in [14.2, 25.5, 13.8, 13.9, 9.5]] if False else [],
        "pb_hist": [round(v, 2) for v in [1.67, 1.69, 1.45, 1.25, 2.14]] if False else [],
        "pe_quarters": pe_arr,
        "pb_quarters": pb_arr,
        "quarter_labels": q_labels,
    }

    json_path = os.path.join(data_dir, f"{TICKER}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"[OK] JSON summary saved: {json_path}")


# ── MAIN ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("   HPG ANALYSIS - Excel Model + PDF Report")
    print(f"   Date: {MONTH}")
    print("=" * 60)

    build_excel()
    make_charts()
    try:
        build_pdf()
    except Exception as e:
        print(f"[WARN] PDF generation failed: {e}")
    save_json_summary()

    print(f"\n  COMPLETE")
    print(f"  Excel: {EXCEL_FILE}")
    print(f"  PDF:   {PDF_FILE}")
    print(f"  Charts: {CHART_DIR}\\")
