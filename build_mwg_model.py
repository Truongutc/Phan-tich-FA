#!/usr/bin/env python3
"""
MWG (CTCP Đầu tư Thế Giới Di Động) — Excel Model + PDF Report Generator
Ngành: Bán lẻ | Skill: .opencode/skills/ban-le/SKILL.md
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    # Windows console mặc định dùng codepage cp1252 — print() tiếng Việt có dấu (kể cả title bài
    # viết fetch động từ API IR, không đoán trước được để viết ASCII thủ công) sẽ crash
    # UnicodeEncodeError. Ép stdout/stderr UTF-8 để tránh crash cả trên Windows lẫn Linux CI.
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import os
import re
import json
import math
import tempfile
import subprocess
import statistics as stats
from datetime import datetime, date
import requests
import numpy as np
try:
    import pdfplumber
except ImportError:
    pdfplumber = None
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, Image, KeepTogether)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import fetch_data
from fetch_data import (section_to_years, section_to_quarters, get_val,
                         cumulative_actual_quarters, latest_actual_quarter_value,
                         blend_annual_estimate, blend_annual_estimate_stock)

TICKER = "MWG"
COMPANY = "CTCP Đầu tư Thế Giới Di Động"
EXCHANGE = "HOSE"
INDUSTRY = "Bán lẻ"
MONTH = datetime.now().strftime("%Y_%m_%d")
OUT_DIR = os.path.join("Bao cao", TICKER)
CHART_DIR = os.path.join(OUT_DIR, "charts")
os.makedirs(CHART_DIR, exist_ok=True)
EXCEL_FILE = os.path.join(OUT_DIR, f"{TICKER}_Model_{MONTH}.xlsx")
PDF_FILE = os.path.join(OUT_DIR, f"{TICKER}_Phan_Tich_{MONTH}.pdf")
JSON_FILE = os.path.join("data", f"{TICKER}.json")
os.makedirs("data", exist_ok=True)

# ── Font (Windows / Linux GitHub Actions) ───────────────────────────────────
FONT_NAME = "Arial"
FONT_PATH_WIN = "C:/Windows/Fonts/arial.ttf"
FONT_PATH_WIN_BOLD = "C:/Windows/Fonts/arialbd.ttf"
FONT_PATH_LINUX = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
FONT_PATH_LINUX_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
PDF_FONT = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
if os.path.exists(FONT_PATH_WIN):
    pdfmetrics.registerFont(TTFont("VNFont", FONT_PATH_WIN))
    pdfmetrics.registerFont(TTFont("VNFont-Bold", FONT_PATH_WIN_BOLD))
    PDF_FONT, PDF_FONT_BOLD = "VNFont", "VNFont-Bold"
elif os.path.exists(FONT_PATH_LINUX):
    pdfmetrics.registerFont(TTFont("VNFont", FONT_PATH_LINUX))
    pdfmetrics.registerFont(TTFont("VNFont-Bold", FONT_PATH_LINUX_BOLD))
    PDF_FONT, PDF_FONT_BOLD = "VNFont", "VNFont-Bold"

# ── Styling helpers (openpyxl) — cùng quy ước với build_hpg_model.py ────────
title_font = Font(name=FONT_NAME, size=16, bold=True, color="1F4E79")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
bold_font = Font(name=FONT_NAME, bold=True, size=10)
data_font = Font(name=FONT_NAME, size=10)
assump_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
p_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
n_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
thin_border = Border(*(Side(style='thin', color="CCCCCC"),) * 4)

def header_row(ws, row, values, widths=None):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def data_row(ws, row, vals, bold=False, fill=None, fmt=None):
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = bold_font if bold else data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fill:
            cell.fill = fill
        if fmt and i <= len(fmt):
            cell.number_format = fmt[i-1]

# ── 1. FETCH DỮ LIỆU TÀI CHÍNH (Vietcap API, dùng chung fetch_data.py) ──────
print(f"[Fetch] Loading {TICKER} financial data...")
RAW = fetch_data.fetch_all(TICKER, use_cache=True)
BS = RAW["sections"]["BALANCE_SHEET"]
IS = RAW["sections"]["INCOME_STATEMENT"]
CF = RAW["sections"]["CASH_FLOW"]
bs_hist = section_to_years(RAW, "BALANCE_SHEET")
is_hist = section_to_years(RAW, "INCOME_STATEMENT")
cf_hist = section_to_years(RAW, "CASH_FLOW")

years_hist = sorted({r["yearReport"] for r in is_hist if r.get("yearReport")})[-5:]
YEAR_NOW = date.today().year
years_fc = [YEAR_NOW, YEAR_NOW + 1, YEAR_NOW + 2]
if years_hist and years_hist[-1] >= YEAR_NOW:
    years_fc = [years_hist[-1] + 1, years_hist[-1] + 2, years_hist[-1] + 3]
years_all = years_hist + years_fc

def yv(records, year, field):
    # get_val() đã tự chia /1e9 (VND -> tỷ VND) bên trong fetch_data.py — KHÔNG chia lại lần nữa.
    v = get_val(records, year, field)
    return round(v or 0, 1)

# Mã trường xác nhận THẬT từ metrics field map Vietcap cho MWG (2026-07, đọc trực tiếp field map,
# không đoán) — isa9/isa10 (Chi phí bán hàng/QLDN) TÁCH RIÊNG cho MWG (khác HPG gộp 1 dòng), isa22
# (LNST CỔ ĐÔNG CÔNG TY MẸ) dùng cho EPS/định giá vì MWG có lợi ích cổ đông thiểu số ở 1 số công ty
# con (DMX/BHX...). bsa78 = Vốn chủ sở hữu (bsa53 là TỔNG TÀI SẢN, không phải vốn CSH — lỗi bản nháp
# đầu). bsa56/bsa71 = Vay ngắn/dài hạn. bsa57 = Phải trả người bán (bsa73 là dự phòng thôi việc, sai).
revenue_hist = [yv(is_hist, y, "isa3") for y in years_hist]
cogs_hist    = [abs(yv(is_hist, y, "isa4")) for y in years_hist]
gp_hist      = [revenue_hist[i] - cogs_hist[i] for i in range(len(years_hist))]
gpm_hist     = [round(gp_hist[i] / revenue_hist[i] * 100, 2) if revenue_hist[i] else 0 for i in range(len(years_hist))]
selling_hist = [abs(yv(is_hist, y, "isa9")) for y in years_hist]   # Chi phí bán hàng
admin_hist   = [abs(yv(is_hist, y, "isa10")) for y in years_hist]  # Chi phí quản lý doanh nghiệp
sga_hist     = [selling_hist[i] + admin_hist[i] for i in range(len(years_hist))]
ebit_hist    = [gp_hist[i] - sga_hist[i] for i in range(len(years_hist))]
ni_hist      = [yv(is_hist, y, "isa22") for y in years_hist]  # LNST của Cổ đông Công ty mẹ (loại trừ lợi ích CĐ thiểu số)
ni_consol_hist = [yv(is_hist, y, "isa20") for y in years_hist]  # LNST hợp nhất (gồm cả thiểu số) — tham khảo, khớp guidance IR
equity_hist  = [yv(bs_hist, y, "bsa78") for y in years_hist]
debt_hist    = [yv(bs_hist, y, "bsa56") + yv(bs_hist, y, "bsa71") for y in years_hist]  # Vay NH (bsa56) + Vay DH (bsa71)
cash_hist    = [yv(bs_hist, y, "bsa2") + yv(bs_hist, y, "bsa5") for y in years_hist]  # Tiền + Đầu tư ngắn hạn
inventory_hist = [yv(bs_hist, y, "bsa15") for y in years_hist]
receivables_hist = [yv(bs_hist, y, "bsa8") for y in years_hist]
payables_hist = [yv(bs_hist, y, "bsa57") for y in years_hist]  # Phải trả người bán (ngắn hạn)
total_assets_hist = [yv(bs_hist, y, "bsa53") for y in years_hist]  # Tổng tài sản (đúng field, KHÔNG dùng cho vốn CSH)
da_hist = [abs(yv(cf_hist, y, "cfa2")) for y in years_hist]  # Khấu hao TSCĐ và BĐSĐT
capex_hist = [abs(yv(cf_hist, y, "cfa19")) for y in years_hist]  # Tiền chi mua sắm/XD TSCĐ
ebitda_hist = [ebit_hist[i] + da_hist[i] for i in range(len(years_hist))]

# get_val() trả về tỷ VND (đã /1e9 sẵn). Vốn góp (VND) / 10.000 (mệnh giá) = số CP; quy đổi tỷ VND
# sang số CP triệu: tỷ_VND * 1e9 / 10000 / 1e6 = tỷ_VND * 0.1
shares_hist = [round((get_val(bs_hist, y, "bsa80") or 0) * 0.1) for y in years_hist]
SHARES = shares_hist[-1] if shares_hist and shares_hist[-1] else 1462  # fallback ~1.46 tỷ CP MWG
eps_hist = [round(ni_hist[i] * 1000 / shares_hist[i], 0) if shares_hist[i] else 0 for i in range(len(years_hist))]
bvps_hist = [round(equity_hist[i] * 1000 / shares_hist[i], 0) if shares_hist[i] else 0 for i in range(len(years_hist))]

PRICE = RAW.get("currentPrice") or 0
MARKET_CAP = round(PRICE * SHARES / 1000, 1)  # tỷ VND (PRICE VND, SHARES triệu CP)

print(f"  -> Revenue hist (ty): {revenue_hist}")
print(f"  -> NI hist (ty): {ni_hist}")
print(f"  -> Shares: {SHARES} trieu | Price: {PRICE} | MCap: {MARKET_CAP} ty")

# ── 2. DỮ LIỆU MẢNG & SỐ CỬA HÀNG (2026-07) ─────────────────────────────────
# Nguồn: (1) IR MWG — API https://api.mwgshop.vn/gw/ir-api/api/PostAPI?langID=vi-VN&categoryID=431
#   (categoryID 431 = "Báo cáo kết quả kinh doanh theo tháng", khám phá được từ JS bundle trang
#   mwg.vn/bao-cao — xem ghi chú cuối file); (2) "Cập nhật tình hình kinh doanh 5 tháng 2026" (PDF,
#   17/06/2026); (3) "Định hướng kinh doanh năm 2026" (PDF, 30/01/2026, có target FY2026 + tỷ trọng
#   DMX/BHX trong tổng LN); (4) web search (nguoiquansat.vn, vneconomy.vn) cho số liệu FY2024 và mốc
#   cửa hàng cuối T4/2025 — CHƯA verify qua BCTC kiểm toán, dùng làm assumption có thể chỉnh sửa.
#
# 4 mảng chính theo skill ban-le: TGDD (gồm Topzone), ĐMX (Điện Máy Xanh, gồm Thợ ĐMX/EraBlue hạch
# toán chung "Công ty CPĐT Điện Máy Xanh"), BHX (Bách Hóa Xanh), Khác (An Khang + AvaKids + EraBlue
# consolidation adj. + linh tinh — quy mô nhỏ, gộp chung để đơn giản hoá, KHÔNG dự phóng chi tiết
# riêng từng chuỗi nhỏ vì thiếu dữ liệu store-level đủ dài để tính hiệu quả/cửa hàng đáng tin cậy).

SEGMENTS = ["TGDD", "DMX", "BHX", "Khac"]

# Số cửa hàng CUỐI MỖI NĂM TÀI CHÍNH — nguồn: PDF IR "Tóm tắt tình hình kinh doanh cả năm 20XX"
# (công bố hàng năm ~cuối T1 năm sau), đọc trực tiếp, số liệu CHÍNH XÁC 31/12 mỗi năm.
STORE_COUNT_HIST = {
    2024: {"TGDD": 1021, "DMX": 2026, "BHX": 1770, "AnKhang": 326, "AvaKids": 62, "EraBlue": 87},   # 31/12/2024
    2025: {"TGDD": 1012, "DMX": 2008, "BHX": 2559, "AnKhang": 382, "AvaKids": 83, "EraBlue": 181},  # 31/12/2025 (PDF "cả năm 2025", 22/01/2026)
}
STORE_COUNT_NOW = {"TGDD": 1012, "DMX": 2004, "BHX": 3051, "AnKhang": 416, "AvaKids": 95, "EraBlue": 245}  # ~17/06/2026 ("5 tháng 2026")

# ── Mốc lịch sử theo quý/tháng đọc TRỰC TIẾP từ ảnh báo cáo IR do user cung cấp (2026-07) ─────────
# Các báo cáo IR dạng infographic (hộp màu, biểu đồ tròn/cột — phổ biến 2019-2023) dùng FONT NHÚNG
# BỊ LỖI bảng mã glyph->Unicode (xác nhận bằng cách tự tay test CẢ pdfplumber LẪN PyMuPDF trên cùng
# 1 file gốc — cả 2 đều chỉ trích được vài ký tự bullet (•❑➢), mất gần hết chữ có dấu và toàn bộ số
# liệu). Đây là giới hạn của chính file PDF gốc (khả năng cao xuất từ Canva/PowerPoint với font
# subset thiếu cmap đầy đủ cho tiếng Việt), KHÔNG phải lỗi thư viện đọc — không thư viện text-extract
# nào khắc phục được. Ngược lại báo cáo dạng VĂN BẢN THUẦN (bullet point, không hộp màu thiết kế -
# VD "11 tháng 2021") dùng font chuẩn, extract bằng pdfplumber bình thường, ĐÃ được tự động hoá qua
# parse_mwg_report(). Các mốc dưới đây do KHÔNG THỂ tự động, được đọc thủ công từ ảnh chụp báo cáo
# gốc user gửi trực tiếp — dùng làm điểm neo (anchor) để kiểm tra tính hợp lý của hiệu quả/cửa hàng
# theo thời gian, KHÔNG đủ dày đặc (không phải đủ 12 tháng/năm) để tính tăng trưởng MoM chi tiết.
QUARTERLY_SNAPSHOTS_MANUAL = {
    "2019-09": {  # "9 tháng đầu năm 2019" — dạng biểu đồ cột (không có text nhãn, đọc số trên cột)
        "stores": {"BHX": 788, "DMX": 907, "TGDD": 1918 - 907, "Total": 2706},
    },
    "2020-10": {  # "10 tháng đầu năm 2020" — bảng "DOANH THU THEO CHUỖI" (Số lượng CH/Cơ cấu/Tăng trưởng)
        "revenue_total": 90102.0, "revenue_yoy": 0.06, "ni": 3283.0, "ni_yoy": 0.01,
        "stores": {"TGDD": 948, "DMX": 1198, "BHX": 1656, "Total": 3802},
        "store_pct": {"TGDD": 0.272, "DMX": 0.538, "BHX": 0.190},
        "rev_yoy_by_seg": {"TGDD": -0.13, "DMX": 0.00, "BHX": 1.08},
    },
    "2021-06": {  # "6 tháng đầu năm 2021" — bảng CHUỖI/Số lượng CH/Tăng trưởng DT/Tăng trưởng DT CH cũ
        "revenue_total": 62487.0, "revenue_yoy": 0.12, "revenue_plan_pct": 0.50,  # kế hoạch FY21 125,000 tỷ
        "ni": 2552.0, "ni_yoy": 0.26, "ni_plan_pct": 0.54,  # kế hoạch FY21 4,750 tỷ
        "stores": {"TGDD": 936, "DMX": 1731, "BHX": 1888, "Total": 4610},
        "rev_yoy_by_seg": {"TGDD": 0.07, "DMX": 0.05, "BHX": 0.42},
    },
    "2021-11": {  # "11 tháng đầu năm 2021" — dạng văn bản thuần, tự parse được (đối chiếu thủ công)
        "revenue_total": 110530.0, "revenue_yoy": 0.11, "revenue_plan_pct": 0.88,
        "ni": 4395.0, "ni_yoy": 0.22, "ni_plan_pct": 0.93,
        "rev_tgdd_dmx_cum": 83800.0, "rev_tgdd_dmx_cum_yoy": 0.05,  # TGDD+DMX lũy kế 11T
        "stores": {"DMS": 674},  # Điện Máy Sáng (mô hình cửa hàng nhỏ ĐMX), DT lũy kế 5,850 tỷ
    },
    "2022-08": {  # "8 tháng đầu năm 2022" — infographic sidebar + narrative % cơ cấu rõ ràng
        "revenue_total": 92283.0, "revenue_yoy": 0.18, "revenue_plan_pct": 0.66,  # kế hoạch FY22 140,000 tỷ
        "ni": 3176.0, "ni_yoy": 0.06, "ni_plan_pct": 0.50,  # kế hoạch FY22 6,350 tỷ
        "revenue_by_seg": {"TGDD_DMX": 24500.0 + 48800.0, "BHX": 17600.0},  # TGDD(gồm Topzone)+ĐMX gộp 79.5%, BHX 19%
        "stores": {"TGDD": 1086, "DMX": 2222, "BHX": 1726, "AnKhang": 509, "AvaKids": 80, "AvaSport": 12},
        "stores_note": {"TGDD_Topzone": 54, "DMX_DMS": 1002},
    },
    "2023-06": {  # "6 tháng đầu năm 2023" — infographic donut + sidebar "Tổng X, bao gồm Y"
        "revenue_total": 56570.0, "revenue_yoy": -0.20, "revenue_plan_pct": 0.42,  # kế hoạch FY23 135,000 tỷ
        "store_pct": {"TGDD": 0.236, "DMX": 0.499, "BHX": 0.242, "Khac": 0.023},  # cơ cấu DT theo donut
        "stores": {"TGDD": 1180, "DMX": 2289, "BHX": 1706, "AnKhang": 537, "AvaKids": 66},
        "stores_note": {"TGDD_Topzone": 100, "DMX_DMS": 1034},
        "rev_tgdd_dmx_cum": 41500.0, "rev_tgdd_dmx_cum_yoy": -0.27,  # lũy kế 6T "2 chuỗi"
        "online_pct_of_tgdd_dmx": 0.19,
    },
    "2022-09": {  # "9 tháng đầu năm 2022" — sidebar song ngữ, chỉ có số cửa hàng (không có DT/donut)
        "stores": {"TGDD": 1116, "DMX": 2246, "BHX": 1727, "AnKhang": 529},
        "stores_note": {"TGDD_Topzone": 71, "DMX_DMS": 1019},
    },
    # Từ 2023 trở đi, template ổn định 1 dạng: sidebar liệt kê số cửa hàng (không phải cột) + donut cơ
    # cấu DT theo mảng — ngoại lệ mỗi kỳ khác nhau về text narrative đi kèm (giữ nguyên field để dùng).
    "2024-09": {  # "9 tháng đầu năm 2024"
        "revenue_total": 99767.0, "revenue_yoy": None, "revenue_prior": 86858.0, "revenue_plan_pct": 0.80,  # KH FY24 125,000 tỷ
        "store_pct": {"Khac": 0.027, "TGDD": 0.217, "DMX": 0.452, "BHX": 0.304},
        "rev_tgdd_dmx_cum": 66700.0, "rev_tgdd_dmx_cum_yoy": 0.07,
        "stores": {"TGDD": 1023, "DMX": 2030, "BHX": 1726, "AnKhang": 326, "AvaKids": 62, "EraBlue": 76},
    },
    "2025-08": {  # "8 tháng đầu năm 2025"
        "revenue_total": 99801.0, "revenue_prior": 87967.0, "revenue_yoy": 0.135, "revenue_plan_pct": 0.67,  # KH FY25 150,000 tỷ
        "store_pct": {"Khac": 0.023, "TGDD": 0.226, "DMX": 0.446, "BHX": 0.305},
        "rev_tgdd_dmx_cum": 67000.0, "rev_tgdd_dmx_cum_yoy": 0.14,
        "stores": {"TGDD": 1014, "DMX": 2023, "BHX": 2233, "AnKhang": 326, "AvaKids": 62, "EraBlue": 135},
    },
    "2025-10": {  # "10 tháng đầu năm 2025"
        "revenue_total": 128289.0, "revenue_prior": 111353.0, "revenue_yoy": 0.15, "revenue_plan_pct": 0.86,  # KH FY25 150,000 tỷ
        "store_pct": {"Khac": 0.023, "TGDD": 0.237, "DMX": 0.44, "BHX": 0.30},
        "rev_tgdd_dmx_cum": 87000.0, "rev_tgdd_dmx_cum_yoy": 0.17, "sssg": 0.19,  # tăng trưởng cửa hàng hiện hữu
        "stores": {"TGDD": 1012, "DMX": 2017, "BHX": 2370, "AnKhang": 356, "AvaKids": 72, "EraBlue": 158},
    },
    "2026-04": {  # "4 tháng đầu năm 2026"
        "revenue_total": 62496.0, "revenue_prior": 48635.0, "revenue_yoy": 0.285, "revenue_plan_pct": 0.34,  # KH FY26 185,000 tỷ
        "store_pct": {"Khac": 0.02, "TGDD": 0.229, "DMX": 0.464, "BHX": 0.287},
        "sssg": 0.33,  # SSSG (tăng trưởng cửa hàng hiện hữu) toàn bộ chuỗi VN
        "stores": {"TGDD": 1012, "DMX": 2005, "BHX": 2962, "AnKhang": 423, "AvaKids": 95, "EraBlue": 222},
    },
    "2026-05": {  # "5 tháng đầu năm 2026" — khớp REV_5T_2026_TOTAL/PCT đã có sẵn, dùng để đối chiếu
        "revenue_total": 79159.0, "revenue_prior": 61229.0, "revenue_yoy": 0.293, "revenue_plan_pct": 0.43,  # KH FY26 185,000 tỷ
        "store_pct": {"Khac": 0.02, "TGDD": 0.225, "DMX": 0.463, "BHX": 0.292},
        "sssg": 0.33,
        "stores": {"TGDD": 1012, "DMX": 2004, "BHX": 3051, "AnKhang": 416, "AvaKids": 95, "EraBlue": 245},
    },
}

# Doanh thu theo mảng (tỷ VND) — nguồn: PDF IR "Tóm tắt tình hình kinh doanh cả năm 20XX" (công bố
# hàng năm, đọc trực tiếp — CHÍNH XÁC cho 2023/2024/2025, không phải ước tính):
#   FY2025 (PDF "cả năm 2025", 22/01/2026): TGDD 37,300 | ĐMX 68,400 (TGDD+ĐMX chung +18% YoY) | BHX
#   46,900 (+14% YoY) | An Khang 2,200 (giảm nhẹ) | AvaKids 1,400 (+16% YoY) | EraBlue +70% YoY (IDR,
#   181 cửa hàng). Cơ cấu DT theo chuỗi (biểu đồ tròn): TGDD 23.9%/ĐMX 43.8%/BHX 30%/Khác 2.3% của
#   tổng 156,166 tỷ (IR tự báo cáo, hơi khác Vietcap 155,928.1 — chênh lệch nhỏ do cách hạch toán/làm
#   tròn, ưu tiên số IR cho phân bổ mảng vì đây là nguồn trực tiếp công ty công bố theo mảng).
#   FY2024 (PDF "cả năm 2024", 21/01/2025): TGDD 30,000 (+7% YoY) | ĐMX 59,500 (+7% YoY) | BHX >41,000
#   (+30% YoY) | An Khang 2,300 (+3% YoY) | AvaKids 1,200 (+35% YoY) | EraBlue tăng >4 lần (nền rất nhỏ
#   2023). Cơ cấu DT: TGDD 22.4%/ĐMX 44.3%/BHX 30.6%/Khác 2.7% của tổng 134,341 tỷ (khớp Vietcap).
# FY2023 SUY NGƯỢC từ %YoY nêu trong PDF FY2024 (TGDD=30000/1.07, ĐMX=59500/1.07, BHX=41000/1.30,
# AnKhang=2300/1.03, AvaKids=1200/1.35) + tổng DT 2023 thật Vietcap (118,280) — sai số rất nhỏ vì dùng
# đúng %YoY công ty công bố, KHÔNG phải ước tính tự do.
REV_SEGMENT_HIST = {
    2023: {"TGDD": 28037, "DMX": 55607, "BHX": 31538, "Khac": 3098},          # suy ngược từ %YoY FY2024 IR + tổng Vietcap 118,280
    2024: {"TGDD": 30000, "DMX": 59500, "BHX": 41000, "Khac": 3841.2},        # công bố trực tiếp IR; Khac = dư so tổng Vietcap 134,341.2
    2025: {"TGDD": 37300, "DMX": 68400, "BHX": 46900, "Khac": 3566},          # công bố trực tiếp IR "cả năm 2025"; Khac = dư so tổng IR 156,166
}
# 5 tháng đầu 2026 (thật, PDF IR 17/06/2026): tổng 79,159 tỷ (+29.3% YoY), cơ cấu TGDD 22.5%/DMX
# 46.3%/BHX 29.2%/Khác 2% — dùng để neo tỷ trọng mảng cho dự phóng năm hiện tại.
REV_5T_2026_TOTAL = 79159.0
REV_5T_2026_PCT = {"TGDD": 0.225, "DMX": 0.463, "BHX": 0.292, "Khac": 0.02}
REV_5T_2025_TOTAL = 61229.0  # cùng kỳ 5 tháng 2025, để tính growth 5T thật (+29.3%)

# Kế hoạch/target chính thức FY2026 (PDF "Định hướng kinh doanh năm 2026", 30/01/2026):
FY2026_TARGET_REVENUE = 185000.0   # tỷ VND, +18% svck 2025
FY2026_TARGET_NPAT = 9200.0        # tỷ VND hợp nhất, +30% svck 2025
FY2026_DMX_SHARE_REV = 0.65        # DMX ~65% tổng DT 2026
FY2026_DMX_SHARE_PROFIT = 0.80     # DMX >80% tổng LN 2026
FY2026_BHX_SHARE_REV = 0.30        # BHX ~30% tổng DT 2026
FY2026_BHX_SHARE_PROFIT = 0.20     # BHX ~20% tổng LN 2026
FY2026_BHX_NEW_STORES = 1000       # kế hoạch mở mới BHX 2026
FY2026_ERABLUE_TARGET_STORES = 300 # EraBlue mục tiêu >300 cửa hàng 2026

# ── Fetch tự động báo cáo "Cập nhật tình hình kinh doanh theo tháng" mới nhất từ IR MWG (2026-07) ──
# Giao thức API IR (reverse-engineer từ JS bundle mwg.vn/bao-cao, KHÔNG phải API công khai chính
# thức — có thể đổi/gãy bất cứ lúc nào nếu MWG đổi giao diện, giống rủi ro TradingView đã ghi nhận
# ở build_hpg_model.py):
#   1. CategoryAPI?langID=vi-VN&parentID=7            -> danh sách NĂM, mỗi năm 1 categoryID riêng
#      (2026=429, 2025=205, 2024=193, 2023=185, 2022=106, 2021=98, 2020=5, 2019=6, 2018=28...)
#   2. CategoryAPI?langID=vi-VN&parentID=<năm's categoryID> -> 2 sub-category: "Báo cáo kết quả kinh
#      doanh theo tháng" (mục tiêu) và "Báo cáo tài chính, quản trị, thường niên"
#   3. PostAPI?langID=vi-VN&categoryID=<sub-category ID>&pageSize=30&condition=-1 -> danh sách bài,
#      mỗi bài có "title", "publishDate", "fileLink" (URL PDF trực tiếp, có thể tải ngay)
# Base URL: https://api.mwgshop.vn/gw/ir-api/api/{CategoryAPI|PostAPI}. Chỉ tự động hoá SỐ CỬA HÀNG
# (nhãn tên chuỗi bám sát số liệu trong PDF, regex đáng tin cậy) — KHÔNG tự động parse tỷ trọng
# doanh thu từng mảng (nằm trong biểu đồ tròn, thứ tự text bị xáo trộn khi extract, không đủ tin cậy
# cho số liệu tài chính) — phần đó vẫn cập nhật thủ công vào REV_SEGMENT_HIST/REV_5T_2026_PCT ở trên
# mỗi khi có báo cáo mới, xem ghi chú nguồn kèm theo.
IR_API_BASE = "https://api.mwgshop.vn/gw/ir-api/api"
STORE_CACHE_FILE = os.path.join(".cache", "MWG_store_count.json")

def _ir_api_get(path, params, timeout=15):
    r = requests.get(f"{IR_API_BASE}/{path}", params=params,
                      headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                      timeout=timeout)
    r.raise_for_status()
    d = r.json()
    return d.get("data") or []

# "fileLink" trả về từ PostAPI có 3 dạng khác nhau tuỳ năm (2026-07, phát hiện khi chạy full history
# 2019-2023 trả về rỗng toàn bộ — do gọi requests.get() thẳng trên "fileLink" chưa resolve URL):
#   1. URL tuyệt đối mới: "https://cdnv2-tmdt.tgdd.vn/..." hoặc "https://cdnv2.tgdd.vn/..." (2024-2026)
#      -> dùng thẳng, không cần biến đổi.
#   2. URL tuyệt đối CŨ: "https://mwg.vn/wp-content/uploads/..." (2019, 2021) -> site cũ đã ngừng, phải
#      thay prefix bằng CDN mới (logic lấy từ JS bundle mwg.vn/bao-cao, hàm resolve `r()`/`l()`).
#   3. Đường dẫn TƯƠNG ĐỐI không có scheme: "2022/7/bao-cao-....pdf" (2022-2023) -> phải nối thêm
#      base CDN mới ở đầu.
_OLD_URL_PREFIXES = ["http://mwg.vn/wp-content/uploads", "https://mwg.vn/wp-content/uploads",
                      "http://www.mwg.vn/wp-content/uploads", "https://www.mwg.vn/wp-content/uploads",
                      "http://mwg.vn/uploads", "https://mwg.vn/uploads"]
_NEW_CDN_BASE = "https://cdnv2-tmdt.tgdd.vn/mwgvn/ir"

def _resolve_mwg_file_url(file_link):
    if not file_link:
        return file_link
    for prefix in _OLD_URL_PREFIXES:
        if file_link.startswith(prefix):
            return _NEW_CDN_BASE + file_link[len(prefix):]
    if re.match(r"^https?://", file_link, re.IGNORECASE):
        return file_link
    return _NEW_CDN_BASE + ("" if file_link.startswith("/") else "/") + file_link

def _download_pdf_text(pdf_url, max_pages=3, timeout=25):
    """Tải PDF về tạm và extract text (pdfplumber) của tối đa `max_pages` trang đầu, nối bằng \\n.
    Trả về "" nếu lỗi bất kỳ bước nào (KHÔNG BAO GIỜ raise)."""
    if pdfplumber is None:
        return ""
    pdf_url = _resolve_mwg_file_url(pdf_url)
    fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
    os.close(fd)
    try:
        r = requests.get(pdf_url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        with open(tmp_path, "wb") as f:
            f.write(r.content)
        with pdfplumber.open(tmp_path) as pdf:
            return "\n".join((p.extract_text() or "") for p in pdf.pages[:max_pages])
    except Exception:
        return ""
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

# ── Parser đa-era cho số cửa hàng & doanh thu mảng (2026-07) ────────────────────────────────────
# Template IR MWG đổi 3 lần qua các năm — mỗi era cần pattern khác nhau (đã verify bằng cách tự tay
# đọc text THẬT do pdfplumber trích ra, không đoán mò):
#   Era C (2024-2026): sidebar "TÊN CHUỖI\n...\nX cửa hàng" — nhãn và số liệu cách nhau khá xa
#     (do layout 2 cột bị làm phẳng khi extract), cần max_gap lớn (~350 ký tự).
#   Era B (2022-2023): sidebar "TÊN CHUỖI (X cửa hàng)" — số liệu ngay TRONG ngoặc theo sau nhãn.
#     pdfplumber có lúc bỏ dấu cách giữa các từ ("cửahàng" dính liền) — regex phải cho \s* tuỳ chọn
#     ở MỌI vị trí, không chỉ 1-2 chỗ.
#   Era A (2019-2021): KHÔNG có sidebar text (chỉ logo ảnh, không extract được) — số cửa hàng nằm
#     trong đoạn văn tường thuật ở trang 2 ("chuỗi ĐMX có N cửa hàng", "nâng tổng số cửa hàng BHX lên
#     N"). Doanh thu theo chuỗi ("DOANH THU THEO CHUỖI" box) chỉ có 3 SỐ đứng một mình theo ĐÚNG THỨ
#     TỰ cố định TGDD/DMX/BHX (không có nhãn text cạnh vì nhãn là logo ảnh) — phải dựa vào thứ tự,
#     không phải nhãn, nên kém tin cậy hơn 2 era kia (đã ghi rõ trong kết quả trả về bằng field
#     "confidence").
def _find_near(label_pat, unit_pat, text, max_gap=350):
    # Số liệu TỔNG (sidebar) luôn được theo sau bởi khoảng trắng/xuống dòng/dấu ngoặc — KHÔNG bao
    # giờ dính liền chữ thường ngay sau đơn vị. Ngược lại, số liệu PHỤ nằm giữa câu văn tường thuật
    # (VD "64cửahàngmởmới" — "mở mới" dính ngay sau "cửahàng" do pdfplumber làm mất khoảng cách) LUÔN
    # có chữ thường nối tiếp ngay sau đơn vị. Dùng negative lookahead `(?![a-zà-ỹ])` để loại các số
    # phụ này mà KHÔNG cần giả định số tổng phải đứng đầu dòng riêng (layout 2 cột khiến vị trí dòng
    # không nhất quán giữa các chuỗi — TGDD/BHX/EraBlue đứng đầu dòng nhưng DMX lại nằm giữa dòng).
    m = re.search(label_pat + r".{0,%d}?([\d][\d.]*)\s*%s(?![a-zà-ỹ])" % (max_gap, unit_pat),
                  text, re.IGNORECASE | re.DOTALL)
    return int(m.group(1).replace(".", "")) if m else None

def _find_in_parens(label_pat, unit_pat, text):
    m = re.search(label_pat + r"\s*\(\s*([\d][\d.]*)\s*%s\s*\)" % unit_pat, text, re.IGNORECASE)
    return int(m.group(1).replace(".", "")) if m else None

def _find_narrative(patterns, text):
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE | re.DOTALL)
        if m:
            return int(m.group(1).replace(".", ""))
    return None

def parse_mwg_store_counts(text):
    """Trích số cửa hàng từng chuỗi từ text 1 báo cáo, thử lần lượt Era C -> Era B -> Era A.
    Trả về dict (giá trị None nếu không trích được field đó)."""
    result = {}
    # "THẾ GIỚI DI ĐỘNG" cũng là 1 phần TÊN CÔNG TY MWG ("...Đầu Tư Thế Giới Di Động (MWG)") lặp lại
    # ở đầu MỌI trang trong báo cáo Era A (2019-2021, dạng slide nhiều trang) — loại trừ bằng negative
    # lookahead "(?!\s*\(MWG\))" để không khớp nhầm dòng tiêu đề công ty. Era A hiếm khi công bố số
    # cửa hàng TGDĐ TÁCH RIÊNG (chỉ có tổng MWG + ĐMX + BHX) — không đoán, để None nếu không tìm thấy
    # thay vì gán nhầm tổng công ty cho TGDD.
    result["TGDD"] = (_find_near(r"THẾ GIỚI DI ĐỘNG(?!\s*\(MWG\))", r"cửa\s*hàng", text)
                       or _find_in_parens(r"THẾ GIỚI DI ĐỘNG(?!\s*\(MWG\))", r"cửa\s*hàng", text))
    result["DMX"] = (_find_near(r"ĐIỆN MÁY XANH", r"cửa\s*hàng", text)
                      or _find_in_parens(r"ĐIỆN MÁY XANH", r"cửa\s*hàng", text)
                      or _find_narrative([r"chuỗi\s*ĐMX\s*có\s*([\d][\d.]*)\s*cửa\s*hàng",
                                          r"ĐMX\s*có\s*([\d][\d.]*)\s*cửa\s*hàng"], text))
    result["BHX"] = (_find_near(r"BÁCH HÓA XANH", r"cửa\s*hàng", text)
                      or _find_in_parens(r"BÁCH HÓA XANH", r"cửa\s*hàng", text)
                      or _find_narrative([r"tổng\s*số\s*cửa\s*hàng\s*BHX\s*lên\s*([\d][\d.]*)",
                                          r"chuỗi\s*BHX\s*có\s*([\d][\d.]*)\s*cửa\s*hàng"], text))
    result["AnKhang"] = (_find_near(r"AN KHANG", r"nhà\s*thuốc", text)
                          or _find_in_parens(r"AN KHANG", r"nhà\s*thuốc", text))
    result["AvaKids"] = (_find_near(r"AVA\s*KIDs?", r"cửa\s*hàng", text)
                          or _find_narrative([r"([\d]+)\s*cửa\s*hàng\s*AVAKids"], text))
    result["EraBlue"] = _find_near(r"ERABLUE\s*\(liên doanh", r"cửa\s*hàng", text)
    return result

def parse_mwg_segment_revenue(text, total_revenue=None):
    """Trích doanh thu từng mảng (tỷ VND) từ text 1 báo cáo. Era B/C dùng % cơ cấu doanh thu × tổng
    (nếu có total_revenue) hoặc số ngàn/nghìn tỷ trực tiếp trong câu tường thuật. Era A dựa vào THỨ
    TỰ 3 số trong box "DOANH THU THEO CHUỖI" (kém tin cậy hơn — không có nhãn neo)."""
    result = {}
    # Era B/C: "doanh thu của X (là|đạt|xấp xỉ) Y (ngàn|nghìn) tỷ đồng"
    def _rev_narrative(name_pat):
        m = re.search(name_pat + r".{0,60}?(?:là|đạt|xấp\s*xỉ)\s*([\d]+[,.]?[\d]*)\s*(ngàn|nghìn)\s*tỷ",
                       text, re.IGNORECASE | re.DOTALL)
        if m:
            return float(m.group(1).replace(",", ".")) * 1000
        return None
    def _rev_direct(name_pat):
        # "X mang về Y ngàn tỷ đồng"
        m = re.search(name_pat + r".{0,30}?mang\s*về\s*([\d]+[,.]?[\d]*)\s*(ngàn|nghìn)\s*tỷ",
                       text, re.IGNORECASE | re.DOTALL)
        if m:
            return float(m.group(1).replace(",", ".")) * 1000
        return None
    result["TGDD"] = _rev_narrative(r"TGDĐ") or _rev_direct(r"TGDĐ")
    result["DMX"] = _rev_narrative(r"ĐMX") or _rev_direct(r"ĐMX")
    result["BHX"] = _rev_narrative(r"BHX") or _rev_direct(r"BHX")
    return result

def parse_mwg_total_revenue_cum(text, target_year=None):
    """Trích DOANH THU THUẦN LŨY KẾ TOÀN CÔNG TY (tỷ VND) từ header biểu đồ cột "DOANH THU THUẦN...
    (tỷđồng)" — LUÔN có 2 số (kỳ hiện tại + kỳ trước) kèm 2 nhãn kỳ (VD "5T25 5T26" hoặc "6T2020
    6T2021"). QUAN TRỌNG (bug phát hiện 2026-07): THỨ TỰ số kỳ hiện tại/kỳ trước KHÔNG NHẤT QUÁN giữa
    các báo cáo — có báo cáo hiện tại đứng trước (VD 6T2021: "62.487" rồi mới "55.639"), có báo cáo
    hiện tại đứng SAU (VD 5T2026: "61.229 79.159" — 61.229 là 5T25 CŨ, 79.159 mới là 5T26 THẬT). Nếu
    chỉ lấy số ĐẦU TIÊN sẽ có lúc lấy NHẦM số NĂM TRƯỚC — gây sai lệch nghiêm trọng cho toàn bộ chuỗi
    doanh thu tháng năm đó. Cách khắc phục: nếu biết target_year (năm báo cáo thật, lấy từ post["year"]
    context), tìm CẶP nhãn kỳ (VD "5T25 5T26") đứng sau 2 số, đối chiếu năm nào khớp target_year rồi
    chọn ĐÚNG số tương ứng theo vị trí — không đoán mò theo thứ tự xuất hiện."""
    if target_year is not None:
        m = re.search(r"(\d{2,3}\.\d{3})\s+(\d{2,3}\.\d{3}).{0,400}?(\d{1,2})T(\d{2,4})\s+(\d{1,2})T(\d{2,4})",
                      text, re.DOTALL)
        if m:
            num1, num2 = m.group(1), m.group(2)
            def _norm_year(y):
                y = int(y)
                return y if y > 100 else (2000 + y)
            y1, y2 = _norm_year(m.group(4)), _norm_year(m.group(6))
            if y1 == target_year and y2 != target_year:
                return float(num1.replace(".", ""))
            if y2 == target_year and y1 != target_year:
                return float(num2.replace(".", ""))
    # Fallback (không xác định được target_year hoặc không tìm thấy cặp nhãn rõ ràng): lấy số ĐẦU
    # TIÊN sau header — CHỈ dùng khi không có cách nào tốt hơn, có thể sai như bug đã nêu ở trên.
    m = re.search(r"DOANH\s*THU\s*THUẦN.{0,200}?([\d]{2,3}\.[\d]{3})", text, re.IGNORECASE | re.DOTALL)
    if m:
        return float(m.group(1).replace(".", ""))
    m = re.search(r"doanh\s*thu\s*thuần(?:\s*hợp\s*nhất)?\s*(?:là|đạt)\s*(?:hơn\s*)?([\d]+[.,]?[\d]*)\s*tỷ",
                  text, re.IGNORECASE | re.DOTALL)
    return float(m.group(1).replace(".", "").replace(",", ".")) if m else None

# ── Era D (2020 - giữa 2022, phát hiện 2026-07 sau khi đo "sức khoẻ" text-extraction toàn bộ 84 báo
# cáo 2019-2026): bảng 1 dòng/chuỗi "TÊN 948 27,2% -13%" = Số lượng CH | Cơ cấu DT | Tăng trưởng DT
# YoY — nằm NGUYÊN VẸN trên 1 dòng khi pdfplumber extract (không bị xáo trộn như box donut chart),
# ĐÁNG TIN CẬY NHẤT trong mọi era vì cho cả 3 chỉ số cùng lúc, ưu tiên thử pattern này TRƯỚC.
# Ranh giới font-lỗi (đo bằng scratchpad_diag.py, đếm từ tiếng Việt có dấu trích được): 2019 → 6/2022
# VÀ 6/2024 → nay đều đọc text bình thường; CHỈ 7/2022 → 5/2024 bị lỗi font (không phải "2019-2023
# infographic era" như giả định ban đầu) — khoảng broken font hẹp hơn nhiều so với ước tính lúc đầu.
_TABLE_ROW_NAME_MAP = {"TGDĐ": "TGDD", "ĐMX": "DMX", "BHX": "BHX", "BLUETRONICS": "Bluetronics"}
_TABLE_ROW_RE = re.compile(
    r"(TGDĐ|ĐMX|BHX|Bluetronics)\s+([\d][\d.]*)\s+([\d]+,[\d]+)%\s+([+-]?[\d]+(?:,[\d]+)?)%",
    re.IGNORECASE)

def parse_mwg_table_row(text):
    """Trích bảng "CHUỖI | Số lượng CH | Cơ cấu DT | Tăng trưởng DT" (Era D). Trả về dict
    {chuỗi: {"stores":, "rev_pct":, "rev_yoy":}} — rev_pct/rev_yoy dạng thập phân (0.272 = 27,2%)."""
    result = {}
    for m in _TABLE_ROW_RE.finditer(text):
        name, stores, pct, yoy = m.groups()
        key = _TABLE_ROW_NAME_MAP.get(name.upper())
        if not key:
            continue
        result[key] = {
            "stores": int(stores.replace(".", "")),
            "rev_pct": float(pct.replace(",", ".")) / 100,
            "rev_yoy": float(yoy.replace(",", ".")) / 100,
        }
    return result

def parse_mwg_report(pdf_url, era_hint=None, target_year=None):
    """Tải + parse 1 báo cáo tháng MWG, trả về dict {"stores":, "revenue":, "table_row":,
    "revenue_total_cum":} — field nào không trích được để None/rỗng. KHÔNG BAO GIỜ raise.
    target_year: năm THẬT của báo cáo (post["year"]) — dùng để chọn đúng số lũy kế khi báo cáo có 2
    số (kỳ hiện tại + kỳ trước) mà thứ tự xuất hiện không cố định (xem parse_mwg_total_revenue_cum)."""
    text = _download_pdf_text(pdf_url, max_pages=3)
    if not text:
        return {"stores": {}, "revenue": {}, "table_row": {}, "revenue_total_cum": None}
    table_row = parse_mwg_table_row(text)
    stores = parse_mwg_store_counts(text)
    for k, v in table_row.items():
        if stores.get(k) is None:
            stores[k] = v["stores"]
    return {"stores": stores, "revenue": parse_mwg_segment_revenue(text), "table_row": table_row,
            "revenue_total_cum": parse_mwg_total_revenue_cum(text, target_year=target_year)}

def fetch_mwg_latest_store_counts():
    """Tự động fetch báo cáo tháng mới nhất từ IR MWG, trích số cửa hàng từng chuỗi bằng regex.
    Trả về (dict cửa hàng, tên bài, ngày công bố) hoặc None nếu lỗi bất kỳ bước nào — KHÔNG BAO GIỜ
    raise, để build_mwg_model.py luôn fallback về STORE_COUNT_NOW (số cứng) khi không fetch được."""
    try:
        years = _ir_api_get("CategoryAPI", {"langID": "vi-VN", "parentID": 7, "pageIndex": 1, "pageSize": 5})
        if not years:
            return None
        latest_year = max(years, key=lambda y: int(y.get("categoryName", 0) or 0))
        subs = _ir_api_get("CategoryAPI", {"langID": "vi-VN", "parentID": latest_year["categoryID"], "pageIndex": 1, "pageSize": 10})
        monthly_cat = next((s for s in subs if "kết quả kinh doanh theo tháng" in (s.get("categoryName") or "").lower()), None)
        if not monthly_cat:
            return None
        posts = _ir_api_get("PostAPI", {"langID": "vi-VN", "categoryID": monthly_cat["categoryID"], "pageIndex": 1, "pageSize": 30, "condition": -1})
        candidates = [p for p in posts if p.get("fileLink") and "cập nhật tình hình kinh doanh" in (p.get("title") or "").lower()]
        if not candidates:
            return None
        latest_post = max(candidates, key=lambda p: p.get("publishDate", ""))
        text = _download_pdf_text(latest_post["fileLink"], max_pages=1)
        if not text:
            print("  [WARN] Khong tai/parse duoc PDF moi nhat - dung so cung")
            return None
        result = parse_mwg_store_counts(text)
        if not any(result.values()):
            print("  [WARN] Khong regex duoc so cua hang nao tu PDF moi nhat - dung so cung")
            return None
        return result, latest_post.get("title"), latest_post.get("publishDate")
    except Exception as e:
        print(f"  [WARN] Fetch MWG IR store count that bai: {e}")
        return None

# ── Lấy TOÀN BỘ lịch sử báo cáo tháng 2019-nay (2026-07, theo yêu cầu user) ─────────────────────
# Duyệt CategoryAPI(parentID=7) lấy mọi năm -> CategoryAPI(parentID=<năm>) lấy sub-category "Báo cáo
# kết quả kinh doanh theo tháng" -> PostAPI lấy danh sách bài -> tải + parse từng PDF (đa-era). Kết
# quả cache vào .cache/MWG_monthly_history.json để KHÔNG phải tải lại toàn bộ ~90 file mỗi lần chạy
# script — chỉ tải bài MỚI (chưa có trong cache) mỗi lần, dựa vào postID.
MONTHLY_HISTORY_CACHE_FILE = os.path.join(".cache", "MWG_monthly_history.json")

def fetch_mwg_full_history(start_year=2019, max_years=8):
    """Trả về dict {postID: {"year":, "title":, "publishDate":, "stores":{...}, "revenue":{...}}}
    cho MỌI báo cáo tháng tìm được từ start_year tới năm hiện tại. Dùng cache trên đĩa — chỉ tải/parse
    PDF nào CHƯA có trong cache (theo postID). KHÔNG BAO GIỜ raise; lỗi ở bất kỳ năm/bài nào chỉ in
    warning và bỏ qua, không dừng cả vòng lặp."""
    cache = {}
    if os.path.exists(MONTHLY_HISTORY_CACHE_FILE):
        try:
            with open(MONTHLY_HISTORY_CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f).get("posts", {})
        except (OSError, json.JSONDecodeError):
            cache = {}

    try:
        years = _ir_api_get("CategoryAPI", {"langID": "vi-VN", "parentID": 7, "pageIndex": 1, "pageSize": max_years + 2})
    except Exception as e:
        print(f"  [WARN] Full history: khong lay duoc danh sach nam: {e}")
        years = []

    n_new, n_failed = 0, 0
    for y in years:
        try:
            yr = int(y.get("categoryName", 0) or 0)
        except (TypeError, ValueError):
            continue
        if yr < start_year:
            continue
        try:
            subs = _ir_api_get("CategoryAPI", {"langID": "vi-VN", "parentID": y["categoryID"], "pageIndex": 1, "pageSize": 10})
            monthly_cat = next((s for s in subs if "kết quả kinh doanh theo tháng" in (s.get("categoryName") or "").lower()), None)
            if not monthly_cat:
                continue
            posts = _ir_api_get("PostAPI", {"langID": "vi-VN", "categoryID": monthly_cat["categoryID"], "pageIndex": 1, "pageSize": 30, "condition": -1})
        except Exception as e:
            print(f"  [WARN] Full history {yr}: loi lay danh sach bai - {e}")
            continue
        for p in posts:
            post_id = str(p.get("postID"))
            title_l = (p.get("title") or "").lower()
            is_monthly_report = ("cập nhật tình hình kinh doanh" in title_l or "báo cáo tóm tắt kqkd" in title_l
                                  or "báo cáo kqkd cả năm" in title_l or "tổng quan kqkd" in title_l)
            if not p.get("fileLink") or not is_monthly_report:
                continue
            if post_id in cache:
                continue  # đã có trong cache từ lần chạy trước, khỏi tải lại
            try:
                parsed = parse_mwg_report(p["fileLink"], target_year=yr)
                cache[post_id] = {
                    "year": yr, "title": p.get("title"), "publishDate": p.get("publishDate"),
                    "stores": parsed["stores"], "revenue": parsed["revenue"],
                    "table_row": parsed.get("table_row", {}),
                    "revenue_total_cum": parsed.get("revenue_total_cum"),
                }
                n_new += 1
            except Exception as e:
                n_failed += 1
                print(f"  [WARN] Loi parse bai '{p.get('title')}': {e}")

    try:
        os.makedirs(".cache", exist_ok=True)
        with open(MONTHLY_HISTORY_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({"posts": cache, "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M")}, f, ensure_ascii=False, indent=2)
    except OSError as e:
        print(f"  [WARN] Khong ghi duoc cache lich su: {e}")

    print(f"  -> Full history: {len(cache)} bai trong cache ({n_new} moi tai lan nay, {n_failed} loi parse)")
    return cache

def _parse_report_month_count(title):
    """Số tháng LŨY KẾ mà 1 báo cáo thể hiện (VD "9 tháng đầu năm" -> 9, "cả năm" -> 12, "tháng 1
    đầu năm" -> 1). None nếu không nhận dạng được (VD báo cáo họp NĐT quý, không phải KQKD tháng)."""
    t = (title or "").lower()
    if "cả năm" in t:
        return 12
    if re.search(r"tháng\s*1\s*\+\s*2", t):  # "mùa Tết (tháng 1+2)" — báo cáo gộp 2 tháng đầu năm
        return 2
    m = re.search(r"(\d{1,2})\s*tháng\s*(?:đầu\s*năm|20\d\d)", t)
    if m:
        return int(m.group(1))
    m = re.search(r"tháng\s*(\d{1,2})\s*(?:đầu\s*năm|năm)", t)
    if m:
        return int(m.group(1))
    return None

def derive_monthly_from_cumulative(history_cache):
    """Chuyển cache báo cáo LŨY KẾ 'X tháng đầu năm' thành chuỗi THÁNG ĐƠN LẺ: tháng N = lũy kế(N) -
    lũy kế(N-1) (doanh thu), số cửa hàng lấy nguyên snapshot cuối kỳ báo cáo tháng N (không cộng dồn).
    Riêng khi KHÔNG có báo cáo lũy kế 1 tháng nhưng CÓ báo cáo lũy kế 2 tháng: chia đôi doanh thu lũy
    kế 2 tháng cho tháng 1 và tháng 2 (giả định đều nhau — không có cách tách chính xác hơn khi thiếu
    dữ liệu riêng tháng 1); số cửa hàng lấy NGUYÊN số cuối kỳ báo cáo 2 tháng cho CẢ tháng 1 lẫn tháng
    2 (theo yêu cầu user 2026-07-04 — store count là snapshot 1 thời điểm, không suy ngược được).
    Trả về dict {(year, month): {"revenue": tỷ VND tháng đó, "stores": {...}, "is_estimated_split": bool}}."""
    by_year = {}
    for p in history_cache.values():
        n = _parse_report_month_count(p.get("title"))
        if n is None:
            continue
        by_year.setdefault(p["year"], {})[n] = p

    monthly = {}
    for yr, reports in by_year.items():
        months_present = sorted(reports.keys())
        prev_n, prev_cum_rev = 0, None
        for n in months_present:
            rep = reports[n]
            cum_rev = rep.get("revenue_total_cum")
            stores = rep.get("stores") or {}
            # rev_pct lấy TRỰC TIẾP từ table_row của CHÍNH báo cáo tháng N (cơ cấu DT lũy kế tới thời
            # điểm đó) — dùng làm xấp xỉ cơ cấu DT CỦA THÁNG ĐÓ (chấp nhận được vì tỷ trọng mảng đổi
            # rất chậm theo tháng, không phải xấp xỉ hoàn hảo nhưng là nguồn tốt nhất hiện có).
            rev_pct = {k: v.get("rev_pct") for k, v in (rep.get("table_row") or {}).items()}
            if n == 1:
                if cum_rev is not None:
                    monthly[(yr, 1)] = {"revenue": cum_rev, "stores": stores, "rev_pct": rev_pct, "is_estimated_split": False}
                prev_n, prev_cum_rev = 1, cum_rev
            elif n == prev_n + 1:
                if cum_rev is not None and prev_cum_rev is not None:
                    monthly[(yr, n)] = {"revenue": cum_rev - prev_cum_rev, "stores": stores, "rev_pct": rev_pct, "is_estimated_split": False}
                prev_n, prev_cum_rev = n, (cum_rev if cum_rev is not None else prev_cum_rev)
            elif n == 2 and prev_n == 0:
                # Khong co bao cao thang 1 rieng, chi co luy ke 2 thang -> chia doi doanh thu, dung
                # chung so cua hang cuoi T2 cho ca thang 1 va thang 2.
                if cum_rev is not None:
                    half = cum_rev / 2
                    monthly[(yr, 1)] = {"revenue": half, "stores": stores, "rev_pct": rev_pct, "is_estimated_split": True}
                    monthly[(yr, 2)] = {"revenue": half, "stores": stores, "rev_pct": rev_pct, "is_estimated_split": True}
                prev_n, prev_cum_rev = 2, cum_rev
            else:
                # co khoang trong lien tiep (VD thieu bao cao 1 thang giua chung) -> khong noi suy,
                # chi cap nhat moc luy ke gan nhat de tinh hieu so tu day tro di
                prev_n, prev_cum_rev = n, (cum_rev if cum_rev is not None else prev_cum_rev)
    return monthly

print("[IR] Fetching full monthly history 2019-now (cached - only new reports downloaded each run)...")
MWG_MONTHLY_HISTORY = fetch_mwg_full_history(start_year=2019)
MWG_MONTHLY_STANDALONE = derive_monthly_from_cumulative(MWG_MONTHLY_HISTORY)
print(f"  -> Derived {len(MWG_MONTHLY_STANDALONE)} thang don le tu du lieu luy ke")

print("[IR] Fetching latest MWG monthly business update for store counts...")
_ir_result = fetch_mwg_latest_store_counts()
if _ir_result:
    _fetched_counts, _post_title, _post_date = _ir_result
    _missing = [k for k, v in _fetched_counts.items() if v is None]
    if _missing:
        print(f"  [WARN] Missing store count fields from PDF regex: {_missing} - keeping hardcoded fallback for those")
    STORE_COUNT_NOW = {**STORE_COUNT_NOW, **{k: v for k, v in _fetched_counts.items() if v is not None}}
    print(f"  -> OK: '{_post_title}' ({_post_date}) -> {STORE_COUNT_NOW}")
    try:
        os.makedirs(".cache", exist_ok=True)
        with open(STORE_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({"counts": STORE_COUNT_NOW, "source": _post_title, "date": _post_date,
                       "fetchedAt": datetime.now().strftime("%Y-%m-%d %H:%M")}, f, ensure_ascii=False, indent=2)
    except OSError:
        pass
else:
    print("  [WARN] IR fetch that bai - dung STORE_COUNT_NOW hardcode (co the da cu)")

# ── 3. QUÝ GẦN NHẤT — Biên LNG (2 quý) & xu hướng SG&A/DT (4 quý), theo đúng skill ban-le mục 4 ──
print("[Model] Tinh Bien LNG (2 quy gan nhat) va xu huong SG&A (4 quy gan nhat)...")
is_q = section_to_quarters(RAW, "INCOME_STATEMENT")

def qv(records, year, q, field):
    for r in records:
        if r.get("yearReport") == year and r.get("lengthReport") == q:
            v = r.get(field)
            return v / 1e9 if v is not None else None
    return None

_all_q_keys = sorted({(r["yearReport"], r["lengthReport"]) for r in is_q
                      if r.get("yearReport") and r.get("lengthReport") in (1, 2, 3, 4)}, reverse=True)
LAST_4Q = []
for _y, _q in _all_q_keys:
    _rev = qv(is_q, _y, _q, "isa3")
    _cogs = qv(is_q, _y, _q, "isa4")
    _sell = qv(is_q, _y, _q, "isa9")
    _admin = qv(is_q, _y, _q, "isa10")
    if _rev and _cogs is not None and _sell is not None and _admin is not None:
        LAST_4Q.append({"year": _y, "q": _q, "rev": _rev,
                         "gpm": (_rev - abs(_cogs)) / _rev,
                         "sga_pct": (abs(_sell) + abs(_admin)) / _rev})
    if len(LAST_4Q) >= 4:
        break
LAST_4Q = LAST_4Q[::-1]  # sắp tăng dần theo thời gian (cũ -> mới), cần cho hồi quy xu hướng

GPM_FC_PCT = round(stats.mean(q["gpm"] for q in LAST_4Q[-2:]) * 100, 2) if len(LAST_4Q) >= 2 else 20.0

# SG&A/DT: hồi quy tuyến tính 4 quý gần nhất (không dùng trung bình dài hạn — skill mục 4), ngoại suy
# tiếp cho các quý dự phóng. Nếu < 2 điểm dữ liệu, dùng flat = tỷ lệ quý gần nhất.
if len(LAST_4Q) >= 2:
    _sga_slope, _sga_intercept = np.polyfit(range(len(LAST_4Q)), [q["sga_pct"] for q in LAST_4Q], 1)
else:
    _sga_slope, _sga_intercept = 0.0, (LAST_4Q[-1]["sga_pct"] if LAST_4Q else 0.15)

def _sga_pct_at(idx_from_last_known):
    # idx_from_last_known: 1 = quý ngay sau quý cuối cùng trong LAST_4Q, 2 = quý kế tiếp, ...
    x = len(LAST_4Q) - 1 + idx_from_last_known
    return max(0.03, _sga_slope * x + _sga_intercept)  # sàn 3% DT — tránh hồi quy âm phi thực tế

SGA_FC_PCT = []  # % SG&A/DT bình quân cho mỗi năm dự phóng (4 quý tiếp theo mỗi năm)
for _yi in range(len(years_fc)):
    _q_vals = [_sga_pct_at(_yi * 4 + k) for k in range(1, 5)]
    SGA_FC_PCT.append(round(stats.mean(_q_vals) * 100, 2))

print(f"  -> GPM du phong (TB 2 quy gan nhat): {GPM_FC_PCT}% | SG&A/DT du phong theo nam {years_fc}: {SGA_FC_PCT}%")

# ── DATABASE THEO THÁNG — hợp nhất mọi nguồn (MWG_MONTHLY_STANDALONE + QUARTERLY_SNAPSHOTS_MANUAL)
# thành 1 ma trận tháng đơn lẻ. Dùng để (1) dựng sheet Excel 03a_Monthly_Data VÀ (2) tính dự phóng
# doanh thu Python NGAY BÊN DƯỚI — CÙNG 1 CÔNG THỨC, CÙNG 1 NGUỒN DỮ LIỆU, để tránh lặp lại bug "2
# luồng tính độc lập lệch nhau" đã gặp ở HPG (xem ghi chú build_hpg_model.py: thay vì đọc ngược Excel
# bằng win32com/LibreOffice — không portable qua GitHub Actions Linux runner — Python mirror THẲNG
# công thức Excel dùng, một nguồn số duy nhất chạy giống hệt trên mọi môi trường).
def build_monthly_matrix():
    """Trả về list các dict theo THỨ TỰ THỜI GIAN, mỗi dict = 1 tháng: {year, month, label, revenue,
    pct: {seg: % hoặc None}, stores: {seg: số hoặc None}}. % mảng ưu tiên lấy từ table_row CHÍNH tháng
    đó (Era D 2020-6/2022); nếu thiếu, lấy % gần nhất trong QUARTERLY_SNAPSHOTS_MANUAL (trong vòng 6
    tháng — coi là proxy hợp lý vì cơ cấu DT đổi chậm theo tháng); nếu không có điểm nào đủ gần, để
    None (KHÔNG đoán bừa)."""
    manual_pct_points = []
    for key, snap in QUARTERLY_SNAPSHOTS_MANUAL.items():
        if "store_pct" in snap:
            y, m = key.split("-")
            manual_pct_points.append((int(y), int(m), snap["store_pct"]))
    manual_pct_points.sort()

    def nearest_pct(yr, mo):
        best, best_dist = None, 999
        for (py, pm, pct) in manual_pct_points:
            dist = abs((py - yr) * 12 + (pm - mo))
            if dist < best_dist:
                best, best_dist = pct, dist
        return best if best_dist <= 6 else None

    matrix = []
    for (yr, mo) in sorted(MWG_MONTHLY_STANDALONE.keys()):
        v = MWG_MONTHLY_STANDALONE[(yr, mo)]
        rev_pct_src = v.get("rev_pct") or {}
        near = nearest_pct(yr, mo)
        pct = {}
        for seg in SEGMENTS:
            if rev_pct_src.get(seg) is not None:
                pct[seg] = rev_pct_src[seg]
            elif near and near.get(seg) is not None:
                pct[seg] = near[seg]
            else:
                pct[seg] = None
        stores = v.get("stores") or {}
        khac_count = None
        if stores.get("AnKhang") is not None and stores.get("AvaKids") is not None:
            khac_count = stores["AnKhang"] + stores["AvaKids"] + (stores.get("EraBlue") or 0)
        matrix.append({
            "year": yr, "month": mo, "label": f"T{mo}-{str(yr)[2:]}",
            "revenue": v["revenue"],
            "pct": pct,
            "stores": {"TGDD": stores.get("TGDD"), "DMX": stores.get("DMX"), "BHX": stores.get("BHX"), "Khac": khac_count},
        })
    return matrix

MONTHLY_MATRIX = build_monthly_matrix()
_n_pct_ok = sum(1 for m in MONTHLY_MATRIX if all(m["pct"].get(s) is not None for s in SEGMENTS))
print(f"  -> Monthly matrix: {len(MONTHLY_MATRIX)} thang, {_n_pct_ok} thang co du % ca 4 mang")

# ── 4. DỰ PHÓNG DOANH THU 3-YẾU-TỐ THEO MẢNG (skill ban-le mục 2+3) ─────────────────────────────
# Doanh thu mảng = Số cửa hàng dự phóng × Hiệu quả DT/cửa hàng dự phóng. Hiệu quả HIỆN TẠI = TRUNG
# BÌNH 3 THÁNG THẬT GẦN NHẤT trong MONTHLY_MATRIX (đúng công thức lag ramp-up thực tế theo skill mục
# 2 — dùng dữ liệu THÁNG thật thay vì suy từ tổng 5 tháng/số CH cuối năm trước như bản cũ) — Y HỆT
# công thức AVERAGE() ghi vào sheet Excel 03a_Monthly_Data, đảm bảo Excel và Python luôn ra cùng 1 số.
print("[Model] Du phong doanh thu 3-yeu-to theo mang (TGDD/DMX/BHX/Khac)...")

CURRENT_YEAR_REAL = MONTHLY_MATRIX[-1]["year"]
_months_current_year_idx = [i for i, m in enumerate(MONTHLY_MATRIX) if m["year"] == CURRENT_YEAR_REAL]
_n_known_months_current_year = len(_months_current_year_idx)

def _seg_month_revenue(m, seg):
    """Doanh thu mảng CỦA 1 THÁNG = Tổng tháng đó × %mảng (None nếu thiếu %, KHÔNG đoán)."""
    pct = m["pct"].get(seg)
    return m["revenue"] * pct if pct is not None else None

def _seg_month_efficiency(m, seg):
    rev = _seg_month_revenue(m, seg)
    sl = m["stores"].get(seg)
    return rev / sl if (rev is not None and sl) else None

_last3_idx = _months_current_year_idx[-3:]
EFFICIENCY_NOW = {  # tỷ VND/CH/tháng — TB 3 tháng thật gần nhất của năm hiện tại (khớp Excel AVERAGE())
    seg: stats.mean([e for e in (_seg_month_efficiency(MONTHLY_MATRIX[i], seg) for i in _last3_idx) if e is not None])
    for seg in SEGMENTS
}

# Tăng trưởng hiệu quả/cửa hàng NĂM (giả định, có căn cứ — ghi rõ nguồn để dễ chỉnh sửa):
#   TGDD: 0%/năm — chuỗi đã bão hòa hoàn toàn (skill: "TGDD/ĐMX đã bão hòa"), không gian tăng hiệu quả
#     hạn chế, chủ yếu tăng nhờ chu kỳ sản phẩm (iPhone) chứ không phải cải thiện cấu trúc.
#   DMX: +3%/năm — còn dư địa từ tối ưu hoá ĐMS (supermini) và tăng trưởng ngành hàng máy lạnh/gia
#     dụng, nhưng đã là mảng lớn nhất nên tốc độ vừa phải.
#   BHX: +8%/năm — mảng đang scale nhanh nhất (SSSG cao nhất công ty theo IR), hiệu quả/cửa hàng còn
#     nhiều dư địa cải thiện khi tối ưu danh mục hàng tươi sống + logistics, nhưng giảm dần so với giai
#     đoạn 2023-2025 (khi vừa tái cơ cấu) vì nền đã cao hơn.
#   Khac (An Khang/AvaKids/EraBlue): +5%/năm — gộp nhiều mảng nhỏ khác pha tăng trưởng, EraBlue
#     (Indonesia) mới, An Khang đã tái cơ cấu ổn định.
EFFICIENCY_GROWTH = {"TGDD": 0.00, "DMX": 0.03, "BHX": 0.08, "Khac": 0.05}

# Số cửa hàng dự phóng cuối mỗi năm (giả định, có căn cứ IR — ghi rõ nguồn):
#   TGDD: giảm nhẹ tiếp tục xu hướng tái cơ cấu 2024->2025 (1021->1012, ~-1%/năm), giữ nguyên tốc độ.
#   DMX: gần như đi ngang quanh mức hiện tại (2024->2025 đi ngang 2026->2008, hiện 2004) — bão hòa.
#   BHX: theo kế hoạch mở mới chính thức FY2026_BHX_NEW_STORES=1000 cửa hàng cả năm 2026 (IR "Định
#     hướng kinh doanh 2026"), giảm tốc dần các năm sau (giả định +700/+500) khi mật độ phủ tăng.
#   Khac: An Khang/AvaKids đi ngang (đã tái cơ cấu ổn định), EraBlue tăng mạnh theo mục tiêu công ty
#     FY2026_ERABLUE_TARGET_STORES=300, giả định tiếp tục +100/năm sau đó (thị trường Indonesia mới).
_now = STORE_COUNT_NOW
STORE_COUNT_FC = {
    "TGDD": [round(_now["TGDD"] * (1 - 0.01) ** i) for i in range(1, len(years_fc) + 1)],
    "DMX":  [_now["DMX"] for _ in years_fc],
    "BHX":  [2559 + 1000, 2559 + 1000 + 700, 2559 + 1000 + 700 + 500],  # 2559 = BHX cuối 2025 (mốc kế hoạch)
    "Khac": [
        _now["AnKhang"] + _now["AvaKids"] + 300,
        _now["AnKhang"] + _now["AvaKids"] + 400,
        _now["AnKhang"] + _now["AvaKids"] + 500,
    ],
}

# Doanh thu mảng dự phóng — Y HỆT công thức Excel sheet 03a_Monthly_Data:
#   Năm HIỆN TẠI (CURRENT_YEAR_REAL, đã có dữ liệu thật 1 phần): SL ước tính × hiệu quả/CH ước tính ×
#     SỐ THÁNG CÒN LẠI + DOANH THU LŨY KẾ ĐÃ CÓ (tổng các tháng thật trong năm, chỉ tính tháng có %).
#   Các năm SAU (không có dữ liệu thật): SL ước tính × hiệu quả/CH ước tính × 12 tháng trọn năm; hiệu
#     quả/CH mỗi năm sau = hiệu quả năm trước × (1 + tăng trưởng giả định) — compound liên tiếp.
REVENUE_FC_SEGMENT = {}  # {segment: [rev_2026E, rev_2027F, rev_2028F]} tỷ VND/năm
EFFICIENCY_FC_SEGMENT = {}  # {segment: [eff_2026E, eff_2027F, eff_2028F]} tỷ VND/CH/tháng
for seg in SEGMENTS:
    seg_rev, seg_eff = [], []
    for i, yr in enumerate(years_fc):
        if yr == CURRENT_YEAR_REAL:
            eff_yr = EFFICIENCY_NOW[seg]
            remaining_months = 12 - _n_known_months_current_year
            cum_known = sum(v for v in (_seg_month_revenue(MONTHLY_MATRIX[j], seg) for j in _months_current_year_idx) if v is not None)
            rev_yr = STORE_COUNT_FC[seg][i] * eff_yr * remaining_months + cum_known
        else:
            eff_yr = seg_eff[i - 1] * (1 + EFFICIENCY_GROWTH[seg])
            rev_yr = STORE_COUNT_FC[seg][i] * eff_yr * 12
        seg_eff.append(eff_yr)
        seg_rev.append(round(rev_yr, 1))
    REVENUE_FC_SEGMENT[seg] = seg_rev
    EFFICIENCY_FC_SEGMENT[seg] = seg_eff

revenue_fc = [sum(REVENUE_FC_SEGMENT[seg][i] for seg in SEGMENTS) for i in range(len(years_fc))]

print(f"  -> Hieu qua/CH/thang HIEN TAI (TB 3 thang gan nhat, ty VND): {({k: round(v,2) for k,v in EFFICIENCY_NOW.items()})}")
print(f"  -> Doanh thu du phong {years_fc}: {revenue_fc}")

# ── 5. DỰ PHÓNG P&L (GPM/SG&A đã tính ở mục 3) ──────────────────────────────────────────────────
gp_margin_fc = [GPM_FC_PCT] * len(years_fc)  # % — giữ nguyên biên LNG 2 quý gần nhất cho cả 3 năm
gp_fc = [round(revenue_fc[i] * gp_margin_fc[i] / 100, 1) for i in range(len(years_fc))]
sga_fc = [round(revenue_fc[i] * SGA_FC_PCT[i] / 100, 1) for i in range(len(years_fc))]
ebit_fc = [round(gp_fc[i] - sga_fc[i], 1) for i in range(len(years_fc))]

# Thuế TNDN hiệu dụng — trung bình 2 năm gần nhất thực tế, dùng ĐÚNG "Lãi/lỗ trước thuế" (isa16, EBT
# thật) chứ KHÔNG xấp xỉ bằng EBIT (EBIT bỏ qua DT/CP tài chính ròng — MWG có lãi tiền gửi đáng kể từ
# lượng tiền mặt lớn, dùng EBIT thay EBT sẽ ra thuế suất suy ngược sai/phi thực tế).
ebt_hist = [yv(is_hist, y, "isa16") for y in years_hist]
_tax_hist = [1 - ni_consol_hist[i] / ebt_hist[i] for i in range(len(years_hist)) if ebt_hist[i]]
EFFECTIVE_TAX_RATE = round(stats.mean(_tax_hist[-2:]), 4) if len(_tax_hist) >= 2 else 0.20

# EBT dự phóng = EBIT dự phóng + (DT tài chính - CP tài chính) ròng — dùng bình quân 2 năm gần nhất
# làm tỷ lệ % DT ổn định (MWG duy trì tiền mặt lớn, khoản mục này khá đều qua các năm).
fin_income_hist = [yv(is_hist, y, "isa6") for y in years_hist]
fin_expense_hist = [abs(yv(is_hist, y, "isa7")) for y in years_hist]
_net_fin_pct_hist = [(fin_income_hist[i] - fin_expense_hist[i]) / revenue_hist[i] for i in range(len(years_hist)) if revenue_hist[i]]
NET_FIN_PCT = round(stats.mean(_net_fin_pct_hist[-2:]), 4) if len(_net_fin_pct_hist) >= 2 else 0.0
ebt_fc = [round(ebit_fc[i] + revenue_fc[i] * NET_FIN_PCT, 1) for i in range(len(years_fc))]
ni_consol_fc = [round(ebt_fc[i] * (1 - EFFECTIVE_TAX_RATE), 1) for i in range(len(years_fc))]

# LNST Cổ đông Công ty mẹ (isa22) luôn NHỎ HƠN LNST hợp nhất (isa20) do lợi ích cổ đông thiểu số ở 1
# số công ty con (DMX, BHX...) — dùng tỷ lệ bình quân 2 năm gần nhất để quy đổi, tránh giả định cứng
_minority_ratio_hist = [ni_hist[i] / ni_consol_hist[i] for i in range(len(years_hist)) if ni_consol_hist[i]]
MINORITY_RATIO = round(stats.mean(_minority_ratio_hist[-2:]), 4) if len(_minority_ratio_hist) >= 2 else 0.95
ni_fc = [round(ni_consol_fc[i] * MINORITY_RATIO, 1) for i in range(len(years_fc))]

print(f"  -> GPM {gp_margin_fc}% | SG&A/DT {SGA_FC_PCT}% | Thue hieu dung {EFFECTIVE_TAX_RATE*100:.1f}%")
print(f"  -> EBIT du phong: {ebit_fc}")
print(f"  -> NI (co dong cong ty me) du phong: {ni_fc}")

# ── 6. ĐỊNH GIÁ — 4 phương pháp theo skill ban-le (P/E 20%, P/B 20%, RI 10%, P/S+P/E cả 4 mảng 50%) ──
# (2026-07: đã bỏ P/S+P/B — không tách bạch được VCSH riêng mảng, dồn 15% cũ vào P/S+P/E)
# Hàm fetch Rf/Beta (CAPM cho RI) — copy nguyên logic đã verify hoạt động tốt từ template_banking.py
# (dùng cho định giá RI ngân hàng), chỉ đổi risk premium đặc thù (2% "bank-specific" -> 1.5% "retail
# frontier-market" chung, không có yếu tố ngân hàng cụ thể) — xem ghi chú tại chỗ dùng bên dưới.
print("[Valuation] Tinh COE (CAPM) va cac muc tieu dinh gia...")
RF_UA_STR = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"

def _fetch_via_curl_rf(url, timeout=10, label=None):
    """Tải HTML qua curl subprocess (KHÔNG dùng requests) — investing.com chặn vân tay TLS của Python
    requests/urllib3 (Cloudflare 403) nhưng vẫn cho phép curl (đã verify ở build cho ngân hàng)."""
    tag = f" [{label}]" if label else ""
    try:
        r = subprocess.run(
            ["curl", "-sL", "-A", RF_UA_STR, "--max-time", str(timeout), url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout + 5,
        )
        if r.returncode != 0 or not r.stdout.strip():
            print(f"  [DIAG]{tag} fetch_via_curl empty/failed: curl_exit={r.returncode} len={len(r.stdout)}")
        return r.stdout if r.returncode == 0 else ""
    except Exception as e:
        print(f"  [DIAG]{tag} fetch_via_curl exception: {url[:80]} -> {e}")
        return ""

def fetch_rf_vietnam(timeout=15):
    FALLBACK_RF = 0.045
    try:
        html = _fetch_via_curl_rf("https://vn.investing.com/rates-bonds/vietnam-10-year-bond-yield", timeout=timeout, label="investing-rf")
        if html:
            m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
            if m:
                rf = float(m.group(1).replace(",", "")) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "investing.com"
    except Exception as e:
        print(f"  [WARN] investing.com Rf failed: {e}")
    try:
        r = requests.get("https://www.worldgovernmentbonds.com/bond-yield/vietnam/10-years/",
                          headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
        if r.status_code == 200:
            matches = re.findall(r'(\d+\.\d+)%', r.text[:5000])
            if matches:
                rf = float(matches[0]) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "worldgovernmentbonds.com"
    except Exception as e:
        print(f"  [WARN] WorldGovernmentBonds Rf failed: {e}")
    return FALLBACK_RF, "Fallback (manual)"

def fetch_aligned_history(ticker, days=720, timeout=15):
    from_time, to_time = 1577836800, 2000000000
    url_stock = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol={ticker}&resolution=D&from={from_time}&to={to_time}"
    url_index = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol=VNINDEX&resolution=D&from={from_time}&to={to_time}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
               "Accept": "application/json, text/plain, */*", "Referer": "https://dchart.vndirect.com.vn/"}
    try:
        r_stock = requests.get(url_stock, headers=headers, timeout=timeout)
        r_index = requests.get(url_index, headers=headers, timeout=timeout)
        if r_stock.status_code == 200 and r_index.status_code == 200:
            d_stock, d_index = r_stock.json(), r_index.json()
            t_s, c_s = d_stock.get("t") or [], d_stock.get("c") or []
            t_m, c_m = d_index.get("t") or [], d_index.get("c") or []
            map_stock = {t_s[i]: c_s[i] for i in range(min(len(t_s), len(c_s))) if c_s[i] is not None and c_s[i] > 0}
            map_index = {t_m[i]: c_m[i] for i in range(min(len(t_m), len(c_m))) if c_m[i] is not None and c_m[i] > 0}
            common_t = sorted(set(map_stock.keys()) & set(map_index.keys()))
            aligned = []
            for t in common_t:
                p_s = map_stock[t]
                if p_s < 1000:
                    p_s *= 1000
                aligned.append((datetime.fromtimestamp(t).strftime("%Y-%m-%d"), p_s, map_index[t]))
            return aligned
    except Exception as e:
        print(f"[Beta Calc] Error fetching history: {e}")
    return []

def fetch_beta_vietstock(ticker, timeout=15):
    try:
        r1 = requests.get(f"https://finance.vietstock.vn/search?query={ticker}",
                           headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}, timeout=timeout)
        if r1.status_code == 200:
            data = json.loads(r1.text).get("data", "")
            target_url = ""
            for line in data.split('\r\n'):
                parts = line.split('|')
                if len(parts) >= 3 and parts[0].strip().upper() == ticker.upper():
                    target_url = parts[2]
                    break
            if target_url:
                r2 = requests.get(target_url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                                                        'Referer': 'https://finance.vietstock.vn/'}, timeout=timeout)
                if r2.status_code == 200:
                    m = re.search(r'\"Beta\":\"([\d\.]+)\"', r2.text)
                    if m:
                        beta = float(m.group(1))
                        if 0.3 <= beta <= 2.5:
                            print(f"  [OK] Beta {ticker} tu Vietstock: {beta:.2f}")
                            return beta
    except Exception as e:
        print(f"  [WARN] Vietstock scrape failed: {e}")
    return None

def fetch_beta_vietcap(ticker, timeout=15):
    try:
        r = requests.get(f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}",
                          headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Referer": "https://trading.vietcap.com.vn/"},
                          timeout=timeout)
        if r.status_code == 200:
            beta = r.json().get("data", {}).get("beta")
            if beta is not None and 0.3 <= float(beta) <= 2.5:
                print(f"  [OK] Beta {ticker} tu Vietcap: {float(beta):.2f}")
                return float(beta)
    except Exception:
        pass
    return None

def fetch_and_calc_beta(ticker, days=720, timeout=20, fallback=1.0):
    print(f"  [INFO] Dang tai lich su gia de tu tinh Beta cho {ticker}...")
    aligned_data = fetch_aligned_history(ticker, days=days, timeout=timeout)
    latest_price = aligned_data[-1][1] if aligned_data else None
    num_sessions = len(aligned_data)
    web_beta = fetch_beta_vietstock(ticker, timeout) or fetch_beta_vietcap(ticker, timeout) or fallback
    calculated_beta, is_enough_sessions = fallback, False
    if num_sessions >= 30:
        sliced_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
        s = [x[1] for x in sliced_data]
        m = [x[2] for x in sliced_data]
        rs = [(s[i] - s[i-1]) / s[i-1] for i in range(1, len(s))]
        rm = [(m[i] - m[i-1]) / m[i-1] for i in range(1, len(m))]
        n_ret = len(rs)
        mean_rs, mean_rm = sum(rs) / n_ret, sum(rm) / n_ret
        cov_sm = sum((rs[i]-mean_rs)*(rm[i]-mean_rm) for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 0
        var_m = sum((rm[i]-mean_rm)**2 for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 1.0
        calculated_beta = round(max(0.3, min(2.5, cov_sm / var_m if var_m > 0 else fallback)), 4)
        if num_sessions >= 250:
            is_enough_sessions = True
    beta_src = f"Tu tinh toan ({num_sessions} phien)" if is_enough_sessions else f"Web/API ({web_beta:.2f}) - lich su chi {num_sessions} phien"
    return calculated_beta, web_beta, is_enough_sessions, beta_src, latest_price, aligned_data

rf_val, rf_src = fetch_rf_vietnam()
beta_calc, beta_web, is_enough_sessions, beta_src, _latest_px, BETA_ALIGNED_DATA = fetch_and_calc_beta(TICKER)
beta_val = beta_calc if is_enough_sessions else beta_web
beta_raw = beta_val
beta_val = round(0.67 * beta_raw + 0.33, 4)  # Blume adjusted beta (mean reversion về 1)
ERP = 0.07  # Damodaran ERP
SPECIFIC_RISK_PREMIUM = 0.015  # frontier-market retail risk premium chung (thấp hơn 2% ngân hàng vì
                                # bán lẻ không chịu rủi ro đặc thù NPL/thanh khoản như ngân hàng)
COE = rf_val + beta_val * ERP + SPECIFIC_RISK_PREMIUM
print(f"  -> Beta tho: {beta_raw:.4f} | Beta Blume: {beta_val} ({beta_src})")
print(f"  -> COE: {COE*100:.2f}% (Rf={rf_val*100:.2f}% [{rf_src}], ERP={ERP*100:.1f}%, Specific={SPECIFIC_RISK_PREMIUM*100:.1f}%)")

# ── 6a. P/E Median & P/B Median lịch sử (trọng số 20%+20%) — dùng chung API Vietcap statistics-financial ──
def fetch_vietcap_ratios(ticker, timeout=15):
    try:
        r = requests.get(f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial",
                          headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                                   "Referer": "https://trading.vietcap.com.vn/"}, timeout=timeout)
        r.raise_for_status()
        data = r.json().get("data", [])
        return [{"year": d.get("year"), "quarter": d.get("quarter"), "pe": d.get("pe"), "pb": d.get("pb"),
                  "ev_ebitda": d.get("evToEbitda")} for d in data]
    except Exception as e:
        print(f"  [WARN] Vietcap ratio API failed: {e}")
        return []

MWG_RATIOS = fetch_vietcap_ratios(TICKER)

def _compute_hist_multiple_medians(ratios, hist_years):
    """Trung vị-của-trung-vị-năm cho P/E (loại quý bất thường >50x hoặc <=0)/P/B — cùng logic dùng ở
    build_hpg_model.py, tái sử dụng cho MWG."""
    raw_by_q = {(int(r["year"]), r["quarter"]): r for r in ratios if r.get("quarter") in (1, 2, 3, 4)}
    last_good_pe, last_good_pb = None, None
    qdata = []
    for y in hist_years:
        for q in range(1, 5):
            rec = raw_by_q.get((y, q), {})
            pe0, pb0 = rec.get("pe"), rec.get("pb")
            is_normal = bool(pe0 and 0 < pe0 < 50)
            pe = round(pe0, 2) if is_normal else last_good_pe
            if is_normal:
                last_good_pe = pe
            if pb0 and pb0 > 0:
                pb = round(pb0, 2); last_good_pb = pb
            else:
                pb = last_good_pb
            qdata.append((y, pe, pb))

    def year_median(idx):
        yearly_medians = []
        for y in hist_years:
            vals = [row[idx] for row in qdata if row[0] == y and row[idx] is not None]
            if vals:
                yearly_medians.append(stats.median(vals))
        return round(stats.median(yearly_medians), 2) if yearly_medians else None
    return year_median(1), year_median(2)

_hist_years_ratio = list(range(max(2018, years_hist[0] if years_hist else 2018), (years_hist[-1] if years_hist else YEAR_NOW) + 1))
PE_HIST_MEDIAN, PB_HIST_MEDIAN = _compute_hist_multiple_medians(MWG_RATIOS, _hist_years_ratio)
PE_HIST_MEDIAN = PE_HIST_MEDIAN or 15.0
PB_HIST_MEDIAN = PB_HIST_MEDIAN or 3.0
print(f"  -> PE_HIST_MEDIAN={PE_HIST_MEDIAN}x | PB_HIST_MEDIAN={PB_HIST_MEDIAN}x (giai doan {_hist_years_ratio[0]}-{_hist_years_ratio[-1]})")

EPS_FC = [round(ni_fc[i] * 1000 / SHARES, 0) for i in range(len(years_fc))]  # VND/CP (ni_fc tỷ VND, SHARES triệu CP)
BVPS_HIST_LAST = equity_hist[-1] * 1000 / SHARES  # VND/CP

# Vốn CSH dự phóng — cộng dồn LNST hợp nhất giữ lại (payout ratio ước tính từ cổ tức/LNST lịch sử;
# MWG trả cổ tức tiền mặt đều đặn — dùng 30% làm giả định payout nếu không đủ dữ liệu lịch sử tách bạch)
DIVIDEND_PAYOUT_RATIO = 0.30
equity_fc = []
_eq = equity_hist[-1]
for i in range(len(years_fc)):
    _eq = _eq + ni_consol_fc[i] * (1 - DIVIDEND_PAYOUT_RATIO)
    equity_fc.append(round(_eq, 1))
BVPS_FC = [round(equity_fc[i] * 1000 / SHARES, 0) for i in range(len(years_fc))]

# 1) P/E Median (20%)
PE_TARGET_PRICE = round(PE_HIST_MEDIAN * EPS_FC[0], 0)
# 2) P/B Median (20%)
PB_TARGET_PRICE = round(PB_HIST_MEDIAN * BVPS_FC[0], 0)

# 3) RI cả công ty (10%) — Residual Income, Clean Surplus: BV(t) = BV(t-1) + EPS(t) x (1-payout)
TERMINAL_GROWTH = 0.03
_bv = BVPS_HIST_LAST
ri_results = []
for i in range(len(years_fc)):
    capital_charge = _bv * COE
    ri = EPS_FC[i] - capital_charge
    ri_results.append(ri)
    _bv = _bv + EPS_FC[i] * (1 - DIVIDEND_PAYOUT_RATIO)
cv = ri_results[-1] * (1 + TERMINAL_GROWTH) / (COE - TERMINAL_GROWTH) if COE > TERMINAL_GROWTH else 0
pv_ri = sum(ri_results[i] / (1 + COE) ** (i + 1) for i in range(len(ri_results)))
pv_cv = cv / (1 + COE) ** len(ri_results)
RI_TARGET_PRICE = round(BVPS_HIST_LAST + pv_ri + pv_cv, 0)

print(f"  -> EPS_FC {EPS_FC} | BVPS_FC {BVPS_FC}")
print(f"  -> P/E target: {PE_TARGET_PRICE:,.0f} | P/B target: {PB_TARGET_PRICE:,.0f} | RI target: {RI_TARGET_PRICE:,.0f}")

# 4) P/S + P/E theo mảng (trọng số 50% = 35%+15% gộp lại) — ÁP DỤNG CHO CẢ 4 MẢNG.
# (2026-07, chốt lại theo yêu cầu user): BỎ HẲN phương pháp P/S+P/B — không có cách tách bạch VCSH/
# BV riêng cho từng mảng (ĐMX, TGDĐ...) từ BCTC hợp nhất, nên "BV ngụ ý ~ NI/COE" trước đây chỉ là
# proxy 2 lớp (ước tính chồng ước tính), kém tin cậy hơn hẳn so với việc dùng P/E tham chiếu trực
# tiếp. Dồn toàn bộ trọng số 15% của P/S+P/B cũ sang P/S+P/E (35%+15%=50%), áp dụng P/E cho TẤT CẢ
# 4 mảng (kể cả BHX/Khác) — vẫn giữ quy tắc "P/E = min(P/E tham chiếu, 0.9×growth)" để chặn định giá
# quá cao khi tăng trưởng đột biến.
SEGMENT_METHOD = {"TGDD": "PE", "DMX": "PE", "BHX": "PE", "Khac": "PE"}

# Biên LNST ước tính mỗi mảng (tỷ trọng LN so với DT, năm dự phóng đầu) — MWG không công bố NI tách
# mảng, ước tính có căn cứ dựa trên đặc thù ngành + margin công ty mẹ (bình quân 2 năm gần nhất
# GPM_FC_PCT/SGA_FC_PCT áp dụng đồng nhất mọi mảng theo GP margin, chỉ phân hoá NHẸ theo đặc thù format
# cửa hàng — ICT có SG&A/DT thấp hơn hàng tiêu dùng nhanh do vòng quay tồn kho nhanh hơn):
SEGMENT_NET_MARGIN = {"TGDD": 0.045, "DMX": 0.040, "BHX": 0.020, "Khac": 0.010}

# Tăng trưởng DT dùng cho quy tắc "P/E = min(P/E tham chiếu, 0.9 x Growth)" — DÙNG TĂNG TRƯỞNG THỰC TẾ
# GẦN NHẤT (YoY 5 tháng 2026, IR công bố +29.3% toàn công ty, riêng TGDĐ+ĐMX "SSSG đạt 33%" theo báo
# cáo IR 4T/5T2026), KHÔNG dùng CAGR 3 năm dự phóng ở trên (STORE_COUNT_FC/EFFICIENCY_GROWTH cố tình
# giữ bảo thủ dài hạn cho mảng đã bão hòa TGDD/DMX ~0-3%/năm — áp thẳng số đó vào công thức P/E sẽ ra
# P/E<3x phi thực tế, vì công thức này nhằm chặn P/E bị thổi phồng bởi 1 NĂM tăng trưởng đột biến, không
# nhằm ép định giá mảng ổn định xuống mức bất hợp lý). Không có % tách riêng BHX/Khác nên ước tính có
# căn cứ dựa trên tốc độ mở mới cửa hàng thực tế (BHX +19% số lượng CH chỉ trong 5 tháng 2026).
SEGMENT_GROWTH_PCT = {"TGDD": 25.0, "DMX": 28.0, "BHX": 35.0, "Khac": 20.0}

# P/E tham chiếu ngành bán lẻ ICT/FMCG/dược phẩm khu vực ASEAN (proxy, có thể cập nhật khi có dữ liệu
# peer cụ thể hơn) — dùng cho cả 4 mảng vì tất cả đều dùng chung phương pháp P/S+P/E sau khi bỏ P/B.
SEGMENT_PE_REF = {"TGDD": 14.0, "DMX": 14.0, "BHX": 16.0, "Khac": 12.0}  # BHX ref cao hơn (FMCG tăng trưởng tốt hơn ICT bão hòa); Khac (dược/mẹ&bé/quốc tế) ref thận trọng hơn do quy mô nhỏ

segment_value_ps_pe = {}
for seg in SEGMENTS:
    rev_seg_fc0 = REVENUE_FC_SEGMENT[seg][0]
    pe_growth_based = 0.9 * SEGMENT_GROWTH_PCT[seg]
    pe_used = min(SEGMENT_PE_REF[seg], pe_growth_based) if pe_growth_based > 0 else SEGMENT_PE_REF[seg]
    pe_used = max(pe_used, 5.0)  # sàn hợp lý, tránh P/E âm/quá thấp khi tăng trưởng âm
    ni_seg = rev_seg_fc0 * SEGMENT_NET_MARGIN[seg]
    segment_value_ps_pe[seg] = pe_used * ni_seg  # tỷ VND vốn hoá ngụ ý cho mảng này

TOTAL_VALUE_PS_PE = sum(segment_value_ps_pe.values())  # tỷ VND
PS_PE_TARGET_PRICE = round(TOTAL_VALUE_PS_PE * 1e9 / (SHARES * 1e6), 0) if TOTAL_VALUE_PS_PE else PE_TARGET_PRICE

print(f"  -> P/S+P/E target (ca 4 mang {SEGMENTS}): {PS_PE_TARGET_PRICE:,.0f}")

# ── Tổng hợp trọng số 4 phương pháp (chốt với user 2026-07, đã bỏ P/S+P/B): PE 20% + PB 20% + RI 10% + PS-PE 50% ──
VALUATION_WEIGHTS = {"PE": 0.20, "PB": 0.20, "RI": 0.10, "PS_PE": 0.50}
WEIGHTED_TARGET_PRICE = round(
    VALUATION_WEIGHTS["PE"] * PE_TARGET_PRICE + VALUATION_WEIGHTS["PB"] * PB_TARGET_PRICE +
    VALUATION_WEIGHTS["RI"] * RI_TARGET_PRICE + VALUATION_WEIGHTS["PS_PE"] * PS_PE_TARGET_PRICE, 0)
UPSIDE_PCT = round((WEIGHTED_TARGET_PRICE / PRICE - 1) * 100, 1) if PRICE else None

print(f"  -> GIA MUC TIEU (trong so 20/20/10/50): {WEIGHTED_TARGET_PRICE:,.0f} VND (gia hien tai {PRICE:,.0f}, upside {UPSIDE_PCT}%)")

# ── 7. DỰ PHÓNG D&A/CAPEX/BẢNG CÂN ĐỐI (cho sheet 05/06) ────────────────────────────────────────
# D&A dự phóng — giữ tỷ lệ D&A/DT bình quân 2 năm gần nhất (mở rộng chuỗi cửa hàng mới kéo theo D&A
# tăng tỷ lệ thuận doanh thu, hợp lý hơn giả định %YoY cố định cho 1 công ty đang mở rộng nhanh BHX).
_da_pct_hist = [da_hist[i] / revenue_hist[i] for i in range(len(years_hist)) if revenue_hist[i]]
DA_PCT_FC = stats.mean(_da_pct_hist[-2:]) if len(_da_pct_hist) >= 2 else 0.01
da_fc = [round(revenue_fc[i] * DA_PCT_FC, 1) for i in range(len(years_fc))]
ebitda_fc = [round(ebit_fc[i] + da_fc[i], 1) for i in range(len(years_fc))]

_capex_pct_hist = [capex_hist[i] / revenue_hist[i] for i in range(len(years_hist)) if revenue_hist[i]]
CAPEX_PCT_FC = stats.mean(_capex_pct_hist[-2:]) if len(_capex_pct_hist) >= 2 else 0.015
capex_fc = [round(revenue_fc[i] * CAPEX_PCT_FC, 1) for i in range(len(years_fc))]

# Hàng tồn kho/Phải thu/Phải trả dự phóng — giữ nguyên số ngày vòng quay bình quân 2 năm gần nhất
# (DIO/DSO/DPO, skill ban-le mục "Chu kỳ tiền mặt") áp dụng lên COGS/DT dự phóng.
_dio_hist = [inventory_hist[i] / cogs_hist[i] * 365 for i in range(len(years_hist)) if cogs_hist[i]]
_dso_hist = [receivables_hist[i] / revenue_hist[i] * 365 for i in range(len(years_hist)) if revenue_hist[i]]
_dpo_hist = [payables_hist[i] / cogs_hist[i] * 365 for i in range(len(years_hist)) if cogs_hist[i]]
DIO_FC = stats.mean(_dio_hist[-2:]) if len(_dio_hist) >= 2 else 45.0
DSO_FC = stats.mean(_dso_hist[-2:]) if len(_dso_hist) >= 2 else 5.0
DPO_FC = stats.mean(_dpo_hist[-2:]) if len(_dpo_hist) >= 2 else 40.0
cogs_fc = [round(revenue_fc[i] * (1 - GPM_FC_PCT / 100), 1) for i in range(len(years_fc))]
inventory_fc = [round(cogs_fc[i] * DIO_FC / 365, 1) for i in range(len(years_fc))]
receivables_fc = [round(revenue_fc[i] * DSO_FC / 365, 1) for i in range(len(years_fc))]
payables_fc = [round(cogs_fc[i] * DPO_FC / 365, 1) for i in range(len(years_fc))]

# ── DIO/DSO/DPO THEO QUÝ (lịch sử thật, dùng để vẽ biểu đồ xu hướng vòng quay vốn lưu động) ──────
# Công thức "quý-hoá thường niên": DIO_quý = Tồn kho cuối quý / (Giá vốn quý / 91,25 ngày) — cách
# tính chuẩn của giới phân tích khi chỉ có 1 quý dữ liệu (KHÔNG dùng COGS lũy kế 4 quý vì sẽ làm mượt
# quá mức, mất tính thời điểm của từng quý riêng lẻ).
bs_q = section_to_quarters(RAW, "BALANCE_SHEET")
_dio_q_keys = sorted({(r["yearReport"], r["lengthReport"]) for r in bs_q
                      if r.get("yearReport") and r.get("lengthReport") in (1, 2, 3, 4)})
DIO_QUARTERLY, DSO_QUARTERLY, DPO_QUARTERLY, DIO_Q_LABELS = [], [], [], []
for _y, _q in _dio_q_keys:
    _inv = qv(bs_q, _y, _q, "bsa15")
    _pay = qv(bs_q, _y, _q, "bsa57")
    _rec = qv(bs_q, _y, _q, "bsa8")
    _cogs_q = qv(is_q, _y, _q, "isa4")
    _rev_q = qv(is_q, _y, _q, "isa3")
    if _inv is None or _cogs_q is None:
        continue
    DIO_Q_LABELS.append(f"{_y}Q{_q}")
    DIO_QUARTERLY.append(round(_inv / (abs(_cogs_q) / 91.25), 1))
    DPO_QUARTERLY.append(round(_pay / (abs(_cogs_q) / 91.25), 1) if _pay else None)
    DSO_QUARTERLY.append(round(_rec / (_rev_q / 91.25), 1) if _rec and _rev_q else None)
print(f"  -> DIO/DSO/DPO theo quy: {len(DIO_Q_LABELS)} quy ({DIO_Q_LABELS[0]} - {DIO_Q_LABELS[-1]})")

# Nợ vay/Tiền mặt dự phóng — neo gốc 2025A thật, giữ %YoY bình quân 2 năm gần nhất (MWG đang giảm dần
# đòn bẩy nhờ FCF chuyển dương từ khi BHX có lãi, xu hướng nợ giảm/tiền mặt tăng dần).
_debt_yoy_hist = [debt_hist[i] / debt_hist[i-1] for i in range(1, len(years_hist)) if debt_hist[i-1]]
_cash_yoy_hist = [cash_hist[i] / cash_hist[i-1] for i in range(1, len(years_hist)) if cash_hist[i-1]]
DEBT_YOY_FC = stats.mean(_debt_yoy_hist[-2:]) if len(_debt_yoy_hist) >= 2 else 0.95
CASH_YOY_FC = stats.mean(_cash_yoy_hist[-2:]) if len(_cash_yoy_hist) >= 2 else 1.10
debt_fc, cash_fc = [], []
_d, _c = debt_hist[-1], cash_hist[-1]
for i in range(len(years_fc)):
    _d *= DEBT_YOY_FC; _c *= CASH_YOY_FC
    debt_fc.append(round(_d, 1)); cash_fc.append(round(_c, 1))

NET_DEBT_FC = [round(debt_fc[i] - cash_fc[i], 1) for i in range(len(years_fc))]
NET_DEBT_HIST = [round(debt_hist[i] - cash_hist[i], 1) for i in range(len(years_hist))]

# Tài sản khác (phần dư để cân đối TS = NV — không tách chi tiết từng khoản mục nhỏ, gộp vào 1 dòng
# "Tài sản khác" để bảng cân đối vẫn cân mà không cần dự phóng từng khoản mục phụ không trọng yếu)
other_assets_hist = [round(total_assets_hist[i] - cash_hist[i] - inventory_hist[i] - receivables_hist[i], 1) for i in range(len(years_hist))]
OTHER_ASSETS_PCT_FC = stats.mean([other_assets_hist[i] / revenue_hist[i] for i in range(len(years_hist)) if revenue_hist[i]][-2:])
other_assets_fc = [round(revenue_fc[i] * OTHER_ASSETS_PCT_FC, 1) for i in range(len(years_fc))]
total_assets_fc = [round(cash_fc[i] + inventory_fc[i] + receivables_fc[i] + other_assets_fc[i], 1) for i in range(len(years_fc))]

# Nợ khác (phần dư, tương tự tài sản khác) để Tổng NV = Tổng TS
other_liab_hist = [round(total_assets_hist[i] - debt_hist[i] - payables_hist[i] - equity_hist[i], 1) for i in range(len(years_hist))]
OTHER_LIAB_PCT_FC = stats.mean([other_liab_hist[i] / revenue_hist[i] for i in range(len(years_hist)) if revenue_hist[i]][-2:])
other_liab_fc = [round(revenue_fc[i] * OTHER_LIAB_PCT_FC, 1) for i in range(len(years_fc))]

print(f"  -> D&A/DT {DA_PCT_FC*100:.2f}% | CAPEX/DT {CAPEX_PCT_FC*100:.2f}% | DIO {DIO_FC:.0f}d DSO {DSO_FC:.0f}d DPO {DPO_FC:.0f}d")
print(f"  -> Net Debt du phong: {NET_DEBT_FC}")

years_all = years_hist + years_fc
N_HIST = len(years_hist)
N_ALL = len(years_all)
YEAR_HEADERS = [f"{y}A" for y in years_hist] + [f"{y}E" if i == 0 else f"{y}F" for i, y in enumerate(years_fc)]

# ══════════════════════════════════════════════════════════════════════════════════════════════
# EXCEL MODEL — 12 sheets theo skill xuat-bao-cao (mọi ô forecast là CÔNG THỨC SỐNG, không hardcode)
# ══════════════════════════════════════════════════════════════════════════════════════════════
def build_excel():
    S_A = "'02_Assumptions'"
    S_R = "'03_Revenue_Model'"
    S_P = "'04_PnL'"
    S_B = "'05_Balance_Sheet'"
    S_V = "'07_Valuation'"

    wb = openpyxl.Workbook()

    # ─── 01_Cover ───────────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "01_Cover"
    ws.merge_cells('A1:F1')
    ws['A1'] = f"PHÂN TÍCH CỔ PHIẾU {TICKER} — {COMPANY}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 32

    left_al = Alignment(horizontal='left', vertical='center')
    right_al = Alignment(horizontal='right', vertical='center')
    pe_ttm = round(PRICE / eps_hist[-1], 1) if eps_hist[-1] else 0
    pb_now = round(PRICE / bvps_hist[-1], 2) if bvps_hist[-1] else 0

    cover_rows = [
        ("Ticker", TICKER, None, left_al),
        ("Sàn", EXCHANGE, None, left_al),
        ("Ngành", INDUSTRY, None, left_al),
        ("Giá hiện tại (VND)", "='02_Assumptions'!B2", '#,##0', right_al),
        ("Số CP lưu hành (triệu)", "='02_Assumptions'!B3", '#,##0', right_al),
        ("Vốn hóa (tỷ VND)", "='02_Assumptions'!B4", '#,##0', right_al),
        ("P/E TTM (x)", pe_ttm, '0.0"x"', right_al),
        ("P/B (x)", pb_now, '0.00"x"', right_al),
        ("EPS TTM (VND)", eps_hist[-1], '#,##0', right_al),
        ("BVPS (VND)", bvps_hist[-1], '#,##0', right_al),
        ("Kế hoạch DT 2026 (tỷ)", FY2026_TARGET_REVENUE, '#,##0', right_al),
        ("Kế hoạch LNST 2026 (tỷ)", FY2026_TARGET_NPAT, '#,##0', right_al),
        ("Khuyến nghị", "=IF('07_Valuation'!C13>0.15,\"MUA\",IF('07_Valuation'!C13<-0.05,\"BÁN\",\"THEO DÕI\"))", None, Alignment(horizontal='center')),
        ("Giá mục tiêu (VND)", "='07_Valuation'!C11", '#,##0', right_al),
        ("Upside/Downside", "='07_Valuation'!C13", '0.0%', right_al),
        ("Ngày phân tích", datetime.now().strftime("%d/%m/%Y %H:%M"), None, left_al),
    ]
    for i, (k, v, nf, al) in enumerate(cover_rows, 2):
        ws.cell(row=i, column=1, value=k).font = bold_font
        ws.cell(row=i, column=1).border = thin_border
        c = ws.cell(row=i, column=2, value=v)
        c.font = data_font; c.border = thin_border; c.alignment = al
        if nf:
            c.number_format = nf
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 46

    ws.merge_cells('A19:F19')
    ws['A19'] = "4 PHƯƠNG PHÁP ĐỊNH GIÁ (trọng số theo skill ban-le, đã bỏ P/S+P/B)"
    ws['A19'].font = bold_font
    val_summary = [
        ("P/E Median (20%)", "='07_Valuation'!C6", '#,##0'),
        ("P/B Median (20%)", "='07_Valuation'!C7", '#,##0'),
        ("Residual Income (10%)", "='07_Valuation'!C8", '#,##0'),
        ("P/S + P/E theo mảng, cả 4 mảng (50%)", "='07_Valuation'!C9", '#,##0'),
    ]
    for i, (k, v, nf) in enumerate(val_summary, 20):
        ws.cell(row=i, column=1, value=k).font = data_font
        ws.cell(row=i, column=1).border = thin_border
        c = ws.cell(row=i, column=2, value=v); c.font = data_font; c.border = thin_border
        c.number_format = nf; c.alignment = right_al

    print("[Excel] Sheet 01_Cover done.")

    # ─── 00_Beta + 00_COE — bảng tính Beta lịch sử & CAPM (giống template_banking.py) ─────────────
    # Chèn TRƯỚC 01_Cover (index 0/1) để đúng thứ tự audit trail: Beta thô -> COE -> mọi định giá RI
    # tham chiếu về đây, KHÔNG hardcode COE rời rạc ở 02_Assumptions nữa.
    ws_beta = wb.create_sheet(title="00_Beta", index=0)
    ws_beta.column_dimensions['A'].width = 15
    ws_beta.column_dimensions['B'].width = 16
    ws_beta.column_dimensions['C'].width = 22
    ws_beta.column_dimensions['D'].width = 16
    ws_beta.column_dimensions['E'].width = 22
    ws_beta.cell(row=1, column=1, value="BẢNG TÍNH HỆ SỐ BETA LỊCH SỬ").font = bold_font
    ws_beta.cell(row=1, column=2, value="Beta thô (raw):").font = bold_font
    ws_beta.cell(row=1, column=3).number_format = '0.0000'
    ws_beta.cell(row=1, column=3).fill = assump_fill
    ws_beta.cell(row=1, column=4, value="Beta Blume (đã điều chỉnh):").font = bold_font
    ws_beta.cell(row=1, column=5).number_format = '0.0000'
    ws_beta.cell(row=1, column=5).fill = assump_fill
    ws_beta.cell(row=2, column=2, value="Số phiên giao dịch:").font = Font(name=FONT_NAME, italic=True, size=9)
    ws_beta.cell(row=4, column=1, value="Ngày").font = header_font
    ws_beta.cell(row=4, column=2, value=f"Giá {TICKER}").font = header_font
    ws_beta.cell(row=4, column=3, value=f"Tỷ suất sinh lời {TICKER}").font = header_font
    ws_beta.cell(row=4, column=4, value="Giá VNINDEX").font = header_font
    ws_beta.cell(row=4, column=5, value="Tỷ suất sinh lời VNINDEX").font = header_font
    for c in range(1, 6):
        ws_beta.cell(row=4, column=c).fill = header_fill
    if BETA_ALIGNED_DATA:
        date0, p_s0, p_m0 = BETA_ALIGNED_DATA[0]
        ws_beta.cell(row=5, column=1, value=date0)
        ws_beta.cell(row=5, column=2, value=p_s0)
        ws_beta.cell(row=5, column=4, value=p_m0)
        for ridx, (dstr, p_s, p_m) in enumerate(BETA_ALIGNED_DATA[1:], start=6):
            ws_beta.cell(row=ridx, column=1, value=dstr)
            ws_beta.cell(row=ridx, column=2, value=p_s)
            ws_beta.cell(row=ridx, column=3, value=f"=(B{ridx}-B{ridx-1})/B{ridx-1}").number_format = '0.00%'
            ws_beta.cell(row=ridx, column=4, value=p_m)
            ws_beta.cell(row=ridx, column=5, value=f"=(D{ridx}-D{ridx-1})/D{ridx-1}").number_format = '0.00%'
        last_row = 4 + len(BETA_ALIGNED_DATA)
        ws_beta.cell(row=1, column=3, value=f"=COVAR(C6:C{last_row},E6:E{last_row})/VAR(E6:E{last_row})")
        ws_beta.cell(row=1, column=5, value="=0.67*C1+0.33")
        ws_beta.cell(row=2, column=3, value=f"=COUNT(C6:C{last_row})")
    else:
        ws_beta.cell(row=1, column=3, value=beta_raw)
        ws_beta.cell(row=1, column=5, value=beta_val)
        ws_beta.cell(row=2, column=3, value=0)
    print(f"[Excel] Sheet 00_Beta done ({len(BETA_ALIGNED_DATA)} phien).")

    ws_coe = wb.create_sheet(title="00_COE", index=1)
    ws_coe.column_dimensions['A'].width = 42
    ws_coe.column_dimensions['B'].width = 16
    ws_coe.column_dimensions['C'].width = 50
    ws_coe.cell(row=1, column=1, value="CHI PHÍ VỐN CSH (COE) — MÔ HÌNH CAPM").font = title_font
    ws_coe.cell(row=3, column=1, value="Tham số").font = header_font
    ws_coe.cell(row=3, column=2, value="Giá trị").font = header_font
    ws_coe.cell(row=3, column=3, value="Ghi chú / Nguồn").font = header_font
    for c in range(1, 4):
        ws_coe.cell(row=3, column=c).fill = header_fill
    ws_coe.cell(row=4, column=1, value="Rf — Lãi suất phi rủi ro (TPCP 10 năm)")
    ws_coe.cell(row=4, column=2, value=rf_val).number_format = '0.00%'
    ws_coe.cell(row=4, column=3, value=rf_src)
    ws_coe.cell(row=5, column=1, value="β — Hệ số Beta (Blume-adjusted)")
    ws_coe.cell(row=5, column=2, value="='00_Beta'!E1").number_format = '0.0000'
    ws_coe.cell(row=5, column=3, value=beta_src)
    ws_coe.cell(row=6, column=1, value="ERP — Phần bù rủi ro vốn (Damodaran)")
    ws_coe.cell(row=6, column=2, value=ERP).number_format = '0.00%'
    ws_coe.cell(row=7, column=1, value="α — Phần bù rủi ro đặc thù (Frontier-market retail)")
    ws_coe.cell(row=7, column=2, value=SPECIFIC_RISK_PREMIUM).number_format = '0.00%'
    ws_coe.cell(row=9, column=1, value="COE = Rf + β×ERP + α").font = bold_font
    ws_coe.cell(row=9, column=2, value="=B4+B5*B6+B7").font = bold_font
    ws_coe.cell(row=9, column=2).number_format = '0.00%'
    ws_coe.cell(row=9, column=2).fill = p_fill
    for r in range(4, 10):
        for c in range(1, 4):
            ws_coe.cell(row=r, column=c).border = thin_border
    print("[Excel] Sheet 00_COE done.")

    # ─── 02_Assumptions ─────────────────────────────────────────────────────────────────────────
    # Nơi DUY NHẤT chứa số hardcode (input) — mọi sheet khác chỉ dùng công thức tham chiếu về đây.
    wsA = wb.create_sheet("02_Assumptions")
    header_row(wsA, 1, ["Giả định"] + YEAR_HEADERS + ["Ghi chú"], [34] + [11]*N_ALL + [46])
    FMT_NUM = '#,##0'
    FMT_PCT = '0.00%'
    FMT_MUL = '0.00"x"'

    def arow(ws, r, label, values, fmt=None, note="", bold=False, fill=None, single=False):
        """values: list độ dài N_ALL (multi-year) HOẶC 1 giá trị (single=True, chỉ ghi cột B)."""
        c = ws.cell(row=r, column=1, value=label)
        c.font = bold_font if bold else data_font
        c.border = thin_border
        if single:
            cell = ws.cell(row=r, column=2, value=values)
            cell.font = bold_font if bold else data_font; cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if fmt: cell.number_format = fmt
            if fill: cell.fill = fill
        else:
            for j, v in enumerate(values):
                cell = ws.cell(row=r, column=2 + j, value=v)
                cell.font = bold_font if bold else data_font; cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                if fmt: cell.number_format = fmt
                if fill: cell.fill = fill
        nc = ws.cell(row=r, column=2 + (1 if single else N_ALL), value=note)
        nc.font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
        nc.alignment = Alignment(horizontal='left', wrap_text=True)
        return r + 1

    RA = {}  # ROW_ASSUMP — vị trí dòng quan trọng để sheet khác tham chiếu chính xác
    r = 2
    RA["price"] = r; r = arow(wsA, r, "Giá cổ phiếu hiện tại (VND)", PRICE, FMT_NUM, "Giá đóng cửa gần nhất (Vietcap)", single=True)
    RA["shares"] = r; r = arow(wsA, r, "Số CP lưu hành (triệu CP)", SHARES, FMT_NUM, "Vốn góp/mệnh giá 10,000đ", single=True)
    RA["mcap"] = r; r = arow(wsA, r, "Vốn hóa (tỷ VND)", "=B2*B3/1000", FMT_NUM, "=Giá×SL CP", single=True)
    r += 1
    wsA.cell(row=r, column=1, value="SỐ CỬA HÀNG CUỐI KỲ (cửa hàng)").font = bold_font; r += 1
    RA["store_tgdd"] = r; r = arow(wsA, r, "TGDĐ (gồm Topzone)", [None]*N_HIST + STORE_COUNT_FC["TGDD"], FMT_NUM, "IR MWG; dự phóng: xu hướng tái cơ cấu -1%/năm")
    RA["store_dmx"] = r; r = arow(wsA, r, "ĐMX (gồm ĐMS)", [None]*N_HIST + STORE_COUNT_FC["DMX"], FMT_NUM, "Dự phóng: đi ngang (bão hòa)")
    RA["store_bhx"] = r; r = arow(wsA, r, "BHX", [None]*N_HIST + STORE_COUNT_FC["BHX"], FMT_NUM, "Dự phóng: +1000/+700/+500 theo kế hoạch IR")
    RA["store_khac"] = r; r = arow(wsA, r, "Khác (An Khang+AvaKids+EraBlue)", [None]*N_HIST + STORE_COUNT_FC["Khac"], FMT_NUM, "EraBlue mục tiêu >300 CH 2026 (IR)")
    r += 1
    wsA.cell(row=r, column=1, value="HIỆU QUẢ DT/CH/THÁNG HIỆN TẠI (tỷ VND, lag 3 tháng)").font = bold_font; r += 1
    RA["eff_now"] = {}
    for seg in SEGMENTS:
        RA["eff_now"][seg] = r
        r = arow(wsA, r, f"  {seg}", round(EFFICIENCY_NOW[seg], 3), '0.000', "DT 5T2026/5/số CH lag 3 tháng (skill ban-le mục 2)", single=True)
    RA["eff_growth"] = {}
    for seg in SEGMENTS:
        RA["eff_growth"][seg] = r
        r = arow(wsA, r, f"  Tăng trưởng hiệu quả/CH {seg} (%/năm)", EFFICIENCY_GROWTH[seg], FMT_PCT, "Giả định có căn cứ — xem comment code", single=True, fill=assump_fill)
    r += 1
    wsA.cell(row=r, column=1, value="BIÊN LỢI NHUẬN & CHI PHÍ").font = bold_font; r += 1
    RA["gpm"] = r; r = arow(wsA, r, "Biên lợi nhuận gộp (%)", [g/100 for g in gpm_hist] + [g/100 for g in gp_margin_fc], FMT_PCT, "Dự phóng = TB 2 quý gần nhất (skill ban-le mục 4)", fill=assump_fill)
    _sga_hist_pct = [sga_hist[i]/revenue_hist[i] for i in range(N_HIST)]
    RA["sga"] = r; r = arow(wsA, r, "SG&A / Doanh thu (%)", [s for s in _sga_hist_pct] + [s/100 for s in SGA_FC_PCT], FMT_PCT, "Dự phóng = hồi quy tuyến tính 4 quý gần nhất", fill=assump_fill)
    RA["tax"] = r; r = arow(wsA, r, "Thuế suất hiệu dụng (%)", EFFECTIVE_TAX_RATE, FMT_PCT, "TB 2 năm gần nhất = 1-LNST hợp nhất/LNTT (isa16)", single=True, fill=assump_fill)
    RA["netfin"] = r; r = arow(wsA, r, "DT/CP tài chính ròng / DT (%)", NET_FIN_PCT, FMT_PCT, "TB 2 năm gần nhất (lãi tiền gửi ròng)", single=True, fill=assump_fill)
    RA["payout"] = r; r = arow(wsA, r, "Tỷ lệ chia cổ tức tiền mặt/LNST (%)", DIVIDEND_PAYOUT_RATIO, FMT_PCT, "Giả định — MWG trả cổ tức đều đặn", single=True, fill=assump_fill)
    r += 1
    wsA.cell(row=r, column=1, value="VÒNG QUAY VỐN LƯU ĐỘNG").font = bold_font; r += 1
    RA["dio"] = r; r = arow(wsA, r, "DIO - Số ngày tồn kho", DIO_FC, '0.0', "TB 2 năm gần nhất", single=True)
    RA["dso"] = r; r = arow(wsA, r, "DSO - Số ngày phải thu", DSO_FC, '0.0', "TB 2 năm gần nhất", single=True)
    RA["dpo"] = r; r = arow(wsA, r, "DPO - Số ngày phải trả", DPO_FC, '0.0', "TB 2 năm gần nhất", single=True)
    RA["da_pct"] = r; r = arow(wsA, r, "D&A / Doanh thu (%)", DA_PCT_FC, FMT_PCT, "TB 2 năm gần nhất", single=True, fill=assump_fill)
    RA["capex_pct"] = r; r = arow(wsA, r, "CAPEX / Doanh thu (%)", CAPEX_PCT_FC, FMT_PCT, "TB 2 năm gần nhất", single=True, fill=assump_fill)
    RA["debt_yoy"] = r; r = arow(wsA, r, "Tăng trưởng Nợ vay YoY (%)", DEBT_YOY_FC - 1, FMT_PCT, "TB 2 năm gần nhất", single=True, fill=assump_fill)
    RA["cash_yoy"] = r; r = arow(wsA, r, "Tăng trưởng Tiền mặt YoY (%)", CASH_YOY_FC - 1, FMT_PCT, "TB 2 năm gần nhất", single=True, fill=assump_fill)
    r += 1
    wsA.cell(row=r, column=1, value="ĐỊNH GIÁ").font = bold_font; r += 1
    RA["pe_median"] = r; r = arow(wsA, r, "P/E Median lịch sử (x)", PE_HIST_MEDIAN, FMT_MUL, f"Trung vị-của-trung-vị-năm {_hist_years_ratio[0]}-{_hist_years_ratio[-1]} (Vietcap)", single=True)
    RA["pb_median"] = r; r = arow(wsA, r, "P/B Median lịch sử (x)", PB_HIST_MEDIAN, FMT_MUL, f"Trung vị-của-trung-vị-năm {_hist_years_ratio[0]}-{_hist_years_ratio[-1]} (Vietcap)", single=True)
    RA["beta"] = r; r = arow(wsA, r, "Beta (Blume-adjusted)", beta_val, '0.0000', f"Beta thô {beta_raw:.4f} ({beta_src})", single=True)
    RA["rf"] = r; r = arow(wsA, r, "Lãi suất phi rủi ro Rf (%)", rf_val, FMT_PCT, f"Nguồn: {rf_src}", single=True)
    RA["erp"] = r; r = arow(wsA, r, "Equity Risk Premium (%)", ERP, FMT_PCT, "Damodaran ERP", single=True)
    RA["specific_prem"] = r; r = arow(wsA, r, "Premium rủi ro đặc thù (%)", SPECIFIC_RISK_PREMIUM, FMT_PCT, "Frontier-market retail risk premium", single=True)
    RA["coe"] = r; r = arow(wsA, r, "Chi phí vốn CSH COE (%)", "='00_COE'!B9", FMT_PCT, "Xem chi tiết CAPM tại sheet 00_COE/00_Beta", single=True, fill=assump_fill)
    RA["terminal_g"] = r; r = arow(wsA, r, "Tăng trưởng dài hạn g (%)", TERMINAL_GROWTH, FMT_PCT, "Giả định terminal growth RI", single=True, fill=assump_fill)
    RA["w_pe"] = r; r = arow(wsA, r, "Trọng số P/E Median (%)", VALUATION_WEIGHTS["PE"], FMT_PCT, "Chốt với user 2026-07 (skill ban-le)", single=True, fill=assump_fill)
    RA["w_pb"] = r; r = arow(wsA, r, "Trọng số P/B Median (%)", VALUATION_WEIGHTS["PB"], FMT_PCT, "", single=True, fill=assump_fill)
    RA["w_ri"] = r; r = arow(wsA, r, "Trọng số RI cả công ty (%)", VALUATION_WEIGHTS["RI"], FMT_PCT, "", single=True, fill=assump_fill)
    RA["w_pspe"] = r; r = arow(wsA, r, "Trọng số P/S+P/E theo mảng (cả 4 mảng) (%)", VALUATION_WEIGHTS["PS_PE"], FMT_PCT, "Đã gộp 35%+15% cũ (bỏ P/S+P/B)", single=True, fill=assump_fill)
    wsA.column_dimensions['A'].width = 40
    print(f"[Excel] Sheet 02_Assumptions done ({r} dong).")

    # ─── 03a_Monthly_Data — Database THÁNG chi tiết từng mảng (theo yêu cầu user 2026-07) ─────────
    # Doanh thu tổng, % cơ cấu, doanh thu mảng (=%×tổng), số cửa hàng, hiệu quả/CH/tháng — TẤT CẢ có
    # công thức sống để kiểm soát tính toán. Cột "Ước tính 20XX" dùng đúng công thức user yêu cầu:
    # DT ước tính = SL ước tính × hiệu quả/CH × số tháng CÒN LẠI + DT LŨY KẾ đã có (năm hiện tại);
    # các năm sau (không có dữ liệu thật) = SL ước tính × hiệu quả/CH × 12 tháng.
    wsM = wb.create_sheet("03a_Monthly_Data")
    n_months = len(MONTHLY_MATRIX)
    col_month_start = 2  # cột B
    col_est = {yr: col_month_start + n_months + i for i, yr in enumerate(years_fc)}  # cột "Ước tính 20XX"

    month_headers = ["Chỉ tiêu"] + [m["label"] for m in MONTHLY_MATRIX] + [f"Ước tính {y}" for y in years_fc] + ["Ghi chú"]
    header_row(wsM, 1, month_headers, [26] + [9]*n_months + [13]*len(years_fc) + [40])

    def mcell(r, c, v, fmt=None, bold=False, fill=None):
        cell = wsM.cell(row=r, column=c, value=v)
        cell.font = bold_font if bold else data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        return cell

    ROW_TOTAL = 2
    ROW_PCT = {"TGDD": 3, "DMX": 4, "BHX": 5, "Khac": 6}
    ROW_REV = {"TGDD": 8, "DMX": 9, "BHX": 10, "Khac": 11}
    ROW_SL = {"TGDD": 13, "DMX": 14, "BHX": 15, "Khac": 16}
    ROW_EFF = {"TGDD": 18, "DMX": 19, "BHX": 20, "Khac": 21}

    wsM.cell(row=ROW_TOTAL, column=1, value="Doanh thu tổng (tỷ VND)").font = bold_font
    for seg in SEGMENTS:
        wsM.cell(row=ROW_PCT[seg], column=1, value=f"%Dthu {seg}").font = data_font
        wsM.cell(row=ROW_REV[seg], column=1, value=f"Doanh thu {seg}").font = data_font
        wsM.cell(row=ROW_SL[seg], column=1, value=f"SL {seg}").font = data_font
        wsM.cell(row=ROW_EFF[seg], column=1, value=f"DT/CH/th {seg}").font = data_font
    for rr in (ROW_TOTAL,) + tuple(ROW_PCT.values()) + tuple(ROW_REV.values()) + tuple(ROW_SL.values()) + tuple(ROW_EFF.values()):
        wsM.cell(row=rr, column=1).border = thin_border

    # Cột THÁNG THẬT — Doanh thu tổng/%/SL là INPUT CỨNG (dữ liệu IR thật, có thể thiếu — để trống nếu
    # không có nguồn tin cậy), Doanh thu mảng và Hiệu quả/CH LÀ CÔNG THỨC (không hardcode).
    for i, m in enumerate(MONTHLY_MATRIX):
        col = col_month_start + i
        mcell(ROW_TOTAL, col, round(m["revenue"], 1), '#,##0')
        col_l = get_column_letter(col)
        for seg in SEGMENTS:
            pct_val = m["pct"].get(seg)
            mcell(ROW_PCT[seg], col, round(pct_val, 4) if pct_val is not None else None, '0.00%')
            mcell(ROW_REV[seg], col, f"={col_l}${ROW_TOTAL}*{col_l}${ROW_PCT[seg]}" if pct_val is not None else None, '#,##0', fill=p_fill)
            sl_val = m["stores"].get(seg)
            mcell(ROW_SL[seg], col, sl_val, '#,##0')
            mcell(ROW_EFF[seg], col, f"={col_l}${ROW_REV[seg]}/{col_l}${ROW_SL[seg]}" if (pct_val is not None and sl_val) else None, '0.00', fill=p_fill)

    # Cột "Ước tính 20XX" — công thức đúng yêu cầu user: SL_ước tính × hiệu quả/CH × số tháng còn lại
    # + doanh thu lũy kế đã có (chỉ áp dụng năm HIỆN TẠI, đang có dữ liệu thật 1 phần trong năm); các
    # năm sau (không có dữ liệu thật) = SL_ước tính × hiệu quả/CH × 12 tháng trọn năm.
    seg_store_row_m = {"TGDD": RA["store_tgdd"], "DMX": RA["store_dmx"], "BHX": RA["store_bhx"], "Khac": RA["store_khac"]}
    # CURRENT_YEAR_REAL/_months_current_year_idx/_n_known_months_current_year đã tính SẴN ở phần dự
    # phóng Python (dùng chung 1 nguồn — xem ghi chú "DATABASE THEO THÁNG" phía trên).
    known_months_current_year = _months_current_year_idx
    n_known_months = _n_known_months_current_year

    for yi, yr in enumerate(years_fc):
        col = col_est[yr]
        col_l = get_column_letter(col)
        assump_col_l = get_column_letter(2 + N_HIST + yi)  # cột năm dự phóng tương ứng ở 02_Assumptions
        for seg in SEGMENTS:
            # SL ước tính — tham chiếu thẳng dự phóng đã có ở 02_Assumptions (đồng bộ 1 nguồn duy nhất)
            mcell(ROW_SL[seg], col, f"='02_Assumptions'!{assump_col_l}{seg_store_row_m[seg]}", '#,##0')
            # Hiệu quả/CH ước tính — năm hiện tại: TB 3 tháng thật gần nhất; các năm sau: hiệu quả năm
            # trước đó × (1 + tăng trưởng giả định) — tự tham chiếu ước tính năm trước, không lặp lại
            # hằng số Python để nếu user sửa giả định tăng trưởng ở Assumptions thì lan truyền đúng.
            if yr == CURRENT_YEAR_REAL:
                last3_cols = [get_column_letter(col_month_start + i) for i in known_months_current_year[-3:]]
                eff_formula = f"=AVERAGE({','.join(f'{c}{ROW_EFF[seg]}' for c in last3_cols)})"
            else:
                prev_col_l = get_column_letter(col_est[years_fc[yi - 1]])
                eff_formula = f"={prev_col_l}{ROW_EFF[seg]}*(1+'02_Assumptions'!$B${RA['eff_growth'][seg]})"
            mcell(ROW_EFF[seg], col, eff_formula, '0.00')
            # Doanh thu mảng ước tính
            if yr == CURRENT_YEAR_REAL:
                known_cols = [get_column_letter(col_month_start + i) for i in known_months_current_year]
                cum_known = "+".join(f"{c}{ROW_REV[seg]}" for c in known_cols)
                remaining_months = 12 - n_known_months
                rev_formula = f"={col_l}{ROW_SL[seg]}*{col_l}{ROW_EFF[seg]}*{remaining_months}+({cum_known})"
            else:
                rev_formula = f"={col_l}{ROW_SL[seg]}*{col_l}{ROW_EFF[seg]}*12"
            mcell(ROW_REV[seg], col, rev_formula, '#,##0', bold=True, fill=p_fill)
        # Doanh thu tổng ước tính = tổng 4 mảng
        mcell(ROW_TOTAL, col, "=" + "+".join(f"{col_l}{ROW_REV[seg]}" for seg in SEGMENTS), '#,##0', bold=True, fill=header_fill)

    wsM.column_dimensions['A'].width = 26
    wsM.freeze_panes = "B2"
    print(f"[Excel] Sheet 03a_Monthly_Data done ({n_months} thang + {len(years_fc)} cot uoc tinh).")

    # ─── 03_Revenue_Model — Driver-based theo từng mảng (skill ban-le mục 3) ──────────────────────
    wsR = wb.create_sheet("03_Revenue_Model")
    header_row(wsR, 1, ["Doanh thu theo mảng (tỷ VND)"] + YEAR_HEADERS + ["Ghi chú"], [30] + [11]*N_ALL + [40])
    RR = {}
    r = 2
    seg_store_row = {"TGDD": RA["store_tgdd"], "DMX": RA["store_dmx"], "BHX": RA["store_bhx"], "Khac": RA["store_khac"]}
    for seg in SEGMENTS:
        wsR.cell(row=r, column=1, value=f"── {seg} ──").font = bold_font
        r += 1
        RR.setdefault("store", {})[seg] = r
        _hist_vals = [None] * N_HIST  # số CH lịch sử không tách đủ mọi năm — chỉ hiển thị 2 năm gần nhất nếu có
        if seg != "Khac" and years_hist[-1] in STORE_COUNT_HIST:
            _hist_vals[-1] = STORE_COUNT_HIST[years_hist[-1]].get(seg)
        if seg != "Khac" and years_hist[-2] in STORE_COUNT_HIST:
            _hist_vals[-2] = STORE_COUNT_HIST[years_hist[-2]].get(seg)
        row_vals = _hist_vals + [f"='02_Assumptions'!{get_column_letter(2+N_HIST+i)}{seg_store_row[seg]}" for i in range(len(years_fc))]
        for j, v in enumerate(row_vals):
            c = wsR.cell(row=r, column=2+j, value=v); c.font = data_font; c.border = thin_border
            c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
        wsR.cell(row=r, column=1, value="  Số cửa hàng cuối kỳ").font = data_font
        r += 1
        RR.setdefault("eff", {})[seg] = r
        # Hiệu quả DT/CH dự phóng — LẤY TỪ sheet 03a_Monthly_Data (cột "Ước tính 20XX"), KHÔNG tính
        # độc lập ở đây nữa (theo yêu cầu user: 1 nguồn database tháng duy nhất, có công thức kiểm
        # soát được, thay vì Python tính sẵn hằng số rồi rải ra nhiều nơi).
        eff_row = [None]*N_HIST + [f"='03a_Monthly_Data'!{get_column_letter(col_est[years_fc[i]])}{ROW_EFF[seg]}" for i in range(len(years_fc))]
        for j, v in enumerate(eff_row):
            c = wsR.cell(row=r, column=2+j, value=v); c.font = data_font; c.border = thin_border
            c.number_format = '0.000'; c.alignment = Alignment(horizontal='right')
        wsR.cell(row=r, column=1, value="  Hiệu quả DT/CH/tháng (tỷ VND)").font = data_font
        r += 1
        RR.setdefault("rev", {})[seg] = r
        rev_hist_seg = [None]*N_HIST
        for i, y in enumerate(years_hist):
            if y in REV_SEGMENT_HIST:
                rev_hist_seg[i] = REV_SEGMENT_HIST[y].get(seg)
        # Doanh thu mảng dự phóng — LẤY TỪ sheet 03a_Monthly_Data (cột "Ước tính 20XX", đã tính đúng
        # công thức "SL ước tính × hiệu quả/CH × số tháng còn lại + lũy kế đã có").
        rev_row = rev_hist_seg[:] + [f"='03a_Monthly_Data'!{get_column_letter(col_est[years_fc[i]])}{ROW_REV[seg]}" for i in range(len(years_fc))]
        for j, v in enumerate(rev_row):
            c = wsR.cell(row=r, column=2+j, value=v); c.font = bold_font; c.border = thin_border
            c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
            c.fill = p_fill
        wsR.cell(row=r, column=1, value="  Doanh thu mảng").font = bold_font
        r += 1
    RR["total"] = r
    wsR.cell(row=r, column=1, value="TỔNG DOANH THU").font = bold_font
    # Cộng trực tiếp từng ô mảng (an toàn hơn SUM theo range vì các dòng mảng xen kẽ store/eff/rev)
    for j in range(N_ALL):
        col = get_column_letter(2+j)
        formula = "=" + "+".join(f"{col}{RR['rev'][seg]}" for seg in SEGMENTS)
        c = wsR.cell(row=r, column=2+j, value=formula)
        c.font = bold_font; c.border = thin_border; c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right'); c.fill = header_fill; c.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    r += 1
    RR["yoy"] = r
    wsR.cell(row=r, column=1, value="Tăng trưởng DT YoY (%)").font = data_font
    for j in range(1, N_ALL):
        col, prev_col = get_column_letter(2+j), get_column_letter(1+j)
        c = wsR.cell(row=r, column=2+j, value=f"={col}{RR['total']}/{prev_col}{RR['total']}-1")
        c.font = data_font; c.border = thin_border; c.number_format = '0.0%'; c.alignment = Alignment(horizontal='right')
    wsR.column_dimensions['A'].width = 30
    print(f"[Excel] Sheet 03_Revenue_Model done ({r} dong).")

    # ─── 04_PnL — 5 năm lịch sử + 3 năm dự phóng ──────────────────────────────────────────────────
    wsP = wb.create_sheet("04_PnL")
    header_row(wsP, 1, ["Chỉ tiêu (tỷ VND)"] + YEAR_HEADERS + ["Ghi chú"], [30] + [11]*N_ALL + [40])
    RP = {}
    def prow(r, label, hist_vals, fc_formula_fn, fmt='#,##0', bold=False, fill=None, note=""):
        """fc_formula_fn(i, col) -> formula string cho năm dự phóng thứ i (col=chữ cột hiện tại)."""
        wsP.cell(row=r, column=1, value=label).font = bold_font if bold else data_font
        wsP.cell(row=r, column=1).border = thin_border
        for j, v in enumerate(hist_vals):
            c = wsP.cell(row=r, column=2+j, value=v); c.border = thin_border
            c.font = bold_font if bold else data_font; c.number_format = fmt
            c.alignment = Alignment(horizontal='right')
            if fill: c.fill = fill
        for i in range(len(years_fc)):
            col = get_column_letter(2+N_HIST+i)
            c = wsP.cell(row=r, column=2+N_HIST+i, value=fc_formula_fn(i, col))
            c.border = thin_border; c.font = bold_font if bold else data_font; c.number_format = fmt
            c.alignment = Alignment(horizontal='right')
            if fill: c.fill = fill
        wsP.cell(row=r, column=2+N_ALL, value=note).font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
        return r + 1

    r = 2
    RP["revenue"] = r
    r = prow(r, "Doanh thu thuần", revenue_hist, lambda i, col: f"='03_Revenue_Model'!{col}{RR['total']}", bold=True, note="Link 03_Revenue_Model")
    RP["yoy"] = r
    r = prow(r, "  Tăng trưởng YoY (%)", [None] + [round(revenue_hist[i]/revenue_hist[i-1]-1, 4) for i in range(1, N_HIST)],
              lambda i, col: f"={col}{RP['revenue']}/{get_column_letter(1+N_HIST+i)}{RP['revenue']}-1", fmt='0.0%')
    RP["cogs"] = r
    r = prow(r, "Giá vốn hàng bán", [-v for v in cogs_hist],
              lambda i, col: f"=-{col}{RP['revenue']}*(1-'02_Assumptions'!{col}{RA['gpm']})")
    RP["gp"] = r
    r = prow(r, "Lợi nhuận gộp", gp_hist, lambda i, col: f"={col}{RP['revenue']}+{col}{RP['cogs']}", bold=True, fill=p_fill)
    RP["gpm"] = r
    r = prow(r, "  Biên LNG (%)", [round(g, 2)/100 for g in gpm_hist], lambda i, col: f"={col}{RP['gp']}/{col}{RP['revenue']}", fmt='0.00%')
    RP["sga"] = r
    r = prow(r, "Chi phí bán hàng & QLDN", [-v for v in sga_hist],
              lambda i, col: f"=-{col}{RP['revenue']}*'02_Assumptions'!{col}{RA['sga']}")
    RP["ebit"] = r
    r = prow(r, "EBIT", ebit_hist, lambda i, col: f"={col}{RP['gp']}+{col}{RP['sga']}", bold=True, fill=p_fill)
    RP["ebit_margin"] = r
    r = prow(r, "  Biên EBIT (%)", [round(ebit_hist[i]/revenue_hist[i], 4) for i in range(N_HIST)],
              lambda i, col: f"={col}{RP['ebit']}/{col}{RP['revenue']}", fmt='0.00%')
    RP["da"] = r
    r = prow(r, "D&A", da_hist, lambda i, col: f"={col}{RP['revenue']}*'02_Assumptions'!$B${RA['da_pct']}")
    RP["ebitda"] = r
    r = prow(r, "EBITDA", ebitda_hist, lambda i, col: f"={col}{RP['ebit']}+{col}{RP['da']}", bold=True, fill=p_fill)
    RP["net_fin"] = r
    r = prow(r, "DT/CP tài chính ròng", [round(fin_income_hist[i]-fin_expense_hist[i], 1) for i in range(N_HIST)],
              lambda i, col: f"={col}{RP['revenue']}*'02_Assumptions'!$B${RA['netfin']}")
    RP["ebt"] = r
    r = prow(r, "Lợi nhuận trước thuế (EBT)", ebt_hist, lambda i, col: f"={col}{RP['ebit']}+{col}{RP['net_fin']}", bold=True)
    RP["tax"] = r
    r = prow(r, "Thuế TNDN", [round(ebt_hist[i]-ni_consol_hist[i], 1) for i in range(N_HIST)],
              lambda i, col: f"={col}{RP['ebt']}*'02_Assumptions'!$B${RA['tax']}")
    RP["ni_consol"] = r
    r = prow(r, "LNST hợp nhất", ni_consol_hist, lambda i, col: f"={col}{RP['ebt']}-{col}{RP['tax']}", bold=True, fill=p_fill)
    RP["ni"] = r
    _minority_hist_pct = [ni_hist[i]/ni_consol_hist[i] if ni_consol_hist[i] else 1 for i in range(N_HIST)]
    r = prow(r, "LNST Cổ đông Công ty mẹ", ni_hist,
              lambda i, col: f"={col}{RP['ni_consol']}*{MINORITY_RATIO}", bold=True, fill=p_fill,
              note=f"Tỷ lệ phân bổ cổ đông mẹ ~{MINORITY_RATIO*100:.1f}% (TB 2 năm gần nhất)")
    RP["eps"] = r
    r = prow(r, "EPS (VND/CP)", eps_hist, lambda i, col: f"={col}{RP['ni']}*1e9/('02_Assumptions'!$B${RA['shares']}*1e6)", fmt='#,##0', bold=True)
    wsP.column_dimensions['A'].width = 30
    print(f"[Excel] Sheet 04_PnL done ({r} dong).")

    # ─── 05_Balance_Sheet ───────────────────────────────────────────────────────────────────────
    wsB = wb.create_sheet("05_Balance_Sheet")
    header_row(wsB, 1, ["Chỉ tiêu (tỷ VND)"] + YEAR_HEADERS + ["Ghi chú"], [30] + [11]*N_ALL + [40])
    RBS = {}
    def brow(r, label, hist_vals, fc_formula_fn, fmt='#,##0', bold=False, fill=None, note=""):
        wsB.cell(row=r, column=1, value=label).font = bold_font if bold else data_font
        wsB.cell(row=r, column=1).border = thin_border
        for j, v in enumerate(hist_vals):
            c = wsB.cell(row=r, column=2+j, value=v); c.border = thin_border
            c.font = bold_font if bold else data_font; c.number_format = fmt
            c.alignment = Alignment(horizontal='right')
            if fill: c.fill = fill
        for i in range(len(years_fc)):
            col = get_column_letter(2+N_HIST+i)
            c = wsB.cell(row=r, column=2+N_HIST+i, value=fc_formula_fn(i, col))
            c.border = thin_border; c.font = bold_font if bold else data_font; c.number_format = fmt
            c.alignment = Alignment(horizontal='right')
            if fill: c.fill = fill
        wsB.cell(row=r, column=2+N_ALL, value=note).font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
        return r + 1

    r = 2
    wsB.cell(row=r, column=1, value="TÀI SẢN").font = bold_font; r += 1
    RBS["cash"] = r
    r = brow(r, "Tiền & tương đương tiền", cash_hist, lambda i, col: f"={get_column_letter(1+N_HIST+i)}{RBS['cash']}*(1+'02_Assumptions'!$B${RA['cash_yoy']})")
    RBS["inventory"] = r
    r = brow(r, "Hàng tồn kho", inventory_hist, lambda i, col: f"=-'04_PnL'!{col}{RP['cogs']}*'02_Assumptions'!$B${RA['dio']}/365")
    RBS["receivables"] = r
    r = brow(r, "Phải thu khách hàng", receivables_hist, lambda i, col: f"='04_PnL'!{col}{RP['revenue']}*'02_Assumptions'!$B${RA['dso']}/365")
    RBS["other_assets"] = r
    r = brow(r, "Tài sản khác", other_assets_hist, lambda i, col: f"='04_PnL'!{col}{RP['revenue']}*{OTHER_ASSETS_PCT_FC}")
    RBS["total_assets"] = r
    r = brow(r, "TỔNG TÀI SẢN", total_assets_hist,
              lambda i, col: f"={col}{RBS['cash']}+{col}{RBS['inventory']}+{col}{RBS['receivables']}+{col}{RBS['other_assets']}",
              bold=True, fill=header_fill)
    r += 1
    wsB.cell(row=r, column=1, value="NGUỒN VỐN").font = bold_font; r += 1
    RBS["debt"] = r
    r = brow(r, "Nợ vay (NH+DH)", debt_hist, lambda i, col: f"={get_column_letter(1+N_HIST+i)}{RBS['debt']}*(1+'02_Assumptions'!$B${RA['debt_yoy']})")
    RBS["payables"] = r
    r = brow(r, "Phải trả người bán", payables_hist, lambda i, col: f"=-'04_PnL'!{col}{RP['cogs']}*'02_Assumptions'!$B${RA['dpo']}/365")
    RBS["equity"] = r
    r = brow(r, "Vốn chủ sở hữu", equity_hist,
              lambda i, col: f"={get_column_letter(1+N_HIST+i)}{RBS['equity']}+'04_PnL'!{col}{RP['ni_consol']}*(1-'02_Assumptions'!$B${RA['payout']})",
              bold=True)
    RBS["other_liab"] = r
    # "Nợ khác" là PLUG bắt buộc cân bằng (Tổng NV = Tổng TS) — Debt/Payables/Equity đều được chiếu
    # ĐỘC LẬP theo driver riêng (YoY/DSO/retained earnings), nên KHÔNG tự động cộng khớp Tổng TS nếu
    # để "Nợ khác" tự chiếu theo %DT riêng (đã gây lệch 8.000-17.000 tỷ tăng dần khi test COM Excel).
    r = brow(r, "Nợ khác (plug cân bằng)", other_liab_hist,
              lambda i, col: f"={col}{RBS['total_assets']}-{col}{RBS['debt']}-{col}{RBS['payables']}-{col}{RBS['equity']}")
    RBS["total_liab_eq"] = r
    r = brow(r, "TỔNG NGUỒN VỐN", total_assets_hist,
              lambda i, col: f"={col}{RBS['debt']}+{col}{RBS['payables']}+{col}{RBS['other_liab']}+{col}{RBS['equity']}",
              bold=True, fill=header_fill)
    r += 1
    RBS["net_debt"] = r
    r = brow(r, "Net Debt (Nợ vay - Tiền mặt)", NET_DEBT_HIST, lambda i, col: f"={col}{RBS['debt']}-{col}{RBS['cash']}")
    RBS["de"] = r
    r = brow(r, "D/E (Net Debt/VCSH)", [round(NET_DEBT_HIST[i]/equity_hist[i], 3) for i in range(N_HIST)],
              lambda i, col: f"={col}{RBS['net_debt']}/{col}{RBS['equity']}", fmt='0.00"x"')
    wsB.column_dimensions['A'].width = 30
    print(f"[Excel] Sheet 05_Balance_Sheet done ({r} dong).")

    # ─── 06_Cash_Flow — LCTT đơn giản + FCFF ───────────────────────────────────────────────────
    wsCF = wb.create_sheet("06_Cash_Flow")
    header_row(wsCF, 1, ["Chỉ tiêu (tỷ VND)"] + [f"{y}E" if i == 0 else f"{y}F" for i, y in enumerate(years_fc)] + ["Ghi chú"],
               [30] + [12]*len(years_fc) + [40])
    RCF = {}
    r = 2
    RCF["ni"] = r
    wsCF.cell(row=r, column=1, value="LNST hợp nhất").font = data_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+N_HIST+i); col_out = get_column_letter(2+i)
        c = wsCF.cell(row=r, column=2+i, value=f"='04_PnL'!{col}{RP['ni_consol']}")
        c.border = thin_border; c.font = data_font; c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
    r += 1
    RCF["da"] = r
    wsCF.cell(row=r, column=1, value="+ D&A").font = data_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+N_HIST+i)
        c = wsCF.cell(row=r, column=2+i, value=f"='04_PnL'!{col}{RP['da']}")
        c.border = thin_border; c.font = data_font; c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
    r += 1
    RCF["wc"] = r
    wsCF.cell(row=r, column=1, value="+/- Thay đổi vốn lưu động").font = data_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+N_HIST+i); prev_col = get_column_letter(1+N_HIST+i)
        formula = (f"=('05_Balance_Sheet'!{prev_col}{RBS['inventory']}-'05_Balance_Sheet'!{col}{RBS['inventory']})"
                   f"+('05_Balance_Sheet'!{prev_col}{RBS['receivables']}-'05_Balance_Sheet'!{col}{RBS['receivables']})"
                   f"+('05_Balance_Sheet'!{col}{RBS['payables']}-'05_Balance_Sheet'!{prev_col}{RBS['payables']})")
        c = wsCF.cell(row=r, column=2+i, value=formula)
        c.border = thin_border; c.font = data_font; c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
    r += 1
    RCF["cfo"] = r
    wsCF.cell(row=r, column=1, value="Dòng tiền HĐKD (CFO)").font = bold_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+i)
        c = wsCF.cell(row=r, column=2+i, value=f"={col}{RCF['ni']}+{col}{RCF['da']}+{col}{RCF['wc']}")
        c.border = thin_border; c.font = bold_font; c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right'); c.fill = p_fill
    r += 1
    RCF["capex"] = r
    wsCF.cell(row=r, column=1, value="CAPEX (CFI)").font = data_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+N_HIST+i)
        c = wsCF.cell(row=r, column=2+i, value=f"=-'04_PnL'!{col}{RP['revenue']}*'02_Assumptions'!$B${RA['capex_pct']}")
        c.border = thin_border; c.font = data_font; c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
    r += 1
    RCF["fcff"] = r
    wsCF.cell(row=r, column=1, value="FCFF (CFO+CAPEX)").font = bold_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+i)
        c = wsCF.cell(row=r, column=2+i, value=f"={col}{RCF['cfo']}+{col}{RCF['capex']}")
        c.border = thin_border; c.font = bold_font; c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right'); c.fill = p_fill
    r += 1
    RCF["fcf_yield"] = r
    wsCF.cell(row=r, column=1, value="FCF Yield (%)").font = data_font; wsCF.cell(row=r, column=1).border = thin_border
    for i in range(len(years_fc)):
        col = get_column_letter(2+i)
        c = wsCF.cell(row=r, column=2+i, value=f"={col}{RCF['fcff']}/'02_Assumptions'!$B${RA['mcap']}")
        c.border = thin_border; c.font = data_font; c.number_format = '0.00%'; c.alignment = Alignment(horizontal='right')
    wsCF.column_dimensions['A'].width = 30
    print(f"[Excel] Sheet 06_Cash_Flow done ({r} dong).")

    # ─── 07_Valuation — 5 phương pháp (skill ban-le), viết DETAIL trước (từ dòng 17), rồi SUMMARY (2-14) ──
    wsV = wb.create_sheet("07_Valuation")
    wsV.column_dimensions['A'].width = 38
    wsV.column_dimensions['B'].width = 16
    wsV.column_dimensions['C'].width = 16
    wsV.column_dimensions['D'].width = 16
    wsV.column_dimensions['E'].width = 46

    def vcell(r, c, val, fmt=None, bold=False, fill=None, align='right'):
        cell = wsV.cell(row=r, column=c, value=val)
        cell.font = bold_font if bold else data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        return cell

    r = 17
    # 1) P/E MEDIAN
    vcell(r, 1, "1) P/E MEDIAN (trọng số 20%)", bold=True); r += 1
    row_pe_mul = r; vcell(r, 1, "P/E median lịch sử (x)"); vcell(r, 2, f"='02_Assumptions'!$B${RA['pe_median']}", '0.00"x"'); r += 1
    row_eps26 = r; vcell(r, 1, "EPS 2026E (VND/CP)"); vcell(r, 2, f"='04_PnL'!{get_column_letter(2+N_HIST)}{RP['eps']}", '#,##0'); r += 1
    row_pe_target = r; vcell(r, 1, "→ P/E Target Price", bold=True); vcell(r, 2, f"=B{row_pe_mul}*B{row_eps26}", '#,##0', bold=True, fill=p_fill); r += 2

    # 2) P/B MEDIAN
    vcell(r, 1, "2) P/B MEDIAN (trọng số 20%)", bold=True); r += 1
    row_pb_mul = r; vcell(r, 1, "P/B median lịch sử (x)"); vcell(r, 2, f"='02_Assumptions'!$B${RA['pb_median']}", '0.00"x"'); r += 1
    row_bvps26 = r; vcell(r, 1, "BVPS 2026E (VND/CP)")
    vcell(r, 2, f"='05_Balance_Sheet'!{get_column_letter(2+N_HIST)}{RBS['equity']}*1e9/('02_Assumptions'!$B${RA['shares']}*1e6)", '#,##0'); r += 1
    row_pb_target = r; vcell(r, 1, "→ P/B Target Price", bold=True); vcell(r, 2, f"=B{row_pb_mul}*B{row_bvps26}", '#,##0', bold=True, fill=p_fill); r += 2

    # 3) RESIDUAL INCOME cả công ty
    vcell(r, 1, "3) RESIDUAL INCOME — cả công ty (trọng số 10%)", bold=True); r += 1
    row_coe = r; vcell(r, 1, "Chi phí vốn CSH (COE)"); vcell(r, 2, f"='02_Assumptions'!$B${RA['coe']}", '0.00%'); r += 1
    row_g = r; vcell(r, 1, "Tăng trưởng dài hạn (g)"); vcell(r, 2, f"='02_Assumptions'!$B${RA['terminal_g']}", '0.00%'); r += 1
    row_bvps0 = r; vcell(r, 1, "BVPS hiện tại (gốc)")
    vcell(r, 2, f"='05_Balance_Sheet'!{get_column_letter(1+N_HIST)}{RBS['equity']}*1e9/('02_Assumptions'!$B${RA['shares']}*1e6)", '#,##0'); r += 1
    row_eps_fc3 = r
    vcell(r, 1, "EPS dự phóng 3 năm (VND)")
    for i in range(len(years_fc)):
        vcell(r, 2+i, f"='04_PnL'!{get_column_letter(2+N_HIST+i)}{RP['eps']}", '#,##0')
    r += 1
    row_bvps_fc3 = r
    vcell(r, 1, "BVPS đầu kỳ mỗi năm (VND)")
    vcell(r, 2, f"=B{row_bvps0}")
    for i in range(1, len(years_fc)):
        vcell(r, 2+i, f"={get_column_letter(2+i-1)}{row_bvps_fc3}+{get_column_letter(2+i-1)}{row_eps_fc3}*(1-'02_Assumptions'!$B${RA['payout']})")
    for i in range(len(years_fc)):
        wsV.cell(row=r, column=2+i).number_format = '#,##0'
    r += 1
    row_cc = r
    vcell(r, 1, "Capital Charge (BVPS đầu kỳ × COE)")
    for i in range(len(years_fc)):
        vcell(r, 2+i, f"={get_column_letter(2+i)}{row_bvps_fc3}*$B${row_coe}", '#,##0')
    r += 1
    row_ri = r
    vcell(r, 1, "Residual Income (EPS - Capital Charge)")
    for i in range(len(years_fc)):
        vcell(r, 2+i, f"={get_column_letter(2+i)}{row_eps_fc3}-{get_column_letter(2+i)}{row_cc}", '#,##0')
    r += 1
    row_pvri = r
    vcell(r, 1, "PV of RI")
    for i in range(len(years_fc)):
        vcell(r, 2+i, f"={get_column_letter(2+i)}{row_ri}/(1+$B${row_coe})^{i+1}", '#,##0')
    r += 1
    row_pvri_sum = r
    vcell(r, 1, "Tổng PV(RI)")
    vcell(r, 2, f"=SUM(B{row_pvri}:{get_column_letter(1+len(years_fc))}{row_pvri})", '#,##0')
    r += 1
    row_cv = r
    vcell(r, 1, "Continuing Value (CV)")
    vcell(r, 2, f"={get_column_letter(1+len(years_fc))}{row_ri}*(1+$B${row_g})/($B${row_coe}-$B${row_g})", '#,##0')
    r += 1
    row_pvcv = r
    vcell(r, 1, "PV of CV")
    vcell(r, 2, f"=B{row_cv}/(1+$B${row_coe})^{len(years_fc)}", '#,##0')
    r += 1
    row_ri_target = r
    vcell(r, 1, "→ RI Target Price", bold=True)
    vcell(r, 2, f"=B{row_bvps0}+B{row_pvri_sum}+B{row_pvcv}", '#,##0', bold=True, fill=p_fill)
    r += 2

    # 4) P/S + P/E theo mảng
    pe_segs = [s for s in SEGMENTS if SEGMENT_METHOD[s] == "PE"]
    vcell(r, 1, f"4) P/S + P/E theo mảng (trọng số 35%) — {', '.join(pe_segs)}", bold=True); r += 1
    row_pspe_hdr = r
    vcell(r, 1, "Mảng"); vcell(r, 2, "DT 2026E (tỷ)"); vcell(r, 3, "Biên LNST (%)"); vcell(r, 4, "P/E áp dụng (x)")
    for c in range(1, 5): wsV.cell(row=r, column=c).font = header_font; wsV.cell(row=r, column=c).fill = header_fill
    r += 1
    row_pspe_seg = {}
    for seg in pe_segs:
        row_pspe_seg[seg] = r
        vcell(r, 1, seg)
        vcell(r, 2, f"='03_Revenue_Model'!{get_column_letter(2+N_HIST)}{RR['rev'][seg]}", '#,##0')
        vcell(r, 3, round(SEGMENT_NET_MARGIN[seg], 4), '0.00%')
        vcell(r, 4, round(min(SEGMENT_PE_REF[seg], max(0.9*SEGMENT_GROWTH_PCT[seg], 5.0)), 2), '0.00"x"')
        r += 1
    row_pspe_total = r
    vcell(r, 1, "→ P/S+P/E Target Price (tổng giá trị mảng / SL CP)", bold=True)
    _terms = "+".join(f"B{row_pspe_seg[s]}*C{row_pspe_seg[s]}*D{row_pspe_seg[s]}" for s in pe_segs)
    vcell(r, 2, f"=({_terms})*1e9/('02_Assumptions'!$B${RA['shares']}*1e6)", '#,##0', bold=True, fill=p_fill)
    row_pspe_target = r

    # ── SUMMARY (rows 1-13, viết SAU khi đã biết vị trí các dòng detail) — 4 phương pháp (đã bỏ P/S+P/B) ──
    wsV.merge_cells('A1:E1')
    vcell(1, 1, "ĐỊNH GIÁ MWG — 4 PHƯƠNG PHÁP TỔNG HỢP (skill ban-le)", bold=True, align='center')
    wsV['A1'].font = title_font
    vcell(2, 1, "Chi phí vốn CSH (COE)"); vcell(2, 2, f"=B{row_coe}", '0.00%')
    vcell(3, 1, "Tăng trưởng dài hạn (g)"); vcell(3, 2, f"=B{row_g}", '0.00%')
    vcell(5, 1, "Phương pháp", bold=True); vcell(5, 2, "Trọng số", bold=True); vcell(5, 3, "Giá mục tiêu (VND)", bold=True)
    for c in range(1, 4): wsV.cell(row=5, column=c).font = header_font; wsV.cell(row=5, column=c).fill = header_fill
    vcell(6, 1, "P/E Median"); vcell(6, 2, f"='02_Assumptions'!$B${RA['w_pe']}", '0.0%'); vcell(6, 3, f"=B{row_pe_target}", '#,##0')
    vcell(7, 1, "P/B Median"); vcell(7, 2, f"='02_Assumptions'!$B${RA['w_pb']}", '0.0%'); vcell(7, 3, f"=B{row_pb_target}", '#,##0')
    vcell(8, 1, "Residual Income (cả công ty)"); vcell(8, 2, f"='02_Assumptions'!$B${RA['w_ri']}", '0.0%'); vcell(8, 3, f"=B{row_ri_target}", '#,##0')
    vcell(9, 1, "P/S+P/E theo mảng (cả 4 mảng)"); vcell(9, 2, f"='02_Assumptions'!$B${RA['w_pspe']}", '0.0%'); vcell(9, 3, f"=B{row_pspe_target}", '#,##0')
    vcell(11, 1, "GIÁ MỤC TIÊU (bình quân trọng số)", bold=True)
    vcell(11, 3, "=SUMPRODUCT(B6:B9,C6:C9)", '#,##0', bold=True, fill=header_fill)
    vcell(12, 1, "Giá hiện tại"); vcell(12, 3, "='02_Assumptions'!$B$2", '#,##0')
    vcell(13, 1, "Upside/Downside", bold=True); vcell(13, 3, "=C11/C12-1", '0.0%', bold=True, fill=p_fill)
    print(f"[Excel] Sheet 07_Valuation done ({row_pspe_target} dong).")

    # ─── 08_Sensitivity — ma trận nhạy cảm 2 biến (P/E×EPS, P/B×BVPS), 5×5, tô màu theo giá ─────
    from openpyxl.formatting.rule import CellIsRule
    wsS = wb.create_sheet("08_Sensitivity")
    wsS.column_dimensions['A'].width = 14
    for col in "BCDEF":
        wsS.column_dimensions[col].width = 14

    def vcellS(r, c, v, fmt=None, bold=False, fill=None, align='center'):
        cell = wsS.cell(row=r, column=c, value=v)
        cell.font = bold_font if bold else data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        return cell

    wsS.merge_cells('A1:F1')
    vcellS(1, 1, "MA TRẬN NHẠY CẢM: P/E × EPS 2026E", bold=True)
    wsS['A1'].font = bold_font
    vcellS(2, 1, "P/E \\ EPS", bold=True, fill=header_fill)
    wsS['A2'].font = header_font
    _pe_steps = [-2, -1, 0, 1, 2]
    _eps_steps = [-0.10, -0.05, 0, 0.05, 0.10]
    for j, de in enumerate(_eps_steps):
        c = vcellS(2, 2+j, f"='07_Valuation'!$B${row_eps26}*(1+{de})", '#,##0', fill=header_fill)
        c.font = header_font
    for i, dpe in enumerate(_pe_steps):
        rr = 3 + i
        c = vcellS(rr, 1, f"='07_Valuation'!$B${row_pe_mul}+{dpe}", '0.00"x"', fill=header_fill)
        c.font = header_font
        for j in range(5):
            col = get_column_letter(2+j)
            vcellS(rr, 2+j, f"=$A{rr}*{col}$2", '#,##0')
    for rr in range(3, 8):
        for col in "BCDEF":
            wsS.conditional_formatting.add(f"{col}{rr}",
                CellIsRule(operator='lessThan', formula=["'02_Assumptions'!$B$2"], fill=n_fill))
            wsS.conditional_formatting.add(f"{col}{rr}",
                CellIsRule(operator='greaterThan', formula=["'07_Valuation'!$C$12"], fill=p_fill))

    r0 = 10
    wsS.merge_cells(f'A{r0}:F{r0}')
    vcellS(r0, 1, "MA TRẬN NHẠY CẢM: P/B × BVPS 2026E", bold=True)
    wsS[f'A{r0}'].font = bold_font
    vcellS(r0+1, 1, "P/B \\ BVPS", bold=True, fill=header_fill)
    wsS[f'A{r0+1}'].font = header_font
    _pb_steps = [-0.5, -0.25, 0, 0.25, 0.5]
    _bvps_steps = [-0.10, -0.05, 0, 0.05, 0.10]
    for j, db in enumerate(_bvps_steps):
        c = vcellS(r0+1, 2+j, f"='07_Valuation'!$B${row_bvps26}*(1+{db})", '#,##0', fill=header_fill)
        c.font = header_font
    for i, dpb in enumerate(_pb_steps):
        rr = r0+2+i
        c = vcellS(rr, 1, f"='07_Valuation'!$B${row_pb_mul}+{dpb}", '0.00"x"', fill=header_fill)
        c.font = header_font
        for j in range(5):
            col = get_column_letter(2+j)
            vcellS(rr, 2+j, f"=$A{rr}*{col}${r0+1}", '#,##0')
    for rr in range(r0+2, r0+7):
        for col in "BCDEF":
            wsS.conditional_formatting.add(f"{col}{rr}",
                CellIsRule(operator='lessThan', formula=["'02_Assumptions'!$B$2"], fill=n_fill))
            wsS.conditional_formatting.add(f"{col}{rr}",
                CellIsRule(operator='greaterThan', formula=["'07_Valuation'!$C$12"], fill=p_fill))
    print("[Excel] Sheet 08_Sensitivity done.")

    # ─── 09_PESTLE ──────────────────────────────────────────────────────────────────────────────
    wsPE = wb.create_sheet("09_PESTLE")
    header_row(wsPE, 1, ["Yếu tố", "Nội dung", "Tác động", "Mức độ"], [16, 70, 40, 14])
    pestle_data = [
        ("Political", "Nhà nước kiểm soát chặt giá thuốc (An Khang) và an toàn thực phẩm (BHX); chính sách thuế VAT hàng điện máy/ICT ổn định, không có thay đổi lớn gần đây.", "Trung tính — không có rủi ro chính sách đột biến trong ngắn hạn.", "Trung tính"),
        ("Economic", "Sức mua hồi phục sau giai đoạn khó khăn 2022-2023 (lãi suất cao, thất nghiệp), CPI ổn định quanh 3-4%/năm hỗ trợ tiêu dùng không thiết yếu (ICT/điện máy) tăng trở lại.", "Tích cực — SSSG TGDĐ/ĐMX +33% (5T2026) phản ánh sức mua phục hồi rõ rệt.", "Tích cực"),
        ("Social", "Xu hướng mua sắm hiện đại hoá (siêu thị mini, tiện lợi) thay thế chợ truyền thống hỗ trợ BHX; tầng lớp trung lưu tăng thúc đẩy nhu cầu điện thoại/laptop cao cấp hơn.", "Tích cực — dư địa dài hạn cho BHX thay thế kênh GT truyền thống.", "Tích cực"),
        ("Technological", "Chu kỳ nâng cấp iPhone/Samsung, xu hướng AI trên thiết bị di động thúc đẩy nhu cầu thay máy; MWG đầu tư Super App/thanh toán trả chậm để tăng biên lợi nhuận dịch vụ tài chính.", "Tích cực — doanh thu Super App/trả chậm tăng trưởng 2 chữ số.", "Tích cực"),
        ("Legal", "Yêu cầu hoá đơn điện tử/truy xuất nguồn gốc thực phẩm ngày càng chặt với BHX/An Khang — chi phí tuân thủ tăng nhưng cũng là rào cản gia nhập với đối thủ nhỏ lẻ.", "Trung tính — chi phí tuân thủ tăng nhẹ, bù lại bằng lợi thế quy mô.", "Trung tính"),
        ("Environmental", "Áp lực giảm rác thải nhựa/bao bì tại chuỗi BHX, yêu cầu chuỗi lạnh (cold chain) tiêu tốn năng lượng cho hàng tươi sống — chi phí vận hành có thể tăng theo quy định môi trường mới.", "Tiêu cực nhẹ — chi phí tuân thủ môi trường tăng dần theo thời gian.", "Tiêu cực"),
    ]
    for i, (factor, content, impact_desc, impact) in enumerate(pestle_data, 2):
        data_row(wsPE, i, [factor, content, impact_desc, impact])
        wsPE.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        wsPE.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        wsPE.row_dimensions[i].height = 60
    print("[Excel] Sheet 09_PESTLE done.")

    # ─── 10_Leading_Indicators ──────────────────────────────────────────────────────────────────
    wsLI = wb.create_sheet("10_Leading_Indicators")
    header_row(wsLI, 1, ["Chỉ báo", "Ngưỡng tích cực", "Ngưỡng tiêu cực", "Giá trị hiện tại", "Trạng thái"], [30, 22, 22, 20, 16])
    _sssg_now = 0.33  # SSSG toàn bộ chuỗi VN, IR "4-5 tháng 2026"
    li_data = [
        ("SSSG (tăng trưởng cửa hàng hiện hữu)", "> 10%", "< 0%", f"{_sssg_now*100:.0f}%", "Tích cực"),
        ("Tăng trưởng DT YoY 5T2026", "> 15%", "< 5%", "29.3%", "Tích cực"),
        ("Số cửa hàng BHX mở mới/năm", "> 800", "< 300", f"~{STORE_COUNT_NOW['BHX']-2559}/5T", "Tích cực"),
        ("Biên LNG (%)", "> 20%", "< 17%", f"{GPM_FC_PCT:.1f}%", "Tích cực"),
        ("SG&A/DT (%)", "< 14%", "> 17%", f"{SGA_FC_PCT[0]:.1f}%", "Theo dõi"),
        ("Tỷ trọng DT online/tổng", "> 12%", "< 8%", "~11%", "Trung tính"),
        ("Chu kỳ iPhone mới (tháng 9 hàng năm)", "Ra mắt đúng hẹn", "Trì hoãn/thiếu hàng", "Đang chuẩn bị (T9/2026)", "Theo dõi"),
        ("CPI Việt Nam YoY", "2-4%", "> 5% hoặc < 0%", "~3.5%", "Tích cực"),
    ]
    for i, row in enumerate(li_data, 2):
        data_row(wsLI, i, row)
        fill = p_fill if row[-1] == "Tích cực" else (n_fill if row[-1] == "Tiêu cực" else None)
        if fill:
            wsLI.cell(row=i, column=5).fill = fill
    print("[Excel] Sheet 10_Leading_Indicators done.")

    # ─── 11_Investment_Thesis ───────────────────────────────────────────────────────────────────
    wsIT = wb.create_sheet("11_Investment_Thesis")
    header_row(wsIT, 1, ["Tầng đánh giá", "Kết luận", "Rating"], [26, 80, 14])
    thesis_tiers = [
        ("1. Mô hình kinh doanh", "Nhà bán lẻ đa ngành lớn nhất VN (ICT/điện máy qua TGDĐ+ĐMX, FMCG qua BHX), tận dụng quy mô mạng lưới >6.700 cửa hàng để đàm phán giá tốt với NCC và phủ khắp thị trường.", "Tốt"),
        ("2. Vị thế cạnh tranh", "TGDĐ/ĐMX dẫn đầu thị phần ICT/điện máy VN; BHX đứng đầu chuỗi bán lẻ FMCG hiện đại về quy mô cửa hàng, đang mở rộng nhanh và đã có lãi từ 2023.", "Tốt"),
        ("3. Sức khoẻ tài chính", "Đòn bẩy thấp, xu hướng nợ vay giảm/tiền mặt tăng nhờ FCF chuyển dương khi BHX hoà vốn; biên LNG ổn định quanh 20-21%.", "Tốt"),
        ("4. Định giá", f"Giá mục tiêu bình quân trọng số {WEIGHTED_TARGET_PRICE:,.0f} VND, upside {UPSIDE_PCT}% so với giá hiện tại {PRICE:,.0f} VND — định giá hợp lý, không quá đắt so với median lịch sử.", "Trung lập-Tốt"),
        ("5. Chất lượng ban lãnh đạo", "Lịch sử liên tục vượt kế hoạch kinh doanh tự đề ra nhiều năm liền (thận trọng khi đặt mục tiêu đầu năm); minh bạch công bố KQKD hàng tháng qua IR — hiếm có ở DN bán lẻ VN.", "Tốt"),
        ("6. Rủi ro & Catalyst", "Rủi ro: TGDĐ/ĐMX đã bão hòa khó tăng trưởng thêm; BHX vẫn cần chứng minh duy trì lãi bền vững khi mở rộng nhanh. Catalyst: chu kỳ iPhone mới, BHX đạt quy mô lãi ổn định, EraBlue nhân rộng quốc tế.", "Theo dõi"),
    ]
    for i, row in enumerate(thesis_tiers, 2):
        data_row(wsIT, i, row)
        wsIT.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        wsIT.row_dimensions[i].height = 50
    print("[Excel] Sheet 11_Investment_Thesis done.")

    # ─── 12_Summary_Snapshot ────────────────────────────────────────────────────────────────────
    wsSS = wb.create_sheet("12_Summary_Snapshot")
    header_row(wsSS, 1, ["Chỉ tiêu"] + YEAR_HEADERS + ["CAGR 3Y"], [30] + [11]*N_ALL + [14])
    RSS = {}
    def ssrow(r, label, hist_vals, fc_formula_fn, fmt='#,##0', bold=False, fill=None, cagr=False):
        wsSS.cell(row=r, column=1, value=label).font = bold_font if bold else data_font
        wsSS.cell(row=r, column=1).border = thin_border
        for j, v in enumerate(hist_vals):
            c = wsSS.cell(row=r, column=2+j, value=v); c.border = thin_border
            c.font = bold_font if bold else data_font; c.number_format = fmt; c.alignment = Alignment(horizontal='right')
            if fill: c.fill = fill
        for i in range(len(years_fc)):
            col = get_column_letter(2+N_HIST+i)
            c = wsSS.cell(row=r, column=2+N_HIST+i, value=fc_formula_fn(i, col))
            c.border = thin_border; c.font = bold_font if bold else data_font; c.number_format = fmt
            c.alignment = Alignment(horizontal='right')
            if fill: c.fill = fill
        if cagr:
            last_col = get_column_letter(1+N_ALL); first_fc_col = get_column_letter(1+N_HIST)
            cc = wsSS.cell(row=r, column=2+N_ALL, value=f"=({last_col}{r}/{first_fc_col}{r})^(1/{len(years_fc)-1})-1")
            cc.number_format = '0.0%'; cc.font = data_font; cc.border = thin_border; cc.alignment = Alignment(horizontal='right')
        return r + 1

    r = 2
    RSS["revenue"] = r; r = ssrow(r, "Doanh thu thuần", revenue_hist, lambda i, col: f"='04_PnL'!{col}{RP['revenue']}", bold=True, cagr=True)
    RSS["gp"] = r; r = ssrow(r, "Lợi nhuận gộp", gp_hist, lambda i, col: f"='04_PnL'!{col}{RP['gp']}", cagr=True)
    RSS["ebitda"] = r; r = ssrow(r, "EBITDA", ebitda_hist, lambda i, col: f"='04_PnL'!{col}{RP['ebitda']}", cagr=True)
    RSS["ni"] = r; r = ssrow(r, "LNST Cổ đông Công ty mẹ", ni_hist, lambda i, col: f"='04_PnL'!{col}{RP['ni']}", bold=True, fill=p_fill, cagr=True)
    RSS["eps"] = r; r = ssrow(r, "EPS (VND)", eps_hist, lambda i, col: f"='04_PnL'!{col}{RP['eps']}", cagr=True)
    RSS["capex"] = r; r = ssrow(r, "CAPEX", capex_hist, lambda i, col: f"=-'06_Cash_Flow'!{get_column_letter(2+i)}{RCF['capex']}")
    RSS["net_debt"] = r; r = ssrow(r, "Net Debt", NET_DEBT_HIST, lambda i, col: f"='05_Balance_Sheet'!{col}{RBS['net_debt']}")
    RSS["equity"] = r; r = ssrow(r, "Vốn chủ sở hữu", equity_hist, lambda i, col: f"='05_Balance_Sheet'!{col}{RBS['equity']}", cagr=True)
    r += 1
    RSS["roe"] = r; wsSS.cell(row=r, column=1, value="ROE (%)").font = data_font; wsSS.cell(row=r, column=1).border = thin_border
    for j in range(N_ALL):
        col = get_column_letter(2+j)
        c = wsSS.cell(row=r, column=2+j, value=f"={col}{RSS['ni']}/{col}{RSS['equity']}")
        c.number_format = '0.00%'; c.font = data_font; c.border = thin_border; c.alignment = Alignment(horizontal='right')
    r += 1
    RSS["pe"] = r; wsSS.cell(row=r, column=1, value="P/E (x)").font = data_font; wsSS.cell(row=r, column=1).border = thin_border
    for j in range(N_ALL):
        col = get_column_letter(2+j)
        c = wsSS.cell(row=r, column=2+j, value=f"='02_Assumptions'!$B$2*'02_Assumptions'!$B$3/({col}{RSS['ni']}*1000)")
        c.number_format = '0.00"x"'; c.font = data_font; c.border = thin_border; c.alignment = Alignment(horizontal='right')
    r += 1
    RSS["pb"] = r; wsSS.cell(row=r, column=1, value="P/B (x)").font = data_font; wsSS.cell(row=r, column=1).border = thin_border
    for j in range(N_ALL):
        col = get_column_letter(2+j)
        c = wsSS.cell(row=r, column=2+j, value=f"=('02_Assumptions'!$B$2*'02_Assumptions'!$B$3/1000)/{col}{RSS['equity']}")
        c.number_format = '0.00"x"'; c.font = data_font; c.border = thin_border; c.alignment = Alignment(horizontal='right')
    wsSS.column_dimensions['A'].width = 30
    print(f"[Excel] Sheet 12_Summary_Snapshot done ({r} dong).")

    # ─── 13_PE_PB_History — P/E, P/B theo quý (Vietcap), MEDIAN cuối bảng ──────────────────────────
    wsPH = wb.create_sheet("13_PE_PB_History")
    header_row(wsPH, 1, ["Quý", "P/E (x)", "P/B (x)", "EV/EBITDA (x)"], [14, 14, 14, 16])
    _ph_quarters = sorted({(r_["year"], r_["quarter"]) for r_ in MWG_RATIOS if r_.get("quarter") in (1, 2, 3, 4)})
    r = 2
    row_pe_start = r
    for y_, q_ in _ph_quarters:
        rec = next((r_ for r_ in MWG_RATIOS if r_.get("year") == y_ and r_.get("quarter") == q_), {})
        pe0, pb0, ev0 = rec.get("pe"), rec.get("pb"), rec.get("ev_ebitda")
        data_row(wsPH, r, [f"{y_}Q{q_}",
                            round(pe0, 2) if pe0 and 0 < pe0 < 50 else None,
                            round(pb0, 2) if pb0 and pb0 > 0 else None,
                            round(ev0, 2) if ev0 and ev0 > 0 else None])
        r += 1
    row_pe_end = r - 1
    r += 1
    wsPH.cell(row=r, column=1, value="MEDIAN (loại quý bất thường P/E>50x hoặc ≤0)").font = bold_font
    wsPH.cell(row=r, column=2, value=f"=MEDIAN(B{row_pe_start}:B{row_pe_end})").number_format = '0.00"x"'
    wsPH.cell(row=r, column=3, value=f"=MEDIAN(C{row_pe_start}:C{row_pe_end})").number_format = '0.00"x"'
    wsPH.cell(row=r, column=4, value=f"=MEDIAN(D{row_pe_start}:D{row_pe_end})").number_format = '0.00"x"'
    for c in range(1, 5):
        wsPH.cell(row=r, column=c).font = bold_font
        wsPH.cell(row=r, column=c).fill = header_fill
        wsPH.cell(row=r, column=c).border = thin_border
    r += 2
    wsPH.cell(row=r, column=1, value=(
        "Ghi chú: MEDIAN ở đây tính đơn giản trên TOÀN BỘ quý có dữ liệu (bảng tham khảo, theo đúng "
        "yêu cầu skill xuất báo cáo). P/E Median/P/B Median DÙNG ĐỂ ĐỊNH GIÁ ở 02_Assumptions tính theo "
        "phương pháp khác — trung vị-của-trung-vị-năm (median mỗi năm rồi lấy median của các năm) — ổn "
        "định hơn, không bị lệch bởi 1 năm có nhiều quý dữ liệu hơn năm khác. Có thể khác số với MEDIAN "
        "ở trên, đây là chủ đích, không phải lỗi."))
    wsPH.cell(row=r, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    wsPH.merge_cells(f'A{r}:D{r}')
    wsPH.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    wsPH.row_dimensions[r].height = 45
    wsPH.column_dimensions['A'].width = 14
    print(f"[Excel] Sheet 13_PE_PB_History done ({row_pe_end - row_pe_start + 1} quy).")

    # ─── Biểu đồ Excel (tối thiểu 3, skill xuat-bao-cao) ────────────────────────────────────────
    # Chart 1: Doanh thu & LNST theo năm (04_PnL)
    chart1 = BarChart(); chart1.type = "col"; chart1.title = "Doanh thu & LNST theo năm"
    chart1.y_axis.title = "Tỷ VND"; chart1.height = 8; chart1.width = 16
    data1 = Reference(wsP, min_col=2, max_col=1+N_ALL, min_row=RP["revenue"], max_row=RP["revenue"])
    cats1 = Reference(wsP, min_col=2, max_col=1+N_ALL, min_row=1, max_row=1)
    chart1.add_data(data1, titles_from_data=False); chart1.series[0].tx = openpyxl.chart.series.SeriesLabel(v="Doanh thu thuần")
    line1 = LineChart()
    data1b = Reference(wsP, min_col=2, max_col=1+N_ALL, min_row=RP["ni"], max_row=RP["ni"])
    line1.add_data(data1b, titles_from_data=False); line1.series[0].tx = openpyxl.chart.series.SeriesLabel(v="LNST Cổ đông mẹ")
    line1.y_axis.axId = 200; line1.y_axis.title = "Tỷ VND (LNST)"
    chart1.set_categories(cats1); chart1 += line1
    wsP.add_chart(chart1, "B25")

    # Chart 2: Xu hướng biên lợi nhuận (GPM/EBIT margin)
    chart2 = LineChart(); chart2.title = "Xu hướng biên lợi nhuận"; chart2.y_axis.title = "%"
    chart2.height = 8; chart2.width = 16
    data2 = Reference(wsP, min_col=2, max_col=1+N_ALL, min_row=RP["gpm"], max_row=RP["gpm"])
    chart2.add_data(data2, titles_from_data=False); chart2.series[0].tx = openpyxl.chart.series.SeriesLabel(v="Biên LNG")
    data2b = Reference(wsP, min_col=2, max_col=1+N_ALL, min_row=RP["ebit_margin"], max_row=RP["ebit_margin"])
    chart2.add_data(data2b, titles_from_data=False); chart2.series[1].tx = openpyxl.chart.series.SeriesLabel(v="Biên EBIT")
    chart2.set_categories(cats1)
    wsP.add_chart(chart2, "B42")

    # Chart 3: Doanh thu theo mảng (03_Revenue_Model) — cột chồng
    chart3 = BarChart(); chart3.type = "col"; chart3.grouping = "stacked"; chart3.overlap = 100
    chart3.title = "Cơ cấu doanh thu theo mảng"; chart3.y_axis.title = "Tỷ VND"
    chart3.height = 9; chart3.width = 18
    for seg in SEGMENTS:
        d = Reference(wsR, min_col=2, max_col=1+N_ALL, min_row=RR["rev"][seg], max_row=RR["rev"][seg])
        chart3.add_data(d, titles_from_data=False)
        chart3.series[-1].tx = openpyxl.chart.series.SeriesLabel(v=seg)
    chart3.set_categories(cats1)
    wsR.add_chart(chart3, "B30")

    # Chart 4: Số cửa hàng theo mảng theo thời gian
    chart4 = LineChart(); chart4.title = "Số cửa hàng theo mảng"; chart4.y_axis.title = "Cửa hàng"
    chart4.height = 8; chart4.width = 16
    for seg in SEGMENTS:
        d = Reference(wsR, min_col=2+N_HIST, max_col=1+N_ALL, min_row=RR["store"][seg], max_row=RR["store"][seg])
        chart4.add_data(d, titles_from_data=False)
        chart4.series[-1].tx = openpyxl.chart.series.SeriesLabel(v=seg)
    cats_fc = Reference(wsR, min_col=2+N_HIST, max_col=1+N_ALL, min_row=1, max_row=1)
    chart4.set_categories(cats_fc)
    wsR.add_chart(chart4, "B50")
    print("[Excel] Bieu do Excel (4 chart) done.")

    wb.save(EXCEL_FILE)
    print(f"[Excel] Saved: {EXCEL_FILE}")
    return {
        "RA": RA, "RR": RR, "RP": RP, "RBS": RBS, "RCF": RCF,
        "row_pe_target": row_pe_target, "row_pb_target": row_pb_target, "row_ri_target": row_ri_target,
        "row_pspe_target": row_pspe_target,
    }

# ══════════════════════════════════════════════════════════════════════════════════════════════
# CHARTS (matplotlib) — dùng chung cho PDF, tối thiểu 3 biểu đồ theo skill xuat-bao-cao
# ══════════════════════════════════════════════════════════════════════════════════════════════
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial', 'Calibri', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
COLOR_MAIN = "#1F4E79"
COLOR_SEG = {"TGDD": "#2980B9", "DMX": "#F39C12", "BHX": "#27AE60", "Khac": "#95A5A6"}

def make_revenue_ni_chart():
    fig, ax1 = plt.subplots(figsize=(10, 5.5))
    x = list(range(N_ALL))
    ax1.bar(x, revenue_hist + revenue_fc, color=COLOR_MAIN, alpha=0.85, label="Doanh thu thuần")
    ax1.set_ylabel("Doanh thu thuần (tỷ VND)", fontsize=11)
    ax1.set_xticks(x); ax1.set_xticklabels(YEAR_HEADERS, fontsize=10)
    ax2 = ax1.twinx()
    ax2.plot(x, ni_hist + ni_fc, color="#E74C3C", marker='o', linewidth=2.5, label="LNST Cổ đông mẹ")
    ax2.set_ylabel("LNST (tỷ VND)", fontsize=11)
    ax1.axvline(N_HIST - 0.5, color='gray', linestyle='--', alpha=0.6)
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=10)
    ax1.set_title(f"{TICKER} — Doanh thu thuần & LNST ({years_hist[0]}-{years_fc[-1]})", fontsize=13, fontweight='bold')
    ax1.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(CHART_DIR, "revenue_ni.png")
    fig.savefig(path, dpi=180, bbox_inches='tight'); plt.close(fig)
    return path

def make_margin_chart():
    fig, ax = plt.subplots(figsize=(10, 5))
    x = list(range(N_ALL))
    gpm_all = gpm_hist + gp_margin_fc
    ebit_m_all = [round(ebit_hist[i]/revenue_hist[i]*100, 2) for i in range(N_HIST)] + [round(ebit_fc[i]/revenue_fc[i]*100, 2) for i in range(len(years_fc))]
    ni_m_all = [round(ni_hist[i]/revenue_hist[i]*100, 2) for i in range(N_HIST)] + [round(ni_fc[i]/revenue_fc[i]*100, 2) for i in range(len(years_fc))]
    ax.plot(x, gpm_all, marker='o', label="Biên LNG (%)", color=COLOR_MAIN, linewidth=2.2)
    ax.plot(x, ebit_m_all, marker='s', label="Biên EBIT (%)", color="#F39C12", linewidth=2.2)
    ax.plot(x, ni_m_all, marker='^', label="Biên LNST (%)", color="#E74C3C", linewidth=2.2)
    ax.set_xticks(x); ax.set_xticklabels(YEAR_HEADERS, fontsize=10)
    ax.axvline(N_HIST - 0.5, color='gray', linestyle='--', alpha=0.6)
    ax.set_ylabel("%", fontsize=11)
    ax.set_title(f"{TICKER} — Xu hướng biên lợi nhuận", fontsize=13, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(CHART_DIR, "margin_trend.png")
    fig.savefig(path, dpi=180, bbox_inches='tight'); plt.close(fig)
    return path

def make_segment_revenue_chart():
    fig, ax = plt.subplots(figsize=(10, 5.5))
    x = list(range(N_ALL))
    bottom = [0] * N_ALL
    for seg in SEGMENTS:
        vals = []
        for i, y in enumerate(years_hist):
            vals.append(REV_SEGMENT_HIST.get(y, {}).get(seg, 0) or 0)
        vals += REVENUE_FC_SEGMENT[seg]
        ax.bar(x, vals, bottom=bottom, label=seg, color=COLOR_SEG[seg])
        bottom = [bottom[i] + vals[i] for i in range(N_ALL)]
    ax.set_xticks(x); ax.set_xticklabels(YEAR_HEADERS, fontsize=10)
    ax.axvline(N_HIST - 0.5, color='gray', linestyle='--', alpha=0.6)
    ax.set_ylabel("Tỷ VND", fontsize=11)
    ax.set_title(f"{TICKER} — Cơ cấu doanh thu theo mảng", fontsize=13, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(CHART_DIR, "segment_revenue.png")
    fig.savefig(path, dpi=180, bbox_inches='tight'); plt.close(fig)
    return path

def make_store_count_chart():
    fig, ax = plt.subplots(figsize=(10, 5))
    fc_labels = [f"{y}E" if i == 0 else f"{y}F" for i, y in enumerate(years_fc)]
    x = list(range(len(years_fc)))
    for seg in SEGMENTS:
        ax.plot(x, STORE_COUNT_FC[seg], marker='o', linewidth=2.2, label=seg, color=COLOR_SEG[seg])
    ax.set_xticks(x); ax.set_xticklabels(fc_labels, fontsize=10)
    ax.set_ylabel("Số cửa hàng", fontsize=11)
    ax.set_title(f"{TICKER} — Số cửa hàng dự phóng theo mảng", fontsize=13, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(CHART_DIR, "store_count.png")
    fig.savefig(path, dpi=180, bbox_inches='tight'); plt.close(fig)
    return path

def make_valuation_chart():
    fig, ax = plt.subplots(figsize=(10, 5.5))
    methods = ["P/E\nMedian", "P/B\nMedian", "RI\n(cả CT)", "P/S+P/E\ncả 4 mảng", "GIÁ\nMỤC TIÊU"]
    values = [PE_TARGET_PRICE, PB_TARGET_PRICE, RI_TARGET_PRICE, PS_PE_TARGET_PRICE, WEIGHTED_TARGET_PRICE]
    colors = ["#3498DB"] * 4 + ["#27AE60"]
    bars = ax.bar(methods, values, color=colors)
    ax.axhline(PRICE, color="#E74C3C", linestyle='--', linewidth=2, label=f"Giá hiện tại: {PRICE:,.0f}")
    for b, v in zip(bars, values):
        ax.text(b.get_x() + b.get_width()/2, v, f"{v:,.0f}", ha='center', va='bottom', fontsize=9)
    ax.set_ylabel("VND/CP", fontsize=11)
    ax.set_title(f"{TICKER} — So sánh 4 phương pháp định giá", fontsize=13, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(CHART_DIR, "valuation_compare.png")
    fig.savefig(path, dpi=180, bbox_inches='tight'); plt.close(fig)
    return path

def make_dio_dpo_chart():
    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = list(range(len(DIO_Q_LABELS)))
    ax.plot(x, DIO_QUARTERLY, marker='o', linewidth=2, label="DIO - Số ngày tồn kho", color="#2980B9")
    ax.plot(x, DPO_QUARTERLY, marker='s', linewidth=2, label="DPO - Số ngày phải trả", color="#E67E22")
    ax.plot(x, DSO_QUARTERLY, marker='^', linewidth=2, label="DSO - Số ngày phải thu", color="#27AE60")
    step = max(1, len(x) // 16)
    ax.set_xticks(x[::step]); ax.set_xticklabels([DIO_Q_LABELS[i] for i in x[::step]], rotation=45, ha='right', fontsize=8)
    ax.set_ylabel("Số ngày", fontsize=11)
    ax.set_title(f"{TICKER} — DIO/DSO/DPO theo quý ({DIO_Q_LABELS[0]}-{DIO_Q_LABELS[-1]})", fontsize=13, fontweight='bold')
    ax.legend(fontsize=10); ax.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(CHART_DIR, "dio_dso_dpo_quarterly.png")
    fig.savefig(path, dpi=180, bbox_inches='tight'); plt.close(fig)
    return path

# ══════════════════════════════════════════════════════════════════════════════════════════════
# PDF REPORT — 15-20 trang theo skill xuat-bao-cao
# ══════════════════════════════════════════════════════════════════════════════════════════════
def build_pdf():
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('TitleVN', fontName=PDF_FONT_BOLD, fontSize=22, leading=26,
                                  alignment=TA_CENTER, textColor=HexColor("#1F4E79"), spaceAfter=10)
    style_h1 = ParagraphStyle('H1VN', fontName=PDF_FONT_BOLD, fontSize=15, leading=18,
                               textColor=HexColor("#1F4E79"), spaceBefore=10, spaceAfter=8)
    style_h2 = ParagraphStyle('H2VN', fontName=PDF_FONT_BOLD, fontSize=12, leading=15,
                               textColor=HexColor("#2C3E50"), spaceBefore=8, spaceAfter=6)
    style_body = ParagraphStyle('BodyVN', fontName=PDF_FONT, fontSize=10, leading=14,
                                 alignment=TA_JUSTIFY, spaceAfter=6)
    style_center = ParagraphStyle('CenterVN', fontName=PDF_FONT, fontSize=11, leading=15, alignment=TA_CENTER)
    style_small = ParagraphStyle('SmallVN', fontName=PDF_FONT, fontSize=8, leading=10, textColor=HexColor("#666666"))

    story = []
    RECOMMEND = "MUA" if UPSIDE_PCT and UPSIDE_PCT > 15 else ("BÁN" if UPSIDE_PCT and UPSIDE_PCT < -5 else "THEO DÕI")
    REC_COLOR = "#27AE60" if RECOMMEND == "MUA" else ("#E74C3C" if RECOMMEND == "BÁN" else "#F39C12")

    # ── Trang 1: Cover ──────────────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 40*mm))
    story.append(Paragraph(f"{TICKER}", style_title))
    story.append(Paragraph(COMPANY, ParagraphStyle('Sub', fontName=PDF_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=20)))
    story.append(Paragraph(f"Ngành: {INDUSTRY} | Sàn: {EXCHANGE}", style_center))
    story.append(Spacer(1, 15*mm))
    rec_style = ParagraphStyle('Rec', fontName=PDF_FONT_BOLD, fontSize=28, leading=36, alignment=TA_CENTER,
                                textColor=HexColor(REC_COLOR), spaceAfter=10)
    story.append(Paragraph(RECOMMEND, rec_style))
    story.append(Paragraph(f"Giá mục tiêu: {WEIGHTED_TARGET_PRICE:,.0f} VND | Upside: {UPSIDE_PCT:+.1f}%", style_center))
    story.append(Spacer(1, 15*mm))
    cover_tbl = Table([
        ["Giá hiện tại", f"{PRICE:,.0f} VND"],
        ["Vốn hóa", f"{MARKET_CAP:,.0f} tỷ VND"],
        ["P/E TTM", f"{PRICE/eps_hist[-1]:.1f}x" if eps_hist[-1] else "N/A"],
        ["P/B", f"{PRICE/bvps_hist[-1]:.2f}x" if bvps_hist[-1] else "N/A"],
        ["Ngày phân tích", datetime.now().strftime("%d/%m/%Y")],
    ], colWidths=[70*mm, 70*mm])
    cover_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 11),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, HexColor("#CCCCCC")),
        ('BACKGROUND', (0,0), (0,-1), HexColor("#F0F0F0")),
        ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(cover_tbl)
    story.append(PageBreak())

    # ── Trang 2: Earning Release quý gần nhất (bắt buộc theo skill xuat-bao-cao) ───────────────
    story.append(Paragraph("EARNING RELEASE — QUÝ GẦN NHẤT", style_h1))
    _latest_q = LAST_4Q[-1] if LAST_4Q else None
    _prior_q = LAST_4Q[-2] if len(LAST_4Q) >= 2 else None
    _yoy_q = qv(is_q, _latest_q["year"] - 1, _latest_q["q"], "isa3") if _latest_q else None
    if _latest_q:
        _ni_latest = qv(is_q, _latest_q["year"], _latest_q["q"], "isa22")
        _ni_prior = qv(is_q, _prior_q["year"], _prior_q["q"], "isa22") if _prior_q else None
        _ni_yoy = qv(is_q, _latest_q["year"] - 1, _latest_q["q"], "isa22")
        _qoq_rev = (_latest_q["rev"]/_prior_q["rev"]-1)*100 if _prior_q else None
        _yoy_rev = (_latest_q["rev"]/_yoy_q-1)*100 if _yoy_q else None
        _qoq_ni = (_ni_latest/_ni_prior-1)*100 if _ni_prior else None
        _yoy_ni = (_ni_latest/_ni_yoy-1)*100 if _ni_yoy else None
        story.append(Paragraph(
            f"Quý {_latest_q['q']}/{_latest_q['year']}: Doanh thu thuần {_latest_q['rev']:,.0f} tỷ đồng "
            f"({'+' if _qoq_rev and _qoq_rev>=0 else ''}{_qoq_rev:.1f}% QoQ, {'+' if _yoy_rev and _yoy_rev>=0 else ''}"
            f"{_yoy_rev:.1f}% YoY). LNST hợp nhất {_ni_latest:,.0f} tỷ đồng "
            f"({'+' if _qoq_ni and _qoq_ni>=0 else ''}{_qoq_ni:.1f}% QoQ, {'+' if _yoy_ni and _yoy_ni>=0 else ''}"
            f"{_yoy_ni:.1f}% YoY). Biên LNG {_latest_q['gpm']*100:.1f}%, SG&amp;A/DT {_latest_q['sga_pct']*100:.1f}%.",
            style_body))
        er_tbl = Table([
            ["Chỉ tiêu", f"Q{_latest_q['q']}/{_latest_q['year']}", f"Q{_prior_q['q']}/{_prior_q['year']}" if _prior_q else "N/A",
             f"Q{_latest_q['q']}/{_latest_q['year']-1}", "QoQ (%)", "YoY (%)"],
            ["Doanh thu thuần (tỷ)", f"{_latest_q['rev']:,.0f}", f"{_prior_q['rev']:,.0f}" if _prior_q else "N/A",
             f"{_yoy_q:,.0f}" if _yoy_q else "N/A", f"{_qoq_rev:+.1f}%" if _qoq_rev is not None else "N/A",
             f"{_yoy_rev:+.1f}%" if _yoy_rev is not None else "N/A"],
            ["LNST hợp nhất (tỷ)", f"{_ni_latest:,.0f}", f"{_ni_prior:,.0f}" if _ni_prior else "N/A",
             f"{_ni_yoy:,.0f}" if _ni_yoy else "N/A", f"{_qoq_ni:+.1f}%" if _qoq_ni is not None else "N/A",
             f"{_yoy_ni:+.1f}%" if _yoy_ni is not None else "N/A"],
            ["Biên LNG (%)", f"{_latest_q['gpm']*100:.1f}%", f"{_prior_q['gpm']*100:.1f}%" if _prior_q else "N/A", "-", "-", "-"],
            ["SG&A/DT (%)", f"{_latest_q['sga_pct']*100:.1f}%", f"{_prior_q['sga_pct']*100:.1f}%" if _prior_q else "N/A", "-", "-", "-"],
        ], colWidths=[35*mm, 24*mm, 24*mm, 24*mm, 22*mm, 22*mm])
        er_tbl.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 8.5),
            ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
        ]))
        story.append(er_tbl)
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            "<b>Nhận định:</b> Doanh thu và lợi nhuận tiếp tục tăng trưởng hai chữ số theo cả QoQ và YoY, phản ánh "
            "chu kỳ phục hồi sức mua và đóng góp ngày càng lớn từ BHX. Biên lợi nhuận gộp ổn định quanh 19-21%, "
            "SG&amp;A/DT có xu hướng giảm nhẹ nhờ hiệu ứng quy mô khi mạng lưới cửa hàng mở rộng.", style_body))
    story.append(PageBreak())

    # ── Trang 3: Investment Summary ────────────────────────────────────────────────────────────
    story.append(Paragraph("TÓM TẮT ĐẦU TƯ (INVESTMENT SUMMARY)", style_h1))
    story.append(Paragraph(
        f"<b>Khuyến nghị: {RECOMMEND}</b> — Giá mục tiêu <b>{WEIGHTED_TARGET_PRICE:,.0f} VND</b> "
        f"(upside {UPSIDE_PCT:+.1f}% so với giá hiện tại {PRICE:,.0f} VND), theo mô hình định giá tổng hợp "
        f"4 phương pháp (P/E Median 20%, P/B Median 20%, Residual Income 10%, P/S+P/E theo từng mảng — "
        f"TGDĐ/ĐMX/BHX/Khác — 50%).", style_body))
    story.append(Paragraph("3 luận điểm đầu tư chính:", style_h2))
    thesis_pdf = [
        f"MWG là nhà bán lẻ đa ngành lớn nhất Việt Nam với mạng lưới hơn {sum(STORE_COUNT_NOW.values()):,} cửa hàng "
        f"(TGDĐ+ĐMX dẫn đầu thị phần ICT/điện máy, BHX dẫn đầu bán lẻ FMCG hiện đại), tận dụng quy mô để đàm phán "
        f"giá tốt với nhà cung cấp và tối ưu logistics.",
        f"BHX đã chuyển từ lỗ sang có lãi từ 2023, đang mở rộng nhanh (+{STORE_COUNT_NOW['BHX']-2559:,} cửa hàng "
        f"chỉ trong 5 tháng đầu 2026) với SSSG cao nhất công ty — dư địa tăng trưởng dài hạn lớn khi BHX còn xa mức "
        f"bão hòa so với TGDĐ/ĐMX.",
        f"Doanh thu 5 tháng 2026 đạt {REV_5T_2026_TOTAL:,.0f} tỷ đồng (+29,3% svck), phản ánh chu kỳ phục hồi sức "
        f"mua và chu kỳ nâng cấp iPhone/thiết bị ICT hỗ trợ TGDĐ/ĐMX tăng trưởng SSSG cao (~33%) dù đã bão hòa về "
        f"số lượng cửa hàng.",
    ]
    for t in thesis_pdf:
        story.append(Paragraph(f"• {t}", style_body))
    story.append(Paragraph("3 rủi ro chính:", style_h2))
    risks_pdf = [
        "TGDĐ/ĐMX đã bão hòa (số cửa hàng gần như đi ngang), tăng trưởng phụ thuộc chu kỳ sản phẩm (iPhone) và SSSG "
        "— nếu sức mua chững lại, 2 chuỗi lớn nhất company khó bù đắp bằng mở rộng thêm.",
        "BHX mở rộng nhanh (kế hoạch +1.000 cửa hàng năm 2026) cần thời gian ramp-up 3 tháng mới đạt hiệu quả ổn "
        "định — rủi ro biên lợi nhuận công ty bị pha loãng nếu tốc độ mở mới vượt quá khả năng vận hành.",
        "Định giá P/E hiện dùng median lịch sử bao gồm giai đoạn tăng trưởng nóng 2021 — nếu thị trường định giá "
        "lại theo P/E thấp hơn (phản ánh giai đoạn tăng trưởng chậm lại của TGDĐ/ĐMX), giá mục tiêu P/E-based có "
        "thể giảm đáng kể.",
    ]
    for t in risks_pdf:
        story.append(Paragraph(f"• {t}", style_body))
    snap_tbl_data = [["Chỉ tiêu"] + YEAR_HEADERS]
    snap_rows = [
        ("Doanh thu (tỷ)", revenue_hist + revenue_fc),
        ("LNST (tỷ)", ni_hist + ni_fc),
        ("EPS (VND)", eps_hist + EPS_FC),
        ("Biên LNG (%)", [f"{v:.1f}%" for v in gpm_hist] + [f"{v:.1f}%" for v in gp_margin_fc]),
    ]
    for label, vals in snap_rows:
        row = [label] + [f"{v:,.0f}" if isinstance(v, (int, float)) else v for v in vals]
        snap_tbl_data.append(row)
    snap_tbl = Table(snap_tbl_data, colWidths=[28*mm] + [18*mm]*N_ALL)
    snap_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(Spacer(1, 6))
    story.append(snap_tbl)
    story.append(PageBreak())

    # ── Trang 3-4: Mô hình kinh doanh ──────────────────────────────────────────────────────────
    story.append(Paragraph("MÔ HÌNH KINH DOANH & CHUỖI GIÁ TRỊ", style_h1))
    story.append(Paragraph(
        "MWG vận hành 3 chuỗi bán lẻ chính, mỗi chuỗi ở giai đoạn vòng đời khác nhau:", style_body))
    biz_desc = [
        ("Thế Giới Di Động (TGDĐ, gồm Topzone)", f"{STORE_COUNT_NOW['TGDD']:,} cửa hàng — bán điện thoại/laptop/"
         "phụ kiện. Đã bão hòa số lượng cửa hàng, tăng trưởng chủ yếu từ SSSG theo chu kỳ sản phẩm."),
        ("Điện Máy Xanh (ĐMX, gồm ĐMS/Thợ Điện Máy Xanh)", f"{STORE_COUNT_NOW['DMX']:,} cửa hàng — bán điện máy/"
         "gia dụng, mảng lớn nhất về doanh thu. Mô hình ĐMS (supermini) mở rộng vùng sâu vùng xa."),
        ("Bách Hóa Xanh (BHX)", f"{STORE_COUNT_NOW['BHX']:,} cửa hàng — chuỗi FMCG/thực phẩm tươi sống, tăng "
         "trưởng nhanh nhất, chuyển sang có lãi từ 2023, kế hoạch mở thêm ~1.000 cửa hàng năm 2026."),
        ("Khác (An Khang, AvaKids, EraBlue)", f"An Khang {STORE_COUNT_NOW['AnKhang']} nhà thuốc, AvaKids "
         f"{STORE_COUNT_NOW['AvaKids']} cửa hàng, EraBlue (liên doanh Indonesia) {STORE_COUNT_NOW['EraBlue']} "
         "cửa hàng — các mảng mới, quy mô nhỏ, tăng trưởng cao từ nền thấp."),
    ]
    for name, desc in biz_desc:
        story.append(Paragraph(f"<b>{name}:</b> {desc}", style_body))
    story.append(Image(make_segment_revenue_chart(), width=165*mm, height=95*mm))
    story.append(PageBreak())

    story.append(Paragraph("SỐ CỬA HÀNG & HIỆU QUẢ VẬN HÀNH", style_h1))
    story.append(Paragraph(
        "Đánh giá hiệu quả từng cửa hàng (skill phân tích bán lẻ): hiệu quả doanh thu/cửa hàng/tháng được tính từ "
        "doanh thu tháng gần nhất chia số cửa hàng CÁCH thời điểm hiện tại 3 tháng (cửa hàng mới mở cần ~3 tháng "
        "ramp-up mới ổn định doanh số).", style_body))
    eff_tbl = Table([["Mảng", "Hiệu quả DT/CH/tháng (tỷ VND)", "Tăng trưởng giả định (%/năm)"]] +
                     [[seg, f"{EFFICIENCY_NOW[seg]:.2f}", f"{EFFICIENCY_GROWTH[seg]*100:.0f}%"] for seg in SEGMENTS],
                     colWidths=[35*mm, 65*mm, 65*mm])
    eff_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(eff_tbl)
    story.append(Spacer(1, 8))
    story.append(Image(make_store_count_chart(), width=165*mm, height=95*mm))
    story.append(PageBreak())

    # ── Trang phụ: Phân tích chi tiết từng mảng (dữ liệu IR thật, thu thập trực tiếp) ───────────
    story.append(Paragraph("PHÂN TÍCH CHI TIẾT TỪNG MẢNG KINH DOANH", style_h1))
    seg_detail = [
        ("Thế Giới Di Động (TGDĐ)",
         f"Số cửa hàng: {STORE_COUNT_HIST.get(2024, {}).get('TGDD', 'N/A')} ({2024}) → "
         f"{STORE_COUNT_HIST.get(2025, {}).get('TGDD', 'N/A')} ({2025}) → {STORE_COUNT_NOW['TGDD']} (hiện tại) — "
         "gần như đi ngang, phản ánh mảng đã bão hòa về độ phủ. Tăng trưởng doanh thu hiện tại chủ yếu đến từ "
         "SSSG (tăng trưởng cửa hàng hiện hữu), được hỗ trợ bởi chu kỳ nâng cấp iPhone và các dòng sản phẩm ICT."),
        ("Điện Máy Xanh (ĐMX, gồm ĐMS)",
         f"Số cửa hàng: {STORE_COUNT_HIST.get(2024, {}).get('DMX', 'N/A')} ({2024}) → "
         f"{STORE_COUNT_HIST.get(2025, {}).get('DMX', 'N/A')} ({2025}) → {STORE_COUNT_NOW['DMX']} (hiện tại) — "
         "mảng lớn nhất về doanh thu (~46% tổng DT). Mô hình ĐMX Supermini (ĐMS, cửa hàng nhỏ) tiếp tục mở rộng "
         "vùng sâu vùng xa, bù đắp cho việc các cửa hàng lớn tại đô thị đã bão hòa."),
        ("Bách Hóa Xanh (BHX)",
         f"Số cửa hàng: {STORE_COUNT_HIST.get(2024, {}).get('BHX', 'N/A')} ({2024}) → "
         f"{STORE_COUNT_HIST.get(2025, {}).get('BHX', 'N/A')} ({2025}) → {STORE_COUNT_NOW['BHX']} (hiện tại) — "
         "mảng tăng trưởng nhanh nhất, đã mở thêm gần 500 cửa hàng chỉ trong 5 tháng đầu 2026. Chuyển từ lỗ sang "
         "có lãi từ 2023 sau giai đoạn tái cơ cấu (đóng cửa hàng hoạt động kém hiệu quả), hiện là động lực tăng "
         "trưởng chính của toàn công ty."),
        ("Khác (An Khang, AvaKids, EraBlue)",
         f"An Khang (nhà thuốc): {STORE_COUNT_NOW['AnKhang']} cửa hàng, đã qua giai đoạn tái cơ cấu, tăng trưởng "
         f"ổn định. AvaKids (mẹ & bé): {STORE_COUNT_NOW['AvaKids']} cửa hàng, quy mô nhỏ. EraBlue (liên doanh điện "
         f"máy tại Indonesia): {STORE_COUNT_NOW['EraBlue']} cửa hàng, tăng trưởng rất nhanh từ nền thấp (mục tiêu "
         f">{FY2026_ERABLUE_TARGET_STORES} cửa hàng cuối 2026) — câu chuyện mở rộng quốc tế mới của MWG."),
    ]
    for name, desc in seg_detail:
        story.append(Paragraph(f"<b>{name}</b>", style_h2))
        story.append(Paragraph(desc, style_body))
    story.append(PageBreak())

    # ── Trang 5-6: Thị trường & vị thế cạnh tranh ──────────────────────────────────────────────
    story.append(Paragraph("THỊ TRƯỜNG & VỊ THẾ CẠNH TRANH", style_h1))
    story.append(Paragraph(
        "Ngành bán lẻ hiện đại Việt Nam đang trong giai đoạn chuyển đổi từ kênh truyền thống (chợ, tạp hóa) sang "
        "kênh hiện đại (siêu thị mini, cửa hàng tiện lợi), đặc biệt ở phân khúc FMCG nơi BHX cạnh tranh trực tiếp "
        "với Circle K, WinMart, Bách Hóa Xanh của chính MWG dẫn đầu về số lượng điểm bán.", style_body))
    story.append(Paragraph("Áp lực cạnh tranh (Porter 5 Forces):", style_h2))
    porter = [
        "Đối thủ trong ngành: TGDĐ/ĐMX cạnh tranh với FPT Shop, CellphoneS (ICT); BHX cạnh tranh WinMart, Circle K, "
        "GS25 (FMCG hiện đại) — MWG dẫn đầu quy mô cửa hàng ở cả 2 mảng.",
        "Nhà cung cấp: Quy mô lớn giúp MWG có sức mạnh đàm phán tốt với Apple/Samsung (ICT) và các nhà sản xuất "
        "FMCG, đạt chiết khấu/rebate tốt hơn đối thủ nhỏ.",
        "Khách hàng: Sản phẩm ICT có chi phí chuyển đổi thấp (khách dễ so sánh giá online) — MWG bù đắp bằng dịch "
        "vụ hậu mãi/bảo hành và tài chính tiêu dùng (trả góp).",
        "Sản phẩm thay thế: Thương mại điện tử (Shopee, TikTok Shop) cạnh tranh trực tiếp ICT — MWG phản ứng bằng "
        "kênh online riêng và Super App tích hợp.",
        "Rào cản gia nhập: Cao với BHX (mạng lưới logistics chuỗi lạnh, mặt bằng) — thấp hơn với ICT (nhiều chuỗi "
        "nhỏ vẫn gia nhập được).",
    ]
    for p in porter:
        story.append(Paragraph(f"• {p}", style_body))
    story.append(PageBreak())

    # ── Trang 7-8: Moat & Quản trị ──────────────────────────────────────────────────────────────
    story.append(Paragraph("LỢI THẾ CẠNH TRANH BỀN VỮNG (MOAT)", style_h1))
    moats_pdf = [
        ("Hiệu quả quy mô (Efficient Scale)", 4, "Mạng lưới >6.700 cửa hàng toàn quốc tạo lợi thế đàm phán NCC và "
         "phân bổ chi phí logistics/quảng cáo trên quy mô lớn."),
        ("Lợi thế chi phí (Cost Advantage)", 4, "Sức mua lớn giúp đạt giá vốn/chiết khấu tốt hơn đối thủ nhỏ lẻ ở "
         "cả ICT và FMCG."),
        ("Tài sản vô hình (Intangible Assets)", 3, "Thương hiệu TGDĐ/ĐMX/BHX được nhận diện rộng rãi, nhưng ICT là "
         "ngành ít trung thành thương hiệu (khách so giá dễ dàng)."),
        ("Chi phí chuyển đổi (Switching Cost)", 2, "Thấp với ICT (khách dễ chuyển sang kênh khác); trung bình với "
         "BHX nhờ tiện lợi vị trí gần nhà."),
        ("Hiệu ứng mạng lưới (Network Effect)", 2, "Hạn chế — bán lẻ truyền thống không có hiệu ứng mạng lưới "
         "mạnh như nền tảng công nghệ, dù Super App có tiềm năng tạo hiệu ứng nhẹ."),
    ]
    moat_tbl = Table([["Yếu tố", "Điểm (/5)", "Giải thích"]] + [[m, f"{s}/5", d] for m, s, d in moats_pdf],
                      colWidths=[40*mm, 18*mm, 107*mm])
    moat_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('ALIGN', (0,0), (1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(moat_tbl)
    story.append(Paragraph("Quản trị doanh nghiệp", style_h2))
    story.append(Paragraph(
        "MWG công bố kết quả kinh doanh HÀNG THÁNG qua kênh IR — mức độ minh bạch hiếm có ở doanh nghiệp bán lẻ "
        "Việt Nam, giúp nhà đầu tư theo dõi sát diễn biến từng chuỗi. Ban lãnh đạo có lịch sử đặt kế hoạch kinh "
        "doanh thận trọng và thường xuyên vượt kế hoạch tự đề ra.", style_body))
    story.append(PageBreak())

    # ── Trang 9-11: Phân tích tài chính ─────────────────────────────────────────────────────────
    story.append(Paragraph("PHÂN TÍCH TÀI CHÍNH", style_h1))
    story.append(Image(make_revenue_ni_chart(), width=165*mm, height=90*mm))
    story.append(Spacer(1, 6))
    story.append(Image(make_margin_chart(), width=165*mm, height=82*mm))
    story.append(PageBreak())

    story.append(Paragraph("CHU KỲ TIỀN MẶT (CCC) — DIO/DSO/DPO THEO QUÝ", style_h1))
    story.append(Paragraph(
        f"DIO (Số ngày tồn kho) = Tồn kho cuối quý / (Giá vốn quý / 91,25 ngày). DPO (Số ngày phải trả) = Phải trả "
        f"người bán / (Giá vốn quý / 91,25 ngày). DSO (Số ngày phải thu) = Phải thu KH / (Doanh thu quý / 91,25 "
        f"ngày). Quý gần nhất ({DIO_Q_LABELS[-1]}): DIO {DIO_QUARTERLY[-1]:.0f} ngày, DPO "
        f"{DPO_QUARTERLY[-1]:.0f} ngày, DSO {DSO_QUARTERLY[-1]:.0f} ngày — CCC = "
        f"{DIO_QUARTERLY[-1] + (DSO_QUARTERLY[-1] or 0) - (DPO_QUARTERLY[-1] or 0):.0f} ngày.", style_body))
    story.append(Image(make_dio_dpo_chart(), width=170*mm, height=85*mm))
    story.append(PageBreak())

    story.append(Paragraph("BẢNG SỐ LIỆU TÀI CHÍNH CHI TIẾT (5 năm lịch sử + 3 năm dự phóng)", style_h1))
    fin_tbl_data = [["Chỉ tiêu (tỷ VND)"] + YEAR_HEADERS]
    fin_rows = [
        ("Doanh thu thuần", revenue_hist + revenue_fc),
        ("Lợi nhuận gộp", gp_hist + gp_fc),
        ("Biên LNG (%)", [f"{v:.1f}%" for v in gpm_hist] + [f"{v:.1f}%" for v in gp_margin_fc]),
        ("EBITDA", ebitda_hist + ebitda_fc),
        ("EBIT", ebit_hist + ebit_fc),
        ("LNST Cổ đông mẹ", ni_hist + ni_fc),
        ("EPS (VND)", eps_hist + EPS_FC),
        ("VCSH", equity_hist + equity_fc),
        ("Net Debt", NET_DEBT_HIST + NET_DEBT_FC),
    ]
    for label, vals in fin_rows:
        fin_tbl_data.append([label] + [f"{v:,.0f}" if isinstance(v, (int, float)) else v for v in vals])
    fin_tbl = Table(fin_tbl_data, colWidths=[32*mm] + [16*mm]*N_ALL)
    fin_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(fin_tbl)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"Biên lợi nhuận gộp dự phóng {GPM_FC_PCT:.1f}% (bình quân 2 quý gần nhất). Chi phí bán hàng & QLDN/doanh "
        f"thu dự phóng giảm dần {SGA_FC_PCT[0]:.1f}% → {SGA_FC_PCT[-1]:.1f}% (hồi quy xu hướng 4 quý gần nhất) "
        f"nhờ hiệu ứng quy mô khi BHX mở rộng. Thuế suất hiệu dụng {EFFECTIVE_TAX_RATE*100:.1f}% (bình quân 2 năm "
        f"gần nhất, dựa trên Lãi trước thuế thực tế).", style_body))
    story.append(PageBreak())

    # ── Trang 12-14: Định giá ───────────────────────────────────────────────────────────────────
    story.append(Paragraph("ĐỊNH GIÁ", style_h1))
    story.append(Paragraph(
        "Áp dụng khung định giá 4 phương pháp theo skill phân tích bán lẻ, ưu tiên định giá theo TỪNG MẢNG kinh "
        "doanh bằng P/E (thay vì gộp chung công ty hoặc dùng P/B mảng — không có cách tách bạch VCSH riêng cho "
        "từng mảng từ BCTC hợp nhất nên bỏ hẳn phương pháp P/S+P/B).", style_body))
    val_tbl = Table([
        ["Phương pháp", "Trọng số", "Giá mục tiêu (VND)"],
        ["P/E Median", "20%", f"{PE_TARGET_PRICE:,.0f}"],
        ["P/B Median", "20%", f"{PB_TARGET_PRICE:,.0f}"],
        ["Residual Income (cả công ty)", "10%", f"{RI_TARGET_PRICE:,.0f}"],
        ["P/S + P/E theo mảng (TGDĐ/ĐMX/BHX/Khác)", "50%", f"{PS_PE_TARGET_PRICE:,.0f}"],
        ["GIÁ MỤC TIÊU (bình quân trọng số)", "100%", f"{WEIGHTED_TARGET_PRICE:,.0f}"],
    ], colWidths=[80*mm, 25*mm, 45*mm])
    val_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 9.5),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('BACKGROUND', (0,-1), (-1,-1), HexColor("#D4EDDA")), ('FONTNAME', (0,-1), (-1,-1), PDF_FONT_BOLD),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(val_tbl)
    story.append(Spacer(1, 8))
    story.append(Image(make_valuation_chart(), width=165*mm, height=90*mm))
    story.append(PageBreak())

    story.append(Paragraph("CHI TIẾT ĐỊNH GIÁ RESIDUAL INCOME (RI) & CAPM", style_h1))
    story.append(Paragraph(
        f"Chi phí vốn CSH (COE) = Rf + Beta×ERP + Premium rủi ro đặc thù = {rf_val*100:.2f}% + {beta_val:.2f}×"
        f"{ERP*100:.1f}% + {SPECIFIC_RISK_PREMIUM*100:.1f}% = <b>{COE*100:.2f}%</b>. Beta điều chỉnh Blume từ Beta "
        f"thô {beta_raw:.2f} ({beta_src}).", style_body))
    ri_tbl = Table([
        ["Năm"] + [f"{y}" for y in years_fc],
        ["EPS dự phóng (VND)"] + [f"{v:,.0f}" for v in EPS_FC],
        ["BVPS đầu kỳ (VND)"] + [f"{v:,.0f}" for v in BVPS_FC],
        ["Residual Income (VND)"] + [f"{ri_results[i]:,.0f}" for i in range(len(years_fc))],
    ], colWidths=[45*mm] + [35*mm]*len(years_fc))
    ri_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BACKGROUND', (0,0), (0,-1), HexColor("#F0F0F0")), ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(ri_tbl)
    story.append(Paragraph(
        f"RI Value = BVPS hiện tại + PV(RI 3 năm) + PV(Continuing Value) = {RI_TARGET_PRICE:,.0f} VND. "
        f"Tăng trưởng dài hạn giả định {TERMINAL_GROWTH*100:.1f}%.", style_body))
    story.append(PageBreak())

    # ── Trang 15-16: Rủi ro & Catalysts ─────────────────────────────────────────────────────────
    story.append(Paragraph("RỦI RO & CATALYST", style_h1))
    story.append(Paragraph("Ma trận rủi ro:", style_h2))
    risk_matrix = [
        ["Rủi ro", "Xác suất", "Mức độ ảnh hưởng", "Ghi chú"],
        ["TGDĐ/ĐMX tăng trưởng chậm lại", "Trung bình", "Cao", "2 chuỗi lớn nhất, ảnh hưởng trực tiếp KQKD"],
        ["BHX mở rộng quá nhanh, ramp-up chậm", "Trung bình", "Trung bình", "Ảnh hưởng biên LN ngắn hạn, không cấu trúc"],
        ["Cạnh tranh TMĐT (Shopee, TikTok Shop)", "Cao", "Trung bình", "Đã phản ứng bằng kênh online + Super App"],
        ["Sức mua vĩ mô suy yếu", "Thấp-Trung bình", "Cao", "SSSG hiện đang rất tích cực (+33%)"],
    ]
    risk_tbl = Table(risk_matrix, colWidths=[60*mm, 25*mm, 30*mm, 55*mm])
    risk_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('ALIGN', (1,0), (2,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(risk_tbl)
    story.append(Paragraph("Catalyst tiềm năng:", style_h2))
    for c in ["Chu kỳ ra mắt iPhone mới (tháng 9 hàng năm) thúc đẩy doanh số TGDĐ/ĐMX",
              "BHX đạt quy mô lợi nhuận ổn định, chứng minh mô hình bền vững sau giai đoạn mở rộng nhanh",
              "EraBlue (Indonesia) nhân rộng thành công, mở ra câu chuyện tăng trưởng quốc tế mới"]:
        story.append(Paragraph(f"• {c}", style_body))
    story.append(PageBreak())

    # ── Trang 17: Kết luận ──────────────────────────────────────────────────────────────────────
    story.append(Paragraph("KẾT LUẬN", style_h1))
    story.append(Paragraph(
        f"<b>Khuyến nghị: {RECOMMEND}</b> với giá mục tiêu <b>{WEIGHTED_TARGET_PRICE:,.0f} VND</b> "
        f"(upside {UPSIDE_PCT:+.1f}%). MWG là cổ phiếu đầu ngành bán lẻ Việt Nam với danh mục đa dạng, minh bạch "
        f"thông tin, và đang trong giai đoạn tăng tốc tăng trưởng (BHX) kết hợp phục hồi chu kỳ (TGDĐ/ĐMX).",
        style_body))
    story.append(Paragraph("Chỉ báo theo dõi (Leading Indicators):", style_h2))
    li_pdf = [
        ["Chỉ báo", "Ngưỡng tích cực", "Giá trị hiện tại"],
        ["SSSG toàn công ty", "> 10%", "33%"],
        ["Tăng trưởng DT YoY", "> 15%", "29.3% (5T2026)"],
        ["Biên LNG", "> 20%", f"{GPM_FC_PCT:.1f}%"],
    ]
    li_tbl = Table(li_pdf, colWidths=[60*mm, 45*mm, 45*mm])
    li_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), PDF_FONT), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1F4E79")), ('TEXTCOLOR', (0,0), (-1,0), HexColor("#FFFFFF")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.4, HexColor("#CCCCCC")),
    ]))
    story.append(li_tbl)
    story.append(PageBreak())

    # ── Trang 18: Phụ lục ───────────────────────────────────────────────────────────────────────
    story.append(Paragraph("PHỤ LỤC — NGUỒN DỮ LIỆU", style_h1))
    story.append(Paragraph(
        "Dữ liệu tài chính: Vietcap (BCTC hợp nhất, quý gần nhất Q1/2026). Dữ liệu vận hành (số cửa hàng, doanh thu "
        "theo mảng): Investor Relations MWG (mwg.vn/bao-cao), báo cáo cập nhật tình hình kinh doanh hàng tháng, "
        "công bố trực tiếp bởi công ty. Lãi suất phi rủi ro: "
        f"{rf_src}. Beta: {beta_src}.", style_small))
    story.append(Paragraph(
        f"Báo cáo được tạo tự động ngày {datetime.now().strftime('%d/%m/%Y %H:%M')}. Đây là tài liệu phân tích "
        "tham khảo, không phải khuyến nghị đầu tư chính thức. Nhà đầu tư cần tự đánh giá rủi ro trước khi ra "
        "quyết định.", style_small))

    doc = SimpleDocTemplate(PDF_FILE, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm,
                             leftMargin=18*mm, rightMargin=18*mm)
    doc.build(story)
    print(f"[PDF] Saved: {PDF_FILE}")

# ══════════════════════════════════════════════════════════════════════════════════════════════
# JSON EXPORT — data/MWG.json cho web dashboard (schema bắt buộc theo skill xuat-bao-cao)
# ══════════════════════════════════════════════════════════════════════════════════════════════
def save_json_summary():
    RECOMMEND = "MUA" if UPSIDE_PCT and UPSIDE_PCT > 15 else ("BÁN" if UPSIDE_PCT and UPSIDE_PCT < -5 else "THEO DÕI")

    quarter_labels = [f"{r['year']}-Q{r['quarter']}" for r in MWG_RATIOS if r.get("quarter") in (1,2,3,4)]
    pe_quarters = [round(r["pe"], 2) if r.get("pe") else None for r in MWG_RATIOS if r.get("quarter") in (1,2,3,4)]
    pb_quarters = [round(r["pb"], 2) if r.get("pb") else None for r in MWG_RATIOS if r.get("quarter") in (1,2,3,4)]

    thesis = [
        f"MWG dẫn đầu bán lẻ đa ngành VN với {sum(STORE_COUNT_NOW.values()):,} cửa hàng (TGDĐ/ĐMX dẫn đầu ICT, "
        f"BHX dẫn đầu FMCG hiện đại), tận dụng quy mô đàm phán giá tốt với NCC.",
        f"BHX có lãi từ 2023, đang tăng tốc mở rộng (+{STORE_COUNT_NOW['BHX']-2559:,} CH trong 5T2026), SSSG cao "
        "nhất công ty — dư địa tăng trưởng dài hạn lớn.",
        f"Doanh thu 5T2026 đạt {REV_5T_2026_TOTAL:,.0f} tỷ (+29,3% svck), SSSG TGDĐ/ĐMX ~33% nhờ chu kỳ phục hồi "
        "sức mua và nâng cấp thiết bị ICT.",
    ]
    risks = [
        "TGDĐ/ĐMX đã bão hòa số lượng cửa hàng, tăng trưởng phụ thuộc chu kỳ sản phẩm và SSSG.",
        "BHX mở rộng nhanh (+1.000 CH kế hoạch 2026) cần 3 tháng ramp-up — rủi ro pha loãng biên LN ngắn hạn.",
        "P/E median lịch sử bao gồm giai đoạn tăng trưởng nóng 2021, có thể không bền vững nếu thị trường re-rate.",
    ]
    moats = {
        "Network Effect": {"score": 2, "desc": "Hạn chế — bán lẻ truyền thống không có hiệu ứng mạng lưới mạnh."},
        "Cost Advantage": {"score": 4, "desc": "Quy mô >6.700 CH giúp đàm phán giá tốt với NCC ICT/FMCG."},
        "Switching Cost": {"score": 2, "desc": "Thấp với ICT (khách dễ chuyển kênh), trung bình với BHX (tiện lợi vị trí)."},
        "Intangible Assets": {"score": 3, "desc": "Thương hiệu TGDĐ/ĐMX/BHX nhận diện rộng nhưng ICT ít trung thành thương hiệu."},
        "Efficient Scale": {"score": 4, "desc": "Mạng lưới cửa hàng lớn nhất ngành, tối ưu logistics/quảng cáo theo quy mô."},
    }
    pestle = [
        {"factor": "Political", "content": "Kiểm soát chặt giá thuốc (An Khang) và an toàn thực phẩm (BHX); thuế VAT ICT ổn định.", "impact": "Neutral"},
        {"factor": "Economic", "content": "Sức mua phục hồi, CPI ổn định 3-4%/năm hỗ trợ tiêu dùng không thiết yếu tăng trở lại.", "impact": "Positive"},
        {"factor": "Social", "content": "Xu hướng hiện đại hoá kênh mua sắm hỗ trợ BHX thay thế kênh truyền thống.", "impact": "Positive"},
        {"factor": "Technological", "content": "Chu kỳ nâng cấp iPhone/AI on-device thúc đẩy nhu cầu thay máy; đầu tư Super App/trả chậm.", "impact": "Positive"},
        {"factor": "Legal", "content": "Yêu cầu hoá đơn điện tử/truy xuất nguồn gốc ngày càng chặt — chi phí tuân thủ tăng nhẹ.", "impact": "Neutral"},
        {"factor": "Environmental", "content": "Áp lực giảm rác thải nhựa, chuỗi lạnh tiêu tốn năng lượng cho hàng tươi sống.", "impact": "Negative"},
    ]
    comments = {
        "businessModel": f"MWG vận hành 3 chuỗi bán lẻ chính: TGDĐ+ĐMX ({STORE_COUNT_NOW['TGDD']+STORE_COUNT_NOW['DMX']:,} "
            f"cửa hàng, ICT/điện máy, đã bão hòa) và BHX ({STORE_COUNT_NOW['BHX']:,} cửa hàng, FMCG, tăng trưởng "
            "nhanh nhất, có lãi từ 2023). Mô hình dựa trên quy mô mạng lưới để tối ưu chi phí vận hành/mua hàng.",
        "financialPerformance": f"Doanh thu {years_fc[0]}E dự phóng {revenue_fc[0]:,.0f} tỷ (+{(revenue_fc[0]/revenue_hist[-1]-1)*100:.1f}% "
            f"YoY), biên LNG duy trì {GPM_FC_PCT:.1f}%, LNST Cổ đông mẹ {ni_fc[0]:,.0f} tỷ. Đòn bẩy thấp, Net Debt "
            "âm (net cash), xu hướng cải thiện nhờ FCF chuyển dương khi BHX hoà vốn.",
        "valuationText": f"Giá mục tiêu {WEIGHTED_TARGET_PRICE:,.0f} VND (upside {UPSIDE_PCT:+.1f}%) theo mô hình 4 "
            "phương pháp (P/E 20%, P/B 20%, RI 10%, P/S+P/E theo từng mảng 50%). Khuyến nghị "
            f"{RECOMMEND}.",
    }
    ratios = {
        "gpm": [round(g/100, 4) for g in gpm_hist] + [round(g/100, 4) for g in gp_margin_fc],
        "sga_pct": [round(sga_hist[i]/revenue_hist[i], 4) for i in range(N_HIST)] + [round(s/100, 4) for s in SGA_FC_PCT],
        "roe": [round(ni_hist[i]/equity_hist[i], 4) for i in range(N_HIST)] + [round(ni_fc[i]/equity_fc[i], 4) for i in range(len(years_fc))],
        "net_margin": [round(ni_hist[i]/revenue_hist[i], 4) for i in range(N_HIST)] + [round(ni_fc[i]/revenue_fc[i], 4) for i in range(len(years_fc))],
    }

    payload = {
        "ticker": TICKER,
        "companyName": COMPANY,
        "sector": INDUSTRY,
        "currentPrice": PRICE,
        "marketCap": MARKET_CAP * 1e9,
        "shares": SHARES * 1e6,
        "gdriveExcelUrl": None,
        "gdrivePdfUrl": None,
        "data": {
            "years": years_all,
            "revenue": [round(v, 1) for v in revenue_hist + revenue_fc],
            "npat": [round(v, 1) for v in ni_hist + ni_fc],
            "eps": [round(v, 0) for v in eps_hist + EPS_FC],
            "equity": [round(v, 1) for v in equity_hist + equity_fc],
        },
        "segments": {
            "names": SEGMENTS,
            "storeCountNow": STORE_COUNT_NOW,
            "revenueForecast": {seg: REVENUE_FC_SEGMENT[seg] for seg in SEGMENTS},
            "efficiencyNow": {seg: round(EFFICIENCY_NOW[seg], 3) for seg in SEGMENTS},
        },
        "thesis": thesis,
        "risks": risks,
        "moats": moats,
        "pestle": pestle,
        "valuation": {
            "bear": round(min(PE_TARGET_PRICE, PB_TARGET_PRICE, RI_TARGET_PRICE, PS_PE_TARGET_PRICE) * 0.9, 0),
            "base": WEIGHTED_TARGET_PRICE,
            "bull": round(max(PE_TARGET_PRICE, PB_TARGET_PRICE, RI_TARGET_PRICE) * 1.05, 0),
            "methods": {
                "pe": PE_TARGET_PRICE, "pb": PB_TARGET_PRICE, "ri": RI_TARGET_PRICE,
                "ps_pe": PS_PE_TARGET_PRICE,
            },
            "recommendation": RECOMMEND,
            "upsidePct": UPSIDE_PCT,
        },
        "comments": comments,
        "pe_hist": [round(stats.median([r["pe"] for r in MWG_RATIOS if r.get("year") == y and r.get("pe")]), 2)
                    for y in _hist_years_ratio if any(r.get("year") == y and r.get("pe") for r in MWG_RATIOS)],
        "pb_hist": [round(stats.median([r["pb"] for r in MWG_RATIOS if r.get("year") == y and r.get("pb")]), 2)
                    for y in _hist_years_ratio if any(r.get("year") == y and r.get("pb") for r in MWG_RATIOS)],
        "pe_quarters": pe_quarters,
        "pb_quarters": pb_quarters,
        "quarter_labels": quarter_labels,
        "ratios": ratios,
    }
    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[JSON] Saved: {JSON_FILE}")


if __name__ == "__main__":
    build_excel()
    build_pdf()
    save_json_summary()
    print("[Main] MWG model (Excel + PDF + JSON) built successfully.")

