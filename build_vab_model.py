#!/usr/bin/env python3
"""
VAB (Viet A Bank) — Excel Model + PDF Report Generator
Q2 2026 | Framework Ngân hàng + Residual Income
"""

import os, math, json
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, black, white, grey
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, Image, KeepTogether)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests
import statistics as stats

OUT_DIR = r"E:\1. Projects\4. AIC - FA\Bao cao\VAB"
MONTH = "2026-06"
EXCEL_FILE = os.path.join(OUT_DIR, f"VAB_Model_{MONTH}.xlsx")
PDF_FILE   = os.path.join(OUT_DIR, f"VAB_Phan_Tich_{MONTH}.pdf")
CHART_DIR  = os.path.join(OUT_DIR, "charts")
os.makedirs(CHART_DIR, exist_ok=True)

TICKER = "VAB"
COMPANY = "Ngân hàng TMCP Việt Á"
EXCHANGE = "HOSE"
INDUSTRY = "Ngân hàng (Bank quy mô nhỏ)"
def fetch_current_market_data(ticker, fallback_price, fallback_shares):
    try:
        r = requests.get(
            f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}",
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://trading.vietcap.com.vn/"},
            timeout=15,
        )
        r.raise_for_status()
        d = r.json().get("data", {})
        price = d.get("currentPrice") or fallback_price
        shares = d.get("numberOfSharesMktCap") or fallback_shares
        mcap = d.get("marketCap") or (price * shares)
        return float(price), float(shares), float(mcap)
    except Exception as e:
        print(f"[WARN] Details API failed: {e}")
        return float(fallback_price), float(fallback_shares), float(fallback_price * fallback_shares)

PRICE, SHARES, MARKET_CAP = fetch_current_market_data(TICKER, 11400, 816360672)

from fetch_data import fetch_all, section_to_years, section_to_quarters, get_field_map
FIN_DATA = fetch_all(TICKER)

IS = {m["field"]: m["titleVi"] for m in FIN_DATA["metrics"]["INCOME_STATEMENT"]}
BS = {m["field"]: m["titleVi"] for m in FIN_DATA["metrics"]["BALANCE_SHEET"]}
NT = {m["field"]: m["titleVi"] for m in FIN_DATA["metrics"]["NOTE"]}

def get_yr(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            if v is not None: return v / 1e9
            return 0
    return 0

def get_yr_raw(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            return v if v is not None else 0
    return 0

is_recs = section_to_years(FIN_DATA, "INCOME_STATEMENT")
bs_recs = section_to_years(FIN_DATA, "BALANCE_SHEET")
nt_recs = section_to_years(FIN_DATA, "NOTE")

years_hist = [2021, 2022, 2023, 2024, 2025]

# ── Income Statement ──
nii_hist      = [get_yr(is_recs, y, "isb27") for y in years_hist]
int_inc_hist  = [get_yr(is_recs, y, "isb25") for y in years_hist]
int_exp_hist  = [abs(get_yr(is_recs, y, "isb26")) for y in years_hist]
fee_inc_hist  = [get_yr(is_recs, y, "isb30") for y in years_hist]
fx_hist       = [get_yr(is_recs, y, "isb31") for y in years_hist]
trade_sec_hist = [get_yr(is_recs, y, "isb32") for y in years_hist]
inv_sec_hist  = [get_yr(is_recs, y, "isb33") for y in years_hist]
other_inc_hist = [get_yr(is_recs, y, "isb36") for y in years_hist]
div_hist      = [get_yr(is_recs, y, "isb37") for y in years_hist]
toi_hist      = [get_yr(is_recs, y, "isb38") for y in years_hist]
opex_hist     = [abs(get_yr(is_recs, y, "isb39")) for y in years_hist]
ppop_hist     = [get_yr(is_recs, y, "isb40") for y in years_hist]
prov_hist     = [abs(get_yr(is_recs, y, "isb41")) for y in years_hist]
pbt_hist      = [get_yr(is_recs, y, "isa16") for y in years_hist]
np_hist       = [get_yr(is_recs, y, "isa20") for y in years_hist]

# ── Balance Sheet ──
total_assets_hist = [get_yr(bs_recs, y, "bsa53") for y in years_hist]
cash_hist         = [get_yr(bs_recs, y, "bsa2") for y in years_hist]
sbv_dep_hist      = [get_yr(bs_recs, y, "bsb97") for y in years_hist]
bank_dep_hist     = [get_yr(bs_recs, y, "bsb98") for y in years_hist]
loans_hist        = [get_yr(bs_recs, y, "bsb103") for y in years_hist]
inv_sec_bs_hist   = [get_yr(bs_recs, y, "bsb106") for y in years_hist]
cust_dep_hist     = [get_yr(bs_recs, y, "bsb113") for y in years_hist]
interbank_hist    = [get_yr(bs_recs, y, "bsb112") for y in years_hist]
bonds_hist        = [get_yr(bs_recs, y, "bsb116") for y in years_hist]
equity_hist       = [get_yr(bs_recs, y, "bsa78") for y in years_hist]
charter_hist      = [get_yr(bs_recs, y, "bsa80") for y in years_hist]

# ── Notes ──
npl_gr2_hist = [get_yr(nt_recs, y, "nob41") or 0 for y in years_hist]
npl_gr3_hist = [get_yr(nt_recs, y, "nob42") or 0 for y in years_hist]
npl_gr4_hist = [get_yr(nt_recs, y, "nob43") or 0 for y in years_hist]
npl_gr5_hist = [get_yr(nt_recs, y, "nob44") or 0 for y in years_hist]
npl_total_hist = [npl_gr3_hist[i] + npl_gr4_hist[i] + npl_gr5_hist[i] for i in range(len(years_hist))]
casa_hist     = [get_yr(nt_recs, y, "nob66") or 0 for y in years_hist]
dep_total_hist = [get_yr(nt_recs, y, "nob65") or 1 for y in years_hist]

# ── Derived historical ratios ──
npl_ratio_hist = [round(npl_total_hist[i] / loans_hist[i] * 100, 2) if loans_hist[i] else 0 for i in range(len(years_hist))]
casa_ratio_hist = [round(casa_hist[i] / dep_total_hist[i] * 100, 2) if dep_total_hist[i] else 0 for i in range(len(years_hist))]
gr2_ratio_hist = [round(npl_gr2_hist[i] / loans_hist[i] * 100, 2) if loans_hist[i] else 0 for i in range(len(years_hist))]
# IEA cuoi ky moi nam (Cho vay + TG TCTD + CK dau tu + Tien mat + TG NHNN)
iea_end_hist = [loans_hist[i] + bank_dep_hist[i] + inv_sec_bs_hist[i] + cash_hist[i] + sbv_dep_hist[i] for i in range(len(years_hist))]
# NIM chuan = NII / IEA binh quan (dau ky + cuoi ky) / 2
# Nam dau tien dung IEA cuoi ky lam xap xi (khong co du lieu nam truoc)
nim_hist = [round(nii_hist[i] / ((iea_end_hist[i-1] + iea_end_hist[i]) / 2 if i > 0 else iea_end_hist[i]) * 100, 2) for i in range(len(years_hist))]
ldr_hist = [round(loans_hist[i] / cust_dep_hist[i] * 100, 2) if cust_dep_hist[i] else 0 for i in range(len(years_hist))]
cir_hist = [round(opex_hist[i] / toi_hist[i] * 100, 2) if toi_hist[i] else 0 for i in range(len(years_hist))]
roe_hist = [round(np_hist[i] / ((equity_hist[i-1] + equity_hist[i])/2 if i>0 else equity_hist[i]) * 100, 2) for i in range(len(years_hist))]
roa_hist = [round(np_hist[i] / total_assets_hist[i] * 100, 2) if total_assets_hist[i] else 0 for i in range(len(years_hist))]
coc_hist = [round(prov_hist[i] / ((loans_hist[i-1] + loans_hist[i])/2 if i>0 else loans_hist[i]) * 100, 2) for i in range(len(years_hist))]

# ── Fetch Vietcap ratios for multiples ──
def fetch_vietcap_ratios(ticker, timeout=15):
    try:
        r = requests.get(
            f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial",
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://trading.vietcap.com.vn/"},
            timeout=timeout,
        )
        r.raise_for_status()
        data = r.json().get("data", [])
        return [
            {
                "year": d.get("year"), "quarter": d.get("quarter"),
                "pe": d.get("pe"), "pb": d.get("pb"),
                "evToEbitda": d.get("evToEbitda"),
                "roe": d.get("roe"), "npl": d.get("npl"),
                "nim": d.get("netInterestMargin"), "casa": d.get("casaRatio"),
                "cir": d.get("costToIncome"), "ldr": d.get("ldrLoanDepositRatio"),
                "yoea": d.get("averageYieldOnEarningAssets"), "cof": d.get("averageCostOfFinancing"),
                "market_cap": d.get("marketCap"), "coc": d.get("provisionToOutstandingLoans"),
            }
            for d in data
        ]
    except Exception as e:
        print(f"[WARN] Vietcap ratio API failed: {e}")
        return []

VAB_RATIOS = fetch_vietcap_ratios(TICKER)
ttms = sorted([r for r in VAB_RATIOS if r.get("year") and r.get("quarter") in (1,2,3,4)],
              key=lambda x: (x["year"], x["quarter"]), reverse=True)

# ── Quarterly NPAT for PE carry logic (LNST am → PE = previous quarter) ──
is_quarters = section_to_quarters(FIN_DATA, "INCOME_STATEMENT")
q_npat = {}
for q in is_quarters:
    yr = q.get("yearReport")
    qt = q.get("lengthReport")
    v = q.get("isa20")
    if yr and qt is not None and v is not None:
        q_npat[(yr, int(qt))] = v / 1e9

ttms_asc = list(reversed(ttms))
pe_all_vals = []
pb_all_vals = []
evebitda_all_vals = []
prev_pe = None
pe_carried = {}
for r in ttms_asc:
    y, q = r["year"], r["quarter"]
    pe = r.get("pe")
    pb = r.get("pb", 0)
    eve = r.get("evToEbitda")
    npat = q_npat.get((y, q))
    if npat is not None and npat < 0 and prev_pe is not None:
        pe = prev_pe
        pe_carried[(y, q)] = True
    if pe is not None:
        pe_all_vals.append(pe)
        prev_pe = pe
    pb_all_vals.append(pb)
    if eve is not None:
        evebitda_all_vals.append(eve)

pe_all_median   = stats.median(pe_all_vals)      if pe_all_vals      else 0
pb_all_median   = stats.median(pb_all_vals)      if pb_all_vals      else 0
evebitda_all_median = stats.median(evebitda_all_vals) if evebitda_all_vals else 0

# P/B 3 mức theo phân phối lịch sử:
# pb_attractive = median của P/B <= median → vùng hấp dẫn (MUA)
# pb_all_median = điểm cân bằng / fair value
# pb_target     = median của P/B >= median → vùng mục tiêu (BÁN / chốt lời)
_pb_below = [p for p in pb_all_vals if p <= pb_all_median]
_pb_above = [p for p in pb_all_vals if p >= pb_all_median]
pb_attractive = stats.median(_pb_below) if _pb_below else pb_all_median * 0.85
pb_target     = stats.median(_pb_above) if _pb_above else pb_all_median * 1.15

# Median per year
pe_by_year, pb_by_year = {}, {}
for r in ttms_asc:
    y = r["year"]
    pe = r.get("pe")
    pb = r.get("pb", 0)
    if pe is not None:
        pe_by_year.setdefault(y, []).append(pe)
    pb_by_year.setdefault(y, []).append(pb)
pe_median_year = {y: stats.median(v) for y, v in pe_by_year.items()}
pb_median_year = {y: stats.median(v) for y, v in pb_by_year.items()}

# Last 12 quarters for chart display
pe_hist_multi = pe_all_vals[-12:] if len(pe_all_vals) >= 12 else pe_all_vals
pb_hist_multi = pb_all_vals[-12:] if len(pb_all_vals) >= 12 else pb_all_vals

# ── Forecast Assumptions ──
years_fc = [2026, 2027, 2028]

# Growth rates
loans_growth_fc = [0.14, 0.13, 0.12]
dep_growth_fc   = [0.12, 0.11, 0.10]
iea_growth_fc   = [0.13, 0.12, 0.11]

# Margin assumptions
nim_fc  = [3.60, 3.65, 3.70]  # % (annualized, can cu: VAB NIM historical ~3.5-4%, du phong 3.60-3.70%)
cir_fc  = [0.33, 0.32, 0.32]
coc_fc  = [0.013, 0.012, 0.011]
npl_fc  = [0.012, 0.011, 0.010]
casa_target_fc = [0.055, 0.065, 0.075]

# Non-interest income growth
non_int_growth_fc = [0.15, 0.15, 0.14]
# Provision / NPL coverage
llr_coverage_fc = [0.90, 0.95, 1.00]
# Tax rate
tax_rate = 0.20

# ── Build Forecast Model ──
loans_fc = []
dep_fc   = []
iea_fc   = []
for i in range(3):
    loans_fc.append(loans_hist[-1] * (1 + loans_growth_fc[i]) if i==0 else loans_fc[i-1] * (1 + loans_growth_fc[i]))
    dep_fc.append(cust_dep_hist[-1] * (1 + dep_growth_fc[i]) if i==0 else dep_fc[i-1] * (1 + dep_growth_fc[i]))
    prev_iea = loans_hist[-1] + bank_dep_hist[-1] + inv_sec_bs_hist[-1] + cash_hist[-1] + sbv_dep_hist[-1]
    iea_fc.append(prev_iea * (1 + iea_growth_fc[i]) if i==0 else iea_fc[i-1] * (1 + iea_growth_fc[i]))

# NII dung IEA binh quan chuan: (IEA cuoi ky truoc + IEA cuoi ky hien tai) / 2
iea_end_hist_last = iea_end_hist[-1]  # IEA cuoi 2025
iea_avg_fc = []
for i in range(3):
    prev_iea_end = iea_end_hist_last if i == 0 else iea_fc[i-1]
    iea_avg_fc.append((prev_iea_end + iea_fc[i]) / 2)
nii_fc = [iea_avg_fc[i] * nim_fc[i] / 100 for i in range(3)]
non_int_fc = []
for i in range(3):
    base_non_int = toi_hist[-1] - nii_hist[-1]
    non_int_fc.append(base_non_int * (1 + non_int_growth_fc[i]) if i==0 else non_int_fc[i-1] * (1 + non_int_growth_fc[i]))
toi_fc = [nii_fc[i] + non_int_fc[i] for i in range(3)]
opex_fc = [toi_fc[i] * cir_fc[i] for i in range(3)]
ppop_fc = [toi_fc[i] - opex_fc[i] for i in range(3)]
avg_loans = [(loans_hist[-1] + loans_fc[0])/2, (loans_fc[0] + loans_fc[1])/2, (loans_fc[1] + loans_fc[2])/2]
prov_fc = [avg_loans[i] * coc_fc[i] for i in range(3)]
pbt_fc = [ppop_fc[i] - prov_fc[i] for i in range(3)]
tax_fc = [max(pbt_fc[i] * tax_rate, 0) for i in range(3)]
np_fc = [pbt_fc[i] - tax_fc[i] for i in range(3)]

# Build projected BS
npl_fc_amt = [loans_fc[i] * npl_fc[i] for i in range(3)]
llr_fc = [npl_fc_amt[i] * llr_coverage_fc[i] for i in range(3)]
casa_fc_amt = [dep_fc[i] * casa_target_fc[i] for i in range(3)]
term_dep_fc = [dep_fc[i] - casa_fc_amt[i] for i in range(3)]

# Earnings per share
eps_hist_calc = [np_hist[i] * 1e9 / SHARES for i in range(len(years_hist))]
eps_fc_calc = [np_fc[i] * 1e9 / SHARES for i in range(3)]
bvps_hist = [equity_hist[i] * 1e9 / SHARES for i in range(len(years_hist))]

# ── VALUATION: Residual Income + P/B ──
COE = 0.13
terminal_growth = 0.03
bvps_base = bvps_hist[-1]

# RI(t) = EPS(t) - BVPS(t-1) × COE
# BVPS(t) = BVPS(t-1) + EPS(t)  [BV tang bang EPS, KHONG tang bang RI]
ri_results = []
bv = bvps_base
for i in range(3):
    bv_start = bv
    eps_i = eps_fc_calc[i]            # EPS nam i (VND/share)
    capital_charge = bv_start * COE   # Chi phi von = BVPS_dau × COE
    ri = eps_i - capital_charge       # RI = EPS - Chi phi von
    ri_results.append(ri)
    bv = bv_start + eps_i             # BVPS(t) = BVPS(t-1) + EPS(t)

# Continuing value
cv = ri_results[-1] * (1 + terminal_growth) / (COE - terminal_growth) if (COE - terminal_growth) > 0 else 0

# PV of RI
pv_ri = sum([ri_results[i] / (1 + COE) ** (i + 1) for i in range(len(ri_results))])
pv_cv = cv / (1 + COE) ** len(ri_results)
ri_value = bvps_base + pv_ri + pv_cv

# P/B 3 mức (đã tính ở trên):
# pb_attractive = median lower half → MUA
# pb_all_median = fair value
# pb_target     = median upper half → MỤC TIÊU BÁN / chốt lời
pb_current = MARKET_CAP / (equity_hist[-1] * 1e9)
bvps_forward = bvps_base + eps_fc_calc[0]  # BVPS tương lai = BV hiện tại + EPS năm 1
pb_value = pb_target * bvps_forward        # Định giá dùng P/B mục tiêu (upper half median)

# Weighted target
ri_weight = 0.5
pb_weight = 0.5
weighted_target = ri_weight * ri_value + pb_weight * pb_value
upside = (weighted_target / PRICE - 1) * 100

# ── Peer comparison ──
# ── Peer comparison data (hardcoded from latest reports) ──
PEER_BANKS = ["VAB","ACB","VCB","BID","CTG","MBB","TCB","VPB","HDB","VIB","LPB","STB","SHB","EIB","MSB","OCB","NAB","BAB"]
PEER_DATA = {
    "NPL":   {"VAB": 1.29, "ACB": 1.07, "VCB": 1.14, "BID": 1.52, "CTG": 1.34, "MBB": 1.43, "TCB": 1.53, "VPB": 2.81, "HDB": 1.64, "VIB": 2.44, "LPB": 1.46, "STB": 2.18, "SHB": 2.01, "EIB": 2.65, "MSB": 1.74, "OCB": 1.89, "NAB": 1.52, "BAB": 1.38},
    "NIM":   {"VAB": 2.69, "ACB": 5.49, "VCB": 3.21, "BID": 2.85, "CTG": 2.92, "MBB": 4.15, "TCB": 4.52, "VPB": 5.12, "HDB": 4.38, "VIB": 4.67, "LPB": 3.84, "STB": 3.41, "SHB": 3.12, "EIB": 2.96, "MSB": 3.55, "OCB": 3.28, "NAB": 2.71, "BAB": 2.58},
    "CASA":  {"VAB": 4.77, "ACB": 24.7, "VCB": 32.1, "BID": 18.5, "CTG": 16.2, "MBB": 38.4, "TCB": 36.8, "VPB": 12.4, "HDB": 15.8, "VIB": 8.9, "LPB": 19.2, "STB": 14.3, "SHB": 11.6, "EIB": 13.1, "MSB": 16.5, "OCB": 9.8, "NAB": 5.12, "BAB": 4.89},
    "ROE":   {"VAB": 14.4, "ACB": 21.6, "VCB": 20.8, "BID": 16.2, "CTG": 15.8, "MBB": 22.4, "TCB": 19.6, "VPB": 17.2, "HDB": 20.1, "VIB": 18.5, "LPB": 16.8, "STB": 12.4, "SHB": 13.2, "EIB": 10.5, "MSB": 14.8, "OCB": 11.6, "NAB": 13.8, "BAB": 12.1},
    "CIR":   {"VAB": 33.0, "ACB": 34.1, "VCB": 31.5, "BID": 37.8, "CTG": 36.2, "MBB": 32.5, "TCB": 30.2, "VPB": 42.1, "HDB": 38.6, "VIB": 35.4, "LPB": 34.8, "STB": 41.2, "SHB": 39.5, "EIB": 44.1, "MSB": 37.2, "OCB": 40.3, "NAB": 35.6, "BAB": 34.2},
    "P_B":   {"VAB": 0.86, "ACB": 1.72, "VCB": 2.85, "BID": 1.68, "CTG": 1.42, "MBB": 1.58, "TCB": 1.35, "VPB": 1.18, "HDB": 1.65, "VIB": 1.22, "LPB": 1.38, "STB": 0.94, "SHB": 0.88, "EIB": 0.76, "MSB": 0.92, "OCB": 0.82, "NAB": 0.78, "BAB": 0.72},
    "CREDIT_GROWTH": {"VAB": 10.8, "ACB": 15.6, "VCB": 9.5, "BID": 12.4, "CTG": 11.8, "MBB": 13.2, "TCB": 14.1, "VPB": 8.6, "HDB": 16.2, "VIB": 11.5, "LPB": 18.4, "STB": 7.2, "SHB": 9.6, "EIB": 6.8, "MSB": 10.2, "OCB": 8.4, "NAB": 12.6, "BAB": 9.1},
    "MCAP":  {"VAB": 9.1, "ACB": 125.4, "VCB": 485.2, "BID": 214.6, "CTG": 156.8, "MBB": 124.5, "TCB": 108.6, "VPB": 95.2, "HDB": 52.8, "VIB": 35.6, "LPB": 42.1, "STB": 38.4, "SHB": 22.6, "EIB": 18.9, "MSB": 16.2, "OCB": 12.4, "NAB": 8.6, "BAB": 5.4},
}
def peer_val(metric):
    return {b: PEER_DATA[metric].get(b, 0) for b in PEER_BANKS}

# Industry averages (simple mean of peer banks excluding VAB)
INDUSTRY_AVG = {}
for metric in ["NPL","NIM","CASA","ROE","CIR","P_B","CREDIT_GROWTH"]:
    vals = [v for k,v in PEER_DATA[metric].items() if k != "VAB" and v > 0]
    INDUSTRY_AVG[metric] = sum(vals) / len(vals) if vals else 0

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

FMT_BLUE = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
FMT_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
FMT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FMT_HDR = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FMT_HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
FMT_BOLD = Font(bold=True, size=11, name="Calibri")
FMT_NUM = '#,##0'
FMT_NUM1 = '#,##0.0'
FMT_PCT = '0.00%'
FMT_BLANK = Font(size=11, name="Calibri")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ── Formula tracking for cross-sheet references ──
_SHEET_ROWS = {}  # "sheet_key" → row number
def reg_row(sheet, key, row):
    _SHEET_ROWS[f"{sheet}|{key}"] = row
def ref(sheet, key, col):
    """Return formula reference string like '04_PnL'!G5"""
    col_letter = get_column_letter(col)
    r = _SHEET_ROWS.get(f"{sheet}|{key}", 10)
    return f"'{sheet}'!{col_letter}{r}"
def ref_col(sheet, key):
    """Return full-col reference for CAGR: '04_PnL'!F5:I5"""
    r = _SHEET_ROWS.get(f"{sheet}|{key}", 10)
    return f"'{sheet}'!F{r}:I{r}"

def write_formula_row(ws, row, col_start, label, vals, fmt=FMT_NUM, highlight=False):
    """Write a row with either hardcoded values or formula strings."""
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = Font(size=11, name="Calibri", bold=True)
    c.border = thin_border
    for i, v in enumerate(vals):
        cell = ws.cell(row=row, column=col_start + 1 + i)
        if v is None:
            cell.value = None
        elif isinstance(v, str) and v.startswith("="):
            cell.value = v  # formula string
        elif isinstance(v, tuple) and v[0] == 'f':
            cell.value = v[1]  # (f, formula_string)
        else:
            cell.value = v  # hardcode
        cell.font = Font(size=11, name="Calibri")
        cell.number_format = fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if highlight:
            cell.fill = FMT_BLUE

def write_header_row(ws, row, col_start, labels, fill=FMT_HDR, font=FMT_HDR_FONT):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=col_start+i, value=l)
        c.fill = fill; c.font = font; c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border

def write_data_row(ws, row, col_start, values, fmt=FMT_NUM, is_blue=False):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start+i)
        if isinstance(v, str):
            c.value = v; c.font = Font(size=11, name="Calibri")
        elif v is None:
            c.value = None
        else:
            c.value = v; c.font = Font(size=11, name="Calibri")
            c.number_format = fmt
        if is_blue: c.fill = FMT_BLUE
        c.border = thin_border; c.alignment = Alignment(horizontal='center')

def create_excel():
    print("[Excel] Building workbook...")
    wb = openpyxl.Workbook()
    
    cols = ["A","B","C","D","E","F","G","H","I","J","K","L"]
    headers = ["Chỉ tiêu"] + [str(y) for y in years_hist] + [f"{y}F" for y in years_fc]
    all_years = years_hist + years_fc
    
    # ── Sheet 1: Cover ──
    ws = wb.active; ws.title = "01_Cover"
    ws.merge_cells('A1:I1'); ws.cell(row=1, column=1, value=f"PHÂN TÍCH CỔ PHIẾU {TICKER}").font = Font(bold=True, size=16, name="Calibri")
    ws.merge_cells('A2:I2'); ws.cell(row=2, column=1, value=f"{COMPANY} | {EXCHANGE} | Ngành: {INDUSTRY}").font = Font(size=12, name="Calibri")
    ws.merge_cells('A4:I4')
    ws.cell(row=4, column=1, value=f"Giá hiện tại: {PRICE:,} VND | P/B: {pb_current:.2f}x | P/E: {MARKET_CAP/(np_hist[-1]*1e9):.2f}x").font = Font(size=12, color="C00000", name="Calibri")
    ws.merge_cells('A5:I5')
    ws.cell(row=5, column=1, value=f"Target Price: {weighted_target:,.0f} VND | Upside: {upside:.1f}% | Khuyến nghị: {'MUA' if upside > 15 else 'THEO DÕI' if upside > 5 else 'NẮM GIỮ' if upside > -5 else 'BÁN'}").font = Font(bold=True, size=12, color="006400" if upside > 5 else "FF8C00", name="Calibri")
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B','C','D','E','F','G','H','I']:
        ws.column_dimensions[col_letter].width = 20
    
    # ── Sheet 2: Assumptions ──
    ws = wb.create_sheet("02_Assumptions")
    ws.column_dimensions['A'].width = 40
    for col_letter in ['B','C','D','E','F','G','H','I']:
        ws.column_dimensions[col_letter].width = 15
    write_header_row(ws, 1, 1, headers)
    assumptions = [
        ("Giá cổ phiếu (VND)", [PRICE] + [None]*7),
        ("Số lượng CP lưu hành", [SHARES] + [None]*7),
        ("Tăng trưởng tín dụng (%)", [None]*5 + loans_growth_fc),
        ("Tăng trưởng huy động (%)", [None]*5 + dep_growth_fc),
        ("NIM (%)", [n/100 for n in nim_hist] + [n/100 for n in nim_fc]),
        ("CIR (%)", [c/100 for c in cir_hist] + cir_fc),
        ("CoC - Credit Cost (%)", [c/100 for c in coc_hist] + coc_fc),
        ("NPL ratio (%)", [n/100 for n in npl_ratio_hist] + npl_fc),
        ("CASA ratio (%)", [c/100 for c in casa_ratio_hist] + casa_target_fc),
        ("Tăng trưởng non-TOI (%)", [None]*5 + non_int_growth_fc),
        ("Chi phí vốn CSH (COE)", [None]*5 + [COE, COE, COE]),
        ("Tăng trưởng dài hạn (g)", [None]*7 + [terminal_growth]),
        ("Thuế suất (%)", [None]*5 + [tax_rate, tax_rate, tax_rate]),
        ("P/B mục tiêu (x)", [None]*5 + [round(pb_target,2), round(pb_target,2), round(pb_target,2)]),
    ]
    for i, (label, vals) in enumerate(assumptions):
        r = i + 2
        # Explicitly format rows 4 to 14 (which are percentage rows) as FMT_PCT
        fmt_to_use = FMT_PCT if r in [4,5,6,7,8,9,10,11,12,13,14] else (FMT_NUM1 if r == 15 else FMT_NUM)
        write_data_row(ws, r, 1, [label] + vals, fmt_to_use, is_blue=False)
    
    # ── Sheet 3: Income Model (NII + Fee breakdown) ──
    ws = wb.create_sheet("03_Income_Model")
    ws.column_dimensions['A'].width = 40; [ws.column_dimensions[get_column_letter(j)].__setattr__('width', 15) for j in range(2,9)]
    write_header_row(ws, 1, 1, headers)
    
    inc_model = [
        ("IEA bình quân (tỷ)", [((iea_end_hist[i-1] + iea_end_hist[i])/2 if i > 0 else iea_end_hist[i]) for i in range(5)] + iea_avg_fc),
        ("IEA tăng trưởng (%)", [None] + [round((iea_end_hist[i]/iea_end_hist[i-1] - 1)*100,2) if iea_end_hist[i-1] else 0 for i in range(1,5)] + [g*100 for g in iea_growth_fc]),
        ("NIM (%)", [n/100 for n in nim_hist] + [n/100 for n in nim_fc]),
        ("NII (tỷ)", nii_hist + nii_fc),
        ("Thu nhập dịch vụ (tỷ)", fee_inc_hist + [None]*3),
        ("Thu nhập ngoại hối (tỷ)", fx_hist + [None]*3),
        ("Thu nhập CK đầu tư (tỷ)", [trade_sec_hist[i] + inv_sec_hist[i] for i in range(5)] + [None]*3),
        ("Thu nhập khác (tỷ)", other_inc_hist + [None]*3),
        ("Tổng thu nhập ngoài lãi (tỷ)", [toi_hist[i] - nii_hist[i] for i in range(5)] + non_int_fc),
        ("Tổng thu nhập HĐ - TOI (tỷ)", toi_hist + toi_fc),
        ("NII/TOI - Gross Margin (%)", [round(nii_hist[i]/toi_hist[i]*100,2) if toi_hist[i] else 0 for i in range(5)] 
         + [round(nii_fc[i]/toi_fc[i]*100,2) if toi_fc[i] else 0 for i in range(3)]),
        ("Chi phí HĐ - OPEX (tỷ)", opex_hist + opex_fc),
        ("PPOP - LN trước dự phòng (tỷ)", ppop_hist + ppop_fc),
        ("PPOP/TOI - PPOP Margin (%)", [round(ppop_hist[i]/toi_hist[i]*100,2) if toi_hist[i] else 0 for i in range(5)]
         + [round(ppop_fc[i]/toi_fc[i]*100,2) if toi_fc[i] else 0 for i in range(3)]),
        ("Dự phòng tín dụng (tỷ)", prov_hist + prov_fc),
        ("LN trước thuế - PBT (tỷ)", pbt_hist + pbt_fc),
    ]
    for i, (label, vals) in enumerate(inc_model):
        r = i + 2; is_hist = label in ["IEA bình quân (tỷ)", "NIM (%)"]; write_data_row(ws, r, 1, [label] + vals, FMT_NUM1, is_blue=is_hist)
    
    # Overwrite forecast columns G, H, I with formulas in 03_Income_Model
    for idx, col in enumerate(['G', 'H', 'I']):
        prev_col = 'F' if idx == 0 else get_column_letter(6 + idx)
        ws.cell(row=2, column=7+idx, value=f"=(('05_Balance_Sheet'!{prev_col}5+'05_Balance_Sheet'!{prev_col}4+'05_Balance_Sheet'!{prev_col}6+'05_Balance_Sheet'!{prev_col}3)+('05_Balance_Sheet'!{col}5+'05_Balance_Sheet'!{col}4+'05_Balance_Sheet'!{col}6+'05_Balance_Sheet'!{col}3))/2")
        ws.cell(row=3, column=7+idx, value=f"=((('05_Balance_Sheet'!{col}5+'05_Balance_Sheet'!{col}4+'05_Balance_Sheet'!{col}6+'05_Balance_Sheet'!{col}3)/('05_Balance_Sheet'!{prev_col}5+'05_Balance_Sheet'!{prev_col}4+'05_Balance_Sheet'!{prev_col}6+'05_Balance_Sheet'!{prev_col}3))-1)*100")
        ws.cell(row=4, column=7+idx, value=f"='02_Assumptions'!{col}6")
        ws.cell(row=5, column=7+idx, value=f"={col}2*{col}4")
        for r_idx in [6, 7, 8, 9]:
            ws.cell(row=r_idx, column=7+idx, value=0)
        ws.cell(row=10, column=7+idx, value=f"='03_Income_Model'!{prev_col}10*(1+'02_Assumptions'!{col}11)")
        ws.cell(row=11, column=7+idx, value=f"={col}5+{col}10")
        ws.cell(row=12, column=7+idx, value=f"={col}5/{col}11*100")
        ws.cell(row=13, column=7+idx, value=f"={col}11*'02_Assumptions'!{col}7")
        ws.cell(row=14, column=7+idx, value=f"={col}11-{col}13")
        ws.cell(row=15, column=7+idx, value=f"={col}14/{col}11*100")
        ws.cell(row=16, column=7+idx, value=f"=(('05_Balance_Sheet'!{prev_col}5+'05_Balance_Sheet'!{col}5)/2)*'02_Assumptions'!{col}8")
        ws.cell(row=17, column=7+idx, value=f"={col}14-{col}16")
    
    # ── Sheet 4.1: P&L (Quarterly) ──
    ws_pq = wb.create_sheet("04_PnL_Quarterly")
    is_q_recs = sorted(section_to_quarters(FIN_DATA, "INCOME_STATEMENT"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-20:]
    pq_headers = ["Chỉ tiêu"] + [f"Q{r.get('lengthReport')}/{r['yearReport']}" for r in is_q_recs]
    ws_pq.column_dimensions['A'].width = 40; [ws_pq.column_dimensions[get_column_letter(j)].__setattr__('width', 14) for j in range(2, 2 + len(is_q_recs))]
    write_header_row(ws_pq, 1, 1, pq_headers)
    
    pq_data = [
        ("Thu nhập lãi thuần - NII", [(r.get("isb27") or 0) / 1e9 for r in is_q_recs]),
        ("Thu nhập ngoài lãi", [((r.get("isb38") or 0) - (r.get("isb27") or 0)) / 1e9 for r in is_q_recs]),
        ("Tổng thu nhập HĐ - TOI", [(r.get("isb38") or 0) / 1e9 for r in is_q_recs]),
        ("Chi phí hoạt động - OPEX", [abs(r.get("isb39") or 0) / 1e9 for r in is_q_recs]),
        ("LN trước dự phòng - PPOP", [(r.get("isb40") or 0) / 1e9 for r in is_q_recs]),
        ("Dự phòng tín dụng", [abs(r.get("isb41") or 0) / 1e9 for r in is_q_recs]),
        ("LN trước thuế - PBT", [(r.get("isa16") or 0) / 1e9 for r in is_q_recs]),
        ("LN sau thuế - NPAT", [(r.get("isa20") or 0) / 1e9 for r in is_q_recs]),
    ]
    for i, (label, vals) in enumerate(pq_data):
        write_data_row(ws_pq, i + 2, 1, [label] + vals, FMT_NUM1)

    # ── Sheet 4.2: P&L (Yearly) ──
    ws = wb.create_sheet("04_PnL")
    ws.column_dimensions['A'].width = 40; [ws.column_dimensions[get_column_letter(j)].__setattr__('width', 16) for j in range(2,9)]
    write_header_row(ws, 1, 1, headers)
    pnl_data = [
        ("Thu nhập lãi thuần - NII", nii_hist + nii_fc),
        ("Thu nhập ngoài lãi", [toi_hist[i] - nii_hist[i] for i in range(len(years_hist))] + non_int_fc),
        ("Tổng thu nhập HĐ - TOI", toi_hist + toi_fc),
        ("Chi phí hoạt động - OPEX", opex_hist + opex_fc),
        ("LN trước dự phòng - PPOP", ppop_hist + ppop_fc),
        ("Dự phòng tín dụng", prov_hist + prov_fc),
        ("LN trước thuế - PBT", pbt_hist + pbt_fc),
        ("Thuế TNDN", [round(pbt_hist[i] * tax_rate, 1) for i in range(len(years_hist))] + tax_fc),
        ("LN sau thuế - NPAT", np_hist + np_fc),
        ("EPS (VND)", eps_hist_calc + eps_fc_calc),
    ]
    for i, (label, vals) in enumerate(pnl_data):
        r = i + 2; is_blue = i < 3 or i == 8
        write_data_row(ws, r, 1, [label] + vals, FMT_NUM1, is_blue=is_blue)
        
    # Overwrite forecast columns G, H, I with formulas in 04_PnL
    for idx, col in enumerate(['G', 'H', 'I']):
        ws.cell(row=2, column=7+idx, value=f"='03_Income_Model'!{col}5")
        ws.cell(row=3, column=7+idx, value=f"='03_Income_Model'!{col}10")
        ws.cell(row=4, column=7+idx, value=f"=SUM({col}2:{col}3)")
        ws.cell(row=5, column=7+idx, value=f"='03_Income_Model'!{col}13")
        ws.cell(row=6, column=7+idx, value=f"={col}4-{col}5")
        ws.cell(row=7, column=7+idx, value=f"='03_Income_Model'!{col}16")
        ws.cell(row=8, column=7+idx, value=f"={col}6-{col}7")
        ws.cell(row=9, column=7+idx, value=f"=MAX({col}8*'02_Assumptions'!{col}14, 0)")
        ws.cell(row=10, column=7+idx, value=f"={col}8-{col}9")
        ws.cell(row=11, column=7+idx, value=f"={col}10*1e9/'02_Assumptions'!$B$3")
        
    # ── Sheet 5.1: Balance Sheet (Quarterly) ──
    ws_bq = wb.create_sheet("05_Balance_Sheet_Quarterly")
    bs_q_recs = sorted(section_to_quarters(FIN_DATA, "BALANCE_SHEET"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-20:]
    nt_q_recs = sorted(section_to_quarters(FIN_DATA, "NOTE"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-20:]
    
    # Matching note quarters to balance sheet quarters
    def get_nt_q_val(y, q, field):
        for r in nt_q_recs:
            if r.get("yearReport") == y and r.get("lengthReport") == q:
                return (r.get(field) or 0) / 1e9
        return 0

    bq_headers = ["Chỉ tiêu"] + [f"Q{r.get('lengthReport')}/{r['yearReport']}" for r in bs_q_recs]
    ws_bq.column_dimensions['A'].width = 40; [ws_bq.column_dimensions[get_column_letter(j)].__setattr__('width', 14) for j in range(2, 2 + len(bs_q_recs))]
    write_header_row(ws_bq, 1, 1, bq_headers)
    bq_data = [
        ("Tổng tài sản", [(r.get("bsa53") or 0) / 1e9 for r in bs_q_recs]),
        ("Tiền mặt & NHNN", [(r.get("bsa2") or 0) / 1e9 for r in bs_q_recs]),
        ("TG các TCTD khác", [(r.get("bsb98") or 0) / 1e9 for r in bs_q_recs]),
        ("Cho vay khách hàng", [(r.get("bsb103") or 0) / 1e9 for r in bs_q_recs]),
        ("CK đầu tư", [(r.get("bsb106") or 0) / 1e9 for r in bs_q_recs]),
        ("Tiền gửi khách hàng", [(r.get("bsb113") or 0) / 1e9 for r in bs_q_recs]),
        ("  CASA", [get_nt_q_val(r['yearReport'], r.get('lengthReport'), "nob66") for r in bs_q_recs]),
        ("Vốn chủ sở hữu", [(r.get("bsa78") or 0) / 1e9 for r in bs_q_recs]),
        ("Nợ nhóm 2", [get_nt_q_val(r['yearReport'], r.get('lengthReport'), "nob41") for r in bs_q_recs]),
        ("Nợ nhóm 3", [get_nt_q_val(r['yearReport'], r.get('lengthReport'), "nob42") for r in bs_q_recs]),
        ("Nợ nhóm 4", [get_nt_q_val(r['yearReport'], r.get('lengthReport'), "nob43") for r in bs_q_recs]),
        ("Nợ nhóm 5", [get_nt_q_val(r['yearReport'], r.get('lengthReport'), "nob44") for r in bs_q_recs]),
    ]
    for i, (label, vals) in enumerate(bq_data):
        write_data_row(ws_bq, i + 2, 1, [label] + vals, FMT_NUM1)

    # ── Sheet 5.2: Balance Sheet (Yearly) ──
    ws = wb.create_sheet("05_Balance_Sheet")
    ws.column_dimensions['A'].width = 40; [ws.column_dimensions[get_column_letter(j)].__setattr__('width', 16) for j in range(2,9)]
    write_header_row(ws, 1, 1, headers)
    bs_data = [
        ("Tổng tài sản", total_assets_hist + [total_assets_hist[-1] * (1.12)**(i+1) for i in range(3)]),
        ("Tiền mặt & NHNN", cash_hist + [cash_hist[-1] * (1.05)**(i+1) for i in range(3)]),
        ("TG các TCTD khác", bank_dep_hist + [bank_dep_hist[-1] * (1.10)**(i+1) for i in range(3)]),
        ("Cho vay khách hàng", loans_hist + loans_fc),
        ("CK đầu tư", inv_sec_bs_hist + [inv_sec_bs_hist[-1] * (1.15)**(i+1) for i in range(3)]),
        ("TS khác", [total_assets_hist[i] - loans_hist[i] - bank_dep_hist[i] - inv_sec_bs_hist[i] - cash_hist[i] - sbv_dep_hist[i] for i in range(len(years_hist))] + [0,0,0]),
        ("", [None]*8),
        ("Tiền gửi khách hàng", cust_dep_hist + dep_fc),
        ("  CASA", casa_hist + casa_fc_amt),
        ("  Có kỳ hạn", [cust_dep_hist[i] - casa_hist[i] for i in range(len(years_hist))] + term_dep_fc),
        ("TG & vay TCTD khác", interbank_hist + [interbank_hist[-1] * (1.08)**(i+1) for i in range(3)]),
        ("GTCT phát hành", bonds_hist + [bonds_hist[-1] * (1.20)**(i+1) for i in range(3)]),
        ("Nợ khác", [total_assets_hist[i] - cust_dep_hist[i] - interbank_hist[i] - bonds_hist[i] - equity_hist[i] for i in range(len(years_hist))] + [0,0,0]),
        ("Vốn chủ sở hữu", equity_hist + [equity_hist[-1] + sum(np_fc[:i+1]) for i in range(3)]),
        ("  Vốn điều lệ", charter_hist + [charter_hist[-1]]*3),
        ("Nợ nhóm 2 (tỷ)", npl_gr2_hist + [0, 0, 0]),
        ("Nợ nhóm 3 (tỷ)", npl_gr3_hist + [0, 0, 0]),
        ("Nợ nhóm 4 (tỷ)", npl_gr4_hist + [0, 0, 0]),
        ("Nợ nhóm 5 (tỷ)", npl_gr5_hist + [0, 0, 0]),
    ]
    for i, (label, vals) in enumerate(bs_data):
        r = i + 2; is_blue = label in ["Tổng tài sản", "Cho vay khách hàng", "Tiền gửi khách hàng", "Vốn chủ sở hữu"]
        write_data_row(ws, r, 1, [label] + vals, FMT_NUM1, is_blue=is_blue)
        
    # Overwrite forecast columns G, H, I with formulas in 05_Balance_Sheet
    for idx, col in enumerate(['G', 'H', 'I']):
        prev_col = 'F' if idx == 0 else get_column_letter(6 + idx)
        ws.cell(row=2, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}2*1.12")
        ws.cell(row=3, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}3*1.05")
        ws.cell(row=4, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}4*1.10")
        ws.cell(row=5, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}5*(1+'02_Assumptions'!{col}4)")
        ws.cell(row=6, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}6*1.15")
        ws.cell(row=7, column=7+idx, value=f"={col}2-SUM({col}3:{col}6)")
        ws.cell(row=8, column=7+idx, value=None)
        ws.cell(row=9, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}9*(1+'02_Assumptions'!{col}5)")
        ws.cell(row=10, column=7+idx, value=f"={col}9*'02_Assumptions'!{col}10")
        ws.cell(row=11, column=7+idx, value=f"={col}9-{col}10")
        ws.cell(row=12, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}12*1.08")
        ws.cell(row=13, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}13*1.20")
        ws.cell(row=15, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}15+'04_PnL'!{col}10")
        ws.cell(row=16, column=7+idx, value=f"='05_Balance_Sheet'!{prev_col}16")
        ws.cell(row=14, column=7+idx, value=f"={col}2-{col}9-{col}12-{col}13-{col}15")
        ws.cell(row=17, column=7+idx, value=0.0)
        ws.cell(row=18, column=7+idx, value=f"=({col}5*'02_Assumptions'!{col}9)*0.2")
        ws.cell(row=19, column=7+idx, value=f"=({col}5*'02_Assumptions'!{col}9)*0.3")
        ws.cell(row=20, column=7+idx, value=f"=({col}5*'02_Assumptions'!{col}9)*0.5")

    # ── Sheet 6: Ratios (FORMULA-BASED for BOTH historical and forecast cells) ──
    ws = wb.create_sheet("06_Ratios")
    ws.column_dimensions['A'].width = 40; [ws.column_dimensions[get_column_letter(j)].__setattr__('width', 16) for j in range(2,9)]
    write_header_row(ws, 1, 1, headers)
    
    # Tạo công thức động cho từng năm
    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'] # 2021 đến 2028
    nim_formulas = []
    npl_formulas = []
    gr2_formulas = []
    casa_formulas = []
    cir_formulas = []
    roe_formulas = []
    roa_formulas = []
    ldr_formulas = []
    coc_formulas = []
    llr_formulas = []
    
    for idx, col in enumerate(cols):
        # NIM: NII / (Cho vay + TG TCTD + CK đầu tư + Tiền mặt)
        if idx < 5: # Lịch sử
            nim_formulas.append(f"=('04_PnL'!{col}2 / ('05_Balance_Sheet'!{col}5 + '05_Balance_Sheet'!{col}4 + '05_Balance_Sheet'!{col}6 + '05_Balance_Sheet'!{col}3)) * 100")
            npl_formulas.append(f"=(SUM('05_Balance_Sheet'!{col}18:'05_Balance_Sheet'!{col}20) / '05_Balance_Sheet'!{col}5) * 100")
            gr2_formulas.append(f"=('05_Balance_Sheet'!{col}17 / '05_Balance_Sheet'!{col}5) * 100")
            casa_formulas.append(f"=('05_Balance_Sheet'!{col}10 / '05_Balance_Sheet'!{col}9) * 100")
            llr_formulas.append(f"=('04_PnL'!{col}7 / MAX(SUM('05_Balance_Sheet'!{col}18:'05_Balance_Sheet'!{col}20), 0.1)) * 100")
        else: # Dự báo
            fc_year_idx = idx - 5
            nim_formulas.append(f"=\'03_Income_Model\'!{col}4*100")
            npl_formulas.append(f"='02_Assumptions'!{col}9")
            gr2_formulas.append("0.0")
            casa_formulas.append(f"='02_Assumptions'!{col}10")
            llr_formulas.append(f"={llr_coverage_fc[fc_year_idx]*100:.1f}")
            
        cir_formulas.append(f"=('04_PnL'!{col}5 / '04_PnL'!{col}4) * 100")
        
        # ROE
        if idx == 0:
            roe_formulas.append(f"=('04_PnL'!{col}10 / '05_Balance_Sheet'!{col}15) * 100")
            coc_formulas.append(f"=('04_PnL'!{col}7 / '05_Balance_Sheet'!{col}5) * 100")
        else:
            prev_col = cols[idx-1]
            roe_formulas.append(f"=('04_PnL'!{col}10 / (('05_Balance_Sheet'!{prev_col}15 + '05_Balance_Sheet'!{col}15) / 2)) * 100")
            coc_formulas.append(f"=('04_PnL'!{col}7 / (('05_Balance_Sheet'!{prev_col}5 + '05_Balance_Sheet'!{col}5) / 2)) * 100")
            
        roa_formulas.append(f"=('04_PnL'!{col}10 / '05_Balance_Sheet'!{col}2) * 100")
        ldr_formulas.append(f"=('05_Balance_Sheet'!{col}5 / '05_Balance_Sheet'!{col}9) * 100")

    ratio_rows = [
        ("NIM (%)", nim_formulas),
        ("NPL (%)", npl_formulas),
        ("Nợ nhóm 2 (%)", gr2_formulas),
        ("CASA (%)", casa_formulas),
        ("CIR (%)", cir_formulas),
        ("ROE (%)", roe_formulas),
        ("ROA (%)", roa_formulas),
        ("LDR (%)", ldr_formulas),
        ("CoC - Credit Cost (%)", coc_formulas),
        ("LLR / NPL (%)", llr_formulas),
        ("YOEA (%)", [7.76, 7.83, 7.76, 7.8, 7.9] + [7.8, 7.9, 8.0]), # Lấy mẫu 5 năm lịch sử
        ("COF (%)", [5.19, 5.21, 4.69, 4.8, 4.7] + [5.0, 4.8, 4.6]),
        ("IEA bình quân (tỷ)", [0]*5 + [round(x, 1) for x in iea_fc]),
        ("P/B (x)", [pb_median_year.get(y, 0) for y in years_hist] + ["='02_Assumptions'!G15", "='02_Assumptions'!H15", "='02_Assumptions'!I15"]),
        ("P/E (x)", [pe_median_year.get(y, 0) for y in years_hist] + [
            "=('02_Assumptions'!$B$2*'02_Assumptions'!$B$3)/('04_PnL'!G10*1e9)",
            "=('02_Assumptions'!$B$2*'02_Assumptions'!$B$3)/('04_PnL'!H10*1e9)",
            "=('02_Assumptions'!$B$2*'02_Assumptions'!$B$3)/('04_PnL'!I10*1e9)"
        ]),
    ]
    
    for i, (label, vals) in enumerate(ratio_rows):
        r = i + 2
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(size=11, name="Calibri", bold=True); c.border = thin_border
        for j, v in enumerate(vals):
            cell = ws.cell(row=r, column=2+j)
            cell.font = Font(size=11, name="Calibri"); cell.number_format = FMT_NUM1
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.value = v
        # ── Sheet 7: Valuation (FORMULA-BASED) ──
    ws = wb.create_sheet("07_Valuation")
    ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 20
    ws.cell(row=1, column=1, value="PHUONG PHAP DINH GIA").font = FMT_BOLD
    ws.cell(row=2, column=1, value="Chi phi von CSH (COE)").font = Font(size=11, name="Calibri")
    ws.cell(row=2, column=2, value=f"='02_Assumptions'!G12").number_format = FMT_PCT
    ws.cell(row=3, column=1, value="Tang truong dai han (g)")
    ws.cell(row=3, column=2, value=f"={terminal_growth}").number_format = FMT_PCT
    ws.cell(row=5, column=1, value="RESIDUAL INCOME MODEL").font = FMT_BOLD
    ws.cell(row=6, column=1, value="BV/share hien tai (VND)")
    ws.cell(row=6, column=2, value=f"='05_Balance_Sheet'!F15*1e9/'02_Assumptions'!B3").number_format = FMT_NUM
    ws.cell(row=7, column=1, value="PV cua RI 3 nam (VND)")
    ws.cell(row=7, column=2, value="=SUM(B31:D31)").number_format = FMT_NUM
    ws.cell(row=8, column=1, value="PV cua Continuing Value (VND)")
    ws.cell(row=8, column=2, value="=(D29*(1+B3)/(B2-B3))*D30").number_format = FMT_NUM
    ws.cell(row=9, column=1, value="GIA TRI RI (VND)"); ws.cell(row=9, column=2, value=f"=B6+B7+B8").number_format = FMT_NUM
    ws.cell(row=9, column=2).font = FMT_BOLD
    
    ws.cell(row=11, column=1, value="P/B MULTIPLE — 3 muc phan phoi lich su").font = FMT_BOLD
    ws.cell(row=12, column=1, value="P/B hien tai (x)")
    ws.cell(row=12, column=2, value="='02_Assumptions'!B2/B6").number_format = '0.00'
    ws.cell(row=13, column=1, value="P/B hap dan (x) — MUA")  # lower half median
    ws.cell(row=13, column=2, value=f"={pb_attractive:.2f}").number_format = '0.00'
    ws.cell(row=13, column=2).font = Font(color="006400", bold=True, name="Calibri")
    ws.cell(row=13, column=3, value="Median P/B nua duoi (vung mua hap dan)").font = Font(size=9, color="006400", name="Calibri")
    ws.cell(row=14, column=1, value="P/B median all-time (x) — FAIR VALUE")
    ws.cell(row=14, column=2, value=f"={pb_all_median:.2f}").number_format = '0.00'
    ws.cell(row=14, column=3, value="Median toan bo lich su (diem can bang)").font = Font(size=9, color="808080", name="Calibri")
    ws.cell(row=15, column=1, value="P/B muc tieu (x) — BAN / CHOT LOI")  # upper half median
    ws.cell(row=15, column=2, value=f"='02_Assumptions'!G15").number_format = '0.00'
    ws.cell(row=15, column=2).font = Font(color="C00000", bold=True, name="Calibri")
    ws.cell(row=15, column=3, value="Median P/B nua tren (vung muc tieu ban)").font = Font(size=9, color="C00000", name="Calibri")
    ws.cell(row=16, column=1, value="BV/share tuong lai (2026F)")
    ws.cell(row=16, column=2, value="=B6+B26").number_format = FMT_NUM
    ws.cell(row=17, column=1, value="GIA TRI P/B (VND)")
    ws.cell(row=17, column=2, value="=B15*B16").number_format = FMT_NUM   # dung pb_target (row 15)
    ws.cell(row=17, column=2).font = FMT_BOLD
    
    ws.cell(row=19, column=1, value="WEIGHTED TARGET PRICE").font = Font(bold=True, size=13, name="Calibri")
    ws.cell(row=20, column=1, value="Trong so: 50% RI + 50% P/B")
    ws.cell(row=21, column=1, value="Target Price (VND)")
    ws.cell(row=21, column=2, value=f"=B9*0.5+B17*0.5").number_format = FMT_NUM   # B17 = P/B value
    ws.cell(row=21, column=2).font = FMT_BOLD
    ws.cell(row=22, column=1, value="Gia hien tai (VND)"); ws.cell(row=22, column=2, value=f"='02_Assumptions'!B2")
    ws.cell(row=23, column=1, value="UPSIDE (%)")
    ws.cell(row=23, column=2, value=f"=B21/B22-1").number_format = '0.00%'
    ws.cell(row=23, column=2).font = Font(bold=True, size=12, color="006400" if upside > 0 else "C00000", name="Calibri")

    # Working row for Residual Income per Share details
    ws.cell(row=25, column=1, value="--- RI DETAIL ---").font = Font(bold=True, color="2E75B6", name="Calibri")
    ws.cell(row=26, column=1, value="EPS (VND)")
    ws.cell(row=27, column=1, value="BVPS dau ky (VND)")
    ws.cell(row=28, column=1, value="Capital Charge (VND)")
    ws.cell(row=29, column=1, value="Residual Income (VND)")
    ws.cell(row=30, column=1, value="Discount Factor")
    ws.cell(row=31, column=1, value="PV of RI")

    for idx, (yr, col) in enumerate(zip([2026, 2027, 2028], ['G', 'H', 'I'])):
        col_letter = get_column_letter(2 + idx)
        ws.cell(row=25, column=2+idx, value=f"{yr}F").font = FMT_BOLD
        ws.cell(row=26, column=2+idx, value=f"='04_PnL'!{col}10*1e9/'02_Assumptions'!$B$3").number_format = FMT_NUM
        prev_col = 'F' if idx == 0 else get_column_letter(6 + idx)
        ws.cell(row=27, column=2+idx, value=f"='05_Balance_Sheet'!{prev_col}15*1e9/'02_Assumptions'!$B$3").number_format = FMT_NUM
        ws.cell(row=28, column=2+idx, value=f"={col_letter}27*'07_Valuation'!$B$2").number_format = FMT_NUM
        ws.cell(row=29, column=2+idx, value=f"={col_letter}26-{col_letter}28").number_format = FMT_NUM
        ws.cell(row=30, column=2+idx, value=f"=1/(1+'07_Valuation'!$B$2)^{idx+1}").number_format = '0.0000'
        ws.cell(row=31, column=2+idx, value=f"={col_letter}29*{col_letter}30").number_format = FMT_NUM
    
    # ── Sheet 8: Sensitivity ──
    ws = wb.create_sheet("08_Sensitivity")
    ws.column_dimensions['A'].width = 20
    sensitivities = [0.10, 0.115, 0.13, 0.145, 0.16]  # COE scenarios
    term_gs = [0.01, 0.02, 0.03, 0.04, 0.05]  # g scenarios
    ws.cell(row=1, column=1, value="Sensitivity: COE x g").font = FMT_BOLD
    ws.cell(row=2, column=1, value="COE \\ g")
    
    # Write numeric headers for g
    for j, g in enumerate(term_gs):
        cell = ws.cell(row=2, column=j+2, value=g)
        cell.font = FMT_BOLD; cell.number_format = '0%'
        
    # Write numeric headers for COE and formula cells
    for i, coe in enumerate(sensitivities):
        r = i + 3
        cell = ws.cell(row=r, column=1, value=coe)
        cell.font = FMT_BOLD; cell.number_format = '0.0%'
        for j, g in enumerate(term_gs):
            col_letter = get_column_letter(j+2)
            # Standard Residual Income formula using cell references
            formula = f"='07_Valuation'!$B$6 + (('07_Valuation'!$B$26 - '07_Valuation'!$B$27*$A{r})/(1+$A{r})^1 + ('07_Valuation'!$C$26 - '07_Valuation'!$C$27*$A{r})/(1+$A{r})^2 + ('07_Valuation'!$D$26 - '07_Valuation'!$D$27*$A{r})/(1+$A{r})^3) + (('07_Valuation'!$D$26 - '07_Valuation'!$D$27*$A{r})*(1+{col_letter}$2)/($A{r}-{col_letter}$2))/(1+$A{r})^3"
            c = ws.cell(row=r, column=j+2, value=formula)
            c.number_format = FMT_NUM
            cv_ij = ri_results[-1] * (1 + g) / (coe - g) if ri_results else 0
            pv_ij = sum([ri_results[k] / (1 + coe) ** (k + 1) for k in range(len(ri_results))])
            pv_cv_ij = cv_ij / (1 + coe) ** len(ri_results)
            val = bvps_base + pv_ij + pv_cv_ij
            if val < PRICE: c.font = Font(color="C00000")
            elif val > PRICE * 1.2: c.font = Font(color="006400")
    
    # ── Sheet 9: PESTLE ──
    ws = wb.create_sheet("09_PESTLE")
    ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 60; ws.column_dimensions['C'].width = 15
    write_header_row(ws, 1, 1, ["Yếu tố", "Nội dung", "Tác động"])
    pestle_data = [
        ("Chính trị - PL", "SBV giữ LS điều hành 4.5%. Môi trường chính trị ổn định. Thông tư 22 siết room tín dụng.", "Tích cực"),
        ("Kinh tế", "GDP 2026E ~6.5%. LS huy động nhích nhẹ (5-5.5%). Tín dụng toàn ngành 14-15%.", "Tích cực"),
        ("Xã hội", "Thu nhập tăng, ngân hàng số phát triển. Thanh toán không dùng TM mở rộng.", "Trung tính"),
        ("Công nghệ", "Chuyển đổi số ngành NH. Ngân hàng nhỏ chậm hơn trong đầu tư CN.", "Trung tính"),
        ("Môi trường", "Tín dụng xanh được khuyến khích. Ảnh hưởng không lớn với ngân hàng TMCP.", "Trung tính"),
        ("Pháp lý khác", "Thông tư 06/2025 sửa đổi về phân loại nợ. NHNN siết tỷ lệ vốn ngắn hạn cho vay dài hạn.", "Tiêu cực"),
    ]
    for i, (factor, content, impact) in enumerate(pestle_data):
        r = i + 2
        ws.cell(row=r, column=1, value=factor).border = thin_border
        ws.cell(row=r, column=2, value=content).border = thin_border
        ws.cell(row=r, column=3, value=impact).border = thin_border
        if impact == "Tích cực": ws.cell(row=r, column=3).font = Font(color="006400")
        elif impact == "Tiêu cực": ws.cell(row=r, column=3).font = Font(color="C00000")
    
    # ── Sheet 10: Leading Indicators ──
    ws = wb.create_sheet("10_Leading_Indicators")
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15; ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 15
    write_header_row(ws, 1, 1, ["Chỉ báo", "Ngưỡng tích cực", "Ngưỡng tiêu cực", "Giá trị hiện tại", "Trạng thái"])
    indicators = [
        ("CASA ratio (%)", 10, 3, casa_ratio_hist[-1], "Cần cải thiện"),
        ("NIM (%)", 3.5, 2.5, nim_hist[-1], "Thấp"),
        ("NPL (%)", 1.0, 2.0, npl_ratio_hist[-1], "Khá"),
        ("ROE (%)", 18, 12, roe_hist[-1], "Trung bình"),
        ("P/B (x)", 1.5, 0.8, pb_current, "Rẻ"),
        ("CIR (%)", 30, 45, cir_hist[-1], "Tốt"),
        ("LDR (%)", 85, 95, ldr_hist[-1], "Ổn định"),
        ("CAR (%)", 12, 9, 9.4, "Cần chú ý"),
        ("Room tín dụng NHNN (%)", 15, 10, 14, "Tích cực"),
        ("KLGD BQ 20 phiên (CP)", 500000, 100000, 200000, "Thanh khoản thấp"),
    ]
    for i, (name, pos, neg, val, status) in enumerate(indicators):
        r = i + 2
        ws.cell(row=r, column=1, value=name).border = thin_border
        ws.cell(row=r, column=2, value=pos).border = thin_border
        ws.cell(row=r, column=3, value=neg).border = thin_border
        ws.cell(row=r, column=4, value=val).border = thin_border
        ws.cell(row=r, column=5, value=status).border = thin_border
        if "Tốt" in status or "Tích cực" in status: ws.cell(row=r, column=5).font = Font(color="006400")
        elif "Cần" in status or "chú ý" in status: ws.cell(row=r, column=5).font = Font(color="FF8C00")
        elif "Thấp" in status: ws.cell(row=r, column=5).font = Font(color="C00000")
    
    # ── Sheet 11: Investment Thesis ──
    ws = wb.create_sheet("11_Investment_Thesis")
    ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 60; ws.column_dimensions['C'].width = 15
    write_header_row(ws, 1, 1, ["Tầng đánh giá", "Kết luận", "Rating"])
    thesis = [
        ("Chất lượng TS", f"NPL {npl_ratio_hist[-1]}% (tốt). Nợ nhóm 2 ~0%. LLR {prov_hist[-1]/max(npl_total_hist[-1],1)*100:.0f}% (dưới 100%). CAR ~9.4% thấp.", "Trung bình"),
        ("Tăng trưởng", f"Tín dụng CAGR 13%. LNST CAGR 33%. Tăng vốn charter 51% năm 2025. Quy mô nhỏ.", "Tốt"),
        ("Lợi nhuận", f"PPOP CAGR 32% - tăng trưởng thực. NIM {nim_hist[-1]}% thấp. CASA {casa_ratio_hist[-1]}% rất thấp.", "Trung bình"),
        ("Hiệu quả", f"ROE {roe_hist[-1]}% (dưới 15%). CIR {cir_hist[-1]}% (tốt nhưng cần verify nguồn gốc).", "Trung bình"),
        ("Định giá", f"P/B {pb_current:.2f}x (dưới 1x - rẻ). P/E {MARKET_CAP/(np_hist[-1]*1e9):.1f}x. Target {weighted_target:,.0f} VND (+{upside:.0f}%).", "Tốt"),
        ("Vĩ mô & Ngành", "LS huy động nhích tăng ảnh hưởng COF. NHNN room 14% cho bank tốt. CASA thấp gây áp lực NIM.", "Trung bình"),
    ]
    for i, (tier, conclusion, rating) in enumerate(thesis):
        r = i + 2
        ws.cell(row=r, column=1, value=tier).border = thin_border; ws.cell(row=r, column=1).font = FMT_BOLD
        ws.cell(row=r, column=2, value=conclusion).border = thin_border; ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=rating).border = thin_border
        if rating == "Tốt": ws.cell(row=r, column=3).font = Font(color="006400", bold=True)
        elif rating == "Trung bình": ws.cell(row=r, column=3).font = Font(color="FF8C00", bold=True)
        else: ws.cell(row=r, column=3).font = Font(color="C00000", bold=True)
        ws.row_dimensions[r].height = 50
    
    # ── Sheet 12: Summary ──
    ws = wb.create_sheet("12_Summary_Snapshot")
    ws.column_dimensions['A'].width = 30; [ws.column_dimensions[get_column_letter(j)].__setattr__('width', 16) for j in range(2,9)]
    write_header_row(ws, 1, 1, ["Chỉ tiêu"] + [str(y) for y in all_years])
    sum_items = [
        ("Tổng tài sản (tỷ)", total_assets_hist + [round(x, 1) for x in [total_assets_hist[-1]*(1.12), total_assets_hist[-1]*(1.12)**2, total_assets_hist[-1]*(1.12)**3]]),
        ("Cho vay KH (tỷ)", loans_hist + [round(x, 1) for x in loans_fc]),
        ("TG khách hàng (tỷ)", cust_dep_hist + [round(x, 1) for x in dep_fc]),
        ("VCSH (tỷ)", equity_hist + [round(equity_hist[-1] + sum(np_fc[:i+1]), 1) for i in range(3)]),
        ("NII (tỷ)", nii_hist + [round(x, 1) for x in nii_fc]),
        ("PPOP (tỷ)", ppop_hist + [round(x, 1) for x in ppop_fc]),
        ("LNST (tỷ)", np_hist + [round(x, 1) for x in np_fc]),
        ("EPS (VND)", [round(x) for x in eps_hist_calc + eps_fc_calc]),
        ("BVPS (VND)", [round(x) for x in bvps_hist + [round(bvps_base + sum(np_fc[:i+1])*1e9/SHARES, 0) for i in range(3)]]),
        ("", [None]*8),
        ("NIM (%)", [n/100 for n in nim_hist] + [n/100 for n in nim_fc]),
        ("NPL (%)", npl_ratio_hist + [round(n * 100, 2) for n in npl_fc]),
        ("CASA (%)", casa_ratio_hist + [round(c * 100, 2) for c in casa_target_fc]),
        ("CIR (%)", cir_hist + [round(c * 100, 2) for c in cir_fc]),
        ("ROE (%)", roe_hist + [round(np_fc[0]/equity_hist[-1]*100,2)] + [round(np_fc[i]/(equity_hist[-1]+sum(np_fc[:i]))*100,2) for i in range(1,3)]),
        ("P/B (x)", [pb_current] + [pb_target]*3),
        ("P/E (x)", [round(MARKET_CAP/(np_hist[-1]*1e9),2)] + [round(MARKET_CAP/(np_fc[i]*1e9),2) for i in range(3)]),
    ]
    # Add CAGR row
    for i, (label, vals) in enumerate(sum_items):
        r = i + 2
        write_data_row(ws, r, 1, [label] + vals, FMT_NUM1, is_blue=True)
    
    # Add CAGR
    cagr_row = len(sum_items) + 3
    ws.cell(row=cagr_row, column=1, value="CAGR 2023-2025 (%)").font = FMT_BOLD
    cagr_items = [
        ("Tổng tài sản", total_assets_hist),
        ("Cho vay KH", loans_hist),
        ("TG khách hàng", cust_dep_hist),
        ("NII", nii_hist),
        ("PPOP", ppop_hist),
        ("LNST", np_hist),
    ]
    for i, (label, vals) in enumerate(cagr_items):
        r = cagr_row + 1 + i
        cagr_val = (vals[-1] / vals[0]) ** (1/2) - 1 if vals[0] else 0
        ws.cell(row=r, column=1, value=label).border = thin_border
        ws.cell(row=r, column=2, value=cagr_val).number_format = '0.00%'; ws.cell(row=r, column=2).border = thin_border
    
    # ── Sheet 13: PE/PB History (historical data only, NO forecast) ──
    ws = wb.create_sheet("13_PE_PB_History")
    ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15; ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15; ws.column_dimensions['F'].width = 15
    pe_periods = sorted([r for r in ttms if r.get("pe")], key=lambda x: (x["year"], x["quarter"]))
    ws.cell(row=1, column=1, value="PE/PB/EVEBITDA HISTORY").font = FMT_BOLD
    ws.cell(row=2, column=1, value="Ky"); ws.cell(row=2, column=2, value="P/E (x)")
    ws.cell(row=2, column=3, value="P/B (x)"); ws.cell(row=2, column=4, value="EV/EBITDA (x)")
    ws.cell(row=2, column=5, value="Ghi chu"); ws.cell(row=2, column=6, value="")
    for hc in range(1, 7):
        ws.cell(row=2, column=hc).font = Font(bold=True, size=11, name="Calibri")
        ws.cell(row=2, column=hc).border = thin_border
        ws.cell(row=2, column=hc).fill = FMT_HDR
        ws.cell(row=2, column=hc).alignment = Alignment(horizontal='center')
    for i, rp in enumerate(pe_periods):
        row = i + 3
        yq = (rp["year"], rp["quarter"])
        label = f"Q{rp['quarter']}/{rp['year']}"
        pe_val_raw = rp.get("pe")
        pe_val = pe_val_raw if pe_val_raw else 0
        pb_val = rp.get("pb") or 0
        eve_val = rp.get("evToEbitda")
        is_carried = yq in pe_carried
        ws.cell(row=row, column=1, value=label); ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=pe_val); ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=2).border = thin_border
        if is_carried:
            ws.cell(row=row, column=2).font = Font(color="FF0000", name="Calibri", size=11)
        ws.cell(row=row, column=3, value=pb_val); ws.cell(row=row, column=3).number_format = '0.00'
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=4, value=eve_val); ws.cell(row=row, column=4).number_format = '0.0' if eve_val else '@'
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5, value="P/E dieu chinh (LNST am)" if is_carried else ""); ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6, value=None); ws.cell(row=row, column=6).border = thin_border
    # MEDIAN rows at bottom
    median_start = len(pe_periods) + 4
    ws.cell(row=median_start, column=1, value="MEDIAN ALL-TIME").font = Font(bold=True, size=11, name="Calibri", color="2F5496")
    ws.cell(row=median_start, column=2, value=round(pe_all_median, 1))
    ws.cell(row=median_start, column=2).number_format = '0.0'
    ws.cell(row=median_start, column=2).font = Font(bold=True, size=11, name="Calibri")
    ws.cell(row=median_start, column=3, value=round(pb_all_median, 2))
    ws.cell(row=median_start, column=3).number_format = '0.00'
    ws.cell(row=median_start, column=3).font = Font(bold=True, size=11, name="Calibri")
    ws.cell(row=median_start, column=4, value=round(evebitda_all_median, 1) if evebitda_all_median else None)
    ws.cell(row=median_start, column=4).font = Font(bold=True, size=11, name="Calibri")
    for mc in range(1, 5):
        ws.cell(row=median_start, column=mc).border = thin_border
    # ── CHARTS ──
    # Chart 1: Revenue & NPAT
    chart1 = BarChart(); chart1.type = "col"; chart1.grouping = "clustered"
    chart1.title = "NII & LNST (tỷ đồng)"; chart1.y_axis.title = "tỷ đồng"
    years_str = [str(y) for y in all_years]
    data_ref = openpyxl.Workbook()
    data_ws = data_ref.active
    data_ws.append(["Năm"] + years_str)
    data_ws.append(["NII"] + nii_hist + [round(x, 1) for x in nii_fc])
    data_ws.append(["LNST"] + np_hist + [round(x, 1) for x in np_fc])
    data_ws.append(["NIM"] + nim_hist + nim_fc)
    cats = Reference(data_ws, min_col=2, max_col=len(years_str)+1, min_row=1)
    d1 = Reference(data_ws, min_col=1, max_col=1, min_row=2); d1.name = "NII"; chart1.add_data(d1, from_rows=True)
    d2 = Reference(data_ws, min_col=1, max_col=1, min_row=3); d2.name = "LNST"; chart1.add_data(d2, from_rows=True)
    if hasattr(chart1, 'set_categories'): chart1.set_categories(cats)
    chart1.width = 20; chart1.height = 12
    ws2 = wb.create_sheet("Charts")
    ws2.add_chart(chart1, "A1")
    
    # ── Sheet 6.1: Ratios Quarterly (FORMULA-BASED) ──
    ws_rq = wb.create_sheet("06_Ratios_Quarterly")
    rq_headers = ["Chỉ tiêu"] + [f"Q{r.get('lengthReport')}/{r.get('yearReport')}" for r in bs_q_recs]
    ws_rq.column_dimensions['A'].width = 40
    [ws_rq.column_dimensions[get_column_letter(j)].__setattr__('width', 14) for j in range(2, 2 + len(bs_q_recs))]
    write_header_row(ws_rq, 1, 1, rq_headers)
    
    cols_q = [get_column_letter(j) for j in range(2, 2 + len(bs_q_recs))]
    rq_nim = []
    rq_roe = []
    rq_ldr = []
    rq_casa = []
    rq_cir = []
    rq_loan_grow = []
    rq_dep_grow = []
    
    for idx, col in enumerate(cols_q):
        # NIM quarterly (annualized: * 400)
        rq_nim.append(f"=('04_PnL_Quarterly'!{col}2 / ('05_Balance_Sheet_Quarterly'!{col}5 + '05_Balance_Sheet_Quarterly'!{col}4 + '05_Balance_Sheet_Quarterly'!{col}6 + '05_Balance_Sheet_Quarterly'!{col}3)) * 400")
        # ROE quarterly (annualized: * 400)
        rq_roe.append(f"=('04_PnL_Quarterly'!{col}9 / '05_Balance_Sheet_Quarterly'!{col}9) * 400")
        # LDR
        rq_ldr.append(f"=('05_Balance_Sheet_Quarterly'!{col}5 / '05_Balance_Sheet_Quarterly'!{col}7) * 100")
        # CASA
        rq_casa.append(f"=('05_Balance_Sheet_Quarterly'!{col}8 / '05_Balance_Sheet_Quarterly'!{col}7) * 100")
        # CIR
        rq_cir.append(f"=('04_PnL_Quarterly'!{col}5 / '04_PnL_Quarterly'!{col}4) * 100")
        # Loan growth QoQ
        if idx == 0:
            rq_loan_grow.append("0.0")
            rq_dep_grow.append("0.0")
        else:
            prev_col = cols_q[idx-1]
            rq_loan_grow.append(f"=('05_Balance_Sheet_Quarterly'!{col}5 / '05_Balance_Sheet_Quarterly'!{prev_col}5 - 1) * 100")
            rq_dep_grow.append(f"=('05_Balance_Sheet_Quarterly'!{col}7 / '05_Balance_Sheet_Quarterly'!{prev_col}7 - 1) * 100")

    rq_rows = [
        ("NIM (%)", rq_nim),
        ("ROE (%)", rq_roe),
        ("LDR (%)", rq_ldr),
        ("CASA (%)", rq_casa),
        ("CIR (%)", rq_cir),
        ("Tăng trưởng tín dụng QoQ (%)", rq_loan_grow),
        ("Tăng trưởng huy động QoQ (%)", rq_dep_grow),
    ]
    for i, (label, vals) in enumerate(rq_rows):
        r = i + 2
        c = ws_rq.cell(row=r, column=1, value=label)
        c.font = Font(size=11, name="Calibri", bold=True); c.border = thin_border
        for j, v in enumerate(vals):
            cell = ws_rq.cell(row=r, column=2+j)
            cell.font = Font(size=11, name="Calibri"); cell.number_format = FMT_NUM1
            cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            cell.value = v

    wb.save(EXCEL_FILE)
    print(f"[Excel] Saved to {EXCEL_FILE}")
    return wb

# ═══════════════════════════════════════════════════════════════════════════════
# PDF EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def register_vn_fonts():
    font_dir = "C:/Windows/Fonts"
    fonts_to_try = [("arial.ttf", "Arial"), ("arialbd.ttf", "Arial-Bold"),
                    ("ariali.ttf", "Arial-Italic"), ("arialbi.ttf", "Arial-BoldItalic"),
                    ("times.ttf", "TimesNewRoman"), ("timesbd.ttf", "TimesNewRoman-Bold")]
    found = {}
    for fname, freg in fonts_to_try:
        path = os.path.join(font_dir, fname)
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(freg, path))
                found[freg] = path
            except:
                pass
    return found

def calc_yoea_cof():
    """Calculate YOEA and COF from available data."""
    iea_hist = []
    for i in range(len(years_hist)):
        iea = loans_hist[i] + bank_dep_hist[i] + inv_sec_bs_hist[i] + cash_hist[i] + sbv_dep_hist[i]
        iea_hist.append(iea)
    funding_hist = [cust_dep_hist[i] + interbank_hist[i] + bonds_hist[i] for i in range(len(years_hist))]
    yoea = [round(int_inc_hist[i] / max(iea_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    cof = [round(int_exp_hist[i] / max(funding_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    iea_fc_calc = [(loans_hist[-1] + bank_dep_hist[-1] + inv_sec_bs_hist[-1] + cash_hist[-1] + sbv_dep_hist[-1]) * (1 + iea_growth_fc[i]) for i in range(3)]
    funding_fc = [dep_fc[i] + interbank_hist[-1] * (1.08)**(i+1) + bonds_hist[-1] * (1.20)**(i+1) for i in range(3)]
    yoea_fc = [round(yoea[-1] + 0.1*i, 2) for i in range(3)]
    cof_fc = [round(cof[-1] - 0.1*i, 2) for i in range(3)]
    return yoea + yoea_fc, cof + cof_fc

def create_charts():
    """Generate 12 chart images for PDF (Types A-L from MBS format)."""
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial']
    
    years = [str(y) for y in (years_hist + years_fc)]
    x = range(len(years))
    hist3 = [str(y) for y in years_hist]
    x3 = range(len(years_hist))
    # x3 = range(3)
    yoea_cof, cof_vals = calc_yoea_cof()
    nim_all = nim_hist + [round(x,2) for x in nim_fc]
    
    # ── CHART A (1): NIM Decomposition (YOEA, COF, NIM) ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(x, yoea_cof, 'o-', color='#4472C4', linewidth=2, markersize=5, label='YOEA (%)')
    ax.plot(x, cof_vals, 's-', color='#ED7D31', linewidth=2, markersize=5, label='COF (%)')
    ax.plot(x, nim_all, 'D-', color='#70AD47', linewidth=3, markersize=6, label='NIM (%)')
    for i, v in enumerate(yoea_cof):
        ax.text(i, v+0.1, f'{v:.2f}%', ha='center', fontsize=7, color='#4472C4')
    for i, v in enumerate(cof_vals):
        ax.text(i, v+0.1, f'{v:.2f}%', ha='center', fontsize=7, color='#ED7D31')
    for i, v in enumerate(nim_all):
        ax.text(i, v+0.1, f'{v:.2f}%', ha='center', fontsize=7, color='#70AD47', fontweight='bold')
    ax.axvline(x=2.5, color='gray', linestyle='--', alpha=0.5)
    ax.set_xticks(x); ax.set_xticklabels(years, fontsize=8)
    ax.set_ylabel('%', fontsize=9)
    ax.legend(fontsize=9, loc='upper left')
    plt.title(f'{TICKER}: NIM Decomposition — YOEA, COF & NIM', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartA_nim_decomp.png'), dpi=150)
    plt.close()
    
    # ── CHART B (2): Peer NPL Comparison ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    npl_peers = peer_val("NPL")
    sorted_banks = sorted(npl_peers.items(), key=lambda kv: kv[1])
    bank_labels = [b[0] for b in sorted_banks]
    bank_npl = [b[1] for b in sorted_banks]
    colors_b = ['#C00000' if b == TICKER else '#4472C4' for b in bank_labels]
    bars_b = ax.barh(range(len(bank_labels)), bank_npl, color=colors_b, height=0.6)
    for i, (bar, v) in enumerate(zip(bars_b, bank_npl)):
        ax.text(bar.get_width() + 0.02, bar.get_y() + bar.get_height()/2, f'{v:.2f}%', va='center', fontsize=7, color='#333')
    ax.set_yticks(range(len(bank_labels))); ax.set_yticklabels(bank_labels, fontsize=8)
    ax.set_xlabel('NPL (%)', fontsize=9)
    ax.invert_yaxis()
    plt.title(f'Peer NPL Comparison (%)', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartB_peer_npl.png'), dpi=150)
    plt.close()
    
    # ── CHART C (3): Peer Credit Growth ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    cred_peers = peer_val("CREDIT_GROWTH")
    sorted_cred = sorted(cred_peers.items(), key=lambda kv: kv[1])
    cred_labels = [c[0] for c in sorted_cred]
    cred_vals = [c[1] for c in sorted_cred]
    colors_c = ['#70AD47' if c == TICKER else '#4472C4' for c in cred_labels]
    bars_c = ax.barh(range(len(cred_labels)), cred_vals, color=colors_c, height=0.6)
    for i, (bar, v) in enumerate(zip(bars_c, cred_vals)):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height()/2, f'{v:.1f}%', va='center', fontsize=7, color='#333')
    ax.set_yticks(range(len(cred_labels))); ax.set_yticklabels(cred_labels, fontsize=8)
    ax.set_xlabel('Credit Growth YTD (%)', fontsize=9)
    ax.invert_yaxis()
    plt.title(f'Peer Credit Growth (% YTD)', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartC_peer_credit.png'), dpi=150)
    plt.close()
    
    # ── CHART D (4): Bank vs Industry Average ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    metrics_d = ['NIM', 'ROE', 'CIR', 'CASA', 'NPL']
    bank_vals_d = [PEER_DATA[m].get(TICKER, 0) for m in metrics_d]
    ind_vals_d = [INDUSTRY_AVG[m] for m in metrics_d]
    x_d = range(len(metrics_d))
    w = 0.35
    bars_d1 = ax.bar([i - w/2 for i in x_d], bank_vals_d, width=w, color='#2F5496', label=TICKER)
    bars_d2 = ax.bar([i + w/2 for i in x_d], ind_vals_d, width=w, color='#A5A5A5', label='Ngành (TB)')
    for bar in bars_d1:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{bar.get_height():.1f}', ha='center', fontsize=7, color='#2F5496', fontweight='bold')
    ax.set_xticks(x_d); ax.set_xticklabels(metrics_d, fontsize=9)
    ax.set_ylabel('%', fontsize=9)
    ax.legend(fontsize=9)
    plt.title(f'{TICKER} vs Industry Average', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartD_vs_industry.png'), dpi=150)
    plt.close()
    
    # ── CHART E (5): Earning Assets Structure (STACKED AREA CHART / BIỂU ĐỒ MIỀN CHỒNG 100%) ──
    ea_components = ['Cash & SBV', 'Bank Dep', 'Loans', 'Inv Securities']
    fig, ax = plt.subplots(figsize=(8, 4.5))
    
    pct_cash_sbv, pct_bank_dep, pct_loans, pct_inv_sec = [], [], [], []
    for i in range(len(years_hist)):
        ea_vals = [cash_hist[i] + sbv_dep_hist[i], bank_dep_hist[i], loans_hist[i], inv_sec_bs_hist[i]]
        total_ea = sum(ea_vals)
        pct_cash_sbv.append((cash_hist[i] + sbv_dep_hist[i]) / max(total_ea, 1) * 100)
        pct_bank_dep.append(bank_dep_hist[i] / max(total_ea, 1) * 100)
        pct_loans.append(loans_hist[i] / max(total_ea, 1) * 100)
        pct_inv_sec.append(inv_sec_bs_hist[i] / max(total_ea, 1) * 100)
        
    y = np.row_stack((pct_cash_sbv, pct_bank_dep, pct_loans, pct_inv_sec))
    colors_e = ['#FFC000', '#A5A5A5', '#4472C4', '#70AD47']
    
    # Plot Stacked Area
    polys = ax.stackplot(years_hist, y, labels=ea_components, colors=colors_e, alpha=0.85)
    
    # Add values text on vertical year lines
    for i, yr in enumerate(years_hist):
        cum_y = 0
        for val, color in zip([pct_cash_sbv[i], pct_bank_dep[i], pct_loans[i], pct_inv_sec[i]], colors_e):
            if val > 5: # Only label if height is significant
                ax.text(yr, cum_y + val/2, f"{val:.1f}%", ha='center', va='center', fontsize=8, fontweight='bold')
            cum_y += val
            
    ax.set_title(f'{TICKER}: Earning Assets Structure (2021-2025)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Cơ cấu (%)', fontsize=9)
    ax.set_xticks(years_hist)
    ax.set_xticklabels([str(y) for y in years_hist], fontsize=8)
    ax.set_ylim(0, 100)
    ax.grid(axis='y', linestyle='--', alpha=0.5)
    ax.legend(loc='lower center', bbox_to_anchor=(0.5, -0.2), ncol=4, fontsize=8)
    plt.tight_layout(rect=[0, 0.08, 1, 0.95])
    plt.savefig(os.path.join(CHART_DIR, 'chartE_earning_assets.png'), dpi=150)
    plt.close()
    
    # ── CHART F (6): Non-Interest Income Breakdown ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    nonii_components = {
        'Fee': fee_inc_hist,
        'FX': fx_hist,
        'Trading Sec': trade_sec_hist,
        'Inv Sec': inv_sec_hist,
        'Other': other_inc_hist,
    }
    bottom = [0] * len(years_hist)
    colors_f = ['#4472C4', '#ED7D31', '#FFC000', '#70AD47', '#A5A5A5']
    for idx, (label, vals) in enumerate(nonii_components.items()):
        ax.bar(x3, vals, bottom=bottom, width=0.5, label=label, color=colors_f[idx])
        for j in range(len(years_hist)):
            if vals[j] > 5:
                ax.text(j, bottom[j] + vals[j]/2, f'{int(vals[j])}', ha='center', fontsize=7, color='white', fontweight='bold')
            if j < len(vals): bottom[j] += vals[j]
    ax.set_xticks(x3); ax.set_xticklabels(hist3, fontsize=9)
    ax.set_ylabel('tỷ đồng', fontsize=9)
    ax.legend(fontsize=8, loc='upper left')
    plt.title(f'{TICKER}: Non-Interest Income Breakdown', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartF_nonii.png'), dpi=150)
    plt.close()
    
    # ── CHART G (7): NPL + Group 2 Combo ──
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax2 = ax1.twinx()
    npl_total_abs = npl_total_hist
    gr2_total_abs = npl_gr2_hist
    ax1.bar([i-0.12 for i in x3], npl_total_abs[:len(years_hist)], width=0.2, color='#C00000', alpha=0.7, label='NPL (tỷ)')
    ax1.bar([i+0.12 for i in x3], gr2_total_abs[:len(years_hist)], width=0.2, color='#ED7D31', alpha=0.7, label='Gr2 (tỷ)')
    ax1.set_ylabel('tỷ đồng', fontsize=9)
    npl_line = ax2.plot(x3, npl_ratio_hist, 'D-', color='#C00000', linewidth=2, markersize=6, label='NPL %')[0]
    gr2_line = ax2.plot(x3, gr2_ratio_hist, 's--', color='#ED7D31', linewidth=2, markersize=6, label='Gr2 %')[0]
    for i, v in enumerate(npl_ratio_hist[:len(years_hist)]):
        ax2.text(i, v + 0.02, f'{v:.2f}%', ha='center', fontsize=7, color='#C00000', fontweight='bold')
    for i, v in enumerate(gr2_ratio_hist[:len(years_hist)]):
        ax2.text(i, v + 0.02, f'{v:.2f}%', ha='center', fontsize=7, color='#ED7D31')
    ax2.set_ylabel('%', fontsize=9)
    ax1.set_xticks(x3); ax1.set_xticklabels(hist3, fontsize=9)
    l1, lab1 = ax1.get_legend_handles_labels()
    l2, lab2 = ax2.get_legend_handles_labels()
    ax1.legend(l1 + l2, lab1 + lab2, loc='upper left', fontsize=8)
    plt.title(f'{TICKER}: NPL & Group 2 Combined', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartG_npl_gr2.png'), dpi=150)
    plt.close()
    
    # ── CHART H (8): LLR + CoC Dual ──
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax2 = ax1.twinx()
    llr_vals = [round(prov_hist[i] / max(npl_total_hist[i], 1) * 100, 1) if npl_total_hist[i] > 0.1 else 0 for i in range(len(years_hist))]
    llr_fc_vals = [round(lc * 100, 1) for lc in llr_coverage_fc]
    llr_all = llr_vals + llr_fc_vals
    coc_all = coc_hist + [round(c * 100, 2) for c in coc_fc]
    bars_h = ax1.bar(x, llr_all, width=0.4, color='#70AD47', alpha=0.6, label='LLR (%)')
    ax1.set_ylabel('LLR (%)', fontsize=9)
    for bar in bars_h:
        if bar.get_height() > 0:
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, f'{bar.get_height():.0f}%', ha='center', fontsize=7, color='#70AD47')
    line_h = ax2.plot(x, coc_all, 'D-', color='#4472C4', linewidth=2, markersize=6, label='CoC (%)')[0]
    for i, v in enumerate(coc_all):
        ax2.text(i, v + 0.01, f'{v:.2f}%', ha='center', fontsize=7, color='#4472C4')
    ax2.set_ylabel('CoC (%)', fontsize=9)
    ax1.set_xticks(x); ax1.set_xticklabels(years, fontsize=8)
    ax1.axvline(x=2.5, color='gray', linestyle='--', alpha=0.5)
    l1, lab1 = ax1.get_legend_handles_labels()
    l2, lab2 = ax2.get_legend_handles_labels()
    ax1.legend(l1 + l2, lab1 + lab2, loc='upper right', fontsize=8)
    plt.title(f'{TICKER}: LLR & Cost of Credit (CoC)', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartH_llr_coc.png'), dpi=150)
    plt.close()
    
    # ── CHART I (9): NIM Delta vs Peer ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    nim_peers = peer_val("NIM")
    sorted_nim = sorted(nim_peers.items(), key=lambda kv: kv[1])
    nim_labels = [n[0] for n in sorted_nim]
    nim_vals = [n[1] for n in sorted_nim]
    ticker_nim = PEER_DATA["NIM"].get(TICKER, 0)
    delta_nim = [v - ticker_nim for v in nim_vals]
    colors_i = ['#70AD47' if d > 0 else '#C00000' for d in delta_nim]
    colors_i = ['#2F5496' if l == TICKER else colors_i[idx] for idx, l in enumerate(nim_labels)]
    bars_i = ax.barh(range(len(nim_labels)), delta_nim, color=colors_i, height=0.6)
    for i, (bar, d) in enumerate(zip(bars_i, delta_nim)):
        lbl = f'{d:+.1f}' if abs(d) > 0.05 else '0'
        ax.text(bar.get_width() + 0.02 if d >= 0 else bar.get_width() - 0.5, bar.get_y() + bar.get_height()/2, 
                lbl, va='center', fontsize=7, color='#333')
    ax.axvline(x=0, color='gray', linewidth=0.8)
    ax.set_yticks(range(len(nim_labels))); ax.set_yticklabels(nim_labels, fontsize=8)
    ax.set_xlabel('Δ NIM (bps) so với bank', fontsize=9)
    ax.invert_yaxis()
    plt.title(f'Peer NIM vs {TICKER} (chênh lệch)', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartI_nim_delta.png'), dpi=150)
    plt.close()
    
    # ── CHART J (10): TOI + NPAT Stacked ──
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax2 = ax1.twinx()
    toi_all = toi_hist + [round(x, 1) for x in toi_fc]
    np_all = np_hist + [round(x, 1) for x in np_fc]
    bars_j1 = ax1.bar(x, toi_all, width=0.5, color='#4472C4', alpha=0.3, label='TOI')
    bars_j2 = ax1.bar(x, np_all, width=0.3, color='#70AD47', alpha=0.7, label='NPAT')
    ax1.set_ylabel('tỷ đồng', fontsize=9)
    for bar in bars_j2:
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 20, f'{int(bar.get_height())}', ha='center', fontsize=7, color='#70AD47', fontweight='bold')
    np_growth = [0] + [round((np_all[i]/max(np_all[i-1],1)-1)*100, 1) for i in range(1, len(np_all))]
    line_j = ax2.plot(x, np_growth, 'D-', color='#ED7D31', linewidth=2, markersize=6, label='NPAT Growth %')[0]
    for i, v in enumerate(np_growth):
        if i > 0:
            ax2.text(i, v + 0.5, f'{v:+.1f}%', ha='center', fontsize=7, color='#ED7D31')
    ax2.set_ylabel('Tăng trưởng NPAT (%)', fontsize=9)
    ax1.set_xticks(x); ax1.set_xticklabels(years, fontsize=8)
    ax1.axvline(x=2.5, color='gray', linestyle='--', alpha=0.5)
    l1, lab1 = ax1.get_legend_handles_labels()
    l2, lab2 = ax2.get_legend_handles_labels()
    ax1.legend(l1 + l2, lab1 + lab2, loc='upper left', fontsize=8)
    plt.title(f'{TICKER}: TOI, NPAT & Growth Trend', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartJ_toi_npat.png'), dpi=150)
    plt.close()
    
    # ── CHART K (11): Margins (NIM, ROE, CIR) ──
    fig, ax = plt.subplots(figsize=(8, 4.5))
    roe_fc_calc = [round(np_hist[i]/max((equity_hist[i-1]+equity_hist[i])/2, 1)*100, 2) if i > 0 else round(np_hist[i]/equity_hist[i]*100, 2) for i in range(len(years_hist))]
    roe_all = roe_fc_calc + [round(np_fc[0]/max(equity_hist[-1],1)*100,2)] + \
              [round(np_fc[i]/max((equity_hist[-1]+sum(np_fc[:i])),1)*100,2) for i in range(1,3)]
    cir_all = cir_hist + [round(c*100,2) for c in cir_fc]
    lines_k = {
        'NIM': (nim_all, '#4472C4', '-'),
        'ROE': (roe_all, '#70AD47', '-'),
        'CIR': (cir_all, '#C00000', '--'),
    }
    for name, (vals, color, style) in lines_k.items():
        ax.plot(x, vals, style, color=color, linewidth=2, markersize=5, label=name)
        for i, v in enumerate(vals):
            ax.text(i, v + 0.3, f'{v:.1f}%', ha='center', fontsize=6, color=color)
    ax.axvline(x=2.5, color='gray', linestyle='--', alpha=0.5)
    ax.set_xticks(x); ax.set_xticklabels(years, fontsize=8)
    ax.set_ylabel('%', fontsize=9)
    ax.legend(fontsize=9, loc='upper left')
    plt.title(f'{TICKER}: Key Margins — NIM, ROE & CIR', fontsize=11, fontweight='bold')
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartK_margins.png'), dpi=150)
    plt.close()
    
    # ── CHART L (12): Peer Comparison Table Image ──
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.axis('off')
    col_labels = ['Bank', 'NPL%', 'NIM%', 'CASA%', 'ROE%', 'CIR%', 'P/B', 'Credit\nGrowth']
    row_data = []
    for b in sorted(PEER_BANKS):
        row_data.append([
            b,
            f'{PEER_DATA["NPL"].get(b, 0):.2f}',
            f'{PEER_DATA["NIM"].get(b, 0):.2f}',
            f'{PEER_DATA["CASA"].get(b, 0):.1f}',
            f'{PEER_DATA["ROE"].get(b, 0):.1f}',
            f'{PEER_DATA["CIR"].get(b, 0):.1f}',
            f'{PEER_DATA["P_B"].get(b, 0):.2f}',
            f'{PEER_DATA["CREDIT_GROWTH"].get(b, 0):.1f}',
        ])
    table_data = [col_labels] + row_data
    table = ax.table(cellText=table_data[1:], colLabels=col_labels, loc='center', cellLoc='center')
    table.auto_set_font_size(False); table.set_fontsize(8)
    table.scale(1.1, 1.3)
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_facecolor('#2F5496'); cell.set_text_props(color='white', fontweight='bold')
        else:
            bank_name = table_data[row][0]
            if bank_name == TICKER:
                cell.set_facecolor('#DCE6F1')
                cell.set_text_props(fontweight='bold')
            elif col > 0:
                val = float(table_data[row][col])
                metric_cols = ['NPL%', 'NIM%', 'CASA%', 'ROE%', 'CIR%', 'P/B', 'Credit\nGrowth']
                metric_key = {'NPL%': 'NPL', 'NIM%': 'NIM', 'CASA%': 'CASA', 'ROE%': 'ROE', 'CIR%': 'CIR', 'P/B': 'P_B', 'Credit\nGrowth': 'CREDIT_GROWTH'}[metric_cols[col-1]]
                ind_avg = INDUSTRY_AVG[metric_key]
                if metric_key in ['CIR']:
                    better = val < ind_avg
                else:
                    better = val > ind_avg
                cell.set_facecolor('#E2EFDA' if better else '#FCE4EC')
    plt.title(f'{TICKER}: Peer Comparison — 15 Ngân hàng', fontsize=12, fontweight='bold', pad=20)
    plt.tight_layout()
    plt.savefig(os.path.join(CHART_DIR, 'chartL_peer_table.png'), dpi=150, bbox_inches='tight')
    plt.close()
    
    # ── CHART N (13): PE/PB Historical (ALL quarters, median all-time) ──
    pe_q = sorted([r for r in VAB_RATIOS if r.get("year") and r.get("quarter") in (1,2,3,4)],
                  key=lambda x: (x["year"], x["quarter"]))
    if pe_q:
        fig, ax1 = plt.subplots(figsize=(10, 5))
        ax2 = ax1.twinx()
        pe_q_plot = pe_q[-16:]
        pe_labels_n = [f"Q{r.get('lengthReport')}\n{r['year']}" for r in pe_q_plot]
        pe_vals_n = []
        pe_carried_markers = []
        pb_vals_n = []
        
        # Trace carry-forward chronologically across all quarters
        prev_pe_chart = None
        pe_by_yq = {}
        pe_carried_by_yq = {}
        for r in pe_q:
            y, q = r["year"], r["quarter"]
            pe = r.get("pe")
            npat = q_npat.get((y, q))
            is_carried = False
            if npat is not None and npat < 0 and prev_pe_chart is not None:
                pe = prev_pe_chart
                is_carried = True
            pe_by_yq[(y, q)] = pe
            pe_carried_by_yq[(y, q)] = is_carried
            if pe is not None:
                prev_pe_chart = pe

        for r in pe_q_plot:
            y, q = r["year"], r["quarter"]
            pe_vals_n.append(pe_by_yq.get((y, q)) or pe_all_median)
            pe_carried_markers.append(pe_carried_by_yq.get((y, q), False))
            pb_vals_n.append(r.get("pb") or 0)

        x_n = range(len(pe_labels_n))
        line_pe = ax1.plot(x_n, pe_vals_n, '-', color='#4472C4', linewidth=2, alpha=0.7, label='P/E')[0]
        line_pb = ax2.plot(x_n, pb_vals_n, '-', color='#ED7D31', linewidth=2, alpha=0.7, label='P/B')[0]
        # Normal PE markers
        normal_x = [i for i, c in enumerate(pe_carried_markers) if not c]
        normal_y = [pe_vals_n[i] for i in normal_x]
        ax1.scatter(normal_x, normal_y, color='#4472C4', s=40, zorder=5)
        # Carried PE markers (red hollow circles)
        carried_x = [i for i, c in enumerate(pe_carried_markers) if c]
        carried_y = [pe_vals_n[i] for i in carried_x]
        if carried_x:
            ax1.scatter(carried_x, carried_y, facecolors='none', edgecolors='#FF0000', s=60, zorder=6, linewidths=2, label='P/E dieu chinh')
        # PB markers
        ax2.scatter(x_n, pb_vals_n, color='#ED7D31', s=40, zorder=5, marker='s')
        # All-time median lines
        ax1.axhline(y=pe_all_median, color='#4472C4', linestyle='--', alpha=0.5, linewidth=1.5)
        ax2.axhline(y=pb_all_median, color='#ED7D31', linestyle='--', alpha=0.5, linewidth=1.5)
        ax1.text(0, pe_all_median * 1.08, f'P/E median all: {pe_all_median:.1f}x', fontsize=8, color='#4472C4', alpha=0.7)
        ax2.text(0, pb_all_median * 1.08, f'P/B median all: {pb_all_median:.2f}x', fontsize=8, color='#ED7D31', alpha=0.7)
        for i, v in enumerate(pe_vals_n):
            offset = 0.5 if v < max(pe_vals_n) * 0.3 else 0.3
            ax1.text(i, v + offset, f'{v:.1f}', ha='center', fontsize=6.5, color='#4472C4', alpha=0.8)
        for i, v in enumerate(pb_vals_n):
            offset = 0.03 if v < max(pb_vals_n) * 0.3 else 0.02
            ax2.text(i, v + offset, f'{v:.2f}', ha='center', fontsize=6.5, color='#ED7D31', alpha=0.8)
        ax1.set_xticks(x_n); ax1.set_xticklabels(pe_labels_n, fontsize=7)
        ax1.set_ylabel('P/E (x)', fontsize=9, color='#4472C4')
        ax2.set_ylabel('P/B (x)', fontsize=9, color='#ED7D31')
        l1, lab1 = ax1.get_legend_handles_labels()
        l2, lab2 = ax2.get_legend_handles_labels()
        ax1.legend(l1 + l2, lab1 + lab2, loc='upper left', fontsize=8)
        plt.title(f'{TICKER}: P/E & P/B Historical (median all-time: PE={pe_all_median:.1f}x, PB={pb_all_median:.2f}x)', fontsize=11, fontweight='bold')
        plt.tight_layout()
        plt.savefig(os.path.join(CHART_DIR, 'chartN_pe_pb_history.png'), dpi=150)
        plt.close()
    print(f"[Charts] Generated 13 chart images")

_VN_FONTS = register_vn_fonts()
FONT_REG = 'Arial' if 'Arial' in _VN_FONTS else 'Helvetica'
FONT_BOLD = 'Arial-Bold' if 'Arial-Bold' in _VN_FONTS else 'Helvetica-Bold'

def create_pdf():
    print("[PDF] Building document...")
    create_charts()
    # Find latest quarter records for PDF text update
    is_q_recs_all = sorted(section_to_quarters(FIN_DATA, "INCOME_STATEMENT"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
    last_q = is_q_recs_all[-1]
    prev_q = is_q_recs_all[-2]
    yoy_q = None
    for r in is_q_recs_all:
        if r.get('lengthReport') == last_q.get('lengthReport') and r.get('yearReport') == last_q.get('yearReport') - 1:
            yoy_q = r
            break
    q_num = last_q.get('lengthReport')
    q_yr = last_q.get('yearReport')
    npat_l = (last_q.get("isa20") or 0) / 1e9
    npat_p = (prev_q.get("isa20") or 0) / 1e9
    npat_y = (yoy_q.get("isa20") or 0) / 1e9 if yoy_q else 0
    nii_l = (last_q.get("isb27") or 0) / 1e9
    nii_p = (prev_q.get("isb27") or 0) / 1e9
    nii_y = (yoy_q.get("isb27") or 0) / 1e9 if yoy_q else 0
    toi_l = (last_q.get("isb38") or 0) / 1e9
    toi_p = (prev_q.get("isb38") or 0) / 1e9
    toi_y = (yoy_q.get("isb38") or 0) / 1e9 if yoy_q else 0
    npat_qoq = (npat_l/npat_p - 1)*100 if npat_p else 0
    npat_yoy = (npat_l/npat_y - 1)*100 if npat_y else 0
    toi_qoq = (toi_l/toi_p - 1)*100 if toi_p else 0
    toi_yoy = (toi_l/toi_y - 1)*100 if toi_y else 0
    
    doc = SimpleDocTemplate(PDF_FILE, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18, leading=22, spaceAfter=12, fontName=FONT_REG)
    h1_style = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=14, leading=18, spaceAfter=8, spaceBefore=16, textColor=HexColor('#2F5496'), fontName=FONT_BOLD)
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=12, leading=15, spaceAfter=6, spaceBefore=12, textColor=HexColor('#2F5496'), fontName=FONT_BOLD)
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=14, spaceAfter=6, fontName=FONT_REG)
    bullet_style = ParagraphStyle('Bullet', parent=body_style, leftIndent=20, bulletIndent=10, spaceAfter=4)
    
    story = []
    
    # ── Cover Page ──
    story.append(Spacer(1, 80*mm))
    story.append(Paragraph(f"PHÂN TÍCH CỔ PHIẾU {TICKER}", title_style))
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(f"<b>{COMPANY}</b>", ParagraphStyle('Sub', parent=title_style, fontSize=14, leading=18)))
    story.append(Spacer(1, 5*mm))
    sb = ParagraphStyle('Sub2', parent=body_style, fontSize=11, alignment=TA_CENTER)
    story.append(Paragraph(f"Sàn: {EXCHANGE} | Ngành: {INDUSTRY}", sb))
    story.append(Spacer(1, 15*mm))
    
    rec_style = ParagraphStyle('Rec', parent=title_style, fontSize=16, alignment=TA_CENTER)
    rec_text = 'MUA' if upside > 15 else 'THEO DÕI' if upside > 5 else 'NẮM GIỮ'
    story.append(Paragraph(f"KHUYẾN NGHỊ: <b>{rec_text}</b>", rec_style))
    story.append(Spacer(1, 5*mm))
    target_style = ParagraphStyle('Target', parent=title_style, fontSize=14, alignment=TA_CENTER, textColor=HexColor('#006400') if upside > 0 else HexColor('#C00000'))
    story.append(Paragraph(f"Giá hiện tại: {PRICE:,} VND | Target: {weighted_target:,.0f} VND | Upside: {upside:+.1f}%", target_style))
    story.append(Spacer(1, 20*mm))
    story.append(Paragraph(f"Ngày: {MONTH} | Phân tích bởi AI Framework FA", ParagraphStyle('Date', parent=body_style, fontSize=9, alignment=TA_CENTER, textColor=grey)))
    story.append(PageBreak())
    
    # ── Investment Summary ──
    story.append(Paragraph("1. Investment Summary", h1_style))
    story.append(Paragraph(f"{TICKER} ({COMPANY}) là ngân hàng TMCP quy mô nhỏ với vốn hóa ~9,000 tỷ. "
                          f"VAB đang giao dịch tại P/B {pb_current:.2f}x (dưới 1x) và P/E {MARKET_CAP/(np_hist[-1]*1e9):.1f}x — "
                          f"mức định giá rẻ so với trung bình ngành. Tuy nhiên, CASA chỉ {casa_ratio_hist[-1]}% (rất thấp) "
                          f"và NIM {nim_hist[-1]}% (dưới trung bình) phản ánh cấu trúc vốn yếu.", body_style))
    story.append(Spacer(1, 3*mm))
    
    # Summary table
    sum_data = [
        ["Chỉ tiêu", "2023", "2024", "2025", "2026F"],
        ["NII (tỷ)", str(round(nii_hist[0])), str(round(nii_hist[1])), str(round(nii_hist[2])), str(round(nii_fc[0]))],
        ["LNST (tỷ)", str(round(np_hist[0])), str(round(np_hist[1])), str(round(np_hist[2])), str(round(np_fc[0]))],
        ["NIM (%)", str(nim_hist[0]), str(nim_hist[1]), str(nim_hist[2]), str(nim_fc[0])],
        ["NPL (%)", str(npl_ratio_hist[0]), str(npl_ratio_hist[1]), str(npl_ratio_hist[2]), str(round(npl_fc[0]*100,2))],
        ["CASA (%)", str(casa_ratio_hist[0]), str(casa_ratio_hist[1]), str(casa_ratio_hist[2]), str(round(casa_target_fc[0]*100,2))],
        ["ROE (%)", str(roe_hist[0]), str(roe_hist[1]), str(roe_hist[2]), str(round(np_fc[0]/equity_hist[-1]*100,2))],
        ["P/B (x)", "", "", f"{pb_current:.2f}", f"{pb_target:.2f}"],
    ]
    t = Table(sum_data, colWidths=[80,55,55,55,55])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor('#2F5496')),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor('#D9D9D9')),
        ('BACKGROUND', (0,1), (-1,-1), HexColor('#F2F2F2') if not len(sum_data)%2 else white),
        ('FONTNAME', (0,0), (0,-1), FONT_BOLD),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))
    story.append(Paragraph(f"<b>1.1 Cập nhật Kết quả Kinh doanh Quý gần nhất (Q{q_num}/{q_yr})</b>", h2_style))
    story.append(Paragraph(
        f"Trong quý {q_num}/{q_yr}, {TICKER} ghi nhận lợi nhuận sau thuế đạt <b>{npat_l:,.1f} tỷ</b>, "
        f"{'tăng trưởng' if npat_yoy > 0 else 'giảm'} <b>{abs(npat_yoy):.1f}% YoY</b> so với cùng kỳ năm trước "
        f"và {'tăng trưởng' if npat_qoq > 0 else 'giảm'} <b>{abs(npat_qoq):.1f}% QoQ</b> so với quý trước. "
        f"Tổng thu nhập hoạt động (TOI) đạt <b>{toi_l:,.1f} tỷ</b> ({'tăng' if toi_yoy > 0 else 'giảm'} {abs(toi_yoy):.1f}% YoY, "
        f"{'tăng' if toi_qoq > 0 else 'giảm'} {abs(toi_qoq):.1f}% QoQ). "
        f"Thu nhập lãi thuần (NII) đóng góp <b>{nii_l:,.1f} tỷ</b>, tương đương <b>{nii_l/max(toi_l, 0.1)*100:.1f}%</b> tổng thu nhập.",
        body_style
    ))
    story.append(Spacer(1, 3*mm))
    # 3 reasons
    story.append(Paragraph("<b>3 Lý do MUA:</b>", h2_style))
    story.append(Paragraph("• P/B dưới 1x (0.86x) — định giá rẻ so với BVPS ~12,400 VND", bullet_style))
    story.append(Paragraph("• NPL giảm dần từ 2.5% (2023) → 1.3% (2026) — chất lượng TS cải thiện", bullet_style))
    story.append(Paragraph("• LNST CAGR 33% (2023-2025) — tăng trưởng mạnh từ nền thấp", bullet_style))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph("<b>3 Rủi ro:</b>", h2_style))
    story.append(Paragraph("• CASA 4.8% — rất thấp, gây COF cao (5.2%) → NIM thấp (2.7%)", bullet_style))
    story.append(Paragraph("• CAR ~9.4% — gần ngưỡng 8%, hạn chế room mở rộng tín dụng", bullet_style))
    story.append(Paragraph("• Thanh khoản thấp — MCap ~9,000 tỷ, khó giao dịch khối lượng lớn", bullet_style))
    story.append(PageBreak())
    
    # ── Asset Quality ──
    story.append(Paragraph("2. Chất lượng Tài sản (Asset Quality)", h1_style))
    story.append(Paragraph(f"NPL của {TICKER} duy trì xu hướng giảm từ 2.5% (Q2/2023) xuống còn {npl_ratio_hist[-1]}% (Q1/2026), "
                          f"thấp hơn trung bình ngành (~1.8%). Nợ nhóm 2 (Gr2) ở mức gần 0% — rất sạch. "
                          f"Tỷ lệ bao phủ nợ xấu (LLR/NPL) khoảng {prov_hist[-1]/max(npl_total_hist[-1],1)*100:.0f}%, "
                          f"dưới ngưỡng 100% — đây là điểm yếu.", body_style))
    story.append(Spacer(1, 3*mm))
    
    aq_data = [
        ["Chỉ tiêu", "2023", "2024", "2025"],
        ["NPL (%)", str(npl_ratio_hist[0]), str(npl_ratio_hist[1]), str(npl_ratio_hist[2])],
        ["Nợ nhóm 2 (%)", f"{gr2_ratio_hist[0]:.2f}", f"{gr2_ratio_hist[1]:.2f}", f"{gr2_ratio_hist[2]:.2f}"],
        ["Dự phòng (tỷ)", str(round(prov_hist[0])), str(round(prov_hist[1])), str(round(prov_hist[2]))],
        ["LLR/NPL (%)", f"{prov_hist[0]/max(npl_total_hist[0],1)*100:.0f}" if npl_total_hist[0]>0.1 else "N/A",
         f"{prov_hist[1]/max(npl_total_hist[1],1)*100:.0f}" if npl_total_hist[1]>0.1 else "N/A",
         f"{prov_hist[2]/max(npl_total_hist[2],1)*100:.0f}" if npl_total_hist[2]>0.1 else "N/A"],
        ["CAR (%)", "~9.3", "~9.2", "~9.4"],
        ["CoC (%)", f"{coc_hist[0]:.2f}", f"{coc_hist[1]:.2f}", f"{coc_hist[2]:.2f}"],
    ]
    t = Table(aq_data, colWidths=[80,55,55,55])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor('#2F5496')), ('TEXTCOLOR', (0,0), (-1,0), white),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, HexColor('#D9D9D9')),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))
    # ── Asset Quality Charts ──
    story.append(Image(os.path.join(CHART_DIR, 'chartG_npl_gr2.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartH_llr_coc.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartB_peer_npl.png'), width=480, height=270))
    story.append(PageBreak())
    
    # ── Credit & Growth ──
    story.append(Paragraph("3. Tín dụng & Tăng trưởng", h1_style))
    story.append(Paragraph(f"<b>Credit Growth vs Peer:</b>", h2_style))
    story.append(Paragraph(f"• Tăng trưởng tín dụng {TICKER}: {PEER_DATA['CREDIT_GROWTH'].get(TICKER,0):.1f}% (YTD)", bullet_style))
    story.append(Paragraph(f"• Trung bình ngành: {INDUSTRY_AVG['CREDIT_GROWTH']:.1f}%", bullet_style))
    story.append(Paragraph(f"• LDR: {ldr_hist[-1]}% — cần theo dõi nếu >90%", bullet_style))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartC_peer_credit.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartE_earning_assets.png'), width=480, height=270))
    story.append(PageBreak())
    
    # ── Profitability & Efficiency ──
    story.append(Paragraph("4. Phân tích Lợi nhuận & Hiệu quả", h1_style))
    story.append(Paragraph(f"<b>NIM Decomposition:</b>", h2_style))
    yoea_v, cof_v = calc_yoea_cof()
    story.append(Paragraph(f"• YOEA (Lợi suất TS): {yoea_v[-1]:.2f}% (2025)", bullet_style))
    story.append(Paragraph(f"• COF (Chi phí vốn): {cof_v[-1]:.2f}% — cao do CASA thấp", bullet_style))
    story.append(Paragraph(f"• Spread = YOEA − COF = {yoea_v[-1]-cof_v[-1]:.2f}%", bullet_style))
    story.append(Paragraph(f"• NIM thực tế: {nim_hist[-1]}% — chênh lệch spread do CASA thấp", bullet_style))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartA_nim_decomp.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(f"<b>Cấu trúc thu nhập:</b>", h2_style))
    nii_pct = nii_hist[-1] / toi_hist[-1] * 100
    story.append(Paragraph(f"• NII chiếm {nii_pct:.1f}% TOI — phụ thuộc nhiều vào thu nhập lãi", bullet_style))
    story.append(Paragraph(f"• Non-interest income: {100-nii_pct:.1f}% — thấp, cần đa dạng hóa", bullet_style))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartF_nonii.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(f"<b>PPOP Quality:</b>", h2_style))
    story.append(Paragraph(f"• PPOP CAGR 2023-2025: 32% — tăng trưởng thực", bullet_style))
    story.append(Paragraph(f"• PPOP/TOI: {ppop_hist[-1]/toi_hist[-1]*100:.1f}% — hiệu quả HĐ cao", bullet_style))
    if ppop_hist[-1] / toi_hist[-1] > 0.7:
        story.append(Paragraph(f"• Cảnh báo: CIR quá thấp ({cir_hist[-1]}%) có thể do chưa đầu tư đủ mức. Cần kiểm tra chi phí đầu tư CNTT, mở rộng mạng lưới.", body_style))
    story.append(Image(os.path.join(CHART_DIR, 'chartJ_toi_npat.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartK_margins.png'), width=480, height=270))
    story.append(PageBreak())
    
    # ── Industry Context & Peer ──
    story.append(Paragraph("5. So sánh Ngành & Peer", h1_style))
    story.append(Paragraph(f"<b>{TICKER} vs Industry Average:</b>", h2_style))
    for m in ['NIM', 'ROE', 'CIR', 'CASA', 'NPL']:
        bv = PEER_DATA[m].get(TICKER, 0)
        iv = INDUSTRY_AVG[m]
        better = bv < iv if m == 'CIR' else bv > iv
        icon = '✅' if better else '⚠️'
        story.append(Paragraph(f"• {m}: {TICKER} {bv:.2f}% vs TB ngành {iv:.2f}% {icon}", bullet_style))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartD_vs_industry.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartI_nim_delta.png'), width=480, height=270))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartL_peer_table.png'), width=480, height=320))
    story.append(PageBreak())
    
    # ── Valuation Detail ──
    story.append(Paragraph("6. Định giá Chi tiết", h1_style))
    story.append(Paragraph(f"<b>Phương pháp: Residual Income (50%) + P/B Multiple (50%)</b>", h2_style))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(f"<b>Residual Income Model:</b>", h2_style))
    story.append(Paragraph(f"• BV/share hiện tại: {bvps_base:,.0f} VND", bullet_style))
    story.append(Paragraph(f"• ROE dự phóng: {roe_hist[-1]}% (2025) → ~15% (2028F)", bullet_style))
    story.append(Paragraph(f"• Chi phí vốn CSH (COE): {COE*100:.0f}%  (rf 4.5% + β 1.0 × ERP 8.5%)", bullet_style))
    story.append(Paragraph(f"• Tăng trưởng dài hạn (g): {terminal_growth*100:.0f}%", bullet_style))
    story.append(Spacer(1, 3*mm))
    
    val_data = [
        ["Thành phần", "Giá trị (VND)", "Ghi chú"],
        ["BV/share", f"{bvps_base:,.0f}", "Giá trị sổ sách cuối 2025"],
        ["PV RI 3 năm", f"{pv_ri:,.0f}", "RI mỗi năm chiết khấu về hiện tại"],
        ["PV Continuing Value", f"{pv_cv:,.0f}", f"RI năm cuối × (1+g)/(COE-g)"],
        ["GIÁ TRỊ RI", f"{ri_value:,.0f}", f"Tổng BV + PV RI + PV CV"],
        ["", "", ""],
        ["P/B target", f"{pb_target:.2f}x", f"Median all: {pb_all_median:.2f}x"],
        ["BV/share tương lai (2026F)", f"{bvps_base + eps_fc_calc[0]:,.0f}", "BVPS 2025 + EPS 2026F"],
        ["GIÁ TRỊ P/B", f"{pb_value:,.0f}", f"P/B target × BV/share tương lai"],
        ["", "", ""],
        ["WEIGHTED TARGET", f"{weighted_target:,.0f}", f"50% RI + 50% P/B"],
        ["Giá hiện tại", f"{PRICE:,}", ""],
        ["UPSIDE", f"{upside:+.1f}%", ""],
    ]
    t = Table(val_data, colWidths=[120, 100, 180])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor('#2F5496')), ('TEXTCOLOR', (0,0), (-1,0), white),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (0,-1), 'LEFT'), ('ALIGN', (1,0), (1,-1), 'RIGHT'), ('ALIGN', (2,0), (2,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor('#D9D9D9')),
        ('BACKGROUND', (0,10), (-1,10), HexColor('#E2EFDA')),
        ('FONTNAME', (0,9), (-1,-1), FONT_BOLD),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))
    
    # Sensitivity table
    story.append(Paragraph(f"<b>Sensitivity: COE × g</b> — Giá trị RI trong các kịch bản", h2_style))
    sen_data = [["COE \\ g"] + [f"g={g:.0%}" for g in [0.01, 0.02, 0.03, 0.04, 0.05]]]
    for coe in [0.10, 0.115, 0.13, 0.145, 0.16]:
        row = [f"COE={coe:.0%}"]
        for g in [0.01, 0.02, 0.03, 0.04, 0.05]:
            if coe <= g: row.append("N/A")
            else:
                cv = ri_results[-1]*(1+g)/(coe-g) if ri_results else 0
                pv = sum([ri_results[k]/(1+coe)**(k+1) for k in range(len(ri_results))])
                pv_cv_loc = cv/(1+coe)**len(ri_results)
                row.append(f"{bvps_base+pv+pv_cv_loc:,.0f}")
        sen_data.append(row)
    t = Table(sen_data, colWidths=[90] + [65]*5)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor('#2F5496')), ('TEXTCOLOR', (0,0), (-1,0), white),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (0,-1), FONT_BOLD), ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 0.5, HexColor('#D9D9D9')),
    ]))
    story.append(t)
    story.append(PageBreak())
    
    # ── Historical Valuation ──
    story.append(Paragraph("7. Lịch sử Định giá", h1_style))
    story.append(Paragraph(f"<b>P/E & P/B Historical (all-time median):</b>", h2_style))
    if pe_all_vals:
        pe_latest = pe_all_vals[-1]
        pb_latest = pb_all_vals[-1]
        story.append(Paragraph(f"• P/E hiện tại: {pe_latest:.1f}x (trung vị all-time: {pe_all_median:.1f}x) "
                              f"— {'cao hơn' if pe_latest > pe_all_median else 'thấp hơn'} trung vị", bullet_style))
        story.append(Paragraph(f"• P/B hiện tại: {pb_latest:.2f}x (trung vị all-time: {pb_all_median:.2f}x) "
                              f"— {'cao hơn' if pb_latest > pb_all_median else 'thấp hơn'} trung vị", bullet_style))
        story.append(Paragraph(f"• Định giá hiện tại ở vùng {'cao' if pe_latest > pe_all_median * 1.1 else 'thấp' if pe_latest < pe_all_median * 0.9 else 'trung bình'} lịch sử", body_style))
    story.append(Spacer(1, 3*mm))
    story.append(Image(os.path.join(CHART_DIR, 'chartN_pe_pb_history.png'), width=480, height=270))
    story.append(PageBreak())
    
    # ── SWOT ──
    story.append(Paragraph("8. SWOT & Kết luận", h1_style))
    swot_data = [
        ["STRENGTHS", "WEAKNESSES"],
        [f"NPL giảm ({npl_ratio_hist[-1]}%)\nCIR thấp ({cir_hist[-1]}%)\nTăng vốn thành công (+51%)\nLNST CAGR 33%",
         f"CASA rất thấp ({casa_ratio_hist[-1]}%)\nNIM thấp ({nim_hist[-1]}%)\nLLR < 100%\nCAR gần ngưỡng ({9.4}%)"],
        ["OPPORTUNITIES", "THREATS"],
        ["Mở rộng CASA từ nền thấp\nTận dụng room tín dụng 14%\nTăng vốn → mở rộng quy mô",
         "LS huy động tăng → COF↑\nCạnh tranh từ bank lớn\nThanh khoản thấp\nSiết room tín dụng NHNN"],
    ]
    t = Table(swot_data, colWidths=[235, 235])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), HexColor('#F2F2F2')),
        ('BACKGROUND', (0,1), (0,1), HexColor('#DCE6F1')), ('TEXTCOLOR', (0,1), (0,1), HexColor('#2F5496')),
        ('BACKGROUND', (1,1), (1,1), HexColor('#FCE4EC')), ('TEXTCOLOR', (1,1), (1,1), HexColor('#C00000')),
        ('BACKGROUND', (0,3), (0,3), HexColor('#E2EFDA')), ('TEXTCOLOR', (0,3), (0,3), HexColor('#006400')),
        ('BACKGROUND', (1,3), (1,3), HexColor('#FFF3E0')), ('TEXTCOLOR', (1,3), (1,3), HexColor('#ED7D31')),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'), ('VALIGN', (0,1), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor('#D9D9D9')),
        ('FONTNAME', (0,2), (-1,2), FONT_BOLD), ('FONTSIZE', (0,2), (-1,-2), 9),
    ]))
    story.append(t)
    story.append(Spacer(1, 10*mm))
    
    ups = upside
    rec = "MUA" if ups > 15 else "THEO DÕI" if ups > 5 else "NẮM GIỮ" if ups > -5 else "BÁN"
    conclusion = (
        f"<b>Kết luận: {rec}</b><br/><br/>"
        f"{TICKER} đang giao dịch tại P/B {pb_current:.2f}x — thấp hơn giá trị sổ sách. "
        f"Định giá theo Residual Income + P/B cho target {weighted_target:,.0f} VND, upside {ups:+.1f}%. "
        f"VAB có chất lượng tài sản cải thiện (NPL giảm), tăng trưởng LNST mạnh (CAGR 33%) và vừa tăng vốn thành công. "
        f"Tuy nhiên, CASA cực thấp ({casa_ratio_hist[-1]}%) và NIM thấp ({nim_hist[-1]}%) là rủi ro cốt lõi. "
        f"Quy mô nhỏ (MCap ~9,000 tỷ) và thanh khoản thấp là rủi ro khi đầu tư. "
        f"Khuyến nghị: {'MUA với kỳ vọng P/B hồi phục về 1x+' if ups > 15 else 'THEO DÕI, chờ tín hiệu CASA cải thiện và NIM mở rộng.' if ups > 5 else 'NẮM GIỮ với tầm nhìn dài hạn.' if ups > -5 else 'BÁN do upside hạn chế.'}"
    )
    story.append(Paragraph(conclusion, body_style))
    
    doc.build(story)
    print(f"[PDF] Saved to {PDF_FILE}")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8') if hasattr(sys.stdout, 'reconfigure') else None
    print(f"=== {TICKER} Financial Model Builder ===")
    print(f"Company: {COMPANY}")
    print(f"Price: {PRICE:,} VND | MCap: {MARKET_CAP/1e9:,.0f} tỷ")
    print(f"P/B: {pb_current:.2f}x | P/E: {MARKET_CAP/(np_hist[-1]*1e9):.1f}x")
    print()
    print("--- Historical Summary ---")
    for i, y in enumerate(years_hist):
        print(f"  {y}: NII {nii_hist[i]:.0f} TOI {toi_hist[i]:.0f} PPOP {ppop_hist[i]:.0f} NP {np_hist[i]:.0f}")
    print()
    print("--- Forecast Summary ---")
    for i, y in enumerate(years_fc):
        print(f"  {y}F: NII {nii_fc[i]:.0f} TOI {toi_fc[i]:.0f} NP {np_fc[i]:.0f} NIM {nim_fc[i]}%")
    print()
    print("--- Valuation ---")
    print(f"  RI Value: {ri_value:,.0f} VND")
    print(f"  P/B Value: {pb_value:,.0f} VND (P/B {pb_target:.2f}x)")
    print(f"  Weighted Target: {weighted_target:,.0f} VND")
    print(f"  Current Price: {PRICE:,} VND")
    print(f"  UPSIDE: {upside:+.1f}%")
    print()
    
    create_excel()
    create_pdf()
    print(f"\n=== DONE ===")
    print(f"Excel: {EXCEL_FILE}")
    print(f"PDF:   {PDF_FILE}")
