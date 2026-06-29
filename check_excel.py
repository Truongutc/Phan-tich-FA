import openpyxl
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'E:\1. Projects\4. AIC - FA\Bao cao\VIB\VIB_Model_2026-06.xlsx')

ws_rat = wb['06_Ratios']
print('=== 06_Ratios ===')
for r in range(1, 12):
    v = ws_rat.cell(row=r, column=1).value
    if v:
        print(f'  Row {r}: {repr(v)[:70]}')

ws_bs = wb['05_Balance_Sheet']
print('\n=== 05_Balance_Sheet ===')
for r in range(1, 8):
    v = ws_bs.cell(row=r, column=1).value
    if v:
        print(f'  Row {r}: {repr(v)[:70]}')

ws_val = wb['07_Valuation']
print('\n=== 07_Valuation key cells ===')
for col_idx, col_name in [(2,'B'), (3,'C'), (4,'D')]:
    cell = ws_val.cell(row=26, column=col_idx)
    print(f'  {col_name}26 EPS: {repr(cell.value)[:80]}')
print(f'  B6 BVPS hien tai: {repr(ws_val.cell(row=6, column=2).value)}')
print(f'  B16 BVPS forward: {repr(ws_val.cell(row=16, column=2).value)}')
print(f'  B9  RI Value: {repr(ws_val.cell(row=9, column=2).value)}')
print(f'  B17 PB Value: {repr(ws_val.cell(row=17, column=2).value)}')
print(f'  B21 Target Price: {repr(ws_val.cell(row=21, column=2).value)}')
print(f'  B13 PB attractive: {ws_val.cell(row=13, column=2).value}')
print(f'  B14 PB median: {ws_val.cell(row=14, column=2).value}')
print(f'  B15 PB target: {repr(ws_val.cell(row=15, column=2).value)}')

ws_pest = wb['09_PESTLE']
print('\n=== 09_PESTLE ===')
cnt = 0
for r in range(2, 10):
    v = ws_pest.cell(row=r, column=1).value
    if v:
        cnt += 1
        print(f'  Row {r}: {repr(v)[:60]}')
print(f'  TOTAL: {cnt} factors (expect 6)')

ws_sens = wb['08_Sensitivity']
print('\n=== 08_Sensitivity ===')
print(f'  Title: {repr(ws_sens.cell(row=1, column=1).value)}')
print(f'  COE A3: {ws_sens.cell(row=3, column=1).value}')
print(f'  Value B3 (COE=0.08, g=0.01): {ws_sens.cell(row=3, column=2).value}')
print(f'  Value C5 (COE=0.10, g=0.02): {ws_sens.cell(row=5, column=3).value}')

ws_ass = wb['02_Assumptions']
print('\n=== 02_Assumptions row 12 (COE) ===')
for col in range(7, 10):
    v = ws_ass.cell(row=12, column=col).value
    print(f'  Col {col}: {v} (type={type(v).__name__})')
