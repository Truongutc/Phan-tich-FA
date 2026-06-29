import openpyxl

def check_file(path):
    print(f"\n=== FILE: {path} ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['07_Valuation']
    
    print("Row 2 (COE):", ws['B2'].value)
    print("Row 3 (g):", ws['B3'].value)
    print("Row 6 (BVPS HT):", ws['B6'].value)
    print("Row 7 (PV RI):", ws['B7'].value)
    print("Row 8 (PV CV):", ws['B8'].value)
    print("Row 9 (RI Value):", ws['B9'].value)
    
    print("D26 (EPS Y3):", ws['D26'].value)
    print("D27 (BVPS Y3):", ws['D27'].value)
    print("D28 (Capital Charge Y3):", ws['D28'].value)
    print("D29 (RI Y3):", ws['D29'].value)
    print("D30 (DF Y3):", ws['D30'].value)
    print("D31 (PV RI Y3):", ws['D31'].value)

check_file(r'E:\1. Projects\4. AIC - FA\Bao cao\VIB\VIB_Model_2026-06.xlsx')
check_file(r'E:\1. Projects\4. AIC - FA\Bao cao\TCB\TCB_Model_2026-06.xlsx')
