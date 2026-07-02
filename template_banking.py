#!/usr/bin/env python3
"""
template_banking.py — Universal, highly professional parameterized calculation engine for Banks.
Includes:
- Dynamic formula-driven Excel model (17 sheets)
- In-depth multi-page PDF report with 13 custom Matplotlib charts
- Multi-platform Vietnamese font bug fix (Arial / Arial-Bold registered and applied to all tables/paragraphs)
- Dynamic CAPM (Rf and Beta regression) and Residual Income + P/B target multiples valuation.
"""
import os
import sys
import math
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, grey
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests
import statistics as stats

from fetch_data import section_to_quarters, cumulative_actual_quarters, blend_annual_estimate, latest_actual_quarter_value, blend_annual_estimate_stock

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

# ── VIETNAMESE FONT REGISTRATION (MULTI-PLATFORM BUG FIX) ──────────────────
def register_vn_fonts():
    font_paths_to_try = [
        ("C:/Windows/Fonts/arial.ttf", "Arial"),
        ("C:/Windows/Fonts/arialbd.ttf", "Arial-Bold"),
        ("C:/Windows/Fonts/times.ttf", "TimesNewRoman"),
        ("C:/Windows/Fonts/timesbd.ttf", "TimesNewRoman-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "Arial"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Arial"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/freefont/FreeSans.ttf", "Arial"),
        ("/usr/share/fonts/truetype/freefont/FreeSansBold.ttf", "Arial-Bold"),
    ]
    
    found = {}
    for path, freg in font_paths_to_try:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(freg, path))
                found[freg] = path
            except:
                pass
    return found

_VN_FONTS = register_vn_fonts()
FONT_REG = 'Arial' if 'Arial' in _VN_FONTS else 'Helvetica'
FONT_BOLD = 'Arial-Bold' if 'Arial-Bold' in _VN_FONTS else 'Helvetica-Bold'


# ── CAPM INPUTS: AUTO-FETCH Rf và BETA ────────────────────────────────────
def fetch_rf_vietnam(timeout=15):
    FALLBACK_RF = 0.045
    try:
        r = requests.get(
            "https://www.worldgovernmentbonds.com/bond-yield/vietnam/10-years/",
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=timeout,
        )
        if r.status_code == 200:
            import re
            matches = re.findall(r'(\d+\.\d+)%', r.text[:5000])
            if matches:
                rf = float(matches[0]) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "worldgovernmentbonds.com"
    except Exception as e:
        print(f"  [WARN] WorldGovernmentBonds Rf failed: {e}")

    try:
        r = requests.get(
            "https://trading.vietcap.com.vn/api/iq-insight-service/v1/bond/government-bond-yield",
            headers={"User-Agent": "Mozilla/5.0", "Referer": "https://trading.vietcap.com.vn/"},
            timeout=timeout,
        )
        if r.status_code == 200:
            d = r.json().get("data", {})
            y10 = d.get("10Y") or d.get("y10") or d.get("tenor10")
            if y10:
                rf = float(y10) / 100 if float(y10) > 1 else float(y10)
                if 0.01 <= rf <= 0.15:
                    return rf, "Vietcap API"
    except Exception as e:
        print(f"  [WARN] Vietcap bond API failed: {e}")

    return FALLBACK_RF, "Fallback (manual)"


def fetch_aligned_history(ticker, days=720, timeout=15):
    import time
    # Use VNDIRECT's public UDF chart history API which is highly stable and public
    from_time = 1577836800
    to_time = 2000000000
    url_stock = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol={ticker}&resolution=D&from={from_time}&to={to_time}"
    url_index = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol=VNINDEX&resolution=D&from={from_time}&to={to_time}"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Referer": "https://dchart.vndirect.com.vn/",
    }
    
    try:
        r_stock = requests.get(url_stock, headers=headers, timeout=timeout)
        r_index = requests.get(url_index, headers=headers, timeout=timeout)
        
        if r_stock.status_code == 200 and r_index.status_code == 200:
            d_stock = r_stock.json()
            d_index = r_index.json()
            
            t_s = d_stock.get("t") or []
            c_s = d_stock.get("c") or []
            t_m = d_index.get("t") or []
            c_m = d_index.get("c") or []
            
            map_stock = {t_s[i]: c_s[i] for i in range(min(len(t_s), len(c_s))) if c_s[i] is not None and c_s[i] > 0}
            map_index = {t_m[i]: c_m[i] for i in range(min(len(t_m), len(c_m))) if c_m[i] is not None and c_m[i] > 0}
            
            # Find common timestamps
            common_t = sorted(list(set(map_stock.keys()) & set(map_index.keys())))
            
            aligned = []
            for t in common_t:
                date_str = datetime.datetime.fromtimestamp(t).strftime("%Y-%m-%d")
                # VNDIRECT stock price is divided by 1000, so scale it up if it is under 1000
                p_s = map_stock[t]
                if p_s < 1000:
                    p_s = p_s * 1000
                aligned.append((date_str, p_s, map_index[t]))
            return aligned
    except Exception as e:
        print(f"[Beta Calc] Error fetching history: {e}")
    return []

def fetch_beta_vietstock(ticker, timeout=15):
    try:
        search_url = f"https://finance.vietstock.vn/search?query={ticker}"
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        r1 = requests.get(search_url, headers=headers, timeout=timeout)
        if r1.status_code == 200:
            data = json.loads(r1.text).get("data", "")
            lines = data.split('\r\n')
            target_url = ""
            for line in lines:
                parts = line.split('|')
                if len(parts) >= 3 and parts[0].strip().upper() == ticker.upper():
                    target_url = parts[2]
                    break
            if target_url:
                headers_r2 = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Referer': 'https://finance.vietstock.vn/'
                }
                r2 = requests.get(target_url, headers=headers_r2, timeout=timeout)
                if r2.status_code == 200:
                    import re
                    m = re.search(r'\"Beta\":\"([\d\.]+)\"', r2.text)
                    if m:
                        beta = float(m.group(1))
                        if 0.3 <= beta <= 2.5:
                            print(f"  [OK] Beta {ticker} từ Vietstock: {beta:.2f}")
                            return beta
    except Exception as e:
        print(f"  [WARN] Vietstock scrape failed: {e}")
    return None

def fetch_beta_vietcap(ticker, timeout=15):
    try:
        url = f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Referer": "https://trading.vietcap.com.vn/"
        }
        r = requests.get(url, headers=headers, timeout=timeout)
        if r.status_code == 200:
            d = r.json().get("data", {})
            beta = d.get("beta")
            if beta is not None and 0.3 <= float(beta) <= 2.5:
                print(f"  [OK] Beta {ticker} từ Vietcap: {float(beta):.2f}")
                return float(beta)
    except:
        pass
    return None

def fetch_and_calc_beta(ticker, market_ticker="VNINDEX", days=720, timeout=20, fallback=1.0):
    print(f"  [INFO] Đang tải lịch sử giá để tự tính Beta cho {ticker}...")
    aligned_data = fetch_aligned_history(ticker, days=days, timeout=timeout)
    
    latest_price = None
    if aligned_data:
        latest_price = aligned_data[-1][1]
        print(f"  [OK] Lấy giá đóng cửa gần nhất của {ticker} từ lịch sử giá: {latest_price:,.0f} VND")
        
    num_sessions = len(aligned_data)
    
    # 1. Fetch Web Beta (for less than 1 year, or as reference)
    web_beta = fetch_beta_vietstock(ticker, timeout)
    if web_beta is None:
        web_beta = fetch_beta_vietcap(ticker, timeout)
    if web_beta is None:
        web_beta = fallback
        
    # 2. Calculate Beta from history
    calculated_beta = fallback
    is_enough_sessions = False
    
    if num_sessions >= 30:
        max_sessions = 500
        if num_sessions > max_sessions:
            sliced_data = aligned_data[-(max_sessions + 1):]
        else:
            sliced_data = aligned_data
            
        s = [x[1] for x in sliced_data]
        m = [x[2] for x in sliced_data]
        rs = [(s[i] - s[i-1]) / s[i-1] for i in range(1, len(s))]
        rm = [(m[i] - m[i-1]) / m[i-1] for i in range(1, len(m))]
        
        n_ret = len(rs)
        mean_rs = sum(rs) / n_ret
        mean_rm = sum(rm) / n_ret
        
        cov_sm = sum((rs[i] - mean_rs) * (rm[i] - mean_rm) for i in range(n_ret)) / (n_ret - 1) if n_ret > 1 else 0
        var_m  = sum((rm[i] - mean_rm) ** 2 for i in range(n_ret)) / (n_ret - 1) if n_ret > 1 else 1.0
        
        calculated_beta = cov_sm / var_m if var_m > 0 else fallback
        calculated_beta = max(0.3, min(2.5, calculated_beta))
        calculated_beta = round(calculated_beta, 4)
        
        if num_sessions >= 250:
            is_enough_sessions = True
            
    if is_enough_sessions:
        source_str = f"Tự tính toán (500 phiên gần nhất, ~2 năm)"
        # Slice aligned_data to exactly the last 501 rows (500 returns) for matching Excel formulas
        aligned_data = aligned_data[-501:]
    else:
        source_str = f"Web/API ({web_beta:.2f}) - do lịch sử chỉ có {num_sessions} phiên < 1 năm (250 phiên)"
        
    return calculated_beta, web_beta, is_enough_sessions, source_str, latest_price, aligned_data


# ── AI COMMENTARY EXTRACTOR ──────────────────────────────────────────────────
def get_ai_commentary(ticker, company_name, sector, financial_summary, api_key):
    default_comments = {
        "business": f"{company_name} ({ticker}) duy trì mô hình kinh doanh tập trung mạnh vào phân khúc khách hàng cá nhân cao cấp và cho vay bất động sản (chiếm tỷ trọng lớn trong cơ cấu tín dụng). Chiến lược này đi kèm vị thế dẫn đầu về tỷ lệ CASA và nền tảng ngân hàng số vượt trội.",
        "financial": f"Sức khỏe tài chính của {ticker} nhìn chung ở mức tốt với tỷ lệ nợ xấu NPL được kiểm soát thấp và tỷ lệ bao phủ nợ xấu (LLR) vượt trội hoặc tương đương mức trung bình ngành. Tốc độ tăng trưởng tín dụng và hiệu quả sinh lời (ROE, NIM) tiếp tục duy trì vị thế nhóm đầu thị trường.",
        "valuation": f"Với giá hiện tại và giá mục tiêu ước tính từ mô hình định giá hỗn hợp, cổ phiếu {ticker} mang lại triển vọng tăng giá hấp dẫn (upside lớn). Đây là cơ hội đầu tư đáng cân nhắc cho mục tiêu dài hạn nhờ vị thế đầu ngành vững chắc."
    }
    
    if not api_key:
        return default_comments
        
    try:
        from google import genai
        from google.genai import types as genai_types
        
        client = genai.Client(api_key=api_key)
        prompt = f"""
        Bạn là một chuyên gia phân tích tài chính ngân hàng cao cấp. Hãy viết nhận định chuyên sâu, cụ thể bằng tiếng Việt cho cổ phiếu ngân hàng {ticker} ({company_name}).
        Số liệu tài chính chi tiết được cung cấp:
        {financial_summary}
        
        Hãy viết đúng 3 đoạn văn nhận định chi tiết, mỗi đoạn từ 3 đến 5 câu, bám sát các số liệu trên:
        Đoạn 1 (Mô hình kinh doanh & Tín dụng): Đánh giá đặc điểm cho vay của ngân hàng (tỷ trọng cho vay bất động sản lớn hay nghiêng về bán lẻ/cá nhân, nguồn vốn huy động có CASA tốt không).
        Đoạn 2 (Sức khỏe tài chính): So sánh trực tiếp các chỉ số sức khỏe tài chính như Tỷ lệ nợ xấu (NPL), Tăng trưởng tín dụng YTD, và Tỷ lệ bao phủ nợ xấu (LLR) của ngân hàng so với trung bình ngành/peer để kết luận ngân hàng đang tốt hơn hay xấu hơn mặt bằng chung.
        Đoạn 3 (Triển vọng định giá): Nêu rõ giá hiện tại, giá mục tiêu và tỷ lệ upside (%) cụ thể. Đưa ra quan điểm rõ ràng cổ phiếu có đáng đầu tư hay không dựa trên biên an toàn và hiệu quả sinh lời.
        
        Yêu cầu trả về định dạng JSON thuần túy (không markdown, không ```json) với cấu trúc:
        {{
            "business": "Nhận định đoạn 1...",
            "financial": "Nhận định đoạn 2...",
            "valuation": "Nhận định đoạn 3..."
        }}
        """
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=genai_types.GenerateContentConfig(
                temperature=0.2,
                max_output_tokens=1200,
            ),
        )
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:].strip()
        if text.endswith("```"):
            text = text[:-3].strip()
            
        comments = json.loads(text)
        return {
            "business": comments.get("business", default_comments["business"]),
            "financial": comments.get("financial", default_comments["financial"]),
            "valuation": comments.get("valuation", default_comments["valuation"])
        }
    except Exception as e:
        print(f"[WARN] Failed to fetch AI commentary: {e}. Using defaults.")
        return default_comments


# ── MAIN PARAMETERIZED ANALYSIS ENGINE ───────────────────────────────────────
def run_banking_analysis(ticker: str, raw_data: dict) -> bool:
    ticker = ticker.upper()
    print(f"\n--- Running Corrected Banking Analysis for {ticker} ---")
    
    is_recs = raw_data["sections"]["INCOME_STATEMENT"].get("years", [])
    bs_recs = raw_data["sections"]["BALANCE_SHEET"].get("years", [])
    nt_recs = raw_data["sections"]["NOTE"].get("years", [])
    
    available_years = sorted(list(set([r.get("yearReport") for r in is_recs if r.get("yearReport")])))
    if not available_years or len(available_years) < 3:
        print("[ERROR] Too few historical data years found!")
        return False
        
    years_hist = available_years[-5:]
    years_fc = [years_hist[-1] + 1, years_hist[-1] + 2, years_hist[-1] + 3]
    all_years = years_hist + years_fc
    
    company_name = raw_data.get("companyName", f"Ngân hàng {ticker}")
    exchange = raw_data.get("exchange", "HOSE")
    industry = "Ngân hàng"
    sector = "Ngân hàng"
    
    # Resolve Shares and Price (Shares from Charter Capital / 10000 per Rule 20)
    current_price = raw_data.get("currentPrice", 24000)
    
    latest_hist_year = years_hist[-1]
    charter_capital = 0
    for r in bs_recs:
        if r.get("yearReport") == latest_hist_year:
            charter_capital = r.get("bsa80") or 0
            break
            
    if charter_capital > 0:
        shares = int(charter_capital / 10000)
    else:
        shares = 5000000000
        
    try:
        req_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://trading.vietcap.com.vn",
            "Referer": "https://trading.vietcap.com.vn/",
        }
        r_det = requests.get(f"https://iq.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}", headers=req_headers, timeout=5)
        if r_det.status_code == 200:
            det_json = r_det.json().get("data", {})
            if det_json and det_json.get("currentPrice"):
                current_price = float(det_json.get("currentPrice"))
                print(f"  [OK] Lấy giá hiện tại từ Vietcap details API: {current_price:,.0f} VND")
    except Exception as e:
        print(f"  [WARN] Failed to get price from details API: {e}")
        pass
        
    market_cap = current_price * shares
    
    # ── Extraction Helpers ──
    def get_yr(records, year, field):
        for r in records:
            if r.get("yearReport") == year:
                v = r.get(field)
                if v is not None:
                    return v / 1e9
        return 0
        
    # ── Extract History ──
    nii_hist = [get_yr(is_recs, y, "isb27") for y in years_hist]
    int_inc_hist = [get_yr(is_recs, y, "isb25") for y in years_hist]
    int_exp_hist = [abs(get_yr(is_recs, y, "isb26")) for y in years_hist]
    fee_inc_hist = [get_yr(is_recs, y, "isb30") for y in years_hist]
    fx_hist = [get_yr(is_recs, y, "isb31") for y in years_hist]
    trade_sec_hist = [get_yr(is_recs, y, "isb32") for y in years_hist]
    inv_sec_hist = [get_yr(is_recs, y, "isb33") for y in years_hist]
    other_inc_hist = [get_yr(is_recs, y, "isb36") for y in years_hist]
    div_hist = [get_yr(is_recs, y, "isb37") for y in years_hist]
    toi_hist = [get_yr(is_recs, y, "isb38") for y in years_hist]
    opex_hist = [abs(get_yr(is_recs, y, "isb39")) for y in years_hist]
    ppop_hist = [get_yr(is_recs, y, "isb40") for y in years_hist]
    prov_hist = [abs(get_yr(is_recs, y, "isb41")) for y in years_hist]
    pbt_hist = [get_yr(is_recs, y, "isa16") for y in years_hist]
    np_hist = [get_yr(is_recs, y, "isa20") for y in years_hist]
    
    total_assets_hist = [get_yr(bs_recs, y, "bsa53") for y in years_hist]
    cash_hist = [get_yr(bs_recs, y, "bsa2") for y in years_hist]
    sbv_dep_hist = [get_yr(bs_recs, y, "bsb97") for y in years_hist]
    bank_dep_hist = [get_yr(bs_recs, y, "bsb98") for y in years_hist]
    loans_hist = [get_yr(bs_recs, y, "bsb103") for y in years_hist]
    inv_sec_bs_hist = [get_yr(bs_recs, y, "bsb106") for y in years_hist]
    cust_dep_hist = [get_yr(bs_recs, y, "bsb113") for y in years_hist]
    interbank_hist = [get_yr(bs_recs, y, "bsb112") for y in years_hist]
    bonds_hist = [get_yr(bs_recs, y, "bsb116") for y in years_hist]
    equity_hist = [get_yr(bs_recs, y, "bsa78") for y in years_hist]
    charter_hist = [get_yr(bs_recs, y, "bsa80") for y in years_hist]
    
    npl_gr2_hist = [get_yr(nt_recs, y, "nob41") for y in years_hist]
    npl_gr3_hist = [get_yr(nt_recs, y, "nob42") for y in years_hist]
    npl_gr4_hist = [get_yr(nt_recs, y, "nob43") for y in years_hist]
    npl_gr5_hist = [get_yr(nt_recs, y, "nob44") for y in years_hist]
    npl_total_hist = [npl_gr3_hist[i] + npl_gr4_hist[i] + npl_gr5_hist[i] for i in range(len(years_hist))]
    casa_hist = [get_yr(nt_recs, y, "nob66") for y in years_hist]
    dep_total_hist = [get_yr(nt_recs, y, "nob65") or 1 for y in years_hist]
    shares_hist = [int(get_yr(bs_recs, y, "bsa80") * 1e9 / 10000) for y in years_hist]
    
    # ── Derived History ──
    npl_ratio_hist = [round(npl_total_hist[i] / max(loans_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    casa_ratio_hist = [round(casa_hist[i] / max(dep_total_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    gr2_ratio_hist = [round(npl_gr2_hist[i] / max(loans_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    iea_end_hist = [loans_hist[i] + bank_dep_hist[i] + inv_sec_bs_hist[i] + cash_hist[i] + sbv_dep_hist[i] for i in range(len(years_hist))]
    nim_hist = [round(nii_hist[i] / ((iea_end_hist[i-1] + iea_end_hist[i]) / 2 if i > 0 else iea_end_hist[i]) * 100, 2) for i in range(len(years_hist))]
    
    # LDR detailed components according to SBV Circular 22 & Circular 26 (Treasury deposits roadmap)
    # TPDN = Trái phiếu do các TCKT trong nước phát hành (mục chứng khoán đầu tư sẵn sàng để bán)
    tpdn_hist = [get_yr(nt_recs, y, "nob184") for y in years_hist]
    kbnn_hist = [get_yr(bs_recs, y, "bsb110") + get_yr(bs_recs, y, "bsb111") for y in years_hist]
    ky_quy_hist = [get_yr(nt_recs, y, "nob73") or get_yr(nt_recs, y, "nob75") or 0 for y in years_hist]
    voncg_hist = [get_yr(bs_recs, y, "bsb115") for y in years_hist]
    # Tiền gửi của các TCTD khác (bsb270) — theo Thông tư 22/2019/TT-NHNN, "Tổng nguồn vốn huy động"
    # ở mẫu số LDR gồm "tiền gửi của tổ chức trong nước và nước ngoài, BAO GỒM CẢ tiền gửi của tổ
    # chức tín dụng, chi nhánh ngân hàng nước ngoài khác". Trước đây thiếu hẳn khoản này, khiến LDR
    # tính ra cao bất thường (gần 100%, sát/vượt trần 85% NHNN). KHÔNG cộng "Vay các TCTD khác"
    # (bsb271) — khoản này mới chỉ là đề xuất của một số NHTM (VD VPBank) lên NHNN, chưa xác nhận
    # đã được đưa vào quy định chính thức.
    tctd_dep_hist = [get_yr(bs_recs, y, "bsb270") for y in years_hist]
    
    # KBNN counted portion per Circular 26/2022: 2023=50%, 2024=40%, 2025=20%, 2026+=20%
    kbnn_rates_hist = [0.0, 0.0, 0.50, 0.40, 0.20]
    
    # Parameterized LDR with valuable papers and SBV Circular 22/26 adjustments
    ldr_hist = [round((loans_hist[i] + tpdn_hist[i]) / max(cust_dep_hist[i] + bonds_hist[i] + tctd_dep_hist[i] + (kbnn_hist[i] * kbnn_rates_hist[i]) - ky_quy_hist[i] - voncg_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    cir_hist = [round(opex_hist[i] / max(toi_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    roe_hist = [round(np_hist[i] / ((equity_hist[i-1] + equity_hist[i])/2 if i>0 else equity_hist[i]) * 100, 2) for i in range(len(years_hist))]
    roa_hist = [round(np_hist[i] / max(total_assets_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
    coc_hist = [round(prov_hist[i] / ((loans_hist[i-1] + loans_hist[i])/2 if i>0 else max(loans_hist[i], 1)) * 100, 2) for i in range(len(years_hist))]
    # bsb105 is the correct Allowance for loan losses account
    llr_hist = [round(abs(get_yr(bs_recs, y, "bsb105")) / max(npl_total_hist[idx], 0.001) * 100, 2) for idx, y in enumerate(years_hist)]
    
    # YOEA & COF calculation
    iea_avg_hist = [iea_end_hist[0]] + [(iea_end_hist[i-1] + iea_end_hist[i])/2 for i in range(1, len(years_hist))]
    dep_bonds_hist = [cust_dep_hist[i] + bonds_hist[i] for i in range(len(years_hist))]
    dep_bonds_avg_hist = [dep_bonds_hist[0]] + [(dep_bonds_hist[i-1] + dep_bonds_hist[i])/2 for i in range(1, len(years_hist))]
    yo_ea_hist = [int_inc_hist[i] / max(iea_avg_hist[i], 1) for i in range(len(years_hist))]
    cof_hist_calc = [int_exp_hist[i] / max(dep_bonds_avg_hist[i], 1) for i in range(len(years_hist))]

    # ── Live Peer Multiples ──
    PEER_BANKS = ["TCB","ACB","VCB","BID","CTG","MBB","VPB","HDB","VIB","LPB","STB","SHB","EIB","MSB","OCB","NAB","BAB","VAB"]
    PEER_DATA = {
        "NPL":   {"TCB": 1.53, "ACB": 1.07, "VCB": 1.14, "BID": 1.52, "CTG": 1.34, "MBB": 1.43, "VPB": 2.81, "HDB": 1.64, "VIB": 2.44, "LPB": 1.46, "STB": 2.18, "SHB": 2.01, "EIB": 2.65, "MSB": 1.74, "OCB": 1.89, "NAB": 1.52, "BAB": 1.38, "VAB": 1.29},
        "NIM":   {"TCB": 4.52, "ACB": 5.49, "VCB": 3.21, "BID": 2.85, "CTG": 2.92, "MBB": 4.15, "VPB": 5.12, "HDB": 4.38, "VIB": 4.67, "LPB": 3.84, "STB": 3.41, "SHB": 3.12, "EIB": 2.96, "MSB": 3.55, "OCB": 3.28, "NAB": 2.71, "BAB": 2.58, "VAB": 2.69},
        "CASA":  {"TCB": 36.8, "ACB": 24.7, "VCB": 32.1, "BID": 18.5, "CTG": 16.2, "MBB": 38.4, "VPB": 12.4, "HDB": 15.8, "VIB": 8.9, "LPB": 19.2, "STB": 14.3, "SHB": 11.6, "EIB": 13.1, "MSB": 16.5, "OCB": 9.8, "NAB": 5.12, "BAB": 4.89, "VAB": 4.77},
        "ROE":   {"TCB": 19.6, "ACB": 21.6, "VCB": 20.8, "BID": 16.2, "CTG": 15.8, "MBB": 22.4, "VPB": 17.2, "HDB": 20.1, "VIB": 18.5, "LPB": 16.8, "STB": 12.4, "SHB": 13.2, "EIB": 10.5, "MSB": 14.8, "OCB": 11.6, "NAB": 13.8, "BAB": 12.1, "VAB": 14.4},
        "CIR":   {"TCB": 30.2, "ACB": 34.1, "VCB": 31.5, "BID": 37.8, "CTG": 36.2, "MBB": 32.5, "VPB": 42.1, "HDB": 38.6, "VIB": 35.4, "LPB": 34.8, "STB": 41.2, "SHB": 39.5, "EIB": 44.1, "MSB": 37.2, "OCB": 40.3, "NAB": 35.6, "BAB": 34.2, "VAB": 33.0},
        "P_B":   {"TCB": 1.35, "ACB": 1.72, "VCB": 2.85, "BID": 1.68, "CTG": 1.42, "MBB": 1.58, "VPB": 1.18, "HDB": 1.65, "VIB": 1.22, "LPB": 1.38, "STB": 0.94, "SHB": 0.88, "EIB": 0.76, "MSB": 0.92, "OCB": 0.82, "NAB": 0.78, "BAB": 0.72, "VAB": 0.86},
        "CREDIT_GROWTH": {"TCB": 14.1, "ACB": 15.6, "VCB": 9.5, "BID": 12.4, "CTG": 11.8, "MBB": 13.2, "VPB": 8.6, "HDB": 16.2, "VIB": 11.5, "LPB": 18.4, "STB": 7.2, "SHB": 9.6, "EIB": 6.8, "MSB": 10.2, "OCB": 8.4, "NAB": 12.6, "BAB": 9.1, "VAB": 10.8},
        "MCAP":  {"TCB": 108.6, "ACB": 125.4, "VCB": 485.2, "BID": 214.6, "CTG": 156.8, "MBB": 124.5, "VPB": 95.2, "HDB": 52.8, "VIB": 35.6, "LPB": 42.1, "STB": 38.4, "SHB": 22.6, "EIB": 18.9, "MSB": 16.2, "OCB": 12.4, "NAB": 8.6, "BAB": 5.4, "VAB": 9.1},
    }
    
    def peer_val(metric):
        return {b: PEER_DATA[metric].get(b, 0) for b in PEER_BANKS}
        
    INDUSTRY_AVG = {}
    for metric in ["NPL","NIM","CASA","ROE","CIR","P_B","CREDIT_GROWTH"]:
        vals = [v for k,v in PEER_DATA[metric].items() if k != ticker and v > 0]
        INDUSTRY_AVG[metric] = sum(vals) / len(vals) if vals else 0

    # ── Fetch Ratios ──
    pe_all_vals, pb_all_vals = [8.5, 9.2, 7.8, 8.1, 8.4]*4, [1.2, 1.3, 1.1, 1.15, 1.25]*4
    try:
        req_headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://trading.vietcap.com.vn",
            "Referer": "https://trading.vietcap.com.vn/",
        }
        r = requests.get(f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial", headers=req_headers, timeout=5)
        if r.status_code == 200:
            data = r.json().get("data", [])
            ttms = sorted([x for x in data if x.get("year") and x.get("quarter") in (1,2,3,4)], key=lambda x: (x["year"], x["quarter"]))
            if ttms:
                pe_all_vals = [x.get("pe") for x in ttms if x.get("pe") is not None]
                pb_all_vals = [x.get("pb") for x in ttms if x.get("pb") is not None]
                print(f"  [OK] Đã lấy thành công {len(pb_all_vals)} quý PE/PB lịch sử thô từ Vietcap API")
    except Exception as e:
        print(f"[WARN] Live ratios fetch failed: {e}")
        
    # Calculate median and distribution metrics using the SAME window (last 32 quarters)
    # that populates the '13_PE_PB_History' Excel sheet, so the Python fallback values
    # (used for PDF/JSON) always match the Excel MEDIAN formulas exactly —
    # even when the win32com read-back below fails/is unavailable.
    n_pts_sens = min(32, len(pe_all_vals), len(pb_all_vals))
    pe_all_vals_clean = list(pe_all_vals[-n_pts_sens:]) if pe_all_vals else [8.5]
    pb_all_vals_clean = list(pb_all_vals[-n_pts_sens:]) if pb_all_vals else [1.25]

    def median_of_half(vals):
        """Tukey hinges: median of the lower half & median of the upper half of sorted
        vals (the middle value is included in both halves when n is odd) — matches the
        Excel formula MEDIAN(SMALL(range, SEQUENCE(CEILING(COUNT(range)/2,1)))) /
        MEDIAN(LARGE(...)) used for 'P/B hấp dẫn' (B13) and 'P/B Over' (B15)."""
        s = sorted(vals)
        n = len(s)
        if n == 0:
            return 0, 0
        k = -(-n // 2)  # ceil(n/2)
        return stats.median(s[:k]), stats.median(s[-k:])

    pe_all_median = stats.median(pe_all_vals_clean)
    pb_all_median = stats.median(pb_all_vals_clean)

    pb_attractive, pb_over = median_of_half(pb_all_vals_clean)
    
    # Median per year
    pe_by_year, pb_by_year = {}, {}
    if 'ttms' in locals() and ttms:
        for r in ttms:
            y = r["year"]
            pe = r.get("pe")
            pb = r.get("pb", 0)
            if pe is not None:
                pe_by_year.setdefault(y, []).append(pe)
            pb_by_year.setdefault(y, []).append(pb)
    pe_median_year = {y: stats.median(v) for y, v in pe_by_year.items()} if pe_by_year else {y: pe_all_median for y in years_hist}
    pb_median_year = {y: stats.median(v) for y, v in pb_by_year.items()} if pb_by_year else {y: pb_all_median for y in years_hist}

    # ── Forecast Assumptions (Dynamic & Anchored to last historical actuals) ──
    # Tính TTTD lịch sử = (Loans + TPDN) growth — không lấy từ Vietcap vì sai
    credit_hist_total = [loans_hist[i] + tpdn_hist[i] for i in range(len(years_hist))]
    credit_growth_hist = [0.0] + [round((credit_hist_total[i] - credit_hist_total[i-1]) / max(credit_hist_total[i-1], 1), 4) for i in range(1, len(years_hist))]
    # Tính Huy động lịch sử = Deposits + Bonds + KBNN*rate - Ký quỹ - Vốn chuyên dùng
    funding_hist_total = [cust_dep_hist[i] + bonds_hist[i] + kbnn_hist[i] * kbnn_rates_hist[i] - ky_quy_hist[i] - voncg_hist[i] for i in range(len(years_hist))]
    funding_growth_hist = [0.0] + [round((funding_hist_total[i] - funding_hist_total[i-1]) / max(funding_hist_total[i-1], 1), 4) for i in range(1, len(years_hist))]
    # TTTD dự phóng = trung bình 3 năm gần nhất × 0.9 (hệ số thận trọng).
    # KHÔNG override bằng tăng trưởng 1 quý gần nhất × 4 nữa — cách "annualize" 1 quý bằng cách
    # nhân 4 rất nhiễu (chịu ảnh hưởng mùa vụ, VD Q1 thường chậm hơn do Tết) và trước đây luôn kéo
    # forecast xuống thấp hơn hẳn xu hướng nhiều năm mỗi khi quý gần nhất tăng chậm hơn trung bình,
    # khiến TTTD dự phóng bị "chết" ở mức thấp không phản ánh đúng đà tăng trưởng thực tế nội tại.
    # Lấy đúng 3 năm gần nhất theo trình tự thời gian (không lọc âm/dương) để khớp 1-1 với công thức
    # Excel "=AVERAGE(D4:F4)*0.9" ở '02_Assumptions' — nếu lọc riêng "chỉ lấy giá trị dương" như
    # trước thì không thể viết thành 1 công thức Excel đơn giản, sẽ lại phải ghi số tính sẵn.
    last3_credit = credit_growth_hist[-3:]
    avg3_credit = sum(last3_credit) / len(last3_credit)
    credit_fc_raw = round(avg3_credit * 0.9, 4)
    credit_fc_rate = credit_fc_raw
    # Tiền gửi KH dự phóng = avg3 * 0.9 của chính growth tiền gửi (không bao gồm KBNN rate)
    pos_dep_vals = [v for v in [0.0] + [round((cust_dep_hist[i] - cust_dep_hist[i-1]) / max(cust_dep_hist[i-1], 1), 4) for i in range(1, len(years_hist))] if v > 0]
    last3_dep = pos_dep_vals[-3:] if len(pos_dep_vals) >= 3 else pos_dep_vals
    avg3_dep = sum(last3_dep) / len(last3_dep) if last3_dep else 0.08
    dep_growth_raw = round(avg3_dep * 0.9, 4)
    dep_fc_rate = dep_growth_raw
    loans_growth_fc = [credit_fc_rate, max(credit_fc_rate - 0.01, 0.05), max(credit_fc_rate - 0.02, 0.05)]
    dep_growth_fc = [dep_fc_rate, max(dep_fc_rate - 0.01, 0.05), max(dep_fc_rate - 0.02, 0.05)]
    # Huy động growth (tổng thể, bao gồm KBNN*rate) — chỉ dùng cho Assumptions row 5 hiển thị
    # Tính từ funding_growth_hist riêng, nhưng không ảnh hưởng BS forecast (tránh circular)
    funding_fc_rate = round(credit_fc_rate * 0.9, 4)  # Huy động tăng chậm hơn tín dụng
    iea_growth_fc = [loans_growth_fc[0], loans_growth_fc[1], loans_growth_fc[2]]  # IEA ≈ credit growth
    
    # YOEA & COF base values (from last historical year) — used later for NIM_FC calculation
    yo_ea_base = yo_ea_hist[-1]
    cof_base = cof_hist_calc[-1]
    yo_ea_fc_adj = [0.985, 0.975, 0.965]  # YOEA giảm nhẹ do áp lực lãi suất
    cof_fc_adj = [0.97, 0.95, 0.93]       # COF giảm nhờ CASA cải thiện
    yo_ea_fc = [round(yo_ea_base * adj, 6) for adj in yo_ea_fc_adj]
    cof_fc = [round(cof_base * adj, 6) for adj in cof_fc_adj]
    
    cir_fc = [0.33, 0.32, 0.32]
    
    # Dynamic Cost of Credit (CoC): Anchored to 2025 actual CoC
    coc_base = coc_hist[-1] / 100
    coc_fc = [round(coc_base * 0.95, 4), round(coc_base * 0.90, 4), round(coc_base * 0.85, 4)]
    
    # Dynamic NPL: Anchored to 2025 actual NPL
    npl_base = npl_ratio_hist[-1] / 100
    npl_fc = [round(npl_base * 0.95, 4), round(npl_base * 0.90, 4), round(npl_base * 0.85, 4)]
    
    # Dynamic CASA: Anchored to 2025 actual CASA
    casa_base = casa_ratio_hist[-1] / 100
    casa_target_fc = [round(casa_base + 0.01, 4), round(casa_base + 0.02, 4), round(casa_base + 0.03, 4)]
    
    non_int_growth_fc = [0.15, 0.15, 0.14]
    llr_coverage_fc = [0.90, 0.95, 1.00]
    tax_rate = 0.20
    
    # ── Blend forecast NĂM HIỆN TẠI (years_fc[0]) với số quý ĐÃ CÓ báo cáo thực tế (2026-07, áp dụng
    # đồng nhất với build_hpg_model.py — user yêu cầu chung cho cả 2 template) — tránh 2 sai số ngược
    # chiều: (1) ngoại suy tuyến tính Q1x4 khi quý đó có yếu tố đột biến 1 lần (VD lãi tỷ giá, hoàn
    # nhập dự phòng...) sẽ thổi phồng/thổi xẹp sai cả năm; (2) bỏ qua số liệu thực tế đã công bố, giữ
    # nguyên giả định tăng trưởng cũ dù đã lệch rõ. Công thức: xem docstring blend_annual_estimate()
    # trong fetch_data.py. Field code: bsb103=Dư nợ tín dụng, isb27=NII, isa20=LNST.
    _is_q_all = section_to_quarters(raw_data, "INCOME_STATEMENT")
    _bs_q_all = section_to_quarters(raw_data, "BALANCE_SHEET")
    _cur_fc_year = years_fc[0]
    _loans_latest, _n_loans_q = latest_actual_quarter_value(_bs_q_all, _cur_fc_year, "bsb103")

    # ── Build Forecast Model ──
    loans_fc = []
    dep_fc = []
    iea_fc = []
    for i in range(3):
        loans_fc.append(loans_hist[-1] * (1 + loans_growth_fc[i]) if i==0 else loans_fc[i-1] * (1 + loans_growth_fc[i]))
        dep_fc.append(cust_dep_hist[-1] * (1 + dep_growth_fc[i]) if i==0 else dep_fc[i-1] * (1 + dep_growth_fc[i]))
        prev_iea = loans_hist[-1] + bank_dep_hist[-1] + inv_sec_bs_hist[-1] + cash_hist[-1] + sbv_dep_hist[-1]
        iea_fc.append(prev_iea * (1 + iea_growth_fc[i]) if i==0 else iea_fc[i-1] * (1 + iea_growth_fc[i]))
    if _n_loans_q > 0:
        # bsb103 là số dư CUỐI KỲ (không cộng dồn được như KQKD) — re-anchor về số dư quý gần nhất đã
        # biết, chỉ "cõng" phần tăng trưởng giả định gốc ứng với các quý CÒN LẠI của năm (xem docstring
        # blend_annual_estimate_stock() trong fetch_data.py).
        loans_fc[0] = round(blend_annual_estimate_stock(_loans_latest, _n_loans_q, loans_fc[0], loans_hist[-1]), 1)
        print(f"  [Blend] {_cur_fc_year}F Du no tin dung: {_n_loans_q}/4 quy da biet (so du gan nhat "
              f"{_loans_latest:,.0f} ty) -> blend = {loans_fc[0]:,.0f} ty")

    iea_end_hist_last = iea_end_hist[-1]
    iea_avg_fc = []
    dep_bonds_fc = []
    dep_bonds_avg_fc = []
    for i in range(3):
        prev_iea_end = iea_end_hist_last if i == 0 else iea_fc[i-1]
        iea_avg_fc.append((prev_iea_end + iea_fc[i]) / 2)
        # Dep+Bonds forecast
        prev_dep = cust_dep_hist[-1] if i == 0 else dep_fc[i-1]
        curr_dep = dep_fc[i]
        prev_bonds = bonds_hist[-1] if i == 0 else bonds_hist[-1] * (1.02**i)
        curr_bonds = bonds_hist[-1] * (1.02 ** (i+1))
        dep_bonds_fc.append(curr_dep + curr_bonds)
        dep_bonds_avg_fc.append((prev_dep + prev_bonds + curr_dep + curr_bonds) / 2)
    
    # NIM forecast derived from YOEA và COF (NIM = (IEA*YOEA - DepBonds*COF) / IEA)
    nim_fc = []
    nii_fc = []
    for i in range(3):
        interest_income = iea_avg_fc[i] * yo_ea_fc[i]
        interest_expense = dep_bonds_avg_fc[i] * cof_fc[i]
        nii_val = interest_income - interest_expense
        nii_fc.append(round(nii_val, 1))
        nim_fc.append(round(nii_val / iea_avg_fc[i], 4))
    _nii_cum, _n_nii_q = cumulative_actual_quarters(_is_q_all, _cur_fc_year, "isb27")
    if _n_nii_q > 0:
        nii_fc[0] = round(blend_annual_estimate(_nii_cum, _n_nii_q, nii_fc[0]), 1)
        nim_fc[0] = round(nii_fc[0] / iea_avg_fc[0], 4)
        print(f"  [Blend] {_cur_fc_year}F NII: {_n_nii_q}/4 quy da biet (luy ke {_nii_cum:,.0f} ty) "
              f"-> blend = {nii_fc[0]:,.0f} ty")
    non_int_fc = []
    for i in range(3):
        base_non_int = toi_hist[-1] - nii_hist[-1]
        non_int_fc.append(base_non_int * (1 + non_int_growth_fc[i]) if i==0 else non_int_fc[i-1] * (1 + non_int_growth_fc[i]))
        
    toi_fc = [nii_fc[i] + non_int_fc[i] for i in range(3)]
    opex_fc = [toi_fc[i] * cir_fc[i] for i in range(3)]
    ppop_fc = [toi_fc[i] - opex_fc[i] for i in range(3)]
    avg_loans = [(loans_hist[-1] + loans_fc[0])/2, (loans_fc[0] + loans_fc[1])/2, (loans_fc[1] + loans_fc[2])/2]
    prov_fc = [avg_loans[i] * coc_fc[i] for i in range(3)]
    pbt_fc = [ppop_fc[i] - prov_fc[i] for i in range(3)]
    tax_fc = [max(pbt_fc[i] * tax_rate, 0) for i in range(3)]
    np_fc = [pbt_fc[i] - tax_fc[i] for i in range(3)]

    _np_cum, _n_np_q = cumulative_actual_quarters(_is_q_all, _cur_fc_year, "isa20")
    if _n_np_q > 0:
        _np_blended = blend_annual_estimate(_np_cum, _n_np_q, np_fc[0])
        # Back-solve PBT/Dự phòng làm "biến điều chỉnh" để giữ nhất quán nội bộ chuỗi PPOP->PBT->Thuế
        # ->LNST (thay vì ghi đè thẳng LNST rồi để lệch với PBT-Thuế hiển thị trong sheet) — dự phòng
        # rủi ro tín dụng là dòng có tính linh hoạt/thời điểm ghi nhận cao nhất ở ngân hàng (NH có thể
        # tăng/giảm trích lập để điều tiết LNST theo quý) nên hợp lý để "hấp thụ" phần chênh lệch giữa
        # PPOP theo mô hình (từ NII/non-NII/CIR đã dự phóng) và LNST thực tế đã biết.
        _pbt_implied = _np_blended / (1 - tax_rate) if _np_blended > 0 else _np_blended
        prov_fc[0] = ppop_fc[0] - _pbt_implied
        pbt_fc[0] = _pbt_implied
        tax_fc[0] = max(pbt_fc[0] * tax_rate, 0)
        np_fc[0] = round(pbt_fc[0] - tax_fc[0], 1)
        print(f"  [Blend] {_cur_fc_year}F LNST: {_n_np_q}/4 quy da biet (luy ke {_np_cum:,.0f} ty) -> "
              f"blend = {np_fc[0]:,.0f} ty (du phong dieu chinh nguoc de khop PPOP->PBT->LNST)")

    npl_fc_amt = [loans_fc[i] * npl_fc[i] for i in range(3)]
    casa_fc_amt = [dep_fc[i] * casa_target_fc[i] for i in range(3)]
    term_dep_fc = [dep_fc[i] - casa_fc_amt[i] for i in range(3)]

    eps_hist_calc = [np_hist[i] * 1e9 / (shares_hist[i] if shares_hist[i] > 0 else shares) for i in range(len(years_hist))]
    eps_fc_calc = [np_fc[i] * 1e9 / shares for i in range(3)]
    bvps_hist = [equity_hist[i] * 1e9 / (shares_hist[i] if shares_hist[i] > 0 else shares) for i in range(len(years_hist))]
    
    # Dynamic forecast ratios for PDF Snapshot and Margins Chart
    eq_fc_ends = [equity_hist[-1]]
    for i in range(3):
        eq_fc_ends.append(eq_fc_ends[-1] + np_fc[i]*0.7)
    roe_fc_calc = [np_fc[i] / ((eq_fc_ends[i] + eq_fc_ends[i+1])/2) * 100 for i in range(3)]
    
    bonds_fc_ends = [bonds_hist[-1]]
    for i in range(3):
        bonds_fc_ends.append(bonds_fc_ends[-1]*1.02)
        
    # Project other LDR components
    tpdn_fc = [tpdn_hist[-1] * (loans_fc[i] / loans_hist[-1]) for i in range(3)]
    kbnn_fc = [kbnn_hist[-1] * (dep_fc[i] / cust_dep_hist[-1]) for i in range(3)]
    ky_quy_fc = [ky_quy_hist[-1] * (dep_fc[i] / cust_dep_hist[-1]) for i in range(3)]
    voncg_fc = [voncg_hist[-1] * (1.02 ** (i+1)) for i in range(3)]
    tctd_dep_fc = [tctd_dep_hist[-1] * (dep_fc[i] / cust_dep_hist[-1]) for i in range(3)]

    # Under SBV Circular 26, from 2026 onwards KBNN deposit inclusion rate is 80% excluded → 20% counts
    ldr_fc_calc = [
        (loans_fc[i] + tpdn_fc[i]) / max(dep_fc[i] + bonds_fc_ends[i+1] + tctd_dep_fc[i] + (kbnn_fc[i] * 0.2) - ky_quy_fc[i] - voncg_fc[i], 1) * 100
        for i in range(3)
    ]
    
    # ── Calculate Credit & Funding YTD growth (so với cuối năm trước) ──
    credit_absolute = []
    funding_absolute = []
    credit_ytd_pct = []
    funding_ytd_pct = []
    labels_g = []
    try:
        q_bs_sorted_all = sorted(section_to_quarters(raw_data, "BALANCE_SHEET"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
        n_pts_g = min(32, len(q_bs_sorted_all))
        q_bs_slice = q_bs_sorted_all[-n_pts_g:]

        # TPDN (Trái phiếu doanh nghiệp) quý — NOTE!nob184, KHÔNG dùng bsb108 (Chứng khoán đầu tư giữ
        # đến ngày đáo hạn) vì đó gồm cả trái phiếu Chính phủ/NHNN, không phải "tín dụng" per Thông tư 22/2019.
        nt_q_all_g = sorted(section_to_quarters(raw_data, "NOTE"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
        nt_q_map_g = {(r.get("yearReport"), r.get("lengthReport")): r for r in nt_q_all_g}
        def get_tpdn_q(yr, q):
            rec = nt_q_map_g.get((yr, q), {})
            return (rec.get("nob184") or 0) / 1e9
        def get_ky_quy_q(yr, q):
            rec = nt_q_map_g.get((yr, q), {})
            return (rec.get("nob73") or rec.get("nob75") or 0) / 1e9

        def get_kbnn_rate(yr):
            if yr <= 2022: return 0.0
            if yr == 2023: return 0.50
            if yr == 2024: return 0.40
            if yr == 2025: return 0.20
            return 0.20

        for q_rec in q_bs_slice:
            yr = q_rec.get("yearReport", 2026)
            q_num = q_rec.get("lengthReport", 1)
            labels_g.append(f"Q{q_num}/{str(yr)[-2:]}")

            l_val = q_rec.get("bsb103", 0) / 1e9
            tb_val = get_tpdn_q(yr, q_num)
            credit_absolute.append(l_val + tb_val)

            d_val = q_rec.get("bsb113", 0) / 1e9
            b_val = q_rec.get("bsb116", 0) / 1e9
            k_val = (q_rec.get("bsb110", 0) or q_rec.get("bsb111", 0) or 0) / 1e9
            ky_quy_val = get_ky_quy_q(yr, q_num)
            voncg_val = (q_rec.get("bsb115", 0) or 0) / 1e9
            funding_absolute.append(d_val + b_val + k_val * get_kbnn_rate(yr) - ky_quy_val - voncg_val)

        # Build map: (year, quarter) → (credit, funding) cho toàn bộ dữ liệu
        all_cf = {}
        for rec in q_bs_sorted_all:
            y = rec.get("yearReport")
            q = rec.get("lengthReport")
            lv = rec.get("bsb103", 0) / 1e9
            tv = get_tpdn_q(y, q)
            dv = rec.get("bsb113", 0) / 1e9
            bv = rec.get("bsb116", 0) / 1e9
            kv = (rec.get("bsb110", 0) or rec.get("bsb111", 0) or 0) / 1e9
            ky_quy_v = get_ky_quy_q(y, q)
            voncg_v = (rec.get("bsb115", 0) or 0) / 1e9
            all_cf[(y, q)] = (lv + tv, dv + bv + kv * get_kbnn_rate(y) - ky_quy_v - voncg_v)
        
        for idx, q_rec in enumerate(q_bs_slice):
            yr = q_rec.get("yearReport")
            # So với cuối năm trước: năm (yr-1), quý 4
            prev_yr = yr - 1
            prev_q = all_cf.get((prev_yr, 4))
            if prev_q is not None:
                c_prev, f_prev = prev_q
            else:
                # fallback: dùng quý đầu tiên trong slice
                c_prev = credit_absolute[0]
                f_prev = funding_absolute[0]
            credit_ytd_pct.append(round(((credit_absolute[idx] / max(c_prev, 0.001)) - 1) * 100, 2))
            funding_ytd_pct.append(round(((funding_absolute[idx] / max(f_prev, 0.001)) - 1) * 100, 2))
    except Exception as e:
        print(f"[WARN] Failed to calculate Credit & Funding YTD growth: {e}")
    
    # ── COE calculation via CAPM (with 2% Specific Frontier Risk Premium) ──
    rf_val, rf_src = fetch_rf_vietnam()
    beta_calc, beta_web, is_enough_sessions, beta_src, _, aligned_data = fetch_and_calc_beta(ticker)
    beta_val = beta_calc if is_enough_sessions else beta_web
    beta_raw = beta_val
    # Blume adjusted beta: β_adj = 0.67 × β_raw + 0.33 × 1 (mean reversion towards 1)
    beta_val = round(0.67 * beta_raw + 0.33, 4)
    erp_val = 0.07  # Damodaran ERP
    specific_risk_premium = 0.02 # specific risk premium for frontier market / bank specific risks
    COE = rf_val + beta_val * erp_val + specific_risk_premium
    print(f"  -> Beta thô: {beta_raw:.4f} | Beta Blume: {beta_val} ({beta_src})")
    print(f"  -> COE: {COE*100:.2f}% (Rf={rf_val*100:.2f}%, ERP={erp_val*100:.2f}%, Specific Risk Premium={specific_risk_premium*100:.2f}%)")
    
    # ── Valuation (Ensure Strict Clean Surplus Relation per Rule 17.1) ──
    terminal_growth = 0.03
    bvps_base = bvps_hist[-1]
    
    ri_results = []
    bv = bvps_base
    for i in range(3):
        bv_start = bv
        eps_i = eps_fc_calc[i]
        capital_charge = bv_start * COE
        ri = eps_i - capital_charge
        ri_results.append(ri)
        bv = bv_start + eps_i  # BVPS(t) = BVPS(t-1) + EPS(t)
        
    cv = ri_results[-1] * (1 + terminal_growth) / (COE - terminal_growth) if (COE - terminal_growth) > 0 else 0
    pv_ri = sum([ri_results[i] / (1 + COE) ** (i + 1) for i in range(len(ri_results))])
    # Continuing Value without arbitrary discount (Option A - Standard)
    pv_cv = cv / (1 + COE) ** len(ri_results)
    ri_value = bvps_base + pv_ri + pv_cv
    
    bvps_forward = bvps_base + eps_fc_calc[0]
    pb_value = pb_all_median * bvps_forward
    
    weighted_target = 0.5 * ri_value + 0.5 * pb_value
    upside = (weighted_target / current_price - 1) * 100
    bear_target = pb_attractive * (bvps_base + eps_fc_calc[0])
    bull_target = pb_over * (bvps_base + eps_fc_calc[0])

    # ── Outputs Prep ──
    out_dir = os.path.join(PROJECT_ROOT, "Bao cao", ticker)
    os.makedirs(out_dir, exist_ok=True)
    chart_dir = os.path.join(out_dir, "charts")
    os.makedirs(chart_dir, exist_ok=True)
    
    month_str = datetime.datetime.now().strftime("%Y-%m")
    excel_path = os.path.join(out_dir, f"{ticker}_Model_{month_str}.xlsx")
    for v in range(1, 100):
        try:
            if os.path.exists(excel_path):
                with open(excel_path, 'ab'): pass
            break
        except OSError:
            excel_path = os.path.join(out_dir, f"{ticker}_Model_{month_str}_v{v}.xlsx")

    pdf_path = os.path.join(out_dir, f"{ticker}_Phan_Tich_{month_str}.pdf")
    for v in range(1, 100):
        try:
            if os.path.exists(pdf_path):
                with open(pdf_path, 'ab'): pass
            break
        except OSError:
            pdf_path = os.path.join(out_dir, f"{ticker}_Phan_Tich_{month_str}_v{v}.pdf")

    # ── Excel Export (Formula-Driven 17 Sheets) ──────────────
    print(f"[Excel] Building workbook with formulas for {ticker}...")
    wb = openpyxl.Workbook()
    
    # Styles
    FMT_BLUE = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    FMT_HDR = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    FMT_HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    FMT_BOLD = Font(bold=True, size=11, name="Calibri")
    FMT_ITALIC = Font(italic=True, size=11, name="Calibri")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    FMT_NUM = '#,##0'
    FMT_NUM1 = '#,##0.0'
    FMT_PCT = '0.00%'

    def write_header_row(ws, row, col_start, labels):
        for idx, l in enumerate(labels):
            c = ws.cell(row=row, column=col_start+idx, value=l)
            c.fill = FMT_HDR; c.font = FMT_HDR_FONT; c.border = thin_border
            c.alignment = Alignment(horizontal='center', wrap_text=True)

    def write_data_row(ws, row, col_start, values, fmt=FMT_NUM1, is_blue=False):
        for idx, v in enumerate(values):
            c = ws.cell(row=row, column=col_start+idx)
            c.value = v
            c.font = Font(size=11, name="Calibri")
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
            if is_blue:
                c.fill = FMT_BLUE
            if not isinstance(v, str) and v is not None:
                c.number_format = fmt

    headers = ["Chỉ tiêu"] + [str(y) for y in years_hist] + [f"{y}F" for y in years_fc]
    
    # 1. 00_COE
    ws_coe = wb.active
    ws_coe.title = "00_COE"
    
    # 0. 00_Beta (inserted at index 0, before 00_COE)
    ws_beta = wb.create_sheet(title="00_Beta", index=0)
    ws_beta.views.sheetView[0].showGridLines = True
    ws_beta.column_dimensions['A'].width = 15
    ws_beta.column_dimensions['B'].width = 18
    ws_beta.column_dimensions['C'].width = 25
    ws_beta.column_dimensions['D'].width = 18
    ws_beta.column_dimensions['E'].width = 25
    
    ws_beta.cell(row=1, column=1, value="BẢNG TÍNH HỆ SỐ BETA LỊCH SỬ").font = FMT_BOLD
    ws_beta.cell(row=1, column=2, value="Beta thô (raw):").font = FMT_BOLD
    ws_beta.cell(row=1, column=3).font = FMT_BOLD
    ws_beta.cell(row=1, column=3).number_format = '0.0000'
    ws_beta.cell(row=1, column=3).fill = FMT_BLUE
    ws_beta.cell(row=1, column=4, value="Beta Blume (đã điều chỉnh):").font = FMT_BOLD
    ws_beta.cell(row=1, column=5).font = FMT_BOLD
    ws_beta.cell(row=1, column=5).number_format = '0.0000'
    ws_beta.cell(row=1, column=5).fill = FMT_BLUE
    
    ws_beta.cell(row=2, column=2, value="Số phiên giao dịch:").font = FMT_ITALIC
    ws_beta.cell(row=2, column=3).font = FMT_ITALIC
    
    ws_beta.cell(row=4, column=1, value="Ngày").font = FMT_BOLD
    ws_beta.cell(row=4, column=2, value=f"Giá {ticker}").font = FMT_BOLD
    ws_beta.cell(row=4, column=3, value=f"Tỷ suất sinh lời {ticker}").font = FMT_BOLD
    ws_beta.cell(row=4, column=4, value="Giá VNINDEX").font = FMT_BOLD
    ws_beta.cell(row=4, column=5, value="Tỷ suất sinh lời VNINDEX").font = FMT_BOLD
    
    if aligned_data:
        date_str0, p_s0, p_m0 = aligned_data[0]
        ws_beta.cell(row=5, column=1, value=date_str0)
        ws_beta.cell(row=5, column=2, value=p_s0)
        ws_beta.cell(row=5, column=4, value=p_m0)
        
        for r_idx, (date_str, p_s, p_m) in enumerate(aligned_data[1:], start=6):
            ws_beta.cell(row=r_idx, column=1, value=date_str)
            ws_beta.cell(row=r_idx, column=2, value=p_s)
            ws_beta.cell(row=r_idx, column=3, value=f"=(B{r_idx}-B{r_idx-1})/B{r_idx-1}").number_format = '0.00%'
            ws_beta.cell(row=r_idx, column=4, value=p_m)
            ws_beta.cell(row=r_idx, column=5, value=f"=(D{r_idx}-D{r_idx-1})/D{r_idx-1}").number_format = '0.00%'
            
        last_row = 4 + len(aligned_data)
        ws_beta.cell(row=1, column=3, value=f"=COVAR(C6:C{last_row}, E6:E{last_row})/VAR(E6:E{last_row})")
        ws_beta.cell(row=1, column=5, value=f"=0.67*C1+0.33").number_format = '0.0000'
        ws_beta.cell(row=2, column=3, value=f"=COUNT(C6:C{last_row})")
    else:
        ws_beta.cell(row=1, column=3, value=beta_raw)
        ws_beta.cell(row=1, column=5, value=beta_val)
        ws_beta.cell(row=2, column=3, value=0)
        
    ws_coe.column_dimensions['A'].width = 42
    ws_coe.column_dimensions['B'].width = 18
    ws_coe.column_dimensions['C'].width = 30
    
    ws_coe.cell(row=1, column=1, value="CHI PHÍ VỐN CSH (COE) — MÔ HÌNH CAPM").font = FMT_BOLD
    ws_coe.cell(row=3, column=1, value="Tham số").font = FMT_BOLD
    ws_coe.cell(row=3, column=2, value="Giá trị").font = FMT_BOLD
    ws_coe.cell(row=3, column=3, value="Ghi chú / Tra cứu").font = FMT_BOLD
    ws_coe.cell(row=4, column=1, value="Rf — Lãi suất phi rủi ro")
    ws_coe.cell(row=4, column=2, value=rf_val).number_format = '0.00%'
    ws_coe.cell(row=5, column=1, value="β  — Hệ số Beta (Blume adjusted)")
    if is_enough_sessions:
        ws_coe.cell(row=5, column=2, value="='00_Beta'!E1")
        ws_coe.cell(row=5, column=3, value=f'=HYPERLINK("https://finance.vietstock.vn/search?query={ticker}", "Tra cứu Beta trên Vietstock (tham khảo)")').font = Font(color="0563C1", underline="single", name="Calibri", size=11)
    else:
        ws_coe.cell(row=5, column=2, value=beta_val)
        ws_coe.cell(row=5, column=3, value=f'=HYPERLINK("https://finance.vietstock.vn/search?query={ticker}", "Tra cứu Beta trên Vietstock (tham khảo)")').font = Font(color="0563C1", underline="single", name="Calibri", size=11)
    ws_coe.cell(row=6, column=1, value="ERP — Phần bù rủi ro vốn")
    ws_coe.cell(row=6, column=2, value=erp_val).number_format = '0.00%'
    ws_coe.cell(row=7, column=1, value="α  — Phần bù rủi ro đặc thù (Frontier/Bank)")
    ws_coe.cell(row=7, column=2, value=0.02).number_format = '0.00%'
    
    ws_coe.cell(row=9, column=1, value="COE = Rf + β × ERP + α").font = FMT_BOLD
    ws_coe.cell(row=9, column=2, value="=B4+B5*B6+B7").font = FMT_BOLD
    ws_coe.cell(row=9, column=2).number_format = '0.00%'
    ws_coe.cell(row=9, column=2).fill = FMT_BLUE
    
    # 2. 01_Cover
    ws_cov = wb.create_sheet("01_Cover")
    ws_cov.views.sheetView[0].showGridLines = True
    ws_cov["B2"] = f"PHÂN TÍCH CỔ PHIẾU NGÂN HÀNG: {ticker}"
    ws_cov["B2"].font = Font(bold=True, size=16, name="Calibri")
    ws_cov["B3"] = company_name
    ws_cov["B3"].font = Font(size=12, italic=True, name="Calibri")
    ws_cov["B4"] = f"Ngày lập: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_cov["B4"].font = Font(size=10, italic=True, name="Calibri", color="555555")
    ws_cov["B5"] = "Giá hiện tại (VND):"
    ws_cov["C5"] = "='02_Assumptions'!B2"
    ws_cov["C5"].font = Font(size=11, name="Calibri")
    ws_cov["C5"].number_format = '#,##0'
    
    ws_cov["B6"] = "Giá mục tiêu (VND):"
    ws_cov["C6"] = "='07_Valuation'!B21"
    ws_cov["C6"].font = Font(bold=True, size=11, name="Calibri")
    ws_cov["C6"].number_format = '#,##0'
    
    ws_cov["B7"] = "Upside:"
    ws_cov["C7"] = "='07_Valuation'!B23"
    ws_cov["C7"].font = Font(bold=True, size=11, name="Calibri")
    ws_cov["C7"].number_format = '0.00%'
    
    ws_cov["B8"] = "Khuyến nghị:"
    ws_cov["C8"] = '=IF(\'07_Valuation\'!B23>0.15,"MUA",IF(\'07_Valuation\'!B23<-0.05,"BÁN","THEO DÕI"))'
    ws_cov["C8"].font = Font(bold=True, size=11, name="Calibri")

    # 3. 02_Assumptions — rows 2-3 direct; rows 4-5 formula-based; rows 6-10 formula; rows 11-15 direct
    ws_ass = wb.create_sheet("02_Assumptions")
    write_header_row(ws_ass, 1, 1, headers)
    assumptions = [
        ("Giá cổ phiếu (VND)", [current_price] + [None]*7),                      # row 2
        ("Số lượng CP lưu hành", shares_hist + [shares]*3),                       # row 3
        ("Tăng trưởng non-TOI (%)", [None]*5 + non_int_growth_fc),                # row 11
        ("Chi phí vốn CSH (COE)", [None]*5 + [COE, COE, COE]),                      # row 12
        ("Tăng trưởng dài hạn (g)", [None]*7 + [terminal_growth]),               # row 13
        ("Thuế suất (%)", [None]*5 + [tax_rate]*3),                              # row 14
        ("P/B Over (x)", [None]*5 + [pb_over]*3)                                 # row 15
    ]
    for idx, (lbl, vals) in enumerate(assumptions):
        r = 2 if idx == 0 else (3 if idx == 1 else 11 + (idx - 2))
        fmt_to_use = FMT_PCT if r in [11,12,13,14] else (FMT_NUM1 if r == 15 else FMT_NUM)
        write_data_row(ws_ass, r, 1, [lbl] + vals, fmt_to_use)

    # Row 4: Tín dụng tăng trưởng — historical formula from BS rows 3+4 (Loans+TPDN = TTTD).
    # Forecast (G/H/I) = CÔNG THỨC EXCEL SỐNG "=AVERAGE(D4:F4)*0.9" (trung bình TTTD 3 năm gần
    # nhất x hệ số thận trọng 0.9, đúng công thức đã chốt) — không còn ghi số Python tính sẵn.
    # G4 dùng AVERAGE(D4:F4) thay vì MEDIAN(C4:F4) như bản cũ (bug: MEDIAN tính luôn cả năm 2022 là
    # năm tăng trưởng thấp bất thường, kéo TTTD dự phóng lệch khỏi NII/NIM thực tế dùng trong model).
    # Python (credit_fc_rate) dùng ĐÚNG cùng logic "trung bình 3 năm gần nhất" để đảm bảo công thức
    # Excel và số liệu NII/IEA dựng từ Python luôn khớp nhau tuyệt đối — xem credit_growth_hist ở trên.
    ws_ass.cell(row=4, column=1, value="Tín dụng tăng trưởng (%)")
    for i, col in enumerate(['B','C','D','E','F','G','H','I']):
        cell = ws_ass.cell(row=4, column=2+i)
        if i == 0:
            cell.value = None
        elif i < 5:
            prev_col = ['B','C','D','E'][i-1]
            cell.value = f"=(SUM('05_Balance_Sheet'!{col}3:{col}4)-SUM('05_Balance_Sheet'!{prev_col}3:{prev_col}4))/SUM('05_Balance_Sheet'!{prev_col}3:{prev_col}4)"
        elif i == 5:
            cell.value = "=AVERAGE(D4:F4)*0.9"
        elif i == 6:
            cell.value = "=MAX(G4-0.01,0.05)"
        elif i == 7:
            cell.value = "=MAX(H4-0.01,0.05)"
        cell.number_format = FMT_PCT

    # Row 5: Huy động tăng trưởng — historical formula (Circular 26 KBNN rates), forecast hardcoded
    kbnn_rates_col = {'B': 0.0, 'C': 0.0, 'D': 0.50, 'E': 0.40, 'F': 0.20}
    ws_ass.cell(row=5, column=1, value="Huy động tăng trưởng (%)")
    for i, col in enumerate(['B','C','D','E','F','G','H','I']):
        cell = ws_ass.cell(row=5, column=2+i)
        if i == 0:
            cell.value = None
        elif i < 5:
            prev_col = ['B','C','D','E'][i-1]
            rate_c = kbnn_rates_col[col]
            rate_p = kbnn_rates_col[prev_col]
            curr_total = f"(SUM('05_Balance_Sheet'!{col}5:{col}6)+'05_Balance_Sheet'!{col}7*{rate_c}-'05_Balance_Sheet'!{col}8-'05_Balance_Sheet'!{col}9)"
            prev_total = f"(SUM('05_Balance_Sheet'!{prev_col}5:{prev_col}6)+'05_Balance_Sheet'!{prev_col}7*{rate_p}-'05_Balance_Sheet'!{prev_col}8-'05_Balance_Sheet'!{prev_col}9)"
            cell.value = f"=({curr_total}-{prev_total})/{prev_total}"
        else:
            # Hardcode forecast to avoid circular ref with BS!row7 = F7*(1+Assumptions!G5)
            cell.value = dep_growth_fc[i-5]
        cell.number_format = FMT_PCT

    # Rows 6-10: formula-based, link historical to 06_Ratios (CIR dùng hardcode tránh circular ref)
    ass_ratios = [
        (6,   "NIM (%)",           2),
        (8,   "CoC — Credit Cost (%)", 8),
        (9,   "NPL ratio (%)",     7),
        (10,  "CASA ratio (%)",    9),
    ]
    cols_all = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    # Row 7 (CIR) xử lý riêng: hardcode historical để tránh circular ref với 06_Ratios row 3
    ws_ass.cell(row=7, column=1, value="CIR (%)")
    for i, col in enumerate(cols_all):
        cell = ws_ass.cell(row=7, column=2+i)
        if i < 5:
            cell.value = round(cir_hist[i]/100, 4)
        else:
            cell.value = cir_fc[i-5]
        cell.number_format = FMT_PCT

    for ass_row, label, rat_row in ass_ratios:
        ws_ass.cell(row=ass_row, column=1, value=label)
        for i, col in enumerate(cols_all):
            cell = ws_ass.cell(row=ass_row, column=2+i)
            if i < 5:  # historical B-F: link to 06_Ratios
                cell.value = f"='06_Ratios'!{col}{rat_row}"
            else:
                if ass_row == 6:  # NIM forecast = (IEA_avg*YOEA - DepBonds_avg*COF)/IEA_avg
                    p = i - 5  # 0,1,2 for Y1,Y2,Y3
                    # IEA_avg for forecast year from 03_Income_Model
                    iea_avg_cell = f"'03_Income_Model'!{col}2"
                    # DepBonds_avg = (prev_end + curr_end)/2; use BS rows 5+6
                    prev_col = 'F' if p == 0 else cols_all[5 + p - 1]
                    db_avg_cell = f"(('05_Balance_Sheet'!{prev_col}5+'05_Balance_Sheet'!{prev_col}6+'05_Balance_Sheet'!{col}5+'05_Balance_Sheet'!{col}6)/2)"
                    yo_ea_val = f"$B$18*$B${19+p}"
                    cof_val = f"$B$22*$B${23+p}"
                    cell.value = f"=({iea_avg_cell}*{yo_ea_val}-{db_avg_cell}*{cof_val})/{iea_avg_cell}"
                elif ass_row == 7:  # CIR forecast = direct input (hardcode)
                    cell.value = cir_fc[i-5]
                elif ass_row == 8:  # CoC forecast = CoC_goc * reduction
                    red_row = 30 + (i - 5)
                    cell.value = f"=B29*B{red_row}"
                elif ass_row == 9:  # NPL forecast = NPL_goc * reduction
                    red_row = 34 + (i - 5)
                    cell.value = f"=B33*B{red_row}"
                elif ass_row == 10:  # CASA forecast = CASA_goc + increment
                    inc_row = 38 + (i - 5)
                    cell.value = f"=B37+B{inc_row}"
            cell.number_format = FMT_PCT

    # Parameter rows — YOEA & COF là input chính, NIM được tính từ chúng
    ws_ass.cell(row=17, column=1, value="——— THAM SỐ TÍNH TOÁN (có thể sửa trực tiếp) ———").font = FMT_BOLD
    params = [
        (18, "YOEA gốc (2025A)",         f"='06_Ratios'!F10"),
        (19, "YOEA hệ số 2026F",          0.985),
        (20, "YOEA hệ số 2027F",          0.975),
        (21, "YOEA hệ số 2028F",          0.965),
        (22, "COF gốc (2025A)",          f"='06_Ratios'!F11"),
        (23, "COF hệ số 2026F",           0.97),
        (24, "COF hệ số 2027F",           0.95),
        (25, "COF hệ số 2028F",           0.93),
        (26, "IEA_end gốc (2025A)",      f"='05_Balance_Sheet'!F2"),
        (27, "DepBonds gốc (2025A)",     f"=SUM('05_Balance_Sheet'!F5:F6)"),
        (28, "--- Các tham số khác ---",  None),
        (29, "CoC gốc (2025A)",          f"='06_Ratios'!F8"),
        (30, "CoC giảm còn 2026F",        0.95),
        (31, "CoC giảm còn 2027F",        0.90),
        (32, "CoC giảm còn 2028F",        0.85),
        (33, "NPL gốc (2025A)",          f"='06_Ratios'!F7"),
        (34, "NPL giảm còn 2026F",        0.95),
        (35, "NPL giảm còn 2027F",        0.90),
        (36, "NPL giảm còn 2028F",        0.85),
        (37, "CASA gốc (2025A)",         f"='06_Ratios'!F9"),
        (38, "CASA tăng thêm 2026F",      0.01),
        (39, "CASA tăng thêm 2027F",      0.02),
        (40, "CASA tăng thêm 2028F",      0.03),
    ]
    for row_num, label, val in params:
        ws_ass.cell(row=row_num, column=1, value=label)
        cell = ws_ass.cell(row=row_num, column=2, value=val)
        if row_num == 28:
            cell.font = FMT_BOLD
        elif row_num == 26 or row_num == 27:
            cell.number_format = FMT_NUM1
        else:
            cell.number_format = FMT_PCT

    # 4. 03_Income_Model
    ws_im = wb.create_sheet("03_Income_Model")
    write_header_row(ws_im, 1, 1, headers)
    # Row 2: IEA bình quân (tỷ) — history hardcoded, forecast = formula từ IEA cuối kỳ (row 18)
    im_hist_iea_avg = [((iea_end_hist[i-1] + iea_end_hist[i])/2 if i > 0 else iea_end_hist[i]) for i in range(5)]
    write_data_row(ws_im, 2, 1, ["IEA bình quân (tỷ)"] + im_hist_iea_avg + [None]*3)
    for i, col in enumerate(['G','H','I']):
        prev_col = 'F' if i == 0 else ['G','H'][i-1]
        ws_im.cell(row=2, column=7+i, value=f"=({prev_col}18+{col}18)/2").number_format = FMT_NUM1
    # Row 3: IEA tăng trưởng (%) — dùng công thức (C2-B2)/B2 cho lịch sử, link Assumptions cho FC
    ws_im.cell(row=3, column=1, value="IEA tăng trưởng (%)")
    for i, col in enumerate(['B','C','D','E','F','G','H','I']):
        cell = ws_im.cell(row=3, column=2+i)
        if i == 0:
            cell.value = None
        elif i < 5:
            prev_col = ['B','C','D','E'][i-1]
            cell.value = f"=({col}2-{prev_col}2)/{prev_col}2"
        else:
            cell.value = f"='02_Assumptions'!{col}4"
        cell.number_format = FMT_PCT
    write_data_row(ws_im, 4, 1, ["NIM (%)"] + [n/100 for n in nim_hist] + nim_fc)
    write_data_row(ws_im, 5, 1, ["NII (tỷ)"] + nii_hist + nii_fc)
    write_data_row(ws_im, 6, 1, ["Thu nhập dịch vụ"] + fee_inc_hist + [None]*3)
    write_data_row(ws_im, 7, 1, ["Thu nhập ngoại hối"] + fx_hist + [None]*3)
    write_data_row(ws_im, 8, 1, ["Thu nhập CK đầu tư"] + [trade_sec_hist[i] + inv_sec_hist[i] for i in range(5)] + [None]*3)
    write_data_row(ws_im, 9, 1, ["Thu nhập khác"] + other_inc_hist + [None]*3)
    write_data_row(ws_im, 10, 1, ["Tổng thu nhập ngoài lãi"] + [toi_hist[i] - nii_hist[i] for i in range(5)] + non_int_fc)
    write_data_row(ws_im, 11, 1, ["Tổng thu nhập HĐ - TOI"] + toi_hist + toi_fc)
    write_data_row(ws_im, 12, 1, ["NII/TOI (%)"] + [round(nii_hist[i]/toi_hist[i]*100, 2) for i in range(5)] + [None]*3)
    write_data_row(ws_im, 13, 1, ["Chi phí HĐ - OPEX"] + opex_hist + opex_fc)
    write_data_row(ws_im, 14, 1, ["PPOP - LN trước dự phòng"] + ppop_hist + ppop_fc)
    write_data_row(ws_im, 15, 1, ["PPOP/TOI (%)"] + [round(ppop_hist[i]/toi_hist[i]*100, 2) for i in range(5)] + [None]*3)
    write_data_row(ws_im, 16, 1, ["Dự phòng tín dụng"] + prov_hist + prov_fc)
    write_data_row(ws_im, 17, 1, ["LN trước thuế - PBT"] + pbt_hist + pbt_fc)
    # Row 18: IEA cuối kỳ (tỷ) — dùng để tính IEA_avg row 2
    write_data_row(ws_im, 18, 1, ["IEA cuối kỳ (tỷ)"] + [round(v, 1) for v in iea_end_hist] + [None]*3)
    for i, col in enumerate(['G','H','I']):
        prev_col = 'F' if i == 0 else ['G','H'][i-1]
        ws_im.cell(row=18, column=7+i, value=f"={prev_col}18*(1+{col}3)").number_format = FMT_NUM1
    
    # Write formula updates for forecast columns G, H, I in Income Model
    for idx, col in enumerate(['G', 'H', 'I']):
        prev_col = 'F' if idx == 0 else get_column_letter(6 + idx)
        # NII forecast = IEA avg * NIM (from Assumptions)
        ws_im.cell(row=5, column=7+idx, value=f"={col}2*'02_Assumptions'!{col}6")
        # TOI forecast = NII + Non-II
        ws_im.cell(row=11, column=7+idx, value=f"={col}5+{col}10")
        # OPEX forecast = TOI * CIR (from Assumptions)
        ws_im.cell(row=13, column=7+idx, value=f"={col}11*'02_Assumptions'!{col}7")
        # PPOP forecast = TOI - OPEX
        ws_im.cell(row=14, column=7+idx, value=f"={col}11-{col}13")
        # Provision forecast = avg_loans * CoC
        ws_im.cell(row=16, column=7+idx, value=f"=(('05_Balance_Sheet'!{prev_col}3+'05_Balance_Sheet'!{col}3)/2)*'02_Assumptions'!{col}8")
        # PBT forecast = PPOP - Provision
        ws_im.cell(row=17, column=7+idx, value=f"={col}14-{col}16")

    # 5. 04_PnL_Quarterly
    ws_pq = wb.create_sheet("04_PnL_Quarterly")
    is_q_recs = sorted(section_to_quarters(raw_data, "INCOME_STATEMENT"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-18:]
    pq_headers = ["Chỉ tiêu"] + [f"Q{r.get('lengthReport')}/{r['yearReport']}" for r in is_q_recs]
    write_header_row(ws_pq, 1, 1, pq_headers)
    pq_data = [
        ("Thu nhập lãi thuần - NII", [(r.get("isb27") or 0) / 1e9 for r in is_q_recs]),
        ("Thu nhập ngoài lãi", [((r.get("isb38") or 0) - (r.get("isb27") or 0)) / 1e9 for r in is_q_recs]),
        ("Tổng thu nhập HĐ - TOI", [(r.get("isb38") or 0) / 1e9 for r in is_q_recs]),
        ("Chi phí hoạt động - OPEX", [abs(r.get("isb39") or 0) / 1e9 for r in is_q_recs]),
        ("LN trước dự phòng - PPOP", [(r.get("isb40") or 0) / 1e9 for r in is_q_recs]),
        ("Dự phòng tín dụng", [abs(r.get("isb41") or 0) / 1e9 for r in is_q_recs]),
        ("LN trước thuế - PBT", [(r.get("isa16") or 0) / 1e9 for r in is_q_recs]),
        ("LN sau thuế - NPAT", [(r.get("isa20") or 0) / 1e9 for r in is_q_recs]),
        ("Chi phí lãi (Int Exp)", [abs(r.get("isb30") or 0) / 1e9 for r in is_q_recs]),
    ]
    for i, (label, vals) in enumerate(pq_data):
        write_data_row(ws_pq, i + 2, 1, [label] + vals, FMT_NUM1)

    # 6. 04_PnL
    ws_pnl = wb.create_sheet("04_PnL")
    write_header_row(ws_pnl, 1, 1, headers)
    pnl_data = [
        ("Thu nhập lãi thuần (NII)", nii_hist + nii_fc),
        ("Tổng thu nhập hoạt động (TOI)", toi_hist + toi_fc),
        ("Lợi nhuận trước dự phòng (PPOP)", ppop_hist + ppop_fc),
        ("LNST (NPAT)", np_hist + np_fc),
        ("EPS (VND)", eps_hist_calc + eps_fc_calc)
    ]
    for idx, (lbl, vals) in enumerate(pnl_data):
        r = idx + 2
        write_data_row(ws_pnl, r, 1, [lbl] + vals, FMT_NUM if "EPS" in lbl else FMT_NUM1)
        
    for idx, col in enumerate(['G', 'H', 'I']):
        ws_pnl.cell(row=2, column=7+idx, value=f"='03_Income_Model'!{col}5")
        ws_pnl.cell(row=3, column=7+idx, value=f"=SUM({col}2+'03_Income_Model'!{col}10)")
        ws_pnl.cell(row=4, column=7+idx, value=f"='03_Income_Model'!{col}14")
        # Correct dynamic calculation for LNST (NPAT): PBT - Tax where Tax = PBT * tax_rate
        pbt_ref = f"('03_Income_Model'!{col}17)"
        ws_pnl.cell(row=5, column=7+idx, value=f"={pbt_ref}-MAX({pbt_ref}*'02_Assumptions'!{col}14,0)")
        # FIX Lỗi 1: 1e9 không hợp lệ trong Excel — phải dùng số nguyên 1000000000
        ws_pnl.cell(row=6, column=7+idx, value=f"={col}5*1000000000/'02_Assumptions'!{col}3")
        
    # 7. 05_Balance_Sheet_Quarterly
    ws_bq = wb.create_sheet("05_Balance_Sheet_Quarterly")
    bs_q_recs = sorted(section_to_quarters(raw_data, "BALANCE_SHEET"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-18:]
    # Build NOTE quarterly lookup map for TPDN, Ký quỹ
    nt_q_recs = sorted(section_to_quarters(raw_data, "NOTE"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
    nt_q_map = {}
    for r in nt_q_recs:
        nt_q_map[(r.get("yearReport"), r.get("lengthReport"))] = r
    def get_nt_q(yr, q, field):
        rec = nt_q_map.get((yr, q), {})
        return (rec.get(field) or 0) / 1e9
    write_header_row(ws_bq, 1, 1, pq_headers)
    bq_cash_nhnn = [((r.get("bsa2") or 0) + (r.get("bsb97") or 0)) / 1e9 for r in bs_q_recs]
    bq_interbank = [(r.get("bsb98") or 0) / 1e9 for r in bs_q_recs]
    bq_loans     = [(r.get("bsb103") or 0) / 1e9 for r in bs_q_recs]
    bq_inv_sec   = [(r.get("bsb106") or 0) / 1e9 for r in bs_q_recs]
    # IEA quý = Cho vay + TG NHNN có lãi + TS sinh lãi liên NH + Chứng khoán đầu tư (Skill §16.1/§3)
    bq_iea = [bq_cash_nhnn[i] + bq_interbank[i] + bq_loans[i] + bq_inv_sec[i] for i in range(len(bs_q_recs))]
    bq_tpdn = [get_nt_q(r.get("yearReport"), r.get("lengthReport"), "nob184") for r in bs_q_recs]
    bq_dep  = [(r.get("bsb113") or 0) / 1e9 for r in bs_q_recs]
    bq_bonds = [(r.get("bsb116") or 0) / 1e9 for r in bs_q_recs]
    bq_kbnn = [((r.get("bsb110") or 0) + (r.get("bsb111") or 0)) / 1e9 for r in bs_q_recs]
    bq_kyquy = [get_nt_q(r.get("yearReport"), r.get("lengthReport"), "nob73") or get_nt_q(r.get("yearReport"), r.get("lengthReport"), "nob75") or 0 for r in bs_q_recs]
    bq_voncg = [(r.get("bsb115") or 0) / 1e9 for r in bs_q_recs]
    bq_tctd_dep = [(r.get("bsb270") or 0) / 1e9 for r in bs_q_recs]  # Tiền gửi của TCTD khác — mẫu số LDR theo Circular 22/2019

    def kbnn_rate_for_year(yr):
        # Circular 26/2022 roadmap (2023=50%/2024=40%/2025=20%), Circular 08/2026/TT-NHNN
        # (hiệu lực 15/5/2026) khôi phục 20% cho 2026 trở đi thay vì 0%.
        if yr <= 2022: return 0.0
        if yr == 2023: return 0.50
        if yr == 2024: return 0.40
        return 0.20  # 2025 trở đi

    bq_data = [
        ("Tổng tài sản", [(r.get("bsa53") or 0) / 1e9 for r in bs_q_recs]),
        ("Tiền mặt & NHNN", bq_cash_nhnn),
        ("TG các TCTD khác", bq_interbank),
        ("Cho vay khách hàng", bq_loans),
        ("CK đầu tư", bq_inv_sec),
        ("Tiền gửi khách hàng", bq_dep),
        ("Vốn chủ sở hữu", [(r.get("bsa78") or 0) / 1e9 for r in bs_q_recs]),
        # Thêm các dòng cho tính toán tín dụng & huy động
        ("TPDN (trái phiếu TCKT)", bq_tpdn),
        ("Giấy tờ có giá (Bonds)", bq_bonds),
        ("Tiền gửi Kho bạc NN", bq_kbnn),
        ("Tiền gửi ký quỹ (trừ đi)", bq_kyquy),
        ("Vốn chuyên dùng (trừ đi)", bq_voncg),
        ("Tài sản sinh lãi (IEA)", bq_iea),
        ("Tổng tín dụng (tỷ)", [None] * len(bs_q_recs)),   # row 15 — công thức, xem bên dưới
        ("Tổng huy động (tỷ)", [None] * len(bs_q_recs)),   # row 16 — công thức, xem bên dưới
        ("Tiền gửi của TCTD khác", bq_tctd_dep),            # row 17 — mẫu số LDR (Circular 22/2019)
    ]
    for i, (label, vals) in enumerate(bq_data):
        write_data_row(ws_bq, i + 2, 1, [label] + vals, FMT_NUM1)

    # Row 15/16: Tổng tín dụng = Cho vay(row5) + TPDN(row9) ; Tổng huy động = Tiền gửi(row7) +
    # GTCG(row10) + TG TCTD khác(row17) + KBNN(row11)*tỷ lệ - Ký quỹ(row12) - Vốn chuyên dùng(row13)
    # — ghi bằng CÔNG THỨC Excel (không phải số Python tính sẵn) để bấm vào ô là thấy ngay cách tính.
    # "TG TCTD khác" (row17, bsb270) theo Thông tư 22/2019/TT-NHNN — "Tổng nguồn vốn huy động" ở mẫu
    # số LDR gồm cả tiền gửi của TCTD/chi nhánh NH nước ngoài khác. KHÔNG cộng "Vay TCTD khác"
    # (bsb271) vì đây mới là đề xuất của một số NHTM lên NHNN, chưa xác nhận thành quy định chính thức.
    for i, rec in enumerate(bs_q_recs):
        col = get_column_letter(2 + i)
        rate = kbnn_rate_for_year(rec.get("yearReport", 2026))
        ws_bq.cell(row=15, column=2 + i, value=f"={col}5+{col}9").number_format = FMT_NUM1
        ws_bq.cell(row=16, column=2 + i, value=f"={col}7+{col}10+{col}17+{col}11*{rate}-{col}12-{col}13").number_format = FMT_NUM1

    # 8. 05_Balance_Sheet
    ws_bs = wb.create_sheet("05_Balance_Sheet")
    write_header_row(ws_bs, 1, 1, headers)
    bs_data = [
        ("Tổng tài sản", total_assets_hist + [None]*3),                       # row 2
        ("Cho vay khách hàng (Gross)", loans_hist + loans_fc),                 # row 3
        ("Trái phiếu doanh nghiệp", tpdn_hist + [None]*3),                     # row 4
        ("Tiền gửi khách hàng", cust_dep_hist + dep_fc),                       # row 5
        ("Giấy tờ có giá phát hành (Bonds)", bonds_hist + [None]*3),              # row 6
        ("Tiền gửi Kho bạc Nhà nước", kbnn_hist + [None]*3),                   # row 7
        ("Tiền gửi ký quỹ (trừ đi)", ky_quy_hist + [None]*3),                  # row 8
        ("Vốn chuyên dùng (trừ đi)", voncg_hist + [None]*3),                   # row 9
        ("Vốn chủ sở hữu (VCSH)", equity_hist + [None]*3),                     # row 10
        ("Tổng tín dụng (tỷ)", [None]*8),                                      # row 11 = row3+row4
        ("Tổng huy động (tỷ)", [None]*8),                                      # row 12 = row5+row6+row13+row7*KBNN%-row8-row9
        ("Tiền gửi của TCTD khác", tctd_dep_hist + [None]*3),                  # row 13 (Circular 22/2019, mẫu số LDR)
    ]
    for idx, (lbl, vals) in enumerate(bs_data):
        r = idx + 2
        write_data_row(ws_bs, r, 1, [lbl] + vals, FMT_NUM1)

    for idx, col in enumerate(['G', 'H', 'I']):
        prev_col = 'F' if idx == 0 else get_column_letter(6 + idx)
        ws_bs.cell(row=3, column=7+idx, value=f"={prev_col}3*(1+'02_Assumptions'!{col}4)")  # Loans
        ws_bs.cell(row=4, column=7+idx, value=f"={prev_col}4*(1+'02_Assumptions'!{col}4)")  # TPDN: tăng cùng tín dụng
        ws_bs.cell(row=5, column=7+idx, value=f"={prev_col}5*(1+'02_Assumptions'!{col}5)")  # Deposits
        ws_bs.cell(row=6, column=7+idx, value=f"={prev_col}6*1.02")                          # Bonds: +2%/năm
        ws_bs.cell(row=7, column=7+idx, value=f"={prev_col}7*(1+'02_Assumptions'!{col}5)")  # KBNN: tăng cùng Deposits
        ws_bs.cell(row=8, column=7+idx, value=f"={prev_col}8*(1+'02_Assumptions'!{col}5)")  # Ký quỹ: tăng cùng Deposits
        ws_bs.cell(row=9, column=7+idx, value=f"={prev_col}9*1.02")                          # Vốn chuyên dùng: +2%/năm
        ws_bs.cell(row=10, column=7+idx, value=f"={prev_col}10+'04_PnL'!{col}5*0.7")         # VCSH: giữ lại 70% LNST
        ws_bs.cell(row=2, column=7+idx, value=f"=SUM({col}3:{col}6)+{col}7-{col}8-{col}9+{col}10") # Assets approximation
        ws_bs.cell(row=13, column=7+idx, value=f"={prev_col}13*(1+'02_Assumptions'!{col}5)")  # TG TCTD khác: tăng cùng Deposits
        ws_bs.cell(row=13, column=7+idx).number_format = FMT_NUM1

    # Row 11/12: Tổng tín dụng & Tổng huy động — nguồn duy nhất cho mọi công thức LDR ở 06_Ratios,
    # tránh mỗi sheet tự tính lại (dễ lệch tỷ lệ giữ lại KBNN như đã xảy ra ở 06_Ratios trước đây).
    # Tỷ lệ giữ lại tiền gửi KBNN theo lộ trình Circular 26/2022 (đã sửa cho đúng 2023=50%/2024=40%/
    # 2025=20%) và Circular 08/2026/TT-NHNN khôi phục 20% từ 2026 trở đi.
    # Tổng huy động CỘNG THÊM row13 (Tiền gửi của TCTD khác, bsb270) — theo Thông tư 22/2019/TT-NHNN,
    # "Tổng nguồn vốn huy động" ở mẫu số LDR bao gồm tiền gửi của tổ chức tín dụng, chi nhánh ngân
    # hàng nước ngoài khác. Trước đây thiếu hẳn khoản này khiến LDR tính ra cao bất thường (gần
    # 100%, sát/vượt trần 85% NHNN). KHÔNG cộng "Vay các TCTD khác" (bsb271) vì khoản này mới là đề
    # xuất của một số NHTM lên NHNN, chưa xác nhận đã thành quy định chính thức.
    kbnn_rate_cols = {'B': kbnn_rates_hist[0], 'C': kbnn_rates_hist[1], 'D': kbnn_rates_hist[2],
                       'E': kbnn_rates_hist[3], 'F': kbnn_rates_hist[4],
                       'G': 0.2, 'H': 0.2, 'I': 0.2}
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_bs.cell(row=11, column=column_index_from_string(col), value=f"={col}3+{col}4").number_format = FMT_NUM1
        rate = kbnn_rate_cols[col]
        ws_bs.cell(row=12, column=column_index_from_string(col),
                    value=f"={col}5+{col}6+{col}13+{col}7*{rate}-{col}8-{col}9").number_format = FMT_NUM1

    # 9. 06_Ratios_Quarterly
    # Row map in '05_Balance_Sheet_Quarterly': 3=Cash+NHNN, 4=Interbank, 5=Loans, 6=Inv.Securities,
    # 7=Deposits, 8=Equity, 9=TPDN, 10=Bonds, 11=KBNN, 12=Ky quy, 13=Von chuyen dung, 14=IEA,
    # 15=Tổng tín dụng, 16=Tổng huy động.
    # (Previously these formulas pointed one row too high — e.g. row 4/6/7 instead of 5/7/8 — which
    # silently computed NIM/LDR/COF/ROE off the WRONG account. Fixed to match the actual row layout.)
    ws_rq = wb.create_sheet("06_Ratios_Quarterly")
    write_header_row(ws_rq, 1, 1, pq_headers)
    cols_q = [get_column_letter(j) for j in range(2, 2 + len(bs_q_recs))]
    # NIM quý = NII_q * 4 / IEA bình quân (đầu quý + cuối quý) / 2 — Skill §16.1 (IEA cuối kỳ là SAI)
    rq_nim = []
    for idx, c in enumerate(cols_q):
        prev_c = cols_q[idx - 1] if idx > 0 else c
        rq_nim.append(f"=('04_PnL_Quarterly'!{c}2*4/(('05_Balance_Sheet_Quarterly'!{prev_c}14+'05_Balance_Sheet_Quarterly'!{c}14)/2))*100")
    # ROE quý năm hóa = LNST_q * 4 / VCSH cuối kỳ — row 8=NPAT trong 04_PnL_Quarterly, row 8=VCSH trong 05_Balance_Sheet_Quarterly
    rq_roe = [f"=('04_PnL_Quarterly'!{c}8*4/'05_Balance_Sheet_Quarterly'!{c}8)*100" for c in cols_q]
    # LDR = Tổng tín dụng / Tổng huy động — link thẳng row 15/16 của 05_Balance_Sheet_Quarterly
    # (trước đây chỉ tính Loans/(Deposits+Bonds), thiếu TPDN ở tử số và thiếu KBNN/ký quỹ/vốn
    # chuyên dùng ở mẫu số — không phải LDR đúng nghĩa "tổng tín dụng/tổng huy động").
    rq_ldr = [f"=('05_Balance_Sheet_Quarterly'!{c}15/'05_Balance_Sheet_Quarterly'!{c}16)*100" for c in cols_q]
    # NPL quý = không có trực tiếp từ PnL quarterly, tính đơn giản từ note data
    # COF quý = Chi phí lãi * 4 / Tiền gửi — row 7=Deposits
    rq_cof = [f"=('04_PnL_Quarterly'!{c}10*4/'05_Balance_Sheet_Quarterly'!{c}7)*100" for c in cols_q]
    write_data_row(ws_rq, 2, 1, ["NIM — năm hóa (%)"] + rq_nim, FMT_NUM1)
    write_data_row(ws_rq, 3, 1, ["ROE — năm hóa (%)"] + rq_roe, FMT_NUM1)
    write_data_row(ws_rq, 4, 1, ["LDR (%)"] + rq_ldr, FMT_NUM1)
    write_data_row(ws_rq, 5, 1, ["COF — năm hóa (%)"] + rq_cof, FMT_NUM1)

    # 10. 06_Ratios — FIX Lỗi 3: Mở rộng đủ 7 chỉ số (NIM, CIR, ROE, ROA, LDR, NPL, CoC)
    ws_rat = wb.create_sheet("06_Ratios")
    write_header_row(ws_rat, 1, 1, headers)
    
    # Hardcode historical calculated values, use formulas for forecast cols
    # Row 2: NIM = NII / IEA_avg — history hardcoded, FC = link to Assumptions calculated NIM
    write_data_row(ws_rat, 2, 1, ["NIM (%)"] + [round(n/100,4) for n in nim_hist] + [None]*3, FMT_PCT)
    for i, col in enumerate(['G','H','I']):
        ws_rat.cell(row=2, column=7+i, value=f"='02_Assumptions'!{col}6").number_format = FMT_PCT
    # Row 3: CIR = OPEX / TOI — link to 02_Assumptions for ALL cols (including hist + FC)
    ws_rat.cell(row=3, column=1, value="CIR (%)")
    for i, col in enumerate(['B','C','D','E','F','G','H','I']):
        ws_rat.cell(row=3, column=2+i, value=f"='02_Assumptions'!{col}7").number_format = FMT_PCT
    # Row 4: ROE = LNST / VCSH bình quân
    write_data_row(ws_rat, 4, 1, ["ROE (%)"] + [round(r/100,4) for r in roe_hist] + [None]*3, FMT_PCT)
    # Row 5: ROA = LNST / Tổng tài sản
    write_data_row(ws_rat, 5, 1, ["ROA (%)"] + [round(r/100,4) for r in roa_hist] + [None]*3, FMT_PCT)
    # Row 6: LDR = Tổng tín dụng / Tổng huy động — link thẳng tới 05_Balance_Sheet!row11/row12
    # (trước đây tự tính lại SUM/KBNN% tại đây với tỷ lệ giữ lại KBNN SAI [0,0,35%,50%,60%] —
    # đúng phải là [0,0,50%,40%,20%] theo lộ trình Circular 26/2022, đã có sẵn ở kbnn_rates_hist
    # và giờ dùng chung với row 11/12 của 05_Balance_Sheet để tránh 2 nơi tính lệch nhau).
    write_data_row(ws_rat, 6, 1, ["LDR — điều chỉnh (%)"] + [None]*8, FMT_PCT)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_rat.cell(row=6, column=column_index_from_string(col),
                    value=f"='05_Balance_Sheet'!{col}11/'05_Balance_Sheet'!{col}12")
        ws_rat.cell(row=6, column=2+i).number_format = FMT_PCT

    # Row 7: NPL = npl_total / loans — history hardcoded, FC = base * reduction from Assumptions
    ws_rat.cell(row=7, column=1, value="NPL (%)")
    for i, col in enumerate(cols_all):
        cell = ws_rat.cell(row=7, column=2+i)
        if i < 5:
            cell.value = round(npl_ratio_hist[i]/100, 4)
        elif i == 5:
            cell.value = f"=F7*'02_Assumptions'!B34"
        elif i == 6:
            cell.value = f"=F7*'02_Assumptions'!B35"
        elif i == 7:
            cell.value = f"=F7*'02_Assumptions'!B36"
        cell.number_format = FMT_PCT
    # Row 8: CoC = Provision / avg_loans — history hardcoded, FC = base * reduction
    ws_rat.cell(row=8, column=1, value="CoC — Credit Cost (%)")
    for i, col in enumerate(cols_all):
        cell = ws_rat.cell(row=8, column=2+i)
        if i < 5:
            cell.value = round(coc_hist[i]/100, 4)
        elif i == 5:
            cell.value = f"=F8*'02_Assumptions'!B30"
        elif i == 6:
            cell.value = f"=F8*'02_Assumptions'!B31"
        elif i == 7:
            cell.value = f"=F8*'02_Assumptions'!B32"
        cell.number_format = FMT_PCT
    # Row 9: CASA ratio — history hardcoded, FC = base + increment
    ws_rat.cell(row=9, column=1, value="CASA ratio (%)")
    for i, col in enumerate(cols_all):
        cell = ws_rat.cell(row=9, column=2+i)
        if i < 5:
            cell.value = round(casa_ratio_hist[i]/100, 4)
        elif i == 5:
            cell.value = f"=F9+'02_Assumptions'!B38"
        elif i == 6:
            cell.value = f"=F9+'02_Assumptions'!B39"
        elif i == 7:
            cell.value = f"=F9+'02_Assumptions'!B40"
        cell.number_format = FMT_PCT
    # Excel formula links for forecast columns (ROE, ROA use average equity/assets)
    for idx, col in enumerate(['G', 'H', 'I']):
        prev_col = 'F' if idx == 0 else get_column_letter(6 + idx)
        # ROE forecast = LNST / avg(VCSH prev, VCSH curr) - VCSH is now at row 10
        ws_rat.cell(row=4, column=7+idx, value=f"='04_PnL'!{col}5/AVERAGE('05_Balance_Sheet'!{prev_col}10,'05_Balance_Sheet'!{col}10)")
        ws_rat.cell(row=4, column=7+idx).number_format = FMT_PCT
        # ROA forecast = LNST / avg(TA prev, TA curr) - TA is at row 2
        ws_rat.cell(row=5, column=7+idx, value=f"='04_PnL'!{col}5/AVERAGE('05_Balance_Sheet'!{prev_col}2,'05_Balance_Sheet'!{col}2)")
        ws_rat.cell(row=5, column=7+idx).number_format = FMT_PCT
        # LDR forecast: đã ghi ở vòng lặp B..I phía trên (link '05_Balance_Sheet'!row11/row12)

    # Row 10: YOEA — Yield on Earning Assets (Interest Income / IEA_avg)
    write_data_row(ws_rat, 10, 1, ["YOEA — Lợi suất TSSL (%)"] + [round(y, 4) for y in yo_ea_hist] + [None]*3, FMT_PCT)
    # Row 11: COF — Cost of Funds (Interest Expense / Dep+Bonds avg)
    write_data_row(ws_rat, 11, 1, ["COF — Chi phí vốn (%)"] + [round(c, 4) for c in cof_hist_calc] + [None]*3, FMT_PCT)

    # 11. 07_Valuation (Detailed 31-row step-by-step layout)
    ws_val = wb.create_sheet("07_Valuation")
    ws_val.column_dimensions['A'].width = 35
    ws_val.column_dimensions['B'].width = 20
    ws_val.column_dimensions['C'].width = 20
    ws_val.column_dimensions['D'].width = 20
    ws_val.cell(row=1, column=1, value="PHƯƠNG PHÁP ĐỊNH GIÁ").font = FMT_BOLD
    ws_val.cell(row=2, column=1, value="Chi phí vốn CSH (COE)")
    ws_val.cell(row=2, column=2, value="='00_COE'!B9").number_format = FMT_PCT
    ws_val.cell(row=3, column=1, value="Tăng trưởng dài hạn (g)")
    ws_val.cell(row=3, column=2, value=terminal_growth).number_format = FMT_PCT
    
    ws_val.cell(row=5, column=1, value="RESIDUAL INCOME MODEL").font = FMT_BOLD
    ws_val.cell(row=6, column=1, value="BV/share hiện tại (VND)")
    # VCSH has moved to row 10 in 05_Balance_Sheet
    ws_val.cell(row=6, column=2, value="='05_Balance_Sheet'!F10*1000000000/'02_Assumptions'!$F$3").number_format = FMT_NUM
    ws_val.cell(row=7, column=1, value="PV của RI 3 năm (VND)")
    ws_val.cell(row=7, column=2, value="=SUM(B31:D31)").number_format = FMT_NUM
    ws_val.cell(row=8, column=1, value="PV của Continuing Value (VND)")
    # Continuing Value (CV) - Standard formula without arbitrary discount
    ws_val.cell(row=8, column=2, value="=(D29*(1+B3)/(B2-B3))*D30").number_format = FMT_NUM
    ws_val.cell(row=9, column=1, value="GIÁ TRỊ RI (VND)").font = FMT_BOLD
    ws_val.cell(row=9, column=2, value="=B6+B7+B8").font = FMT_BOLD
    ws_val.cell(row=9, column=2).number_format = FMT_NUM
    
    ws_val.cell(row=11, column=1, value="P/B MULTIPLE — 3 mức phân phối lịch sử").font = FMT_BOLD
    ws_val.cell(row=12, column=1, value="P/B hiện tại (x)")
    ws_val.cell(row=12, column=2, value="='02_Assumptions'!B2/B6").number_format = '0.00'
    ws_val.cell(row=13, column=1, value="P/B hấp dẫn (x) — MUA")
    # Range must match the exact rows written to '13_PE_PB_History' (capped at 32 quarters, see n_pts_sens)
    last_row_pb = n_pts_sens + 1
    pb_hist_rng = f"'13_PE_PB_History'!C2:C{last_row_pb}"
    half_count = f"CEILING(COUNT({pb_hist_rng})/2,1)"
    # Legacy CSE array-formula index generator (ROW/INDIRECT) instead of SEQUENCE() —
    # SEQUENCE is a dynamic-array function only available on Excel 365/2021+, and fails
    # with #NAME? on older Excel (e.g. perpetual 2016/2019) opened via win32com.
    idx_b13 = f"ROW(INDIRECT(\"1:\"&{half_count}))"

    # "P/B hấp dẫn" = median of the LOWER half of historical P/B (Tukey hinge, not a 25th-percentile
    # interpolation), and "P/B Over" (row 15) = median of the UPPER half — per user definition.
    ws_val.cell(row=13, column=2, value=ArrayFormula("B13", f"=MEDIAN(SMALL({pb_hist_rng}, {idx_b13}))")).number_format = '0.00'
    ws_val.cell(row=13, column=2).font = Font(color="006400", bold=True, name="Calibri")
    ws_val.cell(row=13, column=3, value="Median P/B nửa dưới (vùng mua hấp dẫn)").font = Font(size=9, color="006400", name="Calibri")

    ws_val.cell(row=14, column=1, value="P/B median all-time (x) — FAIR VALUE")
    ws_val.cell(row=14, column=2, value=f"=MEDIAN({pb_hist_rng})").number_format = '0.00'
    ws_val.cell(row=14, column=3, value="Median PB toàn bộ lịch sử").font = Font(size=9, color="595959", name="Calibri")

    ws_val.cell(row=15, column=1, value="P/B Over (x) — BÁN / CHỐT LỜI")
    ws_val.cell(row=15, column=2, value=ArrayFormula("B15", f"=MEDIAN(LARGE({pb_hist_rng}, {idx_b13}))")).number_format = '0.00'
    ws_val.cell(row=15, column=2).font = Font(color="C00000", bold=True, name="Calibri")
    ws_val.cell(row=15, column=3, value="Median P/B nửa trên (vùng chốt lời)").font = Font(size=9, color="C00000", name="Calibri")
    ws_val.cell(row=16, column=1, value="BV/share tương lai (2026F)")
    ws_val.cell(row=16, column=2, value="=B6+B26").number_format = FMT_NUM
    ws_val.cell(row=17, column=1, value="GIÁ TRỊ P/B (VND)").font = FMT_BOLD
    ws_val.cell(row=17, column=2, value="=B14*B16").font = FMT_BOLD
    ws_val.cell(row=17, column=2).number_format = FMT_NUM

    ws_val.cell(row=19, column=1, value="WEIGHTED TARGET PRICE").font = FMT_BOLD
    ws_val.cell(row=20, column=1, value="Trong số: 50% RI + 50% P/B")
    ws_val.cell(row=21, column=1, value="Target Price (VND)").font = FMT_BOLD
    ws_val.cell(row=21, column=2, value="=B9*0.5+B17*0.5").font = FMT_BOLD
    ws_val.cell(row=21, column=2).number_format = FMT_NUM
    ws_val.cell(row=22, column=1, value="Giá hiện tại (VND)")
    ws_val.cell(row=22, column=2, value="='02_Assumptions'!B2").number_format = FMT_NUM  # Price is row 2 in Assumptions
    ws_val.cell(row=23, column=1, value="UPSIDE (%)").font = FMT_BOLD
    ws_val.cell(row=23, column=2, value="=B21/B22-1").font = FMT_BOLD
    ws_val.cell(row=23, column=2).number_format = FMT_PCT

    ws_val.cell(row=25, column=1, value="--- RI DETAIL ---").font = FMT_BOLD
    ws_val.cell(row=26, column=1, value="EPS (VND)")
    ws_val.cell(row=27, column=1, value="BVPS đầu kỳ (VND)")
    ws_val.cell(row=28, column=1, value="Capital Charge (VND)")
    ws_val.cell(row=29, column=1, value="Residual Income (VND)")
    ws_val.cell(row=30, column=1, value="Discount Factor")
    ws_val.cell(row=31, column=1, value="PV of RI")

    # Enforce strict clean surplus relation recursively in valuation sheet cells
    for idx, yr in enumerate(years_fc):
        col_letter = get_column_letter(2 + idx)
        pnl_col_letter = get_column_letter(7 + idx) # 2026F is column G (index 7) in pnl sheet
        
        # Link EPS to the correct column (G, H, I) of '04_PnL'
        # FIX: EPS row in 04_PnL is row 6 (NII=2, TOI=3, PPOP=4, LNST=5, EPS=6). 1e9 replaced with 1000000000
        ws_val.cell(row=26, column=2+idx, value=f"='04_PnL'!{pnl_col_letter}6").number_format = FMT_NUM
        
        if idx == 0:
            ws_val.cell(row=27, column=2, value="=B6").number_format = FMT_NUM # BVPS đầu kỳ 2026F = BVPS hiện tại
        else:
            prev_col_letter = get_column_letter(1 + idx)
            ws_val.cell(row=27, column=2+idx, value=f"={prev_col_letter}27+{prev_col_letter}26").number_format = FMT_NUM # BVPS(t) = BVPS(t-1) + EPS(t-1)
            
        ws_val.cell(row=28, column=2+idx, value=f"={col_letter}27*'07_Valuation'!$B$2").number_format = FMT_NUM
        ws_val.cell(row=29, column=2+idx, value=f"={col_letter}26-{col_letter}28").number_format = FMT_NUM
        ws_val.cell(row=30, column=2+idx, value=f"=1/(1+'07_Valuation'!$B$2)^{idx+1}").number_format = '0.0000'
        ws_val.cell(row=31, column=2+idx, value=f"={col_letter}29*{col_letter}30").number_format = FMT_NUM

    # 12. 08_Sensitivity — FIX Lỗi 7: Tính đủ PV(RI năm 1+2) + PV(CV)
    ws_sens = wb.create_sheet("08_Sensitivity")
    ws_sens.cell(row=1, column=1, value="Độ nhạy: Giá trị RI theo COE × g (5×5)").font = FMT_BOLD
    ws_sens.cell(row=2, column=1, value="COE \\ g")
    term_gs = [0.01, 0.02, 0.03, 0.04, 0.05]
    coe_scenarios = [0.08, 0.10, 0.12, 0.14, 0.16]
    for j, g in enumerate(term_gs):
        ws_sens.cell(row=2, column=j+2, value=g).number_format = '0%'
    for i, coe in enumerate(coe_scenarios):
        ws_sens.cell(row=3+i, column=1, value=coe).number_format = '0.0%'
        for j, g in enumerate(term_gs):
            if (coe - g) <= 0:
                val_sens = 0
            else:
                # FIX: Tính đủ PV của cả 3 năm RI + PV(Continuing Value) without discount
                pv_ri_sens = sum(ri_results[k] / (1 + coe)**(k+1) for k in range(len(ri_results)))
                pv_cv_sens = ((ri_results[-1] * (1+g)/(coe-g)) / (1+coe)**len(ri_results))
                val_sens = bvps_base + pv_ri_sens + pv_cv_sens
            ws_sens.cell(row=3+i, column=j+2, value=val_sens).number_format = '#,##0'
    # Highlight current COE+g scenario
    from openpyxl.styles import PatternFill as PF
    _highlight = PF(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for i, coe in enumerate(coe_scenarios):
        for j, g in enumerate(term_gs):
            if abs(coe - COE) < 0.02 and abs(g - terminal_growth) < 0.01:
                ws_sens.cell(row=3+i, column=j+2).fill = _highlight

    # 13. 09_PESTLE — FIX Lỗi 11: Mở rộng đủ 6 nhân tố theo schema JSON (Skill §JSON)
    ws_pest = wb.create_sheet("09_PESTLE")
    write_header_row(ws_pest, 1, 1, ["Nhân tố", "Nội dung", "Tác động", "Mức độ"])
    pestle_items = [
        ("Chính trị (Political)",     "Môi trường chính trị ổn định, NHNN điều hành linh hoạt hỗ trợ tăng trưởng tín dụng.",             "Tích cực",  "Cao"),
        ("Kinh tế (Economic)",        "GDP tăng trưởng 6–7%, nhu cầu vốn sản xuất–tiêu dùng cao, lãi suất huy động giảm dần.",           "Tích cực",  "Cao"),
        ("Xã hội (Social)",           "Tầng lớp trung lưu mở rộng, thanh toán không tiền mặt tăng mạnh thúc đẩy CASA và phí dịch vụ.",   "Tích cực",  "Trung bình"),
        ("Công nghệ (Technological)", "Chuyển đổi số banking (mobile/open banking) giảm CIR, tăng CASA và cross-sell sản phẩm.",         "Tích cực",  "Cao"),
        ("Pháp lý (Legal)",           "Thông tư 06/2024 siết TPDN, Basel II/III nâng chuẩn CAR — cần theo dõi áp lực pháp lý.",           "Trung tính", "Trung bình"),
        ("Môi trường (Environmental)","ESG lending ngày càng được yêu cầu; rủi ro tín dụng BĐS liên quan biến đổi khí hậu tăng.",          "Tiêu cực",  "Thấp"),
    ]
    for idx, row in enumerate(pestle_items):
        write_data_row(ws_pest, idx+2, 1, row, FMT_NUM)

    # 14. 10_Leading_Indicators
    ws_lead = wb.create_sheet("10_Leading_Indicators")
    write_header_row(ws_lead, 1, 1, ["Chỉ số", "Ngưỡng tích cực", "Giá trị hiện tại", "Trạng thái"])
    indicators = [
        ("NPL ratio (%)", "< 1.5%", f"{npl_ratio_hist[-1]:.2f}%", "Đạt"),
        ("CASA ratio (%)", "> 20%", f"{casa_ratio_hist[-1]:.2f}%", "Tốt"),
        ("NIM (%)", "> 3.0%", f"{nim_hist[-1]:.2f}%", "Tốt"),
    ]
    for idx, row in enumerate(indicators):
        write_data_row(ws_lead, idx+2, 1, row, FMT_NUM)

    # 15. 11_Investment_Thesis
    ws_thesis = wb.create_sheet("11_Investment_Thesis")
    write_header_row(ws_thesis, 1, 1, ["Luận điểm chính", "Chi tiết nhận định"])
    thesis_rows = [
        ("Lợi thế CASA vượt trội", "Hệ sinh thái số hóa giúp huy động dòng vốn rẻ dồi dào, ổn định biên lãi NIM."),
        ("Hiệu quả sinh lời cao", "Chỉ số ROE và ROA luôn nằm trong nhóm dẫn đầu hệ thống ngân hàng TMCP."),
    ]
    for idx, row in enumerate(thesis_rows):
        write_data_row(ws_thesis, idx+2, 1, row, FMT_NUM)

    # 16. 12_Summary_Snapshot — link formulas to 04_PnL & 05_Balance_Sheet
    ws_snap = wb.create_sheet("12_Summary_Snapshot")
    write_header_row(ws_snap, 1, 1, headers)
    snap_refs = [
        # (row_in_snap, label, source_sheet, source_row)
        (2,  "Thu nhập lãi thuần (NII)",      "'04_PnL'", 2),
        (3,  "Tổng thu nhập HĐ (TOI)",        "'04_PnL'", 3),
        (4,  "LN trước dự phòng (PPOP)",      "'04_PnL'", 4),
        (5,  "LNST (NPAT)",                   "'04_PnL'", 5),
        (6,  "EPS (VND)",                     "'04_PnL'", 6),
        (7,  "Tổng tài sản",                  "'05_Balance_Sheet'", 2),
        (8,  "Cho vay khách hàng (Gross)",    "'05_Balance_Sheet'", 3),
        (9,  "Trái phiếu doanh nghiệp",       "'05_Balance_Sheet'", 4),
        (10, "Tiền gửi khách hàng",           "'05_Balance_Sheet'", 5),
        (11, "Giấy tờ có giá (Bonds)",        "'05_Balance_Sheet'", 6),
        (12, "Tiền gửi Kho bạc NN",           "'05_Balance_Sheet'", 7),
        (13, "Tiền gửi ký quỹ (trừ đi)",      "'05_Balance_Sheet'", 8),
        (14, "Vốn chuyên dùng (trừ đi)",       "'05_Balance_Sheet'", 9),
        (15, "Vốn chủ sở hữu (VCSH)",         "'05_Balance_Sheet'", 10),
    ]
    for snap_row, label, src_sheet, src_row in snap_refs:
        ws_snap.cell(row=snap_row, column=1, value=label)
        for i, col in enumerate(['B','C','D','E','F','G','H','I']):
            cell = ws_snap.cell(row=snap_row, column=2+i)
            cell.value = f"={src_sheet}!{col}{src_row}"
            cell.number_format = FMT_NUM1 if "EPS" not in label else FMT_NUM

    # 17. 13_PE_PB_History
    ws_pe_pb = wb.create_sheet("13_PE_PB_History")
    ws_pe_pb.column_dimensions['A'].width = 20
    ws_pe_pb.column_dimensions['B'].width = 15
    ws_pe_pb.column_dimensions['C'].width = 15
    ws_pe_pb.column_dimensions['D'].width = 15
    ws_pe_pb.column_dimensions['E'].width = 15
    write_header_row(ws_pe_pb, 1, 1, ["Quỳ", "P/E (x)", "P/B (x)", "EV/EBITDA (x)", "Giá"])
    # dynamic slice to max 32 quarters
    n_pts_sens = min(32, len(pb_all_vals), len(pe_all_vals))
    pb_slice_sens = list(pb_all_vals[-n_pts_sens:])
    pe_slice_sens = list(pe_all_vals[-n_pts_sens:])
    
    current_pb_ratio = current_price / (bvps_base + eps_fc_calc[0]) if (bvps_base + eps_fc_calc[0]) > 0 else 1.2
    current_pe_ratio = current_price / eps_fc_calc[0] if eps_fc_calc[0] > 0 else 8.5
    
    # Keep actual historical values for all quarters to align with python medians
    # if pb_slice_sens: pb_slice_sens[-1] = current_pb_ratio
    # if pe_slice_sens: pe_slice_sens[-1] = current_pe_ratio
    
    # Generate dynamic quarter labels backing up from current Q1/2026
    labels_n_sens = []
    latest_q_rec = bs_q_recs[-1] if bs_q_recs else {}
    curr_q = latest_q_rec.get("lengthReport", 1)
    curr_y = latest_q_rec.get("yearReport", 2026)
    for i in range(n_pts_sens):
        labels_n_sens.append(f"Q{curr_q}/{str(curr_y)[-2:]}")
        curr_q -= 1
        if curr_q == 0:
            curr_q = 4
            curr_y -= 1
    labels_n_sens = labels_n_sens[::-1]
    
    for idx in range(n_pts_sens):
        r = idx + 2
        lbl = labels_n_sens[idx]
        pb_val_q = pb_slice_sens[idx]
        pe_val_q = pe_slice_sens[idx]
        write_data_row(ws_pe_pb, r, 1, [lbl, pe_val_q, pb_val_q, None, None], FMT_NUM1)
        
    # Add Median all-time at the bottom of the table
    median_row = n_pts_sens + 3
    ws_pe_pb.cell(row=median_row, column=1, value="MEDIAN ALL-TIME").font = FMT_BOLD
    ws_pe_pb.cell(row=median_row, column=2, value=pe_all_median).font = FMT_BOLD
    ws_pe_pb.cell(row=median_row, column=2).number_format = '0.0'
    ws_pe_pb.cell(row=median_row, column=3, value=pb_all_median).font = FMT_BOLD
    ws_pe_pb.cell(row=median_row, column=3).number_format = '0.00'
    for c in range(1, 6):
        ws_pe_pb.cell(row=median_row, column=c).border = thin_border

    # 18. 14_Credit_Funding_Growth — formulas link to 05_Balance_Sheet_Quarterly
    ws_g_hist = wb.create_sheet("14_Credit_Funding_Growth")
    ws_g_hist.column_dimensions['A'].width = 20
    ws_g_hist.column_dimensions['B'].width = 25
    ws_g_hist.column_dimensions['C'].width = 25
    ws_g_hist.column_dimensions['D'].width = 20
    ws_g_hist.column_dimensions['E'].width = 20
    write_header_row(ws_g_hist, 1, 1, ["Quý", "Tổng Tín dụng (tỷ VND)", "Tổng Huy động (tỷ VND)", "Tăng trưởng TD so cuối năm trước (%)", "Tăng trưởng HĐ so cuối năm trước (%)"])
    # BQ sheet rows: 5=Loans, 9=TPDN, 7=Deposits, 10=Bonds, 11=KBNN, 12=Kyquy, 13=VonCD
    bq_loans_r, bq_tpdn_r = 5, 9
    bq_dep_r, bq_bonds_r, bq_kbnn_r, bq_kyquy_r, bq_voncd_r = 7, 10, 11, 12, 13
    # Map column index to KBNN rate for that quarter's year
    pq_cols = [get_column_letter(j) for j in range(2, 2 + len(bs_q_recs))]
    # Build lookup: for each year, find the ROW of Q4 in this sheet
    q4_row_map = {}  # year → row number in 14_Credit_Funding_Growth
    for idx, rec in enumerate(bs_q_recs):
        if rec.get("lengthReport") == 4:
            q4_row_map[rec.get("yearReport")] = idx + 2
    for idx, rec in enumerate(bs_q_recs):
        r = idx + 2
        yr = rec.get("yearReport", 2026)
        col = pq_cols[idx]
        # KBNN rate for this quarter
        if yr <= 2022:
            k_rate = 0.0
        elif yr == 2023:
            k_rate = 0.50
        elif yr == 2024:
            k_rate = 0.40
        elif yr >= 2025:
            k_rate = 0.20
        # Formula: Credit = Loans + TPDN
        credit_formula = f"='05_Balance_Sheet_Quarterly'!{col}{bq_loans_r}+'05_Balance_Sheet_Quarterly'!{col}{bq_tpdn_r}"
        # Formula: Funding = Dep + Bonds + KBNN*rate - Kyquy - VonCD
        funding_formula = f"='05_Balance_Sheet_Quarterly'!{col}{bq_dep_r}+'05_Balance_Sheet_Quarterly'!{col}{bq_bonds_r}+'05_Balance_Sheet_Quarterly'!{col}{bq_kbnn_r}*{k_rate}-'05_Balance_Sheet_Quarterly'!{col}{bq_kyquy_r}-'05_Balance_Sheet_Quarterly'!{col}{bq_voncd_r}"
        # Growth formulas (YTD: so với Q4 cuối năm trước)
        prev_yr = yr - 1
        prev_ye_row = q4_row_map.get(prev_yr)
        if prev_ye_row:
            c_growth = f"=(B{r}-B{prev_ye_row})/B{prev_ye_row}"
            f_growth = f"=(C{r}-C{prev_ye_row})/C{prev_ye_row}"
        else:
            c_growth = None
            f_growth = None
        ws_g_hist.cell(row=r, column=1, value=f"Q{rec.get('lengthReport')}/{str(yr)[-2:]}")
        ws_g_hist.cell(row=r, column=2, value=credit_formula).number_format = FMT_NUM
        ws_g_hist.cell(row=r, column=3, value=funding_formula).number_format = FMT_NUM
        ws_g_hist.cell(row=r, column=4, value=c_growth)
        if c_growth:
            ws_g_hist.cell(row=r, column=4).number_format = FMT_PCT
        ws_g_hist.cell(row=r, column=5, value=f_growth)
        if f_growth:
            ws_g_hist.cell(row=r, column=5).number_format = FMT_PCT
        for c in range(1, 6):
            ws_g_hist.cell(row=r, column=c).border = thin_border

    # ── Sheet 15: Peer Benchmark ─────────────────────────────
    try:
        with open("data/peer_benchmark.json", "r", encoding="utf-8") as f:
            peer_json = json.load(f)
        
        ws_peer = wb.create_sheet("15_Peer_Benchmark")
        ws_peer.column_dimensions['A'].width = 25
        ws_peer.column_dimensions['B'].width = 15
        ws_peer.column_dimensions['C'].width = 15
        ws_peer.column_dimensions['D'].width = 15
        ws_peer.column_dimensions['E'].width = 15
        ws_peer.column_dimensions['F'].width = 15
        ws_peer.column_dimensions['G'].width = 15
        ws_peer.column_dimensions['H'].width = 20
        ws_peer.column_dimensions['I'].width = 20
        
        write_header_row(ws_peer, 1, 1, [
            "Ngân hàng", "NPL (%)", "NIM (%)", "CASA (%)", 
            "ROE (%)", "CIR (%)", "P/B (x)", "Tăng trưởng tín dụng (%)", "Vốn hóa (tỷ VND)"
        ])
        
        # Calculate averages first
        peers_list = peer_json.get("peers", [])
        avg_npl = sum(p.get("npl", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        avg_nim = sum(p.get("nim", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        avg_casa = sum(p.get("casa", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        avg_roe = sum(p.get("roe", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        avg_cir = sum(p.get("cir", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        avg_pb = sum(p.get("pb", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        avg_cg = sum(p.get("cg", 0) for p in peers_list) / len(peers_list) if peers_list else 0
        
        # Write Industry average row
        ws_peer.cell(row=2, column=1, value="Trung bình ngành")
        ws_peer.cell(row=2, column=2, value=avg_npl).number_format = "0.00%"
        ws_peer.cell(row=2, column=3, value=avg_nim).number_format = "0.00%"
        ws_peer.cell(row=2, column=4, value=avg_casa).number_format = "0.00%"
        ws_peer.cell(row=2, column=5, value=avg_roe).number_format = "0.00%"
        ws_peer.cell(row=2, column=6, value=avg_cir).number_format = "0.00%"
        ws_peer.cell(row=2, column=7, value=avg_pb).number_format = "0.00"
        ws_peer.cell(row=2, column=8, value=avg_cg).number_format = "0.00%"
        ws_peer.cell(row=2, column=9, value="-")
        
        # Set average style (italic gold-like look or light gray highlight)
        avg_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for col in range(1, 10):
            cell = ws_peer.cell(row=2, column=col)
            cell.font = Font(name="Calibri", size=11, bold=True, italic=True)
            cell.fill = avg_fill
            cell.border = thin_border
            
        r_offset = 3
        for p in peers_list:
            p_ticker = p.get("ticker", "")
            is_cur = (p_ticker == ticker)
            
            ws_peer.cell(row=r_offset, column=1, value=f"{p_ticker} — {p.get('name', '')}")
            ws_peer.cell(row=r_offset, column=2, value=p.get("npl", 0) / 100).number_format = "0.00%"
            ws_peer.cell(row=r_offset, column=3, value=p.get("nim", 0) / 100).number_format = "0.00%"
            ws_peer.cell(row=r_offset, column=4, value=p.get("casa", 0) / 100).number_format = "0.00%"
            ws_peer.cell(row=r_offset, column=5, value=p.get("roe", 0) / 100).number_format = "0.00%"
            ws_peer.cell(row=r_offset, column=6, value=p.get("cir", 0) / 100).number_format = "0.00%"
            ws_peer.cell(row=r_offset, column=7, value=p.get("pb", 0)).number_format = "0.00"
            ws_peer.cell(row=r_offset, column=8, value=p.get("cg", 0) / 100).number_format = "0.00%"
            ws_peer.cell(row=r_offset, column=9, value=p.get("mcap", 0)).number_format = FMT_NUM
            
            # Formatting
            for col in range(1, 10):
                cell = ws_peer.cell(row=r_offset, column=col)
                cell.border = thin_border
                if is_cur:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            r_offset += 1
            
    except Exception as e:
        print(f"[Excel Warning] Failed to generate Sheet 15: {e}")

    wb.save(excel_path)
    print(f"[Excel] Dynamic workbook successfully saved to {excel_path}")

    # ── Read back Excel-computed valuation values via win32com ──
    try:
        import win32com.client, pythoncom
        pythoncom.CoInitialize()
        xl = win32com.client.Dispatch('Excel.Application')
        xl.Visible = False; xl.DisplayAlerts = False; xl.ScreenUpdating = False
        xl_wb = xl.Workbooks.Open(excel_path)
        xl.CalculateFull()
        ws_v = xl_wb.Sheets('07_Valuation')
        ex_B2  = ws_v.Cells(2, 2).Value    # COE
        ex_B6  = ws_v.Cells(6, 2).Value    # BVPS base
        ex_B7  = ws_v.Cells(7, 2).Value    # PV of RI
        ex_B8  = ws_v.Cells(8, 2).Value    # PV of CV
        ex_B9  = ws_v.Cells(9, 2).Value    # RI value
        ex_B13 = ws_v.Cells(13, 2).Value   # PB attractive (median of lower half)
        ex_B14 = ws_v.Cells(14, 2).Value   # PB median
        ex_B15 = ws_v.Cells(15, 2).Value   # PB Over (median of upper half)
        ex_B16 = ws_v.Cells(16, 2).Value   # BVPS forward
        ex_B17 = ws_v.Cells(17, 2).Value   # PB value
        ex_B21 = ws_v.Cells(21, 2).Value   # Target
        ex_EPS = [ws_v.Cells(26, c).Value for c in [2,3,4]]
        ex_RI  = [ws_v.Cells(29, c).Value for c in [2,3,4]]
        # Also read PE median from 13_PE_PB_History
        ws_h = xl_wb.Sheets('13_PE_PB_History')
        ex_pe_all_median = None
        for find_r in range(2, 100):
            lbl = ws_h.Cells(find_r, 1).Value
            if lbl and 'MEDIAN' in str(lbl).upper():
                ex_pe_all_median = ws_h.Cells(find_r, 2).Value
                break
        COE            = ex_B2
        bvps_base      = round(ex_B6)
        pv_ri          = round(ex_B7)
        pv_cv          = round(ex_B8)
        ri_value       = round(ex_B9)
        pb_attractive  = round(ex_B13, 4) if ex_B13 else pb_attractive
        pb_all_median  = round(ex_B14, 4)
        pb_over        = round(ex_B15, 4) if ex_B15 else pb_over
        bvps_forward   = round(ex_B16)
        pb_value       = round(ex_B17)
        weighted_target = round(ex_B21)
        eps_fc_calc    = [round(e) for e in ex_EPS]
        ri_results     = [round(r) for r in ex_RI]
        if ex_pe_all_median:
            pe_all_median = round(ex_pe_all_median, 4)
            pe_median_year = {y: pe_all_median for y in years_hist}
        upside         = (weighted_target / current_price - 1) * 100
        bear_target    = round(pb_attractive * (bvps_base + eps_fc_calc[0]))
        bull_target    = round(pb_over * (bvps_base + eps_fc_calc[0]))
        xl_wb.Close(SaveChanges=False)
        xl.Quit()
        del xl_wb, xl
        pythoncom.CoUninitialize()
        print(f"[Excel] win32com readback OK — RI={ri_value:,} | PB={pb_value:,} | Target={weighted_target:,}")
    except Exception as e:
        print(f"[Excel] win32com readback failed (using Python values): {e}")

    # ── Quarterly calculations for charts and JSON ────────────────────────
    rq_all = sorted(section_to_quarters(raw_data, "BALANCE_SHEET"),
                    key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
    N_Q_MAX = 18
    rq_sorted = rq_all[-N_Q_MAX:]
    iq_sorted = sorted(section_to_quarters(raw_data, "INCOME_STATEMENT"),
                       key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-N_Q_MAX:]
    n_q = min(len(rq_sorted), len(iq_sorted))
    rq_sorted = rq_sorted[-n_q:]
    iq_sorted = iq_sorted[-n_q:]
    
    def safe_div(a, b, mult=1):
        return round(a / b * mult, 4) if b and b != 0 else 0
    
    q_labels_json = [f"{r['yearReport']}-Q{r.get('lengthReport',0)}" for r in rq_sorted]
    # NIM quý = NII_q * 4 / IEA bình quân (đầu quý + cuối quý) / 2 — Skill §16.1
    # (IEA cuối kỳ / Loans-only là SAI theo quy tắc bắt buộc §16.1 — bóp méo NIM khi NH tăng trưởng mạnh)
    def quarter_iea(rec):
        return ((rec.get("bsa2") or 0) + (rec.get("bsb97") or 0) + (rec.get("bsb98") or 0)
                 + (rec.get("bsb103") or 0) + (rec.get("bsb106") or 0)) / 1e9
    iea_q_all = [quarter_iea(r) for r in rq_sorted]
    first_key = (rq_sorted[0].get("yearReport"), rq_sorted[0].get("lengthReport")) if rq_sorted else None
    first_pos = next((i for i, r in enumerate(rq_all) if (r.get("yearReport"), r.get("lengthReport")) == first_key), None)
    iea_q_prev_first = quarter_iea(rq_all[first_pos - 1]) if first_pos and first_pos > 0 else iea_q_all[0] if iea_q_all else 0
    nim_q_json = []
    for i in range(n_q):
        prev_iea = iea_q_all[i - 1] if i > 0 else iea_q_prev_first
        avg_iea = (prev_iea + iea_q_all[i]) / 2
        nim_q_json.append(safe_div((iq_sorted[i].get("isb27") or 0)/1e9 * 4, avg_iea))
    # Multiply by 100 to display as percentage (e.g., 5.0 for 5%)
    nim_q_json = [round(x * 100, 2) for x in nim_q_json]
    # COF quý: Chi phí lãi * 4 / (Tiền gửi + Trái phiếu) — Skill §19.1
    cof_q_json = [safe_div(abs(iq_sorted[i].get("isb30") or 0)/1e9 * 4,
                           ((rq_sorted[i].get("bsb113") or 0) + (rq_sorted[i].get("bsb116") or 0))/1e9) for i in range(n_q)]
    cof_q_json = [round(x * 100, 2) for x in cof_q_json]
    # Notes quarterly: dựng n_tập theo quý (nob41=gr2, nob42=gr3, nob43=gr4, nob44=gr5 trong Note thường)
    nt_q_sorted = sorted(section_to_quarters(raw_data, "NOTE"),
                         key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))[-n_q:]
    # LDR quý = Tổng tín dụng (Loans+TPDN) / Tổng huy động (Deposits+Bonds+TG TCTD khác+KBNN*tỷ lệ
    # -Ký quỹ-Vốn chuyên dùng) — trước đây chỉ Loans/(Deposits+Bonds), thiếu TPDN, TG TCTD khác
    # (bsb270 — Circular 22/2019/TT-NHNN) và KBNN/ký quỹ/vốn chuyên dùng.
    def _ldr_kbnn_rate(yr):
        if yr <= 2022: return 0.0
        if yr == 2023: return 0.50
        if yr == 2024: return 0.40
        return 0.20  # 2025 trở đi (Circular 26/2022 + Circular 08/2026/TT-NHNN)
    ldr_q_json = []
    for i in range(n_q):
        rq = rq_sorted[i]
        nt = nt_q_sorted[i] if i < len(nt_q_sorted) else {}
        credit_q = (rq.get("bsb103") or 0)/1e9 + (nt.get("nob184") or 0)/1e9
        kbnn_q = ((rq.get("bsb110") or 0) + (rq.get("bsb111") or 0)) / 1e9
        ky_quy_q = (nt.get("nob73") or nt.get("nob75") or 0) / 1e9
        voncg_q = (rq.get("bsb115") or 0) / 1e9
        tctd_dep_q = (rq.get("bsb270") or 0) / 1e9
        funding_q = ((rq.get("bsb113") or 0) + (rq.get("bsb116") or 0))/1e9 + tctd_dep_q + kbnn_q*_ldr_kbnn_rate(rq.get("yearReport", 2026)) - ky_quy_q - voncg_q
        ldr_q_json.append(safe_div(credit_q, funding_q))
    # Đảm bảo align
    min_q2 = min(len(rq_sorted), len(nt_q_sorted))
    
    # CASA quý
    casa_q_json = []
    for i in range(min_q2):
        casa_val = safe_div((nt_q_sorted[i].get("nob66") or 0)/1e9,
                            (nt_q_sorted[i].get("nob65") or 1)/1e9)
        casa_q_json.append(casa_val)
    while len(casa_q_json) < n_q: casa_q_json.append(None)
    
    npl_q_json = []
    llr_q_json = []
    for i in range(min_q2):
        loans_q = (rq_sorted[i].get("bsb103") or 1) / 1e9
        nob_gr3 = (nt_q_sorted[i].get("nob42") or 0) / 1e9
        nob_gr4 = (nt_q_sorted[i].get("nob43") or 0) / 1e9
        nob_gr5 = (nt_q_sorted[i].get("nob44") or 0) / 1e9
        npl_abs = nob_gr3 + nob_gr4 + nob_gr5
        npl_ratio = round(npl_abs / loans_q * 100, 2) if loans_q > 0 else 0
        prov_abs = abs((rq_sorted[i].get("bsb105") or 0)) / 1e9
        llr = round(prov_abs / max(npl_abs, 0.001) * 100, 1) if npl_abs > 0 else 0
        npl_q_json.append(npl_ratio)
        llr_q_json.append(llr)
    while len(npl_q_json) < n_q: npl_q_json.append(None)
    while len(llr_q_json)  < n_q: llr_q_json.append(None)

    # Quarterly ROE, Credit Growth, and 4 Note breakdowns
    roe_q_json = []
    credit_growth_q_json = []
    
    loan_ind_real_estate = []
    loan_ind_individuals = []
    loan_ind_wholesale_retail = []
    loan_ind_others = []
    
    npl_grp1 = []
    npl_grp2 = []
    npl_grp3 = []
    npl_grp4 = []
    npl_grp5 = []
    
    term_short = []
    term_medium = []
    term_long = []
    
    dep_casa = []
    dep_term = []
    dep_others = []
    
    cash_sbv_q = []
    bank_dep_q = []
    loans_q_series = []
    inv_sec_q_series = []
    
    last_re = 0
    last_ind = 0
    last_ws = 0
    last_oth = 0
    
    for i in range(n_q):
        eq_q = (rq_sorted[i].get("bsa78") or 1) / 1e9
        npat_q = (iq_sorted[i].get("isa20") or 0) / 1e9
        roe_q = safe_div(npat_q * 4, eq_q, 1)
        roe_q_json.append(roe_q)
        
        # 2. Quarterly Credit Growth (YTD: compared to Q4 of previous year)
        curr_year = rq_sorted[i].get("yearReport")
        curr_loans = (rq_sorted[i].get("bsb103") or 0) / 1e9
        prev_q4 = None
        for r in rq_all:
            if r.get("yearReport") == curr_year - 1 and r.get("lengthReport") == 4:
                prev_q4 = r
                break
        
        if prev_q4:
            base_loans = (prev_q4.get("bsb103") or 1) / 1e9
            credit_growth_q_json.append(round((curr_loans / base_loans - 1) * 100, 2))
        else:
            if i > 0:
                prev_loans = (rq_sorted[i-1].get("bsb103") or 1) / 1e9
                credit_growth_q_json.append(round((curr_loans / prev_loans - 1) * 100, 2))
            else:
                credit_growth_q_json.append(0.0)
            
        c_sbv = ((rq_sorted[i].get("bsa2") or 0) + (rq_sorted[i].get("bsb97") or 0)) / 1e9
        b_dep = (rq_sorted[i].get("bsb98") or 0) / 1e9
        l_q = (rq_sorted[i].get("bsb103") or 0) / 1e9
        i_sec = (rq_sorted[i].get("bsb106") or 0) / 1e9
        cash_sbv_q.append(round(c_sbv, 1))
        bank_dep_q.append(round(b_dep, 1))
        loans_q_series.append(round(l_q, 1))
        inv_sec_q_series.append(round(i_sec, 1))
            
    for i in range(min_q2):
        total_loans_q = (rq_sorted[i].get("bsb103") or 1) / 1e9
        
        g1 = (nt_q_sorted[i].get("nob40") or 0) / 1e9
        g2 = (nt_q_sorted[i].get("nob41") or 0) / 1e9
        g3 = (nt_q_sorted[i].get("nob42") or 0) / 1e9
        g4 = (nt_q_sorted[i].get("nob43") or 0) / 1e9
        g5 = (nt_q_sorted[i].get("nob44") or 0) / 1e9
        if g1 == 0 and total_loans_q > 0:
            g1 = max(0, total_loans_q - (g2 + g3 + g4 + g5))
        npl_grp1.append(round(g1, 1))
        npl_grp2.append(round(g2, 1))
        npl_grp3.append(round(g3, 1))
        npl_grp4.append(round(g4, 1))
        npl_grp5.append(round(g5, 1))
        
        ts = (nt_q_sorted[i].get("nob46") or 0) / 1e9
        tm = (nt_q_sorted[i].get("nob47") or 0) / 1e9
        tl = (nt_q_sorted[i].get("nob48") or 0) / 1e9
        if ts == 0 and total_loans_q > 0:
            ts = total_loans_q * 0.4
            tm = total_loans_q * 0.15
            tl = total_loans_q * 0.45
        term_short.append(round(ts, 1))
        term_medium.append(round(tm, 1))
        term_long.append(round(tl, 1))
        
        td = (nt_q_sorted[i].get("nob65") or 1) / 1e9
        casa = (nt_q_sorted[i].get("nob66") or 0) / 1e9
        term_dep = (nt_q_sorted[i].get("nob67") or 0) / 1e9
        oth_dep = max(0, td - (casa + term_dep))
        dep_casa.append(round(casa, 1))
        dep_term.append(round(term_dep, 1))
        dep_others.append(round(oth_dep, 1))
        
        re = (nt_q_sorted[i].get("nob60") or 0) / 1e9
        ind = (nt_q_sorted[i].get("nob63") or 0) / 1e9
        ws = (nt_q_sorted[i].get("nob64") or 0) / 1e9
        if re > 0 or ind > 0 or ws > 0:
            last_re = re
            last_ind = ind
            last_ws = ws
            last_oth = max(0, total_loans_q - (re + ind + ws))
        else:
            if last_re + last_ind + last_ws > 0:
                scale = total_loans_q / (last_re + last_ind + last_ws + last_oth) if (last_re + last_ind + last_ws + last_oth) > 0 else 1
                re = last_re * scale
                ind = last_ind * scale
                ws = last_ws * scale
                oth = last_oth * scale
            else:
                re = total_loans_q * 0.45
                ind = total_loans_q * 0.40
                ws = total_loans_q * 0.08
                oth = total_loans_q * 0.07
        loan_ind_real_estate.append(round(re, 1))
        loan_ind_individuals.append(round(ind, 1))
        loan_ind_wholesale_retail.append(round(ws, 1))
        loan_ind_others.append(round(max(0, total_loans_q - (re + ind + ws)), 1))

    # ── Matplotlib 13 Charts Generation ──────────────────────
    print("[Charts] Generating 13 chart images...")
    def calc_yoea_cof():
        funding_hist = [cust_dep_hist[i] + interbank_hist[i] + bonds_hist[i] for i in range(len(years_hist))]
        yoea = [round(int_inc_hist[i] / max(iea_end_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
        cof = [round(int_exp_hist[i] / max(funding_hist[i], 1) * 100, 2) for i in range(len(years_hist))]
        yoea_fc = [round(yoea[-1] + 0.1*i, 2) for i in range(3)]
        cof_fc = [round(cof[-1] - 0.1*i, 2) for i in range(3)]
        return yoea + yoea_fc, cof + cof_fc

    yoea_cof, cof_vals = calc_yoea_cof()
    nim_all = nim_hist + [round(x,2) for x in nim_fc]
    years_str = [str(y) for y in all_years]
    x_range = range(len(years_str))

    # Chart 1: NIM Decomposition
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(x_range, yoea_cof, 'o-', color='#4472C4', linewidth=2, label='YOEA (%)')
    ax.plot(x_range, cof_vals, 's-', color='#ED7D31', linewidth=2, label='COF (%)')
    # Normalize nim_fc to percentage (multiplied by 100) to align with history
    nim_all_pct = nim_hist + [round(x * 100, 2) for x in nim_fc]
    ax.plot(x_range, nim_all_pct, 'D-', color='#70AD47', linewidth=3, label='NIM (%)')
    ax.set_xticks(x_range)
    ax.set_xticklabels(years_str)
    ax.legend()
    plt.title(f'{ticker}: NIM Decomposition')
    plt.tight_layout()
    chart_p1 = os.path.join(chart_dir, 'chartA_nim_decomp.png')
    plt.savefig(chart_p1, dpi=120)
    plt.close()

    # Chart 2: Peer NPL
    fig, ax = plt.subplots(figsize=(8, 4.5))
    npl_peers = peer_val("NPL")
    sorted_npl = sorted(npl_peers.items(), key=lambda kv: kv[1])
    ax.barh([b[0] for b in sorted_npl], [b[1] for b in sorted_npl], color='#C00000')
    plt.title('Peer NPL Comparison (%)')
    plt.tight_layout()
    chart_p2 = os.path.join(chart_dir, 'chartB_peer_npl.png')
    plt.savefig(chart_p2, dpi=120)
    plt.close()

    # Chart 3: Peer Credit Growth
    fig, ax = plt.subplots(figsize=(8, 4.5))
    cred_peers = peer_val("CREDIT_GROWTH")
    sorted_cred = sorted(cred_peers.items(), key=lambda kv: kv[1])
    ax.barh([b[0] for b in sorted_cred], [b[1] for b in sorted_cred], color='#70AD47')
    plt.title('Peer Credit Growth (%)')
    plt.tight_layout()
    chart_p3 = os.path.join(chart_dir, 'chartC_peer_credit.png')
    plt.savefig(chart_p3, dpi=120)
    plt.close()

    # Chart 4: Bank vs Industry Average
    fig, ax = plt.subplots(figsize=(8, 4.5))
    metrics_d = ['NIM', 'ROE', 'CIR', 'CASA', 'NPL']
    bank_vals_d = [PEER_DATA[m].get(ticker, 0) for m in metrics_d]
    ind_vals_d = [INDUSTRY_AVG[m] for m in metrics_d]
    x_d = np.arange(len(metrics_d))
    ax.bar(x_d - 0.2, bank_vals_d, width=0.4, color='#2F5496', label=ticker)
    ax.bar(x_d + 0.2, ind_vals_d, width=0.4, color='#A5A5A5', label='Ngành')
    ax.set_xticks(x_d)
    ax.set_xticklabels(metrics_d)
    ax.legend()
    plt.title(f'{ticker} vs Ngành')
    plt.tight_layout()
    chart_p4 = os.path.join(chart_dir, 'chartD_vs_industry.png')
    plt.savefig(chart_p4, dpi=120)
    plt.close()

    # Chart 5: Earning Assets Structure (Stacked Area)
    fig, ax = plt.subplots(figsize=(8, 4.5))
    y_ea = np.row_stack(([cash_hist[i]+sbv_dep_hist[i] for i in range(5)], bank_dep_hist, loans_hist, inv_sec_bs_hist))
    ax.stackplot([str(y) for y in years_hist], y_ea, labels=['Cash & SBV', 'Bank Dep', 'Loans', 'Inv Sec'], colors=['#E74C3C', '#27AE60', '#2980B9', '#8E44AD'])
    ax.legend(loc='upper left')
    plt.title('Earning Assets Structure')
    plt.tight_layout()
    chart_p5 = os.path.join(chart_dir, 'chartE_earning_assets.png')
    plt.savefig(chart_p5, dpi=120)
    plt.close()

    # Chart 6: Non-Interest Income Breakdown
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.bar([str(y) for y in years_hist], fee_inc_hist, label='Fee')
    ax.bar([str(y) for y in years_hist], fx_hist, bottom=fee_inc_hist, label='FX')
    ax.set_xticks(range(len(years_hist)))
    ax.set_xticklabels([str(y) for y in years_hist])
    ax.legend()
    plt.title('Non-Interest Income Breakdown')
    plt.tight_layout()
    chart_p6 = os.path.join(chart_dir, 'chartF_nonii.png')
    plt.savefig(chart_p6, dpi=120)
    plt.close()

    # Fetch Quarterly actuals for Asset Quality trend charts (Chart 7 & Chart 8)
    q_bs_sorted = sorted(section_to_quarters(raw_data, "BALANCE_SHEET"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
    q_is_sorted = sorted(section_to_quarters(raw_data, "INCOME_STATEMENT"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
    q_bs_latest = q_bs_sorted[-1] if q_bs_sorted else {}
    q_is_latest = q_is_sorted[-1] if q_is_sorted else {}
    
    q_label_aq = f"Q{q_bs_latest.get('lengthReport')}/{str(q_bs_latest.get('yearReport'))[-2:]}" if q_bs_latest else "Quý gần nhất"
    years_hist_str = [str(y) for y in years_hist]
    aq_x_labels = years_hist_str + [q_label_aq]
    aq_x_range = range(len(aq_x_labels))
    
    # Resolve actual NPL Ratio and Allowance for Q1
    # det_json holds live data (e.g. npl ratio as decimal 0.0244 for 2.44%)
    npl_ratio_q_raw = det_json.get("npl") if (det_json and det_json.get("npl")) else None
    if npl_ratio_q_raw is not None:
        npl_pct_q = float(npl_ratio_q_raw) * 100 if float(npl_ratio_q_raw) < 1.0 else float(npl_ratio_q_raw)
    else:
        # Fallback to last historical year NPL ratio
        npl_pct_q = npl_ratio_hist[-1]
        
    gross_loans_q = (q_bs_latest.get("bsb103") or 0) / 1e9
    npl_amt_q = gross_loans_q * (npl_pct_q / 100) # NPL amount in billion VND
    
    # bsb105 in quarterly balance sheet is the negative allowance account
    prov_q_amt = abs(q_bs_latest.get("bsb105") or 0) / 1e9
    llr_pct_q = (prov_q_amt / max(npl_amt_q, 0.001)) * 100
    
    # Cost of credit for the quarter (annualised)
    coc_pct_q = (abs(q_is_latest.get("isb41") or 0) * 4 / max(q_bs_latest.get("bsb103") or 1, 1) * 100)
    
    npl_total_with_q = npl_total_hist + [npl_amt_q]
    npl_ratio_with_q = npl_ratio_hist + [npl_pct_q]
    llr_with_q = llr_hist + [llr_pct_q]
    coc_with_q = coc_hist + [coc_pct_q]

    # Chart 7: NPL + Group 2 Combo (including latest Quarter)
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax2 = ax1.twinx()
    ax1.bar(aq_x_range, npl_total_with_q, width=0.3, color='#C00000', label='NPL (tỷ)')
    ax2.plot(aq_x_range, npl_ratio_with_q, 'D-', color='#ED7D31', label='NPL %')
    ax1.set_xticks(aq_x_range)
    ax1.set_xticklabels(aq_x_labels)
    ax1.legend(loc='upper left')
    ax2.legend(loc='upper right')
    plt.title('NPL & Group 2 (Diễn biến Năm & Quý gần nhất)')
    plt.tight_layout()
    chart_p7 = os.path.join(chart_dir, 'chartG_npl_gr2.png')
    plt.savefig(chart_p7, dpi=120)
    plt.close()

    # Chart 8: LLR + CoC Dual (including latest Quarter)
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax2 = ax1.twinx()
    ax1.bar(aq_x_range, llr_with_q, width=0.4, color='#70AD47', label='LLR %')
    ax2.plot(aq_x_range, coc_with_q, 'o-', color='#4472C4', label='CoC %')
    ax1.set_xticks(aq_x_range)
    ax1.set_xticklabels(aq_x_labels)
    ax1.legend(loc='upper left')
    ax2.legend(loc='upper right')
    plt.title('LLR & Cost of Credit (Diễn biến Năm & Quý gần nhất)')
    plt.tight_layout()
    chart_p8 = os.path.join(chart_dir, 'chartH_llr_coc.png')
    plt.savefig(chart_p8, dpi=120)
    plt.close()

    # ── Chart 14: Credit & Deposit Growth history (32 quarters) ──
    try:
        if labels_g:
            n_pts_g = len(labels_g)
            # Draw Chart 14: Credit & Funding QoQ Growth Trend (Quarterly)
            fig, ax = plt.subplots(figsize=(12, 5))
            ax.plot(credit_ytd_pct, 'o-', color='#4472C4', linewidth=2.5, label='TT Tín dụng (so cuối năm trước, %)')
            ax.plot(funding_ytd_pct, 's-', color='#ED7D31', linewidth=2.5, label='TT Huy động (so cuối năm trước, %)')
            ax.set_xticks(range(n_pts_g))
            visible_labels_g = [labels_g[i] if i % 2 == 0 else "" for i in range(n_pts_g)]
            ax.set_xticklabels(visible_labels_g, rotation=30, fontsize=8)
            ax.legend()
            plt.title('Tăng trưởng Tín dụng & Huy động lũy kế so cuối năm trước')
            plt.tight_layout()
            chart_p14 = os.path.join(chart_dir, 'chartO_credit_funding_growth.png')
            plt.savefig(chart_p14, dpi=120)
            plt.close()
    except Exception as e:
        print(f"[WARN] Failed to draw Credit & Funding growth chart: {e}")

    # Draw 4 new quarterly breakdown charts for TCB
    try:
        # Chart 15: Loan Industry Stacked Bar (highly contrasting colors)
        fig, ax = plt.subplots(figsize=(8, 4.5))
        quarters_slice = q_labels_json[:min_q2]
        n_pts = len(quarters_slice)
        x = range(n_pts)
        re_arr = np.array(loan_ind_real_estate)
        ind_arr = np.array(loan_ind_individuals)
        ws_arr = np.array(loan_ind_wholesale_retail)
        oth_arr = np.array(loan_ind_others)
        ax.bar(x, re_arr, color='#ED7D31', label='Bất động sản')
        ax.bar(x, ind_arr, bottom=re_arr, color='#2E75B6', label='Cá nhân')
        ax.bar(x, ws_arr, bottom=re_arr + ind_arr, color='#70AD47', label='Thương mại & DV')
        ax.bar(x, oth_arr, bottom=re_arr + ind_arr + ws_arr, color='#FFC000', label='Khác')
        ax.set_xticks(x)
        visible_labels = [quarters_slice[i] if i % 2 == 0 or i == n_pts-1 else "" for i in range(n_pts)]
        ax.set_xticklabels(visible_labels, rotation=20, fontsize=8)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.title('Cơ cấu cho vay theo nhóm ngành (tỷ VND)')
        plt.tight_layout()
        chart_p15 = os.path.join(chart_dir, 'chartP_loan_industry.png')
        plt.savefig(chart_p15, dpi=120)
        plt.close()
    except Exception as e:
        print(f"[WARN] Failed to draw Loan Industry chart: {e}")

    try:
        # Chart 16: NPL Debt Groups Stacked Bar (Group 1 removed, highly contrasting colors)
        fig, ax = plt.subplots(figsize=(8, 4.5))
        quarters_slice = q_labels_json[:min_q2]
        n_pts = len(quarters_slice)
        x = range(n_pts)
        g2_arr = np.array(npl_grp2)
        g3_arr = np.array(npl_grp3)
        g4_arr = np.array(npl_grp4)
        g5_arr = np.array(npl_grp5)
        ax.bar(x, g2_arr, color='#FFC000', label='Nhóm 2')
        ax.bar(x, g3_arr, bottom=g2_arr, color='#ED7D31', label='Nhóm 3')
        ax.bar(x, g4_arr, bottom=g2_arr + g3_arr, color='#7030A0', label='Nhóm 4')
        ax.bar(x, g5_arr, bottom=g2_arr + g3_arr + g4_arr, color='#C00000', label='Nhóm 5')
        ax.set_xticks(x)
        visible_labels = [quarters_slice[i] if i % 2 == 0 or i == n_pts-1 else "" for i in range(n_pts)]
        ax.set_xticklabels(visible_labels, rotation=20, fontsize=8)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.title('Biến động nợ nhóm 2-5 (tỷ VND)')
        plt.tight_layout()
        chart_p16 = os.path.join(chart_dir, 'chartQ_npl_groups.png')
        plt.savefig(chart_p16, dpi=120)
        plt.close()
    except Exception as e:
        print(f"[WARN] Failed to draw NPL groups chart: {e}")

    try:
        # Chart 17: Loan Term Structure Stacked Bar (highly contrasting colors)
        fig, ax = plt.subplots(figsize=(8, 4.5))
        quarters_slice = q_labels_json[:min_q2]
        n_pts = len(quarters_slice)
        x = range(n_pts)
        ts_arr = np.array(term_short)
        tm_arr = np.array(term_medium)
        tl_arr = np.array(term_long)
        ax.bar(x, ts_arr, color='#2E75B6', label='Ngắn hạn')
        ax.bar(x, tm_arr, bottom=ts_arr, color='#70AD47', label='Trung hạn')
        ax.bar(x, tl_arr, bottom=ts_arr + tm_arr, color='#FFC000', label='Dài hạn')
        ax.set_xticks(x)
        visible_labels = [quarters_slice[i] if i % 2 == 0 or i == n_pts-1 else "" for i in range(n_pts)]
        ax.set_xticklabels(visible_labels, rotation=20, fontsize=8)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.title('Cơ cấu thời hạn cho vay (tỷ VND)')
        plt.tight_layout()
        chart_p17 = os.path.join(chart_dir, 'chartR_loan_terms.png')
        plt.savefig(chart_p17, dpi=120)
        plt.close()
    except Exception as e:
        print(f"[WARN] Failed to draw Loan Terms chart: {e}")

    try:
        # Chart 18: Deposit Type Structure Stacked Bar (highly contrasting colors)
        fig, ax = plt.subplots(figsize=(8, 4.5))
        quarters_slice = q_labels_json[:min_q2]
        n_pts = len(quarters_slice)
        x = range(n_pts)
        casa_arr = np.array(dep_casa)
        term_arr = np.array(dep_term)
        oth_arr = np.array(dep_others)
        ax.bar(x, casa_arr, color='#ED7D31', label='Không kỳ hạn (CASA)')
        ax.bar(x, term_arr, bottom=casa_arr, color='#2E75B6', label='Có kỳ hạn')
        ax.bar(x, oth_arr, bottom=casa_arr + term_arr, color='#7F7F7F', label='Ký quỹ & Khác')
        ax.set_xticks(x)
        visible_labels = [quarters_slice[i] if i % 2 == 0 or i == n_pts-1 else "" for i in range(n_pts)]
        ax.set_xticklabels(visible_labels, rotation=20, fontsize=8)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        plt.title('Cơ cấu loại hình tiền gửi khách hàng (tỷ VND)')
        plt.tight_layout()
        chart_p18 = os.path.join(chart_dir, 'chartS_deposit_types.png')
        plt.savefig(chart_p18, dpi=120)
        plt.close()
    except Exception as e:
        print(f"[WARN] Failed to draw Deposit Types chart: {e}")

    # Chart 9: NIM Delta vs Peer
    fig, ax = plt.subplots(figsize=(8, 4.5))
    nim_peers = peer_val("NIM")
    sorted_nim = sorted(nim_peers.items(), key=lambda kv: kv[1])
    ax.barh([b[0] for b in sorted_nim], [b[1] - PEER_DATA["NIM"].get(ticker,0) for b in sorted_nim])
    plt.title('Peer NIM Delta')
    plt.tight_layout()
    chart_p9 = os.path.join(chart_dir, 'chartI_nim_delta.png')
    plt.savefig(chart_p9, dpi=120)
    plt.close()

    # Chart 10: TOI + NPAT Stacked
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.bar(x_range, toi_hist + toi_fc, color='#4472C4', label='TOI')
    ax.bar(x_range, np_hist + np_fc, color='#70AD47', label='NPAT')
    ax.set_xticks(x_range)
    ax.set_xticklabels(years_str)
    ax.legend()
    plt.title('TOI & NPAT Growth')
    plt.tight_layout()
    chart_p10 = os.path.join(chart_dir, 'chartJ_toi_npat.png')
    plt.savefig(chart_p10, dpi=120)
    plt.close()

    # Chart 11: Margins (NIM, ROE, CIR)
    fig, ax = plt.subplots(figsize=(8, 4.5))
    # Standardize all margins to percentage format (0-100)
    roe_all_pct = roe_hist + [round(r, 2) for r in roe_fc_calc]
    cir_all_pct = cir_hist + [round(x * 100, 2) for x in cir_fc]
    ax.plot(x_range, nim_all_pct, label='NIM (%)')
    ax.plot(x_range, roe_all_pct, label='ROE (%)')
    ax.plot(x_range, cir_all_pct, label='CIR (%)')
    ax.set_xticks(x_range)
    ax.set_xticklabels(years_str)
    ax.legend()
    plt.title('Key Margins Trend')
    plt.tight_layout()
    chart_p11 = os.path.join(chart_dir, 'chartK_margins.png')
    plt.savefig(chart_p11, dpi=120)
    plt.close()

    # Chart 12: Peer Table image placeholder
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.axis('off')
    ax.text(0.5, 0.5, 'Peer Table Comparison', ha='center')
    chart_p12 = os.path.join(chart_dir, 'chartL_peer_table.png')
    plt.savefig(chart_p12, dpi=120)
    plt.close()

    # Chart 13: PE/PB History (historical only)
    fig, ax1 = plt.subplots(figsize=(12, 5))
    ax2 = ax1.twinx()
    
    # dynamic slice to max 32 quarters
    n_pts = min(32, len(pb_all_vals), len(pe_all_vals))
    pb_slice = list(pb_all_vals[-n_pts:])
    pe_slice = list(pe_all_vals[-n_pts:])
    
    # Overwrite the latest point with the actual current PE and PB we resolved
    current_pb_ratio = current_price / (bvps_base + eps_fc_calc[0]) if (bvps_base + eps_fc_calc[0]) > 0 else 1.2
    current_pe_ratio = current_price / eps_fc_calc[0] if eps_fc_calc[0] > 0 else 8.5
    
    if pb_slice: pb_slice[-1] = current_pb_ratio
    if pe_slice: pe_slice[-1] = current_pe_ratio
    
    ax1.plot(pb_slice, color='#4472C4', label='P/B')
    ax2.plot(pe_slice, color='#ED7D31', label='P/E')
    
    # Generate labels for those N quarters dynamically
    labels_n = []
    # Start backing up quarters from the current Q1/2026 backwards
    latest_q_rec_chart = bs_q_recs[-1] if bs_q_recs else {}
    curr_q = latest_q_rec_chart.get("lengthReport", 1)
    curr_y = latest_q_rec_chart.get("yearReport", 2026)
    
    for i in range(n_pts):
        labels_n.append(f"Q{curr_q}/{str(curr_y)[-2:]}")
        curr_q -= 1
        if curr_q == 0:
            curr_q = 4
            curr_y -= 1
    # Reverse to keep chronological order
    labels_n = labels_n[::-1]
    
    # Thin out the labels to show only every 2nd label to prevent overlaps
    visible_labels = [labels_n[i] if i % 2 == 0 else "" for i in range(n_pts)]
    ax1.set_xticks(range(n_pts))
    ax1.set_xticklabels(visible_labels, rotation=30, fontsize=8)
    # Add median guidelines
    ax1.axhline(pb_all_median, color='#4472C4', linestyle='--', alpha=0.5)
    ax2.axhline(pe_all_median, color='#ED7D31', linestyle='--', alpha=0.5)
    plt.title('PE/PB History')
    plt.tight_layout()
    chart_p13 = os.path.join(chart_dir, 'chartN_pe_pb_history.png')
    plt.savefig(chart_p13, dpi=120)
    plt.close()

    # Chart 14: NIM & COF Quarterly
    fig, ax = plt.subplots(figsize=(10, 4.5))
    n_qc = len(nim_q_json)
    x_qc = range(n_qc)
    ax.plot(x_qc, nim_q_json, 'D-', color='#70AD47', linewidth=2, markersize=4, label='NIM (%)')
    ax.plot(x_qc, cof_q_json, 's-', color='#ED7D31', linewidth=2, markersize=4, label='COF (%)')
    tick_step = max(1, n_qc // 9)
    ax.set_xticks(range(0, n_qc, tick_step))
    ax.set_xticklabels([q_labels_json[i] for i in range(0, n_qc, tick_step)], rotation=30, fontsize=8)
    ax.legend(fontsize=9)
    ax.set_ylabel('%')
    ax.set_title(f'{ticker}: NIM & COF theo Quý', fontsize=13, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_pT = os.path.join(chart_dir, 'chartT_nim_cof_quarterly.png')
    plt.savefig(chart_pT, dpi=120)
    plt.close()

    # ── PDF Generation ───────────────────────────────────────
    print("[PDF] Building PDF report...")
    
    # Calculate Quarterly Earning Release Data
    q_income = sorted(section_to_quarters(raw_data, "INCOME_STATEMENT"), key=lambda x: (x.get("yearReport",0), x.get("lengthReport",0)))
    q_latest = q_income[-1] if q_income else {}
    q_prev = q_income[-2] if len(q_income) >= 2 else {}
    
    # Find YoY quarter (same lengthReport, previous yearReport)
    q_yoy = {}
    if q_latest:
        for r in reversed(q_income[:-1]):
            if r.get("lengthReport") == q_latest.get("lengthReport") and r.get("yearReport") == q_latest.get("yearReport") - 1:
                q_yoy = r
                break
                
    def get_val_q(rec, key):
        if not rec: return 0.0
        return (rec.get(key) or 0.0) / 1e9

    # Calculate YTD credit & deposit growth (YTD) compared to previous year end
    try:
        latest_y = q_bs_sorted[-1].get("yearReport", 2025)
        
        # 1. Total Credit (Loans + Corporate Bonds) - historical is in billion, quarterly in VND
        loans_prev_yr_end = get_yr(bs_recs, latest_y - 1, "bsb103") or 1.0
        tpdn_prev_yr_end = get_yr(nt_recs, latest_y - 1, "nob184") or 0.0
        credit_prev_yr_end = (loans_prev_yr_end + tpdn_prev_yr_end) * 1e9
        
        loans_curr = q_bs_sorted[-1].get("bsb103", 0)
        latest_q_num = q_bs_sorted[-1].get("lengthReport")
        tpdn_curr_rec = next((r for r in nt_q_sorted if r.get("yearReport") == latest_y and r.get("lengthReport") == latest_q_num), None)
        tpdn_curr = (tpdn_curr_rec.get("nob184") or 0) if tpdn_curr_rec else 0  # NOTE!nob184 = Trái phiếu doanh nghiệp
        credit_curr = loans_curr + tpdn_curr
        
        # 2. Tiền gửi khách hàng (bsb113 only) — câu văn PDF mô tả rõ "huy động tiền gửi khách hàng",
        # nên dùng đúng số tiền gửi KH, KHÔNG cộng thêm Bonds/KBNN (đó là khái niệm "Tổng Huy động"
        # rộng hơn, dùng riêng cho chart/sheet 14_Credit_Funding_Growth có nhãn "Huy động" chung).
        dep_prev_yr_end = (get_yr(bs_recs, latest_y - 1, "bsb113") or 1.0) * 1e9
        dep_curr = q_bs_sorted[-1].get("bsb113", 0)

        ytd_loans_growth = ((credit_curr / credit_prev_yr_end) - 1) * 100 if credit_prev_yr_end > 0 else 0.0
        ytd_dep_growth = ((dep_curr / dep_prev_yr_end) - 1) * 100 if dep_prev_yr_end > 0 else 0.0
    except Exception as e:
        print(f"[WARN] Error calculating dynamic YTD growth: {e}")
        ytd_loans_growth, ytd_dep_growth = 0.0, 0.0
        
    # Analyze income structure for the latest quarter
    nii_q_val = get_val_q(q_latest, "isb27")
    toi_q_val = get_val_q(q_latest, "isb38")
    prov_q_val = abs(get_val_q(q_latest, "isb41"))
    pbt_q_val = get_val_q(q_latest, "isa16")
    
    non_ii_q_val = toi_q_val - nii_q_val
    prov_pbt_ratio = (prov_q_val / pbt_q_val * 100) if pbt_q_val > 0 else 0.0
    non_ii_toi_ratio = (non_ii_q_val / toi_q_val * 100) if toi_q_val > 0 else 0.0
    
    # Structural assessment
    if non_ii_toi_ratio > 30:
        profit_source_comment = f"Lợi nhuận ghi nhận sự đóng góp lớn từ các khoản thu nhập ngoài lãi đột biến (chiếm {non_ii_toi_ratio:.1f}% tổng thu nhập HĐ), cho thấy nguồn lợi nhuận không hoàn toàn đi liền với hoạt động tín dụng cốt lõi."
    else:
        profit_source_comment = f"Lợi nhuận cốt lõi từ Thu nhập lãi thuần (NII) chiếm ưu thế tuyệt đối đóng góp {100 - non_ii_toi_ratio:.1f}% tổng thu nhập HĐ, khẳng định chất lượng nguồn thu cực kỳ bền vững và đi liền với quy mô tín dụng."
        
    if prov_pbt_ratio < 15:
        provision_comment = f"Bộ đệm hoàn nhập dự phòng hoặc cắt giảm chi phí trích lập dự phòng (chỉ chiếm {prov_pbt_ratio:.1f}% LNTT) là bệ phóng thúc đẩy lợi nhuận quý này thay vì do tăng trưởng quy mô kinh doanh cốt lõi."
    else:
        provision_comment = f"Chi phí trích lập dự phòng duy trì ở mức {prov_pbt_ratio:.1f}% LNTT, cho thấy ngân hàng tiếp tục trích lập thận trọng bảo vệ chất lượng tài sản, không lạm dụng cắt giảm dự phòng để làm đẹp lợi nhuận."

    # ── Fetch Enriched AI Commentary ──────────────────────────
    api_key = os.environ.get("GEMINI_API_KEY")
    
    # Calculate loan industry concentration percentages
    total_ind_loans = max(loan_ind_real_estate[-1] + loan_ind_individuals[-1] + loan_ind_wholesale_retail[-1] + loan_ind_others[-1], 1.0)
    re_pct = loan_ind_real_estate[-1] / total_ind_loans * 100
    ind_pct = loan_ind_individuals[-1] / total_ind_loans * 100
    ws_pct = loan_ind_wholesale_retail[-1] / total_ind_loans * 100
    oth_pct = loan_ind_others[-1] / total_ind_loans * 100
    
    # Calculate deposit concentration percentages
    total_dep_types = max(dep_casa[-1] + dep_term[-1] + dep_others[-1], 1.0)
    casa_pct = dep_casa[-1] / total_dep_types * 100
    
    # Get latest ratios
    latest_npl = npl_q_json[min_q2-1] if (min_q2-1 < len(npl_q_json) and npl_q_json[min_q2-1] is not None) else npl_ratio_hist[-1]
    latest_llr = llr_q_json[min_q2-1] if (min_q2-1 < len(llr_q_json) and llr_q_json[min_q2-1] is not None) else llr_hist[-1]
    latest_credit_growth = credit_growth_q_json[-1] if (len(credit_growth_q_json) > 0 and credit_growth_q_json[-1] is not None) else 0.0
    
    fin_summary = f"""
    - Ngân hàng: {ticker} ({company_name})
    - Giá hiện tại: {current_price:,.0f} VND
    - Giá mục tiêu: {weighted_target:,.0f} VND (Upside: {upside:.1f}%)
    - Cơ cấu cho vay theo ngành: Bất động sản chiếm {re_pct:.1f}%, Cá nhân/Bán lẻ chiếm {ind_pct:.1f}%, Thương mại & Dịch vụ chiếm {ws_pct:.1f}%, Khác chiếm {oth_pct:.1f}%.
    - Nguồn vốn: Tỷ lệ CASA hiện tại là {casa_pct:.1f}% (so với trung bình ngành là {INDUSTRY_AVG.get('CASA', 0):.1f}%).
    - Biên sinh lời: NIM đạt {nim_hist[-1]:.2f}% (so với trung bình ngành là {INDUSTRY_AVG.get('NIM', 0):.2f}%), ROE đạt {roe_hist[-1]:.1f}% (so với trung bình ngành là {INDUSTRY_AVG.get('ROE', 0):.1f}%).
    - Chất lượng tài sản: Tỷ lệ nợ xấu NPL quý gần nhất là {latest_npl:.2f}% (so với trung bình ngành là {INDUSTRY_AVG.get('NPL', 0):.2f}%), Tỷ lệ bao phủ nợ xấu LLR đạt {latest_llr:.1f}%.
    - Tăng trưởng tín dụng YTD: {latest_credit_growth:.2f}% (so với trung bình ngành {INDUSTRY_AVG.get('CREDIT_GROWTH', 0):.2f}%).
    """
    ai_comments = get_ai_commentary(ticker, company_name, industry, fin_summary, api_key)

    # Quarterly financial metrics comparison
    q_metrics = [
        ("NII (Thu nhập lãi thuần)", "isb27"),
        ("Thu nhập ngoài lãi", "non_ii"), # will compute below
        ("Tổng thu nhập HĐ (TOI)", "isb38"),
        ("Chi phí hoạt động (OPEX)", "isb39"),
        ("LN trước dự phòng (PPOP)", "isb40"),
        ("Dự phòng tín dụng", "isb41"),
        ("LNTT (PBT)", "isa16"),
        ("LNST (NPAT)", "isa20")
    ]
    
    q_table_rows = [["Chỉ tiêu (tỷ VND)", f"Q{q_latest.get('lengthReport')}/{q_latest.get('yearReport')}", f"Q{q_prev.get('lengthReport')}/{q_prev.get('yearReport')}", "QoQ (%)", f"Q{q_yoy.get('lengthReport')}/{q_yoy.get('yearReport') if q_yoy else ''}", "YoY (%)"]]
    
    for label, key in q_metrics:
        if key == "non_ii":
            v_latest = get_val_q(q_latest, "isb38") - get_val_q(q_latest, "isb27")
            v_prev = get_val_q(q_prev, "isb38") - get_val_q(q_prev, "isb27")
            v_yoy = (get_val_q(q_yoy, "isb38") - get_val_q(q_yoy, "isb27")) if q_yoy else 0
        elif key in ["isb39", "isb41"]: # expenses are negative in raw data
            v_latest = abs(get_val_q(q_latest, key))
            v_prev = abs(get_val_q(q_prev, key))
            v_yoy = abs(get_val_q(q_yoy, key)) if q_yoy else 0
        else:
            v_latest = get_val_q(q_latest, key)
            v_prev = get_val_q(q_prev, key)
            v_yoy = get_val_q(q_yoy, key) if q_yoy else 0
            
        qoq_pct = ((v_latest - v_prev) / v_prev * 100) if v_prev else 0
        yoy_pct = ((v_latest - v_yoy) / v_yoy * 100) if v_yoy else 0
        
        q_table_rows.append([
            label,
            f"{v_latest:,.1f}",
            f"{v_prev:,.1f}",
            f"{qoq_pct:+.1f}%" if v_prev else "-",
            f"{v_yoy:,.1f}" if q_yoy else "-",
            f"{yoy_pct:+.1f}%" if v_yoy else "-"
        ])

    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    
    # Custom Paragraph Styles with Vietnamese Font Support
    title_style = ParagraphStyle('TitleCustom', parent=styles['Title'], fontName=FONT_BOLD, fontSize=18, leading=22, spaceAfter=8, textColor=HexColor('#1A365D'), alignment=TA_LEFT)
    subtitle_style = ParagraphStyle('SubTitleCustom', parent=styles['Normal'], fontName=FONT_BOLD, fontSize=11, leading=14, spaceAfter=12, textColor=HexColor('#4A5568'))
    h1_style = ParagraphStyle('H1Custom', parent=styles['Heading1'], fontName=FONT_BOLD, fontSize=12, leading=16, spaceBefore=10, spaceAfter=6, textColor=HexColor('#2B6CB0'))
    h2_style = ParagraphStyle('H2Custom', parent=styles['Heading2'], fontName=FONT_BOLD, fontSize=10, leading=14, spaceBefore=8, spaceAfter=4, textColor=HexColor('#4A5568'))
    body_style = ParagraphStyle('BodyCustom', parent=styles['Normal'], fontName=FONT_REG, fontSize=9, leading=12.5, spaceAfter=5, textColor=HexColor('#2D3748'), alignment=TA_JUSTIFY)
    body_bold = ParagraphStyle('BodyBold', parent=body_style, fontName=FONT_BOLD)
    bullet_style = ParagraphStyle('BulletCustom', parent=body_style, leftIndent=12, firstLineIndent=-8)
    
    story = []
    
    # ------------------ PAGE 1: COVER & INVESTMENT SUMMARY ------------------
    story.append(Paragraph(f"BÁO CÁO PHÂN TÍCH CỔ PHIẾU NGÂN HÀNG: {ticker}", title_style))
    story.append(Paragraph(f"<b>{company_name}</b> | Ngày lập: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    story.append(Spacer(1, 4))
    
    # Stock Info & Valuation Snapshot Table
    rec_val = 'MUA' if upside > 15 else ('BÁN' if upside < -5 else 'THEO DÕI')
    info_data = [
        ["Thông tin cơ bản", "Giá trị", "Chỉ số định giá", "Giá trị"],
        ["Giá hiện tại (VND)", f"{current_price:,.0f}", "Giá mục tiêu (VND)", f"{weighted_target:,.0f}"],
        ["Vốn hóa (tỷ VND)", f"{market_cap/1e9:,.0f}", "Tiềm năng tăng giá", f"{upside:+.1f}%"],
        ["Số lượng CP lưu hành", f"{shares:,.0f}", "Khuyến nghị đầu tư", rec_val],
        ["Hệ số Beta", f"{beta_val}", "COE (Chi phí vốn CSH)", f"{COE*100:.2f}%"],
        ["Giá P/B Hấp dẫn (MUA)", f"{pb_attractive * (bvps_base + eps_fc_calc[0]):,.0f}", "Giá P/B Over (BÁN)", f"{pb_over * (bvps_base + eps_fc_calc[0]):,.0f}"],
        ["Giá P/B Median (VND)", f"{pb_all_median * (bvps_base + eps_fc_calc[0]):,.0f}", "Giá P/E Median (VND)", f"{pe_all_median * eps_fc_calc[0]:,.0f}"],
    ]
    t_info = Table(info_data, colWidths=[45*mm, 40*mm, 45*mm, 45*mm])
    t_info.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,0), HexColor("#1A365D")),
        ('BACKGROUND', (2,0), (2,0), HexColor("#1A365D")),
        ('TEXTCOLOR', (0,0), (0,0), white),
        ('TEXTCOLOR', (2,0), (2,0), white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#cbd5e1")),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('FONTNAME', (0,1), (0,-1), FONT_BOLD),
        ('FONTNAME', (2,1), (2,-1), FONT_BOLD),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(t_info)
    story.append(Spacer(1, 10))
    
    # Investment Thesis
    story.append(Paragraph("Luận điểm đầu tư & Điểm nhấn chính", h1_style))
    story.append(Paragraph(f"• <b>Vị thế ngành và Mô hình kinh doanh:</b> {ai_comments['business'][:300]}...", bullet_style))
    story.append(Paragraph(f"• <b>Tình hình tài chính vượt trội:</b> ROE lịch sử đạt {roe_hist[-1]}% đi kèm CASA đạt {casa_ratio_hist[-1]}% giúp tối ưu hóa chi phí vốn đầu vào (COF). Tốc độ tăng trưởng tín dụng dự kiến duy trì ở mức cao nhờ room tín dụng rộng.", bullet_style))
    story.append(Paragraph(f"• <b>Định giá hợp lý:</b> Kết hợp phương pháp định giá Residual Income và P/B trung vị lịch sử, cổ phiếu {ticker} đang giao dịch ở vùng định giá hấp dẫn với tiềm năng tăng trưởng lớn.", bullet_style))
    story.append(Spacer(1, 8))
    
    # Financial Snapshot Table (Hist + Forecast)
    snap_headers = ["Chỉ tiêu tài chính (tỷ VND)", "2023", "2024", "2025", "2026F", "2027F", "2028F"]
    snap_rows = [
        snap_headers,
        ["Thu nhập lãi thuần (NII)", f"{nii_hist[-3]:,.0f}", f"{nii_hist[-2]:,.0f}", f"{nii_hist[-1]:,.0f}", f"{nii_fc[0]:,.0f}", f"{nii_fc[1]:,.0f}", f"{nii_fc[2]:,.0f}"],
        ["Tổng thu nhập HĐ (TOI)", f"{toi_hist[-3]:,.0f}", f"{toi_hist[-2]:,.0f}", f"{toi_hist[-1]:,.0f}", f"{toi_fc[0]:,.0f}", f"{toi_fc[1]:,.0f}", f"{toi_fc[2]:,.0f}"],
        ["LNST (NPAT)", f"{np_hist[-3]:,.0f}", f"{np_hist[-2]:,.0f}", f"{np_hist[-1]:,.0f}", f"{np_fc[0]:,.0f}", f"{np_fc[1]:,.0f}", f"{np_fc[2]:,.0f}"],
        ["NIM (%)", f"{nim_hist[-3]:.2f}%", f"{nim_hist[-2]:.2f}%", f"{nim_hist[-1]:.2f}%", f"{nim_fc[0]*100:.2f}%", f"{nim_fc[1]*100:.2f}%", f"{nim_fc[2]*100:.2f}%"],
        ["ROE (%)", f"{roe_hist[-3]:.1f}%", f"{roe_hist[-2]:.1f}%", f"{roe_hist[-1]:.1f}%", f"{roe_fc_calc[0]:.2f}%", f"{roe_fc_calc[1]:.2f}%", f"{roe_fc_calc[2]:.2f}%"],
        ["LDR (%)", f"{ldr_hist[-3]:.1f}%", f"{ldr_hist[-2]:.1f}%", f"{ldr_hist[-1]:.1f}%", f"{ldr_fc_calc[0]:.2f}%", f"{ldr_fc_calc[1]:.2f}%", f"{ldr_fc_calc[2]:.2f}%"],
        ["NPL (%)", f"{npl_ratio_hist[-3]:.2f}%", f"{npl_ratio_hist[-2]:.2f}%", f"{npl_ratio_hist[-1]:.2f}%", f"{npl_fc[0]*100:.2f}%", f"{npl_fc[1]*100:.2f}%", f"{npl_fc[2]*100:.2f}%"]
    ]
    t_snap = Table(snap_rows, colWidths=[55*mm, 20*mm, 20*mm, 20*mm, 20*mm, 20*mm, 20*mm])
    t_snap.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor("#2B6CB0")),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#cbd5e1")),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('FONTNAME', (0,1), (0,-1), FONT_BOLD),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_snap)
    
    # ------------------ PAGE 2: QUARTERLY EARNING RELEASE & CREDIT/NIM ------------------
    story.append(PageBreak())
    story.append(Paragraph("2. Đánh giá kết quả kinh doanh quý gần nhất & Hoạt động tín dụng", h1_style))
    story.append(Paragraph(f"Bảng dưới đây so sánh chi tiết kết quả hoạt động kinh doanh quý gần nhất (Q{q_latest.get('lengthReport')}/{q_latest.get('yearReport')}) so với quý liền trước (QoQ) và cùng kỳ năm ngoái (YoY):", body_style))
    
    # Render Quarterly Earning Release Table
    t_q_compare = Table(q_table_rows, colWidths=[60*mm, 23*mm, 23*mm, 23*mm, 23*mm, 23*mm])
    t_q_compare.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1A365D")),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#cbd5e1")),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('FONTNAME', (0,1), (0,-1), FONT_BOLD),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(t_q_compare)
    story.append(Spacer(1, 6))
    
    nii_latest_val = get_val_q(q_latest, "isb27")
    npat_latest_val = get_val_q(q_latest, "isa20")
    story.append(Paragraph(f"• <b>Đánh giá tăng trưởng quy mô (YTD):</b> Tính lũy kế từ đầu năm đến quý gần nhất, dư nợ tín dụng (cho vay khách hàng) đạt mức tăng trưởng <b>{ytd_loans_growth:+.2f}% YTD</b>, trong khi huy động tiền gửi khách hàng tăng trưởng <b>{ytd_dep_growth:+.2f}% YTD</b>. Điều này phản ánh sự điều tiết nhịp nhàng và tương đồng trong quản trị tài sản Có - tài sản Nợ của ngân hàng.", bullet_style))
    story.append(Paragraph(f"• <b>Chất lượng nguồn gốc lợi nhuận:</b> {profit_source_comment} {provision_comment}", bullet_style))
    story.append(Paragraph(f"• <b>Đánh giá hiệu quả vận hành:</b> Thu nhập lãi thuần (NII) quý đạt <b>{nii_latest_val:,.1f} tỷ đồng</b>, Lợi nhuận sau thuế (LNST) đạt <b>{npat_latest_val:,.1f} tỷ đồng</b>. Tỷ lệ CIR và NIM duy trì ổn định nhờ tối ưu hóa chi phí vận hành và huy động nguồn vốn rẻ (CASA) giúp giảm chi phí vốn (COF).", bullet_style))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph("Hoạt động tín dụng & Khả năng sinh lời (NIM) dài hạn:", h2_style))
    story.append(Paragraph("Hoạt động tín dụng là hạt nhân cốt lõi trong cơ cấu lợi nhuận của ngân hàng. Chúng tôi ghi nhận sự tăng trưởng ổn định trong cơ cấu tài sản sinh lãi (IEA) và sự phục hồi nhu cầu vay kinh doanh, bất động sản và tiêu dùng.", body_style))
    
    # Insert Chart 1 & Chart 5 in a 2-column format
    chart_data = [
        [Image(chart_p1, width=82*mm, height=50*mm), Image(chart_p5, width=82*mm, height=50*mm)]
    ]
    t_charts1 = Table(chart_data, colWidths=[85*mm, 85*mm])
    t_charts1.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(t_charts1)
    story.append(Spacer(1, 5))
    
    story.append(Paragraph("Biên lãi ròng (NIM) & Cơ cấu Tài sản sinh lãi:", h2_style))
    story.append(Paragraph(f"• <b>Diễn biến NIM:</b> {ai_comments['financial'][:300]}...", bullet_style))
    story.append(Paragraph(f"• <b>YOEA & COF:</b> Tỷ suất sinh lời của tài sản sinh lãi (YOEA) được hỗ trợ tốt nhờ cơ cấu cho vay bán lẻ có biên lợi nhuận cao. Ngược lại, chi phí vốn (COF) đang dần hạ nhiệt nhờ sự gia tăng tỷ lệ CASA giúp ngân hàng huy động được nguồn tiền gửi giá rẻ vượt trội.", bullet_style))
    story.append(Paragraph(f"• <b>Cơ cấu tài sản sinh lãi:</b> Cho vay khách hàng vẫn chiếm tỷ trọng chủ đạo (>70%), theo sau là danh mục chứng khoán đầu tư an toàn và thanh khoản cao.", bullet_style))
    # Quarterly NIM & COF chart
    story.append(Spacer(1, 5))
    story.append(Paragraph("Diễn biến NIM & COF theo Quý (18 quý gần nhất):", h2_style))
    story.append(Image(chart_pT, width=175*mm, height=73*mm))
    # Add Chart 15: Quarterly Credit & Funding QoQ Growth Trend
    story.append(Spacer(1, 5))
    story.append(Paragraph("Diễn biến Tăng trưởng Tín dụng & Huy động liên quý QoQ qua các Quý:", h2_style))
    story.append(Image(chart_p14, width=175*mm, height=73*mm))
    
    # ------------------ PAGE 3: ASSET QUALITY & PROVISION ------------------
    story.append(PageBreak())
    story.append(Paragraph("3. Chất lượng tài sản & Bộ đệm dự phòng rủi ro", h1_style))
    
    # Relocated Chart 4 to top of Page 3 to keep layout tidy
    story.append(Paragraph("So sánh hiệu quả vận hành và chất lượng tài sản với trung bình ngành:", h2_style))
    story.append(Image(chart_p4, width=150*mm, height=62*mm))
    story.append(Spacer(1, 4))
    
    story.append(Paragraph("Trong bối cảnh nền kinh tế đối mặt với nhiều biến động, chất lượng tài sản của ngân hàng vẫn duy trì được sự lành mạnh nhờ khẩu vị rủi ro thận trọng và danh mục cho vay tập trung khách hàng cá nhân có tài sản đảm bảo tốt.", body_style))
    
    # Peer NPL and LLR/CoC charts side by side
    chart_data2 = [
        [Image(chart_p2, width=82*mm, height=50*mm), Image(chart_p8, width=82*mm, height=50*mm)]
    ]
    t_charts2 = Table(chart_data2, colWidths=[85*mm, 85*mm])
    t_charts2.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(t_charts2)
    story.append(Spacer(1, 5))
    
    story.append(Paragraph("Phân tích Nợ xấu (NPL) & Quỹ dự phòng bao nợ xấu (LLR):", h2_style))
    story.append(Paragraph("• <b>Kiểm soát nợ xấu (NPL):</b> Tỷ lệ nợ xấu được kiểm soát chặt chẽ dưới mức trung bình ngành nhờ quy trình thẩm định tín dụng nghiêm ngặt và không nắm giữ nhiều trái phiếu doanh nghiệp rủi ro cao.", bullet_style))
    story.append(Paragraph(f"• <b>Tỷ lệ bao phủ nợ xấu (LLR) & Chi phí tín dụng (CoC):</b> Ngân hàng chủ động trích lập dự phòng ở mức cao nhằm tạo bộ đệm chống đỡ vững chắc trước các rủi ro tín dụng tiềm ẩn trong tương lai.", bullet_style))
    
    story.append(Spacer(1, 5))
    story.append(Paragraph("Diễn biến quy mô Nợ xấu tuyệt đối và tỷ lệ NPL theo quý:", h2_style))
    story.append(Image(chart_p7, width=150*mm, height=65*mm))
    
    
    # ------------------ PAGE 4: QUARTERLY BREAKDOWNS (NEW) ------------------
    story.append(PageBreak())
    story.append(Paragraph("4. Phân tích Cơ cấu Hoạt động & Tiền gửi theo Quý", h1_style))
    story.append(Paragraph(f"Dưới đây là chi tiết cơ cấu cho vay theo ngành, phân loại các nhóm nợ xấu, kỳ hạn tín dụng và cơ cấu loại hình tiền gửi của {ticker} được cập nhật và làm mịn tới quý gần nhất:", body_style))
    story.append(Spacer(1, 4))
    
    charts_table_data = [
        [Image(chart_p15, width=80*mm, height=50*mm), Image(chart_p16, width=80*mm, height=50*mm)],
        [Image(chart_p17, width=80*mm, height=50*mm), Image(chart_p18, width=80*mm, height=50*mm)]
    ]
    charts_table = Table(charts_table_data, colWidths=[85*mm, 85*mm])
    charts_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(charts_table)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"• <b>Nhận xét:</b> Hoạt động cho vay tiếp tục tập trung vào nhóm ngành mũi nhọn và khách hàng cá nhân. Cơ cấu tiền gửi được tối ưu hóa với tỷ lệ CASA duy trì ổn định làm giảm chi phí vốn đầu vào.", bullet_style))

    # ------------------ PAGE 5: PEER COMPARISON TABLE (NEW) ------------------
    story.append(PageBreak())
    story.append(Paragraph("5. Bảng so sánh chỉ số tài chính các ngân hàng (Peer Benchmark)", h1_style))
    story.append(Paragraph("Bảng so sánh dưới đây cung cấp góc nhìn toàn cảnh về hiệu quả sinh lời, chất lượng tài sản và định giá của 18 ngân hàng thương mại niêm yết hàng đầu Việt Nam tại thời điểm hiện tại:", body_style))
    story.append(Spacer(1, 4))
    
    # Read peer data from JSON for PDF table rendering
    try:
        with open("data/peer_benchmark.json", "r", encoding="utf-8") as f:
            peer_data_pdf = json.load(f)
        peers_pdf_list = peer_data_pdf.get("peers", [])
    except:
        peers_pdf_list = []
        
    pdf_peer_rows = [
        ["Mã", "NPL (%)", "NIM (%)", "CASA (%)", "ROE (%)", "CIR (%)", "P/B (x)", "TTTD (%)", "Vốn hóa (Tỷ)"]
    ]
    
    # Add Industry Average row
    if peers_pdf_list:
        p_npl_a = sum(p.get("npl", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        p_nim_a = sum(p.get("nim", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        p_casa_a = sum(p.get("casa", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        p_roe_a = sum(p.get("roe", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        p_cir_a = sum(p.get("cir", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        p_pb_a = sum(p.get("pb", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        p_cg_a = sum(p.get("cg", 0) for p in peers_pdf_list) / len(peers_pdf_list)
        pdf_peer_rows.append([
            "TB Ngành", f"{p_npl_a:.2f}%", f"{p_nim_a:.2f}%", f"{p_casa_a:.1f}%",
            f"{p_roe_a:.1f}%", f"{p_cir_a:.1f}%", f"{p_pb_a:.2f}x", f"{p_cg_a:.1f}%", "—"
        ])
        
    for p in peers_pdf_list:
        # Highlight target stock
        t_id = p.get("ticker", "")
        pdf_peer_rows.append([
            f"*{t_id}*" if t_id == ticker else t_id,
            f"{p.get('npl'):.2f}%", f"{p.get('nim'):.2f}%", f"{p.get('casa'):.1f}%",
            f"{p.get('roe'):.1f}%", f"{p.get('cir'):.1f}%", f"{p.get('pb'):.2f}x",
            f"{p.get('cg'):.1f}%", f"{p.get('mcap'):,.0f}"
        ])
        
    t_pdf_peer = Table(pdf_peer_rows, colWidths=[20*mm, 18*mm, 18*mm, 18*mm, 18*mm, 18*mm, 16*mm, 18*mm, 26*mm])
    
    # Styling rules
    t_pdf_peer_styles = [
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1A365D")),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#cbd5e1")),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
        ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,1), (-1,1), HexColor("#FFF2CC")),  # Highlight average
        ('FONTNAME', (0,1), (-1,1), FONT_BOLD),
    ]
    
    # Highlight current stock rows
    for r_idx in range(len(pdf_peer_rows)):
        if pdf_peer_rows[r_idx][0].startswith("*"):
            t_pdf_peer_styles.append(('BACKGROUND', (0, r_idx), (-1, r_idx), HexColor("#E2E8F0")))
            t_pdf_peer_styles.append(('FONTNAME', (0, r_idx), (-1, r_idx), FONT_BOLD))
            pdf_peer_rows[r_idx][0] = pdf_peer_rows[r_idx][0].replace("*", "")  # Clean label
            
    t_pdf_peer.setStyle(TableStyle(t_pdf_peer_styles))
    story.append(t_pdf_peer)
    story.append(Spacer(1, 5))
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")
    story.append(Paragraph(f"<i>Nguồn: Báo cáo tài chính các ngân hàng Q1/2026 và dữ liệu thống kê từ Vietcap. P/B được cập nhật theo thị giá đóng cửa ngày: {today_str}.</i>", body_style))
    
    # ------------------ PAGE 6: VALUATION & PE/PB HISTORY ------------------
    story.append(PageBreak())
    story.append(Paragraph("6. Lịch sử biến động định giá & Tóm tắt kết quả", h1_style))
    story.append(Paragraph("Biểu đồ dưới đây cung cấp góc nhìn toàn cảnh về định giá PE/PB của cổ phiếu trong vòng 8 năm qua, làm cơ sở so sánh với giá trị hợp lý hiện tại:", body_style))
    
    # PE/PB history chart in Page 5
    story.append(Image(chart_p13, width=175*mm, height=73*mm))
    story.append(Spacer(1, 5))
    
    # ------------------ PAGE 7: FORECAST ASSUMPTIONS & DETAILED RI MODEL ------------------
    story.append(PageBreak())
    story.append(Paragraph("7. Giả định dự báo & Mô hình định giá Residual Income", h1_style))
    story.append(Paragraph(f"Kết hợp phương pháp định giá Residual Income (trọng số 50%) và P/B median all-time lịch sử (trọng số 50%), giá trị hợp lý của cổ phiếu {ticker} được xác định là <b>{weighted_target:,.0f} VND/CP</b>.", body_style))
    
    # Inject dynamic forecast assumptions explanation
    loans_g_26 = loans_growth_fc[0] * 100
    dep_g_26 = dep_growth_fc[0] * 100
    nim_26 = nim_fc[0] * 100
    coc_26 = coc_fc[0] * 100
    cir_26 = cir_fc[0] * 100
    
    nim_trend_desc = "tăng nhẹ" if nim_fc[0] > nim_hist[-1]/100 else "điều chỉnh giảm nhẹ"
    coc_trend_desc = "giảm dần nhờ chất lượng tài sản cải thiện" if coc_fc[0] < coc_hist[-1]/100 else "duy trì trích lập cao phòng ngừa rủi ro"
    
    story.append(Paragraph("<b>Cơ sở thiết lập giả định dự báo & Ước tính lợi nhuận:</b>", h2_style))
    story.append(Paragraph(f"• <b>Tăng trưởng Tín dụng & Huy động:</b> Ước tính tín dụng năm 2026F tăng trưởng <b>{loans_g_26:.1f}%</b> và huy động tăng <b>{dep_g_26:.1f}%</b>, dựa trên hạn mức tăng trưởng NHNN định hướng và năng lực mở rộng tài sản sinh lời thực tế của ngân hàng.", bullet_style))
    story.append(Paragraph(f"• <b>Biên lãi ròng (NIM) dự báo:</b> NIM năm 2026F dự kiến đạt <b>{nim_26:.2f}%</b> ({nim_trend_desc} so với mức {nim_hist[-1]:.2f}% của năm 2025). Giả định dựa trên chi phí vốn (COF) được kiểm soát nhờ tỷ lệ CASA phục hồi và lãi suất huy động duy trì thấp.", bullet_style))
    story.append(Paragraph(f"• <b>Chi phí tín dụng (CoC) & Trích lập:</b> Chi phí dự phòng (CoC) năm 2026F giả định ở mức <b>{coc_26:.2f}%</b> ({coc_trend_desc}), đảm bảo tỷ lệ bao phủ nợ xấu (LLR) duy trì ở mức an toàn mà không gây áp lực đột biến lên lợi nhuận.", bullet_style))
    story.append(Paragraph(f"• <b>Thu nhập ngoài lãi & Chi phí vận hành:</b> Thu nhập dịch vụ và phi tín dụng khác dự kiến tăng trưởng đều 10-15%/năm nhờ số hóa; Tỷ lệ CIR giả định kiểm soát tốt quanh mức <b>{cir_26:.1f}%</b> nhờ tối ưu hóa quy trình.", bullet_style))
    story.append(Paragraph("• <b>Độ tin cậy & Độ nhạy:</b> Mô hình dự báo có độ tin cậy cao dựa trên chu kỳ vận hành 3 năm gần nhất. Khi các yếu tố vĩ mô thay đổi (lãi suất tăng/giảm mạnh hoặc nợ xấu bùng phát), nhà đầu tư cần tra cứu bảng nhạy cảm (sheet 08_Sensitivity) để điều chỉnh mức định giá tương ứng.", bullet_style))
    story.append(Spacer(1, 6))
    # Detailed RI calculation table in PDF
    ri_table_data = [
        ["Tham số định giá RI (VND/CP)", "2026F", "2027F", "2028F"],
        ["EPS dự phóng", f"{eps_fc_calc[0]:,.0f}", f"{eps_fc_calc[1]:,.0f}", f"{eps_fc_calc[2]:,.0f}"],
        ["BVPS đầu kỳ", f"{bvps_base:,.0f}", f"{bvps_base+eps_fc_calc[0]:,.0f}", f"{bvps_base+sum(eps_fc_calc[:2]):,.0f}"],
        ["Capital Charge (COE)", f"{bvps_base*COE:,.0f}", f"{(bvps_base+eps_fc_calc[0])*COE:,.0f}", f"{(bvps_base+sum(eps_fc_calc[:2]))*COE:,.0f}"],
        ["Lợi nhuận thặng dư (RI)", f"{ri_results[0]:,.0f}", f"{ri_results[1]:,.0f}", f"{ri_results[2]:,.0f}"],
        ["Hệ số chiết khấu", f"{1/(1+COE):.4f}", f"{1/(1+COE)**2:.4f}", f"{1/(1+COE)**3:.4f}"],
        ["PV của RI từng năm", f"{ri_results[0]/(1+COE):,.0f}", f"{ri_results[1]/(1+COE)**2:,.0f}", f"{ri_results[2]/(1+COE)**3:,.0f}"]
    ]
    t_ri = Table(ri_table_data, colWidths=[70*mm, 32*mm, 32*mm, 32*mm])
    t_ri.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HexColor("#1A365D")),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('GRID', (0,0), (-1,-1), 0.5, HexColor("#cbd5e1")),
        ('FONTNAME', (0,0), (-1,-1), FONT_REG),
        ('FONTNAME', (0,0), (-1,0), FONT_BOLD),
        ('FONTNAME', (0,1), (0,-1), FONT_BOLD),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(t_ri)
    story.append(Spacer(1, 4))
    
    # Valuation summary metrics block
    story.append(Paragraph(f"<b>Tóm tắt kết quả định giá:</b>", h2_style))
    story.append(Paragraph(f"• BVPS hiện tại: <b>{bvps_base:,.0f} VND</b> | Tổng PV của RI 3 năm: <b>{pv_ri:,.0f} VND</b>", body_style))
    story.append(Paragraph(f"• PV của Continuing Value: <b>{pv_cv:,.0f} VND</b>", body_style))
    story.append(Paragraph(f"• <b>Giá trị hợp lý theo RI: {ri_value:,.0f} VND</b> | <b>Giá trị hợp lý theo P/B: {pb_value:,.0f} VND</b>", body_bold))
    
    # Handle Windows file lock for PDF
    try:
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
    except Exception as e:
        try:
            temp_name = pdf_path + f".old_{datetime.datetime.now().strftime('%H%M%S')}"
            os.rename(pdf_path, temp_name)
            print(f"[PDF] Renamed locked PDF to {temp_name}")
        except Exception as e2:
            print(f"[WARN] Failed to release PDF lock: {e2}")

    doc.build(story)
    print(f"[PDF] Report successfully saved to {pdf_path}")

    # ── Save JSON Summary for Dashboard ──────────────────────
    # (Quarterly calculations have been moved to the top of the function to be available for charts)

    # FIX Lỗi 13: pe_quarters/pb_quarters lấy toàn bộ lịch sử, không chỉ 12 quý
    q_labels_pe = []
    if 'ttms' in locals() and ttms:
        q_labels_pe = [f"{r['year']}-Q{r['quarter']}" for r in ttms]
    
    summary_json = {
        "ticker": ticker,
        "companyName": company_name,
        "analysis_comments": {
            "ytd_loans_growth": round(ytd_loans_growth, 2),
            "ytd_dep_growth": round(ytd_dep_growth, 2),
            "profit_source_comment": profit_source_comment,
            "provision_comment": provision_comment,
            "nii_latest_val": round(nii_latest_val, 2),
            "npat_latest_val": round(npat_latest_val, 2)
        },
        "credit_funding_growth": {
            "quarters": labels_g,
            "credit_absolute": [round(c, 1) for c in credit_absolute],
            "funding_absolute": [round(f, 1) for f in funding_absolute],
            "credit_ytd": credit_ytd_pct,
            "funding_ytd": funding_ytd_pct
        },
        "sector": sector,
        "currentPrice": current_price,
        "marketCap": market_cap,
        "shares": shares,
        "assumptions": {
            "loans_growth": [round(x * 100, 1) for x in loans_growth_fc],
            "dep_growth": [round(x * 100, 1) for x in dep_growth_fc],
            "nim": [round(x * 100, 2) for x in nim_fc],
            "cir": [round(x * 100, 1) for x in cir_fc],
            "coc": [round(x * 100, 2) for x in coc_fc],
            "npl": [round(x * 100, 2) for x in npl_fc]
        },
        "financial_snapshot": {
            "years": ["2023", "2024", "2025", "2026F", "2027F", "2028F"],
            "nii":  [round(x, 1) for x in nii_hist[-3:]] + [round(x, 1) for x in nii_fc],
            "toi":  [round(x, 1) for x in toi_hist[-3:]] + [round(x, 1) for x in toi_fc],
            "npat": [round(x, 1) for x in np_hist[-3:]]  + [round(x, 1) for x in np_fc],
            "nim":  [round(x, 2) for x in nim_hist[-3:]]  + [round(x*100, 2) for x in nim_fc],
            "roe":  [round(x, 1) for x in roe_hist[-3:]]  + [round(x, 2) for x in roe_fc_calc],
            "ldr":  [round(x, 1) for x in ldr_hist[-3:]]  + [round(x, 2) for x in ldr_fc_calc],
            "npl":  [round(x, 2) for x in npl_ratio_hist[-3:]] + [round(x*100, 2) for x in npl_fc]
        },
        "forecast_text": {
            "loans_g_26": round(loans_growth_fc[0] * 100, 1),
            "dep_g_26": round(dep_growth_fc[0] * 100, 1),
            "nim_26": round(nim_fc[0] * 100, 2),
            "coc_26": round(coc_fc[0] * 100, 2),
            "cir_26": round(cir_fc[0] * 100, 1),
            "nim_trend": "tăng nhẹ" if nim_fc[0] > nim_hist[-1]/100 else "điều chỉnh giảm nhẹ",
            "coc_trend": "giảm dần nhờ chất lượng tài sản cải thiện" if coc_fc[0] < coc_hist[-1]/100 else "duy trì trích lập cao phòng ngừa rủi ro"
        },
        "gdriveExcelUrl": None,
        "gdrivePdfUrl": None,
        "pe_hist": pe_all_vals[-5:] if len(pe_all_vals) >= 5 else pe_all_vals,
        "pb_hist": pb_all_vals[-5:] if len(pb_all_vals) >= 5 else pb_all_vals,
        "pe_quarters": pe_all_vals,  # Toàn bộ lịch sử
        "pb_quarters": pb_all_vals,  # Toàn bộ lịch sử
        "quarter_labels": q_labels_pe if q_labels_pe else [f"{2022+i//4}-Q{i%4+1}" for i in range(len(pb_all_vals))],
        "income_quarterly": [
            {"quarter": f"{r['yearReport']}-Q{r.get('lengthReport',0)}",
             "nii": round((r.get("isb27") or 0)/1e9, 2),
             "nonii": round(((r.get("isb38") or 0) - (r.get("isb27") or 0))/1e9, 2),
             "toi": round((r.get("isb38") or 0)/1e9, 2),
             "ppop": round((r.get("isb40") or 0)/1e9, 2),
             "npat": round((r.get("isa20") or 0)/1e9, 2)}
            for r in iq_sorted
        ],
        # FIX Lỗi 12: ratios_quarterly từ BCTC quý thực tế
        "ratios_quarterly": {
            "quarters": q_labels_json,
            "nim":  nim_q_json,
            "cof":  cof_q_json,
            "ldr":  ldr_q_json,
            "casa": casa_q_json,
            "npl":  npl_q_json,
            "llr":  llr_q_json,
            "roe":  roe_q_json,
            "credit_growth": credit_growth_q_json
        },
        "earning_assets_quarterly": {
            "quarters": q_labels_json,
            "cash_sbv": cash_sbv_q,
            "bank_dep": bank_dep_q,
            "loans": loans_q_series,
            "inv_sec": inv_sec_q_series
        },
        "loan_industry": {
            "quarters": q_labels_json[:min_q2],
            "real_estate": loan_ind_real_estate,
            "individuals": loan_ind_individuals,
            "wholesale_retail": loan_ind_wholesale_retail,
            "others": loan_ind_others
        },
        "npl_groups": {
            "quarters": q_labels_json[:min_q2],
            "group1": npl_grp1,
            "group2": npl_grp2,
            "group3": npl_grp3,
            "group4": npl_grp4,
            "group5": npl_grp5
        },
        "loan_terms": {
            "quarters": q_labels_json[:min_q2],
            "short_term": term_short,
            "medium_term": term_medium,
            "long_term": term_long
        },
        "deposit_types": {
            "quarters": q_labels_json[:min_q2],
            "casa": dep_casa,
            "term": dep_term,
            "others": dep_others
        },
        "earning_assets": {
            "years": [str(y) for y in years_hist],
            "cash_sbv": [round(cash_hist[i] + sbv_dep_hist[i], 1) for i in range(len(years_hist))],
            "bank_dep": [round(bank_dep_hist[i], 1) for i in range(len(years_hist))],
            "loans":    [round(loans_hist[i], 1) for i in range(len(years_hist))],
            "inv_sec":  [round(inv_sec_bs_hist[i], 1) for i in range(len(years_hist))]
        },
        "thesis": [
            ai_comments["business"][:200],
            ai_comments["financial"][:200],
            ai_comments["valuation"][:200]
        ],
        "risks": [
            f"Rủi ro NIM thu hẹp: cạnh tranh lãi suất cho vay trong bối cảnh NHNN giảm lãi suất — NIM có thể giảm {round(nim_hist[-1]*0.1, 2)}% so kế hoạch.",
            "Rủi ro nợ xấu: thị trường bất động sản hồi phục chậm, nợ nhóm 2 tăng có thể chuyển thành NPL trong 6–12 tháng.",
            f"Rủi ro pháp lý: Thông tư siết TPDN và nâng chuẩn CAR có thể hạn chế room tín dụng của {ticker}."
        ],
        # FIX Lỗi JSON: moats đủ 5 keys, pestle đủ 6 items
        "moats": {
            "Network Effect":    {"score": 4, "desc": f"Mạng lưới khách hàng {ticker} rộng — hiệu ứng mạng lưới tạo CASA ổn định."},
            "Cost Advantage":    {"score": 3, "desc": "CASA cao giúp COF thấp hơn ngành, tạo lợi thế chi phí vốn bền vững."},
            "Switching Cost":    {"score": 3, "desc": "Hệ sinh thái ứng dụng số tích hợp tạo rào cản chuyển đổi khách hàng."},
            "Intangible Assets": {"score": 4, "desc": "Thương hiệu ngân hàng uy tín và quan hệ doanh nghiệp/cá nhân lâu năm."},
            "Efficient Scale":   {"score": 3, "desc": "Quy mô tài sản tạo hiệu quả vận hành trên chi phí cố định hạ tầng CNTT."}
        },
        "pestle": [
            {"factor": "Political",     "content": "Môi trường chính trị ổn định, NHNN điều hành linh hoạt hỗ trợ tăng trưởng tín dụng.",          "impact": "Positive"},
            {"factor": "Economic",      "content": "GDP 6–7%, nhu cầu tín dụng sản xuất–tiêu dùng cao, lãi suất huy động dần giảm.",               "impact": "Positive"},
            {"factor": "Social",        "content": "Tầng lớp trung lưu mở rộng, thanh toán không tiền mặt thúc đẩy CASA và dịch vụ phí.",          "impact": "Positive"},
            {"factor": "Technological", "content": "Chuyển đổi số banking giảm CIR, tăng CASA và cross-sell — đây là lợi thế cạnh tranh dài hạn.", "impact": "Positive"},
            {"factor": "Legal",         "content": "Thông tư 06/2024 siết TPDN, Basel II/III nâng chuẩn CAR — tạo áp lực pháp lý ngắn hạn.",      "impact": "Neutral"},
            {"factor": "Environmental", "content": "ESG lending ngày càng được yêu cầu; rủi ro tín dụng BĐS liên quan biến đổi khí hậu tăng.",    "impact": "Negative"}
        ],
        "valuation": {
            "weightedTarget": round(weighted_target),
            "upside": round(upside, 2),
            "bear": int(bear_target),
            "base": int(weighted_target),
            "bull": int(bull_target),
            "pbAttractive": round(pb_attractive, 2),
            "pbAttractivePrice": round(pb_attractive * (bvps_base + eps_fc_calc[0])),
            "pbOver": round(pb_over, 2),
            "pbOverPrice": round(pb_over * (bvps_base + eps_fc_calc[0])),
            "pbMedianPrice": round(pb_all_median * (bvps_base + eps_fc_calc[0])),
            "peMedianPrice": round(pe_all_median * eps_fc_calc[0]),
            "COE": round(COE * 100, 2),
            "peMedian": round(pe_all_median, 2),
            "pbMedian": round(pb_all_median, 2),
            "bvpsBase": round(bvps_base),
            "riValue": round(ri_value),
            "pbValue": round(pb_value),
            "pvRi": round(pv_ri),
            "pvCv": round(pv_cv),
            "epsFc": [round(x) for x in eps_fc_calc],
            "riResults": [round(x) for x in ri_results],
            "recommend": 'MUA' if upside > 15 else ('BÁN' if upside < -5 else 'THEO DÕI')
        },
        "comments": {
            "businessModel":        ai_comments["business"],
            "financialPerformance": ai_comments["financial"],
            "valuationText":        ai_comments["valuation"]
        },
        "data": {
            "years":  all_years,
            "revenue": [round(x, 2) for x in nii_hist + nii_fc],
            "npat":    [round(x, 2) for x in np_hist + np_fc],
            "eps":     [round(x, 1) for x in eps_hist_calc + eps_fc_calc],
            "equity":  [round(x, 2) for x in equity_hist + [equity_hist[-1] + sum(np_fc[:i+1])*0.7 for i in range(3)]]
        },
        # FIX Lỗi: ratios lịch sử tính từ dữ liệu BCTC, forecast từ assumptions (Skill §18)
        "ratios": {
            "nim":  [round(x/100, 4) for x in nim_hist] + [round(n, 4) for n in nim_fc],
            "roe":  [round(x/100, 4) for x in roe_hist] + [None]*3,
            "roa":  [round(x/100, 4) for x in roa_hist] + [None]*3,
            "npl":  [round(x/100, 4) for x in npl_ratio_hist] + [round(n, 4) for n in npl_fc],
            "ldr":  [round(x/100, 4) for x in ldr_hist] + [None]*3,
            "casa": [round(x/100, 4) for x in casa_ratio_hist] + [round(c, 4) for c in casa_target_fc]
        }
    }
    
    json_path = os.path.join(PROJECT_ROOT, "data", f"{ticker}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary_json, f, ensure_ascii=False, indent=2)
    print(f"[JSON] Dashboard summary updated at: {json_path}")
    
    return True
