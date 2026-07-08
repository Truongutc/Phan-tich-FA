#!/usr/bin/env python3
"""
template_kcn.py — Universal calculation engine for Vietnamese Industrial Park (KCN)
real-estate leasing companies: IDC, SIP, PHR, SZC, KBC, BCM, NTC, DPR...

Mô hình: doanh thu/giá vốn/lợi nhuận gộp theo TỪNG MẢNG kinh doanh (điện, cho thuê
đất & hạ tầng KCN, dịch vụ tiện ích, BOT, xây dựng, BĐS, cao su, ...), lấy TRỰC TIẾP
từ thuyết minh BCTC hợp nhất (kho dữ liệu data/segments_kcn/<TICKER>.json — KHÔNG
ước tính tỷ trọng/biên lợi nhuận). Xem quy trình thu thập dữ liệu trong skill `bds-kcn`.

Định giá CỐ ĐỊNH cho mọi mã KCN: 40% P/E + 40% P/B + 20% Residual Income (RI).
RI dùng đúng công thức skill `ngan-hang` (RI = EPS − COE×BVPS, PV 3 năm + giá trị
tiếp diễn) nhưng KHÔNG cộng phần bù rủi ro đặc thù (COE = Rf + β×ERP), và roll-forward
BVPS có trừ cổ tức (payout ratio lịch sử) vì nhóm KCN thường chia cổ tức tiền mặt lớn.

Output: Excel model (15 sheets, formula-driven) + PDF report + data/<TICKER>.json.
"""
import os
import sys

# Fix Windows console encoding (cp1252 không hỗ trợ tiếng Việt)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import json
import datetime
import subprocess
import statistics as stats
import statistics
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, grey, black
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests

from fetch_data import section_to_years, section_to_quarters

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
STORE_DIR = os.path.join(PROJECT_ROOT, "data", "segments_kcn")


# ══════════════════════════════════════════════════════════════════════════
# VIETNAMESE FONT REGISTRATION (copy nguyên từ template_securities.py)
# ══════════════════════════════════════════════════════════════════════════
def register_vn_fonts():
    font_paths_to_try = [
        ("C:/Windows/Fonts/arial.ttf", "Arial"),
        ("C:/Windows/Fonts/arialbd.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "Arial"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Arial"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", "Arial-Bold"),
    ]
    found = {}
    for path, freg in font_paths_to_try:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(freg, path))
                found[freg] = path
            except Exception:
                pass
    return found


_VN_FONTS = register_vn_fonts()
FONT_REG = 'Arial' if 'Arial' in _VN_FONTS else 'Helvetica'
FONT_BOLD = 'Arial-Bold' if 'Arial-Bold' in _VN_FONTS else 'Helvetica-Bold'


# ══════════════════════════════════════════════════════════════════════════
# CAPM INPUTS: Rf + Beta (COPY NGUYÊN VẸN từ template_securities.py/template_banking.py)
# ══════════════════════════════════════════════════════════════════════════
RF_UA_STR = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"


def _fetch_via_curl_rf(url, timeout=10, label=None):
    tag = f" [{label}]" if label else ""
    try:
        r = subprocess.run(
            ["curl", "-sL", "-A", RF_UA_STR, "--max-time", str(timeout), url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout + 5,
        )
        return r.stdout if r.returncode == 0 else ""
    except Exception as e:
        print(f"  [DIAG]{tag} fetch_via_curl exception: {e}")
        return ""


def fetch_rf_vietnam(timeout=15):
    FALLBACK_RF = 0.045
    try:
        html = _fetch_via_curl_rf("https://vn.investing.com/rates-bonds/vietnam-10-year-bond-yield", timeout=timeout, label="investing-rf")
        if html:
            import re
            m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
            if m:
                rf = float(m.group(1).replace(",", "")) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "investing.com"
    except Exception as e:
        print(f"  [WARN] investing.com Rf failed: {e}")
    try:
        r = requests.get("https://www.worldgovernmentbonds.com/bond-yield/vietnam/10-years/",
                          headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
        if r.status_code == 200:
            import re
            matches = re.findall(r'(\d+\.\d+)%', r.text[:5000])
            if matches:
                rf = float(matches[0]) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "worldgovernmentbonds.com"
    except Exception as e:
        print(f"  [WARN] WorldGovernmentBonds Rf failed: {e}")
    return FALLBACK_RF, "Fallback (manual)"


def fetch_aligned_history(ticker, days=720, timeout=15):
    from_time = 1577836800
    to_time = 2000000000
    url_stock = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol={ticker}&resolution=D&from={from_time}&to={to_time}"
    url_index = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol=VNINDEX&resolution=D&from={from_time}&to={to_time}"
    headers = {
        "User-Agent": RF_UA_STR, "Accept": "application/json, text/plain, */*",
        "Referer": "https://dchart.vndirect.com.vn/",
    }
    try:
        r_stock = requests.get(url_stock, headers=headers, timeout=timeout)
        r_index = requests.get(url_index, headers=headers, timeout=timeout)
        if r_stock.status_code == 200 and r_index.status_code == 200:
            d_stock, d_index = r_stock.json(), r_index.json()
            t_s, c_s = d_stock.get("t") or [], d_stock.get("c") or []
            t_m, c_m = d_index.get("t") or [], d_index.get("c") or []
            map_stock = {t_s[i]: c_s[i] for i in range(min(len(t_s), len(c_s))) if c_s[i] is not None and c_s[i] > 0}
            map_index = {t_m[i]: c_m[i] for i in range(min(len(t_m), len(c_m))) if c_m[i] is not None and c_m[i] > 0}
            common_t = sorted(list(set(map_stock.keys()) & set(map_index.keys())))
            aligned = []
            for t in common_t:
                date_str = datetime.datetime.fromtimestamp(t).strftime("%Y-%m-%d")
                p_s = map_stock[t]
                if p_s < 1000:
                    p_s = p_s * 1000
                aligned.append((date_str, p_s, map_index[t]))
            return aligned
    except Exception as e:
        print(f"[Beta Calc] Error fetching history: {e}")
    return []


def fetch_beta_vietstock(ticker, timeout=15):
    try:
        search_url = f"https://finance.vietstock.vn/search?query={ticker}"
        headers = {'User-Agent': RF_UA_STR}
        r1 = requests.get(search_url, headers=headers, timeout=timeout)
        if r1.status_code == 200:
            data = json.loads(r1.text).get("data", "")
            target_url = ""
            for line in data.split('\r\n'):
                parts = line.split('|')
                if len(parts) >= 3 and parts[0].strip().upper() == ticker.upper():
                    target_url = parts[2]
                    break
            if target_url:
                r2 = requests.get(target_url, headers={'User-Agent': RF_UA_STR, 'Referer': 'https://finance.vietstock.vn/'}, timeout=timeout)
                if r2.status_code == 200:
                    import re
                    m = re.search(r'\"Beta\":\"([\d\.]+)\"', r2.text)
                    if m:
                        beta = float(m.group(1))
                        if 0.3 <= beta <= 2.5:
                            return beta
    except Exception as e:
        print(f"  [WARN] Vietstock scrape failed: {e}")
    return None


def fetch_beta_vietcap(ticker, timeout=15):
    try:
        url = f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}"
        r = requests.get(url, headers={"User-Agent": RF_UA_STR, "Referer": "https://trading.vietcap.com.vn/"}, timeout=timeout)
        if r.status_code == 200:
            d = r.json().get("data", {})
            beta = d.get("beta")
            if beta is not None and 0.3 <= float(beta) <= 2.5:
                return float(beta)
    except Exception:
        pass
    return None


def fetch_and_calc_beta(ticker, days=720, timeout=20, fallback=1.1):
    print(f"  [INFO] Đang tải lịch sử giá để tự tính Beta cho {ticker}...")
    aligned_data = fetch_aligned_history(ticker, days=days, timeout=timeout)
    latest_price = aligned_data[-1][1] if aligned_data else None
    num_sessions = len(aligned_data)
    web_beta = fetch_beta_vietstock(ticker, timeout) or fetch_beta_vietcap(ticker, timeout) or fallback
    calculated_beta, is_enough_sessions = fallback, False
    if num_sessions >= 30:
        sliced_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
        s = [x[1] for x in sliced_data]
        m = [x[2] for x in sliced_data]
        rs = [(s[i] - s[i-1]) / s[i-1] for i in range(1, len(s))]
        rm = [(m[i] - m[i-1]) / m[i-1] for i in range(1, len(m))]
        n_ret = len(rs)
        mean_rs, mean_rm = sum(rs) / n_ret, sum(rm) / n_ret
        cov_sm = sum((rs[i]-mean_rs)*(rm[i]-mean_rm) for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 0
        var_m = sum((rm[i]-mean_rm)**2 for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 1.0
        calculated_beta = round(max(0.3, min(2.5, cov_sm / var_m if var_m > 0 else fallback)), 4)
        if num_sessions >= 250:
            is_enough_sessions = True
        aligned_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
    beta_src = f"Tự tính toán ({num_sessions} phiên)" if is_enough_sessions else f"Web/API ({web_beta:.2f}) - lịch sử chỉ {num_sessions} phiên"
    return calculated_beta, web_beta, is_enough_sessions, beta_src, latest_price, aligned_data


# ══════════════════════════════════════════════════════════════════════════
# FIELD MAP VIETCAP — CÔNG TY THƯỜNG (isa/bsa/cfa), verify trực tiếp qua cache IDC 2026-07
# ══════════════════════════════════════════════════════════════════════════
IS_GEN = {
    "revenue": "isa3",         # Doanh thu thuần
    "cogs": "isa4",            # Giá vốn hàng bán
    "gross_profit": "isa5",    # Lợi nhuận gộp
    "fin_income": "isa6",      # Doanh thu hoạt động tài chính
    "fin_expense": "isa7",     # Chi phí tài chính
    "interest_expense": "isa8",# Chi phí lãi vay
    "sga_sales": "isa9",       # Chi phí bán hàng
    "sga_admin": "isa10",      # Chi phí quản lý doanh nghiệp
    "operating_result": "isa11",
    "other_income_net": "isa14",
    "pbt": "isa16",            # Lãi/(lỗ) trước thuế
    "tax_current": "isa17",
    "tax_deferred": "isa18",
    "npat": "isa20",           # Lãi/(lỗ) thuần sau thuế
    "nci_income": "isa21",     # Lợi ích của cổ đông thiểu số (P&L)
    "npat_parent": "isa22",    # LNST của cổ đông công ty mẹ
    "eps_basic": "isa23",
}
BS_GEN = {
    "cash": "bsa2",             # Tiền và tương đương tiền
    "total_assets": "bsa53",    # TỔNG CỘNG TÀI SẢN
    "total_liab": "bsa54",      # NỢ PHẢI TRẢ
    "short_borrow": "bsa56",    # Vay ngắn hạn
    "long_borrow": "bsa71",     # Vay dài hạn
    "equity_total": "bsa78",    # VỐN CHỦ SỞ HỮU (đã bao gồm NCI — xem note bên dưới)
    "charter_capital": "bsa80", # Vốn góp
    "nci": "bsa210",            # Lợi ích cổ đông không kiểm soát (nested trong bsa78)
    "total_capital": "bsa96",   # TỔNG CỘNG NGUỒN VỐN
}
CF_GEN = {
    "depreciation": "cfa2",     # Khấu hao TSCĐ và BĐSĐT
    "dividends_paid": "cfa32",  # Cổ tức, lợi nhuận đã trả cho chủ sở hữu (âm)
}
# NOTE QUAN TRỌNG (verify với .cache/IDC_bctc.json 2026-07):
#   bsa54 (Nợ phải trả) + bsa78 (Vốn chủ sở hữu) == bsa96 (Tổng cộng nguồn vốn) — ĐÚNG.
#   => bsa78 là TỔNG VCSH đã bao gồm NCI (bsa210). VCSH thuộc về cổ đông công ty mẹ (dùng cho
#      BVPS/EPS/RI) PHẢI tính = bsa78 - bsa210 (đã verify: IDC bsa210 khớp đúng số NCI trong TM 30).


def _get_yr(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            if v is not None:
                return v / 1e9
    return 0.0


def _get_q(records, year, quarter, field):
    for r in records:
        if r.get("yearReport") == year and r.get("lengthReport") == quarter:
            v = r.get(field)
            if v is not None:
                return v / 1e9
    return 0.0


# ══════════════════════════════════════════════════════════════════════════
# STYLING CONSTANTS (giống template_securities.py để đồng nhất giao diện Excel)
# ══════════════════════════════════════════════════════════════════════════
FONT_NAME = "Calibri"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name=FONT_NAME, size=11, bold=True)
ITALIC_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")
DATA_FONT = Font(name=FONT_NAME, size=10)
ASSUMP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LINK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
P_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                      top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
FMT_NUM = '#,##0'
FMT_PCT = '0.00%'
FMT_MUL = '0.00"x"'
FMT_PRICE = '#,##0'


def header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def data_row(ws, row, label, values, fmt=None, note=None, bold=False, fill=None):
    c0 = ws.cell(row=row, column=1, value=label)
    c0.font = BOLD_FONT if bold else DATA_FONT
    for i, v in enumerate(values, start=2):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BOLD_FONT if bold else DATA_FONT
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
    if note:
        ws.cell(row=row, column=len(values) + 3, value=note).font = ITALIC_FONT


# ══════════════════════════════════════════════════════════════════════════
# KHO DỮ LIỆU MẢNG (data/segments_kcn/<TICKER>.json) — xem skill bds-kcn
# ══════════════════════════════════════════════════════════════════════════
def load_segments_kcn(ticker):
    path = os.path.join(STORE_DIR, f"{ticker.upper()}.json")
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Chưa có kho dữ liệu mảng cho {ticker} tại {path}. "
            f"Chạy quy trình thu thập dữ liệu mảng (xem skill bds-kcn: bctc_pdf_tool.py "
            f"list/plan-downloads/download/render + segments_kcn_tool.py merge/validate) trước khi phân tích."
        )
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def check_segment_consistency(store, is_recs_years, is_recs_quarters):
    """Đối chiếu Σ doanh thu/giá vốn các mảng với tổng DT thuần/GVHB Vietcap (isa3/isa4).
    Trả về dict {period_key: {"gapRevPct":..., "gapCogsPct":..., "status": OK/WARN/FAIL}}."""
    result = {}
    result = {}
    for year_str, seg_data in store.get("yearly", {}).items():
        # Trích xuất số năm từ định dạng YYYY hoặc YYYY(CN)
        clean_yr_str = year_str.split("(")[0]
        try:
            year = int(clean_yr_str)
        except ValueError:
            print(f"  [WARN] Không thể parse năm {year_str} thành int, bỏ qua đối chiếu.")
            continue
        rev_sum = sum((v.get("revenue") or 0) for v in seg_data.values())
        cogs_sum = sum((v.get("cogs") or 0) for v in seg_data.values() if v.get("cogs") is not None)
        vc_rev = _get_yr(is_recs_years, year, IS_GEN["revenue"])
        vc_cogs = _get_yr(is_recs_years, year, IS_GEN["cogs"])
        gap_rev = abs(rev_sum - vc_rev) / vc_rev * 100 if vc_rev else None
        gap_cogs = abs(cogs_sum - vc_cogs) / vc_cogs * 100 if vc_cogs else None
        status = "OK"
        for g in (gap_rev, gap_cogs):
            if g is not None:
                if g > 10:
                    status = "FAIL"
                elif g > 3 and status != "FAIL":
                    status = "WARN"
        result[year_str] = {"gapRevPct": round(gap_rev, 1) if gap_rev is not None else None,
                             "gapCogsPct": round(gap_cogs, 1) if gap_cogs is not None else None,
                             "status": status}
    for qkey, seg_data in store.get("quarterly", {}).items():
        year, q = int(qkey[:4]), int(qkey[-1])
        rev_sum = sum((v.get("revenue") or 0) for v in seg_data.values())
        cogs_sum = sum((v.get("cogs") or 0) for v in seg_data.values() if v.get("cogs") is not None)
        vc_rev = _get_q(is_recs_quarters, year, q, IS_GEN["revenue"])
        vc_cogs = _get_q(is_recs_quarters, year, q, IS_GEN["cogs"])
        gap_rev = abs(rev_sum - vc_rev) / vc_rev * 100 if vc_rev else None
        gap_cogs = abs(cogs_sum - vc_cogs) / vc_cogs * 100 if vc_cogs else None
        status = "OK"
        for g in (gap_rev, gap_cogs):
            if g is not None:
                if g > 10:
                    status = "FAIL"
                elif g > 3 and status != "FAIL":
                    status = "WARN"
        result[qkey] = {"gapRevPct": round(gap_rev, 1) if gap_rev is not None else None,
                         "gapCogsPct": round(gap_cogs, 1) if gap_cogs is not None else None,
                         "status": status}
    for k, v in result.items():
        if v["status"] != "OK":
            print(f"  [{v['status']}] Segment check {k}: gap DT={v['gapRevPct']}% GV={v['gapCogsPct']}%")
    max_gap = max([abs(v["gapRevPct"]) for v in result.values() if v["gapRevPct"] is not None] or [0])
    return result, max_gap


def build_segment_history(store, is_recs_y, is_recs_q):
    """Trả về (seg_names theo order, seg_labels, seg_colors, yearly{seg:{year:{rev,cogs,gp,gm}}},
    quarterly{seg:{qkey:{...}}}) từ kho dữ liệu mảng, có tự động sửa tỷ lệ đơn vị nếu vượt quá tổng DT."""
    seg_defs = store.get("segments", {})
    seg_names = sorted(seg_defs.keys(), key=lambda s: seg_defs[s].get("order", 99))
    seg_labels = {s: seg_defs[s].get("label", s) for s in seg_names}
    seg_colors = {s: seg_defs[s].get("color", "#94a3b8") for s in seg_names}

    # Xây dựng bản đồ tổng doanh thu từ Vietcap để đối chiếu đơn vị
    rev_map = {}
    for r in is_recs_y:
        y = r.get("yearReport")
        val = r.get("isa3")  # Doanh thu thuần (tỷ VND)
        if y and val:
            rev_map[f"{y}(CN)"] = val
    for r in is_recs_q:
        y = r.get("yearReport")
        q = r.get("lengthReport")
        val = r.get("isa3")  # Doanh thu thuần (tỷ VND)
        if y and q and val:
            rev_map[f"{y}Q{q}"] = val

    yearly = {s: {} for s in seg_names}
    for year_str, seg_data in store.get("yearly", {}).items():
        total_ref = rev_map.get(year_str)
        for seg, v in seg_data.items():
            rev, cogs = v.get("revenue"), v.get("cogs")
            if rev is not None and total_ref is not None and rev > total_ref * 1.05:
                # Tự động sửa lỗi đơn vị (chia 1000)
                while rev > total_ref * 1.05:
                    rev /= 1000.0
                    if cogs is not None:
                        cogs /= 1000.0
                rev = round(rev, 2)
                if cogs is not None:
                    cogs = round(cogs, 2)

            gp = (rev - cogs) if (rev is not None and cogs is not None) else None
            gm = (gp / rev) if (gp is not None and rev) else None
            yearly.setdefault(seg, {})[year_str] = {"revenue": rev, "cogs": cogs, "grossProfit": gp, "grossMargin": gm}

    quarterly = {s: {} for s in seg_names}
    for qkey, seg_data in store.get("quarterly", {}).items():
        total_ref = rev_map.get(qkey)
        for seg, v in seg_data.items():
            rev, cogs = v.get("revenue"), v.get("cogs")
            if rev is not None and total_ref is not None and rev > total_ref * 1.05:
                # Tự động sửa lỗi đơn vị (chia 1000)
                while rev > total_ref * 1.05:
                    rev /= 1000.0
                    if cogs is not None:
                        cogs /= 1000.0
                rev = round(rev, 2)
                if cogs is not None:
                    cogs = round(cogs, 2)

            gp = (rev - cogs) if (rev is not None and cogs is not None) else None
            gm = (gp / rev) if (gp is not None and rev) else None
            quarterly.setdefault(seg, {})[qkey] = {"revenue": rev, "cogs": cogs, "grossProfit": gp, "grossMargin": gm}

    return seg_names, seg_labels, seg_colors, yearly, quarterly


def _cagr(values):
    """CAGR từ dãy giá trị dương theo thời gian tăng dần, kẹp [-10%, +20%]."""
    vals = [v for v in values if v is not None and v > 0]
    if len(vals) < 2:
        return 0.03
    n = len(vals) - 1
    g = (vals[-1] / vals[0]) ** (1 / n) - 1
    return max(-0.10, min(0.20, g))


DRIVER_GROWTH_CAP = {
    "power": 0.05, "bot": 0.05, "rubber": 0.0, "utilities": 0.07,
    "leasing": None, "construction": None, "realestate": None,
    "wastewater": 0.06, "goods": None, "other": None,
}


def forecast_segments(store, seg_names, yearly, years_hist_avail, n_fc=3, quarterly=None):
    """Dự phóng doanh thu + biên LNG mỗi mảng dựa trên CAGR lịch sử (kẹp theo driver) +
    biên LNG bình quân 2 kỳ gần nhất. Trả về {seg: {"g": float, "gm_fc": float,
    "cap": float|None, "driver": str, "ytd_annualised": float|None}}.
    
    Nếu có dữ liệu quý (quarterly) của năm lịch sử mới nhất, tính run-rate YTD và so sánh
    với năm trước cùng kỳ để xác định tốc độ tăng trưởng thực tế hơn CAGR thuần.
    """
    fc = {}
    for seg in seg_names:
        seg_years = sorted(yearly.get(seg, {}).keys())
        rev_series = [yearly[seg][y]["revenue"] for y in seg_years if yearly[seg][y]["revenue"] is not None]
        gm_series = [yearly[seg][y]["grossMargin"] for y in seg_years if yearly[seg][y].get("grossMargin") is not None]
        driver = store["segments"].get(seg, {}).get("driver", "other")
        cap = DRIVER_GROWTH_CAP.get(driver)
        g = _cagr(rev_series) if len(rev_series) >= 2 else None
        gm_fc = round(stats.mean(gm_series[-2:]), 4) if gm_series else None
        
        # ── YTD anchoring: nếu có dữ liệu quý năm hiện tại → dùng run-rate ──
        ytd_annualised = None
        ytd_growth = None
        if quarterly is not None:
            q_keys_seg = sorted(quarterly.get(seg, {}).keys())
            if q_keys_seg:
                # Tìm các quý của năm mới nhất trong kho quarterly
                latest_q_year = max(int(k[:4]) for k in q_keys_seg if k[:4].isdigit())
                latest_q_data = {k: quarterly[seg][k] for k in q_keys_seg if k.startswith(str(latest_q_year))}
                if latest_q_data:
                    ytd_rev = sum(
                        v.get("revenue") or 0
                        for v in latest_q_data.values()
                        if v.get("revenue") is not None
                    )
                    n_qtrs = len(latest_q_data)
                    if n_qtrs > 0 and ytd_rev > 0:
                        ytd_annualised = round(ytd_rev * 4 / n_qtrs, 2)  # run-rate hoá lên 4Q
                        # So với năm trước cùng kỳ (same number of quarters)
                        prev_year = latest_q_year - 1
                        prev_q_data = {k: quarterly[seg][k] for k in q_keys_seg if k.startswith(str(prev_year))}
                        if prev_q_data:
                            same_qtrs = sorted(prev_q_data.keys())[:n_qtrs]
                            prev_ytd = sum(
                                quarterly[seg][k].get("revenue") or 0
                                for k in same_qtrs
                                if quarterly[seg][k].get("revenue") is not None
                            )
                            if prev_ytd > 0:
                                ytd_growth = round(ytd_rev / prev_ytd - 1, 4)
                
                # Nếu YTD growth mạnh hơn CAGR thì dùng YTD growth cho năm fc đầu tiên
                if ytd_growth is not None and g is not None:
                    # Chỉ thay thế nếu sai lệch lớn hơn 5pp để tránh volatility 1 kỳ
                    if abs(ytd_growth - g) > 0.05:
                        g_ytd_capped = max(-0.10, min(0.30, ytd_growth))  # kẹp rộng hơn CAGR
                        g = round(0.5 * g + 0.5 * g_ytd_capped, 4)  # blend 50/50
        
        # Áp dụng cap theo driver
        if cap is not None and g is not None:
            g = min(g, cap)
            
        fc[seg] = {"g": g, "gm_fc": gm_fc, "cap": cap, "driver": driver, 
                   "ytd_annualised": ytd_annualised, "ytd_growth": ytd_growth}
    return fc


# ══════════════════════════════════════════════════════════════════════════
# AI COMMENTARY (Gemini 2.5 Flash) — giống template_generic.py
# ══════════════════════════════════════════════════════════════════════════
def get_ai_commentary_kcn(ticker, company_name, fin_summary):
    """Gọi Gemini sinh nhận định chuyên sâu về doanh nghiệp KCN. Trả về dict
    {business, financial, valuation}. Nếu thiếu API key hoặc lỗi thì dùng default."""
    defaults = {
        "business": (
            f"{company_name} ({ticker}) là doanh nghiệp bất động sản khu công nghiệp với mô hình đa mảng "
            f"(cho thuê đất, tiện ích, điện, BOT…). Doanh thu mảng hạ tầng KCN mang tính chu kỳ nhưng ổn định "
            f"nhờ hợp đồng thuê dài hạn; mảng điện và tiện ích tạo dòng tiền đều đặn hàng quý."
        ),
        "financial": (
            f"Doanh thu và lợi nhuận gộp phân hóa rõ giữa các mảng: hạ tầng KCN và dịch vụ tiện ích "
            f"thường có biên cao nhất, trong khi xây dựng và hàng hóa có biên thấp. Payout ratio cổ tức "
            f"tiền mặt lịch sử lớn đòi hỏi kiểm soát chặt dòng tiền tự do và vay nợ."
        ),
        "valuation": (
            f"Định giá theo tổ hợp 40% P/E + 40% P/B + 20% RI phù hợp với đặc thù nhóm KCN: biên lợi nhuận "
            f"tương đối ổn định, tài sản cố định lớn và cổ tức tiền mặt đều đặn. "
            f"Rủi ro chính: lấp đầy KCN chậm hơn kỳ vọng, lãi suất tăng ảnh hưởng COE và giá trị RI."
        ),
    }
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        return defaults
    try:
        from google import genai
        from google.genai import types as genai_types
        client = genai.Client(api_key=api_key)
        prompt = (
            f"Bạn là chuyên gia phân tích tài chính cổ phiếu Việt Nam. Hãy viết nhận định chuyên sâu "
            f"bằng tiếng Việt cho cổ phiếu {ticker} ({company_name}) — nhóm bất động sản khu công nghiệp.\n"
            f"Tóm tắt tài chính: {fin_summary}\n\n"
            f"Viết 3 đoạn ngắn (~3-4 câu mỗi đoạn):\n"
            f"1. Mô hình kinh doanh & Vị thế cạnh tranh.\n"
            f"2. Sức khỏe tài chính & Phân tích doanh thu/biên lợi nhuận theo mảng.\n"
            f"3. Triển vọng định giá & Rủi ro cốt lõi.\n"
            f"Trả về JSON thuần (không markdown): "
            f'{{\"business\":\"...\",\"financial\":\"...\",\"valuation\":\"...\"}}'
        )
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=genai_types.GenerateContentConfig(temperature=0.2, max_output_tokens=900),
        )
        text = resp.text.strip().lstrip("```json").lstrip("```").rstrip("```").strip()
        parsed = json.loads(text)
        return {
            "business":  parsed.get("business",  defaults["business"]),
            "financial": parsed.get("financial", defaults["financial"]),
            "valuation": parsed.get("valuation", defaults["valuation"]),
        }
    except Exception as e:
        print(f"  [WARN] AI commentary failed: {e}. Dùng nhận định mặc định.")
        return defaults


# ══════════════════════════════════════════════════════════════════════════
# ĐỊNH GIÁ: 40% P/E + 40% P/B + 20% RI
# ══════════════════════════════════════════════════════════════════════════
def _bvps_parent(bs_recs, year, shares):
    """BVPS thuộc về cổ đông công ty mẹ = (bsa78 - bsa210) / shares (VND/cp).
    bsa78 = VCSH tổng (gồm NCI), bsa210 = lợi ích cổ đông không kiểm soát."""
    for r in bs_recs:
        if r.get("yearReport") == year:
            equity_total = (r.get("bsa78") or 0)
            nci = (r.get("bsa210") or 0)
            vcsh_parent = equity_total - nci
            if vcsh_parent > 0 and shares > 0:
                return vcsh_parent / shares  # VND/cp
    return None


def _eps_parent(is_recs, year):
    """EPS cơ bản từ isa23 (VND/cp). Dùng isa22/shares làm fallback nếu isa23 = 0."""
    for r in is_recs:
        if r.get("yearReport") == year:
            eps = r.get("isa23")
            if eps is not None and eps != 0:
                return float(eps)
    return None


def _payout_ratio(cf_recs, is_recs, year):
    """Payout ratio = |cổ tức đã trả (cfa32)| / LNST cổ đông mẹ (isa22), kẹp [0, 1]."""
    div_paid, npat_parent = 0.0, 0.0
    for r in cf_recs:
        if r.get("yearReport") == year:
            v = r.get("cfa32")
            if v is not None:
                div_paid = abs(float(v))
    for r in is_recs:
        if r.get("yearReport") == year:
            v = r.get("isa22")
            if v is not None:
                npat_parent = abs(float(v))
    if npat_parent > 0:
        return min(1.0, div_paid / npat_parent)
    return 0.40  # fallback 40%


def calc_valuation_kcn(ticker, is_recs_y, bs_recs_y, cf_recs_y, hist_years, shares,
                        current_price, rf, beta, erp=0.07, specific_risk_premium=0.02, n_ri=3,
                        target_pe=None, target_pb=None, eps_floor=0.0):
    """Tính định giá theo tổ hợp 40%P/E + 40%P/B + 20%RI.
    Trả về dict chứa đủ inputs/outputs để ghi vào Excel và JSON."""
    coe = rf + beta * erp + specific_risk_premium  # COE = Rf + β × ERP + specific_risk_premium

    # --- Lấy số liệu lịch sử ---
    eps_vals  = [_eps_parent(is_recs_y, y) for y in hist_years]
    bvps_vals = [_bvps_parent(bs_recs_y, y, shares) for y in hist_years]
    payout_vals = [_payout_ratio(cf_recs_y, is_recs_y, y) for y in hist_years]

    # EPS và BVPS forward (dùng năm lịch sử gần nhất có dữ liệu)
    eps_valid  = [(y, v) for y, v in zip(hist_years, eps_vals)  if v is not None and v > 0]
    bvps_valid = [(y, v) for y, v in zip(hist_years, bvps_vals) if v is not None and v > 0]

    eps_last   = eps_valid[-1][1]  if eps_valid  else 1000.0
    bvps_last  = bvps_valid[-1][1] if bvps_valid else 10000.0
    avg_payout = stats.mean([p for p in payout_vals if p is not None] or [0.40])

    # --- P/E target: median P/E lịch sử, kẹp [8, 25] ---
    pe_hist = []
    for y, eps in eps_valid:
        if current_price > 0 and eps > 0:
            pe_hist.append(current_price / eps)
    if target_pe is None:
        target_pe = round(max(8.0, min(25.0, stats.median(pe_hist))) if pe_hist else 14.0, 1)

    # --- P/B target: median P/B lịch sử, kẹp [0.6, 4] ---
    pb_hist = []
    for y, bvps in bvps_valid:
        if current_price > 0 and bvps > 0:
            pb_hist.append(current_price / bvps)
    if target_pb is None:
        target_pb = round(max(0.6, min(4.0, stats.median(pb_hist))) if pb_hist else 1.4, 1)

    # --- EPS/BVPS dự phóng 1 năm tới (tăng trưởng EPS từ CAGR 3 năm) ---
    if len(eps_valid) >= 2:
        eps_cagr = _cagr([v for _, v in eps_valid[-3:]])
    else:
        eps_cagr = 0.08
    eps_fc1   = eps_last * (1 + eps_cagr)
    if eps_floor > 0:
        eps_fc1 = max(eps_fc1, eps_floor)
    roe_est   = eps_fc1 / bvps_last if bvps_last > 0 else 0.10
    bvps_fc1  = bvps_last * (1 + roe_est * (1 - avg_payout))

    fair_pe = target_pe * eps_fc1
    fair_pb = target_pb * bvps_fc1

    # --- Residual Income (RI) model ---
    # RI_t = EPS_t − COE × BVPS_{t-1}
    # Roll-forward BVPS_{t} = BVPS_{t-1} × (1 + ROE × (1 - payout))
    ri_list, bvps_t = [], bvps_last
    eps_t = eps_last
    for i in range(n_ri):
        eps_t  = eps_t * (1 + eps_cagr)
        ri_t   = eps_t - coe * bvps_t
        ri_list.append(ri_t / ((1 + coe) ** (i + 1)))
        bvps_t = bvps_t + eps_t * (1 - avg_payout)
    # Terminal value: RI_{n+1} / (COE - g_terminal), g_terminal = max(min(eps_cagr*0.5, 0.04), 0)
    g_term = max(0.0, min(0.04, eps_cagr * 0.5))
    ri_terminal = ri_list[-1] * (1 + coe) / max(coe - g_term, 0.01)  # un-discount then re-PV
    ri_terminal_pv = ri_terminal / ((1 + coe) ** n_ri)
    fair_ri = bvps_last + sum(ri_list) + ri_terminal_pv

    # --- Tổng hợp ---
    fair_blend = 0.40 * fair_pe + 0.40 * fair_pb + 0.20 * fair_ri
    upside = (fair_blend - current_price) / current_price if current_price > 0 else 0.0

    return {
        # inputs
        "rf": rf, "beta": beta, "erp": erp, "coe": coe,
        "target_pe": target_pe, "target_pb": target_pb,
        "eps_last": eps_last, "bvps_last": bvps_last,
        "avg_payout": avg_payout, "eps_cagr": eps_cagr, "roe_est": roe_est,
        "eps_fc1": eps_fc1, "bvps_fc1": bvps_fc1,
        # outputs
        "fair_pe": round(fair_pe), "fair_pb": round(fair_pb), "fair_ri": round(fair_ri),
        "fair_blend": round(fair_blend),
        "upside": round(upside, 4),
        "ri_list": [round(r, 1) for r in ri_list],
        "ri_terminal_pv": round(ri_terminal_pv, 1),
        # lịch sử cho chart
        "pe_hist": pe_hist, "pb_hist": pb_hist,
        "eps_vals": [round(v, 0) if v else None for v in eps_vals],
        "bvps_vals": [round(v, 0) if v else None for v in bvps_vals],
        "hist_years": list(hist_years),
    }


# ══════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER — 8 sheets
# ══════════════════════════════════════════════════════════════════════════
def _ws_freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_beta_coe_sheets(wb, ticker, beta_raw, beta_val, beta_src, aligned_data, rf_val, rf_src, erp, specific_rp, coe):
    ws_beta = wb.create_sheet("00_Beta")
    ws_beta.column_dimensions['A'].width = 15
    ws_beta.column_dimensions['B'].width = 16
    ws_beta.column_dimensions['C'].width = 22
    ws_beta.column_dimensions['D'].width = 16
    ws_beta.column_dimensions['E'].width = 22
    ws_beta.cell(row=1, column=1, value="BẢNG TÍNH HỆ SỐ BETA LỊCH SỬ").font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    ws_beta.cell(row=1, column=2, value="Beta thô (raw):").font = BOLD_FONT
    ws_beta.cell(row=1, column=4, value="Beta Blume (điều chỉnh):").font = BOLD_FONT
    ws_beta.cell(row=2, column=2, value="Số phiên giao dịch:").font = ITALIC_FONT
    header_row(ws_beta, 4, ["Ngày", f"Giá {ticker}", f"Tỷ suất sinh lời {ticker}", "Giá VNINDEX", "Tỷ suất sinh lời VNINDEX"])
    if aligned_data:
        date0, p_s0, p_m0 = aligned_data[0]
        ws_beta.cell(row=5, column=1, value=date0).font = DATA_FONT
        ws_beta.cell(row=5, column=2, value=p_s0).font = DATA_FONT
        ws_beta.cell(row=5, column=4, value=p_m0).font = DATA_FONT
        for ridx, (dstr, p_s, p_m) in enumerate(aligned_data[1:], start=6):
            ws_beta.cell(row=ridx, column=1, value=dstr).font = DATA_FONT
            ws_beta.cell(row=ridx, column=2, value=p_s).font = DATA_FONT
            ws_beta.cell(row=ridx, column=3, value=f"=(B{ridx}-B{ridx-1})/B{ridx-1}").number_format = '0.00%'
            ws_beta.cell(row=ridx, column=3).font = DATA_FONT
            ws_beta.cell(row=ridx, column=4, value=p_m).font = DATA_FONT
            ws_beta.cell(row=ridx, column=5, value=f"=(D{ridx}-D{ridx-1})/D{ridx-1}").number_format = '0.00%'
            ws_beta.cell(row=ridx, column=5).font = DATA_FONT
        last_row = 4 + len(aligned_data)
        n_sessions = len(aligned_data)
        window_start = max(6, last_row - 499) if n_sessions > 500 else 6
        ws_beta.cell(row=1, column=3, value=f"=COVAR(C{window_start}:C{last_row},E{window_start}:E{last_row})/VAR(E{window_start}:E{last_row})").number_format = '0.0000'
        ws_beta.cell(row=1, column=3).font = BOLD_FONT
        ws_beta.cell(row=1, column=5, value="=0.67*C1+0.33").number_format = '0.0000'
        ws_beta.cell(row=1, column=5).font = BOLD_FONT
        ws_beta.cell(row=2, column=3, value=f"=COUNT(C{window_start}:C{last_row})").font = ITALIC_FONT
    else:
        ws_beta.cell(row=1, column=3, value=beta_raw).number_format = '0.0000'
        ws_beta.cell(row=1, column=3).font = BOLD_FONT
        ws_beta.cell(row=1, column=5, value=beta_val).number_format = '0.0000'
        ws_beta.cell(row=1, column=5).font = BOLD_FONT
        ws_beta.cell(row=2, column=3, value=0).font = ITALIC_FONT
        
    for r in range(4, 4 + len(aligned_data) + 1):
        for c in range(1, 6):
            ws_beta.cell(row=r, column=c).border = THIN_BORDER
    _ws_freeze(ws_beta, "B5")
    print(f"[Excel] Sheet 00_Beta done ({len(aligned_data)} phiên).")

    ws_coe = wb.create_sheet("00_COE")
    ws_coe.column_dimensions['A'].width = 42
    ws_coe.column_dimensions['B'].width = 16
    ws_coe.column_dimensions['C'].width = 50
    ws_coe.cell(row=1, column=1, value="CHI PHÍ VỐN CHỦ SỞ HỮU (COE) - MÔ HÌNH CAPM").font = Font(name=FONT_NAME, size=12, bold=True, color="1F4E78")
    header_row(ws_coe, 3, ["Tham số", "Giá trị", "Ghi chú / Nguồn"])
    
    ws_coe.cell(row=4, column=1, value="Rf - Lãi suất phi rủi ro (TPCP 10 năm)").font = DATA_FONT
    ws_coe.cell(row=4, column=2, value=rf_val).number_format = FMT_PCT
    ws_coe.cell(row=4, column=2).font = DATA_FONT
    ws_coe.cell(row=4, column=3, value=rf_src).font = ITALIC_FONT
    
    ws_coe.cell(row=5, column=1, value="  Hệ số Beta (Blume-adjusted)").font = DATA_FONT
    ws_coe.cell(row=5, column=2, value="='00_Beta'!E1").number_format = '0.0000'
    ws_coe.cell(row=5, column=2).font = DATA_FONT
    ws_coe.cell(row=5, column=3, value=beta_src).font = ITALIC_FONT
    
    ws_coe.cell(row=6, column=1, value="ERP - Phần bù rủi ro vốn (Damodaran)").font = DATA_FONT
    ws_coe.cell(row=6, column=2, value=erp).number_format = FMT_PCT
    ws_coe.cell(row=6, column=2).font = DATA_FONT
    ws_coe.cell(row=6, column=3, value="Damodaran ERP cho Việt Nam").font = ITALIC_FONT
    
    ws_coe.cell(row=7, column=1, value="  Phần bù rủi ro đặc thù (Quốc gia/Ngành)").font = DATA_FONT
    ws_coe.cell(row=7, column=2, value=specific_rp).number_format = FMT_PCT
    ws_coe.cell(row=7, column=2).font = DATA_FONT
    ws_coe.cell(row=7, column=3, value="Phần bù rủi ro quốc gia/KCN đặc thù").font = ITALIC_FONT
    
    ws_coe.cell(row=9, column=1, value="COE = Rf + β × ERP + PBĐT").font = BOLD_FONT
    ws_coe.cell(row=9, column=2, value="=B4+B5*B6+B7").font = BOLD_FONT
    ws_coe.cell(row=9, column=2).number_format = FMT_PCT
    ws_coe.cell(row=9, column=2).fill = P_FILL
    ws_coe.cell(row=9, column=3, value="Chi phí vốn chủ sở hữu (CAPM điều chỉnh)").font = ITALIC_FONT
    
    for r in range(3, 10):
        for c in range(1, 4):
            ws_coe.cell(row=r, column=c).border = THIN_BORDER
    print("[Excel] Sheet 00_COE done.")


def build_pe_pb_history_sheet(wb, ticker, quarter_labels, pe_quarters, pb_quarters):
    ws = wb.create_sheet("03_PE_PB_History")
    header_row(ws, 1, ["Quý", "P/E (x)", "P/B (x)"], [14, 14, 14])
    r = 2
    row_start = r
    for i, label in enumerate(quarter_labels):
        ws.cell(row=r, column=1, value=label).font = DATA_FONT
        pe_v = pe_quarters[i] if i < len(pe_quarters) else None
        pb_v = pb_quarters[i] if i < len(pb_quarters) else None
        
        c_pe = ws.cell(row=r, column=2, value=pe_v)
        c_pe.font = DATA_FONT
        if pe_v is not None:
            c_pe.number_format = "0.00"
            
        c_pb = ws.cell(row=r, column=3, value=pb_v)
        c_pb.font = DATA_FONT
        if pb_v is not None:
            c_pb.number_format = "0.00"
        
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    row_end = r - 1
    
    # MEDIAN row
    ws.cell(row=r, column=1, value="MEDIAN").font = BOLD_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    
    c_pe_med = ws.cell(row=r, column=2, value=f"=MEDIAN(B{row_start}:B{row_end})")
    c_pe_med.font = BOLD_FONT
    c_pe_med.fill = HEADER_FILL
    c_pe_med.number_format = "0.00"
    
    c_pb_med = ws.cell(row=r, column=3, value=f"=MEDIAN(C{row_start}:C{row_end})")
    c_pb_med.font = BOLD_FONT
    c_pb_med.fill = HEADER_FILL
    c_pb_med.number_format = "0.00"
    
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = THIN_BORDER
    _ws_freeze(ws, "B2")
    print(f"[Excel] Sheet 03_PE_PB_History done ({len(quarter_labels)} quý).")


def build_excel_kcn(wb, ticker, company_name, current_price, shares,
                    hist_years, fc_years,
                    is_recs_y, bs_recs_y, cf_recs_y,
                    seg_names, seg_labels, seg_colors, yearly_seg, quarterly_seg,
                    seg_fc, val, ai_comments, beta_src, latest_price_hist, aligned_data,
                    quarter_labels, pe_quarters, pb_quarters):
    """Xây dựng các sheets Excel cho mô hình KCN."""
    # Xóa sheet mặc định ban đầu
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    # Tạo các sheet Beta và COE trước tiên
    beta_raw = val.get("beta")
    rf_val = val.get("rf")
    erp = val.get("erp", 0.07)
    specific_rp = 0.02
    coe = val.get("coe")
    build_beta_coe_sheets(wb, ticker, beta_raw, beta_raw, beta_src, aligned_data, rf_val, "TPCP 10Y VN", erp, specific_rp, coe)
    all_years = list(hist_years) + list(fc_years)
    n_hist = len(hist_years)
    n_fc   = len(fc_years)

    # ── Sheet 1: Cover ───────────────────────────────────────────────
    ws1 = wb.create_sheet("01_Cover")
    ws1["B2"] = f"MÔ HÌNH PHÂN TÍCH TÀI CHÍNH & ĐỊNH GIÁ — {ticker}"
    ws1["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
    ws1["B3"] = company_name
    ws1["B3"].font = Font(name=FONT_NAME, size=11, italic=True, color="4A5568")
    ws1["B4"] = f"Ngày lập: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["B4"].font = ITALIC_FONT

    meta = [
        ("Giá thị trường (VND)", current_price, FMT_PRICE),
        ("Giá lịch sử cuối (VND)", latest_price_hist or current_price, FMT_PRICE),
        ("Số CP lưu hành", shares, "#,##0"),
        ("Vốn hóa (tỷ VND)", f"=B6*B7/1e9", "#,##0.0"),
        ("Beta (CAPM)", "='02_Assumptions'!B3", "0.000"),
        ("Rf (%/năm)", "='02_Assumptions'!B2", FMT_PCT),
        ("COE (%/năm)", "='02_Assumptions'!B6", FMT_PCT),
        ("Nguồn Beta", beta_src, None),
    ]
    for i, (label, value, fmt) in enumerate(meta, start=6):
        ws1.cell(row=i, column=2, value=label).font = DATA_FONT
        c = ws1.cell(row=i, column=3, value=value)
        if fmt:
            c.number_format = fmt
        c.font = DATA_FONT

    # Định giá tóm tắt
    ws1["B15"] = "─── ĐỊNH GIÁ TỔNG HỢP ───"
    ws1["B15"].font = BOLD_FONT
    val_rows = [
        ("Giá hợp lý P/E (40%)",  val["fair_pe"]),
        ("Giá hợp lý P/B (40%)",  val["fair_pb"]),
        ("Giá hợp lý RI (20%)",   val["fair_ri"]),
        ("Giá mục tiêu blend",     val["fair_blend"]),
        ("Upside/Downside",        val["upside"]),
    ]
    for i, (lbl, v) in enumerate(val_rows, start=16):
        ws1.cell(row=i, column=2, value=lbl).font = DATA_FONT
        c = ws1.cell(row=i, column=3, value=v)
        c.font = BOLD_FONT if lbl.startswith("Giá mục tiêu") else DATA_FONT
        c.number_format = FMT_PCT if lbl.startswith("Upside") else FMT_PRICE
        if lbl.startswith("Giá mục tiêu"):
            c.fill = P_FILL

    ws1["B22"] = "Nhận định AI:"
    ws1["B22"].font = BOLD_FONT
    ws1["B23"] = ai_comments["business"]
    ws1["B23"].alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[23].height = 60
    _set_col_widths(ws1, [5, 35, 18, 5])
    _ws_freeze(ws1, "B2")

    # ── Sheet 2: Assumptions ─────────────────────────────────────────
    ws2 = wb.create_sheet("02_Assumptions")
    header_row(ws2, 1, ["Tham số giả định", "Giá trị", "Ghi chú"], [32, 14, 35])
    
    # Tạo sheet PE/PB History trước tiên
    build_pe_pb_history_sheet(wb, ticker, quarter_labels, pe_quarters, pb_quarters)

    ass_rows = [
        ("Rf (lãi TPCP 10Y, %/năm)",     "='00_COE'!B4",    "TPCP 10Y VN"),
        ("Beta (CAPM)",                   "='00_COE'!B5",    beta_src),
        ("ERP (phần bù rủi ro thị trường)", "='00_COE'!B6",    "Damodaran ERP cho VN"),
        ("Phần bù rủi ro đặc thù",        "='00_COE'!B7",    "Phần bù rủi ro quốc gia/ngành đặc thù"),
        ("COE = Rf + β × ERP + PBĐT",     "='00_COE'!B9",    "Cost of Equity (CAPM điều chỉnh)"),
        ("EPS lịch sử gần nhất (VND)",    val["eps_last"],  "Từ isa23 Vietcap"),
        ("BVPS mẹ lịch sử gần nhất (VND)", val["bvps_last"], "(bsa78 − bsa210) / shares"),
        ("Payout ratio trung bình",        val["avg_payout"], "Trung bình lịch sử cfa32/isa22"),
        ("CAGR EPS dự phóng",             val["eps_cagr"],  "CAGR 3 năm lịch sử, kẹp [-10%, +20%]"),
        ("ROE ước tính",                  "=B7/B8",         "EPS_last / BVPS_last"),
        ("P/E mục tiêu",                  f"=MAX(8, MIN(25, '03_PE_PB_History'!B{len(quarter_labels)+2}))", "Median P/E lịch sử, kẹp [8, 25]"),
        ("P/B mục tiêu",                  f"=MAX(0.6, MIN(4, '03_PE_PB_History'!C{len(quarter_labels)+2}))", "Median P/B lịch sử, kẹp [0.6, 4]"),
        ("Số năm RI dự phóng (n)",        3,                "PV RI 3 năm + terminal value"),
        ("g terminal RI",                 "=MAX(0, MIN(0.04, B10 * 0.5))", "Formula: min(CAGR_EPS × 0.5, 4%)"),
        ("Số CP lưu hành",                shares,           ""),
        ("Giá thị trường (VND)",          current_price,    "Giá cuối phiên gần nhất"),
    ]
    for i, (lbl, v, note) in enumerate(ass_rows, start=2):
        ws2.cell(row=i, column=1, value=lbl).font = DATA_FONT
        c = ws2.cell(row=i, column=2, value=v)
        c.fill = ASSUMP_FILL
        c.font = DATA_FONT
        if isinstance(v, float) and 0 < v < 1 and "Beta" not in lbl:
            c.number_format = FMT_PCT
        elif isinstance(v, float):
            c.number_format = "0.00"
        elif str(v).startswith("="):
            # Formula string
            if "%" in lbl or "COE" in lbl or "Rf" in lbl or "ERP" in lbl or "g terminal" in lbl:
                c.number_format = FMT_PCT
            elif "Beta" in lbl or "P/E" in lbl or "P/B" in lbl:
                c.number_format = "0.00"
        else:
            c.number_format = "#,##0"
        ws2.cell(row=i, column=3, value=note).font = ITALIC_FONT
    _ws_freeze(ws2, "A2")

    # ── Sheet 3: PnL lịch sử & dự phóng ─────────────────────────────
    ws3 = wb.create_sheet("03_PnL")
    col_labels = ["Chỉ tiêu P&L (tỷ VND)"] + [f"{y}A" for y in hist_years] + [f"{y}E" for y in fc_years]
    header_row(ws3, 1, col_labels, [32] + [12] * (n_hist + n_fc))

    def _yr(field, yr):
        return _get_yr(is_recs_y, yr, IS_GEN[field])

    def _yr_bs(field, yr):
        return _get_yr(bs_recs_y, yr, BS_GEN[field])

    rev_h   = [_yr("revenue", y) for y in hist_years]
    cogs_h  = [_yr("cogs", y) for y in hist_years]
    gp_h    = [r - c for r, c in zip(rev_h, cogs_h)]
    pbt_h   = [_yr("pbt", y) for y in hist_years]
    npat_h  = [_yr("npat", y) for y in hist_years]
    npat_p_h = [_yr("npat_parent", y) for y in hist_years]

    # Dự phóng đơn giản: tổng tất cả mảng
    def _fc_rev(base_rev, n):
        """Cộng dự phóng tất cả mảng, fallback constant nếu không đủ lịch sử."""
        rows = []
        cur_seg_rev = {seg: (max([yearly_seg[seg][y]["revenue"] for y in sorted(yearly_seg[seg]) if yearly_seg[seg][y].get("revenue")] or [0]) if yearly_seg[seg] else 0)
                       for seg in seg_names}
        for step in range(1, n + 1):
            total = 0
            for seg in seg_names:
                g = seg_fc[seg].get("g") or 0.05
                cap = seg_fc[seg].get("cap")
                if cap is not None:
                    g = min(g, cap)
                total += cur_seg_rev[seg] * ((1 + g) ** step)
            rows.append(round(total, 2))
        # Fallback nếu kho mảng trống: dùng CAGR tổng công ty
        if not any(rows):
            last_rev = rev_h[-1] if rev_h else base_rev
            rows = [round(last_rev * ((1 + 0.07) ** s), 2) for s in range(1, n + 1)]
        return rows

    rev_fc_list = _fc_rev(rev_h[-1] if rev_h else 1000, n_fc)
    # Giả định biên gộp = trung bình 2 năm lịch sử gần nhất
    gm_avg = stats.mean([gp_h[i] / rev_h[i] for i in range(len(rev_h)) if rev_h[i] > 0][-2:]) if rev_h else 0.25
    nm_avg = stats.mean([npat_h[i] / rev_h[i] for i in range(len(rev_h)) if rev_h[i] > 0][-2:]) if rev_h else 0.12
    cogs_fc_list  = [round(r * (1 - gm_avg), 2) for r in rev_fc_list]
    gp_fc_list    = [round(r * gm_avg, 2) for r in rev_fc_list]
    npat_fc_list  = [round(r * nm_avg, 2) for r in rev_fc_list]
    
    # Dự phóng LNTT và LNST cổ đông mẹ
    pbt_ratio = stats.mean([pbt_h[i] / gp_h[i] for i in range(len(gp_h)) if gp_h[i] > 0][-2:]) if (gp_h and len(gp_h) >= 2) else 0.20
    pbt_fc_list = [round(gp * pbt_ratio, 2) for gp in gp_fc_list]
    npat_p_ratio = stats.mean([npat_p_h[i] / npat_h[i] for i in range(len(npat_h)) if npat_h[i] > 0][-2:]) if (npat_h and len(npat_h) >= 2) else 0.85
    npat_p_fc_list = [round(npat * npat_p_ratio, 2) for npat in npat_fc_list]

    pnl_rows = [
        ("Doanh thu thuần",          rev_h,    rev_fc_list,  True),
        ("Giá vốn hàng bán",         cogs_h,   cogs_fc_list, False),
        ("Lợi nhuận gộp",            gp_h,     gp_fc_list,   True),
        ("Lợi nhuận trước thuế",     pbt_h,    pbt_fc_list,  False),
        ("LNST (hợp nhất)",          npat_h,   npat_fc_list, False),
        ("LNST cổ đông mẹ",          npat_p_h, npat_p_fc_list,  True),
    ]
    for ri, (label, hist_vals, fc_vals, bold) in enumerate(pnl_rows, start=2):
        row = [label] + [round(v, 2) if v else 0 for v in hist_vals] + [(round(v, 2) if v else 0) for v in fc_vals]
        ws3.append(row)
        ws3.cell(row=ri, column=1).font = BOLD_FONT if bold else DATA_FONT
        for ci in range(2, len(all_years) + 2):
            c = ws3.cell(row=ri, column=ci)
            c.number_format = FMT_NUM
            c.font = BOLD_FONT if bold else DATA_FONT
            if ci > n_hist + 1:  # Forecast columns
                c.fill = ASSUMP_FILL

    # Biên lợi nhuận gộp
    gm_row = ["Biên lợi nhuận gộp (%)"]
    for i, y in enumerate(hist_years):
        gm_row.append(round(gp_h[i] / rev_h[i], 4) if rev_h[i] else 0)
    for i in range(n_fc):
        col_idx = n_hist + i + 2
        col_letter = get_column_letter(col_idx)
        row3_gp  = get_column_letter(col_idx)
        row3_rev = get_column_letter(col_idx)
        gm_row.append(f"={row3_gp}{ri-3}/{row3_rev}{ri-5}" if ri > 6 else round(gm_avg, 4))
    ws3.append(gm_row)
    gm_ri = ws3.max_row
    ws3.cell(row=gm_ri, column=1).font = ITALIC_FONT
    for ci in range(2, len(all_years) + 2):
        ws3.cell(row=gm_ri, column=ci).number_format = FMT_PCT

    _ws_freeze(ws3, "B2")
    _set_col_widths(ws3, [32] + [12] * (n_hist + n_fc))

    # ── Sheet 4: Segments — Doanh thu theo mảng ──────────────────────
    ws4 = wb.create_sheet("04_Segments_DT")
    # Lấy tất cả kỳ có dữ liệu (năm + quý, sắp xếp)
    all_yearly_keys  = sorted(set().union(*[yearly_seg[s].keys() for s in seg_names]))
    all_q_keys       = sorted(set().union(*[quarterly_seg[s].keys() for s in seg_names]))

    header_row(ws4, 1, ["Mảng kinh doanh"] + [f"{k}A" for k in all_yearly_keys] + [f"{y}E" for y in fc_years],
               [30] + [11] * (len(all_yearly_keys) + n_fc))
    for ri_s, seg in enumerate(seg_names, start=2):
        ws4.cell(row=ri_s, column=1, value=seg_labels.get(seg, seg)).font = DATA_FONT
        
        # Historical
        for ci, yk in enumerate(all_yearly_keys, start=2):
            rev_val = round(yearly_seg[seg].get(yk, {}).get("revenue") or 0, 2)
            c = ws4.cell(row=ri_s, column=ci, value=rev_val)
            c.number_format = FMT_NUM
            c.font = DATA_FONT

        # Forecast formulas
        hist_last_ci = 1 + len(all_yearly_keys)
        g = seg_fc[seg].get("g") or 0.05
        cap = seg_fc[seg].get("cap")
        if cap is not None:
            g = min(g, cap)
            
        for step, fy in enumerate(fc_years, start=1):
            target_ci = hist_last_ci + step
            prev_col = get_column_letter(target_ci - 1)
            c = ws4.cell(row=ri_s, column=target_ci, value=f"={prev_col}{ri_s} * (1 + {g:.4f})")
            c.number_format = FMT_NUM
            c.font = DATA_FONT
            c.fill = ASSUMP_FILL

    # Tổng DT
    tot_row = len(seg_names) + 2
    ws4.cell(row=tot_row, column=1, value="TỔNG DOANH THU").font = BOLD_FONT
    ws4.cell(row=tot_row, column=1).fill = LINK_FILL
    for ci in range(2, len(all_yearly_keys) + n_fc + 2):
        col = get_column_letter(ci)
        c = ws4.cell(row=tot_row, column=ci, value=f"=SUM({col}2:{col}{len(seg_names)+1})")
        c.font = BOLD_FONT
        c.fill = LINK_FILL
        c.number_format = FMT_NUM

    # Ghi chú giả định tăng trưởng và YTD run-rate
    g_note_row = tot_row + 1
    ytd_note_row = tot_row + 2
    ws4.cell(row=g_note_row, column=1, value="Giả định g (CAGR/YTD blend)").font = ITALIC_FONT
    ws4.cell(row=ytd_note_row, column=1, value="YTD run-rate (tỷ, annualised)").font = ITALIC_FONT
    for ri_s_note, seg in enumerate(seg_names, start=2):
        g_used = seg_fc[seg].get("g") or 0.05
        cap_used = seg_fc[seg].get("cap")
        if cap_used is not None:
            g_used = min(g_used, cap_used)
        ytd_src = seg_fc[seg].get("ytd_growth")
        ytd_ann = seg_fc[seg].get("ytd_annualised")
        # Hiển thị g theo từng cột forecast
        hist_last_ci_note = 1 + len(all_yearly_keys)
        for step in range(1, n_fc + 1):
            ci_note = hist_last_ci_note + step
            c_g = ws4.cell(row=g_note_row, column=ci_note, value=g_used)
            c_g.number_format = "0.0%"
            c_g.font = ITALIC_FONT
            if ytd_src is not None:
                c_g.comment = None  # openpyxl basic — no comment API needed
        # YTD run-rate chỉ điền vào cột forecast đầu tiên
        if ytd_ann is not None:
            c_ytd = ws4.cell(row=ytd_note_row, column=hist_last_ci_note + 1, value=ytd_ann)
            c_ytd.number_format = FMT_NUM
            c_ytd.font = ITALIC_FONT
    _ws_freeze(ws4, "B2")
    _set_col_widths(ws4, [30] + [11] * (len(all_yearly_keys) + n_fc))

    # ── Sheet 5: Segments — Giá vốn theo mảng ────────────────────────
    ws5 = wb.create_sheet("05_Segments_GV")
    header_row(ws5, 1, ["Mảng kinh doanh (Giá vốn)"] + [f"{k}A" for k in all_yearly_keys] + [f"{y}E" for y in fc_years],
               [30] + [11] * (len(all_yearly_keys) + n_fc))
    for ri_s, seg in enumerate(seg_names, start=2):
        ws5.cell(row=ri_s, column=1, value=seg_labels.get(seg, seg)).font = DATA_FONT
        
        # Historical
        for ci, yk in enumerate(all_yearly_keys, start=2):
            cogs_val = round(yearly_seg[seg].get(yk, {}).get("cogs") or 0, 2)
            c = ws5.cell(row=ri_s, column=ci, value=cogs_val)
            c.number_format = FMT_NUM
            c.font = DATA_FONT
            
        # Forecast formula
        for step, fy in enumerate(fc_years, start=1):
            target_ci = hist_last_ci + step
            col = get_column_letter(target_ci)
            c = ws5.cell(row=ri_s, column=target_ci, value=f"='04_Segments_DT'!{col}{ri_s} * (1 - '06_Segments_GM'!{col}{ri_s})")
            c.number_format = FMT_NUM
            c.font = DATA_FONT
            c.fill = ASSUMP_FILL
            
    # Tổng GV
    ws5.cell(row=len(seg_names)+2, column=1, value="TỔNG GIÁ VỐN").font = BOLD_FONT
    ws5.cell(row=len(seg_names)+2, column=1).fill = LINK_FILL
    for ci in range(2, len(all_yearly_keys) + n_fc + 2):
        col = get_column_letter(ci)
        c = ws5.cell(row=len(seg_names)+2, column=ci, value=f"=SUM({col}2:{col}{len(seg_names)+1})")
        c.font = BOLD_FONT
        c.fill = LINK_FILL
        c.number_format = FMT_NUM
    _ws_freeze(ws5, "B2")

    # ── Sheet 6: Segments — Biên lợi nhuận gộp ───────────────────────
    ws6 = wb.create_sheet("06_Segments_GM")
    header_row(ws6, 1, ["Mảng kinh doanh (Biên LNG %)"] + [f"{k}A" for k in all_yearly_keys] + [f"{y}E" for y in fc_years],
               [30] + [11] * (len(all_yearly_keys) + n_fc))
    for ri_s, seg in enumerate(seg_names, start=2):
        ws6.cell(row=ri_s, column=1, value=seg_labels.get(seg, seg)).font = DATA_FONT
        
        # Historical formula
        for ci, yk in enumerate(all_yearly_keys, start=2):
            col = get_column_letter(ci)
            c = ws6.cell(row=ri_s, column=ci, value=f"=IF('04_Segments_DT'!{col}{ri_s}=0, 0, ('04_Segments_DT'!{col}{ri_s} - '05_Segments_GV'!{col}{ri_s}) / '04_Segments_DT'!{col}{ri_s})")
            c.number_format = FMT_PCT
            c.font = DATA_FONT
            
        # Forecast target GM
        for step, fy in enumerate(fc_years, start=1):
            target_ci = hist_last_ci + step
            gm_fc = seg_fc[seg].get("gm_fc") or 0.20
            c = ws6.cell(row=ri_s, column=target_ci, value=round(gm_fc, 4))
            c.number_format = FMT_PCT
            c.font = DATA_FONT
            c.fill = ASSUMP_FILL
    _ws_freeze(ws6, "B2")

    # ── Sheet 7: Balance Sheet ────────────────────────────────────────
    ws7 = wb.create_sheet("07_BalanceSheet")
    bs_col_labels = ["Chỉ tiêu BS (tỷ VND)"] + [f"{y}A" for y in hist_years]
    header_row(ws7, 1, bs_col_labels, [32] + [12] * n_hist)
    bs_items = [
        ("Tiền & tương đương", "cash"),
        ("Tổng tài sản", "total_assets"),
        ("Vay ngắn hạn", "short_borrow"),
        ("Vay dài hạn", "long_borrow"),
        ("Tổng nợ phải trả", "total_liab"),
        ("VCSH (bao gồm NCI)", "equity_total"),
        ("Vốn góp (charter capital)", "charter_capital"),
        ("NCI (cổ đông thiểu số)", "nci"),
    ]
    for ri_b, (label, fkey) in enumerate(bs_items, start=2):
        row = [label] + [round(_get_yr(bs_recs_y, y, BS_GEN[fkey]), 2) for y in hist_years]
        ws7.append(row)
        ws7.cell(row=ri_b, column=1).font = DATA_FONT
        for ci in range(2, n_hist + 2):
            ws7.cell(row=ri_b, column=ci).number_format = FMT_NUM

    # VCSH mẹ = bsa78 - bsa210 (cho BVPS/RI)
    vcsh_parent_row = ["VCSH cổ đông mẹ (= VCSH − NCI)"]
    for y in hist_years:
        eq_total = _get_yr(bs_recs_y, y, BS_GEN["equity_total"])
        nci_val  = _get_yr(bs_recs_y, y, BS_GEN["nci"])
        vcsh_parent_row.append(round(eq_total - nci_val, 2))
    ws7.append(vcsh_parent_row)
    last_r = ws7.max_row
    ws7.cell(last_r, 1).font = BOLD_FONT
    ws7.cell(last_r, 1).fill = P_FILL
    for ci in range(2, n_hist + 2):
        ws7.cell(last_r, ci).number_format = FMT_NUM
        ws7.cell(last_r, ci).font = BOLD_FONT
        ws7.cell(last_r, ci).fill = P_FILL

    # BVPS, EPS theo năm
    ws7.append([])
    bvps_row = ["BVPS mẹ (VND/cp)"]
    eps_row  = ["EPS cơ bản (VND/cp)"]
    payout_row = ["Payout ratio lịch sử"]
    for y in hist_years:
        bvps_row.append(round(_bvps_parent(bs_recs_y, y, shares) or 0, 0))
        eps_row.append(round(_eps_parent(is_recs_y, y) or 0, 0))
        payout_row.append(round(_payout_ratio(cf_recs_y, is_recs_y, y), 4))
    for row_data, fmt in [(bvps_row, FMT_PRICE), (eps_row, FMT_PRICE), (payout_row, FMT_PCT)]:
        ws7.append(row_data)
        r = ws7.max_row
        ws7.cell(r, 1).font = DATA_FONT
        for ci in range(2, n_hist + 2):
            ws7.cell(r, ci).number_format = fmt
    _ws_freeze(ws7, "B2")
    _set_col_widths(ws7, [32] + [12] * n_hist)

    # ── Sheet 8: Định giá ─────────────────────────────────────────────
    ws8 = wb.create_sheet("08_Valuation")
    header_row(ws8, 1, ["Định giá KCN (40%P/E + 40%P/B + 20%RI)", "Giá trị", "Ghi chú"], [38, 16, 40])
    val_items = [
        ("─ COE & CAPM ─",          "",     ""),
        ("Rf (lãi suất phi rủi ro)", "='00_COE'!B4",           "TPCP 10Y VN"),
        ("Beta",                     "='00_COE'!B5",           beta_src),
        ("ERP",                      "='00_COE'!B6",           "Phần bù rủi ro TTCK VN"),
        ("Phần bù rủi ro đặc thù",    "='00_COE'!B7",           "Phần bù rủi ro quốc gia/ngành đặc thù"),
        ("COE = Rf + β × ERP + PBĐT", "='00_COE'!B9",           "Chi phí vốn chủ sở hữu"),
        ("─ P/E ─",                  "",     ""),
        ("EPS forward 1Y (VND)",     round(val["eps_fc1"]),    "Formula: EPS_last × (1 + CAGR EPS)"),
        ("P/E mục tiêu",             "=02_Assumptions!B12",    "Median P/E lịch sử, kẹp [8,25]"),
        ("Giá hợp lý P/E",           "=B9*B10",                "Formula: P/E × EPS_fc1"),
        ("─ P/B ─",                  "",     ""),
        ("BVPS forward 1Y (VND)",    round(val["bvps_fc1"]),   "Roll-forward trừ cổ tức"),
        ("P/B mục tiêu",             "=02_Assumptions!B13",    "Median P/B lịch sử, kẹp [0.6,4]"),
        ("Giá hợp lý P/B",           "=B13*B14",                "Formula: P/B × BVPS_fc1"),
        ("─ Residual Income ─",      "",     ""),
        ("BVPS hiện tại (VND)",      "=02_Assumptions!B8",     "bsa78 − bsa210 / shares"),
        ("CAGR EPS",                 "=02_Assumptions!B10",    "Kẹp [−10%, +20%]"),
        ("RI Năm 1 (VND)",           val["ri_list"][0] if val["ri_list"] else 0, "PV EPS1 − COE×BVPS0"),
        ("RI Năm 2 (VND)",           val["ri_list"][1] if len(val["ri_list"]) > 1 else 0, "PV EPS2 − COE×BVPS1"),
        ("RI Năm 3 (VND)",           val["ri_list"][2] if len(val["ri_list"]) > 2 else 0, "PV EPS3 − COE×BVPS2"),
        ("RI Terminal (PV, VND)",    val["ri_terminal_pv"],    "Formula: TV/(COE-g)"),
        ("Giá hợp lý RI (VND)",      "=B17+B19+B20+B21+B22",   "Formula: BVPS + ΣRI_PV + Terminal_PV"),
        ("─ KẾT QUẢ ─",             "",     ""),
        ("Giá thị trường (VND)",     "=02_Assumptions!B17",    ""),
        ("Giá mục tiêu (40+40+20)",  "=0.4*B11+0.4*B15+0.2*B23", "Tổng hợp 3 phương pháp"),
        ("Upside / Downside",        "=(B26-B25)/B25",         "+/- % so với giá thị trường"),
    ]
    for ri_v, (label, value, note) in enumerate(val_items, start=2):
        bold = label.startswith("─") or label in ("Giá mục tiêu (40+40+20)", "Upside / Downside")
        ws8.cell(row=ri_v, column=1, value=label).font = BOLD_FONT if bold else DATA_FONT
        c = ws8.cell(row=ri_v, column=2, value=value)
        c.font = BOLD_FONT if bold else DATA_FONT
        if label.startswith("─"):
            c.fill = HEADER_FILL
            ws8.cell(ri_v, 1).fill = HEADER_FILL
            ws8.cell(ri_v, 1).font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        elif label in ("Giá mục tiêu (40+40+20)", "Upside / Downside"):
            c.fill = P_FILL
            ws8.cell(ri_v, 1).fill = P_FILL
            
        if isinstance(value, (int, float)):
            if "%" in label or label in ("COE = Rf + β × ERP + PBĐT", "CAGR EPS"):
                c.number_format = FMT_PCT
            elif isinstance(value, float) and 0 < abs(value) < 10:
                c.number_format = "0.00"
            else:
                c.number_format = FMT_PRICE
        elif str(value).startswith("="):
            if "%" in label or label == "Upside / Downside" or "COE" in label:
                c.number_format = FMT_PCT
            elif "Beta" in label or "P/E" in label or "P/B" in label:
                c.number_format = "0.00"
            else:
                c.number_format = FMT_PRICE
        ws8.cell(row=ri_v, column=3, value=note).font = ITALIC_FONT
    _set_col_widths(ws8, [38, 16, 45])

    # ── Sheet 9: Quarterly — Doanh thu ──────────────────────────────
    ws9 = wb.create_sheet("09_Quarterly_DT")
    all_qkeys = sorted(set().union(*[quarterly_seg[s].keys() for s in seg_names]))
    header_row(ws9, 1, ["Mảng / Quý"] + all_qkeys + ["Trung bình"],
               [28] + [12] * len(all_qkeys) + [14])
    for ri_q, seg in enumerate(seg_names, start=2):
        ws9.cell(row=ri_q, column=1, value=seg_labels.get(seg, seg)).font = DATA_FONT
        for ci, qk in enumerate(all_qkeys, start=2):
            rev_q = quarterly_seg[seg].get(qk, {}).get("revenue")
            c = ws9.cell(row=ri_q, column=ci, value=round(rev_q, 2) if rev_q is not None else "")
            c.number_format = FMT_NUM
            c.font = DATA_FONT
        # average formula
        last_col = get_column_letter(len(all_qkeys) + 1)
        avg_c = ws9.cell(row=ri_q, column=len(all_qkeys) + 2, value=f"=AVERAGE(B{ri_q}:{last_col}{ri_q})")
        avg_c.number_format = FMT_NUM
        avg_c.font = BOLD_FONT
        avg_c.fill = LINK_FILL
    _ws_freeze(ws9, "B2")

    # ── Sheet 10: Quarterly — Giá vốn ───────────────────────────────
    ws10 = wb.create_sheet("10_Quarterly_GV")
    header_row(ws10, 1, ["Mảng / Quý"] + all_qkeys + ["Trung bình"],
               [28] + [12] * len(all_qkeys) + [14])
    for ri_q, seg in enumerate(seg_names, start=2):
        ws10.cell(row=ri_q, column=1, value=seg_labels.get(seg, seg)).font = DATA_FONT
        for ci, qk in enumerate(all_qkeys, start=2):
            cogs_q = quarterly_seg[seg].get(qk, {}).get("cogs")
            c = ws10.cell(row=ri_q, column=ci, value=round(cogs_q, 2) if cogs_q is not None else "")
            c.number_format = FMT_NUM
            c.font = DATA_FONT
        # average formula
        last_col = get_column_letter(len(all_qkeys) + 1)
        avg_c = ws10.cell(row=ri_q, column=len(all_qkeys) + 2, value=f"=AVERAGE(B{ri_q}:{last_col}{ri_q})")
        avg_c.number_format = FMT_NUM
        avg_c.font = BOLD_FONT
        avg_c.fill = LINK_FILL
    _ws_freeze(ws10, "B2")

    # ── Sheet 11: Quarterly — Biên lợi nhuận gộp ─────────────────────
    ws11 = wb.create_sheet("11_Quarterly_GM")
    header_row(ws11, 1, ["Mảng / Quý (Biên LNG %)"] + all_qkeys + ["Trung bình"],
               [28] + [12] * len(all_qkeys) + [14])
    for ri_q, seg in enumerate(seg_names, start=2):
        ws11.cell(row=ri_q, column=1, value=seg_labels.get(seg, seg)).font = DATA_FONT
        for ci, qk in enumerate(all_qkeys, start=2):
            col = get_column_letter(ci)
            c = ws11.cell(row=ri_q, column=ci, value=f"=IF('09_Quarterly_DT'!{col}{ri_q}=0, 0, ('09_Quarterly_DT'!{col}{ri_q} - '10_Quarterly_GV'!{col}{ri_q}) / '09_Quarterly_DT'!{col}{ri_q})")
            c.number_format = FMT_PCT
            c.font = DATA_FONT
        # average formula
        last_col = get_column_letter(len(all_qkeys) + 1)
        avg_c = ws11.cell(row=ri_q, column=len(all_qkeys) + 2, value=f"=AVERAGE(B{ri_q}:{last_col}{ri_q})")
        avg_c.number_format = FMT_PCT
        avg_c.font = BOLD_FONT
        avg_c.fill = LINK_FILL
    _ws_freeze(ws11, "B2")

    return wb


# ══════════════════════════════════════════════════════════════════════════
# CHART GENERATOR
# ══════════════════════════════════════════════════════════════════════════
def build_charts_kcn(out_dir, ticker, hist_years, fc_years,
                     rev_h, gp_h, npat_h,
                     seg_names, seg_labels, seg_colors, yearly_seg,
                     quarterly_seg, val):
    """Tạo 4 biểu đồ matplotlib → PNG. Trả về dict {chart_name: path}."""
    charts = {}
    all_years = list(hist_years) + list(fc_years)

    # ── Biểu đồ 1: Doanh thu & LNST toàn công ty (cột + đường) ──────
    p1 = os.path.join(out_dir, f"{ticker}_chart1_revenue_npat.png")
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    x = range(len(hist_years))
    ax1.bar(x, rev_h, color="#1F4E78", alpha=0.85, label="Doanh thu (tỷ)")
    ax1.bar(x, gp_h, color="#2E75B6", alpha=0.75, label="LN gộp (tỷ)")
    ax2 = ax1.twinx()
    ax2.plot(x, npat_h, color="#C00000", marker="o", linewidth=2, markersize=5, label="LNST (tỷ)")
    ax1.set_xticks(list(x))
    ax1.set_xticklabels([str(y) for y in hist_years], fontsize=9)
    ax1.set_ylabel("Tỷ VND", fontsize=9)
    ax2.set_ylabel("LNST (tỷ VND)", fontsize=9, color="#C00000")
    ax1.set_title(f"Doanh thu & Lợi nhuận gộp — {ticker}", fontsize=11, fontweight="bold")
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=8)
    plt.tight_layout()
    plt.savefig(p1, dpi=130)
    plt.close()
    charts["revenue_npat"] = p1

    # ── Biểu đồ 2: Cơ cấu doanh thu theo mảng (stacked bar, năm) ────
    p2 = os.path.join(out_dir, f"{ticker}_chart2_segments_stacked.png")
    all_yk = sorted(set().union(*[yearly_seg[s].keys() for s in seg_names]))
    if all_yk:
        fig, ax = plt.subplots(figsize=(8, 4.5))
        bottoms = [0.0] * len(all_yk)
        x_idx = range(len(all_yk))
        for seg in seg_names:
            vals = [yearly_seg[seg].get(yk, {}).get("revenue") or 0 for yk in all_yk]
            color = seg_colors.get(seg, "#94a3b8")
            ax.bar(x_idx, vals, bottom=bottoms, label=seg_labels.get(seg, seg), color=color, alpha=0.85)
            bottoms = [bottoms[i] + vals[i] for i in range(len(all_yk))]
        ax.set_xticks(list(x_idx))
        ax.set_xticklabels(all_yk, fontsize=9)
        ax.set_ylabel("Tỷ VND", fontsize=9)
        ax.set_title(f"Cơ cấu Doanh thu theo Mảng — {ticker}", fontsize=11, fontweight="bold")
        ax.legend(loc="upper left", fontsize=7, ncol=2)
        plt.tight_layout()
        plt.savefig(p2, dpi=130)
        plt.close()
        charts["segments_stacked"] = p2

    # ── Biểu đồ 3: Biên LNG mỗi mảng theo quý (line chart) ──────────
    p3 = os.path.join(out_dir, f"{ticker}_chart3_segment_gm_quarterly.png")
    all_qk = sorted(set().union(*[quarterly_seg[s].keys() for s in seg_names]))
    if all_qk and len(all_qk) >= 2:
        fig, ax = plt.subplots(figsize=(9, 4.5))
        for seg in seg_names:
            gm_vals = []
            for qk in all_qk:
                gm = quarterly_seg[seg].get(qk, {}).get("grossMargin")
                gm_vals.append(gm if gm is not None else float("nan"))
            color = seg_colors.get(seg, "#94a3b8")
            ax.plot(range(len(all_qk)), gm_vals, marker="o", linewidth=1.8,
                    markersize=4, label=seg_labels.get(seg, seg), color=color)
        ax.set_xticks(range(len(all_qk)))
        ax.set_xticklabels(all_qk, fontsize=8, rotation=30, ha="right")
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y*100:.0f}%"))
        ax.set_ylabel("Biên LNG (%)", fontsize=9)
        ax.set_title(f"Biên Lợi Nhuận Gộp theo Mảng (quý) — {ticker}", fontsize=11, fontweight="bold")
        ax.legend(loc="upper right", fontsize=7, ncol=2)
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        plt.tight_layout()
        plt.savefig(p3, dpi=130)
        plt.close()
        charts["segment_gm"] = p3

    # ── Biểu đồ 4: Định giá — waterfall / bar ────────────────────────
    p4 = os.path.join(out_dir, f"{ticker}_chart4_valuation.png")
    labels_val = ["Giá TT", "Fair P/E", "Fair P/B", "Fair RI", "Mục tiêu"]
    values_val = [val["fair_blend"] / val.get("fair_pe", val["fair_blend"]),  # ratio placeholder
                  val["fair_pe"], val["fair_pb"], val["fair_ri"], val["fair_blend"]]
    values_abs = [round(val.get("current_price_chart", 0)), val["fair_pe"], val["fair_pb"], val["fair_ri"], val["fair_blend"]]
    colors_val = ["#475569", "#2563eb", "#0891b2", "#7c3aed", "#C00000"]
    fig, ax = plt.subplots(figsize=(7, 4))
    bars = ax.bar(labels_val, [0, val["fair_pe"], val["fair_pb"], val["fair_ri"], val["fair_blend"]],
                  color=colors_val, alpha=0.85, width=0.5)
    for bar, v in zip(bars, [0, val["fair_pe"], val["fair_pb"], val["fair_ri"], val["fair_blend"]]):
        if v > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, v + max(val["fair_blend"] * 0.01, 100),
                    f"{v:,.0f}", ha="center", va="bottom", fontsize=8.5, fontweight="bold")
    ax.set_ylabel("VND/cp", fontsize=9)
    ax.set_title(f"So sánh Định giá — {ticker}", fontsize=11, fontweight="bold")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y/1000:.0f}k"))
    plt.tight_layout()
    plt.savefig(p4, dpi=130)
    plt.close()
    charts["valuation"] = p4

    return charts


# ══════════════════════════════════════════════════════════════════════════
# PDF REPORT BUILDER
# ══════════════════════════════════════════════════════════════════════════
def build_pdf_kcn(pdf_path, ticker, company_name, current_price, shares,
                  hist_years, all_years, rev_h, gp_h, npat_h,
                  seg_names, seg_labels, yearly_seg, quarterly_seg,
                  val, ai_comments, charts, beta_src):
    """Tạo báo cáo PDF đầy đủ cho cổ phiếu KCN."""
    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()

    # Style definitions
    title_st = ParagraphStyle("KCN_Title", parent=styles["Heading1"],
        fontName=FONT_BOLD, fontSize=18, leading=22,
        textColor=HexColor("#1F4E78"), spaceAfter=12)
    h1_st = ParagraphStyle("KCN_H1", parent=styles["Heading2"],
        fontName=FONT_BOLD, fontSize=13, leading=17,
        textColor=HexColor("#2E75B6"), spaceBefore=14, spaceAfter=7)
    h2_st = ParagraphStyle("KCN_H2", parent=styles["Heading3"],
        fontName=FONT_BOLD, fontSize=11, leading=15,
        textColor=HexColor("#404040"), spaceBefore=8, spaceAfter=4)
    body_st = ParagraphStyle("KCN_Body", parent=styles["Normal"],
        fontName=FONT_REG, fontSize=10, leading=14,
        textColor=HexColor("#2D3748"), spaceAfter=6)
    italic_st = ParagraphStyle("KCN_Italic", parent=styles["Normal"],
        fontName=FONT_REG, fontSize=9, leading=12,
        textColor=HexColor("#718096"), italic=True)

    BLUE_DARK  = HexColor("#1F4E78")
    BLUE_MID   = HexColor("#2E75B6")
    LIGHT_BLUE = HexColor("#DDEBF7")
    YELLOW     = HexColor("#FFF2CC")
    RED        = HexColor("#C00000")

    def tbl_style(header_col=BLUE_DARK, alt_col=LIGHT_BLUE, font_size=9):
        return TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), header_col),
            ("TEXTCOLOR",     (0, 0), (-1, 0), white),
            ("FONTNAME",      (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE",      (0, 0), (-1, -1), font_size),
            ("FONTNAME",      (0, 1), (-1, -1), FONT_REG),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [white, alt_col]),
            ("ALIGN",         (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",         (0, 0), (0, -1), "LEFT"),
            ("GRID",          (0, 0), (-1, -1), 0.4, HexColor("#CBD5E1")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ])

    story = []

    # ── Trang bìa ──────────────────────────────────────────────────
    story.append(Paragraph(f"PHÂN TÍCH & ĐỊNH GIÁ CỔ PHIẾU: {ticker}", title_st))
    story.append(Paragraph(
        f"<b>{company_name}</b> | Nhóm: Bất động sản khu công nghiệp (KCN) | "
        f"Ngày lập: {datetime.datetime.now().strftime('%d/%m/%Y')}",
        body_st))
    story.append(Spacer(1, 8))

    # Bảng tóm tắt định giá
    upside_pct = f"{val['upside']*100:+.1f}%"
    upside_color = "#16a34a" if val["upside"] >= 0 else "#dc2626"
    summary_data = [
        ["Mã CP", "Giá TT (VND)", "Vốn hóa (tỷ)", "Mục tiêu (VND)", "Upside/Down"],
        [ticker,
         f"{current_price:,.0f}",
         f"{shares * current_price / 1e9:,.1f}",
         f"{val['fair_blend']:,.0f}",
         upside_pct],
    ]
    t_sum = Table(summary_data, colWidths=[22*mm, 35*mm, 32*mm, 40*mm, 32*mm])
    t_sum.setStyle(tbl_style())
    story.append(t_sum)
    story.append(Spacer(1, 12))

    # ── 1. Mô hình kinh doanh ──────────────────────────────────────
    story.append(Paragraph("1. Mô hình kinh doanh & Vị thế cạnh tranh", h1_st))
    story.append(Paragraph(ai_comments["business"], body_st))

    # ── 2. Kết quả tài chính lịch sử ──────────────────────────────
    story.append(Paragraph("2. Kết quả tài chính lịch sử", h1_st))
    story.append(Paragraph(ai_comments["financial"], body_st))

    # Bảng P&L
    pnl_header = ["Chỉ tiêu (tỷ VND)"] + [str(y) for y in hist_years]
    pnl_rows_pdf = [
        ["Doanh thu thuần"] + [f"{v:,.1f}" for v in rev_h],
        ["Lợi nhuận gộp"]  + [f"{v:,.1f}" for v in gp_h],
        ["Biên LNG (%)"]   + [f"{gp_h[i]/rev_h[i]*100:.1f}%" if rev_h[i] else "—" for i in range(len(hist_years))],
        ["LNST"]           + [f"{v:,.1f}" for v in npat_h],
    ]
    col_w_pnl = [42*mm] + [23*mm] * len(hist_years)
    t_pnl = Table([pnl_header] + pnl_rows_pdf, colWidths=col_w_pnl)
    t_pnl.setStyle(tbl_style())
    story.append(t_pnl)
    story.append(Spacer(1, 8))

    if "revenue_npat" in charts:
        story.append(Paragraph("Biểu đồ doanh thu & lợi nhuận:", h2_st))
        story.append(Image(charts["revenue_npat"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 10))

    # ── 3. Phân tích theo mảng ────────────────────────────────────
    story.append(Paragraph("3. Phân tích doanh thu & biên LNG theo mảng", h1_st))
    all_yk = sorted(set().union(*[yearly_seg[s].keys() for s in seg_names]))
    if all_yk:
        seg_header = ["Mảng"] + list(all_yk) + ["Biên LNG TB (%)"]
        seg_rows_pdf = []
        for seg in seg_names:
            row = [seg_labels.get(seg, seg)]
            gm_list = []
            for yk in all_yk:
                rv = yearly_seg[seg].get(yk, {}).get("revenue")
                row.append(f"{rv:,.1f}" if rv is not None else "—")
                gm = yearly_seg[seg].get(yk, {}).get("grossMargin")
                if gm is not None:
                    gm_list.append(gm)
            row.append(f"{stats.mean(gm_list)*100:.1f}%" if gm_list else "—")
            seg_rows_pdf.append(row)
        col_w_seg = [38*mm] + [max(14*mm, 18*mm) for _ in all_yk] + [22*mm]
        t_seg = Table([seg_header] + seg_rows_pdf, colWidths=col_w_seg)
        t_seg.setStyle(tbl_style())
        story.append(t_seg)
        story.append(Spacer(1, 8))

    if "segments_stacked" in charts:
        story.append(Image(charts["segments_stacked"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 8))
    if "segment_gm" in charts:
        story.append(Paragraph("Biên lợi nhuận gộp theo mảng (quý):", h2_st))
        story.append(Image(charts["segment_gm"], width=155*mm, height=78*mm))
        story.append(Spacer(1, 10))

    # ── 4. Dữ liệu quý gần nhất ──────────────────────────────────
    all_qk = sorted(set().union(*[quarterly_seg[s].keys() for s in seg_names]))
    if all_qk:
        story.append(Paragraph("4. Dữ liệu quý gần nhất theo mảng (doanh thu, tỷ VND)", h1_st))
        latest_qk = all_qk[-min(4, len(all_qk)):]
        q_header = ["Mảng"] + list(latest_qk)
        q_rows_pdf = []
        for seg in seg_names:
            row = [seg_labels.get(seg, seg)]
            for qk in latest_qk:
                rv = quarterly_seg[seg].get(qk, {}).get("revenue")
                row.append(f"{rv:,.1f}" if rv is not None else "—")
            q_rows_pdf.append(row)
        col_w_q = [40*mm] + [max(18*mm, 22*mm) for _ in latest_qk]
        t_q = Table([q_header] + q_rows_pdf, colWidths=col_w_q)
        t_q.setStyle(tbl_style())
        story.append(t_q)
        story.append(Spacer(1, 10))

    # ── 5. Định giá ──────────────────────────────────────────────
    story.append(Paragraph("5. Định giá (40% P/E + 40% P/B + 20% RI)", h1_st))
    story.append(Paragraph(ai_comments["valuation"], body_st))
    story.append(Spacer(1, 6))

    val_tbl_data = [
        ["Phương pháp", "Trọng số", "Giá hợp lý (VND)", "Ghi chú"],
        ["P/E", "40%", f"{val['fair_pe']:,.0f}", f"P/E mt = {val['target_pe']:.1f}x, EPS fc = {val['eps_fc1']:,.0f}"],
        ["P/B", "40%", f"{val['fair_pb']:,.0f}", f"P/B mt = {val['target_pb']:.1f}x, BVPS fc = {val['bvps_fc1']:,.0f}"],
        ["Residual Income", "20%", f"{val['fair_ri']:,.0f}", f"COE = {val['coe']*100:.2f}%, n = 3 năm"],
        [f"GIÁ MỤC TIÊU — {ticker}", "100%", f"{val['fair_blend']:,.0f}", f"Upside: {val['upside']*100:+.1f}%"],
    ]
    t_val = Table(val_tbl_data, colWidths=[38*mm, 20*mm, 38*mm, 65*mm])
    ts_val = tbl_style()
    ts_val.add("BACKGROUND", (0, 4), (-1, 4), YELLOW)
    ts_val.add("FONTNAME",   (0, 4), (-1, 4), FONT_BOLD)
    t_val.setStyle(ts_val)
    story.append(t_val)
    story.append(Spacer(1, 8))

    if "valuation" in charts:
        story.append(Image(charts["valuation"], width=130*mm, height=74*mm))
        story.append(Spacer(1, 8))

    # CAPM detail
    story.append(Paragraph(
        f"CAPM: Rf = {val['rf']*100:.2f}% (TPCP 10Y VN) | Beta = {val['beta']:.3f} ({beta_src}) "
        f"| ERP = 5.5% | COE = {val['coe']*100:.2f}% | Payout avg = {val['avg_payout']*100:.1f}%",
        italic_st))
    story.append(Spacer(1, 12))

    # ── Footer ──────────────────────────────────────────────────────
    story.append(Paragraph(
        "Báo cáo được tạo tự động bởi hệ thống Phân tích FA. Thông tin chỉ phục vụ mục đích tham khảo, "
        "không phải khuyến nghị đầu tư. Dữ liệu nguồn: Vietcap IQ API, BCTC hợp nhất (CafeF FileBCTC.ashx), "
        "lịch sử giá VNDirect.",
        italic_st))

    doc.build(story)
    print(f"  [OK] PDF: {pdf_path}")


# ══════════════════════════════════════════════════════════════════════════
# JSON SAVER — cho dashboard kcn.html / app_kcn.js
# ══════════════════════════════════════════════════════════════════════════
def save_json_kcn(ticker, company_name, current_price, shares,
                  hist_years, rev_h, gp_h, npat_h,
                  seg_names, seg_labels, seg_colors,
                  yearly_seg, quarterly_seg,
                  val, ai_comments, excel_url=None, pdf_url=None):
    """Ghi data/TICKER.json cho dashboard web."""
    # Cơ cấu doanh thu theo mảng (năm)
    all_yk = sorted(set().union(*[yearly_seg[s].keys() for s in seg_names]))
    seg_yearly_out = {}
    for seg in seg_names:
        seg_yearly_out[seg] = {
            "label": seg_labels.get(seg, seg),
            "color": seg_colors.get(seg, "#94a3b8"),
            "data": {yk: {
                "revenue":     yearly_seg[seg].get(yk, {}).get("revenue"),
                "cogs":        yearly_seg[seg].get(yk, {}).get("cogs"),
                "grossProfit": yearly_seg[seg].get(yk, {}).get("grossProfit"),
                "grossMargin": yearly_seg[seg].get(yk, {}).get("grossMargin"),
            } for yk in all_yk}
        }

    all_qk = sorted(set().union(*[quarterly_seg[s].keys() for s in seg_names]))
    seg_quarterly_out = {}
    for seg in seg_names:
        seg_quarterly_out[seg] = {
            "label": seg_labels.get(seg, seg),
            "color": seg_colors.get(seg, "#94a3b8"),
            "data": {qk: {
                "revenue":     quarterly_seg[seg].get(qk, {}).get("revenue"),
                "cogs":        quarterly_seg[seg].get(qk, {}).get("cogs"),
                "grossProfit": quarterly_seg[seg].get(qk, {}).get("grossProfit"),
                "grossMargin": quarterly_seg[seg].get(qk, {}).get("grossMargin"),
            } for qk in all_qk}
        }

    out = {
        "ticker":      ticker,
        "companyName": company_name,
        "sector":      "Bất động sản khu công nghiệp",
        "currentPrice": current_price,
        "marketCap":   shares * current_price,
        "shares":      shares,
        "gdriveExcelUrl": excel_url,
        "gdrivePdfUrl":   pdf_url,
        # Tài chính tổng hợp
        "data": {
            "years":   list(hist_years),
            "revenue": [round(v, 2) for v in rev_h],
            "grossProfit": [round(v, 2) for v in gp_h],
            "npat":    [round(v, 2) for v in npat_h],
        },
        # Định giá
        "valuation": {
            "fair_pe":    val["fair_pe"],
            "fair_pb":    val["fair_pb"],
            "fair_ri":    val["fair_ri"],
            "fair_blend": val["fair_blend"],
            "upside":     val["upside"],
            "target_pe":  val["target_pe"],
            "target_pb":  val["target_pb"],
            "coe":        round(val["coe"], 4),
            "rf":         round(val["rf"], 4),
            "beta":       round(val["beta"], 3),
            "eps_last":   round(val["eps_last"], 0),
            "bvps_last":  round(val["bvps_last"], 0),
            "avg_payout": round(val["avg_payout"], 4),
        },
        # Mảng kinh doanh
        "segments_yearly":    seg_yearly_out,
        "segments_quarterly": seg_quarterly_out,
        "yearly_keys":    all_yk,
        "quarterly_keys": all_qk,
        # AI
        "comments": {
            "overall":   ai_comments["business"],
            "financial": ai_comments["financial"],
            "valuation": ai_comments["valuation"],
        },
        # Moat & thesis (placeholder — chuẩn cho dashboard)
        "thesis": [
            f"Doanh nghiệp BĐS KCN với mô hình đa mảng và dòng tiền ổn định.",
            ai_comments["business"][:160],
        ],
        "risks": [
            "Tỷ lệ lấp đầy KCN chậm hơn kỳ vọng ảnh hưởng ghi nhận doanh thu cho thuê đất.",
            "Lãi suất tăng làm tăng COE và giảm giá trị RI model.",
            "Rủi ro NCI lớn: LNST cổ đông thiểu số có thể ảnh hưởng EPS mẹ.",
        ],
        "moats": {
            "Switching Cost":  {"score": 4, "desc": "Hợp đồng thuê đất KCN thường 30-50 năm, rào cản rời bỏ cao."},
            "Efficient Scale": {"score": 4, "desc": "Hạ tầng KCN lớn — barrier to entry về vốn rất cao."},
            "Cost Advantage":  {"score": 3, "desc": "Tiện ích điện/nước/xử lý nước thải tự vận hành → tiết kiệm chi phí."},
            "Network Effect":  {"score": 2, "desc": "Cụm KCN mạnh thu hút thêm khách thuê, hiệu ứng tích cực nhẹ."},
        },
        "pestle": {
            "Political":     "Chính sách thu hút FDI và phát triển KCN tiếp tục được ưu tiên.",
            "Economic":      "Làn sóng dịch chuyển chuỗi cung ứng từ Trung Quốc sang VN thúc đẩy nhu cầu KCN.",
            "Social":        "Yêu cầu nhà ở công nhân, hạ tầng xã hội trong KCN ngày càng được coi trọng.",
            "Technological": "Xu hướng KCN thông minh (smart industrial park) đòi hỏi đầu tư số hóa.",
            "Legal":         "Luật Đất đai sửa đổi và quy định môi trường nghiêm ngặt hơn.",
            "Environmental": "Áp lực ESG từ nhà đầu tư FDI nước ngoài trong KCN.",
        },
    }

    json_path = os.path.join(PROJECT_ROOT, "data", f"{ticker}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"  [OK] JSON: {json_path}")
    return json_path


# ══════════════════════════════════════════════════════════════════════════
# ENTRY POINT — run_kcn_analysis(ticker)
# ══════════════════════════════════════════════════════════════════════════
def run_kcn_analysis(ticker, use_cache=True):
    """
    Entry point chính. Chạy toàn bộ pipeline phân tích & định giá KCN:
      1. Fetch dữ liệu Vietcap (cache hoặc live)
      2. Load kho dữ liệu mảng (data/segments_kcn/TICKER.json)
      3. Tính định giá 40%P/E + 40%P/B + 20%RI
      4. AI commentary (Gemini nếu có GEMINI_API_KEY)
      5. Xuất Excel (8 sheets) + PDF + JSON dashboard
    """
    ticker = ticker.upper()
    print(f"\n{'='*60}")
    print(f"  KCN ANALYSIS ENGINE — {ticker}")
    print(f"{'='*60}")

    # ── 1. Fetch Vietcap BCTC ──────────────────────────────────────
    from fetch_data import fetch_all
    raw = fetch_all(ticker, use_cache=use_cache)

    company_name  = raw.get("companyName") or raw.get("ticker") or ticker
    current_price = raw.get("currentPrice") or raw.get("info", {}).get("currentPrice") or 0

    is_recs_y = raw["sections"]["INCOME_STATEMENT"].get("years", [])
    bs_recs_y = raw["sections"]["BALANCE_SHEET"].get("years", [])
    cf_recs_y = raw["sections"]["CASH_FLOW"].get("years", [])
    is_recs_q = raw["sections"]["INCOME_STATEMENT"].get("quarters", [])

    if not is_recs_y:
        print(f"[ERROR] Không có dữ liệu IS năm cho {ticker}. Dừng.")
        return False

    # Lấy 5 năm lịch sử gần nhất có dữ liệu
    hist_years = sorted({r["yearReport"] for r in is_recs_y if r.get("yearReport")})[-5:]
    fc_years   = [hist_years[-1] + 1, hist_years[-1] + 2, hist_years[-1] + 3]

    # ── 2. Shares — từ charter_capital / BVPS ────────────────────
    shares = 0
    for r in bs_recs_y:
        cap = r.get(BS_GEN["charter_capital"])
        if cap and cap > 0:
            # Vốn điều lệ đơn vị VND — chia 10,000 = số cp (mệnh giá 10k)
            shares = int(cap / 10_000)
            break
    if shares <= 0:
        # Fallback: estimate from market cap
        mcap = raw.get("info", {}).get("marketCap") or 0
        if mcap > 0 and current_price > 0:
            shares = int(mcap / current_price)
        else:
            shares = 100_000_000  # 100 triệu cp mặc định

    print(f"  Company : {company_name}")
    print(f"  Price   : {current_price:,.0f} VND")
    print(f"  Shares  : {shares:,.0f} cp")
    print(f"  Hist YR : {hist_years}")

    # ── 3. Load kho dữ liệu mảng ─────────────────────────────────
    # Tự động lập kế hoạch, tải PDF và parse các kỳ BCTC còn thiếu
    try:
        print("  [INFO] Đang kiểm tra các kỳ BCTC thiếu để tự động tải & parse số liệu mảng...")
        import subprocess
        # 1. Gọi plan-downloads để xem có file nào cần tải không
        res = subprocess.run(
            [sys.executable, "bctc_pdf_tool.py", "plan-downloads", ticker],
            cwd=PROJECT_ROOT,
            capture_output=True,
            text=True,
            encoding="utf-8"
        )
        output = res.stdout
        
        # Parse output để tìm các link cần tải dạng: "CN/2024" hoặc "Q1/2024" và link tương ứng
        download_tasks = []
        for line in output.split("\n"):
            if "|" in line and "http" in line:
                parts = [p.strip() for p in line.split("|")]
                if len(parts) >= 3:
                    period_raw = parts[0]  # vd: "2024(CN)" hoặc "2024Q1"
                    url = parts[2]
                    
                    # Chuẩn hóa period_key tương thích với kho dữ liệu
                    # vd: "2024(CN)" -> "2024(CN)", "2024Q1" -> "2024Q1"
                    period_key = period_raw
                    # Xác định tên file đầu ra
                    pdf_filename = f"{ticker}_{period_key.replace('(','_').replace(')','')}.pdf"
                    pdf_path = os.path.join(PROJECT_ROOT, "BCTC_PDF", ticker, pdf_filename)
                    download_tasks.append((url, pdf_path, period_key))
                    
        if download_tasks:
            print(f"  [INFO] Phát hiện {len(download_tasks)} kỳ BCTC thiếu dữ liệu mảng. Tiến hành tải & trích xuất...")
            for url, out_path, period_key in download_tasks:
                # Kiểm tra xem kỳ này đã được cập nhật dữ liệu đầy đủ chưa (tránh tải trùng do trích xuất chéo)
                try:
                    curr_store = load_segments_kcn(ticker)
                    p_data = curr_store.get("yearly", {}).get(period_key) or curr_store.get("quarterly", {}).get(period_key)
                    if p_data:
                        non_khac = [s for s in p_data if s != "Khac" and (p_data[s].get("revenue", 0) > 0 or p_data[s].get("cogs", 0) > 0)]
                        if non_khac:
                            print(f"    -> [SKIP] {period_key} đã được cập nhật đầy đủ (trích xuất chéo từ kỳ khác).")
                            continue
                except Exception:
                    pass

                print(f"    -> Đang tải & parse {period_key}: {os.path.basename(out_path)}...")
                # Gọi lệnh download của bctc_pdf_tool.py (lệnh này tự convert và gọi parser)
                subprocess.run(
                    [sys.executable, "bctc_pdf_tool.py", "download", url, "--out", out_path, "--ticker", ticker, "--period", period_key],
                    cwd=PROJECT_ROOT
                )
            print("  [INFO] Đã hoàn thành tải và trích xuất dữ liệu mảng tự động.")
        else:
            print("  [INFO] Kho dữ liệu mảng đã đầy đủ hoặc không có kỳ mới cần tải.")
            
    except Exception as e:
        print(f"  [WARN] Lỗi khi tự động tải & parse dữ liệu mảng: {e}")

    # Tự động suy ra Q4 từ Yearly và Q1, Q2, Q3 cho các năm lịch sử nếu có đủ số liệu
    try:
        for y in range(2021, 2027):
            subprocess.run(
                [sys.executable, "segments_kcn_tool.py", "derive-q4", ticker, str(y), "--force"],
                cwd=PROJECT_ROOT,
                capture_output=True
            )
    except Exception:
        pass

    store = load_segments_kcn(ticker)
    seg_names, seg_labels, seg_colors, yearly_seg, quarterly_seg = build_segment_history(store, is_recs_y, is_recs_q)

    # Check consistency (log only, không block)
    check_segment_consistency(store, is_recs_y, is_recs_q)

    # Forecast mảng (dùng quarterly_seg để anchor YTD nếu có dữ liệu mới)
    years_hist_avail = sorted(set().union(*[yearly_seg[s].keys() for s in seg_names]))
    seg_fc = forecast_segments(store, seg_names, yearly_seg, years_hist_avail, n_fc=3, quarterly=quarterly_seg)

    # ── 4. Beta & Rf ─────────────────────────────────────────────
    print("  [INFO] Fetching Rf & Beta...")
    rf, rf_src     = fetch_rf_vietnam()
    calc_beta, web_beta, is_enough, beta_src, latest_price_hist, aligned_data = fetch_and_calc_beta(ticker)
    beta_raw = calc_beta if is_enough else web_beta
    beta = round(0.67 * beta_raw + 0.33, 4)
    print(f"  Rf={rf*100:.2f}% ({rf_src}) | Beta thô={beta_raw:.3f} | Beta Blume={beta:.3f} ({beta_src})")

    if current_price <= 0 and latest_price_hist:
        current_price = latest_price_hist
        print(f"  [INFO] Dùng giá lịch sử cuối: {current_price:,.0f} VND")

    # ── 5. P&L tổng hợp lịch sử ──────────────────────────────────
    rev_h   = [_get_yr(is_recs_y, y, IS_GEN["revenue"])    for y in hist_years]
    gp_h    = [_get_yr(is_recs_y, y, IS_GEN["gross_profit"]) for y in hist_years]
    npat_h  = [_get_yr(is_recs_y, y, IS_GEN["npat"])       for y in hist_years]

    # ── 5.5. Fetch P/E & P/B lịch sử theo quý từ Vietcap ────────
    print("  [INFO] Fetching quarterly P/E & P/B history...")
    quarter_labels, pe_quarters, pb_quarters = [], [], []
    try:
        import requests as _requests
        _rf_ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        r_stat = _requests.get(
            f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial",
            headers={"User-Agent": _rf_ua, "Referer": "https://trading.vietcap.com.vn/"},
            timeout=15
        )
        if r_stat.status_code == 200:
            data_stat = r_stat.json().get("data", [])
            ttms = sorted(
                [x for x in data_stat if x.get("year") and x.get("quarter") in (1, 2, 3, 4)],
                key=lambda x: (x["year"], x["quarter"])
            )
            for x in ttms:
                quarter_labels.append(f"{x['year']}-Q{x['quarter']}")
                pe_quarters.append(round(x["pe"], 2) if x.get("pe") else None)
                pb_quarters.append(round(x["pb"], 2) if x.get("pb") else None)
            # Lọc giá trị P/E hợp lệ để tính median cho định giá
            _pe_valid = [v for v in pe_quarters if v and 0 < v < 60]
            _pb_valid = [v for v in pb_quarters if v and v > 0]
            pe_median = round(statistics.median(_pe_valid), 2) if _pe_valid else None
            pb_median = round(statistics.median(_pb_valid), 2) if _pb_valid else None
            print(f"    -> Đã lấy {len(quarter_labels)} quý P/E, P/B (P/E median={pe_median}, P/B median={pb_median})")
            # Cập nhật định giá với PE/PB median từ dữ liệu thực
            if pe_median or pb_median:
                val = calc_valuation_kcn(
                    ticker, is_recs_y, bs_recs_y, cf_recs_y,
                    hist_years, shares, current_price, rf, beta,
                    target_pe=pe_median, target_pb=pb_median,
                )
                val["current_price_chart"] = current_price
    except Exception as e_stat:
        print(f"    [WARN] Không lấy được P/E, P/B lịch sử: {e_stat}")

    # ── 6. Định giá ──────────────────────────────────────────────
    print("  [INFO] Calculating valuation...")
    if not quarter_labels:
        # fallback nếu không fetch được PE/PB history
        val = calc_valuation_kcn(
            ticker, is_recs_y, bs_recs_y, cf_recs_y,
            hist_years, shares, current_price, rf, beta,
        )
        val["current_price_chart"] = current_price
    print(f"  P/E={val['fair_pe']:,.0f} | P/B={val['fair_pb']:,.0f} | RI={val['fair_ri']:,.0f} | Blend={val['fair_blend']:,.0f} ({val['upside']*100:+.1f}%)")

    # ── 7. AI Commentary ─────────────────────────────────────────
    fin_sum = (
        f"DT {rev_h[-1]:.0f} tỷ, LNG {gp_h[-1]:.0f} tỷ, LNST {npat_h[-1]:.0f} tỷ năm {hist_years[-1]}; "
        f"EPS {val['eps_last']:.0f} VND, BVPS {val['bvps_last']:.0f} VND; "
        f"Upside {val['upside']*100:+.1f}%"
    )
    ai_comments = get_ai_commentary_kcn(ticker, company_name, fin_sum)

    # ── 8. Output paths ──────────────────────────────────────────
    out_dir = os.path.join(PROJECT_ROOT, "Bao cao", ticker)
    os.makedirs(out_dir, exist_ok=True)
    month_str  = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_path = os.path.join(out_dir, f"{ticker}_KCN_Model_{month_str}.xlsx")
    pdf_path   = os.path.join(out_dir, f"{ticker}_KCN_Report_{month_str}.pdf")

    # ── 9. Excel ─────────────────────────────────────────────────
    print("  [INFO] Building Excel...")
    wb = openpyxl.Workbook()
    build_excel_kcn(
        wb, ticker, company_name, current_price, shares,
        hist_years, fc_years,
        is_recs_y, bs_recs_y, cf_recs_y,
        seg_names, seg_labels, seg_colors, yearly_seg, quarterly_seg,
        seg_fc, val, ai_comments, beta_src, latest_price_hist, aligned_data,
        quarter_labels, pe_quarters, pb_quarters,
    )
    wb.save(excel_path)
    print(f"  [OK] Excel: {excel_path}")

    # ── 10. Charts ───────────────────────────────────────────────
    print("  [INFO] Building charts...")
    charts = build_charts_kcn(
        out_dir, ticker, hist_years, fc_years,
        rev_h, gp_h, npat_h,
        seg_names, seg_labels, seg_colors, yearly_seg, quarterly_seg, val,
    )

    # ── 11. PDF ──────────────────────────────────────────────────
    print("  [INFO] Building PDF...")
    build_pdf_kcn(
        pdf_path, ticker, company_name, current_price, shares,
        hist_years, list(hist_years) + fc_years,
        rev_h, gp_h, npat_h,
        seg_names, seg_labels, yearly_seg, quarterly_seg,
        val, ai_comments, charts, beta_src,
    )

    # ── 12. JSON dashboard ───────────────────────────────────────
    print("  [INFO] Saving JSON dashboard...")
    save_json_kcn(
        ticker, company_name, current_price, shares,
        hist_years, rev_h, gp_h, npat_h,
        seg_names, seg_labels, seg_colors,
        yearly_seg, quarterly_seg,
        val, ai_comments,
    )

    # ── 13. Cleanup chart PNGs ───────────────────────────────────
    for p in charts.values():
        try:
            os.remove(p)
        except Exception:
            pass

    # ── 14. Cleanup BCTC PDF/Markdown files to keep repository light ──
    # Tạm thời comment để giữ lại Markdown kiểm tra lỗi
    # try:
    #     print("  [INFO] Cleaning up BCTC PDF/Markdown temporary files...")
    #     import subprocess
    #     # Run clean command dynamically
    #     subprocess.run(
    #         [sys.executable, "bctc_pdf_tool.py", "clean", ticker],
    #         cwd=os.path.dirname(os.path.abspath(__file__)),
    #         capture_output=True
    #     )
    # except Exception as e:
    #     print(f"  [WARN] Failed to clean up BCTC PDF/Markdown files: {e}")

    print(f"\n{'='*60}")
    print(f"  HOÀN THÀNH — {ticker}")
    print(f"  Excel : {excel_path}")
    print(f"  PDF   : {pdf_path}")
    print(f"{'='*60}\n")
    return True


# ══════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Phân tích & định giá cổ phiếu BĐS KCN (40%P/E + 40%P/B + 20%RI).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ví dụ:
  python template_kcn.py IDC
  python template_kcn.py SIP --no-cache
  python template_kcn.py PHR --no-cache
        """,
    )
    parser.add_argument("ticker", help="Mã cổ phiếu KCN (vd: IDC, SIP, PHR, SZC, KBC)")
    parser.add_argument("--no-cache", action="store_true", help="Bỏ qua cache, fetch mới từ Vietcap API")
    args = parser.parse_args()
    success = run_kcn_analysis(args.ticker, use_cache=not args.no_cache)
    sys.exit(0 if success else 1)
