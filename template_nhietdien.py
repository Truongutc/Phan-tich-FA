#!/usr/bin/env python3
"""
template_nhietdien.py — Universal calculation engine for Vietnamese thermal power
generation companies: POW, NT2, PPC, QTP.

Giai đoạn 1 (xem "Hướng dẫn Định giá Doanh nghiệp Nhóm Nhiệt điện Việt Nam" trong
Logic phan tich cac nganh/): định giá từ BCTC chuẩn (Vietcap) + giá nhiên liệu than/
khí/dầu/tỷ giá (cào qua curl, không AI). CHƯA cào sản lượng điện (MWh)/hệ số tải (%)
từ IR từng công ty/EVN — để dành giai đoạn 2 sau khi lõi định giá đã chạy ổn định
(giống HPG: nguồn sản lượng thép mất nhiều vòng lặp thực tế mới ổn định).

REE bị loại khỏi nhóm này (công ty đa ngành, không phải nhiệt điện thuần) — không xử
lý trong file này.

Định giá: DCF 50% + P/E 20% + P/B 15% + Asset-based 15% (khác KCN: 40%P/E+40%P/B+20%RI
— đây là DCF thật đầu tiên trong hệ thống, các template khác chưa có WACC/DCF nào).
Asset-based diễn giải = Book Value of Equity thuộc cổ đông mẹ (bsa78-bsa210), TƯƠNG
ĐƯƠNG "P/B sàn 1.0x" — không chép nguyên công thức ví dụ trong tài liệu hướng dẫn (ví
dụ gốc tự mâu thuẫn: trừ nợ 2 lần vì "giá trị sổ sách" trong ví dụ đã là VCSH, vốn đã
net nợ sẵn).

Output: Excel model + PDF report + data/<TICKER>.json (cho web dashboard nhietdien.html).
"""
import os
import sys

# Fix Windows console encoding (cp1252 không hỗ trợ tiếng Việt)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import re
import json
import datetime
import subprocess
import statistics as stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, grey, black
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

NHIETDIEN_COMPANY_NAMES = {
    "POW": "Tổng Công ty Điện lực Dầu khí Việt Nam - CTCP",
    "NT2": "CTCP Điện lực Dầu khí Nhơn Trạch 2",
    "PPC": "CTCP Nhiệt điện Phả Lại",
    "QTP": "CTCP Nhiệt điện Quảng Ninh",
}


# ══════════════════════════════════════════════════════════════════════════
# STYLING CONSTANTS (copy nguyên từ template_kcn.py để đồng nhất giao diện Excel)
# ══════════════════════════════════════════════════════════════════════════
FONT_NAME = "Calibri"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name=FONT_NAME, size=11, bold=True)
ITALIC_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")
DATA_FONT = Font(name=FONT_NAME, size=10)
ASSUMP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LINK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
P_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                      top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
FMT_NUM = '#,##0'
FMT_PCT = '0.00%'
FMT_MUL = '0.00"x"'
FMT_PRICE = '#,##0'


def header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def data_row(ws, row, label, values, fmt=None, note=None, bold=False, fill=None):
    c0 = ws.cell(row=row, column=1, value=label)
    c0.font = BOLD_FONT if bold else DATA_FONT
    for i, v in enumerate(values, start=2):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BOLD_FONT if bold else DATA_FONT
        if fmt:
            c.number_format = fmt
        if fill:
            c.fill = fill
    if note:
        ws.cell(row=row, column=len(values) + 3, value=note).font = ITALIC_FONT


# ══════════════════════════════════════════════════════════════════════════
# VIETNAMESE FONT REGISTRATION (copy nguyên từ template_kcn.py)
# ══════════════════════════════════════════════════════════════════════════
def register_vn_fonts():
    font_paths_to_try = [
        ("C:/Windows/Fonts/arial.ttf", "Arial"),
        ("C:/Windows/Fonts/arialbd.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "Arial"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Arial"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", "Arial-Bold"),
    ]
    found = {}
    for path, freg in font_paths_to_try:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(freg, path))
                found[freg] = path
            except Exception:
                pass
    return found


_VN_FONTS = register_vn_fonts()
FONT_REG = 'Arial' if 'Arial' in _VN_FONTS else 'Helvetica'
FONT_BOLD = 'Arial-Bold' if 'Arial-Bold' in _VN_FONTS else 'Helvetica-Bold'


# ══════════════════════════════════════════════════════════════════════════
# CAPM INPUTS: Rf + Beta (COPY NGUYÊN VẸN từ template_kcn.py — không có module dùng
# chung, mỗi template tự copy khối này, xem plan)
# ══════════════════════════════════════════════════════════════════════════
UA_STR = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"


def fetch_via_curl(url, timeout=10, label=None):
    """Tải HTML/text qua curl subprocess (KHÔNG dùng thư viện requests) — investing.com
    chặn vân tay TLS của Python requests/urllib3 (Cloudflare 403) nhưng vẫn cho phép
    curl. Copy nguyên từ build_hpg_model.py — KHÔNG import trực tiếp file đó vì nó chạy
    toàn bộ pipeline HPG ngay khi import (side-effect ở module scope)."""
    tag = f" [{label}]" if label else ""
    try:
        r = subprocess.run(
            ["curl", "-sL", "-A", UA_STR, "--max-time", str(timeout), "-w", "\n__HTTP_STATUS__:%{http_code}", url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout + 5,
        )
        out = r.stdout
        status = None
        marker = "\n__HTTP_STATUS__:"
        if marker in out:
            out, status_str = out.rsplit(marker, 1)
            status = status_str.strip()
        if r.returncode != 0 or not out.strip() or status not in (None, "200"):
            print(f"  [DIAG]{tag} fetch_via_curl weak/empty result: curl_exit={r.returncode} http_status={status} "
                  f"len={len(out)} url={url[:90]}" + (f" stderr={r.stderr[:150].strip()}" if r.stderr else ""))
        return out if r.returncode == 0 else ""
    except Exception as e:
        print(f"  [DIAG]{tag} fetch_via_curl exception: {url[:90]} -> {e}")
        return ""


# ══════════════════════════════════════════════════════════════════════════
# GIÁ NHIÊN LIỆU (than/khí/dầu) + TỶ GIÁ USD/VND — cùng kỹ thuật fetch_via_curl +
# regex data-test="instrument-price-last" trên investing.com đã dùng thành công cho
# HRC/quặng sắt/than luyện cốc trong build_hpg_model.py (đã verify chạy tốt trên GitHub
# Actions — xem log thật phiên trước). KHÔNG import build_hpg_model.py (chạy side-effect
# toàn bộ pipeline HPG ngay khi import) — copy độc lập.
#
# LƯU Ý MÔI TRƯỜNG LOCAL (2026-07): máy dev đang bị nhà mạng (FPT) chặn DNS investing.com
# (vn.investing.com bị trả về 127.0.0.1 — xác nhận qua nslookup, KHÔNG phải lỗi code/URL).
# Không verify được bằng curl thật tại đây — sẽ verify khi chạy GitHub Actions (mạng khác,
# đã xác nhận investing.com hoạt động bình thường ở đó qua log thật). Nguồn/URL bên dưới
# tham khảo cùng cấu trúc trang đã dùng cho than luyện cốc/quặng sắt/HRC, chỉ đổi slug
# sản phẩm — cùng độ tin cậy với các fetcher đã proven trong build_hpg_model.py.
# ══════════════════════════════════════════════════════════════════════════
def fetch_usd_vnd_rate(fallback=26200.0, timeout=15):
    """Tỷ giá USD/VND từ investing.com — dùng để quy đổi giá nhiên liệu quốc tế (thường
    niêm yết USD) sang VND khi cần, và tham chiếu trong PDF (tài liệu hướng dẫn mục 9.3:
    'Giá bán điện thường tính bằng USD... nếu tỷ giá tăng, doanh thu VND tăng')."""
    html = fetch_via_curl("https://vn.investing.com/currencies/usd-vnd", timeout=timeout, label="investing-usdvnd")
    if html:
        m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
        if m:
            try:
                rate = float(m.group(1).replace(",", ""))
                if 15000 <= rate <= 35000:
                    return rate, "investing.com"
            except ValueError:
                pass
    return fallback, "Fallback (manual)"


def fetch_coal_price(fallback=110.0, timeout=15):
    """Giá than nhiệt (Newcastle coal), USD/tấn — investing.com. Dùng ước tính chi phí
    nhiên liệu than cho POW/PPC/QTP (nhà máy than) và ngữ cảnh rủi ro giá nhiên liệu
    (tài liệu hướng dẫn mục 6.2/9.2.1)."""
    html = fetch_via_curl("https://vn.investing.com/commodities/coal-futures", timeout=timeout, label="investing-coal")
    if html:
        m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
        if m:
            try:
                price = float(m.group(1).replace(",", ""))
                if 30 <= price <= 500:
                    return price, "investing.com (Newcastle coal, USD/tấn)"
            except ValueError:
                pass
    return fallback, "Fallback (manual)"


def fetch_gas_price(fallback=3.0, timeout=15):
    """Giá khí tự nhiên (Henry Hub), USD/MMBtu — investing.com. Dùng ước tính chi phí
    nhiên liệu khí cho NT2 (nhà máy khí) và ngữ cảnh rủi ro giá nhiên liệu."""
    html = fetch_via_curl("https://vn.investing.com/commodities/natural-gas", timeout=timeout, label="investing-gas")
    if html:
        m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
        if m:
            try:
                price = float(m.group(1).replace(",", ""))
                if 0.5 <= price <= 30:
                    return price, "investing.com (Henry Hub, USD/MMBtu)"
            except ValueError:
                pass
    return fallback, "Fallback (manual)"


def fetch_oil_price(fallback=75.0, timeout=15):
    """Giá dầu thô Brent, USD/thùng — investing.com. Một số nhà máy dùng dầu FO/DO thay
    thế khi than/khí không đủ (tài liệu hướng dẫn mục 2.3.3)."""
    html = fetch_via_curl("https://vn.investing.com/commodities/brent-oil", timeout=timeout, label="investing-oil")
    if html:
        m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
        if m:
            try:
                price = float(m.group(1).replace(",", ""))
                if 20 <= price <= 200:
                    return price, "investing.com (Brent, USD/thùng)"
            except ValueError:
                pass
    return fallback, "Fallback (manual)"


def fetch_rf_vietnam(timeout=15):
    FALLBACK_RF = 0.045
    try:
        html = fetch_via_curl("https://vn.investing.com/rates-bonds/vietnam-10-year-bond-yield", timeout=timeout, label="investing-rf")
        if html:
            m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
            if m:
                rf = float(m.group(1).replace(",", "")) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "investing.com"
    except Exception as e:
        print(f"  [WARN] investing.com Rf failed: {e}")
    try:
        r = requests.get("https://www.worldgovernmentbonds.com/bond-yield/vietnam/10-years/",
                          headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
        if r.status_code == 200:
            matches = re.findall(r'(\d+\.\d+)%', r.text[:5000])
            if matches:
                rf = float(matches[0]) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "worldgovernmentbonds.com"
    except Exception as e:
        print(f"  [WARN] WorldGovernmentBonds Rf failed: {e}")
    return FALLBACK_RF, "Fallback (manual)"


def fetch_aligned_history(ticker, days=720, timeout=15):
    from_time = 1577836800
    to_time = 2000000000
    url_stock = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol={ticker}&resolution=D&from={from_time}&to={to_time}"
    url_index = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol=VNINDEX&resolution=D&from={from_time}&to={to_time}"
    headers = {
        "User-Agent": UA_STR, "Accept": "application/json, text/plain, */*",
        "Referer": "https://dchart.vndirect.com.vn/",
    }
    try:
        r_stock = requests.get(url_stock, headers=headers, timeout=timeout)
        r_index = requests.get(url_index, headers=headers, timeout=timeout)
        if r_stock.status_code == 200 and r_index.status_code == 200:
            d_stock, d_index = r_stock.json(), r_index.json()
            t_s, c_s = d_stock.get("t") or [], d_stock.get("c") or []
            t_m, c_m = d_index.get("t") or [], d_index.get("c") or []
            map_stock = {t_s[i]: c_s[i] for i in range(min(len(t_s), len(c_s))) if c_s[i] is not None and c_s[i] > 0}
            map_index = {t_m[i]: c_m[i] for i in range(min(len(t_m), len(c_m))) if c_m[i] is not None and c_m[i] > 0}
            common_t = sorted(list(set(map_stock.keys()) & set(map_index.keys())))
            aligned = []
            for t in common_t:
                date_str = datetime.datetime.fromtimestamp(t).strftime("%Y-%m-%d")
                p_s = map_stock[t]
                if p_s < 1000:
                    p_s = p_s * 1000
                aligned.append((date_str, p_s, map_index[t]))
            return aligned
    except Exception as e:
        print(f"[Beta Calc] Error fetching history: {e}")
    return []


def fetch_beta_vietstock(ticker, timeout=15):
    try:
        search_url = f"https://finance.vietstock.vn/search?query={ticker}"
        headers = {'User-Agent': UA_STR}
        r1 = requests.get(search_url, headers=headers, timeout=timeout)
        if r1.status_code == 200:
            data = json.loads(r1.text).get("data", "")
            target_url = ""
            for line in data.split('\r\n'):
                parts = line.split('|')
                if len(parts) >= 3 and parts[0].strip().upper() == ticker.upper():
                    target_url = parts[2]
                    break
            if target_url:
                r2 = requests.get(target_url, headers={'User-Agent': UA_STR, 'Referer': 'https://finance.vietstock.vn/'}, timeout=timeout)
                if r2.status_code == 200:
                    m = re.search(r'\"Beta\":\"([\d\.]+)\"', r2.text)
                    if m:
                        beta = float(m.group(1))
                        if 0.3 <= beta <= 2.5:
                            return beta
    except Exception as e:
        print(f"  [WARN] Vietstock scrape failed: {e}")
    return None


def fetch_beta_vietcap(ticker, timeout=15):
    try:
        url = f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}"
        r = requests.get(url, headers={"User-Agent": UA_STR, "Referer": "https://trading.vietcap.com.vn/"}, timeout=timeout)
        if r.status_code == 200:
            d = r.json().get("data", {})
            beta = d.get("beta")
            if beta is not None and 0.3 <= float(beta) <= 2.5:
                return float(beta)
    except Exception:
        pass
    return None


def fetch_and_calc_beta(ticker, days=720, timeout=20, fallback=1.1):
    print(f"  [INFO] Đang tải lịch sử giá để tự tính Beta cho {ticker}...")
    aligned_data = fetch_aligned_history(ticker, days=days, timeout=timeout)
    latest_price = aligned_data[-1][1] if aligned_data else None
    num_sessions = len(aligned_data)
    web_beta = fetch_beta_vietstock(ticker, timeout) or fetch_beta_vietcap(ticker, timeout) or fallback
    calculated_beta, is_enough_sessions = fallback, False
    if num_sessions >= 30:
        sliced_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
        s = [x[1] for x in sliced_data]
        m = [x[2] for x in sliced_data]
        rs = [(s[i] - s[i-1]) / s[i-1] for i in range(1, len(s))]
        rm = [(m[i] - m[i-1]) / m[i-1] for i in range(1, len(m))]
        n_ret = len(rs)
        mean_rs, mean_rm = sum(rs) / n_ret, sum(rm) / n_ret
        cov_sm = sum((rs[i]-mean_rs)*(rm[i]-mean_rm) for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 0
        var_m = sum((rm[i]-mean_rm)**2 for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 1.0
        calculated_beta = round(max(0.3, min(2.5, cov_sm / var_m if var_m > 0 else fallback)), 4)
        if num_sessions >= 250:
            is_enough_sessions = True
        aligned_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
    beta_src = f"Tự tính toán ({num_sessions} phiên)" if is_enough_sessions else f"Web/API ({web_beta:.2f}) - lịch sử chỉ {num_sessions} phiên"
    return calculated_beta, web_beta, is_enough_sessions, beta_src, latest_price, aligned_data


# ══════════════════════════════════════════════════════════════════════════
# FIELD MAP VIETCAP (isa/bsa/cfa) — verify trực tiếp qua cache POW 2026-07
# CAPEX = cfa19 xác nhận CHÉO qua build_hpg_model.py + build_mwg_model.py (2 template
# khác đã dùng đúng field này), KHÔNG đoán mới — tránh lặp lại sai lầm dò field code
# một mình không kiểm chứng.
# ══════════════════════════════════════════════════════════════════════════
IS_GEN = {
    "revenue": "isa3",          # Doanh thu thuần
    "cogs": "isa4",             # Giá vốn hàng bán
    "gross_profit": "isa5",     # Lợi nhuận gộp
    "fin_income": "isa6",       # Doanh thu hoạt động tài chính
    "fin_expense": "isa7",      # Chi phí tài chính
    "interest_expense": "isa8", # Chi phí lãi vay
    "sga_sales": "isa9",        # Chi phí bán hàng
    "sga_admin": "isa10",       # Chi phí quản lý doanh nghiệp
    "operating_result": "isa11",
    "other_income_net": "isa14",
    "pbt": "isa16",             # Lãi/(lỗ) trước thuế
    "tax_current": "isa17",
    "tax_deferred": "isa18",
    "npat": "isa20",            # Lãi/(lỗ) thuần sau thuế
    "nci_income": "isa21",      # Lợi ích của cổ đông thiểu số (P&L)
    "npat_parent": "isa22",     # LNST của cổ đông công ty mẹ
    "eps_basic": "isa23",
}
BS_GEN = {
    "cash": "bsa2",              # Tiền và tương đương tiền
    "total_assets": "bsa53",     # TỔNG CỘNG TÀI SẢN
    "total_liab": "bsa54",       # NỢ PHẢI TRẢ
    "short_borrow": "bsa56",     # Vay ngắn hạn
    "long_borrow": "bsa71",      # Vay dài hạn
    "equity_total": "bsa78",     # VỐN CHỦ SỞ HỮU (đã bao gồm NCI)
    "charter_capital": "bsa80",  # Vốn góp
    "nci": "bsa210",             # Lợi ích cổ đông không kiểm soát (nested trong bsa78)
    "total_capital": "bsa96",    # TỔNG CỘNG NGUỒN VỐN
}
CF_GEN = {
    "depreciation": "cfa2",      # Khấu hao TSCĐ và BĐSĐT
    "capex": "cfa19",            # Tiền chi mua sắm/XD TSCĐ (âm) — verify chéo build_hpg_model.py/build_mwg_model.py
    "dividends_paid": "cfa32",   # Cổ tức, lợi nhuận đã trả cho chủ sở hữu (âm)
}


def _get_yr(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            if v is not None:
                return v / 1e9
    return 0.0


def _get_q(records, year, quarter, field):
    for r in records:
        if r.get("yearReport") == year and r.get("lengthReport") == quarter:
            v = r.get(field)
            if v is not None:
                return v / 1e9
    return 0.0


# ══════════════════════════════════════════════════════════════════════════
# ĐỊNH GIÁ: DCF 50% + P/E 20% + P/B 15% + Asset-based 15%
# (xem "Hướng dẫn Định giá..." mục 4-5 — đây là DCF THẬT đầu tiên trong hệ thống,
# tham khảo vòng lặp chiết khấu/terminal value của calc_valuation_kcn nhưng đổi từ
# Residual Income sang FCFF chiết khấu bằng WACC)
# ══════════════════════════════════════════════════════════════════════════
def _cagr_nhietdien(values):
    """CAGR từ dãy giá trị dương theo thời gian tăng dần, kẹp [-5%, +15%] — nhiệt điện
    tăng trưởng chậm & ổn định hơn KCN (KCN dùng [-10%,+20%]) vì sản lượng bị giới hạn
    bởi công suất lắp đặt cố định, không mở rộng quỹ đất/dự án như BĐS KCN."""
    vals = [v for v in values if v is not None and v > 0]
    if len(vals) < 2:
        return 0.02
    n = len(vals) - 1
    g = (vals[-1] / vals[0]) ** (1 / n) - 1
    return max(-0.05, min(0.15, g))


def _eps_parent(is_recs, year):
    """EPS cơ bản từ isa23 (VND/cp). Dùng isa22/shares làm fallback nếu isa23 = 0."""
    for r in is_recs:
        if r.get("yearReport") == year:
            eps = r.get("isa23")
            if eps is not None and eps != 0:
                return float(eps)
    return None


def _bvps_parent(bs_recs, year, shares):
    """BVPS thuộc về cổ đông công ty mẹ = (bsa78 - bsa210) / shares (VND/cp)."""
    for r in bs_recs:
        if r.get("yearReport") == year:
            equity_total = (r.get("bsa78") or 0)
            nci = (r.get("bsa210") or 0)
            vcsh_parent = equity_total - nci
            if vcsh_parent > 0 and shares > 0:
                return vcsh_parent / shares
    return None


def calc_wacc(rf, beta, erp, current_price, shares, debt_last, debt_prev, interest_expense, tax_rate=0.20):
    """WACC = tỷ trọng vốn CSH×COE + tỷ trọng nợ vay×COD×(1-thuế).
    COE = CAPM chuẩn (Rf + β×ERP, không cộng specific risk premium — dòng tiền PPA
    tương đối ổn định, khác CTCK biến động cao). COD = lãi vay / nợ vay bình quân,
    kẹp [2%, 12%] để tránh nhiễu khi nợ vay quá nhỏ/= 0 (vd PPC không vay nợ).
    Tất cả input tài chính đơn vị TỶ VND, current_price đơn vị VND/cp, shares đơn vị cp."""
    coe = rf + beta * erp
    avg_debt = (debt_last + debt_prev) / 2 if debt_prev else debt_last
    cod_raw = (interest_expense / avg_debt) if avg_debt > 0 else 0.0
    cod = max(0.02, min(0.12, cod_raw))
    market_cap = current_price * shares / 1e9  # tỷ VND
    total_cap = market_cap + debt_last
    if total_cap <= 0:
        return {"wacc": coe, "coe": coe, "cod": cod, "w_e": 1.0, "w_d": 0.0, "market_cap": market_cap}
    w_e = market_cap / total_cap
    w_d = debt_last / total_cap
    wacc_raw = w_e * coe + w_d * cod * (1 - tax_rate)
    # Sàn 8% (không phải 6%) — cổ phiếu nhiệt điện VN thanh khoản thấp thường cho beta đo
    # được RẤT thấp (vd QTP beta~0.36) khiến CAPM/WACC gần bằng Rf phi rủi ro, làm Terminal
    # Value nổ số phi lý (WACC-g_term quá mỏng). Sàn 8% ép ERP tối thiểu ~3.5-4% dù beta đo
    # được thấp — phát hiện thực tế 2026-07 khi test QTP ra upside +396% (đã sửa, xem
    # calc_dcf_valuation dùng thêm exit-multiple để tránh phụ thuộc 1 mình Gordon-growth).
    wacc = max(0.08, min(0.18, wacc_raw))
    return {"wacc": wacc, "coe": coe, "cod": cod, "w_e": w_e, "w_d": w_d, "market_cap": market_cap}


def project_fcf(hist_years, is_recs_y, cf_recs_y, n_fc=5, tax_rate=0.20, is_recs_q=None):
    """Dự phóng FCFF n_fc năm tới. EBIT tính từ Lợi nhuận gộp CỘNG Chi phí bán hàng+QLDN
    (isa5 + isa9 + isa10 — isa9/isa10 Vietcap lưu ÂM nên CỘNG mới đúng, xem comment
    sga_hist bên dưới) — KHÔNG dùng isa11 (Lợi nhuận thuần HĐKD theo VAS) vì isa11 đã trừ
    cả doanh thu/chi phí tài chính (gồm lãi vay), trong khi FCFF cần EBIT THUẦN HOẠT ĐỘNG
    (chưa trừ lãi vay) — ảnh hưởng vốn vay đã nằm trong WACC, không được trừ trùng ở đây.
    Biên EBIT/D&A/CAPEX dự phóng = bình quân lịch sử (D&A và CAPEX không có driver vật lý
    rõ ràng ở Giai đoạn 1 nên neo theo % doanh thu).

    NĂM DỰ PHÓNG ĐẦU TIÊN (năm hiện tại, ngay sau hist_years[-1]) — nếu đã có báo cáo quý
    thực tế trong năm đó (is_recs_q), dùng blend_annual_estimate() (fetch_data.py, DÙNG
    CHUNG mọi ngành trong hệ thống) để BLEND: lũy kế THỰC TẾ các quý đã biết + ước tính
    theo công thức gốc × phần quý CÒN LẠI — tránh 2 lỗi ngược chiều: (1) có dữ liệu đột
    biến 1 quý mà không cập nhật vào dự phóng cả năm dẫn tới định giá sai (yêu cầu user),
    (2) suy diễn 1 quý bất thường ra nguyên cả năm. Áp dụng cho revenue VÀ ebit độc lập
    (không suy ebit từ % margin cũ khi đã có ebit quý thực tế)."""
    from fetch_data import cumulative_actual_quarters, blend_annual_estimate

    revenue_hist = [_get_yr(is_recs_y, y, IS_GEN["revenue"]) for y in hist_years]
    gp_hist = [_get_yr(is_recs_y, y, IS_GEN["gross_profit"]) for y in hist_years]
    # isa9 (sga_sales) và isa10 (sga_admin) được Vietcap lưu ÂM (verify trực tiếp cache
    # POW 2025: isa9=-1.5 tỷ, isa10=-1245.7 tỷ) — cùng quy ước với isa4 (cogs, cũng âm,
    # Revenue+COGS=GrossProfit). EBIT = GP + SGA (CỘNG, không trừ) mới đúng — bug thật đã
    # phát hiện qua kiểm tra chéo Excel: công thức "GP-SGA" cũ cho EBIT > Gross Profit,
    # vô lý (thêm chi phí bán hàng/QLDN thay vì trừ đi).
    sga_hist = [_get_yr(is_recs_y, y, IS_GEN["sga_sales"]) + _get_yr(is_recs_y, y, IS_GEN["sga_admin"]) for y in hist_years]
    ebit_hist = [gp_hist[i] + sga_hist[i] for i in range(len(hist_years))]
    da_hist = [_get_yr(cf_recs_y, y, CF_GEN["depreciation"]) for y in hist_years]
    capex_hist = [abs(_get_yr(cf_recs_y, y, CF_GEN["capex"])) for y in hist_years]

    rev_g = _cagr_nhietdien(revenue_hist)
    ebit_margins = [ebit_hist[i] / revenue_hist[i] for i in range(len(hist_years)) if revenue_hist[i] > 0]
    ebit_margin_fc = stats.mean(ebit_margins[-3:]) if len(ebit_margins) >= 2 else (ebit_margins[-1] if ebit_margins else 0.10)
    da_pct_hist = [da_hist[i] / revenue_hist[i] for i in range(len(hist_years)) if revenue_hist[i] > 0]
    da_pct_fc = stats.mean(da_pct_hist) if da_pct_hist else 0.08
    capex_pct_hist = [capex_hist[i] / revenue_hist[i] for i in range(len(hist_years)) if revenue_hist[i] > 0]
    capex_pct_fc = stats.mean(capex_pct_hist) if capex_pct_hist else 0.10

    current_year = (hist_years[-1] + 1) if hist_years else None
    ytd_blend_info = None
    if is_recs_q and current_year:
        rev_ytd, n_q_rev = cumulative_actual_quarters(is_recs_q, current_year, IS_GEN["revenue"])
        gp_ytd, n_q_gp = cumulative_actual_quarters(is_recs_q, current_year, IS_GEN["gross_profit"])
        sga1_ytd, n_q_sga1 = cumulative_actual_quarters(is_recs_q, current_year, IS_GEN["sga_sales"])
        sga2_ytd, n_q_sga2 = cumulative_actual_quarters(is_recs_q, current_year, IS_GEN["sga_admin"])
        if n_q_rev > 0:
            ytd_blend_info = {
                "current_year": current_year, "n_known_quarters": n_q_rev,
                "revenue_ytd": rev_ytd, "ebit_ytd": (gp_ytd + sga1_ytd + sga2_ytd) if n_q_gp == n_q_rev else None,
            }

    fc_rows = []
    rev_t = revenue_hist[-1] if revenue_hist else 0.0
    for i in range(n_fc):
        rev_formula = rev_t * (1 + rev_g)
        ebit_formula = rev_formula * ebit_margin_fc
        if i == 0 and ytd_blend_info:
            rev_i = blend_annual_estimate(ytd_blend_info["revenue_ytd"], ytd_blend_info["n_known_quarters"], rev_formula)
            if ytd_blend_info["ebit_ytd"] is not None:
                ebit_i = blend_annual_estimate(ytd_blend_info["ebit_ytd"], ytd_blend_info["n_known_quarters"], ebit_formula)
            else:
                ebit_i = rev_i * ebit_margin_fc
        else:
            rev_i, ebit_i = rev_formula, ebit_formula
        rev_t = rev_i  # năm sau lấy mốc từ năm đã blend (nếu có), không lùi về số formula thuần
        nopat_t = ebit_i * (1 - tax_rate)
        da_t = rev_i * da_pct_fc
        capex_t = rev_i * capex_pct_fc
        fcff_t = nopat_t + da_t - capex_t
        fc_rows.append({
            "year_offset": i + 1, "revenue": rev_i, "ebit": ebit_i, "nopat": nopat_t,
            "da": da_t, "capex": capex_t, "fcff": fcff_t,
        })
    assumptions = {
        "rev_g": rev_g, "ebit_margin_fc": ebit_margin_fc,
        "da_pct_fc": da_pct_fc, "capex_pct_fc": capex_pct_fc,
        "revenue_hist": revenue_hist, "ebit_hist": ebit_hist,
        "da_hist": da_hist, "capex_hist": capex_hist, "ytd_blend_info": ytd_blend_info,
    }
    return fc_rows, assumptions


def calc_dcf_valuation(fc_rows, wacc, shares, net_debt, rev_g, exit_ev_ebitda=None):
    """PV(FCFF 5 năm) + Terminal Value chiết khấu về hiện tại ở WACC.

    Terminal Value = TRUNG BÌNH của 2 phương pháp độc lập, không chỉ dùng 1 mình
    Gordon-growth — phát hiện thực tế 2026-07 (test QTP): khi WACC thấp (beta đo được
    thấp do thanh khoản cổ phiếu kém) và mẫu số (WACC-g_term) quá mỏng, Gordon-growth
    một mình cho Terminal Value nổ số phi lý (QTP: TV chiếm 81% Enterprise Value, upside
    +396%). Neo thêm bằng EXIT-MULTIPLE (EBITDA năm cuối × EV/EBITDA mục tiêu — CÙNG bội
    số đã tính ở calc_ev_ebitda_valuation, neo vào thị trường thay vì thuần công thức
    toán) để 2 phương pháp tự kiểm chứng chéo lẫn nhau, giảm rủi ro lệch về 1 phía.
    g_term kẹp [0%, 3%] theo 40% tốc độ tăng trưởng doanh thu dự phóng — nhiệt điện là
    ngành trưởng thành, g dài hạn phải thấp hơn nhiều tăng trưởng ngắn hạn (cùng logic
    g_term của calc_valuation_kcn, chỉ đổi biên trên 4%->3% vì tăng trưởng nền thấp hơn).
    Mẫu số WACC-g_term kẹp tối thiểu 3% (không phải 1%) để tránh nổ số khi WACC sát sàn."""
    n = len(fc_rows)
    pv_sum = 0.0
    for row in fc_rows:
        pv = row["fcff"] / ((1 + wacc) ** row["year_offset"])
        row["fcff_pv"] = pv
        pv_sum += pv
    g_term = max(0.0, min(0.03, rev_g * 0.4))
    last_fcff = fc_rows[-1]["fcff"] if fc_rows else 0.0
    tv_gordon = last_fcff * (1 + g_term) / max(wacc - g_term, 0.03)

    tv_exit = None
    if exit_ev_ebitda and fc_rows:
        last_ebitda = fc_rows[-1]["ebit"] + fc_rows[-1]["da"]
        tv_exit = last_ebitda * exit_ev_ebitda
        terminal_value = (tv_gordon + tv_exit) / 2
    else:
        terminal_value = tv_gordon

    terminal_value_pv = terminal_value / ((1 + wacc) ** n) if n else terminal_value
    enterprise_value = pv_sum + terminal_value_pv
    equity_value = enterprise_value - net_debt
    fair_dcf = (equity_value * 1e9) / shares if shares > 0 else 0.0
    return {
        "pv_sum": pv_sum, "g_term": g_term, "tv_gordon": tv_gordon, "tv_exit": tv_exit,
        "terminal_value": terminal_value,
        "terminal_value_pv": terminal_value_pv, "enterprise_value": enterprise_value,
        "net_debt": net_debt, "equity_value": equity_value, "fair_dcf": fair_dcf,
    }


def _get_shares_from_raw(raw, bs_recs_y):
    """Số lượng cổ phiếu = Vốn điều lệ (bsa80, năm báo cáo MỚI NHẤT) / 10.000 (mệnh giá
    10.000 VND/cp) — fallback numberOfSharesMktCap từ Vietcap API nếu năm nào cũng thiếu
    bsa80. KHÔNG suy từ market cap/giá (dễ sai nếu giá cổ phiếu bất thường)."""
    shares = 0
    bs_sorted_desc = sorted(bs_recs_y, key=lambda r: r.get("yearReport", 0), reverse=True)
    for r in bs_sorted_desc:
        cap = r.get(BS_GEN["charter_capital"])
        if cap and cap > 0:
            shares = int(cap / 10_000)
            break
    if shares <= 0:
        shares_api = raw.get("numberOfSharesMktCap", 0)
        shares = int(shares_api) if shares_api > 0 else 0
    return shares


NHIETDIEN_TICKERS = ["POW", "NT2", "PPC", "QTP"]


def fetch_peer_multiples(tickers=None, use_cache=True):
    """Tính P/E, P/B, EV/EBITDA của TỪNG MÃ trong nhóm nhiệt điện tại kỳ báo cáo năm gần
    nhất, rồi lấy MEDIAN CẢ NHÓM làm target multiple áp dụng đồng nhất cho tất cả — đây
    là cách làm 'peer comps' chuẩn dùng SỐ LIỆU THẬT của toàn nhóm thay vì tự đặt biên số
    kiểu 8-12x/0.8-1.2x không có căn cứ (yêu cầu user: 'median đầy đủ, không estimate').
    Trả về {'per_ticker': {...từng mã...}, 'peer_median': {pe, pb, ev_ebitda, n_pe, n_pb,
    n_ev_ebitda}} — n_* = số mã có dữ liệu hợp lệ đóng góp vào median (để biết median có
    đáng tin không, vd chỉ 1-2/4 mã có lãi dương thì P/E median kém tin cậy hơn)."""
    from fetch_data import fetch_all
    tickers = tickers or NHIETDIEN_TICKERS
    per_ticker = {}
    for t in tickers:
        try:
            raw = fetch_all(t, use_cache=use_cache)
        except Exception as e:
            print(f"  [WARN] Không fetch được BCTC {t} cho peer comps: {e}")
            continue
        is_y = raw["sections"]["INCOME_STATEMENT"].get("years", [])
        bs_y = raw["sections"]["BALANCE_SHEET"].get("years", [])
        cf_y = raw["sections"]["CASH_FLOW"].get("years", [])
        hist_years = sorted({r["yearReport"] for r in is_y if r.get("yearReport")})
        if not hist_years:
            continue
        y_last = hist_years[-1]
        current_price = raw.get("currentPrice") or 0
        shares = _get_shares_from_raw(raw, bs_y)
        if shares <= 0 or current_price <= 0:
            continue

        eps = _eps_parent(is_y, y_last)
        bvps = _bvps_parent(bs_y, y_last, shares)
        gp = _get_yr(is_y, y_last, IS_GEN["gross_profit"])
        sga = _get_yr(is_y, y_last, IS_GEN["sga_sales"]) + _get_yr(is_y, y_last, IS_GEN["sga_admin"])
        da = _get_yr(cf_y, y_last, CF_GEN["depreciation"])
        ebitda = (gp + sga) + da  # sga (isa9+isa10) đã âm sẵn — CỘNG mới đúng, xem project_fcf
        debt = _get_yr(bs_y, y_last, BS_GEN["short_borrow"]) + _get_yr(bs_y, y_last, BS_GEN["long_borrow"])
        cash = _get_yr(bs_y, y_last, BS_GEN["cash"])
        net_debt = debt - cash
        market_cap = current_price * shares / 1e9
        ev = market_cap + net_debt

        pe = current_price / eps if eps and eps > 0 else None
        pb = current_price / bvps if bvps and bvps > 0 else None
        ev_ebitda = ev / ebitda if ebitda and ebitda > 0 else None

        per_ticker[t] = {
            "year": y_last, "current_price": current_price, "eps": eps, "bvps": bvps,
            "ebitda": ebitda, "market_cap": market_cap, "net_debt": net_debt, "ev": ev,
            "pe": pe, "pb": pb, "ev_ebitda": ev_ebitda,
        }

    pe_vals = [v["pe"] for v in per_ticker.values() if v["pe"] and v["pe"] > 0]
    pb_vals = [v["pb"] for v in per_ticker.values() if v["pb"] and v["pb"] > 0]
    ev_vals = [v["ev_ebitda"] for v in per_ticker.values() if v["ev_ebitda"] and v["ev_ebitda"] > 0]
    peer_median = {
        "pe": round(stats.median(pe_vals), 2) if pe_vals else None,
        "pb": round(stats.median(pb_vals), 2) if pb_vals else None,
        "ev_ebitda": round(stats.median(ev_vals), 2) if ev_vals else None,
        "n_pe": len(pe_vals), "n_pb": len(pb_vals), "n_ev_ebitda": len(ev_vals),
    }
    return {"per_ticker": per_ticker, "peer_median": peer_median}


def calc_ev_ebitda_valuation(hist_years, is_recs_y, cf_recs_y, bs_recs_y, current_price, shares,
                              fc_rows, target_ev_ebitda=None, ev_low=4.5, ev_high=8.0):
    """Định giá theo EV/EBITDA lịch sử — thay cho P/E trong công thức blend (theo yêu cầu
    user 2026-07): nhóm nhiệt điện có KHẤU HAO và LÃI VAY rất lớn (nhà máy vốn đầu tư
    khủng, khấu hao nhanh; nhiều mã vay nợ lớn tài trợ dự án) nên LNST (dùng cho P/E) bị
    bóp méo mạnh bởi 2 khoản này — EBITDA (= EBIT + D&A, TRƯỚC lãi vay+khấu hao) phản ánh
    đúng hơn khả năng sinh tiền hoạt động thực. EV lịch sử = market cap HIỆN TẠI + net
    debt của TỪNG NĂM lịch sử (cùng quy ước với cách P/E-lịch sử/P/B-lịch sử trong
    template_kcn.py dùng current_price áp vào EPS/BVPS quá khứ — không có giá cổ phiếu
    lịch sử nên phải xấp xỉ vậy). Biên [4.5x, 8.0x] là giả định chung hợp lý cho ngành
    điện VN (không có số liệu ngành cụ thể trong tài liệu hướng dẫn — nếu có số liệu
    ngành thực tế sau này nên thay bằng biên đó)."""
    market_cap = current_price * shares / 1e9  # tỷ VND
    ebitda_hist_pairs = []
    for y in hist_years:
        revenue = _get_yr(is_recs_y, y, IS_GEN["revenue"])
        gp = _get_yr(is_recs_y, y, IS_GEN["gross_profit"])
        sga = _get_yr(is_recs_y, y, IS_GEN["sga_sales"]) + _get_yr(is_recs_y, y, IS_GEN["sga_admin"])
        da = _get_yr(cf_recs_y, y, CF_GEN["depreciation"])
        ebit = gp + sga  # sga (isa9+isa10) đã âm sẵn — CỘNG mới đúng, xem project_fcf
        ebitda = ebit + da
        debt_y = _get_yr(bs_recs_y, y, BS_GEN["short_borrow"]) + _get_yr(bs_recs_y, y, BS_GEN["long_borrow"])
        cash_y = _get_yr(bs_recs_y, y, BS_GEN["cash"])
        net_debt_y = debt_y - cash_y
        ev_y = market_cap + net_debt_y
        if ebitda > 0:
            ebitda_hist_pairs.append((y, ebitda, ev_y / ebitda))

    ev_ebitda_hist = [m for _, _, m in ebitda_hist_pairs]
    if target_ev_ebitda is None:
        # Không có target từ peer-group (fetch_peer_multiples) -> fallback tự-lịch sử của
        # chính mã, kẹp biên giả định chung (ev_low/ev_high) — CHỈ dùng khi gọi hàm này
        # độc lập/test 1 mã, bình thường luôn nên truyền target_ev_ebitda từ peer median.
        target_ev_ebitda = round(max(ev_low, min(ev_high, stats.median(ev_ebitda_hist))), 2) if ev_ebitda_hist else round((ev_low + ev_high) / 2, 2)

    ebitda_fc1 = (fc_rows[0]["ebit"] + fc_rows[0]["da"]) if fc_rows else (ebitda_hist_pairs[-1][1] if ebitda_hist_pairs else 0.0)
    debt_last = _get_yr(bs_recs_y, hist_years[-1], BS_GEN["short_borrow"]) + _get_yr(bs_recs_y, hist_years[-1], BS_GEN["long_borrow"])
    cash_last = _get_yr(bs_recs_y, hist_years[-1], BS_GEN["cash"])
    net_debt_last = debt_last - cash_last

    implied_ev = target_ev_ebitda * ebitda_fc1
    implied_equity = implied_ev - net_debt_last
    fair_ev_ebitda = (implied_equity * 1e9) / shares if shares > 0 else 0.0

    return {
        "ebitda_hist_pairs": ebitda_hist_pairs, "target_ev_ebitda": target_ev_ebitda,
        "ebitda_fc1": ebitda_fc1, "implied_ev": implied_ev, "net_debt_last": net_debt_last,
        "implied_equity": implied_equity, "fair_ev_ebitda": fair_ev_ebitda,
    }


def calc_valuation_nhietdien(ticker, is_recs_y, bs_recs_y, cf_recs_y, hist_years, shares,
                              current_price, rf, beta, erp=0.07, tax_rate=0.20, n_fc=5,
                              peer_multiples=None, is_recs_q=None):
    """Entry point tính toán định giá: DCF 50% + EV/EBITDA 20% + P/B 15% + Asset-based 15%.
    (Đổi từ P/E sang EV/EBITDA theo yêu cầu user — xem docstring calc_ev_ebitda_valuation:
    D&A/lãi vay của nhóm nhiệt điện quá lớn khiến LNST/P/E bị méo). P/E vẫn được tính và
    trả về làm THAM CHIẾU hiển thị (không nằm trong trọng số blend).

    `peer_multiples` (dict "peer_median" từ fetch_peer_multiples(), khuyến nghị LUÔN
    truyền vào khi chạy thật) — target P/E, P/B, EV/EBITDA lấy từ MEDIAN CẢ NHÓM (POW+
    NT2+PPC+QTP thực tế), KHÔNG tự đặt biên số áng chừng (yêu cầu user: 'median đầy đủ,
    link đầy đủ không estimate'). Nếu None (test độc lập 1 mã không có dữ liệu 3 mã kia)
    thì fallback về median lịch sử CHÍNH MÃ đó, kẹp biên giả định — CHỈ để không crash khi
    test riêng lẻ, sản phẩm thật luôn phải chạy qua fetch_peer_multiples() trước.

    Asset-based diễn giải = Book Value of Equity mẹ hiện tại (bsa78-bsa210)/shares —
    KHÔNG chép công thức ví dụ trong tài liệu hướng dẫn (tự mâu thuẫn, trừ nợ 2 lần).
    Trả về dict đủ để ghi Excel/JSON."""
    peer_multiples = peer_multiples or {}
    # --- WACC ---
    debt_last = _get_yr(bs_recs_y, hist_years[-1], BS_GEN["short_borrow"]) + _get_yr(bs_recs_y, hist_years[-1], BS_GEN["long_borrow"])
    debt_prev = 0.0
    if len(hist_years) >= 2:
        debt_prev = _get_yr(bs_recs_y, hist_years[-2], BS_GEN["short_borrow"]) + _get_yr(bs_recs_y, hist_years[-2], BS_GEN["long_borrow"])
    interest_expense = abs(_get_yr(is_recs_y, hist_years[-1], IS_GEN["interest_expense"]))
    wacc_info = calc_wacc(rf, beta, erp, current_price, shares, debt_last, debt_prev, interest_expense, tax_rate)

    # --- Dự phóng FCFF (dùng chung cho cả DCF và EV/EBITDA forward) ---
    fc_rows, fc_assump = project_fcf(hist_years, is_recs_y, cf_recs_y, n_fc=n_fc, tax_rate=tax_rate, is_recs_q=is_recs_q)
    cash_last = _get_yr(bs_recs_y, hist_years[-1], BS_GEN["cash"])
    net_debt = debt_last - cash_last

    # --- EV/EBITDA (thay P/E trong blend) — tính TRƯỚC DCF vì Terminal Value của DCF cần
    # target_ev_ebitda làm neo exit-multiple (xem calc_dcf_valuation) ---
    ev_ebitda_result = calc_ev_ebitda_valuation(hist_years, is_recs_y, cf_recs_y, bs_recs_y, current_price, shares, fc_rows,
                                                 target_ev_ebitda=peer_multiples.get("ev_ebitda"))

    # --- DCF (Terminal Value neo cả Gordon-growth lẫn exit-multiple EV/EBITDA) ---
    dcf_result = calc_dcf_valuation(fc_rows, wacc_info["wacc"], shares, net_debt, fc_assump["rev_g"],
                                     exit_ev_ebitda=ev_ebitda_result["target_ev_ebitda"])

    # --- P/E (chỉ để tham chiếu hiển thị) & P/B (nằm trong blend) — target lấy từ peer
    # median nếu có, fallback tự-lịch sử kẹp biên khi test độc lập (xem docstring) ---
    eps_vals = [(y, _eps_parent(is_recs_y, y)) for y in hist_years]
    bvps_vals = [(y, _bvps_parent(bs_recs_y, y, shares)) for y in hist_years]
    eps_valid = [(y, v) for y, v in eps_vals if v is not None and v > 0]
    bvps_valid = [(y, v) for y, v in bvps_vals if v is not None and v > 0]
    eps_last = eps_valid[-1][1] if eps_valid else 1000.0
    bvps_last = bvps_valid[-1][1] if bvps_valid else 10000.0

    pe_hist = [current_price / eps for _, eps in eps_valid if current_price > 0 and eps > 0]
    target_pe = peer_multiples.get("pe") or (round(max(8.0, min(12.0, stats.median(pe_hist))), 1) if pe_hist else 10.0)
    pb_hist = [current_price / bvps for _, bvps in bvps_valid if current_price > 0 and bvps > 0]
    target_pb = peer_multiples.get("pb") or (round(max(0.8, min(1.2, stats.median(pb_hist))), 2) if pb_hist else 1.0)

    eps_cagr = _cagr_nhietdien([v for _, v in eps_valid[-3:]]) if len(eps_valid) >= 2 else 0.02
    eps_fc1 = eps_last * (1 + eps_cagr)
    fair_pe = target_pe * eps_fc1  # tham chiếu, không dùng trong blend
    fair_pb = target_pb * bvps_last

    # --- Asset-based (= Book Value of Equity mẹ hiện tại, xem docstring) ---
    fair_asset = bvps_last

    # --- Blend: DCF 50% + EV/EBITDA 20% + P/B 15% + Asset 15% ---
    fair_dcf = dcf_result["fair_dcf"]
    fair_ev_ebitda = ev_ebitda_result["fair_ev_ebitda"]
    fair_blend = 0.50 * fair_dcf + 0.20 * fair_ev_ebitda + 0.15 * fair_pb + 0.15 * fair_asset
    upside = (fair_blend - current_price) / current_price if current_price > 0 else 0.0

    return {
        "rf": rf, "beta": beta, "erp": erp,
        "wacc": wacc_info, "fc_rows": fc_rows, "fc_assumptions": fc_assump, "dcf": dcf_result,
        "ev_ebitda": ev_ebitda_result,
        "target_pe": target_pe, "target_pb": target_pb, "eps_last": eps_last, "bvps_last": bvps_last,
        "eps_cagr": eps_cagr, "eps_fc1": eps_fc1,
        "fair_dcf": fair_dcf, "fair_ev_ebitda": fair_ev_ebitda, "fair_pe": fair_pe,
        "fair_pb": fair_pb, "fair_asset": fair_asset,
        "fair_blend": fair_blend, "current_price": current_price, "upside": upside,
    }


# ══════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER — FORMULA-DRIVEN (yêu cầu user: "file excel sẽ phải đầy đủ công thức
# tính, việc tính định giá DCF... cũng phải có phương pháp luận và công thức tính đầy
# đủ"). MỌI Ô TÍNH TOÁN (WACC, FCFF, PV, Terminal Value, Fair Value...) là CÔNG THỨC
# EXCEL THẬT tham chiếu qua lại giữa các ô/sheet — sửa 1 giả định (vd Beta, growth rate)
# ở ô màu vàng là cả bảng tính lại tự động, không phải paste số đã tính sẵn từ Python.
# Chỉ dữ liệu LỊCH SỬ THÔ (revenue/COGS/D&A... lấy thẳng từ Vietcap) là giá trị tĩnh —
# đây là INPUT gốc, không phải kết quả tính toán, không có công thức nào để thay thế.
# ══════════════════════════════════════════════════════════════════════════
def _ws_freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def build_excel_nhietdien(ticker, company_name, current_price, shares, hist_years,
                           is_recs_y, bs_recs_y, cf_recs_y, val, peer_result, fuel_prices):
    """Xây workbook 6 sheet, công thức sống hoàn toàn cho phần tính toán:
    00_TongQuan, 01_LichSuTaiChinh, 02_WACC_DCF, 03_PeerComps_EV_PE_PB,
    04_GiaNhienLieu, 05_DinhGia."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    n_hist = len(hist_years)
    year_cols = [get_column_letter(2 + i) for i in range(n_hist)]  # B, C, D...

    # ── Sheet 00: Tổng quan ─────────────────────────────────────────────
    ws0 = wb.create_sheet("00_TongQuan")
    ws0.cell(row=1, column=1, value=f"{ticker} — {company_name}").font = TITLE_FONT
    ws0.cell(row=2, column=1, value="Ngành: Nhiệt điện (Thermal Power Generation)").font = ITALIC_FONT
    ws0.cell(row=3, column=1, value=f"Ngày lập báo cáo: {datetime.date.today().isoformat()}").font = ITALIC_FONT

    r = 5
    ws0.cell(row=r, column=1, value="Giá thị trường hiện tại (VND/cp)").font = BOLD_FONT
    ws0.cell(row=r, column=2, value=current_price).number_format = FMT_PRICE
    R_PRICE = r
    r += 1
    ws0.cell(row=r, column=1, value="Số lượng cổ phiếu lưu hành (cp)").font = BOLD_FONT
    ws0.cell(row=r, column=2, value=shares).number_format = FMT_NUM
    ws0.cell(row=r, column=3, value="= Vốn điều lệ (bsa80, năm gần nhất) / 10.000 VND mệnh giá").font = ITALIC_FONT
    R_SHARES = r
    r += 1
    ws0.cell(row=r, column=1, value="Vốn hóa thị trường (tỷ VND)").font = BOLD_FONT
    c = ws0.cell(row=r, column=2, value=f"=B{R_PRICE}*B{R_SHARES}/1000000000")
    c.number_format = FMT_NUM
    r += 2

    ws0.cell(row=r, column=1, value="TÓM TẮT ĐỊNH GIÁ").font = BOLD_FONT
    r += 1
    header_row(ws0, r, ["Phương pháp", "Trọng số", "Fair Value (VND/cp)", "Nguồn công thức"], [26, 10, 20, 46])
    r += 1
    R_SUMMARY_START = r
    rows_summary = [
        ("DCF (FCFF chiết khấu WACC)", 0.50, "='02_WACC_DCF'!$B$" ),  # placeholder, patched below
        ("EV/EBITDA (peer median)", 0.20, "='03_PeerComps_EV_PE_PB'!$B$"),
        ("P/B (peer median)", 0.15, "='02_WACC_DCF'!$B$"),
        ("Asset-based (Book Value of Equity)", 0.15, "='02_WACC_DCF'!$B$"),
    ]
    # (các ô tham chiếu chính xác được patch lại bên dưới sau khi biết đúng vị trí dòng
    # trên sheet 02/03 — xem phần "PATCH liên kết 00_TongQuan" cuối hàm)
    for label, weight, _ in rows_summary:
        ws0.cell(row=r, column=1, value=label).font = DATA_FONT
        wcell = ws0.cell(row=r, column=2, value=weight)
        wcell.number_format = FMT_PCT
        r += 1
    R_SUMMARY_END = r - 1
    ws0.cell(row=r, column=1, value="FAIR VALUE BLEND").font = BOLD_FONT
    c = ws0.cell(row=r, column=2, value=f"=SUMPRODUCT(B{R_SUMMARY_START}:B{R_SUMMARY_END},C{R_SUMMARY_START}:C{R_SUMMARY_END})")
    c.number_format = FMT_PRICE
    c.font = BOLD_FONT
    R_FAIRBLEND = r
    r += 1
    ws0.cell(row=r, column=1, value="Upside/Downside so với giá hiện tại").font = BOLD_FONT
    c = ws0.cell(row=r, column=2, value=f"=B{R_FAIRBLEND}/B{R_PRICE}-1")
    c.number_format = FMT_PCT
    c.font = BOLD_FONT
    ws0.column_dimensions["A"].width = 34
    ws0.column_dimensions["B"].width = 16
    _ws_freeze(ws0)

    # ── Sheet 01: Lịch sử tài chính (input thô Vietcap + tỷ lệ công thức) ────
    ws1 = wb.create_sheet("01_LichSuTaiChinh")
    ws1.cell(row=1, column=1, value=f"{ticker} — Lịch sử tài chính (tỷ VND, nguồn Vietcap)").font = TITLE_FONT
    r = 3
    header_row(ws1, r, ["Chỉ tiêu"] + [str(y) for y in hist_years] + ["Nguồn/Ghi chú"], [30] + [12]*n_hist + [30])
    r += 1

    def _row_vals(field_map, key):
        return [round(_get_yr(is_recs_y, y, field_map[key]), 2) for y in hist_years]

    revenue_vals = _row_vals(IS_GEN, "revenue")
    cogs_vals = [round(-_get_yr(is_recs_y, y, IS_GEN["cogs"]), 2) if _get_yr(is_recs_y, y, IS_GEN["cogs"]) < 0
                 else round(_get_yr(is_recs_y, y, IS_GEN["cogs"]), 2) for y in hist_years]
    gp_vals = _row_vals(IS_GEN, "gross_profit")
    sga_vals = [round(_get_yr(is_recs_y, y, IS_GEN["sga_sales"]) + _get_yr(is_recs_y, y, IS_GEN["sga_admin"]), 2) for y in hist_years]
    da_vals = [round(_get_yr(cf_recs_y, y, CF_GEN["depreciation"]), 2) for y in hist_years]
    capex_vals = [round(abs(_get_yr(cf_recs_y, y, CF_GEN["capex"])), 2) for y in hist_years]
    interest_vals = [round(abs(_get_yr(is_recs_y, y, IS_GEN["interest_expense"])), 2) for y in hist_years]
    npat_vals = _row_vals(IS_GEN, "npat_parent")
    equity_vals = [round(_get_yr(bs_recs_y, y, BS_GEN["equity_total"]) - _get_yr(bs_recs_y, y, BS_GEN["nci"]), 2) for y in hist_years]
    debt_vals = [round(_get_yr(bs_recs_y, y, BS_GEN["short_borrow"]) + _get_yr(bs_recs_y, y, BS_GEN["long_borrow"]), 2) for y in hist_years]
    cash_vals = [round(_get_yr(bs_recs_y, y, BS_GEN["cash"]), 2) for y in hist_years]

    R_REV = r; data_row(ws1, r, "Doanh thu thuần (isa3)", revenue_vals, FMT_NUM, "Vietcap — BCTC hợp nhất"); r += 1
    R_COGS = r; data_row(ws1, r, "Giá vốn hàng bán (isa4)", cogs_vals, FMT_NUM, "Vietcap"); r += 1
    R_GP = r; data_row(ws1, r, "Lợi nhuận gộp (isa5)", gp_vals, FMT_NUM, "Vietcap"); r += 1
    R_SGA = r; data_row(ws1, r, "CP bán hàng + QLDN (isa9+isa10, ÂM)", sga_vals, FMT_NUM, "Vietcap — đã âm sẵn"); r += 1
    R_EBIT = r
    # isa9/isa10 Vietcap lưu ÂM (verify cache thật) — CỘNG mới đúng EBIT, "-" sẽ cho EBIT
    # > Gross Profit (vô lý) vì cộng ngược chi phí thay vì trừ.
    ebit_formulas = [f"={year_cols[i]}{R_GP}+{year_cols[i]}{R_SGA}" for i in range(n_hist)]
    data_row(ws1, r, "EBIT (= LN gộp + CP bán hàng+QLDN, đã âm)", ebit_formulas, FMT_NUM,
             "CÔNG THỨC — KHÔNG dùng isa11 (đã trừ lãi vay, xem project_fcf)", bold=True); r += 1
    R_EBITMARGIN = r
    ebit_margin_formulas = [f"={year_cols[i]}{R_EBIT}/{year_cols[i]}{R_REV}" for i in range(n_hist)]
    data_row(ws1, r, "Biên EBIT (%)", ebit_margin_formulas, FMT_PCT, "= EBIT / Doanh thu"); r += 1
    R_GM = r
    gm_formulas = [f"={year_cols[i]}{R_GP}/{year_cols[i]}{R_REV}" for i in range(n_hist)]
    data_row(ws1, r, "Biên lợi nhuận gộp (%)", gm_formulas, FMT_PCT, "= LN gộp / Doanh thu"); r += 1
    R_DA = r; data_row(ws1, r, "Khấu hao (cfa2)", da_vals, FMT_NUM, "Vietcap — Báo cáo LCTT"); r += 1
    R_EBITDA = r
    ebitda_formulas = [f"={year_cols[i]}{R_EBIT}+{year_cols[i]}{R_DA}" for i in range(n_hist)]
    data_row(ws1, r, "EBITDA (= EBIT + Khấu hao)", ebitda_formulas, FMT_NUM, "CÔNG THỨC", bold=True); r += 1
    R_CAPEX = r; data_row(ws1, r, "CAPEX (cfa19, trị tuyệt đối)", capex_vals, FMT_NUM, "Vietcap — Báo cáo LCTT"); r += 1
    R_INTEREST = r; data_row(ws1, r, "Chi phí lãi vay (isa8, trị tuyệt đối)", interest_vals, FMT_NUM, "Vietcap"); r += 1
    R_NPAT = r; data_row(ws1, r, "LNST cổ đông công ty mẹ (isa22)", npat_vals, FMT_NUM, "Vietcap"); r += 1
    R_NPATMARGIN = r
    npm_formulas = [f"={year_cols[i]}{R_NPAT}/{year_cols[i]}{R_REV}" for i in range(n_hist)]
    data_row(ws1, r, "Biên LNST (%)", npm_formulas, FMT_PCT, "= LNST mẹ / Doanh thu"); r += 1
    R_EQUITY = r; data_row(ws1, r, "VCSH thuộc cổ đông mẹ (bsa78-bsa210)", equity_vals, FMT_NUM, "Vietcap — Bảng CĐKT"); r += 1
    R_DEBT = r; data_row(ws1, r, "Tổng nợ vay (bsa56+bsa71)", debt_vals, FMT_NUM, "Vietcap"); r += 1
    R_CASH = r; data_row(ws1, r, "Tiền và tương đương tiền (bsa2)", cash_vals, FMT_NUM, "Vietcap"); r += 1
    R_ROE = r
    roe_formulas = [f"={year_cols[i]}{R_NPAT}/{year_cols[i]}{R_EQUITY}" for i in range(n_hist)]
    data_row(ws1, r, "ROE (= LNST mẹ / VCSH mẹ)", roe_formulas, FMT_PCT, "CÔNG THỨC", bold=True); r += 1
    ws1.column_dimensions["A"].width = 38
    for col in year_cols:
        ws1.column_dimensions[col].width = 13
    _ws_freeze(ws1)

    # ── Sheet 02: WACC & DCF (toàn bộ công thức sống) ────────────────────
    ws2 = wb.create_sheet("02_WACC_DCF")
    ws2.cell(row=1, column=1, value=f"{ticker} — WACC & Định giá DCF").font = TITLE_FONT
    r = 3
    ws2.cell(row=r, column=1, value="① CHI PHÍ VỐN (WACC)").font = BOLD_FONT
    r += 1
    ws2.cell(row=r, column=1, value="Lãi suất phi rủi ro Rf").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["rf"]); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    ws2.cell(row=r, column=3, value="Nguồn: investing.com VN 10Y bond yield / worldgovernmentbonds.com").font = ITALIC_FONT
    R_RF = r; r += 1
    ws2.cell(row=r, column=1, value="Beta (β)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["beta"]); c.number_format = "0.000"; c.fill = ASSUMP_FILL
    ws2.cell(row=r, column=3, value="Tự tính hồi quy giá CP vs VNINDEX (≥250 phiên) hoặc Vietstock/Vietcap API").font = ITALIC_FONT
    R_BETA = r; r += 1
    ws2.cell(row=r, column=1, value="Phần bù rủi ro thị trường (ERP)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["erp"]); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    ws2.cell(row=r, column=3, value="Giả định chuẩn thị trường VN (~7%/năm)").font = ITALIC_FONT
    R_ERP = r; r += 1
    ws2.cell(row=r, column=1, value="Chi phí vốn CSH (COE = Rf + β×ERP)").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_RF}+B{R_BETA}*B{R_ERP}"); c.number_format = FMT_PCT
    R_COE = r; r += 1
    ws2.cell(row=r, column=1, value=f"Nợ vay năm {hist_years[-1]} (tỷ VND)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"='01_LichSuTaiChinh'!{year_cols[-1]}{R_DEBT}"); c.number_format = FMT_NUM
    R_DEBTLAST = r; r += 1
    ws2.cell(row=r, column=1, value=f"Chi phí lãi vay năm {hist_years[-1]} (tỷ VND)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"='01_LichSuTaiChinh'!{year_cols[-1]}{R_INTEREST}"); c.number_format = FMT_NUM
    R_INTLAST = r; r += 1
    ws2.cell(row=r, column=1, value="Chi phí nợ vay (COD = Lãi vay / Nợ vay, kẹp [2%,12%])").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=MAX(0.02,MIN(0.12,B{R_INTLAST}/MAX(B{R_DEBTLAST},1)))"); c.number_format = FMT_PCT
    R_COD = r; r += 1
    ws2.cell(row=r, column=1, value="Thuế suất TNDN").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=0.20); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    R_TAX = r; r += 1
    ws2.cell(row=r, column=1, value="Vốn hóa thị trường (tỷ VND)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"='00_TongQuan'!$B${R_PRICE}*'00_TongQuan'!$B${R_SHARES}/1000000000"); c.number_format = FMT_NUM
    R_MCAP = r; r += 1
    ws2.cell(row=r, column=1, value="Tỷ trọng vốn CSH (E/(D+E))").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_MCAP}/(B{R_MCAP}+B{R_DEBTLAST})"); c.number_format = FMT_PCT
    R_WE = r; r += 1
    ws2.cell(row=r, column=1, value="Tỷ trọng nợ vay (D/(D+E))").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"=1-B{R_WE}"); c.number_format = FMT_PCT
    R_WD = r; r += 1
    ws2.cell(row=r, column=1, value="WACC = we×COE + wd×COD×(1-thuế), sàn 8%/trần 18%").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=MAX(0.08,MIN(0.18,B{R_WE}*B{R_COE}+B{R_WD}*B{R_COD}*(1-B{R_TAX})))")
    c.number_format = FMT_PCT; c.font = BOLD_FONT; c.fill = LINK_FILL
    R_WACC = r; r += 2

    ws2.cell(row=r, column=1, value="② GIẢ ĐỊNH DỰ PHÓNG (từ mô hình kinh doanh — không phải ước lượng tùy ý)").font = BOLD_FONT
    r += 1
    ws2.cell(row=r, column=1, value="Tăng trưởng doanh thu dài hạn (CAGR lịch sử, kẹp [-5%,+15%])").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["fc_assumptions"]["rev_g"]); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    ws2.cell(row=r, column=3, value=f"CAGR doanh thu {n_hist} năm lịch sử ('01_LichSuTaiChinh')").font = ITALIC_FONT
    R_REVG = r; r += 1
    ws2.cell(row=r, column=1, value="Biên EBIT dự phóng (TB 2-3 năm gần nhất)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["fc_assumptions"]["ebit_margin_fc"]); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    R_EBITM = r; r += 1
    ws2.cell(row=r, column=1, value="Khấu hao / Doanh thu (TB lịch sử)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["fc_assumptions"]["da_pct_fc"]); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    R_DAPCT = r; r += 1
    ws2.cell(row=r, column=1, value="CAPEX / Doanh thu (TB lịch sử)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=val["fc_assumptions"]["capex_pct_fc"]); c.number_format = FMT_PCT; c.fill = ASSUMP_FILL
    R_CAPEXPCT = r; r += 1
    ytd = val["fc_assumptions"].get("ytd_blend_info")
    if ytd:
        ws2.cell(row=r, column=1,
                 value=f"⚠ Năm {ytd['current_year']} đã có {ytd['n_known_quarters']}/4 quý báo cáo thực tế "
                       f"(lũy kế DT={ytd['revenue_ytd']:.0f} tỷ). Bảng dự phóng bên dưới dùng công thức "
                       f"THUẦN (Doanh thu năm trước × (1+g)) để giữ minh bạch/dễ kiểm tra công thức — "
                       f"số Fair Value công bố trên PDF/web có thể lệch nhẹ vì đã BLEND thêm lũy kế thực "
                       f"tế của năm {ytd['current_year']} (xem blend_annual_estimate trong code).").font = ITALIC_FONT
        r += 1
    r += 1

    ws2.cell(row=r, column=1, value="③ DỰ PHÓNG FCFF 5 NĂM (công thức sống)").font = BOLD_FONT
    r += 1
    n_fc = len(val["fc_rows"])
    fc_cols = [get_column_letter(2 + i) for i in range(n_fc)]
    header_row(ws2, r, ["Chỉ tiêu"] + [f"Năm +{i+1}" for i in range(n_fc)], [40] + [14]*n_fc)
    r += 1
    R_FC_REV = r
    ws2.cell(row=r, column=1, value="Doanh thu (tỷ VND)").font = DATA_FONT
    for i in range(n_fc):
        prev_ref = f"'01_LichSuTaiChinh'!{year_cols[-1]}{R_REV}" if i == 0 else f"{fc_cols[i-1]}{R_FC_REV}"
        c = ws2.cell(row=r, column=2+i, value=f"={prev_ref}*(1+B{R_REVG})")
        c.number_format = FMT_NUM
    r += 1
    R_FC_EBIT = r
    ws2.cell(row=r, column=1, value="EBIT (= Doanh thu × Biên EBIT)").font = DATA_FONT
    for i in range(n_fc):
        c = ws2.cell(row=r, column=2+i, value=f"={fc_cols[i]}{R_FC_REV}*B{R_EBITM}")
        c.number_format = FMT_NUM
    r += 1
    R_FC_NOPAT = r
    ws2.cell(row=r, column=1, value="NOPAT (= EBIT × (1-thuế))").font = DATA_FONT
    for i in range(n_fc):
        c = ws2.cell(row=r, column=2+i, value=f"={fc_cols[i]}{R_FC_EBIT}*(1-B{R_TAX})")
        c.number_format = FMT_NUM
    r += 1
    R_FC_DA = r
    ws2.cell(row=r, column=1, value="Khấu hao (= Doanh thu × %DT)").font = DATA_FONT
    for i in range(n_fc):
        c = ws2.cell(row=r, column=2+i, value=f"={fc_cols[i]}{R_FC_REV}*B{R_DAPCT}")
        c.number_format = FMT_NUM
    r += 1
    R_FC_CAPEX = r
    ws2.cell(row=r, column=1, value="CAPEX (= Doanh thu × %DT)").font = DATA_FONT
    for i in range(n_fc):
        c = ws2.cell(row=r, column=2+i, value=f"={fc_cols[i]}{R_FC_REV}*B{R_CAPEXPCT}")
        c.number_format = FMT_NUM
    r += 1
    R_FC_FCFF = r
    ws2.cell(row=r, column=1, value="FCFF (= NOPAT + Khấu hao - CAPEX)").font = BOLD_FONT
    for i in range(n_fc):
        c = ws2.cell(row=r, column=2+i, value=f"={fc_cols[i]}{R_FC_NOPAT}+{fc_cols[i]}{R_FC_DA}-{fc_cols[i]}{R_FC_CAPEX}")
        c.number_format = FMT_NUM; c.font = BOLD_FONT
    r += 1
    R_FC_PV = r
    ws2.cell(row=r, column=1, value="PV(FCFF) chiết khấu về hiện tại ở WACC").font = DATA_FONT
    for i in range(n_fc):
        c = ws2.cell(row=r, column=2+i, value=f"={fc_cols[i]}{R_FC_FCFF}/(1+$B${R_WACC})^{i+1}")
        c.number_format = FMT_NUM
    r += 2

    ws2.cell(row=r, column=1, value="④ TERMINAL VALUE (trung bình Gordon-growth + Exit-multiple)").font = BOLD_FONT
    r += 1
    ws2.cell(row=r, column=1, value="Tăng trưởng dài hạn g_term (=40% growth ngắn hạn, kẹp [0%,3%])").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"=MAX(0,MIN(0.03,B{R_REVG}*0.4))"); c.number_format = FMT_PCT
    R_GTERM = r; r += 1
    ws2.cell(row=r, column=1, value="Terminal Value (Gordon-growth) = FCFF cuối×(1+g)/(WACC-g)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"={fc_cols[-1]}{R_FC_FCFF}*(1+B{R_GTERM})/MAX(B{R_WACC}-B{R_GTERM},0.03)")
    c.number_format = FMT_NUM
    R_TVGORDON = r; r += 1
    ws2.cell(row=r, column=1, value="EV/EBITDA mục tiêu (peer median — xem sheet 03)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value="='03_PeerComps_EV_PE_PB'!$B$3"); c.number_format = FMT_MUL
    R_TARGETEVEBITDA_REF = r; r += 1
    ws2.cell(row=r, column=1, value="EBITDA năm cuối dự phóng (= EBIT + Khấu hao)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"={fc_cols[-1]}{R_FC_EBIT}+{fc_cols[-1]}{R_FC_DA}"); c.number_format = FMT_NUM
    R_EBITDALAST = r; r += 1
    ws2.cell(row=r, column=1, value="Terminal Value (Exit-multiple) = EBITDA cuối × EV/EBITDA mục tiêu").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_EBITDALAST}*B{R_TARGETEVEBITDA_REF}"); c.number_format = FMT_NUM
    R_TVEXIT = r; r += 1
    ws2.cell(row=r, column=1, value="Terminal Value (trung bình 2 phương pháp)").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=AVERAGE(B{R_TVGORDON},B{R_TVEXIT})"); c.number_format = FMT_NUM; c.font = BOLD_FONT
    R_TV = r; r += 1
    ws2.cell(row=r, column=1, value="PV(Terminal Value)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_TV}/(1+B{R_WACC})^{n_fc}"); c.number_format = FMT_NUM
    R_TVPV = r; r += 2

    ws2.cell(row=r, column=1, value="⑤ TỪ ENTERPRISE VALUE RA FAIR VALUE/CP").font = BOLD_FONT
    r += 1
    ws2.cell(row=r, column=1, value="Enterprise Value (= ΣPV(FCFF) + PV(Terminal Value))").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=SUM({fc_cols[0]}{R_FC_PV}:{fc_cols[-1]}{R_FC_PV})+B{R_TVPV}")
    c.number_format = FMT_NUM; c.font = BOLD_FONT
    R_EV = r; r += 1
    ws2.cell(row=r, column=1, value=f"Nợ vay ròng năm {hist_years[-1]} (= Nợ vay - Tiền mặt)").font = DATA_FONT
    c = ws2.cell(row=r, column=2,
                 value=f"='01_LichSuTaiChinh'!{year_cols[-1]}{R_DEBT}-'01_LichSuTaiChinh'!{year_cols[-1]}{R_CASH}")
    c.number_format = FMT_NUM
    R_NETDEBT = r; r += 1
    ws2.cell(row=r, column=1, value="Equity Value (= Enterprise Value - Nợ vay ròng)").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_EV}-B{R_NETDEBT}"); c.number_format = FMT_NUM; c.font = BOLD_FONT
    R_EQVALUE = r; r += 1
    ws2.cell(row=r, column=1, value="Fair Value DCF (VND/cp) = Equity Value × 1 tỷ / Số CP").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_EQVALUE}*1000000000/'00_TongQuan'!$B${R_SHARES}")
    c.number_format = FMT_PRICE; c.font = BOLD_FONT; c.fill = LINK_FILL
    R_FAIRDCF = r; r += 1

    ws2.cell(row=r, column=1, value="P/B target (peer median — xem sheet 03)").font = DATA_FONT
    c = ws2.cell(row=r, column=2, value="='03_PeerComps_EV_PE_PB'!$C$3"); c.number_format = FMT_MUL
    R_TARGETPB = r; r += 1
    ws2.cell(row=r, column=1, value=f"BVPS hiện tại (= VCSH mẹ năm {hist_years[-1]} × 1 tỷ / Số CP)").font = DATA_FONT
    c = ws2.cell(row=r, column=2,
                 value=f"='01_LichSuTaiChinh'!{year_cols[-1]}{R_EQUITY}*1000000000/'00_TongQuan'!$B${R_SHARES}")
    c.number_format = FMT_PRICE
    R_BVPS = r; r += 1
    ws2.cell(row=r, column=1, value="Fair Value P/B (= BVPS × P/B target)").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_BVPS}*B{R_TARGETPB}"); c.number_format = FMT_PRICE; c.font = BOLD_FONT
    R_FAIRPB = r; r += 1
    ws2.cell(row=r, column=1, value="Fair Value Asset-based (= BVPS hiện tại, sàn thanh lý)").font = BOLD_FONT
    c = ws2.cell(row=r, column=2, value=f"=B{R_BVPS}"); c.number_format = FMT_PRICE; c.font = BOLD_FONT
    R_FAIRASSET = r; r += 1
    ws2.cell(row=r, column=1, value="P/E target (peer median, THAM CHIẾU — không dùng trong blend)").font = ITALIC_FONT
    c = ws2.cell(row=r, column=2, value="='03_PeerComps_EV_PE_PB'!$D$3"); c.number_format = FMT_MUL
    r += 1

    ws1.column_dimensions["A"].width = 38
    ws2.column_dimensions["A"].width = 52
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 50
    _ws_freeze(ws2)

    # ── Sheet 03: Peer Comps (P/E, P/B, EV/EBITDA cả 4 mã + MEDIAN công thức Excel) ──
    ws3 = wb.create_sheet("03_PeerComps_EV_PE_PB")
    ws3.cell(row=1, column=1, value="So sánh tương quan nhóm Nhiệt điện (POW/NT2/PPC/QTP)").font = TITLE_FONT
    r = 3
    ws3.cell(row=r, column=1, value="MEDIAN CẢ NHÓM (target multiple áp dụng đồng nhất — không tự đặt biên số)").font = BOLD_FONT
    r += 1
    header_row(ws3, r, ["", "EV/EBITDA", "P/B", "P/E"], [30, 14, 14, 14])
    r += 1
    peer_tickers = list(peer_result["per_ticker"].keys())
    R_PEER_DATA_START = r + 1
    R_PEER_DATA_END = R_PEER_DATA_START + len(peer_tickers) - 1
    ws3.cell(row=r, column=1, value="MEDIAN").font = BOLD_FONT
    c = ws3.cell(row=r, column=2, value=f"=MEDIAN(B{R_PEER_DATA_START}:B{R_PEER_DATA_END})"); c.number_format = FMT_MUL; c.font = BOLD_FONT; c.fill = LINK_FILL
    c = ws3.cell(row=r, column=3, value=f"=MEDIAN(C{R_PEER_DATA_START}:C{R_PEER_DATA_END})"); c.number_format = FMT_MUL; c.font = BOLD_FONT; c.fill = LINK_FILL
    c = ws3.cell(row=r, column=4, value=f"=MEDIAN(D{R_PEER_DATA_START}:D{R_PEER_DATA_END})"); c.number_format = FMT_MUL; c.font = BOLD_FONT; c.fill = LINK_FILL
    r += 1
    header_row(ws3, r, ["Mã", "EV/EBITDA", "P/B", "P/E", "Năm BC", "Giá (VND)", "Vốn hóa (tỷ)"], [12, 14, 14, 14, 10, 14, 16])
    r += 1
    for pt in peer_tickers:
        pv = peer_result["per_ticker"][pt]
        ws3.cell(row=r, column=1, value=pt).font = BOLD_FONT if pt == ticker else DATA_FONT
        c = ws3.cell(row=r, column=2, value=round(pv["ev_ebitda"], 2) if pv["ev_ebitda"] else None); c.number_format = FMT_MUL
        c = ws3.cell(row=r, column=3, value=round(pv["pb"], 2) if pv["pb"] else None); c.number_format = FMT_MUL
        c = ws3.cell(row=r, column=4, value=round(pv["pe"], 2) if pv["pe"] else None); c.number_format = FMT_MUL
        ws3.cell(row=r, column=5, value=pv["year"])
        c = ws3.cell(row=r, column=6, value=pv["current_price"]); c.number_format = FMT_PRICE
        c = ws3.cell(row=r, column=7, value=round(pv["market_cap"], 1)); c.number_format = FMT_NUM
        r += 1
    ws3.cell(row=r+1, column=1,
             value="Nguồn: Vietcap BCTC hợp nhất kỳ báo cáo gần nhất mỗi mã + giá thị trường hiện tại — "
                   "tính P/E=Giá/EPS, P/B=Giá/BVPS, EV/EBITDA=(Vốn hóa+Nợ vay ròng)/EBITDA.").font = ITALIC_FONT
    ws3.column_dimensions["A"].width = 30
    _ws_freeze(ws3)

    # ── Sheet 04: Giá nhiên liệu & tỷ giá ────────────────────────────────
    ws4 = wb.create_sheet("04_GiaNhienLieu")
    ws4.cell(row=1, column=1, value="Giá nhiên liệu đầu vào & tỷ giá (tham chiếu rủi ro)").font = TITLE_FONT
    r = 3
    header_row(ws4, r, ["Chỉ tiêu", "Giá trị", "Đơn vị", "Nguồn"], [30, 14, 14, 40])
    r += 1
    fuel_rows = [
        ("Giá than nhiệt (Newcastle coal)", fuel_prices["coal"][0], "USD/tấn", fuel_prices["coal"][1]),
        ("Giá khí tự nhiên (Henry Hub)", fuel_prices["gas"][0], "USD/MMBtu", fuel_prices["gas"][1]),
        ("Giá dầu thô (Brent)", fuel_prices["oil"][0], "USD/thùng", fuel_prices["oil"][1]),
        ("Tỷ giá USD/VND", fuel_prices["usdvnd"][0], "VND/USD", fuel_prices["usdvnd"][1]),
    ]
    for label, value, unit, source in fuel_rows:
        ws4.cell(row=r, column=1, value=label).font = DATA_FONT
        c = ws4.cell(row=r, column=2, value=round(value, 2)); c.number_format = FMT_NUM
        ws4.cell(row=r, column=3, value=unit).font = DATA_FONT
        ws4.cell(row=r, column=4, value=source).font = ITALIC_FONT
        r += 1
    r += 1
    ws4.cell(row=r, column=1,
             value="Lưu ý (tài liệu hướng dẫn mục 6.2): than tăng 20% → LNST giảm ước 30-40%; khí đắt hơn "
                   "than nên nhà máy khí (NT2) nhạy cảm hơn với giá LNG/Henry Hub.").font = ITALIC_FONT
    ws4.column_dimensions["A"].width = 34
    _ws_freeze(ws4)

    # ── Sheet 05: Định giá — kết nối trọng số + PATCH lại tham chiếu sheet 00 ──
    ws5 = wb.create_sheet("05_DinhGia")
    ws5.cell(row=1, column=1, value=f"{ticker} — Định giá kết hợp: DCF 50% + EV/EBITDA 20% + P/B 15% + Asset 15%").font = TITLE_FONT
    r = 3
    header_row(ws5, r, ["Phương pháp", "Trọng số", "Fair Value (VND/cp)", "Ghi chú"], [30, 12, 20, 50])
    r += 1
    R5_DCF = r
    ws5.cell(row=r, column=1, value="DCF (FCFF chiết khấu WACC)").font = DATA_FONT
    c = ws5.cell(row=r, column=2, value=0.50); c.number_format = FMT_PCT
    c = ws5.cell(row=r, column=3, value=f"='02_WACC_DCF'!$B${R_FAIRDCF}"); c.number_format = FMT_PRICE
    ws5.cell(row=r, column=4, value="Terminal Value = TB(Gordon-growth, Exit-multiple EV/EBITDA)").font = ITALIC_FONT
    r += 1
    R5_EVEBITDA = r
    ws5.cell(row=r, column=1, value="EV/EBITDA (peer median)").font = DATA_FONT
    c = ws5.cell(row=r, column=2, value=0.20); c.number_format = FMT_PCT
    ev_fair_ref = f"={round(val['fair_ev_ebitda'], 2)}"  # tính lại bằng công thức bên dưới thay vì giá trị cứng
    # Fair EV/EBITDA = (target EV/EBITDA × EBITDA năm+1 - Net debt) × 1 tỷ / shares — công thức sống:
    c = ws5.cell(row=r, column=3,
                 value=(f"=('03_PeerComps_EV_PE_PB'!$B$3*'02_WACC_DCF'!${fc_cols[0]}${R_FC_EBIT}"
                        f"+'03_PeerComps_EV_PE_PB'!$B$3*'02_WACC_DCF'!${fc_cols[0]}${R_FC_DA}"
                        f"-'02_WACC_DCF'!$B${R_NETDEBT})*1000000000/'00_TongQuan'!$B${R_SHARES}"))
    c.number_format = FMT_PRICE
    ws5.cell(row=r, column=4, value="= (EV/EBITDA mục tiêu × EBITDA năm+1 - Nợ vay ròng) × 1 tỷ / Số CP").font = ITALIC_FONT
    r += 1
    R5_PB = r
    ws5.cell(row=r, column=1, value="P/B (peer median)").font = DATA_FONT
    c = ws5.cell(row=r, column=2, value=0.15); c.number_format = FMT_PCT
    c = ws5.cell(row=r, column=3, value=f"='02_WACC_DCF'!$B${R_FAIRPB}"); c.number_format = FMT_PRICE
    ws5.cell(row=r, column=4, value="= BVPS hiện tại × P/B mục tiêu (peer median)").font = ITALIC_FONT
    r += 1
    R5_ASSET = r
    ws5.cell(row=r, column=1, value="Asset-based (Book Value of Equity)").font = DATA_FONT
    c = ws5.cell(row=r, column=2, value=0.15); c.number_format = FMT_PCT
    c = ws5.cell(row=r, column=3, value=f"='02_WACC_DCF'!$B${R_FAIRASSET}"); c.number_format = FMT_PRICE
    ws5.cell(row=r, column=4, value="= BVPS mẹ hiện tại (bsa78-bsa210)/CP — sàn giá trị thanh lý").font = ITALIC_FONT
    r += 1
    ws5.cell(row=r, column=1, value="FAIR VALUE BLEND").font = BOLD_FONT
    c = ws5.cell(row=r, column=2, value=f"=SUM(B{R5_DCF}:B{R5_ASSET})"); c.number_format = FMT_PCT; c.font = BOLD_FONT
    c = ws5.cell(row=r, column=3, value=f"=SUMPRODUCT(B{R5_DCF}:B{R5_ASSET},C{R5_DCF}:C{R5_ASSET})")
    c.number_format = FMT_PRICE; c.font = BOLD_FONT; c.fill = LINK_FILL
    R5_BLEND = r; r += 1
    ws5.cell(row=r, column=1, value="Giá thị trường hiện tại").font = DATA_FONT
    c = ws5.cell(row=r, column=3, value="='00_TongQuan'!$B$" + str(R_PRICE)); c.number_format = FMT_PRICE
    R5_PRICE = r; r += 1
    ws5.cell(row=r, column=1, value="Upside/Downside").font = BOLD_FONT
    c = ws5.cell(row=r, column=3, value=f"=C{R5_BLEND}/C{R5_PRICE}-1"); c.number_format = FMT_PCT; c.font = BOLD_FONT
    ws5.column_dimensions["A"].width = 32
    _ws_freeze(ws5)

    # ── PATCH: sheet 00_TongQuan tóm tắt trỏ đúng về sheet 05 (đã biết vị trí) ──
    ws0.cell(row=R_SUMMARY_START, column=3, value=f"='05_DinhGia'!$C${R5_DCF}").number_format = FMT_PRICE
    ws0.cell(row=R_SUMMARY_START + 1, column=3, value=f"='05_DinhGia'!$C${R5_EVEBITDA}").number_format = FMT_PRICE
    ws0.cell(row=R_SUMMARY_START + 2, column=3, value=f"='05_DinhGia'!$C${R5_PB}").number_format = FMT_PRICE
    ws0.cell(row=R_SUMMARY_START + 3, column=3, value=f"='05_DinhGia'!$C${R5_ASSET}").number_format = FMT_PRICE

    return wb


# ══════════════════════════════════════════════════════════════════════════
# CHART GENERATOR — matplotlib PNG, dùng cho PDF report
# ══════════════════════════════════════════════════════════════════════════
def build_charts_nhietdien(out_dir, ticker, hist_years, is_recs_y, bs_recs_y, cf_recs_y,
                            is_recs_q, val, peer_result, fuel_prices, shares):
    """7 biểu đồ: (1) Doanh thu & biên LNG, (2) LNST & biên LNST theo năm, (2b) LNST theo
    quý, (3) ROE & P/B theo năm, (4) EV/EBITDA & P/E lịch sử, (5) Giá nhiên liệu hiện tại,
    (6) FCF dự phóng + PV, (7) So sánh 4 phương pháp định giá. Trả về dict {name: path}."""
    charts = {}
    year_labels = [f"{y}A" for y in hist_years]

    revenue_hist = [_get_yr(is_recs_y, y, IS_GEN["revenue"]) for y in hist_years]
    gp_hist = [_get_yr(is_recs_y, y, IS_GEN["gross_profit"]) for y in hist_years]
    gm_hist = [gp_hist[i] / revenue_hist[i] if revenue_hist[i] else 0 for i in range(len(hist_years))]
    npat_hist = [_get_yr(is_recs_y, y, IS_GEN["npat_parent"]) for y in hist_years]
    npm_hist = [npat_hist[i] / revenue_hist[i] if revenue_hist[i] else 0 for i in range(len(hist_years))]
    equity_hist = [_get_yr(bs_recs_y, y, BS_GEN["equity_total"]) - _get_yr(bs_recs_y, y, BS_GEN["nci"]) for y in hist_years]
    roe_hist = [npat_hist[i] / equity_hist[i] if equity_hist[i] else 0 for i in range(len(hist_years))]

    # ── Chart 1: Doanh thu & Biên LNG ───────────────────────────────────
    p1 = os.path.join(out_dir, f"{ticker}_chart1_revenue_gm.png")
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax1.bar(range(len(hist_years)), revenue_hist, color="#1F4E78", alpha=0.85, label="Doanh thu (tỷ)")
    ax2 = ax1.twinx()
    ax2.plot(range(len(hist_years)), [g*100 for g in gm_hist], color="#C00000", marker="o", linewidth=2, label="Biên LNG (%)")
    ax1.set_xticks(range(len(hist_years))); ax1.set_xticklabels(year_labels, fontsize=9)
    ax1.set_ylabel("Tỷ VND", fontsize=9); ax2.set_ylabel("Biên LNG (%)", fontsize=9, color="#C00000")
    ax1.set_title(f"Doanh thu & Biên Lợi nhuận gộp — {ticker}", fontsize=11, fontweight="bold")
    l1, lb1 = ax1.get_legend_handles_labels(); l2, lb2 = ax2.get_legend_handles_labels()
    ax1.legend(l1+l2, lb1+lb2, loc="upper left", fontsize=8)
    plt.tight_layout(); plt.savefig(p1, dpi=130); plt.close()
    charts["revenue_gm"] = p1

    # ── Chart 2: LNST & Biên LNST theo năm ──────────────────────────────
    p2 = os.path.join(out_dir, f"{ticker}_chart2_npat_year.png")
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax1.bar(range(len(hist_years)), npat_hist, color="#2E75B6", alpha=0.85, label="LNST mẹ (tỷ)")
    ax2 = ax1.twinx()
    ax2.plot(range(len(hist_years)), [m*100 for m in npm_hist], color="#C00000", marker="o", linewidth=2, label="Biên LNST (%)")
    ax1.set_xticks(range(len(hist_years))); ax1.set_xticklabels(year_labels, fontsize=9)
    ax1.set_ylabel("Tỷ VND", fontsize=9); ax2.set_ylabel("Biên LNST (%)", fontsize=9, color="#C00000")
    ax1.set_title(f"LNST & Biên LNST theo năm — {ticker}", fontsize=11, fontweight="bold")
    l1, lb1 = ax1.get_legend_handles_labels(); l2, lb2 = ax2.get_legend_handles_labels()
    ax1.legend(l1+l2, lb1+lb2, loc="upper left", fontsize=8)
    plt.tight_layout(); plt.savefig(p2, dpi=130); plt.close()
    charts["npat_year"] = p2

    # ── Chart 2b: LNST theo quý (nếu có dữ liệu quý) ────────────────────
    if is_recs_q:
        q_keys = sorted({(r["yearReport"], r["lengthReport"]) for r in is_recs_q
                          if r.get("yearReport") and r.get("lengthReport")})[-12:]
        if len(q_keys) >= 4:
            q_labels = [f"{y}Q{q}" for y, q in q_keys]
            npat_q = [_get_q(is_recs_q, y, q, IS_GEN["npat_parent"]) for y, q in q_keys]
            p2b = os.path.join(out_dir, f"{ticker}_chart2b_npat_qtr.png")
            fig, ax = plt.subplots(figsize=(9, 4.5))
            colors = ["#2E75B6" if v >= 0 else "#C00000" for v in npat_q]
            ax.bar(range(len(q_keys)), npat_q, color=colors, alpha=0.85)
            ax.set_xticks(range(len(q_keys))); ax.set_xticklabels(q_labels, fontsize=8, rotation=30, ha="right")
            ax.set_ylabel("Tỷ VND", fontsize=9)
            ax.set_title(f"LNST theo Quý (12 quý gần nhất) — {ticker}", fontsize=11, fontweight="bold")
            ax.grid(axis="y", linestyle="--", alpha=0.4)
            plt.tight_layout(); plt.savefig(p2b, dpi=130); plt.close()
            charts["npat_qtr"] = p2b

    # ── Chart 3: ROE & P/B theo năm ──────────────────────────────────────
    p3 = os.path.join(out_dir, f"{ticker}_chart3_roe_pb.png")
    current_price = val["current_price"]
    # P/B lịch sử = giá hiện tại / BVPS từng năm — cùng quy ước fair_pb/pe_hist (giá cổ
    # phiếu lịch sử không có sẵn nên xấp xỉ bằng giá hiện tại, giống calc_valuation_nhietdien).
    bvps_hist = [(equity_hist[i]*1e9/shares) if shares > 0 else None for i in range(len(hist_years))]
    pb_hist_chart = [(current_price/b if b and b > 0 else None) for b in bvps_hist]
    fig, ax1 = plt.subplots(figsize=(8, 4.5))
    ax1.plot(range(len(hist_years)), [r*100 for r in roe_hist], color="#10b981", marker="o", linewidth=2, label="ROE (%)")
    ax1.set_xticks(range(len(hist_years))); ax1.set_xticklabels(year_labels, fontsize=9)
    ax1.set_ylabel("ROE (%)", fontsize=9, color="#10b981")
    ax2 = ax1.twinx()
    ax2.plot(range(len(hist_years)), pb_hist_chart, color="#7c3aed", marker="s", linewidth=2, linestyle="--", label="P/B (x)")
    ax2.set_ylabel("P/B (x)", fontsize=9, color="#7c3aed")
    ax1.set_title(f"ROE & P/B theo năm — {ticker}", fontsize=11, fontweight="bold")
    l1, lb1 = ax1.get_legend_handles_labels(); l2, lb2 = ax2.get_legend_handles_labels()
    ax1.legend(l1+l2, lb1+lb2, loc="upper left", fontsize=8)
    ax1.grid(axis="y", linestyle="--", alpha=0.4)
    plt.tight_layout(); plt.savefig(p3, dpi=130); plt.close()
    charts["roe_pb"] = p3

    # ── Chart 4: EV/EBITDA & P/E lịch sử (chính mã, giá hiện tại áp vào quá khứ) ──
    ev = val["ev_ebitda"]
    if ev["ebitda_hist_pairs"]:
        p4 = os.path.join(out_dir, f"{ticker}_chart4_ev_pe.png")
        yrs_ev = [y for y, _, _ in ev["ebitda_hist_pairs"]]
        mult_ev = [m for _, _, m in ev["ebitda_hist_pairs"]]
        fig, ax = plt.subplots(figsize=(8, 4.5))
        ax.plot(range(len(yrs_ev)), mult_ev, color="#7c3aed", marker="o", linewidth=2, label="EV/EBITDA (x)")
        ax.axhline(y=ev["target_ev_ebitda"], color="#7c3aed", linestyle="--", alpha=0.5,
                    label=f"Peer median {ev['target_ev_ebitda']}x")
        ax.set_xticks(range(len(yrs_ev))); ax.set_xticklabels([str(y) for y in yrs_ev], fontsize=9)
        ax.set_ylabel("EV/EBITDA (x)", fontsize=9)
        ax.set_title(f"EV/EBITDA lịch sử — {ticker}", fontsize=11, fontweight="bold")
        ax.legend(loc="upper left", fontsize=8)
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        plt.tight_layout(); plt.savefig(p4, dpi=130); plt.close()
        charts["ev_ebitda"] = p4

    # ── Chart 5: Giá nhiên liệu hiện tại (snapshot) ─────────────────────
    p5 = os.path.join(out_dir, f"{ticker}_chart5_fuel_prices.png")
    fuel_labels = ["Than (USD/tấn)", "Khí (USD/MMBtu)", "Dầu Brent (USD/thùng)"]
    fuel_vals = [fuel_prices["coal"][0], fuel_prices["gas"][0], fuel_prices["oil"][0]]
    fig, ax = plt.subplots(figsize=(7, 4))
    bars = ax.bar(fuel_labels, fuel_vals, color=["#475569", "#0891b2", "#f59e0b"], alpha=0.85)
    for bar, v in zip(bars, fuel_vals):
        ax.text(bar.get_x()+bar.get_width()/2, v, f"{v:,.1f}", ha="center", va="bottom", fontsize=9, fontweight="bold")
    ax.set_title(f"Giá nhiên liệu đầu vào hiện tại (tỷ giá {fuel_prices['usdvnd'][0]:,.0f} VND/USD)",
                 fontsize=10, fontweight="bold")
    plt.tight_layout(); plt.savefig(p5, dpi=130); plt.close()
    charts["fuel_prices"] = p5

    # ── Chart 6: FCF dự phóng + PV ────────────────────────────────────────
    p6 = os.path.join(out_dir, f"{ticker}_chart6_fcf.png")
    fc_rows = val["fc_rows"]
    fc_labels = [f"Năm +{row['year_offset']}" for row in fc_rows]
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.bar([i-0.2 for i in range(len(fc_rows))], [r["fcff"] for r in fc_rows], width=0.4, color="#2563eb", label="FCFF")
    ax.bar([i+0.2 for i in range(len(fc_rows))], [r["fcff_pv"] for r in fc_rows], width=0.4, color="#94a3b8", label="PV(FCFF)")
    ax.set_xticks(range(len(fc_rows))); ax.set_xticklabels(fc_labels, fontsize=9)
    ax.set_ylabel("Tỷ VND", fontsize=9)
    ax.set_title(f"FCFF dự phóng & Giá trị hiện tại — {ticker}", fontsize=11, fontweight="bold")
    ax.legend(loc="upper left", fontsize=8)
    plt.tight_layout(); plt.savefig(p6, dpi=130); plt.close()
    charts["fcf"] = p6

    # ── Chart 7: So sánh các phương pháp định giá ────────────────────────
    p7 = os.path.join(out_dir, f"{ticker}_chart7_valuation.png")
    labels_val = ["Giá TT", "DCF", "EV/EBITDA", "P/B", "Asset", "Blend"]
    values_val = [current_price, val["fair_dcf"], val["fair_ev_ebitda"], val["fair_pb"], val["fair_asset"], val["fair_blend"]]
    colors_val = ["#475569", "#2563eb", "#0891b2", "#7c3aed", "#f59e0b", "#C00000"]
    fig, ax = plt.subplots(figsize=(8, 4.5))
    bars = ax.bar(labels_val, values_val, color=colors_val, alpha=0.85)
    for bar, v in zip(bars, values_val):
        ax.text(bar.get_x()+bar.get_width()/2, v, f"{v:,.0f}", ha="center", va="bottom", fontsize=8.5, fontweight="bold")
    ax.set_ylabel("VND/cp", fontsize=9)
    ax.set_title(f"So sánh Định giá — {ticker}", fontsize=11, fontweight="bold")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y/1000:.0f}k"))
    plt.tight_layout(); plt.savefig(p7, dpi=130); plt.close()
    charts["valuation"] = p7

    return charts


# ══════════════════════════════════════════════════════════════════════════
# PDF REPORT BUILDER (reportlab) — cấu trúc mirror template_kcn.py::build_pdf_kcn
# ══════════════════════════════════════════════════════════════════════════
def build_pdf_nhietdien(pdf_path, ticker, company_name, current_price, shares, hist_years,
                         is_recs_y, bs_recs_y, val, peer_result, fuel_prices, charts):
    """Tạo báo cáo PDF đầy đủ cho cổ phiếu nhiệt điện."""
    doc = SimpleDocTemplate(
        pdf_path, pagesize=A4,
        rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("ND_Title", parent=styles["Heading1"], fontName=FONT_BOLD,
                               fontSize=18, leading=22, textColor=HexColor("#1F4E78"), spaceAfter=12)
    h1_st = ParagraphStyle("ND_H1", parent=styles["Heading2"], fontName=FONT_BOLD,
                            fontSize=13, leading=17, textColor=HexColor("#2E75B6"), spaceBefore=14, spaceAfter=7)
    h2_st = ParagraphStyle("ND_H2", parent=styles["Heading3"], fontName=FONT_BOLD,
                            fontSize=11, leading=15, textColor=HexColor("#404040"), spaceBefore=8, spaceAfter=4)
    body_st = ParagraphStyle("ND_Body", parent=styles["Normal"], fontName=FONT_REG,
                              fontSize=10, leading=14, textColor=HexColor("#2D3748"), spaceAfter=6)
    italic_st = ParagraphStyle("ND_Italic", parent=styles["Normal"], fontName=FONT_REG,
                                fontSize=9, leading=12, textColor=HexColor("#718096"), italic=True)

    BLUE_DARK = HexColor("#1F4E78")
    LIGHT_BLUE = HexColor("#DDEBF7")

    def tbl_style(header_col=BLUE_DARK, alt_col=LIGHT_BLUE, font_size=9):
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_col),
            ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("FONTNAME", (0, 1), (-1, -1), FONT_REG),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, alt_col]),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, HexColor("#CBD5E1")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])

    story = []

    # ── Trang bìa ──────────────────────────────────────────────────
    story.append(Paragraph(f"PHÂN TÍCH & ĐỊNH GIÁ CỔ PHIẾU: {ticker}", title_st))
    story.append(Paragraph(
        f"<b>{company_name}</b> | Ngành: Nhiệt điện (Thermal Power Generation) | "
        f"Ngày lập: {datetime.datetime.now().strftime('%d/%m/%Y')}", body_st))
    story.append(Spacer(1, 8))

    upside_pct = f"{val['upside']*100:+.1f}%"
    summary_data = [
        ["Mã CP", "Giá TT (VND)", "Vốn hóa (tỷ)", "Fair Value (VND)", "Upside/Down"],
        [ticker, f"{current_price:,.0f}", f"{shares*current_price/1e9:,.1f}",
         f"{val['fair_blend']:,.0f}", upside_pct],
    ]
    t_sum = Table(summary_data, colWidths=[22*mm, 32*mm, 32*mm, 38*mm, 32*mm])
    t_sum.setStyle(tbl_style())
    story.append(t_sum)
    story.append(Spacer(1, 12))

    # ── 1. Mô hình kinh doanh ──────────────────────────────────────
    story.append(Paragraph("1. Mô hình kinh doanh", h1_st))
    story.append(Paragraph(
        f"{company_name} ({ticker}) hoạt động trong ngành nhiệt điện — vận hành nhà máy phát điện "
        f"từ than/khí đốt, bán điện chủ yếu cho EVN theo hợp đồng mua bán điện (PPA). Doanh thu phụ "
        f"thuộc vào sản lượng điện phát ra, giá bán điện bình quân, và chi phí nhiên liệu đầu vào "
        f"(than/khí/dầu). Xem \"Hướng dẫn Định giá Doanh nghiệp Nhóm Nhiệt điện Việt Nam\" để biết "
        f"chi tiết phương pháp luận.", body_st))

    # ── 2. Kết quả tài chính lịch sử ─────────────────────────────
    story.append(Paragraph("2. Kết quả tài chính lịch sử", h1_st))
    revenue_h = [_get_yr(is_recs_y, y, IS_GEN["revenue"]) for y in hist_years]
    gp_h = [_get_yr(is_recs_y, y, IS_GEN["gross_profit"]) for y in hist_years]
    sga_h = [_get_yr(is_recs_y, y, IS_GEN["sga_sales"]) + _get_yr(is_recs_y, y, IS_GEN["sga_admin"]) for y in hist_years]
    ebit_h = [gp_h[i] + sga_h[i] for i in range(len(hist_years))]
    npat_h = [_get_yr(is_recs_y, y, IS_GEN["npat_parent"]) for y in hist_years]
    pnl_header = ["Chỉ tiêu (tỷ VND)"] + [str(y) for y in hist_years]
    pnl_rows = [
        ["Doanh thu thuần"] + [f"{v:,.0f}" for v in revenue_h],
        ["Lợi nhuận gộp"] + [f"{v:,.0f}" for v in gp_h],
        ["Biên LNG (%)"] + [f"{gp_h[i]/revenue_h[i]*100:.1f}%" if revenue_h[i] else "—" for i in range(len(hist_years))],
        ["EBIT"] + [f"{v:,.0f}" for v in ebit_h],
        ["LNST cổ đông mẹ"] + [f"{v:,.0f}" for v in npat_h],
        ["Biên LNST (%)"] + [f"{npat_h[i]/revenue_h[i]*100:.1f}%" if revenue_h[i] else "—" for i in range(len(hist_years))],
    ]
    col_w_pnl = [38*mm] + [max(15*mm, 90*mm/len(hist_years))]*len(hist_years)
    t_pnl = Table([pnl_header] + pnl_rows, colWidths=col_w_pnl)
    t_pnl.setStyle(tbl_style())
    story.append(t_pnl)
    story.append(Spacer(1, 8))
    if "revenue_gm" in charts:
        story.append(Image(charts["revenue_gm"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 6))
    if "npat_year" in charts:
        story.append(Image(charts["npat_year"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 6))
    if "npat_qtr" in charts:
        story.append(Paragraph("LNST theo quý (12 quý gần nhất):", h2_st))
        story.append(Image(charts["npat_qtr"], width=150*mm, height=75*mm))
        story.append(Spacer(1, 6))
    if "roe_pb" in charts:
        story.append(Image(charts["roe_pb"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 10))

    # ── 3. Peer comps (P/E, P/B, EV/EBITDA cả nhóm) ────────────────
    story.append(Paragraph("3. So sánh tương quan nhóm Nhiệt điện", h1_st))
    pm = peer_result["peer_median"]
    story.append(Paragraph(
        f"Target multiple dùng trong định giá lấy từ MEDIAN thực tế cả nhóm (POW/NT2/PPC/QTP), "
        f"không tự đặt biên số: EV/EBITDA={pm['ev_ebitda']}x, P/B={pm['pb']}x, P/E={pm['pe']}x "
        f"(tham chiếu).", body_st))
    peer_header = ["Mã", "EV/EBITDA", "P/B", "P/E", "Giá (VND)", "Vốn hóa (tỷ)"]
    peer_rows = []
    for pt, pv in peer_result["per_ticker"].items():
        mark = f"{pt} ★" if pt == ticker else pt
        peer_rows.append([
            mark,
            f"{pv['ev_ebitda']:.2f}x" if pv['ev_ebitda'] else "—",
            f"{pv['pb']:.2f}x" if pv['pb'] else "—",
            f"{pv['pe']:.2f}x" if pv['pe'] else "—",
            f"{pv['current_price']:,.0f}",
            f"{pv['market_cap']:,.0f}",
        ])
    t_peer = Table([peer_header] + peer_rows, colWidths=[24*mm, 26*mm, 22*mm, 22*mm, 28*mm, 28*mm])
    t_peer.setStyle(tbl_style())
    story.append(t_peer)
    story.append(Spacer(1, 8))
    if "ev_ebitda" in charts:
        story.append(Image(charts["ev_ebitda"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 10))

    # ── 4. Định giá DCF + Blend ──────────────────────────────────
    story.append(Paragraph("4. Định giá: DCF 50% + EV/EBITDA 20% + P/B 15% + Asset-based 15%", h1_st))
    w = val["wacc"]
    story.append(Paragraph(
        f"WACC = {w['wacc']*100:.2f}% (COE={w['coe']*100:.2f}% từ CAPM Rf={val['rf']*100:.2f}%+"
        f"β={val['beta']:.2f}×ERP={val['erp']*100:.1f}%; COD={w['cod']*100:.2f}%; tỷ trọng vốn CSH "
        f"{w['w_e']*100:.1f}%/nợ vay {w['w_d']*100:.1f}%). FCFF chiết khấu 5 năm + Terminal Value = "
        f"trung bình phương pháp Gordon-growth và Exit-multiple (EV/EBITDA mục tiêu) — xem sheet "
        f"Excel \"02_WACC_DCF\" để biết công thức đầy đủ, có thể chỉnh giả định và tự tính lại.",
        body_st))
    val_header = ["Phương pháp", "Trọng số", "Fair Value (VND/cp)"]
    val_rows = [
        ["DCF (FCFF chiết khấu WACC)", "50%", f"{val['fair_dcf']:,.0f}"],
        [f"EV/EBITDA (target {val['ev_ebitda']['target_ev_ebitda']}x)", "20%", f"{val['fair_ev_ebitda']:,.0f}"],
        [f"P/B (target {val['target_pb']}x)", "15%", f"{val['fair_pb']:,.0f}"],
        ["Asset-based (Book Value of Equity)", "15%", f"{val['fair_asset']:,.0f}"],
        [f"P/E (target {val['target_pe']}x) — THAM CHIẾU, không trong blend", "—", f"{val['fair_pe']:,.0f}"],
    ]
    t_val = Table([val_header] + val_rows, colWidths=[75*mm, 25*mm, 45*mm])
    t_val.setStyle(tbl_style())
    story.append(t_val)
    story.append(Spacer(1, 8))
    if "fcf" in charts:
        story.append(Image(charts["fcf"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 6))
    if "valuation" in charts:
        story.append(Image(charts["valuation"], width=145*mm, height=82*mm))
        story.append(Spacer(1, 10))

    # ── 5. Giá nhiên liệu & rủi ro ────────────────────────────────
    story.append(Paragraph("5. Giá nhiên liệu đầu vào & Rủi ro", h1_st))
    fuel_header = ["Chỉ tiêu", "Giá trị", "Đơn vị", "Nguồn"]
    fuel_rows_pdf = [
        ["Than nhiệt (Newcastle)", f"{fuel_prices['coal'][0]:,.1f}", "USD/tấn", fuel_prices['coal'][1]],
        ["Khí tự nhiên (Henry Hub)", f"{fuel_prices['gas'][0]:,.2f}", "USD/MMBtu", fuel_prices['gas'][1]],
        ["Dầu thô (Brent)", f"{fuel_prices['oil'][0]:,.1f}", "USD/thùng", fuel_prices['oil'][1]],
        ["Tỷ giá USD/VND", f"{fuel_prices['usdvnd'][0]:,.0f}", "VND/USD", fuel_prices['usdvnd'][1]],
    ]
    t_fuel = Table([fuel_header] + fuel_rows_pdf, colWidths=[45*mm, 30*mm, 30*mm, 45*mm])
    t_fuel.setStyle(tbl_style())
    story.append(t_fuel)
    story.append(Spacer(1, 6))
    if "fuel_prices" in charts:
        story.append(Image(charts["fuel_prices"], width=120*mm, height=68*mm))
        story.append(Spacer(1, 8))

    story.append(Paragraph("Các nhóm rủi ro chính (theo Hướng dẫn Định giá Nhóm Nhiệt điện):", h2_st))
    risk_items = [
        "<b>Chính sách năng lượng:</b> Việt Nam ưu tiên phát triển năng lượng tái tạo (điện mặt trời, "
        "gió) có thể giảm hệ số tải nhiệt điện; các quy định môi trường (CO2, NOx, SO2) có thể tăng "
        "chi phí vận hành.",
        "<b>Giá nhiên liệu:</b> Than/khí/dầu biến động lớn theo thị trường quốc tế — giá than tăng "
        "~20% có thể làm LNST giảm ước 30-40% (độ nhạy cao nhất trong các biến số).",
        "<b>Hệ số tải:</b> Phụ thuộc điều độ của EVN và nhu cầu điện — nhu cầu giảm hoặc ưu tiên năng "
        "lượng tái tạo sẽ giảm sản lượng phát, giảm doanh thu.",
        "<b>Kỹ thuật:</b> Nhà máy cũ (như PPC) có chi phí vận hành cao hơn, hiệu suất thấp hơn; bảo "
        "trì lớn định kỳ có thể yêu cầu dừng máy, giảm sản lượng tạm thời.",
        "<b>Tài chính:</b> Nợ vay cao (đặc biệt các dự án đang xây dựng như NT3/NT4 của POW) làm tăng "
        "rủi ro lãi suất; cấu trúc vốn ảnh hưởng trực tiếp tới WACC và định giá.",
    ]
    for item in risk_items:
        story.append(Paragraph(f"• {item}", body_st))
    story.append(Spacer(1, 8))

    story.append(Paragraph(
        "⚠ Giai đoạn 1: định giá dựa trên BCTC chuẩn (Vietcap) + giá nhiên liệu quốc tế — CHƯA tích "
        "hợp trực tiếp sản lượng điện (MWh)/hệ số tải (%)/giá bán bình quân theo IR từng công ty/EVN "
        "(sẽ bổ sung ở giai đoạn 2). Dự phóng FCFF dùng tăng trưởng CAGR lịch sử + biên EBIT bình "
        "quân — với POW đang đầu tư thêm NT3/NT4 (tăng công suất đáng kể), fair value DCF có thể "
        "CHƯA phản ánh đủ tiềm năng tăng trưởng sản lượng từ dự án mới.", italic_st))

    doc.build(story)
    return pdf_path


if __name__ == "__main__":
    # Smoke-test Bước 1+2 của kế hoạch: fetch BCTC + Rf/Beta thật, in số liệu thô +
    # kết quả định giá WACC/DCF/blend để đối chiếu với ví dụ tính tay trong tài liệu
    # hướng dẫn mục 5.3 (không kỳ vọng khớp tuyệt đối vì input thực tế khác giả định,
    # nhưng thứ tự độ lớn phải hợp lý). Sẽ thay bằng CLI thật ở bước cuối cùng.
    from fetch_data import fetch_all
    t = sys.argv[1].upper() if len(sys.argv) > 1 else "POW"
    raw = fetch_all(t, use_cache=True)
    is_y = raw["sections"]["INCOME_STATEMENT"].get("years", [])
    bs_y = raw["sections"]["BALANCE_SHEET"].get("years", [])
    cf_y = raw["sections"]["CASH_FLOW"].get("years", [])
    hist_years = sorted({r["yearReport"] for r in is_y if r.get("yearReport")})[-5:]
    current_price = raw.get("currentPrice") or 0

    shares = 0
    bs_sorted_desc = sorted(bs_y, key=lambda r: r.get("yearReport", 0), reverse=True)
    for r in bs_sorted_desc:
        cap = r.get(BS_GEN["charter_capital"])
        if cap and cap > 0:
            shares = int(cap / 10_000)
            break
    if shares <= 0:
        shares_api = raw.get("numberOfSharesMktCap", 0)
        shares = int(shares_api) if shares_api > 0 else 100_000_000

    print(f"\n{t} — {raw.get('companyName')}")
    print(f"Current price: {current_price} | Shares: {shares:,}")
    print(f"Hist years: {hist_years}")
    for y in hist_years:
        rev = _get_yr(is_y, y, IS_GEN["revenue"])
        cogs = _get_yr(is_y, y, IS_GEN["cogs"])
        npat = _get_yr(is_y, y, IS_GEN["npat_parent"])
        da = _get_yr(cf_y, y, CF_GEN["depreciation"])
        capex = abs(_get_yr(cf_y, y, CF_GEN["capex"]))
        equity = _get_yr(bs_y, y, BS_GEN["equity_total"]) - _get_yr(bs_y, y, BS_GEN["nci"])
        debt = _get_yr(bs_y, y, BS_GEN["short_borrow"]) + _get_yr(bs_y, y, BS_GEN["long_borrow"])
        int_exp = _get_yr(is_y, y, IS_GEN["interest_expense"])
        print(f"  {y}: DT={rev:8.1f} GVHB={cogs:8.1f} LNST_me={npat:7.1f} D&A={da:6.1f} "
              f"CAPEX={capex:7.1f} VCSH_me={equity:8.1f} No_vay={debt:7.1f} LaiVay={int_exp:5.1f}")

    is_q = raw["sections"]["INCOME_STATEMENT"].get("quarters", [])

    print("\n[Fetching Rf/Beta...]")
    rf, rf_src = fetch_rf_vietnam()
    calc_beta, web_beta, is_enough, beta_src, _, _ = fetch_and_calc_beta(t)
    beta = calc_beta if is_enough else web_beta
    print(f"Rf={rf*100:.2f}% ({rf_src}) | Beta={beta:.3f} ({beta_src})")

    print("\n[Fetching peer multiples cả nhóm POW/NT2/PPC/QTP...]")
    peer_result = fetch_peer_multiples()
    for pt, pv in peer_result["per_ticker"].items():
        print(f"  {pt} ({pv['year']}): P/E={pv['pe']} P/B={pv['pb']} EV/EBITDA={pv['ev_ebitda']}")
    pm = peer_result["peer_median"]
    print(f"  -> PEER MEDIAN: P/E={pm['pe']}x (n={pm['n_pe']}) | P/B={pm['pb']}x (n={pm['n_pb']}) | "
          f"EV/EBITDA={pm['ev_ebitda']}x (n={pm['n_ev_ebitda']})")

    val = calc_valuation_nhietdien(t, is_y, bs_y, cf_y, hist_years, shares, current_price, rf, beta,
                                    peer_multiples=pm, is_recs_q=is_q)
    w = val["wacc"]
    print(f"\nWACC={w['wacc']*100:.2f}% (COE={w['coe']*100:.2f}%, COD={w['cod']*100:.2f}%, "
          f"w_e={w['w_e']*100:.1f}%, w_d={w['w_d']*100:.1f}%)")
    print("FCFF dự phóng 5 năm (tỷ VND):")
    for row in val["fc_rows"]:
        print(f"  Năm +{row['year_offset']}: DT={row['revenue']:8.1f} EBIT={row['ebit']:7.1f} "
              f"NOPAT={row['nopat']:7.1f} D&A={row['da']:6.1f} CAPEX={row['capex']:6.1f} "
              f"FCFF={row['fcff']:7.1f} PV={row['fcff_pv']:7.1f}")
    d = val["dcf"]
    print(f"Terminal Value: Gordon={d['tv_gordon']:.0f} | Exit-multiple={d['tv_exit'] if d['tv_exit'] is None else round(d['tv_exit'])} "
          f"-> TB={d['terminal_value']:.0f} (g_term={d['g_term']*100:.1f}%)")
    print(f"PV(FCFF)={d['pv_sum']:.0f} + TV_PV={d['terminal_value_pv']:.0f} "
          f"= EV={d['enterprise_value']:.0f} - NetDebt={d['net_debt']:.0f} = EquityValue={d['equity_value']:.0f} tỷ VND")
    ev = val["ev_ebitda"]
    print(f"\nEV/EBITDA lịch sử: {[(y, round(m,2)) for y, _, m in ev['ebitda_hist_pairs']]} "
          f"-> target={ev['target_ev_ebitda']}x | EBITDA dự phóng năm+1={ev['ebitda_fc1']:.0f} tỷ VND")
    print(f"Fair DCF       = {val['fair_dcf']:,.0f} VND/cp  (trọng số 50%)")
    print(f"Fair EV/EBITDA = {val['fair_ev_ebitda']:,.0f} VND/cp  (trọng số 20%, target {ev['target_ev_ebitda']}x)")
    print(f"Fair P/B       = {val['fair_pb']:,.0f} VND/cp  (trọng số 15%, target P/B={val['target_pb']}x)")
    print(f"Fair Asset     = {val['fair_asset']:,.0f} VND/cp  (trọng số 15%)")
    print(f"Fair P/E (chỉ tham chiếu, KHÔNG trong blend) = {val['fair_pe']:,.0f} VND/cp (target P/E={val['target_pe']}x)")
    print(f"=> FAIR BLEND = {val['fair_blend']:,.0f} VND/cp | Giá hiện tại = {current_price:,.0f} | "
          f"Upside = {val['upside']*100:+.1f}%")
