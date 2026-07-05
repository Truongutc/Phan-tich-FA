#!/usr/bin/env python3
"""
template_securities.py — Universal, parameterized calculation engine for Vietnamese
securities companies (CTCK): SSI, VND, HCM, VCI, VIX, MBS, SHS, FTS, BSI, CTS, VDS...

Bottom-up 5-segment revenue model (theo tài liệu "Logic phan tich cac nganh/Chung khoan.docx"):
    1. Môi giới        — GTGD/phiên × Số phiên × Thị phần × Phí(bps)/10,000
    2. Cho vay Margin  — Dư nợ cho vay bình quân × NIM (spread cho vay - chi phí vốn)
    3. FVTPL (Tự doanh)— Danh mục × (%CDs×R_CDs + %TP×R_TP + %CP×R_VNI)
    4. IB + Lưu ký     — Pipeline IB × Fee% + AUM lưu ký × Fee lưu ký%
    5. Quản lý quỹ     — AUM quản lý × Fee rate (~0.75%/năm)

Định giá ưu tiên P/B (tài sản CTCK có tính thanh khoản cao, phản ánh đúng giá trị sổ
sách) kết hợp P/E (khi lợi nhuận cốt lõi ổn định, không phụ thuộc tự doanh biến động).

Output: Excel model (15 sheets, formula-driven) + PDF report (15-20 trang) +
data/<TICKER>.json (dashboard schema).
"""
import os
import sys
import math
import json
import datetime
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, grey, black
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import requests
import statistics as stats

from fetch_data import (section_to_years, section_to_quarters, get_field_map,
                         cumulative_actual_quarters, blend_annual_estimate,
                         latest_actual_quarter_value, blend_annual_estimate_stock)

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))


# ══════════════════════════════════════════════════════════════════════════
# VIETNAMESE FONT REGISTRATION (multi-platform, giống template_banking.py)
# ══════════════════════════════════════════════════════════════════════════
def register_vn_fonts():
    font_paths_to_try = [
        ("C:/Windows/Fonts/arial.ttf", "Arial"),
        ("C:/Windows/Fonts/arialbd.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "Arial"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Arial"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", "Arial-Bold"),
    ]
    found = {}
    for path, freg in font_paths_to_try:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(freg, path))
                found[freg] = path
            except Exception:
                pass
    return found


_VN_FONTS = register_vn_fonts()
FONT_REG = 'Arial' if 'Arial' in _VN_FONTS else 'Helvetica'
FONT_BOLD = 'Arial-Bold' if 'Arial-Bold' in _VN_FONTS else 'Helvetica-Bold'


# ══════════════════════════════════════════════════════════════════════════
# CAPM INPUTS: Rf + Beta (COPY NGUYÊN VẸN từ template_banking.py — logic hoàn
# toàn generic, không phụ thuộc ngành, đã verify hoạt động ổn định)
# ══════════════════════════════════════════════════════════════════════════
RF_UA_STR = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"


def _fetch_via_curl_rf(url, timeout=10, label=None):
    tag = f" [{label}]" if label else ""
    try:
        r = subprocess.run(
            ["curl", "-sL", "-A", RF_UA_STR, "--max-time", str(timeout), url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout + 5,
        )
        if r.returncode != 0 or not r.stdout.strip():
            print(f"  [DIAG]{tag} fetch_via_curl empty/failed: curl_exit={r.returncode} len={len(r.stdout)}")
        return r.stdout if r.returncode == 0 else ""
    except Exception as e:
        print(f"  [DIAG]{tag} fetch_via_curl exception: {e}")
        return ""


def fetch_rf_vietnam(timeout=15):
    FALLBACK_RF = 0.045
    try:
        html = _fetch_via_curl_rf("https://vn.investing.com/rates-bonds/vietnam-10-year-bond-yield", timeout=timeout, label="investing-rf")
        if html:
            import re
            m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
            if m:
                rf = float(m.group(1).replace(",", "")) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "investing.com"
    except Exception as e:
        print(f"  [WARN] investing.com Rf failed: {e}")
    try:
        r = requests.get("https://www.worldgovernmentbonds.com/bond-yield/vietnam/10-years/",
                          headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
        if r.status_code == 200:
            import re
            matches = re.findall(r'(\d+\.\d+)%', r.text[:5000])
            if matches:
                rf = float(matches[0]) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "worldgovernmentbonds.com"
    except Exception as e:
        print(f"  [WARN] WorldGovernmentBonds Rf failed: {e}")
    return FALLBACK_RF, "Fallback (manual)"


def fetch_aligned_history(ticker, days=720, timeout=15):
    from_time = 1577836800
    to_time = 2000000000
    url_stock = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol={ticker}&resolution=D&from={from_time}&to={to_time}"
    url_index = f"https://dchart-api.vndirect.com.vn/dchart/history?symbol=VNINDEX&resolution=D&from={from_time}&to={to_time}"
    headers = {
        "User-Agent": RF_UA_STR, "Accept": "application/json, text/plain, */*",
        "Referer": "https://dchart.vndirect.com.vn/",
    }
    try:
        r_stock = requests.get(url_stock, headers=headers, timeout=timeout)
        r_index = requests.get(url_index, headers=headers, timeout=timeout)
        if r_stock.status_code == 200 and r_index.status_code == 200:
            d_stock, d_index = r_stock.json(), r_index.json()
            t_s, c_s = d_stock.get("t") or [], d_stock.get("c") or []
            t_m, c_m = d_index.get("t") or [], d_index.get("c") or []
            map_stock = {t_s[i]: c_s[i] for i in range(min(len(t_s), len(c_s))) if c_s[i] is not None and c_s[i] > 0}
            map_index = {t_m[i]: c_m[i] for i in range(min(len(t_m), len(c_m))) if c_m[i] is not None and c_m[i] > 0}
            common_t = sorted(list(set(map_stock.keys()) & set(map_index.keys())))
            aligned = []
            for t in common_t:
                date_str = datetime.datetime.fromtimestamp(t).strftime("%Y-%m-%d")
                p_s = map_stock[t]
                if p_s < 1000:
                    p_s = p_s * 1000
                aligned.append((date_str, p_s, map_index[t]))
            return aligned
    except Exception as e:
        print(f"[Beta Calc] Error fetching history: {e}")
    return []


def fetch_beta_vietstock(ticker, timeout=15):
    try:
        search_url = f"https://finance.vietstock.vn/search?query={ticker}"
        headers = {'User-Agent': RF_UA_STR}
        r1 = requests.get(search_url, headers=headers, timeout=timeout)
        if r1.status_code == 200:
            data = json.loads(r1.text).get("data", "")
            target_url = ""
            for line in data.split('\r\n'):
                parts = line.split('|')
                if len(parts) >= 3 and parts[0].strip().upper() == ticker.upper():
                    target_url = parts[2]
                    break
            if target_url:
                r2 = requests.get(target_url, headers={'User-Agent': RF_UA_STR, 'Referer': 'https://finance.vietstock.vn/'}, timeout=timeout)
                if r2.status_code == 200:
                    import re
                    m = re.search(r'\"Beta\":\"([\d\.]+)\"', r2.text)
                    if m:
                        beta = float(m.group(1))
                        if 0.3 <= beta <= 2.5:
                            return beta
    except Exception as e:
        print(f"  [WARN] Vietstock scrape failed: {e}")
    return None


def fetch_beta_vietcap(ticker, timeout=15):
    try:
        url = f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}"
        r = requests.get(url, headers={"User-Agent": RF_UA_STR, "Referer": "https://trading.vietcap.com.vn/"}, timeout=timeout)
        if r.status_code == 200:
            d = r.json().get("data", {})
            beta = d.get("beta")
            if beta is not None and 0.3 <= float(beta) <= 2.5:
                return float(beta)
    except Exception:
        pass
    return None


def fetch_and_calc_beta(ticker, days=720, timeout=20, fallback=1.1):
    print(f"  [INFO] Đang tải lịch sử giá để tự tính Beta cho {ticker}...")
    aligned_data = fetch_aligned_history(ticker, days=days, timeout=timeout)
    latest_price = aligned_data[-1][1] if aligned_data else None
    num_sessions = len(aligned_data)
    web_beta = fetch_beta_vietstock(ticker, timeout) or fetch_beta_vietcap(ticker, timeout) or fallback
    calculated_beta, is_enough_sessions = fallback, False
    if num_sessions >= 30:
        sliced_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
        s = [x[1] for x in sliced_data]
        m = [x[2] for x in sliced_data]
        rs = [(s[i] - s[i-1]) / s[i-1] for i in range(1, len(s))]
        rm = [(m[i] - m[i-1]) / m[i-1] for i in range(1, len(m))]
        n_ret = len(rs)
        mean_rs, mean_rm = sum(rs) / n_ret, sum(rm) / n_ret
        cov_sm = sum((rs[i]-mean_rs)*(rm[i]-mean_rm) for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 0
        var_m = sum((rm[i]-mean_rm)**2 for i in range(n_ret)) / (n_ret-1) if n_ret > 1 else 1.0
        calculated_beta = round(max(0.3, min(2.5, cov_sm / var_m if var_m > 0 else fallback)), 4)
        if num_sessions >= 250:
            is_enough_sessions = True
        aligned_data = aligned_data[-501:] if num_sessions > 500 else aligned_data
    beta_src = f"Tự tính toán ({num_sessions} phiên)" if is_enough_sessions else f"Web/API ({web_beta:.2f}) - lịch sử chỉ {num_sessions} phiên"
    return calculated_beta, web_beta, is_enough_sessions, beta_src, latest_price, aligned_data


# ══════════════════════════════════════════════════════════════════════════
# TRƯỜNG DỮ LIỆU VIETCAP RIÊNG CHO CTCK (income-statement-securities "iss" +
# balance-sheet-securities "bss") — xác nhận trực tiếp qua live API (SSI, 2026-07)
# ══════════════════════════════════════════════════════════════════════════
SEG = {
    "brokerage_rev": "iss42",     # Doanh thu nghiệp vụ môi giới chứng khoán
    "brokerage_cost": "iss133",   # Chi phí nghiệp vụ môi giới chứng khoán (âm)
    "margin_rev": "iss120",       # Lãi từ các khoản cho vay và phải thu
    "fvtpl_gain": "iss115",       # Lãi từ tài sản tài chính FVTPL
    "fvtpl_loss": "iss124",       # Lỗ tài sản tài chính FVTPL (âm)
    "fvtpl_cost": "iss132",       # Chi phí hoạt động tự doanh (âm)
    "ib_underwrite": "iss44",     # Doanh thu bảo lãnh phát hành
    "ib_advisory": "iss46",       # Doanh thu tư vấn đầu tư chứng khoán
    "ib_finadvisory": "iss123",   # Doanh thu tư vấn tài chính
    "custody_rev": "iss47",       # Doanh thu lưu ký chứng khoán
    "ib_cost": "iss134",          # Chi phí bảo lãnh, đại lý phát hành (âm)
    "advisory_cost": "iss135",    # Chi phí tư vấn đầu tư (âm)
    "finadvisory_cost": "iss138", # Chi phí tư vấn tài chính (âm)
    "custody_cost": "iss137",     # Chi phí nghiệp vụ lưu ký (âm)
    "other_rev": "iss50",         # Doanh thu khác (bao gồm QLQ nếu có, phần dư)
    "other_cost": "iss139",       # Chi phí các dịch vụ khác (âm)
    "fin_income": "iss141",       # Doanh thu hoạt động tài chính
    "fin_expense": "iss146",      # Chi phí tài chính (âm, gồm lãi vay funding margin+tự doanh)
    "interest_expense": "iss148", # Chi phí lãi vay (âm)
}
IS_TOTAL = {
    "total_rev": "isa1",     # DOANH THU HOẠT ĐỘNG
    "total_cost": "isa4",    # CHI PHÍ HOẠT ĐỘNG (âm)
    "gross_profit": "isa5",  # LỢI NHUẬN GỘP
    "sga": "isa10",          # CHI PHÍ QUẢN LÝ CÔNG TY CHỨNG KHOÁN (âm)
    "operating_result": "isa11",  # KẾT QUẢ HOẠT ĐỘNG
    "pbt": "isa16",          # LNTT (nhãn gốc Vietcap ghi nhầm "CHI PHÍ THUẾ..." nhưng đúng vị trí LNTT)
    "tax_current": "isa17",  # Chi phí thuế hiện hành (âm)
    "tax_deferred": "isa18", # Chi phí thuế hoãn lại (âm)
    "npat": "isa20",         # LNST
    "npat_parent": "isa22",  # LNST phân bổ chủ sở hữu
    "eps_basic": "isa23",    # EPS cơ bản
}
BS = {
    "fvtpl_portfolio": "bsa6",   # Tài sản tài chính FVTPL (danh mục tự doanh)
    "margin_loans": "bss215",   # Các khoản cho vay (dư nợ margin + ứng trước)
    "total_assets": "bsa53",    # TỔNG CỘNG TÀI SẢN
    "total_liab": "bsa54",      # NỢ PHẢI TRẢ
    "short_borrow": "bsa56",    # Vay và nợ thuê tài sản tài chính ngắn hạn
    "long_borrow": "bsa71",     # Vay và nợ thuê tài sản tài chính dài hạn
    "equity": "bsa78",          # VỐN CHỦ SỞ HỮU
    "charter_capital": "bsa80", # Vốn điều lệ
    "cash": "bsa2",  # Tiền và tương đương tiền
}
SEGMENT_NAMES = ["MoiGioi", "Margin", "TuDoanh", "IB_LuuKy", "QLQ"]
SEGMENT_LABELS_VI = {
    "MoiGioi": "Môi giới", "Margin": "Cho vay Margin", "TuDoanh": "Tự doanh (FVTPL)",
    "IB_LuuKy": "IB + Lưu ký", "QLQ": "Quản lý quỹ",
}
SEGMENT_COLORS = {"MoiGioi": "#3b82f6", "Margin": "#f59e0b", "TuDoanh": "#10b981",
                  "IB_LuuKy": "#8b5cf6", "QLQ": "#ec4899"}


def _get_yr(records, year, field):
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            if v is not None:
                return v / 1e9
    return 0.0


def _get_yr_raw(records, year, field):
    """Giống _get_yr nhưng trả về 0 nếu field không tồn tại trong record (không phải None do thiếu dữ
    liệu) — dùng để phân biệt 'CTCK này không có dòng doanh thu này' (hợp lệ, VD không có mảng QLQ)
    với 'lỗi fetch dữ liệu'."""
    for r in records:
        if r.get("yearReport") == year:
            v = r.get(field)
            return (v or 0.0) / 1e9
    return 0.0


# ══════════════════════════════════════════════════════════════════════════
# STYLING CONSTANTS
# ══════════════════════════════════════════════════════════════════════════
FONT_NAME = "Calibri"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F4E78")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name=FONT_NAME, size=11, bold=True)
ITALIC_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")
DATA_FONT = Font(name=FONT_NAME, size=10)
ASSUMP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LINK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
P_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                      top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
FMT_NUM = '#,##0'
FMT_PCT = '0.00%'
FMT_MUL = '0.00"x"'
FMT_PRICE = '#,##0'


def header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def data_row(ws, row, label, values, fmt=None, note=None, bold=False, fill=None):
    c0 = ws.cell(row=row, column=1, value=label)
    c0.font = BOLD_FONT if bold else DATA_FONT
    for i, v in enumerate(values, start=2):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BOLD_FONT if bold else DATA_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="right")
        f = fmt
        if isinstance(fmt, list):
            f = fmt[i - 2] if i - 2 < len(fmt) else None
        if f:
            c.number_format = f
        if fill:
            c.fill = fill
    if note:
        note_col = len(values) + 2
        nc = ws.cell(row=row, column=note_col, value=note)
        nc.font = ITALIC_FONT
        nc.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


# ══════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════
def run_securities_analysis(ticker: str, raw_data: dict) -> bool:
    ticker = ticker.upper()
    print(f"\n--- Running Securities (CTCK) Analysis for {ticker} ---")

    is_recs = section_to_years(raw_data, "INCOME_STATEMENT")
    bs_recs = section_to_years(raw_data, "BALANCE_SHEET")
    is_q = section_to_quarters(raw_data, "INCOME_STATEMENT")
    bs_q = section_to_quarters(raw_data, "BALANCE_SHEET")

    available_years = sorted(set(r.get("yearReport") for r in is_recs if r.get("yearReport")))
    if not available_years or len(available_years) < 3:
        print("[ERROR] Too few historical data years found!")
        return False

    years_hist = available_years[-5:]
    years_fc = [years_hist[-1] + 1, years_hist[-1] + 2, years_hist[-1] + 3]
    all_years = years_hist + years_fc
    N_HIST = len(years_hist)
    company_name = raw_data.get("companyName") or f"CTCP Chứng khoán {ticker}"

    # ── Company details: giá hiện tại, số CP, tên công ty (live fetch, giống pattern banking) ──
    current_price, shares, industry_full = None, None, "Chứng khoán"
    try:
        r_det = requests.get(f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/details?ticker={ticker}",
                              headers={"User-Agent": RF_UA_STR, "Referer": "https://trading.vietcap.com.vn/"}, timeout=10)
        if r_det.status_code == 200:
            d = r_det.json().get("data", {})
            if d.get("currentPrice"):
                current_price = float(d["currentPrice"])
            if d.get("numberOfSharesMktCap"):
                shares = float(d["numberOfSharesMktCap"])
            if d.get("viOrganName"):
                company_name = d["viOrganName"]
    except Exception as e:
        print(f"  [WARN] Failed to fetch company details: {e}")

    latest_hist_year = years_hist[-1]
    if not shares:
        charter_capital = _get_yr(bs_recs, latest_hist_year, BS["charter_capital"])
        shares = (charter_capital * 1e9 / 10000) if charter_capital > 0 else 300_000_000
    shares = int(shares)
    if not current_price:
        current_price = 25000.0
    market_cap = current_price * shares / 1e9  # tỷ VND

    print(f"  -> {ticker}: {company_name} | Giá {current_price:,.0f} VND | {shares:,.0f} CP | Vốn hóa {market_cap:,.0f} tỷ")

    # ══════════════════════════════════════════════════════════════════
    # 1. LỊCH SỬ 5 MẢNG KINH DOANH (từ Vietcap iss field, xem SEG dict)
    # ══════════════════════════════════════════════════════════════════
    brokerage_rev_hist = [_get_yr(is_recs, y, SEG["brokerage_rev"]) for y in years_hist]
    brokerage_cost_hist = [_get_yr(is_recs, y, SEG["brokerage_cost"]) for y in years_hist]
    margin_rev_hist = [_get_yr(is_recs, y, SEG["margin_rev"]) for y in years_hist]
    fvtpl_gain_hist = [_get_yr(is_recs, y, SEG["fvtpl_gain"]) for y in years_hist]
    fvtpl_loss_hist = [_get_yr(is_recs, y, SEG["fvtpl_loss"]) for y in years_hist]
    fvtpl_net_hist = [fvtpl_gain_hist[i] + fvtpl_loss_hist[i] for i in range(N_HIST)]
    fvtpl_cost_hist = [_get_yr(is_recs, y, SEG["fvtpl_cost"]) for y in years_hist]
    ib_rev_hist = [_get_yr(is_recs, y, SEG["ib_underwrite"]) + _get_yr(is_recs, y, SEG["ib_advisory"])
                   + _get_yr(is_recs, y, SEG["ib_finadvisory"]) for y in years_hist]
    custody_rev_hist = [_get_yr(is_recs, y, SEG["custody_rev"]) for y in years_hist]
    ib_custody_rev_hist = [ib_rev_hist[i] + custody_rev_hist[i] for i in range(N_HIST)]
    other_rev_hist = [_get_yr(is_recs, y, SEG["other_rev"]) for y in years_hist]
    # QLQ: Vietcap không có field riêng cho hầu hết CTCK (mảng QLQ thường nằm ở công ty con AM hạch
    # toán hợp nhất) — coi "other_rev" là proxy QLQ + phụ trợ khác, tách 40% làm ước tính QLQ tối thiểu
    # để driver AUM×fee có cơ sở lịch sử tham chiếu (KHÔNG chính xác tuyệt đối, ghi rõ trong Assumptions).
    qlq_rev_hist = [round(other_rev_hist[i] * 0.4, 2) for i in range(N_HIST)]
    other_rev_adj_hist = [round(other_rev_hist[i] * 0.6, 2) for i in range(N_HIST)]

    total_rev_hist = [_get_yr(is_recs, y, IS_TOTAL["total_rev"]) for y in years_hist]
    total_cost_hist = [_get_yr(is_recs, y, IS_TOTAL["total_cost"]) for y in years_hist]
    gross_profit_hist = [_get_yr(is_recs, y, IS_TOTAL["gross_profit"]) for y in years_hist]
    sga_hist = [_get_yr(is_recs, y, IS_TOTAL["sga"]) for y in years_hist]
    fin_income_hist = [_get_yr(is_recs, y, SEG["fin_income"]) for y in years_hist]
    fin_expense_hist = [_get_yr(is_recs, y, SEG["fin_expense"]) for y in years_hist]
    pbt_hist = [_get_yr(is_recs, y, IS_TOTAL["pbt"]) for y in years_hist]
    tax_hist = [_get_yr(is_recs, y, IS_TOTAL["tax_current"]) + _get_yr(is_recs, y, IS_TOTAL["tax_deferred"]) for y in years_hist]
    npat_hist = [_get_yr(is_recs, y, IS_TOTAL["npat"]) for y in years_hist]
    npat_parent_hist = [_get_yr(is_recs, y, IS_TOTAL["npat_parent"]) for y in years_hist]

    # ── Bảng cân đối — driver cho margin/tự doanh ──
    fvtpl_portfolio_hist = [_get_yr(bs_recs, y, BS["fvtpl_portfolio"]) for y in years_hist]
    margin_loans_hist = [_get_yr(bs_recs, y, BS["margin_loans"]) for y in years_hist]
    total_assets_hist = [_get_yr(bs_recs, y, BS["total_assets"]) for y in years_hist]
    total_liab_hist = [_get_yr(bs_recs, y, BS["total_liab"]) for y in years_hist]
    short_borrow_hist = [_get_yr(bs_recs, y, BS["short_borrow"]) for y in years_hist]
    long_borrow_hist = [_get_yr(bs_recs, y, BS["long_borrow"]) for y in years_hist]
    equity_hist = [_get_yr(bs_recs, y, BS["equity"]) for y in years_hist]
    charter_hist = [_get_yr(bs_recs, y, BS["charter_capital"]) for y in years_hist]
    borrow_total_hist = [short_borrow_hist[i] + long_borrow_hist[i] for i in range(N_HIST)]

    # ── Dư nợ/danh mục bình quân (đầu kỳ+cuối kỳ)/2 ──
    def avg_series(hist):
        return [hist[0]] + [(hist[i-1] + hist[i]) / 2 for i in range(1, len(hist))]

    margin_avg_hist = avg_series(margin_loans_hist)
    fvtpl_avg_hist = avg_series(fvtpl_portfolio_hist)
    borrow_avg_hist = avg_series(borrow_total_hist)

    # ── NIM cho vay margin lịch sử = Lãi cho vay / Dư nợ bình quân (gộp cả phần bù đắp chi phí vốn —
    # để tách riêng chi phí vốn cần phân bổ chi phí lãi vay theo tỷ trọng margin/tự doanh, vốn không
    # tách bạch được từ BCTC công khai — dùng NIM GỘP (net yield) làm driver, đã net chi phí vốn ngầm
    # định qua chính biên độ quan sát được lịch sử, đúng tinh thần "spread cho vay" trong tài liệu) ──
    margin_nim_hist = [round(margin_rev_hist[i] / margin_avg_hist[i], 4) if margin_avg_hist[i] else 0 for i in range(N_HIST)]
    fvtpl_yield_hist = [round(fvtpl_net_hist[i] / fvtpl_avg_hist[i], 4) if fvtpl_avg_hist[i] else 0 for i in range(N_HIST)]
    cost_of_borrow_hist = [round(abs(_get_yr(is_recs, years_hist[i], SEG["interest_expense"])) / borrow_avg_hist[i], 4)
                            if borrow_avg_hist[i] else 0 for i in range(N_HIST)]

    print(f"  -> NIM Margin lịch sử: {[f'{x*100:.1f}%' for x in margin_nim_hist]}")
    print(f"  -> Hiệu suất tự doanh lịch sử: {[f'{x*100:.1f}%' for x in fvtpl_yield_hist]}")

    # ══════════════════════════════════════════════════════════════════
    # 2. CAPM: Rf + Beta + COE
    # ══════════════════════════════════════════════════════════════════
    rf_val, rf_src = fetch_rf_vietnam()
    beta_calc, beta_web, is_enough_sessions, beta_src, _latest_px, BETA_ALIGNED_DATA = fetch_and_calc_beta(ticker)
    beta_raw = beta_calc if is_enough_sessions else beta_web
    beta_val = round(0.67 * beta_raw + 0.33, 4)  # Blume adjustment
    ERP = 0.07
    SPECIFIC_RISK_PREMIUM = 0.02  # CTCK rủi ro thị trường cao hơn (đòn bẩy tự doanh + margin)
    COE = rf_val + beta_val * ERP + SPECIFIC_RISK_PREMIUM
    print(f"  -> Beta thô: {beta_raw:.4f} | Beta Blume: {beta_val} ({beta_src})")
    print(f"  -> COE: {COE*100:.2f}% (Rf={rf_val*100:.2f}% [{rf_src}], ERP={ERP*100:.1f}%, Specific={SPECIFIC_RISK_PREMIUM*100:.1f}%)")

    # ══════════════════════════════════════════════════════════════════
    # 3. P/E, P/B LỊCH SỬ (median toàn bộ quý, cùng chuẩn skill ngan-hang/thep)
    # ══════════════════════════════════════════════════════════════════
    pe_quarters, pb_quarters, quarter_labels = [], [], []
    try:
        r = requests.get(f"https://trading.vietcap.com.vn/api/iq-insight-service/v1/company/{ticker}/statistics-financial",
                          headers={"User-Agent": RF_UA_STR, "Referer": "https://trading.vietcap.com.vn/"}, timeout=15)
        if r.status_code == 200:
            data = r.json().get("data", [])
            ttms = sorted([x for x in data if x.get("year") and x.get("quarter") in (1, 2, 3, 4)],
                          key=lambda x: (x["year"], x["quarter"]))
            for x in ttms:
                quarter_labels.append(f"{x['year']}-Q{x['quarter']}")
                pe_quarters.append(round(x["pe"], 2) if x.get("pe") else None)
                pb_quarters.append(round(x["pb"], 2) if x.get("pb") else None)
            print(f"  -> Đã lấy {len(quarter_labels)} quý P/E, P/B lịch sử từ Vietcap")
    except Exception as e:
        print(f"  [WARN] Live P/E, P/B fetch failed: {e}")

    _pe_valid = [v for v in pe_quarters if v and 0 < v < 60]
    _pb_valid = [v for v in pb_quarters if v and v > 0]
    PE_HIST_MEDIAN = round(stats.median(_pe_valid), 2) if _pe_valid else 12.0
    PB_HIST_MEDIAN = round(stats.median(_pb_valid), 2) if _pb_valid else 1.3
    PE_PB_MEDIAN_ROW = len(quarter_labels) + 3
    print(f"  -> PE_HIST_MEDIAN={PE_HIST_MEDIAN}x | PB_HIST_MEDIAN={PB_HIST_MEDIAN}x (median {len(_pe_valid)}/{len(_pb_valid)} quý)")

    # ══════════════════════════════════════════════════════════════════
    # 4. GIẢ ĐỊNH DỰ PHÓNG — 5 MẢNG (Bottom-up, theo đúng công thức tài liệu)
    # ══════════════════════════════════════════════════════════════════
    # ── (1) Môi giới: GTGD/phiên × Số phiên × Thị phần × Phí(bps)/10,000 ──
    # GTGD toàn TT bình quân/phiên (tỷ VND) — dữ liệu THỰC TẾ từ HOSE/SSC (cập nhật 07/2026).
    # Nguồn: HOSE thống kê hàng tháng, SSC báo cáo thị trường hàng quý.
    # ⚠ Cần cập nhật khi có số liệu mới hơn từ HOSE/SSC — không dùng số giả định tuyến tính.
    # ----------------------------------------------------------------
    # Phương pháp tính ADTV năm: bình quân có trọng số tháng (năm đủ) hoặc ước tính
    # từ xu hướng nửa năm đã công bố (năm hiện tại).
    #   2021: Bùng nổ COVID — bình quân ~23,500 tỷ/phiên (HOSE thống kê)
    #   2022: Điều chỉnh mạnh — bình quân ~16,800 tỷ/phiên (HOSE)
    #   2023: Phục hồi chậm — bình quân ~14,800 tỷ/phiên (HOSE)
    #   2024: Hồi phục — bình quân ~20,500 tỷ/phiên (HOSE)
    #   2025: Tăng mạnh — bình quân ~28,500 tỷ/phiên (HOSE ước tính)
    #   2026: Q1=35,004 tỷ; T4=24,101 tỷ (đã hạ nhiệt) → ước tính cả năm ~28,000 tỷ
    _ADTV_ACTUAL = {
        2021: 23500.0, 2022: 16800.0, 2023: 14800.0,
        2024: 20500.0, 2025: 28500.0, 2026: 28000.0,
    }
    # Năm nào không có trong bảng → dùng nội suy tuyến tính từ 2 năm gần nhất
    _adtv_sorted = sorted(_ADTV_ACTUAL.keys())
    def _adtv_for_year(y):
        if y in _ADTV_ACTUAL:
            return _ADTV_ACTUAL[y]
        if y < _adtv_sorted[0]:
            return _ADTV_ACTUAL[_adtv_sorted[0]]
        if y > _adtv_sorted[-1]:
            # extrapolate từ 2 năm cuối
            y1, y2 = _adtv_sorted[-2], _adtv_sorted[-1]
            slope = (_ADTV_ACTUAL[y2] - _ADTV_ACTUAL[y1]) / (y2 - y1)
            return _ADTV_ACTUAL[y2] + slope * (y - y2)
        # interpolate
        for j in range(len(_adtv_sorted) - 1):
            y1, y2 = _adtv_sorted[j], _adtv_sorted[j + 1]
            if y1 <= y <= y2:
                t = (y - y1) / (y2 - y1)
                return _ADTV_ACTUAL[y1] * (1 - t) + _ADTV_ACTUAL[y2] * t
        return 20000.0
    MARKET_ADTV_HIST = {y: round(_adtv_for_year(y), 0) for y in years_hist}
    TRADING_DAYS = 250
    # Thị phần môi giới ngụ ý lịch sử = DT môi giới / (ADTV × Số phiên × phí bình quân giả định 0.15%)
    ASSUMED_FEE_BPS = 15.0  # bps, ~0.15% — mức phí môi giới bình quân phổ biến hiện nay
    brokerage_share_hist = [round(brokerage_rev_hist[i] / (MARKET_ADTV_HIST[years_hist[i]] * TRADING_DAYS * ASSUMED_FEE_BPS / 10000), 4)
                             if MARKET_ADTV_HIST[years_hist[i]] > 0 else 0 for i in range(N_HIST)]
    market_share_fc = round(stats.mean(brokerage_share_hist[-2:]), 4) if len(brokerage_share_hist) >= 2 else brokerage_share_hist[-1]

    # ── FALLBACK OLS cho CTCK nhỏ ──
    # Khi thị phần ngụ ý < 0.5%: CTCK quá nhỏ, độ biến động của market_share lớn và khó nhất quán.
    # Thay bằng hồi quy OLS qua gốc toạ độ: brokerage_rev = β × ADTV  (nội suy tỷ lệ thanh khoản).
    # β = Σ(brok_i × adtv_i) / Σ(adtv_i²)  — OLS qua gốc, tả nhiên hơn khi brok ~ const × ADTV.
    MIN_RELIABLE_SHARE = 0.005   # ngưỡng 0.5% thị phần → dưới mức này dùng OLS
    _brok_adtv_pairs = [(brokerage_rev_hist[i], MARKET_ADTV_HIST[years_hist[i]])
                        for i in range(N_HIST)
                        if MARKET_ADTV_HIST[years_hist[i]] > 0 and brokerage_rev_hist[i] > 0]
    _use_ols = (market_share_fc < MIN_RELIABLE_SHARE) and (len(_brok_adtv_pairs) >= 2)
    if _use_ols:
        _num = sum(b * a for b, a in _brok_adtv_pairs)
        _den = sum(a * a for _, a in _brok_adtv_pairs)
        _beta_ols = _num / _den if _den > 0 else 0.0
        _brok_model = f'OLS β={_beta_ols:.6f}'
        _ms_ols_equiv = round(_beta_ols * TRADING_DAYS * ASSUMED_FEE_BPS / 10000, 4)
        print(f"  [Brokerage] {ticker}: OLS fallback (β={_beta_ols:.6f}) — thị phần ngụ ý {market_share_fc:.2%} < {MIN_RELIABLE_SHARE:.1%} ngưỡng")
        print(f"    Implied share tương đương OLS: {_ms_ols_equiv:.2%}")
    else:
        _beta_ols = None
        _brok_model = f'Market share {market_share_fc:.2%}'
        print(f"  [Brokerage] {ticker}: Market share {market_share_fc:.2%} — đủ tin cậy (≥ {MIN_RELIABLE_SHARE:.1%})")

    # Tăng trưởng GTGD toàn TT GIẢM DẦN qua các năm dự phóng (8%→6%→5%) — tránh dự phóng quá lạc quan/
    # "quá đà" cho các năm xa, đúng nguyên tắc thận trọng hóa dần theo thời gian.
    # ⚠ Điều chỉnh taper nếu quan điểm vĩ mô thay đổi (xem sheet 09_PESTLE/khung đánh giá vĩ mô).
    ADTV_GROWTH_TAPER = [0.08, 0.06, 0.05]
    market_adtv_fc = []
    _adtv_prev = MARKET_ADTV_HIST[years_hist[-1]]
    for _g in ADTV_GROWTH_TAPER:
        _adtv_prev = _adtv_prev * (1 + _g)
        market_adtv_fc.append(round(_adtv_prev, 0))

    # ── (2) Cho vay Margin: Dư nợ bình quân × NIM ──
    margin_nim_fc_base = round(stats.mean(margin_nim_hist[-2:]), 4) if len(margin_nim_hist) >= 2 else margin_nim_hist[-1]
    # Tăng trưởng dư nợ Margin GIẢM DẦN (15%→10%→8%) — năm 1 neo theo đà hiện tại, các năm sau thận
    # trọng hơn vì dư nợ margin có TRẦN PHÁP LÝ 2,0x VCSH (xem sheet 10_Hieu_Qua_Mang), không thể
    # duy trì tốc độ cao liên tục nhiều năm nếu công ty chưa có kế hoạch tăng vốn tương ứng
    margin_loan_growth_fc = [0.15, 0.10, 0.08]

    # ── (3) FVTPL (Tự doanh): Danh mục × (%CDs×R_CDs + %TP×R_TP + %CP×R_VNI) ──
    # Tỷ trọng CDs/TP/CP PHÂN BIỆT THEO TỪNG CTCK dựa trên đặc điểm kinh doanh:
    #   SSI: thiên về CP chủ động & phái sinh (~45%), tỷ trọng CDs thấp hơn do chủ yếu dùng vốn vay
    #        (SSI XTRADE/SSI Finvest), TP ~25% phòng thủ. Nguồn: SSI BCTC 2023-2025, thuyết minh mục IV.
    #   VCI: thiên về CDs & TP cố định (~85%), rất ít CP chủ động (~15%) — chiến lược bảo thủ, ưu tiên
    #        thu nhập lãi cố định. Nguồn: VCI BCTC 2023-2025, thuyết minh mục "Chứng khoán kinh doanh".
    #   _DEFAULT: dùng khi không có thông tin thuyết minh cụ thể — tỷ trọng ngành bình quân.
    # ⚠ BẮT BUỘC cập nhật khi có thuyết minh BCTC mới nhất (mỗi quý/năm) — con số này ảnh hưởng
    #   TRỰC TIẾP tới dự phóng doanh thu tự doanh và qua đó ảnh hưởng LN & định giá.
    _FVTPL_MIX_OVERRIDE = {
        "SSI":      {"CDs": 0.30, "TP": 0.25, "CP": 0.45},  # SSI: tự doanh chủ động, CP cao
        "VCI":      {"CDs": 0.50, "TP": 0.35, "CP": 0.15},  # VCI: bảo thủ, CDs+TP >80%
        "HCM":      {"CDs": 0.35, "TP": 0.30, "CP": 0.35},  # HCM: cân bằng
        "MBS":      {"CDs": 0.40, "TP": 0.35, "CP": 0.25},
        "_DEFAULT": {"CDs": 0.45, "TP": 0.25, "CP": 0.30},
    }
    fvtpl_mix_fc = _FVTPL_MIX_OVERRIDE.get(ticker, _FVTPL_MIX_OVERRIDE["_DEFAULT"])
    print(f"  [FVTPL Mix] {ticker}: CDs={fvtpl_mix_fc['CDs']:.0%} / TP={fvtpl_mix_fc['TP']:.0%} / CP={fvtpl_mix_fc['CP']:.0%}"
          f" {'(per-ticker override)' if ticker in _FVTPL_MIX_OVERRIDE else '(default)'}")
    # R_CDs = 7.0%: kiểm chứng qua lãi suất CDs 12 tháng thực tế 2026-07 (Vietcombank 7.4-7.9%, BVBank 7.2%,
    # dao động 5.5%-7.9% tùy NH — neo mức trung bình-thận trọng, KHÔNG dùng mức đỉnh).
    # R_TP = 7.0%: coupon trái phiếu NH bình quân phát hành 2026 là 6.4-6.5%, một số NH phát hành 8-8.9%
    # — nằm trong dải hợp lý.
    # R_VNI = 10%: KHÔNG kiểm chứng được (kỳ vọng lợi suất tương lai mang tính chủ quan, không có "đúng/sai") —
    # đối chiếu với COE (00_COE) hoặc quan điểm thị trường của người phân tích khi điều chỉnh.
    fvtpl_r_fc = {"R_CDs": 0.070, "R_TP": 0.07, "R_VNI": 0.10}  # lãi suất/lợi suất kỳ vọng mỗi loại tài sản
    _FVTPL_RATE_KEY = {"CDs": "R_CDs", "TP": "R_TP", "CP": "R_VNI"}
    fvtpl_expected_yield_fc = sum(fvtpl_mix_fc[k] * fvtpl_r_fc[_FVTPL_RATE_KEY[k]] for k in fvtpl_mix_fc)
    # Tăng trưởng danh mục Tự doanh GIẢM DẦN (10%→8%→6%)
    fvtpl_portfolio_growth_fc = [0.10, 0.08, 0.06]

    # ── (4) IB + Lưu ký: Pipeline IB × Fee% + AUM lưu ký × Fee lưu ký% ──
    ib_pipeline_fc = round(stats.mean(ib_rev_hist[-2:]) / 0.02, 1) if len(ib_rev_hist) >= 2 and stats.mean(ib_rev_hist[-2:]) else 3000.0
    # ⚠ 2%: phí M&A/IPO tại VN KHÔNG công bố công khai (đàm phán riêng từng deal) — KHÔNG kiểm chứng được,
    # neo theo thông lệ quốc tế phổ biến (M&A advisory ~1-3% giá trị deal).
    ib_fee_pct_fc = 0.02  # phí IB bình quân trên giá trị deal
    custody_growth_fc = round(stats.mean([custody_rev_hist[i] / custody_rev_hist[i-1] - 1 for i in range(1, N_HIST) if custody_rev_hist[i-1]] or [0.1]), 4)

    # ── (5) Quản lý quỹ: AUM × Fee rate (~0.75%/năm) ──
    # ⚠ 0.75%: phí quản lý quỹ mở tại VN dao động rộng 0.1%-2%/năm tùy loại quỹ (tiền tệ/trái phiếu thấp,
    # cổ phiếu chủ động cao hơn) — 0.75% là mức trung bình-thận trọng cho danh mục hỗn hợp, KHÔNG kiểm
    # chứng theo từng công ty quản lý quỹ cụ thể.
    qlq_fee_rate_fc = 0.0075
    qlq_aum_fc0 = round(qlq_rev_hist[-1] / qlq_fee_rate_fc, 1) if qlq_rev_hist[-1] > 0 else 2000.0
    # Tăng trưởng AUM QLQ GIẢM DẦN (20%→15%→12%)
    qlq_aum_growth_fc = [0.20, 0.15, 0.12]

    # ── Chi phí, thuế ──
    sga_pct_rev_hist = [abs(sga_hist[i]) / total_rev_hist[i] if total_rev_hist[i] else 0 for i in range(N_HIST)]
    sga_pct_fc = round(stats.mean(sga_pct_rev_hist[-2:]), 4) if len(sga_pct_rev_hist) >= 2 else 0.03
    cost_pct_of_rev_hist = [abs(total_cost_hist[i]) / total_rev_hist[i] if total_rev_hist[i] else 0 for i in range(N_HIST)]
    cost_pct_fc = round(stats.mean(cost_pct_of_rev_hist[-2:]), 4) if len(cost_pct_of_rev_hist) >= 2 else 0.35
    tax_rate_hist = [abs(tax_hist[i]) / pbt_hist[i] if pbt_hist[i] else 0.2 for i in range(N_HIST)]
    tax_rate_fc = min(max(round(stats.mean([t for t in tax_rate_hist if 0 < t < 0.3] or [0.2]), 4), 0.15), 0.22)

    # ══════════════════════════════════════════════════════════════════
    # 5. DỰ PHÓNG DOANH THU 5 MẢNG (Python mirror — Excel sẽ có công thức sống riêng)
    # ══════════════════════════════════════════════════════════════════
    brokerage_rev_fc, margin_rev_fc, fvtpl_rev_fc, ib_custody_rev_fc, qlq_rev_fc = [], [], [], [], []
    margin_loans_fc, fvtpl_portfolio_fc, qlq_aum_fc = [], [], []

    _margin_prev = margin_loans_hist[-1]
    _fvtpl_prev = fvtpl_portfolio_hist[-1]
    _qlq_aum_prev = qlq_aum_fc0
    for i in range(3):
        # (1) Môi giới — 2 phương pháp tùy độ tin cậy thị phần:
        if _use_ols:
            # OLS: DT môi giới = β × ADTV_fc (nội suy tuyến tính tỷ lệ thanh khoản)
            brokerage_rev_fc.append(round(_beta_ols * market_adtv_fc[i], 1))
        else:
            # Market share: ADTV × số_phiên × thị_phần × phí_bps / 10000
            brokerage_rev_fc.append(round(market_adtv_fc[i] * TRADING_DAYS * market_share_fc * ASSUMED_FEE_BPS / 10000, 1))
        # (2) Margin
        margin_loans_fc.append(round(_margin_prev * (1 + margin_loan_growth_fc[i]), 1))
        margin_avg_fc_i = (_margin_prev + margin_loans_fc[i]) / 2
        margin_rev_fc.append(round(margin_avg_fc_i * margin_nim_fc_base, 1))
        _margin_prev = margin_loans_fc[i]
        # (3) FVTPL
        fvtpl_portfolio_fc.append(round(_fvtpl_prev * (1 + fvtpl_portfolio_growth_fc[i]), 1))
        fvtpl_avg_fc_i = (_fvtpl_prev + fvtpl_portfolio_fc[i]) / 2
        fvtpl_rev_fc.append(round(fvtpl_avg_fc_i * fvtpl_expected_yield_fc, 1))
        _fvtpl_prev = fvtpl_portfolio_fc[i]
        # (4) IB + Lưu ký: IB phẳng theo Pipeline×Fee% (Assumptions không đổi theo năm) + Lưu ký tăng trưởng kép
        ib_rev_i = ib_pipeline_fc * ib_fee_pct_fc
        custody_rev_i = custody_rev_hist[-1] * ((1 + custody_growth_fc) ** (i + 1))
        ib_custody_rev_fc.append(round(ib_rev_i + custody_rev_i, 1))
        # (5) QLQ
        qlq_aum_fc.append(round(_qlq_aum_prev * (1 + qlq_aum_growth_fc[i]), 1))
        qlq_rev_fc.append(round(qlq_aum_fc[i] * qlq_fee_rate_fc, 1))
        _qlq_aum_prev = qlq_aum_fc[i]

    # ── Blend mảng Môi giới năm hiện tại với dữ liệu lũy kế thực tế (segment-level blend) ──
    # Trước đây chỉ blend tổng DT và LNST. Giờ blend thêm mảng Môi giới riêng để capture xu hướng
    # thanh khoản thực tế quý gần nhất (thị trường có thể khác ADTV bảng cứng).
    # n=0 → blend no-op (giữ giá trị gốc), nhất quán với hàm blend_annual_estimate.
    _cur_fc_year = years_fc[0]
    _brok_cum, _n_brok_q = cumulative_actual_quarters(is_q, _cur_fc_year, SEG["brokerage_rev"])
    _brok_base0 = brokerage_rev_fc[0]
    brokerage_rev_fc[0] = round(blend_annual_estimate(_brok_cum, _n_brok_q, _brok_base0), 1)
    if _n_brok_q > 0:
        _brok_annual = _brok_cum * (4 / _n_brok_q)
        print(f"  [Blend-Môi giới] {_cur_fc_year}F: {_n_brok_q}/4 quý thực (lũy kế {_brok_cum:,.1f} tỷ, annualized {_brok_annual:,.1f}) "
              f"+ mô hình {_brok_base0:,.1f} → blend = {brokerage_rev_fc[0]:,.1f} tỷ (độ lệch mô hình: {(_brok_annual/_brok_base0-1)*100:+.1f}%)")

    total_rev_fc = [brokerage_rev_fc[i] + margin_rev_fc[i] + fvtpl_rev_fc[i] + ib_custody_rev_fc[i] + qlq_rev_fc[i] for i in range(3)]

    # ── Blend năm hiện tại (years_fc[0]) với số quý ĐÃ CÓ báo cáo thực tế ──
    # Công thức (thống nhất với HPG/MWG): Ước tính chuẩn = Lũy kế thực tế n quý đã biết + Base ước tính
    # gốc/4 × (4-n). blend_annual_estimate() tự động no-op (trả về Base gốc) khi n=0, nên gọi KHÔNG ĐIỀU
    # KIỆN — khớp với công thức SỐNG hiển thị ở sheet 04b_Dien_Bien_Quy (Excel/PDF/JSON dùng chung 1 số).
    _cur_fc_year = years_fc[0]
    _rev_cum, _n_rev_q = cumulative_actual_quarters(is_q, _cur_fc_year, IS_TOTAL["total_rev"])
    _rev_base0 = total_rev_fc[0]
    total_rev_fc[0] = round(blend_annual_estimate(_rev_cum, _n_rev_q, _rev_base0), 1)
    print(f"  [Blend] {_cur_fc_year}F Doanh thu HĐ: {_n_rev_q}/4 quý đã biết (lũy kế {_rev_cum:,.0f} tỷ, base gốc {_rev_base0:,.0f} tỷ) -> blend = {total_rev_fc[0]:,.0f} tỷ")

    total_cost_fc = [-round(total_rev_fc[i] * cost_pct_fc, 1) for i in range(3)]
    gross_profit_fc = [round(total_rev_fc[i] + total_cost_fc[i], 1) for i in range(3)]
    sga_fc = [-round(total_rev_fc[i] * sga_pct_fc, 1) for i in range(3)]
    operating_result_fc = [round(gross_profit_fc[i] + sga_fc[i], 1) for i in range(3)]
    pbt_fc = operating_result_fc[:]  # đơn giản hoá: bỏ qua phần "lãi/lỗ liên doanh liên kết" nhỏ

    tax_fc = [-round(max(pbt_fc[i], 0) * tax_rate_fc, 1) for i in range(3)]
    npat_fc = [round(pbt_fc[i] + tax_fc[i], 1) for i in range(3)]

    _npat_cum, _n_npat_q = cumulative_actual_quarters(is_q, _cur_fc_year, IS_TOTAL["npat_parent"])
    _npat_base0 = npat_fc[0]
    npat_fc[0] = round(blend_annual_estimate(_npat_cum, _n_npat_q, _npat_base0), 1)
    print(f"  [Blend] {_cur_fc_year}F LNST: {_n_npat_q}/4 quý đã biết (lũy kế {_npat_cum:,.0f} tỷ, base gốc {_npat_base0:,.0f} tỷ) -> blend = {npat_fc[0]:,.0f} tỷ")
    npat_parent_fc = npat_fc[:]  # đơn giản hoá: hầu hết CTCK sở hữu 100% (bỏ qua lợi ích thiểu số)

    eps_hist = [round(npat_parent_hist[i] * 1e9 / shares) for i in range(N_HIST)]
    eps_fc = [round(npat_parent_fc[i] * 1e9 / shares) for i in range(3)]
    equity_fc = [equity_hist[-1]]
    for i in range(3):
        equity_fc.append(equity_fc[-1] + npat_fc[i] * 0.8)  # giữ lại 80% LN, chia cổ tức 20%
    equity_fc = equity_fc[1:]
    bvps_hist = [round(equity_hist[i] * 1e9 / shares) for i in range(N_HIST)]
    bvps_fc = [round(equity_fc[i] * 1e9 / shares) for i in range(3)]

    print(f"  -> Doanh thu dự phóng {years_fc}: {[round(v) for v in total_rev_fc]} tỷ")
    print(f"  -> LNST dự phóng: {[round(v) for v in npat_fc]} tỷ | EPS: {eps_fc}")

    # ── Tỉ trọng đóng góp từng mảng & cảnh báo tập trung vốn ──
    seg_pct_fc0 = {
        "MoiGioi": brokerage_rev_fc[0] / total_rev_fc[0] if total_rev_fc[0] else 0,
        "Margin": margin_rev_fc[0] / total_rev_fc[0] if total_rev_fc[0] else 0,
        "TuDoanh": fvtpl_rev_fc[0] / total_rev_fc[0] if total_rev_fc[0] else 0,
        "IB_LuuKy": ib_custody_rev_fc[0] / total_rev_fc[0] if total_rev_fc[0] else 0,
        "QLQ": qlq_rev_fc[0] / total_rev_fc[0] if total_rev_fc[0] else 0,
    }
    concentration_warning = max(seg_pct_fc0, key=seg_pct_fc0.get)
    concentration_flag = "Cảnh báo" if seg_pct_fc0[concentration_warning] > 0.45 else "Bình thường"

    # ══════════════════════════════════════════════════════════════════
    # 6. ĐỊNH GIÁ — P/B ưu tiên số 1 (90%), P/E làm bộ lọc chống nhiễu (10%) CỐ ĐỊNH cho
    # MỌI CTCK — không hạ về 0% (mất khả năng đối chiếu chéo khi P/B bị nhiễu) nhưng cũng
    # không để cao (P/E không phải trọng số chính khi định giá tài sản CTCK — tài sản có
    # tính thanh khoản cao, BVPS đáng tin cậy hơn LNST vốn dễ biến động vì Tự doanh/IB).
    # ══════════════════════════════════════════════════════════════════
    pb_target_price = round(PB_HIST_MEDIAN * bvps_fc[0])
    pe_target_price = round(PE_HIST_MEDIAN * eps_fc[0])
    VALUATION_WEIGHTS = {"PB": 0.90, "PE": 0.10}
    weighted_target = round(VALUATION_WEIGHTS["PB"] * pb_target_price + VALUATION_WEIGHTS["PE"] * pe_target_price)
    upside_pct = round((weighted_target / current_price - 1) * 100, 1) if current_price else 0
    bear_target = round(min(pb_target_price, pe_target_price) * 0.85)
    bull_target = round(max(pb_target_price, pe_target_price) * 1.15)
    recommend = "MUA" if upside_pct > 15 else ("BÁN" if upside_pct < -5 else "THEO DÕI")
    print(f"  -> P/B target: {pb_target_price:,.0f} | P/E target: {pe_target_price:,.0f}")
    print(f"  -> Trọng số P/B {VALUATION_WEIGHTS['PB']*100:.0f}%/P/E {VALUATION_WEIGHTS['PE']*100:.0f}% (cố định cho mọi CTCK)")
    print(f"  -> GIÁ MỤC TIÊU: {weighted_target:,.0f} VND (giá hiện tại {current_price:,.0f}, upside {upside_pct}%)")

    # ── Tổng tỷ suất sinh lời (Upside + Tỷ suất cổ tức) — khớp cách trình bày báo cáo CTCK thực tế ──
    DIVIDEND_PAYOUT_RATIO = 0.20
    dps_fc0 = round(npat_fc[0] * DIVIDEND_PAYOUT_RATIO * 1e9 / shares)
    dividend_yield_pct = round(dps_fc0 / current_price * 100, 2) if current_price else 0
    total_return_pct = round(upside_pct + dividend_yield_pct, 1)
    print(f"  -> Cổ tức dự kiến: {dps_fc0:,.0f} VND/CP ({dividend_yield_pct}%) -> Tổng tỷ suất sinh lời: {total_return_pct}%")

    # ── Sensitivity: nếu VN-Index thay đổi X%, FVTPL biến động bao nhiêu ──
    vnindex_sens_range = [-0.20, -0.10, 0, 0.10, 0.20]
    fvtpl_cp_book_fc0 = fvtpl_portfolio_fc[0] * fvtpl_mix_fc["CP"]
    fvtpl_sens_impact = [round(fvtpl_cp_book_fc0 * s, 1) for s in vnindex_sens_range]
    fvtpl_sens_npat_impact_pct = [round(impact * (1 - tax_rate_fc) / npat_fc[0] * 100, 1) if npat_fc[0] else 0 for impact in fvtpl_sens_impact]

    # ══════════════════════════════════════════════════════════════════
    # 7. XÂY DỰNG EXCEL
    # ══════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_beta_coe_sheets(wb, ticker, beta_raw, beta_val, beta_src, BETA_ALIGNED_DATA, rf_val, rf_src, ERP,
                           SPECIFIC_RISK_PREMIUM, COE)
    build_cover_sheet(wb, ticker, company_name, current_price, market_cap, shares, weighted_target, upside_pct,
                       recommend, pb_target_price, pe_target_price, PE_HIST_MEDIAN, PB_HIST_MEDIAN)
    RA = build_assumptions_sheet(wb, ticker, years_hist, years_fc, current_price, shares, COE,
                                  MARKET_ADTV_HIST, TRADING_DAYS, ASSUMED_FEE_BPS, brokerage_share_hist, market_share_fc,
                                  market_adtv_fc, margin_nim_hist, margin_nim_fc_base, margin_loan_growth_fc,
                                  fvtpl_mix_fc, fvtpl_r_fc, fvtpl_expected_yield_fc, fvtpl_portfolio_growth_fc,
                                  ib_pipeline_fc, ib_fee_pct_fc, custody_growth_fc, qlq_fee_rate_fc, qlq_aum_fc0,
                                  qlq_aum_growth_fc, sga_pct_fc, cost_pct_fc, tax_rate_fc, PE_HIST_MEDIAN, PB_HIST_MEDIAN,
                                  PE_PB_MEDIAN_ROW)
    RR = build_revenue_model_sheet(wb, ticker, years_hist, years_fc, RA,
                                    brokerage_rev_hist, margin_rev_hist, fvtpl_net_hist, ib_custody_rev_hist, qlq_rev_hist,
                                    margin_loans_hist, fvtpl_portfolio_hist, qlq_aum_fc0,
                                    brokerage_rev_fc, margin_rev_fc, fvtpl_rev_fc, ib_custody_rev_fc, qlq_rev_fc,
                                    margin_loans_fc, fvtpl_portfolio_fc, qlq_aum_fc, total_rev_hist, total_rev_fc,
                                    ib_rev_hist, custody_rev_hist)
    RP = build_pnl_sheet(wb, ticker, years_hist, years_fc, RR, RA,
                          total_rev_hist, total_cost_hist, gross_profit_hist, sga_hist, pbt_hist, tax_hist,
                          npat_hist, npat_parent_hist, eps_hist,
                          total_rev_fc, total_cost_fc, gross_profit_fc, sga_fc, pbt_fc, tax_fc, npat_fc, npat_parent_fc, eps_fc)
    build_quarterly_blend_sheet(wb, ticker, is_q, years_fc, get_column_letter(2 + N_HIST), RR, RP)
    RBS = build_balance_sheet(wb, ticker, years_hist, years_fc, RP,
                               fvtpl_portfolio_hist, margin_loans_hist, total_assets_hist, total_liab_hist,
                               borrow_total_hist, equity_hist, fvtpl_portfolio_fc, margin_loans_fc, equity_fc)
    build_cash_flow_sheet(wb, ticker, years_hist, years_fc, RP, RBS, npat_hist, npat_fc)
    RV = build_valuation_sheet(wb, ticker, RA, RP, RBS, N_HIST, PE_HIST_MEDIAN, PB_HIST_MEDIAN, current_price,
                                pb_target_price, pe_target_price, weighted_target, upside_pct, VALUATION_WEIGHTS,
                                bear_target, bull_target, COE, dividend_yield_pct, total_return_pct)
    build_sensitivity_sheet(wb, ticker, RV, RP, vnindex_sens_range, fvtpl_sens_impact, fvtpl_sens_npat_impact_pct,
                             npat_fc, tax_rate_fc)
    build_pestle_sheet(wb, ticker)
    build_segment_efficiency_sheet(wb, ticker, years_hist, years_fc, RA, RR, RBS)
    build_thesis_sheet(wb, ticker, company_name, seg_pct_fc0, upside_pct, recommend)
    build_summary_snapshot(wb, ticker, years_hist, years_fc, total_rev_hist, total_rev_fc, npat_parent_hist, npat_parent_fc,
                            eps_hist, eps_fc, bvps_hist, bvps_fc, equity_hist, equity_fc, current_price, shares)
    build_pe_pb_history_sheet(wb, ticker, quarter_labels, pe_quarters, pb_quarters, PE_PB_MEDIAN_ROW)
    build_segment_quarterly_sheet(wb, ticker, is_q)
    build_peer_benchmark_sheet(wb, ticker)

    out_dir = os.path.join(PROJECT_ROOT, "Bao cao", ticker)
    os.makedirs(out_dir, exist_ok=True)
    month_str = datetime.datetime.now().strftime("%Y-%m-%d")
    excel_path = os.path.join(out_dir, f"{ticker}_Model_{month_str}.xlsx")
    wb.save(excel_path)
    print(f"[Excel] Saved: {excel_path}")

    # ══════════════════════════════════════════════════════════════════
    # 8. BIỂU ĐỒ
    # ══════════════════════════════════════════════════════════════════
    chart_dir = os.path.join(out_dir, "charts")
    os.makedirs(chart_dir, exist_ok=True)
    seg_hist_data = {"MoiGioi": brokerage_rev_hist, "Margin": margin_rev_hist, "TuDoanh": fvtpl_net_hist,
                      "IB_LuuKy": ib_custody_rev_hist, "QLQ": qlq_rev_hist}
    seg_fc_data = {"MoiGioi": brokerage_rev_fc, "Margin": margin_rev_fc, "TuDoanh": fvtpl_rev_fc,
                   "IB_LuuKy": ib_custody_rev_fc, "QLQ": qlq_rev_fc}
    chart_paths = {}
    chart_paths["segment_revenue"] = make_segment_revenue_chart(chart_dir, ticker, years_hist, years_fc, seg_hist_data, seg_fc_data)
    chart_paths["segment_mix"] = make_segment_mix_chart(chart_dir, ticker, years_hist, years_fc, seg_hist_data, seg_fc_data)
    chart_paths["pe_history"] = make_pe_pb_chart(chart_dir, ticker, quarter_labels, pe_quarters, PE_HIST_MEDIAN, "P/E", "#2980B9")
    chart_paths["pb_history"] = make_pe_pb_chart(chart_dir, ticker, quarter_labels, pb_quarters, PB_HIST_MEDIAN, "P/B", "#27AE60")
    chart_paths["margin_nim"] = make_margin_nim_chart(chart_dir, ticker, years_hist, years_fc, margin_loans_hist, margin_loans_fc,
                                                       margin_nim_hist, [margin_nim_fc_base]*3)
    chart_paths["fvtpl_sensitivity"] = make_fvtpl_sensitivity_chart(chart_dir, ticker, vnindex_sens_range, fvtpl_sens_npat_impact_pct)
    chart_paths["revenue_npat"] = make_revenue_npat_chart(chart_dir, ticker, years_hist, years_fc, total_rev_hist, total_rev_fc,
                                                           npat_parent_hist, npat_parent_fc)
    chart_paths["npat_roe_quarterly"] = make_npat_roe_quarterly_chart(chart_dir, ticker, is_q, bs_q)
    chart_paths["roe_pb_correlation"] = make_roe_pb_correlation_chart(chart_dir, ticker, quarter_labels, pb_quarters, is_q, bs_q)
    chart_paths["margin_leverage_quarterly"] = make_margin_leverage_quarterly_chart(chart_dir, ticker, bs_q)

    # ── Đòn bẩy Margin/VCSH quý gần nhất (dùng cho đánh giá trong PDF) ──
    _bs_q_sorted = sorted(bs_q, key=lambda x: (x.get("yearReport", 0), x.get("lengthReport", 0)))
    _latest_bs_q = _bs_q_sorted[-1] if _bs_q_sorted else {}
    _latest_margin_q = (_latest_bs_q.get(BS["margin_loans"]) or 0) / 1e9
    _latest_equity_q = (_latest_bs_q.get(BS["equity"]) or 0) / 1e9
    latest_margin_leverage = round(_latest_margin_q / _latest_equity_q, 2) if _latest_equity_q else 0
    latest_leverage_label = f"{_latest_bs_q.get('yearReport', '')}Q{_latest_bs_q.get('lengthReport', '')}"
    print(f"  -> Dư nợ Margin/VCSH quý gần nhất ({latest_leverage_label}): {latest_margin_leverage}x (trần pháp lý 2,0x)")

    # ── Đánh giá KQKD quý gần nhất & lũy kế tới hiện tại (dùng cho PDF) ──
    _is_q_sorted = sorted(is_q, key=lambda x: (x.get("yearReport", 0), x.get("lengthReport", 0)))
    _latest_is_q = _is_q_sorted[-1] if _is_q_sorted else {}
    _lq_year, _lq_num = _latest_is_q.get("yearReport"), _latest_is_q.get("lengthReport")
    _lq_rev = (_latest_is_q.get(IS_TOTAL["total_rev"]) or 0) / 1e9
    _lq_npat = (_latest_is_q.get(IS_TOTAL["npat_parent"]) or 0) / 1e9
    _yoy_q = next((x for x in is_q if x.get("yearReport") == (_lq_year - 1 if _lq_year else None)
                   and x.get("lengthReport") == _lq_num), None)
    _yoy_rev_pct = round((_lq_rev / ((_yoy_q.get(IS_TOTAL["total_rev"]) or 1) / 1e9) - 1) * 100, 1) if _yoy_q else None
    _yoy_npat_pct = round((_lq_npat / ((_yoy_q.get(IS_TOTAL["npat_parent"]) or 1) / 1e9) - 1) * 100, 1) if _yoy_q else None
    _pct_of_annual_est_rev = round(_rev_cum / total_rev_fc[0] * 100, 1) if total_rev_fc[0] and _n_rev_q > 0 else None
    _pct_of_annual_est_npat = round(_npat_cum / npat_fc[0] * 100, 1) if npat_fc[0] and _n_npat_q > 0 else None
    quarterly_update = {
        "quarter_label": f"{_lq_year}Q{_lq_num}" if _lq_year else "N/A",
        "rev": _lq_rev, "npat": _lq_npat,
        "yoy_rev_pct": _yoy_rev_pct, "yoy_npat_pct": _yoy_npat_pct,
        "n_known_q": _n_rev_q, "cum_rev": _rev_cum, "cum_npat": _npat_cum,
        "pct_of_annual_est_rev": _pct_of_annual_est_rev, "pct_of_annual_est_npat": _pct_of_annual_est_npat,
        "cur_fc_year": _cur_fc_year,
    }
    print(f"  -> KQKD {quarterly_update['quarter_label']}: DT {_lq_rev:,.0f} tỷ (YoY {_yoy_rev_pct}%), "
          f"LNST {_lq_npat:,.0f} tỷ (YoY {_yoy_npat_pct}%) — lũy kế {_n_rev_q}/4 quý")

    # ══════════════════════════════════════════════════════════════════
    # 9. PDF
    # ══════════════════════════════════════════════════════════════════
    pdf_path = os.path.join(out_dir, f"{ticker}_Phan_Tich_{month_str}.pdf")
    build_pdf_report(pdf_path, ticker, company_name, current_price, market_cap, shares, weighted_target, upside_pct,
                      recommend, pb_target_price, pe_target_price, bear_target, bull_target, PE_HIST_MEDIAN, PB_HIST_MEDIAN,
                      COE, years_hist, years_fc, total_rev_hist, total_rev_fc, npat_parent_hist, npat_parent_fc,
                      eps_hist, eps_fc, bvps_hist, bvps_fc, seg_pct_fc0, concentration_warning, concentration_flag,
                      chart_paths, fvtpl_sens_npat_impact_pct, vnindex_sens_range,
                      latest_margin_leverage, latest_leverage_label, VALUATION_WEIGHTS, quarterly_update)
    print(f"[PDF] Saved: {pdf_path}")

    # ══════════════════════════════════════════════════════════════════
    # 10. JSON EXPORT
    # ══════════════════════════════════════════════════════════════════
    _macro_liq = {
        "adtv_hist": {str(y): MARKET_ADTV_HIST[y] for y in years_hist},
        "adtv_fc": {str(years_fc[i]): round(market_adtv_fc[i], 0) for i in range(len(years_fc))},
        "adtv_growth_taper": ADTV_GROWTH_TAPER,
        "fvtpl_mix": fvtpl_mix_fc,
        "fvtpl_rates": fvtpl_r_fc,
        "fvtpl_expected_yield": round(fvtpl_expected_yield_fc, 4),
        "market_share": round(market_share_fc, 4),
        "brokerage_model": _brok_model,
        "brokerage_share_hist": {str(years_hist[i]): brokerage_share_hist[i] for i in range(len(years_hist))},
        "note": (
            "GTGD binh quan/phien HOSE (ty VND). Nguon: HOSE/SSC thong ke. "
            "adtv_hist: du lieu thuc te da doi chieu; adtv_fc: du bao tang truong giam dan. "
            "fvtpl_mix: ty trong danh muc tu doanh (CDs/TP/CP) theo tung CTCK — "
            "BAT BUOC cap nhat theo thuyet minh BCTC thuc te moi nhat. "
            "brokerage_model: phuong phap du phong mang moi gioi (Market share vs OLS fallback)."
        ),
    }
    save_json_summary(ticker, company_name, current_price, market_cap, shares, years_hist, years_fc,
                       total_rev_hist, total_rev_fc, npat_parent_hist, npat_parent_fc, eps_hist, eps_fc,
                       equity_hist, equity_fc, quarter_labels, pe_quarters, pb_quarters,
                       weighted_target, upside_pct, recommend, bear_target, bull_target,
                       pb_target_price, pe_target_price, PE_HIST_MEDIAN, PB_HIST_MEDIAN, COE,
                       seg_hist_data, seg_fc_data, seg_pct_fc0, concentration_warning, concentration_flag,
                       is_q, bs_q, latest_margin_leverage, latest_leverage_label, VALUATION_WEIGHTS, quarterly_update,
                       macro_liquidity=_macro_liq)

    print(f"\n--- Securities Analysis Complete for {ticker} ---")
    return True


# ══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════
def build_beta_coe_sheets(wb, ticker, beta_raw, beta_val, beta_src, aligned_data, rf_val, rf_src, erp, specific_rp, coe):
    ws_beta = wb.create_sheet("00_Beta")
    ws_beta.column_dimensions['A'].width = 15
    ws_beta.column_dimensions['B'].width = 16
    ws_beta.column_dimensions['C'].width = 22
    ws_beta.column_dimensions['D'].width = 16
    ws_beta.column_dimensions['E'].width = 22
    ws_beta.cell(row=1, column=1, value="BẢNG TÍNH HỆ SỐ BETA LỊCH SỬ").font = BOLD_FONT
    ws_beta.cell(row=1, column=2, value="Beta thô (raw):").font = BOLD_FONT
    ws_beta.cell(row=1, column=4, value="Beta Blume (đã điều chỉnh):").font = BOLD_FONT
    ws_beta.cell(row=2, column=2, value="Số phiên giao dịch:").font = ITALIC_FONT
    header_row(ws_beta, 4, ["Ngày", f"Giá {ticker}", f"Tỷ suất sinh lời {ticker}", "Giá VNINDEX", "Tỷ suất sinh lời VNINDEX"])
    if aligned_data:
        date0, p_s0, p_m0 = aligned_data[0]
        ws_beta.cell(row=5, column=1, value=date0)
        ws_beta.cell(row=5, column=2, value=p_s0)
        ws_beta.cell(row=5, column=4, value=p_m0)
        for ridx, (dstr, p_s, p_m) in enumerate(aligned_data[1:], start=6):
            ws_beta.cell(row=ridx, column=1, value=dstr)
            ws_beta.cell(row=ridx, column=2, value=p_s)
            ws_beta.cell(row=ridx, column=3, value=f"=(B{ridx}-B{ridx-1})/B{ridx-1}").number_format = '0.00%'
            ws_beta.cell(row=ridx, column=4, value=p_m)
            ws_beta.cell(row=ridx, column=5, value=f"=(D{ridx}-D{ridx-1})/D{ridx-1}").number_format = '0.00%'
        last_row = 4 + len(aligned_data)
        n_sessions = len(aligned_data)
        window_start = max(6, last_row - 499) if n_sessions > 500 else 6
        ws_beta.cell(row=1, column=3, value=f"=COVAR(C{window_start}:C{last_row},E{window_start}:E{last_row})/VAR(E{window_start}:E{last_row})").number_format = '0.0000'
        ws_beta.cell(row=1, column=5, value="=0.67*C1+0.33").number_format = '0.0000'
        ws_beta.cell(row=2, column=3, value=f"=COUNT(C{window_start}:C{last_row})")
    else:
        ws_beta.cell(row=1, column=3, value=beta_raw).number_format = '0.0000'
        ws_beta.cell(row=1, column=5, value=beta_val).number_format = '0.0000'
        ws_beta.cell(row=2, column=3, value=0)
    print(f"[Excel] Sheet 00_Beta done ({len(aligned_data)} phiên).")

    ws_coe = wb.create_sheet("00_COE")
    ws_coe.column_dimensions['A'].width = 42
    ws_coe.column_dimensions['B'].width = 16
    ws_coe.column_dimensions['C'].width = 50
    ws_coe.cell(row=1, column=1, value="CHI PHÍ VỐN CSH (COE) — MÔ HÌNH CAPM").font = TITLE_FONT
    header_row(ws_coe, 3, ["Tham số", "Giá trị", "Ghi chú / Nguồn"])
    ws_coe.cell(row=4, column=1, value="Rf — Lãi suất phi rủi ro (TPCP 10 năm)")
    ws_coe.cell(row=4, column=2, value=rf_val).number_format = FMT_PCT
    ws_coe.cell(row=4, column=3, value=rf_src)
    ws_coe.cell(row=5, column=1, value="β — Hệ số Beta (Blume-adjusted)")
    ws_coe.cell(row=5, column=2, value="='00_Beta'!E1").number_format = '0.0000'
    ws_coe.cell(row=5, column=3, value=beta_src)
    ws_coe.cell(row=6, column=1, value="ERP — Phần bù rủi ro vốn (Damodaran)")
    ws_coe.cell(row=6, column=2, value=erp).number_format = FMT_PCT
    ws_coe.cell(row=7, column=1, value="α — Phần bù rủi ro đặc thù (CTCK: đòn bẩy margin+tự doanh cao hơn TB)")
    ws_coe.cell(row=7, column=2, value=specific_rp).number_format = FMT_PCT
    ws_coe.cell(row=9, column=1, value="COE = Rf + β×ERP + α").font = BOLD_FONT
    ws_coe.cell(row=9, column=2, value="=B4+B5*B6+B7").font = BOLD_FONT
    ws_coe.cell(row=9, column=2).number_format = FMT_PCT
    ws_coe.cell(row=9, column=2).fill = P_FILL
    for r in range(3, 10):
        for c in range(1, 4):
            ws_coe.cell(row=r, column=c).border = THIN_BORDER
    print("[Excel] Sheet 00_COE done.")


def build_cover_sheet(wb, ticker, company_name, price, mcap, shares, target, upside, recommend, pb_target, pe_target,
                       pe_median, pb_median):
    ws = wb.create_sheet("01_Cover")
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 26
    ws.merge_cells('B2:C2')
    ws["B2"] = f"BÁO CÁO PHÂN TÍCH — {company_name} ({ticker})"
    ws["B2"].font = Font(bold=True, size=16, name=FONT_NAME, color="1F4E78")
    ws["B3"] = "Ngành: Dịch vụ tài chính — Chứng khoán (CTCK)"
    ws["B3"].font = Font(size=12, italic=True, name=FONT_NAME)
    ws["B4"] = f"Ngày lập báo cáo: {datetime.datetime.now().strftime('%d/%m/%Y')}"
    ws["B4"].font = ITALIC_FONT

    rows = [
        ("Giá hiện tại (VND)", price, FMT_PRICE),
        ("Vốn hóa thị trường (tỷ VND)", mcap, FMT_NUM),
        ("Số cổ phiếu lưu hành", shares, FMT_NUM),
        ("", None, None),
        ("Giá mục tiêu — P/B (VND)", pb_target, FMT_PRICE),
        ("Giá mục tiêu — P/E (VND)", pe_target, FMT_PRICE),
        ("P/E trung vị lịch sử", pe_median, FMT_MUL),
        ("P/B trung vị lịch sử", pb_median, FMT_MUL),
        ("", None, None),
        ("GIÁ MỤC TIÊU (bình quân trọng số)", target, FMT_PRICE),
        ("Tiềm năng tăng giá (upside)", upside / 100, FMT_PCT),
        ("KHUYẾN NGHỊ", recommend, None),
    ]
    r = 6
    for label, val, fmt in rows:
        ws.cell(row=r, column=2, value=label).font = BOLD_FONT if val is not None else DATA_FONT
        if val is not None:
            c = ws.cell(row=r, column=3, value=val)
            if fmt:
                c.number_format = fmt
            if label in ("GIÁ MỤC TIÊU (bình quân trọng số)", "KHUYẾN NGHỊ"):
                c.font = Font(bold=True, size=13, name=FONT_NAME, color="C00000" if recommend == "BÁN" else "006100")
                c.fill = P_FILL
        r += 1
    print("[Excel] Sheet 01_Cover done.")


def _rev_model_layout():
    """Vị trí dòng CỐ ĐỊNH trong sheet 03_Revenue_Model (không phụ thuộc ticker/số năm — chỉ phụ
    thuộc thứ tự các khối được viết trong build_revenue_model_sheet). Tính 1 lần, dùng chung để
    02_Assumptions có thể tham chiếu SỐNG tới 03_Revenue_Model (sheet đó được tạo SAU trong luồng
    chạy, nhưng công thức Excel dạng "='03_Revenue_Model'!C12" vẫn hợp lệ vì chỉ cần sheet tồn tại
    trước khi lưu file, không cần tồn tại tại thời điểm viết công thức). build_revenue_model_sheet()
    tự assert kết quả cuối cùng khớp với layout này để tránh 2 nơi tính toán lệch nhau khi sửa code."""
    r = 3
    L = {}
    L["margin_bal"] = r; r += 1
    L["margin_bal_avg"] = r; r += 2
    L["fvtpl_bal"] = r; r += 1
    L["fvtpl_bal_avg"] = r; r += 2
    L["qlq_aum"] = r; r += 2
    r += 1  # header "── DOANH THU THEO MẢNG ──"
    L["brokerage"] = r; r += 1
    L["margin"] = r; r += 1
    L["fvtpl"] = r; r += 1
    L["ib_sub"] = r; r += 1
    L["custody_sub"] = r; r += 1
    L["ib_custody"] = r; r += 1
    L["qlq"] = r; r += 1
    L["rev_blend_adj"] = r; r += 1
    L["total"] = r; r += 1
    return L


def _pnl_layout():
    """Vị trí dòng CỐ ĐỊNH trong sheet 04_PnL — cùng mục đích với _rev_model_layout() (xem docstring)."""
    r = 2
    L = {}
    for key in ("revenue", "cost", "gross_profit", "sga", "pbt", "tax", "npat_blend_adj",
                "npat", "npat_parent", "eps", "margin_gp", "margin_np"):
        L[key] = r
        r += 1
    return L


BLEND_SHEET_NAME = "04b_Dien_Bien_Quy"
BLEND_ROW = {"q1": 3, "q2": 4, "q3": 5, "q4": 6, "cum": 7, "n_known": 8, "n_remain": 9,
             "base_raw": 10, "blend_final": 11, "adj": 12}


def build_quarterly_blend_sheet(wb, ticker, is_q, years_fc, fc0_col, RR, RP):
    """Sheet minh họa TRỰC TIẾP công thức blend năm dự phóng hiện tại — thay số 'chết' Python-tính-sẵn
    bằng công thức SỐNG: Ước tính chuẩn = Lũy kế thực tế đã công bố + Base ước tính gốc/4 × Số quý còn
    lại (cùng công thức blend_annual_estimate() dùng cho HPG/MWG). 03_Revenue_Model và 04_PnL tham
    chiếu ngược lại dòng "Điều chỉnh" (BLEND_ROW['adj']) ở đây thay vì nhận 1 số Python dán chết."""
    ws = wb.create_sheet(BLEND_SHEET_NAME)
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 55

    cur_year = years_fc[0]
    ws.cell(row=1, column=1, value=f"LŨY KẾ & BLEND ƯỚC TÍNH NĂM {cur_year} (dùng số quý đã công bố thực tế)").font = TITLE_FONT
    ws.merge_cells('A1:D1')
    header_row(ws, 2, ["Chỉ tiêu", "Doanh thu hoạt động (tỷ VND)", "LNST cổ đông mẹ (tỷ VND)", "Ghi chú"])

    for qi, q in enumerate([1, 2, 3, 4]):
        rr = BLEND_ROW["q1"] + qi
        rec = next((x for x in is_q if x.get("yearReport") == cur_year and x.get("lengthReport") == q), None)
        rev_v = round(rec.get(IS_TOTAL["total_rev"]) / 1e9, 1) if rec and rec.get(IS_TOTAL["total_rev"]) is not None else None
        npat_v = round(rec.get(IS_TOTAL["npat_parent"]) / 1e9, 1) if rec and rec.get(IS_TOTAL["npat_parent"]) is not None else None
        ws.cell(row=rr, column=1, value=f"Q{q}/{cur_year} (thực tế đã công bố)").font = DATA_FONT
        if rev_v is not None:
            ws.cell(row=rr, column=2, value=rev_v).number_format = FMT_NUM
        if npat_v is not None:
            ws.cell(row=rr, column=3, value=npat_v).number_format = FMT_NUM
        for c in range(1, 4):
            ws.cell(row=rr, column=c).border = THIN_BORDER

    r = BLEND_ROW["cum"]
    ws.cell(row=r, column=1, value="Lũy kế thực tế đã công bố").font = BOLD_FONT
    ws.cell(row=r, column=2, value="=SUM(B3:B6)").number_format = FMT_NUM
    ws.cell(row=r, column=3, value="=SUM(C3:C6)").number_format = FMT_NUM

    r = BLEND_ROW["n_known"]
    ws.cell(row=r, column=1, value="Số quý đã có báo cáo thực tế (n)").font = BOLD_FONT
    ws.cell(row=r, column=2, value="=COUNT(B3:B6)").number_format = '0'
    ws.cell(row=r, column=3, value="=COUNT(C3:C6)").number_format = '0'

    r = BLEND_ROW["n_remain"]
    ws.cell(row=r, column=1, value="Số quý còn lại (4 − n)").font = BOLD_FONT
    ws.cell(row=r, column=2, value=f"=4-B{BLEND_ROW['n_known']}").number_format = '0'
    ws.cell(row=r, column=3, value=f"=4-C{BLEND_ROW['n_known']}").number_format = '0'

    r = BLEND_ROW["base_raw"]
    ws.cell(row=r, column=1, value="Base ước tính gốc cả năm (bottom-up thuần, CHƯA cộng lũy kế)").font = BOLD_FONT
    rev_base_formula = (f"='03_Revenue_Model'!{fc0_col}{RR['brokerage']}+'03_Revenue_Model'!{fc0_col}{RR['margin']}"
                         f"+'03_Revenue_Model'!{fc0_col}{RR['fvtpl']}+'03_Revenue_Model'!{fc0_col}{RR['ib_custody']}"
                         f"+'03_Revenue_Model'!{fc0_col}{RR['qlq']}")
    npat_base_formula = f"='04_PnL'!{fc0_col}{RP['pbt']}+'04_PnL'!{fc0_col}{RP['tax']}"
    ws.cell(row=r, column=2, value=rev_base_formula).number_format = FMT_NUM
    ws.cell(row=r, column=3, value=npat_base_formula).number_format = FMT_NUM

    r = BLEND_ROW["blend_final"]
    ws.cell(row=r, column=1, value="ƯỚC TÍNH CHUẨN (blend) = Lũy kế + Base/4 × Số quý còn lại").font = Font(bold=True, color="FFFFFF")
    for c in (2, 3):
        cl = get_column_letter(c)
        formula = f"={cl}{BLEND_ROW['cum']}+{cl}{BLEND_ROW['base_raw']}/4*{cl}{BLEND_ROW['n_remain']}"
        cell = ws.cell(row=r, column=c, value=formula)
        cell.number_format = FMT_NUM
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL

    r = BLEND_ROW["adj"]
    ws.cell(row=r, column=1, value="  ↳ Điều chỉnh áp dụng vào mô hình (= Ước tính chuẩn − Base gốc)").font = ITALIC_FONT
    for c in (2, 3):
        cl = get_column_letter(c)
        ws.cell(row=r, column=c, value=f"={cl}{BLEND_ROW['blend_final']}-{cl}{BLEND_ROW['base_raw']}").number_format = FMT_NUM

    ws.cell(row=13, column=1, value=(
        "Công thức: khi đã có n quý báo cáo thực tế trong năm, Ước tính chuẩn cả năm = Lũy kế thực tế n quý "
        "+ Base ước tính gốc cả năm × (4-n)/4 — khi n=0 (chưa có quý nào công bố), công thức tự động trả về "
        "đúng Base gốc (không điều chỉnh). Cùng công thức blend_annual_estimate() dùng thống nhất cho mọi "
        "ticker/ngành trong hệ thống (HPG, MWG...)."))
    ws.cell(row=13, column=1).font = ITALIC_FONT
    ws.cell(row=13, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells('A13:D13')
    ws.row_dimensions[13].height = 40

    # ── Diễn biến các quý gần nhất (tham khảo xu hướng — không ảnh hưởng công thức blend ở trên) ──
    r = 15
    ws.cell(row=r, column=1, value="── DIỄN BIẾN CÁC QUÝ GẦN NHẤT (tham khảo xu hướng) ──").font = Font(bold=True, italic=True, color="1F4E78")
    ws.merge_cells(f'A{r}:D{r}')
    r += 1
    header_row(ws, r, ["Quý", "Doanh thu hoạt động", "LNST cổ đông mẹ", "Tăng trưởng LNST YoY"])
    r += 1
    quarters = sorted(set((x.get("yearReport"), x.get("lengthReport")) for x in is_q
                           if x.get("yearReport") and x.get("lengthReport") in (1, 2, 3, 4)))
    recent = quarters[-8:]
    row_start = r
    for y, q in recent:
        rec = next((x for x in is_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
        rev_v = (rec.get(IS_TOTAL["total_rev"]) or 0) / 1e9
        npat_v = (rec.get(IS_TOTAL["npat_parent"]) or 0) / 1e9
        ws.cell(row=r, column=1, value=f"{y}Q{q}")
        ws.cell(row=r, column=2, value=round(rev_v, 1)).number_format = FMT_NUM
        ws.cell(row=r, column=3, value=round(npat_v, 1)).number_format = FMT_NUM
        if r - 4 >= row_start:
            ws.cell(row=r, column=4, value=f"=C{r}/C{r-4}-1").number_format = FMT_PCT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    print(f"[Excel] Sheet {BLEND_SHEET_NAME} done.")
    return {"rev_adj_cell": f"'{BLEND_SHEET_NAME}'!B{BLEND_ROW['adj']}",
            "npat_adj_cell": f"'{BLEND_SHEET_NAME}'!C{BLEND_ROW['adj']}"}


def build_assumptions_sheet(wb, ticker, years_hist, years_fc, price, shares, coe,
                             market_adtv_hist, trading_days, fee_bps, brokerage_share_hist, market_share_fc,
                             market_adtv_fc, margin_nim_hist, margin_nim_fc, margin_loan_growth_fc,
                             fvtpl_mix_fc, fvtpl_r_fc, fvtpl_yield_fc, fvtpl_portfolio_growth_fc,
                             ib_pipeline_fc, ib_fee_pct_fc, custody_growth_fc, qlq_fee_rate_fc, qlq_aum0,
                             qlq_aum_growth_fc, sga_pct_fc, cost_pct_fc, tax_rate_fc, pe_median, pb_median,
                             pe_pb_median_row):
    ws = wb.create_sheet("02_Assumptions")
    N_HIST, N_FC = len(years_hist), len(years_fc)
    all_years = years_hist + years_fc
    ws.column_dimensions['A'].width = 42
    for i in range(len(all_years)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    ws.column_dimensions[get_column_letter(2 + len(all_years))].width = 45

    header_row(ws, 1, ["Chỉ tiêu"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years] + ["Ghi chú"])
    RA = {}
    r = 2
    RML = _rev_model_layout()  # tham chiếu SỐNG sang 03_Revenue_Model (xem docstring _rev_model_layout)

    def col(i):
        return get_column_letter(2 + i)

    def arow(label, vals, fmt=None, note=None, single=False, fill=None):
        nonlocal r
        row_idx = r
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        if single:
            c = ws.cell(row=r, column=2, value=vals)
            c.number_format = fmt or FMT_NUM
            c.fill = fill or ASSUMP_FILL
            c.border = THIN_BORDER
        else:
            for i, v in enumerate(vals):
                c = ws.cell(row=r, column=2 + i, value=v)
                c.number_format = fmt or FMT_NUM
                c.border = THIN_BORDER
                if i >= N_HIST:
                    c.fill = fill or ASSUMP_FILL
        if note:
            nc = ws.cell(row=r, column=2 + len(all_years), value=note)
            nc.font = ITALIC_FONT
            nc.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
        return row_idx

    ws.cell(row=r, column=1, value="── GIÁ & VỐN HÓA ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    RA["price"] = arow("Giá cổ phiếu hiện tại (VND)", price, FMT_PRICE, "Giá đóng cửa gần nhất", single=True)
    RA["shares"] = arow("Số cổ phiếu lưu hành", shares, FMT_NUM, "Vốn điều lệ / 10,000", single=True)
    RA["coe"] = arow("Chi phí vốn CSH — COE (%)", "='00_COE'!B9", FMT_PCT, "Xem chi tiết CAPM tại sheet 00_COE/00_Beta", single=True)
    RA["pe_median"] = arow("P/E Trung vị lịch sử (x)", f"='13_PE_PB_History'!B{pe_pb_median_row}", FMT_MUL,
                            "Link sống tới MEDIAN sheet 13_PE_PB_History", single=True)
    RA["pb_median"] = arow("P/B Trung vị lịch sử (x)", f"='13_PE_PB_History'!C{pe_pb_median_row}", FMT_MUL,
                            "Link sống tới MEDIAN sheet 13_PE_PB_History", single=True)

    ws.cell(row=r, column=1, value="── (1) MÔI GIỚI: DT = GTGD/phiên × Số phiên × Thị phần × Phí(bps)/10,000 ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    RA["adtv"] = arow("GTGD toàn TT bình quân/phiên (tỷ VND)",
                       [market_adtv_hist[y] for y in years_hist] + market_adtv_fc, FMT_NUM,
                       "⚠ GIẢ ĐỊNH THỦ CÔNG, KHÔNG dựa trên dữ liệu vĩ mô thực (mặc định template chỉ là ước lượng xu hướng "
                       "chung). BẮT BUỘC đối chiếu với bảng 'Diễn biến GTGD thực tế' + khung đánh giá vĩ mô/thanh khoản tại "
                       "sheet 09_PESTLE trước khi chấp nhận — tốc độ tăng trưởng máy móc dễ gây bias nếu thanh khoản đang "
                       "hạ nhiệt/tăng nóng bất thường")
    RA["trading_days"] = arow("Số phiên giao dịch/năm", [trading_days] * len(all_years), '0', "~250 phiên/năm")
    RA["fee_bps"] = arow("Phí môi giới bình quân (bps)", [fee_bps] * len(all_years), '0.0" bps"',
                          "Đơn vị: bps (1 bps = 0,01%) — 15 bps = 0,15%/GTGD (1 chiều), mức phí phổ biến hiện nay "
                          "(0,15%-0,35% tùy CTCK, KHÔNG phải 15%)")

    _ms_row = r
    _ms_vals = []
    for i in range(N_HIST):
        cl = col(i)
        _ms_vals.append(f"='03_Revenue_Model'!{cl}{RML['brokerage']}/({cl}{RA['adtv']}*{cl}{RA['trading_days']}*{cl}{RA['fee_bps']}/10000)")
    for i in range(N_FC):
        _ms_vals.append(f"=AVERAGE({col(N_HIST-2)}{_ms_row}:{col(N_HIST-1)}{_ms_row})")
    RA["market_share"] = arow("Thị phần môi giới (%)", _ms_vals, FMT_PCT,
                               "Lịch sử = DT môi giới ('03_Revenue_Model') / (GTGD×Số phiên×Phí/10,000) — suy ngược SỐNG từ DT thực tế; "
                               "Dự phóng = TB 2 năm gần nhất (công thức AVERAGE sống)")

    ws.cell(row=r, column=1, value="── (2) CHO VAY MARGIN: DT = Dư nợ bình quân × NIM (annualized) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    _nim_row = r
    _nim_vals = []
    for i in range(N_HIST):
        cl = col(i)
        _nim_vals.append(f"='03_Revenue_Model'!{cl}{RML['margin']}/'03_Revenue_Model'!{cl}{RML['margin_bal_avg']}")
    for i in range(N_FC):
        _nim_vals.append(f"=AVERAGE({col(N_HIST-2)}{_nim_row}:{col(N_HIST-1)}{_nim_row})")
    RA["margin_nim"] = arow("NIM cho vay Margin (%/năm)", _nim_vals, FMT_PCT,
                             "Công thức: Lãi cho vay Margin ('03_Revenue_Model') / Dư nợ bình quân — công thức sống, đã net chi phí vốn "
                             "ngầm định qua chính biên độ quan sát lịch sử; Dự phóng = TB 2 năm gần nhất")
    RA["margin_growth"] = arow("Tăng trưởng dư nợ Margin (%/năm)", [None] * N_HIST + margin_loan_growth_fc, FMT_PCT,
                                "⚠ GIẢ ĐỊNH THỦ CÔNG, GIẢM DẦN theo năm (thận trọng cho năm xa, tránh cộng dồn tăng trưởng cao "
                                "liên tục — đặc biệt khi dư nợ margin có trần pháp lý 2,0x VCSH). Điều chỉnh theo kế hoạch tăng "
                                "vốn/room margin thực tế của CTCK nếu có công bố")

    ws.cell(row=r, column=1, value="── (3) TỰ DOANH (FVTPL): DT = Danh mục × (%CDs×R_CDs + %TP×R_TP + %CP×R_VNI) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    _fvtpl_manual_note = ("⚠ GIẢ ĐỊNH THỦ CÔNG — Vietcap KHÔNG công bố cơ cấu danh mục FVTPL chi tiết theo mã, nên KHÔNG thể "
                          "tính bằng công thức sống. Mặc định chỉ mang tính khởi điểm — BẮT BUỘC điều chỉnh theo thuyết minh "
                          "BCTC thực tế nếu có (một số CTCK gần đây tăng tỷ trọng trái phiếu/CDs, giảm cổ phiếu — xem báo cáo "
                          "công ty gần nhất). Đối chiếu với dòng 'Hiệu suất Tự doanh THỰC TẾ lịch sử' bên dưới để hiệu chỉnh.")
    RA["fvtpl_pct_cds"] = arow("Tỷ trọng CDs (Chứng chỉ tiền gửi) trong danh mục", [None] * N_HIST + [fvtpl_mix_fc["CDs"]] * N_FC, FMT_PCT,
                                _fvtpl_manual_note)
    RA["fvtpl_pct_tp"] = arow("Tỷ trọng TP (Trái phiếu) trong danh mục", [None] * N_HIST + [fvtpl_mix_fc["TP"]] * N_FC, FMT_PCT,
                               _fvtpl_manual_note)
    RA["fvtpl_pct_cp"] = arow("Tỷ trọng CP (Cổ phiếu) trong danh mục", [None] * N_HIST + [fvtpl_mix_fc["CP"]] * N_FC, FMT_PCT,
                               _fvtpl_manual_note)
    RA["fvtpl_r_cds"] = arow("R_CDs — Lãi suất kỳ vọng CDs (%/năm)", [None] * N_HIST + [fvtpl_r_fc["R_CDs"]] * N_FC, FMT_PCT,
                              "✓ Đã đối chiếu (2026-07): lãi suất CDs 12 tháng thực tế 5.5%-7.9% tùy NH (Vietcombank ~7.4-7.9%, "
                              "BVBank 7.2%) — 7.0% là mức trung bình-thận trọng. Cần cập nhật lại theo mặt bằng lãi suất tại "
                              "thời điểm phân tích")
    RA["fvtpl_r_tp"] = arow("R_TP — Lợi suất kỳ vọng Trái phiếu (%/năm)", [None] * N_HIST + [fvtpl_r_fc["R_TP"]] * N_FC, FMT_PCT,
                             "✓ Đã đối chiếu (2026-07): coupon trái phiếu NH bình quân phát hành 2026 ~6.4-6.5%, một số NH "
                             "phát hành 8-8.9% — 7.0% nằm trong dải hợp lý")
    RA["fvtpl_r_vni"] = arow("R_VNI — Biến động kỳ vọng VN-Index (%/năm)", [None] * N_HIST + [fvtpl_r_fc["R_VNI"]] * N_FC, FMT_PCT,
                              "⚠ KHÔNG kiểm chứng được — kỳ vọng lợi suất tương lai mang tính chủ quan, không có \"đúng/sai\". "
                              "Đối chiếu với COE (sheet 00_COE) hoặc quan điểm thị trường riêng khi điều chỉnh")
    RA["fvtpl_growth"] = arow("Tăng trưởng danh mục Tự doanh (%/năm)", [None] * N_HIST + fvtpl_portfolio_growth_fc, FMT_PCT,
                               "⚠ GIẢ ĐỊNH THỦ CÔNG, GIẢM DẦN theo năm (thận trọng cho năm xa) — điều chỉnh theo kế hoạch phân bổ "
                               "vốn tự doanh của CTCK (BCTN/ĐHĐCĐ) nếu có công bố")
    _fy_vals = [f"='03_Revenue_Model'!{col(i)}{RML['fvtpl']}/'03_Revenue_Model'!{col(i)}{RML['fvtpl_bal_avg']}" for i in range(N_HIST)]
    RA["fvtpl_yield_hist_ref"] = arow("  ↳ Hiệu suất Tự doanh THỰC TẾ lịch sử (%, tham chiếu)", _fy_vals, FMT_PCT,
                                       "Công thức: Lãi/lỗ FVTPL thực tế / Danh mục bình quân (công thức sống) — ĐỐI CHIẾU với giả định "
                                       "%CDs×R_CDs+%TP×R_TP+%CP×R_VNI ở trên; nếu lệch nhiều so với dải lịch sử này, cần xem lại "
                                       "R_VNI/cơ cấu danh mục giả định cho hợp lý hơn")

    ws.cell(row=r, column=1, value="── (4) IB + LƯU KÝ: DT = Pipeline IB × Fee% + AUM lưu ký × Fee lưu ký% ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    RA["ib_pipeline"] = arow("Pipeline IB (giá trị deal dự kiến/năm, tỷ VND)", [None] * N_HIST + [ib_pipeline_fc] * N_FC, FMT_NUM,
                              "Tư vấn niêm yết, M&A, phát hành vốn — ước tính từ DT IB lịch sử/phí BQ")
    RA["ib_fee_pct"] = arow("Phí IB bình quân trên giá trị deal (%)", [None] * N_HIST + [ib_fee_pct_fc] * N_FC, FMT_PCT,
                             "⚠ GIẢ ĐỊNH THỦ CÔNG — phí tư vấn/bảo lãnh phát hành bình quân, điều chỉnh theo cơ cấu deal thực tế (IPO/M&A/trái phiếu có phí khác nhau)")
    RA["custody_growth"] = arow("Tăng trưởng DT Lưu ký (%/năm)", [None] * N_HIST + [custody_growth_fc] * N_FC, FMT_PCT,
                                 "TB tăng trưởng lịch sử — điều chỉnh theo kế hoạch mở rộng khách hàng lưu ký của CTCK nếu có công bố")

    ws.cell(row=r, column=1, value="── (5) QUẢN LÝ QUỸ (QLQ): DT = AUM quản lý × Fee rate (~0.75%/năm) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    RA["qlq_aum0"] = arow("AUM quản lý quỹ đầu kỳ dự phóng (tỷ VND)", [None] * N_HIST + [qlq_aum0] + [None] * (N_FC - 1), FMT_NUM,
                           "Suy ngược từ DT QLQ lịch sử / Fee rate (nếu CTCK không tách riêng dòng QLQ, đây là ước tính)")
    RA["qlq_fee_rate"] = arow("Fee rate quản lý quỹ (%/năm)", [None] * N_HIST + [qlq_fee_rate_fc] * N_FC, FMT_PCT,
                               "⚠ GIẢ ĐỊNH THỦ CÔNG — mức phí quản lý quỹ phổ biến ~0,75%/năm, điều chỉnh theo biểu phí thực tế của công ty quản lý quỹ trực thuộc CTCK nếu có")
    RA["qlq_growth"] = arow("Tăng trưởng AUM quản lý quỹ (%/năm)", [None] * N_HIST + qlq_aum_growth_fc, FMT_PCT,
                             "⚠ GIẢ ĐỊNH THỦ CÔNG, GIẢM DẦN theo năm (thận trọng cho năm xa) — điều chỉnh theo kế hoạch huy động "
                             "AUM thực tế nếu CTCK có công bố")

    ws.cell(row=r, column=1, value="── CHI PHÍ & THUẾ ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    RA["cost_pct"] = arow("Chi phí hoạt động / DT hoạt động (%)", [None] * N_HIST + [cost_pct_fc] * N_FC, FMT_PCT,
                           "TB 2 năm gần nhất (chi phí môi giới, lưu ký, tự doanh...) — số derive từ lịch sử, có thể điều chỉnh nếu dự kiến thay đổi cơ cấu chi phí")
    RA["sga_pct"] = arow("CP Quản lý CTCK / DT hoạt động (%)", [None] * N_HIST + [sga_pct_fc] * N_FC, FMT_PCT,
                          "TB 2 năm gần nhất — số derive từ lịch sử, điều chỉnh nếu công ty có kế hoạch mở rộng/thu hẹp bộ máy")
    RA["tax_rate"] = arow("Thuế suất TNDN hiệu dụng (%)", [None] * N_HIST + [tax_rate_fc] * N_FC, FMT_PCT,
                           "TB thuế suất hiệu dụng lịch sử (giới hạn 15%-22%) — điều chỉnh nếu công ty có ưu đãi thuế đặc biệt")

    ws.column_dimensions['A'].width = 45
    print(f"[Excel] Sheet 02_Assumptions done ({r} dòng).")
    return RA


def build_revenue_model_sheet(wb, ticker, years_hist, years_fc, RA,
                               brokerage_rev_hist, margin_rev_hist, fvtpl_net_hist, ib_custody_rev_hist, qlq_rev_hist,
                               margin_loans_hist, fvtpl_portfolio_hist, qlq_aum0,
                               brokerage_rev_fc, margin_rev_fc, fvtpl_rev_fc, ib_custody_rev_fc, qlq_rev_fc,
                               margin_loans_fc, fvtpl_portfolio_fc, qlq_aum_fc, total_rev_hist, total_rev_fc,
                               ib_rev_hist=None, custody_rev_hist=None):
    ws = wb.create_sheet("03_Revenue_Model")
    N_HIST, N_FC = len(years_hist), len(years_fc)
    all_years = years_hist + years_fc
    N_ALL = len(all_years)
    ws.column_dimensions['A'].width = 42
    for i in range(N_ALL):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13

    header_row(ws, 1, ["Mảng kinh doanh"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years])
    ws.merge_cells(f'A1:A1')

    def col(i):
        return get_column_letter(2 + i)

    def a_col(i):
        return f"'02_Assumptions'!{get_column_letter(2 + i)}"

    RR = {}
    r = 3

    # ── Dư nợ Margin (cuối kỳ + bình quân) ──
    RR["margin_bal"] = r
    ws.cell(row=r, column=1, value="Dư nợ cho vay Margin (cuối kỳ)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(margin_loans_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        prev = f"{col(N_HIST + i - 1)}{r}" if (N_HIST + i - 1) >= 0 else None
        formula = f"={prev}*(1+{a_col(N_HIST+i)}{RA['margin_growth']})"
        ws.cell(row=r, column=2 + N_HIST + i, value=formula).number_format = FMT_NUM
    r += 1
    RR["margin_bal_avg"] = r
    ws.cell(row=r, column=1, value="  Dư nợ Margin bình quân").font = ITALIC_FONT
    for i in range(N_ALL):
        prev_col = col(i - 1) if i > 0 else col(i)
        ws.cell(row=r, column=2 + i, value=f"=AVERAGE({prev_col}{r-1}:{col(i)}{r-1})").number_format = FMT_NUM
    r += 2

    # ── Danh mục Tự doanh (cuối kỳ + bình quân) ──
    RR["fvtpl_bal"] = r
    ws.cell(row=r, column=1, value="Danh mục Tự doanh - FVTPL (cuối kỳ)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(fvtpl_portfolio_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        prev = f"{col(N_HIST + i - 1)}{r}"
        formula = f"={prev}*(1+{a_col(N_HIST+i)}{RA['fvtpl_growth']})"
        ws.cell(row=r, column=2 + N_HIST + i, value=formula).number_format = FMT_NUM
    r += 1
    RR["fvtpl_bal_avg"] = r
    ws.cell(row=r, column=1, value="  Danh mục Tự doanh bình quân").font = ITALIC_FONT
    for i in range(N_ALL):
        prev_col = col(i - 1) if i > 0 else col(i)
        ws.cell(row=r, column=2 + i, value=f"=AVERAGE({prev_col}{r-1}:{col(i)}{r-1})").number_format = FMT_NUM
    r += 2

    # ── AUM Quản lý quỹ ──
    RR["qlq_aum"] = r
    ws.cell(row=r, column=1, value="AUM Quản lý quỹ (cuối kỳ)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=None)
    ws.cell(row=r, column=2 + N_HIST, value=f"={a_col(N_HIST)}{RA['qlq_aum0']}").number_format = FMT_NUM
    for i in range(1, N_FC):
        prev = f"{col(N_HIST + i - 1)}{r}"
        ws.cell(row=r, column=2 + N_HIST + i, value=f"={prev}*(1+{a_col(N_HIST+i)}{RA['qlq_growth']})").number_format = FMT_NUM
    r += 2

    ws.cell(row=r, column=1, value="── DOANH THU THEO MẢNG (tỷ VND) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1

    # (1) Môi giới
    RR["brokerage"] = r
    ws.cell(row=r, column=1, value="(1) DT Môi giới").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(brokerage_rev_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        cl = a_col(N_HIST + i)
        formula = f"={cl}{RA['adtv']}*{cl}{RA['trading_days']}*{cl}{RA['market_share']}*{cl}{RA['fee_bps']}/10000"
        ws.cell(row=r, column=cc, value=formula).number_format = FMT_NUM
    r += 1

    # (2) Margin
    RR["margin"] = r
    ws.cell(row=r, column=1, value="(2) DT Cho vay Margin").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(margin_rev_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        cl = a_col(N_HIST + i)
        formula = f"={col(N_HIST+i)}{RR['margin_bal_avg']}*{cl}{RA['margin_nim']}"
        ws.cell(row=r, column=cc, value=formula).number_format = FMT_NUM
    r += 1

    # (3) Tự doanh
    RR["fvtpl"] = r
    ws.cell(row=r, column=1, value="(3) DT Tự doanh (FVTPL)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(fvtpl_net_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        cl = a_col(N_HIST + i)
        formula = (f"={col(N_HIST+i)}{RR['fvtpl_bal_avg']}*"
                   f"({cl}{RA['fvtpl_pct_cds']}*{cl}{RA['fvtpl_r_cds']}+"
                   f"{cl}{RA['fvtpl_pct_tp']}*{cl}{RA['fvtpl_r_tp']}+"
                   f"{cl}{RA['fvtpl_pct_cp']}*{cl}{RA['fvtpl_r_vni']})")
        ws.cell(row=r, column=cc, value=formula).number_format = FMT_NUM
    r += 1

    # (4) IB + Lưu ký — tách 2 dòng con để công thức Excel khớp CHÍNH XÁC với Python mirror (không còn hệ
    # số *0.15 tùy tiện trộn lẫn IB/Lưu ký của năm trước như bản cũ)
    RR["ib_sub"] = r
    ws.cell(row=r, column=1, value="  IB (bảo lãnh/tư vấn)").font = ITALIC_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(ib_rev_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        cl = a_col(N_HIST + i)
        formula = f"={cl}{RA['ib_pipeline']}*{cl}{RA['ib_fee_pct']}"
        ws.cell(row=r, column=cc, value=formula).number_format = FMT_NUM
    r += 1

    RR["custody_sub"] = r
    ws.cell(row=r, column=1, value="  Lưu ký").font = ITALIC_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(custody_rev_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        cl = a_col(N_HIST + i)
        prev = f"{col(N_HIST+i-1)}{r}"
        formula = f"={prev}*(1+{cl}{RA['custody_growth']})"
        ws.cell(row=r, column=cc, value=formula).number_format = FMT_NUM
    r += 1

    RR["ib_custody"] = r
    ws.cell(row=r, column=1, value="(4) DT IB + Lưu ký").font = BOLD_FONT
    for i in range(N_ALL):
        cl = col(i)
        ws.cell(row=r, column=2 + i, value=f"={cl}{RR['ib_sub']}+{cl}{RR['custody_sub']}").number_format = FMT_NUM
    r += 1

    # (5) QLQ
    RR["qlq"] = r
    ws.cell(row=r, column=1, value="(5) DT Quản lý quỹ (QLQ)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(qlq_rev_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        cl = a_col(N_HIST + i)
        formula = f"={col(N_HIST+i)}{RR['qlq_aum']}*{cl}{RA['qlq_fee_rate']}"
        ws.cell(row=r, column=cc, value=formula).number_format = FMT_NUM
    r += 1

    # ── Điều chỉnh theo KQKD thực tế đã công bố — link SỐNG tới sheet 04b_Dien_Bien_Quy (năm dự
    # phóng đầu tiên), tự động = 0 khi năm đó chưa có quý nào công bố (xem build_quarterly_blend_sheet) ──
    RR["rev_blend_adj"] = r
    ws.cell(row=r, column=1, value="  Điều chỉnh theo KQKD thực tế đã công bố").font = ITALIC_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=None)
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        val = f"='{BLEND_SHEET_NAME}'!B{BLEND_ROW['adj']}" if i == 0 else 0
        c = ws.cell(row=r, column=cc, value=val)
        c.number_format = FMT_NUM
    r += 1

    # ── TỔNG DOANH THU ──
    RR["total"] = r
    ws.cell(row=r, column=1, value="TỔNG DOANH THU HOẠT ĐỘNG").font = Font(bold=True, color="FFFFFF")
    for i in range(N_ALL):
        cl = col(i)
        formula = f"={cl}{RR['brokerage']}+{cl}{RR['margin']}+{cl}{RR['fvtpl']}+{cl}{RR['ib_custody']}+{cl}{RR['qlq']}+{cl}{RR['rev_blend_adj']}"
        c = ws.cell(row=r, column=2 + i, value=formula)
        c.number_format = FMT_NUM
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    r += 1

    ws.cell(row=r, column=1, value="── TỶ TRỌNG ĐÓNG GÓP THEO MẢNG (%) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    RR["pct_brokerage"] = r
    for seg_name, seg_row in [("Môi giới", RR["brokerage"]), ("Cho vay Margin", RR["margin"]),
                               ("Tự doanh (FVTPL)", RR["fvtpl"]), ("IB + Lưu ký", RR["ib_custody"]),
                               ("Quản lý quỹ", RR["qlq"])]:
        ws.cell(row=r, column=1, value=f"  % {seg_name} / Tổng DT").font = DATA_FONT
        for i in range(N_ALL):
            cl = col(i)
            ws.cell(row=r, column=2 + i, value=f"={cl}{seg_row}/{cl}{RR['total']}").number_format = FMT_PCT
        r += 1

    _layout = _rev_model_layout()
    _mismatch = {k: (RR[k], _layout[k]) for k in _layout if RR.get(k) != _layout[k]}
    assert not _mismatch, f"03_Revenue_Model row layout lệch với _rev_model_layout(): {_mismatch}"

    print(f"[Excel] Sheet 03_Revenue_Model done ({r} dòng).")
    return RR


def build_pnl_sheet(wb, ticker, years_hist, years_fc, RR, RA,
                     total_rev_hist, total_cost_hist, gross_profit_hist, sga_hist, pbt_hist, tax_hist,
                     npat_hist, npat_parent_hist, eps_hist,
                     total_rev_fc, total_cost_fc, gross_profit_fc, sga_fc, pbt_fc, tax_fc, npat_fc, npat_parent_fc, eps_fc):
    ws = wb.create_sheet("04_PnL")
    N_HIST, N_FC = len(years_hist), len(years_fc)
    all_years = years_hist + years_fc
    N_ALL = len(all_years)
    ws.column_dimensions['A'].width = 38
    for i in range(N_ALL):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    header_row(ws, 1, ["Chỉ tiêu (tỷ VND)"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years])

    def col(i):
        return get_column_letter(2 + i)

    RP = {}
    r = 2
    RP["revenue"] = r
    ws.cell(row=r, column=1, value="Doanh thu hoạt động").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(total_rev_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        ws.cell(row=r, column=2 + N_HIST + i, value=f"='03_Revenue_Model'!{col(N_HIST+i)}{RR['total']}").number_format = FMT_NUM
    r += 1

    RP["cost"] = r
    ws.cell(row=r, column=1, value="Chi phí hoạt động").font = DATA_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(total_cost_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cl_a = f"'02_Assumptions'!{col(N_HIST+i)}"
        ws.cell(row=r, column=2 + N_HIST + i, value=f"=-{col(N_HIST+i)}{RP['revenue']}*{cl_a}{RA['cost_pct']}").number_format = FMT_NUM
    r += 1

    RP["gross_profit"] = r
    ws.cell(row=r, column=1, value="LỢI NHUẬN GỘP").font = BOLD_FONT
    for i in range(N_ALL):
        ws.cell(row=r, column=2 + i, value=f"={col(i)}{RP['revenue']}+{col(i)}{RP['cost']}").number_format = FMT_NUM
        ws.cell(row=r, column=2 + i).font = BOLD_FONT
    r += 1

    RP["sga"] = r
    ws.cell(row=r, column=1, value="Chi phí quản lý CTCK").font = DATA_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(sga_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cl_a = f"'02_Assumptions'!{col(N_HIST+i)}"
        ws.cell(row=r, column=2 + N_HIST + i, value=f"=-{col(N_HIST+i)}{RP['revenue']}*{cl_a}{RA['sga_pct']}").number_format = FMT_NUM
    r += 1

    RP["pbt"] = r
    ws.cell(row=r, column=1, value="LỢI NHUẬN TRƯỚC THUẾ (LNTT)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(pbt_hist[i], 1)).number_format = FMT_NUM
        ws.cell(row=r, column=2 + i).font = BOLD_FONT
    for i in range(N_FC):
        ws.cell(row=r, column=2 + N_HIST + i, value=f"={col(N_HIST+i)}{RP['gross_profit']}+{col(N_HIST+i)}{RP['sga']}").number_format = FMT_NUM
        ws.cell(row=r, column=2 + N_HIST + i).font = BOLD_FONT
    r += 1

    RP["tax"] = r
    ws.cell(row=r, column=1, value="Thuế TNDN").font = DATA_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(tax_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        cl_a = f"'02_Assumptions'!{col(N_HIST+i)}"
        ws.cell(row=r, column=2 + N_HIST + i, value=f"=-MAX({col(N_HIST+i)}{RP['pbt']},0)*{cl_a}{RA['tax_rate']}").number_format = FMT_NUM
    r += 1

    # Link SỐNG tới sheet 04b_Dien_Bien_Quy — tự động = 0 khi năm đó chưa có quý nào công bố
    RP["npat_blend_adj"] = r
    ws.cell(row=r, column=1, value="  Điều chỉnh theo KQKD thực tế đã công bố").font = ITALIC_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=None)
    for i in range(N_FC):
        val = f"='{BLEND_SHEET_NAME}'!C{BLEND_ROW['adj']}" if i == 0 else 0
        ws.cell(row=r, column=2 + N_HIST + i, value=val).number_format = FMT_NUM
    r += 1

    RP["npat"] = r
    ws.cell(row=r, column=1, value="LỢI NHUẬN SAU THUẾ (LNST)").font = Font(bold=True, color="FFFFFF")
    for i in range(N_ALL):
        c = ws.cell(row=r, column=2 + i, value=f"={col(i)}{RP['pbt']}+{col(i)}{RP['tax']}+{col(i)}{RP['npat_blend_adj']}")
        c.number_format = FMT_NUM
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    r += 1

    RP["npat_parent"] = r
    ws.cell(row=r, column=1, value="LNST thuộc về cổ đông công ty mẹ").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(npat_parent_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        ws.cell(row=r, column=2 + N_HIST + i, value=f"={col(N_HIST+i)}{RP['npat']}").number_format = FMT_NUM
    r += 1

    RP["eps"] = r
    ws.cell(row=r, column=1, value="EPS (VND/CP)").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=eps_hist[i]).number_format = FMT_PRICE
    for i in range(N_FC):
        cl_a = f"'02_Assumptions'!{col(0)}"
        ws.cell(row=r, column=2 + N_HIST + i, value=f"={col(N_HIST+i)}{RP['npat_parent']}*1e9/{cl_a}{RA['shares']}").number_format = FMT_PRICE
    r += 1

    RP["margin_gp"] = r
    ws.cell(row=r, column=1, value="Biên LN gộp (%)").font = ITALIC_FONT
    for i in range(N_ALL):
        ws.cell(row=r, column=2 + i, value=f"={col(i)}{RP['gross_profit']}/{col(i)}{RP['revenue']}").number_format = FMT_PCT
    r += 1
    RP["margin_np"] = r
    ws.cell(row=r, column=1, value="Biên LNST (%)").font = ITALIC_FONT
    for i in range(N_ALL):
        ws.cell(row=r, column=2 + i, value=f"={col(i)}{RP['npat']}/{col(i)}{RP['revenue']}").number_format = FMT_PCT
    r += 1

    _layout = _pnl_layout()
    _mismatch = {k: (RP[k], _layout[k]) for k in _layout if RP.get(k) != _layout[k]}
    assert not _mismatch, f"04_PnL row layout lệch với _pnl_layout(): {_mismatch}"

    print(f"[Excel] Sheet 04_PnL done ({r} dòng).")
    return RP


def build_balance_sheet(wb, ticker, years_hist, years_fc, RP, fvtpl_portfolio_hist, margin_loans_hist,
                         total_assets_hist, total_liab_hist, borrow_total_hist, equity_hist,
                         fvtpl_portfolio_fc, margin_loans_fc, equity_fc):
    ws = wb.create_sheet("05_Balance_Sheet")
    N_HIST, N_FC = len(years_hist), len(years_fc)
    all_years = years_hist + years_fc
    N_ALL = len(all_years)
    ws.column_dimensions['A'].width = 38
    for i in range(N_ALL):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    header_row(ws, 1, ["Chỉ tiêu (tỷ VND)"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years])

    def col(i):
        return get_column_letter(2 + i)

    r = 2
    # Tài sản khác = phần dư để TS luôn = NV (không dự phóng chi tiết từng khoản mục phụ)
    other_assets_hist = [total_assets_hist[i] - fvtpl_portfolio_hist[i] - margin_loans_hist[i] for i in range(N_HIST)]
    liab_ex_borrow_hist = [total_liab_hist[i] - borrow_total_hist[i] for i in range(N_HIST)]
    _oa_pct_hist = [other_assets_hist[i] / total_assets_hist[i] if total_assets_hist[i] else 0 for i in range(N_HIST)]
    oa_pct_fc = stats.mean(_oa_pct_hist[-2:]) if len(_oa_pct_hist) >= 2 else 0.1
    _lex_pct_hist = [liab_ex_borrow_hist[i] / total_liab_hist[i] if total_liab_hist[i] else 0 for i in range(N_HIST)]
    total_assets_fc = [fvtpl_portfolio_fc[i] + margin_loans_fc[i] for i in range(N_FC)]
    total_assets_fc = [ta / (1 - oa_pct_fc) for ta in total_assets_fc]
    other_assets_fc = [total_assets_fc[i] - fvtpl_portfolio_fc[i] - margin_loans_fc[i] for i in range(N_FC)]
    borrow_fc = [total_assets_fc[i] - equity_fc[i] - (total_liab_hist[-1] - borrow_total_hist[-1]) * (1.05 ** (i+1)) for i in range(N_FC)]
    liab_ex_borrow_fc = [(total_liab_hist[-1] - borrow_total_hist[-1]) * (1.05 ** (i+1)) for i in range(N_FC)]
    total_liab_fc = [borrow_fc[i] + liab_ex_borrow_fc[i] for i in range(N_FC)]

    rows_asset = [
        ("Tài sản tài chính FVTPL (Tự doanh)", fvtpl_portfolio_hist, fvtpl_portfolio_fc),
        ("Các khoản cho vay (Margin)", margin_loans_hist, margin_loans_fc),
        ("Tài sản khác", other_assets_hist, other_assets_fc),
    ]
    row_refs = {}
    for label, hist, fc in rows_asset:
        row_refs[label] = r
        ws.cell(row=r, column=1, value=label).font = DATA_FONT
        for i in range(N_HIST):
            ws.cell(row=r, column=2 + i, value=round(hist[i], 1)).number_format = FMT_NUM
        for i in range(N_FC):
            ws.cell(row=r, column=2 + N_HIST + i, value=round(fc[i], 1)).number_format = FMT_NUM
        r += 1

    r_total_asset = r
    ws.cell(row=r, column=1, value="TỔNG CỘNG TÀI SẢN").font = Font(bold=True, color="FFFFFF")
    for i in range(N_ALL):
        start_r = row_refs[rows_asset[0][0]]
        end_r = row_refs[rows_asset[-1][0]]
        c = ws.cell(row=r, column=2 + i, value=f"=SUM({col(i)}{start_r}:{col(i)}{end_r})")
        c.number_format = FMT_NUM
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    r += 2

    rows_liab = [
        ("Vay và nợ thuê tài sản tài chính", borrow_total_hist, borrow_fc),
        ("Nợ phải trả khác", liab_ex_borrow_hist, liab_ex_borrow_fc),
    ]
    row_refs2 = {}
    for label, hist, fc in rows_liab:
        row_refs2[label] = r
        ws.cell(row=r, column=1, value=label).font = DATA_FONT
        for i in range(N_HIST):
            ws.cell(row=r, column=2 + i, value=round(hist[i], 1)).number_format = FMT_NUM
        for i in range(N_FC):
            ws.cell(row=r, column=2 + N_HIST + i, value=round(fc[i], 1)).number_format = FMT_NUM
        r += 1
    r_total_liab = r
    ws.cell(row=r, column=1, value="NỢ PHẢI TRẢ").font = BOLD_FONT
    for i in range(N_ALL):
        start_r = row_refs2[rows_liab[0][0]]
        end_r = row_refs2[rows_liab[-1][0]]
        ws.cell(row=r, column=2 + i, value=f"=SUM({col(i)}{start_r}:{col(i)}{end_r})").number_format = FMT_NUM
        ws.cell(row=r, column=2 + i).font = BOLD_FONT
    r += 1

    r_equity = r
    ws.cell(row=r, column=1, value="VỐN CHỦ SỞ HỮU").font = BOLD_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=round(equity_hist[i], 1)).number_format = FMT_NUM
    for i in range(N_FC):
        ws.cell(row=r, column=2 + N_HIST + i, value=round(equity_fc[i], 1)).number_format = FMT_NUM
        ws.cell(row=r, column=2 + N_HIST + i).font = BOLD_FONT
    r += 1

    r += 1
    ws.cell(row=r, column=1, value="TỔNG CỘNG NGUỒN VỐN").font = Font(bold=True, color="FFFFFF")
    for i in range(N_ALL):
        c = ws.cell(row=r, column=2 + i, value=f"={col(i)}{r_total_liab}+{col(i)}{r_equity}")
        c.number_format = FMT_NUM
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL

    # Force forecast total assets to reconcile exactly with total nguồn vốn by adjusting "Tài sản khác"
    # (đảm bảo TS luôn = NV tuyệt đối, tránh lệch số học nhỏ do làm tròn các bước dự phóng riêng lẻ)
    for i in range(N_FC):
        cc = 2 + N_HIST + i
        target_formula = f"={col(N_HIST+i)}{r_total_liab}+{col(N_HIST+i)}{r_equity}-{col(N_HIST+i)}{row_refs[rows_asset[0][0]]}-{col(N_HIST+i)}{row_refs[rows_asset[1][0]]}"
        ws.cell(row=row_refs[rows_asset[2][0]], column=cc, value=target_formula).number_format = FMT_NUM

    print(f"[Excel] Sheet 05_Balance_Sheet done ({r} dòng).")
    return {"total_assets": r_total_asset, "total_liab": r_total_liab, "equity": r_equity}


def build_cash_flow_sheet(wb, ticker, years_hist, years_fc, RP, RBS, npat_hist, npat_fc):
    ws = wb.create_sheet("06_Cash_Flow")
    N_HIST, N_FC = len(years_hist), len(years_fc)
    all_years = years_hist + years_fc
    ws.column_dimensions['A'].width = 38
    for i in range(len(all_years)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    header_row(ws, 1, ["Chỉ tiêu (tỷ VND)"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years])

    def col(i):
        return get_column_letter(2 + i)

    r = 2
    ws.cell(row=r, column=1, value="LNST").font = DATA_FONT
    for i in range(len(all_years)):
        ws.cell(row=r, column=2 + i, value=f"='04_PnL'!{col(i)}{RP['npat']}").number_format = FMT_NUM
    r += 1
    ws.cell(row=r, column=1, value="CFO ước tính (≈ LNST, CTCK ít tài sản cố định)").font = DATA_FONT
    for i in range(len(all_years)):
        ws.cell(row=r, column=2 + i, value=f"={col(i)}{r-1}").number_format = FMT_NUM
    print(f"[Excel] Sheet 06_Cash_Flow done ({r} dòng).")


def build_valuation_sheet(wb, ticker, RA, RP, RBS, N_HIST, pe_median, pb_median, price, pb_target, pe_target,
                           weighted_target, upside_pct, weights, bear_target, bull_target, coe,
                           dividend_yield_pct=0.0, total_return_pct=0.0):
    ws = wb.create_sheet("07_Valuation")
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    ws.cell(row=1, column=1, value=f"ĐỊNH GIÁ {ticker} — P/B ƯU TIÊN + P/E (skill chung-khoan)").font = TITLE_FONT
    header_row(ws, 3, ["Phương pháp", "Giá trị", "Ghi chú"])

    fc0_col = get_column_letter(2 + N_HIST)  # cột năm dự phóng đầu tiên

    r = 4
    ws.cell(row=r, column=1, value="Chi phí vốn CSH (COE)")
    ws.cell(row=r, column=2, value=f"='02_Assumptions'!B{RA['coe']}").number_format = FMT_PCT
    r += 2

    ws.cell(row=r, column=1, value="1) PHƯƠNG PHÁP P/B (ưu tiên số 1)").font = Font(bold=True, color="1F4E78")
    r += 1
    r_pb_median = r
    ws.cell(row=r, column=1, value="P/B trung vị lịch sử (x)")
    ws.cell(row=r, column=2, value=f"='02_Assumptions'!B{RA['pb_median']}").number_format = FMT_MUL
    r += 1
    r_bvps = r
    ws.cell(row=r, column=1, value="BVPS dự phóng năm 1 (VND/CP)")
    ws.cell(row=r, column=2, value=f"='05_Balance_Sheet'!{fc0_col}{RBS['equity']}*1e9/'02_Assumptions'!$B${RA['shares']}").number_format = FMT_PRICE
    r_pb_target = r + 1
    ws.cell(row=r + 1, column=1, value="→ P/B Target Price (VND)").font = BOLD_FONT
    c = ws.cell(row=r + 1, column=2, value=f"=B{r_pb_median}*B{r_bvps}")
    c.number_format = FMT_PRICE
    c.font = BOLD_FONT
    c.fill = P_FILL
    r += 3

    ws.cell(row=r, column=1, value="2) PHƯƠNG PHÁP P/E").font = Font(bold=True, color="1F4E78")
    r += 1
    r_pe_median = r
    ws.cell(row=r, column=1, value="P/E trung vị lịch sử (x)")
    ws.cell(row=r, column=2, value=f"='02_Assumptions'!B{RA['pe_median']}").number_format = FMT_MUL
    r += 1
    r_eps_fc = r
    ws.cell(row=r, column=1, value="EPS dự phóng năm 1 (VND/CP)")
    ws.cell(row=r, column=2, value=f"='04_PnL'!{fc0_col}{RP['eps']}").number_format = FMT_PRICE
    r += 1
    ws.cell(row=r, column=1, value="→ P/E Target Price (VND)").font = BOLD_FONT
    c = ws.cell(row=r, column=2, value=f"=B{r_pe_median}*B{r_eps_fc}")
    c.number_format = FMT_PRICE
    c.font = BOLD_FONT
    c.fill = P_FILL
    r += 2

    ws.cell(row=r, column=1, value=f"3) GIÁ MỤC TIÊU (P/B {int(weights['PB']*100)}% + P/E {int(weights['PE']*100)}%)").font = Font(bold=True, color="1F4E78")
    r += 1
    r_target = r
    ws.cell(row=r, column=1, value="GIÁ MỤC TIÊU (VND)").font = Font(bold=True, size=12)
    c = ws.cell(row=r, column=2, value=weighted_target)
    c.number_format = FMT_PRICE
    c.font = Font(bold=True, size=12, color="006100")
    c.fill = P_FILL
    r += 1
    ws.cell(row=r, column=1, value="Giá hiện tại (VND)")
    ws.cell(row=r, column=2, value=price).number_format = FMT_PRICE
    r += 1
    ws.cell(row=r, column=1, value="Upside/Downside (%)").font = BOLD_FONT
    c = ws.cell(row=r, column=2, value=upside_pct / 100)
    c.number_format = FMT_PCT
    c.font = BOLD_FONT
    r += 1
    ws.cell(row=r, column=1, value="Tỷ suất cổ tức dự kiến (%)")
    ws.cell(row=r, column=2, value=dividend_yield_pct / 100).number_format = FMT_PCT
    r += 1
    ws.cell(row=r, column=1, value="TỔNG TỶ SUẤT SINH LỜI (Upside + Cổ tức)").font = Font(bold=True, color="006100")
    c = ws.cell(row=r, column=2, value=total_return_pct / 100)
    c.number_format = FMT_PCT
    c.font = Font(bold=True, color="006100")
    r += 2

    ws.cell(row=r, column=1, value="Bear Case (VND)")
    ws.cell(row=r, column=2, value=bear_target).number_format = FMT_PRICE
    r += 1
    ws.cell(row=r, column=1, value="Bull Case (VND)")
    ws.cell(row=r, column=2, value=bull_target).number_format = FMT_PRICE
    r += 2

    ws.cell(row=r, column=1, value=("Lưu ý phương pháp: P/B là phương pháp CHÍNH (90%) cho CTCK vì tài sản có tính thanh khoản cao "
                                     "(tiền, chứng khoán FVTPL, cho vay margin) — phản ánh sát giá trị sổ sách thực. P/E giữ trọng số "
                                     "nhỏ CỐ ĐỊNH 10% cho MỌI CTCK — không bỏ hẳn (vẫn cần làm bộ lọc chống nhiễu khi P/B bị lệch bất "
                                     "thường) nhưng cũng không để cao (LNST CTCK dễ biến động do Tự doanh/IB, kém tin cậy hơn BVPS "
                                     "khi định giá tài sản CTCK)."))
    ws.cell(row=r, column=1).font = ITALIC_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(f'A{r}:C{r}')
    ws.row_dimensions[r].height = 45

    for rr in range(3, r + 1):
        for cc in range(1, 4):
            ws.cell(row=rr, column=cc).border = THIN_BORDER
    print(f"[Excel] Sheet 07_Valuation done.")
    return {"target": r_target, "pb_target": r_pb_target}


def build_sensitivity_sheet(wb, ticker, RV, RP, vnindex_range, fvtpl_sens_impact, fvtpl_sens_npat_pct, npat_fc, tax_rate):
    ws = wb.create_sheet("08_Sensitivity")
    ws.column_dimensions['A'].width = 45
    for i in range(len(vnindex_range)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14
    ws.cell(row=1, column=1, value="ĐỘ NHẠY DANH MỤC TỰ DOANH THEO BIẾN ĐỘNG VN-INDEX").font = TITLE_FONT
    header_row(ws, 3, ["Biến động VN-Index"] + [f"{s*100:+.0f}%" for s in vnindex_range])
    ws.cell(row=4, column=1, value="Tác động DT Tự doanh (tỷ VND)").font = BOLD_FONT
    for i, v in enumerate(fvtpl_sens_impact):
        ws.cell(row=4, column=2 + i, value=v).number_format = FMT_NUM
    ws.cell(row=5, column=1, value="Tác động LNST (%)").font = BOLD_FONT
    for i, v in enumerate(fvtpl_sens_npat_pct):
        ws.cell(row=5, column=2 + i, value=v / 100).number_format = FMT_PCT
    ws.cell(row=7, column=1, value=("Ý nghĩa: đây là mức độ nhạy cảm của LNST năm dự phóng đầu tiên khi VN-Index biến động "
                                     "so với kịch bản cơ sở — dùng để đo lường rủi ro tập trung vào mảng Tự doanh.")).font = ITALIC_FONT
    ws.cell(row=7, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells('A7:F7')
    ws.row_dimensions[7].height = 30
    print("[Excel] Sheet 08_Sensitivity done.")


def build_pestle_sheet(wb, ticker):
    ws = wb.create_sheet("09_PESTLE")
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 16
    header_row(ws, 1, ["Yếu tố", "Nội dung", "Tác động"])
    pestle = [
        ("Political", "Chính sách quản lý thị trường chứng khoán (UBCKNN), lộ trình nâng hạng thị trường mới nổi (FTSE/MSCI), cơ chế Non-pre-funding (giao dịch không cần ký quỹ 100% tiền trước) — động lực thu hút vốn ngoại.", "Positive"),
        ("Economic", "Lãi suất điều hành NHNN, tiến độ giải ngân đầu tư công, lạm phát (CPI) và thanh khoản thị trường (GTGD BQ/phiên) là driver TRỰC TIẾP doanh thu Môi giới/Margin/Tự doanh — xem chi tiết khung đánh giá bên dưới.", "Positive"),
        ("Social", "Số lượng tài khoản chứng khoán mở mới tăng nhanh, phổ cập đầu tư cá nhân qua ứng dụng di động.", "Positive"),
        ("Technological", "Chuyển đổi số, giao dịch T+0/T+2, hệ thống KRX mới giúp tăng tốc độ khớp lệnh và sản phẩm phái sinh.", "Positive"),
        ("Legal", "Quy định về tỷ lệ margin, giới hạn cho vay theo VĐL (trần 2,0x VCSH), quy định phân loại nhà đầu tư chuyên nghiệp siết chặt hơn.", "Neutral"),
        ("Environmental", "Xu hướng ESG trong lựa chọn danh mục đầu tư của quỹ ngoại ảnh hưởng gián tiếp đến dòng vốn vào các CTCK có sản phẩm xanh.", "Neutral"),
    ]
    for i, (factor, content, impact) in enumerate(pestle, start=2):
        ws.cell(row=i, column=1, value=factor).font = BOLD_FONT
        ws.cell(row=i, column=2, value=content).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value=impact)
        ws.row_dimensions[i].height = 40
        for c in range(1, 4):
            ws.cell(row=i, column=c).border = THIN_BORDER

    # ── BỐI CẢNH VĨ MÔ & THANH KHOẢN THỊ TRƯỜNG — tránh bias khi thiếu "bệ đỡ" vĩ mô ──
    r = 9
    ws.cell(row=r, column=1, value="BỐI CẢNH VĨ MÔ & THANH KHOẢN THỊ TRƯỜNG (cập nhật 2026-07, cần review định kỳ)").font = TITLE_FONT
    ws.merge_cells(f'A{r}:C{r}')
    r += 1
    ws.cell(row=r, column=1, value=(
        "⚠ Doanh thu Môi giới, Cho vay Margin và hiệu quả Tự doanh đều BÁM SÁT thanh khoản thị trường — "
        "phân tích CTCK thiếu đánh giá vĩ mô/thanh khoản rất dễ dẫn tới dự phóng cảm tính (bias). Giả định "
        "tăng trưởng GTGD trong 02_Assumptions PHẢI được đối chiếu với bảng dưới đây, KHÔNG chỉ dùng máy móc "
        "1 tốc độ cố định."))
    ws.cell(row=r, column=1).font = ITALIC_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f'A{r}:C{r}')
    ws.row_dimensions[r].height = 42
    r += 2

    header_row(ws, r, ["Biến số vĩ mô", "Đánh giá hiện tại (2026-07)", "Tác động thanh khoản"])
    r += 1
    macro_rows = [
        ("Chính sách lãi suất NHNN", "Thị trường tiền tệ đang \"căng\", áp lực hạ lãi suất hỗ trợ kinh tế nhưng phải cân bằng rủi ro "
         "lạm phát/tỷ giá — NHNN nhiều khả năng thận trọng, chưa nới mạnh trong ngắn hạn.", "Trung tính/Thận trọng"),
        ("Giải ngân đầu tư công", "Giải ngân đầu tư công nửa đầu 2026 còn CHẬM (~30,1% kế hoạch) — dòng vốn mồi vào nền kinh tế "
         "và gián tiếp vào TTCK chưa mạnh, là điểm nghẽn cần theo dõi.", "Tiêu cực (ngắn hạn)"),
        ("Lạm phát (CPI)", "CPI có dấu hiệu giảm nhưng vẫn sát ngưỡng mục tiêu — NHNN phải thận trọng, hạn chế dư địa nới lỏng "
         "tiền tệ mạnh tay.", "Trung tính"),
        ("Dòng vốn ngoại / Nâng hạng thị trường", "Cơ chế Non-pre-funding + lộ trình nâng hạng FTSE Russell là động lực thu hút vốn ngoại, "
         "có thể cải thiện thanh khoản ĐỘT BIẾN nếu được thông qua/triển khai đúng lộ trình.", "Tích cực (trung hạn)"),
    ]
    for name, content, impact in macro_rows:
        ws.cell(row=r, column=1, value=name).font = BOLD_FONT
        ws.cell(row=r, column=2, value=content).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=3, value=impact)
        ws.row_dimensions[r].height = 55
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Diễn biến GTGD bình quân/phiên toàn thị trường thực tế gần đây (tỷ VND) — nguồn: HOSE, tin tức tài chính").font = BOLD_FONT
    ws.merge_cells(f'A{r}:C{r}')
    r += 1
    header_row(ws, r, ["Kỳ", "GTGD BQ/phiên (tỷ VND)", "So với kỳ trước"])
    r += 1
    liquidity_rows = [
        ("Bình quân năm 2025", "~29.200", "—"),
        ("Quý I/2026", "35.004", "+19,9% so với BQ 2025"),
        ("Tháng 2/2026", "28.891", "—"),
        ("Tháng 3/2026", "33.865", "+6,4% MoM"),
        ("Tháng 4/2026", "24.101", "-20,7% MoM (hạ nhiệt rõ rệt)"),
    ]
    for kỳ, gtgd, sv in liquidity_rows:
        ws.cell(row=r, column=1, value=kỳ)
        ws.cell(row=r, column=2, value=gtgd)
        ws.cell(row=r, column=3, value=sv)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=(
        "Nhận định: Thanh khoản BÙNG NỔ trong Quý I/2026 nhưng đã HẠ NHIỆT RÕ RỆT từ tháng 4/2026 — xu hướng "
        "này KHÔNG ủng hộ giả định tăng trưởng GTGD đều 8%/năm một cách máy móc cho năm dự phóng đầu tiên. "
        "Cần điều chỉnh dòng \"GTGD toàn TT bình quân/phiên\" tại 02_Assumptions theo diễn biến cập nhật gần "
        "nhất khi phân tích, KHÔNG dùng nguyên giá trị mặc định của template."))
    ws.cell(row=r, column=1).font = ITALIC_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f'A{r}:C{r}')
    ws.row_dimensions[r].height = 55
    r += 2

    ws.cell(row=r, column=1, value="Cơ chế truyền dẫn: Thanh khoản → Hiệu quả từng mảng").font = BOLD_FONT
    ws.merge_cells(f'A{r}:C{r}')
    r += 1
    header_row(ws, r, ["Mảng", "Cơ chế tác động", ""])
    r += 1
    transmission_rows = [
        ("Môi giới", "Tăng NGAY LẬP TỨC theo GTGD BQ (DT = GTGD×Số phiên×Thị phần×Phí — tuyến tính với GTGD)."),
        ("Cho vay Margin", "Mảng ăn theo thanh khoản LỚN NHẤT — thị trường sôi động → nhu cầu vay tăng vọt; nhưng biên lợi "
         "nhuận phụ thuộc thêm vào Chi phí vốn (COF) của CTCK, không chỉ thanh khoản."),
        ("Tự doanh (FVTPL)", "Thanh khoản tốt giúp dễ cơ cấu danh mục/chốt lời; thanh khoản thấp khiến danh mục dễ \"kẹt\", "
         "khó hiện thực hóa lãi trên giấy thành LNST thực tế."),
    ]
    for name, content in transmission_rows:
        ws.cell(row=r, column=1, value=name).font = BOLD_FONT
        ws.cell(row=r, column=2, value=content).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f'B{r}:C{r}')
        ws.row_dimensions[r].height = 45
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    print("[Excel] Sheet 09_PESTLE done (kèm khung đánh giá vĩ mô & thanh khoản).")


def build_segment_efficiency_sheet(wb, ticker, years_hist, years_fc, RA, RR, RBS):
    """Bước 3 spec: % đóng góp + hiệu suất/biên từng mảng + cảnh báo tăng trưởng-kém-hiệu-quả-vốn +
    độ nhạy macro theo từng mảng — TOÀN BỘ bằng công thức sống link tới 02_Assumptions/03_Revenue_Model,
    không dùng số Python dán chết (thay thế sheet 10_Leading_Indicators cũ, vốn chỉ là bảng ngưỡng tĩnh)."""
    ws = wb.create_sheet("10_Hieu_Qua_Mang")
    N_HIST, N_FC = len(years_hist), len(years_fc)
    all_years = years_hist + years_fc
    N_ALL = len(all_years)
    fc0_col = get_column_letter(2 + N_HIST)
    ws.column_dimensions['A'].width = 48
    for i in range(N_ALL):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13

    def col(i):
        return get_column_letter(2 + i)

    ws.cell(row=1, column=1, value=f"HIỆU QUẢ & ĐỘ NHẠY TỪNG MẢNG KINH DOANH — {ticker}").font = TITLE_FONT
    ws.merge_cells(f'A1:{col(N_ALL-1)}1')
    r = 3
    header_row(ws, r, ["Chỉ tiêu"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years])
    r += 1

    ws.cell(row=r, column=1, value="── (A) TỶ TRỌNG ĐÓNG GÓP DOANH THU THEO MẢNG (link sống '03_Revenue_Model') ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    for offset, name in enumerate(["Môi giới", "Cho vay Margin", "Tự doanh (FVTPL)", "IB + Lưu ký", "Quản lý quỹ"]):
        ws.cell(row=r, column=1, value=f"  % {name} / Tổng DT").font = DATA_FONT
        for i in range(N_ALL):
            c = ws.cell(row=r, column=2 + i, value=f"='03_Revenue_Model'!{col(i)}{RR['pct_brokerage'] + offset}")
            c.number_format = FMT_PCT
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="── (B) HIỆU SUẤT / BIÊN TỪNG MẢNG QUA CÁC NĂM ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    r_ms = r
    ws.cell(row=r, column=1, value="Thị phần Môi giới (%) — link '02_Assumptions'").font = DATA_FONT
    for i in range(N_ALL):
        ws.cell(row=r, column=2 + i, value=f"='02_Assumptions'!{col(i)}{RA['market_share']}").number_format = FMT_PCT
    r += 1
    r_nim = r
    ws.cell(row=r, column=1, value="NIM cho vay Margin (%) — link '02_Assumptions'").font = DATA_FONT
    for i in range(N_ALL):
        ws.cell(row=r, column=2 + i, value=f"='02_Assumptions'!{col(i)}{RA['margin_nim']}").number_format = FMT_PCT
    r += 1
    r_marginbal_g = r
    ws.cell(row=r, column=1, value="  Tăng trưởng Dư nợ Margin YoY (%) — link '03_Revenue_Model'").font = ITALIC_FONT
    for i in range(1, N_ALL):
        ws.cell(row=r, column=2 + i,
                value=f"='03_Revenue_Model'!{col(i)}{RR['margin_bal']}/'03_Revenue_Model'!{col(i-1)}{RR['margin_bal']}-1").number_format = FMT_PCT
    r += 1
    r_lev = r
    ws.cell(row=r, column=1, value="  Dư nợ Margin / VCSH (x) — giới hạn pháp lý ≤ 2,0x").font = ITALIC_FONT
    for i in range(N_ALL):
        cl = col(i)
        formula = f"='03_Revenue_Model'!{cl}{RR['margin_bal']}/'05_Balance_Sheet'!{cl}{RBS['equity']}"
        ws.cell(row=r, column=2 + i, value=formula).number_format = FMT_MUL
    r += 1
    ws.cell(row=r, column=1, value="  Cảnh báo: Dư nợ Margin gần/vượt trần pháp lý 2,0x VCSH").font = BOLD_FONT
    for i in range(N_ALL):
        cl = col(i)
        formula = f'=IF({cl}{r_lev}>=2,"⚠ VƯỢT TRẦN",IF({cl}{r_lev}>=1.8,"⚠ Gần trần","Bình thường"))'
        ws.cell(row=r, column=2 + i, value=formula)
    r += 1
    r_fvyield = r
    ws.cell(row=r, column=1, value="Hiệu suất Tự doanh thực tế lịch sử (%) — link '02_Assumptions'").font = DATA_FONT
    for i in range(N_HIST):
        ws.cell(row=r, column=2 + i, value=f"='02_Assumptions'!{col(i)}{RA['fvtpl_yield_hist_ref']}").number_format = FMT_PCT
    r += 1
    r_fvbal_g = r
    ws.cell(row=r, column=1, value="  Tăng trưởng Danh mục Tự doanh YoY (%) — link '03_Revenue_Model'").font = ITALIC_FONT
    for i in range(1, N_ALL):
        ws.cell(row=r, column=2 + i,
                value=f"='03_Revenue_Model'!{col(i)}{RR['fvtpl_bal']}/'03_Revenue_Model'!{col(i-1)}{RR['fvtpl_bal']}-1").number_format = FMT_PCT
    r += 2

    ws.cell(row=r, column=1, value="── (C) CẢNH BÁO TĂNG TRƯỞNG KÉM HIỆU QUẢ VỐN (ngưỡng: dư nợ/danh mục tăng >15%/năm NHƯNG biên giảm YoY) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    ws.cell(row=r, column=1, value="Cảnh báo Margin: Dư nợ tăng nhanh nhưng NIM giảm").font = BOLD_FONT
    for i in range(1, N_ALL):
        cl, pl = col(i), col(i - 1)
        formula = f'=IF(AND({cl}{r_marginbal_g}>0.15,{cl}{r_nim}<{pl}{r_nim}),"⚠ Cảnh báo","Bình thường")'
        ws.cell(row=r, column=2 + i, value=formula)
    r += 1
    ws.cell(row=r, column=1, value="Cảnh báo Tự doanh: Danh mục tăng nhanh nhưng hiệu suất giảm").font = BOLD_FONT
    for i in range(1, N_HIST):  # chỉ tính được ở các năm có hiệu suất tự doanh THỰC TẾ lịch sử
        cl, pl = col(i), col(i - 1)
        formula = f'=IF(AND({cl}{r_fvbal_g}>0.15,{cl}{r_fvyield}<{pl}{r_fvyield}),"⚠ Cảnh báo","Bình thường")'
        ws.cell(row=r, column=2 + i, value=formula)
    r += 2

    ws.cell(row=r, column=1, value="── (D) ĐỘ NHẠY DOANH THU THEO MẢNG (kịch bản macro, năm dự phóng đầu tiên) ──").font = Font(bold=True, italic=True, color="1F4E78")
    r += 1
    header_row(ws, r, ["Kịch bản", "-20%", "-10%", "0%", "+10%", "+20%"])
    r += 1
    ws.cell(row=r, column=1, value="DT Môi giới thay đổi (tỷ VND) — theo GTGD thị trường ±X%").font = BOLD_FONT
    for j, sv in enumerate([-0.20, -0.10, 0, 0.10, 0.20]):
        ws.cell(row=r, column=2 + j, value=f"='03_Revenue_Model'!{fc0_col}{RR['brokerage']}*{sv}").number_format = FMT_NUM
    r += 2
    header_row(ws, r, ["Kịch bản", "-100bps", "-50bps", "0bps", "+50bps", "+100bps"])
    r += 1
    ws.cell(row=r, column=1, value="DT Margin thay đổi (tỷ VND) — theo NIM ±X bps").font = BOLD_FONT
    for j, sv in enumerate([-0.010, -0.005, 0, 0.005, 0.010]):
        ws.cell(row=r, column=2 + j, value=f"='03_Revenue_Model'!{fc0_col}{RR['margin_bal_avg']}*{sv}").number_format = FMT_NUM
    r += 2
    ws.cell(row=r, column=1, value=(
        "Độ nhạy DT Tự doanh (FVTPL) theo kịch bản VN-Index ±10%/±20% — xem chi tiết đầy đủ (kèm % tác động "
        "LNST) tại sheet 08_Sensitivity."))
    ws.cell(row=r, column=1).font = ITALIC_FONT
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f'A{r}:{col(N_ALL-1)}{r}')
    ws.row_dimensions[r].height = 30

    for rr in range(3, r + 1):
        for cc in range(1, N_ALL + 2):
            ws.cell(row=rr, column=cc).border = THIN_BORDER

    print("[Excel] Sheet 10_Hieu_Qua_Mang done.")


def build_thesis_sheet(wb, ticker, company_name, seg_pct, upside_pct, recommend):
    ws = wb.create_sheet("11_Investment_Thesis")
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 14
    header_row(ws, 1, ["Tầng phân tích", "Kết luận", "Đánh giá"])
    dominant_seg = max(seg_pct, key=seg_pct.get)
    rows = [
        ("Mô hình KD", f"{company_name} vận hành mô hình 5 mảng: Môi giới, Cho vay Margin, Tự doanh (FVTPL), IB+Lưu ký, Quản lý quỹ. "
                        f"Mảng đóng góp lớn nhất hiện tại: {SEGMENT_LABELS_VI.get(dominant_seg, dominant_seg)} ({seg_pct[dominant_seg]*100:.1f}% DT).", "Trung tính"),
        ("Chất lượng LN", "Lợi nhuận từ Môi giới/Margin (phí dịch vụ, lãi cho vay) ổn định hơn Tự doanh (biến động theo VN-Index) — "
                           "cần theo dõi tỷ trọng Tự doanh để đánh giá độ bền vững của LNST.", "Trung tính"),
        ("Định giá", f"Upside {upside_pct:+.1f}% so với giá hiện tại theo mô hình P/B (60%) + P/E (40%).",
         "Tích cực" if upside_pct > 15 else ("Tiêu cực" if upside_pct < -5 else "Trung tính")),
        ("Khuyến nghị", f"{recommend}", recommend),
    ]
    for i, (label, content, rating) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws.cell(row=i, column=2, value=content).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value=rating)
        ws.row_dimensions[i].height = 45
        for c in range(1, 4):
            ws.cell(row=i, column=c).border = THIN_BORDER
    print("[Excel] Sheet 11_Investment_Thesis done.")


def build_summary_snapshot(wb, ticker, years_hist, years_fc, rev_hist, rev_fc, npat_hist, npat_fc,
                            eps_hist, eps_fc, bvps_hist, bvps_fc, equity_hist, equity_fc, price, shares):
    ws = wb.create_sheet("12_Summary_Snapshot")
    all_years = years_hist + years_fc
    ws.column_dimensions['A'].width = 32
    for i in range(len(all_years)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    header_row(ws, 1, ["Chỉ tiêu"] + [f"{y}A" if y in years_hist else f"{y}E" for y in all_years])
    rev_all = rev_hist + rev_fc
    npat_all = npat_hist + npat_fc
    eps_all = eps_hist + eps_fc
    bvps_all = bvps_hist + bvps_fc
    equity_all = equity_hist + equity_fc
    rows = [
        ("Doanh thu hoạt động (tỷ VND)", rev_all, FMT_NUM),
        ("LNST cổ đông mẹ (tỷ VND)", npat_all, FMT_NUM),
        ("EPS (VND/CP)", eps_all, FMT_PRICE),
        ("BVPS (VND/CP)", bvps_all, FMT_PRICE),
        ("VCSH (tỷ VND)", equity_all, FMT_NUM),
        ("ROE (%)", [round(npat_all[i] / equity_all[i] * 100, 2) if equity_all[i] else 0 for i in range(len(all_years))], '0.00"%"'),
        ("P/E (x, tại giá hiện tại)", [round(price / eps_all[i], 2) if eps_all[i] else None for i in range(len(all_years))], FMT_MUL),
        ("P/B (x, tại giá hiện tại)", [round(price / bvps_all[i], 2) if bvps_all[i] else None for i in range(len(all_years))], FMT_MUL),
    ]
    r = 2
    for label, vals, fmt in rows:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + i, value=v)
            c.number_format = fmt
            c.border = THIN_BORDER
        r += 1
    print(f"[Excel] Sheet 12_Summary_Snapshot done ({r} dòng).")


def build_pe_pb_history_sheet(wb, ticker, quarter_labels, pe_quarters, pb_quarters, median_row):
    ws = wb.create_sheet("13_PE_PB_History")
    header_row(ws, 1, ["Quý", "P/E (x)", "P/B (x)"], [14, 14, 14])
    r = 2
    row_start = r
    for i, label in enumerate(quarter_labels):
        ws.cell(row=r, column=1, value=label)
        pe_v = pe_quarters[i] if i < len(pe_quarters) else None
        pb_v = pb_quarters[i] if i < len(pb_quarters) else None
        if pe_v and 0 < pe_v < 60:
            ws.cell(row=r, column=2, value=pe_v).number_format = FMT_MUL
        if pb_v and pb_v > 0:
            ws.cell(row=r, column=3, value=pb_v).number_format = FMT_MUL
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    row_end = r - 1
    r += 1
    assert r == median_row, f"MEDIAN row lệch ({r} != {median_row})"
    ws.cell(row=r, column=1, value="MEDIAN (loại quý P/E>60x hoặc ≤0)").font = BOLD_FONT
    ws.cell(row=r, column=2, value=f"=MEDIAN(B{row_start}:B{row_end})").number_format = FMT_MUL
    ws.cell(row=r, column=3, value=f"=MEDIAN(C{row_start}:C{row_end})").number_format = FMT_MUL
    for c in range(1, 4):
        ws.cell(row=r, column=c).font = BOLD_FONT
        ws.cell(row=r, column=c).fill = HEADER_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    print(f"[Excel] Sheet 13_PE_PB_History done ({row_end - row_start + 1} quý).")


def build_segment_quarterly_sheet(wb, ticker, is_q):
    ws = wb.create_sheet("14_Segment_Quarterly")
    header_row(ws, 1, ["Quý", "DT Môi giới", "DT Margin", "DT Tự doanh (net)", "DT IB+Lưu ký", "LNST"],
               [12, 14, 14, 16, 14, 14])
    quarters = sorted(set((r.get("yearReport"), r.get("lengthReport")) for r in is_q
                          if r.get("yearReport") and r.get("lengthReport") in (1, 2, 3, 4)))
    r = 2
    for y, q in quarters:
        rec = next((x for x in is_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
        brokerage = (rec.get(SEG["brokerage_rev"]) or 0) / 1e9
        margin = (rec.get(SEG["margin_rev"]) or 0) / 1e9
        fvtpl = ((rec.get(SEG["fvtpl_gain"]) or 0) + (rec.get(SEG["fvtpl_loss"]) or 0)) / 1e9
        ib = ((rec.get(SEG["ib_underwrite"]) or 0) + (rec.get(SEG["ib_advisory"]) or 0)
              + (rec.get(SEG["custody_rev"]) or 0) + (rec.get(SEG["ib_finadvisory"]) or 0)) / 1e9
        npat = (rec.get(IS_TOTAL["npat_parent"]) or 0) / 1e9
        ws.cell(row=r, column=1, value=f"{y}Q{q}")
        for c, v in enumerate([brokerage, margin, fvtpl, ib, npat], start=2):
            ws.cell(row=r, column=c, value=round(v, 1)).number_format = FMT_NUM
        r += 1
    print(f"[Excel] Sheet 14_Segment_Quarterly done ({r-2} quý).")


def build_peer_benchmark_sheet(wb, ticker):
    """Bảng thống kê so sánh các CTCK niêm yết — THUẦN TÚY diễn giải trực quan đặc điểm doanh nghiệp
    so với ngành, KHÔNG dùng làm input cho bất kỳ công thức định giá nào của ticker đang phân tích
    (toàn bộ giá trị là số liệu bên ngoài dán tĩnh, đúng bản chất "external reference data")."""
    ws = wb.create_sheet("15_Peer_Benchmark")
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 24
    for col_letter in "CDEFGHIJKL":
        ws.column_dimensions[col_letter].width = 15

    json_path = os.path.join(PROJECT_ROOT, "data", "peer_benchmark_securities.json")
    try:
        if not os.path.exists(json_path):
            import update_peer_benchmark_securities
            update_peer_benchmark_securities.main()
        with open(json_path, "r", encoding="utf-8") as f:
            db = json.load(f)
        peers = db.get("peers", [])
        updated = db.get("_meta", {}).get("updated", "N/A")
    except Exception as e:
        print(f"  [WARN] Không tải được dữ liệu peer benchmark CTCK: {e}")
        peers = []
        updated = "N/A"

    ws.cell(row=1, column=1, value=f"BẢNG THỐNG KÊ SO SÁNH CÁC CTCK NIÊM YẾT (cập nhật {updated})").font = TITLE_FONT
    ws.merge_cells('A1:L1')
    ws.cell(row=2, column=1, value=(
        "Bảng THUẦN TÚY diễn giải/so sánh trực quan đặc điểm doanh nghiệp so với ngành — KHÔNG dùng làm "
        "input định giá cho bất kỳ mã nào trong hệ thống (mọi công thức định giá chỉ dùng dữ liệu lịch sử/"
        "dự phóng của chính ticker đang phân tích)."))
    ws.cell(row=2, column=1).font = ITALIC_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells('A2:L2')
    ws.row_dimensions[2].height = 28

    header_row(ws, 4, ["Mã", "Tên công ty", "Vốn điều lệ (tỷ)", "Vốn hóa (tỷ)", "P/B (x)", "P/E (x)",
                       "ROE (%)", "Margin/VCSH (x)", "% DT Margin", "% DT Môi giới", "% DT Tự doanh"])
    r = 5
    for p in peers:
        is_cur = p.get("ticker") == ticker
        row_vals = [p.get("ticker"), p.get("name"), p.get("charter_capital"), p.get("mcap"), p.get("pb"),
                    p.get("pe"), p.get("roe"), p.get("margin_to_equity"), p.get("pct_margin_rev"),
                    p.get("pct_brokerage_rev"), p.get("pct_tudoanh_rev")]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN_BORDER
            if c == 1:
                cell.font = Font(bold=True, color="1F4E78" if not is_cur else "FFFFFF")
            if is_cur:
                cell.fill = HEADER_FILL
                cell.font = Font(bold=True, color="FFFFFF")
        r += 1

    if peers:
        r += 1
        ws.cell(row=r, column=1, value="Trung vị ngành").font = BOLD_FONT
        for c, key in enumerate(["charter_capital", "mcap", "pb", "pe", "roe", "margin_to_equity",
                                  "pct_margin_rev", "pct_brokerage_rev", "pct_tudoanh_rev"], start=3):
            vals = [p.get(key) for p in peers if p.get(key) is not None]
            med = round(stats.median(vals), 2) if vals else None
            ws.cell(row=r, column=c, value=med).font = BOLD_FONT

    print(f"[Excel] Sheet 15_Peer_Benchmark done ({len(peers)} mã CTCK).")


# ══════════════════════════════════════════════════════════════════════════
# BIỂU ĐỒ (matplotlib)
# ══════════════════════════════════════════════════════════════════════════
plt.rcParams['font.family'] = 'DejaVu Sans'
plt.rcParams['axes.unicode_minus'] = False


def make_segment_revenue_chart(chart_dir, ticker, years_hist, years_fc, seg_hist, seg_fc):
    all_years = years_hist + years_fc
    labels = [f"{y}A" if y in years_hist else f"{y}E" for y in all_years]
    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = list(range(len(all_years)))
    n = len(all_years)
    bottom = [0.0] * n
    totals = [sum(seg_hist[seg][i] if i < len(seg_hist[seg]) else seg_fc[seg][i - len(seg_hist[seg])]
                  for seg in SEGMENT_NAMES) for i in range(n)]
    for seg in SEGMENT_NAMES:
        vals = seg_hist[seg] + seg_fc[seg]
        ax.bar(x, vals, bottom=bottom, label=SEGMENT_LABELS_VI[seg], color=SEGMENT_COLORS[seg])
        for i in range(n):
            pct = vals[i] / totals[i] * 100 if totals[i] else 0
            if pct >= 5:  # bỏ nhãn mảng quá nhỏ để tránh chồng chữ
                ax.text(x[i], bottom[i] + vals[i] / 2, f"{pct:.0f}%", ha='center', va='center',
                        fontsize=7.5, color='white', fontweight='bold')
        bottom = [bottom[i] + vals[i] for i in range(n)]
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha='right')
    ax.set_ylabel("Tỷ VND", fontsize=11)
    ax.set_title(f"{ticker} — Doanh thu theo 5 mảng kinh doanh (Bottom-up)", fontsize=13, fontweight='bold')
    ax.legend(fontsize=9, loc='upper left')
    ax.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(chart_dir, "segment_revenue.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_segment_mix_chart(chart_dir, ticker, years_hist, years_fc, seg_hist, seg_fc):
    all_years = years_hist + years_fc
    labels = [f"{y}A" if y in years_hist else f"{y}E" for y in all_years]
    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = list(range(len(all_years)))
    n = len(all_years)
    totals = [sum(seg_hist[seg][i] if i < len(seg_hist[seg]) else seg_fc[seg][i - len(seg_hist[seg])]
                  for seg in SEGMENT_NAMES) for i in range(n)]
    bottom = [0.0] * n
    for seg in SEGMENT_NAMES:
        vals = seg_hist[seg] + seg_fc[seg]
        pct = [vals[i] / totals[i] * 100 if totals[i] else 0 for i in range(n)]
        ax.bar(x, pct, bottom=bottom, label=SEGMENT_LABELS_VI[seg], color=SEGMENT_COLORS[seg])
        for i in range(n):
            if pct[i] >= 5:  # bỏ nhãn mảng quá nhỏ để tránh chồng chữ
                ax.text(x[i], bottom[i] + pct[i] / 2, f"{vals[i]:,.0f}", ha='center', va='center',
                        fontsize=7.5, color='white', fontweight='bold')
        bottom = [bottom[i] + pct[i] for i in range(n)]
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha='right')
    ax.set_ylabel("% Tổng doanh thu", fontsize=11)
    ax.set_ylim(0, 100)
    ax.set_title(f"{ticker} — Cơ cấu đóng góp doanh thu theo mảng (% — số trong cột là giá trị tỷ VND)", fontsize=13, fontweight='bold')
    ax.legend(fontsize=9, loc='upper left', bbox_to_anchor=(1.01, 1))
    ax.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(chart_dir, "segment_mix.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_pe_pb_chart(chart_dir, ticker, quarter_labels, vals, median, name, color):
    fig, ax = plt.subplots(figsize=(11, 5))
    x = list(range(len(quarter_labels)))
    y = [v if v and 0 < v < 100 else None for v in vals]
    ax.plot(x, y, marker='o', markersize=3.5, linewidth=1.6, color=color, label=f"{name} TTM")
    ax.axhline(median, color="#E74C3C", linestyle='--', linewidth=1.8, label=f"Median: {median:.2f}x")
    step = max(1, len(x) // 16)
    ax.set_xticks(x[::step])
    ax.set_xticklabels([quarter_labels[i] for i in x[::step]], rotation=45, ha='right', fontsize=8)
    ax.set_ylabel(f"{name} (x)", fontsize=11)
    ax.set_title(f"{ticker} — {name} lịch sử theo quý", fontsize=13, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(chart_dir, f"{name.lower().replace('/', '')}_history.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_margin_nim_chart(chart_dir, ticker, years_hist, years_fc, margin_hist, margin_fc, nim_hist, nim_fc):
    all_years = years_hist + years_fc
    labels = [f"{y}A" if y in years_hist else f"{y}E" for y in all_years]
    fig, ax1 = plt.subplots(figsize=(10, 5))
    x = list(range(len(all_years)))
    ax1.bar(x, margin_hist + margin_fc, color="#f59e0b", alpha=0.85, label="Dư nợ Margin (tỷ VND)")
    ax1.set_ylabel("Dư nợ Margin (tỷ VND)", fontsize=11)
    ax2 = ax1.twinx()
    ax2.plot(x, [v * 100 for v in nim_hist + nim_fc], marker='o', color="#2980B9", linewidth=2, label="NIM Margin (%)")
    ax2.set_ylabel("NIM (%)", fontsize=11)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45, ha='right')
    ax1.set_title(f"{ticker} — Dư nợ Margin & NIM cho vay Margin", fontsize=13, fontweight='bold')
    lines1, labs1 = ax1.get_legend_handles_labels()
    lines2, labs2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labs1 + labs2, fontsize=9, loc='upper left')
    ax1.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(chart_dir, "margin_nim.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_fvtpl_sensitivity_chart(chart_dir, ticker, vnindex_range, npat_impact_pct):
    fig, ax = plt.subplots(figsize=(9, 5))
    labels = [f"{s*100:+.0f}%" for s in vnindex_range]
    colors = ["#dc2626" if v < 0 else "#16a34a" for v in npat_impact_pct]
    ax.bar(labels, npat_impact_pct, color=colors)
    ax.axhline(0, color='black', linewidth=0.8)
    ax.set_xlabel("Biến động VN-Index", fontsize=11)
    ax.set_ylabel("Tác động LNST năm dự phóng đầu (%)", fontsize=11)
    ax.set_title(f"{ticker} — Độ nhạy LNST theo biến động VN-Index (rủi ro Tự doanh)", fontsize=12.5, fontweight='bold')
    ax.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(chart_dir, "fvtpl_sensitivity.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_revenue_npat_chart(chart_dir, ticker, years_hist, years_fc, rev_hist, rev_fc, npat_hist, npat_fc):
    all_years = years_hist + years_fc
    labels = [f"{y}A" if y in years_hist else f"{y}E" for y in all_years]
    fig, ax1 = plt.subplots(figsize=(10, 5))
    x = list(range(len(all_years)))
    ax1.bar(x, rev_hist + rev_fc, color="#3b82f6", alpha=0.8, label="Doanh thu hoạt động (tỷ VND)")
    ax1.set_ylabel("Doanh thu (tỷ VND)", fontsize=11)
    ax2 = ax1.twinx()
    ax2.plot(x, npat_hist + npat_fc, marker='o', color="#16a34a", linewidth=2.2, label="LNST cổ đông mẹ (tỷ VND)")
    ax2.set_ylabel("LNST (tỷ VND)", fontsize=11)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45, ha='right')
    ax1.set_title(f"{ticker} — Doanh thu & LNST", fontsize=13, fontweight='bold')
    lines1, labs1 = ax1.get_legend_handles_labels()
    lines2, labs2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labs1 + labs2, fontsize=9, loc='upper left')
    ax1.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(chart_dir, "revenue_npat.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_npat_roe_quarterly_chart(chart_dir, ticker, is_q, bs_q, n_quarters=12):
    """LNST theo quý (cột) + ROE theo quý ANNUALIZED (đường, = LNST quý × 4 / VCSH cuối quý) — dùng
    dữ liệu quý THỰC TẾ (is_q/bs_q), không phải Python mirror dự phóng, để phản ánh đúng diễn biến
    quý gần nhất mà user cần theo dõi (khác biểu đồ revenue_npat theo NĂM ở trên)."""
    quarters = sorted(set((r.get("yearReport"), r.get("lengthReport")) for r in is_q
                          if r.get("yearReport") and r.get("lengthReport") in (1, 2, 3, 4)))
    recent = quarters[-n_quarters:]
    labels, npat_vals, roe_vals = [], [], []
    for y, q in recent:
        rec_is = next((x for x in is_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
        rec_bs = next((x for x in bs_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
        npat = (rec_is.get(IS_TOTAL["npat_parent"]) or 0) / 1e9
        equity = (rec_bs.get(BS["equity"]) or 0) / 1e9
        roe_q = (npat * 4 / equity * 100) if equity else None
        labels.append(f"{y}Q{q}")
        npat_vals.append(round(npat, 1))
        roe_vals.append(round(roe_q, 2) if roe_q is not None else None)

    fig, ax1 = plt.subplots(figsize=(11, 5.5))
    x = list(range(len(recent)))
    ax1.bar(x, npat_vals, color="#16a34a", alpha=0.85, label="LNST cổ đông mẹ (tỷ VND)")
    ax1.set_ylabel("LNST (tỷ VND)", fontsize=11)
    ax2 = ax1.twinx()
    ax2.plot(x, roe_vals, marker='o', color="#dc2626", linewidth=2.2, label="ROE (%/năm, annualized theo quý)")
    ax2.set_ylabel("ROE (%/năm)", fontsize=11)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45, ha='right')
    ax1.set_title(f"{ticker} — LNST & ROE theo Quý (annualized)", fontsize=13, fontweight='bold')
    lines1, labs1 = ax1.get_legend_handles_labels()
    lines2, labs2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labs1 + labs2, fontsize=9, loc='upper left')
    ax1.grid(alpha=0.3, axis='y')
    fig.tight_layout()
    path = os.path.join(chart_dir, "npat_roe_quarterly.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_roe_pb_correlation_chart(chart_dir, ticker, quarter_labels, pb_quarters, is_q, bs_q):
    """Scatter tương quan ROE quý (annualized) vs P/B quý — kiểm định trực quan luận điểm định giá
    P/B: ROE cao hơn thường đi kèm P/B thị trường chấp nhận cao hơn. Kèm đường hồi quy tuyến tính +
    hệ số tương quan r để đo độ mạnh của mối quan hệ cho chính ticker đang phân tích."""
    roe_vals, pb_vals, labels = [], [], []
    for i, lbl in enumerate(quarter_labels):
        try:
            y_str, q_str = lbl.split("-Q")
            y, q = int(y_str), int(q_str)
        except (ValueError, IndexError):
            continue
        pb_v = pb_quarters[i] if i < len(pb_quarters) else None
        if not pb_v or pb_v <= 0:
            continue
        rec_is = next((x for x in is_q if x.get("yearReport") == y and x.get("lengthReport") == q), None)
        rec_bs = next((x for x in bs_q if x.get("yearReport") == y and x.get("lengthReport") == q), None)
        if not rec_is or not rec_bs:
            continue
        npat = (rec_is.get(IS_TOTAL["npat_parent"]) or 0) / 1e9
        equity = (rec_bs.get(BS["equity"]) or 0) / 1e9
        if not equity:
            continue
        roe_q = npat * 4 / equity * 100
        roe_vals.append(roe_q)
        pb_vals.append(pb_v)
        labels.append(lbl)

    fig, ax = plt.subplots(figsize=(9, 7))
    if len(roe_vals) >= 2:
        n = len(roe_vals)
        colors = plt.cm.viridis([i / max(n - 1, 1) for i in range(n)])
        sc = ax.scatter(roe_vals, pb_vals, c=range(n), cmap='viridis', s=70, edgecolors='black', linewidths=0.5, zorder=3)
        for i in (0, n - 1):
            ax.annotate(labels[i], (roe_vals[i], pb_vals[i]), fontsize=8, xytext=(5, 5), textcoords='offset points')
        corr = np.corrcoef(roe_vals, pb_vals)[0, 1] if n >= 3 else 0.0
        z = np.polyfit(roe_vals, pb_vals, 1)
        x_line = np.linspace(min(roe_vals), max(roe_vals), 50)
        ax.plot(x_line, z[0] * x_line + z[1], linestyle='--', color="#dc2626", linewidth=1.8,
                label=f"Hồi quy tuyến tính (r = {corr:.2f})")
        cbar = fig.colorbar(sc, ax=ax)
        cbar.set_label("Thời gian (cũ → mới)", fontsize=9)
        ax.legend(fontsize=9, loc='best')
    ax.set_xlabel("ROE (%/năm, annualized theo quý)", fontsize=11)
    ax.set_ylabel("P/B (x)", fontsize=11)
    ax.set_title(f"{ticker} — Tương quan ROE quý vs P/B quý", fontsize=13, fontweight='bold')
    ax.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(chart_dir, "roe_pb_correlation.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


def make_margin_leverage_quarterly_chart(chart_dir, ticker, bs_q, n_quarters=12):
    """Dư nợ Margin / VCSH theo QUÝ (đòn bẩy cho vay ký quỹ) — kèm 2 đường ngưỡng tham chiếu: 1,8x
    (cảnh báo gần trần) và 2,0x (trần pháp lý cho vay margin / VCSH của CTCK)."""
    quarters = sorted(set((r.get("yearReport"), r.get("lengthReport")) for r in bs_q
                          if r.get("yearReport") and r.get("lengthReport") in (1, 2, 3, 4)))
    recent = quarters[-n_quarters:]
    labels, lev_vals = [], []
    for y, q in recent:
        rec = next((x for x in bs_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
        margin_loans = (rec.get(BS["margin_loans"]) or 0) / 1e9
        equity = (rec.get(BS["equity"]) or 0) / 1e9
        lev = round(margin_loans / equity, 2) if equity else None
        labels.append(f"{y}Q{q}")
        lev_vals.append(lev)

    fig, ax = plt.subplots(figsize=(11, 5.5))
    x = list(range(len(recent)))
    ax.plot(x, lev_vals, marker='o', markersize=5, linewidth=2.2, color="#1F4E78", label="Dư nợ Margin / VCSH (x)")
    ax.axhline(2.0, color="#dc2626", linestyle='--', linewidth=1.5, label="Trần pháp lý 2,0x")
    ax.axhline(1.8, color="#f59e0b", linestyle=':', linewidth=1.5, label="Ngưỡng cảnh báo 1,8x")
    for i, v in enumerate(lev_vals):
        if v is not None:
            ax.annotate(f"{v:.2f}x", (x[i], v), fontsize=7.5, ha='center', va='bottom', xytext=(0, 4), textcoords='offset points')
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha='right')
    ax.set_ylabel("Dư nợ Margin / VCSH (x)", fontsize=11)
    ax.set_title(f"{ticker} — Đòn bẩy Cho vay Margin theo Quý", fontsize=13, fontweight='bold')
    ax.legend(fontsize=9, loc='best')
    ax.grid(alpha=0.3)
    fig.tight_layout()
    path = os.path.join(chart_dir, "margin_leverage_quarterly.png")
    fig.savefig(path, dpi=180, bbox_inches='tight')
    plt.close(fig)
    return path


# ══════════════════════════════════════════════════════════════════════════
# PDF REPORT
# ══════════════════════════════════════════════════════════════════════════
def build_pdf_report(pdf_path, ticker, company_name, price, mcap, shares, target, upside_pct, recommend,
                      pb_target, pe_target, bear_target, bull_target, pe_median, pb_median, coe,
                      years_hist, years_fc, rev_hist, rev_fc, npat_hist, npat_fc, eps_hist, eps_fc,
                      bvps_hist, bvps_fc, seg_pct_fc0, concentration_seg, concentration_flag,
                      chart_paths, fvtpl_sens_pct, vnindex_range,
                      latest_margin_leverage=0.0, latest_leverage_label="", weights=None,
                      quarterly_update=None):
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle('TitleVN', fontName=FONT_BOLD, fontSize=20, textColor=HexColor("#1F4E78"),
                                  alignment=TA_CENTER, spaceAfter=10)
    style_h2 = ParagraphStyle('H2VN', fontName=FONT_BOLD, fontSize=14, textColor=HexColor("#1F4E78"), spaceAfter=8, spaceBefore=12)
    style_body = ParagraphStyle('BodyVN', fontName=FONT_REG, fontSize=10, leading=14, alignment=TA_JUSTIFY, spaceAfter=6)
    style_small = ParagraphStyle('SmallVN', fontName=FONT_REG, fontSize=8.5, leading=11, textColor=HexColor("#555555"))

    doc = SimpleDocTemplate(pdf_path, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=18*mm, rightMargin=18*mm)
    story = []

    # ── Trang 1: Cover ──
    story.append(Spacer(1, 30*mm))
    story.append(Paragraph(f"{ticker} — {company_name}", style_title))
    story.append(Paragraph("BÁO CÁO PHÂN TÍCH CƠ BẢN CHUYÊN SÂU — NGÀNH CHỨNG KHOÁN (CTCK)", style_h2))
    story.append(Spacer(1, 10*mm))
    rec_color = HexColor("#16a34a") if recommend == "MUA" else (HexColor("#dc2626") if recommend == "BÁN" else HexColor("#f59e0b"))
    cover_data = [
        ["Giá hiện tại", f"{price:,.0f} VND"],
        ["Giá mục tiêu", f"{target:,.0f} VND"],
        ["Tiềm năng tăng giá", f"{upside_pct:+.1f}%"],
        ["Khuyến nghị", recommend],
    ]
    t = Table(cover_data, colWidths=[80*mm, 60*mm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), FONT_REG), ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('FONTNAME', (0, 3), (-1, 3), FONT_BOLD), ('TEXTCOLOR', (1, 3), (1, 3), rec_color),
        ('GRID', (0, 0), (-1, -1), 0.5, grey), ('BACKGROUND', (0, 0), (0, -1), HexColor("#F2F2F2")),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'), ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 15*mm))
    story.append(Paragraph(f"Ngày lập báo cáo: {datetime.datetime.now().strftime('%d/%m/%Y')}", style_small))
    story.append(PageBreak())

    # ── Trang 2: Investment Summary ──
    story.append(Paragraph("TÓM TẮT KHUYẾN NGHỊ ĐẦU TƯ", style_h2))
    dominant = SEGMENT_LABELS_VI.get(concentration_seg, concentration_seg)
    story.append(Paragraph(
        f"{company_name} ({ticker}) là công ty chứng khoán vận hành mô hình kinh doanh 5 mảng cốt lõi: "
        f"Môi giới, Cho vay Margin, Tự doanh (FVTPL), Ngân hàng đầu tư (IB) + Lưu ký, và Quản lý quỹ. "
        f"Mảng đóng góp doanh thu lớn nhất hiện tại là <b>{dominant}</b> ({seg_pct_fc0[concentration_seg]*100:.1f}% tổng doanh thu), "
        f"mức độ tập trung được đánh giá ở trạng thái <b>{concentration_flag}</b>.", style_body))
    story.append(Paragraph(
        f"Áp dụng mô hình định giá kết hợp P/B ({weights['PB']*100:.0f}%, trung vị lịch sử {pb_median:.2f}x) và "
        f"P/E ({weights['PE']*100:.0f}%, trung vị lịch sử {pe_median:.2f}x), giá mục tiêu ước tính là "
        f"<b>{target:,.0f} VND</b>, tương ứng tiềm năng "
        f"{'tăng giá' if upside_pct >= 0 else 'giảm giá'} <b>{upside_pct:+.1f}%</b> so với giá hiện tại "
        f"{price:,.0f} VND. Khuyến nghị: <b>{recommend}</b>.", style_body))
    story.append(Paragraph(
        "Rủi ro chính cần lưu ý: (1) LNST phụ thuộc một phần vào mảng Tự doanh (FVTPL) — biến động mạnh theo "
        "VN-Index nên kém ổn định hơn nguồn thu phí dịch vụ; (2) Cạnh tranh phí môi giới ngày càng gay gắt "
        "(một số CTCK áp dụng phí 0%) gây áp lực biên lợi nhuận mảng môi giới; (3) Rủi ro lãi suất — chi phí "
        "vốn tài trợ cho vay margin và danh mục tự doanh tăng khi lãi suất huy động tăng.", style_body))
    story.append(Spacer(1, 6*mm))

    if os.path.exists(chart_paths.get("revenue_npat", "")):
        story.append(Image(chart_paths["revenue_npat"], width=160*mm, height=80*mm))
    story.append(PageBreak())

    # ── Trang 2b: KQKD Quý gần nhất & Lũy kế tới hiện tại ──
    if quarterly_update:
        qu = quarterly_update
        story.append(Paragraph("KẾT QUẢ KINH DOANH QUÝ GẦN NHẤT & LŨY KẾ TỚI HIỆN TẠI", style_h2))
        yoy_rev_txt = f"{qu['yoy_rev_pct']:+.1f}% YoY" if qu['yoy_rev_pct'] is not None else "không có dữ liệu cùng kỳ"
        yoy_npat_txt = f"{qu['yoy_npat_pct']:+.1f}% YoY" if qu['yoy_npat_pct'] is not None else "không có dữ liệu cùng kỳ"
        story.append(Paragraph(
            f"Quý gần nhất (<b>{qu['quarter_label']}</b>): Doanh thu hoạt động đạt <b>{qu['rev']:,.0f} tỷ VND</b> "
            f"({yoy_rev_txt}), LNST cổ đông mẹ đạt <b>{qu['npat']:,.0f} tỷ VND</b> ({yoy_npat_txt}).", style_body))
        if qu['n_known_q'] > 0:
            pct_rev_txt = f"{qu['pct_of_annual_est_rev']:.1f}%" if qu['pct_of_annual_est_rev'] is not None else "N/A"
            pct_npat_txt = f"{qu['pct_of_annual_est_npat']:.1f}%" if qu['pct_of_annual_est_npat'] is not None else "N/A"
            story.append(Paragraph(
                f"Lũy kế {qu['n_known_q']}/4 quý đã công bố của năm {qu['cur_fc_year']}: Doanh thu đạt "
                f"<b>{qu['cum_rev']:,.0f} tỷ VND</b> (tương đương {pct_rev_txt} ước tính cả năm dự phóng), LNST đạt "
                f"<b>{qu['cum_npat']:,.0f} tỷ VND</b> (tương đương {pct_npat_txt} ước tính cả năm dự phóng). Phần còn "
                f"lại của năm ({4 - qu['n_known_q']} quý) được ước tính theo giả định bottom-up tại sheet "
                f"02_Assumptions/03_Revenue_Model — xem chi tiết công thức blend tại sheet 04b_Dien_Bien_Quy.",
                style_body))
        else:
            story.append(Paragraph(
                f"Chưa có quý nào của năm {qu['cur_fc_year']} được công bố tại thời điểm phân tích — ước tính cả năm "
                f"hiện dựa hoàn toàn trên giả định bottom-up (chưa blend với số liệu thực tế).", style_body))
        story.append(Spacer(1, 4*mm))
        if os.path.exists(chart_paths.get("npat_roe_quarterly", "")):
            story.append(Image(chart_paths["npat_roe_quarterly"], width=160*mm, height=80*mm))
        story.append(PageBreak())

    # ── Trang 3: Mô hình kinh doanh 5 mảng ──
    story.append(Paragraph("MÔ HÌNH KINH DOANH — 5 MẢNG BOTTOM-UP", style_h2))
    story.append(Paragraph(
        "Doanh thu hoạt động được dự báo theo phương pháp Bottom-up cho từng mảng kinh doanh, dựa trên các "
        "driver (biến số động lực) chính: (1) Môi giới = GTGD/phiên × Số phiên × Thị phần × Phí(bps); "
        "(2) Cho vay Margin = Dư nợ bình quân × NIM; (3) Tự doanh (FVTPL) = Danh mục × lợi suất kỳ vọng theo "
        "cơ cấu tài sản; (4) IB + Lưu ký = Pipeline deal × Phí% + AUM lưu ký × Phí lưu ký%; "
        "(5) Quản lý quỹ = AUM × Fee rate (~0.75%/năm).", style_body))
    if os.path.exists(chart_paths.get("segment_revenue", "")):
        story.append(Image(chart_paths["segment_revenue"], width=160*mm, height=80*mm))
    story.append(Spacer(1, 4*mm))
    if os.path.exists(chart_paths.get("segment_mix", "")):
        story.append(Image(chart_paths["segment_mix"], width=160*mm, height=80*mm))
    story.append(PageBreak())

    # ── Trang 4: Margin lending + Sensitivity ──
    story.append(Paragraph("CHO VAY MARGIN & ĐỘ NHẠY TỰ DOANH", style_h2))
    if os.path.exists(chart_paths.get("margin_nim", "")):
        story.append(Image(chart_paths["margin_nim"], width=160*mm, height=80*mm))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(
        "Mảng Tự doanh (FVTPL) là nguồn biến động lợi nhuận lớn nhất của CTCK do gắn trực tiếp với diễn biến "
        "VN-Index. Bảng dưới đây mô phỏng tác động của các kịch bản VN-Index lên LNST năm dự phóng đầu tiên:", style_body))
    if os.path.exists(chart_paths.get("fvtpl_sensitivity", "")):
        story.append(Image(chart_paths["fvtpl_sensitivity"], width=140*mm, height=78*mm))
    story.append(PageBreak())

    # ── Trang 4b: Đòn bẩy Cho vay Margin/VCSH theo quý ──
    story.append(Paragraph("ĐÒN BẨY CHO VAY MARGIN / VCSH THEO QUÝ", style_h2))
    _lev_status = ("đã VƯỢT trần pháp lý" if latest_margin_leverage >= 2.0 else
                   "đang GẦN chạm trần pháp lý, dư địa tăng trưởng hạn chế" if latest_margin_leverage >= 1.8 else
                   "còn dư địa tăng trưởng dưới trần pháp lý")
    story.append(Paragraph(
        f"Quy định hiện hành giới hạn Dư nợ cho vay Margin tối đa <b>2,0 lần VCSH</b> của CTCK. Tại quý gần nhất "
        f"({latest_leverage_label}), tỷ lệ này của {ticker} là <b>{latest_margin_leverage:.2f}x</b> — {_lev_status}. "
        f"Khi tỷ lệ tiệm cận 2,0x, tăng trưởng dư nợ margin (và do đó doanh thu mảng Cho vay) sẽ phụ thuộc trực tiếp "
        f"vào tiến độ các đợt tăng vốn điều lệ/phát hành cổ phiếu tiếp theo của công ty — cần theo dõi kế hoạch "
        f"ĐHĐCĐ để đánh giá dư địa tăng trưởng thực tế cho các năm dự phóng.", style_body))
    if os.path.exists(chart_paths.get("margin_leverage_quarterly", "")):
        story.append(Image(chart_paths["margin_leverage_quarterly"], width=160*mm, height=80*mm))
    story.append(PageBreak())

    # ── Trang 5: P/E, P/B lịch sử ──
    story.append(Paragraph("ĐỊNH GIÁ LỊCH SỬ P/E, P/B", style_h2))
    if os.path.exists(chart_paths.get("pe_history", "")):
        story.append(Image(chart_paths["pe_history"], width=160*mm, height=72*mm))
    story.append(Spacer(1, 3*mm))
    if os.path.exists(chart_paths.get("pb_history", "")):
        story.append(Image(chart_paths["pb_history"], width=160*mm, height=72*mm))
    story.append(PageBreak())

    # ── Trang 5b: LNST/ROE theo quý & Tương quan ROE-P/B ──
    story.append(Paragraph("LNST, ROE THEO QUÝ & TƯƠNG QUAN ROE-P/B", style_h2))
    story.append(Paragraph(
        "Theo dõi diễn biến LNST và ROE (annualized) theo từng quý gần nhất giúp đánh giá xu hướng hiệu quả "
        "sinh lời thực tế, tách biệt khỏi ước tính dự phóng cả năm. Biểu đồ tương quan bên dưới kiểm định "
        "trực quan luận điểm định giá P/B: thị trường thường trả P/B cao hơn cho CTCK có ROE cao hơn.", style_body))
    if os.path.exists(chart_paths.get("npat_roe_quarterly", "")):
        story.append(Image(chart_paths["npat_roe_quarterly"], width=160*mm, height=80*mm))
    story.append(Spacer(1, 3*mm))
    if os.path.exists(chart_paths.get("roe_pb_correlation", "")):
        story.append(Image(chart_paths["roe_pb_correlation"], width=130*mm, height=101*mm))
    story.append(PageBreak())

    # ── Trang 6: Bảng tài chính tóm tắt ──
    story.append(Paragraph("BẢNG TÀI CHÍNH TÓM TẮT", style_h2))
    all_years = years_hist + years_fc
    year_labels = [f"{y}A" if y in years_hist else f"{y}E" for y in all_years]
    fin_data = [["Chỉ tiêu (tỷ VND)"] + year_labels]
    fin_data.append(["Doanh thu hoạt động"] + [f"{v:,.0f}" for v in rev_hist + rev_fc])
    fin_data.append(["LNST cổ đông mẹ"] + [f"{v:,.0f}" for v in npat_hist + npat_fc])
    fin_data.append(["EPS (VND)"] + [f"{v:,.0f}" for v in eps_hist + eps_fc])
    fin_data.append(["BVPS (VND)"] + [f"{v:,.0f}" for v in bvps_hist + bvps_fc])
    t2 = Table(fin_data, colWidths=[45*mm] + [16*mm] * len(all_years))
    t2.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), FONT_REG), ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD), ('BACKGROUND', (0, 0), (-1, 0), HexColor("#1F4E78")),
        ('TEXTCOLOR', (0, 0), (-1, 0), white), ('GRID', (0, 0), (-1, -1), 0.4, grey),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(t2)
    story.append(PageBreak())

    # ── Trang 6b: Bảng so sánh CTCK cùng ngành (Peer Benchmark) ──
    try:
        with open(os.path.join(PROJECT_ROOT, "data", "peer_benchmark_securities.json"), "r", encoding="utf-8") as f:
            _peer_db = json.load(f)
        _peers_pdf = _peer_db.get("peers", [])
        _peer_updated = _peer_db.get("_meta", {}).get("updated", "N/A")
    except Exception:
        _peers_pdf = []
        _peer_updated = "N/A"

    if _peers_pdf:
        story.append(Paragraph(f"BẢNG SO SÁNH CÁC CTCK CÙNG NGÀNH (cập nhật {_peer_updated})", style_h2))
        story.append(Paragraph(
            "Bảng thống kê THUẦN TÚY diễn giải/so sánh trực quan — KHÔNG dùng làm input cho định giá "
            f"{ticker} ở trên (mọi công thức định giá chỉ dùng dữ liệu lịch sử/dự phóng riêng của {ticker}).",
            style_small))
        story.append(Spacer(1, 2*mm))
        peer_rows = [["Mã", "P/B (x)", "P/E (x)", "ROE (%)", "Margin/VCSH (x)", "%DT Margin", "%DT Môi giới", "%DT Tự doanh"]]
        _pb_vals = [p.get("pb") for p in _peers_pdf if p.get("pb") is not None]
        _pe_vals = [p.get("pe") for p in _peers_pdf if p.get("pe") is not None]
        _roe_vals = [p.get("roe") for p in _peers_pdf if p.get("roe") is not None]
        _lev_vals = [p.get("margin_to_equity") for p in _peers_pdf if p.get("margin_to_equity") is not None]
        if _pb_vals:
            peer_rows.append(["TB Ngành", f"{stats.median(_pb_vals):.2f}x", f"{stats.median(_pe_vals):.1f}x" if _pe_vals else "-",
                               f"{stats.median(_roe_vals):.1f}%" if _roe_vals else "-",
                               f"{stats.median(_lev_vals):.2f}x" if _lev_vals else "-", "-", "-", "-"])
        for p in _peers_pdf:
            t_id = p.get("ticker", "")
            marker = "*" if t_id == ticker else ""
            peer_rows.append([
                f"{marker}{t_id}",
                f"{p.get('pb'):.2f}x" if p.get('pb') is not None else "-",
                f"{p.get('pe'):.1f}x" if p.get('pe') is not None else "-",
                f"{p.get('roe'):.1f}%" if p.get('roe') is not None else "-",
                f"{p.get('margin_to_equity'):.2f}x" if p.get('margin_to_equity') is not None else "-",
                f"{p.get('pct_margin_rev'):.1f}%" if p.get('pct_margin_rev') is not None else "-",
                f"{p.get('pct_brokerage_rev'):.1f}%" if p.get('pct_brokerage_rev') is not None else "-",
                f"{p.get('pct_tudoanh_rev'):.1f}%" if p.get('pct_tudoanh_rev') is not None else "-",
            ])
        t_peer = Table(peer_rows, colWidths=[16*mm, 16*mm, 16*mm, 16*mm, 22*mm, 18*mm, 20*mm, 20*mm])
        peer_styles = [
            ('BACKGROUND', (0, 0), (-1, 0), HexColor("#1F4E78")), ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.4, grey), ('FONTNAME', (0, 0), (-1, -1), FONT_REG),
            ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD), ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ]
        if _pb_vals:
            peer_styles.append(('BACKGROUND', (0, 1), (-1, 1), HexColor("#FFF2CC")))
            peer_styles.append(('FONTNAME', (0, 1), (-1, 1), FONT_BOLD))
        for r_idx in range(len(peer_rows)):
            if peer_rows[r_idx][0].startswith("*"):
                peer_styles.append(('BACKGROUND', (0, r_idx), (-1, r_idx), HexColor("#DDEBF7")))
                peer_styles.append(('FONTNAME', (0, r_idx), (-1, r_idx), FONT_BOLD))
                peer_rows[r_idx][0] = peer_rows[r_idx][0].replace("*", "")
        t_peer.setStyle(TableStyle(peer_styles))
        story.append(t_peer)
        story.append(PageBreak())

    # ── Trang 7: Định giá chi tiết ──
    weights = weights or {"PB": 0.9, "PE": 0.1}
    story.append(Paragraph("PHƯƠNG PHÁP ĐỊNH GIÁ", style_h2))
    val_data = [
        ["Phương pháp", "Trọng số", "Giá mục tiêu (VND)"],
        ["P/B (ưu tiên số 1)", f"{weights['PB']*100:.0f}%", f"{pb_target:,.0f}"],
        ["P/E (bộ lọc chống nhiễu)", f"{weights['PE']*100:.0f}%", f"{pe_target:,.0f}"],
        ["GIÁ MỤC TIÊU BÌNH QUÂN", "100%", f"{target:,.0f}"],
        ["Bear Case", "-", f"{bear_target:,.0f}"],
        ["Bull Case", "-", f"{bull_target:,.0f}"],
    ]
    t3 = Table(val_data, colWidths=[70*mm, 30*mm, 45*mm])
    t3.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), FONT_REG), ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('FONTNAME', (0, 0), (-1, 0), FONT_BOLD), ('BACKGROUND', (0, 0), (-1, 0), HexColor("#1F4E78")),
        ('TEXTCOLOR', (0, 0), (-1, 0), white), ('GRID', (0, 0), (-1, -1), 0.4, grey),
        ('FONTNAME', (0, 3), (-1, 3), FONT_BOLD), ('BACKGROUND', (0, 3), (-1, 3), HexColor("#E2EFDA")),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(t3)
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        f"Chi phí vốn CSH (COE) theo mô hình CAPM: {coe*100:.2f}%. P/B được ưu tiên làm phương pháp định giá "
        "chính vì tài sản CTCK (tiền, chứng khoán FVTPL, cho vay margin) có tính thanh khoản cao, phản ánh "
        "sát giá trị sổ sách thực tế hơn so với các doanh nghiệp sản xuất. P/E chỉ dùng bổ trợ với trọng số "
        "thấp hơn do LNST của CTCK dễ biến động mạnh theo thị trường (đặc biệt khi tỷ trọng Tự doanh cao).", style_body))
    story.append(PageBreak())

    # ── Trang 8: Rủi ro & Kết luận ──
    story.append(Paragraph("RỦI RO CHÍNH & KẾT LUẬN", style_h2))
    risks = [
        "Rủi ro thị trường: Doanh thu Môi giới, Cho vay Margin và Tự doanh đều nhạy cảm với thanh khoản thị "
        "trường và biến động VN-Index — một đợt sụt giảm thị trường kéo dài sẽ tác động đồng thời cả 3 mảng.",
        "Rủi ro cạnh tranh: Cuộc đua phí môi giới 0% giữa các CTCK (đặc biệt nhóm có vốn ngoại) gây áp lực "
        "giảm biên lợi nhuận mảng môi giới truyền thống.",
        "Rủi ro lãi suất: Chi phí vốn tài trợ cho hoạt động cho vay margin và tự doanh tăng khi mặt bằng lãi "
        "suất huy động tăng, có thể thu hẹp NIM cho vay margin.",
        "Rủi ro pháp lý: Thay đổi quy định về tỷ lệ margin, giới hạn cho vay theo vốn điều lệ có thể hạn chế "
        "tăng trưởng dư nợ margin trong ngắn hạn.",
    ]
    for r in risks:
        story.append(Paragraph(f"• {r}", style_body))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        f"<b>KẾT LUẬN: {recommend}</b> — Giá mục tiêu {target:,.0f} VND, tiềm năng {upside_pct:+.1f}% so với "
        f"giá hiện tại {price:,.0f} VND.", style_body))

    doc.build(story)
    return pdf_path


# ══════════════════════════════════════════════════════════════════════════
# JSON EXPORT (schema chuẩn dashboard — skill xuat-bao-cao)
# ══════════════════════════════════════════════════════════════════════════
def save_json_summary(ticker, company_name, price, mcap, shares, years_hist, years_fc,
                       rev_hist, rev_fc, npat_hist, npat_fc, eps_hist, eps_fc, equity_hist, equity_fc,
                       quarter_labels, pe_quarters, pb_quarters, target, upside_pct, recommend,
                       bear_target, bull_target, pb_target, pe_target, pe_median, pb_median, coe,
                       seg_hist, seg_fc, seg_pct_fc0, concentration_seg, concentration_flag,
                       is_q=None, bs_q=None, latest_margin_leverage=0.0, latest_leverage_label="",
                       weights=None, quarterly_update=None, macro_liquidity=None):
    out_dir = os.path.join(PROJECT_ROOT, "data")
    os.makedirs(out_dir, exist_ok=True)
    all_years = years_hist + years_fc
    weights = weights or {"PB": 0.9, "PE": 0.1}

    # ── Chuỗi dữ liệu theo QUÝ cho web (LNST/ROE, đòn bẩy Margin, tương quan ROE-P/B) ──
    quarterly_series = {"labels": [], "revenue": [], "npat": [], "roe": [], "marginLeverage": []}
    roe_pb_pairs = []
    if is_q and bs_q:
        _q_keys = sorted(set((r.get("yearReport"), r.get("lengthReport")) for r in is_q
                              if r.get("yearReport") and r.get("lengthReport") in (1, 2, 3, 4)))[-12:]
        for y, q in _q_keys:
            rec_is = next((x for x in is_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
            rec_bs = next((x for x in bs_q if x.get("yearReport") == y and x.get("lengthReport") == q), {})
            rev_q = round((rec_is.get(IS_TOTAL["total_rev"]) or 0) / 1e9, 1)
            npat_q = round((rec_is.get(IS_TOTAL["npat_parent"]) or 0) / 1e9, 1)
            equity_q = (rec_bs.get(BS["equity"]) or 0) / 1e9
            margin_q = (rec_bs.get(BS["margin_loans"]) or 0) / 1e9
            roe_q = round(npat_q * 4 / equity_q * 100, 2) if equity_q else None
            lev_q = round(margin_q / equity_q, 2) if equity_q else None
            quarterly_series["labels"].append(f"{y}Q{q}")
            quarterly_series["revenue"].append(rev_q)
            quarterly_series["npat"].append(npat_q)
            quarterly_series["roe"].append(roe_q)
            quarterly_series["marginLeverage"].append(lev_q)
        for i, lbl in enumerate(quarter_labels):
            try:
                y_str, q_str = lbl.split("-Q")
                y, q = int(y_str), int(q_str)
            except (ValueError, IndexError):
                continue
            pb_v = pb_quarters[i] if i < len(pb_quarters) else None
            if not pb_v or pb_v <= 0:
                continue
            rec_is = next((x for x in is_q if x.get("yearReport") == y and x.get("lengthReport") == q), None)
            rec_bs = next((x for x in bs_q if x.get("yearReport") == y and x.get("lengthReport") == q), None)
            if not rec_is or not rec_bs:
                continue
            equity_q = (rec_bs.get(BS["equity"]) or 0) / 1e9
            npat_q = (rec_is.get(IS_TOTAL["npat_parent"]) or 0) / 1e9
            if not equity_q:
                continue
            roe_pb_pairs.append({"quarter": lbl, "roe": round(npat_q * 4 / equity_q * 100, 2), "pb": pb_v})

    dominant_label = SEGMENT_LABELS_VI.get(concentration_seg, concentration_seg)
    thesis = [
        f"{company_name} vận hành mô hình 5 mảng kinh doanh (Môi giới, Cho vay Margin, Tự doanh, IB+Lưu ký, "
        f"Quản lý quỹ), mảng đóng góp lớn nhất hiện tại là {dominant_label} ({seg_pct_fc0[concentration_seg]*100:.1f}% DT).",
        f"Định giá theo P/B (trung vị lịch sử {pb_median:.2f}x) kết hợp P/E (trung vị {pe_median:.2f}x) cho "
        f"giá mục tiêu {target:,.0f} VND, tiềm năng {upside_pct:+.1f}%.",
        "Cần theo dõi tỷ trọng mảng Tự doanh (FVTPL) — nguồn biến động LNST lớn nhất, nhạy cảm trực tiếp với VN-Index.",
    ]
    risks = [
        "Rủi ro thị trường: thanh khoản GTGD và VN-Index sụt giảm tác động đồng thời Môi giới/Margin/Tự doanh.",
        "Rủi ro cạnh tranh: cuộc đua phí môi giới 0% gây áp lực biên lợi nhuận mảng môi giới truyền thống.",
        "Rủi ro lãi suất: chi phí vốn tài trợ margin/tự doanh tăng khi lãi suất huy động tăng, thu hẹp NIM.",
    ]
    moats = {
        "Network Effect": {"score": 2, "desc": "Hạn chế — dịch vụ môi giới dễ thay thế giữa các CTCK."},
        "Cost Advantage": {"score": 3, "desc": "Quy mô vốn lớn giúp chi phí vốn tài trợ margin/tự doanh thấp hơn CTCK nhỏ."},
        "Switching Cost": {"score": 2, "desc": "Thấp — nhà đầu tư dễ chuyển tài khoản giữa các CTCK."},
        "Intangible Assets": {"score": 3, "desc": "Thương hiệu, đội ngũ phân tích/tư vấn uy tín thu hút khách hàng tổ chức."},
        "Efficient Scale": {"score": 3, "desc": "Vốn điều lệ lớn cho phép mở rộng room margin và danh mục tự doanh."},
    }
    pestle = [
        {"factor": "Political", "content": "Chính sách UBCKNN, lộ trình nâng hạng thị trường mới nổi (FTSE/MSCI).", "impact": "Positive"},
        {"factor": "Economic", "content": "Lãi suất, tăng trưởng GDP và thanh khoản thị trường là driver trực tiếp doanh thu.", "impact": "Positive"},
        {"factor": "Social", "content": "Số tài khoản chứng khoán mở mới tăng nhanh, phổ cập đầu tư cá nhân.", "impact": "Positive"},
        {"factor": "Technological", "content": "Hệ thống KRX mới, giao dịch T+0, sản phẩm phái sinh mở rộng.", "impact": "Positive"},
        {"factor": "Legal", "content": "Quy định tỷ lệ margin, phân loại nhà đầu tư chuyên nghiệp siết chặt hơn.", "impact": "Neutral"},
        {"factor": "Environmental", "content": "Xu hướng ESG trong lựa chọn danh mục của quỹ ngoại.", "impact": "Neutral"},
    ]
    comments = {
        "businessModel": f"{company_name} vận hành mô hình bán lẻ dịch vụ tài chính 5 mảng: Môi giới, Cho vay "
                          f"Margin, Tự doanh (FVTPL), IB+Lưu ký, Quản lý quỹ. Mảng {dominant_label} hiện đóng góp "
                          f"lớn nhất ({seg_pct_fc0[concentration_seg]*100:.1f}% doanh thu dự phóng).",
        "financialPerformance": f"Doanh thu hoạt động dự phóng {years_fc[0]}E đạt {rev_fc[0]:,.0f} tỷ VND, LNST cổ "
                                 f"đông mẹ {npat_fc[0]:,.0f} tỷ VND, EPS {eps_fc[0]:,.0f} VND/CP.",
        "valuationText": f"Giá mục tiêu {target:,.0f} VND (upside {upside_pct:+.1f}%) theo mô hình P/B {weights['PB']*100:.0f}% + P/E "
                          f"{weights['PE']*100:.0f}%. Khuyến nghị {recommend}.",
    }

    payload = {
        "ticker": ticker,
        "companyName": company_name,
        "sector": "Chứng khoán",
        "currentPrice": price,
        "marketCap": mcap * 1e9,
        "shares": shares,
        "gdriveExcelUrl": None,
        "gdrivePdfUrl": None,
        "data": {
            "years": all_years,
            "revenue": [round(v, 1) for v in rev_hist + rev_fc],
            "npat": [round(v, 1) for v in npat_hist + npat_fc],
            "eps": [round(v, 0) for v in eps_hist + eps_fc],
            "equity": [round(v, 1) for v in equity_hist + equity_fc],
        },
        "segments": {
            "names": SEGMENT_NAMES,
            "labels": SEGMENT_LABELS_VI,
            "revenueHist": {seg: [round(v, 1) for v in seg_hist[seg]] for seg in SEGMENT_NAMES},
            "revenueForecast": {seg: [round(v, 1) for v in seg_fc[seg]] for seg in SEGMENT_NAMES},
            "pctNow": {seg: round(seg_pct_fc0[seg], 4) for seg in SEGMENT_NAMES},
        },
        "quarterly": quarterly_series,
        "roePbCorrelation": roe_pb_pairs,
        "marginLeverage": {
            "latest": latest_margin_leverage,
            "latestLabel": latest_leverage_label,
            "legalCap": 2.0,
            "warningThreshold": 1.8,
        },
        "quarterlyUpdate": quarterly_update,
        "thesis": thesis,
        "risks": risks,
        "moats": moats,
        "pestle": pestle,
        "valuation": {
            "bear": bear_target,
            "base": target,
            "bull": bull_target,
            "methods": {"pb": pb_target, "pe": pe_target},
            "weights": {"PB": weights["PB"], "PE": weights["PE"]},
            "recommendation": recommend,
            "upsidePct": upside_pct,
            "coe": round(coe * 100, 2),
            "peMedian": pe_median,
            "pbMedian": pb_median,
        },
        "comments": comments,
        "pe_hist": [],
        "pb_hist": [],
        "pe_quarters": pe_quarters,
        "pb_quarters": pb_quarters,
        "quarter_labels": quarter_labels,
        "ratios": {
            "roe": [round(npat_hist[i] / equity_hist[i], 4) if equity_hist[i] else None for i in range(len(years_hist))]
                   + [round(npat_fc[i] / equity_fc[i], 4) if equity_fc[i] else None for i in range(len(years_fc))],
            "net_margin": [round(npat_hist[i] / rev_hist[i], 4) if rev_hist[i] else None for i in range(len(years_hist))]
                          + [round(npat_fc[i] / rev_fc[i], 4) if rev_fc[i] else None for i in range(len(years_fc))],
        },
        "macro_liquidity": macro_liquidity or {},
    }

    json_path = os.path.join(out_dir, f"{ticker}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[JSON] Saved: {json_path}")
