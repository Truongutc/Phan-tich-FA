#!/usr/bin/env python3
"""
template_vimo.py — Phân tích Vĩ mô Kinh tế Việt Nam (không gắn với mã cổ phiếu nào).

Dựa theo "Hướng Dẫn Phân Tích Vĩ Mô Kinh Tế Việt Nam & Ứng Dụng Đầu Tư Chứng Khoán" trong
Logic phan tich cac nganh/ — khung Top-Down 7 chương: theo dõi ~20 chỉ báo (tăng trưởng, lạm
phát, tiền tệ, thương mại/vốn, tài khóa, lao động) + áp lực bên ngoài (Fed, DXY, dầu, Trung
Quốc) + định giá thị trường (P/E, P/B, ERP VN-Index), tổng hợp thành Scorecard Vĩ Mô 6 nhóm
(-1/0/+1 mỗi nhóm, xem SCORECARD_GROUPS) và ma trận quyết định phân bổ vốn (Chương 6).

Dữ liệu THÔ nằm ở data/vimo_raw.json (được fetch_macro_data.py cập nhật) — file này CHỈ tính
toán/trình bày (trend, scorecard, ERP, ma trận quyết định, PDF, JSON dashboard), KHÔNG tự fetch
dữ liệu thô.

Excel lịch sử chỉ số theo thời gian: Bao cao/VIMO/VIMO_Lich_Su_Chi_So.xlsx (xem
update_excel_history_vimo()) — mỗi hàng 1 chỉ báo, mỗi cột 1 kỳ báo cáo, TÁCH RIÊNG SHEET theo tần
suất cập nhật (Theo năm/Theo quý/Theo tháng/Theo ngày/Lũy kế H1-9M-FY) để append cột mới không bao
giờ ảnh hưởng tới vị trí ô của sheet khác hay cột cũ (an toàn cho công thức/biểu đồ user tự gắn
vào file). Tự tích lũy qua từng lần Action chạy (KHÔNG phải Excel model định giá kiểu sector
template theo ticker — chỉ là log lịch sử số liệu thô).

Giai đoạn 1 (skeleton sơ bộ, theo chỉ đạo user "hoàn thiện khung trước, tinh chỉnh sau"): một số
chỉ báo (PMI, lãi suất liên ngân hàng, tín dụng, M2, dự trữ ngoại hối, dòng tiền khối ngoại/margin)
mới chỉ có seed dữ liệu thưa hoặc chưa có nguồn tự động — xem "auto_source" trong vimo_raw.json.
"""
import os
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import re
import json
import datetime
import subprocess
import requests

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
VIMO_RAW_PATH = os.path.join(PROJECT_ROOT, "data", "vimo_raw.json")


# ══════════════════════════════════════════════════════════════════════════
# FONTS (copy pattern từ template_nhietdien.py — Arial nếu có, else Helvetica)
# ══════════════════════════════════════════════════════════════════════════
def register_vn_fonts():
    font_paths_to_try = [
        ("C:/Windows/Fonts/arial.ttf", "Arial"),
        ("C:/Windows/Fonts/arialbd.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "Arial"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", "Arial-Bold"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Arial"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", "Arial-Bold"),
    ]
    found = {}
    for path, freg in font_paths_to_try:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(freg, path))
                found[freg] = path
            except Exception:
                pass
    return found


_VN_FONTS = register_vn_fonts()
FONT_REG = 'Arial' if 'Arial' in _VN_FONTS else 'Helvetica'
FONT_BOLD = 'Arial-Bold' if 'Arial-Bold' in _VN_FONTS else 'Helvetica-Bold'

UA_STR = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
          "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")


# ══════════════════════════════════════════════════════════════════════════
# Rf (copy từ template_nhietdien.py — cần cho ERP = E/P - Rf)
# ══════════════════════════════════════════════════════════════════════════
def fetch_via_curl(url, timeout=10, label=None):
    tag = f" [{label}]" if label else ""
    try:
        r = subprocess.run(
            ["curl", "-sL", "-A", UA_STR, "--max-time", str(timeout), "-w", "\n__HTTP_STATUS__:%{http_code}", url],
            capture_output=True, text=True, encoding="utf-8", errors="ignore", timeout=timeout + 5,
        )
        out = r.stdout
        status = None
        marker = "\n__HTTP_STATUS__:"
        if marker in out:
            out, status_str = out.rsplit(marker, 1)
            status = status_str.strip()
        if r.returncode != 0 or not out.strip() or status not in (None, "200"):
            print(f"  [DIAG]{tag} fetch_via_curl weak/empty: curl_exit={r.returncode} http_status={status} url={url[:90]}")
        return out if r.returncode == 0 else ""
    except Exception as e:
        print(f"  [DIAG]{tag} fetch_via_curl exception: {url[:90]} -> {e}")
        return ""


def fetch_rf_vietnam(timeout=15):
    FALLBACK_RF = 0.045
    try:
        html = fetch_via_curl("https://vn.investing.com/rates-bonds/vietnam-10-year-bond-yield", timeout=timeout, label="investing-rf")
        if html:
            m = re.search(r'data-test="instrument-price-last"[^>]*>([\d.,]+)', html)
            if m:
                rf = float(m.group(1).replace(",", "")) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "investing.com"
    except Exception as e:
        print(f"  [WARN] investing.com Rf failed: {e}")
    try:
        r = requests.get("https://www.worldgovernmentbonds.com/bond-yield/vietnam/10-years/",
                          headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
        if r.status_code == 200:
            matches = re.findall(r'(\d+\.\d+)%', r.text[:5000])
            if matches:
                rf = float(matches[0]) / 100
                if 0.01 <= rf <= 0.15:
                    return rf, "worldgovernmentbonds.com"
    except Exception as e:
        print(f"  [WARN] WorldGovernmentBonds Rf failed: {e}")
    return FALLBACK_RF, "Fallback (manual)"


# ══════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ══════════════════════════════════════════════════════════════════════════
def load_vimo_raw():
    with open(VIMO_RAW_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_vimo_raw(raw):
    with open(VIMO_RAW_PATH, "w", encoding="utf-8") as f:
        json.dump(raw, f, ensure_ascii=False, indent=2)


GROUP_LABELS = {
    "growth": "Tăng trưởng",
    "inflation": "Lạm phát",
    "monetary": "Tiền tệ & Lãi suất",
    "trade": "Thương mại & Vốn",
    "fiscal": "Tài khóa",
    "labor": "Lao động",
    "external": "Áp lực bên ngoài",
    "market": "Thị trường chứng khoán",
    "intl_us": "Quốc tế — Mỹ",
    "intl_eu": "Quốc tế — Châu Âu",
    "intl_uk": "Quốc tế — Anh",
    "intl_jp": "Quốc tế — Nhật Bản",
    "intl_cn": "Quốc tế — Trung Quốc",
}

# Ánh xạ 8 nhóm dữ liệu (vimo_raw.json) -> 6 nhóm Scorecard Vĩ Mô theo đúng Chương 4.1/6.1 tài
# liệu hướng dẫn (KHÔNG trùng 1-1 với 8 nhóm dữ liệu — vd tỷ giá thuộc "monetary" trong dữ liệu
# nhưng thuộc "Tỷ giá & Dòng vốn ngoại" trong scorecard, theo đúng cách tài liệu phân loại).
#
# Lịch sử thay đổi (theo phản hồi user, xem trao đổi 2026-07-13):
# - Bỏ nhóm "Tâm lý thị trường" (trước đây chỉ có vnindex_pe, luôn 0 phiếu) — đây là tín hiệu ĐỊNH
#   GIÁ, không phải "tâm lý" theo nghĩa phân tích kỹ thuật, và đã có chỗ riêng ở marketValuation.
# - Tách "Bên ngoài" (8 chỉ báo khác bản chất, gộp thành 1 phiếu bầu trung bình dễ thiên kiến)
#   thành 2 nhóm độc lập theo cơ chế tác động: (1) tỷ giá/dòng vốn ngoại, (2) thương mại/hàng hóa.
# - Tách tiếp "Lạm phát & Lãi suất" thành 2 nhóm riêng "Lạm phát" / "Lãi suất" (user: "lạm phát và
#   lãi suất nên cho ra 2 tiêu chí khác nhau, gộp vào thì không đánh giá được tổng thể") — Scorecard
#   giờ có 6 nhóm (không còn 5) — xem vimo.html/index.html đã cập nhật lại text "5 nhóm" -> "6 nhóm".
# - Thêm deposit_rate_12m_market_avg (lãi suất huy động BÌNH QUÂN THỊ TRƯỜNG ~38 NH qua 24hmoney,
#   xem fetch_market_deposit_rate_12m() trong fetch_macro_data.py) vào nhóm "Lãi suất" — trước đây
#   nhóm gộp chỉ có cpi_yoy thực sự có phiếu (core_inflation/refinancing_rate/interbank_rate_3m
#   hiếm khi đủ dữ liệu tính trend), nên KHÔNG phản ánh áp lực lãi suất thị trường THẬT (~6-9%,
#   cao hơn nhiều lãi suất điều hành/liên ngân hàng ổn định thấp).
SCORECARD_GROUPS = {
    "Tăng trưởng": ["gdp_growth", "iip_growth", "pmi_manufacturing", "retail_sales_growth", "unemployment_rate"],
    "Lạm phát": ["cpi_yoy", "core_inflation"],
    "Lãi suất": ["refinancing_rate", "interbank_rate_3m", "deposit_rate_12m_market_avg"],
    # credit_growth ĐÃ BỎ khỏi nhóm này (2026-07-24, theo phản hồi user): tín dụng tăng KHÔNG phải
    # bằng chứng thanh khoản dồi dào — ngược lại, tín dụng tăng nhanh hơn huy động là NGUYÊN NHÂN
    # gây căng thanh khoản (xem note/impact của credit_growth). Đếm "tín dụng tăng = +1 tốt" cho
    # nhóm Thanh khoản là sai chiều, che mất tín hiệu căng thẳng thật (đường cong liên ngân hàng
    # dốc lên, NHNN chỉ bơm không hút tín phiếu suốt nhiều tháng). credit_growth vẫn hiển thị đầy
    # đủ trong báo cáo, chỉ không góp phiếu ở đây nữa (giống nhiều chỉ báo tham chiếu khác).
    "Thanh khoản": ["m2_growth", "forex_reserves", "omo_rate_7d"],
    "Tỷ giá & Dòng vốn ngoại": ["usdvnd", "dxy_proxy", "fed_funds_rate", "fdi_disbursed",
                                 "fdi_registered_usd_bn", "fii_net_flow_hose"],
    "Thương mại & Hàng hóa": ["trade_balance", "export_growth", "brent_oil", "china_gdp_growth"],
}

# Tên NGẮN dùng riêng cho dòng "lý do cốt lõi" của Scorecard (user 2026-07-13: muốn nêu thẳng tên
# chỉ báo + nhận định, kiểu "GDP tốt lên, PMI xấu đi", KHÔNG phải đếm số phiếu) — khác với "label"
# đầy đủ dùng ở các chỗ khác (bảng PDF, thẻ chỉ báo web).
SHORT_LABEL = {
    "gdp_growth": "GDP", "iip_growth": "IIP", "pmi_manufacturing": "PMI",
    "retail_sales_growth": "Bán lẻ", "unemployment_rate": "Thất nghiệp",
    "cpi_yoy": "CPI", "core_inflation": "Lạm phát lõi",
    "refinancing_rate": "LS điều hành", "interbank_rate_3m": "LS liên NH 3T",
    "deposit_rate_12m_market_avg": "LS huy động thị trường",
    "credit_growth": "Tín dụng", "m2_growth": "Cung tiền M2", "forex_reserves": "Dự trữ ngoại hối",
    "omo_rate_7d": "OMO",
    "usdvnd": "Tỷ giá", "dxy_proxy": "DXY", "fed_funds_rate": "Lãi suất Fed", "fdi_disbursed": "FDI giải ngân",
    "fdi_registered_usd_bn": "FDI đăng ký", "fii_net_flow_hose": "Khối ngoại HOSE",
    "trade_balance": "Cán cân TM", "export_growth": "Xuất khẩu", "brent_oil": "Dầu Brent",
    "china_gdp_growth": "GDP Trung Quốc",
}

# Ngưỡng mục tiêu lạm phát CHÍNH THỨC (Quốc hội/Chính phủ đặt cho giai đoạn hiện tại — hiện tại:
# Nghị quyết Quốc hội 2026 đặt mục tiêu CPI bình quân dưới 4,5%). CẬP NHẬT hằng số này khi có mục
# tiêu năm mới. Dùng để chấm phiếu THEO MỨC (level), ĐỘC LẬP với phiếu THEO XU HƯỚNG (calc_trend)
# đã có — theo đúng yêu cầu user (2026-07-13): "cao hơn mục tiêu chính phủ = không tốt; nếu CPI
# đang TĂNG (và vượt mục tiêu) = cực xấu; nếu đang GIẢM (nhưng vẫn vượt mục tiêu) = hạ nhiệt, chưa
# thể coi là tốt". Cơ chế 2-phiếu-độc-lập-cộng-trung-bình xử lý đúng cả 2 tình huống: vượt mục
# tiêu + đang tăng -> phiếu mức (-1) VÀ phiếu xu hướng (-1) -> điểm nhóm -1 (cực xấu, đúng ý);
# vượt mục tiêu + đang giảm -> phiếu mức (-1) VÀ phiếu xu hướng (+1) -> trung bình 0 (hạ nhiệt,
# không phải "tốt" dù đang cải thiện) — KHÔNG để 1 điểm dữ liệu (chỉ xu hướng) quyết định cả nhóm.
CPI_TARGET_CEILING = 4.5


def _cpi_level_vote(cpi_value):
    if cpi_value is None:
        return None
    return -1 if cpi_value > CPI_TARGET_CEILING else 1


# Chỉ báo nào có NGƯỠNG MỤC TIÊU CHÍNH THỨC được định nghĩa thì mới có thêm phiếu mức (level) —
# hiện chỉ cpi_yoy (có mục tiêu Quốc hội rõ ràng).
LEVEL_VOTE_FUNCS = {"cpi_yoy": _cpi_level_vote}


# CHÊNH LỆCH lãi suất huy động THỰC TẾ (ngoài khung cho phép — deposit_rate_negotiated_max) so với
# lãi suất NIÊM YẾT Big4 (deposit_rate_12m_vcb) — theo đúng cơ chế user mô tả (2026-07-13): "nếu
# lãi suất huy động thực tế bên ngoài cao hơn trên 2% so với lãi suất niêm yết là cao, tức là bank
# đang quá đói vốn và cần huy động cao hơn khung cho phép bằng nhiều biện pháp để có tiền". Đây LÀ
# phiếu MỨC riêng cho nhóm "Lãi suất" — ĐỘC LẬP với xu hướng tăng/giảm của từng chỉ báo niêm yết
# (lãi suất niêm yết đứng yên/giảm nhẹ KHÔNG có nghĩa hệ thống đang khỏe nếu khoảng cách với thực
# tế bên ngoài vẫn giãn rộng). Ngưỡng 2 điểm % là do user chọn — không tự suy diễn số khác.
DEPOSIT_RATE_GAP_CEILING = 2.0


def _deposit_rate_gap_level_vote(raw, trends):
    actual = trends.get("deposit_rate_negotiated_max", {}).get("latest")
    listed = trends.get("deposit_rate_12m_vcb", {}).get("latest")
    if actual is None or listed is None:
        return None
    gap = actual - listed
    vote = -1 if gap > DEPOSIT_RATE_GAP_CEILING else 1
    label = f"Chênh lệch lãi suất thực tế/niêm yết ({gap:+.2f} điểm %)"
    judgment_label = "đói vốn, huy động vượt khung cho phép" if vote < 0 else "huy động trong khung an toàn"
    return {"indicator": "deposit_rate_gap", "label": label, "vote": vote,
            "arrow": "⚠" if vote < 0 else "✓", "judgment_label": judgment_label}


# So lãi suất LIÊN NGÂN HÀNG (interbank_rate_9m — kỳ hạn dài nhất đang theo dõi, gần 12 tháng
# nhất) với lãi suất HUY ĐỘNG NIÊM YẾT Big4 cùng kỳ hạn gần nhất (deposit_rate_12m_vcb) — theo yêu
# cầu user (2026-07-13): "nếu liên ngân hàng cao hơn huy động niêm yết cùng kỳ hạn thì cũng là
# căng thanh khoản và áp lực lãi suất còn lớn". ĐỘC LẬP với _deposit_rate_gap_level_vote ở trên
# (dùng lãi suất thỏa thuận, dữ liệu THƯA vì chỉ có khi báo chí đưa tin) — chỉ báo này dùng dữ liệu
# CÓ MỖI LẦN ACTION CHẠY (không phụ thuộc tin tức), nên đáng tin cậy/ổn định hơn để xác nhận chéo.
def _interbank_vs_deposit_level_vote(raw, trends):
    interbank = trends.get("interbank_rate_9m", {}).get("latest")
    listed = trends.get("deposit_rate_12m_vcb", {}).get("latest")
    if interbank is None or listed is None:
        return None
    vote = -1 if interbank > listed else 1
    label = f"Liên ngân hàng 9 tháng ({interbank:.2f}%) so với huy động niêm yết 12 tháng ({listed:.2f}%)"
    judgment_label = ("cao hơn niêm yết, căng thanh khoản" if vote < 0
                       else "thấp hơn niêm yết, thanh khoản ổn")
    return {"indicator": "interbank_vs_deposit", "label": label, "vote": vote,
            "arrow": "⚠" if vote < 0 else "✓", "judgment_label": judgment_label}


# So lãi suất iPower TCBS (mức CAO NHẤT quảng cáo — kênh 'gửi tiền' thay thế ngoài ngân hàng
# truyền thống) với huy động niêm yết Big4 cùng kỳ hạn gần nhất — user (2026-07-24): "khi thanh
# khoản căng cứng, lãi suất huy động lên cao thì chính sách tiền gửi của TCBS cũng tăng theo...
# khi nào TCBS hạ mức này thì đồng nghĩa áp lực lãi suất bên ngoài đã qua đi". Đáng tin cậy hơn
# deposit_rate_negotiated_max (tin tức, thưa) vì TCBS công bố công khai + có ngày hiệu lực rõ,
# lấy được MỖI LẦN Action chạy (dùng chung ngưỡng DEPOSIT_RATE_GAP_CEILING với gap huy động thật).
def _tcbs_ipower_gap_level_vote(raw, trends):
    tcbs = trends.get("deposit_rate_tcbs_ipower_max", {}).get("latest")
    listed = trends.get("deposit_rate_12m_vcb", {}).get("latest")
    if tcbs is None or listed is None:
        return None
    gap = tcbs - listed
    vote = -1 if gap > DEPOSIT_RATE_GAP_CEILING else 1
    label = f"Lãi suất iPower TCBS ({tcbs:.2f}%) so với huy động niêm yết 12 tháng ({listed:.2f}%)"
    judgment_label = ("cao hơn niêm yết nhiều, thị trường đói vốn" if vote < 0
                       else "chênh lệch trong khung an toàn")
    return {"indicator": "tcbs_ipower_gap", "label": label, "vote": vote,
            "arrow": "⚠" if vote < 0 else "✓", "judgment_label": judgment_label}


# Các "phiếu mức" cần dữ liệu TỪ NHIỀU CHỈ BÁO cùng lúc (không chỉ 1 chỉ báo tự so với ngưỡng của
# chính nó như LEVEL_VOTE_FUNCS) — gắn theo TÊN NHÓM Scorecard (LIST vì 1 nhóm có thể có nhiều
# phép so sánh độc lập), gọi hết trong calc_scorecard. Gắn vào "Thanh khoản" (KHÔNG PHẢI "Lãi
# suất" như trước 2026-07-24): 2 phép so sánh này đo TRỰC TIẾP mức độ căng thẳng thanh khoản hệ
# thống (chênh lệch lãi suất huy động thực tế/liên ngân hàng so với niêm yết) — đúng bản chất
# nhóm "Thanh khoản" hơn là "Lãi suất" (nhóm này giờ chỉ còn phản ánh MỨC lãi suất chính thức/
# thị trường, không lẫn tín hiệu căng thẳng). User (2026-07-24) chỉ ra nhóm Thanh khoản trước đó
# báo "+1 Tốt" trong khi đường cong liên ngân hàng dốc lên và NHNN chỉ bơm không hút — sửa cả
# việc phân nhóm chỉ báo (ở đây) VÀ việc credit_growth bị tính sai chiều (xem SCORECARD_GROUPS).
GROUP_LEVEL_CHECKS = {"Thanh khoản": [_deposit_rate_gap_level_vote, _interbank_vs_deposit_level_vote,
                                        _tcbs_ipower_gap_level_vote]}
SHORT_LABEL["deposit_rate_gap"] = "Ngân hàng"
SHORT_LABEL["interbank_vs_deposit"] = "Liên ngân hàng"
SHORT_LABEL["tcbs_ipower_gap"] = "iPower TCBS"


# ══════════════════════════════════════════════════════════════════════════
# TREND — so latest vs kỳ liền trước + vs kỳ xa hơn (tối đa 4 kỳ trước) theo hướng "tốt" riêng
# của từng chỉ báo.
# ══════════════════════════════════════════════════════════════════════════
def calc_trend(series, good_direction):
    """series: list [{period, value, ...}] đã sort theo thời gian tăng dần (đúng thứ tự lưu
    trong vimo_raw.json). Trả dict: latest, prior, delta, pct_of_prior, label ('up'/'down'/
    'flat'), arrow, is_improving (theo good_direction), n_points."""
    valid = [p for p in series if p.get("value") is not None]
    if not valid:
        return {"latest": None, "n_points": 0, "label": "no_data", "arrow": "—"}
    latest = valid[-1]
    n = len(valid)
    result = {
        "latest": latest["value"], "latest_period": latest["period"],
        "latest_source": latest.get("source_url"), "n_points": n,
    }
    if n < 2:
        result.update({"label": "insufficient", "arrow": "—", "is_improving": None})
        return result

    prior = valid[-2]
    delta = latest["value"] - prior["value"]
    result["prior"] = prior["value"]
    result["prior_period"] = prior["period"]
    result["delta"] = round(delta, 4)

    # So thêm với kỳ xa hơn (tối đa 4 kỳ trước) nếu đủ dữ liệu — cho cái nhìn xu hướng dài hơn
    if n >= 5:
        far = valid[-5]
        result["delta_vs_4back"] = round(latest["value"] - far["value"], 4)
        result["far_period"] = far["period"]

    threshold = abs(prior["value"]) * 0.003 if prior["value"] else 0.01  # ngưỡng "flat" ~0.3%
    if abs(delta) < max(threshold, 1e-6):
        direction = "flat"
    elif delta > 0:
        direction = "up"
    else:
        direction = "down"
    result["direction"] = direction
    # value_arrow = CHIỀU THẬT của số liệu (↑ nghĩa đen là tăng, ↓ là giảm) — KHÔNG mang nghĩa
    # tốt/xấu, để tránh nhầm lẫn (vd Nợ công/GDP giảm là điều TỐT nhưng số liệu vẫn đi xuống).
    result["value_arrow"] = {"up": "↑", "down": "↓", "flat": "→"}[direction]

    if direction == "flat":
        is_improving = None
    elif good_direction == "higher":
        is_improving = direction == "up"
    else:  # "lower"
        is_improving = direction == "down"
    result["is_improving"] = is_improving

    # judgment_* = ĐÁNH GIÁ tốt lên/xấu đi (theo đúng yêu cầu user) — TÁCH RIÊNG khỏi value_arrow
    # ở trên để không lẫn "chiều số liệu" với "ý nghĩa tốt/xấu" của chiều đó.
    if is_improving is None:
        result["label"] = "flat"
        result["arrow"] = "→"
        result["judgment_label"] = "Ổn định"
        result["judgment_color"] = "#f59e0b"
    elif is_improving:
        result["label"] = "improving"
        result["arrow"] = "▲"
        result["judgment_label"] = "Tốt lên"
        result["judgment_color"] = "#10b981"
    else:
        result["label"] = "worsening"
        result["arrow"] = "▼"
        result["judgment_label"] = "Xấu đi"
        result["judgment_color"] = "#ef4444"
    return result


# ══════════════════════════════════════════════════════════════════════════
# SCORECARD VĨ MÔ — 6 nhóm, mỗi nhóm -1/0/+1 (Chương 4.1/6.1, đã tách Lạm phát/Lãi suất riêng)
# ══════════════════════════════════════════════════════════════════════════
def calc_scorecard(raw, trends):
    """trends: {indicator_key: calc_trend(...) result}. Trả {group_name: {score, detail, n}}.
    Mỗi chỉ báo góp 1 PHIẾU XU HƯỚNG (như cũ); các chỉ báo có ngưỡng mục tiêu chính thức
    (LEVEL_VOTE_FUNCS) góp THÊM 1 PHIẾU MỨC độc lập — 1 chỉ báo có thể góp 2 phiếu vào cùng 1
    nhóm, nên n_votes đếm SỐ PHIẾU chứ không còn nhất thiết = số chỉ báo."""
    scorecard = {}
    for group_name, keys in SCORECARD_GROUPS.items():
        votes = []
        detail = []
        for k in keys:
            t = trends.get(k)
            if t and t.get("is_improving") is not None:
                votes.append(1 if t["is_improving"] else -1)
                detail.append({"indicator": k, "label": raw[k]["label"], "vote": votes[-1],
                                "arrow": t["arrow"], "judgment_label": t["judgment_label"], "kind": "trend"})
            level_func = LEVEL_VOTE_FUNCS.get(k)
            if level_func and t and t.get("latest") is not None:
                level_vote = level_func(t["latest"])
                if level_vote is not None:
                    votes.append(level_vote)
                    detail.append({
                        "indicator": k, "label": f"{raw[k]['label']} (so với mục tiêu)",
                        "vote": level_vote, "arrow": "⚠" if level_vote < 0 else "✓",
                        "judgment_label": "Vượt mục tiêu" if level_vote < 0 else "Trong mục tiêu",
                        "kind": "level",
                    })
        for group_level_check in GROUP_LEVEL_CHECKS.get(group_name, []):
            group_vote = group_level_check(raw, trends)
            if group_vote:
                group_vote.setdefault("kind", "level")
                votes.append(group_vote["vote"])
                detail.append(group_vote)
        if not votes:
            score = 0
        else:
            avg = sum(votes) / len(votes)
            score = 1 if avg > 0.2 else (-1 if avg < -0.2 else 0)
            # CHẶN "Tốt" giả: nếu có BẤT KỲ phiếu MỨC nào cho thấy đang VI PHẠM ngưỡng mục tiêu
            # chính thức (vd CPI vượt 4,5%, huy động vượt khung cho phép), nhóm KHÔNG được lên
            # +1 "Tốt" dù trung bình xu hướng dương — "đang cải thiện từ mức xấu" KHÁC "đang ở mức
            # tốt" (user 2026-07-24: "CPI vượt ngưỡng... đang tốt dần lên thôi" không nên đọc thành
            # "được kiểm soát trong ngưỡng rồi"). Chỉ HẠ điểm, không có chiều ngược lại (phiếu mức
            # tốt không ép nhóm phải dương nếu các chỉ báo khác đang xấu đi thật).
            if score == 1 and any(d["vote"] == -1 and d.get("kind") == "level" for d in detail):
                score = 0
        scorecard[group_name] = {"score": score, "detail": detail, "n_votes": len(votes),
                                  "reason": _scorecard_group_reason(detail, raw, trends)}
    total = sum(g["score"] for g in scorecard.values())
    return scorecard, total


def _scorecard_group_reason(detail, raw, trends):
    """1 dòng NGẮN GỌN NHẤT nêu LUẬN ĐIỂM CỐT LÕI đứng sau điểm nhóm — user (2026-07-13): không
    muốn đếm phiếu ("4/5 phiếu tốt lên"), muốn biết THẲNG chỉ báo nào tốt/xấu (vd "GDP tốt lên,
    PMI xấu đi"). Với chỉ báo có CẢ phiếu mức lẫn phiếu xu hướng (vd CPI: vượt/trong mục tiêu VÀ
    đang tăng/giảm), ghép thành 1 câu đủ nghĩa (vd "CPI vượt mục tiêu nhưng đang giảm") thay vì
    liệt kê 2 dòng rời rạc cho cùng 1 chỉ báo."""
    if not detail:
        return "Chưa đủ dữ liệu để chấm điểm."

    by_indicator = {}
    for d in detail:
        by_indicator.setdefault(d["indicator"], []).append(d)

    _DIRECTION_WORD = {"up": "đang tăng", "down": "đang giảm", "flat": "đi ngang"}
    phrases = []
    for ind_key, votes in by_indicator.items():
        # deposit_rate_gap (từ GROUP_LEVEL_CHECKS) không phải key thật trong raw — dùng label có
        # sẵn trong chính detail làm fallback thay vì tra raw[ind_key] (sẽ KeyError).
        short = SHORT_LABEL.get(ind_key) or (raw[ind_key]["label"] if ind_key in raw else votes[0]["label"])
        level_v = next((v for v in votes if "mục tiêu" in v["judgment_label"]), None)
        trend_v = next((v for v in votes if v is not level_v), None)
        if level_v and trend_v:
            direction = trends.get(ind_key, {}).get("direction")
            trend_txt = _DIRECTION_WORD.get(direction, trend_v["judgment_label"].lower())
            level_txt = level_v["judgment_label"].lower()
            joiner = "và" if level_v["vote"] == trend_v["vote"] else "nhưng"
            phrases.append(f"{short} {level_txt} {joiner} {trend_txt}")
        else:
            v = votes[0]
            phrases.append(f"{short} {v['judgment_label'].lower()}")
    return "; ".join(phrases) + "."


# ══════════════════════════════════════════════════════════════════════════
# ĐỊNH GIÁ THỊ TRƯỜNG — ERP = E/P (từ VN-Index P/E) - Rf (Chương 5.2/6.1)
# ══════════════════════════════════════════════════════════════════════════
# Nhãn "risk_compensation['label']" (khung 4 bậc, có so với lãi suất huy động THỊ TRƯỜNG THẬT,
# xem _calc_risk_compensation) -> valuation_label (nhãn hiển thị đầu báo cáo + đưa vào Ma trận
# quyết định). TRƯỚC 2026-07-25: valuation_label tự tính riêng bằng ngưỡng ERP>3%/<1% (chỉ so
# Rf TPCP LÝ THUYẾT) — độc lập với risk_compensation nên có thể ra 2 kết luận MÂU THUẪN NHAU
# trong cùng báo cáo (vd ERP=3.51%>3% -> "Rẻ/Hấp dẫn" nhưng risk_compensation cùng lúc kết luận
# "Bù đắp tối thiểu, CHƯA có biên an toàn" vì so với lãi suất gửi tiết kiệm thực tế thì chưa đạt
# 1,5 lần). User (2026-07-25) chỉ thẳng: "rẻ thì rẻ thật nhưng hấp dẫn ở đâu?" — hợp nhất về 1
# nguồn sự thật duy nhất (risk_compensation, khung nghiêm ngặt hơn vì có thêm mốc so sánh thực
# tế), tránh đọc 2 kết luận khác nhau cho cùng 1 con số ERP.
_RISK_COMP_TO_VALUATION_LABEL = {
    "Bù đắp tốt": "Rẻ/Hấp dẫn",
    "Bù đắp khá (so TPCP), chưa đủ so với gửi tiết kiệm thực tế": "Hợp lý",
    "Bù đắp tối thiểu, chưa có biên an toàn": "Hợp lý",
    "KHÔNG đủ bù đắp rủi ro": "Đắt/Kém hấp dẫn",
}


def _calc_valuation_from_pepb(raw, rf, pe, pb, pe_source, pb_source):
    """Lõi tính toán định giá (ERP/risk_compensation/CAPM) DÙNG CHUNG cho cả headline (có VIN) VÀ
    ex-VIN — tách riêng để calc_market_valuation() gọi 2 lần với 2 cặp P/E-P/B khác nhau (user
    2026-07-25: '2 quyết định... nếu nhìn vào VN-Index thì sao, nếu nhìn theo VN-Index no VIN
    thì sao'), tránh lặp code."""
    erp = None
    earnings_yield = None
    valuation_label = "Không xác định"
    risk_compensation = None
    if pe and pe > 0:
        earnings_yield = 1 / pe
        erp = earnings_yield - rf
        risk_compensation = _calc_risk_compensation(raw, rf, earnings_yield)
        valuation_label = _RISK_COMP_TO_VALUATION_LABEL.get(risk_compensation["label"], "Hợp lý")
    return {
        "pe": pe, "pb": pb, "rf": rf, "erp": erp, "valuation_label": valuation_label,
        "pe_source": pe_source, "pb_source": pb_source,
        "risk_compensation": risk_compensation,
        "capm_valuation": _calc_capm_valuation(pe, pb, rf),
    }


def calc_market_valuation(raw, rf):
    # ERP dùng P/E (Earnings Yield - Rf) làm căn cứ chính, KHÔNG đổi công thức/nhãn định giá hiện
    # có (tránh phình phạm vi quyết định phân bổ vốn). P/B (2026-07: nâng cấp lại — nguồn 24hmoney.
    # vn/indices/vn-index cấp CHỈ SỐ, xem fetch_vnindex_pe_pb_24hmoney trong fetch_macro_data.py)
    # chỉ hiển thị THÊM để đối chiếu chéo, không tính vào ERP/quyết định.
    # NÂNG CẤP 2026-07-25: pe/pb ở đây là ex-VIN (nguồn chính, xem vnindex_pe/vnindex_pb trong
    # vimo_raw.json) — dùng cho quyết định "chính". calc_market_valuation_headline() bên dưới
    # tính SONG SONG bằng cặp headline (có VIN) để đối chiếu, không thay thế cái này.
    pe_series = raw["vnindex_pe"]["series"]
    pe = pe_series[-1]["value"] if pe_series else None
    pb_series = raw.get("vnindex_pb", {}).get("series", [])
    pb = pb_series[-1]["value"] if pb_series else None
    return _calc_valuation_from_pepb(
        raw, rf, pe, pb,
        pe_series[-1]["source_url"] if pe_series else None,
        pb_series[-1]["source_url"] if pb_series else None)


def calc_market_valuation_headline(raw, rf):
    """Y HỆT calc_market_valuation() nhưng dùng cặp P/E-P/B HEADLINE (có VIN — vnindex_pe_headline/
    vnindex_pb_headline) thay vì ex-VIN — cho "quyết định thứ 2" đối chiếu song song (user
    2026-07-25). Trả None nếu chưa có dữ liệu headline (vd lần chạy đầu trước khi nâng cấp)."""
    pe_series = raw.get("vnindex_pe_headline", {}).get("series", [])
    pe = pe_series[-1]["value"] if pe_series else None
    pb_series = raw.get("vnindex_pb_headline", {}).get("series", [])
    pb = pb_series[-1]["value"] if pb_series else None
    if pe is None:
        return None
    return _calc_valuation_from_pepb(
        raw, rf, pe, pb,
        pe_series[-1]["source_url"] if pe_series else None,
        pb_series[-1]["source_url"] if pb_series else None)


# ROE VN-Index = P/B ÷ P/E (đồng nhất thức: P/B = P/E × ROE => ROE = P/B/P/E) so với CHI PHÍ VỐN
# CHỦ SỞ HỮU (COE) theo CAPM — user (2026-07-25) yêu cầu, dùng β=1 (định giá CẢ INDEX, không phải
# 1 mã riêng lẻ, nên so với chính nó β=1) và ERP Việt Nam ~7% (mức thường thấy trong các mô hình
# CAPM cho thị trường frontier, KHÁC ERP=E/Y-Rf đang dùng cho risk_compensation ở trên — 2 khái
# niệm khác nhau, xem ERP_COUNTRY_VIETNAM). KHÔNG cộng thêm phần bù rủi ro "đặc thù/Frontier" (α)
# như mô hình định giá 1 cổ phiếu đơn lẻ (vd MWG) — theo thảo luận với user: α đó dùng để bù rủi ro
# RIÊNG của 1 mã/ngành, đã bị triệt tiêu khi phân tán qua cả trăm mã trong 1 INDEX; rủi ro QUỐC GIA
# (Frontier) không mất đi khi phân tán nhưng ĐÃ nằm sẵn trong ERP_COUNTRY_VIETNAM=7% (đó chính là lý
# do ERP Việt Nam cao hơn thị trường phát triển ~4.5-5%) — cộng thêm α nữa sẽ tính trùng rủi ro quốc
# gia 2 lần. Justified P/B dùng công thức Residual Income CÓ TĂNG TRƯỞNG (Gordon Growth):
# P/B hợp lý = (ROE-g)/(COE-g). NÂNG CẤP 2026-07-25: bản đầu dùng g=0 (không tăng trưởng) cho ra
# P/B hợp lý ~1.35x — user kiểm chứng bằng dữ liệu P/B lịch sử THẬT (Vietcap IQ, 2009-2026) và chỉ
# ra mức này THẤP HƠN CẢ ĐÁY của 3 đợt khủng hoảng gần nhất (COVID 3/2020: 1.51x; bear 2022: 1.47x;
# sốc thuế quan 4/2025: 1.41x — thấp nhất từ 2015) — nghĩa là khung g=0 không bao giờ đạt được trong
# thực tế 10 năm qua, VÔ DỤNG để phân biệt rẻ/đắt. Nguyên nhân: bỏ qua tăng trưởng thật của ROE Việt
# Nam (tái đầu tư vào nền kinh tế đang tăng trưởng) — công thức (ROE-g)/(COE-g) rất nhạy với g khi
# (COE-g) nhỏ. GROWTH_VNINDEX_LONG_TERM=5% (mức tăng trưởng dài hạn user chọn) đưa P/B hợp lý lên
# ~1.61x — vẫn cho thấy đang đắt hơn (~19% so với 1.92x thực tế) nhưng ở mức HỢP LÝ hơn nhiều so với
# vùng P/B lịch sử (đáy khủng hoảng 1.41-1.51x, trung vị 17 năm ~1.94x). VẪN CHỈ HIỂN THỊ THAM KHẢO
# (không đưa vào valuation_label/Ma trận quyết định) vì g vẫn là giả định chủ quan, nhạy cảm.
ERP_COUNTRY_VIETNAM = 0.07
VNINDEX_BETA = 1.0
GROWTH_VNINDEX_LONG_TERM = 0.05


def _calc_capm_valuation(pe, pb, rf):
    if not pe or pe <= 0 or not pb or pb <= 0 or rf is None:
        return None
    roe = pb / pe
    coe = rf + VNINDEX_BETA * ERP_COUNTRY_VIETNAM
    g = GROWTH_VNINDEX_LONG_TERM
    if coe <= g:
        return None
    justified_pb = (roe - g) / (coe - g)
    if justified_pb <= 0:
        return None
    premium_pct = pb / justified_pb - 1
    if premium_pct > 0.15:
        label, color = "Đắt hơn nhiều so với P/B hợp lý (CAPM)", "bad"
    elif premium_pct < -0.15:
        label, color = "Rẻ hơn nhiều so với P/B hợp lý (CAPM)", "good"
    else:
        label, color = "Quanh vùng P/B hợp lý (CAPM)", "warn"
    text = (f"ROE VN-Index (P/B÷P/E) = {roe*100:.2f}%. Chi phí vốn chủ sở hữu CAPM "
            f"(Rf {rf*100:.2f}% + β{VNINDEX_BETA:.1f}×ERP Việt Nam {ERP_COUNTRY_VIETNAM*100:.1f}%) = {coe*100:.2f}%. "
            f"P/B hợp lý (Gordon Growth, g={g*100:.0f}%) = (ROE-g)/(COE-g) = {justified_pb:.2f}x, so với P/B thực tế {pb:.2f}x "
            f"=> {'cao hơn' if premium_pct >= 0 else 'thấp hơn'} {abs(premium_pct)*100:.1f}%. "
            f"=> {label}.")
    return {"roe": roe, "coe": coe, "justified_pb": justified_pb, "premium_pct": premium_pct,
            "label": label, "color": color, "text": text}


# Bù đắp rủi ro cổ phiếu vs kênh an toàn hơn (user 2026-07-13: "PE PB nêu ra cho có thì để làm gì —
# tôi cần so sánh định giá với mức hấp dẫn của tiền gửi/trái phiếu, có bù được rủi ro đầu tư cổ
# phiếu hay không"). So Earnings Yield (1/P/E) với NGƯỠNG = bội số của lãi suất tham chiếu — dùng
# CẢ 2 mốc: TPCP 10 năm (Rf lý thuyết, sẵn có) VÀ lãi suất huy động BÌNH QUÂN THỊ TRƯỜNG thực tế
# (deposit_rate_12m_market_avg — kênh thay thế THẬT mà nhà đầu tư VN có thể chọn, không chỉ lý
# thuyết như Rf trái phiếu). 4 mức, xếp từ tốt nhất xuống xấu nhất:
#   1) E/Y >= 1.5x lãi suất gửi tiết kiệm thực tế -> Bù đắp tốt
#   2) E/Y >= 2x Rf TPCP (nhưng chưa tới mức 1) -> Bù đắp khá so TPCP, CHƯA đủ so gửi tiết kiệm
#   3) E/Y >= 1.5x Rf TPCP (nhưng chưa tới mức 2) -> Bù đắp tối thiểu, chưa có biên an toàn
#   4) E/Y < 1.5x Rf TPCP -> KHÔNG đủ bù đắp rủi ro, kể cả so với kênh phi rủi ro lý thuyết
def _calc_risk_compensation(raw, rf, earnings_yield):
    if earnings_yield is None:
        return None
    deposit_series = raw.get("deposit_rate_12m_market_avg", {}).get("series", [])
    deposit_rf = deposit_series[-1]["value"] / 100 if deposit_series else None

    hurdle_1_5x_bond = 1.5 * rf
    hurdle_2x_bond = 2.0 * rf
    hurdle_1_5x_deposit = 1.5 * deposit_rf if deposit_rf is not None else None

    if hurdle_1_5x_deposit is not None and earnings_yield >= hurdle_1_5x_deposit:
        label, color = "Bù đắp tốt", "good"
    elif earnings_yield >= hurdle_2x_bond:
        label, color = "Bù đắp khá (so TPCP), chưa đủ so với gửi tiết kiệm thực tế", "warn"
    elif earnings_yield >= hurdle_1_5x_bond:
        label, color = "Bù đắp tối thiểu, chưa có biên an toàn", "warn"
    else:
        label, color = "KHÔNG đủ bù đắp rủi ro", "bad"

    parts = [f"Earnings Yield (1/P·E) hiện {earnings_yield*100:.2f}%."]
    parts.append(f"So với lãi suất TPCP 10 năm ({rf*100:.2f}%): "
                 f"{'ĐẠT' if earnings_yield >= hurdle_1_5x_bond else 'CHƯA đạt'} mức 1,5 lần ({hurdle_1_5x_bond*100:.2f}%), "
                 f"{'ĐẠT' if earnings_yield >= hurdle_2x_bond else 'CHƯA đạt'} mức 2 lần ({hurdle_2x_bond*100:.2f}%).")
    if hurdle_1_5x_deposit is not None:
        parts.append(f"So với lãi suất huy động BÌNH QUÂN THỊ TRƯỜNG thực tế ({deposit_rf*100:.2f}%): "
                      f"{'ĐẠT' if earnings_yield >= hurdle_1_5x_deposit else 'CHƯA đạt'} mức 1,5 lần ({hurdle_1_5x_deposit*100:.2f}%).")
    else:
        parts.append("Chưa có dữ liệu lãi suất huy động thị trường để so mức 1,5 lần.")
    parts.append(f"=> {label}.")

    return {
        "earnings_yield": earnings_yield, "deposit_rf": deposit_rf,
        "hurdle_1_5x_bond": hurdle_1_5x_bond, "hurdle_2x_bond": hurdle_2x_bond,
        "hurdle_1_5x_deposit": hurdle_1_5x_deposit,
        "label": label, "color": color, "text": " ".join(parts),
    }


# ══════════════════════════════════════════════════════════════════════════
# MA TRẬN QUYẾT ĐỊNH PHÂN BỔ VỐN (Chương 6.2) — 2 trục thật (Scorecard, Định giá); trục "Dòng
# tiền & Tâm lý" (khối ngoại/margin) CHƯA có nguồn dữ liệu tự động trong Giai đoạn 1 — ghi chú
# rõ, không tự bịa số dòng tiền.
# ══════════════════════════════════════════════════════════════════════════
def calc_decision_matrix(scorecard_total, valuation_label, lai_suat_score=0, lai_suat_veto=False,
                          trend_label=None, high_risk=False):
    """scorecard_total ở đây là ĐIỂM ĐÃ CÓ TRỌNG SỐ (weighted) — xem run_vimo_analysis(): nhóm
    Lãi suất cộng thêm 1 lần nữa (hiệu lực x2) trước khi truyền vào đây.

    NÂNG CẤP 2026-07-25 (theo yêu cầu user): đổi triết lý từ "vĩ mô là CỔNG CHÍNH, định giá chỉ
    tinh chỉnh bên trong" sang "ĐỊNH GIÁ dẫn dắt MỨC ĐỘ giải ngân, vĩ mô là ĐIỀU KIỆN để lên mức
    mạnh nhất". Nguyên văn user: "nếu định giá hợp lý thì có thể giải ngân mua 1 phần, khi rẻ thì
    mua tỷ trọng cao, khi các yếu tố bất lợi về vĩ mô đều qua đi thì MỚI được hiện là mua mạnh".
    Nghĩa là: định giá Rẻ/Hợp lý đã đủ để giải ngân (không còn "Phòng thủ, giữ tiền mặt" toàn bộ
    khi macro xấu + định giá hợp lý như bản cũ) — macro xấu chỉ hãm bớt TỶ TRỌNG, không chặn
    hoàn toàn việc mua; macro TỐT (đã qua giai đoạn bất lợi) là điều kiện DUY NHẤT để nâng "Rẻ"
    lên mức mạnh nhất "Mua mạnh".
    SỬA LẠI CÙNG NGÀY (phản hồi tiếp theo): bản đầu dùng nhãn "Mua tỷ trọng cao" cho Rẻ+macro xấu
    — user chỉ ra Scorecard đang -1/6 (nhiều nhóm Xấu: Thanh khoản/Tỷ giá/Thương mại) nên "tỷ
    trọng CAO" nghe quá mạnh so với mức rủi ro thực tế đang có. Hạ 1 bậc nhãn xuống "Nên mua vào"
    (khuyến nghị giải ngân, KHÔNG chỉ định mức tỷ trọng) cho MỌI trường hợp Rẻ mà macro CHƯA thật
    sự tốt (kể cả khi lai_suat_veto chặn dù tổng điểm đã dương) — "tăng tỷ trọng lớn"/"Mua mạnh"
    CHỈ dành riêng cho macro tốt THẬT SỰ (không bị lãi suất phủ quyết).

    NÂNG CẤP TIẾP (user, cùng ngày 2026-07-25): bổ sung tiêu chí XU HƯỚNG (trend_label từ
    calc_overall_verdict — "Đang cải thiện"/"Đang xấu đi"/khác) và MỨC ĐỘ RỦI RO (high_risk —
    nhiều nhóm Scorecard cùng Xấu) vào ma trận, theo đúng logic 7 bậc user đưa ra:
      Rẻ + đã ĐẢO CHIỀU được XÁC NHẬN (macro tốt + trend đang cải thiện) => Mua mạnh (mạnh nhất)
      Rẻ + macro tốt nhưng CHƯA xác nhận xu hướng cải thiện liên tục    => Mua tỷ trọng cao
      Rẻ + macro xấu                                                    => Nên mua vào (1 phần)
      Hợp lý + trend đang cải thiện                                     => Tăng tỷ trọng vừa phải
      Hợp lý + macro tốt (nhưng chưa rõ xu hướng cải thiện)             => Duy trì, chọn lọc
      Hợp lý + macro xấu                                                => Giải ngân một phần (tỷ trọng thấp)
      Đắt + macro tốt                                                   => Giảm tỷ trọng, chờ điều chỉnh (giữ nguyên)
      Đắt + macro xấu + NHIỀU rủi ro đồng thời (high_risk)              => Clear toàn bộ (mạnh nhất, MỚI)
      Đắt + macro xấu + đang XẤU ĐI (trend worsening)                   => Nên bán ra (MỚI)
      Đắt + macro xấu (baseline, chưa tới 2 mức trên)                   => Giảm vốn mạnh (giữ nguyên)
    "Đã đảo chiều xác nhận" ưu tiên cao hơn "chỉ đang cải thiện" vì đòi hỏi CẢ 2 điều kiện cùng lúc
    (mức điểm ĐANG dương + hướng đi ĐANG tốt lên) — chỉ 1 trong 2 (vd đang cải thiện nhưng vẫn âm
    điểm, hoặc dương điểm nhưng đang chững/xấu đi) thì chưa đủ để coi là "xác nhận qua đáy"."""
    macro_good = scorecard_total > 0
    macro_improving = trend_label == "Đang cải thiện"
    macro_worsening = trend_label == "Đang xấu đi"
    confirmed_reversal = macro_good and macro_improving

    if valuation_label == "Rẻ/Hấp dẫn":
        if confirmed_reversal and not lai_suat_veto:
            return ("Mua mạnh",
                    "Vĩ mô đã ĐẢO CHIỀU được XÁC NHẬN (Scorecard đang dương điểm VÀ đang cải thiện so với kỳ trước) "
                    "+ định giá rẻ — giải ngân mạnh, tăng tỷ trọng cổ phiếu lớn, ưu tiên ngành hưởng lợi từ vĩ mô.")
        if macro_good and not lai_suat_veto:
            return ("Mua tỷ trọng cao",
                    "Vĩ mô đã qua giai đoạn bất lợi (Scorecard dương) + định giá rẻ, nhưng CHƯA đủ xác nhận xu hướng "
                    "cải thiện liên tục — mua vào tỷ trọng lớn hơn, ưu tiên ngành hưởng lợi từ vĩ mô, chưa dốc toàn "
                    "lực như khi có xác nhận đảo chiều rõ ràng.")
        if macro_good and lai_suat_veto:
            return ("Nên mua vào",
                    "Vĩ mô tổng thể đã qua bất lợi + định giá rẻ, NHƯNG lãi suất (yếu tố dẫn dắt tăng trưởng tương lai) "
                    "vẫn đang xấu — nên mua vào nhưng CHƯA tăng tỷ trọng lớn cho tới khi áp lực lãi suất hạ nhiệt.")
        return "Nên mua vào", "Định giá đã đủ rẻ để bù đắp rủi ro, NHƯNG vĩ mô còn nhiều điểm xấu (xem Scorecard) — nên mua vào từng phần, CHƯA tăng tỷ trọng lớn cho tới khi các yếu tố bất lợi về vĩ mô qua đi."
    elif valuation_label == "Đắt/Kém hấp dẫn":
        if macro_good:
            return "Giảm tỷ trọng, chờ điều chỉnh", "Vĩ mô tốt nhưng định giá đã đắt — chốt lời một phần, chờ cơ hội mua lại giá tốt hơn."
        if high_risk:
            return ("Clear toàn bộ",
                    "Định giá đắt + vĩ mô xấu với NHIỀU nhóm rủi ro cùng lúc (xem chi tiết Scorecard) — nên thoát HẲN "
                    "vị thế cổ phiếu, ưu tiên tuyệt đối an toàn vốn cho tới khi rủi ro giảm bớt.")
        if macro_worsening:
            return ("Nên bán ra",
                    "Định giá đắt và vĩ mô đang XẤU ĐI (so với kỳ trước) — nên bán ra, giảm tỷ trọng, chờ định giá "
                    "hoặc vĩ mô cải thiện rõ ràng hơn trước khi cân nhắc mua lại.")
        return "Giảm vốn mạnh", "Vĩ mô xấu + định giá đắt — cắt giảm tỷ trọng cổ phiếu, ưu tiên tiền mặt."
    else:  # Hợp lý
        if macro_improving:
            return ("Tăng tỷ trọng vừa phải",
                    "Định giá hợp lý và vĩ mô đang CẢI THIỆN so với kỳ trước — có thể tăng tỷ trọng cổ phiếu vừa "
                    "phải, chọn lọc theo ngành hưởng lợi từ xu hướng cải thiện.")
        if macro_good:
            return "Duy trì, chọn lọc", "Vĩ mô tốt, định giá hợp lý — giữ tỷ trọng hiện tại, chọn cổ phiếu nền tảng tốt."
        return "Giải ngân một phần", "Định giá hợp lý dù vĩ mô còn bất lợi — giải ngân một phần, chọn lọc, chưa dốc toàn lực cho tới khi vĩ mô cải thiện hoặc định giá rẻ hơn."


# ══════════════════════════════════════════════════════════════════════════
# ĐÁNH GIÁ TỔNG THỂ — user yêu cầu rõ: phải trả lời được "vĩ mô đang tốt lên hay xấu đi" (XU
# HƯỚNG theo thời gian, không chỉ mức điểm tại 1 thời điểm), "gam màu xám hay sáng tỏ" (mức độ
# ĐỒNG THUẬN giữa các chỉ báo — nhiều chỉ báo cùng chiều = rõ ràng/sáng, chỉ báo trái chiều nhau =
# xám/hỗn hợp), và "có phù hợp đầu tư không" (map thẳng ma trận quyết định). 3 trục này ĐỘC LẬP
# với nhau — 1 nền kinh tế có thể "tốt" (điểm dương) nhưng "đang xấu đi" (điểm giảm dần so với kỳ
# trước) và "xám" (nhiều chỉ báo trái chiều) cùng lúc — không gộp chung thành 1 con số duy nhất.
# ══════════════════════════════════════════════════════════════════════════
def _closest_history_entry(history, target_date, tolerance_days):
    """Tìm entry trong scorecard_history có 'date' gần target_date nhất, trong phạm vi
    tolerance_days. Action chạy theo workflow_dispatch thủ công (không cố định lịch), nên KHÔNG
    thể giả định entry[-1] là 'kỳ trước' có ý nghĩa — có thể chỉ cách vài giờ. Neo theo mốc thời
    gian THẬT (tháng trước/quý trước/đầu năm) mới phản ánh đúng xu hướng, đúng yêu cầu user."""
    best, best_diff = None, None
    for h in history:
        try:
            hd = datetime.datetime.strptime(h["date"], "%Y-%m-%d")
        except (KeyError, ValueError, TypeError):
            continue
        diff = abs((hd - target_date).days)
        if diff <= tolerance_days and (best_diff is None or diff < best_diff):
            best, best_diff = h, diff
    return best


def calc_overall_verdict(scorecard_total, trends, scorecard_history):
    """scorecard_history: list các lần chạy TRƯỚC (chưa gồm lần này) [{date, total, ...}], lấy từ
    raw['_meta']['scorecard_history'] TRƯỚC KHI ghi thêm điểm của lần chạy hiện tại. So sánh xu
    hướng theo mốc thời gian THẬT (tháng trước ~30 ngày, quý trước ~90 ngày, đầu năm) — KHÔNG so
    với "entry gần nhất" (vô nghĩa nếu Action chạy nhiều lần gần nhau trong cùng ngày/tuần). Trả
    dict {trend_label, trend_arrow, trend_detail, trend_week, trend_month, clarity_label,
    clarity_pct, clarity_detail}. trend_week/trend_month là 2 dòng ĐỘC LẬP {label, arrow, detail,
    available} để dashboard hiển thị riêng (user 2026-08-01), tách biệt trend_label/trend_arrow
    "chính" (ưu tiên tuần > tháng > quý > đầu năm, vẫn dùng cho PDF/ma trận quyết định/headline)."""
    now = datetime.datetime.now()
    year_start = datetime.datetime(now.year, 1, 1)
    # entry_week: Action cập nhật theo LỊCH TUẦN (update_vimo.yml, cron thứ Sáu hàng tuần) — user
    # (2026-08-01) yêu cầu thêm mốc so sánh này vì đây mới là chu kỳ cập nhật THẬT, sát hơn "tháng
    # trước" (~4 chu kỳ trước). Dung sai hẹp (3 ngày) vì lịch chạy đều đặn, entry gần 7 ngày trước
    # gần như luôn khớp chính xác lần chạy tuần trước.
    entry_week = _closest_history_entry(scorecard_history, now - datetime.timedelta(days=7), 3)
    entry_month = _closest_history_entry(scorecard_history, now - datetime.timedelta(days=30), 10)
    entry_quarter = _closest_history_entry(scorecard_history, now - datetime.timedelta(days=90), 15)
    entry_ytd = _closest_history_entry(scorecard_history, year_start, 20) if now > year_start else None

    # LƯU Ý: scorecard_history có thể chứa entry cũ được tính từ TRƯỚC lần đổi số nhóm Scorecard
    # gần nhất (vd 5 nhóm -> 6 nhóm ngày 2026-07-13) — các entry cũ đó có thang điểm khác (vd ±5
    # thay vì ±6). Hiển thị luôn dùng đúng thang HIỆN TẠI (len(SCORECARD_GROUPS)) cho cả 2 vế để
    # không tự mâu thuẫn trong 1 câu, dù về bản chất so sánh 2 thang điểm khác nhau vẫn có sai số
    # nhỏ trong giai đoạn chuyển tiếp (sẽ tự hết khi lịch sử tích lũy đủ dài sau lần đổi).
    n_groups = len(SCORECARD_GROUPS)

    def _cmp_txt(label, entry):
        if not entry:
            return None
        diff = scorecard_total - entry["total"]
        arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "→")
        return f"{label} ({entry['date']}): {entry['total']:+d}/{n_groups} → {scorecard_total:+d}/{n_groups} {arrow}"

    def _trend_from_entry(entry):
        """Trả {label, arrow, detail, available} cho 1 mốc so sánh RIÊNG (tuần/tháng) — dùng để
        hiển thị 2 DÒNG ĐỘC LẬP trên dashboard (user 2026-08-01: "tôi muốn 2 dòng: so với tuần
        trước và so với tháng trước"), tách biệt khỏi trend_label/trend_arrow "chính" (vẫn ưu tiên
        tuần > tháng > quý > đầu năm, dùng cho PDF/ma trận quyết định/headline — KHÔNG đổi để
        tránh ảnh hưởng các chỗ khác đang phụ thuộc)."""
        if not entry:
            return {"label": "Chưa đủ lịch sử", "arrow": "—",
                    "detail": "Chưa có đủ lịch sử Scorecard ở mốc này để so sánh.", "available": False}
        diff = scorecard_total - entry["total"]
        if diff > 0:
            label, arrow = "Đang cải thiện", "▲"
        elif diff < 0:
            label, arrow = "Đang xấu đi", "▼"
        else:
            label, arrow = "Đi ngang, chưa có chuyển biến rõ", "→"
        detail = (f"({entry['date']}): {entry['total']:+d}/{n_groups} → {scorecard_total:+d}/{n_groups} {arrow}")
        return {"label": label, "arrow": arrow, "detail": detail, "available": True}

    trend_week = _trend_from_entry(entry_week)
    trend_month = _trend_from_entry(entry_month)

    comparisons = [c for c in [
        _cmp_txt("So với tuần trước", entry_week),
        _cmp_txt("So với tháng trước", entry_month),
        _cmp_txt("So với quý trước", entry_quarter),
        _cmp_txt("So với đầu năm", entry_ytd),
    ] if c]

    # Xu hướng CHÍNH ưu tiên mốc gần nhất có dữ liệu (tuần trước nhạy nhất vì đúng chu kỳ cập
    # nhật thật; nếu chưa đủ 1 tuần lịch sử thì lùi dần ra tháng trước rồi quý trước rồi đầu năm).
    primary_entry = entry_week or entry_month or entry_quarter or entry_ytd
    if primary_entry:
        diff = scorecard_total - primary_entry["total"]
        if diff > 0:
            trend_label, trend_arrow = "Đang cải thiện", "▲"
        elif diff < 0:
            trend_label, trend_arrow = "Đang xấu đi", "▼"
        else:
            trend_label, trend_arrow = "Đi ngang, chưa có chuyển biến rõ", "→"
        trend_detail = ". ".join(comparisons) + "."
    else:
        trend_label, trend_arrow = "Chưa đủ lịch sử để so sánh", "—"
        trend_detail = ("Chưa có đủ lịch sử Scorecard cách đây khoảng 1 tháng/1 quý/đầu năm để so sánh xu hướng đáng "
                         "tin cậy — hệ thống sẽ tự tích lũy qua các lần cập nhật tiếp theo.")

    n_good = sum(1 for t in trends.values() if t.get("judgment_label") == "Tốt lên")
    n_bad = sum(1 for t in trends.values() if t.get("judgment_label") == "Xấu đi")
    n_judged = n_good + n_bad
    if n_judged < 5:
        clarity_label, clarity_pct = "Chưa đủ dữ liệu để đánh giá mức độ đồng thuận", None
        clarity_detail = f"Mới có {n_judged} chỉ báo có xu hướng rõ (cần tối thiểu 5 để đánh giá đáng tin cậy)."
    else:
        clarity_pct = round(n_good / n_judged * 100, 1)
        if clarity_pct >= 65:
            clarity_label = "Sáng — đa số chỉ báo đồng thuận tích cực"
        elif clarity_pct <= 35:
            clarity_label = "Tối — đa số chỉ báo đồng thuận tiêu cực"
        else:
            clarity_label = "Xám — tín hiệu trái chiều/hỗn hợp"
        clarity_detail = f"{n_good}/{n_judged} chỉ báo có xu hướng rõ đang tốt lên ({clarity_pct:.0f}%), {n_bad}/{n_judged} đang xấu đi."

    return {
        "trend_label": trend_label, "trend_arrow": trend_arrow, "trend_detail": trend_detail,
        "trend_week": trend_week, "trend_month": trend_month,
        "clarity_label": clarity_label, "clarity_pct": clarity_pct, "clarity_detail": clarity_detail,
    }


# ══════════════════════════════════════════════════════════════════════════
# TỔNG HỢP PHÂN TÍCH ĐA CHỈ SỐ — "bức tranh tổng thể" (user yêu cầu rõ: không chỉ liệt kê số
# liệu/Scorecard, mà phải có ĐÁNH GIÁ TÁC ĐỘNG thật sự tới kinh tế Việt Nam và TTCK, viết chuyên
# nghiệp, đủ chi tiết — không phải 1-2 câu ngắn). RULE-BASED, KHÔNG dùng AI/LLM (user: "tôi không
# cần gemini api key, vì tôi không muốn dùng ai") — mọi câu chữ được ráp từ template có điều kiện,
# mọi con số lấy trực tiếp từ raw/trends/scorecard/valuation, không suy diễn ngoài dữ liệu thật.
# ══════════════════════════════════════════════════════════════════════════
def _vn_join(items):
    items = list(items)
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    return ", ".join(items[:-1]) + " và " + items[-1]


def _fmt_period_vn(period):
    """Chuyển nhãn kỳ nội bộ (vd '2026-06', '2026-Q2', '2026-H1', '2026-9M', '2026-FY', '2026')
    thành dạng ngắn gọn dễ đọc để chèn vào câu văn tổng hợp — GIÚP người đọc biết số liệu đang
    trích dẫn là CỦA KỲ NÀO, tránh hiểu nhầm số liệu là 'hiện tại' khi thực ra có thể đã vài
    tháng/quý trước (user 2026-07-24: đọc phần Tổng hợp không biết số liệu cập nhật ngày nào)."""
    if not period:
        return None
    m = re.match(r"^(\d{4})-W(\d{2})$", period)
    if m:
        return f"Tuần {int(m.group(2))}/{m.group(1)}"
    m = re.match(r"^(\d{4})-(\d{2})$", period)
    if m:
        return f"T{int(m.group(2))}/{m.group(1)}"
    m = re.match(r"^(\d{4})-Q(\d)$", period)
    if m:
        return f"Q{m.group(2)}/{m.group(1)}"
    m = re.match(r"^(\d{4})-H(\d)$", period)
    if m:
        return f"6T đầu {m.group(1)}" if m.group(2) == "1" else f"6T cuối {m.group(1)}"
    m = re.match(r"^(\d{4})-9M$", period)
    if m:
        return f"9T đầu {m.group(1)}"
    m = re.match(r"^(\d{4})-FY$", period)
    if m:
        return f"cả năm {m.group(1)}"
    m = re.match(r"^(\d{4})$", period)
    if m:
        return m.group(1)
    return period


def _indicator_txt(raw, trends, key):
    """Trả (label, value_text, judgment_label) cho 1 chỉ báo nếu có dữ liệu, else None. value_text
    LUÔN kèm kỳ báo cáo (vd '8.39% (kỳ Q2/2026)') để người đọc biết số liệu mới tới đâu, không lầm
    tưởng mọi số liệu đều là 'hiện tại' khi các chỉ báo có độ trễ công bố khác nhau."""
    ind = raw.get(key)
    t = trends.get(key, {})
    if not ind or t.get("latest") is None:
        return None
    period_label = _fmt_period_vn(t.get("latest_period"))
    period_clause = f" (kỳ {period_label})" if period_label else ""
    val_txt = f"{t['latest']:.2f} {ind['unit']}{period_clause}"
    return ind["label"], val_txt, t.get("judgment_label")


def _join_indicators(raw, trends, keys):
    parts = []
    for k in keys:
        info = _indicator_txt(raw, trends, k)
        if info:
            label, val, judgment = info
            j = f", {judgment.lower()}" if judgment else ""
            parts.append(f"{label} đạt {val}{j}")
    return parts


def _build_overview(raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text, verdict):
    # Headline — trả lời TRỰC TIẾP 3 câu hỏi user luôn quan tâm khi đọc báo cáo: (1) đang tốt lên
    # hay xấu đi (xu hướng theo thời gian), (2) bức tranh rõ ràng hay xám/hỗn hợp (mức đồng thuận
    # giữa các chỉ báo), (3) có phù hợp đầu tư lúc này không (ma trận quyết định). Đặt NGAY ĐẦU,
    # trước cả phần liệt kê điểm sáng/điểm nghẽn phía dưới.
    trend_sentence = ("Chưa đủ lịch sử để xác định xu hướng theo thời gian"
                       if verdict["trend_arrow"] == "—"
                       else f"Xu hướng vĩ mô đang {verdict['trend_label'].lower()} {verdict['trend_arrow']}")
    headline = (f"TÓM TẮT: {trend_sentence} — {verdict['trend_detail']} Bức tranh tổng thể hiện ở mức "
                f"\"{verdict['clarity_label']}\" — {verdict['clarity_detail']} Về mức độ phù hợp đầu tư: "
                f"khuyến nghị hiện tại là \"{decision_label}\".")

    pos = [g for g, d in scorecard.items() if d["score"] == 1]
    neg = [g for g, d in scorecard.items() if d["score"] == -1]
    neu = [g for g, d in scorecard.items() if d["score"] == 0 and d["n_votes"] > 0]
    no_data = [g for g, d in scorecard.items() if d["n_votes"] == 0]

    posture = "tích cực" if scorecard_total > 0 else ("tiêu cực" if scorecard_total < 0 else "trung tính, chưa nghiêng hẳn về hướng nào")
    sentences = [f"Scorecard Vĩ Mô tổng hợp đạt {scorecard_total:+d}/{len(SCORECARD_GROUPS)} điểm, cho thấy bức tranh vĩ mô Việt Nam hiện nghiêng {posture}."]
    if pos:
        sentences.append(f"Động lực tích cực đến từ nhóm {_vn_join(pos)}.")
    if neg:
        sentences.append(f"Ngược lại, nhóm {_vn_join(neg)} đang là điểm nghẽn, kéo lùi điểm tổng.")
    if neu:
        sentences.append(f"Nhóm {_vn_join(neu)} giữ trạng thái trung tính — các chỉ báo trong nhóm chưa đồng thuận rõ hướng.")
    if no_data:
        sentences.append(f"Riêng nhóm {_vn_join(no_data)} chưa có đủ chỉ báo có xu hướng rõ để tính điểm, nên tạm giữ 0 theo quy ước thận trọng (không suy diễn khi thiếu dữ liệu).")
    para1 = " ".join(sentences)

    # Liệt kê CỤ THỂ điểm sáng/điểm nghẽn từ TOÀN BỘ chỉ báo (không chỉ giới hạn trong Scorecard)
    # — mirror phong cách "ĐÁNH GIÁ CHU KỲ" của báo cáo vĩ mô DSC/VikkiBankS (T6/2026, user cung
    # cấp làm mẫu tham khảo): nêu số cụ thể, đánh số rõ ràng từng điểm thay vì chỉ nói chung chung.
    positives, risks = [], []
    for key, ind in raw.items():
        if key == "_meta":
            continue
        t = trends.get(key, {})
        if t.get("latest") is None:
            continue
        period_label = _fmt_period_vn(t.get("latest_period"))
        period_clause = f" (kỳ {period_label})" if period_label else ""
        item_txt = f"{ind['label']} đạt {t['latest']:.2f} {ind['unit']}{period_clause}"
        if t.get("judgment_label") == "Tốt lên":
            positives.append(item_txt)
        elif t.get("judgment_label") == "Xấu đi":
            risks.append(item_txt)

    para2_parts = []
    if positives:
        para2_parts.append("Các điểm sáng cụ thể: " + "; ".join(f"({i+1}) {p}" for i, p in enumerate(positives)) + ".")
    if risks:
        para2_parts.append("Ngược lại, một số điểm nghẽn/rủi ro cần lưu ý: " + "; ".join(f"({i+1}) {r}" for i, r in enumerate(risks)) + ".")
    para2 = " ".join(para2_parts)

    if valuation.get("erp") is not None:
        pb_clause = f", P/B {valuation['pb']:.2f}x" if valuation.get("pb") is not None else ""
        para3 = (f"Về định giá, VN-Index đang giao dịch ở P/E {valuation['pe']:.2f}x{pb_clause}, tương ứng ERP (lợi suất thu nhập trừ "
                 f"lãi suất phi rủi ro) {valuation['erp']*100:.2f}% — mức định giá được xếp loại \"{valuation['valuation_label']}\". "
                 f"Kết hợp Scorecard vĩ mô ({scorecard_total:+d}/{len(SCORECARD_GROUPS)}) và mức định giá này, khuyến nghị phân bổ vốn hiện tại là "
                 f"\"{decision_label}\": {decision_text}")
    else:
        para3 = (f"Chưa đủ dữ liệu P/E để tính ERP — phần định giá thị trường tạm thời để trống, không suy diễn. "
                 f"Khuyến nghị phân bổ vốn dựa trên Scorecard: \"{decision_label}\" — {decision_text}")

    paras = [headline, para1]
    if para2:
        paras.append(para2)
    paras.append(para3)
    return "\n\n".join(paras)


def _build_economy_impact(raw, trends):
    """Trả list [{heading, text}, ...] — MỖI PHẦN có tiêu đề rõ ràng riêng (user 2026-07-13: đọc
    1 khối văn bản dài không biết đoạn nào đang nói chủ đề gì) — KHÔNG còn gộp chung thành 1 chuỗi
    text duy nhất như trước. Web/PDF phải render mỗi {heading, text} thành 1 khối có heading in
    đậm/nổi bật riêng, không phải nối liền bằng \\n\\n."""
    sections = []

    growth_keys = ["gdp_growth", "iip_growth", "pmi_manufacturing", "retail_sales_growth"]
    inflation_keys = ["cpi_yoy", "core_inflation"]
    g_parts = _join_indicators(raw, trends, growth_keys)
    i_parts = _join_indicators(raw, trends, inflation_keys)
    if g_parts or i_parts:
        txt = ""
        if g_parts:
            txt += "; ".join(g_parts) + ". "
        # Bóc tách ĐỘNG LỰC tăng trưởng thật (theo NSO) thay vì chỉ nêu con số GDP đơn lẻ — trả
        # lời trực tiếp câu hỏi "GDP tăng do đâu, do chi tiêu công hay đầu tư tư nhân".
        industry_series = raw.get("gdp_industry_contribution", {}).get("series", [])
        invest_series = raw.get("gdp_investment_growth", {}).get("series", [])
        if industry_series:
            p = industry_series[-1]
            txt += (f"Theo NSO, động lực tăng trưởng GDP {p['period']} chủ yếu đến từ khu vực Công nghiệp-Xây dựng "
                     f"(đóng góp {p['value']:.2f}% vào mức tăng chung), trong đó ngành chế biến-chế tạo là hạt nhân. ")
        if invest_series:
            p = invest_series[-1]
            txt += (f"Xét theo phía cầu (phương pháp sử dụng GDP), tích lũy tài sản — tức ĐẦU TƯ, gồm cả đầu tư công, "
                     f"tư nhân và FDI — tăng {p['value']:.2f}% ({p['period']}), vượt xa tốc độ tăng tiêu dùng cuối cùng, "
                     "cho thấy tăng trưởng đang được dẫn dắt chủ yếu bởi đầu tư và xuất khẩu hơn là cầu tiêu dùng nội địa. ")
        if i_parts:
            txt += "Ở chiều giá cả, " + "; ".join(i_parts) + ". "
        brent_j = _indicator_txt(raw, trends, "brent_oil")
        if brent_j and brent_j[2] == "Tốt lên":
            txt += ("Giá dầu Brent hạ nhiệt là một kênh giảm áp lực CPI đáng chú ý — tác động trực tiếp nhất qua cấu "
                    "phần chi phí vận tải/nhiên liệu trong rổ CPI (nhóm Giao thông), gián tiếp qua chi phí logistics "
                    "đầu vào của doanh nghiệp sản xuất, chứ không phải do giá lương thực-thực phẩm hạ. ")
        usdvnd_j = _indicator_txt(raw, trends, "usdvnd")
        if usdvnd_j and usdvnd_j[2] == "Xấu đi":
            txt += ("Tỷ giá USD/VND đang chịu áp lực tăng, làm tăng chi phí nhập khẩu nguyên nhiên liệu và có thể cộng "
                    "thêm áp lực lên CPI trong các kỳ tới qua kênh lạm phát nhập khẩu.")
        sections.append({"heading": "Tăng trưởng & Lạm phát", "text": txt.strip()})

    # Thanh khoản/lãi suất — đây là mục QUAN TRỌNG NHẤT cần trung thực: biểu lãi suất huy động
    # niêm yết CÔNG KHAI (VCB/VietinBank/NamABank, xem dep_parts bên dưới) thấp hơn NHIỀU so với
    # thực tế thị trường — user đã chỉ ra và cung cấp nguồn thật (VnExpress, NSO) xác nhận độc
    # lập: KHÔNG được kết luận "vĩ mô ổn định" chỉ vì lãi suất niêm yết thấp và đi ngang.
    omo = _indicator_txt(raw, trends, "omo_rate_7d")
    interbank_on_series = raw.get("interbank_rate_on", {}).get("series", [])
    interbank_9m_series = raw.get("interbank_rate_9m", {}).get("series", [])
    refin = _indicator_txt(raw, trends, "refinancing_rate")
    if omo or interbank_on_series or refin:
        txt2 = ("Lãi suất OMO kỳ hạn 7 ngày là công cụ NHNN dùng để bơm/hút thanh khoản NGẮN "
                "HẠN tại thị trường liên ngân hàng (thị trường 2), KHÔNG phải thay đổi cung tiền M2 dài hạn — ")
        if omo:
            txt2 += f"hiện ở mức {omo[1]}. "
        if interbank_on_series and interbank_9m_series:
            on_v = interbank_on_series[-1]["value"]
            m9_v = interbank_9m_series[-1]["value"]
            slope = "dốc lên" if m9_v > on_v else ("gần như phẳng" if abs(m9_v - on_v) < 0.3 else "đảo ngược nhẹ")
            txt2 += (f"Đường cong lãi suất liên ngân hàng VNIBOR hiện {slope} (O/N {on_v:.2f}% so với 9 tháng {m9_v:.2f}%), "
                     "phản ánh kỳ vọng thị trường về mức độ căng thẳng thanh khoản ở các kỳ hạn dài hơn. ")
        txt2 += ("Tác động từ OMO lan tỏa GIÁN TIẾP sang lãi suất huy động/cho vay tại thị trường 1 (doanh nghiệp và dân "
                 "cư) thông qua kênh truyền dẫn lãi suất liên ngân hàng — không tức thời và không 1-1, thường có độ trễ "
                 "vài tuần đến vài tháng tùy thanh khoản từng ngân hàng.")
        # 2 CHIỀU bơm/hút (user 2026-07-13): OMO ở trên là chiều BƠM — tín phiếu NHNN là chiều HÚT
        # đối lập, cần nêu CẢ 2 mới đánh giá đúng thanh khoản hệ thống đang thừa hay thiếu.
        tin_phieu_series = raw.get("tin_phieu_days_since_issuance", {}).get("series", [])
        if tin_phieu_series:
            days = tin_phieu_series[-1]["value"]
            txt2 += (f" Ở CHIỀU NGƯỢC LẠI (hút thanh khoản), NHNN đã KHÔNG chào bán tín phiếu trong {days:.0f} ngày "
                      "gần nhất — nghĩa là suốt giai đoạn này chỉ có hoạt động BƠM (OMO), không có hoạt động HÚT nào "
                      "diễn ra, nhất quán với bức tranh thanh khoản đang căng thẳng/thiếu vốn (nếu hệ thống dư thừa "
                      "vốn, NHNN sẽ cần hút bớt qua tín phiếu để tránh áp lực mất giá VND, nhưng thực tế không xảy ra).")
        dep_parts = _join_indicators(raw, trends, ["deposit_rate_12m_vcb", "deposit_rate_12m_ctg", "deposit_rate_12m_nab"])
        if dep_parts:
            txt2 += " Biểu niêm yết CHÍNH THỨC kỳ hạn 12 tháng của nhóm ngân hàng lớn: " + "; ".join(dep_parts) + "."
        sections.append({"heading": "Thanh khoản hệ thống & Lãi suất chính thức (OMO/VNIBOR/niêm yết Big4)",
                          "text": txt2.strip()})

        neg_series = raw.get("deposit_rate_negotiated_max", {}).get("series", [])
        market_max_series = raw.get("deposit_rate_12m_market_max", {}).get("series", [])
        market_avg_series = raw.get("deposit_rate_12m_market_avg", {}).get("series", [])
        txt3 = "Không nên chỉ nhìn biểu lãi suất niêm yết để kết luận thanh khoản ổn định. "
        if neg_series:
            p = neg_series[-1]
            txt3 += (f"Theo NSO (tính đến 26/6/2026: huy động toàn hệ thống +5,02% YTD trong khi tín dụng +7,41% YTD, "
                     f"tín dụng vượt huy động khoảng 1,48 lần) và báo chí tài chính (bài gần nhất ghi nhận, "
                     f"{p['period']}): một số ngân hàng đã phải chào lãi suất huy động THỎA THUẬN (ngoài biểu niêm yết) "
                     f"lên tới {p['value']:.1f}%/năm cho khoản tiền gửi lớn (200 triệu - 1 tỷ đồng trở lên) để bù đắp "
                     "khoảng cách này. ")
        if market_max_series and market_avg_series:
            mm, ma = market_max_series[-1], market_avg_series[-1]
            txt3 += (f"Bảng lãi suất ONLINE công khai đa ngân hàng ({mm['period']}, 24hmoney.vn) cũng cho thấy mặt "
                     f"bằng thị trường rộng hơn nhiều: mức cao nhất {mm['value']:.2f}%/năm, trung bình "
                     f"{ma['value']:.2f}%/năm kỳ hạn 12 tháng — cao hơn hẳn mức ~5,9% của riêng nhóm Big4 nêu trên. ")
        if neg_series or market_max_series:
            txt3 += ("Đây là dấu hiệu hệ thống ngân hàng đang THỰC SỰ CĂNG THẲNG thanh khoản để đáp ứng nhu cầu tín "
                     "dụng, khác hẳn ấn tượng ổn định nếu chỉ nhìn lãi suất niêm yết Big4.")
            sections.append({"heading": "⚠️ Lãi suất THỰC TẾ thị trường (khác biểu niêm yết)", "text": txt3.strip()})

    trade_keys = ["export_growth", "import_growth", "trade_balance", "fdi_disbursed"]
    fiscal_keys = ["budget_revenue_growth", "budget_expenditure_growth", "public_investment_growth"]
    t_parts = _join_indicators(raw, trends, trade_keys)
    f_parts = _join_indicators(raw, trends, fiscal_keys)
    if t_parts or f_parts:
        txt3 = ""
        if t_parts:
            txt3 += "; ".join(t_parts) + ". "
            exp_v = trends.get("export_growth", {}).get("latest")
            imp_v = trends.get("import_growth", {}).get("latest")
            if exp_v is not None and imp_v is not None:
                if imp_v > exp_v:
                    txt3 += "Tốc độ tăng nhập khẩu đang cao hơn xuất khẩu, cho thấy áp lực thu hẹp cán cân thương mại nếu xu hướng này kéo dài. "
                else:
                    txt3 += "Xuất khẩu đang tăng nhanh hơn nhập khẩu, hỗ trợ tích cực cho cán cân thương mại và nguồn cung ngoại tệ. "
        if f_parts:
            txt3 += ("Về đầu tư công, " + "; ".join(f_parts) + " — mức độ giải ngân thực tế quyết định hiệu ứng lan tỏa "
                     "tới các ngành xây dựng, vật liệu, hạ tầng.")
        sections.append({"heading": "Thương mại, Dòng vốn & Đầu tư công", "text": txt3.strip()})

    unemp = _indicator_txt(raw, trends, "unemployment_rate")
    ext_keys = ["fed_funds_rate", "brent_oil", "dxy_proxy", "china_gdp_growth"]
    e_parts = _join_indicators(raw, trends, ext_keys)
    if unemp or e_parts:
        txt4 = ""
        if unemp:
            txt4 += f"Về lao động, {unemp[0]} ở mức {unemp[1]}. "
        if e_parts:
            txt4 += "Áp lực từ bên ngoài: " + "; ".join(e_parts) + ". "
            txt4 += ("Lãi suất Fed và chỉ số USD ảnh hưởng trực tiếp tới dòng vốn ngoại và tỷ giá USD/VND; giá dầu Brent "
                     "tác động tới chi phí đầu vào và lạm phát nhập khẩu do Việt Nam là nước nhập khẩu ròng xăng dầu "
                     "tinh chế; tăng trưởng GDP Trung Quốc ảnh hưởng tới nhu cầu nhập khẩu hàng hóa Việt Nam do đây là "
                     "một trong các đối tác thương mại lớn nhất.")
        sections.append({"heading": "Lao động & Áp lực bên ngoài", "text": txt4.strip()})

    if not sections:
        sections.append({"heading": "Chưa đủ dữ liệu", "text": "Chưa đủ dữ liệu chỉ báo có xu hướng rõ để đánh giá tác động tới nền kinh tế."})
    return sections


def _build_market_impact(raw, trends, scorecard_total, valuation, decision_label, decision_text):
    paras = []
    posture = "thuận lợi" if scorecard_total > 0 else ("bất lợi" if scorecard_total < 0 else "trung tính")
    paras.append(f"Với Scorecard vĩ mô {scorecard_total:+d}/{len(SCORECARD_GROUPS)} điểm ({posture} cho kênh cổ phiếu) và định giá VN-Index "
                 f"được xếp loại \"{valuation.get('valuation_label', 'chưa xác định')}\", khuyến nghị phân bổ vốn hiện "
                 f"tại là \"{decision_label}\": {decision_text}")

    if valuation.get("erp") is not None and valuation.get("rf") is not None:
        p2 = (f"Chênh lệch lợi suất thu nhập cổ phiếu so với lãi suất phi rủi ro (ERP) hiện ở mức {valuation['erp']*100:.2f}%, "
              f"trên nền lãi suất phi rủi ro tham chiếu {valuation['rf']*100:.2f}%. ")
        if valuation["erp"] < 0.02:
            p2 += ("Biên độ bù đắp rủi ro này tương đối mỏng — nếu lãi suất huy động/trái phiếu tiếp tục tăng, sức hấp "
                   "dẫn tương đối của cổ phiếu so với kênh tiền gửi/trái phiếu sẽ giảm thêm.")
        else:
            p2 += ("Biên độ này vẫn còn dư địa, cổ phiếu vẫn giữ được sức hấp dẫn tương đối so với kênh tiền gửi/trái "
                   "phiếu ở mức lãi suất hiện tại.")
        neg_series = raw.get("deposit_rate_negotiated_max", {}).get("series", [])
        if neg_series and valuation.get("rf") is not None:
            real_gap = neg_series[-1]["value"] / 100 - valuation["rf"]
            if real_gap > 0:
                p2 += (f" LƯU Ý: Rf tham chiếu ở trên lấy từ lợi suất TPCP — thấp hơn đáng kể lãi suất huy động THỰC TẾ "
                       f"cao nhất đang ghi nhận trên thị trường ({neg_series[-1]['value']:.1f}%/năm, xem chi tiết ở mục "
                       f"tác động kinh tế). Nếu dùng mức lãi suất thực tế này làm chuẩn so sánh chi phí vốn thay vì Rf "
                       f"trái phiếu, sức hấp dẫn tương đối của cổ phiếu sẽ THẤP HƠN NHIỀU so với con số ERP nêu trên — "
                       "không nên chỉ dựa vào ERP tính theo Rf trái phiếu để kết luận định giá đang hấp dẫn.")
        paras.append(p2)

    sector_notes = []
    usdvnd_j = _indicator_txt(raw, trends, "usdvnd")
    if usdvnd_j and usdvnd_j[2] == "Xấu đi":
        sector_notes.append("VND mất giá có lợi tương đối cho nhóm xuất khẩu (dệt may, thủy sản, gỗ) nhờ tăng sức cạnh "
                             "tranh giá, nhưng gây áp lực chi phí cho nhóm nhập khẩu nguyên liệu và doanh nghiệp có dư "
                             "nợ vay ngoại tệ lớn.")
    brent_j = _indicator_txt(raw, trends, "brent_oil")
    if brent_j:
        if brent_j[2] == "Tốt lên":
            sector_notes.append("Giá dầu Brent hạ nhiệt hỗ trợ biên lợi nhuận nhóm vận tải, nhựa, phân bón (chi phí "
                                 "đầu vào giảm), nhưng bất lợi cho nhóm khai thác/dầu khí thượng nguồn.")
        elif brent_j[2] == "Xấu đi":
            sector_notes.append("Giá dầu Brent tăng có lợi cho nhóm dầu khí thượng nguồn nhưng bào mòn biên lợi nhuận "
                                 "nhóm vận tải, nhựa, phân bón do chi phí đầu vào tăng.")
    interbank_j = _indicator_txt(raw, trends, "interbank_rate_3m")
    if interbank_j and interbank_j[2] == "Xấu đi":
        sector_notes.append("Lãi suất liên ngân hàng tăng phản ánh thanh khoản hệ thống thắt chặt hơn — thường bất lợi "
                             "cho nhóm cổ phiếu dùng đòn bẩy cao (bất động sản, xây dựng) do chi phí vốn tăng.")
    if sector_notes:
        paras.append("Ở cấp độ ngành: " + " ".join(sector_notes))

    return "\n\n".join(paras)


def _build_watch_points(raw, trends, scorecard):
    worsening = []
    for gdata in scorecard.values():
        for d in gdata["detail"]:
            if d["judgment_label"] == "Xấu đi":
                worsening.append(d["label"])
    lines = []
    if worsening:
        lines.append(f"- Các chỉ báo đang xấu đi cần theo dõi sát: {_vn_join(worsening)}.")
    no_data_groups = [g for g, d in scorecard.items() if d["n_votes"] == 0]
    if no_data_groups:
        lines.append(f"- Nhóm {_vn_join(no_data_groups)} hiện chưa đủ chỉ báo có xu hướng rõ để đưa vào Scorecard — cần bổ sung nguồn dữ liệu hoặc theo dõi thủ công.")
    lines.append("- Diễn biến đường cong lãi suất liên ngân hàng VNIBOR (đặc biệt chênh lệch O/N so với các kỳ hạn dài) — dấu hiệu sớm về căng thẳng thanh khoản hệ thống.")
    lines.append("- Tỷ giá USD/VND và động thái lãi suất Fed — ảnh hưởng trực tiếp tới dòng vốn ngoại và áp lực lạm phát nhập khẩu.")
    lines.append("- Xu hướng cán cân thương mại (xuất khẩu so với nhập khẩu) trong các kỳ báo cáo tiếp theo.")
    cred_v = trends.get("credit_growth", {}).get("latest")
    dep_growth_v = trends.get("deposit_growth", {}).get("latest")
    if cred_v is not None and dep_growth_v is not None and cred_v > dep_growth_v:
        lines.append(f"- Khoảng cách tăng trưởng tín dụng ({cred_v:.2f}%) so với huy động vốn ({dep_growth_v:.2f}%) — "
                      "nếu tiếp tục nới rộng, lãi suất huy động thực tế/thỏa thuận (đã ghi nhận tới 9%/năm, cao hơn "
                      "nhiều biểu niêm yết công khai) có thể còn tăng thêm, siết chi phí vốn toàn nền kinh tế.")
    return "\n".join(lines)


def build_synthesis_vimo(raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text, verdict):
    """Tổng hợp phân tích đa chỉ số RULE-BASED (không AI) — trả dict {overview, economy_impact,
    market_impact, watch_points, verdict}, mỗi mục text là 1+ đoạn văn ráp từ số liệu thật trong
    raw/trends. verdict giữ nguyên dạng dict (không ráp thành văn) để dashboard render badge riêng."""
    return {
        "verdict": verdict,
        "overview": _build_overview(raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text, verdict),
        "economy_impact": _build_economy_impact(raw, trends),
        "market_impact": _build_market_impact(raw, trends, scorecard_total, valuation, decision_label, decision_text),
        "watch_points": _build_watch_points(raw, trends, scorecard),
    }


# ══════════════════════════════════════════════════════════════════════════
# CHARTS — 1 chart/chỉ báo cho các chỉ báo có ≥4 điểm dữ liệu (đủ để vẽ đường xu hướng có ý
# nghĩa); chỉ báo thưa dữ liệu hơn hiển thị dạng bảng trong PDF/dashboard, không ép vẽ chart.
# ══════════════════════════════════════════════════════════════════════════
def build_charts_vimo(out_dir, raw, min_points=4):
    os.makedirs(out_dir, exist_ok=True)
    charts = {}
    plt.rcParams["font.family"] = "DejaVu Sans"
    for key, ind in raw.items():
        if key == "_meta":
            continue
        valid = [p for p in ind["series"] if p.get("value") is not None]
        if len(valid) < min_points:
            continue
        periods = [p["period"] for p in valid]
        values = [p["value"] for p in valid]
        improving = (values[-1] >= values[0]) if ind["good_direction"] == "higher" else (values[-1] <= values[0])
        color = "#10b981" if improving else "#ef4444"

        fig, ax = plt.subplots(figsize=(7.5, 3.4))
        ax.plot(periods, values, marker="o", color=color, linewidth=2)
        ax.fill_between(periods, values, alpha=0.08, color=color)
        ax.set_title(f"{ind['label']} ({ind['unit']})", fontsize=11, fontweight="bold")
        ax.tick_params(axis="x", rotation=45, labelsize=8)
        ax.grid(alpha=0.25)
        fig.tight_layout()
        path = os.path.join(out_dir, f"vimo_{key}.png")
        fig.savefig(path, dpi=130)
        plt.close(fig)
        charts[key] = path
    return charts


# Kỳ hạn hiển thị trong chart lịch sử liên ngân hàng — khớp INTERBANK_HISTORY_TENORS trong
# app_vimo.js (user yêu cầu 2026-07-13: gộp O/N + 1 tháng vào chung chart 6 tháng theo thời gian;
# 2026-07-28: thêm 1W/2W — nguồn VIRA daily, xem fetch_vira_bulletin() trong fetch_macro_data.py).
INTERBANK_HISTORY_TENORS = [
    ("interbank_rate_on", "O/N", "#f59e0b"),
    ("interbank_rate_1w", "1 Tuần", "#ef4444"),
    ("interbank_rate_2w", "2 Tuần", "#10b981"),
    ("interbank_rate_1m", "1 Tháng", "#a78bfa"),
    ("interbank_rate_6m", "6 Tháng", "#3b82f6"),
]

# Đường cong lợi suất TPCP thứ cấp theo thời gian — chỉ báo MỚI (2026-07-28, nguồn VIRA), khớp
# BOND_YIELD_TENORS trong app_vimo.js.
BOND_YIELD_TENORS = [
    ("govt_bond_yield_3y", "3 Năm", "#f59e0b"),
    ("govt_bond_yield_5y", "5 Năm", "#ef4444"),
    ("govt_bond_yield_7y", "7 Năm", "#10b981"),
    ("govt_bond_yield_10y", "10 Năm", "#a78bfa"),
    ("govt_bond_yield_15y", "15 Năm", "#3b82f6"),
]


# Khớp 2 định dạng period hợp lệ cho chart này: tuần "YYYY-Www" (SBV, dùng cho 6M/9M/3M) và ngày
# "YYYY-MM-DD" (VIRA, dùng cho ON/1W/2W/1M từ 2026-07-28) — KHÔNG khớp định dạng tháng cũ "YYYY-MM"
# (7 ký tự) vốn là điểm lũy kế/snapshot cũ còn sót lại trước khi đổi sang tuần, trộn chung sẽ khiến
# trục thời gian bị kéo phẳng sai (user 2026-07-25).
_INTERBANK_HISTORY_PERIOD_RE = re.compile(r"^\d{4}-(W\d{2}|\d{2}-\d{2})$")


def build_interbank_6m_history_chart(out_dir, raw):
    """1 line chart nhiều dòng: lãi suất liên ngân hàng O/N, 1 tuần, 2 tuần, 1 tháng, 6 tháng THEO
    THỜI GIAN, vẽ TOÀN BỘ lịch sử sẵn có mỗi kỳ hạn (không giới hạn min_points=4 như
    build_charts_vimo()) — thay cho chart so sánh ngân hàng cũ (đã gỡ bỏ theo yêu cầu user
    2026-07-13, xem cùng thay đổi ở app_vimo.js). CHỈ lấy điểm period dạng TUẦN hoặc NGÀY (xem
    _INTERBANK_HISTORY_PERIOD_RE) — bỏ điểm THÁNG cũ còn sót lại từ trước khi đổi sang
    _current_period_weekly() (2026-07-24): trộn chung nhiều định dạng kỳ khác nhau trên CÙNG 1 trục
    thời gian khiến chart bị kéo phẳng/sai lệch trực quan (user 2026-07-25). Điểm THÁNG cũ vẫn giữ
    nguyên trong vimo_raw.json (không xóa dữ liệu), chỉ loại khỏi chart NÀY. Trả path hoặc None nếu
    không kỳ hạn nào có dữ liệu."""
    series_by_tenor = [(label, color, [p for p in raw.get(key, {}).get("series", [])
                                        if p.get("value") is not None and _INTERBANK_HISTORY_PERIOD_RE.match(p["period"])])
                        for key, label, color in INTERBANK_HISTORY_TENORS]
    # _period_sort_key() (không phải sorted() mặc định) — period ở đây TRỘN 2 định dạng (ngày từ
    # VIRA + tuần từ SBV), sort chuỗi mặc định đẩy "2026-W30" ra cuối trục dù thực ra nằm GIỮA
    # khoảng ngày đã có (user 2026-07-30 phát hiện qua bản web, cùng lỗi bên app_vimo.js).
    all_periods = sorted({p["period"] for _, _, s in series_by_tenor for p in s}, key=_period_sort_key)
    if not all_periods:
        return None

    fig, ax = plt.subplots(figsize=(7.5, 3.4))
    for label, color, s in series_by_tenor:
        by_period = {p["period"]: p["value"] for p in s}
        values = [by_period.get(period, float("nan")) for period in all_periods]  # NaN = có khoảng trống, khớp spanGaps ở app_vimo.js
        ax.plot(all_periods, values, marker="o", color=color, linewidth=2, label=label)
    ax.set_title("Lãi suất liên ngân hàng O/N, 1 tuần, 2 tuần, 1 tháng, 6 tháng theo thời gian (%)", fontsize=11, fontweight="bold")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    ax.grid(alpha=0.25)
    ax.legend(fontsize=8)
    fig.tight_layout()
    path = os.path.join(out_dir, "vimo_interbank_rate_6m_history.png")
    fig.savefig(path, dpi=130)
    plt.close(fig)
    return path


def build_bond_yield_history_chart(out_dir, raw):
    """1 line chart nhiều dòng: lợi suất TPCP thứ cấp 3-5-7-10-15 năm THEO THỜI GIAN — cùng kiểu
    trình bày với build_interbank_6m_history_chart() ở trên (chỉ báo MỚI, 2026-07-28, nguồn VIRA).
    Trả path hoặc None nếu không kỳ hạn nào có dữ liệu."""
    series_by_tenor = [(label, color, [p for p in raw.get(key, {}).get("series", []) if p.get("value") is not None])
                        for key, label, color in BOND_YIELD_TENORS]
    all_periods = sorted({p["period"] for _, _, s in series_by_tenor for p in s}, key=_period_sort_key)
    if not all_periods:
        return None

    fig, ax = plt.subplots(figsize=(7.5, 3.4))
    for label, color, s in series_by_tenor:
        by_period = {p["period"]: p["value"] for p in s}
        values = [by_period.get(period, float("nan")) for period in all_periods]
        ax.plot(all_periods, values, marker="o", color=color, linewidth=2, label=label)
    ax.set_title("Lợi suất TPCP thứ cấp 3-5-7-10-15 năm theo thời gian (%)", fontsize=11, fontweight="bold")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    ax.grid(alpha=0.25)
    ax.legend(fontsize=8)
    fig.tight_layout()
    path = os.path.join(out_dir, "vimo_bond_yield_history.png")
    fig.savefig(path, dpi=130)
    plt.close(fig)
    return path


VNINDEX_VALUATION_HISTORY_PATH = os.path.join(PROJECT_ROOT, "data", "vnindex_valuation_history.json")


def build_vnindex_valuation_history_charts(out_dir, rf, capm):
    """4 biểu đồ lịch sử ĐỊNH GIÁ VN-Index THEO NGÀY (~17 năm, nguồn Vietcap IQ cho headline —
    fetch_vietcap_index_valuation() — và GitHub Truongutc/AIC---chart-nganh cho ex-VIN —
    fetch_vnindex_nonvin_data(), dự án khác của user), TÁCH RIÊNG theo yêu cầu user (2026-07-25:
    "4 biểu đồ... đều có so sánh tương quan với các dữ liệu kia") — mỗi biểu đồ 1 chuỗi, tự so
    với dải thống kê VÀ ngưỡng hấp dẫn CỦA CHÍNH CHUỖI ĐÓ (headline dùng dải Vietcap tính sẵn,
    ex-VIN dùng dải tự tính bằng statistics.mean/stdev — xem update_vnindex_valuation_history()):
    (1) P/E VN-Index headline (có VIN) vs dải thống kê headline + 1.5×/2× Rf.
    (2) P/E VN-Index ex-VIN (loại VIC/VHM/VRE/VPL) vs dải thống kê ex-VIN + 1.5×/2× Rf.
    (3) P/B VN-Index headline vs dải thống kê headline + P/B hợp lý (CAPM).
    (4) P/B VN-Index ex-VIN vs dải thống kê ex-VIN + P/B hợp lý (CAPM) — CAPM giờ tính từ ROE/P-B
        ex-VIN (nguồn chính cho vnindex_pe/pb) nên đây là cặp NHẤT QUÁN nhất để đối chiếu.
    Trả dict {"pe": path, "pe_exvin": path, "pb": path, "pb_exvin": path} — key nào không dựng
    được thì giá trị None."""
    if not os.path.exists(VNINDEX_VALUATION_HISTORY_PATH):
        return {}
    try:
        with open(VNINDEX_VALUATION_HISTORY_PATH, encoding="utf-8") as f:
            hist = json.load(f)
    except Exception:
        return {}

    def _plot_one(data, unit_label, title, filename, extra_lines=None):
        if not data or not data.get("values"):
            return None
        dates = [datetime.datetime.strptime(p["date"], "%Y-%m-%d") for p in data["values"]]
        values = [p["value"] for p in data["values"]]
        fig, ax = plt.subplots(figsize=(7.5, 3.6))
        ax.plot(dates, values, color="#3b82f6", linewidth=1.1, label=f"{unit_label} VN-Index")
        band_specs = [
            ("average", "Trung bình lịch sử", "#f59e0b", "--"),
            ("plusOneSD", "+1SD", "#ef4444", ":"),
            ("minusOneSD", "-1SD", "#10b981", ":"),
            ("plusTwoSD", "+2SD", "#ef4444", "-."),
            ("minusTwoSD", "-2SD", "#10b981", "-."),
        ]
        for key, label, color, ls in band_specs:
            v = data.get(key)
            if v is not None:
                ax.axhline(v, color=color, linestyle=ls, linewidth=1, alpha=0.8, label=f"{label} ({v:.2f})")
        for extra_label, extra_val, extra_color, extra_ls in (extra_lines or []):
            ax.axhline(extra_val, color=extra_color, linestyle=extra_ls, linewidth=1.6,
                       label=f"{extra_label} ({extra_val:.2f})")
        ax.set_title(title, fontsize=11, fontweight="bold")
        ax.tick_params(axis="x", labelsize=8)
        ax.grid(alpha=0.25)
        ax.legend(fontsize=7, loc="upper left", ncol=2)
        fig.tight_layout()
        path = os.path.join(out_dir, filename)
        fig.savefig(path, dpi=130)
        plt.close(fig)
        return path

    pe_extra = []
    if rf and rf > 0:
        # 2x Rf = ngưỡng "bù đắp tốt" trong risk_compensation, 1.5x Rf = ngưỡng "trung tính" (mốc
        # giữa "bù đắp tối thiểu" và "không đủ bù đắp") — user (2026-07-25) yêu cầu thêm để phân
        # biệt vùng hấp dẫn/trung tính/đắt trên cùng biểu đồ, khớp đúng 2 ngưỡng đã dùng trong
        # _calc_risk_compensation (hurdle_1_5x_bond/hurdle_2x_bond).
        pe_extra.append((f"P/E hoà vốn (E/Y=2×Rf {rf*100:.2f}%) — hấp dẫn",
                          1 / (2 * rf), "#8b5cf6", "-"))
        pe_extra.append((f"P/E trung tính (E/Y=1.5×Rf {rf*100:.2f}%)",
                          1 / (1.5 * rf), "#f97316", "--"))
    pb_extra = []
    if capm and capm.get("justified_pb"):
        pb_extra.append(("P/B hợp lý (CAPM, Gordon Growth)", capm["justified_pb"], "#8b5cf6", "-"))

    paths = {}
    paths["pe"] = _plot_one(hist.get("pe"), "P/E",
                             "P/E VN-Index (headline, có VIN) vs dải thống kê & 1.5×/2× lợi suất TPCP",
                             "vimo_vnindex_pe_history.png", pe_extra)
    paths["pe_exvin"] = _plot_one(hist.get("pe_exvin"), "P/E",
                                   "P/E VN-Index ex-VIN (loại VIC/VHM/VRE/VPL) vs dải thống kê & 1.5×/2× lợi suất TPCP",
                                   "vimo_vnindex_pe_exvin_history.png", pe_extra)
    paths["pb"] = _plot_one(hist.get("pb"), "P/B",
                             "P/B VN-Index (headline, có VIN) vs dải thống kê & P/B hợp lý (CAPM)",
                             "vimo_vnindex_pb_history.png", pb_extra)
    paths["pb_exvin"] = _plot_one(hist.get("pb_exvin"), "P/B",
                                   "P/B VN-Index ex-VIN (loại VIC/VHM/VRE/VPL) vs dải thống kê & P/B hợp lý (CAPM)",
                                   "vimo_vnindex_pb_exvin_history.png", pb_extra)
    return paths


# Biểu đồ miền (stacked area) DÙNG CHUNG cho cơ cấu GDP theo khu vực VÀ cơ cấu vốn đầu tư theo
# thành phần (user 2026-07-13, khớp renderStackedAreaChart() trong app_vimo.js) — mỗi kỳ báo cáo
# LŨY KẾ (Q1/6 tháng/9 tháng/cả năm) là 1 điểm trên trục X, các thành phần % cộng lại ~100%.
def build_stacked_area_chart(out_dir, raw, keys, title, filename):
    series_by_key = [(label, color, [p for p in raw.get(key, {}).get("series", []) if p.get("value") is not None])
                      for key, label, color in keys]
    # _period_sort_key() (không phải sort() mặc định) — chuỗi Q1/H1/9M/FY sort abc SAI thứ tự thời
    # gian thật trong năm ("H1" < "Q1" < "FY" < "9M"), phát hiện khi cơ cấu GDP có ≥2 điểm/năm.
    all_periods = sorted({p["period"] for _, _, s in series_by_key for p in s}, key=_period_sort_key)
    if not all_periods:
        return None

    values_matrix = []
    labels, colors = [], []
    for label, color, s in series_by_key:
        by_period = {p["period"]: p["value"] for p in s}
        values_matrix.append([by_period.get(period, 0) for period in all_periods])
        labels.append(label)
        colors.append(color)

    fig, ax = plt.subplots(figsize=(7.5, 3.6))
    ax.stackplot(all_periods, values_matrix, labels=labels, colors=colors, alpha=0.85)
    ax.set_title(title, fontsize=11, fontweight="bold")
    ax.set_ylim(0, 100)
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    ax.grid(alpha=0.25, axis="y")
    ax.legend(fontsize=8, loc="upper center", bbox_to_anchor=(0.5, -0.18), ncol=len(labels))
    fig.tight_layout()
    path = os.path.join(out_dir, filename)
    fig.savefig(path, dpi=130)
    plt.close(fig)
    return path


GDP_STRUCTURE_KEYS = [
    ("gdp_share_agri", "Nông-Lâm-Thủy sản", "#10b981"),
    ("gdp_share_industry", "Công nghiệp-Xây dựng", "#3b82f6"),
    ("gdp_share_services", "Dịch vụ", "#f59e0b"),
    ("gdp_share_tax", "Thuế sản phẩm (ròng)", "#a78bfa"),
]
INVESTMENT_STRUCTURE_KEYS = [
    ("investment_share_state", "Nhà nước", "#3b82f6"),
    ("investment_share_private", "Ngoài Nhà nước (tư nhân)", "#10b981"),
    ("investment_share_fdi", "FDI", "#f59e0b"),
]


# Kỳ hạn VNIBOR theo đúng thứ tự tăng dần — khớp key trong vimo_raw.json (fetch_sbv_interest_rates()
# trong fetch_macro_data.py). Kiểu trình bày lấy cảm hứng từ chart lãi suất liên ngân hàng đa kỳ
# hạn của vimo.cuthongthai.vn (user chỉ ra) — nhưng dùng CHÍNH dữ liệu đã cào từ sbv.gov.vn,
# không phụ thuộc gì vào nguồn của họ.
INTERBANK_TENOR_KEYS = [
    ("interbank_rate_on", "O/N"), ("interbank_rate_1w", "1 Tuần"), ("interbank_rate_2w", "2 Tuần"),
    ("interbank_rate_1m", "1 Tháng"), ("interbank_rate_3m", "3 Tháng"),
    ("interbank_rate_6m", "6 Tháng"), ("interbank_rate_9m", "9 Tháng"),
]


def build_interbank_curve_chart(out_dir, raw):
    """1 bar chart đường cong lãi suất liên ngân hàng (VNIBOR) theo 7 kỳ hạn tại cùng 1 thời
    điểm — cho thấy hình dạng đường cong lãi suất (dốc lên = thị trường kỳ vọng thanh khoản căng
    hơn ở kỳ hạn dài). Trả path hoặc None nếu chưa đủ dữ liệu."""
    labels, values = [], []
    for key, tenor_label in INTERBANK_TENOR_KEYS:
        series = raw.get(key, {}).get("series", [])
        if series:
            labels.append(tenor_label)
            values.append(series[-1]["value"])
    if len(values) < 2:
        return None
    fig, ax = plt.subplots(figsize=(7.5, 3.4))
    ax.plot(labels, values, marker="o", color="#8b5cf6", linewidth=2)
    ax.fill_between(labels, values, alpha=0.08, color="#8b5cf6")
    for i, v in enumerate(values):
        ax.text(i, v + 0.08, f"{v:.2f}%", ha="center", fontsize=9, fontweight="bold")
    ax.set_title("Đường cong lãi suất liên ngân hàng VNIBOR theo kỳ hạn (%)", fontsize=11, fontweight="bold")
    ax.grid(alpha=0.25)
    fig.tight_layout()
    path = os.path.join(out_dir, "vimo_interbank_curve.png")
    fig.savefig(path, dpi=130)
    plt.close(fig)
    return path


# ══════════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════════
def build_pdf_vimo(pdf_path, raw, trends, scorecard, scorecard_total, valuation, decision_label,
                    decision_text, charts, synthesis, verdict,
                    valuation_headline=None, decision_label_headline=None, decision_text_headline=None):
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
                             topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("VM_Title", parent=styles["Heading1"], fontName=FONT_BOLD, fontSize=18,
                               leading=22, textColor=HexColor("#1F4E78"), spaceAfter=12)
    h1_st = ParagraphStyle("VM_H1", parent=styles["Heading2"], fontName=FONT_BOLD, fontSize=13,
                            leading=17, textColor=HexColor("#2E75B6"), spaceBefore=14, spaceAfter=7)
    h2_st = ParagraphStyle("VM_H2", parent=styles["Heading3"], fontName=FONT_BOLD, fontSize=11,
                            leading=15, textColor=HexColor("#404040"), spaceBefore=8, spaceAfter=4)
    h3_st = ParagraphStyle("VM_H3", parent=styles["Heading4"], fontName=FONT_BOLD, fontSize=9.5,
                            leading=13, textColor=HexColor("#2E75B6"), spaceBefore=6, spaceAfter=2)
    body_st = ParagraphStyle("VM_Body", parent=styles["Normal"], fontName=FONT_REG, fontSize=10,
                              leading=14, textColor=HexColor("#2D3748"), spaceAfter=6)
    italic_st = ParagraphStyle("VM_Italic", parent=styles["Normal"], fontName=FONT_REG, fontSize=8,
                                leading=11, textColor=HexColor("#718096"), italic=True)
    small_st = ParagraphStyle("VM_Small", parent=styles["Normal"], fontName=FONT_REG, fontSize=8,
                               leading=11, textColor=HexColor("#4A5568"))

    BLUE_DARK = HexColor("#1F4E78")
    LIGHT_BLUE = HexColor("#DDEBF7")

    def tbl_style(header_col=BLUE_DARK, alt_col=LIGHT_BLUE, font_size=9):
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_col), ("TEXTCOLOR", (0, 0), (-1, 0), white),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD), ("FONTSIZE", (0, 0), (-1, -1), font_size),
            ("FONTNAME", (0, 1), (-1, -1), FONT_REG), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, alt_col]),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"), ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, HexColor("#CBD5E1")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])

    story = []
    today_str = datetime.datetime.now().strftime("%d/%m/%Y")
    story.append(Paragraph("PHÂN TÍCH VĨ MÔ KINH TẾ VIỆT NAM", title_st))
    story.append(Paragraph(f"Khung phân tích Top-Down — Scorecard Vĩ Mô & Ma trận Quyết định Phân bổ Vốn | "
                            f"Ngày lập: {today_str}", body_st))
    story.append(Spacer(1, 8))

    summary_data = [
        ["Scorecard Vĩ Mô (tổng)", "Định giá (ex-VIN)", "ERP (ex-VIN)", "Khuyến nghị (ex-VIN)"],
        [f"{scorecard_total:+d} / {len(SCORECARD_GROUPS)}", valuation["valuation_label"],
         f"{valuation['erp']*100:.2f}%" if valuation["erp"] is not None else "N/A", decision_label],
    ]
    t_sum = Table(summary_data, colWidths=[42 * mm, 42 * mm, 32 * mm, 55 * mm])
    t_sum.setStyle(tbl_style())
    story.append(t_sum)
    story.append(Paragraph(
        "<i>Lưu ý: ERP ở bảng trên = Earnings Yield (1/P·E) trừ Rf thực tế — KHÁC với \"ERP Việt Nam\" "
        "trong box CAPM bên dưới (hằng số giả định ~7% dùng riêng cho mô hình CAPM/Gordon Growth), "
        "2 con số phục vụ 2 mục đích khác nhau.</i>", small_st))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"<b>Khuyến nghị phân bổ vốn (ex-VIN):</b> {decision_text}", body_st))

    # 2 QUYẾT ĐỊNH SONG SONG — user (2026-07-25): so sánh trực tiếp góc nhìn headline (có VIN) vs
    # ex-VIN (loại VIC/VHM/VRE/VPL), cùng 1 bảng để thấy ngay chênh lệch kết luận nếu có.
    if valuation_headline and decision_label_headline:
        pe_h, pe_e = valuation_headline.get("pe"), valuation.get("pe")
        pb_h, pb_e = valuation_headline.get("pb"), valuation.get("pb")
        compare_data = [
            ["Góc nhìn", "P/E", "P/B", "Đánh giá định giá", "Khuyến nghị"],
            ["VN-Index (headline, có VIN)", f"{pe_h:.2f}x" if pe_h else "N/A",
             f"{pb_h:.2f}x" if pb_h else "N/A", valuation_headline["valuation_label"], decision_label_headline],
            ["VN-Index ex-VIN (loại VIC/VHM/VRE/VPL)", f"{pe_e:.2f}x" if pe_e else "N/A",
             f"{pb_e:.2f}x" if pb_e else "N/A", valuation["valuation_label"], decision_label],
        ]
        t_compare = Table(compare_data, colWidths=[62 * mm, 20 * mm, 20 * mm, 30 * mm, 39 * mm])
        t_compare.setStyle(tbl_style())
        story.append(Spacer(1, 8))
        story.append(Paragraph("<b>🔍 So sánh 2 góc nhìn định giá — có/không VIN:</b>", body_st))
        story.append(t_compare)
        if decision_label != decision_label_headline:
            story.append(Paragraph(
                f"⚠ 2 góc nhìn cho khuyến nghị KHÁC NHAU — VIN (VIC/VHM/VRE/VPL) đang làm lệch "
                f"kết luận định giá chung của thị trường một cách đáng kể.", body_st))

    rc = valuation.get("risk_compensation")
    if rc:
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"<b>⚖️ Bù đắp rủi ro cổ phiếu (ex-VIN): {rc['label']}</b> — {rc['text']}", body_st))
    capm = valuation.get("capm_valuation")
    if capm:
        story.append(Paragraph(f"<b>📐 P/B hợp lý theo CAPM (ex-VIN): {capm['label']}</b> — {capm['text']}", body_st))
    if "vnindex_pe_history" in charts:
        story.append(Spacer(1, 6))
        story.append(Image(charts["vnindex_pe_history"], width=170 * mm, height=76 * mm))
    if "vnindex_pe_exvin_history" in charts:
        story.append(Image(charts["vnindex_pe_exvin_history"], width=170 * mm, height=76 * mm))
    if "vnindex_pb_history" in charts:
        story.append(Image(charts["vnindex_pb_history"], width=170 * mm, height=76 * mm))
    if "vnindex_pb_exvin_history" in charts:
        story.append(Image(charts["vnindex_pb_exvin_history"], width=170 * mm, height=76 * mm))
    story.append(Spacer(1, 10))

    # ── Banner "Đánh giá tổng thể" — trả lời trực tiếp: đang tốt lên/xấu đi (xu hướng theo thời
    # gian, so với lần chạy trước) và bức tranh rõ ràng/xám (mức đồng thuận giữa các chỉ báo).
    trend_color = (HexColor("#10b981") if verdict["trend_arrow"] == "▲"
                   else HexColor("#ef4444") if verdict["trend_arrow"] == "▼" else HexColor("#f59e0b"))
    clarity_color = (HexColor("#10b981") if "Sáng" in verdict["clarity_label"]
                      else HexColor("#ef4444") if "Tối" in verdict["clarity_label"] else HexColor("#f59e0b"))
    verdict_data = [
        ["Xu hướng (so kỳ trước)", "Mức độ đồng thuận tín hiệu"],
        [f"{verdict['trend_arrow']} {verdict['trend_label']}", verdict["clarity_label"]],
    ]
    t_verdict = Table(verdict_data, colWidths=[85 * mm, 86 * mm])
    t_verdict.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLUE_DARK), ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD), ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 1), (-1, 1), FONT_BOLD), ("FONTSIZE", (0, 1), (-1, 1), 11),
        ("TEXTCOLOR", (0, 1), (0, 1), trend_color), ("TEXTCOLOR", (1, 1), (1, 1), clarity_color),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, HexColor("#CBD5E1")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t_verdict)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<i>{verdict['trend_detail']} {verdict['clarity_detail']}</i>", italic_st))
    story.append(Spacer(1, 10))

    # ── 1. TỔNG HỢP PHÂN TÍCH ĐA CHỈ SỐ (rule-based, dựa trên số liệu thật) — đặt NGAY ĐẦU báo
    # cáo vì đây là phần quan trọng nhất theo yêu cầu user, không phải phần liệt kê số liệu thô.
    story.append(Paragraph("1. Tổng hợp Phân tích Đa Chỉ số — Bức tranh Tổng thể", h1_st))
    story.append(Paragraph("1.1. Tổng quan bức tranh vĩ mô", h2_st))
    story.append(Paragraph(synthesis["overview"], body_st))
    story.append(Paragraph("1.2. Tác động tới kinh tế Việt Nam", h2_st))
    for section in synthesis["economy_impact"]:
        story.append(Paragraph(section["heading"], h3_st))
        story.append(Paragraph(section["text"], body_st))
    story.append(Paragraph("1.3. Tác động tới thị trường chứng khoán", h2_st))
    story.append(Paragraph(synthesis["market_impact"], body_st))
    story.append(Paragraph("1.4. Điểm cần theo dõi tiếp", h2_st))
    story.append(Paragraph(synthesis["watch_points"], body_st))
    story.append(Spacer(1, 10))

    # ── Scorecard chi tiết ──
    story.append(Paragraph(f"2. Scorecard Vĩ Mô — {len(SCORECARD_GROUPS)} nhóm chỉ số (Chương 4.1/6.1)", h1_st))
    sc_rows = [["Nhóm", "Điểm", "Số chỉ báo có xu hướng rõ", "Lý do & Chi tiết"]]
    for gname, gdata in scorecard.items():
        score_txt = {1: "+1 (Tốt)", 0: "0 (Trung tính)", -1: "-1 (Xấu)"}[gdata["score"]]
        detail_txt = ", ".join(f"{d['label']} ({d['judgment_label']})" for d in gdata["detail"])
        cell_txt = f"<b>{gdata['reason']}</b>" + (f" — {detail_txt}" if detail_txt else "")
        sc_rows.append([gname, score_txt, str(gdata["n_votes"]), Paragraph(cell_txt, small_st)])
    t_sc = Table(sc_rows, colWidths=[32 * mm, 26 * mm, 26 * mm, 85 * mm])
    t_sc.setStyle(tbl_style())
    story.append(t_sc)
    story.append(Spacer(1, 10))

    # ── Từng nhóm chỉ báo chi tiết + chart ──
    story.append(Paragraph("3. Chi tiết theo từng nhóm chỉ báo (Chương 3)", h1_st))
    groups_order = ["growth", "inflation", "monetary", "trade", "fiscal", "labor", "external", "market",
                    "intl_us", "intl_eu", "intl_uk", "intl_jp", "intl_cn"]
    for grp in groups_order:
        keys = [k for k, v in raw.items() if k != "_meta" and v["group"] == grp]
        if not keys:
            continue
        story.append(Paragraph(f"3.{groups_order.index(grp)+1}. {GROUP_LABELS[grp]}", h2_st))
        rows = [["Chỉ báo", "Giá trị mới nhất", "Kỳ", "Số liệu", "Đánh giá", "Nguồn cập nhật"]]
        for k in keys:
            t = trends.get(k, {})
            ind = raw[k]
            src_label = {"worldbank": "World Bank API", "imf": "IMF DataMapper API", "fred": "FRED API",
                         "fx_api": "exchangerate-api.com", "pe_ratio_api": "worldperatio.com",
                         "nso_scrape": "nso.gov.vn (báo cáo quý, tự động)",
                         "nso_chart_embed": "nso.gov.vn (biểu đồ tháng, tự động)",
                         "sbv_chart": "sbv.gov.vn (biểu đồ, tự động)",
                         "sbv_table": "sbv.gov.vn (bảng lãi suất, tự động)",
                         "vietnambiz": "data.vietnambiz.vn (tự động)",
                         "bank_page": "Trang NH chính thức (tự động)",
                         "news_rss": "RSS tin tức CafeF/VietStock (tự động, chỉ khi có tin mới)",
                         "market_table": "24hmoney.vn (bảng đa ngân hàng, tự động)",
                         "24hmoney_scrape": "24hmoney.vn (chỉ số P/E-P/B, tự động)",
                         "cafef_ajax": "cafef.vn (khối ngoại HOSE, tự động)",
                         "manual": "Nghiên cứu thủ công",
                         "vbma": "vbma.org.vn (CSV tĩnh, tự động)"}.get(ind["auto_source"], ind["auto_source"])
            val_txt = f"{t['latest']:.2f} {ind['unit']}" if t.get("latest") is not None else "N/A"
            # value_arrow = chiều số liệu THẬT (↑/↓/→), judgment_label = đánh giá tốt lên/xấu đi
            # riêng biệt (vd Nợ công/GDP giảm ↓ nhưng đánh giá là "Tốt lên") — TRÁNH nhầm 2 khái
            # niệm này với nhau, đúng yêu cầu user.
            rows.append([ind["label"], val_txt, t.get("latest_period", "—"),
                         t.get("value_arrow", "—"), t.get("judgment_label", "—"), src_label])
        t_grp = Table(rows, colWidths=[40 * mm, 26 * mm, 18 * mm, 14 * mm, 22 * mm, 35 * mm])
        t_grp.setStyle(tbl_style())
        story.append(t_grp)
        story.append(Spacer(1, 4))
        for k in keys:
            if k in charts:
                story.append(Image(charts[k], width=140 * mm, height=63 * mm))
        if grp == "monetary" and "interbank_curve" in charts:
            story.append(Paragraph("Đường cong lãi suất liên ngân hàng VNIBOR theo kỳ hạn:", small_st))
            story.append(Image(charts["interbank_curve"], width=140 * mm, height=63 * mm))
            story.append(Spacer(1, 4))
        if grp == "monetary" and "interbank_6m_history" in charts:
            story.append(Paragraph("Lãi suất liên ngân hàng O/N, 1 tuần, 2 tuần, 1 tháng, 6 tháng theo thời gian:", small_st))
            story.append(Image(charts["interbank_6m_history"], width=140 * mm, height=63 * mm))
            story.append(Spacer(1, 4))
        if grp == "monetary" and "bond_yield_history" in charts:
            story.append(Paragraph("Lợi suất TPCP thứ cấp 3-5-7-10-15 năm theo thời gian:", small_st))
            story.append(Image(charts["bond_yield_history"], width=140 * mm, height=63 * mm))
        if grp == "growth" and "gdp_structure" in charts:
            story.append(Paragraph("Cơ cấu GDP theo khu vực kinh tế (lũy kế theo kỳ báo cáo):", small_st))
            story.append(Image(charts["gdp_structure"], width=140 * mm, height=67 * mm))
        if grp == "trade" and "investment_structure" in charts:
            story.append(Paragraph("Cơ cấu vốn đầu tư thực hiện toàn xã hội theo thành phần (lũy kế theo kỳ báo cáo):", small_st))
            story.append(Image(charts["investment_structure"], width=140 * mm, height=67 * mm))
        story.append(Spacer(1, 8))

    # ── Nguồn tham khảo đầy đủ ──
    story.append(Paragraph("4. Nguồn dữ liệu đầy đủ (theo từng chỉ báo)", h1_st))
    for key, ind in raw.items():
        if key == "_meta":
            continue
        urls = sorted({p.get("source_url") for p in ind["series"] if p.get("source_url")})
        if not urls:
            continue
        story.append(Paragraph(f"<b>{ind['label']}:</b>", small_st))
        for u in urls[:4]:
            story.append(Paragraph(f"• {u}", small_st))
    story.append(Spacer(1, 8))

    story.append(Paragraph(
        "⚠ Giai đoạn 1 (khung sơ bộ): một số chỉ báo (PMI, lãi suất liên ngân hàng, tăng trưởng "
        "tín dụng, cung tiền M2, dự trữ ngoại hối, lạm phát cơ bản) hiện chỉ có seed dữ liệu thưa "
        "hoặc chưa có nguồn tự động cập nhật — xem cột 'Nguồn cập nhật' ở mỗi bảng. Trục 'Dòng "
        "tiền & Tâm lý' (mua/bán ròng khối ngoại, tỷ lệ margin — Chương 5.3/6.1) CHƯA được tích "
        "hợp vào ma trận quyết định do chưa có nguồn dữ liệu tự động đáng tin cậy — ma trận hiện "
        "tại chỉ dựa trên 2 trục Scorecard Vĩ Mô và Định giá thị trường.", italic_st))

    doc.build(story)
    return pdf_path


# ══════════════════════════════════════════════════════════════════════════
# JSON EXPORT
# ══════════════════════════════════════════════════════════════════════════
def save_json_vimo(raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text,
                    synthesis, pdf_url=None,
                    valuation_headline=None, decision_label_headline=None, decision_text_headline=None,
                    monitoring_table=None):
    out = {
        "sector": "Vĩ mô",
        "gdrivePdfUrl": pdf_url,
        "lastUpdated": raw.get("_meta", {}).get("last_auto_update") or raw.get("_meta", {}).get("seeded_at"),
        "scorecard": {
            "groups": {gname: {"score": g["score"], "nVotes": g["n_votes"],
                                "detail": g["detail"], "reason": g["reason"]} for gname, g in scorecard.items()},
            "total": scorecard_total,
        },
        "marketValuation": valuation,
        "decision": {"label": decision_label, "text": decision_text},
        "marketValuationHeadline": valuation_headline,
        "decisionHeadline": ({"label": decision_label_headline, "text": decision_text_headline}
                              if decision_label_headline else None),
        "synthesis": synthesis,
        "monitoringTable": monitoring_table,
        "indicators": {},
    }
    for key, ind in raw.items():
        if key == "_meta":
            continue
        t = trends.get(key, {})
        out["indicators"][key] = {
            "group": ind["group"], "label": ind["label"], "unit": ind["unit"],
            "goodDirection": ind["good_direction"], "autoSource": ind["auto_source"],
            "series": ind["series"], "trend": t, "note": ind.get("note"),
            "impact": ind.get("impact"),
        }
    json_path = os.path.join(PROJECT_ROOT, "data", "vimo.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"  [OK] JSON: {json_path}")
    return json_path


# ══════════════════════════════════════════════════════════════════════════
# EXCEL LỊCH SỬ CHỈ SỐ THEO THỜI GIAN — mỗi hàng 1 chỉ báo, mỗi cột 1 kỳ báo cáo. TÁCH RIÊNG
# THEO SHEET theo đơn vị tần suất cập nhật (năm/quý/tháng/ngày/lũy kế H1-9M-FY) — theo đúng yêu
# cầu user (2026-07-13): mỗi nhóm tần suất nằm ở 1 sheet riêng, để khi có kỳ mới (tháng mới, quý
# mới...) chỉ APPEND CỘT MỚI Ở CUỐI sheet đó, không bao giờ chèn/di chuyển cột của sheet khác hay
# cột cũ trong CHÍNH sheet đó — vị trí ô luôn ổn định qua từng lần Action chạy (an toàn nếu user
# tự gắn công thức/biểu đồ tham chiếu ô cố định vào file). Vì mỗi sheet chỉ chứa 1 định dạng kỳ
# báo cáo duy nhất và các kỳ luôn phát sinh theo đúng trình tự thời gian thật (Action luôn fetch
# kỳ "hiện tại"), append-ở-cuối tự động = đúng thứ tự thời gian, KHÔNG cần sắp xếp lại cột.
# ══════════════════════════════════════════════════════════════════════════
_PERIOD_SHEET_PATTERNS = [
    (re.compile(r"^\d{4}-Q[1-4]$"), "Theo_Quy"),
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"), "Theo_Ngay"),
    (re.compile(r"^\d{4}-W\d{2}$"), "Theo_Tuan"),
    (re.compile(r"^\d{4}-\d{2}$"), "Theo_Thang"),
    (re.compile(r"^\d{4}-(H[12]|9M|FY)$"), "Luy_Ke"),
    (re.compile(r"^\d{4}$"), "Theo_Nam"),
]
_SHEET_TITLES = {
    "Theo_Nam": "Theo năm", "Theo_Quy": "Theo quý", "Theo_Thang": "Theo tháng",
    "Theo_Tuan": "Theo tuần", "Theo_Ngay": "Theo ngày", "Luy_Ke": "Lũy kế (H1-9M-FY)", "Khac": "Khác",
}


def _classify_period_sheet(period):
    period = str(period)
    for pattern, sheet_name in _PERIOD_SHEET_PATTERNS:
        if pattern.match(period):
            return sheet_name
    return "Khac"  # định dạng kỳ lạ/chưa gặp — vẫn ghi được, không crash, chỉ gom vào 1 sheet riêng


_LUY_KE_RANK = {"H1": 1, "9M": 1.5, "H2": 2, "FY": 3}


def _period_sort_key(period):
    """Parse 1 chuỗi kỳ báo cáo thành tuple có thể sort đúng thời gian thật — CHỈ dùng để sắp xếp
    các cột MỚI trong 1 lần ghi (chưa từng có vị trí cột trước đó), KHÔNG dùng để di chuyển cột đã
    tồn tại (xem ghi chú trong update_excel_history_vimo về lý do không sắp xếp lại cột cũ)."""
    m = re.match(r"^(\d{4})-Q([1-4])$", period)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", period)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"^(\d{4})-W(\d{2})$", period)
    if m:
        # Quy về tuple (năm, tháng, ngày) của thứ Hai đầu tuần ISO đó — để so sánh được với period
        # dạng NGÀY khi 2 định dạng trộn chung 1 trục (vd chart lãi suất liên ngân hàng: ON/1W/2W/
        # 1M dạng ngày từ VIRA + 6M dạng tuần từ SBV — user 2026-07-30 phát hiện "2026-W30" bị đẩy
        # ra cuối trục dù thực ra nằm GIỮA khoảng ngày đã có, vì trước đây rơi vào nhánh (9999,)).
        monday = datetime.date.fromisocalendar(int(m.group(1)), int(m.group(2)), 1)
        return (monday.year, monday.month, monday.day)
    m = re.match(r"^(\d{4})-(\d{2})$", period)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    m = re.match(r"^(\d{4})-(H1|H2|9M|FY)$", period)
    if m:
        return (int(m.group(1)), _LUY_KE_RANK[m.group(2)])
    m = re.match(r"^(\d{4})$", period)
    if m:
        return (int(m.group(1)),)
    return (9999,)  # định dạng lạ (sheet "Khac") — không parse được, giữ nguyên thứ tự gặp


def update_excel_history_vimo(raw, out_dir):
    xlsx_path = os.path.join(out_dir, "VIMO_Lich_Su_Chi_So.xlsx")
    indicator_keys = [k for k in raw.keys() if k != "_meta"]

    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # xóa sheet mặc định trống — sheet thật sẽ được tạo theo nhu cầu bên dưới

    sheet_state = {}  # sheet_name -> [ws, existing_rows{label: row}, existing_cols{period: col}]
    # points_by_sheet[sheet_name][label] = (unit, [(period, val), ...]) — gom TRƯỚC, ghi cột SAU,
    # để có thể sắp xếp các cột MỚI (chưa từng có trong file) theo đúng thời gian thật trước khi
    # gán số cột — tránh tình trạng cột bị lệch thứ tự chỉ vì thứ tự xử lý chỉ báo trong dict.
    points_by_sheet = {}
    for key in indicator_keys:
        ind = raw[key]
        series = ind.get("series", [])
        if not series:
            continue
        label, unit = ind["label"], ind["unit"]
        for pt in series:
            period, val = pt.get("period"), pt.get("value")
            if period is None or val is None:
                continue
            sheet_name = _classify_period_sheet(period)
            points_by_sheet.setdefault(sheet_name, {}).setdefault(label, (unit, []))[1].append((str(period), val))

    def _get_sheet_state(sheet_name):
        if sheet_name in sheet_state:
            return sheet_state[sheet_name]
        # Khớp theo TIÊU ĐỀ sheet thật (vd "Theo quý"), KHÔNG phải internal key ("Theo_Quy") — 2
        # chuỗi này khác nhau, nếu so sai sẽ không bao giờ tìm thấy sheet đã tồn tại và tạo trùng
        # sheet mới mỗi lần chạy (openpyxl tự thêm hậu tố "1", "2"... khi trùng tiêu đề).
        title = _SHEET_TITLES.get(sheet_name, sheet_name)
        if title in wb.sheetnames:
            ws = wb[title]
        else:
            ws = wb.create_sheet(title=title)
            ws.cell(row=1, column=1, value="Chỉ báo")
            ws.cell(row=1, column=2, value="Đơn vị")
        existing_rows = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)
                          if ws.cell(row=r, column=1).value}
        existing_cols = {str(ws.cell(row=1, column=c).value): c for c in range(3, ws.max_column + 1)
                          if ws.cell(row=1, column=c).value}
        state = [ws, existing_rows, existing_cols]
        sheet_state[sheet_name] = state
        return state

    for sheet_name, by_label in points_by_sheet.items():
        ws, existing_rows, existing_cols = _get_sheet_state(sheet_name)

        # Xác định TOÀN BỘ các period sẽ cần ghi trong lần chạy này cho sheet này, tách ra period
        # nào MỚI (chưa có cột) — chỉ nhóm MỚI này mới được sắp xếp theo thời gian trước khi gán
        # cột; các cột đã tồn tại giữ NGUYÊN vị trí (không bao giờ di chuyển).
        needed_periods = set()
        for unit, points in by_label.values():
            for period, _ in points:
                needed_periods.add(period)
        new_periods = sorted((p for p in needed_periods if p not in existing_cols), key=_period_sort_key)
        next_col = ws.max_column + 1 if ws.max_column >= 3 else 3
        for period in new_periods:
            existing_cols[period] = next_col
            ws.cell(row=1, column=next_col, value=period)
            next_col += 1

        for label, (unit, points) in by_label.items():
            if label not in existing_rows:
                # Hàng mới trong sheet này (sheet mới tạo, HOẶC chỉ báo mới lần đầu xuất hiện ở tần
                # suất này) -> backfill TOÀN BỘ lịch sử sẵn có của chỉ báo trong sheet này.
                r = max(existing_rows.values(), default=1) + 1
                existing_rows[label] = r
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=unit)
                for period, val in points:
                    ws.cell(row=r, column=existing_cols[period], value=val)
            else:
                # Hàng đã tồn tại -> chỉ ghi/ghi đè điểm MỚI NHẤT của tần suất này (không đụng các
                # cột lịch sử cũ hơn đã ghi từ các lần chạy trước).
                period, val = points[-1]
                ws.cell(row=existing_rows[label], column=existing_cols[period], value=val)

    for ws in wb.worksheets:
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 10
        for c in range(3, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12
        ws.freeze_panes = "C2"

    wb.save(xlsx_path)
    print(f"  [OK] Excel lịch sử: {xlsx_path} ({len(wb.sheetnames)} sheet: {', '.join(wb.sheetnames)})")
    return xlsx_path


# ══════════════════════════════════════════════════════════════════════════
# CHỈ BÁO PHÁI SINH (derived) — tính RA từ chuỗi LŨY KẾ đã có trong vimo_raw.json, KHÔNG fetch gì
# mới. CỐ Ý KHÔNG ghi các key này vào vimo_raw.json (save_vimo_raw) — file đó là DỮ LIỆU THÔ, theo
# đúng _meta.description ("KHÔNG phải kết quả tính toán") — chỉ merge vào bản `raw` TRONG BỘ NHỚ ở
# run_vimo_analysis() (sau khi đã save_vimo_raw(raw) xong) để tái sử dụng MIỄN PHÍ toàn bộ cơ chế
# generic đã có (auto card+chart ở app_vimo.js, cột Excel theo _classify_period_sheet(), mục lục
# JSON indicators) mà không cần viết thêm chart/Excel riêng — y hệt cách 1 chỉ báo raw bình thường
# được xử lý, chỉ khác ở chỗ không persist xuống đĩa. User (2026-07-28) yêu cầu: "FDI giải ngân...
# là số lũy kế, khó so sánh giữa các kỳ — tạo biểu đồ giá trị THEO KỲ RỜI RẠC, số kỳ lấy lũy kế trừ
# đi" — áp dụng chung cho FDI giải ngân (lũy kế theo QUÝ: Q1/H1/9M/FY) và FDI đăng ký (lũy kế theo
# THÁNG: đã là chuỗi tháng thật, không cần suy luận kỳ).
def _cumulative_to_discrete(series, cum_periods_in_year_order, relabel=None):
    """series: list [{period, value, source_url}] LŨY KẾ (reset đầu mỗi năm). cum_periods_in_year_order:
    thứ tự các nhãn kỳ lũy kế THẬT SỰ xuất hiện trong 1 năm (vd ["Q1","H1","9M","FY"] hoặc
    ["01",.., "12"]) — kỳ đầu tiên trong danh sách lấy nguyên giá trị, các kỳ sau = kỳ này trừ kỳ
    liền trước NẾU CẢ HAI đều có mặt trong CÙNG năm đó (thiếu 1 trong 2 thì bỏ qua, không suy đoán).
    relabel: dict đổi nhãn kỳ GỐC (vd 'H1') sang nhãn kỳ RỜI RẠC hiển thị (vd 'Q2') trong period
    output — CẦN đổi cho đúng _classify_period_sheet() ở update_excel_history_vimo() (nhãn 'H1'/'9M'
    /'FY' bị coi là LŨY KẾ nên sẽ lạc vào sheet "Lũy kế" dù giá trị giờ đã là số RỜI RẠC); None =
    giữ nguyên nhãn gốc. Trả list [{period, value, source_url}]."""
    relabel = relabel or {}
    by_year = {}
    for pt in series:
        period = pt.get("period", "")
        m = re.match(r"^(\d{4})-(.+)$", str(period))
        if not m:
            continue
        year, sub = m.group(1), m.group(2)
        if sub in cum_periods_in_year_order:
            by_year.setdefault(year, {})[sub] = pt
    out = []
    for year in sorted(by_year):
        points = by_year[year]
        prev_val = None
        gap_subs = []  # các kỳ THIẾU liên tiếp kể từ mốc lũy kế hợp lệ gần nhất
        for sub in cum_periods_in_year_order:
            pt = points.get(sub)
            if pt is None or pt.get("value") is None:
                gap_subs.append(sub)
                continue
            if prev_val is None:
                # kỳ đầu tiên có dữ liệu trong năm (bình thường là kỳ đầu năm; nếu KHÔNG phải kỳ
                # đầu (vd thiếu luôn kỳ đầu) thì đây vẫn là điểm THẬT duy nhất có được, chấp nhận
                # lấy nguyên giá trị làm mốc khởi đầu chuỗi của năm đó, không suy đoán ngược).
                value = pt["value"]
                out.append({"period": f"{year}-{relabel.get(sub, sub)}", "value": value,
                            "source_url": pt.get("source_url")})
            elif len(gap_subs) == 0:
                value = round(pt["value"] - prev_val, 4)
                out.append({"period": f"{year}-{relabel.get(sub, sub)}", "value": value,
                            "source_url": pt.get("source_url")})
            elif len(gap_subs) == 1:
                # ĐÚNG 1 kỳ liền trước bị thiếu — user (2026-07-28) yêu cầu: coi kỳ thiếu và kỳ hiện
                # tại "bằng nhau", chia đều (lũy kế hiện tại - lũy kế mốc gần nhất)/2 cho CẢ HAI, để
                # không mất điểm/không mất thông tin thay vì bỏ trống hoàn toàn — ĐÁNH DẤU rõ là ước
                # tính (source_url riêng, KHÔNG dùng nguồn thật của điểm lũy kế) để không lẫn với số
                # thật khi xem lại sau này.
                half = round((pt["value"] - prev_val) / 2, 4)
                est_note = (f"ƯỚC TÍNH: chia đều (lũy kế {sub} − lũy kế mốc gần nhất)/2, "
                            f"vì thiếu lũy kế kỳ liền trước để trừ trực tiếp — xem note chỉ báo.")
                out.append({"period": f"{year}-{relabel.get(gap_subs[0], gap_subs[0])}",
                            "value": half, "source_url": est_note})
                out.append({"period": f"{year}-{relabel.get(sub, sub)}",
                            "value": half, "source_url": est_note})
            # >1 kỳ liên tiếp bị thiếu: không đủ tin cậy để chia đều (sai số dồn quá lớn) -> bỏ qua
            prev_val = pt["value"]
            gap_subs = []
    return out


def _add_derived_indicators(raw, trends):
    derived = {
        "fdi_disbursed_quarterly": {
            "source_key": "fdi_disbursed", "cum_order": ["Q1", "H1", "9M", "FY"],
            "relabel": {"H1": "Q2", "9M": "Q3", "FY": "Q4"},
            "group": "trade", "label": "FDI giải ngân theo quý (rời rạc)", "unit": "tỷ USD",
            "good_direction": "higher",
            "note": ("Suy ra từ fdi_disbursed (lũy kế Q1/H1/9M/FY, nso.gov.vn) bằng cách LẤY KỲ SAU "
                     "TRỪ KỲ LIỀN TRƯỚC trong CÙNG NĂM (Q1 giữ nguyên; Q2=H1-Q1; Q3=9M-H1; Q4=FY-9M) "
                     "— để so sánh được TỪNG QUÝ riêng lẻ thay vì nhìn đường lũy kế luôn đi lên. Chỉ "
                     "là PHÁI SINH tính toán, KHÔNG lưu vào vimo_raw.json (xem _add_derived_indicators)."),
            "impact": "So được tốc độ giải ngân FDI TỪNG QUÝ riêng biệt thay vì chỉ thấy xu hướng lũy kế luôn tăng.",
        },
        "fdi_registered_monthly": {
            "source_key": "fdi_registered_usd_bn",
            "cum_order": [f"{m:02d}" for m in range(1, 13)],
            "group": "trade", "label": "FDI đăng ký theo tháng (rời rạc)", "unit": "tỷ USD",
            "good_direction": "higher",
            "note": ("Suy ra từ fdi_registered_usd_bn (lũy kế theo tháng, nso.gov.vn) bằng cách lấy "
                     "tháng sau trừ tháng liền trước trong CÙNG NĂM (tháng 01 giữ nguyên). Chỉ là "
                     "PHÁI SINH tính toán, KHÔNG lưu vào vimo_raw.json."),
            "impact": "So được dòng vốn FDI đăng ký MỚI trong từng tháng thay vì chỉ thấy tổng lũy kế luôn tăng.",
        },
        "public_investment_disbursement_value_monthly": {
            "source_key": "public_investment_disbursement_value",
            "cum_order": [f"{m:02d}" for m in range(1, 13)],
            "group": "fiscal", "label": "Giá trị giải ngân đầu tư công theo tháng (rời rạc)",
            "unit": "nghìn tỷ đồng", "good_direction": "higher",
            "note": ("Suy ra từ public_investment_disbursement_value (lũy kế theo tháng, nso.gov.vn) "
                     "bằng cách lấy tháng sau trừ tháng liền trước trong CÙNG NĂM (tháng 01 giữ "
                     "nguyên). LƯU Ý: nguồn gốc vốn có khoảng trống ở tháng 3/6/9/12 (báo cáo quý/6 "
                     "tháng/9 tháng/cả năm không nêu số NSNN theo cách này — đã kiểm tra kỹ cả bản "
                     "'báo cáo' lẫn 'thông cáo báo chí' của 2024/2025/2026, không bài quý nào có câu "
                     "này) nên tháng liền SAU khoảng trống (4/7/10/01 năm sau) không có mốc lũy kế "
                     "liền trước để trừ trực tiếp — 2 tháng này (vd tháng 3+4) ƯỚC TÍNH BẰNG NHAU: "
                     "= (lũy kế tháng 4 − lũy kế tháng 2)/2 cho cả 2 tháng (theo yêu cầu user "
                     "2026-07-28, vì Bộ Tài chính có số thật cho các tháng này nhưng NSO không công "
                     "khai qua kênh đang cào). 2 điểm ước tính này có source_url ghi rõ 'ƯỚC TÍNH' "
                     "để phân biệt với các điểm còn lại (số thật 100% từ báo cáo tháng NSO). Chỉ là "
                     "PHÁI SINH tính toán, KHÔNG lưu vào vimo_raw.json."),
            "impact": "So được quy mô giải ngân TỪNG THÁNG riêng lẻ thay vì chỉ thấy đường lũy kế luôn đi lên rồi reset đầu năm.",
        },
    }
    for new_key, cfg in derived.items():
        src = raw.get(cfg["source_key"])
        if not src:
            continue
        points = _cumulative_to_discrete(src["series"], cfg["cum_order"], cfg.get("relabel"))
        if not points:
            continue
        raw[new_key] = {
            "group": cfg["group"], "label": cfg["label"], "unit": cfg["unit"],
            "good_direction": cfg["good_direction"], "auto_source": "derived",
            "series": points, "note": cfg["note"], "impact": cfg["impact"],
        }
        trends[new_key] = calc_trend(points, cfg["good_direction"])
        print(f"  -> {cfg['label']}: {len(points)} điểm")


def _add_customs_yoy_growth(raw, trends):
    """Tính export_growth_customs/import_growth_customs (YoY %) từ export_value_monthly/
    import_value_monthly (Hải quan, theo tháng — xem load_customs_xnk_local() trong
    fetch_macro_data.py) — CHỈ tính được cho tháng nào có ĐỦ CẢ tháng hiện tại VÀ CÙNG THÁNG năm
    trước (vd 2026-01 cần có 2025-01). Đặt tên KHÁC export_growth/import_growth (nguồn vietnambiz,
    chỉ 1 điểm, vẫn giữ nguyên cho Scorecard) để không đụng vào wiring SCORECARD_GROUPS đang dùng
    key đó — chỉ là chỉ báo THAM KHẢO thêm, tần suất dày hơn nhiều. Phái sinh tính toán, KHÔNG lưu
    vào vimo_raw.json."""
    for src_key, new_key, label in [
        ("export_value_monthly", "export_growth_customs", "Xuất khẩu YoY (Hải quan, theo tháng)"),
        ("import_value_monthly", "import_growth_customs", "Nhập khẩu YoY (Hải quan, theo tháng)"),
    ]:
        src = raw.get(src_key)
        if not src:
            continue
        by_period = {p["period"]: p["value"] for p in src["series"] if p.get("value") is not None}
        points = []
        for period in sorted(by_period):
            year, month = period.split("-")
            prior_period = f"{int(year) - 1}-{month}"
            if prior_period in by_period and by_period[prior_period]:
                yoy = round((by_period[period] / by_period[prior_period] - 1) * 100, 2)
                points.append({"period": period, "value": yoy,
                               "source_url": next(p["source_url"] for p in src["series"] if p["period"] == period)})
        if not points:
            continue
        raw[new_key] = {
            "group": "trade", "label": label, "unit": "%", "good_direction": "higher",
            "auto_source": "derived",
            "series": points,
            "note": (f"Suy ra từ {src_key} (Hải quan, xem load_customs_xnk_local() trong "
                     f"fetch_macro_data.py) bằng cách so cùng tháng năm trước (YoY) — CHỈ tính được "
                     f"khi có đủ dữ liệu 2 năm liên tiếp cho cùng tháng đó. Phái sinh tính toán, "
                     f"KHÔNG lưu vào vimo_raw.json — KHÔNG dùng để tính Scorecard (khác "
                     f"{src_key.split('_')[0]}_growth đang dùng cho Scorecard, nguồn vietnambiz)."),
            "impact": "Đối chiếu tần suất dày hơn (theo tháng, có đủ lịch sử) với chỉ báo cùng tên nguồn vietnambiz (thưa hơn).",
        }
        trends[new_key] = calc_trend(points, "higher")
        print(f"  -> {label}: {len(points)} điểm")


def _yoy_from_level_series(level_by_period, source_url):
    """Suy ra YoY thật (so CÙNG THÁNG năm trước, KHÔNG reset theo năm) từ 1 chuỗi GIÁ TRỊ TUYỆT
    ĐỐI theo tháng {period: value} — dùng chung cho credit_growth_yoy_monthly/
    deposit_growth_yoy_monthly (xem _add_credit_derived_indicators)."""
    points = []
    for period in sorted(level_by_period):
        year, month = period.split("-")
        prior_period = f"{int(year) - 1}-{month}"
        if prior_period in level_by_period and level_by_period[prior_period]:
            yoy = round((level_by_period[period] / level_by_period[prior_period] - 1) * 100, 2)
            points.append({"period": period, "value": yoy, "source_url": source_url})
    return points


def _add_credit_derived_indicators(raw, trends):
    """Từ credit_balance_total/deposit_balance_total (giá trị tuyệt đối, theo tháng, vbma) suy ra
    các chỉ báo mới (user 2026-08-01/2026-08-03):
    1. credit_growth_yoy_monthly — tăng trưởng SO CÙNG KỲ năm trước theo tháng (YoY thật) — KHÁC
       credit_growth hiện có (nguồn SBV, thực chất là "so với đầu năm", RESET về ~0-1% mỗi tháng 1
       rồi cộng dồn tới tháng 12 — xem chuỗi 2025-06=9.91%...2025-12=19.07%, 2026-01=1.19%, không
       phải YoY thật). credit_growth_yoy_monthly tính trực tiếp period/period cùng tháng năm trước,
       không reset theo năm, đúng nghĩa "so cùng kỳ" user yêu cầu.
    2. deposit_growth_yoy_monthly — CÙNG PHƯƠNG PHÁP, từ deposit_balance_total (= Tiền gửi TCKT +
       Tiền gửi dân cư, CÙNG file CSV VBMA đang dùng cho cung tiền M2) — bổ sung góc nhìn YoY thật
       cho huy động, đối chiếu deposit_growth hiện có (nguồn vietnambiz, snapshot 1 điểm/lần chạy,
       user chỉ ra qua data.vietnambiz.vn/currency-interest-rate).
    3. credit_to_gdp_ratio — dư nợ tín dụng / GDP danh nghĩa (nominal_gdp_annual, World Bank, theo
       NĂM) x100, khớp GDP theo NĂM gần nhất có sẵn <= năm của tháng tín dụng (GDP thường công bố
       trễ hơn tín dụng vài tháng-1 năm). Chỉ báo đòn bẩy tín dụng kinh điển (BIS/IMF) — tỷ lệ tăng
       nhanh là dấu hiệu cảnh báo rủi ro tích lũy (bong bóng tài sản/nợ xấu).
    Cả 3 đều PHÁI SINH tính toán, KHÔNG lưu vào vimo_raw.json, KHÔNG dùng cho Scorecard (đã có
    credit_growth/deposit_growth nguồn SBV/vietnambiz cho vai trò đó)."""
    credit = raw.get("credit_balance_total")
    if not credit:
        return
    credit_by_period = {p["period"]: p["value"] for p in credit["series"] if p.get("value") is not None}

    # 1. YoY thật theo tháng
    yoy_points = _yoy_from_level_series(credit_by_period, "https://vbma.org.vn/vi/market-data/credit")
    if yoy_points:
        raw["credit_growth_yoy_monthly"] = {
            "group": "monetary", "label": "Tăng trưởng dư nợ tín dụng toàn nền kinh tế YoY (theo tháng)",
            "unit": "%", "good_direction": "higher", "auto_source": "derived",
            "series": yoy_points,
            "note": ("Suy ra từ credit_balance_total (dư nợ tuyệt đối theo tháng, vbma.org.vn) bằng "
                     "cách so CÙNG THÁNG năm trước (YoY thật) — KHÁC credit_growth (nguồn SBV, thực "
                     "chất là tăng trưởng SO VỚI ĐẦU NĂM, reset về gần 0% mỗi tháng 1 rồi cộng dồn "
                     "tới tháng 12, không phải YoY thật). Phái sinh tính toán, KHÔNG lưu vào "
                     "vimo_raw.json — KHÔNG dùng để tính Scorecard (đã có credit_growth cho vai trò đó)."),
            "impact": "So được tốc độ tăng dư nợ THEO ĐÚNG NGHĨA cùng kỳ năm trước, không bị ảnh hưởng bởi hiệu ứng reset đầu năm của số liệu SBV công bố.",
        }
        trends["credit_growth_yoy_monthly"] = calc_trend(yoy_points, "higher")
        print(f"  -> Tăng trưởng dư nợ tín dụng YoY (theo tháng): {len(yoy_points)} điểm")

    # 2. Huy động — YoY thật theo tháng (cùng phương pháp, nguồn deposit_balance_total)
    deposit = raw.get("deposit_balance_total")
    if deposit:
        deposit_by_period = {p["period"]: p["value"] for p in deposit["series"] if p.get("value") is not None}
        deposit_yoy_points = _yoy_from_level_series(deposit_by_period, "https://vbma.org.vn/vi/market-data/money-supply")
        if deposit_yoy_points:
            raw["deposit_growth_yoy_monthly"] = {
                "group": "monetary", "label": "Tăng trưởng huy động vốn toàn nền kinh tế YoY (theo tháng)",
                "unit": "%", "good_direction": "higher", "auto_source": "derived",
                "series": deposit_yoy_points,
                "note": ("Suy ra từ deposit_balance_total (= Tiền gửi TCKT + Tiền gửi dân cư, theo "
                         "tháng, vbma.org.vn) bằng cách so CÙNG THÁNG năm trước (YoY thật). Đối chiếu "
                         "deposit_growth (nguồn vietnambiz, chỉ có snapshot 1 điểm/lần Action chạy). "
                         "Phái sinh tính toán, KHÔNG lưu vào vimo_raw.json — KHÔNG dùng để tính "
                         "Scorecard (đã có deposit_growth cho vai trò đó)."),
                "impact": "So được tốc độ tăng huy động THẬT theo cùng kỳ, đối chiếu trực tiếp với tăng trưởng tín dụng cùng phương pháp để biết ngân hàng đang cho vay vượt khả năng huy động hay không.",
            }
            trends["deposit_growth_yoy_monthly"] = calc_trend(deposit_yoy_points, "higher")
            print(f"  -> Tăng trưởng huy động vốn YoY (theo tháng): {len(deposit_yoy_points)} điểm")

    # 3. Tín dụng / GDP danh nghĩa
    gdp = raw.get("nominal_gdp_annual")
    if gdp:
        gdp_by_year = {p["period"]: p["value"] for p in gdp["series"] if p.get("value") is not None}
        if gdp_by_year:
            available_years = sorted(int(y) for y in gdp_by_year)
            ratio_points = []
            for period in sorted(credit_by_period):
                year = int(period.split("-")[0])
                usable_years = [y for y in available_years if y <= year]
                gdp_year = max(usable_years) if usable_years else min(available_years)
                gdp_value = gdp_by_year[str(gdp_year)]
                if gdp_value:
                    ratio = round(credit_by_period[period] / gdp_value * 100, 2)
                    ratio_points.append({"period": period, "value": ratio,
                                          "source_url": "https://vbma.org.vn/vi/market-data/credit"})
            if ratio_points:
                raw["credit_to_gdp_ratio"] = {
                    "group": "monetary", "label": "Tỷ lệ Tín dụng / GDP danh nghĩa", "unit": "%",
                    "good_direction": "lower", "auto_source": "derived",
                    "series": ratio_points,
                    "note": ("= credit_balance_total (vbma, theo tháng) / nominal_gdp_annual (World "
                             "Bank, theo NĂM gần nhất có sẵn) x100 — GDP năm nào chưa công bố thì "
                             "dùng năm gần nhất trước đó (GDP thường công bố trễ hơn tín dụng). Chỉ "
                             "báo đòn bẩy tín dụng kinh điển BIS/IMF. Phái sinh tính toán, KHÔNG lưu "
                             "vào vimo_raw.json, KHÔNG dùng để tính Scorecard."),
                    "impact": "Tỷ lệ Tín dụng/GDP tăng nhanh và bền vững (credit boom) là 1 trong các chỉ báo cảnh báo sớm khủng hoảng tài chính được BIS/IMF theo dõi sát nhất — cùng một mức tăng trưởng tín dụng % nhưng nếu tỷ lệ này đã cao/tăng nhanh thì rủi ro tích lũy lớn hơn nhiều.",
                }
                trends["credit_to_gdp_ratio"] = calc_trend(ratio_points, "lower")
                print(f"  -> Tỷ lệ Tín dụng/GDP danh nghĩa: {len(ratio_points)} điểm")


# Danh sách CỐ ĐỊNH chỉ báo THEO THÁNG cho bảng giám sát (user 2026-08-03, tham khảo trình bày
# kiểu "Bảng giám sát các chỉ số vĩ mô hàng tháng" của báo cáo phân tích — heatmap màu theo hàng).
# Chỉ chọn chỉ báo có period dạng "YYYY-MM" (không lấy fdi_disbursed dạng Q1/H1/9M/FY, không so
# thẳng hàng được) — PHẢI gọi _build_monitoring_table() SAU KHI đã chạy các hàm phái sinh
# (_add_customs_yoy_growth, _add_credit_derived_indicators...) vì nhiều dòng lấy từ chỉ báo phái
# sinh (không có trong vimo_raw.json, chỉ có trong `raw` lúc đang chạy).
_MONITORING_TABLE_ROWS = [
    ("iip_growth", "Sản xuất công nghiệp (IIP, YoY)"),
    ("pmi_manufacturing", "PMI sản xuất"),
    ("retail_sales_growth", "Doanh số bán lẻ (YoY)"),
    ("export_growth_customs", "Xuất khẩu hàng hóa (YoY)"),
    ("import_growth_customs", "Nhập khẩu hàng hóa (YoY)"),
    ("cpi_yoy", "CPI (YoY)"),
    ("core_inflation", "Lạm phát cơ bản (YoY)"),
    # 5 hàng dưới đây là ĐÓNG GÓP (điểm %) vào CPI chung, KHÔNG phải %YoY riêng nhóm (user
    # 2026-08-03, đối chiếu ảnh bảng CPI theo nhóm hàng) — nguồn vbma.org.vn/vi/market-data/cpi,
    # cộng lại ≈ cpi_yoy. Label ghi rõ "đóng góp" để không nhầm với %YoY riêng nhóm.
    ("cpi_contrib_food", "— Đóng góp: Thực phẩm (điểm %)"),
    ("cpi_contrib_housing_utilities", "— Đóng góp: Nhà, điện, nước (điểm %)"),
    ("cpi_contrib_healthcare", "— Đóng góp: Y tế (điểm %)"),
    ("cpi_contrib_transport", "— Đóng góp: Vận tải (điểm %)"),
    ("cpi_contrib_other", "— Đóng góp: Khác (điểm %)"),
    ("credit_growth_yoy_monthly", "Tăng trưởng tín dụng (YoY)"),
    ("deposit_growth_yoy_monthly", "Tăng trưởng huy động (YoY)"),
    ("m2_growth", "Tăng trưởng cung tiền M2 (YoY)"),
    ("public_investment_disbursement_rate", "Tỷ lệ giải ngân đầu tư công (lũy kế, %KH năm)"),
]
_MONITORING_TABLE_MONTHLY_RE = re.compile(r"^\d{4}-\d{2}$")


def _build_monitoring_table(raw, n_months=13):
    """Bảng giám sát chỉ số vĩ mô hàng tháng (heatmap, user 2026-08-03) — lấy N tháng gần nhất
    (hợp nhất từ mọi chỉ báo trong _MONITORING_TABLE_ROWS), mỗi hàng chỉ đưa vào bảng nếu có ĐỦ
    dữ liệu trong cửa sổ đó (>=6/13 điểm) để tránh hàng gần như trống làm rối bảng — vd
    retail_sales_growth hiện chỉ có 2 điểm (nguồn vietnambiz, tích lũy chậm) sẽ tự động bị loại,
    không cần hardcode danh sách loại trừ, tự động xuất hiện lại khi đủ dữ liệu về sau. Trả
    {"periods": [...], "rows": [{"key","label","unit","goodDirection","values":[v_hoặc_None]}]}
    hoặc None nếu không có chỉ báo nào đủ điều kiện."""
    all_periods = set()
    for key, _ in _MONITORING_TABLE_ROWS:
        ind = raw.get(key)
        if not ind:
            continue
        for p in ind["series"]:
            if _MONITORING_TABLE_MONTHLY_RE.match(p["period"]):
                all_periods.add(p["period"])
    periods = sorted(all_periods)[-n_months:]
    if not periods:
        return None

    rows = []
    for key, label in _MONITORING_TABLE_ROWS:
        ind = raw.get(key)
        if not ind:
            continue
        by_period = {p["period"]: p["value"] for p in ind["series"] if p.get("value") is not None}
        values = [by_period.get(p) for p in periods]
        if sum(1 for v in values if v is not None) < 6:
            continue
        # Màu tính theo min-max của TOÀN BỘ LỊCH SỬ chỉ báo (colorMin/colorMax), KHÔNG phải chỉ
        # riêng N tháng đang hiển thị — user (2026-08-03) chỉ ra tín dụng 18,23% (đã là mức cao so
        # lịch sử) vẫn bị tô đỏ vì cửa sổ hiển thị chỉ có 18-22% (bản thân dải hẹp đã ở mức cao,
        # điểm thấp nhất trong dải hẹp đó vẫn luôn bị tô đỏ dù cao tuyệt đối). Dùng full-history
        # cho thang màu tránh bóp méo kiểu này — vd tín dụng full-history 8,96-21,93% thì 18,23%
        # nằm ở ~72% dải, tô xanh đúng như cảm nhận trực quan.
        all_vals = [p["value"] for p in ind["series"] if p.get("value") is not None]
        rows.append({"key": key, "label": label, "unit": ind["unit"],
                     "goodDirection": ind["good_direction"], "values": values,
                     "colorMin": min(all_vals), "colorMax": max(all_vals)})
    if not rows:
        return None
    return {"periods": periods, "rows": rows}


def _add_us_yield_curve_spread(raw, trends):
    """Tính spread đường cong lợi suất TPCP Mỹ 10Y-2Y và 10Y-3M — 2 chỉ báo đảo ngược đường cong
    kinh điển (spread ÂM = đảo ngược, tín hiệu cảnh báo suy thoái Mỹ được thị trường toàn cầu theo
    dõi sát nhất). Chỉ làm cho Mỹ vì FRED có đủ chuỗi ngày tin cậy cho mọi kỳ hạn (DGS*) — Đức/
    Nhật/Anh/Trung Quốc không có bộ kỳ hạn đầy đủ tương đương. Phái sinh tính toán, KHÔNG lưu vào
    vimo_raw.json (theo đúng quy ước _add_derived_indicators/_add_customs_yoy_growth)."""
    pairs = [
        ("us_yield_spread_10y2y", "us_yield_2y", "2 năm"),
        ("us_yield_spread_10y3m", "us_yield_3m", "3 tháng"),
    ]
    long_key = "us_10y_yield"
    long_src = raw.get(long_key)
    if not long_src:
        return
    long_by_period = {p["period"]: p["value"] for p in long_src["series"] if p.get("value") is not None}
    for new_key, short_key, short_label in pairs:
        short_src = raw.get(short_key)
        if not short_src:
            continue
        short_by_period = {p["period"]: p["value"] for p in short_src["series"] if p.get("value") is not None}
        common_periods = sorted(set(long_by_period) & set(short_by_period))
        if not common_periods:
            continue
        points = [{"period": p, "value": round(long_by_period[p] - short_by_period[p], 4),
                   "source_url": "https://fred.stlouisfed.org/series/T10Y2Y" if short_key == "us_yield_2y"
                                 else "https://fred.stlouisfed.org/series/T10Y3M"}
                  for p in common_periods]
        raw[new_key] = {
            "group": "intl_us", "label": f"Spread lợi suất TPCP Mỹ 10 năm - {short_label}",
            "unit": "điểm %", "good_direction": "higher", "auto_source": "derived",
            "series": points,
            "note": (f"= us_10y_yield − {short_key} (khớp theo period NGÀY, cả 2 đều từ FRED). Spread "
                     f"ÂM (đường cong đảo ngược) lịch sử là tín hiệu cảnh báo suy thoái Mỹ đáng tin "
                     f"cậy — mọi lần suy thoái Mỹ từ 1955 đều có đảo ngược 10Y-2Y trước đó vài quý-"
                     f"vài năm. good_direction=higher vì spread dương (đường cong bình thường, dốc "
                     f"lên) là trạng thái lành mạnh. Phái sinh tính toán, KHÔNG lưu vào vimo_raw.json."),
            "impact": "Đường cong lợi suất Mỹ đảo ngược kéo dài là tín hiệu cảnh báo suy thoái Mỹ sớm nhất được thị trường toàn cầu theo dõi — ảnh hưởng khẩu vị rủi ro chung, gián tiếp tác động dòng vốn ngoại vào TTCK Việt Nam.",
        }
        trends[new_key] = calc_trend(points, "higher")
        print(f"  -> Spread 10Y-{short_label}: {len(points)} điểm")


def _fill_disbursement_gaps(raw, trends):
    """Lấp khoảng trống tháng 3/6/9/12 của CẢ HAI public_investment_disbursement_value (nghìn tỷ)
    VÀ public_investment_disbursement_rate (%) — CHỈ sửa TRONG BỘ NHỚ (không lưu vimo_raw.json).
    Bước 1 (giá trị): user (2026-07-28) yêu cầu coi tháng thiếu và tháng liền sau "bằng nhau", chia
    đều (lũy kế tháng sau − lũy kế tháng trước liền kề)/2 cho cả 2 — CÙNG công thức đã dùng ở
    _cumulative_to_discrete()/public_investment_disbursement_value_monthly, làm lại ở đây để ghi
    ĐƯỢC vào chính series gốc (không chỉ bản phái sinh _monthly).
    Bước 2 (%): user đề xuất suy ngược 'kế hoạch năm' = giá trị ÷ (tỷ lệ/100) rồi áp cho tháng thiếu
    — kiểm tra bằng số liệu thật cho thấy 'kế hoạch năm' KHÔNG cố định (tăng dần suốt năm do Chính
    phủ bổ sung/điều chỉnh giữa năm, vd 2025: từ ~861 nghìn tỷ đầu năm lên ~1020 cuối năm) nên NỘI
    SUY TUYẾN TÍNH kế hoạch giữa 2 mốc THẬT liền kề (đúng 1 tháng cách đều mỗi bên) thay vì dùng
    nguyên 1 mốc, giảm sai số. Cả 2 bước đều ĐÁNH DẤU rõ nguồn là ước tính."""
    months_order = [f"{m:02d}" for m in range(1, 13)]

    def _gap_periods(subs_present):
        """Trả [(gap_sub, prev_sub, next_sub)] cho các kỳ ĐÚNG 1 tháng bị thiếu, có mốc thật 2 bên."""
        out, prev_sub = [], None
        for sub in months_order:
            if sub in subs_present:
                prev_sub = sub
                continue
            next_sub = f"{int(sub) + 1:02d}"
            if prev_sub is not None and next_sub in subs_present:
                out.append((sub, prev_sub, next_sub))
        return out

    value_ind = raw.get("public_investment_disbursement_value")
    rate_ind = raw.get("public_investment_disbursement_rate")
    if not value_ind or not rate_ind:
        return

    value_by_period = {p["period"]: p["value"] for p in value_ind["series"] if p.get("value") is not None}
    rate_by_period = {p["period"]: p["value"] for p in rate_ind["series"] if p.get("value") is not None}

    by_year_value = {}
    for period in value_by_period:
        y, s = period.split("-")
        by_year_value.setdefault(y, set()).add(s)

    est_value_note = "ƯỚC TÍNH: chia đều (lũy kế tháng sau − lũy kế tháng trước liền kề)/2 — xem note chỉ báo."
    new_value_points = []
    for year, subs in by_year_value.items():
        for gap_sub, prev_sub, next_sub in _gap_periods(subs):
            p_before, p_after, p_gap = f"{year}-{prev_sub}", f"{year}-{next_sub}", f"{year}-{gap_sub}"
            half = round((value_by_period[p_after] - value_by_period[p_before]) / 2, 4)
            est_val = round(value_by_period[p_before] + half, 4)
            new_value_points.append({"period": p_gap, "value": est_val, "source_url": est_value_note})
            value_by_period[p_gap] = est_val  # để bước 2 dùng ngay

    if new_value_points:
        merged_value = value_ind["series"] + new_value_points
        merged_value.sort(key=lambda p: p["period"])
        value_ind["series"] = merged_value
        trends["public_investment_disbursement_value"] = calc_trend(merged_value, value_ind["good_direction"])

    by_year_rate = {}
    for period in rate_by_period:
        y, s = period.split("-")
        by_year_rate.setdefault(y, set()).add(s)

    new_rate_points = []
    for year, subs in by_year_rate.items():
        for gap_sub, prev_sub, next_sub in _gap_periods(subs):
            p_before, p_after, p_gap = f"{year}-{prev_sub}", f"{year}-{next_sub}", f"{year}-{gap_sub}"
            if p_gap not in value_by_period:
                continue
            plan_before = value_by_period[p_before] / (rate_by_period[p_before] / 100)
            plan_after = value_by_period[p_after] / (rate_by_period[p_after] / 100)
            plan_interp = (plan_before + plan_after) / 2  # nội suy tuyến tính, gap luôn ở giữa
            est_rate = round(value_by_period[p_gap] / plan_interp * 100, 2)
            new_rate_points.append({
                "period": p_gap, "value": est_rate,
                "source_url": (f"ƯỚC TÍNH: giá trị ước tính ÷ kế hoạch năm nội suy tuyến tính giữa "
                               f"{p_before} và {p_after} (kế hoạch năm KHÔNG cố định — xem note chỉ báo)."),
            })

    if new_rate_points:
        merged_rate = rate_ind["series"] + new_rate_points
        merged_rate.sort(key=lambda p: p["period"])
        rate_ind["series"] = merged_rate
        trends["public_investment_disbursement_rate"] = calc_trend(merged_rate, rate_ind["good_direction"])

    print(f"  -> Lấp {len(new_value_points)} điểm giá trị + {len(new_rate_points)} điểm % giải ngân bị thiếu (ước tính)")


# ══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════
def run_vimo_analysis():
    print("\n" + "=" * 60)
    print("  PHÂN TÍCH VĨ MÔ KINH TẾ VIỆT NAM")
    print("=" * 60)

    raw = load_vimo_raw()

    print("[INFO] Fetching Rf (lãi suất TPCP 10 năm) cho ERP...")
    rf, rf_src = fetch_rf_vietnam()
    print(f"  Rf={rf*100:.2f}% ({rf_src})")

    print("[INFO] Tính xu hướng từng chỉ báo...")
    trends = {}
    for key, ind in raw.items():
        if key == "_meta":
            continue
        trends[key] = calc_trend(ind["series"], ind["good_direction"])
        t = trends[key]
        print(f"  {ind['label']}: {t.get('arrow', '—')} (latest={t.get('latest')}, n={t.get('n_points', 0)})")

    print("[INFO] Tính Scorecard Vĩ Mô...")
    scorecard, scorecard_total = calc_scorecard(raw, trends)
    for gname, g in scorecard.items():
        print(f"  {gname}: {g['score']:+d} ({g['n_votes']} phiếu bầu)")
    print(f"  => TỔNG SCORECARD: {scorecard_total:+d} / {len(SCORECARD_GROUPS)}")

    print("[INFO] Tính định giá thị trường (P/E, ERP) — ex-VIN (nguồn chính)...")
    valuation = calc_market_valuation(raw, rf)
    print(f"  P/E={valuation['pe']} | ERP={valuation['erp']*100:.2f}%" if valuation['erp'] is not None else "  P/E/ERP: N/A")
    print(f"  => {valuation['valuation_label']}")

    # Tính XU HƯỚNG (verdict) TRƯỚC ma trận quyết định — user (2026-07-25) yêu cầu ma trận phân
    # biệt "vĩ mô tốt lên" (trend đang cải thiện) khỏi "vĩ mô đã đảo chiều được XÁC NHẬN" (macro
    # tốt + trend đang cải thiện cùng lúc), nên cần trend_label sẵn sàng trước khi gọi
    # calc_decision_matrix(). scorecard_history dùng để so sánh vẫn là history TRƯỚC lần chạy này
    # (đọc 1 lần, entry hôm nay chỉ ghi vào raw sau khi đã tính xong verdict — xem bên dưới).
    print("[INFO] Đánh giá tổng thể (xu hướng + mức độ đồng thuận)...")
    scorecard_history = raw.get("_meta", {}).get("scorecard_history", [])
    verdict = calc_overall_verdict(scorecard_total, trends, scorecard_history)
    print(f"  Xu hướng: {verdict['trend_label']} {verdict['trend_arrow']}")
    print(f"  Mức độ rõ ràng: {verdict['clarity_label']}")

    # "Nhiều rủi ro đồng thời" (high_risk) — user: "định giá cao + vĩ mô xấu VÀ NHIỀU RỦI RO = clear
    # toàn bộ", khác mức độ với chỉ "vĩ mô xấu" thường. Đo bằng số nhóm Scorecard đang -1 Xấu — từ
    # NỬA tổng số nhóm trở lên (vd >=3/6) coi là rủi ro lan rộng, không chỉ khoanh vùng 1-2 nhóm.
    n_neg_groups = sum(1 for g in scorecard.values() if g["score"] == -1)
    high_risk = n_neg_groups >= (len(SCORECARD_GROUPS) + 1) // 2
    print(f"  Số nhóm Xấu: {n_neg_groups}/{len(SCORECARD_GROUPS)} -> high_risk={high_risk}")

    # Lãi suất được cộng THÊM 1 LẦN NỮA (hiệu lực x2) khi tính quyết định phân bổ vốn — KHÔNG
    # đổi scorecard_total "thô" hiển thị/theo dõi lịch sử phía trên (giữ nguyên tính liên tục so
    # sánh xu hướng qua các kỳ). Nếu nhóm Lãi suất đang -1 Xấu, chặn hẳn "Bung vốn mạnh" (xem
    # calc_decision_matrix) — theo phản hồi user (2026-07-25): lãi suất là yếu tố then chốt nhất.
    lai_suat_score = scorecard.get("Lãi suất", {}).get("score", 0)
    decision_score = scorecard_total + lai_suat_score
    decision_label, decision_text = calc_decision_matrix(
        decision_score, valuation["valuation_label"],
        lai_suat_score=lai_suat_score, lai_suat_veto=(lai_suat_score == -1),
        trend_label=verdict["trend_label"], high_risk=high_risk)
    print(f"[INFO] Ma trận quyết định (ex-VIN): {decision_label} — {decision_text} (điểm quyết định có trọng số: {decision_score:+d})")

    # 2 QUYẾT ĐỊNH SONG SONG — user (2026-07-25): "chia ra 2 quyết định: nếu nhìn vào VN-Index thì
    # quyết định là gì, định giá hấp dẫn không. Nếu nhìn theo VN-Index no VIN thì quyết định là
    # gì". Cùng scorecard_total/lai_suat_score (vĩ mô không đổi theo góc nhìn định giá) nhưng
    # valuation_label khác nhau -> decision_label có thể khác nhau giữa 2 góc nhìn.
    print("[INFO] Tính định giá thị trường — headline (có VIN, đối chiếu song song)...")
    valuation_headline = calc_market_valuation_headline(raw, rf)
    decision_label_headline = decision_text_headline = None
    if valuation_headline:
        print(f"  P/E={valuation_headline['pe']} | ERP={valuation_headline['erp']*100:.2f}%"
              if valuation_headline['erp'] is not None else "  P/E/ERP: N/A")
        print(f"  => {valuation_headline['valuation_label']}")
        decision_label_headline, decision_text_headline = calc_decision_matrix(
            decision_score, valuation_headline["valuation_label"],
            lai_suat_score=lai_suat_score, lai_suat_veto=(lai_suat_score == -1),
            trend_label=verdict["trend_label"], high_risk=high_risk)
        print(f"[INFO] Ma trận quyết định (headline): {decision_label_headline} — {decision_text_headline}")
    else:
        print("  [INFO] Chưa có dữ liệu headline P/E-P/B (vnindex_pe_headline) — bỏ qua đối chiếu song song.")
    today_str_iso = datetime.datetime.now().strftime("%Y-%m-%d")
    today_entry = {
        "date": today_str_iso, "total": scorecard_total,
        "valuation_label": valuation["valuation_label"], "decision_label": decision_label,
    }
    # Ghi đè entry NẾU cùng ngày (Action có thể chạy nhiều lần/ngày qua workflow_dispatch thủ
    # công) — tránh phình lịch sử với các điểm trùng ngày vô nghĩa cho việc so sánh xu hướng.
    if scorecard_history and scorecard_history[-1]["date"] == today_str_iso:
        new_history = scorecard_history[:-1] + [today_entry]
    else:
        new_history = scorecard_history + [today_entry]
    raw.setdefault("_meta", {})["scorecard_history"] = new_history[-60:]
    save_vimo_raw(raw)

    print("[INFO] Lấp khoảng trống giải ngân đầu tư công tháng 3/6/9/12 (ước tính, KHÔNG lưu vào vimo_raw.json)...")
    _fill_disbursement_gaps(raw, trends)

    print("[INFO] Tính chỉ báo phái sinh (FDI giải ngân/đăng ký theo kỳ rời rạc, KHÔNG lưu vào vimo_raw.json)...")
    _add_derived_indicators(raw, trends)

    print("[INFO] Tính XK/NK YoY theo tháng từ Hải quan (KHÔNG lưu vào vimo_raw.json)...")
    _add_customs_yoy_growth(raw, trends)

    print("[INFO] Tính spread đường cong lợi suất TPCP Mỹ 10Y-2Y / 10Y-3M (KHÔNG lưu vào vimo_raw.json)...")
    _add_us_yield_curve_spread(raw, trends)

    print("[INFO] Tính tăng trưởng dư nợ tín dụng YoY (theo tháng) + tỷ lệ Tín dụng/GDP (KHÔNG lưu vào vimo_raw.json)...")
    _add_credit_derived_indicators(raw, trends)

    print("[INFO] Dựng bảng giám sát chỉ số vĩ mô hàng tháng (heatmap)...")
    monitoring_table = _build_monitoring_table(raw)
    if monitoring_table:
        print(f"  -> {len(monitoring_table['rows'])} hàng x {len(monitoring_table['periods'])} tháng")
    else:
        print("  -> Chưa đủ dữ liệu tháng để dựng bảng.")

    print("[INFO] Tổng hợp phân tích đa chỉ số (rule-based, dựa trên số liệu thật)...")
    synthesis = build_synthesis_vimo(raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text, verdict)
    print("  -> Đã sinh phân tích tổng hợp")

    out_dir = os.path.join(PROJECT_ROOT, "Bao cao", "VIMO")
    os.makedirs(out_dir, exist_ok=True)
    print("[INFO] Building charts...")
    charts = build_charts_vimo(out_dir, raw)
    interbank_6m_chart = build_interbank_6m_history_chart(out_dir, raw)
    if interbank_6m_chart:
        charts["interbank_6m_history"] = interbank_6m_chart
    interbank_chart = build_interbank_curve_chart(out_dir, raw)
    if interbank_chart:
        charts["interbank_curve"] = interbank_chart
    bond_yield_chart = build_bond_yield_history_chart(out_dir, raw)
    if bond_yield_chart:
        charts["bond_yield_history"] = bond_yield_chart
    gdp_structure_chart = build_stacked_area_chart(out_dir, raw, GDP_STRUCTURE_KEYS,
                                                    "Cơ cấu GDP theo khu vực kinh tế (%)", "vimo_gdp_structure.png")
    if gdp_structure_chart:
        charts["gdp_structure"] = gdp_structure_chart
    investment_structure_chart = build_stacked_area_chart(
        out_dir, raw, INVESTMENT_STRUCTURE_KEYS,
        "Cơ cấu vốn đầu tư thực hiện toàn xã hội theo thành phần (%)", "vimo_investment_structure.png")
    if investment_structure_chart:
        charts["investment_structure"] = investment_structure_chart
    valhist_charts = build_vnindex_valuation_history_charts(out_dir, rf, valuation.get("capm_valuation"))
    if valhist_charts.get("pe"):
        charts["vnindex_pe_history"] = valhist_charts["pe"]
    if valhist_charts.get("pe_exvin"):
        charts["vnindex_pe_exvin_history"] = valhist_charts["pe_exvin"]
    if valhist_charts.get("pb"):
        charts["vnindex_pb_history"] = valhist_charts["pb"]
    if valhist_charts.get("pb_exvin"):
        charts["vnindex_pb_exvin_history"] = valhist_charts["pb_exvin"]
    print(f"  -> {len(charts)} charts")

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    pdf_path = os.path.join(out_dir, f"VIMO_Report_{date_str}.pdf")
    print("[INFO] Building PDF...")
    build_pdf_vimo(pdf_path, raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text,
                    charts, synthesis, verdict,
                    valuation_headline=valuation_headline, decision_label_headline=decision_label_headline,
                    decision_text_headline=decision_text_headline)
    print(f"  [OK] PDF: {pdf_path}")

    print("[INFO] Saving JSON dashboard...")
    save_json_vimo(raw, trends, scorecard, scorecard_total, valuation, decision_label, decision_text, synthesis,
                    valuation_headline=valuation_headline, decision_label_headline=decision_label_headline,
                    decision_text_headline=decision_text_headline, monitoring_table=monitoring_table)

    print("[INFO] Cập nhật Excel lịch sử chỉ số theo tháng...")
    update_excel_history_vimo(raw, out_dir)

    for p in charts.values():
        try:
            os.remove(p)
        except Exception:
            pass

    print(f"\n{'='*60}")
    print(f"  HOÀN THÀNH — PDF: {pdf_path}")
    print(f"{'='*60}\n")
    return True


if __name__ == "__main__":
    run_vimo_analysis()
